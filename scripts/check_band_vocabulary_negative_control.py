#!/usr/bin/env python3
"""Negative control for [R-CAL-02]'s gate: prove it catches the real defect.

A check nobody has seen fail is not evidence. This reinjects the exact text the
site carried on 24-Aug-2026 -- and a band record deliberately disagreeing with
its panel -- into throwaway copies, and asserts the gate goes red on each. If any
case passes, the gate is asleep and this exits nonzero.

It copies only what the gate READS. The first cut did shutil.copytree of the whole
repository once per case -- 1,627 files and 319 MB each, five times, 33 s per copy
on a cold CI checkout -- to mutate one file, while the gate opens 211 files and
7.7 MB. engine/*_study/ and engine/raw_ohlc/ were copied five times and never
opened. Hence the --root argument on the gate.

Also carries a CLEAN arm: legitimate text that must NOT be flagged, following
check_protocol_text_negative_control.py. Without it a gate that flagged
everything would pass this control.
"""
import os
import re
import shutil
import subprocess
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CASES = [
    ("the FAIL banner", "legacy/ledger.html",
     lambda s: s.replace("</body>",
                         '<p>⚠ INDICATIVE ONLY · FAILED CALIBRATION TEST</p></body>', 1)),
    ("a PARITY verdict in a thesis", "assets/coverage.js",
     lambda s: s.replace("const COVERAGE_EN = [",
                         'const X = "\\u00a73 The calibration is PARITY, not skill.";\n'
                         'const COVERAGE_EN = [', 1)),
    ("a matches-benchmark chip", "method.html",
     lambda s: s.replace("</body>", "<p>◆ Indicative · matches benchmark</p></body>", 1)),
    ("CRPS beside a company", "legacy/savola.html",
     lambda s: s.replace("</body>", "<p>SAVOLA scored +0.9% CRPS skill.</p></body>", 1)),
    ("a band record that disagrees with its panel", "assets/data.js",
     lambda s: re.sub(r'(  SAVOLA: \{mkt:"SA", n:)\d+', r'\g<1>999', s, count=1)),
    ("a hand-edited coverage figure", "assets/data.js",
     lambda s: re.sub(r'(  SAVOLA: \{mkt:"SA", n:\d+, hits:)\d+', r'\g<1>58', s, count=1)),
    ("a span naming no record", "legacy/savola.html",
     lambda s: s.replace('data-band-record="SAVOLA"', 'data-band-record="NOSUCHNAME"', 1)),
    # THE RETIRED VERDICT DESCRIBED RATHER THAN NAMED [added 03-Sep-2026]. ARCC's
    # delivered study published "Its skill against a simple random walk is -1.8% —
    # statistically indistinguishable from zero at every block size tested", TWICE,
    # and every pattern this gate carried looked for the TOKENS. A rule saying a
    # measure may never reach a reader is not enforced by banning the word somebody
    # happened to use for it last time. These four are that sentence and its
    # neighbours, injected on a page this gate already reads.
    ("the skill verdict described rather than named", "legacy/savola.html",
     lambda s: s.replace("</body>",
                         "<p>Its skill against a simple random walk is -1.8%.</p></body>", 1)),
    ("the skill verdict as a bare score", "legacy/savola.html",
     lambda s: s.replace("</body>", "<p>The skill score is negative here.</p></body>", 1)),
    ("a claim to beat the benchmark", "legacy/savola.html",
     lambda s: s.replace("</body>", "<p>The cone beats a random walk.</p></body>", 1)),
    ("a comparative against the benchmark", "legacy/savola.html",
     lambda s: s.replace("</body>",
                         "<p>It is worse than a naive random walk.</p></body>", 1)),
    # The figures are images; only their caption TEMPLATE is readable, and this
    # arm is live only because the ratchet is now empty. It is the one that would
    # catch the verdict creeping back into 93 pictures no text check can read.
    ("the verdict back in the figure caption", "engine/metal_backtest.py",
     lambda s: s.replace('    h2 = ("{cov90} of {n} windows',
                         '    h2 = ("PARITY  \u00b7  {cov90} of {n} windows', 1)),
]

# Text that must NOT trip the gate. A check that flags everything is not a check:
# lowercase "parity" is an ordinary word in this book, and CRPS is taught on the
# methodology page on purpose.
CLEAN = [
    # The benchmark may be NAMED as a construction -- the width ratio is published
    # against it and [R-CAL-02] requires that -- so these must NOT fire. A gate that
    # could not tell them apart would make the disclosure the rule mandates
    # impossible to write.
    ("a carry-anchored random walk as the width benchmark", "legacy/savola.html",
     lambda s: s.replace("</body>",
                         "<p>The band ran 1.4 times the width of a carry-anchored "
                         "random walk.</p></body>", 1)),
    ("a currency peg", "legacy/savola.html",
     lambda s: s.replace("</body>", "<p>the riyal's fixed parity to the dollar</p></body>", 1)),
    ("export parity pricing", "legacy/savola.html",
     lambda s: s.replace("</body>", "<p>subsidised vs. export parity feedstock</p></body>", 1)),
    ("CRPS where the rule is taught", "method.html",
     lambda s: s.replace("</body>", "<p>Forecasts are graded with CRPS.</p></body>", 1)),
]

# Only what the gate opens.
COPY = ["*.html", "legacy/*.html", "assets/*.js", "engine/band_record.py", "engine/panels/*_3m.csv",
        "engine/panel_refresh.py", "engine/mc_v3.py", "engine/primitives.py",
        "engine/market_profiles.py", "engine/fv_overlay.py", "engine/__init__.py",
        "engine/build_depth_audit/band_outstanding.json", "engine/metal_backtest.py",
        "engine/primitives.py",
        "scripts/build_market_registry.py", "scripts/build_band_records.py",
        "scripts/check_band_vocabulary.py",
        # ONE DELIVERED WORKBOOK, so the workbook arm added 05-Sep-2026 is LIVE here
        # rather than merely present. Staging none would make that arm refuse an empty
        # population on EVERY case, clean ones included — the gate going red for the
        # WRONG reason, which reads exactly like going red for the right one and is a
        # mistake this repository has already made once inside a sandbox.
        "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx"]


def stage(dst):
    """Copy only the paths the gate reads, preserving layout."""
    import glob as _g
    n = 0
    for pat in COPY:
        for src in _g.glob(os.path.join(ROOT, pat)):
            rel = os.path.relpath(src, ROOT)
            out = os.path.join(dst, rel)
            os.makedirs(os.path.dirname(out), exist_ok=True)
            shutil.copy2(src, out)
            n += 1
    return n


#: Injecting into a workbook is not a string edit, so it gets its own runner. The
#: retired verdict reached FIVE delivered workbooks at their latest edition — two of
#: them built the same week — because every surface this gate reads was swept and the
#: workbook beside the document was read by nothing at all.
WORKBOOK_CASES = [
    ("PARITY in a delivered workbook",
     "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx",
     "Calibration verdict: PARITY"),
    ("the verdict described rather than named, in a workbook",
     "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx",
     "Skill against a random walk: -1.8%"),
    ("CRPS in a workbook",
     "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx",
     "This name's CRPS skill vs a random walk: -0.13"),
]
WORKBOOK_CLEAN = [
    ("the benchmark named as a width construction, in a workbook",
     "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx",
     "Cone width versus a naive carry-anchored band: 1.18x"),
    ("a band record with its count, in a workbook",
     "engine/tmgh_study/TMGH_Valuation_Model_02092026.xlsx",
     "Over 17 resolved forecasts the price finished inside the 90% band 94% of the time"),
]


def run_workbook_case(label, rel, text, must_fail):
    import openpyxl
    with tempfile.TemporaryDirectory() as tmp:
        work = os.path.join(tmp, "repo")
        stage(work)
        path = os.path.join(work, rel)
        if not os.path.exists(path):
            return f"{label}: {rel} was not staged — control is stale"
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]
        before = ws.cell(row=ws.max_row + 2, column=1).value
        ws.cell(row=ws.max_row + 2, column=1, value=text)
        if before == text:
            return f"{label}: could not inject into {rel} — control is stale"
        wb.save(path)
        # PROVE THE MUTATION LANDED before believing anything about the result: a case
        # that silently modified nothing reports a green that means only that the file
        # was untouched, which this repository has shipped once already.
        chk = openpyxl.load_workbook(path)
        if not any(isinstance(c.value, str) and text in c.value
                   for row in chk.worksheets[0].iter_rows() for c in row):
            return f"{label}: the injected text is not in the saved workbook"
        r = subprocess.run([sys.executable,
                            os.path.join(ROOT, "scripts", "check_band_vocabulary.py"),
                            "--root", work], capture_output=True, text=True)
        red = r.returncode != 0
        if red != must_fail:
            what = "PASSED on an injected defect" if must_fail else "FAILED on legitimate text"
            return f"{label}: gate {what} in {rel}"
        print(f"  {'caught' if must_fail else 'allowed'}: {label}  ({os.path.basename(rel)})")
        return None


def run_empty_population_case():
    """[R-ENF-04] An empty population is not a clean one. Removing the only staged
    workbook must make the gate REFUSE rather than report clean — otherwise the whole
    workbook arm could be switched off by a resolver that quietly finds nothing, which
    is the absent answer wearing a clean one's costume."""
    with tempfile.TemporaryDirectory() as tmp:
        work = os.path.join(tmp, "repo")
        stage(work)
        gone = [os.path.join(dp, f)
                for dp, _, fs in os.walk(os.path.join(work, "engine"))
                for f in fs if f.endswith(".xlsx")]
        if not gone:
            return "empty population: no workbook was staged — control is stale"
        for g in gone:
            os.remove(g)
        r = subprocess.run([sys.executable,
                            os.path.join(ROOT, "scripts", "check_band_vocabulary.py"),
                            "--root", work], capture_output=True, text=True)
        if r.returncode == 0:
            return ("empty population: the gate reported CLEAN having read no workbook "
                    "at all")
        if "no delivered workbook was read" not in r.stdout:
            return ("empty population: the gate went red, but not for the population "
                    "reason — it must NAME why, or it is red for the wrong cause")
        print("  caught: an empty workbook population reported as clean  "
              "(all delivered workbooks removed)")
        return None


def run_case(label, rel, mutate, must_fail):
    with tempfile.TemporaryDirectory() as tmp:
        work = os.path.join(tmp, "repo")
        stage(work)
        path = os.path.join(work, rel)
        src = open(path, encoding="utf-8").read()
        out = mutate(src)
        if out == src:
            return f"{label}: could not inject into {rel} — control is stale"
        open(path, "w", encoding="utf-8").write(out)
        r = subprocess.run([sys.executable,
                            os.path.join(ROOT, "scripts", "check_band_vocabulary.py"),
                            "--root", work], capture_output=True, text=True)
        red = r.returncode != 0
        if red != must_fail:
            what = "PASSED on an injected defect" if must_fail else "FAILED on legitimate text"
            return f"{label}: gate {what} in {rel}"
        print(f"  {'caught' if must_fail else 'allowed'}: {label}  ({rel})")
        return None


def main():
    failures = [f for f in
                [run_case(*c, must_fail=True) for c in CASES] +
                [run_case(*c, must_fail=False) for c in CLEAN] +
                [run_workbook_case(*c, must_fail=True) for c in WORKBOOK_CASES] +
                [run_workbook_case(*c, must_fail=False) for c in WORKBOOK_CLEAN] +
                [run_empty_population_case()]
                if f]
    if failures:
        print("NEGATIVE CONTROL FAILED:")
        for f in failures:
            print("  " + f)
        return 1
    print(f"negative control OK — {len(CASES) + len(WORKBOOK_CASES) + 1} defects caught "
          f"({len(WORKBOOK_CASES)} of them in a delivered workbook, plus an emptied "
          f"workbook population), "
          f"{len(CLEAN) + len(WORKBOOK_CLEAN)} legitimate cases allowed through")
    return 0


if __name__ == "__main__":
    sys.exit(main())
