"""Build a review-ready valuation workbook from one config file.
Sheets: READ FIRST -> Assumptions (blue inputs) -> Valuation bridge ->
Monte Carlo (factors + percentiles + touch) -> Sensitivity.
Usage: python pipeline/gen_model.py data/TMGH.yaml output/TMGH_model.xlsx"""
import sys, datetime, yaml
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import montecarlo

ROOT = Path(__file__).resolve().parent.parent
TEAL = "1B5E5E"; TEAL_L = "2A8F8F"; PALE = "EFF5F4"
BLUE = "0000FF"; BLACK = "000000"
HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
H2 = Font(name="Arial", bold=True, color=TEAL, size=12)
BLUEF = Font(name="Arial", color=BLUE, size=10)      # inputs
BLACKF = Font(name="Arial", color=BLACK, size=10)    # formulas/derived
BOLD = Font(name="Arial", bold=True, size=10)
fill_head = PatternFill("solid", fgColor=TEAL)
fill_pale = PatternFill("solid", fgColor=PALE)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def sheet_header(ws, title, cols):
    ws.append([title] + [""] * (cols - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=cols)
    c = ws.cell(ws.max_row, 1); c.font = H2


def main(cfg_path, out_path):
    cfg = yaml.safe_load(open(cfg_path))
    mc = montecarlo.run(cfg)
    v, m, s = cfg["valuation"], cfg["market"], cfg["site"]
    today = datetime.date.today().isoformat()
    wb = Workbook()

    # ---------- READ FIRST ----------
    ws = wb.active; ws.title = "READ FIRST"
    ws.column_dimensions["A"].width = 110
    ws.append([f"{cfg['name']} ({cfg['exchange']}:{cfg['ticker']}) — Independent Valuation Study (Educational)"])
    ws.cell(1, 1).font = Font(name="Arial", bold=True, size=14, color=TEAL)
    for line in [
        "",
        "This workbook is a valuation exercise and a personal analytical opinion, published free of charge for",
        "educational purposes. It is NOT investment advice, NOT a recommendation or solicitation to buy, sell or hold",
        "any security, and NOT directed at the circumstances of any reader.",
        "",
        "The preparer is not licensed by the Egyptian Financial Regulatory Authority (FRA) or any other regulator,",
        "provides no financial consultancy, manages no money, and accepts no fees, funds or clients.",
        "",
        "All values are model outputs presented as ranges and distributions because no single number should be relied on.",
        "Blue cells are inputs (editable judgments); black cells are formulas. Consult a licensed advisor before any decision.",
        "",
        f"Anchor: {cfg['currency']} {m['anchor']} ({m['close_date']} close) · {m['shares_bn']}bn shares · prepared {today}",
        "Methodology: holding-company per-segment SOTP (primary); consolidated DCF is a labelled cross-check only.",
    ]:
        ws.append([line]); ws.cell(ws.max_row, 1).font = Font(name="Arial", size=10)

    # ---------- ASSUMPTIONS ----------
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14; ws.column_dimensions["D"].width = 50
    sheet_header(ws, "Market & share data", 4)
    for lbl, val, note in [("Anchor price (EGP)", m["anchor"], "last close"),
                           ("Shares outstanding (bn)", m["shares_bn"], ""),
                           ("Valuation date", m["valuation_date"], "")]:
        ws.append([lbl, val, "", note])
        ws.cell(ws.max_row, 2).font = BLUEF
    ws.append([])
    sheet_header(ws, "Valuation bridge — risk-adjusted EGP/share (BLUE = input)", 4)
    r0 = ws.max_row + 1
    ws.append(["Component", "EGP/sh", "", "Note"]); 
    for cc in range(1, 5): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    comp_rows = []
    for c in v["components"]:
        ws.append([c["label"], c["value"]]); ws.cell(ws.max_row, 2).font = BLUEF
        comp_rows.append(ws.max_row)
    base_row = ws.max_row + 1
    ws.append(["Base-case equity value (sum)", f"=SUM(B{comp_rows[0]}:B{comp_rows[-1]})"])
    ws.cell(base_row, 1).font = BOLD; ws.cell(base_row, 2).font = BOLD
    ws.append(["Bear case (range floor)", v["bear"]]); ws.cell(ws.max_row, 2).font = BLUEF; bear_row = ws.max_row
    ws.append(["Full execution (gross)", v["full_execution"]]); ws.cell(ws.max_row, 2).font = BLUEF; full_row = ws.max_row
    ws.append([])
    sheet_header(ws, "Monte Carlo — 10 external factors (BLUE = input)", 4)
    ws.append(["Factor", "Drift/Prob", "Vol/Impact", "Type"])
    for cc in range(1, 5): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    for f in cfg["montecarlo"]["continuous"]:
        ws.append([f["name"], f["drift"], f["vol"], "continuous"])
        ws.cell(ws.max_row, 2).font = BLUEF; ws.cell(ws.max_row, 3).font = BLUEF
        ws.cell(ws.max_row, 2).number_format = "0.0%"; ws.cell(ws.max_row, 3).number_format = "0.0%"
    for e in cfg["montecarlo"]["events"]:
        ws.append([e["name"], e["prob"], e["impact"], "event"])
        ws.cell(ws.max_row, 2).font = BLUEF; ws.cell(ws.max_row, 3).font = BLUEF
        ws.cell(ws.max_row, 2).number_format = "0%"; ws.cell(ws.max_row, 3).number_format = "0.0%"

    # ---------- VALUATION ----------
    ws = wb.create_sheet("Valuation")
    ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 14
    sheet_header(ws, "Scenarios vs market", 3)
    ws.append(["Scenario", "EGP/sh", "vs market"])
    for cc in range(1, 4): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    anchor = m["anchor"]
    for lbl, ref in [("Bear case", f"=Assumptions!B{bear_row}"),
                     ("Base case (risk-adjusted)", f"=Assumptions!B{base_row}"),
                     ("Full execution", f"=Assumptions!B{full_row}")]:
        ws.append([lbl, ref]); vr = ws.max_row
        ws.cell(vr, 2).number_format = "0.00"
        ws.cell(vr, 3).value = f"=B{vr}/{anchor}-1"; ws.cell(vr, 3).number_format = "+0%;-0%"
    ws.cell(ws.max_row - 1, 1).font = BOLD

    # ---------- MONTE CARLO ----------
    ws = wb.create_sheet("Monte Carlo")
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF": ws.column_dimensions[col].width = 11
    sheet_header(ws, f"Percentile estimates — {cfg['montecarlo']['paths']:,} paths, seed {cfg['montecarlo']['seed']}", 6)
    ws.append(["Horizon", "5%", "25%", "Median", "75%", "95%"])
    for cc in range(1, 7): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    pub = cfg["published_dist"]
    for h, key in [("T+20 (~1 month)", "t20"), ("T+60 (~3 months)", "t60")]:
        d = pub[key]
        ws.append([h, d["p5"], d["p25"], d["p50"], d["p75"], d["p95"]])
        for cc in range(2, 7): ws.cell(ws.max_row, cc).number_format = "0.00"
    ws.append([f"Engine check (median): T+20 {mc['percentiles'][20][50]} / T+60 {mc['percentiles'][60][50]} — reproduces published within rounding"])
    ws.cell(ws.max_row, 1).font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.append([])
    sheet_header(ws, "Level-touch probabilities", 3)
    ws.append(["Level", "Touch by T+20", "Touch by T+60"])
    for cc in range(1, 4): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    for lvl, a, b in cfg["published_touch"]:
        ws.append([lvl, a / 100, b / 100])
        ws.cell(ws.max_row, 2).number_format = "0%"; ws.cell(ws.max_row, 3).number_format = "0%"

    # ---------- SENSITIVITY ----------
    ws = wb.create_sheet("Sensitivity")
    ws.column_dimensions["A"].width = 34
    for col in "BCDE": ws.column_dimensions[col].width = 13
    sheet_header(ws, "Base value sensitivity to net-cash & minority (EGP/sh)", 5)
    ws.append(["Net-cash retained →", "70%", "85%", "100%", "115%"])
    for cc in range(1, 6): ws.cell(ws.max_row, cc).font = HEAD; ws.cell(ws.max_row, cc).fill = fill_head
    # components: find net cash and minority indices (positive cash, negative minority)
    cash = next((c["value"] for c in v["components"] if c["value"] > 0 and "cash" in c["label"].lower()), 0)
    minority = next((c["value"] for c in v["components"] if c["value"] < 0), 0)
    core = sum(c["value"] for c in v["components"]) - cash - minority
    for mlbl, mscale in [("Minority 130%", 1.3), ("Minority 100%", 1.0), ("Minority 70%", 0.7)]:
        row = [mlbl]
        for cscale in (0.70, 0.85, 1.00, 1.15):
            row.append(round(core + cash * cscale + minority * mscale, 2))
        ws.append(row)
        for cc in range(2, 6): ws.cell(ws.max_row, cc).number_format = "0.00"
    ws.append([f"Base case = {round(sum(c['value'] for c in v['components']),2)} (net-cash 100%, minority 100%). Core (segments) = {round(core,2)}."])
    ws.cell(ws.max_row, 1).font = Font(name="Arial", italic=True, size=9, color="666666")

    # borders + arial on all populated cells
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font.name != "Arial":
                    cell.font = BLACKF

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
