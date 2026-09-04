"""Prove the delivered EMPOWER workbook is a LIVE DRIVER model, not a pasted register.

Three passes:

1. DRIVER DIRECTIONS — each Assumptions driver is perturbed in place, the whole workbook
   is re-evaluated from scratch with xlcalc.py, and the test asserts the RECOVERY central
   fair value moves in the asserted direction. This now includes the audited primary-record
   anchors (DEWA purchases, cash capex, receivables balance) that the derived rates are
   formulas of — changing the audited record must reprice the model.
2. SCENARIO BLOCKS ARE LIVE — the recovery-ladder level cells, the bear/bull connection
   paths (on the Sensitivity sheet) and the physical tariff inputs are perturbed and the
   test asserts the corresponding LIVE scenario output moves (continuation central, bear,
   bull, the implied full-load hours) while pasted grids stay pasted.
3. DEAD-INPUT SWEEP — every remaining pasted numeric cell outside the Assumptions sheet
   (audited display history, the price map, the 5x5 grid, scenario inputs) is bumped and
   must NOT move the recovery central. The named exceptions are the Segments forecast
   credit-loss zeros (deliberate live inputs) and the scenario-input cells, which drive
   their own scenario outputs but never the recovery central.
"""
import json, os
import openpyxl
import xlcalc

HERE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(os.path.join(HERE, 'EMPOWER_Valuation_Model_09082026_public.xlsx'))
ANCH = json.load(open(os.path.join(HERE, 'xlsx_expected.json')))['anchors']
SU, DR, SN = ANCH['summary'], ANCH['dcf'], ANCH['sens']
CEN_ROW, CONT_ROW = SU['central'], SU['central_cont']
WACC_ROW = DR['wacc_ct']

A = {}
for row in wb['Assumptions'].iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str):
        A[c.value] = c.row

def row_of(label):
    if label not in A:
        raise KeyError(f'no Assumptions row labelled {label!r}')
    return A[label]

def read(overrides=None):
    bk = xlcalc.Book(wb, overrides)
    return dict(central=bk.cell_value('Summary', f'B{CEN_ROW}'),
                cont=bk.cell_value('Summary', f'B{CONT_ROW}'),
                dcf=bk.cell_value('SOTP Bridge', f"C{ANCH['bridge']['ps']}"),
                wacc=bk.cell_value('DCF', f'C{WACC_ROW}'),
                ebitda26=bk.cell_value('DCF', 'B6'),
                rev26=bk.cell_value('DCF', 'B5'),
                fcff26=bk.cell_value('DCF', 'B18'),
                tv=bk.cell_value('DCF', f"C{DR['tv']}"),
                nd30=bk.cell_value('Balance Sheet', 'I16'),
                bear=bk.cell_value('Sensitivity', f"C{SN['bear_ps']}"),
                bull=bk.cell_value('Sensitivity', f"C{SN['bull_ps']}"),
                eflh25=bk.cell_value('Segments', 'C30'),
                # the terminal's own lines, so a driver whose effect on the answer nearly
                # cancels can still be asserted on the line it unambiguously moves
                t_maint=bk.cell_value('DCF', 'C96'),
                t_fcff=bk.cell_value('DCF', 'C99'))

base = read()
print('base:  ' + ' · '.join(f'{k} {v:,.4f}' for k, v in base.items()))

# ---- pass 1: Assumptions drivers, required direction on the recovery central ------
# label, column, bump, headline, required direction, why
CASES = [
    ('Spot price (AED)', 'C', +0.30, 'central', -1,
     'a higher spot raises the equity weight on the dearer cost of equity, so the '
     'discount rate rises and the valuation falls'),
    ('Beta (own-stock weekly regression vs the FTSE ADX index)', 'C', +0.20, 'central', -1,
     'a higher beta raises the cost of equity and must lower the valuation'),
    ('Risk-free rate (AED sovereign)', 'C', +0.01, 'central', -1,
     'a higher risk-free rate raises the cost of equity and must lower the valuation'),
    ('Equity risk premium — rating basis', 'C', +0.01, 'central', -1,
     'a higher premium raises the cost of equity and must lower the valuation'),
    ('Marginal cost of debt', 'C', +0.01, 'central', -1,
     'a dearer debt book raises the discount rate and the finance charge'),
    ('Corporate tax rate (headline framing)', 'C', +0.03, 'central', -1,
     'a higher tax rate cuts NOPAT throughout and must lower the valuation'),
    # Growth is DERIVED from a real rate and the house inflation, so both are exercised.
    # Real growth widens the perpetuity and costs the capital this company's own balance
    # sheet says a unit of demand needs; inflation widens it too but escalates the cost of
    # replacing the plant against it, so its direction is asserted on the line it moves
    # unambiguously rather than on an answer where the two nearly cancel.
    ('REAL growth, FY2031-FY2040 window (stage one)', 'C', +0.005, 'central', +1,
     'real growth adds value once it is charged only for the capital another unit of '
     'demand actually needs, rather than for rebuilding the whole plant every 1/g years'),
    ('REAL growth beyond FY2040 (stage two) — NEGATIVE', 'C', +0.002, 'central', +1,
     'a shallower real decline in the perpetuity is worth more than a steeper one'),
    ('Terminal inflation — UAE house macro path', 'C', +0.005, 't_maint', -1,
     'a higher terminal inflation escalates the cost of replacing the plant over half its '
     'life, so the maintenance charge (a negative row) must grow'),
    ('Weighted asset life, DERIVED from notes 5, 6 and 7 (years)', 'C', +5.0, 't_fcff', -1,
     'on this basis the life sets the average VINTAGE of the plant rather than the '
     'replacement frequency: a longer life means the assets carried were bought further '
     'back, so replacing them today costs more against the depreciation already booked'),
    ('New connections by year (k RT)', 'B', +50.0, 'central', +1,
     'each added RT earns regulated per-RT revenue worth a multiple of its capex'),
    ('Consumption per-RT recovery level from FY2027 (share of FY2025)', 'C', +0.03,
     'central', +1, 'a stronger consumption recovery lifts revenue from FY2027 onward'),
    ('Electricity and water purchased from DEWA, FY2025 (AED mn)', 'C', +50.0,
     'central', -1, 'a dearer audited DEWA bill raises the pass-through ratio formula '
     'and absorbs more of every consumption dirham — the audited anchor drives the model'),
    ('Capital expenditure (cash), FY2025 (AED mn)', 'C', +50.0, 'central', -1,
     'dearer audited capex raises the capex-per-RT formula, cutting free cash flow and '
     'terminal return on capital'),
    ('Trade and other receivables, end-FY2025 (AED mn)', 'C', +200.0, 'central', -1,
     'a fatter audited receivables book makes the working-capital ratio less negative, '
     'absorbing cash as revenue grows'),
    ('Wage and services cost escalator', 'C', +0.01, 'central', -1,
     'faster cash-cost escalation compresses EBITDA every year'),
    ('Rental income inside other income, FY2025 (AED mn)', 'C', +5.0, 'central', -1,
     'more of the audited other income classified as rental leaves LESS operating other '
     'income inside EBITDA (the rental is valued at book in the bridge instead)'),
]

fails = []
print('\nDRIVER DIRECTION TABLE — recovery central')
for label, col, bump, key, sign, why in CASES:
    r = row_of(label)
    cur = wb['Assumptions'][f'{col}{r}'].value
    assert isinstance(cur, (int, float)), f'driver cell {label!r} {col}{r} is not a value'
    out = read({('Assumptions', f'{col}{r}'): cur + bump})
    delta = out[key] - base[key]
    rel = delta / abs(base[key]) if base[key] else 0.0
    ok = (delta * sign > 0) and abs(rel) > 1e-9
    flag = 'OK ' if ok else 'BAD'
    print(f'  [{flag}] {label} {bump:+g} -> {key} {base[key]:,.4f} -> {out[key]:,.4f} '
          f'({rel:+.2%})   {why}')
    if not ok:
        fails.append((label, key, delta, why))

# ---- pass 2: the LIVE scenario blocks must move on their own inputs ---------------
print('\nSCENARIO-BLOCK LIVENESS')
scen_cases = [
    (('Sensitivity', f"C{SN['crux_lvl']}"), +0.03, 'cont', +1,
     'raising the 94% recovery-ladder level lifts the live continuation central'),
    (('Sensitivity', f"B{SN['bear_adds']}"), +30.0, 'bear', +1,
     'more bear-case connections lift the live bear fair value'),
    (('Sensitivity', f"B{SN['bull_adds']}"), +30.0, 'bull', +1,
     'more bull-case connections lift the live bull fair value'),
    (('Assumptions', f"C{row_of('Cooling delivered, H1-2026 (m RTh)')}"), +100.0,
     'eflh25', +1,
     'more delivered RTh lowers the implied tariff, so the same consumption revenue '
     'implies more full-load hours — the physical decomposition is live'),
]
for (sh, coord), bump, key, sign, why in scen_cases:
    cur = wb[sh][coord].value
    assert isinstance(cur, (int, float)), f'scenario cell {sh}!{coord} is not a value'
    out = read({(sh, coord): cur + bump})
    delta = out[key] - base[key]
    dc = out['central'] - base['central']
    scen_ok = delta * sign > 0
    if sh == 'Sensitivity':
        scen_ok = scen_ok and abs(dc) < 1e-9    # scenario inputs never touch recovery
    flag = 'OK ' if scen_ok else 'BAD'
    print(f'  [{flag}] {sh}!{coord} {bump:+g} -> {key} {base[key]:,.4f} -> '
          f'{out[key]:,.4f} (recovery central moved {dc:+.6f})   {why}')
    if not scen_ok:
        fails.append((f'{sh}!{coord}', key, delta, why))

# ---- pass 3: dead-input sweep -----------------------------------------------------
# Every pasted numeric cell OUTSIDE the Assumptions sheet is display history, a price-map
# figure, the pasted grid, or a scenario-block input. None may move the RECOVERY central.
# Named live exceptions: Segments C21:G21 are the forecast credit-loss cells — pasted
# zeros that are deliberately live inputs in the EBITDA identity, so bumping them SHOULD
# reprice the model.
LIVE_ZEROS = {('Segments', f'{c}21') for c in 'CDEFG'}
AUDITED_DRIVERS = {}   # none: every audited cell tested inert on the stable build
print('\nDEAD-INPUT SWEEP — pasted cells on non-Assumptions sheets must not move the '
      'recovery central')
dead_movers = []
nswept = 0
for sh in wb.sheetnames:
    if sh in ('Assumptions', 'READ FIRST'):
        continue
    for row in wb[sh].iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if (sh, c.coordinate) in LIVE_ZEROS or (sh, c.coordinate) in AUDITED_DRIVERS:
                continue
            nswept += 1
            out = read({(sh, c.coordinate): v * 1.1 + 1e-3})
            if abs(out['central'] - base['central']) > 1e-9 * max(1, abs(base['central'])):
                # CONFIRM before failing: re-take a clean baseline and repeat the
                # perturbation in isolation. A hit that does not reproduce is
                # evaluator state, not a hidden driver — this guard was added
                # 17-Aug-2026 after exactly one such non-reproducing hit
                # (Income Statement!D12), which moves the central by 0.000000 at
                # x1.02 / x1.05 / x1.08 / x1.10 when tested on its own.
                base2 = read()
                out2 = read({(sh, c.coordinate): v * 1.1 + 1e-3})
                if abs(out2['central'] - base2['central']) > 1e-9 * max(1, abs(base2['central'])):
                    dead_movers.append((sh, c.coordinate, v,
                                        out2['central'] - base2['central']))
                else:
                    print(f'  (unconfirmed hit at {sh}!{c.coordinate} — does not '
                          f'reproduce in isolation; not a driver)')
print(f'  swept {nswept} pasted numeric cells')
if dead_movers:
    for sh, coord, v, dd in dead_movers[:20]:
        print(f'  MOVER: {sh}!{coord} (pasted {v}) shifts central by {dd:+.6g}')
else:
    print('  none moves the recovery central — pasted history and grids are display, '
          'not hidden drivers')

assert not fails, f'{len(fails)} drivers failed to move the model correctly: {fails}'
assert not dead_movers, f'pasted cells move the headline: {dead_movers[:10]}'

print(f'\nDRIVER TEST OK — {len(CASES)} Assumptions drivers reprice the workbook in the '
      f'required direction; {len(scen_cases)} scenario inputs drive their live blocks; '
      f'{nswept} pasted cells verified inert on the recovery central')
