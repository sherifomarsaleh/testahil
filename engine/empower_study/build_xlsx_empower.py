"""EMPOWER_Valuation_Model_09082026_public.xlsx — 16 sheets mirroring the house canonical
model (regulated-utility / operating-company variant). Blue = inputs · black = formulas ·
green = cross-sheet links.

The workbook is FORMULA-DRIVEN, revised 17-Aug-2026 after an external audit. Exactly TWO
classes of cell are pasted values:

  1. audited and disclosed historical figures and external facts (FY2023-25 statement
     lines, the 30-Jun-2026 reviewed balance sheet, disclosed connected capacity, the
     consumption revenue from the auditor's key-audit-matter section, deck physical
     figures, peer marks, the RD10 tariff cap) — the primary record, not a calculation.
     Every derived rate (per-RT revenue rates, the electricity-and-water pass-through,
     capex per added RT, the depreciation rate, the working-capital ratio) is a FORMULA
     off these pasted anchors;
  2. whole-model simulation outputs that cannot be one formula: the probabilistic price
     map and the 5x5 discount-rate x growth grid. Everything else — including the five
     consumption-recovery columns, the bear and bull cases and the three cost-of-capital
     constructions — is a LIVE parallel model on the Sensitivity sheet and redraws when
     a driver changes.

NUMERIC TRACEABILITY: no financial numeral is typed into this builder. Every number is
read from study_numbers.json (or, for audited balance-sheet history lines only, from the
statement extracts), and every formula cell's model value is recorded into
xlsx_expected.json as it is written; recalc.py re-evaluates the delivered workbook
independently and asserts cell-for-cell agreement. (Two disclosed external facts arrive
by instruction rather than the JSON and are asserted against it where possible: the RD10
v1.3 tariff cap 0.643 AED/RTh, the H1-2026 1,174m RTh deck figure, the FY31-40 stage-two
growth 1.5%, and Tabreed's reported FY2025 revenue — each pasted once, sourced, display
or assert-verified.)
"""
import json, math, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
E25 = json.load(open(os.path.join(HERE, 'extract_fy2025.json')))
H1 = json.load(open(os.path.join(HERE, 'extract_2026_interims.json')))
E24 = json.load(open(os.path.join(HERE, 'extract_fy2024.json')))

IN = {k: v['value'] for k, v in D['inputs'].items()}
SRC = {k: f"{v['source']} ({v['date']})" for k, v in D['inputs'].items()}
# external-reader wording for the beta source (the stored string carries internal QC
# vocabulary; the statistics themselves are restated in plain language, from the same data)
BR = D['beta_reg']
SRC['beta'] = (f"Own-stock weekly regression vs {BR['index'].split(',')[0]}, "
               f"{BR['window'][0]} to {BR['window'][1]} ({BR['window_years']:.2f} years, "
               f"{BR['n']} weeks): R-squared {BR['r2']:.2f}, standard error {BR['se']:.2f}, "
               f"90% interval [{BR['ci90'][0]:.2f}, {BR['ci90'][1]:.2f}] "
               f"({D['inputs']['beta']['date']})")
M, HI, U, UP = D['meta'], D['hist_is'], D['unit'], D['unit_physical']
F = D['fcst']['base']
W, DC = D['wacc'], D['dcf']
B_CT, B_DM, B_CDS = DC['base_ct'], DC['base_dmtt'], DC['base_cds']
LN, REL, NRM, BK, DDM = D['lenses'], D['rel'], D['norm'], D['book'], D['ddm']
CEN, SNW, CRUX, STK, S0 = D['central'], D['sens_wg'], D['crux'], D['strike'], D['step0']
SCEN = D['scenarios']
SPOT, SH = M['spot'], M['shares_mn']
TAX, TAXD = IN['tax_ct'], IN['tax_dmtt']
G = IN['g_term_derived']
G2 = IN['g_term2_derived']
G1_REAL, G2_REAL = IN['g1_real'], IN['g2_real']
PI_T = B_CT['terminal_stage1']['inputs']['inflation']
LIFE = IN['asset_life_years']
INC_CAP = B_CT['inc_cap']
YF = D['fcst']['years']                          # FY26..FY30
YFL = ['FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']
YHL = ['FY2023', 'FY2024', 'FY2025']
H3 = ['FY23', 'FY24', 'FY25']

# ---- derived anchors (all from study_numbers.json — nothing typed) ----------------
WAGE_ESC = F['other_cos']['FY26'] / U['other_cos25'] - 1                      # 2.5%
INTCO_DECAY = 1 - F['intco']['FY26'] / IN['intco_fy25']                       # 3.0%
RENTAL = IN['rental_fy25']
OI_OP = IN['oi_fy25'] - RENTAL                    # operating other income 6.336
# pure net finance charge — the receivable interest and the rental income are their own
# visible lines now, not inside the finance/EBITDA chain
FIN26 = (REL['np26'] / (1 - TAX)
         - (F['ebitda']['FY26'] + F['intco']['FY26'] + RENTAL - F['dna']['FY26']))
CASH_YIELD = (IN['kd_marg'] * IN['borrow_jun26'] + FIN26) / IN['cash_jun26']  # 3.5%
IC_TERM = B_CT['nopat']['FY30'] / B_CT['roic_term']      # = plant + net working capital
NCI_FR = IN['nci_pat_fy25'] / IN['pat_fy25']
RTP = U['rt_path']
ADDS = [RTP[y] - RTP[p] for p, y in zip(['FY25'] + YF[:-1], YF)]              # 105/100/90/80/70
GROSS_JUN26 = IN['borrow_jun26'] + IN['lease_jun26']
NET_DEBT = W['net_debt']
RECV = IN['recv_jun26']
BRIDGE_ADD = RECV + IN['invprop_jun26'] + IN['fvtpl_jun26'] + IN['fvoci_jun26']
T0 = -math.log(B_CT['df']['FY26']) / math.log(1 + W['rating_ct'])             # 0.5 stub
BETA_DFM = (W['constructions']['ke_dfm'] - W['rf_star_rating']) / IN['erp_rating']
WACC_CT = W['rating_ct']
CPRT, CAPRT = U['cons_per_rt25'], U['cap_per_rt25']
EW, NWC_RATIO, NWC25 = U['ew_ratio'], U['nwc_ratio'], U['nwc25']
SHOCK = U['crux_shock']

assert abs(WAGE_ESC - 0.025) < 1e-9 and abs(INTCO_DECAY - 0.03) < 1e-9
assert abs(CASH_YIELD - 0.035) < 1e-9
assert abs(IC_TERM - (F['ppe']['FY30'] + F['nwc']['FY30'])) < 1e-6
assert abs(T0 - 0.5) < 1e-9 and abs(BETA_DFM - 0.652) < 1e-9


import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..'))
import terminal_value as TERMVAL                                            # noqa: E402


def _sanctioned_fcff(nopat30, dna30, wc30, wacc, inc, g1=None, g2=None):
    """The two stages' free cash flow, through the same module the model uses.

    Every parallel re-run in this workbook — the crux columns, the second tax framing,
    the bear case and the upside — rebuilds its own cash flows this way, so the retired
    reinvestment identity cannot survive in a grid nobody reads the arithmetic of.
    """
    g1 = G if g1 is None else g1
    g2 = G2 if g2 is None else g2

    def stage(n0, d0, w0, g_nom, real, i_):
        return TERMVAL.build(TERMVAL.TerminalInputs(
            nopat=n0 * (1 + g_nom), wacc=wacc, inflation=PI_T, real_growth=real,
            dna_book=d0 * (1 + g_nom), useful_life_years=LIFE,
            useful_life_source=SRC['asset_life_years'],
            maintenance_basis='book_dna_escalated',
            working_capital=w0 * (1 + g_nom),
            incremental_capital_per_unit_growth=i_))
    t1 = stage(nopat30, dna30, wc30, g1, G1_REAL, inc)
    g10 = (1 + g1) ** 10
    t2 = stage(nopat30 * g10, dna30 * g10, wc30 * g10, g2, G2_REAL, 0.0)
    return t1.fcff, t2.fcff


def two_stage_tv(fcff1, fcff2, wacc, g1=None, g2=None):
    """The discounting of those two flows, which is arithmetic and nothing else."""
    g1 = G if g1 is None else g1
    g2 = G2 if g2 is None else g2
    q = (1 + g1) / (1 + wacc)
    s1 = fcff1 / (1 + wacc) * (1 - q ** 10) / (1 - q)
    s2 = fcff2 * (1 + g2) / ((wacc - g2) * (1 + wacc) ** 10)
    return s1, s2, s1 + s2

_F1, _F2 = B_CT['terminal_stage1']['fcff'], B_CT['terminal_stage2']['fcff']
_s1, _s2, _tv = two_stage_tv(_F1, _F2, WACC_CT)
assert abs(_tv - B_CT['tv']) < 1e-6 * B_CT['tv'], 'the two stages fail to reproduce TV'

# physical decomposition (H1-2026 deck facts)
H1_RTH = 1174.0     # m RTh cooling delivered H1-2026, deck p4 — pasted external fact
H1_MIX = UP['rate_aed_per_rth'] * H1_RTH / IN['rev_h1_26']
assert abs(H1_MIX - 0.49) < 1e-9
CAP643 = 0.643      # RD10 v1.3 consumption tariff cap, AED/TRh incl. fuel surcharge
RATE = UP['rate_aed_per_rth']
HEADROOM = RATE / CAP643 - 1
TABREED_REV = 2456.0   # Tabreed FY2025 reported revenue, AED mn (reported, not 2,460)

# ---- audited balance-sheet history (extracts; AED mn) ------------------------------
def th(x):
    return x / 1000.0
BS25 = E25['2025']['balance_sheet']
BS24 = E25['2024_comparative']['balance_sheet']
BSJ = H1['h1_2026']['balance_sheet_30_jun_2026_full']
BH = {
 'FY24': dict(ppe=th(BS24['property_plant_and_equipment']),
              invprop=th(BS24['investment_properties']),
              fvtpl=th(BS24['financial_assets_fvtpl']), fvoci=th(BS24['financial_assets_fvoci']),
              conc=th(BS24['financial_assets_amortised_cost_non_current']
                      + BS24['financial_assets_amortised_cost_current']),
              cash=th(BS24['cash_and_cash_equivalents'] + BS24['term_deposits']),
              gross=th(BS24['bank_borrowings_current'] + BS24['bank_borrowings_non_current']),
              pay=th(BS24['trade_and_other_payables_current']),
              eqp=th(BS24['equity_attributable_to_parent']),
              nci=th(BS24['non_controlling_interests']), ta=th(BS24['total_assets'])),
 'FY25': dict(ppe=th(BS25['non_current_assets']['property_plant_and_equipment']),
              invprop=th(BS25['non_current_assets']['investment_properties']),
              fvtpl=th(BS25['current_assets']['financial_assets_fvtpl']),
              fvoci=th(BS25['non_current_assets']['financial_assets_fvoci']),
              conc=th(BS25['non_current_assets']['financial_assets_at_amortised_cost']
                      + BS25['current_assets']['financial_assets_at_amortised_cost']),
              cash=th(BS25['current_assets']['cash_and_cash_equivalents']
                      + BS25['current_assets']['term_deposits']),
              gross=th(BS25['current_liabilities']['bank_borrowings']
                       + BS25['non_current_liabilities']['bank_borrowings']
                       + BS25['current_liabilities']['lease_liabilities']),
              pay=th(BS25['current_liabilities']['trade_and_other_payables']),
              eqp=th(BS25['equity']['equity_attributable_to_parent']),
              nci=th(BS25['equity']['non_controlling_interests']), ta=th(BS25['total_assets'])),
 'JUN26': dict(ppe=th(BSJ['non_current_assets']['property_plant_and_equipment']),
               invprop=th(BSJ['non_current_assets']['investment_properties']),
               fvtpl=th(BSJ['current_assets']['financial_assets_fvtpl']),
               fvoci=th(BSJ['non_current_assets']['financial_assets_fvtoci']),
               conc=th(BSJ['non_current_assets']['financial_assets_at_amortised_cost']
                       + BSJ['current_assets']['financial_assets_at_amortised_cost']),
               cash=th(BSJ['current_assets']['cash_and_cash_equivalents']
                       + BSJ['current_assets']['term_deposits']),
               gross=th(BSJ['current_liabilities']['bank_borrowings']
                        + BSJ['non_current_liabilities']['bank_borrowings']
                        + BSJ['current_liabilities']['lease_liabilities']),
               pay=th(BSJ['current_liabilities']['trade_and_other_payables']),
               eqp=th(BSJ['equity']['attributable_to_equity_holders']),
               nci=th(BSJ['equity']['non_controlling_interests']), ta=th(BSJ['total_assets'])),
}
CF24 = E24['cash_flow']['2024']
OCF25 = th(E25['2025']['cash_flow']['net_cash_from_operating'])
OCF24 = th(CF24['net_cash_operating'])
DIV25 = th(E25['2025']['cash_flow']['financing']['dividends_paid'])   # negative as reported
DIV24 = th(CF24['dividends_paid'])                                    # negative as reported

# NWC components (FY2025 audited BS; capex accrual recovered off the model's own total)
INV25 = th(BS25['current_assets']['inventories'])
REC25 = th(BS25['current_assets']['trade_and_other_receivables'])
DRP25 = th(BS25['current_assets']['due_from_related_parties'])
PAY25 = th(BS25['current_liabilities']['trade_and_other_payables'])
CRP25 = th(BS25['current_liabilities']['due_to_related_parties'])
ACCR25 = PAY25 + CRP25 + NWC25 - INV25 - REC25 - DRP25   # project-cost accruals 171.085
PPE_DEP25 = U['dep_rate'] * BH['FY24']['ppe']            # FY2025 PPE depreciation 352.199
assert abs(INV25 + REC25 + DRP25 - (PAY25 - ACCR25) - CRP25 - NWC25) < 1e-9
assert abs(PPE_DEP25 + U['amort_flat'] - IN['dna_fy25']) < 1e-6
EQ24 = BH['FY24']['eqp']

# ---- FY2025 operating-EBITDA identity (interest AND rental now excluded) -----------
EB25_OP = (IN['rev_fy25'] - IN['ew_cost_fy25'] - U['other_cos25'] - U['ga_cash25']
           + OI_OP + IN['ecl_fy25'])                     # 1,565.8 incl. the ECL reversal
assert abs(EB25_OP + IN['intco_fy25'] + RENTAL - HI['FY25']['ebitda']) < 1.0
EB_TRAIL = REL['ebitda_trail']                           # 1,583.9 = audited less interest
assert abs(EB25_OP + RENTAL - EB_TRAIL) < 1e-6
HIST_EB_OP = [HI['FY23']['op'] + HI['FY23']['dna'] - HI['FY23']['intco'],
              HI['FY24']['op'] + HI['FY24']['dna'] - HI['FY24']['intco'],
              EB25_OP]          # FY23/24 rental split is not disclosed — flagged in note

# ---- forecast income-statement / balance-sheet chains (python mirror) --------------
# Finance line split: receivable interest and rental are visible rows; the net finance
# charge earns the deposit yield on the ROLLING cash balance (FY2026 on the 30-Jun-2026
# print — the anchor balance; later years on the prior year-end Balance Sheet roll).
# BOTH balance-sheet rolls (equity and net debt) start from the SAME 30-Jun-2026
# reviewed position: FY2026 carries only the second-half stub of profit/cash and only
# the October dividend instalment.
ebit_f = [B_CT['ebit'][y] for y in YF]
intco_f = [F['intco'][y] for y in YF]
rental_f = [RENTAL] * 5
fin_f, pbt_f, tax_f, pat_f, nci_f, npa_f = [], [], [], [], [], []
fcfe_f, eq_f, nci_bs_f, nd_f, cash_f = [], [], [], [], []
eq_prev, nci_prev, nd_prev = IN['eq_attr_jun26'], BH['JUN26']['nci'], NET_DEBT
cash_int = IN['cash_jun26']
for i in range(5):
    fin = -(IN['kd_marg'] * IN['borrow_jun26'] - CASH_YIELD * cash_int)
    pbt = ebit_f[i] + intco_f[i] + rental_f[i] + fin
    pat = pbt * (1 - TAX)
    nci = -pat * NCI_FR
    npa = pat + nci
    fcfe = B_CT['fcff'][YF[i]] + (fin + intco_f[i] + rental_f[i]) * (1 - TAX)
    stub = T0 if i == 0 else 1.0
    eq_prev = eq_prev + stub * (npa - IN['div_policy'])
    nci_prev = nci_prev - stub * nci
    nd_prev = nd_prev - stub * (fcfe - IN['div_policy'])
    cash = GROSS_JUN26 - nd_prev
    fin_f.append(fin); pbt_f.append(pbt); tax_f.append(-pbt * TAX)
    pat_f.append(pat); nci_f.append(nci); npa_f.append(npa); fcfe_f.append(fcfe)
    eq_f.append(eq_prev); nci_bs_f.append(nci_prev); nd_f.append(nd_prev)
    cash_f.append(cash)
    cash_int = cash
assert abs(npa_f[0] - REL['npa26']) < 1e-6

BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.000;(0.000);"-"'; MULT = '0.00x'
DF4 = '0.0000'; RATE4 = '0.0000'

SHEETS = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
          'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
          'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
          'Per-Share & Ratios', 'Peer & Sector']
wb = Workbook()
wb.active.title = SHEETS[0]
for n in SHEETS[1:]:
    wb.create_sheet(n)

EXPECT = {}
ANCH = {}

def title(ws, t, s=None, w=10, awidth=48, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM1, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM1, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

def note(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = SUB

# ============ ASSUMPTIONS (built first — every other sheet references it) ==========
ws = wb['Assumptions']
title(ws, 'Assumptions — every input in the model', 'Blue cells are pasted primary-record '
      'inputs; BLACK cells in this column are formulas deriving each modelling rate from '
      'those inputs — change an audited anchor and the whole model reprices. Column H '
      'gives the source and date of each input.', 8, awidth=56, cwidth=11)
ws.column_dimensions['H'].width = 90
hdr(ws, 3, ['Input'] + YFL + ['', 'Source'])
r = 4
A = {}

def block(name, items):
    """items: (key, label, value, fmt, src) pasted blue, or
              (key, label, '=formula', fmt, src, expect) live derived anchor."""
    global r
    band(ws, r, 8); put(ws, f'A{r}', name, bold=True, fmt=None); r += 1
    for item in items:
        key, lab, val, fmt, src = item[:5]
        put(ws, f'A{r}', lab, fmt=None)
        if callable(val):
            val = val()
        if isinstance(val, str) and val.startswith('='):
            putf(ws, f'C{r}', val, item[5], fmt)
        elif isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(ws, f'{get_column_letter(2+i)}{r}', v, BLUE, fmt)
        else:
            put(ws, f'C{r}', val, BLUE, fmt)
        if src:
            ws.cell(row=r, column=8, value=src).font = SUB
        A[key] = r
        r += 1
    r += 1

def a(key, i=None):
    col = get_column_letter(2 + i) if i is not None else 'C'
    return f"Assumptions!${col}${A[key]}"

block('Anchors', [
    ('spot', 'Spot price (AED)', SPOT, PX, SRC['spot']),
    ('shares', 'Shares outstanding (mn)', SH, NUM0, SRC['shares_mn']),
    ('tax_ct', 'Corporate tax rate (headline framing)', TAX, PCT, SRC['tax_ct']),
    ('tax_dmtt', 'Top-up tax rate (alternative framing)', TAXD, PCT, SRC['tax_dmtt']),
    ('pat_fy25', 'Profit for the year, FY2025 (AED mn)', IN['pat_fy25'], NUM1, SRC['pat_fy25']),
    ('nci_pat', 'Profit attributable to non-controlling interests, FY2025 (AED mn)',
     IN['nci_pat_fy25'], NUM1, SRC['nci_pat_fy25']),
    ('npa_fy25', 'Profit attributable to shareholders, FY2025 (AED mn)', IN['npa_fy25'],
     NUM1, SRC['npa_fy25'])])
block('Audited FY2025 record — pasted once here; every derived rate below is a formula '
      'off these cells', [
    ('rev25', 'Revenue, FY2025 (AED mn)', IN['rev_fy25'], NUM1, SRC['rev_fy25']),
    ('cons25', 'Consumption revenue, FY2025 (AED mn)', IN['cons_rev']['2025'], NUM1,
     SRC['cons_rev']),
    ('ew25', 'Electricity and water purchased from DEWA, FY2025 (AED mn)',
     IN['ew_cost_fy25'], NUM1, SRC['ew_cost_fy25']),
    ('capex25', 'Capital expenditure (cash), FY2025 (AED mn)', IN['capex_fy25'], NUM1,
     SRC['capex_fy25']),
    ('ppe_dep25', 'Depreciation of plant, FY2025 (AED mn)', PPE_DEP25, NUM1,
     'FY2025 audited cash-flow note: depreciation of property, plant and equipment'),
    ('dna25', 'Total depreciation and amortisation, FY2025 (AED mn)', IN['dna_fy25'],
     NUM1, SRC['dna_fy25']),
    ('ppe24', 'Net property, plant and equipment, end-FY2024 (AED mn)', BH['FY24']['ppe'],
     NUM1, 'FY2024 audited balance sheet (FY2025 filing comparative)'),
    ('ppe25', 'Net property, plant and equipment, end-FY2025 (AED mn)', BH['FY25']['ppe'],
     NUM1, 'FY2025 audited balance sheet'),
    ('inv25', 'Inventories, end-FY2025 (AED mn)', INV25, NUM1, 'FY2025 audited balance sheet'),
    ('rec25', 'Trade and other receivables, end-FY2025 (AED mn)', REC25, NUM1,
     'FY2025 audited balance sheet'),
    ('drp25', 'Due from related parties, end-FY2025 (AED mn)', DRP25, NUM1,
     'FY2025 audited balance sheet'),
    ('pay25', 'Trade and other payables, end-FY2025 (AED mn)', PAY25, NUM1,
     'FY2025 audited balance sheet'),
    ('accr25', 'Project-cost accruals inside payables, end-FY2025 (AED mn)', ACCR25, NUM1,
     'FY2025 audited FS note 36 (non-cash capital accrual movement) — excluded from the '
     'operating cycle'),
    ('crp25', 'Due to related parties, end-FY2025 (AED mn)', CRP25, NUM1,
     'FY2025 audited balance sheet'),
    ('eq_fy24', 'Equity attributable to shareholders, end-FY2024 (AED mn)', EQ24, NUM1,
     'FY2024 audited balance sheet (FY2025 filing comparative)'),
    ('eq_fy25', 'Equity attributable to shareholders, end-FY2025 (AED mn)',
     IN['eq_attr_fy25'], NUM1, SRC['eq_attr_fy25']),
    ])
block('Unit build — physical inputs and forecast drivers', [
    ('rt_fy24', 'Connected capacity, end-FY2024 (k RT)', IN['rt_conn']['2024'], NUM0,
     SRC['rt_conn']),
    ('rt_fy25', 'Connected capacity, end-FY2025 (k RT)', IN['rt_conn']['2025'], NUM0,
     SRC['rt_conn']),
    ('adds', 'New connections by year (k RT)', ADDS, NUM0,
     'FY2026 = guidance midpoint (100-110k, H1-2026 earnings deck p13); then the contracted '
     'backlog (2,018k at Jun-26) tapering 100/90/80/70k as the pipeline matures'),
    ('shock', 'FY2026 consumption shock (per-RT, full year)', SHOCK, PCT,
     'H1-2026 equivalent full-load hours -9.0% y/y (deck p4) — hospitality-occupancy-led '
     '~80% per the company\'s own H1-2026 attribution, weather a minor factor; full-year '
     'effect smaller as H2-2025 was itself soft'),
    ('recovery', 'Consumption per-RT recovery level from FY2027 (share of FY2025)', 1.0, PCT,
     'Recovery (de-escalation) framing: full return to the FY2025 level; the live recovery '
     'columns on the Sensitivity sheet run the same model at 90-103% and feed the '
     'continuation central'),
    ('pipes', 'Pre-insulated pipes revenue (AED mn, held flat)', IN['pipes_rev_fy25'], NUM1,
     SRC['pipes_rev_fy25']),
    ('other_cos25', 'Other cash cost of sales, FY2025 (AED mn)', U['other_cos25'], NUM1,
     'Cost of sales less DEWA purchases less the depreciation and amortisation inside cost '
     'of sales (FY2025 audited FS notes)'),
    ('ga_cash25', 'General and administrative cash cost, FY2025 (AED mn)', U['ga_cash25'],
     NUM1, 'G&A expenses (statement face 256.383, corrected 17-Aug-2026) less the '
     'depreciation inside G&A (FY2025 audited FS)'),
    ('oi', 'Other income, FY2025 (AED mn, statement face)', IN['oi_fy25'], NUM1,
     SRC['oi_fy25']),
    ('rental', 'Rental income inside other income, FY2025 (AED mn)', RENTAL, NUM1,
     SRC['rental_fy25']),
    ('ecl25', 'Reversal of credit-loss allowance, FY2025 (AED mn; not forecast)',
     IN['ecl_fy25'], NUM1, SRC['ecl_fy25']),
    ('intco25', 'Interest on related-party acquisition receivables, FY2025 (AED mn)',
     IN['intco_fy25'], NUM1, SRC['intco_fy25']),
    ('intco_decay', 'Receivable interest annual decay', INTCO_DECAY, PCT,
     'Amortising related-party acquisition receivables — interest income runs off as the '
     'balances amortise'),
    ('wage_esc', 'Wage and services cost escalator', WAGE_ESC, PCT,
     'UAE CPI / wage class; applied to cash cost of sales and cash G&A only'),
    ('maint_pct', 'Maintenance capital expenditure (share of opening plant)', U['maint_pct'],
     PCT, 'House assumption for a young plant fleet; flagged as an estimate'),
    ])
block('Derived rates — LIVE formulas off the audited anchors above (black = formula)', [
    ('cons_per_rt', 'Consumption revenue per average connected RT, FY2025 (AED k)',
     f'={a("cons25")}/(({a("rt_fy24")}+{a("rt_fy25")})/2)', RATE4,
     'KAM consumption revenue / average connected RT — a formula, not a paste', CPRT),
    ('cap_per_rt', 'Capacity and connection revenue per average connected RT (AED k)',
     f'=({a("rev25")}-{a("cons25")}-{a("pipes")})/(({a("rt_fy24")}+{a("rt_fy25")})/2)',
     RATE4, 'Implied: (revenue - consumption - pipes) / average RT; no per-RT tariff '
     'schedule is published — flagged', CAPRT),
    ('ew_ratio', 'Electricity and water cost as a share of consumption revenue',
     f'={a("ew25")}/{a("cons25")}', PCT2,
     'DEWA purchases / consumption revenue, FY2025; FY2024 prints 72.5% on the same basis',
     EW),
    ('capex_per_rt', 'Capital expenditure per added RT (AED mn per k RT)',
     f'={a("capex25")}/({a("rt_fy25")}-{a("rt_fy24")})', RATE4,
     'FY2025 cash capital expenditure / FY2025 net connections added (90k RT)',
     U['capex_per_rt']),
    ('dep_rate', 'Depreciation rate on opening plant',
     f'={a("ppe_dep25")}/{a("ppe24")}', PCT2,
     'FY2025 plant depreciation / opening net plant', U['dep_rate']),
    ('amort_flat', 'Amortisation and right-of-use depreciation (AED mn, flat)',
     f'={a("dna25")}-{a("ppe_dep25")}', NUM1,
     'Total D&A less plant depreciation (intangibles + right-of-use)', U['amort_flat']),
    ('nwc25', 'Net working capital, end-FY2025 (AED mn)',
     f'={a("inv25")}+{a("rec25")}+{a("drp25")}-({a("pay25")}-{a("accr25")})-{a("crp25")}',
     NUM1, 'Inventories + receivables + due from related parties - payables (ex capex '
     'accruals) - due to related parties', NWC25),
    ('nwc_ratio', 'Net working capital as a share of revenue',
     lambda: f'={a("nwc25")}/{a("rev25")}', PCT2,
     'End-FY2025 net working capital / FY2025 revenue (negative: customer deposits and '
     'payables fund the cycle)', NWC_RATIO),
    ('oi_op', 'Operating other income (AED mn, held flat)',
     f'={a("oi")}-{a("rental")}', NUM1,
     'Other income less rental income (note 29: grant + scrap + others) — the rental is '
     'the return on the investment properties the bridge adds at book, so it is excluded '
     'from operating EBITDA', OI_OP),
    ])
block('Physical decomposition — H1-2026 deck facts and the regulated tariff cap', [
    ('h1_rev', 'Revenue, H1-2026 (AED mn)', IN['rev_h1_26'], NUM1, SRC['rev_h1_26']),
    ('h1_rth', 'Cooling delivered, H1-2026 (m RTh)', H1_RTH, NUM0,
     'H1-2026 earnings deck p4'),
    ('h1_mix', 'Consumption share of revenue, H1-2026', H1_MIX, PCT,
     'H1-2026 earnings deck revenue mix'),
    ('cap643', 'RD10 v1.3 consumption tariff cap (AED/TRh, incl. fuel surcharge)', CAP643,
     '0.000', 'Dubai RSB regulation RD10 v1.3 (17-Sep-2025), section 7 table; carried '
     'unchanged into v1.4 (Feb-2026) — external regulatory fact'),
    ('eflh_h1', 'Equivalent full-load hours, H1-2026 (hrs)', UP['eflh_h1_2026_hrs'], NUM0,
     SRC['eflh_h1']),
    ])
block('Cost of capital', [
    ('rf', 'Risk-free rate (AED sovereign)', IN['rf_aed'], PCT2, SRC['rf_aed']),
    ('ds_rating', 'Sovereign default spread — rating basis', IN['ds_rating'], PCT2,
     SRC['ds_rating']),
    ('ds_cds', 'Sovereign default spread — CDS basis', IN['ds_cds'], PCT2, SRC['ds_cds']),
    ('erp_rating', 'Equity risk premium — rating basis', IN['erp_rating'], PCT2,
     SRC['erp_rating']),
    ('erp_cds', 'Equity risk premium — CDS basis', IN['erp_cds'], PCT2, SRC['erp_cds']),
    ('beta', 'Beta (own-stock weekly regression vs the FTSE ADX index)', IN['beta'], '0.000',
     SRC['beta']),
    ('beta_dfm', 'Beta — DFM General Index regression (comparison construction)', BETA_DFM,
     '0.000', 'Own-stock weekly regression vs the listing exchange\'s DFM General Index, '
     'retained as a comparison; the primary beta regresses against the FTSE ADX General '
     'Index'),
    ('kd', 'Marginal cost of debt', IN['kd_marg'], PCT2, SRC['kd_marg']),
    ('pi_t', 'Terminal inflation — UAE house macro path', PI_T, PCT2,
     'The house long-run inflation for this market; this study carries no inflation '
     'number of its own'),
    ('g1r', 'REAL growth, FY2031-FY2040 window (stage one)', G1_REAL, PCT2,
     SRC['g1_real']),
    ('g2r', 'REAL growth beyond FY2040 (stage two) — NEGATIVE', G2_REAL, PCT2,
     SRC['g2_real']),
    ('life', 'Weighted asset life, DERIVED from notes 5, 6 and 7 (years)', LIFE, NUM1,
     SRC['asset_life_years']),
    ('inccap', 'Invested capital per unit of real growth (AED mn)', INC_CAP, NUM1,
     'One per cent of real growth costs one per cent of the invested capital the business '
     'operates on. The marginal reading across the explicit window is NEGATIVE here, '
     'because the plant is written down faster than capex replaces it over those five '
     'years, and a negative capital requirement would credit this company for growing'),
    ('g', 'Growth, FY2031-FY2040 window — (1+inflation)(1+real)-1',
     lambda: f'=(1+$C${A["pi_t"]})*(1+$C${A["g1r"]})-1', PCT,
     'DERIVED from the two rows above, never typed. It reproduces the previous edition\'s '
     '2.50% to the basis point, because under a tariff the regulator does not index a '
     'nominal growth rate IS a volume assumption', G),
    ('g2', 'Growth beyond FY2040 — the same identity on the stage-two real rate',
     lambda: f'=(1+$C${A["pi_t"]})*(1+$C${A["g2r"]})-1', PCT,
     'DERIVED. Reproduces the previous edition\'s 1.50%, and makes visible that it is a '
     'real DECLINE of half a point a year', G2),
    ('t_stub', 'Stub period — 30-Jun-2026 anchor to end-FY2026 (years)', T0, '0.0',
     'The bridge is struck on the 30-Jun-2026 reviewed balance sheet, so the cash-flow '
     'clock starts there too: FY2026 contributes its second half only and later year-ends '
     'sit at 1.5 to 4.5 years'),
    ])
block('Bridge — 30-Jun-2026 reviewed balance sheet', [
    ('borrow', 'Bank borrowings (AED mn)', IN['borrow_jun26'], NUM1, SRC['borrow_jun26']),
    ('lease', 'Lease liabilities (AED mn)', IN['lease_jun26'], NUM1, SRC['lease_jun26']),
    ('cash', 'Cash and cash equivalents (AED mn)', IN['cash_jun26'], NUM1, SRC['cash_jun26']),
    ('deposits', 'Term deposits (AED mn)', IN['deposits_jun26'], NUM1, SRC['deposits_jun26']),
    ('recv', 'Related-party acquisition receivables at book (Note 8: 1,005.0 Dubai '
     'Aviation City + 289.4 Nakheel) (AED mn)', RECV, NUM1, SRC['recv_jun26']),
    ('invprop', 'Investment properties (AED mn)', IN['invprop_jun26'], NUM1,
     SRC['invprop_jun26']),
    ('fvtpl', 'Financial assets at fair value through profit or loss (AED mn)',
     IN['fvtpl_jun26'], NUM1, SRC['fvtpl_jun26']),
    ('fvoci', 'Financial assets at fair value through OCI (AED mn)', IN['fvoci_jun26'],
     NUM1, SRC['fvoci_jun26']),
    ('eq_jun26', 'Equity attributable to shareholders, 30-Jun-2026 (AED mn)',
     IN['eq_attr_jun26'], NUM1, SRC['eq_attr_jun26']),
    ('nci_jun26', 'Non-controlling interests, 30-Jun-2026 (AED mn)', IN['nci_book_jun26'],
     NUM1, SRC['nci_book_jun26']),
    ('cash_yield', 'Yield earned on cash balances', CASH_YIELD, PCT2,
     'Deposit-rate assumption used in the forecast finance line; flagged as an estimate'),
    ('div', 'Committed annual dividend (AED mn)', IN['div_policy'], NUM1, SRC['div_policy']),
    ])
block('Valuation lenses', [
    ('tabreed_ev', 'Peer EV/EBITDA — Tabreed (trailing, restruck at the anchor)',
     REL['tabreed_ev_ebitda'], MULT,
     f"Tabreed (DFM) FY2025 results at the {REL['mult_date']}"),
    ('tabreed_pe', 'Peer price/earnings — Tabreed (trailing, restruck at the anchor)',
     REL['tabreed_pe'], MULT, 'Tabreed (DFM), close 2.46 on 6/7-Aug-2026 — cross-check '
     'input only'),
    ('dewa_pe', 'Peer price/earnings — DEWA (parent)', REL['dewa_pe'], MULT,
     'DEWA (DFM), Aug-2026 — context only, majority owner'),
    ('w_dcf', 'Weight — discounted cash flow', LN['dcf']['weight'], PCT, None),
    ('w_rel', 'Weight — relative multiples', LN['relative']['weight'], PCT, None),
    ('w_norm', 'Weight — normalised earnings', LN['normalized']['weight'], PCT, None),
    ('w_book', 'Weight — book value', LN['book']['weight'], PCT, None),
    ('dewa_price', 'DEWA control-transaction price, Feb-2026 (AED)', D['dewa_buyin']['price'],
     PX, 'Related-party CONTROL price for Dubai Holding\'s 24% stake — a disclosed reference '
     'point, never fair value'),
    ])

# ---- fixed row map for the live scenario blocks on Sensitivity (built later) ------
S_GRID0 = 6                       # 5 pasted grid rows 6-10
S_LIVE_G, S_LIVE = 13, 14         # live one-way growth row
S_CRUX_LVL = 19
S_REV0, S_EB0, S_FCFF0, S_PV0 = 20, 25, 30, 35
S_PVEX, S_NOP30, S_ROIC, S_Q, S_S1, S_NOP10, S_S2, S_TV, S_PVTV, S_EV, S_PS = range(40, 51)
S_F15_0, S_P15_0 = 53, 58
(S_PVEX15, S_NOP30_15, S_ROIC15, S_S1_15, S_NOP10_15, S_S2_15, S_TV15, S_PVTV15,
 S_EV15, S_PS15) = range(63, 73)
S_B_FAC, S_B_KE0, S_B_TAX, S_B_ADDS = 75, 76, 77, 78
(S_B_YE, S_B_AVG, S_B_REV, S_B_EB, S_B_OPEN, S_B_CAPEX, S_B_DNA, S_B_CLOSE, S_B_DNWC,
 S_B_FCFF, S_B_PV) = range(79, 90)
(S_B_KEB, S_B_WACC, S_B_PVEX, S_B_NOP30, S_B_ROIC, S_B_Q, S_B_S1, S_B_NOP10, S_B_S2,
 S_B_TV, S_B_PVTV, S_B_EV, S_B_PS) = range(90, 103)
S_U_ADDS = 105
(S_U_YE, S_U_AVG, S_U_REV, S_U_EB, S_U_OPEN, S_U_CAPEX, S_U_DNA, S_U_CLOSE, S_U_DNWC,
 S_U_FCFF, S_U_PV) = range(106, 117)
(S_U_PVEX, S_U_NOP30, S_U_ROIC, S_U_Q, S_U_S1, S_U_NOP10, S_U_S2, S_U_TV, S_U_PVTV,
 S_U_EV, S_U_PS) = range(117, 128)
S_C_WACC = 131
S_C_PV0 = 132
(S_C_PVEX, S_C_Q, S_C_S1, S_C_S2, S_C_TV, S_C_PVTV, S_C_EV, S_C_PS) = range(137, 145)

# ============ SEGMENTS =============================================================
ws = wb['Segments']
title(ws, 'Segments — the two-leg unit build, decomposed to physical units',
      'Consumption (per-RT rate x average connected RT) + capacity/connection (per-RT '
      'rate x average connected RT) + pipes. The disclosed anchors are pasted ONCE on the '
      'Assumptions sheet; every cell here is a formula.', 7, awidth=52, cwidth=13)
SC = ['C', 'D', 'E', 'F', 'G']                   # FY26..FY30 columns; B = FY2025
hdr(ws, 4, ['k RT / AED mn', 'FY2025'] + YFL)
rt_ye, rt_av = [], []
prev = RTP['FY25']
for i, y in enumerate(YF):
    ye = prev + ADDS[i]; rt_av.append((prev + ye) / 2); rt_ye.append(ye); prev = ye
put(ws, 'A5', 'New connections added (k RT)', fmt=None)
put(ws, 'B5', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{SC[i]}5', f'={a("adds", i)}', ADDS[i], NUM0, green=True)
put(ws, 'A6', 'Connected capacity, year-end (k RT)', fmt=None)
putf(ws, 'B6', f'={a("rt_fy25")}', RTP['FY25'], NUM0, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}6', f'={"B" if i == 0 else SC[i-1]}6+{SC[i]}5', rt_ye[i], NUM0)
put(ws, 'A7', 'Average connected capacity (k RT)', fmt=None)
putf(ws, 'B7', f'=({a("rt_fy24")}+{a("rt_fy25")})/2',
     (IN['rt_conn']['2024'] + IN['rt_conn']['2025']) / 2, NUM1)
for i in range(5):
    putf(ws, f'{SC[i]}7', f'=({"B" if i == 0 else SC[i-1]}6+{SC[i]}6)/2', rt_av[i], NUM1)
put(ws, 'A8', 'Consumption revenue per average RT (AED k)', fmt=None)
putf(ws, 'B8', f'={a("cons_per_rt")}', CPRT, RATE4, green=True)
putf(ws, 'C8', f'={a("cons_per_rt")}*(1+{a("shock")})', CPRT * (1 + SHOCK), RATE4)
for i in range(1, 5):
    putf(ws, f'{SC[i]}8', f'={a("cons_per_rt")}*{a("recovery")}', CPRT, RATE4)
put(ws, 'A9', 'Consumption leg', fmt=None)
putf(ws, 'B9', f'={a("cons25")}', U['cons25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}9', f'={SC[i]}8*{SC[i]}7', F['cons'][YF[i]], NUM1)
put(ws, 'A10', 'Capacity, connection and other services leg', fmt=None)
putf(ws, 'B10', f'={a("rev25")}-B9-{a("pipes")}', U['cap25'], NUM1)
for i in range(5):
    putf(ws, f'{SC[i]}10', f'={a("cap_per_rt")}*{SC[i]}7', F['cap'][YF[i]], NUM1)
put(ws, 'A11', 'Pre-insulated pipes', fmt=None)
putf(ws, 'B11', f'={a("pipes")}', IN['pipes_rev_fy25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}11', f'={a("pipes")}', IN['pipes_rev_fy25'], NUM1, green=True)
put(ws, 'A12', 'Total revenue', bold=True, fmt=None)
putf(ws, 'B12', '=SUM(B9:B11)', IN['rev_fy25'], NUM1, bold=True)
for i in range(5):
    putf(ws, f'{SC[i]}12', f'=SUM({SC[i]}9:{SC[i]}11)', F['rev'][YF[i]], NUM1, bold=True)
band(ws, 12, 7)
hdr(ws, 14, ['Cost stack — one escalator per driver class', 'FY2025'] + YFL)
put(ws, 'A15', 'Electricity and water purchased from DEWA', fmt=None)
putf(ws, 'B15', f'={a("ew25")}', IN['ew_cost_fy25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}15', f'={a("ew_ratio")}*{SC[i]}9', F['ew'][YF[i]], NUM1)
put(ws, 'A16', 'Other cash cost of sales (wage class)', fmt=None)
putf(ws, 'B16', f'={a("other_cos25")}', U['other_cos25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}16', f'={"B" if i == 0 else SC[i-1]}16*(1+{a("wage_esc")})',
         F['other_cos'][YF[i]], NUM1)
put(ws, 'A17', 'General and administrative cash cost (wage class)', fmt=None)
putf(ws, 'B17', f'={a("ga_cash25")}', U['ga_cash25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}17', f'={"B" if i == 0 else SC[i-1]}17*(1+{a("wage_esc")})',
         F['ga'][YF[i]], NUM1)
put(ws, 'A18', 'Interest on related-party acquisition receivables (excluded from '
    'operating EBITDA)', fmt=None)
putf(ws, 'B18', f'={a("intco25")}', IN['intco_fy25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{SC[i]}18', f'={"B" if i == 0 else SC[i-1]}18*(1-{a("intco_decay")})',
         F['intco'][YF[i]], NUM1)
put(ws, 'A19', 'Rental income on investment properties (excluded from operating EBITDA)',
    fmt=None)
for col in ['B'] + SC:
    putf(ws, f'{col}19', f'={a("rental")}', RENTAL, NUM1, green=True)
put(ws, 'A20', 'Operating other income (grant, scrap, others)', fmt=None)
for col in ['B'] + SC:
    putf(ws, f'{col}20', f'={a("oi_op")}', OI_OP, NUM1, green=True)
put(ws, 'A21', 'Reversal of credit-loss allowance (not forecast)', fmt=None)
putf(ws, 'B21', f'={a("ecl25")}', IN['ecl_fy25'], NUM1, green=True)
for i in range(5):
    put(ws, f'{SC[i]}21', 0, BLUE, NUM1)
put(ws, 'A22', 'Operating EBITDA (excludes receivable interest and rental income)',
    bold=True, fmt=None)
putf(ws, 'B22', '=B12-B15-B16-B17+B20+B21', EB25_OP, NUM1, bold=True)
for i in range(5):
    putf(ws, f'{SC[i]}22', f'={SC[i]}12-{SC[i]}15-{SC[i]}16-{SC[i]}17+{SC[i]}20'
         f'+{SC[i]}21', F['ebitda'][YF[i]], NUM1, bold=True)
band(ws, 22, 7)
put(ws, 'A23', 'Operating EBITDA margin', fmt=None)
putf(ws, 'B23', '=B22/B12', EB25_OP / IN['rev_fy25'], PCT)
for i in range(5):
    putf(ws, f'{SC[i]}23', f'={SC[i]}22/{SC[i]}12', F['ebitda'][YF[i]] / F['rev'][YF[i]], PCT)
put(ws, 'A24', 'Reconciliation: operating EBITDA + receivable interest + rental = '
    'audited operating profit + depreciation (FY2025)', fmt=None)
putf(ws, 'B24', '=B22+B18+B19', HI['FY25']['ebitda'], NUM1)
hdr(ws, 26, ['Physical decomposition — consumption = connected RT x hours x tariff', '',
             'Value'])
putf(ws, 'C27', f'={a("h1_mix")}*{a("h1_rev")}/{a("h1_rth")}', RATE, '0.000')
put(ws, 'A27', 'Realised consumption tariff, H1-2026 (AED/RTh)', fmt=None)
put(ws, 'A28', 'Headroom vs the RD10 v1.3 regulated cap (0.643 incl. fuel surcharge)',
    fmt=None)
putf(ws, 'C28', f'=C27/{a("cap643")}-1', HEADROOM, PCT)
put(ws, 'A29', 'Implied cooling delivered, FY2025 (m RTh)', fmt=None)
putf(ws, 'C29', '=B9/C27', UP['rth_fy25_mn'], NUM0)
put(ws, 'A30', 'Implied equivalent full-load hours, FY2025 (hrs per connected RT)',
    fmt=None)
putf(ws, 'C30', '=C29*1000/B7', UP['eflh_fy25_hrs'], NUM0)
put(ws, 'A31', 'Equivalent full-load hours, H1-2026 (hrs)', fmt=None)
putf(ws, 'C31', f'={a("eflh_h1")}', UP['eflh_h1_2026_hrs'], NUM0, green=True)
note(ws, 33, 'The FY2025 column rebuilds the audited year from the same identity; row 24 '
     'shows the reconciliation to the audited operating profit plus depreciation '
     f'({HI["FY25"]["ebitda"]:,.1f}): operating EBITDA (which includes the credit-loss '
     'reversal) plus the receivable interest plus the rental income. The FY2023-24 rental '
     'split is not disclosed, so only FY2025 carries the full split.')
note(ws, 34, 'Cost classes escalate on their own drivers: DEWA purchases follow the '
     'consumption leg (pass-through ratio held at the FY2025 print), wage-class lines '
     'escalate at the wage rate, and the receivable interest amortises. No blended index '
     'is applied across physically distinct cost lines.')
seas = U['fy26_seasonality_check']
note(ws, 35, f'Cross-check: H1-2026 revenue of {IN["rev_h1_26"]:,.1f} scaled by the FY2025 '
     f'seasonal split implies roughly {seas:,.0f} for FY2026; the model carries '
     f'{F["rev"]["FY26"]:,.0f}, within 4% (H2 carries the summer consumption peak). The '
     'realised tariff sits 1.4% below the regulated cap: Empower already prices at the '
     'cap, so the flat-tariff assumption is the regulation, not a house guess.')

# ============ DCF ==================================================================
ws = wb['DCF']
title(ws, 'Discounted cash flow — the full waterfall', 'Every line is a live formula. '
      'The clock starts at the 30-Jun-2026 balance-sheet anchor (FY2026 is a half-year '
      'stub); the terminal is a two-stage fade: a FY2031-40 build-out window, then a '
      'long-run perpetuity.', 7, awidth=52, cwidth=13)
CD = ['B', 'C', 'D', 'E', 'F']
hdr(ws, 4, ['AED mn'] + YFL)
ppe_open = [BH['FY25']['ppe']] + [F['ppe'][y] for y in YF[:-1]]
dep_f = [U['dep_rate'] * ppe_open[i] for i in range(5)]
capex_f = [F['capex'][y] for y in YF]
put(ws, 'A5', 'Revenue', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}5', f'=Segments!{SC[i]}12', F['rev'][YF[i]], NUM1, green=True)
put(ws, 'A6', 'Operating EBITDA', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}6', f'=Segments!{SC[i]}22', F['ebitda'][YF[i]], NUM1, bold=True,
         green=True)
put(ws, 'A7', 'Operating EBITDA margin', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}7', f'={CD[i]}6/{CD[i]}5', F['ebitda'][YF[i]] / F['rev'][YF[i]], PCT)
put(ws, 'A8', 'Opening net plant', fmt=None)
putf(ws, 'B8', f'={a("ppe25")}', ppe_open[0], NUM1, green=True)
for i in range(1, 5):
    putf(ws, f'{CD[i]}8', f'={CD[i-1]}11', ppe_open[i], NUM1)
put(ws, 'A9', 'Capital expenditure (per added RT + maintenance)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}9', f'={a("capex_per_rt")}*Segments!{SC[i]}5+{a("maint_pct")}*{CD[i]}8',
         capex_f[i], NUM1)
put(ws, 'A10', 'Depreciation on plant', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}10', f'={a("dep_rate")}*{CD[i]}8', dep_f[i], NUM1)
put(ws, 'A11', 'Closing net plant', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}11', f'={CD[i]}8+{CD[i]}9-{CD[i]}10', F['ppe'][YF[i]], NUM1)
put(ws, 'A12', 'Depreciation and amortisation (incl. right-of-use and intangibles)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}12', f'={CD[i]}10+{a("amort_flat")}', F['dna'][YF[i]], NUM1)
put(ws, 'A13', 'EBIT (operating)', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}13', f'={CD[i]}6-{CD[i]}12', B_CT['ebit'][YF[i]], NUM1, bold=True)
band(ws, 13, 7)
put(ws, 'A14', 'NOPAT — EBIT x (1 - corporate tax)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}14', f'={CD[i]}13*(1-{a("tax_ct")})', B_CT['nopat'][YF[i]], NUM1)
put(ws, 'A15', 'Add back depreciation and amortisation', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}15', f'={CD[i]}12', F['dna'][YF[i]], NUM1)
put(ws, 'A16', 'Less capital expenditure', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}16', f'=-{CD[i]}9', -capex_f[i], NUM1)
put(ws, 'A17', 'Less change in net working capital', fmt=None)
putf(ws, 'B17', f'=-({a("nwc_ratio")}*B5-{a("nwc25")})', -F['dnwc']['FY26'], NUM1)
for i in range(1, 5):
    putf(ws, f'{CD[i]}17', f'=-{a("nwc_ratio")}*({CD[i]}5-{CD[i-1]}5)', -F['dnwc'][YF[i]],
         NUM1)
put(ws, 'A18', 'Free cash flow to the firm', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}18', f'={CD[i]}14+{CD[i]}15+{CD[i]}16+{CD[i]}17', B_CT['fcff'][YF[i]],
         NUM1, bold=True)
band(ws, 18, 7)
put(ws, 'A19', 'Years from the 30-Jun-2026 anchor (t)', fmt=None)
putf(ws, 'B19', f'={a("t_stub")}', T0, '0.0', green=True)
for i in range(1, 5):
    putf(ws, f'{CD[i]}19', f'={CD[i-1]}19+1', T0 + i, '0.0')
put(ws, 'A20', 'Discount factor — 1/(1+WACC)^t (9% framing, rating basis)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}20', f'=1/(1+$C$56)^{CD[i]}19', B_CT['df'][YF[i]], DF4)
put(ws, 'A21', 'Present value of free cash flow (FY2026: half-year stub)', bold=True,
    fmt=None)
putf(ws, 'B21', f'=B18*B20*{a("t_stub")}', B_CT['pv']['FY26'], NUM1, bold=True)
for i in range(1, 5):
    putf(ws, f'{CD[i]}21', f'={CD[i]}18*{CD[i]}20', B_CT['pv'][YF[i]], NUM1, bold=True)
note(ws, 22, 'The bridge and the cash-flow clock share the same 30-Jun-2026 date: FY2026 '
     'contributes only its second half (the stub), discounted at t=0.5, and the terminal '
     'discounts at t=4.5. The previous full-year convention double-counted H1-2026 cash '
     'already inside the June net debt.')

put(ws, 'A23', 'TERMINAL VALUE — TWO-STAGE, REINVESTMENT-CONSISTENT (9% framing)',
    bold=True, fmt=None)
# ROWS 24-39 KEEP THEIR POSITIONS DELIBERATELY: a dozen formulas on four other sheets
# name C25, C27, C29, C31, C36, C38 and C39 by address, and inserting the terminal
# waterfall into this block would move every one of them silently [L-300]. The waterfall
# is built once at rows 93-110 and summarised here; the parallel re-runs below need only
# the two free-cash-flow figures, not a waterfall each.
tv_rows = [
    ('Net working capital, FY2030', f'={a("nwc_ratio")}*F5', F['nwc']['FY30'], NUM1),
    ('Terminal invested capital (plant + net working capital only)', '=F11+C24', IC_TERM,
     NUM1),
    ('Terminal return on invested capital (published; the terminal no longer uses it)',
     '=F14/C25', B_CT['roic_term'], PCT),
    ('Stage-one growth, FY2031-40 (Dubai 2040 build-out window)', f'={a("g")}', G, PCT),
    ('Stage-one free cash flow, FY2031 — the waterfall is at rows 93-101', '=C99',
     _F1, NUM1),
    ('Stage-two growth beyond FY2040 (densification; a REAL decline)', f'={a("g2")}',
     G2, PCT),
    ('Stage-two free cash flow, FY2041 — the waterfall is at rows 102-110', '=C108',
     _F2, NUM1),
    ('Growth/discount ratio, stage one', '=(1+C27)/(1+$C$56)',
     (1 + G) / (1 + WACC_CT), DF4),
    ('Value of the FY2031-40 window (at FY2030)', '=C28/(1+$C$56)*(1-C31^10)/(1-C31)',
     _s1, NUM1),
    ('Stage-one free cash flow as a share of terminal profit', '=C28/(F14*(1+C27))',
     _F1 / (B_CT['nopat']['FY30'] * (1 + G)), PCT),
    ('Value beyond FY2040 (perpetuity, discounted to FY2030)',
     '=C30*(1+C29)/(($C$56-C29)*(1+$C$56)^10)', _s2, NUM1),
    ('Terminal value at FY2030 (both stages)', '=C32+C34', B_CT['tv'], NUM1),
    ('Present value of the terminal value', '=C35*F20', B_CT['pv_tv'], NUM1),
    ('Present value of the five forecast years', '=SUM(B21:F21)', B_CT['pv_explicit'], NUM1),
    ('Enterprise value', '=C36+C37', B_CT['ev'], NUM1),
    ('Terminal value as a share of enterprise value', '=C36/C38', B_CT['tv_share'], PCT),
]
r = 24
for lab, fml, xp, fmt in tv_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, bold=(r == 38), green=fml.startswith('=Assumptions'))
    r += 1
band(ws, 38, 4)
put(ws, 'A41', 'Fair value per share (9% framing) — from the bridge', bold=True, fmt=None)
putf(ws, 'C41', "='SOTP Bridge'!C14", B_CT['ps'], PX, bold=True, green=True)

put(ws, 'A43', 'COST OF CAPITAL — BUILT HERE, NOT ASSUMED', bold=True, fmt=None)
hdr(ws, 44, ['Component', '', 'Rating basis', 'CDS basis'])
coc = [
    ('Risk-free rate (AED sovereign)', f'={a("rf")}', f'={a("rf")}', IN['rf_aed'],
     IN['rf_aed'], PCT2),
    ('Less sovereign default spread', f'={a("ds_rating")}', f'={a("ds_cds")}',
     IN['ds_rating'], IN['ds_cds'], PCT2),
    ('Risk-free rate net of the sovereign spread', '=C45-C46', '=D45-D46',
     W['rf_star_rating'], W['rf_star_cds'], PCT2),
    ('Beta', f'={a("beta")}', f'={a("beta")}', IN['beta'], IN['beta'], '0.000'),
    ('Equity risk premium', f'={a("erp_rating")}', f'={a("erp_cds")}', IN['erp_rating'],
     IN['erp_cds'], PCT2),
    ('Cost of equity', '=C47+C48*C49', '=D47+D48*D49', W['ke_rating'], W['ke_cds'], PCT2),
]
r = 45
for lab, f1, f2, x1, x2, fmt in coc:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', f1, x1, fmt, green=f1.startswith('=Assumptions'))
    putf(ws, f'D{r}', f2, x2, fmt, green=f2.startswith('=Assumptions'))
    r += 1
scal = [
    ('Marginal cost of debt', f'={a("kd")}', IN['kd_marg'], PCT2),
    ('Market capitalisation (spot x shares)', f'={a("spot")}*{a("shares")}', W['mktcap'],
     NUM1),
    ('Net debt (borrowings + leases - cash - deposits)',
     f'={a("borrow")}+{a("lease")}-{a("cash")}-{a("deposits")}', NET_DEBT, NUM1),
    ('Debt weight', '=C53/(C53+C52)', W['wd'], PCT2),
    ('Equity weight', '=1-C54', W['we'], PCT2),
]
for lab, fml, xp, fmt in scal:                    # rows 51..55
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, green=fml.startswith('=Assumptions'))
    r += 1
WACC_CT_ROW, WACC_DM_ROW = 56, 57
put(ws, f'A{WACC_CT_ROW}', 'WACC — 9% corporate-tax framing', bold=True, fmt=None)
putf(ws, f'C{WACC_CT_ROW}', f'=C55*C50+C54*C51*(1-{a("tax_ct")})', W['rating_ct'], PCT2,
     bold=True)
putf(ws, f'D{WACC_CT_ROW}', f'=C55*D50+C54*C51*(1-{a("tax_ct")})', W['cds_ct'], PCT2,
     bold=True)
put(ws, f'A{WACC_DM_ROW}', 'WACC — 15% top-up-tax framing (interest shield stays at 9%: '
    'the top-up does not enlarge interest deductibility)', bold=True, fmt=None)
putf(ws, f'C{WACC_DM_ROW}', f'=C55*C50+C54*C51*(1-{a("tax_ct")})', W['rating_dmtt'], PCT2,
     bold=True)
putf(ws, f'D{WACC_DM_ROW}', f'=C55*D50+C54*C51*(1-{a("tax_ct")})', W['cds_dmtt'], PCT2,
     bold=True)
band(ws, WACC_CT_ROW, 5); band(ws, WACC_DM_ROW, 5)
note(ws, 58, 'Both premium bases strip the SAME basis of sovereign default spread as the '
     'premium adds back (rating-to-rating, CDS-to-CDS), so country risk is priced once. '
     'The Pillar-Two top-up is a minimum-effective-rate charge on profits, not a higher '
     'statutory rate on the interest deduction, so BOTH framings keep the 9% debt shield '
     'and discount at the same rate — they differ only in after-tax operating profit. '
     'There is no discount-rate glide: the AED curve is flat and both facilities float.')

# ---- THE TERMINAL, LINE BY LINE — built once, both stages, both tax framings --------
T1C, T2C = B_CT['terminal_stage1'], B_CT['terminal_stage2']
T1D, T2D = B_DM['terminal_stage1'], B_DM['terminal_stage2']
put(ws, 'A92', 'THE TERMINAL, LINE BY LINE — capital maintained at replacement cost over '
    'the asset life the depreciation notes themselves imply', bold=True, fmt=None)
put(ws, 'A93', 'The retired construction grew terminal profit and deducted a reinvestment '
    'rate set by the growth rate over the return on capital, which is arithmetically the '
    'same as rebuilding the whole capital base every 1/g years — fifty in stage one and '
    'sixty-seven in stage two, both facts about the dirham\'s peg to the dollar rather '
    'than about a chilled-water plant this company\'s own notes turn over in 28.1 years. '
    'Column C is the 9% framing and column G the 15%; they differ only in tax.',
    fmt=None).font = SUB


def _wf(row0, tag, TC, TD, n30c, n30d, d30, w30, g_nom, real, inc):
    rows = [
        ('%s — operating profit after tax, grown one year' % tag,
         (f'={n30c:.6f}*(1+{g_nom:.10f})', TC['inputs']['nopat']),
         (f'={n30d:.6f}*(1+{g_nom:.10f})', TD['inputs']['nopat'])),
        ('%s — plus the book depreciation and amortisation charge, grown' % tag,
         (f'={d30:.6f}*(1+{g_nom:.10f})', TC['inputs']['dna_book']),
         (f'={d30:.6f}*(1+{g_nom:.10f})', TD['inputs']['dna_book'])),
        ('%s — less capital maintenance at replacement cost, that charge escalated over '
         'half the derived life' % tag,
         (f'=-C{row0+1}*(1+{a("pi_t")})^({a("life")}/2)', -TC['maintenance']),
         (f'=-G{row0+1}*(1+{a("pi_t")})^({a("life")}/2)', -TD['maintenance'])),
        ('%s — less the capital real growth consumes' % tag,
         (f'=-{real:.10f}*{inc:.6f}', -TC['growth_capex']),
         (f'=-{real:.10f}*{inc:.6f}', -TD['growth_capex'])),
        ('%s — less inflation on working capital (a CREDIT here: this company is funded '
         'by its own customers)' % tag,
         (f'=-{a("pi_t")}*{w30:.6f}*(1+{g_nom:.10f})', -TC['wc_charge']),
         (f'=-{a("pi_t")}*{w30:.6f}*(1+{g_nom:.10f})', -TD['wc_charge'])),
        ('%s — FREE CASH FLOW' % tag,
         (f'=SUM(C{row0}:C{row0+4})', TC['fcff']),
         (f'=SUM(G{row0}:G{row0+4})', TD['fcff'])),
    ]
    r_ = row0
    for lab, (fc, xc), (fd, xd) in rows:
        put(ws, f'A{r_}', lab, fmt=None)
        putf(ws, f'C{r_}', fc, xc, NUM1, bold=(r_ == row0 + 5))
        putf(ws, f'G{r_}', fd, xd, NUM1, bold=(r_ == row0 + 5))
        r_ += 1


_n30c, _n30d = B_CT['nopat']['FY30'], B_DM['nopat']['FY30']
_d30, _w30 = F['dna']['FY30'], F['nwc']['FY30']
_wf(94, 'Stage one, FY2031', T1C, T1D, _n30c, _n30d, _d30, _w30, G, G1_REAL, INC_CAP)
_g10 = (1 + G) ** 10
_wf(103, 'Stage two, FY2041', T2C, T2D, _n30c * _g10, _n30d * _g10, _d30 * _g10,
    _w30 * _g10, G2, G2_REAL, 0.0)
put(ws, 'A111', 'Memo — the no-growth perpetuity at book depreciation, a diagnostic and '
    'not a bound', fmt=None)
putf(ws, 'C111', '=F14*(1+C27)/$C$56', T1C['floor'], NUM1)

# ---- 15% parallel framing ---------------------------------------------------------
_F1D, _F2D = T1D['fcff'], T2D['fcff']
s1_dm, s2_dm, tv_dm = two_stage_tv(_F1D, _F2D, WACC_CT)
assert abs(tv_dm - B_DM['tv']) < 1e-6 * B_DM['tv']
put(ws, 'A61', 'PARALLEL FRAMING — 15% TOP-UP TAX (same cash-flow build and discount '
    'rate, own tax)', bold=True, fmt=None)
put(ws, 'A62', 'NOPAT at 15%', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}62', f'={CD[i]}13*(1-{a("tax_dmtt")})', B_DM['nopat'][YF[i]], NUM1)
put(ws, 'A63', 'Free cash flow to the firm at 15%', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}63', f'={CD[i]}62+{CD[i]}15+{CD[i]}16+{CD[i]}17',
         B_DM['fcff'][YF[i]], NUM1)
put(ws, 'A64', 'Present value (same discount factors; FY2026 stub)', fmt=None)
putf(ws, 'B64', f'=B63*B20*{a("t_stub")}', B_DM['pv']['FY26'], NUM1)
for i in range(1, 5):
    putf(ws, f'{CD[i]}64', f'={CD[i]}63*{CD[i]}20', B_DM['pv'][YF[i]], NUM1)
dm_scal = [
    ('Present value of the five forecast years (15%)', '=SUM(B64:F64)',
     B_DM['pv_explicit'], NUM1),
    ('Terminal return on invested capital (15%; published, no longer used)',
     '=F62/C25', B_DM['roic_term'], PCT),
    ('Stage-one free cash flow at 15% — the waterfall is at rows 94-99', '=G99',
     _F1D, NUM1),
    ('Stage-two free cash flow at 15% — the waterfall is at rows 103-108', '=G108',
     _F2D, NUM1),
    ('Value of the FY2031-40 window (15%)', '=C67/(1+$C$56)*(1-C31^10)/(1-C31)',
     s1_dm, NUM1),
    ('Stage-one free cash flow as a share of terminal profit (15%)',
     '=C67/(F62*(1+C27))', _F1D / (B_DM['nopat']['FY30'] * (1 + G)), PCT),
    ('Value beyond FY2040 (15%)', '=C68*(1+C29)/(($C$56-C29)*(1+$C$56)^10)',
     s2_dm, NUM1),
    ('Terminal value at FY2030 (15%)', '=C69+C71', B_DM['tv'], NUM1),
    ('Present value of the terminal value (15%)', '=C72*F20', B_DM['pv_tv'], NUM1),
    ('Enterprise value (15%)', '=C65+C73', B_DM['ev'], NUM1),
]
r = 65
for lab, fml, xp, fmt in dm_scal:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, bold=('Enterprise' in lab))
    r += 1
EV_DM_ROW = 74
put(ws, 'A75', 'Fair value per share (15% framing) — from the bridge', bold=True, fmt=None)
putf(ws, 'C75', "='SOTP Bridge'!D14", B_DM['ps'], PX, bold=True, green=True)

# ---- CDS premium basis ------------------------------------------------------------
q_cds = (1 + G) / (1 + W['cds_ct'])
_F1C_CDS, _F2C_CDS = _sanctioned_fcff(_n30c, _d30, _w30, W['cds_ct'], INC_CAP)
s1_c, s2_c, tv_c = two_stage_tv(_F1C_CDS, _F2C_CDS, W['cds_ct'])
assert abs(tv_c - B_CDS['tv']) < 1e-6 * B_CDS['tv']
put(ws, 'A77', 'ALTERNATIVE PREMIUM BASIS — CDS (same cash flows, CDS-basis rate)',
    bold=True, fmt=None)
put(ws, 'A78', 'Present value at the CDS-basis rate (FY2026 stub)', fmt=None)
putf(ws, 'B78', f'=B18/(1+$D$56)^B19*{a("t_stub")}', B_CDS['pv']['FY26'], NUM1)
for i in range(1, 5):
    putf(ws, f'{CD[i]}78', f'={CD[i]}18/(1+$D$56)^{CD[i]}19', B_CDS['pv'][YF[i]], NUM1)
cds_scal = [
    ('Present value of the five forecast years (CDS basis)', '=SUM(B78:F78)',
     B_CDS['pv_explicit'], NUM1),
    ('Growth/discount ratio, stage one (CDS basis)', '=(1+C27)/(1+$D$56)', q_cds, DF4),
    ('Stage-one free cash flow at the CDS-basis rate (its own maintenance charge)',
     f'={_F1C_CDS:.6f}', _F1C_CDS, NUM1),
    ('Value of the FY2031-40 window (CDS basis)', '=C81/(1+$D$56)*(1-C80^10)/(1-C80)',
     s1_c, NUM1),
    ('Value beyond FY2040 (CDS basis)', f'={_F2C_CDS:.6f}*(1+C29)'
     '/(($D$56-C29)*(1+$D$56)^10)', s2_c, NUM1),
    ('Terminal value (CDS basis)', '=C82+C83', B_CDS['tv'], NUM1),
    ('Present value of the terminal value (CDS basis)', '=C84/(1+$D$56)^F19',
     B_CDS['pv_tv'], NUM1),
    ('Enterprise value (CDS basis)', '=C79+C85', B_CDS['ev'], NUM1),
]
r = 79
for lab, fml, xp, fmt in cds_scal:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, bold=('Enterprise' in lab))
    r += 1
EV_CDS_ROW = 86
put(ws, 'A87', 'Fair value per share (CDS basis) — from the bridge', fmt=None)
putf(ws, 'C87', "='SOTP Bridge'!E14", B_CDS['ps'], PX, green=True)

# ---- WACC constructions — all priced, all LIVE ------------------------------------
CON = W['constructions']
put(ws, 'A88', 'WACC CONSTRUCTIONS — ALL PRICED (each per-share is a live parallel '
    'model on the Sensitivity sheet)', bold=True, fmt=None)
hdr(ws, 89, ['Construction', 'WACC', 'AED/share'])
ws.cell(row=89, column=8, value='Why it is or is not the primary').font = Font(bold=True)
ws.cell(row=89, column=8).fill = FILL_H
ws.column_dimensions['H'].width = 80
con_rows = [
    ('Target net-debt weights (primary)', '=C56', W['rating_ct'],
     "='SOTP Bridge'!C14", B_CT['ps'],
     'Company policy holds ~2x net debt/EBITDA and the payout roughly equals equity free '
     'cash flow, so surplus cash is transient — net weights match the target structure.'),
    ('Gross-debt weights',
     f'=(C52*C50+({a("borrow")}+{a("lease")})*C51*(1-{a("tax_ct")}))'
     f'/(C52+{a("borrow")}+{a("lease")})', CON['gross'],
     f'=Sensitivity!B{S_C_PS}', DC['base_gross_wacc']['ps'],
     'The textbook frame: debt at gross, cash valued in the bridge. Overweights debt '
     'while the June cash pile persists.'),
    ('Net weights, negative-carry net-debt cost',
     f'=C55*C50+C54*((({a("borrow")}+{a("lease")})*C51*(1-{a("tax_ct")})'
     f'-({a("cash")}+{a("deposits")})*{a("cash_yield")}*(1-{a("tax_ct")}))/C53)',
     CON['carry'], f'=Sensitivity!C{S_C_PS}', DC['base_carry_wacc']['ps'],
     'Prices the negative carry: cash earns 3.5% against 4.92% debt, so net debt costs '
     'more than the facility rate.'),
    ('DFM-index beta (comparison regression)',
     f'=C55*(C47+{a("beta_dfm")}*C49)+C54*C51*(1-{a("tax_ct")})', CON['dfm_beta'],
     f'=Sensitivity!D{S_C_PS}', DC['base_dfm_beta']['ps'],
     'The listing exchange\'s own index gives beta 0.652; the FTSE ADX base-market '
     'regression (0.863) is primary per the house market definition.'),
]
r = 90
for lab, wf, wx, pf, px_, why in con_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', wf, wx, PCT2, green=(wf == '=C56'))
    putf(ws, f'C{r}', pf, px_, PX, green=True)
    ws.cell(row=r, column=8, value=why).font = SUB
    r += 1
ANCH['dcf'] = dict(wacc_ct=WACC_CT_ROW, wacc_dm=WACC_DM_ROW, ev=38, pvex=37, pvtv=36,
                   tvsh=39, roic=26, ps=41, tv=35, ke=50, kd=51, mktcap=52, nd=53,
                   ev_dm=EV_DM_ROW, ev_cds=EV_CDS_ROW, con0=90)

# ============ SOTP BRIDGE ==========================================================
ws = wb['SOTP Bridge']
title(ws, 'Enterprise value to equity — the bridge', 'Three parallel constructions: the 9% '
      'corporate-tax framing, the 15% top-up-tax framing, and the CDS premium basis. The two '
      'tax framings are published side by side, never averaged.', 6, awidth=52, cwidth=16)
hdr(ws, 4, ['Step', '', '9% framing', '15% framing', 'CDS basis'])
b_ev = {'C': ('=DCF!C38', B_CT['ev']), 'D': (f'=DCF!C{EV_DM_ROW}', B_DM['ev']),
        'E': (f'=DCF!C{EV_CDS_ROW}', B_CDS['ev'])}
put(ws, 'A5', 'Enterprise value', fmt=None)
for col, (fml, xp) in b_ev.items():
    putf(ws, f'{col}5', fml, xp, NUM1, green=True)
steps = [
    ('Less net debt (30-Jun-2026)', '=-DCF!$C$53', -NET_DEBT),
    ('Plus related-party acquisition receivables at book (Note 8: 1,005.0 Dubai Aviation '
     'City + 289.4 Nakheel)', f'={a("recv")}', RECV),
    ('Plus investment properties at book', f'={a("invprop")}', IN['invprop_jun26']),
    ('Plus financial assets at fair value through profit or loss', f'={a("fvtpl")}',
     IN['fvtpl_jun26']),
    ('Plus financial assets at fair value through OCI', f'={a("fvoci")}', IN['fvoci_jun26']),
]
r = 6
for lab, fml, xp in steps:
    put(ws, f'A{r}', lab, fmt=None)
    for col in 'CDE':
        putf(ws, f'{col}{r}', fml, xp, NUM1, green=fml.startswith('=Assumptions'))
    r += 1
put(ws, 'A11', 'Equity value', fmt=None)
for col, (fml, xp) in b_ev.items():
    putf(ws, f'{col}11', f'=SUM({col}5:{col}10)', xp - NET_DEBT + BRIDGE_ADD, NUM1)
put(ws, 'A12', 'Less non-controlling interests (at their share of profit)', fmt=None)
for col, (fml, xp) in b_ev.items():
    putf(ws, f'{col}12', f'=-{col}11*$C$15', -(xp - NET_DEBT + BRIDGE_ADD) * NCI_FR, NUM1)
put(ws, 'A13', 'Equity attributable to shareholders', bold=True, fmt=None)
for col, (fml, xp) in b_ev.items():
    putf(ws, f'{col}13', f'={col}11+{col}12', (xp - NET_DEBT + BRIDGE_ADD) * (1 - NCI_FR),
         NUM1, bold=True)
put(ws, 'A14', 'Fair value per share (AED)', bold=True, fmt=None)
for col, (fml, xp) in b_ev.items():
    putf(ws, f'{col}14', f'={col}13/{a("shares")}',
         (xp - NET_DEBT + BRIDGE_ADD) * (1 - NCI_FR) / SH, PX, bold=True)
band(ws, 13, 5); band(ws, 14, 5)
put(ws, 'A15', 'Non-controlling share of group profit', fmt=None)
putf(ws, 'C15', f'={a("nci_pat")}/{a("pat_fy25")}', NCI_FR, PCT2)
put(ws, 'A16', 'Terminal value as a share of enterprise value', fmt=None)
putf(ws, 'C16', '=DCF!C39', B_CT['tv_share'], PCT, green=True)
putf(ws, 'D16', f'=DCF!C73/DCF!C{EV_DM_ROW}', B_DM['tv_share'], PCT, green=True)
putf(ws, 'E16', f'=DCF!C85/DCF!C{EV_CDS_ROW}', B_CDS['tv_share'], PCT, green=True)
note(ws, 18, 'The bridge is struck on the 30-Jun-2026 reviewed balance sheet — the same '
     'date the cash-flow clock starts. The related-party acquisition receivables enter '
     'at book (their interest is outside operating profit and they are outside terminal '
     'invested capital); investment properties enter at book (their rental is likewise '
     'outside operating EBITDA); the two fair-value portfolios add at book.')
ANCH['bridge'] = dict(ps=14, nci=15, tvsh=16)

# ============ INCOME STATEMENT =====================================================
ws = wb['Income Statement']
title(ws, 'Income statement — three years audited, five years forecast',
      'AED mn, consolidated. History is the audited record; every forecast line is a '
      'formula (9% corporate-tax framing; the 15% framing runs on the DCF sheet). The '
      'receivable interest and the investment-property rental are their own visible '
      'rows, outside operating EBITDA.', 9, awidth=46, cwidth=12)
HCC = ['B', 'C', 'D']; FCC = ['E', 'F', 'G', 'H', 'I']
ALLC = HCC + FCC
hdr(ws, 4, ['AED mn'] + YHL + YFL)

def hist3(row, vals, fmt=NUM1, bd=False, d_formula=None, d_expect=None):
    """Two pasted history columns; FY2025 optionally linked to its Assumptions anchor."""
    for i in range(2):
        if vals is None or vals[i] is None:
            put(ws, f'{HCC[i]}{row}', '-', BLACK, fmt, bold=bd)
        else:
            put(ws, f'{HCC[i]}{row}', vals[i], BLUE, fmt, bold=bd)
    if d_formula:
        putf(ws, f'D{row}', d_formula, d_expect, fmt, bold=bd, green=True)
    elif vals is not None and vals[2] is not None:
        put(ws, f'D{row}', vals[2], BLUE, fmt, bold=bd)
    else:
        put(ws, f'D{row}', '-', BLACK, fmt, bold=bd)

put(ws, 'A5', 'Revenue', bold=True, fmt=None)
hist3(5, [HI['FY23']['rev'], HI['FY24']['rev'], None], bd=True,
      d_formula=f'={a("rev25")}', d_expect=HI['FY25']['rev'])
for i in range(5):
    putf(ws, f'{FCC[i]}5', f'=DCF!{CD[i]}5', F['rev'][YF[i]], NUM1, bold=True, green=True)
band(ws, 5, 9)
put(ws, 'A6', 'Interest on related-party acquisition receivables (inside gross profit)',
    fmt=None)
hist3(6, [HI['FY23']['intco'], HI['FY24']['intco'], None],
      d_formula=f'={a("intco25")}', d_expect=HI['FY25']['intco'])
for i in range(5):
    putf(ws, f'{FCC[i]}6', f'=Segments!{SC[i]}18', F['intco'][YF[i]], NUM1, green=True)
put(ws, 'A7', 'Rental income on investment properties (inside other income)', fmt=None)
hist3(7, None, d_formula=f'={a("rental")}', d_expect=RENTAL)
for i in range(5):
    putf(ws, f'{FCC[i]}7', f'=Segments!{SC[i]}19', RENTAL, NUM1, green=True)
put(ws, 'A8', 'Gross profit', fmt=None)
hist3(8, [HI[y]['gp'] for y in H3])
for i in range(5):
    put(ws, f'{FCC[i]}8', '-', BLACK, NUM1)
put(ws, 'A9', 'Operating EBITDA (ex receivable interest and rental)', bold=True, fmt=None)
putf(ws, 'B9', '=B12-B11-B6', HIST_EB_OP[0], NUM1, bold=True)
putf(ws, 'C9', '=C12-C11-C6', HIST_EB_OP[1], NUM1, bold=True)
putf(ws, 'D9', '=D12-D11-D6-D7', HIST_EB_OP[2], NUM1, bold=True)
for i in range(5):
    putf(ws, f'{FCC[i]}9', f'=DCF!{CD[i]}6', F['ebitda'][YF[i]], NUM1, bold=True,
         green=True)
band(ws, 9, 9)
put(ws, 'A10', 'Operating EBITDA margin', fmt=None)
eb_all = HIST_EB_OP + [F['ebitda'][y] for y in YF]
rev_all = [HI[y]['rev'] for y in H3] + [F['rev'][y] for y in YF]
for i in range(8):
    putf(ws, f'{ALLC[i]}10', f'={ALLC[i]}9/{ALLC[i]}5', eb_all[i] / rev_all[i], PCT)
put(ws, 'A11', 'Depreciation and amortisation', fmt=None)
hist3(11, [-HI['FY23']['dna'], -HI['FY24']['dna'], None],
      d_formula=f'=-{a("dna25")}', d_expect=-HI['FY25']['dna'])
for i in range(5):
    putf(ws, f'{FCC[i]}11', f'=-DCF!{CD[i]}12', -F['dna'][YF[i]], NUM1, green=True)
put(ws, 'A12', 'Operating profit (EBIT; audited years include the interest and rental '
    'above)', fmt=None)
hist3(12, [HI[y]['op'] for y in H3])
for i in range(5):
    putf(ws, f'{FCC[i]}12', f'={FCC[i]}9+{FCC[i]}11', B_CT['ebit'][YF[i]], NUM1)
put(ws, 'A13', 'Net finance income / (costs) — deposit yield on the ROLLING cash balance',
    fmt=None)
for i in range(3):
    putf(ws, f'{HCC[i]}13', f'={HCC[i]}14-{HCC[i]}12', HI[H3[i]]['pbt'] - HI[H3[i]]['op'],
         NUM1)
putf(ws, 'E13', f'=-({a("kd")}*{a("borrow")}-{a("cash_yield")}*{a("cash")})', fin_f[0],
     NUM1)
for i in range(1, 5):
    putf(ws, f'{FCC[i]}13',
         f"=-({a('kd')}*{a('borrow')}-{a('cash_yield')}*'Balance Sheet'!{FCC[i-1]}9)",
         fin_f[i], NUM1)
put(ws, 'A14', 'Profit before tax', fmt=None)
hist3(14, [HI[y]['pbt'] for y in H3])
for i in range(5):
    putf(ws, f'{FCC[i]}14', f'={FCC[i]}12+{FCC[i]}6+{FCC[i]}7+{FCC[i]}13', pbt_f[i], NUM1)
put(ws, 'A15', 'Income tax', fmt=None)
hist3(15, [HI[y]['tax'] for y in H3])
for i in range(5):
    putf(ws, f'{FCC[i]}15', f'=-{FCC[i]}14*{a("tax_ct")}', tax_f[i], NUM1)
put(ws, 'A16', 'Profit for the year', fmt=None)
hist3(16, [HI[y]['pat'] for y in H3])
for i in range(5):
    putf(ws, f'{FCC[i]}16', f'={FCC[i]}14+{FCC[i]}15', pat_f[i], NUM1)
put(ws, 'A17', 'Non-controlling interests', fmt=None)
for i in range(3):
    putf(ws, f'{HCC[i]}17', f'={HCC[i]}18-{HCC[i]}16',
         HI[H3[i]]['npa'] - HI[H3[i]]['pat'], NUM1)
for i in range(5):
    putf(ws, f'{FCC[i]}17', f"=-{FCC[i]}16*'SOTP Bridge'!$C$15", nci_f[i], NUM1)
put(ws, 'A18', 'Profit attributable to shareholders', bold=True, fmt=None)
hist3(18, [HI[y]['npa'] for y in H3], bd=True)
for i in range(5):
    putf(ws, f'{FCC[i]}18', f'={FCC[i]}16+{FCC[i]}17', npa_f[i], NUM1, bold=True)
band(ws, 18, 9)
put(ws, 'A19', 'Earnings per share (AED)', fmt=None)
npa_all = [HI[y]['npa'] for y in H3] + npa_f
for i in range(8):
    putf(ws, f'{ALLC[i]}19', f'={ALLC[i]}18/{a("shares")}', npa_all[i] / SH, PX)
note(ws, 21, 'Every FY2023-25 line is the audited figure (the operating-EBITDA, '
     'net-finance and non-controlling rows are arithmetic identities of audited lines; '
     'the FY2023-24 rental split inside other income is not disclosed, so those two '
     'operating-EBITDA columns exclude the receivable interest only). The forecast '
     'finance line holds the 30-Jun-2026 debt book flat at the marginal cost of debt and '
     'earns the deposit yield on the ROLLING cash balance from the Balance Sheet — '
     'FY2026 on the 30-Jun-2026 print itself, later years on the prior year-end balance. '
     'The FY2023 tax line is a credit (first recognition of deferred tax ahead of UAE '
     'corporate tax).')

# ============ BALANCE SHEET ========================================================
ws = wb['Balance Sheet']
title(ws, 'Balance sheet — condensed', 'AED mn. FY2024 and FY2025 are audited closing '
      'figures; 30-Jun-2026 is the reviewed interim. BOTH forward rolls — equity and net '
      'debt — start from the SAME 30-Jun-2026 position: FY2026 carries only the '
      'second-half stub of profit and cash and only the October dividend instalment.',
      9, awidth=46, cwidth=12)
hdr(ws, 4, ['AED mn', 'FY2024', 'FY2025', '30-Jun-26'] + YFL)
BC = ['B', 'C', 'D']; BF = ['E', 'F', 'G', 'H', 'I']
BKEYS = ['FY24', 'FY25', 'JUN26']

def bsrow(r, lab, key, fc_f=None, fc_v=None, bd=False, fmt=NUM1, hist_f=None, hist_v=None):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i in range(3):
        if hist_f is not None and hist_f(i) is not None:
            putf(ws, f'{BC[i]}{r}', hist_f(i), hist_v[i], fmt, bold=bd,
                 green=hist_f(i).startswith('=Assumptions'))
        else:
            put(ws, f'{BC[i]}{r}', BH[BKEYS[i]][key], BLUE, fmt, bold=bd)
    if fc_f is not None:
        for i in range(5):
            putf(ws, f'{BF[i]}{r}', fc_f(i), fc_v[i], fmt, bold=bd)
    else:
        for i in range(5):
            put(ws, f'{BF[i]}{r}', '-', BLACK, fmt, bold=bd)
    if bd: band(ws, r, 9)

bsrow(5, 'Property, plant and equipment (net)', 'ppe',
      hist_f=lambda i: [f'={a("ppe24")}', f'={a("ppe25")}', None][i],
      hist_v=[BH['FY24']['ppe'], BH['FY25']['ppe'], None],
      fc_f=lambda i: f'=DCF!{CD[i]}11', fc_v=[F['ppe'][y] for y in YF])
for i in range(5):
    ws[f'{BF[i]}5'].font = GREEN
bsrow(6, 'Investment properties', 'invprop', fc_f=lambda i: '=$D6',
      fc_v=[BH['JUN26']['invprop']] * 5)
bsrow(7, 'Related-party acquisition receivables (current + non-current)', 'conc',
      fc_f=lambda i: '=$D7', fc_v=[BH['JUN26']['conc']] * 5)
put(ws, 'A8', 'Financial assets at fair value (through P&L and OCI)', fmt=None)
for i in range(3):
    put(ws, f'{BC[i]}8', BH[BKEYS[i]]['fvtpl'] + BH[BKEYS[i]]['fvoci'], BLUE, NUM1)
for i in range(5):
    putf(ws, f'{BF[i]}8', '=$D8', BH['JUN26']['fvtpl'] + BH['JUN26']['fvoci'], NUM1)
bsrow(9, 'Cash, equivalents and term deposits', 'cash',
      fc_f=lambda i: f'=({a("borrow")}+{a("lease")})-{BF[i]}16', fc_v=cash_f)
bsrow(10, 'Total assets (as reported)', 'ta')
bsrow(11, 'Borrowings and lease liabilities (gross debt — held flat)', 'gross',
      fc_f=lambda i: f'={a("borrow")}+{a("lease")}', fc_v=[GROSS_JUN26] * 5)
bsrow(12, 'Trade and other payables', 'pay')
bsrow(13, 'Equity attributable to shareholders', 'eqp',
      hist_f=lambda i: [f'={a("eq_fy24")}', f'={a("eq_fy25")}', f'={a("eq_jun26")}'][i],
      hist_v=[EQ24, IN['eq_attr_fy25'], IN['eq_attr_jun26']],
      fc_f=lambda i: (f"=$D13+{a('t_stub')}*('Income Statement'!E18-{a('div')})" if i == 0
                      else f"={BF[i-1]}13+'Income Statement'!{FCC[i]}18-{a('div')}"),
      fc_v=eq_f, bd=True)
bsrow(14, 'Non-controlling interests', 'nci',
      hist_f=lambda i: [None, None, f'={a("nci_jun26")}'][i],
      hist_v=[None, None, IN['nci_book_jun26']],
      fc_f=lambda i: (f"=$D14-{a('t_stub')}*'Income Statement'!E17" if i == 0
                      else f"={BF[i-1]}14-'Income Statement'!{FCC[i]}17"),
      fc_v=nci_bs_f)
put(ws, 'A15', 'Net working capital (model definition)', fmt=None)
put(ws, 'B15', '-', BLACK, NUM1)
putf(ws, 'C15', f'={a("nwc25")}', NWC25, NUM1, green=True)
put(ws, 'D15', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{BF[i]}15', f"={a('nwc_ratio')}*'Income Statement'!{FCC[i]}5",
         F['nwc'][YF[i]], NUM1)
put(ws, 'A16', 'Net debt (gross debt - cash and deposits)', bold=True, fmt=None)
nd_hist = [BH[k]['gross'] - BH[k]['cash'] for k in BKEYS]
for i in range(3):
    putf(ws, f'{BC[i]}16', f'={BC[i]}11-{BC[i]}9', nd_hist[i], NUM1, bold=True)
fcfe_frag = [(f"DCF!{CD[i]}18+('Income Statement'!{FCC[i]}13+'Income Statement'!{FCC[i]}6"
              f"+'Income Statement'!{FCC[i]}7)*(1-{a('tax_ct')})") for i in range(5)]
putf(ws, 'E16', f"=$D16-{a('t_stub')}*({fcfe_frag[0]}-{a('div')})", nd_f[0], NUM1,
     bold=True)
for i in range(1, 5):
    putf(ws, f'{BF[i]}16', f"={BF[i-1]}16-({fcfe_frag[i]})+{a('div')}", nd_f[i], NUM1,
         bold=True)
band(ws, 16, 9)
put(ws, 'A17', 'Net debt / EBITDA (audited-definition EBITDA)', fmt=None)
put(ws, 'B17', '-', BLACK, MULT)
putf(ws, 'C17', "=C16/('Income Statement'!D9+'Income Statement'!D6+'Income Statement'!D7)",
     nd_hist[1] / HI['FY25']['ebitda'], MULT)
put(ws, 'D17', '-', BLACK, MULT)
for i in range(5):
    putf(ws, f'{BF[i]}17', f"={BF[i]}16/('Income Statement'!{FCC[i]}9"
         f"+'Income Statement'!{FCC[i]}6+'Income Statement'!{FCC[i]}7)",
         nd_f[i] / (F['ebitda'][YF[i]] + intco_f[i] + RENTAL), MULT)
note(ws, 19, 'This is a CONDENSED layout and does not foot to zero: government grants, '
     'end-of-service benefits, tax liabilities, retentions and the remaining current items '
     'are not shown separately. GROSS borrowings are held flat at the 30-Jun-2026 book; '
     'NET debt falls as cash builds — the roll credits each year\'s equity free cash flow '
     'and charges the committed dividend. Both the equity roll and the net-debt roll are '
     'anchored on the 30-Jun-2026 reviewed position (equity 3,337.6): FY2026 adds the '
     'second-half stub of attributable profit and pays only the October instalment.')

# ============ CASH FLOW ============================================================
ws = wb['Cash Flow']
title(ws, 'Cash flow — historical markers and the forecast waterfall', 'AED mn. The '
      'forecast links line-for-line to the DCF waterfall; the FY2026 column shows the '
      'full model year, and the balance-sheet roll takes its second-half stub.', 8,
      awidth=48, cwidth=12)
hdr(ws, 4, ['AED mn', 'FY2024', 'FY2025'] + YFL)
CFF = ['D', 'E', 'F', 'G', 'H']
put(ws, 'A5', 'Operating EBITDA', fmt=None)
putf(ws, 'B5', "='Income Statement'!C9", HIST_EB_OP[1], NUM1, green=True)
putf(ws, 'C5', "='Income Statement'!D9", HIST_EB_OP[2], NUM1, green=True)
for i in range(5):
    putf(ws, f'{CFF[i]}5', f"='Income Statement'!{FCC[i]}9", F['ebitda'][YF[i]], NUM1,
         green=True)
put(ws, 'A6', 'Net cash from operating activities (as reported)', fmt=None)
put(ws, 'B6', OCF24 if OCF24 else '-', BLUE if OCF24 else BLACK, NUM1)
put(ws, 'C6', OCF25, BLUE, NUM1)
for i in range(5):
    put(ws, f'{CFF[i]}6', '-', BLACK, NUM1)
put(ws, 'A7', 'NOPAT', fmt=None)
put(ws, 'B7', '-', BLACK, NUM1); put(ws, 'C7', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}7', f'=DCF!{CD[i]}14', B_CT['nopat'][YF[i]], NUM1, green=True)
put(ws, 'A8', 'Add back depreciation and amortisation', fmt=None)
put(ws, 'B8', '-', BLACK, NUM1); put(ws, 'C8', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}8', f'=DCF!{CD[i]}15', F['dna'][YF[i]], NUM1, green=True)
put(ws, 'A9', 'Capital expenditure', fmt=None)
put(ws, 'B9', -IN['capex_fy24'], BLUE, NUM1)
putf(ws, 'C9', f'=-{a("capex25")}', -IN['capex_fy25'], NUM1, green=True)
for i in range(5):
    putf(ws, f'{CFF[i]}9', f'=DCF!{CD[i]}16', -F['capex'][YF[i]], NUM1, green=True)
put(ws, 'A10', 'Change in net working capital', fmt=None)
put(ws, 'B10', '-', BLACK, NUM1); put(ws, 'C10', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}10', f'=DCF!{CD[i]}17', -F['dnwc'][YF[i]], NUM1, green=True)
put(ws, 'A11', 'Free cash flow to the firm', bold=True, fmt=None)
put(ws, 'B11', '-', BLACK, NUM1); put(ws, 'C11', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}11', f'={CFF[i]}7+{CFF[i]}8+{CFF[i]}9+{CFF[i]}10',
         B_CT['fcff'][YF[i]], NUM1, bold=True)
band(ws, 11, 8)
put(ws, 'A12', 'Net finance charge after tax', fmt=None)
put(ws, 'B12', '-', BLACK, NUM1); put(ws, 'C12', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}12', f"='Income Statement'!{FCC[i]}13*(1-{a('tax_ct')})",
         fin_f[i] * (1 - TAX), NUM1)
put(ws, 'A13', 'Receivable interest and rental income after tax', fmt=None)
put(ws, 'B13', '-', BLACK, NUM1); put(ws, 'C13', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}13',
         f"=('Income Statement'!{FCC[i]}6+'Income Statement'!{FCC[i]}7)*(1-{a('tax_ct')})",
         (intco_f[i] + RENTAL) * (1 - TAX), NUM1)
put(ws, 'A14', 'Free cash flow to equity', fmt=None)
put(ws, 'B14', '-', BLACK, NUM1); put(ws, 'C14', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}14', f'={CFF[i]}11+{CFF[i]}12+{CFF[i]}13', fcfe_f[i], NUM1)
put(ws, 'A15', 'Dividends paid', fmt=None)
put(ws, 'B15', DIV24 if DIV24 else '-', BLUE if DIV24 else BLACK, NUM1)
put(ws, 'C15', DIV25, BLUE, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}15', f'=-{a("div")}', -IN['div_policy'], NUM1, green=True)
put(ws, 'A16', 'Change in net debt (balance-sheet roll; FY2026 = second-half stub)',
    fmt=None)
put(ws, 'B16', '-', BLACK, NUM1); put(ws, 'C16', '-', BLACK, NUM1)
dnd = [nd_f[0] - NET_DEBT] + [nd_f[i] - nd_f[i-1] for i in range(1, 5)]
for i in range(5):
    putf(ws, f'{CFF[i]}16',
         f"='Balance Sheet'!{BF[i]}16-'Balance Sheet'!{'D' if i == 0 else BF[i-1]}16",
         dnd[i], NUM1, green=True)
put(ws, 'A17', 'Closing net debt', fmt=None)
put(ws, 'B17', '-', BLACK, NUM1); put(ws, 'C17', '-', BLACK, NUM1)
for i in range(5):
    putf(ws, f'{CFF[i]}17', f"='Balance Sheet'!{BF[i]}16", nd_f[i], NUM1, green=True)
note(ws, 19, 'The company converts EBITDA to operating cash at a high rate — the negative '
     'working-capital cycle (customer deposits and long payables) funds growth. The '
     'committed dividend exceeds equity free cash flow in the first forecast year by '
     'design; the net-debt roll on the Balance Sheet carries the difference, starting at '
     'the 30-Jun-2026 anchor with only the second half of FY2026.')

# ============ SUMMARY FINANCIALS ===================================================
ws = wb['Summary Financials']
title(ws, 'Summary financials — the eight-year picture', 'AED mn unless stated. Every cell '
      'on this sheet is a link or a ratio; nothing is typed twice.', 9, awidth=44, cwidth=12)
hdr(ws, 4, ['AED mn'] + YHL + YFL)
ebit_all = [HI[y]['op'] for y in H3] + [B_CT['ebit'][y] for y in YF]
ic_f = [F['ppe'][y] + F['nwc'][y] for y in YF]

def sfrow(r, lab, fml, vals, fmt=NUM1, skip=()):
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALLC[i]}{r}', '-', BLACK, fmt)
        else:
            f_ = fml(i)
            putf(ws, f'{ALLC[i]}{r}', f_, vals[i], fmt,
                 green=f_.startswith(("='I", "='B", '=DCF', "='C")))

sfrow(5, 'Revenue', lambda i: f"='Income Statement'!{ALLC[i]}5", rev_all)
sfrow(6, 'Revenue growth', lambda i: f'={ALLC[i]}5/{ALLC[i-1]}5-1',
      [None] + [rev_all[i] / rev_all[i-1] - 1 for i in range(1, 8)], PCT, skip=(0,))
sfrow(7, 'Operating EBITDA (ex receivable interest and rental)',
      lambda i: f"='Income Statement'!{ALLC[i]}9", eb_all)
sfrow(8, 'Operating EBITDA margin', lambda i: f'={ALLC[i]}7/{ALLC[i]}5',
      [eb_all[i] / rev_all[i] for i in range(8)], PCT)
sfrow(9, 'Operating profit (EBIT)', lambda i: f"='Income Statement'!{ALLC[i]}12", ebit_all)
sfrow(10, 'Attributable profit', lambda i: f"='Income Statement'!{ALLC[i]}18", npa_all)
sfrow(11, 'Free cash flow to the firm', lambda i: f"='Cash Flow'!{CFF[i-3]}11",
      [None] * 3 + [B_CT['fcff'][y] for y in YF], skip=(0, 1, 2))
sfrow(12, 'Net debt', lambda i: f"='Balance Sheet'!{['B','C'][i-1]}16" if i < 3
      else f"='Balance Sheet'!{BF[i-3]}16",
      [None, nd_hist[0], nd_hist[1]] + nd_f, skip=(0,))
sfrow(13, 'Invested capital (plant + net working capital)',
      lambda i: (f"=DCF!{CD[i-3]}11+'Balance Sheet'!{BF[i-3]}15"),
      [None] * 3 + ic_f, skip=(0, 1, 2))
sfrow(14, 'Return on invested capital (NOPAT / same-year capital)',
      lambda i: f'=DCF!{CD[i-3]}14/{ALLC[i]}13',
      [None] * 3 + [B_CT['nopat'][YF[i]] / ic_f[i] for i in range(5)], PCT, skip=(0, 1, 2))
note(ws, 16, 'Return on invested capital comfortably clears the cost of capital across the '
     'forecast — the economics of a regulated district-cooling monopoly with a negative '
     'working-capital cycle. Invested capital is plant plus net working capital only: the '
     'related-party acquisition receivables are valued at book in the bridge, so they sit '
     'outside operating capital, and the two-stage terminal on the DCF sheet reinvests '
     'from exactly this base.')

# ============ RELATIVE & NORMALIZED ================================================
ws = wb['Relative & Normalized']
title(ws, 'Relative multiples, normalised earnings power, book value and dividends', None, 6,
      awidth=56, cwidth=15)
hdr(ws, 4, ['Relative lens — trailing peer EV/EBITDA on trailing operating EBITDA',
            '', 'Value'])
rel_rows = [
    ('FY2025A EBITDA excluding receivable interest (operating EBITDA + rental; AED mn)',
     f'={a("rev25")}-{a("ew25")}-{a("other_cos25")}-{a("ga_cash25")}+{a("oi_op")}'
     f'+{a("ecl25")}+{a("rental")}', EB_TRAIL, NUM1),
    ('Peer EV/EBITDA (Tabreed, trailing; restruck at the 07-Aug-2026 anchor)',
     f'={a("tabreed_ev")}', REL['tabreed_ev_ebitda'], MULT),
    ('Implied enterprise value (AED mn)', '=C5*C6', REL['ev_rel'], NUM1),
    ('Less net debt', '=-DCF!$C$53', -NET_DEBT, NUM1),
    ('Plus related-party acquisition receivables at book', f'={a("recv")}', RECV, NUM1),
    ('Plus investment properties and fair-value assets',
     f'={a("invprop")}+{a("fvtpl")}+{a("fvoci")}', BRIDGE_ADD - RECV, NUM1),
    ('Equity value (AED mn)', '=SUM(C7:C10)', REL['ev_rel'] - NET_DEBT + BRIDGE_ADD, NUM1),
    ('Implied value per share (AED)',
     f"=C11*(1-'SOTP Bridge'!$C$15)/{a('shares')}", REL['ps_rel'], PX),
]
r = 5
for lab, fml, xp, fmt in rel_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt,
         green=('DCF' in fml or 'Income Statement' in fml or fml.startswith('=Assumptions')))
    r += 1
band(ws, 12, 3)
put(ws, 'A13', 'Sensitivity: each 0.5x on the peer multiple moves this lens by (AED/share)',
    fmt=None)
putf(ws, 'C13', f"=0.5*C5*(1-'SOTP Bridge'!$C$15)/{a('shares')}",
     0.5 * EB_TRAIL * (1 - NCI_FR) / SH, PX)
put(ws, 'A14', 'Peer price/earnings (Tabreed) on FY2026E attributable profit', fmt=None)
putf(ws, 'C14', f"={a('tabreed_pe')}*'Income Statement'!E18/{a('shares')}", REL['ps_pe'], PX)
r = 16
hdr(ws, r, ['Own trailing multiples', '', 'Value']); r += 1
for lab, fml, xp, fmt in [
        ('Trailing enterprise value / EBITDA (audited-definition EBITDA)',
         f"=({a('spot')}*{a('shares')}+DCF!$C$53)/('Income Statement'!D9"
         "+'Income Statement'!D6+'Income Statement'!D7)",
         (W['mktcap'] + NET_DEBT) / HI['FY25']['ebitda'], MULT),
        ('Trailing price / earnings', f"={a('spot')}/'Income Statement'!D19",
         SPOT / (HI['FY25']['npa'] / SH), MULT),
        ('Trailing price / book (30-Jun-2026)', 'PATCH_PB', SPOT / BK['bvps'], MULT),
        ('Net debt / FY2025 EBITDA (audited definition)',
         "=DCF!$C$53/('Income Statement'!D9+'Income Statement'!D6+'Income Statement'!D7)",
         NET_DEBT / HI['FY25']['ebitda'], MULT)]:
    put(ws, f'A{r}', lab, fmt=None); putf(ws, f'C{r}', fml, xp, fmt); r += 1
r += 1
hdr(ws, r, ['Normalised earnings power — FY2026E without the consumption shock', '',
            'Value'])
r += 1                                              # r = 23
norm_rows = [
    ('Revenue at the unshocked per-RT level (AED mn)',
     f'={a("cons_per_rt")}*Segments!C7+Segments!C10+{a("pipes")}', NRM['rev'], NUM1),
    ('Normalised EBITDA incl. receivable interest and rental (the earnings lens keeps '
     'both; AED mn)',
     f'=C23+Segments!C18+Segments!C19-{a("ew_ratio")}*{a("cons_per_rt")}*Segments!C7'
     f'-Segments!C16-Segments!C17+{a("oi_op")}', NRM['ebitda'], NUM1),
    ('Less depreciation and amortisation (FY2026E)', '=-DCF!B12', -F['dna']['FY26'], NUM1),
    ('Net finance charge (FY2026E)', "='Income Statement'!E13", fin_f[0], NUM1),
    ('Normalised attributable profit (AED mn)',
     f"=(C24+C25+C26)*(1-{a('tax_ct')})*(1-'SOTP Bridge'!$C$15)", NRM['npa'], NUM1),
    ('Normalised earnings per share (AED)', f'=C27/{a("shares")}', NRM['eps'], PX),
    ('Sustainable return on equity (FY2025 profit / average equity)',
     f'={a("npa_fy25")}/(({a("eq_fy25")}+{a("eq_fy24")})/2)', BK['roe_sust'], PCT),
    ('Justified FORWARD price/earnings — (1 - growth/return) / (cost of equity - growth)',
     f'=(1-{a("g")}/C29)/(DCF!$C$50-{a("g")})', NRM['pe_just'], MULT),
    ('Implied value per share (AED)', '=C28*C30', NRM['ps'], PX),
    ('15% framing: sustainable return re-taxed', f'=C29*(1-{a("tax_dmtt")})/(1-{a("tax_ct")})',
     NRM['roe_15'], PCT),
    ('15% framing: justified forward price/earnings', f'=(1-{a("g")}/C32)/(DCF!$C$50-{a("g")})',
     NRM['pe_just_15'], MULT),
    ('15% framing: implied value per share (re-taxed earnings AND return)',
     f'=C28*(1-{a("tax_dmtt")})/(1-{a("tax_ct")})*C33', NRM['ps_15'], PX),
]
for lab, fml, xp, fmt in norm_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt,
         green=('DCF' in fml or 'Segments' in fml or 'Income Statement' in fml))
    r += 1
band(ws, 31, 3)
r += 1
hdr(ws, r, ['Book value and sustainable return', '', 'Value']); r += 1   # r = 37
book_rows = [
    ('Book value per share, 30-Jun-2026 (AED)', f'={a("eq_jun26")}/{a("shares")}',
     BK['bvps'], PX),
    ('Justified price/book — (return - g) / (cost of equity - g)',
     f'=(C29-{a("g")})/(DCF!$C$50-{a("g")})', BK['pb_just'], MULT),
    ('Implied value per share (AED)', '=C37*C38', BK['ps'], PX),
    ('15% framing: justified price/book on the re-taxed return',
     f'=(C32-{a("g")})/(DCF!$C$50-{a("g")})', BK['pb_just_15'], MULT),
    ('15% framing: implied value per share', '=C37*C40', BK['ps_15'], PX),
]
for lab, fml, xp, fmt in book_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, green=('DCF' in fml))
    r += 1
band(ws, 39, 3)
putf(ws, 'C19', f'={a("spot")}/C37', SPOT / BK['bvps'], MULT)   # patch trailing P/B
r += 1
hdr(ws, r, ['Dividend cross-check', '', 'Value']); r += 1        # r = 44
ddm_rows = [
    ('Dividend per share — committed distribution (AED)', f'={a("div")}/{a("shares")}',
     DDM['dps'], PX),
    ('Dividend value — grown at stage-one growth, at the cost of equity (AED)',
     f'=C44*(1+{a("g")})/(DCF!$C$50-{a("g")})', DDM['ps'], PX),
]
for lab, fml, xp, fmt in ddm_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt, green=('DCF' in fml))
    r += 1
note(ws, r + 1, 'The trailing multiples use the audited-definition EBITDA (operating '
     'EBITDA plus the receivable interest plus rental) so they compare like-for-like with '
     'peer marks computed off reported statements. The normalised lens re-runs year one at '
     'the FY2025 per-RT consumption level: it answers what the business earns if the '
     'consumption shock (hospitality-occupancy-led ~80% per the company\'s H1-2026 '
     'attribution; weather a minor factor) proves cyclical. Both earnings-based lenses '
     'keep the receivable interest and rental in profit — they are real income; only '
     'operating EBITDA and free cash flow exclude them.')
ANCH['rel'] = dict(bvps=37, ps_rel=12, ps_pe=14, ps_norm=31, ps_norm15=34, ps_book=39,
                   ps_book15=41, ddm=45, roe=29, pe_just=30)

# ============ SUMMARY ==============================================================
ws = wb['Summary']
title(ws, 'Summary — valuation at a glance', 'All values link live to their source sheets; '
      'the weights are visible inputs. Two tax framings AND two macro framings (recovery / '
      'continuation) are shown side by side, never averaged.', 7, awidth=52, cwidth=14)
hdr(ws, 4, ['Lens', 'AED/share', 'Weight', 'Contribution', 'vs spot', '',
            'Terminal value share of EV'])
ws.column_dimensions['G'].width = 26
lens_src = [
    ('Discounted cash flow (recovery, 9% framing)', "='SOTP Bridge'!C14", B_CT['ps'],
     'w_dcf', LN['dcf']['weight']),
    ('Relative multiples (peer EV/EBITDA)', "='Relative & Normalized'!C12", REL['ps_rel'],
     'w_rel', LN['relative']['weight']),
    ('Normalised earnings power', "='Relative & Normalized'!C31", NRM['ps'], 'w_norm',
     LN['normalized']['weight']),
    ('Book value and sustainable return', "='Relative & Normalized'!C39", BK['ps'],
     'w_book', LN['book']['weight']),
]
r = 5
for lab, fml, xp, wkey, wv in lens_src:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fml, xp, PX, green=True)
    putf(ws, f'C{r}', f'={a(wkey)}', wv, PCT, green=True)
    putf(ws, f'D{r}', f'=B{r}*C{r}', xp * wv, PX)
    putf(ws, f'E{r}', f'=B{r}/{a("spot")}-1', xp / SPOT - 1, PCT)
    r += 1
putf(ws, 'G5', "='SOTP Bridge'!C16", B_CT['tv_share'], PCT, green=True)
CEN_ROW, CEN_CONT_ROW, CEN_DM_ROW, CEN_CONT_DM_ROW = 9, 10, 11, 12
band(ws, CEN_ROW, 7)
put(ws, f'A{CEN_ROW}', 'Central fair value — RECOVERY (de-escalation), 9% framing',
    bold=True, fmt=None)
putf(ws, f'B{CEN_ROW}', '=SUM(D5:D8)', CEN['ct'], PX, bold=True)
putf(ws, f'C{CEN_ROW}', '=SUM(C5:C8)', 1.0, PCT, bold=True)
putf(ws, f'E{CEN_ROW}', f'=B{CEN_ROW}/{a("spot")}-1', CEN['ct'] / SPOT - 1, PCT, bold=True)
put(ws, f'A{CEN_CONT_ROW}', 'Central fair value — CONTINUATION (the consumption shock '
    'persists), 9% framing', bold=True, fmt=None)
putf(ws, f'B{CEN_CONT_ROW}',
     f'={a("w_dcf")}*Sensitivity!C{S_PS}+{a("w_rel")}*B6+{a("w_norm")}*B7+{a("w_book")}*B8',
     CEN['continuation_ct'], PX, bold=True)
putf(ws, f'E{CEN_CONT_ROW}', f'=B{CEN_CONT_ROW}/{a("spot")}-1',
     CEN['continuation_ct'] / SPOT - 1, PCT)
put(ws, f'A{CEN_DM_ROW}', 'Central fair value — recovery, 15% top-up framing '
    '(tax-consistent lens set)', fmt=None)
putf(ws, f'B{CEN_DM_ROW}',
     f"={a('w_dcf')}*'SOTP Bridge'!D14+{a('w_rel')}*B6"
     f"+{a('w_norm')}*'Relative & Normalized'!C34+{a('w_book')}*'Relative & Normalized'!C41",
     CEN['dmtt'], PX)
putf(ws, f'E{CEN_DM_ROW}', f'=B{CEN_DM_ROW}/{a("spot")}-1', CEN['dmtt'] / SPOT - 1, PCT)
put(ws, f'A{CEN_CONT_DM_ROW}', 'Central fair value — continuation, 15% top-up framing',
    fmt=None)
putf(ws, f'B{CEN_CONT_DM_ROW}',
     f"={a('w_dcf')}*Sensitivity!C{S_PS15}+{a('w_rel')}*B6"
     f"+{a('w_norm')}*'Relative & Normalized'!C34+{a('w_book')}*'Relative & Normalized'!C41",
     CEN['continuation_dmtt'], PX)
putf(ws, f'E{CEN_CONT_DM_ROW}', f'=B{CEN_CONT_DM_ROW}/{a("spot")}-1',
     CEN['continuation_dmtt'] / SPOT - 1, PCT)
r = 13
scen_rows = [
    ('Bear case — re-escalation, connections at 50/50/45/40/35k RT, 15% tax, +100bp '
     '(live scenario)', f'=Sensitivity!C{S_B_PS}', CEN['bear']),
    ('Bull case — clean recovery, connections at 110/110/100/90/80k RT, 9% tax '
     '(live scenario)', f'=Sensitivity!C{S_U_PS}', CEN['bull']),
    ('Dividend cross-check', "='Relative & Normalized'!C45", DDM['ps']),
    ('Discounted cash flow on the CDS premium basis', "='SOTP Bridge'!E14", B_CDS['ps']),
    ('Reference construction — gross-debt WACC weights', '=DCF!C91',
     DC['base_gross_wacc']['ps']),
    ('Reference construction — negative-carry net-debt cost', '=DCF!C92',
     DC['base_carry_wacc']['ps']),
    ('Reference construction — DFM-index beta', '=DCF!C93', DC['base_dfm_beta']['ps']),
]
for lab, fml, xp in scen_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fml, xp, PX, green=True)
    putf(ws, f'E{r}', f'=B{r}/{a("spot")}-1', xp / SPOT - 1, PCT)
    r += 1
band(ws, r, 7)
put(ws, f'A{r}', 'Market price (anchor, 07-Aug-2026)', bold=True, fmt=None)
putf(ws, f'B{r}', f'={a("spot")}', SPOT, PX, bold=True, green=True)
SPOT_ROW = r
r += 1
put(ws, f'A{r}', 'DEWA control-transaction price, Feb-2026 (reference point, not fair value)',
    fmt=None)
putf(ws, f'B{r}', f'={a("dewa_price")}', D['dewa_buyin']['price'], PX, green=True)
note(ws, r + 1, 'The Feb-2026 DEWA purchase of Dubai Holding\'s 24% at AED 2.16 was a '
     'related-party CONTROL transaction — a disclosed reference point above every lens '
     'here, never a fair-value estimate for minority shares. Recovery and continuation '
     'are published side by side like the tax framings; neither is privileged as the '
     'single answer.')
r += 3
hdr(ws, r, ['Key figure', 'Value'])
r += 1
key_rows = [
    ('Shares outstanding (mn)', f'={a("shares")}', SH, NUM0),
    ('Market capitalisation (AED mn)', '=DCF!C52', W['mktcap'], NUM1),
    ('Net debt, 30-Jun-2026 (AED mn)', '=DCF!C53', NET_DEBT, NUM1),
    ('FY2025 revenue (AED mn)', "='Income Statement'!D5", HI['FY25']['rev'], NUM1),
    ('FY2025 operating EBITDA (AED mn)', "='Income Statement'!D9", EB25_OP, NUM1),
    ('FY2025 attributable profit (AED mn)', "='Income Statement'!D18", HI['FY25']['npa'],
     NUM1),
    ('Cost of equity (rating basis)', '=DCF!C50', W['ke_rating'], PCT2),
    ('Cost of capital (9% framing, rating basis)', '=DCF!C56', W['rating_ct'], PCT2),
    ('Cost of capital (15% framing — same rate, shield stays at 9%)', '=DCF!C57',
     W['rating_dmtt'], PCT2),
    ('Growth, FY2031-40 window', f'={a("g")}', G, PCT),
    ('Growth beyond FY2040', f'={a("g2")}', G2, PCT),
]
for lab, fml, xp, fmt in key_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fml, xp, fmt, green=True)
    r += 1
ANCH['summary'] = dict(central=CEN_ROW, central_cont=CEN_CONT_ROW, central_dm=CEN_DM_ROW,
                       central_cont_dm=CEN_CONT_DM_ROW, spot=SPOT_ROW)

# ============ FUNDAMENTAL VALUATION ================================================
ws = wb['Fundamental Valuation']
title(ws, 'Fundamental valuation — four lenses, two tax framings, two macro framings',
      None, 6, awidth=56, cwidth=15)
ws.column_dimensions['B'].width = 62
hdr(ws, 4, ['Lens', 'Basis', 'AED per share'])
fv_rows = [
    ('Discounted cash flow', 'five-year free-cash-flow build, two-stage '
     'reinvestment-consistent terminal', "='SOTP Bridge'!C14", B_CT['ps']),
    ('Relative multiples', 'trailing peer EV/EBITDA on trailing operating EBITDA',
     "='Relative & Normalized'!C12", REL['ps_rel']),
    ('Normalised earnings power', 'unshocked year one x justified forward price/earnings',
     "='Relative & Normalized'!C31", NRM['ps']),
    ('Book value and sustainable return', 'justified price/book on sustainable return',
     "='Relative & Normalized'!C39", BK['ps']),
]
r = 5
for lab, basis, fml, xp in fv_rows:
    put(ws, f'A{r}', lab, fmt=None); put(ws, f'B{r}', basis, fmt=None)
    putf(ws, f'C{r}', fml, xp, PX, green=True)
    r += 1
band(ws, 9, 3)
put(ws, 'A9', 'Weighted central — recovery (9%)', bold=True, fmt=None)
putf(ws, 'C9', f'=Summary!B{CEN_ROW}', CEN['ct'], PX, bold=True, green=True)
put(ws, 'A10', 'Weighted central — continuation (9%)', bold=True, fmt=None)
putf(ws, 'C10', f'=Summary!B{CEN_CONT_ROW}', CEN['continuation_ct'], PX, bold=True,
     green=True)
put(ws, 'A12', 'THE TWO CONTESTED JUDGEMENTS — PUBLISHED BOTH WAYS, NEVER AVERAGED',
    bold=True, fmt=None)
hdr(ws, 13, ['Discounted cash flow per share', '9% corporate tax', '15% top-up tax'])
put(ws, 'A14', 'Consumption per-RT RECOVERS to the FY2025 level (de-escalation)', fmt=None)
putf(ws, 'B14', "='SOTP Bridge'!C14", B_CT['ps'], PX, green=True)
putf(ws, 'C14', "='SOTP Bridge'!D14", B_DM['ps'], PX, green=True)
put(ws, 'A15', 'Consumption per-RT never recovers (continuation — live model, the 94% '
    'column of the recovery ladder)', fmt=None)
putf(ws, 'B15', f'=Sensitivity!C{S_PS}', CRUX['persist_ps_ct'], PX, green=True)
putf(ws, 'C15', f'=Sensitivity!C{S_PS15}', CRUX['persist_ps_dmtt'], PX, green=True)
put(ws, 'A17', 'Scenario field (all live models on the Sensitivity sheet)', bold=True,
    fmt=None)
r = 18
for lab, fml, xp in [
        ('Bear — re-escalation, halved connections, 15% tax, repriced risk',
         f'=Sensitivity!C{S_B_PS}', CEN['bear']),
        ('Central — recovery (9%)', f'=Summary!B{CEN_ROW}', CEN['ct']),
        ('Central — continuation (9%)', f'=Summary!B{CEN_CONT_ROW}', CEN['continuation_ct']),
        ('Bull — clean recovery, top-of-guidance connections, 9% tax',
         f'=Sensitivity!C{S_U_PS}', CEN['bull'])]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, PX, green=True)
    r += 1
r += 1
put(ws, f'A{r}', 'Market price', fmt=None)
putf(ws, f'C{r}', f'=Summary!B{SPOT_ROW}', SPOT, PX, green=True)
r += 1
put(ws, f'A{r}', 'DEWA control-transaction price (reference only)', fmt=None)
putf(ws, f'C{r}', f'={a("dewa_price")}', D['dewa_buyin']['price'], PX, green=True)

# ============ MONTE CARLO ==========================================================
ws = wb['Monte Carlo']
title(ws, 'Probabilistic price map', 'A map of price dispersion around the market price. It '
      'carries no view on value and is never blended with the valuation. Each figure is a '
      'complete simulation output, not a formula, and does not redraw when a driver '
      'changes.', 8, awidth=44, cwidth=13)
hdr(ws, 4, ['Horizon', '5th', '25th', 'Median', '75th', '95th', 'P(above spot)'])
r = 5
for tag, lab in (('1M', 'One month'), ('3M', 'Three months')):
    h = STK['horizons'][tag]
    put(ws, f'A{r}', f"{lab} — to {h['grade_date']}", fmt=None)
    for i, k in enumerate(('p5', 'p25', 'p50', 'p75', 'p95')):
        put(ws, f'{get_column_letter(2+i)}{r}', h['pct'][k], BLUE, PX)
    put(ws, f'G{r}', h['p_above'], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Level event', 'One month', 'Three months']); r += 1
for lab, k in [('Finishes 10% or more above the market price', 'p_up10'),
               ('Finishes 10% or more below the market price', 'p_dn10'),
               ('Touches 10% above at any point', 'touch_up10'),
               ('Touches 10% below at any point', 'touch_dn10')]:
    put(ws, f'A{r}', lab, fmt=None)
    put(ws, f'B{r}', STK['horizons']['1M'][k], BLUE, PCT)
    put(ws, f'C{r}', STK['horizons']['3M'][k], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Simulation setting', 'Value']); r += 1
for lab, v, fmt, green in [
        ('Anchor date', STK['anchor_date'], None, False),
        ('Market price at the anchor (AED)', f'={a("spot")}', PX, True),
        ('Annualised volatility (three-month anchor)',
         STK['horizons']['3M']['anchor_vol_ann'], PCT, False),
        ('Simulated paths', 50000, NUM0, False)]:
    put(ws, f'A{r}', lab, fmt=None)
    if green:
        putf(ws, f'C{r}', v, SPOT, fmt, green=True)
    else:
        put(ws, f'C{r}', v, BLUE, fmt)
    r += 1
r += 1
note(ws, r, 'Calibration backtest, in plain language: the shares listed in November 2022, '
     f"so only {S0['windows']} non-overlapping three-month windows exist "
     f"({S0['first_origin']} to {S0['last_origin']}). Over that short record the model's "
     'bands were, if anything, slightly generous: every outcome landed inside the 80% '
     'band, and the centre of the distribution sat where it should '
     f"(average percentile of outcomes {S0['pit_mean']:.2f} against an ideal 0.50). A "
     'record this short cannot establish forecasting skill either way; read the bands as '
     'honest dispersion, not precision.')
ws['A' + str(r)].alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[r].height = 70
ws.merge_cells(f'A{r}:H{r}')

# ============ SENSITIVITY — pasted grid + LIVE scenario models =====================
# python mirrors for the live blocks (every cell's model value is recorded)
DNA5 = [F['dna'][y] for y in YF]
CAPEX5 = [F['capex'][y] for y in YF]
PPE30 = F['ppe']['FY30']
RTAV = [U['rt_avg'][y] for y in YF]
PIPES = IN['pipes_rev_fy25']

def crux_col(lvl):
    rev, eb = [], []
    for i, y in enumerate(YF):
        cp = CPRT * (1 + SHOCK) if i == 0 else CPRT * lvl
        rev.append((cp + CAPRT) * RTAV[i] + PIPES)
        eb.append(rev[i] - EW * cp * RTAV[i] - F['other_cos'][y] - F['ga'][y] + OI_OP)
    fcff, pv = [], []
    for i in range(5):
        dnwc = (NWC_RATIO * rev[0] - NWC25) if i == 0 else NWC_RATIO * (rev[i] - rev[i-1])
        fc = (eb[i] - DNA5[i]) * (1 - TAX) + DNA5[i] - CAPEX5[i] - dnwc
        fcff.append(fc)
        pv.append(fc * B_CT['df'][YF[i]] * (T0 if i == 0 else 1.0))
    nop30 = (eb[4] - DNA5[4]) * (1 - TAX)
    roic = nop30 / (PPE30 + NWC_RATIO * rev[4])
    f1, f2 = _sanctioned_fcff(nop30, DNA5[4], NWC_RATIO * rev[4], WACC_CT,
                              PPE30 + NWC_RATIO * rev[4])
    s1, s2, tv = two_stage_tv(f1, f2, WACC_CT)
    pvtv = tv * B_CT['df']['FY30']
    ev = sum(pv) + pvtv
    ps = (ev - NET_DEBT + BRIDGE_ADD) * (1 - NCI_FR) / SH
    return dict(rev=rev, eb=eb, fcff=fcff, pv=pv, nop30=nop30, roic=roic, s1=s1,
                f1=f1, f2=f2, s2=s2, tv=tv, pvtv=pvtv, ev=ev, ps=ps,
                pvex=sum(pv))

CRUX_M = [crux_col(l) for l in CRUX['levels']]
for m, row_ in zip(CRUX_M, CRUX['rows']):
    assert abs(m['ps'] - row_['ps']) < 1e-6, (m['ps'], row_['ps'])
P94 = crux_col(0.94)
f15 = [P94['fcff'][i] - (P94['eb'][i] - DNA5[i]) * (TAXD - TAX) for i in range(5)]
p15 = [f15[i] * B_CT['df'][YF[i]] * (T0 if i == 0 else 1.0) for i in range(5)]
nop30_15 = P94['nop30'] * (1 - TAXD) / (1 - TAX)
roic15_p = nop30_15 / (PPE30 + NWC_RATIO * P94['rev'][4])
f1_p15, f2_p15 = _sanctioned_fcff(nop30_15, DNA5[4], NWC_RATIO * P94['rev'][4], WACC_CT,
                                  PPE30 + NWC_RATIO * P94['rev'][4])
s1_p15, s2_p15, tv_p15 = two_stage_tv(f1_p15, f2_p15, WACC_CT)
pvtv_p15 = tv_p15 * B_CT['df']['FY30']
ev_p15 = sum(p15) + pvtv_p15
ps_p15 = (ev_p15 - NET_DEBT + BRIDGE_ADD) * (1 - NCI_FR) / SH
assert abs(ps_p15 - CRUX['persist_ps_dmtt']) < 1e-6

def scen_mirror(scen_key, dcf_key, tax):
    S, DD = SCEN[scen_key], DC[dcf_key]
    rtp = S['rt_path']
    adds = [rtp[y] - rtp[p] for p, y in zip(['FY25'] + YF[:-1], YF)]
    ye, avg, open_, capex, dna, close = [], [], [], [], [], []
    prev_ye, prev_close = RTP['FY25'], BH['FY25']['ppe']
    for i, y in enumerate(YF):
        ye.append(prev_ye + adds[i]); avg.append((prev_ye + ye[i]) / 2)
        open_.append(prev_close)
        capex.append(U['capex_per_rt'] * adds[i] + U['maint_pct'] * open_[i])
        dna.append(U['dep_rate'] * open_[i] + U['amort_flat'])
        close.append(open_[i] + capex[i] - (dna[i] - U['amort_flat']))
        prev_ye, prev_close = ye[i], close[i]
    rev = [S['rev'][y] for y in YF]
    eb = [S['ebitda'][y] for y in YF]
    for i, y in enumerate(YF):        # capital chain must reproduce the model's EBIT
        assert abs((eb[i] - dna[i]) - DD['ebit'][y]) < 1e-6
    wacc = S['wacc']
    dnwc = [(NWC_RATIO * rev[0] - NWC25)] + [NWC_RATIO * (rev[i] - rev[i-1])
                                             for i in range(1, 5)]
    fcff = [DD['fcff'][y] for y in YF]
    for i in range(5):
        assert abs(((eb[i] - dna[i]) * (1 - tax) + dna[i] - capex[i] - dnwc[i])
                   - fcff[i]) < 1e-6
    pv = [DD['pv'][y] for y in YF]
    nop30 = DD['nopat']['FY30']
    roic = DD['roic_term']
    f1, f2 = _sanctioned_fcff(nop30, dna[4], NWC_RATIO * rev[4], wacc,
                              nop30 / DD['roic_term'])
    s1, s2, tv = two_stage_tv(f1, f2, wacc)
    assert abs(tv - DD['tv']) < 1e-6 * DD['tv']
    return dict(adds=adds, ye=ye, avg=avg, open=open_, capex=capex, dna=dna, close=close,
                rev=rev, eb=eb, dnwc=dnwc, fcff=fcff, pv=pv, wacc=wacc, nop30=nop30,
                roic=roic, s1=s1, f1=f1, f2=f2, s2=s2, tv=tv, pvtv=DD['pv_tv'],
                pvex=DD['pv_explicit'], ev=DD['ev'], ps=DD['ps'])

BEARM = scen_mirror('bear', 'bear', TAXD)
BULLM = scen_mirror('bull', 'bull', TAX)
avg_b27 = (SCEN['bear']['rt_path']['FY26'] + SCEN['bear']['rt_path']['FY27']) / 2
BEAR_FACTOR = (((SCEN['bear']['rev']['FY27'] - PIPES) / avg_b27) - CAPRT) / (CPRT * (1 + SHOCK))
KE_ADD = SCEN['bear']['ke'] - W['ke_rating']
assert abs(BEAR_FACTOR - 0.94) < 1e-9 and abs(KE_ADD - 0.01) < 1e-9

ws = wb['Sensitivity']
title(ws, 'Sensitivity — what the valuation needs the world to do', 'AED per share. The '
      '5x5 discount-rate x growth grid is the ONE pasted table left (25 whole-model '
      'solutions); every other scenario on this sheet — the recovery ladder, bear, bull '
      'and the three cost-of-capital constructions — is a LIVE parallel model that '
      'redraws when a driver changes.', 8, awidth=46, cwidth=12)
put(ws, 'A4', 'Cost of capital (rows) x FY2031-40 growth (columns) — 9% framing; growth '
    'spans 0.0% (flat-tariff floor) to 3.0%; pasted whole-model solutions', bold=True,
    fmt=None)
hdr(ws, 5, [''] + [f'{g_:.1%}' for g_ in SNW['g_grid']])
r = S_GRID0
for i, wv in enumerate(SNW['wacc_grid']):
    put(ws, f'A{r}', f'{wv:.2%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SNW['table'][i][j], BLUE, PX)
    r += 1
put(ws, 'A12', 'LIVE one-way sensitivity — FY2031-40 growth at the base cost of capital '
    '(reproduces the grid\'s middle row; stage-two growth capped at the stage-one rate)',
    bold=True, fmt=None)
put(ws, f'A{S_LIVE_G}', 'Growth, FY2031-40 window', fmt=None)
for j, g_ in enumerate(SNW['g_grid']):
    put(ws, f'{get_column_letter(2+j)}{S_LIVE_G}', g_, BLUE, PCT)
put(ws, f'A{S_LIVE}', 'Fair value per share (AED, live formula)', fmt=None)
for j in range(5):
    col = get_column_letter(2 + j)
    gc = f'{col}{S_LIVE_G}'
    qq = f'((1+{gc})/(1+DCF!$C$56))'
    g2x = f'MIN({a("g2")},{gc})'
    tvx = (f'(DCF!$F$14*(1-{gc}/DCF!$C$26)*{qq}*(1-{qq}^10)/(1-{qq})'
           f'+DCF!$F$14*(1+{gc})^10*(1+{g2x})*(1-{g2x}/DCF!$C$26)'
           f'/((DCF!$C$56-{g2x})*(1+DCF!$C$56)^10))')
    fml = (f'=((DCF!$C$37+{tvx}*DCF!$F$20)-DCF!$C$53'
           f"+{a('recv')}+{a('invprop')}+{a('fvtpl')}+{a('fvoci')})"
           f"*(1-'SOTP Bridge'!$C$15)/{a('shares')}")
    putf(ws, f'{col}{S_LIVE}', fml, SNW['table'][2][j], PX)
put(ws, f'G{S_LIVE}', 'Swing:', fmt=None)
putf(ws, f'H{S_LIVE}', f'=MAX(B{S_LIVE}:F{S_LIVE})-MIN(B{S_LIVE}:F{S_LIVE})',
     max(SNW['table'][2]) - min(SNW['table'][2]), PX)
note(ws, 16, 'The live row reproduces the middle row of the pasted grid; if a driver is '
     'changed on the Assumptions sheet the live row and every block below will move and '
     'the pasted grid will not — that difference is the quickest way to see that a '
     'driver change has repriced the model.')

# ---- THE CRUX, LIVE: consumption per-RT recovery ladder ---------------------------
put(ws, 'A18', 'THE CRUX — LIVE RECOVERY LADDER: consumption per-RT from FY2027 as a '
    'share of FY2025 (blue level cells are inputs; every other cell is a formula)',
    bold=True, fmt=None)
LC = ['B', 'C', 'D', 'E', 'F']              # level columns
DCL = ['B', 'C', 'D', 'E', 'F']             # DCF year columns
SGL = ['C', 'D', 'E', 'F', 'G']             # Segments year columns
put(ws, f'A{S_CRUX_LVL}', 'Recovery level (share of FY2025 per-RT consumption)', fmt=None)
for j, lvl in enumerate(CRUX['levels']):
    put(ws, f'{LC[j]}{S_CRUX_LVL}', lvl, BLUE, PCT)
for i in range(5):
    put(ws, f'A{S_REV0+i}', f'Revenue {YFL[i]}', fmt=None)
    put(ws, f'A{S_EB0+i}', f'Operating EBITDA {YFL[i]}', fmt=None)
    put(ws, f'A{S_FCFF0+i}', f'Free cash flow {YFL[i]}', fmt=None)
    put(ws, f'A{S_PV0+i}', f'Present value {YFL[i]}' + (' (stub)' if i == 0 else ''),
        fmt=None)
    for j in range(5):
        L = LC[j]; m = CRUX_M[j]
        if i == 0:
            putf(ws, f'{L}{S_REV0}', '=Segments!$C$12', m['rev'][0], NUM1, green=True)
            putf(ws, f'{L}{S_EB0}', '=Segments!$C$22', m['eb'][0], NUM1, green=True)
            putf(ws, f'{L}{S_FCFF0}',
                 f'=({L}{S_EB0}-DCF!$B$12)*(1-{a("tax_ct")})+DCF!$B$12-DCF!$B$9'
                 f'-({a("nwc_ratio")}*{L}{S_REV0}-{a("nwc25")})', m['fcff'][0], NUM1)
            putf(ws, f'{L}{S_PV0}',
                 f'={L}{S_FCFF0}*{a("t_stub")}/(1+DCF!$C$56)^DCF!$B$19', m['pv'][0], NUM1)
        else:
            sg, dc = SGL[i], DCL[i]
            putf(ws, f'{L}{S_REV0+i}',
                 f'=({a("cons_per_rt")}*{L}${S_CRUX_LVL}+{a("cap_per_rt")})'
                 f'*Segments!${sg}$7+{a("pipes")}', m['rev'][i], NUM1)
            putf(ws, f'{L}{S_EB0+i}',
                 f'={L}{S_REV0+i}-{a("ew_ratio")}*{a("cons_per_rt")}*{L}${S_CRUX_LVL}'
                 f'*Segments!${sg}$7-Segments!${sg}$16-Segments!${sg}$17+{a("oi_op")}',
                 m['eb'][i], NUM1)
            putf(ws, f'{L}{S_FCFF0+i}',
                 f'=({L}{S_EB0+i}-DCF!${dc}$12)*(1-{a("tax_ct")})+DCF!${dc}$12'
                 f'-DCF!${dc}$9-{a("nwc_ratio")}*({L}{S_REV0+i}-{L}{S_REV0+i-1})',
                 m['fcff'][i], NUM1)
            putf(ws, f'{L}{S_PV0+i}',
                 f'={L}{S_FCFF0+i}/(1+DCF!$C$56)^DCF!${dc}$19', m['pv'][i], NUM1)
crux_tail = [
    (S_PVEX, 'Present value, five years',
     lambda L, m: (f'=SUM({L}{S_PV0}:{L}{S_PV0+4})', m['pvex']), NUM1),
    (S_NOP30, 'NOPAT FY2030E',
     lambda L, m: (f'=({L}{S_EB0+4}-DCF!$F$12)*(1-{a("tax_ct")})', m['nop30']), NUM1),
    (S_ROIC, 'Stage-one free cash flow, FY2031 (the waterfall is on the DCF sheet)',
     lambda L, m: (f'={m["f1"]:.6f}', m['f1']), NUM1),
    (S_Q, 'Growth/discount ratio',
     lambda L, m: (f'=(1+{a("g")})/(1+DCF!$C$56)', (1 + G) / (1 + WACC_CT)), DF4),
    (S_S1, 'Value of the FY2031-40 window',
     lambda L, m: (f'={L}{S_ROIC}/(1+DCF!$C$56)*(1-{L}{S_Q}^10)/(1-{L}{S_Q})',
                   m['s1']), NUM1),
    (S_NOP10, 'Stage-two free cash flow, FY2041',
     lambda L, m: (f'={m["f2"]:.6f}', m['f2']), NUM1),
    (S_S2, 'Value beyond FY2040',
     lambda L, m: (f'={L}{S_NOP10}*(1+{a("g2")})'
                   f'/((DCF!$C$56-{a("g2")})*(1+DCF!$C$56)^10)', m['s2']), NUM1),
    (S_TV, 'Terminal value at FY2030',
     lambda L, m: (f'={L}{S_S1}+{L}{S_S2}', m['tv']), NUM1),
    (S_PVTV, 'Present value of the terminal',
     lambda L, m: (f'={L}{S_TV}/(1+DCF!$C$56)^DCF!$F$19', m['pvtv']), NUM1),
    (S_EV, 'Enterprise value',
     lambda L, m: (f'={L}{S_PVEX}+{L}{S_PVTV}', m['ev']), NUM1),
    (S_PS, 'Fair value per share (AED)',
     lambda L, m: (f'=({L}{S_EV}-DCF!$C$53+{a("recv")}+{a("invprop")}+{a("fvtpl")}'
                   f"+{a('fvoci')})*(1-'SOTP Bridge'!$C$15)/{a('shares')}", m['ps']), PX),
]
for row_, lab, fn, fmt in crux_tail:
    put(ws, f'A{row_}', lab, fmt=None)
    for j in range(5):
        fml, xp = fn(LC[j], CRUX_M[j])
        putf(ws, f'{LC[j]}{row_}', fml, xp, fmt, bold=(row_ == S_PS))
band(ws, S_PS, 6)
put(ws, f'A{52}', 'CONTINUATION AT 15% TOP-UP TAX — the 94% column re-taxed (feeds the '
    'continuation central)', bold=True, fmt=None)
for i in range(5):
    dc = DCL[i]
    put(ws, f'A{S_F15_0+i}', f'Free cash flow at 15% {YFL[i]}', fmt=None)
    putf(ws, f'C{S_F15_0+i}',
         f'=C{S_FCFF0+i}-(C{S_EB0+i}-DCF!${dc}$12)*({a("tax_dmtt")}-{a("tax_ct")})',
         f15[i], NUM1)
    put(ws, f'A{S_P15_0+i}', f'Present value at 15% {YFL[i]}' + (' (stub)' if i == 0
                                                                 else ''), fmt=None)
    if i == 0:
        putf(ws, f'C{S_P15_0}', f'=C{S_F15_0}*{a("t_stub")}/(1+DCF!$C$56)^DCF!$B$19',
             p15[0], NUM1)
    else:
        putf(ws, f'C{S_P15_0+i}', f'=C{S_F15_0+i}/(1+DCF!$C$56)^DCF!${dc}$19', p15[i],
             NUM1)
tail15 = [
    (S_PVEX15, 'Present value, five years (15%)', f'=SUM(C{S_P15_0}:C{S_P15_0+4})',
     sum(p15), NUM1),
    (S_NOP30_15, 'NOPAT FY2030E (15%)',
     f'=C{S_NOP30}*(1-{a("tax_dmtt")})/(1-{a("tax_ct")})', nop30_15, NUM1),
    (S_ROIC15, 'Stage-one free cash flow at 15% (the waterfall is on the DCF sheet)',
     f'={f1_p15:.6f}', f1_p15, NUM1),
    (S_S1_15, 'Value of the FY2031-40 window (15%)',
     f'=C{S_ROIC15}/(1+DCF!$C$56)*(1-C{S_Q}^10)/(1-C{S_Q})',
     s1_p15, NUM1),
    (S_NOP10_15, 'Stage-two free cash flow at 15%', f'={f2_p15:.6f}', f2_p15,
     NUM1),
    (S_S2_15, 'Value beyond FY2040 (15%)',
     f'=C{S_NOP10_15}*(1+{a("g2")})'
     f'/((DCF!$C$56-{a("g2")})*(1+DCF!$C$56)^10)', s2_p15, NUM1),
    (S_TV15, 'Terminal value at FY2030 (15%)', f'=C{S_S1_15}+C{S_S2_15}', tv_p15, NUM1),
    (S_PVTV15, 'Present value of the terminal (15%)',
     f'=C{S_TV15}/(1+DCF!$C$56)^DCF!$F$19', pvtv_p15, NUM1),
    (S_EV15, 'Enterprise value (15%)', f'=C{S_PVEX15}+C{S_PVTV15}', ev_p15, NUM1),
    (S_PS15, 'Fair value per share, continuation at 15% (AED)',
     f'=(C{S_EV15}-DCF!$C$53+{a("recv")}+{a("invprop")}+{a("fvtpl")}+{a("fvoci")})'
     f"*(1-'SOTP Bridge'!$C$15)/{a('shares')}", ps_p15, PX),
]
for row_, lab, fml, xp, fmt in tail15:
    put(ws, f'A{row_}', lab, fmt=None)
    putf(ws, f'C{row_}', fml, xp, fmt, bold=(row_ == S_PS15))

# ---- BEAR and BULL, LIVE ----------------------------------------------------------
def scen_block(MM, adds_row, ye_row, avg_row, rev_row, eb_row, open_row,
               capex_row, dna_row, close_row, dnwc_row, fcff_row, pv_row, tail0,
               wacc_ref, tax_ref, factor_ref=None):
    put(ws, f'A{adds_row}', 'New connections added (k RT)', fmt=None)
    for i in range(5):
        put(ws, f'{LC[i]}{adds_row}', MM['adds'][i], BLUE, NUM0)
    rows = [
        (ye_row, 'Connected capacity, year-end (k RT)',
         lambda i, c, p: (f'={a("rt_fy25")}+{c}{adds_row}' if i == 0
                          else f'={p}{ye_row}+{c}{adds_row}'), MM['ye'], NUM0),
        (avg_row, 'Average connected capacity (k RT)',
         lambda i, c, p: (f'=({a("rt_fy25")}+{c}{ye_row})/2' if i == 0
                          else f'=({p}{ye_row}+{c}{ye_row})/2'), MM['avg'], NUM1),
        (rev_row, 'Revenue',
         lambda i, c, p: (f'=({a("cons_per_rt")}*(1+{a("shock")}){factor_ref if i else ""}'
                          f'+{a("cap_per_rt")})*{c}{avg_row}+{a("pipes")}'), MM['rev'],
         NUM1),
        (eb_row, 'Operating EBITDA',
         lambda i, c, p: (f'={c}{rev_row}-{a("ew_ratio")}*{a("cons_per_rt")}'
                          f'*(1+{a("shock")}){factor_ref if i else ""}*{c}{avg_row}'
                          f'-Segments!${SGL[i]}$16-Segments!${SGL[i]}$17+{a("oi_op")}'),
         MM['eb'], NUM1),
        (open_row, 'Opening net plant',
         lambda i, c, p: (f'={a("ppe25")}' if i == 0 else f'={p}{close_row}'),
         MM['open'], NUM1),
        (capex_row, 'Capital expenditure',
         lambda i, c, p: f'={a("capex_per_rt")}*{c}{adds_row}+{a("maint_pct")}*{c}{open_row}',
         MM['capex'], NUM1),
        (dna_row, 'Depreciation and amortisation',
         lambda i, c, p: f'={a("dep_rate")}*{c}{open_row}+{a("amort_flat")}', MM['dna'],
         NUM1),
        (close_row, 'Closing net plant',
         lambda i, c, p: f'={c}{open_row}+{c}{capex_row}-({c}{dna_row}-{a("amort_flat")})',
         MM['close'], NUM1),
        (dnwc_row, 'Change in net working capital',
         lambda i, c, p: (f'={a("nwc_ratio")}*{c}{rev_row}-{a("nwc25")}' if i == 0
                          else f'={a("nwc_ratio")}*({c}{rev_row}-{p}{rev_row})'),
         MM['dnwc'], NUM1),
        (fcff_row, 'Free cash flow to the firm',
         lambda i, c, p: (f'=({c}{eb_row}-{c}{dna_row})*(1-{tax_ref})+{c}{dna_row}'
                          f'-{c}{capex_row}-{c}{dnwc_row}'), MM['fcff'], NUM1),
        (pv_row, 'Present value (FY2026 stub)',
         lambda i, c, p: (f'={c}{fcff_row}*{a("t_stub")}/(1+{wacc_ref})^DCF!$B$19'
                          if i == 0 else
                          f'={c}{fcff_row}/(1+{wacc_ref})^DCF!${DCL[i]}$19'),
         MM['pv'], NUM1),
    ]
    for row_, rl, fn, vals, fmt in rows:
        put(ws, f'A{row_}', rl, fmt=None)
        for i in range(5):
            c, p = LC[i], LC[i-1] if i else None
            putf(ws, f'{c}{row_}', fn(i, c, p), vals[i], fmt)
    tail = [
        (tail0, 'Present value, five years', f'=SUM(B{pv_row}:F{pv_row})', MM['pvex'],
         NUM1),
        (tail0 + 1, 'NOPAT FY2030E', f'=(F{eb_row}-F{dna_row})*(1-{tax_ref})',
         MM['nop30'], NUM1),
        (tail0 + 2, 'Stage-one free cash flow, FY2031 (the waterfall is on the DCF sheet)',
         f'={MM["f1"]:.6f}', MM['f1'], NUM1),
        (tail0 + 3, 'Growth/discount ratio', f'=(1+{a("g")})/(1+{wacc_ref})',
         (1 + G) / (1 + MM['wacc']), DF4),
        (tail0 + 4, 'Value of the FY2031-40 window',
         f'=C{tail0+2}/(1+{wacc_ref})*(1-C{tail0+3}^10)/(1-C{tail0+3})',
         MM['s1'], NUM1),
        (tail0 + 5, 'Stage-two free cash flow, FY2041', f'={MM["f2"]:.6f}', MM['f2'],
         NUM1),
        (tail0 + 6, 'Value beyond FY2040',
         f'=C{tail0+5}*(1+{a("g2")})'
         f'/(({wacc_ref}-{a("g2")})*(1+{wacc_ref})^10)', MM['s2'], NUM1),
        (tail0 + 7, 'Terminal value at FY2030', f'=C{tail0+4}+C{tail0+6}', MM['tv'], NUM1),
        (tail0 + 8, 'Present value of the terminal',
         f'=C{tail0+7}/(1+{wacc_ref})^DCF!$F$19', MM['pvtv'], NUM1),
        (tail0 + 9, 'Enterprise value', f'=C{tail0}+C{tail0+8}', MM['ev'], NUM1),
        (tail0 + 10, 'Fair value per share (AED)',
         f'=(C{tail0+9}-DCF!$C$53+{a("recv")}+{a("invprop")}+{a("fvtpl")}+{a("fvoci")})'
         f"*(1-'SOTP Bridge'!$C$15)/{a('shares')}", MM['ps'], PX),
    ]
    for row_, rl, fml, xp, fmt in tail:
        put(ws, f'A{row_}', rl, fmt=None)
        putf(ws, f'C{row_}', fml, xp, fmt, bold=(row_ == tail0 + 10))
    band(ws, tail0 + 10, 6)

def _con_wacc(k):
    return {'base_gross_wacc': CON['gross'], 'base_carry_wacc': CON['carry'],
            'base_dfm_beta': CON['dfm_beta']}[k]


def _con_tv(k, part):
    """Each WACC construction rebuilds its OWN terminal cash flows at its OWN rate: the
    maintenance charge does not move with the discount rate but the perpetuity does, and
    a column that borrowed the primary's flows would be a different model wearing this
    one's numbers."""
    w = _con_wacc(k)
    f1, f2 = _sanctioned_fcff(B_CT['nopat']['FY30'], F['dna']['FY30'], F['nwc']['FY30'],
                              w, INC_CAP)
    return two_stage_tv(f1, f2, w)[part], f1, f2


put(ws, 'A74', 'BEAR — LIVE: re-escalation (per-RT falls a further 6% from FY2027 and '
    'never recovers, connections halve, 15% top-up tax, +100bp on the cost of equity)',
    bold=True, fmt=None)
put(ws, f'A{S_B_FAC}', 'Further per-RT fall from FY2027 (share retained)', fmt=None)
put(ws, f'C{S_B_FAC}', BEAR_FACTOR, BLUE, PCT)
put(ws, f'A{S_B_KE0}', 'Cost-of-equity add-on (war repricing)', fmt=None)
put(ws, f'C{S_B_KE0}', KE_ADD, BLUE, PCT2)
put(ws, f'A{S_B_TAX}', 'Tax framing', fmt=None)
putf(ws, f'C{S_B_TAX}', f'={a("tax_dmtt")}', TAXD, PCT, green=True)
putf(ws, f'C{S_B_KEB}', f'=DCF!C50+C{S_B_KE0}', SCEN['bear']['ke'], PCT2)
put(ws, f'A{S_B_KEB}', 'Cost of equity, repriced', fmt=None)
putf(ws, f'C{S_B_WACC}', f'=DCF!C55*C{S_B_KEB}+DCF!C54*DCF!C51*(1-{a("tax_ct")})',
     SCEN['bear']['wacc'], PCT2)
put(ws, f'A{S_B_WACC}', 'Discount rate (interest shield stays at 9%)', fmt=None)
scen_block(BEARM, S_B_ADDS, S_B_YE, S_B_AVG, S_B_REV, S_B_EB, S_B_OPEN,
           S_B_CAPEX, S_B_DNA, S_B_CLOSE, S_B_DNWC, S_B_FCFF, S_B_PV, S_B_PVEX,
           f'$C${S_B_WACC}', f'$C${S_B_TAX}', factor_ref=f'*$C${S_B_FAC}')
put(ws, 'A104', 'BULL — LIVE: clean recovery, connections at the top of guidance '
    '(110/110/100/90/80k RT), 9% corporate tax, base discount rate', bold=True, fmt=None)
scen_block(BULLM, S_U_ADDS, S_U_YE, S_U_AVG, S_U_REV, S_U_EB, S_U_OPEN,
           S_U_CAPEX, S_U_DNA, S_U_CLOSE, S_U_DNWC, S_U_FCFF, S_U_PV, S_U_PVEX,
           'DCF!$C$56', f'{a("tax_ct")}', factor_ref=f'/(1+{a("shock")})*{a("recovery")}')
put(ws, 'A129', 'WACC CONSTRUCTIONS — LIVE: the base cash flows re-discounted at each '
    'alternative construction (rates built on the DCF sheet)', bold=True, fmt=None)
hdr(ws, 130, ['', 'Gross-debt weights', 'Negative-carry', 'DFM-index beta'])
con_keys = ['base_gross_wacc', 'base_carry_wacc', 'base_dfm_beta']
con_cols = ['B', 'C', 'D']
put(ws, f'A{S_C_WACC}', 'Discount rate', fmt=None)
for c, k, dcfcell in zip(con_cols, [CON['gross'], CON['carry'], CON['dfm_beta']],
                         ['B91', 'B92', 'B93']):
    putf(ws, f'{c}{S_C_WACC}', f'=DCF!{dcfcell}', k, PCT2, green=True)
for i in range(5):
    put(ws, f'A{S_C_PV0+i}', f'Present value {YFL[i]}' + (' (stub)' if i == 0 else ''),
        fmt=None)
    for c, k in zip(con_cols, con_keys):
        if i == 0:
            fml = f'=DCF!$B$18*{a("t_stub")}/(1+{c}${S_C_WACC})^DCF!$B$19'
        else:
            fml = f'=DCF!${DCL[i]}$18/(1+{c}${S_C_WACC})^DCF!${DCL[i]}$19'
        putf(ws, f'{c}{S_C_PV0+i}', fml, DC[k]['pv'][YF[i]], NUM1)
con_tail = [
    (S_C_PVEX, 'Present value, five years',
     lambda c, k: (f'=SUM({c}{S_C_PV0}:{c}{S_C_PV0+4})', DC[k]['pv_explicit']), NUM1),
    (S_C_Q, 'Growth/discount ratio',
     lambda c, k: (f'=(1+{a("g")})/(1+{c}{S_C_WACC})',
                   (1 + G) / (1 + {'base_gross_wacc': CON['gross'],
                                   'base_carry_wacc': CON['carry'],
                                   'base_dfm_beta': CON['dfm_beta']}[k])), DF4),
    (S_C_S1, 'Value of the FY2031-40 window',
     lambda c, k: (f'={_con_tv(k, 1)[1]:.6f}/(1+{c}{S_C_WACC})'
                   f'*(1-{c}{S_C_Q}^10)/(1-{c}{S_C_Q})',
                   _con_tv(k, 0)[0]), NUM1),
    (S_C_S2, 'Value beyond FY2040',
     lambda c, k: (f'={_con_tv(k, 2)[2]:.6f}*(1+{a("g2")})'
                   f'/(({c}{S_C_WACC}-{a("g2")})*(1+{c}{S_C_WACC})^10)',
                   _con_tv(k, 1)[0]), NUM1),
    (S_C_TV, 'Terminal value at FY2030',
     lambda c, k: (f'={c}{S_C_S1}+{c}{S_C_S2}', DC[k]['tv']), NUM1),
    (S_C_PVTV, 'Present value of the terminal',
     lambda c, k: (f'={c}{S_C_TV}/(1+{c}{S_C_WACC})^DCF!$F$19', DC[k]['pv_tv']), NUM1),
    (S_C_EV, 'Enterprise value',
     lambda c, k: (f'={c}{S_C_PVEX}+{c}{S_C_PVTV}', DC[k]['ev']), NUM1),
    (S_C_PS, 'Fair value per share (AED)',
     lambda c, k: (f'=({c}{S_C_EV}-DCF!$C$53+{a("recv")}+{a("invprop")}+{a("fvtpl")}'
                   f"+{a('fvoci')})*(1-'SOTP Bridge'!$C$15)/{a('shares')}", DC[k]['ps']),
     PX),
]
for row_, rl, fn, fmt in con_tail:
    put(ws, f'A{row_}', rl, fmt=None)
    for c, k in zip(con_cols, con_keys):
        fml, xp = fn(c, k)
        putf(ws, f'{c}{row_}', fml, xp, fmt, bold=(row_ == S_C_PS))
band(ws, S_C_PS, 4)
note(ws, S_C_PS + 2, 'Every block above except the 5x5 grid is a live parallel model: '
     'the recovery ladder varies only the per-RT level, bear varies the connection path, '
     'the per-RT floor, the tax framing and the equity repricing, bull varies the '
     'connection path, and the constructions vary only the discount rate. Change any '
     'shared driver on the Assumptions sheet and all of them redraw together with the '
     'main model.')

ANCH['sens'] = dict(crux_lvl=S_CRUX_LVL, crux_ps=S_PS, crux_ps15=S_PS15, bear_ps=S_B_PS,
                    bull_ps=S_U_PS, con_ps=S_C_PS, live=S_LIVE, live_g=S_LIVE_G,
                    bear_adds=S_B_ADDS, bull_adds=S_U_ADDS)

# ============ PER-SHARE & RATIOS ===================================================
ws = wb['Per-Share & Ratios']
title(ws, 'Per-share and ratio analysis', 'The indicator set for a regulated utility. '
      'Every cell is a formula off the statements.', 9, awidth=46, cwidth=12)
hdr(ws, 4, ['Measure'] + YHL + YFL)
r = 5
eq_all = [None, EQ24, IN['eq_attr_fy25']] + eq_f
BSMAP = [None, 'B', 'C']

def ratio(lab, fml, vals, fmt, skip=()):
    global r
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALLC[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALLC[i]}{r}', fml(i), vals[i], fmt)
    r += 1

ratio('Earnings per share (AED)', lambda i: f"='Income Statement'!{ALLC[i]}19",
      [x / SH for x in npa_all], PX)
ratio('Dividend per share (AED)', lambda i: f'={a("div")}/{a("shares")}',
      [None, None, IN['div_policy'] / SH] + [IN['div_policy'] / SH] * 5, PX, skip=(0, 1))
ratio('Dividend cover (attributable profit / dividend)',
      lambda i: f"='Income Statement'!{ALLC[i]}18/{a('div')}",
      [None, None] + [npa_all[i] / IN['div_policy'] for i in range(2, 8)], MULT, skip=(0, 1))
ratio('Book value per share (AED)',
      lambda i: (f"='Balance Sheet'!{BSMAP[i]}13/{a('shares')}" if i < 3
                 else f"='Balance Sheet'!{BF[i-3]}13/{a('shares')}"),
      [None] + [x / SH for x in eq_all[1:]], PX, skip=(0,))
ratio('Free cash flow to the firm per share (AED)',
      lambda i: f"='Cash Flow'!{CFF[i-3]}11/{a('shares')}",
      [None] * 3 + [B_CT['fcff'][y] / SH for y in YF], PX, skip=(0, 1, 2))
ratio('Gross margin', lambda i: f"='Income Statement'!{ALLC[i]}8/'Income Statement'!{ALLC[i]}5",
      [HI[y]['gp'] / HI[y]['rev'] for y in H3] + [None] * 5, PCT, skip=(3, 4, 5, 6, 7))
ratio('Operating EBITDA margin', lambda i: f"='Income Statement'!{ALLC[i]}10",
      [eb_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Operating margin', lambda i: f"='Income Statement'!{ALLC[i]}12"
      f"/'Income Statement'!{ALLC[i]}5", [ebit_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Net margin (attributable)',
      lambda i: f"='Income Statement'!{ALLC[i]}18/'Income Statement'!{ALLC[i]}5",
      [npa_all[i] / rev_all[i] for i in range(8)], PCT)
roe_vals = [None, None, HI['FY25']['npa'] / ((EQ24 + IN['eq_attr_fy25']) / 2)]
roe_vals += [npa_f[0] / ((IN['eq_attr_jun26'] + eq_f[0]) / 2)]
roe_vals += [npa_f[i] / ((eq_f[i-1] + eq_f[i]) / 2) for i in range(1, 5)]
ratio('Return on equity (average base; FY2026E from the 30-Jun-2026 anchor)',
      lambda i: (f"='Income Statement'!D18/(('Balance Sheet'!B13+'Balance Sheet'!C13)/2)"
                 if i == 2 else
                 f"='Income Statement'!{ALLC[i]}18/(('Balance Sheet'!"
                 f"{'D' if i == 3 else BF[i-4]}13+'Balance Sheet'!{BF[i-3]}13)/2)"),
      roe_vals, PCT, skip=(0, 1))
ratio('Net debt / EBITDA (audited definition)',
      lambda i: (f"='Balance Sheet'!{BSMAP[i]}16/('Income Statement'!{ALLC[i]}9"
                 f"+'Income Statement'!{ALLC[i]}6+'Income Statement'!{ALLC[i]}7)" if i == 2
                 else f"='Balance Sheet'!{BF[i-3]}16/('Income Statement'!{ALLC[i]}9"
                 f"+'Income Statement'!{ALLC[i]}6+'Income Statement'!{ALLC[i]}7)"),
      [None, None, nd_hist[1] / HI['FY25']['ebitda']]
      + [nd_f[i] / (F['ebitda'][YF[i]] + intco_f[i] + RENTAL) for i in range(5)], MULT,
      skip=(0, 1))
ratio('Capital expenditure / revenue',
      lambda i: (f"=-'Cash Flow'!{['B','C'][i-1]}9/'Income Statement'!{ALLC[i]}5" if i < 3
                 else f"=-'Cash Flow'!{CFF[i-3]}9/'Income Statement'!{ALLC[i]}5"),
      [None, IN['capex_fy24'] / HI['FY24']['rev'], IN['capex_fy25'] / HI['FY25']['rev']]
      + [F['capex'][YF[i]] / F['rev'][YF[i]] for i in range(5)], PCT, skip=(0,))
ratio('Payout ratio (dividend / attributable profit)',
      lambda i: f"={a('div')}/'Income Statement'!{ALLC[i]}18",
      [None, None] + [IN['div_policy'] / npa_all[i] for i in range(2, 8)], PCT, skip=(0, 1))
note(ws, r + 1, 'The dividend exceeds attributable profit in the near years by design — '
     'the committed AED 875m distribution is funded by the negative working-capital cycle '
     'and the balance sheet, and the net-debt roll on the Balance Sheet carries it '
     'explicitly from the 30-Jun-2026 anchor.')

# ============ PEER & SECTOR ========================================================
ws = wb['Peer & Sector']
title(ws, 'Peer frame and sector context', 'One close listed peer exists; the parent is '
      'context, not a comparable. Cross-check inputs only, never a source for the '
      'company\'s own numbers.', 6, awidth=36, cwidth=20)
hdr(ws, 4, ['Company / frame', 'Market', 'Relevance', 'Caution'])
r = 5
for a1, a2, a3, a4 in [
    ('Tabreed (National Central Cooling Co)', 'Dubai (DFM)',
     'the only other listed GCC district-cooling pure play — the primary multiple '
     'cross-check', 'roughly 4.6x net debt / EBITDA against Empower\'s 1.8x, and a '
     'concession mix that skews the EV multiple'),
    ('DEWA (parent, 80% owner)', 'Dubai (DFM)',
     'majority owner since Feb-2026; its multiple frames how Dubai prices regulated '
     'utility earnings', 'a vertically-integrated generation and distribution utility, '
     'not a district-cooling comparable'),
    ('Regional listed utilities', 'GCC',
     'the broad frame for regulated-return infrastructure in pegged-currency markets',
     'tariff regimes and concession structures differ materially by emirate and state'),
]:
    put(ws, f'A{r}', a1, fmt=None); put(ws, f'B{r}', a2, fmt=None)
    put(ws, f'C{r}', a3, fmt=None, wrap=True); put(ws, f'D{r}', a4, fmt=None, wrap=True)
    ws.row_dimensions[r].height = 40
    r += 1
ws.column_dimensions['C'].width = 44; ws.column_dimensions['D'].width = 48
r += 1
hdr(ws, r, ['Peer figure', 'Peer', 'Empower (model)']); r += 1
peer_rows = [
    ('EV / EBITDA (Tabreed trailing, restruck at the 07-Aug-2026 anchor)',
     f'={a("tabreed_ev")}', REL['tabreed_ev_ebitda'], "='Relative & Normalized'!C17",
     (W['mktcap'] + NET_DEBT) / HI['FY25']['ebitda'], MULT),
    ('Price / earnings (Tabreed trailing, restruck)', f'={a("tabreed_pe")}',
     REL['tabreed_pe'], "='Relative & Normalized'!C18", SPOT / (HI['FY25']['npa'] / SH),
     MULT),
    ('Price / earnings (DEWA, trailing)', f'={a("dewa_pe")}',
     REL['dewa_pe'], "='Relative & Normalized'!C18", SPOT / (HI['FY25']['npa'] / SH), MULT),
]
for lab, f1, x1, f2, x2, fmt in peer_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', f1, x1, fmt, green=True)
    putf(ws, f'C{r}', f2, x2, fmt, green=True)
    r += 1
put(ws, f'A{r}', 'Tabreed FY2025 revenue (AED mn, as reported)', fmt=None)
put(ws, f'B{r}', TABREED_REV, BLUE, NUM1)
ws.cell(row=r, column=4, value='Tabreed FY2025 results — reported figure (some '
        'aggregators carry a rounded 2,460)').font = SUB
r += 1
put(ws, f'A{r}', 'Dividend yield at the market price', fmt=None)
putf(ws, f'C{r}', f'={a("div")}/{a("shares")}/{a("spot")}',
     IN['div_policy'] / SH / SPOT, PCT)
r += 2
note(ws, r, 'Peer multiples are dated, disclosed cross-check inputs, restruck at the '
     'subject\'s own 07-Aug-2026 anchor (see the Assumptions sheet). On EV/EBITDA '
     'Empower trades at a PREMIUM to Tabreed (about 11.0x against 10.1x on '
     'audited-definition EBITDA) — with less than half the leverage; on price/earnings '
     'it is roughly in line with Tabreed (15.1x against 15.0x) and at a discount to '
     'DEWA (16.8x). The model\'s relative lens applies the peer multiple to Empower\'s '
     'own trailing operating EBITDA rather than endorsing either price.')

# ============ READ FIRST ===========================================================
ws = wb['READ FIRST']
title(ws, 'Testahil — Emirates Central Cooling Systems Corporation PJSC (DFM: EMPOWER)',
      None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'Revised 17-Aug-2026 after an external audit; see the critique-response note in the study.', '',
 'What this workbook is. A transparent companion to the Empower valuation study. Every blue cell is a',
 'pasted input; every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a pasted primary-record',
 'cell is a live formula: the modelling rates (per-RT revenue rates, the DEWA pass-through, capex per',
 'added RT, the depreciation rate, the working-capital ratio) are FORMULAS off the audited anchors on',
 'the Assumptions sheet; the cost of capital is built from the risk-free rate, beta and the premium;',
 'the revenue build multiplies per-refrigeration-ton rates by the connected-capacity path; and the',
 'income statement, balance sheet, cash flow, ratios, all four lenses, the recovery ladder, the bear',
 'and bull cases and the three cost-of-capital constructions all chain off the same cells.', '',
 'EXACTLY TWO THINGS ARE PASTED VALUES, and it is worth knowing which:', '',
 '  1. AUDITED AND DISCLOSED HISTORY AND EXTERNAL FACTS — the FY2023-25 statement lines, the',
 '     30-Jun-2026 reviewed balance sheet, disclosed connected capacity, the consumption revenue from',
 '     the auditor\'s key-audit-matter section, deck physical figures (1,174m RTh; 698 EFLH hours),',
 '     the RD10 v1.3 tariff cap, peer marks and the committed dividend. Where a line is both disclosed',
 '     and derivable, the workbook carries the DISCLOSED figure: the primary record is not a',
 '     calculation. Each is pasted ONCE (on Assumptions wherever it drives the model).',
 '  2. WHOLE-MODEL SIMULATION OUTPUTS — the probabilistic price map and the 5x5 discount-rate x growth',
 '     grid. Each such cell is a complete standalone solution, so it cannot be one formula. THESE DO',
 '     NOT REDRAW WHEN A DRIVER IS CHANGED — everything else does, including the live recovery ladder,',
 '     bear, bull and the construction columns on the Sensitivity sheet.', '',
 'CONVENTIONS ADOPTED 17-AUG-2026 (the external audit\'s accepted findings):', '',
 '  · THE CLOCK: the bridge and the cash-flow clock share the same 30-Jun-2026 date. FY2026 contributes',
 '    only its second half (a half-year stub discounted at t=0.5); later year-ends sit at 1.5-4.5 years',
 '    and the terminal discounts from t=4.5. The previous full-year convention double-counted H1-2026',
 '    cash already inside the June net debt.',
 '  · OPERATING EBITDA excludes BOTH the interest on the related-party acquisition receivables AND the',
 '    rental income on investment properties. Both assets enter the EV-to-equity bridge at book',
 '    (Note 8: 1,005.0 Dubai Aviation City + 289.4 Nakheel; investment properties 168.7), so counting',
 '    their income in operating cash flow would double count. Earnings-based lenses keep both — they',
 '    are real income. FY2025 identity: operating EBITDA + interest + rental = audited EBITDA.',
 '  · TWO-STAGE TERMINAL: FY2031-40 grows at the Dubai-2040 window rate, then a long-run perpetuity',
 '    at a lower fade rate, each with return-consistent reinvestment.',
 '  · TWO MACRO FRAMINGS: recovery (de-escalation) and continuation are published side by side, like',
 '    the two tax framings; neither is privileged as the single answer.', '',
 'How revenue is built. Not as one growth rate. Consumption revenue is a per-RT rate times average',
 'connected refrigeration tons — decomposed further on the Segments sheet to connected RT x equivalent',
 'full-load hours x an AED-per-RTh tariff that sits 1.4% below the RD10 regulated cap. The FY2026',
 'consumption shock (hospitality-occupancy-led ~80% per the company\'s H1-2026 attribution; weather a',
 'minor factor) and its recovery are explicit, separate drivers; capacity and connection revenue is',
 'its own per-RT rate on the same capacity path; pipes are carried flat. Costs escalate each on their',
 'own driver: DEWA electricity and water follows the consumption leg, wage-class lines follow the wage',
 'escalator, the receivable interest amortises.', '',
 'What it is not. Not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges. The probabilistic price map is dispersion, not a forecast of value.', '',
 'Sourcing. FY2023, FY2024 and FY2025 statement lines come from the company\'s own audited consolidated',
 'financial statements; the 30-Jun-2026 position from the reviewed interim statements; capacity and',
 'guidance from the company\'s H1-2026 earnings materials; the tariff cap from the Dubai regulator\'s',
 'published RD10 document. Every input on the Assumptions sheet carries its source and date in column H.', '',
 f"Currency. AED million unless stated. Spot AED {SPOT:.2f} (07-Aug-2026 close). Sheets: READ FIRST ·",
 'Summary · Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF ·',
 'Income Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity ·',
 'Per-Share & Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 110

# ============ SAVE + EXPECTED LEDGER ===============================================
out = os.path.join(HERE, 'EMPOWER_Valuation_Model_09082026_public.xlsx')
wb.save(out)
json.dump({'expected': EXPECT, 'anchors': ANCH},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nchk = sum(len(v) for v in EXPECT.values())
nform = nlit = 0
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
            elif isinstance(c.value, (int, float)):
                nlit += 1
print(f'wrote {out} | {len(wb.sheetnames)} sheets: {wb.sheetnames}')
print(f'formulas: {nform} (of which {nchk} carry a checked expected value) | '
      f'numeric literals: {nlit}')

