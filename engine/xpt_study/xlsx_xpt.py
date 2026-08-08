"""XPTUSD_Valuation_Model_20072026_public.xlsx — openpyxl builder.
16 sheets: silver Combined model layout (the metals exemplar) + 'Supply & Cost Curve'
+ 'Ledger Cohorts' to honour the TMPV 16-sheet standard. Blue = input, black = formula,
green = cross-sheet link. MC percentiles are engine outputs pasted as values (labelled)."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D = json.load(open('study_numbers_xpt.json'))
S0 = json.load(open('step0_results.json'))
MC, Z, A, T, PR, PA, RA = D['mc'], D['zone'], D['anchors'], D['tech'], D['price'], D['path'], D['ratio']
spot = D['meta']['spot']

wb = openpyxl.Workbook()
INK = '1C3A36'; GREY = '6E7B77'; PANEL = 'EAF0EE'; CREAM = 'F6F1E6'
H = Font(name='Arial', size=12, bold=True, color=INK)
HD = Font(name='Arial', size=10, bold=True, color=INK)
N = Font(name='Arial', size=10, color=INK)
NB = Font(name='Arial', size=10, bold=True, color=INK)
SM = Font(name='Arial', size=9, color=GREY)
BLUE = Font(name='Arial', size=10, color='0000FF')       # hardcoded inputs
GRN = Font(name='Arial', size=10, color='008000')        # cross-sheet links
FILL_H = PatternFill('solid', fgColor=PANEL)
FILL_C = PatternFill('solid', fgColor=CREAM)
TH = Border(bottom=Side(style='thin', color='C9D4D1'))

def sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = True
    return ws

def put(ws, addr, val, font=N, fill=None, fmt=None, wrap=False, align=None):
    c = ws[addr]; c.value = val; c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if wrap or align:
        c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical='top')
    return c

def widths(ws, ww):
    for col, w in ww.items():
        ws.column_dimensions[col].width = w

# ---------------- READ FIRST ----------------
ws = wb.active; ws.title = 'READ FIRST'
widths(ws, {'A': 118})
rows = [
 ('Testahil · Independent Valuation Study — Educational Analysis · Platinum (XPT/USD) · Not investment advice', H),
 ('Combined 1 / 3 / 12-month edition · anchor 20 Jul 2026 close $1,608.37/oz', HD),
 ('•  This workbook is an educational valuation exercise and an expression of personal analytical opinion — not investment advice, not a recommendation, not directed at any reader\'s objectives or needs.', N),
 ('•  The preparer is not licensed by any securities or commodities regulator, provides no consultancy, manages no money, solicits no funds and accepts no clients.', N),
 ('•  A commodity has no income statement, balance sheet, cash-flow statement, per-share metrics or WACC, so the standard template\'s company sheets are a justified N/A here; their metal analogues are the Platinum Balance, Supply & Cost Curve and Demand Segments sheets.', N),
 ('•  The Assumptions sheet is the single input layer: blue = input; black = formula; green = cross-sheet link. Change an input and the linked sheets recalculate.', N),
 ('•  Monte Carlo percentiles are computed engine outputs (mc_v3 carry-anchored YZ-HAR-t: 50,000 paths, seed 42, ν = Gaussian, width_cal = 0.853 — a PROVISIONAL single-instrument self-fit, flagged) pasted as values and labelled as such; they are not spreadsheet formulas.', N),
 ('•  Step 0 verdict is PARITY under the scale-normalized carry-anchored gate — the engine ties, and does not beat, the carry-anchored random walk on platinum\'s history; metals remain the weakest calibration in the system (gold self-fit; silver borrows gold\'s).', N),
 ('•  All figures are model outputs presented as ranges and probability distributions; no single number is a forecast or a price target.', N),
 ('•  The preparer may hold, and may in the future take or dispose of, a position in the instrument discussed.', N),
]
for i, (t, f) in enumerate(rows, 1):
    put(ws, f'A{i}', t, f, wrap=True)
    ws.row_dimensions[i].height = 30 if i > 2 else 20

# ---------------- Assumptions ----------------
ws = sheet('Assumptions')
widths(ws, {'A': 40, 'B': 16, 'C': 64})
put(ws, 'A1', 'Assumptions — single input layer (blue = input)', H)
put(ws, 'A2', 'Input', HD, FILL_H); put(ws, 'B2', 'Value', HD, FILL_H); put(ws, 'C2', 'Source / note', HD, FILL_H)
rowsA = [
 ('Spot platinum (US$/oz)', spot, '20 Jul 2026 close, attached daily history (post data-quality gate)'),
 ('Gold spot (US$/oz)', RA['gold_spot'], 'Fri 17 Jul 2026 close — one-session mismatch vs the Mon 20 Jul anchor, stated; 20-Jul vendor prints ~4,000–4,030 → centre +0.3%'),
 ('Palladium spot (US$/oz)', RA['pd_spot'], 'spot 17 Jul 2026'),
 ('— Pt/Au ratio anchors —', None, 'attached 2011–26 series; post-2016 = modern regime'),
 ('2-year average ratio', RA['mean_2y'], 'recent regime — maps to ~spot'),
 ('5-year average ratio', RA['mean_5y'], 'the mean-reversion case'),
 ('Post-2016 average ratio', RA['mean_post2016'], 'full modern regime — bull anchor'),
 ('— Consensus targets —', None, 'two vintages, $500 apart'),
 ('UBS end-2026', 1700, '29 Jun 2026, post-crash (spot $1,618)'),
 ('UBS mid-2027', 1800, '29 Jun 2026'),
 ('LBMA 2026 survey average', 2222, '20 Jan 2026 — set at the record; stale-high'),
 ('Valterra LT planning low', 2300, 'Feb 2026 — greenfield incentive, excluded from base'),
 ('Valterra LT planning high', 2500, 'Feb 2026'),
 ('— Supply / demand (koz, WPIC Q1-26) —', None, 'WPIC Platinum Quarterly, 18 May 2026 (Metals Focus)'),
 ('2026f total supply', 7377, '+2% YoY'),
 ('2026f mine supply', 5551, 'South Africa 4,005 (~72%)'),
 ('2026f recycling', 1826, '+9% — the elastic leg'),
 ('2026f total demand', 7674, '−9% YoY'),
 ('2026f balance (deficit)', -297, '4th consecutive deficit'),
 ('Above-ground stocks end-26f (koz)', 1747, '≈ 11 weeks of demand cover — thinnest on record'),
 ('— Cost & carry —', None, None),
 ('AISC 2026f (US$/oz)', 1006, 'S&P Global, +7.7% YoY'),
 ('Real 10Y TIPS (%)', 2.31, '18-Jul-2026 close (Treasury est. via TipsWatch) — ~18-yr high; path 1.99 (12-May, FRED) → 2.07 (3-Jun, TE) → 2.31'),
 ('MC carry rf (%)', 3.63, 'Fed funds midpoint held; FOMC 16–17 Jun 2026, statement 17 Jun; q = 0'),
 ('— Fair-value zone (computed §1.5) —', None, 'weights 30/30/20/20 ratio/consensus/balance/cost — house judgment, Driver Ledger'),
 ('Zone low', Z['lo'], 'weighted bear'),
 ('Zone centre', Z['centre'], 'weighted base'),
 ('Zone high', Z['hi'], 'weighted bull'),
]
r = 3
for a, b, c in rowsA:
    put(ws, f'A{r}', a, NB if a.startswith('—') else N)
    if b is not None:
        put(ws, f'B{r}', b, BLUE, fmt='#,##0.00' if isinstance(b, float) and abs(b) < 20 else '#,##0')
    if c: put(ws, f'C{r}', c, SM, wrap=True)
    r += 1

# ---------------- Summary ----------------
ws = sheet('Summary')
widths(ws, {'A': 34, 'B': 16, 'C': 16, 'D': 16, 'E': 40})
put(ws, 'A1', 'Summary — every read at a glance (US$/oz)', H)
put(ws, 'A2', 'Read', HD, FILL_H); put(ws, 'B2', '1 month', HD, FILL_H)
put(ws, 'C2', '3 months', HD, FILL_H); put(ws, 'D2', '12 months', HD, FILL_H); put(ws, 'E2', 'Note', HD, FILL_H)
put(ws, 'A3', 'MC median (predicted)', N)
for col, ref in (('B', "'Monte Carlo'!D4"), ('C', "'Monte Carlo'!D5"), ('D', "'Monte Carlo'!D6")):
    put(ws, f'{col}3', f"={ref}", GRN, fmt='#,##0')
put(ws, 'A4', 'MC p5', N)
for col, ref in (('B', "'Monte Carlo'!B4"), ('C', "'Monte Carlo'!B5"), ('D', "'Monte Carlo'!B6")):
    put(ws, f'{col}4', f"={ref}", GRN, fmt='#,##0')
put(ws, 'A5', 'MC p95', N)
for col, ref in (('B', "'Monte Carlo'!F4"), ('C', "'Monte Carlo'!F5"), ('D', "'Monte Carlo'!F6")):
    put(ws, f'{col}5', f"={ref}", GRN, fmt='#,##0')
put(ws, 'A6', 'Spot', N); put(ws, 'B6', '=Assumptions!$B$3', GRN, fmt='#,##0.00')
put(ws, 'A7', 'Pt/Au ratio', N); put(ws, 'B7', '=Assumptions!$B$3/Assumptions!$B$4', None, fmt='0.000')
put(ws, 'A8', 'Pt/Pd ratio', N); put(ws, 'B8', '=Assumptions!$B$3/Assumptions!$B$5', None, fmt='0.00')
put(ws, 'A9', 'Fair-value zone (lo / centre / hi)', N)
put(ws, 'B9', '=Assumptions!$B$28', GRN, fmt='#,##0'); put(ws, 'C9', '=Assumptions!$B$29', GRN, fmt='#,##0'); put(ws, 'D9', '=Assumptions!$B$30', GRN, fmt='#,##0')
put(ws, 'A10', 'Verdict', N)
put(ws, 'B10', 'Fairly valued — spot ~1.6% below the weighted centre; two-sided cone; 12-mo drift = carry only', SM, wrap=True)
ws.merge_cells('B10:E10')

# ---------------- Fundamental Valuation ----------------
ws = sheet('Fundamental Valuation')
widths(ws, {'A': 36, 'B': 12, 'C': 16, 'D': 52})
put(ws, 'A1', 'Fundamental Valuation — anchor synthesis (fair value NOW, undated)', H)
for i, t in enumerate(['Anchor', 'Weight', 'Base (US$/oz)', 'Note'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
fv_rows = [
 ('Pt/Au ratio (5y mean × gold)', 0.30, "='Fair-Value Anchors'!B5", 'primary relative lens; still cheap vs gold'),
 ('Analyst consensus (fresh vintage)', 0.30, "='Fair-Value Anchors'!B8", 'UBS 1,700–1,800; stale 2,222 excluded from base'),
 ('Supply/demand balance', 0.20, 1500, 'floor-builder: 4th deficit, 11 weeks cover'),
 ('Cost curve (AISC + margin)', 0.20, 1300, 'AISC 1,006; incentive 2,300–2,500'),
 ('Real-rate carry', 0.00, '— (tilt)', 'timing switch, not a level'),
]
r = 3
for a, w, b, note in fv_rows:
    put(ws, f'A{r}', a, N); put(ws, f'B{r}', w, BLUE, fmt='0%')
    if isinstance(b, str) and b.startswith('='):
        put(ws, f'C{r}', b, GRN, fmt='#,##0')
    elif isinstance(b, (int, float)):
        put(ws, f'C{r}', b, BLUE, fmt='#,##0')
    else:
        put(ws, f'C{r}', b, N)
    put(ws, f'D{r}', note, SM, wrap=True); r += 1
put(ws, 'A8', 'Weighted fair-value centre', NB, FILL_C)
put(ws, 'B8', '=SUM(B3:B7)', NB, FILL_C, fmt='0%')
put(ws, 'C8', '=SUMPRODUCT(B3:B6,C3:C6)', NB, FILL_C, fmt='#,##0')
put(ws, 'D8', 'ties to Assumptions zone centre (rounding)', SM, FILL_C)

# ---------------- Fair-Value Anchors ----------------
ws = sheet('Fair-Value Anchors')
widths(ws, {'A': 38, 'B': 16, 'C': 22, 'D': 40})
put(ws, 'A1', 'Fair-Value Anchors — five lenses (US$/oz)', H)
for i, t in enumerate(['Anchor / scenario', 'Implied Pt', 'Basis', 'Note'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
anch_rows = [
 ('Pt/Au @ 2-year mean', '=Assumptions!$B$4*Assumptions!$B$7', 'gold × ratio', 'recent regime ≈ spot'),
 ('Pt/Au @ today', '=Assumptions!$B$3', 'spot', '12th pctile of post-2016'),
 ('Pt/Au @ 5-year mean', '=Assumptions!$B$4*Assumptions!$B$8', 'gold × ratio', 'the mean-reversion case — BASE'),
 ('Pt/Au @ post-2016 mean', '=Assumptions!$B$4*Assumptions!$B$9', 'gold × ratio', 'bull anchor'),
 ('Consensus — UBS end-26', '=Assumptions!$B$11', '2026 target', 'fresh, post-crash'),
 ('Consensus — blend (BASE)', '=(Assumptions!$B$11+Assumptions!$B$12)/2', 'fresh vintage', 'base consensus anchor 1,750'),
 ('Consensus — LBMA survey', '=Assumptions!$B$13', '2026 survey', 'stale-high (set at record)'),
 ('Cost floor — AISC', '=Assumptions!$B$24', 'marginal cost', 'deep-bear bound; tested 2018–24'),
 ('Incentive price — Valterra', '=(Assumptions!$B$14+Assumptions!$B$15)/2', 'greenfield sanction', 'long-run ceiling logic'),
 ('Weighted fair-value zone', '=Assumptions!$B$28&" – "&Assumptions!$B$30', 'synthesis', 'centre = Assumptions!B29'),
]
r = 3
for a, f, b, note in anch_rows:
    put(ws, f'A{r}', a, N); put(ws, f'B{r}', f, GRN, fmt='#,##0')
    put(ws, f'C{r}', b, N); put(ws, f'D{r}', note, SM, wrap=True); r += 1

# ---------------- Pt-Gold Ratio ----------------
ws = sheet('Pt-Gold Ratio')
widths(ws, {'A': 36, 'B': 14, 'C': 48})
put(ws, 'A1', "Pt/Au Ratio — platinum's primary relative lens", H)
for i, t in enumerate(['Item', 'Value', 'Note'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
ratio_rows = [
 ('Current ratio (Pt ÷ Au)', '=Assumptions!$B$3/Assumptions!$B$4', 'live from Assumptions'),
 ('2-year average', RA['mean_2y'], 'recent regime'),
 ('5-year average', RA['mean_5y'], 'mean-reversion anchor'),
 ('Post-2016 average', RA['mean_post2016'], 'modern regime'),
 ('Full-history average (2011–26)', RA['mean_full'], 'a fossil — pre-2015 premium era'),
 ('April 2025 trough', RA['min_full'], 'platinum cheapest ever vs gold'),
 ('2011 peak', RA['max_full'], 'the old premium world'),
 ('Percentile of today (post-2016)', RA['pctile_now_post2016'], 'still historically cheap'),
 ('Pt/Pd cross', '=Assumptions!$B$3/Assumptions!$B$5', 'first premium since 2017 — the crux driver'),
 ('Implied Pt @ current gold, 5y mean', '=Assumptions!$B$4*Assumptions!$B$8', '+14% vs spot'),
 ('Implied Pt if gold $4,742 @ 5y mean', '=4742*Assumptions!$B$8', 'LBMA gold consensus × ratio held'),
]
r = 3
for a, v, note in ratio_rows:
    put(ws, f'A{r}', a, N)
    if isinstance(v, str):
        put(ws, f'B{r}', v, GRN, fmt='0.000' if 'Pd' not in a and 'Implied' not in a else ('0.00' if 'Pd' in a else '#,##0'))
    else:
        put(ws, f'B{r}', v, BLUE, fmt='0.000' if v < 3 else '#,##0')
    put(ws, f'C{r}', note, SM, wrap=True); r += 1

# ---------------- Platinum Balance ----------------
ws = sheet('Platinum Balance')
widths(ws, {'A': 40, 'B': 11, 'C': 11, 'D': 11, 'E': 11})
put(ws, 'A1', "Platinum Balance — the commodity's financial statements (koz)", H)
put(ws, 'A2', 'WPIC Platinum Quarterly Q1 2026 (18 May 2026; research by Metals Focus)', SM)
hdr = ['Line (koz)', '2023', '2024', '2025', '2026f']
for i, t in enumerate(hdr):
    put(ws, chr(65 + i) + '3', t, HD, FILL_H)
bal = [
 ('South Africa', 3957, 4133, 3957, 4005),
 ('Zimbabwe', 507, 512, 516, 508),
 ('Russia', 674, 677, 677, 646),
 ('North America', 278, 265, 212, 201),
 ('Other', 190, 191, 196, 192),
 ('Rounding / unallocated (ties to WPIC print)', 14, 9, 3, -1),
 ('Total mine supply', '=SUM(B4:B9)', '=SUM(C4:C9)', '=SUM(D4:D9)', '=SUM(E4:E9)'),
 ('Recycling', 1515, 1536, 1679, 1826),
 ('TOTAL SUPPLY', '=B10+B11', '=C10+C11', '=D10+D11', '=E10+E11'),
 ('Automotive', 3204, 3108, 3031, 2959),
 ('Jewellery', 1849, 2008, 2214, 1958),
 ('Industrial', 2491, 2526, 2049, 2238),
 ('  of which hydrogen (memo)', 22, 40, 65, 69),
 ('Investment', 388, 713, 1136, 519),
 ('Rounding / unallocated (ties to WPIC print)', 1, 0, 1, 0),
 ('TOTAL DEMAND', '=B13+B14+B15+B17+B18', '=C13+C14+C15+C17+C18', '=D13+D14+D15+D17+D18', '=E13+E14+E15+E17+E18'),
 ('BALANCE (supply − demand)', '=B12-B19', '=C12-C19', '=D12-D19', '=E12-E19'),
 ('Above-ground stocks (year-end)', 4268, 3235, 2044, 1747),
 ('Weeks of demand cover (WPIC-stated)', 34, 20, 12, 11),
]
r = 4
for row in bal:
    put(ws, f'A{r}', row[0], NB if str(row[0]).isupper() or 'Total' in str(row[0]) or 'BALANCE' in str(row[0]) else N)
    for j, v in enumerate(row[1:], 1):
        col = chr(65 + j)
        if isinstance(v, str):
            put(ws, f'{col}{r}', v, NB if r in (12, 19, 20) else N, FILL_C if r in (12, 19, 20) else None, fmt='#,##0')
        else:
            put(ws, f'{col}{r}', v, BLUE, fmt='#,##0')
    r += 1
put(ws, 'A24', 'Four consecutive deficits ≈ 3.3 Moz drawn from stocks; cover 34 → 11 weeks (WPIC-stated; simple AGS ÷ same-year demand × 52 gives ~28/20/13/12 — definitional difference stated, dual-framing rule). Computed balances −798/−1,032 (2023/24) vs WPIC-printed −799/−1,033: 1 koz of print rounding, stated. Hydrogen is a memo line inside Industrial (not double-counted).', SM, wrap=True)
ws.merge_cells('A24:E24')

# ---------------- Supply & Cost Curve ----------------
ws = sheet('Supply & Cost Curve')
widths(ws, {'A': 40, 'B': 16, 'C': 52})
put(ws, 'A1', 'Supply structure & the cost curve', H)
for i, t in enumerate(['Item', 'Value', 'Note'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
sc_rows = [
 ('South Africa share of mine supply', '=4005/5551', '72% — the most concentrated major metal'),
 ('AISC 2026f (US$/oz)', '=Assumptions!$B$24', 'S&P Global, +7.7% YoY'),
 ('Spot ÷ AISC', '=Assumptions!$B$3/Assumptions!$B$24', 'first sustained margin in a decade (2018–24 ≈ 1×)'),
 ('Incentive price (greenfield)', '=(Assumptions!$B$14+Assumptions!$B$15)/2', 'Valterra planning range 2,300–2,500'),
 ('SA electricity cost since 2008', '>+900%', 'structural cost inflation (Reuters/MarketScreener, Feb-2026)'),
 ('Projects surviving from the 2005–10 boom', '2 of 20', 'why supply cannot respond for years'),
 ('Valterra 2026 guidance (PGM M&C)', '3.0–3.4 Moz', 'unchanged at Q2-26; Amandelbult +116% post-flood'),
 ('Recycling 2026f (koz)', '=Assumptions!$B$19', 'the only elastic supply leg'),
]
r = 3
for a, v, note in sc_rows:
    put(ws, f'A{r}', a, N)
    if isinstance(v, str) and v.startswith('='):
        put(ws, f'B{r}', v, GRN, fmt='0%' if 'share' in a else ('0.00x' if '÷' in a else '#,##0'))
    else:
        put(ws, f'B{r}', v, BLUE if not isinstance(v, str) else N, fmt='#,##0')
    put(ws, f'C{r}', note, SM, wrap=True); r += 1

# ---------------- Demand Segments ----------------
ws = sheet('Demand Segments')
widths(ws, {'A': 26, 'B': 12, 'C': 14, 'D': 10, 'E': 52})
put(ws, 'A1', "Demand Segments — platinum's four legs (2026f)", H)
for i, t in enumerate(['Segment', '2026f (koz)', 'Share', 'YoY', 'Driver / swing'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
seg = [
 ('Automotive', "='Platinum Balance'!E13", "=B3/'Platinum Balance'!E19", '−2%', 'ICE+hybrid volumes; reverse Pt-for-Pd substitution — THE crux'),
 ('Jewellery', "='Platinum Balance'!E14", "=B4/'Platinum Balance'!E19", '−12%', 'China fabrication −42% Q1-26; VAT-rebate removal 1-Nov-25'),
 ('Industrial', "='Platinum Balance'!E15", "=B5/'Platinum Balance'!E19", '+9%', 'chem 612 · glass 377 · medical 332 · H2 69 — the quiet grower'),
 ('Investment', "='Platinum Balance'!E17", "=B6/'Platinum Balance'!E19", '−54%', 'bars +27% (China 13t vs <1t 2019) vs ETF −100, exchange −100'),
]
r = 3
for a, f1, f2, yoy, note in seg:
    put(ws, f'A{r}', a, N); put(ws, f'B{r}', f1, GRN, fmt='#,##0')
    put(ws, f'C{r}', f2, N, fmt='0%'); put(ws, f'D{r}', yoy, BLUE)
    put(ws, f'E{r}', note, SM, wrap=True); r += 1

# ---------------- Technical ----------------
ws = sheet('Technical')
widths(ws, {'A': 30, 'B': 18, 'C': 46})
put(ws, 'A1', 'Technical indicators', H)
put(ws, 'A2', f"Computed from XPT/USD daily history (04 Jan 2011 – 20 Jul 2026, {D['meta']['rows']:,} sessions post-clean)", SM)
for i, t in enumerate(['Indicator', 'Reading', 'Signal'], 0):
    put(ws, chr(65 + i) + '3', t, HD, FILL_H)
tech_rows = [
 ('Close', spot, '—'),
 ('SMA 20', T['sma20'], f"price {T['vs20']*100:+.1f}% vs SMA20 — reclaimed"),
 ('SMA 50', T['sma50'], f"price {T['vs50']*100:+.1f}% vs SMA50 — falling"),
 ('SMA 200', T['sma200'], f"price {T['vs200']*100:+.1f}% vs SMA200"),
 ('50/200 stack', 'death-cross', 'both falling — post-halving structure'),
 ('RSI(14)', T['rsi14'], 'neutral, off the June oversold'),
 ('MACD line / signal', f"{T['macd']:.0f} / {T['macd_sig']:.0f}", 'negative'),
 ('MACD histogram', T['macd_hist'], 'positive but fading (was +17)'),
 ('Avg daily range (30d)', PR['avg_true_range_pct_30d'], 'high-volatility tape'),
 ('52-week high (close)', PR['hi52'], '23 Jan 2026 record'),
 ('52-week low (close)', PR['lo52'], '31 Jul 2025'),
 ('Support', '1,540 → 1,500 → 1,348', 'June–July floor → round → Feb-2021 shelf'),
 ('Resistance', '1,758 → 1,990', 'SMA50 → May-2026 recovery high'),
]
r = 4
for a, v, note in tech_rows:
    put(ws, f'A{r}', a, N)
    put(ws, f'B{r}', v, N, fmt='#,##0.00' if isinstance(v, float) and v < 5 else ('0.0%' if a.startswith('Avg') else '#,##0.0'))
    put(ws, f'C{r}', note, SM, wrap=True); r += 1

# ---------------- Monte Carlo ----------------
ws = sheet('Monte Carlo')
widths(ws, {'A': 22, 'B': 11, 'C': 11, 'D': 11, 'E': 11, 'F': 11, 'G': 11})
put(ws, 'A1', 'Monte Carlo — mc_v3 carry-anchored YZ-HAR-t, 50k paths, seed 42 (engine outputs pasted as values)', H)
put(ws, 'A2', 'Width: gap-aware Yang-Zhang + log-HAR(1/5/22), bias-corrected, 0.8 shrink; drift = ln(1+3.63%)·h/252, q=0; ν=Gaussian, width_cal=0.853 (PROVISIONAL self-fit); no signal, no factor drift.', SM, wrap=True)
ws.merge_cells('A2:G2'); ws.row_dimensions[2].height = 26
for i, t in enumerate(['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95', 'mean'], 0):
    put(ws, chr(65 + i) + '3', t, HD, FILL_H)
for r, key, lab in ((4, 't21', 'T+21 (1 month)'), (5, 't63', 'T+63 (3 months)'), (6, 't252', 'T+252 (12 months)')):
    put(ws, f'A{r}', lab, N)
    for col, k in (('B', 'p5'), ('C', 'p25'), ('D', 'p50'), ('E', 'p75'), ('F', 'p95'), ('G', 'mean')):
        put(ws, f'{col}{r}', MC[key][k], N, fmt='#,##0.00')
put(ws, 'A8', 'Ledger-convention cohorts (graded rows)', HD)
for i, t in enumerate(['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95', 'ann. fwd vol'], 0):
    put(ws, chr(65 + i) + '9', t, HD, FILL_H)
for r, key, lab in ((10, 't20', 'T+20 (grade 17 Aug 26)'), (11, 't60', 'T+60 (grade 12 Oct 26)')):
    put(ws, f'A{r}', lab, N)
    for col, k in (('B', 'p5'), ('C', 'p25'), ('D', 'p50'), ('E', 'p75'), ('F', 'p95')):
        put(ws, f'{col}{r}', MC[key][k], N, fmt='#,##0.00')
    put(ws, f'G{r}', MC[key]['ann_fwd_vol'], N, fmt='0.0%')
put(ws, 'A13', 'Zone vs spot', HD, FILL_H); put(ws, 'B13', 'T+63', HD, FILL_H); put(ws, 'C13', 'T+252', HD, FILL_H)
zz = [('below −20%', 'below_m20'), ('−20…−10%', 'm10_m20'), ('−10…−5%', 'm5_m10'), ('±5%', 'pm5'),
      ('+5…+10%', 'p5_p10'), ('+10…+20%', 'p10_p20'), ('above +20%', 'above_p20')]
r = 14
for lab, k in zz:
    put(ws, f'A{r}', lab, N)
    put(ws, f'B{r}', MC['t63']['zones'][k], N, fmt='0%'); put(ws, f'C{r}', MC['t252']['zones'][k], N, fmt='0%')
    r += 1
put(ws, 'A22', 'Touch (level hit at least once)', HD, FILL_H)
put(ws, 'B22', 'by T+20', HD, FILL_H); put(ws, 'C22', 'by T+60', HD, FILL_H); put(ws, 'D22', 'by T+252', HD, FILL_H)
tl = [('$1,750', '1750'), ('$1,990 (May high)', '1990.5'), ('$2,222 (stale consensus)', '2222'),
      ('$2,925 (record)', '2925'), ('$1,540 (floor)', '1539.6'), ('$1,348 (2021 shelf)', '1348.2'), ('$1,006 (AISC)', '1006')]
r = 23
for lab, k in tl:
    put(ws, f'A{r}', lab, N)
    for col, key in (('B', 't20'), ('C', 't60'), ('D', 't252')):
        put(ws, f'{col}{r}', MC[key]['touch_abs'][k], N, fmt='0%')
    r += 1
put(ws, 'A31', f"Sensitivity (stated, not adopted): under the gold+silver-trained LONO shape (ν=20, width 1.035) the T+63 90% band widens to {MC['t63_lono']['p5']:,.0f}–{MC['t63_lono']['p95']:,.0f}.", SM, wrap=True)
ws.merge_cells('A31:G31')

# ---------------- Sensitivity ----------------
ws = sheet('Sensitivity')
widths(ws, {'A': 16, 'B': 11, 'C': 11, 'D': 11, 'E': 11})
put(ws, 'A1', 'Sensitivity — platinum fair value = gold × Pt/Au ratio (US$/oz)', H)
put(ws, 'A2', 'Rows: Pt/Au ratio. Columns: gold price.', SM)
G = D['ratio_grid']
put(ws, 'A3', 'ratio \\ gold', HD, FILL_H)
for j, g in enumerate(G['gold']):
    put(ws, chr(66 + j) + '3', g, BLUE, FILL_H, fmt='#,##0')
r = 4
for i, rt in enumerate(G['ratio']):
    put(ws, f'A{r}', rt, BLUE, fmt='0.000')
    for j in range(len(G['gold'])):
        col = chr(66 + j)
        put(ws, f'{col}{r}', f"={col}$3*$A{r}", N, fmt='#,##0')
    r += 1
put(ws, 'A10', f"Today: gold ~${RA['gold_spot']:,.0f}, ratio {RA['now']:.3f}× → ≈ spot. Bull = joint move (gold up + ratio re-rates); the two dimensions are NOT independent — reverse substitution pins the ratio (§1.7 of the study).", SM, wrap=True)
ws.merge_cells('A10:E10')

# ---------------- Step 0 Calibration ----------------
ws = sheet('Step 0 Calibration')
widths(ws, {'A': 44, 'B': 52})
put(ws, 'A1', 'Step 0 — calibration gate (current house basis)', H)
put(ws, 'A2', 'Walk-forward h=60, non-overlapping, scale-normalized CRPS vs a CARRY-ANCHORED lognormal RW; robust verdict across bootstrap blocks {2,3,4}; production chain (reproduction check passed against the live gold registry).', SM, wrap=True)
ws.merge_cells('A2:B2'); ws.row_dimensions[2].height = 26
sf = S0['scores']['self_fit']; lg = S0['scores']['lono_gold_silver']; bw = S0['scores']['borrowed_live_metals']; dg = S0['diag_self']
cal_rows = [
 ('Windows (full history / last-5y)', '62 (2012–2026) / 20'),
 ('ADOPTED config (provisional self-fit)', 'ν = Gaussian (sentinel 250), width_cal = 0.853 (MLE scale 0.790, clip floor 0.85 active)'),
 ('Self-fit skill & CI (block 2)', f"{sf['skill']:+.4f}  [{sf['ci_block2'][0]:+.3f}, {sf['ci_block2'][1]:+.3f}] — {sf['verdict']}"),
 ('De-circularized LONO (gold+silver-trained)', f"{lg['skill']:+.4f}  [{lg['ci_block2'][0]:+.3f}, {lg['ci_block2'][1]:+.3f}] — {lg['verdict']}"),
 ('Borrowed live METALS (Gaussian/1.0)', f"{bw['skill']:+.4f}  [{bw['ci_block2'][0]:+.3f}, {bw['ci_block2'][1]:+.3f}] — {bw['verdict']}"),
 ('Coverage 50/80/90 (self-fit)', f"{dg['cov50']*100:.0f}% / {dg['cov80']*100:.0f}% / {dg['cov90']*100:.1f}% vs 50/80/90 — mildly over-covered"),
 ('PIT mean', f"{dg['pit_mean']:.3f} — centred; no drift blemish under the carry anchor"),
 ('Reproduction check (gold, live registry)', 'EXACT: 67 windows, +0.0035, CI[−0.005,+0.013], PARITY'),
 ('Verdict', 'PARITY — the engine ties the carry-anchored RW; platinum does NOT arrive failing. PROVISIONAL single-name fit, flagged circular; pooled 3-metal fit (ν≈20, 0.965, 148 windows) is the likely future config.'),
]
r = 3
for a, b in cal_rows:
    put(ws, f'A{r}', a, NB if a == 'Verdict' else N, FILL_C if a == 'Verdict' else None)
    put(ws, f'B{r}', b, N, FILL_C if a == 'Verdict' else None, wrap=True)
    ws.row_dimensions[r].height = 26 if len(b) > 60 else 14
    r += 1

# ---------------- Ledger cohorts — merged into 'Monte Carlo' (16-sheet standard) ----------------
ws = wb['Monte Carlo']
c1, c2 = D['cohorts']
put(ws, 'A33', 'Ledger cohort fields (anchored at publication; graded later; append-only)', HD, FILL_H)
ws.merge_cells('A33:G33')
led = [
 ('instrument / cycle_no', f"{c1['instrument']} / 1", ''),
 ('anchor_date / anchor_price', f"{c1['anchor_date']} / {c1['anchor_price']:,.2f}", ''),
 ('grade_date T+20 / T+60 (projected)', f"{c1['grade_date']} / {c2['grade_date']}", 'Mon–Fri weekmask; actual grading counts real trading rows'),
 ('touch T+20  +5/+10/−5/−10%', f"{c1['touch']['+5%']:.0%} / {c1['touch']['+10%']:.0%} / {c1['touch']['-5%']:.0%} / {c1['touch']['-10%']:.0%}", ''),
 ('touch T+60  +5/+10/−5/−10%', f"{c2['touch']['+5%']:.0%} / {c2['touch']['+10%']:.0%} / {c2['touch']['-5%']:.0%} / {c2['touch']['-10%']:.0%}", ''),
 ('anchor_vol (ann., HAR) T+20 / T+60', f"{c1['anchor_vol']:.1%} / {c2['anchor_vol']:.1%}", 'forward, width_cal applied'),
 ('status', 'NOT anchored to the live site — publication is a separate, explicitly-requested step (house rule).', ''),
]
r = 34
for a, b, c in led:
    put(ws, f'A{r}', a, N); put(ws, f'B{r}', b, N, wrap=True)
    ws.merge_cells(f'B{r}:E{r}')
    if c: put(ws, f'F{r}', c, SM, wrap=True); ws.merge_cells(f'F{r}:G{r}')
    r += 1

# ---------------- Peer & Sector ----------------
ws = sheet('Peer & Sector')
widths(ws, {'A': 24, 'B': 44, 'C': 44})
put(ws, 'A1', 'Peer & Sector — the investable platinum complex', H)
for i, t in enumerate(['Cohort', 'Names', 'Character'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
peers = [
 ('SA primary producers', 'Valterra (ex-Amplats), Impala, Sibanye-Stillwater, Northam', 'levered to spot & ZAR; AISC ~$1,000; payouts over projects'),
 ('Zimbabwe / other', 'Zimplats, Tharisa', 'expansion optionality, political discount'),
 ('Russian by-product', 'Norilsk Nickel', '~12% of supply; sanctions-shadowed'),
 ('Physical ETFs', 'PPLT (~$1.8bn AUM), PLTM', 'direct bullion; 2026f ETF line −100 koz'),
 ('New venue', 'GFEX platinum futures (Guangzhou)', "China's first PGM derivatives; published stockpiles"),
 ('Platinum (this study)', '—', 'the metal itself — no leverage, no balance sheet, no counterparty'),
]
r = 3
for a, b, c in peers:
    put(ws, f'A{r}', a, N); put(ws, f'B{r}', b, N, wrap=True); put(ws, f'C{r}', c, SM, wrap=True); r += 1

# ---------------- Sweep Register ----------------
ws = sheet('Sweep Register')
widths(ws, {'A': 8, 'B': 17, 'C': 25, 'D': 7, 'E': 60, 'F': 13})
put(ws, 'A1', 'Step 2A Information Sweep — four metal rings, validated (engine/research_sweep.py: 0 errors)', H)
for i, t in enumerate(['ID', 'Ring', 'Category', 'Class', 'Headline', 'Date'], 0):
    put(ws, chr(65 + i) + '2', t, HD, FILL_H)
SW = json.load(open('sweep_register_xpt.json'))
r = 3
for f in SW['findings']:
    put(ws, f'A{r}', f['fid'], N); put(ws, f'B{r}', f['ring'], N)
    put(ws, f'C{r}', f['category'], N, wrap=True); put(ws, f'D{r}', f['klass'][0] if f['klass'] != 'NEGATIVE_SEARCH' else 'NEG', N)
    put(ws, f'E{r}', f['headline'], SM, wrap=True); put(ws, f'F{r}', f['date'], N)
    ws.row_dimensions[r].height = 24
    r += 1
put(ws, f'A{r}', 'Driver gate:', HD)
r += 1
for d_ in SW['drivers']:
    put(ws, f'A{r}', d_['mode'], N); put(ws, f'B{r}', d_['driver'], SM, wrap=True)
    ws.merge_cells(f'B{r}:E{r}'); put(ws, f'F{r}', ','.join(d_['refs']), SM)
    ws.row_dimensions[r].height = 22
    r += 1

# ---------------- Sources ----------------
ws = sheet('Sources')
widths(ws, {'A': 118})
put(ws, 'A1', 'Sources & about', H)
src = [
 '•  WPIC Platinum Quarterly Q1 2026 (18 May 2026; research by Metals Focus) — supply/demand balance, above-ground stocks, lease-rate & China commentary. platinuminvestment.com',
 '•  WPIC/Mining Weekly (17 Jul 2026) — China 15th Five-Year Plan (WPIC characterization; the official strategic-minerals catalogue is undisclosed and reconstructed lists carry no PGMs — Swedish National China Centre, 2026), GFEX, China bar market <1t (2019) → ~13t (2025).',
 '•  UBS platinum note via Yahoo Finance (29 Jun 2026) — $1,700 Sep/Dec-26, $1,800 Mar/Jun-27; surplus risk; falling lease rates.',
 '•  LBMA 2026 Precious Metals Forecast Survey via Kitco (20 Jan 2026) — platinum average $2,222 (stale-high, set at the record); gold $4,742.',
 '•  Valterra Platinum Q2-2026 production report via Reuters/TradingView (17 Jul 2026) — output, guidance, realized prices ($1,966 Q2 avg).',
 '•  Reuters via MarketScreener (11 Feb 2026) — payouts over projects; S&P Global AISC 2026f $1,006/oz (+7.7%); $2,300–2,500 planning range; SA electricity +900% since 2008.',
 '•  WPIC Perspectives (24 Jan 2024) — Pt-for-Pd substitution ~700 koz embedded; ~15% of models/yr; ~7-year platform lock-in.',
 '•  CFTC COT via metalcharts.org (14 Jul 2026) — managed money net long +8.3k (Pt), −6.2k (Pd).',
 '•  10Y TIPS real yield 2.31% at the 18-Jul-2026 close (US Treasury estimate via TipsWatch, 19-Jul-2026); rising path 1.99% (FRED DFII10, 12-May) → 2.07% (TradingEconomics, 3-Jun) → 2.31% — an ~18-year high. Fed funds midpoint 3.63%; FOMC 16–17 Jun 2026, statement 17 Jun (federalreserve.gov).',
 '•  Spot cross-section Fri 17 Jul 2026 (Fortune/quote pages): gold $3,972, palladium $1,239, silver $55 — one-session mismatch vs the Mon 20 Jul platinum anchor, stated. 30-Jan-2026 crash catalyst: the surprise Warsh Fed-chair nomination (Fortune/NBC, 31-Jan-2026 — gold −11.4% settle basis, silver −31.4%); the Strait of Hormuz war began 28-Feb-2026 (Brent to $126) — it POST-dates the crash and supported Q1 prices (Wikipedia timeline).',
 '•  Attached XPT/USD daily OHLC history (Investing.com export), 04 Jan 2011 – 20 Jul 2026; Step 0.0 data-quality gate applied (9 placeholder rows dropped; no corporate-action repairs; 260.0 rows/yr = metals Mon–Fri calendar).',
 '•  Monte Carlo: engine/mc_v3.py carry-anchored YZ-HAR-t; 50,000 paths, seed 42; ν=Gaussian (sentinel), width_cal=0.853 — PROVISIONAL single-instrument self-fit, flagged; Step 0 PARITY.',
 '•  Educational analysis under the Testahil Standing Research Protocol; not investment advice; no ratings, no price targets — fair-value ranges and distributions only.',
]
for i, s_ in enumerate(src, 2):
    put(ws, f'A{i}', s_, N if i < 4 else SM, wrap=True)
    ws.row_dimensions[i].height = 26

# order sheets: READ FIRST first (already), then desired order
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'Fair-Value Anchors',
         'Pt-Gold Ratio', 'Platinum Balance', 'Supply & Cost Curve', 'Demand Segments', 'Technical',
         'Monte Carlo', 'Sensitivity', 'Step 0 Calibration', 'Peer & Sector',
         'Sweep Register', 'Sources']
wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
wb.save('XPTUSD_Valuation_Model_20072026_public.xlsx')
print('xlsx saved:', wb.sheetnames, len(wb.sheetnames), 'sheets')
