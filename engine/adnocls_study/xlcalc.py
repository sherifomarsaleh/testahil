"""A small, explicit spreadsheet evaluator.

The delivered model is verified by evaluating it here rather than by handing it back to the
application that wrote it: arithmetic, SUM/MIN/MAX/MEDIAN/AVERAGE over ranges or over
comma-separated scalars, absolute and relative cell references, and cross-sheet references.
Anything the evaluator does not understand raises rather than being silently skipped.

This started as a workaround — LibreOffice could not load any document here, because only
libreoffice-core was installed and the Writer/Calc import filters were absent. That is fixed
(see engine/make_pdf.py), but the evaluator is KEPT deliberately: an independent
reimplementation that has to agree with the model cell-for-cell is a stronger check than
asking a spreadsheet engine to confirm its own arithmetic, and it is what carries the
expected-value gate in recalc.py.

Used by recalc.py — does the workbook reproduce the model? — and by driver_test.py — does
changing a driver on the Assumptions sheet actually reprice the workbook?

TEXT IS NEVER A NUMBER HERE. An earlier version of this evaluator quietly read the string
'-' as zero and treated any other text operand as if it were nothing at all. That made it
MORE PERMISSIVE THAN THE APPLICATION THE READER OPENS THE FILE IN: a workbook that raised
#VALUE! in 689 cells across twelve sheets in LibreOffice reconciled here cell-for-cell, and
this evaluator reported it clean. A verifier that accepts what the real spreadsheet rejects
is not a verifier. So any text reaching an arithmetic operator — including a display dash
in a cell some formula turns out to read — is a hard failure, raised as TextOperand and
reported like any other unresolvable cell. The two places a spreadsheet genuinely does
ignore text are kept, because they are what Excel and LibreOffice actually do: a BLANK cell
counts as zero in arithmetic, and text inside a SUM/AVERAGE/MIN/MAX RANGE is skipped rather
than propagated. Text passed to one of those functions as a direct scalar argument is not
ignored — it raises, exactly as the applications do.
"""
import datetime
import re
from openpyxl.utils import range_boundaries, get_column_letter

EPOCH = datetime.date(1899, 12, 30)          # the spreadsheet's own day zero


class TextOperand(ValueError):
    """Text has reached an arithmetic operator, where a real spreadsheet raises #VALUE!."""


def as_number(v):
    """A cell value as the number a spreadsheet would use, or None if it is not one.

    A DATE IS A NUMBER. Excel and LibreOffice store a date as the count of days since
    30 December 1899 and subtract one from another to get days; the reader sees a date only
    because of the cell's format. openpyxl hands such a cell back as a datetime, so it is
    converted here rather than rejected — rejecting it would make this evaluator STRICTER
    than the applications, which is its own kind of wrong answer.
    """
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, datetime.datetime):
        return (v.date() - EPOCH).days + (v - datetime.datetime.combine(
            v.date(), datetime.time())).total_seconds() / 86400.0
    if isinstance(v, datetime.date):
        return float((v - EPOCH).days)
    return None

FUNC = re.compile(r'\b(SUM|MIN|MAX|MEDIAN|AVERAGE)\(([^()]*)\)')
# An UNQUOTED sheet name may not contain a hyphen or a space: every sheet in this workbook
# whose name does is written quoted. Allowing them made "C34-Assumptions!$C$45" parse as a
# reference to a sheet called "C34-Assumptions", which silently swallowed the subtraction.
SHEETREF = re.compile(r"(?<![A-Za-z0-9_$!.])(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))"
                      r"!(\$?[A-Z]{1,3}\$?\d+)")
CELLREF = re.compile(r'(?<![A-Z0-9_!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![\d(])')
RANGE = re.compile(r"^(?:(?:'[^']+'|[A-Za-z][A-Za-z0-9_]*)!)?\$?[A-Z]{1,3}\$?\d+"
                   r":\$?[A-Z]{1,3}\$?\d+$")


class Book:
    """Evaluates a workbook's formulas, with an optional override layer over input cells."""

    def __init__(self, wb, overrides=None):
        self.wb = wb
        self.overrides = {(s, c): v for (s, c), v in (overrides or {}).items()}
        self.cache = {}
        self.stack = []

    def cell_value(self, sheet, coord):
        key = (sheet, coord)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise ValueError(f'circular reference at {sheet}!{coord}')
        v = self.wb[sheet][coord].value
        if isinstance(v, str) and v.startswith('='):
            self.stack.append(key)
            try:
                v = self.evaluate(v[1:], sheet)
            finally:
                self.stack.pop()
        elif v is None:
            v = 0.0          # a BLANK cell is zero in arithmetic, in Excel as here
        # a text cell is returned AS TEXT: whether that is fatal depends on what reads it,
        # and the decision is made at the point of use, never by coercing it here
        self.cache[key] = v
        return v

    def numeric(self, sheet, coord, where):
        """A cell's value where a number is required. Text raises, as it does in Excel."""
        v = self.cell_value(sheet, coord)
        n = as_number(v)
        if n is None:
            kind = 'the text' if isinstance(v, str) else ''
            raise TextOperand(
                f'{sheet}!{coord} holds {kind} {v!r}, which {where} reads as a number. '
                f'Excel and LibreOffice raise #VALUE! here and so does this evaluator: no '
                f'text may sit anywhere in an arithmetic chain')
        return n

    def range_values(self, sheet, rng):
        """Text inside a SUM/AVERAGE range is SKIPPED — that is what Excel does — so this
        is the one place text is tolerated, and only because tolerating it is correct."""
        c1, r1, c2, r2 = range_boundaries(rng)
        out = []
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                out.append(as_number(self.cell_value(sheet, f'{get_column_letter(cc)}{rr}')))
        return [x for x in out if x is not None]

    def arg_values(self, arg, sheet):
        """A function argument is either one range, or a comma-separated list of scalars."""
        arg = arg.strip()
        if RANGE.match(arg):
            tgt, rng = sheet, arg
            sm = re.match(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))!(.+)", arg)
            if sm:
                tgt = sm.group(1) or sm.group(2); rng = sm.group(3)
            return self.range_values(tgt, rng.replace('$', ''))
        # a scalar argument is NOT a range: text here is an error, not something to skip
        return [float(self.evaluate(part, sheet)) for part in arg.split(',') if part.strip()]

    def evaluate(self, expr, sheet):
        e = expr
        # functions over ranges or scalar lists (possibly cross-sheet)
        while True:
            m = FUNC.search(e)
            if not m:
                break
            fn = m.group(1)
            vals = self.arg_values(m.group(2), sheet)
            if not vals:
                val = 0.0
            elif fn == 'SUM':
                val = sum(vals)
            elif fn == 'MIN':
                val = min(vals)
            elif fn == 'MAX':
                val = max(vals)
            elif fn == 'AVERAGE':
                val = sum(vals) / len(vals)
            else:
                vs = sorted(vals); n = len(vs)
                val = (vs[n // 2] if n % 2 else (vs[n // 2 - 1] + vs[n // 2]) / 2)
            e = e[:m.start()] + repr(float(val)) + e[m.end():]
        # cross-sheet single-cell references — each one lands in an arithmetic expression,
        # so each one must be a number
        while True:
            m = SHEETREF.search(e)
            if not m:
                break
            tgt = m.group(1) or m.group(2)
            v = self.numeric(tgt, m.group(3).replace('$', ''), f'the formula {expr!r}')
            e = e[:m.start()] + repr(v) + e[m.end():]
        # same-sheet references
        while True:
            m = CELLREF.search(e)
            if not m:
                break
            v = self.numeric(sheet, f'{m.group(2)}{m.group(4)}', f'the formula {expr!r}')
            e = e[:m.start()] + repr(v) + e[m.end():]
        e = e.replace('^', '**')   # Excel exponentiation
        if not re.fullmatch(r'[-+*/(). 0-9eE]+', e.replace('**', '*')):
            raise ValueError(f'unparsed formula fragment: {expr!r} -> {e!r}')
        return eval(e)

    def formula_cells(self):
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith('='):
                        yield ws.title, c.coordinate
