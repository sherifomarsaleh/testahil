"""Recalculate the DELIVERED xlsx and reconcile it cell-by-cell against the model.

Recalculation is done by the explicit evaluator in xlcalc.py over the formula set the
builder actually emits, independently of the library that wrote the file. Anything the
evaluator does not understand is reported as a FAILURE rather than skipped.

Three gates run here, in increasing strength:

  1. every formula in the workbook must evaluate;
  2. EVERY formula cell must reproduce the value the model itself computed for that cell —
     the builder records them in xlsx_expected.json as it writes. This is the gate that
     makes a formula-driven workbook safe: a formula that computes the right thing the
     wrong way, or points one row off, fails here rather than silently shipping a different
     number from the study;
  3. a hand-written set of headline reconciliations against study_numbers.json, kept as an
     independent cross-check on the expected map itself.

Every formula cell must also be COVERED by the expected map. An unchecked formula is a
failure, not a skip.
"""
import datetime as dt
import json, os
import openpyxl
import xlcalc

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'ADNOCLS_Valuation_Model_09082026_public.xlsx')
wb = openpyxl.load_workbook(XLSX)
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
XP = json.load(open(os.path.join(HERE, 'xlsx_expected.json')))
EXPECT, ANCH, ROWS = XP['expected'], XP['anchors'], XP['rows']
SG, RN, DF_ = ROWS['Segments'], ROWS['Relative & Normalized'], ROWS['DCF']
IS, BS, CF = ROWS['Income Statement'], ROWS['Balance Sheet'], ROWS['Cash Flow']
SB, SU, FV = ROWS['SOTP Bridge'], ROWS['Summary'], ROWS['Fundamental Valuation']
DCF, DCFA, LN, WACC = D['dcf'], D['dcf_beta_alt'], D['lenses'], D['wacc']
HI, HB, FC, FIN, FBS = D['hist_is'], D['hist_bs'], D['fcst'], D['fin'], D['fcst_bs']
REL, NRM, BK, SOTP = D['rel'], D['norm'], D['book'], D['sotp']
FL = D['fleet']
V_ = {k: v['value'] for k, v in D['inputs'].items()}
BF = D['beta_framing']
SH = D['meta']['shares_mn']

BK_ = xlcalc.Book(wb)
cell_value = BK_.cell_value

SHEETS = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
          'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
          'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
          'Per-Share & Ratios', 'Peer & Sector']
assert wb.sheetnames == SHEETS, f'sheet list/order wrong: {wb.sheetnames}'

# ---- gate 1: every formula must evaluate -------------------------------------
nform, errors = 0, []
for sh, coord in BK_.formula_cells():
    nform += 1
    try:
        cell_value(sh, coord)
    except Exception as ex:
        errors.append(f'{sh}!{coord}: {ex}')
for e in errors[:20]:
    print('  UNRESOLVABLE', e)

# ---- gate 2: every formula cell must reproduce the model's own value ----------
def tol_for(v):
    return max(2e-4, abs(v) * 5e-6)


nchk, drift = 0, []
for sh, cells in EXPECT.items():
    for coord, want in cells.items():
        nchk += 1
        try:
            got = cell_value(sh, coord)
        except Exception as ex:
            drift.append((sh, coord, f'ERROR {ex}', want)); continue
        if not isinstance(got, (int, float)) or abs(float(got) - want) > tol_for(want):
            drift.append((sh, coord, got, want))
for sh, coord, got, want in drift[:30]:
    g = f'{got:,.6f}' if isinstance(got, (int, float)) else repr(got)
    print(f'  DISAGREES {sh}!{coord}: workbook={g} model={want:,.6f}')

# every formula cell must be covered by the expected map — no unchecked formulas
uncovered = [f'{sh}!{coord}' for sh, coord in BK_.formula_cells()
             if coord not in EXPECT.get(sh, {})]
for u in uncovered[:20]:
    print('  UNCHECKED', u)

# ---- gate 3: headline reconciliations against study_numbers.json --------------
def g(sheet, cell):
    return cell_value(sheet, cell)


checks = [
    ('DCF present value of the explicit years', g('DCF', f"C{DF_['pvex']}"),
     DCF['pv_explicit'], 1.0),
    ('DCF present value of the terminal value', g('DCF', f"C{DF_['pvtv']}"), DCF['pv_tv'],
     1.0),
    ('DCF terminal value share of enterprise value', g('DCF', f"C{DF_['tvshare']}"),
     DCF['tv_share'], 0.0005),
    ('DCF enterprise value', g('DCF', f"C{DF_['ev']}"), DCF['ev'], 1.0),
    ('DCF equity value before the minorities', g('DCF', f"C{DF_['prenci']}"),
     DCF['ev'] - DCF['net_debt'], 1.0),
    ('DCF minority deduction — the contracted slice plus the rest at value',
     g('DCF', f"C{DF_['nci']}"), -DCF['nci'], 1.0),
    ('DCF equity attributable to ordinary shareholders', g('DCF', f"C{DF_['eq']}"),
     DCF['equity'], 1.0),
    ('DCF fair value per share (AED)', g('DCF', f"C{DF_['fvaed']}"), DCF['fv_aed'], 0.005),
    ('DCF terminal return on invested capital', g('DCF', f"C{DF_['roic']}"),
     DCF['roic_terminal'], 0.0005),
    ('DCF required reinvestment rate', g('DCF', f"C{DF_['reinv']}"), DCF['reinvest'],
     0.0005),
    ('Cost of equity', g('DCF', f"C{DF_['ke']}"), WACC['ke'], 0.00005),
    ('Cost of debt — method 1', g('DCF', f"C{DF_['kd1']}"), WACC['kd_method1'], 0.00005),
    ('Cost of debt — method 2', g('DCF', f"C{DF_['kd2']}"), WACC['kd_method2'], 0.00005),
    ('Cost of debt — method 3', g('DCF', f"C{DF_['kd3']}"), WACC['kd_method3'], 0.00005),
    ('Cost of debt — the three averaged', g('DCF', f"C{DF_['kd']}"), WACC['kd'], 0.00005),
    ('Equity weight', g('DCF', f"C{DF_['we']}"), WACC['we'], 0.00005),
    ('Debt weight', g('DCF', f"C{DF_['wd']}"), WACC['wd'], 0.00005),
    ('Perpetual capital securities weight', g('DCF', f"C{DF_['whyb']}"), WACC['wh'],
     0.00005),
    ('The three weights sum to one',
     g('DCF', f"C{DF_['we']}") + g('DCF', f"C{DF_['wd']}") + g('DCF', f"C{DF_['whyb']}"),
     1.0, 1e-9),
    ('Cost of the perpetual capital securities', g('DCF', f"C{DF_['kh']}"), WACC['kh'],
     0.00005),
    ('Cost of the perpetual capital securities — terminal', g('DCF', f"C{DF_['khterm']}"),
     WACC['kh_term'], 0.00005),
    ('Perpetual capital securities at carrying value', g('DCF', f"C{DF_['hybcap']}"),
     WACC['hybrid_cap'], 1.0),
    ('Cost of capital — explicit window', g('DCF', f"C{DF_['wacc']}"), WACC['wacc'],
     0.00005),
    ('Cost of capital — terminal', g('DCF', f"C{DF_['waccterm']}"), WACC['wacc_term'],
     0.00005),
    ('Cost of equity on the composite-index beta', g('DCF', f"C{DF_['kea']}"),
     WACC['ke_beta1'], 0.00005),
    ('Cost of capital on the composite-index beta — explicit window',
     g('DCF', f"C{DF_['wacca']}"), DCFA['wacc'], 0.00005),
    ('Cost of capital on the composite-index beta — terminal',
     g('DCF', f"C{DF_['wactermsa']}"), DCFA['wacc_term'], 0.00005),
    ('Cost of equity at the lower 90% confidence bound', g('DCF', f"C{DF_['kecilo']}"),
     WACC['rf_star'] + V_['beta_ci_lo'] * WACC['erp'], 0.00005),
    ('Cost of equity at the upper 90% confidence bound', g('DCF', f"C{DF_['kecihi']}"),
     WACC['rf_star'] + V_['beta_ci_hi'] * WACC['erp'], 0.00005),
    ('Cost of equity on the lead-lag sum beta', g('DCF', f"C{DF_['kedims']}"),
     WACC['ke_dimson'], 0.00005),
    ('DCF fair value per share — composite-index beta (AED)',
     g('DCF', f"C{DF_['fvaeda']}"), DCFA['fv_aed'], 0.005),
    ('DCF terminal value share — composite-index beta', g('DCF', f"C{DF_['tvsharea']}"),
     DCFA['tv_share'], 0.0005),
    ('Bridge equity attributable', g('SOTP Bridge', f"C{SB['eq']}"), DCF['equity'], 1.0),
    ('Bridge terminal value share', g('SOTP Bridge', f"C{SB['tvshare']}"), DCF['tv_share'],
     0.0005),
    ('Bridge fair value per share (AED)', g('SOTP Bridge', f"C{SB['fvaed']}"),
     DCF['fv_aed'], 0.005),
    ('Sum-of-the-parts enterprise value of the legs', g('SOTP Bridge', f"D{SB['legt']}"),
     SOTP['ev_ops'], 1.0),
    ('Sum-of-the-parts equity', g('SOTP Bridge', f"C{SB['beq']}"), SOTP['equity'], 1.0),
    ('Sum-of-the-parts fair value per share (AED)', g('SOTP Bridge', f"C{SB['bfv']}"),
     SOTP['fv_aed'], 0.005),
    ('Shipping multiple, built on the sheet', g('SOTP Bridge', f"C{SB['mship']}"),
     REL['blend_ev_ebitda'], 0.0005),
    # ---- the tanker leg, solved in the sheet rather than pasted --------------
    ('Implied VLCC spot rate, first quarter 2026 — solved out of the published blend',
     g('Segments', f"F{SG['sp0'] + 4}"), FL['spot_q1_26']['vlcc'], 0.5),
    ('Implied VLCC spot rate, 2025 — the four quarters averaged',
     g("Segments", f"F{SG['sp25']}"), FL['spot_fy25']['vlcc'], 0.5),
    ('Implied VLCC mid-cycle spot anchor', g('Segments', f"F{SG['spmid']}"),
     FL['spot_mid']['vlcc'], 0.5),
    ('Charter vessel-days, long range 2, FY2026', g('Segments', f"B{SG['ycd0'] + 3}"),
     sum(max(0, (min(dt.date(2027, 1, 1), dt.date(*map(int, c['end'].split('-'))))
                 - max(dt.date(2026, 1, 1),
                       dt.date(*map(int, c['start'].split('-'))))).days)
         for c in FL['charters'] if c['klass'] == 'lr2'), 0.5),
    ('Vessel-days in 2025, the basis the running cost was solved on',
     g('Segments', f"B{SG['vdays25']}"), FL['vessel_days_25'], 0.5),
    ('Implied all-in running cost per vessel-day, solved',
     g('Segments', f"B{SG['opexd0']}"), FL['opex_day'], 0.005),
    ('Implied gas-carrier revenue per vessel-day, solved',
     g('Segments', f"B{SG['gasrate0']}"), FL['gas_rate_day'], 0.005),
    ('Segments FY2026E Tankers EBITDA', g('Segments', f"B{SG['teb']}"),
     D['fcst_seg']['Tankers']['ebitda'][0], 1.0),
    ('Segments FY2026E Tankers revenue', g('Segments', f"B{SG['trev']}"),
     D['fcst_seg']['Tankers']['rev'][0], 1.0),
    ('Segments FY2026E Gas Carriers EBITDA', g('Segments', f"B{SG['gaseb']}"),
     D['fcst_seg']['Gas Carriers']['ebitda'][0], 1.0),
    ('Segments FY2026E Services EBITDA, net of the joint-venture share',
     g('Segments', f"B{SG['feb0'] + D['segs'].index('Services')}"),
     D['fcst_seg']['Services']['ebitda'][0], 1.0),
    ('Segments FY2026E total revenue', g('Segments', f"B{SG['frevt']}"), FC['revenue'][0],
     1.0),
    ('Segments FY2030E total EBITDA', g('Segments', f"F{SG['febt']}"), FC['ebitda'][4],
     1.0),
    # ---- the lenses ----------------------------------------------------------
    ('Relative lens value per share', g('Relative & Normalized', f"C{RN['base']}"),
     LN['relative']['base'], 0.005),
    ('Normalised lens value per share', g('Relative & Normalized', f"C{RN['nbase']}"),
     LN['normalized']['base'], 0.005),
    ('Book lens value per share', g('Relative & Normalized', f"C{RN['bbase']}"),
     LN['book']['base'], 0.005),
    # the two bounds are discounted at the two ends of the beta's own confidence interval;
    # built on the alternative index construction they would invert, so they are reconciled
    ('Book lens bear bound', g('Relative & Normalized', f"C{RN['bbear']}"),
     LN['book']['bear'], 0.005),
    ('Book lens bull bound', g('Relative & Normalized', f"D{RN['bbear']}"),
     LN['book']['bull'], 0.005),
    ('Book lens equity value — residual income (USD mn)',
     g('Relative & Normalized', f"C{RN['beq']}"), BK['equity_value'], 0.005),
    ('Book lens present value of the fading remainder (USD mn)',
     g('Relative & Normalized', f"C{RN['bpvtv']}"), BK['pv_terminal'], 0.005),
    ('Book lens residual income, FY2026 (USD mn)',
     g('Relative & Normalized', f"B{RN['bri']}"), BK['detail'][0]['residual_income'],
     0.005),
    ('Book lens opening book, FY2027 (USD mn)',
     g('Relative & Normalized', f"C{RN['bopen']}"), BK['detail'][1]['opening_book'], 0.005),
    ('Book lens return on ordinary equity, FY2030',
     g('Relative & Normalized', f"F{RN['broey']}"), BK['roe_path'][4], 0.00005),
    ('Sustainable return on equity', g('Relative & Normalized', f"C{RN['broe']}"),
     BK['roe_sustainable'], 0.0005),
    ('Implied price / book', g('Relative & Normalized', f"C{RN['bpb']}"), BK['pb_fair'],
     0.005),
    ('Realised vessel price over carrying value',
     g('Relative & Normalized', f"C{RN['vsratio']}"), BK['vessel_value_to_book'], 0.005),
    ('Own enterprise value / trailing EBITDA',
     g('Relative & Normalized', f"C{RN['eveb_ttm']}"), REL['own_ev_ebitda_ttm'], 0.005),
    ('Own enterprise value on the bridge convention',
     g('Relative & Normalized', f"C{RN['evbr']}"), REL['own_ev_bridge'], 1.0),
    ('Own enterprise value / FY2026E EBITDA — bridge convention',
     g('Relative & Normalized', f"C{RN['ebbr_26']}"), REL['own_ev_ebitda_26_bridge'],
     0.005),
    ('Own price / 2025 ordinary earnings', g('Relative & Normalized', f"C{RN['pe_ttm']}"),
     REL['own_pe_ttm'], 0.005),
    ('Summary weighted central', g('Summary', ANCH['summary_central']), D['central'],
     0.005),
    ('Summary weighted central — composite-index beta',
     g('Summary', ANCH['summary_central_beta_alt']), D['central_beta_alt'], 0.005),
    ('Summary expert panel average', g('Summary', f"C{SU['panel']}"), D['panel_centre'],
     0.005),
    # the beta block on the Fundamental Valuation sheet must be the framing the study
    # publishes: both constructions in full, and the interval beside them
    ('Beta — published index of its own exchange (primary)',
     g('Fundamental Valuation', f"C{FV['beta']}"), BF['primary']['beta'], 1e-9),
    ('Fair value on the primary beta', g('Fundamental Valuation', f"C{FV['fv']}"),
     BF['primary']['fv'], 0.005),
    ('Beta — equal-weight composite of the same exchange (alternative)',
     g('Fundamental Valuation', f"C{FV['betaa']}"), BF['alternative']['beta'], 1e-9),
    ('Fair value on the alternative beta', g('Fundamental Valuation', f"C{FV['fva']}"),
     BF['alternative']['fv'], 0.005),
    ('Beta — lower bound of the 90% confidence interval',
     g('Fundamental Valuation', f"C{FV['cilo']}"), BF['ci90'][0], 1e-9),
    ('Beta — upper bound of the 90% confidence interval',
     g('Fundamental Valuation', f"C{FV['cihi']}"), BF['ci90'][1], 1e-9),
    ('Income statement FY2025 EBITDA (operating)', g('Income Statement', f"D{IS['ebitda']}"),
     HI['ebitda_op'][2], 1.0),
    ('Income statement FY2025 EBITDA as reported',
     g('Income Statement', f"D{IS['ebrep']}"), HI['ebitda_reported'][2], 1.0),
    ('Income statement FY2025 profit before tax', g('Income Statement', f"D{IS['pbt']}"),
     HI['pbt'][2], 1.0),
    ('Income statement FY2030E attributable profit',
     g('Income Statement', f"I{IS['npa']}"), FIN['npa'][4], 1.0),
    ('Earnings per ordinary share FY2026E, AFTER the perpetual coupon (USD)',
     g('Income Statement', f"E{IS['eps']}"), FIN['eps'][0], 5e-6),
    ('Earnings per share FY2026E, BEFORE the perpetual coupon (USD)',
     g('Income Statement', f"E{IS['epspre']}"), FIN['eps_pre_coupon'][0], 5e-6),
    ('Earnings attributable to ordinary shareholders FY2026E',
     g('Income Statement', f"E{IS['ordn']}"), FIN['npa_ordinary'][0], 1.0),
    ('Balance sheet FY2025 net working capital', g('Balance Sheet', f"D{BS['nwc']}"),
     HB['nwc'][2], 1.0),
    ('Balance sheet FY2030E net debt', g('Balance Sheet', f"I{BS['nd']}"),
     FIN['net_debt'][4], 1.0),
    ('Balance sheet FY2030E equity attributable', g('Balance Sheet', f"I{BS['eqp']}"),
     FBS[4]['equity_parent'], 1.0),
    ('Balance sheet FY2030E invested capital', g('Balance Sheet', f"I{BS['ic']}"),
     FBS[4]['invested_capital'], 1.0),
    ('Balance sheet FY2026E return on equity', g('Balance Sheet', f"E{BS['roe']}"),
     FBS[0]['roe'], 0.0005),
    ('Cash flow FY2026E free cash flow to the firm', g('Cash Flow', f"E{CF['fcff']}"),
     FC['fcff'][0], 1.0),
    ('Summary financials FY2030E revenue', g('Summary Financials', 'I5'), FC['revenue'][4],
     1.0),
    # ---- the three derived cells that now sit on the driver sheet as formulas ----
    ('Assumptions running cost per vessel-day is a link, not a pasted result',
     g('Assumptions', f"C{ROWS['Assumptions']['opex_day']}"), FL['opex_day'], 0.005),
    ('Assumptions days sales outstanding on REPORTED revenue is derived from the audited '
     'columns', g('Assumptions', f"C{ROWS['Assumptions']['dso_rep']}"), D['ccc']['dso'][2],
     0.005),
    # the re-basing, in the sheet: the reported ratio scaled by the revenue basis the
    # forecast is built on. It is checked against the committed input, which records it to
    # one decimal, and against the arithmetic recomputed here from the committed parts.
    ('Assumptions days sales outstanding, re-based onto the forecast revenue basis',
     g('Assumptions', f"C{ROWS['Assumptions']['dso']}"), V_['dso_days'], 0.05),
    ('The re-basing arithmetic, recomputed independently',
     g('Assumptions', f"C{ROWS['Assumptions']['dso']}"),
     D['ccc']['dso'][2] * V_['rev_fy25']
     / (V_['rev_fy25'] - V_['seg_rev_tankers_fy25']
        + FL['tce_rev_25'] * V_['tnk_grossup_26']), 0.005),
    ('The 2025 gross-up shown beside it — reported tanker revenue over the same fleet\'s '
     'charter-equivalent revenue',
     g('Assumptions', f"C{ROWS['Assumptions']['gu25']}"), V_['tnk_grossup_25'], 0.0005),
    ('The 2026 gross-up shown beside it',
     g('Assumptions', f"C{ROWS['Assumptions']['gu26']}"), V_['tnk_grossup_26'], 1e-9),
    # ---- the fleet purchase announced on the anchor date ----------------------
    ('The eleven vessels bought on 7 August 2026, at the announced price',
     g('Assumptions', f"C{ROWS['Assumptions']['acq_cost']}"), V_['acq_2026_cost'], 1.0),
    ('Crude carriers acquired', g('Assumptions', f"B{ROWS['Assumptions']['acq_vlcc']}"),
     V_['acq_2026_vlcc'], 1e-9),
    ('Gas carriers acquired', g('Assumptions', f"B{ROWS['Assumptions']['acq_gas']}"),
     V_['acq_2026_gas'], 1e-9),
    ('Vessels acquired in total, added up in the sheet',
     g('Assumptions', f"B{ROWS['Assumptions']['acq_total']}"),
     V_['acq_2026_vlcc'] + V_['acq_2026_gas'], 1e-9),
    ('Acquired crude-carrier vessel-days in 2026 — from the delivery date to the year end',
     g('Segments', f"B{SG['yacd0'] + 4}"),
     V_['acq_2026_vlcc'] * (dt.date(2027, 1, 1) - dt.date(2026, 9, 1)).days, 0.5),
    ('Acquired crude-carrier vessel-days in 2027 — a full year of them',
     g('Segments', f"C{SG['yacd0'] + 4}"), V_['acq_2026_vlcc'] * 365, 0.5),
    ('No acquired vessel-days in a class where nothing was bought',
     g('Segments', f"C{SG['yacd0']}"), 0.0, 1e-9),
    ('Gas vessel-years the purchase adds in 2027 — all five, for the whole year',
     g('Assumptions', f"C{ROWS['Assumptions']['gas_vy_acq']}"), V_['acq_2026_gas'], 1e-6),
    ('Gas contracted vessel-years, 2027 — the contract table plus the purchase',
     g('Assumptions', f"C{ROWS['Assumptions']['gas_vy']}"), FL['gas_vessel_years'][1],
     1e-6),
    ('The purchase price is carried in the equity bridge',
     g('DCF', f"C{DF_['acq']}"), -V_['acq_2026_cost'], 1.0),
    ('The purchase price is capitalised in the asset roll',
     g('Balance Sheet', f"B{ROWS['Balance Sheet']['ppeacq']}"), V_['acq_2026_cost'], 1.0),
    ('The purchase price is carried in the opening net debt',
     g('Balance Sheet', f"B{ROWS['Balance Sheet']['ndopen']}"),
     V_['q1_26_netdebt'] + V_['q1_26_pcp'] + V_['acq_2026_cost'], 1.0),
    # ---- the smallest tankers, scaled rather than substituted -----------------
    ('Handysize published blend, first quarter 2026 — the medium-range rate scaled by the '
     'disclosed relative move', g('Segments', f"F{SG['blend0']}"),
     FL['blend_q1_26']['mr'] * V_['handysize_relative'], 0.5),
    ('Handysize 2025 blend, the same scaling', g('Segments', f"B{SG['blend0']}"),
     FL['blend_fy25']['hs'], 0.5),
    # ---- the cost of debt, labelled for what each figure is -------------------
    ('Cost of debt — the balance-weighted construction alone',
     g('DCF', f"C{DF_['kdbal']}"), WACC['kd_balance_weighted'], 0.00005),
    # ---- the earnings multiple on both bases ----------------------------------
    ('Blended forward price/earnings', g('Peer & Sector', f"C{ROWS['Peer & Sector']['pe']}"),
     REL['blend_pe'], 0.0005),
    ('Blended trailing price/earnings, built the same way on the trailing figures',
     g('Peer & Sector', f"C{ROWS['Peer & Sector']['pe_t']}"),
     (1 - REL['spot_weight']) * D['peers'][0]['pe_ttm']
     + REL['spot_weight'] * D['peers'][1]['pe_ttm'], 0.0005),
    ('The company on the trailing basis',
     g('Relative & Normalized', f"C{RN['pe_ttm']}"), REL['own_pe_ttm'], 0.005),
    ('The company on the forward basis — the basis the peer multiples are quoted on',
     g('Relative & Normalized', f"C{RN['pe_fwd']}"),
     WACC['mktcap'] / FIN['npa_ordinary'][0], 0.005),
    ('Assumptions opening net working capital is derived from the audited columns',
     g('Assumptions', f"C{ROWS['Assumptions']['nwc25']}"), HB['nwc'][2], 1.0),
]
bad = 0
for name, got, want, tol in checks:
    ok = got is not None and abs(float(got) - float(want)) <= tol
    if not ok:
        bad += 1
        print(f'  [BAD] {name}: workbook={got:,.4f} model={float(want):,.4f}')

assert not errors, f'{len(errors)} unresolvable formulas'
assert not drift, f'{len(drift)} formula cells disagree with the model'
assert not uncovered, f'{len(uncovered)} formula cells are not checked against the model'
assert bad == 0, f'{bad} reconciliation mismatches'
print(f'{nchk} of {nform} formula cells reproduce the model, 0 unresolvable, 0 unchecked')
print(f'RECALC OK — {len(checks)} headline reconciliations against study_numbers.json '
      f'passed')
