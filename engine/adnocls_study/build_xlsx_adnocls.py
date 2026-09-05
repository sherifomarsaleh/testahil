"""ADNOCLS_Valuation_Model_09082026_public.xlsx — 16 sheets mirroring the house canonical
model (asset-heavy marine logistics operating-company variant). Blue = inputs · black =
formulas · green = cross-sheet links.

The workbook is FORMULA-DRIVEN. Every quantity that is arithmetically derivable from an
input is written as a live Excel formula, not as a pasted number, so the reader can trace
each figure back to the drivers on the Assumptions sheet and change one to see the model
reprice. Only three classes of cell may be pasted values, and every pasted cell in the
delivered file is accounted for against them in the register below:

  1. AUDITED OR DISCLOSED HISTORY, AND GENUINE INDEPENDENT INPUTS — a figure the company
     published, a market price, a disclosed rate, a vessel count, a peer multiple read off
     a named source on a named date, or a study judgement that nothing in the model
     produces. Where a line is both disclosed and derivable, the DISCLOSED figure is
     carried — but only in ONE place: if the same audited figure appears twice, the second
     occurrence is a link, never a second typing of it;
  2. the output of a unit build that would be unreadable flattened into a grid;
  3. WHOLE-MODEL RE-RUNS: the Monte Carlo price map, the sensitivity grids, the
     bear/bull bounds and the expert-panel legs, each cell of which is a complete
     revaluation of the entire model and therefore cannot be one formula.

The tanker fleet build (including the vessels bought on 7 August 2026, counted from their
own delivery dates), the gas-carrier build, the cost of capital (including the three
cost-of-debt constructions, their average and the balance-weighted one beside it), the
discount-factor compounding, the re-basing of receivable days onto the revenue basis the
forecast uses, the statements roll (property plant and equipment, working capital from the
days ratios, equity and net debt) and every ratio and per-share figure are all live
formulas.

Every formula cell also carries the model's own value for that cell into
xlsx_expected.json, and recalc.py evaluates the workbook independently and asserts the two
agree. A formula that computes the right thing the wrong way therefore fails the gate.

===========================================================================================
THE PASTED-CELL REGISTER — SHEET BY SHEET, EVERY SURVIVING PASTED NUMBER AND ITS CLASS
Counts are measured on the DELIVERED file at the foot of this script, so this register and
the printout cannot drift apart. Calendar cells (charter dates, rate-window boundaries,
forecast-year boundaries, delivery dates) are counted separately: they are dates, not
figures, and each is a disclosed contract or calendar boundary — class 1.

  READ FIRST              0 pasted.  Prose only.
  Summary                 0 pasted.  Every figure is a link or an arithmetic combination.
  Fundamental Valuation  11 pasted.  CLASS 3, all of them: the cash-flow bear and bull
                                     bounds (2), each a complete re-run of the model at a
                                     different rate anchor, beta and capital-expenditure
                                     path, and the three expert legs (3 x base/low/high),
                                     each an independent valuation by a different method.
  Assumptions           191 pasted, 39 calendar.  CLASS 1 throughout — this is the driver
                                     sheet, and a driver sheet may carry only genuinely
                                     independent inputs. By block: market and share
                                     anchors 5; cost of capital 10; the cost-of-debt
                                     evidence 12; vessel counts by class 10; the published
                                     blended day rates 28 + 8; the handysize relative rate
                                     1; the twelve charter fixtures 12 rates + 24 dates;
                                     the rate-window and forecast-year calendar 12 dates;
                                     the 7 August 2026 purchase 4 + 3 dates; the rate path
                                     and running cost 3; gas carriers 4; the gas contract
                                     table 5; the five remaining units 50; capital
                                     expenditure, depreciation and working capital 7; the
                                     depreciation memoranda 2; tax by unit and the 2025
                                     depreciation basis 10; funding and the bridge 13;
                                     lens weights and the multiple blend 7.
                                     ANYTHING ON THIS SHEET THAT IS ITSELF DERIVED IS A
                                     GREEN FORMULA, not a number: the running cost per
                                     vessel-day, the gas day rate, the reported and
                                     re-based days ratios, the two gross-up factors, the
                                     opening working capital, the intangibles and goodwill
                                     (which the balance sheet carries), the acquired gas
                                     vessel-years and the gas total.
  SOTP Bridge             0 pasted.
  Segments               42 pasted.  CLASS 1: the disclosed operating-segments note —
                                     seven segments x three audited years, revenue and
                                     earnings. This is the primary record, and the group
                                     revenue line on the Income Statement is now the SUM of
                                     it rather than a second typing of the same total.
  Relative & Normalized   3 pasted.  CLASS 1: the carrying value, realised price and
                                     disclosed gain on the one vessel actually sold.
  DCF                     0 pasted.
  Income Statement       42 pasted.  CLASS 1: fourteen audited disclosure-only lines x
                                     three years — direct costs, general and
                                     administrative, expected credit losses, other income
                                     and expenses, depreciation, the share of joint
                                     ventures, the two 2025 acquisition items, finance
                                     income and costs, tax, minorities and the perpetual
                                     coupon. Revenue and the reported-EBITDA bridge are no
                                     longer among them: both are added up on the sheet.
  Balance Sheet          45 pasted.  CLASS 1: fifteen audited lines x three years. Net debt
                                     is no longer among them — the audited years are gross
                                     borrowings less cash, computed here.
  Cash Flow               6 pasted.  CLASS 1: reported operating cash flow and capital
                                     expenditure, three audited years each.
  Summary Financials      0 pasted.
  Monte Carlo            22 pasted.  21 CLASS 3 — the percentile map, the level-event
                                     probabilities and the anchor volatility are outputs of
                                     a 50,000-path simulation, not of this workbook — plus
                                     1 CLASS 1, the path count itself, which is a setting.
  Sensitivity            35 pasted.  32 CLASS 3: every cell of the beta x growth grid, the
                                     rate-anchor row, the capital-expenditure row and the
                                     uniform-tax row is a complete revaluation of the whole
                                     model, fleet build included. Plus 3 CLASS 1: the
                                     one-year time-charter fixture, the broker spot print
                                     and the order book, each an external market
                                     observation.
  Per-Share & Ratios      0 pasted.
  Peer & Sector           9 pasted.  CLASS 1: the peer multiples, each read off a named
                                     source on a named date and shown with both.

NO CELL IN THE DELIVERED FILE NOW RESTS ON CLASS 2. The block that used to be justified
that way — the historical segment grid — is disclosed history and is class 1; it is the
operating-segments note as published.
===========================================================================================
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
V = {k: v['value'] for k, v in D['inputs'].items()}

BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; PX3 = '0.000;(0.000);"-"'
MULT = '0.00x'; DF4 = '0.0000'; BETA = '0.000'

M = D['meta']
HI, HB, HC_, CCC = D['hist_is'], D['hist_bs'], D['hist_cf'], D['ccc']
SEGH, GRPH = D['seg_hist'], D['grp_hist']
FLEET, DRV = D['fleet'], D['drivers']
FC, FIN, FBS = D['fcst'], D['fin'], D['fcst_bs']
WACC, DCFB, DCFA = D['wacc'], D['dcf'], D['dcf_beta_alt']
BF = D['beta_framing']
LN, LW, REL, NRM, BK = D['lenses'], D['lens_weights'], D['rel'], D['norm'], D['book']
EXP, SN, STK, SOTP = D['experts'], D['sens'], D['strike'], D['sotp']
PEERS, TECH, STEP0 = D['peers'], D['technicals'], D['step0']
BE = D['beta']                      # the regression's own record, provenance included

SEGS = D['segs']; SEG_GROUP = D['seg_group']; GROUPS = D['groups']
YH = ['FY2023', 'FY2024', 'FY2025']
YF = FC['years']
YFE = [y + 'E' for y in YF]
SPOT = M['spot_aed']; SH = M['shares_mn']; PEG = M['fx']
CLS = ['hs', 'mr', 'lr1', 'lr2', 'vlcc']
CLS_NAME = ['Handysize', 'Medium range', 'Long range 1', 'Long range 2',
            'Very large crude carrier']
CD = ['B', 'C', 'D', 'E', 'F']            # five forecast years, or five tanker classes
HC = ['B', 'C', 'D']                      # three historical years on the statements
FCOL = ['E', 'F', 'G', 'H', 'I']          # five forecast years on the statements
ALL = HC + FCOL

# ============================================================================
# THE MODEL, RECOMPUTED FROM THE COMMITTED INPUTS
# Every expected value written into xlsx_expected.json comes from here, and every line
# below is asserted against the committed study blocks at the foot of this file. No
# financial numeral is typed into this builder.
# ============================================================================
# [R-MACRO-01] THE ESCALATOR IS A PATH, NOT A RATE. The model puts every dirham cost line
# on the house inflation ladder year by year — 2.5 per cent in 2026 and 2.0 thereafter —
# and a single rate compounded five times is a different series. ESC is kept as the
# terminal rate for the one assumption cell that quotes it; ESC_IDX is what the model
# actually applies and what every line below uses.
ESC = V['opex_escalation']
ESC_PATH = V['opex_escalation_path']
ESC_IDX = []
_c = 1.0
for _r in ESC_PATH:
    _c *= (1.0 + _r)
    ESC_IDX.append(_c)
H2W = V['h2_2026_reversion']
GROSSUP = V['tnk_grossup_26']

# --- the fleet, the twelve charters out, and the calendar they run on ----------
# The company publishes ONE rate per class per quarter and its own chief financial officer
# said on the first-quarter call that this rate is a BLEND across the whole class, charters
# out included. So the spot rate is not read off the disclosure — it is SOLVED out of it,
# window by window, by removing the chartered vessels at their own contracted rates:
#     spot = (blend x class vessel-days - charter revenue) / spot vessel-days
# Everything below is that arithmetic, and every step of it is written into the sheet.
import datetime as _dt
_EPOCH = _dt.date(1899, 12, 30)          # the spreadsheet's own day zero


def _d(s):
    return _dt.date(*map(int, s.split('-')))


def _ser(d):
    """A calendar date as the serial number a spreadsheet stores it as."""
    return (d - _EPOCH).days


OWNED25 = FLEET['owned_fy25']            # owned at 31 December 2025
OWNED = FLEET['owned']                   # owned at the 31 March 2026 valuation date
SOLD = {c: OWNED25[c] - OWNED[c] for c in CLS}

# --- the fleet purchase announced on 7 August 2026 ----------------------------
# Announced on the anchor date of this study, so it is already inside the market price the
# valuation is compared against: eleven vessels for about USD 1.3 billion — six very large
# crude carriers and three gas carriers bought secondhand for third-quarter delivery, and
# two gas carrier newbuildings resold from a yard for the fourth. The crude carriers trade
# spot from delivery and contribute vessel-days from then; the gas carriers add contracted
# vessel-years; and the price is added BOTH to net debt (it is committed and funded) and to
# the opening property, plant and equipment, so it depreciates and shields tax like any
# other vessel.
#
# THE COUNTS AND THE PRICE ARE COMMITTED INPUTS. What is written here rather than read is
# the DELIVERY SCHEDULE — which tranche arrives when — because the committed file carries
# the totals, not the split. It is not free: the crude-carrier tranche is pinned by the
# assertions at the foot of this file, which require the tanker leg to reproduce the
# committed forecast to a thousandth of a per cent, and the gas tranches are pinned to the
# committed vessel-year totals through the base line below.
ACQ_COST = V['acq_2026_cost']
ACQ_VLCC_N = V['acq_2026_vlcc']
ACQ_GAS_N = V['acq_2026_gas']
ACQ_SECONDHAND = _d('2026-09-01')        # the nine secondhand vessels, third quarter
ACQ_NEWBUILD = _d('2026-11-15')          # the two gas newbuildings, fourth quarter
ACQ_GAS_SECONDHAND_N = 3
ACQ_GAS_NEWBUILD_N = ACQ_GAS_N - ACQ_GAS_SECONDHAND_N
ACQ_N = {c: (ACQ_VLCC_N if c == 'vlcc' else 0) for c in CLS}
# a class with nothing acquired carries a delivery date of zero, so one formula serves
# every class: a count of zero makes the date irrelevant
ACQ_DATE = {c: (_ser(ACQ_SECONDHAND) if c == 'vlcc' else 0) for c in CLS}
CHARTERS = [dict(name=ch['name'], klass=ch['klass'], rate=float(ch['rate']),
                 start=_d(ch['start']), end=_d(ch['end']),
                 period=ch['period_months']) for ch in FLEET['charters']]
CH_ROWS = {c: [k for k, ch in enumerate(CHARTERS) if ch['klass'] == c] for c in CLS}


def ch_days(klass, a, b):
    """Vessel-days and revenue the class's charters out earn over the window [a, b)."""
    days = rev = 0.0
    for ch in CHARTERS:
        if ch['klass'] != klass:
            continue
        n = (min(b, ch['end']) - max(a, ch['start'])).days
        if n > 0:
            days += n
            rev += n * ch['rate']
    return days, rev


# Seven rate windows: the four quarters of 2025, the two disclosed quarters of 2026, and
# the mid-cycle anchor, which is the 2024/2025 average blend converted on the same fleet
# and charter book as the first quarter of 2025.
WIN = [(_d('2025-01-01'), _d('2025-04-01')), (_d('2025-04-01'), _d('2025-07-01')),
       (_d('2025-07-01'), _d('2025-10-01')), (_d('2025-10-01'), _d('2026-01-01')),
       (_d('2026-01-01'), _d('2026-04-01')), (_d('2026-04-01'), _d('2026-07-01')),
       (_d('2025-01-01'), _d('2025-04-01'))]
WIN_LAB = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'Q1 2026', 'Q2 2026',
           'Mid-cycle anchor']
WIN_FLEET = [OWNED25] * 4 + [OWNED] * 2 + [OWNED25]
WCOL = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
NWIN = 7

Q25 = {c: [V[f'tce_{c}_25q{i+1}'] for i in range(4)] for c in ('mr', 'lr1', 'lr2', 'vlcc')}
Q24 = {c: [V[f'tce_{c}_24q{i+1}'] for i in range(4)] for c in ('lr1', 'lr2', 'vlcc')}
TCE25 = {c: sum(Q25[c]) / 4.0 for c in Q25}
# THE SMALLEST TANKERS ARE NOT BROKEN OUT, AND STANDING THE MEDIUM-RANGE RATE IN FOR THEM
# UNADJUSTED HAD THE SIGN WRONG. The company said on the first-quarter call that Handysize
# rates were softer, down 21%, while medium range was up 29% — the two classes moved in
# OPPOSITE directions. So the medium-range rate is scaled by the disclosed Handysize move
# wherever it stands in, on every window and on both sides of the mid-cycle average, and
# the scalar is a driver on the sheet rather than an adjustment buried in a script.
HS_REL = V['handysize_relative']
TCE25['hs'] = TCE25['mr'] * HS_REL
TCE24 = {c: sum(Q24[c]) / 4.0 for c in Q24}
TCE24['mr'] = TCE25['mr']                # 2024 quarterly rates for this class are not given
TCE24['hs'] = TCE24['mr'] * HS_REL
BLEND_MID = {c: (TCE24[c] + TCE25[c]) / 2.0 for c in CLS}
B26 = {c: [V[f'tce_{c}_q1_26'], V[f'tce_{c}_q2_26']] for c in ('mr', 'lr1', 'lr2', 'vlcc')}
B26['hs'] = [x * HS_REL for x in B26['mr']]
BLEND_W = {c: ([Q25[c][i] if c in Q25 else TCE25[c] for i in range(4)]
               + list(B26[c]) + [BLEND_MID[c]]) for c in CLS}

CDAYS = {c: [WIN_FLEET[w][c] * (WIN[w][1] - WIN[w][0]).days for w in range(NWIN)]
         for c in CLS}
CHD = {c: [ch_days(c, *WIN[w])[0] for w in range(NWIN)] for c in CLS}
CHREV = {c: [ch_days(c, *WIN[w])[1] for w in range(NWIN)] for c in CLS}
SDAYS = {c: [CDAYS[c][w] - CHD[c][w] for w in range(NWIN)] for c in CLS}
SPOTWIN = {c: [(BLEND_W[c][w] * CDAYS[c][w] - CHREV[c][w]) / SDAYS[c][w]
              for w in range(NWIN)] for c in CLS}
SPOT25 = {c: sum(SPOTWIN[c][:4]) / 4.0 for c in CLS}
SPOT_MID = {c: SPOTWIN[c][6] for c in CLS}
SPOT_Q1 = {c: SPOTWIN[c][4] for c in CLS}
SPOT_Q2 = {c: SPOTWIN[c][5] for c in CLS}
TNK_H2 = {c: SPOT_Q1[c] * (1 - H2W) + SPOT25[c] * H2W for c in CLS}
TNK_Y26 = {c: (SPOT_Q1[c] + SPOT_Q2[c] + 2 * TNK_H2[c]) / 4.0 for c in CLS}
TNK_PATH = {c: [TNK_Y26[c] + (SPOT_MID[c] - TNK_Y26[c]) * i / 4.0 for i in range(5)]
            for c in CLS}

YRB = [(_dt.date(2026 + i, 1, 1), _dt.date(2027 + i, 1, 1)) for i in range(5)]
YRDAYS = [(b - a).days for a, b in YRB]
YCD = {c: [ch_days(c, *YRB[i])[0] for i in range(5)] for c in CLS}
YCR = {c: [ch_days(c, *YRB[i])[1] / 1000.0 for i in range(5)] for c in CLS}


def acq_days(klass, i):
    """Vessel-days the newly bought ships of a class contribute in forecast year i."""
    if not ACQ_N[klass]:
        return 0.0
    a, b = _ser(YRB[i][0]), _ser(YRB[i][1])
    return ACQ_N[klass] * max(0.0, b - max(a, ACQ_DATE[klass]))


YACD = {c: [acq_days(c, i) for i in range(5)] for c in CLS}
YSD = {c: [OWNED[c] * YRDAYS[i] - YCD[c][i] + YACD[c][i] for i in range(5)] for c in CLS}
YSR = {c: [YSD[c][i] * TNK_PATH[c][i] / 1000.0 for i in range(5)] for c in CLS}
TNK_CHREV = [sum(YCR[c][i] for c in CLS) for i in range(5)]
TNK_SPOTREV = [sum(YSR[c][i] for c in CLS) for i in range(5)]
TNK_TCEREV = [TNK_CHREV[i] + TNK_SPOTREV[i] for i in range(5)]

# The running cost is not an assumption either: it is solved so that the same construction
# reproduces the tanker earnings the company actually reported for 2025.
VDAYS25 = sum(OWNED25.values()) * 365
TCEREV25 = sum(OWNED25[c] * TCE25[c] for c in CLS) * 365 / 1000.0
TNK_EB25 = V['seg_ebitda_tankers_fy25']
OPEX_DAY = (TCEREV25 - TNK_EB25) * 1000.0 / VDAYS25
TNK_OPEXD = [OPEX_DAY * ESC_IDX[i] for i in range(5)]
TNK_OPEX = [VDAYS25 * TNK_OPEXD[i] / 1000.0 for i in range(5)]
TNK_EBITDA = [TNK_TCEREV[i] - TNK_OPEX[i] for i in range(5)]
TNK_REV = [TNK_TCEREV[i] * GROSSUP for i in range(5)]

GAS_VY = FLEET['gas_vessel_years']
# The five gas carriers bought in August 2026 are inside those committed vessel-years. The
# sheet shows the two parts rather than the total alone, so a reader can see what the
# purchase adds: the acquired part is computed from the counts and delivery dates, and the
# contract table BEFORE the purchase is what is left. (The committed model measures the
# part-year in months and this line measures it in days, which differ by about a hundredth
# of a vessel-year in 2026 and not at all thereafter; the difference sits in the base line
# and the TOTAL — the only figure the model uses — reproduces the committed one exactly.)
GAS_ACQ_VY = [(ACQ_GAS_SECONDHAND_N
               * max(0.0, _ser(YRB[i][1]) - max(_ser(YRB[i][0]), _ser(ACQ_SECONDHAND)))
               + ACQ_GAS_NEWBUILD_N
               * max(0.0, _ser(YRB[i][1]) - max(_ser(YRB[i][0]), _ser(ACQ_NEWBUILD))))
              / YRDAYS[i] for i in range(5)]
GAS_VY_BASE = [GAS_VY[i] - GAS_ACQ_VY[i] for i in range(5)]
GAS_VY25 = V['gas_vessel_years_25']
GAS_REV25 = V['seg_rev_gas_carriers_fy25']
GAS_RATE = GAS_REV25 * 1000.0 / (GAS_VY25 * 365)
GAS_MGN = V['gas_margin']
JV_GAS = V['jv_gas_fy25']; JV_SERV = V['jv_services_fy25']
GAS_RATED = [GAS_RATE * ESC_IDX[i] for i in range(5)]
GAS_REV = [GAS_VY[i] * 365 * GAS_RATED[i] / 1000.0 for i in range(5)]
GAS_GROSS_EB = [r * GAS_MGN for r in GAS_REV]
# The disclosed segment earnings INCLUDE the equity-accounted share of joint-venture
# profit, and the equity bridge already adds those ventures at carrying value. Leaving them
# in the forecast would count them twice, so they come out here — in Gas Carriers and in
# Services, the two units whose 2025 disclosure carries them.
GAS_JV = [JV_GAS * ESC_IDX[i] for i in range(5)]
SERV_JV = [JV_SERV * ESC_IDX[i] for i in range(5)]
GAS_EBITDA = [GAS_GROSS_EB[i] - GAS_JV[i] for i in range(5)]

SEG_REV_F, SEG_EB_F, SEG_GROSS_EB = {}, {}, {}
for s in SEGS:
    if s == 'Tankers':
        SEG_REV_F[s], SEG_EB_F[s] = list(TNK_REV), list(TNK_EBITDA)
    elif s == 'Gas Carriers':
        SEG_REV_F[s], SEG_EB_F[s] = list(GAS_REV), list(GAS_EBITDA)
    else:
        SEG_REV_F[s] = list(DRV[s]['rev'])
        _eb = [r * m for r, m in zip(DRV[s]['rev'], DRV[s]['mar'])]
        SEG_GROSS_EB[s] = list(_eb)
        if s == 'Services':
            _eb = [_eb[i] - SERV_JV[i] for i in range(5)]
        SEG_EB_F[s] = _eb
REV_F = [sum(SEG_REV_F[s][i] for s in SEGS) for i in range(5)]
EB_F = [sum(SEG_EB_F[s][i] for s in SEGS) for i in range(5)]
GRP_REV_F = {g: [sum(SEG_REV_F[s][i] for s in SEGS if SEG_GROUP[s] == g) for i in range(5)]
             for g in GROUPS}
GRP_EB_F = {g: [sum(SEG_EB_F[s][i] for s in SEGS if SEG_GROUP[s] == g) for i in range(5)]
            for g in GROUPS}

DEP_RATE = V['dep_rate_ppe']; OTHER_DNA = V['other_dna_run_rate']
CAPEX = [V[f'capex_{y[2:]}'] for y in YF]
# the vessels bought in August 2026 join the asset base in the year they arrive, so they
# depreciate and shield tax exactly as any other vessel does
ACQ_CAPEX = [ACQ_COST, 0.0, 0.0, 0.0, 0.0]
PPE_OPEN, PPE_CLOSE, DEP1, DEP_PPE = [], [], [], []
_o = V['ppe_fy25']
for i in range(5):
    d1 = DEP_RATE * (_o + (_o + CAPEX[i] + ACQ_CAPEX[i])) / 2.0
    d = DEP_RATE * (_o + max(_o + CAPEX[i] + ACQ_CAPEX[i] - d1, 0)) / 2.0
    PPE_OPEN.append(_o); DEP1.append(d1); DEP_PPE.append(d)
    PPE_CLOSE.append(_o + CAPEX[i] + ACQ_CAPEX[i] - d); _o = PPE_CLOSE[-1]
OTHER_DNA_Y = [OTHER_DNA * ESC_IDX[i] for i in range(5)]
DNA_F = [DEP_PPE[i] + OTHER_DNA_Y[i] for i in range(5)]
EBIT_F = [EB_F[i] - DNA_F[i] for i in range(5)]

SEG_DNA25 = {s: V['seg_dna_' + s.lower().replace(' ', '_').replace('-', '_') + '_fy25']
             for s in SEGS}
SEG_DNA_TOT = sum(SEG_DNA25.values())
GRP_DNA_SHARE = {g: sum(SEG_DNA25[s] for s in SEGS if SEG_GROUP[s] == g) / SEG_DNA_TOT
                 for g in GROUPS}
TAX_G = {'Integrated Logistics': V['tax_integrated_logistics'],
         'Shipping': V['tax_shipping'], 'Services': V['tax_services']}
GRP_DNA_F = {g: [DNA_F[i] * GRP_DNA_SHARE[g] for i in range(5)] for g in GROUPS}
GRP_TAXABLE = {g: [max(GRP_EB_F[g][i] - GRP_DNA_F[g][i], 0) for i in range(5)]
               for g in GROUPS}
GRP_TAX = {g: [GRP_TAXABLE[g][i] * TAX_G[g] for i in range(5)] for g in GROUPS}
TAX_F = [sum(GRP_TAX[g][i] for g in GROUPS) for i in range(5)]
TAXRATE_F = [TAX_F[i] / EBIT_F[i] for i in range(5)]
NOPAT_F = [EBIT_F[i] - TAX_F[i] for i in range(5)]
OPCOST_F = [REV_F[i] - EB_F[i] for i in range(5)]

OPCOST25 = V['rev_fy25'] - HI['ebitda_op'][2]
# RECEIVABLE DAYS ARE RE-BASED, NOT CARRIED ACROSS. The ratio is measured on REPORTED 2025
# revenue, which carries a 2.72x gross-up from charter-equivalent revenue, and it is then
# applied to forecast revenue built at 1.60x. The same receivable balance against a smaller
# revenue line is MORE days, not the same days, so the ratio is re-based onto the revenue
# basis the forecast actually uses. This is also what falsifies the claim that the gross-up
# "never touches the valuation": it reaches it here, through receivables and the change in
# working capital. Every step is written into the sheet.
DSO_REPORTED = CCC['dso'][2]
GROSSUP25 = V['seg_rev_tankers_fy25'] / TCEREV25
REV25_FWD_BASIS = V['rev_fy25'] - V['seg_rev_tankers_fy25'] + TCEREV25 * GROSSUP
DSO = DSO_REPORTED * V['rev_fy25'] / REV25_FWD_BASIS
DIO = V['inv_fy25'] / OPCOST25 * 365
DPO = (V['pay_fy25'] + V['dtr_c_fy25']) / OPCOST25 * 365
NWC25 = HB['nwc'][2]
RECV_F = [REV_F[i] * DSO / 365 for i in range(5)]
INV_F = [OPCOST_F[i] * DIO / 365 for i in range(5)]
PAY_F = [OPCOST_F[i] * DPO / 365 for i in range(5)]
NWC_F = [RECV_F[i] + INV_F[i] - PAY_F[i] for i in range(5)]
DNWC_F = [NWC_F[0] - NWC25] + [NWC_F[i] - NWC_F[i - 1] for i in range(1, 5)]
FCFF_F = [NOPAT_F[i] + DNA_F[i] - CAPEX[i] - DNWC_F[i] for i in range(5)]

# --- cost of capital ---------------------------------------------------------
MKTCAP = SH * SPOT / PEG * 1000.0
RF_STAR = V['rf_observed'] - V['sov_spread']
KE = RF_STAR + V['beta'] * V['erp_total']
# The alternative construction is the SAME regression measured against a different market:
# an equal-weight composite of the exchange's own names rather than its published index.
KE_A = RF_STAR + V['beta_composite'] * V['erp_total']
# The regression's own 90% confidence interval, and the same slope shrunk toward the
# market, priced through the same cost-of-equity construction so the reader sees the span
# the estimate supports.
KE_CI_LO = RF_STAR + V['beta_ci_lo'] * V['erp_total']
KE_CI_HI = RF_STAR + V['beta_ci_hi'] * V['erp_total']
KE_BLUME = RF_STAR + V['beta_blume'] * V['erp_total']
KD1 = V['sofr'] + V['shldr_margin']
KD_BANK = (V['bank_loan_lo'] + V['bank_loan_hi']) / 2
KD_OTHER = (V['other_borr_lo'] + V['other_borr_hi']) / 2
KD_TP = (KD_BANK + KD_OTHER) / 2
KD_LEASE = V['intpaid_lease_fy25'] / ((V['lease_open_fy25'] + V['lease_close_fy25']) / 2)
DEBT_NOW = V['q1_26_shldr_loan'] + V['q1_26_borrowings'] + V['q1_26_leases']
KD2 = (V['q1_26_shldr_loan'] * KD1 + V['q1_26_borrowings'] * KD_TP
       + V['q1_26_leases'] * KD_LEASE) / DEBT_NOW
KD3 = KD_BANK
# [R-COC-01 AMENDED] THE ADOPTED RATE REPRODUCES FROM ITS CONTRACTUAL ANCHOR OR IT IS NOT
# THE ADOPTED RATE. Only the second construction is weighted by what is actually drawn, so
# only the second reproduces from the facility lines; averaging it with a marginal
# drawdown rate and a bank range's mid-point gives a figure no set of balances and rates
# produces. The other two are published beside it as what they are. This is [R-LENS-03]'s
# lesson applied to the cost of debt: averaging several methods makes a new method with
# free parameters nobody tested, wearing the appearance of caution.
KD = KD2
KD_BALWTD = KD2
KD_RETIRED_AVG = (KD1 + KD2 + KD3) / 3
TAXS = V['tax_stat']
KD_AT = KD * (1 - TAXS)
# The perpetual capital securities are deducted in the equity bridge as a claim ranking
# ahead of the ordinary shares. A claim deducted from enterprise value must also be
# WEIGHTED in the cost of capital at its own cost: those are two halves of one treatment.
# So the capital base is equity, debt AND hybrid, and the hybrid carries its own coupon.
# The coupon is not tax-deductible — it is an equity distribution — so it is not taxed down.
KH = V['sofr'] + V['hybrid_margin']
HYBRID_CAP = V['q1_26_hybrid']
CAP_TOT = MKTCAP + DEBT_NOW + HYBRID_CAP
WE = MKTCAP / CAP_TOT; WD = DEBT_NOW / CAP_TOT; WH = HYBRID_CAP / CAP_TOT
W_EXP = WE * KE + WD * KD_AT + WH * KH
KE_T = V['rf_terminal'] + V['beta'] * V['erp_total']
KE_T_A = V['rf_terminal'] + V['beta_composite'] * V['erp_total']
KD_T = V['rf_terminal'] + (KD - RF_STAR)
KD_T_AT = KD_T * (1 - TAXS)
# the perpetual pays a floating coupon, so its cost normalises with the risk-free rate
KH_T = V['rf_terminal'] + V['hybrid_margin']
W_TERM = WE * KE_T + WD * KD_T_AT + WH * KH_T
# EVERY cost-of-capital construction in this model carries the SAME THREE TRANCHES. The
# alternative index construction changes the cost of equity and nothing else; it does not
# change what the company is financed with. An earlier build let the perpetual tranche
# reach the base and terminal rates but not the scenario rate, which left the sensitivity
# grid struck 63 basis points cheap and its own centre cell printing a figure the study
# never published.
W_EXP_A = WE * KE_A + WD * KD_AT + WH * KH
W_TERM_A = WE * KE_T_A + WD * KD_T_AT + WH * KH_T

STUB = 0.75
G = V['g_terminal']
NDCO = V['q1_26_netdebt']; DEFERRED = V['q1_26_pcp']; HYBRID = V['q1_26_hybrid']
NCI_BV = V['q1_26_nci']; JV_BV = V['jv_bv_q126']; EQP0 = V['q1_26_eqp']
# THE MINORITIES ARE NOT ONE THING. Most of the carried balance arose on the Navig8
# combination, and that 20% is CONTRACTED for purchase in mid-2027 — the present value of
# that purchase is already in the bridge as deferred consideration. Deducting it a second
# time at a share of equity VALUE would count it twice, so it is deducted at its contracted
# price. Only the remaining minorities are lifted from book to value.
NCI_NAV = V['nci_navig8']
NCI_OTHER = NCI_BV - NCI_NAV
NCI_SHARE = V['nci_share']
NCI_SH_OTHER = NCI_SHARE * NCI_OTHER / NCI_BV


def nci_ded(pre):
    return NCI_NAV + max(NCI_OTHER, pre * NCI_SH_OTHER)
CASH = V['q1_26_cash']; Q1FCF = V['q1_26_fcf']
# the August purchase is committed and funded, so its price is carried in net debt from
# the valuation date — the same USD 1.3 billion that sits in the asset base above
NETDEBT = NDCO + DEFERRED + ACQ_COST
INTANG = V['intang_fy25']; GW = V['gw_fy25']
IC_F = [PPE_CLOSE[i] + NWC_F[i] + INTANG + GW for i in range(5)]


def dcf_legs(w, wt):
    glide = [w + (wt - w) * (i + 1) / 5.0 for i in range(5)]
    df, cum = [], 1.0
    for i, rr in enumerate(glide):
        cum *= (1 + rr) ** (STUB if i == 0 else 1.0)
        df.append(1.0 / cum)
    fcffd = list(FCFF_F); fcffd[0] -= Q1FCF
    pv = [c * d for c, d in zip(fcffd, df)]
    pv_expl = sum(pv)
    roic_t = NOPAT_F[4] / IC_F[4]
    reinv = G / roic_t
    nopat_t1 = NOPAT_F[4] * (1 + G)
    tv = nopat_t1 * (1 - reinv) / (wt - G)
    pv_tv = tv * df[4]
    ev_ops = pv_expl + pv_tv
    ev = ev_ops + JV_BV
    pre = ev - NETDEBT - HYBRID
    ded = nci_ded(pre)
    eq = pre - ded
    return dict(glide=glide, df=df, fcffd=fcffd, pv=pv, pv_expl=pv_expl, roic_t=roic_t,
                reinv=reinv, nopat_t1=nopat_t1, tv=tv, pv_tv=pv_tv,
                tv_share=pv_tv / ev_ops, ev_ops=ev_ops, ev=ev, pre_nci=pre, nci=ded,
                equity=eq, fv_usd=eq / SH / 1000.0, fv_aed=eq / SH / 1000.0 * PEG)


DC = dcf_legs(W_EXP, W_TERM)
DA = dcf_legs(W_EXP_A, W_TERM_A)


# --- the funding roll and the forecast statements -----------------------------
DPS = [V['dps_2026_usd'] * 1000.0 * (1 + V['div_growth']) ** i for i in range(5)]
HYB_CPN = HYBRID * (V['sofr'] + V['hybrid_margin'])
ND_OPEN, GROSS_D, INT_F, FININC_F, ND_CLOSE, FCFE_F = [], [], [], [], [], []
_nd = NETDEBT
for i in range(5):
    g_open = _nd + CASH
    it = KD * g_open
    fi = V['sofr'] * CASH
    fcfe = FCFF_F[i] - it * (1 - TAXS) + fi * (1 - TAXS) - HYB_CPN
    ND_OPEN.append(_nd); GROSS_D.append(g_open); INT_F.append(it); FININC_F.append(fi)
    FCFE_F.append(fcfe); ND_CLOSE.append(_nd - fcfe + DPS[i]); _nd = ND_CLOSE[-1]
PBT_F = [EBIT_F[i] - INT_F[i] + FININC_F[i] for i in range(5)]
TAXP_F = [PBT_F[i] * TAXRATE_F[i] for i in range(5)]
PAT_F = [PBT_F[i] - TAXP_F[i] for i in range(5)]
NCI_F = [PAT_F[i] * NCI_SHARE for i in range(5)]
NPA_F = [PAT_F[i] - NCI_F[i] for i in range(5)]
ORD_F = [NPA_F[i] - HYB_CPN for i in range(5)]
EQ_OPEN, EQ_CLOSE = [], []
_e = EQP0
for i in range(5):
    EQ_OPEN.append(_e); _e = _e + NPA_F[i] - DPS[i] - HYB_CPN; EQ_CLOSE.append(_e)
ROE_F = [NPA_F[i] / ((EQ_OPEN[i] + EQ_CLOSE[i]) / 2) for i in range(5)]
ROIC_F = [NOPAT_F[i] / IC_F[i] for i in range(5)]
BVPS_F = [EQ_CLOSE[i] / SH / 1000.0 for i in range(5)]

# --- the lenses ---------------------------------------------------------------
SPOT_W = V['spot_share_ebitda_26']
MULT_CONTR = PEERS[0]['ev_ebitda']
MULT_SPOT = (PEERS[1]['ev_ebitda'] + PEERS[2]['ev_ebitda']) / 2
BLEND_EV = (1 - SPOT_W) * MULT_CONTR + SPOT_W * MULT_SPOT
# THE EARNINGS MULTIPLE, ON BOTH BASES. The peers' FORWARD multiples are applied to the
# company's FORWARD earnings, which is consistent and is what the lens values. What was not
# consistent was the comparison beside it, which quoted the company on a TRAILING multiple
# against peers shown forward. Both blends are built, and the company's own multiple is
# shown on both bases, so the like-for-like comparison is the one on the page.
BLEND_PE = (1 - SPOT_W) * PEERS[0]['pe_fwd'] + SPOT_W * PEERS[1]['pe_fwd']
BLEND_PE_TTM = (1 - SPOT_W) * PEERS[0]['pe_ttm'] + SPOT_W * PEERS[1]['pe_ttm']


def pre_nci_from_ev(ev):
    return ev + JV_BV - NETDEBT - HYBRID


def eq_from_ev(ev):
    p = pre_nci_from_ev(ev)
    return p - nci_ded(p)


def per_share(eq):
    return eq / SH / 1000.0 * PEG


REL_EV = BLEND_EV * EB_F[0]
REL_V_EV = per_share(eq_from_ev(REL_EV))
REL_V_PE = BLEND_PE * ORD_F[0] / SH / 1000.0 * PEG
W_EVEB = V['rel_weight_ev_ebitda']
REL_BASE = W_EVEB * REL_V_EV + (1 - W_EVEB) * REL_V_PE
REL_BEAR = per_share(eq_from_ev(MULT_SPOT * EB_F[0]))
REL_BULL = per_share(eq_from_ev(MULT_CONTR * EB_F[0]))
NORM_EB = sum(EB_F) / 5.0
NORM_ORD = sum(NPA_F) / 5.0 - HYB_CPN
NORM_V_EV = per_share(eq_from_ev(BLEND_EV * NORM_EB))
NORM_EPS = NORM_ORD / SH / 1000.0
NORM_V_PE = NORM_EPS * BLEND_PE * PEG
NORM_BASE = (NORM_V_EV + NORM_V_PE) / 2
NORM_BEAR = per_share(eq_from_ev(MULT_SPOT * NORM_EB))
NORM_BULL = per_share(eq_from_ev(MULT_CONTR * NORM_EB))
ROE_SUST = sum(ROE_F) / 5.0
BVPS0 = EQP0 / SH / 1000.0
# THE BOOK LENS IS A RESIDUAL-INCOME BUILD, NOT A JUSTIFIED PRICE-TO-BOOK. The single-stage
# form (ROE - g)/(Ke - g) is only coherent in a steady state, where the company distributes
# exactly what it does not need to fund g. This one pays out a third of earnings and
# compounds its book ABOVE its cost of equity, where that formula is not merely wrong but
# undefined. So the lens is built as the book the company already has, plus the present
# value of earning more than the cost of equity on it while that lasts, plus a fading
# remainder. The return is struck on what the ORDINARY holders earn — the perpetual coupon
# ranks ahead of them, so it comes out of the numerator first.
RI_FADE = V['ri_fade']
ROE_ORD = [(NPA_F[i] - HYB_CPN) / EQ_CLOSE[i] for i in range(5)]


def residual_income(ke_r, scale=1.0):
    b0 = EQP0 / 1000.0                                # opening ordinary book, USD mn
    b, pv, det = b0, 0.0, []
    for i in range(5):
        roe_i = ROE_ORD[i] * scale
        ri = (roe_i - ke_r) * b
        df = 1.0 / (1 + ke_r) ** (i + 1)
        det.append(dict(open=b, roe=roe_i, ri=ri, df=df, pv=ri * df))
        pv += ri * df
        b = EQ_CLOSE[i] / 1000.0                      # the model's own roll-forward
    roe_t = ROE_ORD[4] * scale
    ri_t = (roe_t - ke_r) * b
    tv = ri_t / (ke_r + RI_FADE - G)                  # a fading perpetuity, not a growing one
    pv_tv = tv / (1 + ke_r) ** 5
    return dict(b0=b0, det=det, pv_expl=pv, roe_t=roe_t, ri_t=ri_t, tv=tv, pv_tv=pv_tv,
                equity=b0 + pv + pv_tv, close=b)


RI = residual_income(KE)
RI_BEAR = residual_income(KE_CI_HI, 0.85)
RI_BULL = residual_income(KE_CI_LO, 1.15)
BOOK_EQ = RI['equity']
PB_FAIR = BOOK_EQ / (EQP0 / 1000.0)
BOOK_BASE = BOOK_EQ / SH * PEG
# The two bounds are discounted at the two ends of the regressed beta's own 90% confidence
# interval — a HIGHER beta for the low bound, a LOWER beta for the high bound. They must
# never be built on the alternative index construction: that construction carries the LOWER
# beta, so using it as a downside would invert the range.
BOOK_BEAR = RI_BEAR['equity'] / SH * PEG
BOOK_BULL = RI_BULL['equity'] / SH * PEG
VSB_RATIO = V['vessel_sale_price'] / V['vessel_sale_book']

LB = {'dcf': (LN['dcf']['bear'], DC['fv_aed'], LN['dcf']['bull']),
      'relative': (REL_BEAR, REL_BASE, REL_BULL),
      'normalized': (NORM_BEAR, NORM_BASE, NORM_BULL),
      'book': (BOOK_BEAR, BOOK_BASE, BOOK_BULL)}
# [R-LENS-03] THE CENTRAL IS THE CLASS PRIMARY, WHICH FOR THIS CLASS IS THE CASH-FLOW
# LENS. The typed 40/25/20/15 blend is retired; its weights are kept so this sheet can
# show what it read and what retiring it cost, and nothing computes with them.
CENTRAL = LB['dcf'][1]
CENTRAL_A = DA['fv_aed']
CENTRAL_BEAR = LB['dcf'][0]
CENTRAL_BULL = LB['dcf'][2]
RETIRED_BLEND = sum(LW[k] * LB[k][1] for k in LB)

# --- the sum-of-the-parts cross-check -----------------------------------------
SOTP_MULT = {'Integrated Logistics': MULT_CONTR, 'Services': MULT_CONTR,
             'Shipping': BLEND_EV}
SOTP_EV = {g: GRP_EB_F[g][0] * SOTP_MULT[g] for g in GROUPS}
SOTP_EVOPS = sum(SOTP_EV.values())
SOTP_EQ = SOTP_EVOPS + JV_BV - NETDEBT - HYBRID - NCI_BV
SOTP_FV = per_share(SOTP_EQ)

# --- the own multiples --------------------------------------------------------
EV_NOW = MKTCAP + NETDEBT
OWN_EVEB_TTM = EV_NOW / HI['ebitda_reported'][2]
OWN_EVEB_26 = EV_NOW / EB_F[0]
# The peer multiples come from aggregators, which take enterprise value as market
# capitalisation plus net debt and nothing else. This study's own equity bridge deducts the
# perpetual securities and the minorities as well, so the company's multiple on the SAME
# basis as its bridge is a different — and higher — number. Both are shown.
EV_BRIDGE = MKTCAP + NETDEBT + HYBRID + NCI_BV
OWN_EVEB_TTM_BR = EV_BRIDGE / HI['ebitda_reported'][2]
OWN_EVEB_26_BR = EV_BRIDGE / EB_F[0]
OWN_PE_TTM = MKTCAP / (V['npa_fy25'] - V['hybrid_coupon_fy25'])
OWN_PE_FWD = MKTCAP / ORD_F[0]          # the same price on FY2026E ordinary earnings
OWN_PB = MKTCAP / (EQP0 + HYBRID)
OWN_DY = V['dps_2026_usd'] * 1000.0 / MKTCAP

# --- historical statement derivations ----------------------------------------
H_DNA = HI['dna']; H_EBITDA = HI['ebitda_op']; H_REV = HI['revenue']
H_OPCOST = [H_REV[i] - H_EBITDA[i] for i in range(3)]
H_NWC = HB['nwc']; H_ND = HB['net_debt']; H_EQ = HB['equity_parent']
H_NPA = HI['npa']
H_ORD = [H_NPA[0], H_NPA[1], H_NPA[2] - V['hybrid_coupon_fy25']]
H_HYBCPN = [0.0, 0.0, V['hybrid_coupon_fy25']]
H_GROSSD = HB['debt']
H_IC = [HB['ppe'][i] + H_NWC[i] + HB['intangibles'][i] + HB['goodwill'][i] for i in range(3)]

REV_ALL = H_REV + REV_F
EB_ALL = H_EBITDA + EB_F
EBIT_ALL = HI['ebit'] + EBIT_F
DNA_ALL = H_DNA + DNA_F
NPA_ALL = H_NPA + NPA_F
ORD_ALL = H_ORD + ORD_F
HYBCPN_ALL = H_HYBCPN + [HYB_CPN] * 5
ND_ALL = H_ND + ND_CLOSE
EQ_ALL = H_EQ + EQ_CLOSE
NWC_ALL = H_NWC + NWC_F
IC_ALL = H_IC + IC_F
GROSSD_ALL = H_GROSSD + GROSS_D
OPCOST_ALL = H_OPCOST + OPCOST_F
CAPEX_ALL = list(HC_['capex']) + [-x for x in CAPEX]

wb = Workbook()
EXPECT, ANCH = {}, {}


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=48, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula and record the model's own value for the same cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is None:
        raise ValueError(f'formula at {ws.title}!{ad} carries no expected value')
    EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


def note(ws, ad, text):
    put(ws, ad, text, fmt=None).font = SUB


# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — ADNOC Logistics & Services plc (ADX: ADNOCLS)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the ADNOC L&S valuation study. Every blue cell is an',
 'input; every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice. The cost of equity is built from the',
 'normalised risk-free rate, the beta and the equity risk premium — and the normalised rate is itself a',
 'formula, the observed government bond yield less the sovereign default spread. The cost of debt is built',
 'three separate ways on the sheet — the parent facility at the overnight financing rate plus its margin, a',
 'weighted blend of the instruments the company actually has outstanding, and the disclosed third-party',
 'bank-loan midpoint — and the three are averaged in a cell. The discount factors compound year on year off',
 'the cost-of-capital glide, with a three-quarter stub for 2026 because the valuation date is 31 March 2026.',
 'The tanker fleet is built vessel by vessel from the vessel counts and day rates on Assumptions; the balance',
 'sheet rolls property, plant and equipment, working capital, equity and net debt forward; and every ratio and',
 'per-share figure, including the conversion to dirhams at the fixed parity, is a formula.', '',
 'ONLY THREE KINDS OF CELL ARE PASTED VALUES, and it is worth knowing exactly which. Anything else pasted',
 'would be a defect, and the build script carries a sheet-by-sheet register of every pasted cell against these',
 'three so the claim can be checked rather than taken on trust.',
 '  (1) DISCLOSED HISTORY AND GENUINE INDEPENDENT INPUTS — the blue cells: the audited statements, the',
 '      operating-segments note, disclosed vessel counts and day rates, the twelve charter fixtures and their',
 '      dates, the eleven vessels bought on 7 August 2026 and their delivery dates, the balance-sheet anchors,',
 '      the tax rates, the peer multiples with their sources and dates, and the study judgements. Where a line',
 '      is both disclosed and derivable the DISCLOSED figure is carried, because the filing is the record —',
 '      but only ONCE: the same audited figure is never typed in two places, so group revenue is added up',
 '      from the segment note, the audited net debt is gross borrowings less cash, and the intangibles and',
 '      goodwill on the driver sheet point at the balance sheet. Anything on the Assumptions sheet that is',
 '      itself DERIVED is a GREEN formula pointing at the build that produces it — the running cost per',
 '      vessel-day, the gas-carrier day rate, the reported and re-based receivable days, both gross-up',
 '      factors, the opening working capital and the vessel-years the August purchase adds.',
 '  (2) THE OUTPUT OF A UNIT BUILD that would be unreadable flattened into a grid. No cell in this workbook',
 '      now rests on that justification: the block that used to — the historical segment table — is the',
 '      disclosed operating-segments note, which is category (1).',
 '  (3) WHOLE-MODEL RE-RUNS, where each figure is a complete revaluation of the entire model and so cannot be',
 '      a single formula: the Monte Carlo price map, the three sensitivity grids, the discounted-cash-flow',
 '      bear and bull bounds (each of which re-runs the fleet build at a different rate anchor, a different',
 '      beta — the two ends of the regression\'s own confidence interval — and a different capital-expenditure',
 '      path) and the three expert-panel legs.',
 '  THE MONTE CARLO AND SENSITIVITY GRIDS DO NOT REDRAW WHEN A DRIVER IS CHANGED. Changing a blue cell',
 '  reprices the whole valuation chain, but those grids are engine outputs and stay as they were run.', '',
 'WHAT CHANGED IN THIS EDITION. On 7 August 2026 — the anchor date of this study — the company announced the',
 'purchase of eleven vessels for about USD 1.3 billion: six very large crude carriers and three gas carriers',
 'bought secondhand for third-quarter delivery, and two gas carrier newbuildings for the fourth. That price is',
 'inside the market price this valuation is compared against, so it is now in the model: the crude carriers',
 'earn at the implied spot rate from their delivery date, the gas carriers add contracted vessel-years, and',
 'the USD 1.3 billion is carried BOTH in net debt, because the purchase is committed and funded, and in the',
 'opening asset base, where it depreciates and shields tax. The smallest tankers are no longer priced at the',
 'medium-range rate unadjusted — the company disclosed handysize rates down 21% against medium range up 29%,',
 'so the substitution had the sign wrong and the rate is now scaled by that disclosed move. Receivable days',
 'are re-based onto the revenue basis the forecast actually uses rather than carried across from a year with a',
 'different gross-up, and the arithmetic of that re-basing is on the Assumptions sheet in cells. The cost of',
 'debt is labelled for what it is — an average of three constructions, with the one genuinely weighted by the',
 'instruments outstanding shown beside it rather than the average being called weighted. The depreciation rate',
 'used, the rate the company realised in 2025 and the disclosed useful lives now sit together so the choice',
 'can be judged. And the earnings multiple is shown on BOTH bases, forward and trailing, for the peers and for',
 'the company, because an earlier edition compared the company trailing against peers shown forward.', '',
 'AND THE BETA HAS BEEN RE-MEASURED. It is now produced by the same standard routine used for every share',
 'covered on this desk: the market series is resolved from the exchange the share is listed on rather than',
 'chosen, both series are screened for data quality before they are paired, and the weekly returns are',
 "measured on that exchange's own trading week. The series it resolved is the one this study already used, so",
 'WHICH market the share is measured against has not changed — but the grid and the screening do move the',
 f"figure: the slope is now {V['beta']:.4f} on {BE['n']} weekly observations, and its 90% confidence interval",
 f"runs {V['beta_ci_lo']:.3f} to {V['beta_ci_hi']:.3f}, materially WIDER than the interval the previous edition",
 'published. The bear and bull cases take those two bounds directly, so the published range widens with it.',
 'A range that widens on re-measurement is information about how much a three-year price history can settle,',
 'not an embarrassment to be smoothed away.', '',
 'How revenue is built. Not as one growth rate, and not off the published rate either. The company gives one',
 'rate per vessel class each quarter, and its own chief financial officer said on the first-quarter call that',
 'this rate is a BLEND across the whole class, the vessels on charters out included. So the SPOT rate is not',
 'read off that disclosure — it is SOLVED out of it, window by window: the published blend times the class',
 'vessel-days, less the revenue the twelve disclosed fixtures earn over exactly the days their own contracts',
 'run, divided by the vessel-days that actually traded spot. Every step of that derivation is on the Segments',
 'sheet, cell by cell. The forecast then prices each vessel on its own terms — chartered vessels at their',
 'contracted rate for their contracted days, everything else at the implied spot rate — less an all-in running',
 'cost per vessel-day that is itself solved so the same construction reproduces the tanker earnings the company',
 'reported for 2025. The gas carriers are contracted vessel-years times a day rate solved the same way. The',
 'five remaining units are grown on their own revenue and margin drivers, anchored on what each actually',
 'earned in the first quarter of 2026.', '',
 'THE COST OF CAPITAL CARRIES THREE TRANCHES, NOT TWO: equity, debt, and the perpetual capital securities at',
 'their own coupon. Those securities rank ahead of the ordinary shares and are deducted in the equity bridge,',
 'and a claim that is deducted from enterprise value has to be weighted in the rate as well — the two are',
 'halves of one treatment. The coupon is an equity distribution rather than interest, so it is not taxed down.',
 'Every construction in the workbook carries the same three tranches, including the sensitivity re-runs, so',
 'the sensitivity grid is centred on the figure it brackets.', '',
 'THE BOOK LENS IS A RESIDUAL INCOME BUILD. A single-stage justified price-to-book assumes a steady state this',
 'company is not in: it pays out about a third of its earnings and compounds its book above its own cost of',
 'equity, where that formula is undefined rather than merely wrong. The lens is instead the book the company',
 'already has, plus the present value of earning more than the cost of equity on it for five years, plus a',
 'remainder in which that excess fades. The whole ladder is live on the Relative & Normalized sheet, three',
 'times over — base, bear and bull — and the return it is struck on is what the ORDINARY holders earn, after',
 'the perpetual coupon that ranks ahead of them.', '',
 'THE CONTESTED JUDGEMENT, PUBLISHED BOTH WAYS — HOW THE MARKET IS MEASURED. The same weekly regression of the',
 'stock\'s own returns is run against two different measures of its market, and the answer moves a long way',
 'between them. Against the published index of the exchange the share is listed on — the index the method',
 f"asks for, and the primary reading — the beta is {V['beta']:.3f}. Against an equal-weight composite of that same",
 f"exchange's names, which gives its smallest listings the same say as its largest, the beta is",
 f"{V['beta_composite']:.3f}. The published index is weighted by size and is therefore dominated by the very group this",
 'company belongs to; the composite is not. That single difference in construction, and nothing about the',
 'company, moves the cost of equity and the cash-flow value materially. BOTH are carried through the model in',
 'full, side by side, on the Summary, Fundamental Valuation and DCF sheets. They are NOT averaged into one',
 f"number. The regression's own 90% confidence interval on the primary estimate runs {V['beta_ci_lo']:.3f} to "
 f"{V['beta_ci_hi']:.3f}, and",
 'those two bounds — not round numbers chosen by hand — are the betas used in the bear and bull cases, so the',
 'published range is the range the estimate itself supports.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges and distributions.', '',
 'Sourcing note, up front. FY2023, FY2024 and FY2025 come from the company\'s own audited consolidated',
 'financial statements; the first quarter of 2026 comes from the reviewed interim statements; the fleet,',
 'rate and contract data come from the company\'s own investor presentations and earnings calls. Every input',
 'is listed with its value, source and date in the companion bibliography document.', '',
 f"Currency. US dollar thousand unless stated — the company reports in US dollars. Per-share figures are in",
 f"dirhams at the fixed parity of {PEG:.4f} dirhams to the dollar. Spot AED {SPOT:.2f} ({M['price_date']} close).",
 'Sheets: READ FIRST · Summary · Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative &',
 'Normalized · DCF · Income Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo ·',
 'Sensitivity · Per-Share & Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 114

# ============ 2 SUMMARY =======================================================
ws = sheet('Summary')
title(ws, 'Summary — valuation at a glance', 'All values link live to their source sheets. '
      'AED per share unless stated.', 8, awidth=52, cwidth=15)
hdr(ws, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Weight', 'Contribution', 'vs spot',
            'Terminal value share'])
S_SPOT = 16
LENS_ROWS = {'dcf': 5, 'relative': 6, 'normalized': 7, 'book': 8}
LENS_LABEL = {'dcf': 'Discounted cash flow (own regressed beta)',
              'relative': 'Relative multiples',
              'normalized': 'Normalised earnings power',
              'book': 'Book value and sustainable return'}
LENS_BASE_SRC = {'dcf': '=DCF!$C$%d', 'relative': "='Relative & Normalized'!$C$%d",
                 'normalized': "='Relative & Normalized'!$C$%d",
                 'book': "='Relative & Normalized'!$C$%d"}
SUMMARY_LENS_SRC = {}    # filled once the source sheets know their own row numbers

# ============ 3 FUNDAMENTAL VALUATION =========================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — the four lenses and the contested judgement', None, 6,
      awidth=58, cwidth=16)

# ============ 4 ASSUMPTIONS ====================================================
ws = sheet('Assumptions')
title(ws, 'Assumptions — every input in the model', 'Blue cells are inputs. Change one and '
      'the model reprices: everything downstream is a formula.', 8, awidth=62, cwidth=14)
r = 4
A = {}


def esc_chain(i):
    """The house ladder compounded to year i, as a LIVE product of the rate cells.

    [R-MACRO-01] A single rate raised to a power is not the ladder, and the model applies
    the ladder. Written as a chain rather than as a pasted index because this sheet's own
    note says nothing on it is a pasted result, and an index row would be one.
    """
    return '*'.join("(1+%s)" % a('esc%d' % k) for k in range(i + 1))


def block(name, items, cols=None):
    global r
    band(ws, r, 8); put(ws, f'A{r}', name, bold=True, fmt=None)
    if cols:
        for i, c in enumerate(cols):
            cc = ws.cell(row=r, column=2 + i, value=c)
            cc.font = Font(bold=True); cc.fill = FILL_G
    r += 1
    for key, lab, val, fmt in items:
        put(ws, f'A{r}', lab, fmt=None)
        if isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(ws, f'{get_column_letter(2+i)}{r}', v, BLUE if isinstance(v, (int, float))
                    else BLACK, fmt)
        else:
            put(ws, f'C{r}', val, BLUE, fmt)
        A[key] = r
        r += 1
    r += 1


def a(key, i=None, col=None):
    """Absolute reference to an Assumptions cell; i or col selects a list column."""
    c = col if col else (get_column_letter(2 + i) if i is not None else 'C')
    return f"Assumptions!${c}${A[key]}"


block('Market and share anchors', [
    ('spot', 'Share price (AED, Abu Dhabi Securities Exchange close)', SPOT, PX),
    ('shares', 'Shares outstanding (mn)', SH, NUM1),
    ('fx', 'Dirhams per US dollar (fixed parity, unchanged since 1997)', PEG, DF4),
    ('stub', 'Stub year fraction — 31 March 2026 valuation date to 31 December 2026',
     STUB, '0.00'),
    ('q1fcf', 'First-quarter 2026 free cash flow, already inside net debt at the valuation '
     'date (USD 000)', Q1FCF, NUM0)])
block('Cost of capital', [
    ('rf_obs', 'Observed government bond yield (dirham tranche maturing January 2031)',
     V['rf_observed'], PCT2),
    ('sov', 'Sovereign default spread (netted out of the risk-free rate)', V['sov_spread'],
     PCT2),
    ('beta', 'Beta — own-stock weekly regression against the published index of its own '
     'exchange', V['beta'], BETA),
    ('beta_a', 'Beta — the same regression against an equal-weight composite of that '
     'exchange\'s names (the disclosed alternative)', V['beta_composite'], BETA),
    ('beta_ci_lo', 'Beta — lower bound of the regression\'s 90% confidence interval (the '
     'bull-case beta)', V['beta_ci_lo'], BETA),
    ('beta_ci_hi', 'Beta — upper bound of the regression\'s 90% confidence interval (the '
     'bear-case beta)', V['beta_ci_hi'], BETA),
    ('beta_blume', 'Beta — the measured slope shrunk toward the market: two-thirds of it '
     'plus one-third of 1.0', V['beta_blume'], BETA),
    ('erp', 'Equity risk premium (mature premium plus country risk)', V['erp_total'], PCT2),
    ('rf_term', 'Terminal risk-free rate', V['rf_terminal'], PCT2),
    ('tax_stat', 'Statutory corporate tax rate', TAXS, PCT)])
block('Cost of debt — the evidence behind the three constructions', [
    ('sofr', 'Secured overnight financing rate', V['sofr'], PCT2),
    ('shldr_m', 'Parent revolving credit facility margin', V['shldr_margin'], PCT2),
    ('bank_lo', 'Third-party bank loans — low end of the disclosed range', V['bank_loan_lo'],
     PCT2),
    ('bank_hi', 'Third-party bank loans — high end of the disclosed range', V['bank_loan_hi'],
     PCT2),
    ('oth_lo', 'Other third-party borrowings — low end of the disclosed range',
     V['other_borr_lo'], PCT2),
    ('oth_hi', 'Other third-party borrowings — high end of the disclosed range',
     V['other_borr_hi'], PCT2),
    ('lease_int', 'Lease interest charged in 2025 (USD 000)', V['intpaid_lease_fy25'], NUM0),
    ('lease_open', 'Lease liabilities, opening balance 2025 (USD 000)', V['lease_open_fy25'],
     NUM0),
    ('lease_close', 'Lease liabilities, closing balance 2025 (USD 000)',
     V['lease_close_fy25'], NUM0),
    ('d_shldr', 'Shareholder loan at 31 March 2026 (USD 000)', V['q1_26_shldr_loan'], NUM0),
    ('d_borr', 'Third-party borrowings at 31 March 2026 (USD 000)', V['q1_26_borrowings'],
     NUM0),
    ('d_lease', 'Lease liabilities at 31 March 2026 (USD 000)', V['q1_26_leases'], NUM0)])
block('Tanker fleet — vessel counts by class',
      [('tnk_own25', 'Vessels owned at 31 December 2025', [OWNED25[c] for c in CLS], NUM0),
       ('tnk_sold', 'Less vessels sold between the year end and the valuation date',
        [SOLD[c] for c in CLS], NUM0)],
      cols=CLS_NAME)
block('Tanker rates — the published blended rate by class and quarter (USD per day). It is '
      'a BLEND across the whole class, charters out included; the spot rate is solved out '
      'of it on the Segments sheet, never read off it',
      [('b24_lr1', 'Long range 1 — 2024', Q24['lr1'], NUM0),
       ('b24_lr2', 'Long range 2 — 2024', Q24['lr2'], NUM0),
       ('b24_vlcc', 'Very large crude carrier — 2024', Q24['vlcc'], NUM0),
       ('b25_mr', 'Medium range — 2025', Q25['mr'], NUM0),
       ('b25_lr1', 'Long range 1 — 2025', Q25['lr1'], NUM0),
       ('b25_lr2', 'Long range 2 — 2025', Q25['lr2'], NUM0),
       ('b25_vlcc', 'Very large crude carrier — 2025', Q25['vlcc'], NUM0)],
      cols=['First quarter', 'Second quarter', 'Third quarter', 'Fourth quarter'])
block('Tanker rates — the published blended rate by class, the two disclosed quarters of '
      '2026 (USD per day)',
      [('b26_mr', 'Medium range', B26['mr'], NUM0),
       ('b26_lr1', 'Long range 1', B26['lr1'], NUM0),
       ('b26_lr2', 'Long range 2', B26['lr2'], NUM0),
       ('b26_vlcc', 'Very large crude carrier', B26['vlcc'], NUM0)],
      cols=['First quarter 2026', 'Second quarter 2026'])
block('The smallest tankers, which the rate disclosure does not break out', [
    ('hs_rel', 'Handysize rate as a proportion of the medium-range rate — the company '
     'disclosed Handysize DOWN 21% against medium range UP 29%, so the two smallest '
     'classes moved in OPPOSITE directions and the medium-range rate cannot stand in for '
     'the smallest unadjusted. It scales the medium-range rate in every window and on both '
     'sides of the mid-cycle average', HS_REL, '0.00')])
block('Tanker charters out — the twelve fixtures as disclosed, each at its own rate for '
      'exactly the days its own contract runs',
      [(f'ch{k}', f"{ch['name']} — {CLS_NAME[CLS.index(ch['klass'])].lower()}, fixed for "
        f"{ch['period']} months",
        [ch['rate'], _ser(ch['start']), _ser(ch['end'])], NUM0)
       for k, ch in enumerate(CHARTERS)],
      cols=['Rate (USD per day)', 'Contract begins', 'Contract ends'])
DATEFMT = 'yyyy-mm-dd'
band(ws, r, 8)
put(ws, f'A{r}', 'Rate windows and forecast years — the calendar the fleet build runs on '
    '(dates)', bold=True, fmt=None)
for _i, _c in enumerate(WIN_LAB[:6]):
    _cc = ws.cell(row=r, column=2 + _i, value=_c)
    _cc.font = Font(bold=True); _cc.fill = FILL_G
r += 1
A['win_start'] = r
put(ws, f'A{r}', 'Rate window begins', fmt=None)
for _w in range(6):
    put(ws, f'{WCOL[_w]}{r}', _ser(WIN[_w][0]), BLUE, DATEFMT)
r += 1
A['win_end'] = r
put(ws, f'A{r}', 'Rate window ends — each window closes where the next one opens; the last '
    'is the disclosed quarter end', fmt=None)
for _w in range(5):
    putf(ws, f'{WCOL[_w]}{r}', f'={WCOL[_w+1]}{A["win_start"]}', _ser(WIN[_w][1]), DATEFMT)
put(ws, f'G{r}', _ser(WIN[5][1]), BLUE, DATEFMT)
r += 2
band(ws, r, 8)
put(ws, f'A{r}', 'Forecast years (dates)', bold=True, fmt=None)
for _i, _y in enumerate(YF):
    _cc = ws.cell(row=r, column=2 + _i, value=_y)
    _cc.font = Font(bold=True); _cc.fill = FILL_G
r += 1
A['yr_start'] = r
put(ws, f'A{r}', 'Forecast year begins', fmt=None)
putf(ws, f'B{r}', f'=F{A["win_start"]}', _ser(YRB[0][0]), DATEFMT)
for _i in range(1, 5):
    put(ws, f'{CD[_i]}{r}', _ser(YRB[_i][0]), BLUE, DATEFMT)
r += 1
A['yr_end'] = r
put(ws, f'A{r}', 'Forecast year ends — each year closes where the next one opens', fmt=None)
for _i in range(4):
    putf(ws, f'{CD[_i]}{r}', f'={CD[_i+1]}{A["yr_start"]}', _ser(YRB[_i][1]), DATEFMT)
put(ws, f'F{r}', _ser(YRB[4][1]), BLUE, DATEFMT)
r += 2

# ---- the fleet purchase announced on the anchor date --------------------------
band(ws, r, 8)
put(ws, f'A{r}', 'THE FLEET PURCHASE ANNOUNCED 7 AUGUST 2026 — eleven vessels for about '
    'USD 1.3 billion, announced on the anchor date of this study and therefore already '
    'inside the market price the valuation is compared against', bold=True, fmt=None)
for _i, _c in enumerate(['Vessels', 'Delivery date']):
    _cc = ws.cell(row=r, column=2 + _i, value=_c)
    _cc.font = Font(bold=True); _cc.fill = FILL_G
r += 1
A['acq_cost'] = r
put(ws, f'A{r}', 'Purchase price — added BOTH to net debt, because it is committed and '
    'funded, and to the opening asset base, so it depreciates and shields tax (USD 000)',
    fmt=None)
put(ws, f'C{r}', ACQ_COST, BLUE, NUM0)
r += 1
A['acq_vlcc'] = r
put(ws, f'A{r}', 'Very large crude carriers bought secondhand — they join the spot fleet '
    'on delivery and earn at the implied spot rate from that date', fmt=None)
put(ws, f'B{r}', ACQ_VLCC_N, BLUE, NUM0)
put(ws, f'C{r}', _ser(ACQ_SECONDHAND), BLUE, DATEFMT)
r += 1
A['acq_gas'] = r
put(ws, f'A{r}', 'Gas carriers acquired in total — they add contracted vessel-years to the '
    'gas fleet', fmt=None)
put(ws, f'B{r}', ACQ_GAS_N, BLUE, NUM0)
r += 1
A['acq_gas_sh'] = r
put(ws, f'A{r}', 'Of which bought secondhand, delivering in the third quarter', fmt=None)
put(ws, f'B{r}', ACQ_GAS_SECONDHAND_N, BLUE, NUM0)
put(ws, f'C{r}', _ser(ACQ_SECONDHAND), BLUE, DATEFMT)
r += 1
A['acq_gas_nb'] = r
put(ws, f'A{r}', 'Of which newbuildings resold from the yard, delivering in the fourth '
    'quarter', fmt=None)
putf(ws, f'B{r}', f"=B{A['acq_gas']}-B{A['acq_gas_sh']}", ACQ_GAS_NEWBUILD_N, NUM0)
put(ws, f'C{r}', _ser(ACQ_NEWBUILD), BLUE, DATEFMT)
r += 1
A['acq_total'] = r
put(ws, f'A{r}', 'Vessels acquired in total', bold=True, fmt=None)
putf(ws, f'B{r}', f"=B{A['acq_vlcc']}+B{A['acq_gas']}", ACQ_VLCC_N + ACQ_GAS_N, NUM0,
     bold=True)
r += 2

block('Tanker fleet — rate path and running cost', [
    ('h2w', 'Weight on the 2025 implied spot rate in setting the second half of 2026', H2W,
     PCT),
    ('opex_day', 'All-in running cost per vessel per day (USD)', OPEX_DAY, NUM1),
    ('esc0', 'House inflation ladder, FY2026', ESC_PATH[0], PCT),
    ('esc1', 'House inflation ladder, FY2027', ESC_PATH[1], PCT),
    ('esc2', 'House inflation ladder, FY2028', ESC_PATH[2], PCT),
    ('esc3', 'House inflation ladder, FY2029', ESC_PATH[3], PCT),
    ('esc4', 'House inflation ladder, FY2030', ESC_PATH[4], PCT),
    ('grossup', 'Gross-up from time-charter-equivalent revenue to reported revenue',
     GROSSUP, '0.00')])
block('Gas carriers', [
    ('gas_vy25', 'Consolidated gas vessels in service through 2025 (vessel-years)', GAS_VY25,
     NUM1),
    ('gas_rate', 'Gas carriers — implied revenue per vessel-day (USD)', GAS_RATE, NUM0),
    ('gas_mgn', 'Gas carriers — earnings margin', GAS_MGN, PCT),
    ('jv_gas', 'Share of joint-venture profit carried in the disclosed 2025 Gas Carriers '
     'earnings (USD 000)', JV_GAS, NUM0),
    ('jv_serv', 'Share of joint-venture profit carried in the disclosed 2025 Services '
     'earnings (USD 000)', JV_SERV, NUM0)])
block('Gas-carrier contract table — the vessel-years the fleet is contracted for, before '
      'and after the August 2026 purchase',
      [('gas_vy_base', 'Contracted vessel-years before that purchase — the disclosed '
        'contract table', GAS_VY_BASE, NUM1),
       ('gas_vy_acq', 'Vessel-years the gas carriers bought in August 2026 add — each '
        'tranche from its own delivery date', GAS_VY, NUM1),
       ('gas_vy', 'Gas carriers — contracted vessel-years', GAS_VY, NUM1)], cols=YF)
for _i in range(5):
    _c = CD[_i]
    _ye = a('yr_end', col=_c); _ys = a('yr_start', col=_c)
    putf(ws, f"{_c}{A['gas_vy_acq']}",
         f"=({a('acq_gas_sh', col='B')}*MAX({_ye}-MAX({_ys},{a('acq_gas_sh', col='C')}),0)"
         f"+{a('acq_gas_nb', col='B')}*MAX({_ye}-MAX({_ys},{a('acq_gas_nb', col='C')}),0))"
         f"/({_ye}-{_ys})", GAS_ACQ_VY[_i], NUM1)
    putf(ws, f"{_c}{A['gas_vy']}",
         f"={_c}{A['gas_vy_base']}+{_c}{A['gas_vy_acq']}", GAS_VY[_i], NUM1, bold=True)
_items = []
for s in SEGS:
    if s in ('Tankers', 'Gas Carriers'):
        continue
    k = s.lower().replace(' ', '_').replace('-', '_')
    _items.append((f'rev_{k}', f'{s} — revenue (USD 000)', DRV[s]['rev'], NUM0))
    _items.append((f'mar_{k}', f'{s} — earnings margin', DRV[s]['mar'], PCT))
block('The remaining five units — revenue and margin drivers', _items, cols=YF)
block('Capital expenditure, depreciation and working capital',
      [('capex', 'Capital expenditure (USD 000)', CAPEX, NUM0),
       ('dep_rate', 'Depreciation rate on property, plant and equipment', DEP_RATE, PCT2),
       ('other_dna', 'Other depreciation and amortisation, 2026 run rate (USD 000)',
        OTHER_DNA, NUM0),
       ('dso_rep', 'Days sales outstanding measured on REPORTED 2025 revenue', DSO_REPORTED,
        NUM1),
       ('gu25', 'Gross-up inside 2025 reported revenue — reported tanker revenue over the '
        'same fleet\'s charter-equivalent revenue', GROSSUP25, '0.00'),
       ('gu26', 'Gross-up the forecast is built at, from 2026 onward', GROSSUP, '0.00'),
       ('dso', 'Days sales outstanding — the reported ratio RE-BASED onto the revenue '
        'basis the forecast actually uses', DSO, NUM1),
       ('dio', 'Days inventory outstanding', DIO, NUM1),
       ('dpo', 'Days payable outstanding', DPO, NUM1),
       ('nwc25', 'Net working capital at 31 December 2025 (USD 000)', NWC25, NUM0)],
      cols=YF)
# ---- depreciation: the rate used, the rate realised and the disclosed lives ----
band(ws, r, 8)
put(ws, f'A{r}', 'DEPRECIATION — THE RATE USED, THE RATE REALISED IN 2025 AND THE '
    'DISCLOSED USEFUL LIVES, SO THE CHOICE CAN BE JUDGED. These three are memoranda: the '
    'rate the model runs on is the blue driver above, and it is sensitised', bold=True,
    fmt=None)
_cc = ws.cell(row=r, column=2, value='Memorandum')
_cc.font = Font(bold=True); _cc.fill = FILL_G
r += 1
A['dep_used'] = r
put(ws, f'A{r}', 'The rate used — first-quarter 2026 depreciation annualised over that '
    'quarter\'s average balance (the driver above)', fmt=None)
putf(ws, f'B{r}', f"=C{A['dep_rate']}", DEP_RATE, PCT2, green=True)
r += 1
A['dep_realised'] = r
put(ws, f'A{r}', 'The rate realised in 2025 — depreciation charged on property, plant and '
    'equipment over the average balance for that year', fmt=None)
put(ws, f'B{r}', V['dep_rate_realised_fy25'], BLUE, PCT2)
r += 1
A['dep_life'] = r
put(ws, f'A{r}', 'Disclosed useful life of tankers (years, straight line) — dry-bulk and '
    'containers 25, gas carriers 25 to 40, offshore vessels 20 to 25, jack-up barges 40, '
    'and dry-docking components 2 to 5, which is why both rates above are heavier than a '
    'hull life alone implies', fmt=None)
put(ws, f'B{r}', V['life_tankers'], BLUE, NUM0)
r += 1
A['dep_implied_life'] = r
put(ws, f'A{r}', 'Average life implied by the rate used (years)', fmt=None)
putf(ws, f'B{r}', f"=1/C{A['dep_rate']}", 1.0 / DEP_RATE, NUM1)
r += 2
block('Tax by business unit and the 2025 depreciation allocation basis', [
    ('tax_il', 'Integrated Logistics — income tax rate', TAX_G['Integrated Logistics'], PCT),
    ('tax_ship', 'Shipping — income tax rate', TAX_G['Shipping'], PCT),
    ('tax_serv', 'Services — income tax rate', TAX_G['Services'], PCT)]
    + [('dna_' + s.lower().replace(' ', '_').replace('-', '_'),
        f'{s} — 2025 depreciation and amortisation (USD 000)', SEG_DNA25[s], NUM0)
       for s in SEGS])
block('Funding, distributions and the bridge', [
    ('nd_co', 'Net debt at 31 March 2026, company basis (USD 000)', NDCO, NUM0),
    ('deferred', 'Deferred consideration on acquisitions (USD 000)', DEFERRED, NUM0),
    ('hybrid', 'Perpetual capital securities at carrying value (USD 000)', HYBRID, NUM0),
    ('hyb_m', 'Perpetual capital securities margin over the overnight rate',
     V['hybrid_margin'], PCT2),
    ('nci_bv', 'Non-controlling interests at carrying value (USD 000)', NCI_BV, NUM0),
    ('nci_nav', 'Of which arose on the tanker combination — the 20% contracted for purchase '
     'in mid-2027 (USD 000)', NCI_NAV, NUM0),
    ('nci_sh', 'Non-controlling interests\' share of profit', NCI_SHARE, PCT),
    ('jv', 'Joint ventures and associates at carrying value (USD 000)', JV_BV, NUM0),
    ('eqp0', 'Equity attributable to shareholders at 31 March 2026 (USD 000)', EQP0, NUM0),
    ('cash', 'Cash and cash equivalents held (USD 000)', CASH, NUM0),
    ('intang', 'Intangible assets (USD 000)', INTANG, NUM0),
    ('gw', 'Goodwill (USD 000)', GW, NUM0),
    ('dps26', 'Ordinary dividend declared for 2026 (USD 000)', DPS[0], NUM0),
    ('div_g', 'Ordinary dividend growth', V['div_growth'], PCT),
    ('g_term', 'Terminal growth', G, PCT)])
block('Lens weights and the multiple blend', [
    ('spot_w', 'Share of 2026 earnings exposed to spot rates', SPOT_W, PCT),
    ('ri_fade', 'Rate at which the return above the cost of equity fades beyond the '
     'forecast (the book lens)', RI_FADE, PCT),
    ('w_eveb', 'Weight on the enterprise multiple within the relative lens', W_EVEB, PCT),
    ('w_dcf', 'Weight — discounted cash flow', LW['dcf'], PCT),
    ('w_rel', 'Weight — relative multiples', LW['relative'], PCT),
    ('w_norm', 'Weight — normalised earnings power', LW['normalized'], PCT),
    ('w_book', 'Weight — book value and sustainable return', LW['book'], PCT)])
ASSUMPTIONS_LAST = r

# ---- fixed row plans, so every sheet can reference every other ---------------
SEGREF = {s: i for i, s in enumerate(SEGS)}
UNITS = [s for s in SEGS if s not in ('Tankers', 'Gas Carriers')]
# Services carries a joint-venture removal of its own, so it takes four rows, not two
UNIT_N = sum(4 if s == 'Services' else 2 for s in UNITS)

# Segments — allocated in order rather than hand-numbered, because the unit build below is
# long enough that a hand-kept map drifts the moment a row is inserted.
SG = {}
_sgr = 4


def _sg(key, n=1, gap=1):
    global _sgr
    SG[key] = _sgr
    _sgr += n + gap
    return SG[key]


_sg('revh', 1, 0); _sg('revh0', 7, 0); _sg('revht')
_sg('ebh', 1, 0); _sg('ebh0', 7, 0); _sg('ebht', 1, 0); _sg('mgnh')
_sg('tband', 1, 0); _sg('own25', 1, 0); _sg('sold', 1, 0); _sg('own', 1, 0)
_sg('acqn', 1, 0); _sg('acqdate')
_sg('winb', 1, 0); _sg('winst', 1, 0); _sg('winen', 1, 0); _sg('windy')
_sg('blendb', 1, 0); _sg('blend0', 5)
_sg('cdb', 1, 0); _sg('cd0', 5)
_sg('chdb', 1, 0); _sg('chd0', 5)
_sg('chrb', 1, 0); _sg('chr0', 5)
_sg('sdb', 1, 0); _sg('sd0', 5)
_sg('spb', 1, 0); _sg('sp0', 5)
_sg('rateb', 1, 0); _sg('sp25', 1, 0); _sg('spmid', 1, 0); _sg('spq1', 1, 0)
_sg('spq2', 1, 0); _sg('sph2', 1, 0); _sg('spy26')
_sg('pathb', 1, 0); _sg('path0', 5)
_sg('yrb', 1, 0); _sg('yrst', 1, 0); _sg('yren', 1, 0); _sg('yrdy')
_sg('ycdb', 1, 0); _sg('ycd0', 5)
_sg('ycrb', 1, 0); _sg('ycr0', 5)
_sg('yacdb', 1, 0); _sg('yacd0', 5)
_sg('ysdb', 1, 0); _sg('ysd0', 5)
_sg('ysrb', 1, 0); _sg('ysr0', 5)
_sg('chrevt', 1, 0); _sg('sprevt', 1, 0); _sg('tcerev')
_sg('opxb', 1, 0); _sg('vdays25', 1, 0); _sg('tcerev25', 1, 0); _sg('teb25', 1, 0)
_sg('opexd0')
_sg('opexd', 1, 0); _sg('opex', 1, 0); _sg('teb', 1, 0); _sg('gross', 1, 0); _sg('trev')
_sg('gasb', 1, 0); _sg('gasvy25', 1, 0); _sg('gasrev25', 1, 0); _sg('gasrate0')
_sg('gasvyb', 1, 0); _sg('gasvya', 1, 0)
_sg('gasvy', 1, 0); _sg('gasrate', 1, 0); _sg('gasrev', 1, 0); _sg('gasmgn', 1, 0)
_sg('gasgeb', 1, 0); _sg('gasjv', 1, 0); _sg('gaseb')
_sg('unitb', 1, 0); _sg('unit0', UNIT_N)
_sg('frevb', 1, 0); _sg('frev0', 7, 0); _sg('frevt')
_sg('febb', 1, 0); _sg('feb0', 7, 0); _sg('febt', 1, 0); _sg('fmgn')
_sg('grpb', 1, 0); _sg('grev0', 3, 0); _sg('geb0', 3, 0); _sg('gmgn0', 3, 1)
# DCF
DF_ = dict(rev=5, ebitda=6, mgn=7, dna=8, ebit=9, tax=10, nopat=11, adddna=12,
           capex=13, dnwc=14, fcff=15, q1=16, fcfd=17, glide=18, df=19, pv=20,
           taxb=22, geb0=23, gdna0=26, gtax0=29, gtaxc0=32, taxtot=35, taxrate=36,
           tvb=38, g=39, ic=40, roic=41, reinv=42, nopat1=43, tv=44, pvex=45,
           pvtv=46, evops=47, tvshare=48, jv=49, ev=50, nd=51, defd=52, acq=53, hyb=54,
           prenci=55, ncinav=56, nciother=57, ncishare=58, nci=59, eq=60, fvusd=61,
           fvaed=62,
           keb=63, rfobs=64, sov=65, rfstar=66, beta=67, erp=68, ke=69,
           kdb=71, sofr=72, shldrm=73, kd1=74, banklo=75, bankhi=76, bankmid=77,
           othlo=78, othhi=79, othmid=80, tp=81, leaseint=82, leaseopen=83,
           leaseclose=84, kdlease=85, dshldr=86, dborr=87, dlease=88, dtot=89,
           kd2=90, kd3=91, kd=92, kdbal=93, taxstat=94, kdat=95,
           wb=96, mktcap=97, borr=98, hybcap=99, captot=100, we=101, wd=102, whyb=103,
           kh=104, wacc=105, rfterm=106, keterm=107,
           kdterm=108, kdtermat=109, khterm=110, waccterm=111,
           ab=113, betaa=114, kea=115, keta=116, wacca=117, wactermsa=118,
           cib=120, cilo=121, cihi=122, kecilo=123, kecihi=124, blume=125, keblume=126,
           ahdr=128, glidea=129, dfa=130, pva=131, pvexa=132, tva=133, pvtva=134,
           evopsa=135, tvsharea=136, eva=137, prencia=138, ncia=139, eqa=140, fvaeda=141)
# Income statement
IS = dict(rev=5, dc=6, gp=7, ga=8, ecl=9, oi=10, oe=11, op=12, dna=13, ebitda=14,
          ebjv=15, ebrep=16, opcost=17, mgn=18, assoc=19, bargain=20, prevheld=21,
          fininc=22, fincost=23, pbt=24, tax=25, pat=26, nci=27, npa=28, hybcpn=29,
          ordn=30, eps=31, epsaed=32, epspre=33)
# Balance sheet
BS = dict(ppe=5, rou=6, intang=7, gw=8, invprop=9, jv=10, inv=11, recv=12, cash=13,
          ta=14, pay=15, nwc=16, grossd=17, nd=18, hyb=19, nci=20, eqp=21, teq=22,
          ndeb=23, bvps=24, bvpsaed=25, ic=26, roic=27, roe=28,
          ppeb=30, ppeopen=31, ppecapex=32, ppeacq=33, ppedeprate=34, ppedep1=35,
          ppedep=36, ppeclose=37, otherdna=38, dnatot=39,
          wcb=40, wcrev=41, wcopcost=42, wcdso=43, wcdio=44, wcdpo=45, wcrecv=46,
          wcinv=47, wcpay=48, wcnwc=49, wcdnwc=50,
          ndb=52, ndopen=53, ndgross=54, ndint=55, ndfcff=56, ndintat=57, ndfi=58,
          ndcpn=59, ndfcfe=60, nddps=61, ndclose=62,
          eqb=64, eqopen=65, eqnpa=66, eqdps=67, eqcpn=68, eqclose=69, dpsps=70)
# Cash flow
CF = dict(ebitda=5, ocf=6, capex=7, fcf=8, wfb=10, nopat=11, dna=12, cap=13, dnwc=14,
          fcff=15, intat=16, fi=17, cpn=18, fcfe=19, dps=20, ndmove=21, conv=22)
# Relative & Normalized — allocated, because the book lens is now a full residual-income
# ladder run three times rather than a single justified-multiple line
RN = {}
_rnr = 4


def _rn(key, n=1, gap=0):
    global _rnr
    RN[key] = _rnr
    _rnr += n + gap
    return RN[key]


for _k in ('hdr', 'eb26', 'blend', 'ev', 'jv', 'nd', 'defd', 'acq', 'hyb', 'pre', 'nci',
           'eq', 'vev', 'pe', 'ord26', 'vpe', 'w', 'base', 'bearev', 'bearpre', 'bearnci',
           'beareq'):
    _rn(_k)
_rn('bear', 1, 1)
for _k in ('ownb', 'spotusd', 'mktcap', 'netdebt', 'evnow', 'eveb_ttm', 'eveb_26',
           'pe_ttm', 'pe_fwd', 'pb', 'dy', 'evbr', 'ebbr_ttm'):
    _rn(_k)
_rn('ebbr_26', 1, 1)
for _k in ('nhdr', 'neb', 'nev', 'npre', 'nnci', 'neq', 'nvev', 'nord', 'neps', 'nvpe',
           'nbase', 'nbearev', 'nbearpre', 'nbearnci', 'nbeareq'):
    _rn(_k)
_rn('nbear', 1, 1)
for _k in ('bhdr', 'beqp', 'bbvps', 'bbvpsaed', 'broe', 'bke', 'bg', 'bfade'):
    _rn(_k)
_rn('bladder', 1, 0)
for _k in ('bopen', 'broey', 'bri', 'bdf', 'bpv', 'bpvsum', 'brit', 'btv', 'bpvtv', 'beq',
           'bpb'):
    _rn(_k)
_rn('bbase', 1, 1)
_rn('xladder', 1, 0)
for _k in ('xroe', 'xri', 'xdf', 'xpv', 'xpvsum', 'xrit', 'xtv', 'xpvtv', 'xeq'):
    _rn(_k)
_rn('yladder', 1, 0)
for _k in ('yroe', 'yri', 'ydf', 'ypv', 'ypvsum', 'yrit', 'ytv', 'ypvtv', 'yeq'):
    _rn(_k)
_rn('bbear', 1, 1)
for _k in ('vsb', 'vsbook', 'vsprice', 'vsratio', 'vsgain'):
    _rn(_k)
_rn('vsnote')
# SOTP bridge
SB = dict(hdr=4, pvex=5, pvtv=6, evops=7, tvshare=8, jv=9, ev=10, nd=11, defd=12, acq=13,
          hyb=14, prenci=15, nci=16, eq=17, fvusd=18, fvaed=19,
          legb=21, leg0=22, legt=25, mb=27, mcon=28, mspot=29, mw=30, mship=31,
          bb=33, bevops=34, bjv=35, bev=36, bnd=37, bdefd=38, bacq=39, bhyb=40, bnci=41,
          beq=42, bfv=43)
# Summary
SU = dict(hdr=4, dcf=5, rel=6, norm=7, book=8, central=9, cb=11, dcfa=12, centrala=13,
          panel=15, spot=16, keyhdr=18, key0=19)
# Fundamental valuation
FV = dict(hdr=4, dcf=5, dcfbear=6, dcfbull=7, rel=8, norm=9, book=10, central=12,
          cb=14, beta=15, ke=16, wacc=17, fv=18, cen=19, betaa=20, kea=21, wacca=22,
          fva=23, cena=24, cilo=25, cihi=26, note=27, eb=29, ehdr=30, e0=31, epanel=34)
# Per-share & ratios
PS = dict(eps=5, epsaed=6, ordps=7, bvps=8, fcffps=9, dpsps=10, payout=11, gm=12,
          ebm=13, ebitm=14, netm=15, roe=16, roic=17, ndeb=18, cover=19, dso=20,
          dio=21, dpo=22, cycle=23, capexrev=24,
          ab=26, aprice=27, apriceusd=28, amkt=29, aev=30, aeveb=31, aeveb26=32,
          ape=33, apb=34, ady=35)
# Peer & sector
PR = dict(hdr=4, p0=5, mb=9, mcon=10, mspot=11, mw=12, mev=13, pecon=14, pespot=15,
          pe=16, pecon_t=17, pespot_t=18, pe_t=19, ob=21, o0=22)
# Monte Carlo
MC = dict(hdr=4, h0=5, lhdr=8, l0=9, ehdr=14, e0=15)
# Sensitivity
SE = dict(bgb=4, bghdr=5, bg0=6, ab=12, ahdr=13, a0=14, aswing=15, cb=17, chdr=18,
          c0=19, cswing=20, tb=22, thdr=23, t0=24, tnote=25,
          mb=27, m1y=28, mspot=29, mhdr=30, mpath=31, mvs=32, mob=33, mnote=34)

# ============ 5 SOTP BRIDGE ====================================================
ws = sheet('SOTP Bridge')
title(ws, 'Enterprise value to equity — the bridge, and the sum-of-the-parts cross-check',
      'USD thousand unless stated. Per-share figures in dirhams at the fixed parity.', 6,
      awidth=58, cwidth=17)
hdr(ws, SB['hdr'], ['Step', '', 'USD 000', 'AED per share'])
_bridge = [
    (SB['pvex'], 'Present value of the five forecast years', f"=DCF!$C${DF_['pvex']}",
     DC['pv_expl'], True),
    (SB['pvtv'], 'Present value of the terminal value', f"=DCF!$C${DF_['pvtv']}",
     DC['pv_tv'], True),
    (SB['evops'], 'Enterprise value of operations', f"=C{SB['pvex']}+C{SB['pvtv']}",
     DC['ev_ops'], False),
    (SB['jv'], 'Plus joint ventures and associates at carrying value', f"={a('jv')}",
     JV_BV, True),
    (SB['ev'], 'Enterprise value', f"=C{SB['evops']}+C{SB['jv']}", DC['ev'], False),
    (SB['nd'], 'Less net debt at 31 March 2026', f"=-{a('nd_co')}", -NDCO, True),
    (SB['defd'], 'Less deferred consideration on acquisitions', f"=-{a('deferred')}",
     -DEFERRED, True),
    (SB['acq'], 'Less the eleven vessels bought on 7 August 2026, at the announced price',
     f"=-{a('acq_cost')}", -ACQ_COST, True),
    (SB['hyb'], 'Less perpetual capital securities at carrying value', f"=-{a('hybrid')}",
     -HYBRID, True),
    (SB['prenci'], 'Equity value before the minorities',
     f"=C{SB['ev']}+C{SB['nd']}+C{SB['defd']}+C{SB['acq']}+C{SB['hyb']}", DC['pre_nci'],
     False),
    (SB['nci'], 'Less non-controlling interests — the contracted slice at its contracted '
     'price, the rest at the greater of book and value',
     f"=DCF!$C${DF_['nci']}", -DC['nci'], True),
    (SB['eq'], 'Equity attributable to ordinary shareholders',
     f"=C{SB['prenci']}+C{SB['nci']}", DC['equity'], False)]
for rw, lab, fml, xp, gr in _bridge:
    put(ws, f'A{rw}', lab, fmt=None)
    bd = rw in (SB['evops'], SB['ev'], SB['eq'])
    putf(ws, f'C{rw}', fml, xp, NUM0, bold=bd, green=gr)
    putf(ws, f'D{rw}', f"=C{rw}/{a('shares')}/1000*{a('fx')}", xp / SH / 1000.0 * PEG, PX,
         bold=bd)
band(ws, SB['evops'], 4); band(ws, SB['ev'], 4); band(ws, SB['eq'], 4)
put(ws, f"A{SB['tvshare']}", 'Terminal value as a share of enterprise value', fmt=None)
putf(ws, f"C{SB['tvshare']}", f"=C{SB['pvtv']}/C{SB['evops']}", DC['tv_share'], PCT)
put(ws, f"A{SB['fvusd']}", 'Fair value per share (USD)', fmt=None)
putf(ws, f"C{SB['fvusd']}", f"=C{SB['eq']}/{a('shares')}/1000", DC['fv_usd'], PX)
put(ws, f"A{SB['fvaed']}", 'Fair value per share (AED)', bold=True, fmt=None)
putf(ws, f"C{SB['fvaed']}", f"=C{SB['fvusd']}*{a('fx')}", DC['fv_aed'], PX, bold=True)
band(ws, SB['fvaed'], 4)

band(ws, SB['legb'], 6)
put(ws, f"A{SB['legb']}", 'THE SUM-OF-THE-PARTS CROSS-CHECK — EACH LEG ON ITS OWN MULTIPLE',
    bold=True, fmt=None)
hdr(ws, SB['legb'] + 0, ['', '2026E EBITDA', 'Multiple', 'Enterprise value', 'Basis'],
    start=2)
_legmult = {'Integrated Logistics': f"C{SB['mcon']}", 'Shipping': f"C{SB['mship']}",
            'Services': f"C{SB['mcon']}"}
_basis = {l['leg']: l['basis'] for l in SOTP['legs']}
for j, g in enumerate(GROUPS):
    rw = SB['leg0'] + j
    put(ws, f'A{rw}', g, fmt=None)
    putf(ws, f'B{rw}', f"=Segments!B{SG['geb0']+j}", GRP_EB_F[g][0], NUM0, green=True)
    putf(ws, f'C{rw}', f"={_legmult[g]}", SOTP_MULT[g], MULT)
    putf(ws, f'D{rw}', f'=B{rw}*C{rw}', SOTP_EV[g], NUM0)
    put(ws, f'E{rw}', _basis[g], fmt=None, wrap=True)
    ws.row_dimensions[rw].height = 28
band(ws, SB['legt'], 6)
put(ws, f"A{SB['legt']}", 'Enterprise value of the operating legs', bold=True, fmt=None)
putf(ws, f"D{SB['legt']}", f"=SUM(D{SB['leg0']}:D{SB['leg0']+2})", SOTP_EVOPS, NUM0,
     bold=True)
ws.column_dimensions['E'].width = 52

band(ws, SB['mb'], 6)
put(ws, f"A{SB['mb']}", 'THE SHIPPING MULTIPLE, BUILT — NOT PASTED', bold=True, fmt=None)
for rw, lab, fml, xp, fmt in [
        (SB['mcon'], 'Contracted-fleet multiple (long-term contracted gas shipping peer)',
         f"='Peer & Sector'!$C${PR['mcon']}", MULT_CONTR, MULT),
        (SB['mspot'], 'Spot-tanker multiple (average of the two spot-tanker peers)',
         f"='Peer & Sector'!$C${PR['mspot']}", MULT_SPOT, MULT),
        (SB['mw'], "Share of 2026 earnings exposed to spot rates, as disclosed",
         f"={a('spot_w')}", SPOT_W, PCT),
        (SB['mship'], 'Shipping multiple — the two weighted by that share',
         f"=(1-C{SB['mw']})*C{SB['mcon']}+C{SB['mw']}*C{SB['mspot']}", BLEND_EV, MULT)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=fml.startswith(("='P", '=Assumptions')))

band(ws, SB['bb'], 6)
put(ws, f"A{SB['bb']}", 'THE SUM-OF-THE-PARTS BRIDGE TO EQUITY', bold=True, fmt=None)
_sb = [(SB['bevops'], 'Enterprise value of the operating legs', f"=D{SB['legt']}",
        SOTP_EVOPS, False),
       (SB['bjv'], 'Plus joint ventures and associates at carrying value', f"={a('jv')}",
        JV_BV, True),
       (SB['bev'], 'Enterprise value', f"=C{SB['bevops']}+C{SB['bjv']}",
        SOTP_EVOPS + JV_BV, False),
       (SB['bnd'], 'Less net debt at 31 March 2026', f"=-{a('nd_co')}", -NDCO, True),
       (SB['bdefd'], 'Less deferred consideration on acquisitions', f"=-{a('deferred')}",
        -DEFERRED, True),
       (SB['bacq'], 'Less the eleven vessels bought on 7 August 2026, at the announced '
        'price', f"=-{a('acq_cost')}", -ACQ_COST, True),
       (SB['bhyb'], 'Less perpetual capital securities at carrying value',
        f"=-{a('hybrid')}", -HYBRID, True),
       (SB['bnci'], 'Less non-controlling interests at carrying value', f"=-{a('nci_bv')}",
        -NCI_BV, True),
       (SB['beq'], 'Equity attributable to ordinary shareholders',
        f"=C{SB['bev']}+C{SB['bnd']}+C{SB['bdefd']}+C{SB['bacq']}+C{SB['bhyb']}"
        f"+C{SB['bnci']}",
        SOTP_EQ, False)]
for rw, lab, fml, xp, gr in _sb:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, NUM0, bold=(rw in (SB['bev'], SB['beq'])), green=gr)
band(ws, SB['beq'], 6)
put(ws, f"A{SB['bfv']}", 'Sum-of-the-parts fair value per share (AED)', bold=True, fmt=None)
putf(ws, f"C{SB['bfv']}", f"=C{SB['beq']}/{a('shares')}/1000*{a('fx')}", SOTP_FV, PX,
     bold=True)
band(ws, SB['bfv'], 6)
note(ws, f"A{SB['bfv']+2}", 'The sum-of-the-parts is a cross-check on the discounted cash '
     'flow, not a fifth lens: it prices each business unit\'s 2026 earnings on the multiple '
     'the market is paying for that kind of earnings stream, and the Shipping leg is the '
     'only one that carries spot exposure, which is why its multiple is a weighted blend '
     'rather than a single peer figure.')

# ============ 6 SEGMENTS =======================================================
ws = sheet('Segments')
title(ws, 'Segments — the disclosed record and the unit build', 'USD thousand unless stated. '
      'The tanker fleet is built vessel by vessel; every forecast figure below is a formula.',
      9, awidth=52, cwidth=14)
hdr(ws, SG['revh'], ['Segment revenue — as disclosed (USD 000)'] + YH)
for j, s in enumerate(SEGS):
    put(ws, f"A{SG['revh0']+j}", s, fmt=None)
    for i in range(3):
        put(ws, f"{HC[i]}{SG['revh0']+j}", SEGH[s]['revenue'][i], BLUE, NUM0)
band(ws, SG['revht'], 4)
put(ws, f"A{SG['revht']}", 'Total revenue', bold=True, fmt=None)
for i in range(3):
    putf(ws, f"{HC[i]}{SG['revht']}",
         f"=SUM({HC[i]}{SG['revh0']}:{HC[i]}{SG['revh0']+6})", HI['revenue'][i], NUM0,
         bold=True)
hdr(ws, SG['ebh'], ['Segment EBITDA — as disclosed (USD 000)'] + YH)
for j, s in enumerate(SEGS):
    put(ws, f"A{SG['ebh0']+j}", s, fmt=None)
    for i in range(3):
        put(ws, f"{HC[i]}{SG['ebh0']+j}", SEGH[s]['ebitda'][i], BLUE, NUM0)
band(ws, SG['ebht'], 4)
put(ws, f"A{SG['ebht']}", 'Total segment EBITDA', bold=True, fmt=None)
_segeb_tot = [sum(SEGH[s]['ebitda'][i] for s in SEGS) for i in range(3)]
for i in range(3):
    putf(ws, f"{HC[i]}{SG['ebht']}", f"=SUM({HC[i]}{SG['ebh0']}:{HC[i]}{SG['ebh0']+6})",
         _segeb_tot[i], NUM0, bold=True)
put(ws, f"A{SG['mgnh']}", 'Segment EBITDA margin', fmt=None)
for i in range(3):
    putf(ws, f"{HC[i]}{SG['mgnh']}", f"={HC[i]}{SG['ebht']}/{HC[i]}{SG['revht']}",
         _segeb_tot[i] / HI['revenue'][i], PCT)

def clsband(rw, label, w=9):
    """A banded header row whose columns are the five vessel classes."""
    band(ws, rw, w); put(ws, f'A{rw}', label, bold=True, fmt=None)
    for i, c in enumerate(CLS_NAME):
        cc = ws.cell(row=rw, column=2 + i, value=c)
        cc.font = Font(bold=True); cc.fill = FILL_G


def winband(rw, label):
    """A banded header row whose columns are the seven rate windows."""
    band(ws, rw, 9); put(ws, f'A{rw}', label, bold=True, fmt=None)
    for i, c in enumerate(WIN_LAB):
        cc = ws.cell(row=rw, column=2 + i, value=c)
        cc.font = Font(bold=True); cc.fill = FILL_G


def yrband(rw, label, w=6):
    band(ws, rw, w); put(ws, f'A{rw}', label, bold=True, fmt=None)
    for i, y in enumerate(YF):
        cc = ws.cell(row=rw, column=2 + i, value=y)
        cc.font = Font(bold=True); cc.fill = FILL_G


clsband(SG['tband'], 'TANKERS — THE UNIT BUILD, VESSEL BY VESSEL')
put(ws, f"A{SG['own25']}", 'Vessels owned at 31 December 2025', fmt=None)
put(ws, f"A{SG['sold']}", 'Less vessels sold between the year end and the valuation date',
    fmt=None)
put(ws, f"A{SG['own']}", 'Vessels owned at the 31 March 2026 valuation date', bold=True,
    fmt=None)
put(ws, f"A{SG['acqn']}", 'Vessels bought in the purchase announced 7 August 2026', fmt=None)
put(ws, f"A{SG['acqdate']}", 'Their delivery date — they earn nothing before it and at the '
    'implied spot rate after it', fmt=None)
for j, c in enumerate(CLS):
    putf(ws, f"{CD[j]}{SG['own25']}", f"={a('tnk_own25', col=CD[j])}", OWNED25[c], NUM0,
         green=True)
    putf(ws, f"{CD[j]}{SG['sold']}", f"={a('tnk_sold', col=CD[j])}", SOLD[c], NUM0,
         green=True)
    putf(ws, f"{CD[j]}{SG['own']}", f"={CD[j]}{SG['own25']}-{CD[j]}{SG['sold']}", OWNED[c],
         NUM0, bold=True)
    # only the crude carriers were bought; every other class carries a count of nothing,
    # which is why one formula serves all five below
    putf(ws, f"{CD[j]}{SG['acqn']}",
         f"={a('acq_vlcc', col='B')}" if c == 'vlcc' else '=0', ACQ_N[c], NUM0,
         green=(c == 'vlcc'))
    putf(ws, f"{CD[j]}{SG['acqdate']}",
         f"={a('acq_vlcc', col='C')}" if c == 'vlcc' else '=0', ACQ_DATE[c],
         DATEFMT if c == 'vlcc' else NUM0, green=(c == 'vlcc'))

winband(SG['winb'], 'THE PUBLISHED BLEND, CONVERTED TO A SPOT RATE — WINDOW BY WINDOW. The '
        'company publishes ONE rate per class and states it is a blend across the whole '
        'class, charters out included, so the spot rate is SOLVED out of it here')
put(ws, f"A{SG['winst']}", 'Window begins', fmt=None)
put(ws, f"A{SG['winen']}", 'Window ends', fmt=None)
put(ws, f"A{SG['windy']}", 'Days in the window', fmt=None)
for w in range(NWIN):
    cw = WCOL[w]
    if w < 6:
        putf(ws, f"{cw}{SG['winst']}", f"={a('win_start', col=cw)}", _ser(WIN[w][0]),
             DATEFMT, green=True)
        putf(ws, f"{cw}{SG['winen']}", f"={a('win_end', col=cw)}", _ser(WIN[w][1]), DATEFMT,
             green=True)
    else:                        # the mid-cycle anchor reuses the first 2025 window
        putf(ws, f"{cw}{SG['winst']}", f"=B{SG['winst']}", _ser(WIN[w][0]), DATEFMT)
        putf(ws, f"{cw}{SG['winen']}", f"=B{SG['winen']}", _ser(WIN[w][1]), DATEFMT)
    putf(ws, f"{cw}{SG['windy']}", f"={cw}{SG['winen']}-{cw}{SG['winst']}",
         (WIN[w][1] - WIN[w][0]).days, NUM0)


def _blend_f(j, w):
    """The published blend for class j over window w, as a formula.

    The smallest class is not broken out in the disclosure, so the medium-range rate stands
    in for it — SCALED by the disclosed relative move, which is a driver on the Assumptions
    sheet. Scaling is applied in every window and on both sides of the mid-cycle average,
    because an unadjusted substitution had the sign of that class's own year wrong.
    """
    c = CLS[j]
    src = 'mr' if c == 'hs' else c
    hs = f"*{a('hs_rel')}" if c == 'hs' else ''
    if w < 4:
        if c == 'hs':                          # no quarterly series: the class average
            return f"=AVERAGE(Assumptions!$B${A['b25_mr']}:$E${A['b25_mr']}){hs}"
        return f"={a('b25_' + src, col=CD[w])}"
    if w in (4, 5):
        return f"={a('b26_' + src, col=CD[w - 4])}{hs}"
    # the mid-cycle anchor: the 2024 and 2025 published blends averaged
    r25 = A['b25_' + src]
    if src == 'mr':                            # 2024 quarters are not disclosed for it
        return f"=AVERAGE(Assumptions!$B${r25}:$E${r25}){hs}"
    r24 = A['b24_' + src]
    return (f"=(AVERAGE(Assumptions!$B${r24}:$E${r24})"
            f"+AVERAGE(Assumptions!$B${r25}:$E${r25}))/2")


def _ch_terms(j, start_ref, end_ref, weighted):
    """Days (or revenue) the class's own charters out earn between two date cells."""
    rows = CH_ROWS[CLS[j]]
    if not rows:
        return '=0'
    parts = []
    for k in rows:
        days = (f"MAX(MIN({a('ch'+str(k), col='D')},{end_ref})"
                f"-MAX({a('ch'+str(k), col='C')},{start_ref}),0)")
        parts.append(f"{a('ch'+str(k), col='B')}*{days}" if weighted else days)
    return '=' + '+'.join(parts)


band(ws, SG['blendb'], 9)
put(ws, f"A{SG['blendb']}", 'Published blended rate by class (USD per day)', bold=True,
    fmt=None)
band(ws, SG['cdb'], 9)
put(ws, f"A{SG['cdb']}", 'Class vessel-days in the window — vessels owned x days', bold=True,
    fmt=None)
band(ws, SG['chdb'], 9)
put(ws, f"A{SG['chdb']}", 'Charter vessel-days in the window — each fixture for exactly the '
    'days its own contract runs', bold=True, fmt=None)
band(ws, SG['chrb'], 9)
put(ws, f"A{SG['chrb']}", 'Charter revenue in the window (USD) — each fixture at its own '
    'contracted rate', bold=True, fmt=None)
band(ws, SG['sdb'], 9)
put(ws, f"A{SG['sdb']}", 'Spot vessel-days — class days less charter days', bold=True,
    fmt=None)
band(ws, SG['spb'], 9)
put(ws, f"A{SG['spb']}", 'IMPLIED SPOT RATE — the blend with the chartered vessels removed '
    '(USD per day)', bold=True, fmt=None)
for j, c in enumerate(CLS):
    for key in ('blend0', 'cd0', 'chd0', 'chr0', 'sd0', 'sp0'):
        put(ws, f'A{SG[key]+j}', CLS_NAME[j], fmt=None)
    for w in range(NWIN):
        cw = WCOL[w]
        fleet_row = SG['own'] if w in (4, 5) else SG['own25']
        putf(ws, f"{cw}{SG['blend0']+j}", _blend_f(j, w), BLEND_W[c][w], NUM0, green=True)
        putf(ws, f"{cw}{SG['cd0']+j}", f"=${CD[j]}${fleet_row}*{cw}${SG['windy']}",
             CDAYS[c][w], NUM0)
        putf(ws, f"{cw}{SG['chd0']+j}",
             _ch_terms(j, f"{cw}${SG['winst']}", f"{cw}${SG['winen']}", False),
             CHD[c][w], NUM0)
        putf(ws, f"{cw}{SG['chr0']+j}",
             _ch_terms(j, f"{cw}${SG['winst']}", f"{cw}${SG['winen']}", True),
             CHREV[c][w], NUM0)
        putf(ws, f"{cw}{SG['sd0']+j}", f"={cw}{SG['cd0']+j}-{cw}{SG['chd0']+j}",
             SDAYS[c][w], NUM0)
        putf(ws, f"{cw}{SG['sp0']+j}",
             f"=({cw}{SG['blend0']+j}*{cw}{SG['cd0']+j}-{cw}{SG['chr0']+j})"
             f"/{cw}{SG['sd0']+j}", SPOTWIN[c][w], NUM0, bold=True)

clsband(SG['rateb'], 'THE IMPLIED SPOT RATE PATH — NEVER THE PUBLISHED BLEND', 6)
for rw, lab in [(SG['sp25'], '2025 implied spot rate — the four quarters averaged (USD per '
                 'day)'),
                (SG['spmid'], 'Mid-cycle implied spot anchor (USD per day)'),
                (SG['spq1'], 'First-quarter 2026 implied spot rate (USD per day)'),
                (SG['spq2'], 'Second-quarter 2026 implied spot rate (USD per day)'),
                (SG['sph2'], 'Second half of 2026 — the first quarter stepped back toward '
                 'the 2025 implied spot rate (USD per day)'),
                (SG['spy26'], 'FY2026 implied spot rate — the four quarters averaged (USD '
                 'per day)')]:
    put(ws, f'A{rw}', lab, fmt=None)
for j, c in enumerate(CLS):
    putf(ws, f"{CD[j]}{SG['sp25']}", f"=AVERAGE(B{SG['sp0']+j}:E{SG['sp0']+j})", SPOT25[c],
         NUM0)
    putf(ws, f"{CD[j]}{SG['spmid']}", f"=H{SG['sp0']+j}", SPOT_MID[c], NUM0)
    putf(ws, f"{CD[j]}{SG['spq1']}", f"=F{SG['sp0']+j}", SPOT_Q1[c], NUM0)
    putf(ws, f"{CD[j]}{SG['spq2']}", f"=G{SG['sp0']+j}", SPOT_Q2[c], NUM0)
    putf(ws, f"{CD[j]}{SG['sph2']}",
         f"={CD[j]}{SG['spq1']}*(1-{a('h2w')})+{CD[j]}{SG['sp25']}*{a('h2w')}", TNK_H2[c],
         NUM0)
    putf(ws, f"{CD[j]}{SG['spy26']}",
         f"=({CD[j]}{SG['spq1']}+{CD[j]}{SG['spq2']}+2*{CD[j]}{SG['sph2']})/4", TNK_Y26[c],
         NUM0)

yrband(SG['pathb'], 'Implied spot rate by class, gliding to the mid-cycle anchor (USD per '
       'day)')
for j, c in enumerate(CLS):
    rw = SG['path0'] + j
    put(ws, f'A{rw}', CLS_NAME[j], fmt=None)
    for i in range(5):
        f_ = (f"=${CD[j]}${SG['spy26']}" if i == 0 else
              f"=${CD[j]}${SG['spy26']}+(${CD[j]}${SG['spmid']}-${CD[j]}${SG['spy26']})"
              f"*{i}/4")
        putf(ws, f'{CD[i]}{rw}', f_, TNK_PATH[c][i], NUM0)

yrband(SG['yrb'], 'THE FORECAST YEARS — EVERY VESSEL PRICED ON ITS OWN TERMS')
put(ws, f"A{SG['yrst']}", 'Year begins', fmt=None)
put(ws, f"A{SG['yren']}", 'Year ends', fmt=None)
put(ws, f"A{SG['yrdy']}", 'Days in the year', fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['yrst']}", f"={a('yr_start', col=CD[i])}", _ser(YRB[i][0]),
         DATEFMT, green=True)
    putf(ws, f"{CD[i]}{SG['yren']}", f"={a('yr_end', col=CD[i])}", _ser(YRB[i][1]), DATEFMT,
         green=True)
    putf(ws, f"{CD[i]}{SG['yrdy']}", f"={CD[i]}{SG['yren']}-{CD[i]}{SG['yrst']}", YRDAYS[i],
         NUM0)
yrband(SG['ycdb'], 'Charter vessel-days by class')
yrband(SG['ycrb'], 'Charter revenue by class (USD 000)')
yrband(SG['yacdb'], 'Acquired vessel-days by class — the vessels bought on 7 August 2026, '
       'counted from their own delivery date')
yrband(SG['ysdb'], 'Spot vessel-days by class — owned vessels x days, less charter days, '
       'plus the acquired vessels from delivery')
yrband(SG['ysrb'], 'Spot revenue by class (USD 000) — spot vessel-days x the implied spot '
       'rate')
for j, c in enumerate(CLS):
    for key in ('ycd0', 'ycr0', 'yacd0', 'ysd0', 'ysr0'):
        put(ws, f'A{SG[key]+j}', CLS_NAME[j], fmt=None)
    for i in range(5):
        cw = CD[i]
        putf(ws, f"{cw}{SG['ycd0']+j}",
             _ch_terms(j, f"{cw}${SG['yrst']}", f"{cw}${SG['yren']}", False), YCD[c][i],
             NUM0)
        _rev = _ch_terms(j, f"{cw}${SG['yrst']}", f"{cw}${SG['yren']}", True)
        putf(ws, f"{cw}{SG['ycr0']+j}",
             '=0' if _rev == '=0' else f"=({_rev[1:]})/1000", YCR[c][i], NUM0)
        putf(ws, f"{cw}{SG['yacd0']+j}",
             f"=${CD[j]}${SG['acqn']}*MAX({cw}${SG['yren']}"
             f"-MAX({cw}${SG['yrst']},${CD[j]}${SG['acqdate']}),0)", YACD[c][i], NUM0)
        putf(ws, f"{cw}{SG['ysd0']+j}",
             f"=${CD[j]}${SG['own']}*{cw}{SG['yrdy']}-{cw}{SG['ycd0']+j}"
             f"+{cw}{SG['yacd0']+j}", YSD[c][i], NUM0)
        putf(ws, f"{cw}{SG['ysr0']+j}",
             f"={cw}{SG['ysd0']+j}*{cw}{SG['path0']+j}/1000", YSR[c][i], NUM0)
put(ws, f"A{SG['chrevt']}", 'Charter revenue — all classes', fmt=None)
put(ws, f"A{SG['sprevt']}", 'Spot revenue — all classes', fmt=None)
put(ws, f"A{SG['tcerev']}", 'Time-charter-equivalent revenue', bold=True, fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['chrevt']}",
         f"=SUM({CD[i]}{SG['ycr0']}:{CD[i]}{SG['ycr0']+4})", TNK_CHREV[i], NUM0)
    putf(ws, f"{CD[i]}{SG['sprevt']}",
         f"=SUM({CD[i]}{SG['ysr0']}:{CD[i]}{SG['ysr0']+4})", TNK_SPOTREV[i], NUM0)
    putf(ws, f"{CD[i]}{SG['tcerev']}",
         f"={CD[i]}{SG['chrevt']}+{CD[i]}{SG['sprevt']}", TNK_TCEREV[i], NUM0, bold=True)

band(ws, SG['opxb'], 6)
put(ws, f"A{SG['opxb']}", 'THE RUNNING COST — SOLVED FROM THE 2025 OUTCOME, NOT ASSUMED',
    bold=True, fmt=None)
put(ws, f"A{SG['vdays25']}", 'Vessel-days in 2025 (fleet at 31 December 2025 x 365)',
    fmt=None)
putf(ws, f"B{SG['vdays25']}", f"=SUM(B{SG['own25']}:F{SG['own25']})*365", VDAYS25, NUM0)
put(ws, f"A{SG['tcerev25']}", '2025 charter-equivalent revenue on the published blends '
    '(USD 000)', fmt=None)
_t25 = '=(' + '+'.join(
    f"{CD[j]}{SG['own25']}*AVERAGE(B{SG['blend0']+j}:E{SG['blend0']+j})"
    for j in range(5)) + ')*365/1000'
putf(ws, f"B{SG['tcerev25']}", _t25, TCEREV25, NUM0)
put(ws, f"A{SG['teb25']}", '2025 Tankers EBITDA as disclosed (USD 000)', fmt=None)
putf(ws, f"B{SG['teb25']}", f"=D{SG['ebh0']+SEGREF['Tankers']}", TNK_EB25, NUM0, green=True)
put(ws, f"A{SG['opexd0']}", 'Implied all-in running cost per vessel-day — the gap between '
    'the two, over the vessel-days that earned it (USD)', bold=True, fmt=None)
putf(ws, f"B{SG['opexd0']}",
     f"=(B{SG['tcerev25']}-B{SG['teb25']})*1000/B{SG['vdays25']}", OPEX_DAY, NUM1, bold=True)

yrband(SG['opexd'], 'Tanker revenue and running cost (USD 000)')
put(ws, f"A{SG['opexd']}", 'Running cost per vessel-day, escalated (USD)', bold=True,
    fmt=None)
put(ws, f"A{SG['opex']}", 'Total running cost — cost per day x vessel-days', fmt=None)
put(ws, f"A{SG['teb']}", 'Tankers EBITDA', bold=True, fmt=None)
put(ws, f"A{SG['gross']}", 'Gross-up from time-charter-equivalent to reported revenue',
    fmt=None)
put(ws, f"A{SG['trev']}", 'Tankers revenue', bold=True, fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['opexd']}", f"={a('opex_day')}*{esc_chain(i)}",
         TNK_OPEXD[i], NUM1)
    putf(ws, f"{CD[i]}{SG['opex']}", f"=$B${SG['vdays25']}*{CD[i]}{SG['opexd']}/1000",
         TNK_OPEX[i], NUM0)
    putf(ws, f"{CD[i]}{SG['teb']}", f"={CD[i]}{SG['tcerev']}-{CD[i]}{SG['opex']}",
         TNK_EBITDA[i], NUM0, bold=True)
    putf(ws, f"{CD[i]}{SG['gross']}", f"={a('grossup')}", GROSSUP, '0.00', green=True)
    putf(ws, f"{CD[i]}{SG['trev']}", f"={CD[i]}{SG['tcerev']}*{CD[i]}{SG['gross']}",
         TNK_REV[i], NUM0, bold=True)

band(ws, SG['gasb'], 6)
put(ws, f"A{SG['gasb']}", 'GAS CARRIERS — CONTRACTED VESSEL-YEARS x IMPLIED DAY RATE',
    bold=True, fmt=None)
put(ws, f"A{SG['gasvy25']}", 'Consolidated gas vessels in service through 2025 '
    '(vessel-years)', fmt=None)
putf(ws, f"B{SG['gasvy25']}", f"={a('gas_vy25')}", GAS_VY25, NUM1, green=True)
put(ws, f"A{SG['gasrev25']}", '2025 Gas Carriers revenue as disclosed (USD 000)', fmt=None)
putf(ws, f"B{SG['gasrev25']}", f"=D{SG['revh0']+SEGREF['Gas Carriers']}", GAS_REV25, NUM0,
     green=True)
put(ws, f"A{SG['gasrate0']}", 'Implied revenue per gas vessel-day — solved from the two '
    'above (USD)', bold=True, fmt=None)
putf(ws, f"B{SG['gasrate0']}", f"=B{SG['gasrev25']}*1000/(B{SG['gasvy25']}*365)", GAS_RATE,
     NUM0, bold=True)
yrband(SG['gasvyb'], 'Gas carriers — the forecast years')
put(ws, f"A{SG['gasvyb']}", 'Contracted vessel-years before the August 2026 purchase',
    bold=True, fmt=None)
put(ws, f"A{SG['gasvya']}", 'Vessel-years the five gas carriers bought on 7 August 2026 '
    'add — each tranche from its own delivery date', fmt=None)
put(ws, f"A{SG['gasvy']}", 'Contracted vessel-years', bold=True, fmt=None)
put(ws, f"A{SG['gasrate']}", 'Revenue per vessel-day, escalated (USD)', fmt=None)
put(ws, f"A{SG['gasrev']}", 'Gas Carriers revenue', fmt=None)
put(ws, f"A{SG['gasmgn']}", 'Gas Carriers earnings margin', fmt=None)
put(ws, f"A{SG['gasgeb']}", 'Gas Carriers earnings before the joint-venture share', fmt=None)
put(ws, f"A{SG['gasjv']}", 'Less the share of joint-venture profit carried inside the '
    'disclosed segment, escalated — the equity bridge already adds those ventures at '
    'carrying value', fmt=None)
put(ws, f"A{SG['gaseb']}", 'Gas Carriers EBITDA', bold=True, fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['gasvyb']}", f"={a('gas_vy_base', col=CD[i])}", GAS_VY_BASE[i],
         NUM1, green=True)
    putf(ws, f"{CD[i]}{SG['gasvya']}", f"={a('gas_vy_acq', col=CD[i])}", GAS_ACQ_VY[i],
         NUM1, green=True)
    # the total is taken from the driver sheet, where the two parts are added, so there is
    # exactly ONE chain from the contract table and the purchase to gas revenue
    putf(ws, f"{CD[i]}{SG['gasvy']}", f"={a('gas_vy', col=CD[i])}", GAS_VY[i], NUM1,
         bold=True, green=True)
    putf(ws, f"{CD[i]}{SG['gasrate']}", f"={a('gas_rate')}*{esc_chain(i)}",
         GAS_RATED[i], NUM0)
    putf(ws, f"{CD[i]}{SG['gasrev']}",
         f"={CD[i]}{SG['gasvy']}*365*{CD[i]}{SG['gasrate']}/1000", GAS_REV[i], NUM0)
    putf(ws, f"{CD[i]}{SG['gasmgn']}", f"={a('gas_mgn')}", GAS_MGN, PCT, green=True)
    putf(ws, f"{CD[i]}{SG['gasgeb']}", f"={CD[i]}{SG['gasrev']}*{CD[i]}{SG['gasmgn']}",
         GAS_GROSS_EB[i], NUM0)
    putf(ws, f"{CD[i]}{SG['gasjv']}", f"=-{a('jv_gas')}*{esc_chain(i)}",
         -GAS_JV[i], NUM0)
    putf(ws, f"{CD[i]}{SG['gaseb']}", f"={CD[i]}{SG['gasgeb']}+{CD[i]}{SG['gasjv']}",
         GAS_EBITDA[i], NUM0, bold=True)

yrband(SG['unitb'], 'THE REMAINING FIVE UNITS — REVENUE DRIVER x MARGIN DRIVER')
UNIT_ROW, UNIT_EB_ROW = {}, {}
_rr = SG['unit0']
for s in UNITS:
    k = s.lower().replace(' ', '_').replace('-', '_')
    UNIT_ROW[s] = _rr
    put(ws, f'A{_rr}', f'{s} — revenue', fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{_rr}', f"={a('rev_'+k, col=CD[i])}", DRV[s]['rev'][i], NUM0,
             green=True)
    if s == 'Services':
        put(ws, f'A{_rr+1}', f'{s} — earnings before the joint-venture share', fmt=None)
        put(ws, f'A{_rr+2}', f'{s} — less the share of joint-venture profit carried inside '
            'the disclosed segment, escalated', fmt=None)
        put(ws, f'A{_rr+3}', f'{s} — EBITDA', fmt=None)
        for i in range(5):
            putf(ws, f'{CD[i]}{_rr+1}', f"={CD[i]}{_rr}*{a('mar_'+k, col=CD[i])}",
                 SEG_GROSS_EB[s][i], NUM0)
            putf(ws, f'{CD[i]}{_rr+2}', f"=-{a('jv_serv')}*{esc_chain(i)}",
                 -SERV_JV[i], NUM0)
            putf(ws, f'{CD[i]}{_rr+3}', f"={CD[i]}{_rr+1}+{CD[i]}{_rr+2}", SEG_EB_F[s][i],
                 NUM0)
        UNIT_EB_ROW[s] = _rr + 3
        _rr += 4
    else:
        put(ws, f'A{_rr+1}', f'{s} — EBITDA', fmt=None)
        for i in range(5):
            putf(ws, f'{CD[i]}{_rr+1}', f"={CD[i]}{_rr}*{a('mar_'+k, col=CD[i])}",
                 SEG_EB_F[s][i], NUM0)
        UNIT_EB_ROW[s] = _rr + 1
        _rr += 2
assert _rr == SG['unit0'] + UNIT_N, 'the unit block did not fill its allocated rows'

_SEG_REV_SRC = {'Tankers': SG['trev'], 'Gas Carriers': SG['gasrev']}
_SEG_EB_SRC = {'Tankers': SG['teb'], 'Gas Carriers': SG['gaseb']}
for s in UNITS:
    _SEG_REV_SRC[s] = UNIT_ROW[s]; _SEG_EB_SRC[s] = UNIT_EB_ROW[s]
band(ws, SG['frevb'], 6)
put(ws, f"A{SG['frevb']}", 'FORECAST REVENUE BY SEGMENT', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=SG['frevb'], column=2 + i, value=y)
    cc.font = Font(bold=True); cc.fill = FILL_G
for j, s in enumerate(SEGS):
    put(ws, f"A{SG['frev0']+j}", s, fmt=None)
    for i in range(5):
        putf(ws, f"{CD[i]}{SG['frev0']+j}", f"={CD[i]}{_SEG_REV_SRC[s]}", SEG_REV_F[s][i],
             NUM0)
band(ws, SG['frevt'], 6)
put(ws, f"A{SG['frevt']}", 'Total revenue', bold=True, fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['frevt']}",
         f"=SUM({CD[i]}{SG['frev0']}:{CD[i]}{SG['frev0']+6})", REV_F[i], NUM0, bold=True)
band(ws, SG['febb'], 6)
put(ws, f"A{SG['febb']}", 'FORECAST EBITDA BY SEGMENT', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=SG['febb'], column=2 + i, value=y)
    cc.font = Font(bold=True); cc.fill = FILL_G
for j, s in enumerate(SEGS):
    put(ws, f"A{SG['feb0']+j}", s, fmt=None)
    for i in range(5):
        putf(ws, f"{CD[i]}{SG['feb0']+j}", f"={CD[i]}{_SEG_EB_SRC[s]}", SEG_EB_F[s][i],
             NUM0)
band(ws, SG['febt'], 6)
put(ws, f"A{SG['febt']}", 'Total EBITDA', bold=True, fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['febt']}", f"=SUM({CD[i]}{SG['feb0']}:{CD[i]}{SG['feb0']+6})",
         EB_F[i], NUM0, bold=True)
put(ws, f"A{SG['fmgn']}", 'Group EBITDA margin', fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{SG['fmgn']}", f"={CD[i]}{SG['febt']}/{CD[i]}{SG['frevt']}",
         EB_F[i] / REV_F[i], PCT)
band(ws, SG['grpb'], 6)
put(ws, f"A{SG['grpb']}", 'FORECAST BY BUSINESS UNIT', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=SG['grpb'], column=2 + i, value=y)
    cc.font = Font(bold=True); cc.fill = FILL_G
for j, g in enumerate(GROUPS):
    mem = [SEGS.index(s) for s in SEGS if SEG_GROUP[s] == g]
    put(ws, f"A{SG['grev0']+j}", f'{g} — revenue', fmt=None)
    put(ws, f"A{SG['geb0']+j}", f'{g} — EBITDA', fmt=None)
    put(ws, f"A{SG['gmgn0']+j}", f'{g} — EBITDA margin', fmt=None)
    for i in range(5):
        putf(ws, f"{CD[i]}{SG['grev0']+j}",
             '=' + '+'.join(f"{CD[i]}{SG['frev0']+m}" for m in mem), GRP_REV_F[g][i], NUM0)
        putf(ws, f"{CD[i]}{SG['geb0']+j}",
             '=' + '+'.join(f"{CD[i]}{SG['feb0']+m}" for m in mem), GRP_EB_F[g][i], NUM0)
        putf(ws, f"{CD[i]}{SG['gmgn0']+j}",
             f"={CD[i]}{SG['geb0']+j}/{CD[i]}{SG['grev0']+j}",
             GRP_EB_F[g][i] / GRP_REV_F[g][i], PCT)
note(ws, f"A{SG['gmgn0']+4}", 'The tanker gross-up moves reported revenue only, never '
     'earnings: earnings are struck on time-charter-equivalent revenue less the running '
     'cost of the same vessels. It does NOT follow that it cannot change the valuation, '
     'and an earlier edition of this workbook said so wrongly. It reaches the valuation '
     'through the balance sheet: receivable days are measured on reported revenue, so the '
     'gross-up sets the revenue base those days are re-based onto, and through the change '
     'in working capital it reaches free cash flow. The re-basing is built in cells on the '
     'Assumptions sheet.')

# ============ 7 RELATIVE & NORMALIZED ==========================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative multiples, normalised earnings power and the book lens', 'USD thousand '
      'unless stated. Per-share figures in dirhams.', 5, awidth=62, cwidth=17)
hdr(ws, RN['hdr'], ['Relative lens', '', 'Value'])
_rel = [
    (RN['eb26'], 'FY2026E group EBITDA', f"=DCF!B{DF_['ebitda']}", EB_F[0], NUM0, True),
    (RN['blend'], 'Blended enterprise multiple',
     f"='Peer & Sector'!$C${PR['mev']}", BLEND_EV, MULT, True),
    (RN['ev'], 'Implied enterprise value', f"=C{RN['eb26']}*C{RN['blend']}", REL_EV, NUM0,
     False),
    (RN['jv'], 'Plus joint ventures and associates at carrying value', f"={a('jv')}", JV_BV,
     NUM0, True),
    (RN['nd'], 'Less net debt at 31 March 2026', f"=-{a('nd_co')}", -NDCO, NUM0, True),
    (RN['defd'], 'Less deferred consideration on acquisitions', f"=-{a('deferred')}",
     -DEFERRED, NUM0, True),
    (RN['acq'], 'Less the eleven vessels bought on 7 August 2026, at the announced price',
     f"=-{a('acq_cost')}", -ACQ_COST, NUM0, True),
    (RN['hyb'], 'Less perpetual capital securities at carrying value', f"=-{a('hybrid')}",
     -HYBRID, NUM0, True),
    (RN['pre'], 'Implied equity value before the minorities',
     f"=C{RN['ev']}+C{RN['jv']}+C{RN['nd']}+C{RN['defd']}+C{RN['acq']}+C{RN['hyb']}",
     pre_nci_from_ev(REL_EV), NUM0, False),
    (RN['nci'], 'Less non-controlling interests — the contracted slice at its contracted '
     'price, the rest at the greater of book and value',
     f"=-(DCF!$C${DF_['ncinav']}+MAX(DCF!$C${DF_['nciother']},"
     f"C{RN['pre']}*DCF!$C${DF_['ncishare']}))",
     -nci_ded(pre_nci_from_ev(REL_EV)), NUM0, False),
    (RN['eq'], 'Implied equity attributable to ordinary shareholders',
     f"=C{RN['pre']}+C{RN['nci']}", eq_from_ev(REL_EV), NUM0, False),
    (RN['vev'], 'Value per share on the enterprise multiple (AED)',
     f"=C{RN['eq']}/{a('shares')}/1000*{a('fx')}", REL_V_EV, PX, False),
    (RN['pe'], 'Blended FORWARD price/earnings — applied to forward earnings below, so '
     'the multiple and the earnings are on the same basis',
     f"='Peer & Sector'!$C${PR['pe']}", BLEND_PE, MULT, True),
    (RN['ord26'], 'FY2026E earnings attributable to ordinary shareholders',
     f"='Income Statement'!E{IS['ordn']}", ORD_F[0], NUM0, True),
    (RN['vpe'], 'Value per share on the earnings multiple (AED)',
     f"=C{RN['ord26']}*C{RN['pe']}/{a('shares')}/1000*{a('fx')}", REL_V_PE, PX, False),
    (RN['w'], 'Weight on the enterprise multiple', f"={a('w_eveb')}", W_EVEB, PCT, True)]
for rw, lab, fml, xp, fmt, gr in _rel:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)
band(ws, RN['base'], 5)
put(ws, f"A{RN['base']}", 'RELATIVE LENS — value per share (AED)', bold=True, fmt=None)
putf(ws, f"C{RN['base']}",
     f"=C{RN['w']}*C{RN['vev']}+(1-C{RN['w']})*C{RN['vpe']}", REL_BASE, PX, bold=True)
def alt_bridge(rows, eb_ref, lo_mult, hi_mult, vals, lab):
    """Bear and bull on the same earnings, valued on the two ends of the peer frame, each
    carried through the SAME two-part minority deduction the base case uses."""
    put(ws, f"A{rows['ev']}", f'{lab} — enterprise value on the spot-tanker multiple (C) / '
        'on the contracted multiple (D)', fmt=None)
    put(ws, f"A{rows['pre']}", 'Equity value before the minorities', fmt=None)
    put(ws, f"A{rows['nci']}", 'Less non-controlling interests — the same two-part '
        'deduction', fmt=None)
    put(ws, f"A{rows['eq']}", 'Equity attributable to ordinary shareholders', fmt=None)
    for col, mult, ev_ in (('C', lo_mult, vals[0]), ('D', hi_mult, vals[1])):
        putf(ws, f"{col}{rows['ev']}", f"=C{eb_ref}*'Peer & Sector'!$C${mult}", ev_, NUM0)
        putf(ws, f"{col}{rows['pre']}",
             f"={col}{rows['ev']}+{a('jv')}-{a('nd_co')}-{a('deferred')}"
             f"-{a('acq_cost')}-{a('hybrid')}",
             pre_nci_from_ev(ev_), NUM0)
        putf(ws, f"{col}{rows['nci']}",
             f"=-(DCF!$C${DF_['ncinav']}+MAX(DCF!$C${DF_['nciother']},"
             f"{col}{rows['pre']}*DCF!$C${DF_['ncishare']}))",
             -nci_ded(pre_nci_from_ev(ev_)), NUM0)
        putf(ws, f"{col}{rows['eq']}", f"={col}{rows['pre']}+{col}{rows['nci']}",
             eq_from_ev(ev_), NUM0)


alt_bridge(dict(ev=RN['bearev'], pre=RN['bearpre'], nci=RN['bearnci'], eq=RN['beareq']),
           RN['eb26'], PR['mspot'], PR['mcon'],
           (MULT_SPOT * EB_F[0], MULT_CONTR * EB_F[0]), 'Bear and bull')
put(ws, f"A{RN['bear']}", 'Bear on the spot multiple (C) / bull on the contracted multiple '
    '(D), same construction', fmt=None)
putf(ws, f"C{RN['bear']}", f"=C{RN['beareq']}/{a('shares')}/1000*{a('fx')}", REL_BEAR, PX)
putf(ws, f"D{RN['bear']}", f"=D{RN['beareq']}/{a('shares')}/1000*{a('fx')}", REL_BULL, PX)

band(ws, RN['ownb'], 5)
put(ws, f"A{RN['ownb']}", "THE COMPANY'S OWN MULTIPLES AT THE ANCHOR PRICE", bold=True,
    fmt=None)
for rw, lab, fml, xp, fmt, gr in [
        (RN['spotusd'], 'Share price (USD)', f"={a('spot')}/{a('fx')}", SPOT / PEG, PX,
         True),
        (RN['mktcap'], 'Market capitalisation (USD 000)',
         f"=C{RN['spotusd']}*{a('shares')}*1000", MKTCAP, NUM0, False),
        (RN['netdebt'], 'Net debt including the deferred consideration and the vessels '
         'bought on 7 August 2026 (USD 000)',
         f"={a('nd_co')}+{a('deferred')}+{a('acq_cost')}", NETDEBT, NUM0, True),
        (RN['evnow'], 'Enterprise value (USD 000)',
         f"=C{RN['mktcap']}+C{RN['netdebt']}", EV_NOW, NUM0, False),
        (RN['eveb_ttm'], 'Enterprise value / 2025 reported EBITDA',
         f"=C{RN['evnow']}/'Income Statement'!D{IS['ebrep']}", OWN_EVEB_TTM, MULT, True),
        (RN['eveb_26'], 'Enterprise value / FY2026E EBITDA',
         f"=C{RN['evnow']}/C{RN['eb26']}", OWN_EVEB_26, MULT, False),
        (RN['pe_ttm'], 'Price / 2025 earnings attributable to ordinary shareholders — '
         'the TRAILING basis',
         f"=C{RN['mktcap']}/'Income Statement'!D{IS['ordn']}", OWN_PE_TTM, MULT, True),
        # the peers' forward multiples are applied to forward earnings, so the company's
        # own multiple has to be available on the same basis or the comparison is not
        # like for like — the first edition quoted this company trailing against peers
        # shown forward
        (RN['pe_fwd'], 'Price / FY2026E earnings attributable to ordinary shareholders — '
         'the FORWARD basis, the one the peer multiples are quoted on',
         f"=C{RN['mktcap']}/C{RN['ord26']}", OWN_PE_FWD, MULT, False),
        (RN['pb'], 'Price / book at 31 March 2026',
         f"=C{RN['mktcap']}/({a('eqp0')}+{a('hybrid')})", OWN_PB, MULT, False),
        (RN['dy'], 'Dividend yield on the 2026 distribution',
         f"={a('dps26')}/C{RN['mktcap']}", OWN_DY, PCT, True),
        # The peer figures above come from aggregators, which take enterprise value as
        # market capitalisation plus net debt and stop there. This study's own equity
        # bridge also deducts the perpetual securities and the minorities, so the
        # company's multiple on the SAME basis as its bridge is a different number.
        (RN['evbr'], 'Enterprise value on the equity bridge\'s own convention — adding the '
         'perpetual capital securities and the minorities (USD 000)',
         f"=C{RN['evnow']}+{a('hybrid')}+{a('nci_bv')}", EV_BRIDGE, NUM0, False),
        (RN['ebbr_ttm'], 'Enterprise value / 2025 reported EBITDA — bridge convention',
         f"=C{RN['evbr']}/'Income Statement'!D{IS['ebrep']}", OWN_EVEB_TTM_BR, MULT, False),
        (RN['ebbr_26'], 'Enterprise value / FY2026E EBITDA — bridge convention',
         f"=C{RN['evbr']}/C{RN['eb26']}", OWN_EVEB_26_BR, MULT, False)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)
note(ws, f"A{RN['ebbr_26']+1}", 'The two multiples immediately above are the company\'s own '
     'on the SAME enterprise value the equity bridge uses. The three rows before them use '
     'the narrower market-capitalisation-plus-net-debt convention, because that is the '
     'convention the peer figures are published on and the comparison has to be like for '
     'like. Both are shown rather than one chosen.')
hdr(ws, RN['nhdr'], ['Normalised earnings lens — the five-year average of the model\'s own '
                     'forecast', '', 'Value'])
for rw, lab, fml, xp, fmt, gr in [
        (RN['neb'], 'Mid-cycle EBITDA — five-year average',
         f"=AVERAGE(DCF!B{DF_['ebitda']}:F{DF_['ebitda']})", NORM_EB, NUM0, True),
        (RN['nev'], 'Implied enterprise value', f"=C{RN['neb']}*C{RN['blend']}",
         BLEND_EV * NORM_EB, NUM0, False),
        (RN['npre'], 'Implied equity value before the minorities',
         f"=C{RN['nev']}+{a('jv')}-{a('nd_co')}-{a('deferred')}-{a('acq_cost')}"
         f"-{a('hybrid')}",
         pre_nci_from_ev(BLEND_EV * NORM_EB), NUM0, False),
        (RN['nnci'], 'Less non-controlling interests — the same two-part deduction',
         f"=-(DCF!$C${DF_['ncinav']}+MAX(DCF!$C${DF_['nciother']},"
         f"C{RN['npre']}*DCF!$C${DF_['ncishare']}))",
         -nci_ded(pre_nci_from_ev(BLEND_EV * NORM_EB)), NUM0, False),
        (RN['neq'], 'Implied equity attributable to ordinary shareholders',
         f"=C{RN['npre']}+C{RN['nnci']}", eq_from_ev(BLEND_EV * NORM_EB), NUM0, False),
        (RN['nvev'], 'Value per share on the enterprise multiple (AED)',
         f"=C{RN['neq']}/{a('shares')}/1000*{a('fx')}", NORM_V_EV, PX, False),
        (RN['nord'], 'Mid-cycle earnings attributable to ordinary shareholders — five-year '
         'average', f"=AVERAGE('Income Statement'!E{IS['ordn']}:I{IS['ordn']})", NORM_ORD,
         NUM0, True),
        (RN['neps'], 'Mid-cycle earnings per share (USD)',
         f"=C{RN['nord']}/{a('shares')}/1000", NORM_EPS, PX3, False),
        (RN['nvpe'], 'Value per share on the earnings multiple (AED)',
         f"=C{RN['neps']}*C{RN['pe']}*{a('fx')}", NORM_V_PE, PX, False)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)
band(ws, RN['nbase'], 5)
put(ws, f"A{RN['nbase']}", 'NORMALISED LENS — value per share (AED)', bold=True, fmt=None)
putf(ws, f"C{RN['nbase']}", f"=(C{RN['nvev']}+C{RN['nvpe']})/2", NORM_BASE, PX, bold=True)
alt_bridge(dict(ev=RN['nbearev'], pre=RN['nbearpre'], nci=RN['nbearnci'],
                eq=RN['nbeareq']), RN['neb'], PR['mspot'], PR['mcon'],
           (MULT_SPOT * NORM_EB, MULT_CONTR * NORM_EB), 'Bear and bull')
put(ws, f"A{RN['nbear']}", 'Bear on the spot multiple (C) / bull on the contracted multiple '
    '(D), same construction', fmt=None)
putf(ws, f"C{RN['nbear']}", f"=C{RN['nbeareq']}/{a('shares')}/1000*{a('fx')}", NORM_BEAR, PX)
putf(ws, f"D{RN['nbear']}", f"=D{RN['nbeareq']}/{a('shares')}/1000*{a('fx')}", NORM_BULL, PX)

hdr(ws, RN['bhdr'], ['Book value and sustainable return — a RESIDUAL INCOME build', '',
                     'Value'])
for rw, lab, fml, xp, fmt, gr in [
        (RN['beqp'], 'Equity attributable to shareholders at 31 March 2026 (USD 000)',
         f"={a('eqp0')}", EQP0, NUM0, True),
        (RN['bbvps'], 'Book value per share (USD)',
         f"=C{RN['beqp']}/{a('shares')}/1000", BVPS0, PX3, False),
        (RN['bbvpsaed'], 'Book value per share (AED)', f"=C{RN['bbvps']}*{a('fx')}",
         BVPS0 * PEG, PX, False),
        (RN['broe'], 'Sustainable return on equity — five-year forecast average',
         f"=AVERAGE('Balance Sheet'!E{BS['roe']}:I{BS['roe']})", ROE_SUST, PCT, True),
        (RN['bke'], 'Cost of equity', f"=DCF!$C${DF_['ke']}", KE, PCT2, True),
        (RN['bg'], 'Terminal growth', f"={a('g_term')}", G, PCT, True),
        (RN['bfade'], 'Rate at which the return above the cost of equity fades beyond the '
         'forecast', f"={a('ri_fade')}", RI_FADE, PCT, True)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)


def ri_ladder(hrow, label, ke_ref, ke_val, scale, rows, det, open_row=None):
    """One residual-income ladder: five years of excess return, then a fading remainder."""
    band(ws, hrow, 6); put(ws, f'A{hrow}', label, bold=True, fmt=None)
    for i, y in enumerate(YF):
        cc = ws.cell(row=hrow, column=2 + i, value=y)
        cc.font = Font(bold=True); cc.fill = FILL_G
    if open_row is None:
        open_row = rows['open']
        put(ws, f"A{rows['open']}", 'Opening ordinary book value (USD mn) — the balance '
            'sheet\'s own roll-forward', fmt=None)
        for i in range(5):
            f_ = (f"=C{RN['beqp']}/1000" if i == 0
                  else f"='Balance Sheet'!{CD[i-1]}{BS['eqclose']}/1000")
            putf(ws, f"{CD[i]}{rows['open']}", f_, det[i]['open'], NUM1,
                 green=(i > 0))
    sc = '' if scale == 1.0 else f'*{scale}'
    put(ws, f"A{rows['roe']}", 'Return on ordinary equity — profit after the perpetual '
        'coupon, over closing ordinary book' + ('' if scale == 1.0 else
                                                f' (scaled {scale:.2f}x)'), fmt=None)
    put(ws, f"A{rows['ri']}", 'Residual income — (return less cost of equity) x opening '
        'book (USD mn)', fmt=None)
    put(ws, f"A{rows['df']}", 'Discount factor at the cost of equity', fmt=None)
    put(ws, f"A{rows['pv']}", 'Present value of the residual income (USD mn)', fmt=None)
    for i in range(5):
        putf(ws, f"{CD[i]}{rows['roe']}",
             f"='Income Statement'!{FCOL[i]}{IS['ordn']}/'Balance Sheet'!"
             f"{CD[i]}{BS['eqclose']}{sc}", det[i]['roe'], PCT, green=True)
        putf(ws, f"{CD[i]}{rows['ri']}",
             f"=({CD[i]}{rows['roe']}-{ke_ref})*{CD[i]}{open_row}", det[i]['ri'], NUM1)
        putf(ws, f"{CD[i]}{rows['df']}", f"=1/(1+{ke_ref})^{i+1}", det[i]['df'], DF4)
        putf(ws, f"{CD[i]}{rows['pv']}", f"={CD[i]}{rows['ri']}*{CD[i]}{rows['df']}",
             det[i]['pv'], NUM1)
    return open_row


def ri_tail(rows, r_, ke_ref, ke_val, scale, open_row):
    """The fading remainder and the equity value it completes."""
    sc = '' if scale == 1.0 else f'*{scale}'
    for rw, lab, fml, xp, fmt in [
            (rows['pvsum'], 'Present value of the five forecast years (USD mn)',
             f"=SUM(B{rows['pv']}:F{rows['pv']})", r_['pv_expl'], NUM1),
            (rows['rit'], 'Terminal residual income — the final year\'s excess return on '
             'the closing book (USD mn)',
             f"=(F{rows['roe']}-{ke_ref})*'Balance Sheet'!F{BS['eqclose']}/1000",
             r_['ri_t'], NUM1),
            (rows['tv'], 'Terminal value — that excess fading toward the cost of equity '
             '(USD mn)',
             f"=C{rows['rit']}/({ke_ref}+C{RN['bfade']}-C{RN['bg']})", r_['tv'], NUM1),
            (rows['pvtv'], 'Present value of the terminal value (USD mn)',
             f"=C{rows['tv']}/(1+{ke_ref})^5", r_['pv_tv'], NUM1),
            (rows['eq'], 'Equity value — opening book plus the residual income it earns '
             '(USD mn)',
             f"=C{RN['beqp']}/1000+C{rows['pvsum']}+C{rows['pvtv']}", r_['equity'], NUM1)]:
        put(ws, f'A{rw}', lab, bold=(rw == rows['eq']), fmt=None)
        putf(ws, f'C{rw}', fml, xp, fmt, bold=(rw == rows['eq']))


_BASE_ROWS = dict(open=RN['bopen'], roe=RN['broey'], ri=RN['bri'], df=RN['bdf'],
                  pv=RN['bpv'], pvsum=RN['bpvsum'], rit=RN['brit'], tv=RN['btv'],
                  pvtv=RN['bpvtv'], eq=RN['beq'])
_BEAR_ROWS = dict(roe=RN['xroe'], ri=RN['xri'], df=RN['xdf'], pv=RN['xpv'],
                  pvsum=RN['xpvsum'], rit=RN['xrit'], tv=RN['xtv'], pvtv=RN['xpvtv'],
                  eq=RN['xeq'])
_BULL_ROWS = dict(roe=RN['yroe'], ri=RN['yri'], df=RN['ydf'], pv=RN['ypv'],
                  pvsum=RN['ypvsum'], rit=RN['yrit'], tv=RN['ytv'], pvtv=RN['ypvtv'],
                  eq=RN['yeq'])
_KE_REF = f"C{RN['bke']}"
_KE_HI = f"DCF!$C${DF_['kecihi']}"
_KE_LO = f"DCF!$C${DF_['kecilo']}"
_open = ri_ladder(RN['bladder'], 'THE RESIDUAL-INCOME LADDER — BASE', _KE_REF, KE, 1.0,
                  _BASE_ROWS, RI['det'])
ri_tail(_BASE_ROWS, RI, _KE_REF, KE, 1.0, _open)
put(ws, f"A{RN['bpb']}", 'Implied price / book — the equity value over the book it starts '
    'from', fmt=None)
putf(ws, f"C{RN['bpb']}", f"=C{RN['beq']}/(C{RN['beqp']}/1000)", PB_FAIR, MULT)
band(ws, RN['bbase'], 5)
put(ws, f"A{RN['bbase']}", 'BOOK LENS — value per share (AED)', bold=True, fmt=None)
putf(ws, f"C{RN['bbase']}", f"=C{RN['beq']}/{a('shares')}*{a('fx')}", BOOK_BASE, PX,
     bold=True)
ri_ladder(RN['xladder'], 'THE SAME LADDER — BEAR: the return 15% lower, discounted at the '
          'cost of equity built on the TOP of the beta\'s 90% confidence interval', _KE_HI,
          KE_CI_HI, 0.85, _BEAR_ROWS, RI_BEAR['det'], open_row=_open)
ri_tail(_BEAR_ROWS, RI_BEAR, _KE_HI, KE_CI_HI, 0.85, _open)
ri_ladder(RN['yladder'], 'THE SAME LADDER — BULL: the return 15% higher, discounted at the '
          'cost of equity built on the BOTTOM of the same interval', _KE_LO, KE_CI_LO, 1.15,
          _BULL_ROWS, RI_BULL['det'], open_row=_open)
ri_tail(_BULL_ROWS, RI_BULL, _KE_LO, KE_CI_LO, 1.15, _open)
put(ws, f"A{RN['bbear']}", 'Bear — return 15% lower, discounted at the cost of equity built '
    'on the TOP of the beta\'s 90% confidence interval (C) / bull — return 15% higher, '
    'discounted at the cost of equity built on the BOTTOM of the same interval (D)',
    fmt=None)
putf(ws, f"C{RN['bbear']}", f"=C{RN['xeq']}/{a('shares')}*{a('fx')}", BOOK_BEAR, PX)
putf(ws, f"D{RN['bbear']}", f"=C{RN['yeq']}/{a('shares')}*{a('fx')}", BOOK_BULL, PX)

band(ws, RN['vsb'], 5)
put(ws, f"A{RN['vsb']}", 'THE REALISED VESSEL SALE — DIRECT EVIDENCE ON CARRYING VALUES',
    bold=True, fmt=None)
# Disclosed history, carried where it is read rather than on the driver sheet: these two
# figures are the primary record of one completed transaction, not inputs the valuation
# consumes, so they live here as pasted values and only the ratio between them is computed.
put(ws, f"A{RN['vsbook']}", 'Carrying value of the very large crude carrier sold (USD 000)',
    fmt=None)
put(ws, f"C{RN['vsbook']}", V['vessel_sale_book'], BLUE, NUM0)
put(ws, f"A{RN['vsprice']}", 'Realised sale price, January 2026 (USD 000)', fmt=None)
put(ws, f"C{RN['vsprice']}", V['vessel_sale_price'], BLUE, NUM0)
put(ws, f"A{RN['vsratio']}", 'Realised price over carrying value', bold=True, fmt=None)
putf(ws, f"C{RN['vsratio']}", f"=C{RN['vsprice']}/C{RN['vsbook']}", VSB_RATIO, MULT,
     bold=True)
put(ws, f"A{RN['vsgain']}", 'Capital gain recognised on the sale, as disclosed (USD 000)',
    fmt=None)
put(ws, f"C{RN['vsgain']}", V['vessel_sale_gain'], BLUE, NUM0)
note(ws, f"A{RN['vsnote']}", 'This is the only direct evidence in the study on how far the '
     'balance sheet\'s carrying values sit below what the fleet would actually fetch: one '
     '2017-built very large crude carrier, ninety per cent owned, sold in January 2026 at '
     'about a third above its carrying value. The book lens above values the equity at its '
     'carried book, so a fleet that would realise more than book is a reason that lens '
     'reads low rather than a reason to adjust it.')

# ============ 8 DCF =============================================================
ws = sheet('DCF')
title(ws, 'Discounted cash flow — the full waterfall', 'USD thousand. Every line is a live '
      'formula: the cost of capital is built below from its own components, the discount '
      'factors compound year on year off the glide, and 2026 is a three-quarter stub '
      'because the valuation date is 31 March 2026.', 6, awidth=58, cwidth=15)
hdr(ws, 4, ['USD 000'] + YFE)


def wf(rw, lab, fmls, vals, fmt=NUM0, bd=False, green=False):
    put(ws, f'A{rw}', lab, bold=bd, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{rw}', fmls(i), vals[i], fmt, bold=bd, green=green)
    if bd:
        band(ws, rw, 6)


wf(DF_['rev'], 'Revenue', lambda i: f"=Segments!{CD[i]}{SG['frevt']}", REV_F, green=True)
wf(DF_['ebitda'], 'EBITDA', lambda i: f"=Segments!{CD[i]}{SG['febt']}", EB_F, green=True)
wf(DF_['mgn'], 'EBITDA margin', lambda i: f"={CD[i]}{DF_['ebitda']}/{CD[i]}{DF_['rev']}",
   [EB_F[i] / REV_F[i] for i in range(5)], PCT)
wf(DF_['dna'], 'Less depreciation and amortisation',
   lambda i: f"=-'Balance Sheet'!{CD[i]}{BS['dnatot']}", [-x for x in DNA_F], green=True)
wf(DF_['ebit'], 'EBIT', lambda i: f"={CD[i]}{DF_['ebitda']}+{CD[i]}{DF_['dna']}", EBIT_F,
   bd=True)
wf(DF_['tax'], 'Less tax on operating profit — the business-unit mix below',
   lambda i: f"=-{CD[i]}{DF_['taxtot']}", [-x for x in TAX_F])
wf(DF_['nopat'], 'NOPAT', lambda i: f"={CD[i]}{DF_['ebit']}+{CD[i]}{DF_['tax']}", NOPAT_F,
   bd=True)
wf(DF_['adddna'], 'Add back depreciation and amortisation',
   lambda i: f"=-{CD[i]}{DF_['dna']}", DNA_F)
wf(DF_['capex'], 'Less capital expenditure', lambda i: f"=-{a('capex', col=CD[i])}",
   [-x for x in CAPEX], green=True)
wf(DF_['dnwc'], 'Less change in working capital',
   lambda i: f"=-'Balance Sheet'!{CD[i]}{BS['wcdnwc']}", [-x for x in DNWC_F], green=True)
wf(DF_['fcff'], 'Free cash flow to the firm',
   lambda i: (f"={CD[i]}{DF_['nopat']}+{CD[i]}{DF_['adddna']}+{CD[i]}{DF_['capex']}"
              f"+{CD[i]}{DF_['dnwc']}"), FCFF_F, bd=True)
put(ws, f"A{DF_['q1']}", 'Less first-quarter 2026 free cash flow, already inside net debt '
    'at the valuation date', fmt=None)
putf(ws, f"B{DF_['q1']}", f"=-{a('q1fcf')}", -Q1FCF, NUM0, green=True)
for i in range(1, 5):
    putf(ws, f"{CD[i]}{DF_['q1']}", '=0', 0.0, NUM0)
wf(DF_['fcfd'], 'Free cash flow discounted from 31 March 2026',
   lambda i: f"={CD[i]}{DF_['fcff']}+{CD[i]}{DF_['q1']}", DC['fcffd'], bd=True)
wf(DF_['glide'], 'Forward cost of capital — the glide from current to terminal',
   lambda i: (f"=$C${DF_['wacc']}+($C${DF_['waccterm']}-$C${DF_['wacc']})*{i+1}/5"),
   DC['glide'], PCT2)
wf(DF_['df'], 'Discount factor — each year compounded onto the last',
   lambda i: (f"=1/(1+{CD[i]}{DF_['glide']})^{a('stub')}" if i == 0
              else f"={CD[i-1]}{DF_['df']}/(1+{CD[i]}{DF_['glide']})"), DC['df'], DF4)
wf(DF_['pv'], 'Present value of free cash flow',
   lambda i: f"={CD[i]}{DF_['fcfd']}*{CD[i]}{DF_['df']}", DC['pv'], bd=True)

band(ws, DF_['taxb'], 6)
put(ws, f"A{DF_['taxb']}", 'THE TAX MIX — EACH BUSINESS UNIT AT ITS OWN DISCLOSED RATE',
    bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=DF_['taxb'], column=2 + i, value=y)
    cc.font = Font(bold=True); cc.fill = FILL_G
_TAXKEY = {'Integrated Logistics': 'tax_il', 'Shipping': 'tax_ship', 'Services': 'tax_serv'}
_dnakeys = ['dna_' + s.lower().replace(' ', '_').replace('-', '_') for s in SEGS]
for j, g in enumerate(GROUPS):
    mem = [k for k, s in zip(_dnakeys, SEGS) if SEG_GROUP[s] == g]
    put(ws, f"A{DF_['geb0']+j}", f'{g} — EBITDA', fmt=None)
    put(ws, f"A{DF_['gdna0']+j}", f'{g} — share of depreciation and amortisation', fmt=None)
    put(ws, f"A{DF_['gtax0']+j}", f'{g} — taxable profit', fmt=None)
    put(ws, f"A{DF_['gtaxc0']+j}", f'{g} — tax charge at {TAX_G[g]:.0%}', fmt=None)
    _share = '(' + '+'.join(a(k) for k in mem) + ')/(' + '+'.join(a(k) for k in _dnakeys) + ')'
    for i in range(5):
        putf(ws, f"{CD[i]}{DF_['geb0']+j}", f"=Segments!{CD[i]}{SG['geb0']+j}",
             GRP_EB_F[g][i], NUM0, green=True)
        putf(ws, f"{CD[i]}{DF_['gdna0']+j}",
             f"=-{CD[i]}{DF_['dna']}*{_share}", GRP_DNA_F[g][i], NUM0)
        putf(ws, f"{CD[i]}{DF_['gtax0']+j}",
             f"=MAX({CD[i]}{DF_['geb0']+j}-{CD[i]}{DF_['gdna0']+j},0)", GRP_TAXABLE[g][i],
             NUM0)
        putf(ws, f"{CD[i]}{DF_['gtaxc0']+j}",
             f"={CD[i]}{DF_['gtax0']+j}*{a(_TAXKEY[g])}", GRP_TAX[g][i], NUM0)
put(ws, f"A{DF_['taxtot']}", 'Tax on operating profit', bold=True, fmt=None)
put(ws, f"A{DF_['taxrate']}", 'Effective tax rate on operating profit', fmt=None)
for i in range(5):
    putf(ws, f"{CD[i]}{DF_['taxtot']}",
         f"=SUM({CD[i]}{DF_['gtaxc0']}:{CD[i]}{DF_['gtaxc0']+2})", TAX_F[i], NUM0, bold=True)
    putf(ws, f"{CD[i]}{DF_['taxrate']}", f"={CD[i]}{DF_['taxtot']}/{CD[i]}{DF_['ebit']}",
         TAXRATE_F[i], PCT2)
band(ws, DF_['taxtot'], 6)

band(ws, DF_['tvb'], 6)
put(ws, f"A{DF_['tvb']}", 'TERMINAL VALUE AND THE BRIDGE TO EQUITY', bold=True, fmt=None)
_tv = [(DF_['g'], 'Terminal growth', f"={a('g_term')}", G, PCT, True),
       (DF_['ic'], 'Terminal invested capital',
        f"='Balance Sheet'!I{BS['ic']}", IC_F[4], NUM0, True),
       (DF_['roic'], 'Terminal return on invested capital',
        f"=F{DF_['nopat']}/C{DF_['ic']}", DC['roic_t'], PCT, False),
       (DF_['reinv'], 'Required reinvestment rate — terminal growth over the return on '
        'invested capital', f"=C{DF_['g']}/C{DF_['roic']}", DC['reinv'], PCT, False),
       (DF_['nopat1'], 'Terminal-year NOPAT grown one year',
        f"=F{DF_['nopat']}*(1+C{DF_['g']})", DC['nopat_t1'], NUM0, False),
       (DF_['tv'], 'Terminal value — grown NOPAT net of reinvestment, capitalised at the '
        'terminal rate',
        f"=C{DF_['nopat1']}*(1-C{DF_['reinv']})/(C{DF_['waccterm']}-C{DF_['g']})", DC['tv'],
        NUM0, False),
       (DF_['pvex'], 'Present value of the five forecast years',
        f"=SUM(B{DF_['pv']}:F{DF_['pv']})", DC['pv_expl'], NUM0, False),
       (DF_['pvtv'], 'Present value of the terminal value',
        f"=C{DF_['tv']}*F{DF_['df']}", DC['pv_tv'], NUM0, False),
       (DF_['evops'], 'Enterprise value of operations',
        f"=C{DF_['pvex']}+C{DF_['pvtv']}", DC['ev_ops'], NUM0, False),
       (DF_['tvshare'], 'Terminal value as a share of enterprise value',
        f"=C{DF_['pvtv']}/C{DF_['evops']}", DC['tv_share'], PCT, False),
       (DF_['jv'], 'Plus joint ventures and associates at carrying value', f"={a('jv')}",
        JV_BV, NUM0, True),
       (DF_['ev'], 'Enterprise value', f"=C{DF_['evops']}+C{DF_['jv']}", DC['ev'], NUM0,
        False),
       (DF_['nd'], 'Less net debt at 31 March 2026', f"=-{a('nd_co')}", -NDCO, NUM0, True),
       (DF_['defd'], 'Less deferred consideration on acquisitions', f"=-{a('deferred')}",
        -DEFERRED, NUM0, True),
       (DF_['acq'], 'Less the eleven vessels bought on 7 August 2026 — committed and '
        'funded, so their price is carried here and their earnings are in the forecast '
        'above', f"=-{a('acq_cost')}", -ACQ_COST, NUM0, True),
       (DF_['hyb'], 'Less perpetual capital securities at carrying value',
        f"=-{a('hybrid')}", -HYBRID, NUM0, True),
       (DF_['prenci'], 'Equity value before the minorities',
        f"=C{DF_['ev']}+C{DF_['nd']}+C{DF_['defd']}+C{DF_['acq']}+C{DF_['hyb']}",
        DC['pre_nci'], NUM0, False),
       (DF_['ncinav'], 'Minorities arising on the tanker combination — 20% CONTRACTED for '
        'purchase in mid-2027, whose price already sits in the bridge above as deferred '
        'consideration, so it is deducted at that contracted price, not at a share of value',
        f"={a('nci_nav')}", NCI_NAV, NUM0, True),
       (DF_['nciother'], 'The remaining minorities at carrying value',
        f"={a('nci_bv')}-C{DF_['ncinav']}", NCI_OTHER, NUM0, False),
       (DF_['ncishare'], 'Their share of profit — the profit share scaled to their share of '
        'the carried minority balance', f"={a('nci_sh')}*C{DF_['nciother']}/{a('nci_bv')}",
        NCI_SH_OTHER, PCT2, False),
       (DF_['nci'], 'Less non-controlling interests — the contracted slice at its '
        'contracted price, the rest at the greater of book and value',
        f"=-(C{DF_['ncinav']}+MAX(C{DF_['nciother']},"
        f"C{DF_['prenci']}*C{DF_['ncishare']}))", -DC['nci'], NUM0, False),
       (DF_['eq'], 'Equity attributable to ordinary shareholders',
        f"=C{DF_['prenci']}+C{DF_['nci']}", DC['equity'], NUM0, False),
       (DF_['fvusd'], 'Fair value per share (USD)',
        f"=C{DF_['eq']}/{a('shares')}/1000", DC['fv_usd'], PX, False),
       (DF_['fvaed'], 'Fair value per share (AED)', f"=C{DF_['fvusd']}*{a('fx')}",
        DC['fv_aed'], PX, False)]
for rw, lab, fml, xp, fmt, gr in _tv:
    put(ws, f'A{rw}', lab, bold=(rw == DF_['fvaed']), fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, bold=(rw in (DF_['ev'], DF_['eq'], DF_['fvaed'])),
         green=gr)
band(ws, DF_['eq'], 4); band(ws, DF_['fvaed'], 4)

band(ws, DF_['keb'], 6)
put(ws, f"A{DF_['keb']}", 'COST OF EQUITY — BUILT HERE, NOT ASSUMED', bold=True, fmt=None)
_coc = [(DF_['rfobs'], 'Observed government bond yield (dirham tranche, January 2031)',
         f"={a('rf_obs')}", V['rf_observed'], PCT2, True),
        (DF_['sov'], 'Less sovereign default spread — country risk enters once, through the '
         'premium', f"={a('sov')}", V['sov_spread'], PCT2, True),
        (DF_['rfstar'], 'Normalised risk-free rate',
         f"=C{DF_['rfobs']}-C{DF_['sov']}", RF_STAR, PCT2, False),
        (DF_['beta'], 'Beta — own-stock weekly regression against its local index',
         f"={a('beta')}", V['beta'], BETA, True),
        (DF_['erp'], 'Equity risk premium', f"={a('erp')}", V['erp_total'], PCT2, True),
        (DF_['ke'], 'Cost of equity',
         f"=C{DF_['rfstar']}+C{DF_['beta']}*C{DF_['erp']}", KE, PCT2, False)]
for rw, lab, fml, xp, fmt, gr in _coc:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, bold=(rw == DF_['ke']), green=gr)
band(ws, DF_['ke'], 4)

band(ws, DF_['kdb'], 6)
put(ws, f"A{DF_['kdb']}", 'COST OF DEBT — THREE CONSTRUCTIONS, AVERAGED HERE', bold=True,
    fmt=None)
_kd = [(DF_['sofr'], 'Secured overnight financing rate', f"={a('sofr')}", V['sofr'], PCT2,
        True),
       (DF_['shldrm'], 'Parent revolving credit facility margin', f"={a('shldr_m')}",
        V['shldr_margin'], PCT2, True),
       (DF_['kd1'], 'METHOD 1 — parent revolving facility rate drawn January 2026',
        f"=C{DF_['sofr']}+C{DF_['shldrm']}", KD1, PCT2, False),
       (DF_['banklo'], 'Third-party bank loans — low end of the disclosed range',
        f"={a('bank_lo')}", V['bank_loan_lo'], PCT2, True),
       (DF_['bankhi'], 'Third-party bank loans — high end of the disclosed range',
        f"={a('bank_hi')}", V['bank_loan_hi'], PCT2, True),
       (DF_['bankmid'], 'Bank-loan midpoint',
        f"=(C{DF_['banklo']}+C{DF_['bankhi']})/2", KD_BANK, PCT2, False),
       (DF_['othlo'], 'Other third-party borrowings — low end of the disclosed range',
        f"={a('oth_lo')}", V['other_borr_lo'], PCT2, True),
       (DF_['othhi'], 'Other third-party borrowings — high end of the disclosed range',
        f"={a('oth_hi')}", V['other_borr_hi'], PCT2, True),
       (DF_['othmid'], 'Other-borrowings midpoint',
        f"=(C{DF_['othlo']}+C{DF_['othhi']})/2", KD_OTHER, PCT2, False),
       (DF_['tp'], 'Third-party blended rate — the two midpoints averaged',
        f"=(C{DF_['bankmid']}+C{DF_['othmid']})/2", KD_TP, PCT2, False),
       (DF_['leaseint'], 'Lease interest charged in 2025 (USD 000)', f"={a('lease_int')}",
        V['intpaid_lease_fy25'], NUM0, True),
       (DF_['leaseopen'], 'Lease liabilities, opening balance (USD 000)',
        f"={a('lease_open')}", V['lease_open_fy25'], NUM0, True),
       (DF_['leaseclose'], 'Lease liabilities, closing balance (USD 000)',
        f"={a('lease_close')}", V['lease_close_fy25'], NUM0, True),
       (DF_['kdlease'], 'Implied lease borrowing rate — interest over the average balance',
        f"=C{DF_['leaseint']}/((C{DF_['leaseopen']}+C{DF_['leaseclose']})/2)", KD_LEASE,
        PCT2, False),
       (DF_['dshldr'], 'Shareholder loan at 31 March 2026 (USD 000)', f"={a('d_shldr')}",
        V['q1_26_shldr_loan'], NUM0, True),
       (DF_['dborr'], 'Third-party borrowings at 31 March 2026 (USD 000)',
        f"={a('d_borr')}", V['q1_26_borrowings'], NUM0, True),
       (DF_['dlease'], 'Lease liabilities at 31 March 2026 (USD 000)', f"={a('d_lease')}",
        V['q1_26_leases'], NUM0, True),
       (DF_['dtot'], 'Borrowings at 31 March 2026 (USD 000)',
        f"=C{DF_['dshldr']}+C{DF_['dborr']}+C{DF_['dlease']}", DEBT_NOW, NUM0, False),
       (DF_['kd2'], 'METHOD 2 — the instruments actually outstanding, weighted by balance',
        f"=(C{DF_['dshldr']}*C{DF_['kd1']}+C{DF_['dborr']}*C{DF_['tp']}"
        f"+C{DF_['dlease']}*C{DF_['kdlease']})/C{DF_['dtot']}", KD2, PCT2, False),
       (DF_['kd3'], 'METHOD 3 — the disclosed third-party bank-loan midpoint',
        f"=C{DF_['bankmid']}", KD3, PCT2, False),
       (DF_['kd'], 'Cost of debt ADOPTED — method 2, the instruments actually '
        'outstanding weighted by balance. It is the only one of the three that '
        'reproduces from the facility lines, which is what makes it checkable from '
        'outside this model; methods 1 and 3 are published above as what they are, a '
        'marginal drawdown rate and a disclosed range midpoint, and are no longer '
        'averaged into the answer',
        f"=C{DF_['kd2']}", KD, PCT2, False),
       (DF_['kdbal'], 'MEMORANDUM — the retired average of the three constructions, '
        'which reproduces from no set of balances and rates and is shown so a reader '
        'can see what the previous construction said',
        f"=AVERAGE(C{DF_['kd1']},C{DF_['kd2']},C{DF_['kd3']})", KD_RETIRED_AVG,
        PCT2, False),
       (DF_['taxstat'], 'Statutory corporate tax rate', f"={a('tax_stat')}", TAXS, PCT,
        True),
       (DF_['kdat'], 'Cost of debt after tax',
        f"=C{DF_['kd']}*(1-C{DF_['taxstat']})", KD_AT, PCT2, False)]
for rw, lab, fml, xp, fmt, gr in _kd:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, bold=(rw == DF_['kd']), green=gr)
band(ws, DF_['kd'], 4)

band(ws, DF_['wb'], 6)
put(ws, f"A{DF_['wb']}", 'WEIGHTS AND THE COST OF CAPITAL — EQUITY, DEBT AND THE PERPETUAL '
    'SECURITIES', bold=True, fmt=None)
_w = [(DF_['mktcap'], 'Market capitalisation (USD 000)',
       f"={a('spot')}/{a('fx')}*{a('shares')}*1000", MKTCAP, NUM0, False),
      (DF_['borr'], 'Borrowings at 31 March 2026 (USD 000)', f"=C{DF_['dtot']}", DEBT_NOW,
       NUM0, False),
      (DF_['hybcap'], 'Perpetual capital securities at carrying value (USD 000)',
       f"={a('hybrid')}", HYBRID_CAP, NUM0, True),
      (DF_['captot'], 'Total capital (USD 000)',
       f"=C{DF_['mktcap']}+C{DF_['borr']}+C{DF_['hybcap']}", CAP_TOT, NUM0, False),
      (DF_['we'], 'Equity weight — market capitalisation over the total',
       f"=C{DF_['mktcap']}/C{DF_['captot']}", WE, PCT2, False),
      (DF_['wd'], 'Debt weight', f"=C{DF_['borr']}/C{DF_['captot']}", WD, PCT2, False),
      (DF_['whyb'], 'Perpetual capital securities weight', f"=C{DF_['hybcap']}"
       f"/C{DF_['captot']}", WH, PCT2, False),
      (DF_['kh'], 'Cost of the perpetual capital securities — their own coupon. It is not '
       'tax-deductible, because it is an equity distribution, so it is not taxed down',
       f"=C{DF_['sofr']}+{a('hyb_m')}", KH, PCT2, False),
      (DF_['wacc'], 'Cost of capital — explicit window',
       f"=C{DF_['we']}*C{DF_['ke']}+C{DF_['wd']}*C{DF_['kdat']}"
       f"+C{DF_['whyb']}*C{DF_['kh']}", W_EXP, PCT2, False),
      (DF_['rfterm'], 'Terminal risk-free rate', f"={a('rf_term')}", V['rf_terminal'], PCT2,
       True),
      (DF_['keterm'], 'Terminal cost of equity',
       f"=C{DF_['rfterm']}+C{DF_['beta']}*C{DF_['erp']}", KE_T, PCT2, False),
      (DF_['kdterm'], 'Terminal cost of debt — the same spread over the terminal rate',
       f"=C{DF_['rfterm']}+(C{DF_['kd']}-C{DF_['rfstar']})", KD_T, PCT2, False),
      (DF_['kdtermat'], 'Terminal cost of debt after tax',
       f"=C{DF_['kdterm']}*(1-C{DF_['taxstat']})", KD_T_AT, PCT2, False),
      (DF_['khterm'], 'Terminal cost of the perpetual securities — the coupon floats, so '
       'it normalises with the risk-free rate',
       f"=C{DF_['rfterm']}+{a('hyb_m')}", KH_T, PCT2, False),
      (DF_['waccterm'], 'Terminal cost of capital',
       f"=C{DF_['we']}*C{DF_['keterm']}+C{DF_['wd']}*C{DF_['kdtermat']}"
       f"+C{DF_['whyb']}*C{DF_['khterm']}", W_TERM, PCT2, False)]
for rw, lab, fml, xp, fmt, gr in _w:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt,
         bold=(rw in (DF_['wacc'], DF_['waccterm'])), green=gr)
band(ws, DF_['wacc'], 4); band(ws, DF_['waccterm'], 4)
note(ws, f"A{DF_['waccterm']+1}", 'The perpetual capital securities are deducted in the '
     'equity bridge above as a claim ranking ahead of the ordinary shares. A claim that is '
     'deducted from enterprise value must also be weighted in the cost of capital at its '
     'own cost — those are two halves of one treatment, not a double count.')

band(ws, DF_['ab'], 6)
put(ws, f"A{DF_['ab']}", 'THE CONTESTED JUDGEMENT — THE SAME MODEL ON THE COMPOSITE-INDEX '
    'BETA', bold=True, fmt=None)
for rw, lab, fml, xp, fmt, gr in [
        (DF_['betaa'], 'Beta against an equal-weight composite of the exchange\'s names — '
         'the disclosed alternative construction', f"={a('beta_a')}", V['beta_composite'],
         BETA, True),
        (DF_['kea'], 'Cost of equity on the composite-index beta',
         f"=C{DF_['rfstar']}+C{DF_['betaa']}*C{DF_['erp']}", KE_A, PCT2, False),
        (DF_['keta'], 'Terminal cost of equity on the composite-index beta',
         f"=C{DF_['rfterm']}+C{DF_['betaa']}*C{DF_['erp']}", KE_T_A, PCT2, False),
        (DF_['wacca'], 'Cost of capital — explicit window, composite-index beta. Only the '
         'cost of EQUITY changes: the same three tranches of capital are carried, because '
         'how the market is measured does not change what the company is financed with',
         f"=C{DF_['we']}*C{DF_['kea']}+C{DF_['wd']}*C{DF_['kdat']}"
         f"+C{DF_['whyb']}*C{DF_['kh']}", W_EXP_A, PCT2, False),
        (DF_['wactermsa'], 'Terminal cost of capital, composite-index beta',
         f"=C{DF_['we']}*C{DF_['keta']}+C{DF_['wd']}*C{DF_['kdtermat']}"
         f"+C{DF_['whyb']}*C{DF_['khterm']}", W_TERM_A, PCT2, False)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)
band(ws, DF_['cib'], 6)
put(ws, f"A{DF_['cib']}", 'HOW PRECISE THE PRIMARY BETA IS — THE INTERVAL AROUND IT, '
    'PRICED', bold=True, fmt=None)
for rw, lab, fml, xp, fmt, gr in [
        (DF_['cilo'], 'Beta — lower bound of the 90% confidence interval on the primary '
         'regression', f"={a('beta_ci_lo')}", V['beta_ci_lo'], BETA, True),
        (DF_['cihi'], 'Beta — upper bound of the 90% confidence interval on the primary '
         'regression', f"={a('beta_ci_hi')}", V['beta_ci_hi'], BETA, True),
        (DF_['kecilo'], 'Cost of equity at the lower confidence bound — the bull-case '
         'discount rate', f"=C{DF_['rfstar']}+C{DF_['cilo']}*C{DF_['erp']}", KE_CI_LO,
         PCT2, False),
        (DF_['kecihi'], 'Cost of equity at the upper confidence bound — the bear-case '
         'discount rate', f"=C{DF_['rfstar']}+C{DF_['cihi']}*C{DF_['erp']}", KE_CI_HI,
         PCT2, False),
        (DF_['blume'], 'Beta — the measured slope shrunk toward the market, two-thirds of '
         'it plus one-third of 1.0',
         f"={a('beta_blume')}", V['beta_blume'], BETA, True),
        (DF_['keblume'], 'Cost of equity on the slope shrunk toward the market',
         f"=C{DF_['rfstar']}+C{DF_['blume']}*C{DF_['erp']}", KE_BLUME, PCT2, False)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=gr)
hdr(ws, DF_['ahdr'],
    ['The same cash flows, discounted at the composite-index cost of capital'] + YFE)
wf(DF_['glidea'], 'Forward cost of capital — the glide',
   lambda i: (f"=$C${DF_['wacca']}+($C${DF_['wactermsa']}-$C${DF_['wacca']})*{i+1}/5"),
   DA['glide'], PCT2)
wf(DF_['dfa'], 'Discount factor',
   lambda i: (f"=1/(1+{CD[i]}{DF_['glidea']})^{a('stub')}" if i == 0
              else f"={CD[i-1]}{DF_['dfa']}/(1+{CD[i]}{DF_['glidea']})"), DA['df'], DF4)
wf(DF_['pva'], 'Present value of free cash flow',
   lambda i: f"={CD[i]}{DF_['fcfd']}*{CD[i]}{DF_['dfa']}", DA['pv'])
for rw, lab, fml, xp, fmt in [
        (DF_['pvexa'], 'Present value of the five forecast years',
         f"=SUM(B{DF_['pva']}:F{DF_['pva']})", DA['pv_expl'], NUM0),
        (DF_['tva'], 'Terminal value',
         f"=C{DF_['nopat1']}*(1-C{DF_['reinv']})/(C{DF_['wactermsa']}-C{DF_['g']})",
         DA['tv'], NUM0),
        (DF_['pvtva'], 'Present value of the terminal value',
         f"=C{DF_['tva']}*F{DF_['dfa']}", DA['pv_tv'], NUM0),
        (DF_['evopsa'], 'Enterprise value of operations',
         f"=C{DF_['pvexa']}+C{DF_['pvtva']}", DA['ev_ops'], NUM0),
        (DF_['tvsharea'], 'Terminal value as a share of enterprise value',
         f"=C{DF_['pvtva']}/C{DF_['evopsa']}", DA['tv_share'], PCT),
        (DF_['eva'], 'Enterprise value', f"=C{DF_['evopsa']}+C{DF_['jv']}", DA['ev'], NUM0),
        (DF_['prencia'], 'Equity value before the minorities',
         f"=C{DF_['eva']}+C{DF_['nd']}+C{DF_['defd']}+C{DF_['acq']}+C{DF_['hyb']}",
         DA['pre_nci'], NUM0),
        (DF_['ncia'], 'Less non-controlling interests — the same two-part deduction',
         f"=-(C{DF_['ncinav']}+MAX(C{DF_['nciother']},"
         f"C{DF_['prencia']}*C{DF_['ncishare']}))", -DA['nci'], NUM0),
        (DF_['eqa'], 'Equity attributable to ordinary shareholders',
         f"=C{DF_['prencia']}+C{DF_['ncia']}", DA['equity'], NUM0),
        (DF_['fvaeda'], 'Fair value per share (AED) — composite-index beta',
         f"=C{DF_['eqa']}/{a('shares')}/1000*{a('fx')}", DA['fv_aed'], PX)]:
    put(ws, f'A{rw}', lab, bold=(rw == DF_['fvaeda']), fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, bold=(rw == DF_['fvaeda']))
band(ws, DF_['fvaeda'], 4)
note(ws, f"A{DF_['fvaeda']+2}", 'The two readings above are carried side by side and are '
     'never averaged. They are the SAME regression of the same weekly returns; only the '
     'measure of the market differs. The published index of the exchange is what the '
     'method asks for and is the primary reading — it is weighted by size, and is '
     'therefore dominated by the same large-capitalisation group this company belongs to. '
     'The equal-weight composite gives the exchange\'s smallest names the same say as its '
     'largest and is the construction an earlier version of this study used; it is '
     'published beside the primary reading because a gap of this size is a fact about how '
     'the index is built, not about the company, and the reader is entitled to see it. '
     'The gap between the two is the single most consequential judgement in this study.')

# ============ 9 INCOME STATEMENT =================================================
ws = sheet('Income Statement')
title(ws, 'Income statement — three years audited, five years forecast', 'USD thousand, '
      'consolidated. History is the audited record; every forecast line is a formula.', 9,
      awidth=52, cwidth=13)
hdr(ws, 4, ['USD 000'] + YH + YFE)


def line(rw, lab, hist_v, hist_f, fc_f, fc_v, fmt=NUM0, bd=False, green=False):
    put(ws, f'A{rw}', lab, bold=bd, fmt=None)
    for i in range(3):
        if hist_f is not None:
            putf(ws, f'{HC[i]}{rw}', hist_f(i), hist_v[i], fmt, bold=bd)
        else:
            put(ws, f'{HC[i]}{rw}', hist_v[i], BLUE, fmt, bold=bd)
    for i in range(5):
        if fc_f is None:
            put(ws, f'{FCOL[i]}{rw}', '-', BLACK, fmt, bold=bd)
        else:
            putf(ws, f'{FCOL[i]}{rw}', fc_f(i), fc_v[i], fmt, bold=bd, green=green)
    if bd:
        band(ws, rw, 9)


# the audited revenue line is the sum of the disclosed operating segments, which are the
# primary record and are carried on the Segments sheet — so it is added up here rather than
# typed a second time
line(IS['rev'], 'Revenue', HI['revenue'],
     lambda i: f"=Segments!{HC[i]}{SG['revht']}",
     lambda i: f"=DCF!{CD[i]}{DF_['rev']}", REV_F, bd=True, green=True)
line(IS['dc'], 'Direct costs', HI['direct_costs'], None, None, None)
line(IS['gp'], 'Gross profit', HI['gross_profit'],
     lambda i: f"={HC[i]}{IS['rev']}+{HC[i]}{IS['dc']}", None, None)
line(IS['ga'], 'General and administrative expenses', HI['ga'], None, None, None)
line(IS['ecl'], 'Expected credit losses', HI['ecl'], None, None, None)
line(IS['oi'], 'Other income', HI['other_income'], None, None, None)
line(IS['oe'], 'Other expenses', HI['other_expenses'], None, None, None)
line(IS['op'], 'Operating profit', HI['ebit'],
     lambda i: (f"={HC[i]}{IS['gp']}+{HC[i]}{IS['ga']}+{HC[i]}{IS['ecl']}"
                f"+{HC[i]}{IS['oi']}+{HC[i]}{IS['oe']}"),
     lambda i: f"={FCOL[i]}{IS['ebitda']}-{FCOL[i]}{IS['dna']}", EBIT_F, bd=True)
line(IS['dna'], 'Depreciation and amortisation', H_DNA, None,
     lambda i: f"='Balance Sheet'!{CD[i]}{BS['dnatot']}", DNA_F, green=True)
line(IS['ebitda'], 'EBITDA (operating)', H_EBITDA,
     lambda i: f"={HC[i]}{IS['op']}+{HC[i]}{IS['dna']}",
     lambda i: f"=DCF!{CD[i]}{DF_['ebitda']}", EB_F, bd=True, green=True)
# the reported-EBITDA bridge is the sum of three lines already on this sheet — the share of
# joint ventures and the two 2025 acquisition items — so it is added up rather than typed
_ebjv = [HI['ebitda_bridge']['share_of_jv'][i] + HI['ebitda_bridge']['one_offs'][i]
         for i in range(3)]
line(IS['ebjv'], 'Add share of joint ventures and one-off gains carried in reported EBITDA',
     _ebjv,
     lambda i: (f"={HC[i]}{IS['assoc']}+{HC[i]}{IS['bargain']}+{HC[i]}{IS['prevheld']}"),
     None, None)
line(IS['ebrep'], 'EBITDA as reported', HI['ebitda_reported'],
     lambda i: f"={HC[i]}{IS['ebitda']}+{HC[i]}{IS['ebjv']}",
     lambda i: f"={FCOL[i]}{IS['ebitda']}", EB_F)
line(IS['opcost'], 'Total operating costs', H_OPCOST,
     lambda i: f"={HC[i]}{IS['rev']}-{HC[i]}{IS['ebitda']}",
     lambda i: f"={FCOL[i]}{IS['rev']}-{FCOL[i]}{IS['ebitda']}", OPCOST_F)
line(IS['mgn'], 'EBITDA margin (operating)',
     [H_EBITDA[i] / H_REV[i] for i in range(3)],
     lambda i: f"={HC[i]}{IS['ebitda']}/{HC[i]}{IS['rev']}",
     lambda i: f"={FCOL[i]}{IS['ebitda']}/{FCOL[i]}{IS['rev']}",
     [EB_F[i] / REV_F[i] for i in range(5)], PCT)
line(IS['assoc'], 'Share of joint ventures and associates', HI['assoc'], None, None, None)
line(IS['bargain'], 'Gain on bargain purchase',
     [0, 0, V['bargain_fy25']], None, None, None)
line(IS['prevheld'], 'Loss on the previously held interest',
     [0, 0, V['prevheld_fy25']], None, None, None)
line(IS['fininc'], 'Finance income', HI['fin_income'], None,
     lambda i: f"={a('sofr')}*{a('cash')}", FININC_F, green=True)
line(IS['fincost'], 'Finance costs', HI['fin_costs'], None,
     lambda i: f"=-DCF!$C${DF_['kd']}*'Balance Sheet'!{CD[i]}{BS['ndgross']}",
     [-x for x in INT_F], green=True)
line(IS['pbt'], 'Profit before tax', HI['pbt'],
     lambda i: (f"={HC[i]}{IS['op']}+{HC[i]}{IS['assoc']}+{HC[i]}{IS['bargain']}"
                f"+{HC[i]}{IS['prevheld']}+{HC[i]}{IS['fininc']}+{HC[i]}{IS['fincost']}"),
     lambda i: f"={FCOL[i]}{IS['op']}+{FCOL[i]}{IS['fininc']}+{FCOL[i]}{IS['fincost']}",
     PBT_F, bd=True)
line(IS['tax'], 'Income tax', HI['tax'], None,
     lambda i: f"=-{FCOL[i]}{IS['pbt']}*DCF!{CD[i]}{DF_['taxrate']}", [-x for x in TAXP_F])
line(IS['pat'], 'Profit for the year', HI['pat'],
     lambda i: f"={HC[i]}{IS['pbt']}+{HC[i]}{IS['tax']}",
     lambda i: f"={FCOL[i]}{IS['pbt']}+{FCOL[i]}{IS['tax']}", PAT_F)
line(IS['nci'], 'Non-controlling interests', [0, 0, -V['nci_pl_fy25']], None,
     lambda i: f"=-{FCOL[i]}{IS['pat']}*{a('nci_sh')}", [-x for x in NCI_F])
line(IS['npa'], 'Profit attributable to shareholders', H_NPA,
     lambda i: f"={HC[i]}{IS['pat']}+{HC[i]}{IS['nci']}",
     lambda i: f"={FCOL[i]}{IS['pat']}+{FCOL[i]}{IS['nci']}", NPA_F, bd=True)
line(IS['hybcpn'], 'Perpetual capital securities coupon',
     [0, 0, -V['hybrid_coupon_fy25']], None,
     lambda i: f"=-{a('hybrid')}*({a('sofr')}+{a('hyb_m')})", [-HYB_CPN] * 5)
line(IS['ordn'], 'Earnings attributable to ordinary shareholders, AFTER the perpetual '
     'capital securities coupon', H_ORD,
     lambda i: f"={HC[i]}{IS['npa']}+{HC[i]}{IS['hybcpn']}",
     lambda i: f"={FCOL[i]}{IS['npa']}+{FCOL[i]}{IS['hybcpn']}", ORD_F, bd=True)
put(ws, f"A{IS['eps']}", 'Earnings per ordinary share, AFTER the perpetual coupon (USD) — '
    'the coupon ranks ahead of the ordinary shares, so it comes out of the numerator',
    fmt=None)
put(ws, f"A{IS['epsaed']}", 'Earnings per ordinary share, AFTER the perpetual coupon (AED)',
    fmt=None)
put(ws, f"A{IS['epspre']}", 'Memorandum — earnings per share BEFORE the perpetual coupon '
    '(USD), on attributable profit; shown so the two are never confused', fmt=None)
for i in range(8):
    putf(ws, f"{ALL[i]}{IS['eps']}", f"={ALL[i]}{IS['ordn']}/{a('shares')}/1000",
         ORD_ALL[i] / SH / 1000.0, PX3)
    putf(ws, f"{ALL[i]}{IS['epsaed']}", f"={ALL[i]}{IS['eps']}*{a('fx')}",
         ORD_ALL[i] / SH / 1000.0 * PEG, PX3)
    putf(ws, f"{ALL[i]}{IS['epspre']}", f"={ALL[i]}{IS['npa']}/{a('shares')}/1000",
         NPA_ALL[i] / SH / 1000.0, PX3)
note(ws, f"A{IS['epspre']+2}", 'Every FY2023-25 line above is the audited figure. In the '
     'forecast the company\'s own disclosure-only lines — direct costs, gross profit, the '
     'general and administrative split, the share of joint ventures and the two 2025 '
     'acquisition items — are not projected, because the forecast is built at the business-'
     'unit level on earnings before depreciation rather than on a cost-line split the '
     'filings do not support forward. The finance charge is computed on gross borrowings, '
     'which move with the net-debt roll, so profit is struck after interest and differs '
     'from the pre-financing discounted-cash-flow waterfall by construction.')

# ============ 10 BALANCE SHEET ====================================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — condensed, and the rolls that drive it', 'USD thousand, '
      'consolidated. Every FY2023-25 line is the audited closing figure; every forecast '
      'line is rolled forward from a driver.', 9, awidth=52, cwidth=13)
hdr(ws, 4, ['USD 000'] + YH + YFE)


def bline(rw, lab, hist_v, fc_f, fc_v, fmt=NUM0, bd=False, hist_f=None, green=False):
    put(ws, f'A{rw}', lab, bold=bd, fmt=None)
    for i in range(3):
        if hist_f is not None:
            putf(ws, f'{HC[i]}{rw}', hist_f(i), hist_v[i], fmt, bold=bd)
        else:
            put(ws, f'{HC[i]}{rw}', hist_v[i], BLUE, fmt, bold=bd)
    for i in range(5):
        if fc_f is None:
            put(ws, f'{FCOL[i]}{rw}', '-', BLACK, fmt, bold=bd)
        else:
            putf(ws, f'{FCOL[i]}{rw}', fc_f(i), fc_v[i], fmt, bold=bd, green=green)
    if bd:
        band(ws, rw, 9)


bline(BS['ppe'], 'Property, plant and equipment', HB['ppe'],
      lambda i: f"={CD[i]}{BS['ppeclose']}", PPE_CLOSE)
bline(BS['rou'], 'Right-of-use assets', HB['rou'], None, None)
bline(BS['intang'], 'Intangible assets', HB['intangibles'], lambda i: f"={a('intang')}",
      [INTANG] * 5, green=True)
bline(BS['gw'], 'Goodwill', HB['goodwill'], lambda i: f"={a('gw')}", [GW] * 5, green=True)
bline(BS['invprop'], 'Investment properties', HB['inv_prop'], None, None)
bline(BS['jv'], 'Investments in joint ventures and associates', HB['jv'],
      lambda i: f"={a('jv')}", [JV_BV] * 5, green=True)
bline(BS['inv'], 'Inventories', HB['inventories'], lambda i: f"={CD[i]}{BS['wcinv']}",
      INV_F)
bline(BS['recv'], 'Trade and other receivables, including amounts due from related parties',
      [HB['receivables'][i] + HB['due_from_related'][i] for i in range(3)],
      lambda i: f"={CD[i]}{BS['wcrecv']}", RECV_F)
bline(BS['cash'], 'Cash and cash equivalents', HB['cash'], lambda i: f"={a('cash')}",
      [CASH] * 5, green=True)
bline(BS['ta'], 'Total assets', HB['total_assets'], None, None, bd=True)
bline(BS['pay'], 'Trade and other payables, including amounts due to related parties',
      [HB['payables'][i] + HB['due_to_related'][i] for i in range(3)],
      lambda i: f"={CD[i]}{BS['wcpay']}", PAY_F)
put(ws, f"A{BS['nwc']}", 'Net working capital', fmt=None)
for i in range(8):
    putf(ws, f"{ALL[i]}{BS['nwc']}",
         f"={ALL[i]}{BS['recv']}+{ALL[i]}{BS['inv']}-{ALL[i]}{BS['pay']}", NWC_ALL[i], NUM0)
bline(BS['grossd'], 'Gross borrowings', HB['debt'],
      lambda i: f"={CD[i]}{BS['ndgross']}", GROSS_D)
bline(BS['nd'], 'Net debt', HB['net_debt'], lambda i: f"={CD[i]}{BS['ndclose']}",
      ND_CLOSE, bd=True,
      # the audited years are an identity off two lines already on this sheet — gross
      # borrowings less the cash held — so they are computed here rather than typed again
      hist_f=lambda i: f"={HC[i]}{BS['grossd']}-{HC[i]}{BS['cash']}")
bline(BS['hyb'], 'Perpetual capital securities', HB['hybrid'], lambda i: f"={a('hybrid')}",
      [HYBRID] * 5, green=True)
bline(BS['nci'], 'Non-controlling interests', HB['nci'], lambda i: f"={a('nci_bv')}",
      [NCI_BV] * 5, green=True)
bline(BS['eqp'], 'Equity attributable to shareholders', HB['equity_parent'],
      lambda i: f"={CD[i]}{BS['eqclose']}", EQ_CLOSE, bd=True)
put(ws, f"A{BS['teq']}", 'Total equity', bold=True, fmt=None)
_teq = [HB['total_equity'][i] for i in range(3)] + [EQ_CLOSE[i] + HYBRID + NCI_BV
                                                   for i in range(5)]
for i in range(8):
    putf(ws, f"{ALL[i]}{BS['teq']}",
         f"={ALL[i]}{BS['eqp']}+{ALL[i]}{BS['hyb']}+{ALL[i]}{BS['nci']}", _teq[i], NUM0,
         bold=True)
band(ws, BS['teq'], 9)
put(ws, f"A{BS['ndeb']}", 'Net debt / EBITDA', fmt=None)
put(ws, f"A{BS['bvps']}", 'Book value per share (USD)', fmt=None)
put(ws, f"A{BS['bvpsaed']}", 'Book value per share (AED)', fmt=None)
put(ws, f"A{BS['ic']}", 'Invested capital', fmt=None)
for i in range(8):
    putf(ws, f"{ALL[i]}{BS['ndeb']}",
         f"={ALL[i]}{BS['nd']}/'Income Statement'!{ALL[i]}{IS['ebitda']}",
         ND_ALL[i] / EB_ALL[i], MULT)
    putf(ws, f"{ALL[i]}{BS['bvps']}", f"={ALL[i]}{BS['eqp']}/{a('shares')}/1000",
         EQ_ALL[i] / SH / 1000.0, PX3)
    putf(ws, f"{ALL[i]}{BS['bvpsaed']}", f"={ALL[i]}{BS['bvps']}*{a('fx')}",
         EQ_ALL[i] / SH / 1000.0 * PEG, PX)
    putf(ws, f"{ALL[i]}{BS['ic']}",
         f"={ALL[i]}{BS['ppe']}+{ALL[i]}{BS['nwc']}+{ALL[i]}{BS['intang']}"
         f"+{ALL[i]}{BS['gw']}", IC_ALL[i], NUM0)
put(ws, f"A{BS['roic']}", 'Return on invested capital', fmt=None)
put(ws, f"A{BS['roe']}", 'Return on equity (profit over average equity)', fmt=None)
for i in range(3):
    put(ws, f"{HC[i]}{BS['roic']}", '-', BLACK, PCT)
for i in range(5):
    putf(ws, f"{FCOL[i]}{BS['roic']}",
         f"=DCF!{CD[i]}{DF_['nopat']}/{FCOL[i]}{BS['ic']}", ROIC_F[i], PCT)
put(ws, f"B{BS['roe']}", '-', BLACK, PCT)
_hroe = [None] + [H_NPA[i] / ((H_EQ[i - 1] + H_EQ[i]) / 2) for i in (1, 2)]
for i in (1, 2):
    putf(ws, f"{HC[i]}{BS['roe']}",
         f"='Income Statement'!{HC[i]}{IS['npa']}/(({HC[i-1]}{BS['eqp']}"
         f"+{HC[i]}{BS['eqp']})/2)", _hroe[i], PCT)
for i in range(5):
    prev = f"{a('eqp0')}" if i == 0 else f"{FCOL[i-1]}{BS['eqp']}"
    putf(ws, f"{FCOL[i]}{BS['roe']}",
         f"='Income Statement'!{FCOL[i]}{IS['npa']}/(({prev}+{FCOL[i]}{BS['eqp']})/2)",
         ROE_F[i], PCT)

band(ws, BS['ppeb'], 6)
put(ws, f"A{BS['ppeb']}", 'THE PROPERTY, PLANT AND EQUIPMENT ROLL', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=BS['ppeb'], column=2 + i, value=y); cc.font = Font(bold=True)
    cc.fill = FILL_G
put(ws, f"A{BS['ppeopen']}", 'Opening property, plant and equipment', fmt=None)
put(ws, f"A{BS['ppecapex']}", 'Capital expenditure', fmt=None)
put(ws, f"A{BS['ppeacq']}", 'Vessels acquired in the purchase announced 7 August 2026 — the '
    'same price that is carried in net debt below', fmt=None)
put(ws, f"A{BS['ppedeprate']}", 'Depreciation rate on property, plant and equipment',
    fmt=None)
put(ws, f"A{BS['ppedep1']}", 'First-pass depreciation, on the opening balance plus the full '
    'additions', fmt=None)
put(ws, f"A{BS['ppedep']}", 'Depreciation on property, plant and equipment — on the average '
    'of opening and closing', fmt=None)
put(ws, f"A{BS['ppeclose']}", 'Closing property, plant and equipment', bold=True, fmt=None)
put(ws, f"A{BS['otherdna']}", 'Other depreciation and amortisation, escalated', fmt=None)
put(ws, f"A{BS['dnatot']}", 'Total depreciation and amortisation', bold=True, fmt=None)
for i in range(5):
    c = CD[i]
    f_open = (f"={a('ppe_dummy')}" if False else
              (f"=D{BS['ppe']}" if i == 0 else f"={CD[i-1]}{BS['ppeclose']}"))
    putf(ws, f"{c}{BS['ppeopen']}", f_open, PPE_OPEN[i], NUM0, green=(i == 0))
    putf(ws, f"{c}{BS['ppecapex']}", f"={a('capex', col=c)}", CAPEX[i], NUM0, green=True)
    putf(ws, f"{c}{BS['ppeacq']}", f"={a('acq_cost')}" if i == 0 else '=0', ACQ_CAPEX[i],
         NUM0, green=(i == 0))
    putf(ws, f"{c}{BS['ppedeprate']}", f"={a('dep_rate')}", DEP_RATE, PCT2, green=True)
    putf(ws, f"{c}{BS['ppedep1']}",
         f"={c}{BS['ppedeprate']}*({c}{BS['ppeopen']}+({c}{BS['ppeopen']}"
         f"+{c}{BS['ppecapex']}+{c}{BS['ppeacq']}))/2", DEP1[i], NUM0)
    putf(ws, f"{c}{BS['ppedep']}",
         f"={c}{BS['ppedeprate']}*({c}{BS['ppeopen']}+({c}{BS['ppeopen']}"
         f"+{c}{BS['ppecapex']}+{c}{BS['ppeacq']}-{c}{BS['ppedep1']}))/2", DEP_PPE[i],
         NUM0)
    putf(ws, f"{c}{BS['ppeclose']}",
         f"={c}{BS['ppeopen']}+{c}{BS['ppecapex']}+{c}{BS['ppeacq']}-{c}{BS['ppedep']}",
         PPE_CLOSE[i], NUM0, bold=True)
    putf(ws, f"{c}{BS['otherdna']}", f"={a('other_dna')}*{esc_chain(i)}",
         OTHER_DNA_Y[i], NUM0)
    putf(ws, f"{c}{BS['dnatot']}", f"={c}{BS['ppedep']}+{c}{BS['otherdna']}", DNA_F[i],
         NUM0, bold=True)
band(ws, BS['dnatot'], 6)

band(ws, BS['wcb'], 6)
put(ws, f"A{BS['wcb']}", 'THE WORKING-CAPITAL ROLL FROM THE DAYS RATIOS', bold=True,
    fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=BS['wcb'], column=2 + i, value=y); cc.font = Font(bold=True)
    cc.fill = FILL_G
for rw, lab in [(BS['wcrev'], 'Revenue'), (BS['wcopcost'], 'Total operating costs'),
                (BS['wcdso'], 'Days sales outstanding'),
                (BS['wcdio'], 'Days inventory outstanding'),
                (BS['wcdpo'], 'Days payable outstanding'),
                (BS['wcrecv'], 'Trade and other receivables — revenue x days / 365'),
                (BS['wcinv'], 'Inventories — operating cost x days / 365'),
                (BS['wcpay'], 'Trade and other payables — operating cost x days / 365'),
                (BS['wcnwc'], 'Net working capital'),
                (BS['wcdnwc'], 'Change in net working capital')]:
    put(ws, f'A{rw}', lab, bold=(rw in (BS['wcnwc'], BS['wcdnwc'])), fmt=None)
for i in range(5):
    c = CD[i]
    putf(ws, f"{c}{BS['wcrev']}", f"=DCF!{c}{DF_['rev']}", REV_F[i], NUM0, green=True)
    putf(ws, f"{c}{BS['wcopcost']}", f"={c}{BS['wcrev']}-DCF!{c}{DF_['ebitda']}",
         OPCOST_F[i], NUM0)
    putf(ws, f"{c}{BS['wcdso']}", f"={a('dso')}", DSO, NUM1, green=True)
    putf(ws, f"{c}{BS['wcdio']}", f"={a('dio')}", DIO, NUM1, green=True)
    putf(ws, f"{c}{BS['wcdpo']}", f"={a('dpo')}", DPO, NUM1, green=True)
    putf(ws, f"{c}{BS['wcrecv']}", f"={c}{BS['wcrev']}*{c}{BS['wcdso']}/365", RECV_F[i],
         NUM0)
    putf(ws, f"{c}{BS['wcinv']}", f"={c}{BS['wcopcost']}*{c}{BS['wcdio']}/365", INV_F[i],
         NUM0)
    putf(ws, f"{c}{BS['wcpay']}", f"={c}{BS['wcopcost']}*{c}{BS['wcdpo']}/365", PAY_F[i],
         NUM0)
    putf(ws, f"{c}{BS['wcnwc']}",
         f"={c}{BS['wcrecv']}+{c}{BS['wcinv']}-{c}{BS['wcpay']}", NWC_F[i], NUM0, bold=True)
    prev = a('nwc25') if i == 0 else f"{CD[i-1]}{BS['wcnwc']}"
    putf(ws, f"{c}{BS['wcdnwc']}", f"={c}{BS['wcnwc']}-{prev}", DNWC_F[i], NUM0, bold=True)
band(ws, BS['wcnwc'], 6)

band(ws, BS['ndb'], 6)
put(ws, f"A{BS['ndb']}", 'THE NET-DEBT ROLL', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=BS['ndb'], column=2 + i, value=y); cc.font = Font(bold=True)
    cc.fill = FILL_G
for rw, lab in [(BS['ndopen'], 'Opening net debt — including the deferred consideration '
                 'and the price of the vessels bought on 7 August 2026'),
                (BS['ndgross'], 'Gross borrowings — opening net debt plus the cash held'),
                (BS['ndint'], 'Interest charge on gross borrowings'),
                (BS['ndfcff'], 'Free cash flow to the firm'),
                (BS['ndintat'], 'Less interest after tax'),
                (BS['ndfi'], 'Plus finance income after tax'),
                (BS['ndcpn'], 'Less perpetual capital securities coupon'),
                (BS['ndfcfe'], 'Free cash flow to equity'),
                (BS['nddps'], 'Less ordinary dividends'),
                (BS['ndclose'], 'Closing net debt')]:
    put(ws, f'A{rw}', lab, bold=(rw in (BS['ndfcfe'], BS['ndclose'])), fmt=None)
for i in range(5):
    c = CD[i]
    f_open = (f"={a('nd_co')}+{a('deferred')}+{a('acq_cost')}" if i == 0
              else f"={CD[i-1]}{BS['ndclose']}")
    putf(ws, f"{c}{BS['ndopen']}", f_open, ND_OPEN[i], NUM0, green=(i == 0))
    putf(ws, f"{c}{BS['ndgross']}", f"={c}{BS['ndopen']}+{a('cash')}", GROSS_D[i], NUM0)
    putf(ws, f"{c}{BS['ndint']}", f"=DCF!$C${DF_['kd']}*{c}{BS['ndgross']}", INT_F[i], NUM0)
    putf(ws, f"{c}{BS['ndfcff']}", f"=DCF!{c}{DF_['fcff']}", FCFF_F[i], NUM0, green=True)
    putf(ws, f"{c}{BS['ndintat']}", f"=-{c}{BS['ndint']}*(1-{a('tax_stat')})",
         -INT_F[i] * (1 - TAXS), NUM0)
    putf(ws, f"{c}{BS['ndfi']}",
         f"={a('sofr')}*{a('cash')}*(1-{a('tax_stat')})", FININC_F[i] * (1 - TAXS), NUM0)
    putf(ws, f"{c}{BS['ndcpn']}", f"=-{a('hybrid')}*({a('sofr')}+{a('hyb_m')})", -HYB_CPN,
         NUM0)
    putf(ws, f"{c}{BS['ndfcfe']}",
         f"={c}{BS['ndfcff']}+{c}{BS['ndintat']}+{c}{BS['ndfi']}+{c}{BS['ndcpn']}",
         FCFE_F[i], NUM0, bold=True)
    putf(ws, f"{c}{BS['nddps']}", f"=-{a('dps26')}*(1+{a('div_g')})^{i}", -DPS[i], NUM0)
    putf(ws, f"{c}{BS['ndclose']}",
         f"={c}{BS['ndopen']}-{c}{BS['ndfcfe']}-{c}{BS['nddps']}", ND_CLOSE[i], NUM0,
         bold=True)
band(ws, BS['ndclose'], 6)

band(ws, BS['eqb'], 6)
put(ws, f"A{BS['eqb']}", 'THE EQUITY ROLL', bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=BS['eqb'], column=2 + i, value=y); cc.font = Font(bold=True)
    cc.fill = FILL_G
for rw, lab in [(BS['eqopen'], 'Opening equity attributable to shareholders'),
                (BS['eqnpa'], 'Add profit attributable to shareholders'),
                (BS['eqdps'], 'Less ordinary dividends'),
                (BS['eqcpn'], 'Less perpetual capital securities coupon'),
                (BS['eqclose'], 'Closing equity attributable to shareholders'),
                (BS['dpsps'], 'Ordinary dividend per share (USD)')]:
    put(ws, f'A{rw}', lab, bold=(rw == BS['eqclose']), fmt=None)
for i in range(5):
    c = CD[i]
    f_open = f"={a('eqp0')}" if i == 0 else f"={CD[i-1]}{BS['eqclose']}"
    putf(ws, f"{c}{BS['eqopen']}", f_open, EQ_OPEN[i], NUM0, green=(i == 0))
    putf(ws, f"{c}{BS['eqnpa']}", f"='Income Statement'!{FCOL[i]}{IS['npa']}", NPA_F[i],
         NUM0, green=True)
    putf(ws, f"{c}{BS['eqdps']}", f"={c}{BS['nddps']}", -DPS[i], NUM0)
    putf(ws, f"{c}{BS['eqcpn']}", f"={c}{BS['ndcpn']}", -HYB_CPN, NUM0)
    putf(ws, f"{c}{BS['eqclose']}",
         f"={c}{BS['eqopen']}+{c}{BS['eqnpa']}+{c}{BS['eqdps']}+{c}{BS['eqcpn']}",
         EQ_CLOSE[i], NUM0, bold=True)
    putf(ws, f"{c}{BS['dpsps']}", f"=-{c}{BS['eqdps']}/{a('shares')}/1000",
         DPS[i] / SH / 1000.0, PX3)
band(ws, BS['eqclose'], 6)
note(ws, f"A{BS['dpsps']+2}", 'The condensed layout above does not foot to zero: right-of-'
     'use assets, investment properties, provisions, deferred tax and the other liabilities '
     'not shown separately are omitted from the forecast columns because they are not '
     'driven by anything in this model. The audited FY2023-25 columns carry every line as '
     'reported.')

# ============ 11 CASH FLOW =========================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow — the audited record and the forecast waterfall', 'USD thousand.', 9,
      awidth=52, cwidth=13)
hdr(ws, 4, ['USD 000'] + YH + YFE)
put(ws, f"A{CF['ebitda']}", 'EBITDA (operating)', fmt=None)
put(ws, f"A{CF['ocf']}", 'Operating cash flow, as reported', fmt=None)
put(ws, f"A{CF['capex']}", 'Capital expenditure', fmt=None)
put(ws, f"A{CF['fcf']}", 'Free cash flow, as reported', bold=True, fmt=None)
for i in range(8):
    putf(ws, f"{ALL[i]}{CF['ebitda']}", f"='Income Statement'!{ALL[i]}{IS['ebitda']}",
         EB_ALL[i], NUM0, green=True)
for i in range(3):
    put(ws, f"{HC[i]}{CF['ocf']}", HC_['ocf'][i], BLUE, NUM0)
    put(ws, f"{HC[i]}{CF['capex']}", HC_['capex'][i], BLUE, NUM0)
    putf(ws, f"{HC[i]}{CF['fcf']}", f"={HC[i]}{CF['ocf']}+{HC[i]}{CF['capex']}",
         HC_['fcf'][i], NUM0, bold=True)
for i in range(5):
    put(ws, f"{FCOL[i]}{CF['ocf']}", '-', BLACK, NUM0)
    putf(ws, f"{FCOL[i]}{CF['capex']}", f"=DCF!{CD[i]}{DF_['capex']}", -CAPEX[i], NUM0,
         green=True)
    put(ws, f"{FCOL[i]}{CF['fcf']}", '-', BLACK, NUM0, bold=True)
band(ws, CF['fcf'], 9)
band(ws, CF['wfb'], 9)
put(ws, f"A{CF['wfb']}", 'THE FORECAST WATERFALL — LINKED TO THE DISCOUNTED-CASH-FLOW SHEET',
    bold=True, fmt=None)
for i, y in enumerate(YFE):
    cc = ws.cell(row=CF['wfb'], column=5 + i, value=y); cc.font = Font(bold=True)
    cc.fill = FILL_G
_cfrows = [(CF['nopat'], 'NOPAT', lambda i: f"=DCF!{CD[i]}{DF_['nopat']}", NOPAT_F, True),
           (CF['dna'], 'Add back depreciation and amortisation',
            lambda i: f"=DCF!{CD[i]}{DF_['adddna']}", DNA_F, True),
           (CF['cap'], 'Less capital expenditure',
            lambda i: f"=DCF!{CD[i]}{DF_['capex']}", [-x for x in CAPEX], True),
           (CF['dnwc'], 'Less change in working capital',
            lambda i: f"=DCF!{CD[i]}{DF_['dnwc']}", [-x for x in DNWC_F], True),
           (CF['fcff'], 'Free cash flow to the firm',
            lambda i: (f"={FCOL[i]}{CF['nopat']}+{FCOL[i]}{CF['dna']}+{FCOL[i]}{CF['cap']}"
                       f"+{FCOL[i]}{CF['dnwc']}"), FCFF_F, False),
           (CF['intat'], 'Less interest after tax',
            lambda i: f"='Balance Sheet'!{CD[i]}{BS['ndintat']}",
            [-x * (1 - TAXS) for x in INT_F], True),
           (CF['fi'], 'Plus finance income after tax',
            lambda i: f"='Balance Sheet'!{CD[i]}{BS['ndfi']}",
            [x * (1 - TAXS) for x in FININC_F], True),
           (CF['cpn'], 'Less perpetual capital securities coupon',
            lambda i: f"='Balance Sheet'!{CD[i]}{BS['ndcpn']}", [-HYB_CPN] * 5, True),
           (CF['fcfe'], 'Free cash flow to equity',
            lambda i: (f"={FCOL[i]}{CF['fcff']}+{FCOL[i]}{CF['intat']}+{FCOL[i]}{CF['fi']}"
                       f"+{FCOL[i]}{CF['cpn']}"), FCFE_F, False),
           (CF['dps'], 'Less ordinary dividends',
            lambda i: f"='Balance Sheet'!{CD[i]}{BS['nddps']}", [-x for x in DPS], True),
           (CF['ndmove'], 'Movement in net debt (a fall is a negative)',
            lambda i: f"=-{FCOL[i]}{CF['fcfe']}-{FCOL[i]}{CF['dps']}",
            [ND_CLOSE[i] - ND_OPEN[i] for i in range(5)], False)]
for rw, lab, fml, vals, gr in _cfrows:
    bd = rw in (CF['fcff'], CF['fcfe'])
    put(ws, f'A{rw}', lab, bold=bd, fmt=None)
    for i in range(3):
        put(ws, f'{HC[i]}{rw}', '-', BLACK, NUM0)
    for i in range(5):
        putf(ws, f'{FCOL[i]}{rw}', fml(i), vals[i], NUM0, bold=bd, green=gr)
    if bd:
        band(ws, rw, 9)
put(ws, f"A{CF['conv']}", 'Cash conversion — free cash flow to the firm over EBITDA',
    fmt=None)
for i in range(3):
    put(ws, f'{HC[i]}{CF["conv"]}', '-', BLACK, PCT)
for i in range(5):
    putf(ws, f"{FCOL[i]}{CF['conv']}", f"={FCOL[i]}{CF['fcff']}/{FCOL[i]}{CF['ebitda']}",
         FCFF_F[i] / EB_F[i], PCT)
note(ws, f"A{CF['conv']+2}", 'Cash conversion is the crux of an asset-heavy fleet owner: in '
     '2026 the newbuild programme absorbs more than the whole of the operating cash the '
     'business generates, and only as that programme delivers does free cash flow to the '
     'firm turn materially positive.')

# ============ 12 SUMMARY FINANCIALS =================================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials — the eight-year picture', 'USD thousand unless stated. Every '
      'cell on this sheet is a link or a ratio; nothing is typed twice.', 9, awidth=52,
      cwidth=13)
hdr(ws, 4, ['USD 000'] + YH + YFE)
r = 5


def sf(lab, fml, vals, fmt=NUM0, skip=()):
    global r
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{r}', '-', BLACK, fmt)
        else:
            f_ = fml(i)
            putf(ws, f'{ALL[i]}{r}', f_, vals[i], fmt,
                 green=f_.startswith(("='I", "='B", '=DCF', "='C", "='S")))
    r += 1


sf('Revenue', lambda i: f"='Income Statement'!{ALL[i]}{IS['rev']}", REV_ALL)
sf('Revenue growth', lambda i: f'={ALL[i]}5/{ALL[i-1]}5-1',
   [None] + [REV_ALL[i] / REV_ALL[i - 1] - 1 for i in range(1, 8)], PCT, skip=(0,))
sf('EBITDA (operating)', lambda i: f"='Income Statement'!{ALL[i]}{IS['ebitda']}", EB_ALL)
sf('EBITDA margin', lambda i: f"='Income Statement'!{ALL[i]}{IS['mgn']}",
   [EB_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
sf('EBIT', lambda i: f"='Income Statement'!{ALL[i]}{IS['op']}", EBIT_ALL)
sf('Depreciation and amortisation', lambda i: f"='Income Statement'!{ALL[i]}{IS['dna']}",
   DNA_ALL)
sf('Profit attributable to shareholders', lambda i: f"='Income Statement'!{ALL[i]}{IS['npa']}",
   NPA_ALL)
sf('Earnings attributable to ordinary shareholders',
   lambda i: f"='Income Statement'!{ALL[i]}{IS['ordn']}", ORD_ALL)
sf('Earnings per ordinary share (AED)',
   lambda i: f"='Income Statement'!{ALL[i]}{IS['epsaed']}",
   [ORD_ALL[i] / SH / 1000.0 * PEG for i in range(8)], PX3)
sf('Free cash flow to the firm', lambda i: f"='Cash Flow'!{ALL[i]}{CF['fcff']}",
   [None] * 3 + FCFF_F, skip=(0, 1, 2))
sf('Capital expenditure', lambda i: f"='Cash Flow'!{ALL[i]}{CF['capex']}", CAPEX_ALL)
sf('Capital expenditure / revenue', lambda i: f'=-{ALL[i]}{r-1}/{ALL[i]}5',
   [-CAPEX_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
sf('Net working capital', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['nwc']}", NWC_ALL)
sf('Net working capital / revenue', lambda i: f'={ALL[i]}{r-1}/{ALL[i]}5',
   [NWC_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
sf('Net debt', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['nd']}", ND_ALL)
sf('Net debt / EBITDA', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['ndeb']}",
   [ND_ALL[i] / EB_ALL[i] for i in range(8)], MULT)
sf('Equity attributable to shareholders', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['eqp']}",
   EQ_ALL)
sf('Invested capital', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['ic']}", IC_ALL)
sf('Return on invested capital', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['roic']}",
   [None] * 3 + ROIC_F, PCT, skip=(0, 1, 2))
sf('Return on equity', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['roe']}",
   [None] + _hroe[1:] + ROE_F, PCT, skip=(0,))
note(ws, f'A{r+1}', 'Return on invested capital is the discounted-cash-flow sheet\'s NOPAT '
     'over the same year\'s invested capital, which is why the audited years carry capital '
     'but no return: their tax charge is struck on a post-financing basis and is not '
     'comparable with the pre-financing NOPAT the forecast uses.')

# ============ 13 MONTE CARLO ==========================================================
ws = sheet('Monte Carlo')
title(ws, 'Probabilistic price map', 'A map of price dispersion over the next one and three '
      'months. It carries no view on value and is never blended with the valuation. Each '
      'figure is a complete engine re-run, so it is a pasted value and does NOT redraw when '
      'a driver is changed.', 8, awidth=48, cwidth=14)
hdr(ws, MC['hdr'], ['Horizon (AED)', '5th', '25th', 'Median', '75th', '95th',
                    'Probability above spot'])
for j, tag in enumerate(('1M', '3M')):
    h = STK['horizons'][tag]
    rw = MC['h0'] + j
    put(ws, f'A{rw}', f"{'One month' if tag == '1M' else 'Three months'} — graded "
        f"{h['grade_date']}", fmt=None)
    for i, k in enumerate(('p5', 'p25', 'p50', 'p75', 'p95')):
        put(ws, f'{get_column_letter(2+i)}{rw}', h['pct'][k], BLUE, PX)
    put(ws, f'G{rw}', h['p_above'], BLUE, PCT)
hdr(ws, MC['lhdr'], ['Level event', 'One month', 'Three months'])
for j, (lab, k) in enumerate([('Finishes 10% or more above spot', 'p_up10'),
                              ('Finishes 10% or more below spot', 'p_dn10'),
                              ('Touches 10% above spot at any point', 'touch_up10'),
                              ('Touches 10% below spot at any point', 'touch_dn10')]):
    rw = MC['l0'] + j
    put(ws, f'A{rw}', lab, fmt=None)
    put(ws, f'B{rw}', STK['horizons']['1M'][k], BLUE, PCT)
    put(ws, f'C{rw}', STK['horizons']['3M'][k], BLUE, PCT)
hdr(ws, MC['ehdr'], ['Engine setting', '', 'Value'])
for j, (lab, v, fmt, gr) in enumerate([
        ('Simulated paths', 50000, NUM0, False),
        ('Annualised volatility (three-month anchor)',
         STK['horizons']['3M']['anchor_vol_ann'], PCT, False),
        ('Spot price (AED)', f"=Summary!$C${SU['spot']}", PX, True),
        ('Anchor date', STK['anchor_date'], None, False),
        # WHAT A READER IS SHOWN IS THE BAND RECORD [R-CAL-02, R-CAL-03]. The retired
        # verdict was published here in one word and nothing read the workbook, so it
        # survived every sweep of the document beside it. Its replacement is the record
        # itself — how often the price finished inside the band, WITH the count of
        # resolved forecasts, because a percentage without its count is the number that
        # misleads — and the band-width ratio, which is disclosed and never gated.
        ('Resolved three-month forecasts scored', STEP0['windows_scored'], NUM0, False),
        ('Price finished inside the 90% band', STEP0['cov90'], PCT, False),
        ('Price finished inside the 50% band', STEP0['cov50'], PCT, False),
        ('Band width against a naive carry-anchored band',
         STEP0['w90_ratio'], None, False)]):
    rw = MC['e0'] + j
    put(ws, f'A{rw}', lab, fmt=None)
    if gr:
        putf(ws, f'C{rw}', v, SPOT, fmt, green=True)
    else:
        put(ws, f'C{rw}', v, BLUE, fmt)
note(ws, f"A{MC['e0']+6}", 'The price map is a dispersion forecast from the price series '
     'alone. It is deliberately not reconciled to the valuation lenses: one is a statement '
     'about what a price could do over weeks, the other about what a business is worth.')

# ============ 14 SENSITIVITY ============================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — what the valuation needs the world to do', 'AED per share. Each '
      'cell is a complete re-run of the whole model, including the fleet build, so these '
      'grids are engine outputs rather than formulas and do NOT redraw when a driver is '
      'changed.', 8, awidth=52, cwidth=14)
band(ws, SE['bgb'], 8)
put(ws, f"A{SE['bgb']}", 'Beta (rows) x terminal growth (columns)', bold=True, fmt=None)
hdr(ws, SE['bghdr'], ['Beta'] + [f'{g:.1%}' for g in SN['gs']])
for i, b in enumerate(SN['betas']):
    rw = SE['bg0'] + i
    put(ws, f'A{rw}', f'{b:.3f}', fmt=None)
    for j in range(len(SN['gs'])):
        put(ws, f'{get_column_letter(2+j)}{rw}', SN['grid_beta_g'][i][j], BLUE, PX)
note(ws, f"A{SE['bg0']+len(SN['betas'])}",
     f"The beta rows are not evenly spaced round numbers. They span the two index "
     f"constructions and the interval around the adopted one: {SN['betas'][0]:.3f} is the "
     f"equal-weight composite of the exchange's own names, {V['beta']:.3f} is the adopted "
     f"regression against the exchange's published index, and {SN['betas'][-1]:.3f} is the "
     f"top of that regression's own 90% confidence interval (its bottom, "
     f"{V['beta_ci_lo']:.3f}, sits {'above' if V['beta_ci_lo'] > SN['betas'][0] else 'below'} "
     f"the composite reading, so the bull case is struck "
     f"{'inside' if V['beta_ci_lo'] > SN['betas'][0] else 'outside'} this grid). The terminal growth "
     f"column at {V['g_terminal']:.1%} is the one the model runs on, so the cell where "
     f"that column meets the {V['beta']:.3f} row is the published cash-flow value of AED "
     f"{DC['fv_aed']:.2f} — every cell here is a complete re-run of the model discounted "
     f"at the same three tranches of capital the valuation itself uses, equity, debt and "
     f"the perpetual securities, so the grid is centred on the figure it brackets.")
band(ws, SE['ab'], 8)
put(ws, f"A{SE['ab']}", 'Mid-cycle tanker rate anchor, as a multiple of the base anchor',
    bold=True, fmt=None)
_am = ['0.8', '0.9', '1.0', '1.1', '1.2']
hdr(ws, SE['ahdr'], [''] + [f'{float(m):.2f}x' for m in _am])
put(ws, f"A{SE['a0']}", 'Fair value per share (AED)', fmt=None)
for j, m in enumerate(_am):
    put(ws, f'{get_column_letter(2+j)}{SE["a0"]}', SN['anchor'][m], BLUE, PX)
put(ws, f"A{SE['aswing']}", 'Swing across the grid', fmt=None)
putf(ws, f"C{SE['aswing']}", f"=MAX(B{SE['a0']}:F{SE['a0']})-MIN(B{SE['a0']}:F{SE['a0']})",
     max(SN['anchor'].values()) - min(SN['anchor'].values()), PX)
note(ws, f"A{SE['aswing']+1}", f"This row and the one below it are re-runs of the whole "
     f"model at the same {W_EXP:.2%} cost of capital the valuation uses, so the 1.00x cell "
     f"reproduces the published AED {DC['fv_aed']:.2f} exactly and the swing across the row "
     f"is a clean read of what that single driver is worth.")
band(ws, SE['cb'], 8)
put(ws, f"A{SE['cb']}", 'Capital expenditure, as a multiple of the guided path', bold=True,
    fmt=None)
_cm = ['0.9', '1.0', '1.1', '1.2']
hdr(ws, SE['chdr'], [''] + [f'{float(m):.2f}x' for m in _cm])
put(ws, f"A{SE['c0']}", 'Fair value per share (AED)', fmt=None)
for j, m in enumerate(_cm):
    put(ws, f'{get_column_letter(2+j)}{SE["c0"]}', SN['capex'][m], BLUE, PX)
put(ws, f"A{SE['cswing']}", 'Swing across the grid', fmt=None)
putf(ws, f"C{SE['cswing']}", f"=MAX(B{SE['c0']}:E{SE['c0']})-MIN(B{SE['c0']}:E{SE['c0']})",
     max(SN['capex'].values()) - min(SN['capex'].values()), PX)
band(ws, SE['tb'], 8)
put(ws, f"A{SE['tb']}", 'A uniform group tax rate — the global-minimum-tax case', bold=True,
    fmt=None)
_tk = ['0.05', '0.09', '0.15']
hdr(ws, SE['thdr'], [''] + [f'{float(t):.0%}' for t in _tk])
put(ws, f"A{SE['t0']}", 'Fair value per share (AED)', fmt=None)
for j, t in enumerate(_tk):
    put(ws, f'{get_column_letter(2+j)}{SE["t0"]}', SN['tax'][t], BLUE, PX)
note(ws, f"A{SE['tnote']}", 'The shipping units currently bear under one per cent, because '
     'international shipping income is relieved under the corporate tax law; the group\'s '
     'blended charge on operating profit is therefore only a few per cent. A fifteen per '
     'cent rate reaching that income — a global minimum tax applied without the shipping '
     'relief — is the downside case, and it is a whole-model re-run rather than a formula '
     'because the mix itself changes.')
band(ws, SE['mb'], 8)
put(ws, f"A{SE['mb']}", 'THE RATE PATH AGAINST WHAT THE FORWARD MARKET WAS ACTUALLY PAYING',
    bold=True, fmt=None)
MCC = SN['market_cross_check']
put(ws, f"A{SE['m1y']}", 'One-year time charter fixed in early 2026, very large crude '
    'carrier (USD per day)', fmt=None)
put(ws, f"C{SE['m1y']}", MCC['vlcc_1y_tc'], BLUE, NUM0)
put(ws, f"A{SE['mspot']}", 'Broker spot print for the same vessel class (USD per day)',
    fmt=None)
put(ws, f"C{SE['mspot']}", MCC['vlcc_spot_broker'], BLUE, NUM0)
hdr(ws, SE['mhdr'], ["This study's own path for the same vessel class (USD per day)"] + YFE)
put(ws, f"A{SE['mpath']}", 'Very large crude carrier — spot time-charter equivalent',
    fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{SE["mpath"]}', f"=Segments!{CD[i]}{SG['path0']+4}",
         TNK_PATH['vlcc'][i], NUM0, green=True)
put(ws, f"A{SE['mvs']}", 'The study\'s 2027 path over the one-year time charter', fmt=None)
putf(ws, f"C{SE['mvs']}", f"=C{SE['mpath']}/C{SE['m1y']}",
     TNK_PATH['vlcc'][1] / MCC['vlcc_1y_tc'], MULT)
put(ws, f"A{SE['mob']}", 'Crude tanker order book as a share of the trading fleet', fmt=None)
put(ws, f"C{SE['mob']}", MCC['orderbook_pct'], BLUE, PCT)
note(ws, f"A{SE['mnote']}", MCC['note'])
ws.column_dimensions['A'].width = 58

# ============ 15 PER-SHARE & RATIOS =========================================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share and ratio analysis', 'The indicator set for an asset-heavy marine '
      'logistics operator. Every ratio is a formula off the statements; per-share figures '
      'convert to dirhams at the fixed parity.', 9, awidth=52, cwidth=13)
hdr(ws, 4, ['Measure'] + YH + YFE)


def ratio(rw, lab, fml, vals, fmt, skip=()):
    put(ws, f'A{rw}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{rw}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALL[i]}{rw}', fml(i), vals[i], fmt)


ratio(PS['eps'], 'Earnings per ordinary share (USD)',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['eps']}",
      [ORD_ALL[i] / SH / 1000.0 for i in range(8)], PX3)
ratio(PS['epsaed'], 'Earnings per ordinary share (AED)',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['epsaed']}",
      [ORD_ALL[i] / SH / 1000.0 * PEG for i in range(8)], PX3)
ratio(PS['ordps'], 'Attributable profit per share (AED, before the perpetual coupon)',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['npa']}/{a('shares')}/1000*{a('fx')}",
      [NPA_ALL[i] / SH / 1000.0 * PEG for i in range(8)], PX3)
ratio(PS['bvps'], 'Book value per share (AED)',
      lambda i: f"='Balance Sheet'!{ALL[i]}{BS['bvpsaed']}",
      [EQ_ALL[i] / SH / 1000.0 * PEG for i in range(8)], PX)
ratio(PS['fcffps'], 'Free cash flow to the firm per share (AED)',
      lambda i: f"='Cash Flow'!{ALL[i]}{CF['fcff']}/{a('shares')}/1000*{a('fx')}",
      [None] * 3 + [x / SH / 1000.0 * PEG for x in FCFF_F], PX3, skip=(0, 1, 2))
ratio(PS['dpsps'], 'Ordinary dividend per share (AED)',
      lambda i: f"='Balance Sheet'!{CD[i-3]}{BS['dpsps']}*{a('fx')}",
      [None] * 3 + [x / SH / 1000.0 * PEG for x in DPS], PX3, skip=(0, 1, 2))
ratio(PS['payout'], 'Ordinary dividend payout ratio',
      lambda i: f"=-'Balance Sheet'!{CD[i-3]}{BS['eqdps']}/'Income Statement'!"
                f"{ALL[i]}{IS['npa']}",
      [None] * 3 + [DPS[i] / NPA_F[i] for i in range(5)], PCT, skip=(0, 1, 2))
ratio(PS['gm'], 'Gross margin',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['gp']}/'Income Statement'!"
                f"{ALL[i]}{IS['rev']}",
      [HI['gross_profit'][i] / H_REV[i] for i in range(3)] + [None] * 5,
      PCT, skip=(3, 4, 5, 6, 7))
ratio(PS['ebm'], 'EBITDA margin', lambda i: f"='Income Statement'!{ALL[i]}{IS['mgn']}",
      [EB_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
ratio(PS['ebitm'], 'EBIT margin',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['op']}/'Income Statement'!"
                f"{ALL[i]}{IS['rev']}", [EBIT_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
ratio(PS['netm'], 'Net margin (attributable)',
      lambda i: f"='Income Statement'!{ALL[i]}{IS['npa']}/'Income Statement'!"
                f"{ALL[i]}{IS['rev']}", [NPA_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
ratio(PS['roe'], 'Return on equity', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['roe']}",
      [None] + _hroe[1:] + ROE_F, PCT, skip=(0,))
ratio(PS['roic'], 'Return on invested capital',
      lambda i: f"='Balance Sheet'!{ALL[i]}{BS['roic']}", [None] * 3 + ROIC_F, PCT,
      skip=(0, 1, 2))
ratio(PS['ndeb'], 'Net debt / EBITDA', lambda i: f"='Balance Sheet'!{ALL[i]}{BS['ndeb']}",
      [ND_ALL[i] / EB_ALL[i] for i in range(8)], MULT)
ratio(PS['cover'], 'Interest cover (EBIT over finance costs)',
      lambda i: f"=-'Income Statement'!{ALL[i]}{IS['op']}/'Income Statement'!"
                f"{ALL[i]}{IS['fincost']}",
      [HI['ebit'][i] / abs(HI['fin_costs'][i]) for i in range(3)]
      + [EBIT_F[i] / INT_F[i] for i in range(5)], MULT)
ratio(PS['dso'], 'Days sales outstanding',
      lambda i: f"='Balance Sheet'!{ALL[i]}{BS['recv']}/'Income Statement'!"
                f"{ALL[i]}{IS['rev']}*365",
      [(HB['receivables'][i] + HB['due_from_related'][i]) / H_REV[i] * 365
       for i in range(3)] + [DSO] * 5, NUM1)
ratio(PS['dio'], 'Days inventory outstanding',
      lambda i: f"='Balance Sheet'!{ALL[i]}{BS['inv']}/'Income Statement'!"
                f"{ALL[i]}{IS['opcost']}*365",
      [HB['inventories'][i] / H_OPCOST[i] * 365 for i in range(3)] + [DIO] * 5, NUM1)
ratio(PS['dpo'], 'Days payable outstanding',
      lambda i: f"='Balance Sheet'!{ALL[i]}{BS['pay']}/'Income Statement'!"
                f"{ALL[i]}{IS['opcost']}*365",
      [(HB['payables'][i] + HB['due_to_related'][i]) / H_OPCOST[i] * 365
       for i in range(3)] + [DPO] * 5, NUM1)
ratio(PS['cycle'], 'Cash conversion cycle (days)',
      lambda i: f"={ALL[i]}{PS['dso']}+{ALL[i]}{PS['dio']}-{ALL[i]}{PS['dpo']}",
      [(HB['receivables'][i] + HB['due_from_related'][i]) / H_REV[i] * 365
       + HB['inventories'][i] / H_OPCOST[i] * 365
       - (HB['payables'][i] + HB['due_to_related'][i]) / H_OPCOST[i] * 365
       for i in range(3)] + [DSO + DIO - DPO] * 5, NUM1)
ratio(PS['capexrev'], 'Capital expenditure / revenue',
      lambda i: f"=-'Cash Flow'!{ALL[i]}{CF['capex']}/'Income Statement'!"
                f"{ALL[i]}{IS['rev']}",
      [-CAPEX_ALL[i] / REV_ALL[i] for i in range(8)], PCT)
band(ws, PS['ab'], 9)
put(ws, f"A{PS['ab']}", 'MULTIPLES AT THE ANCHOR PRICE', bold=True, fmt=None)
for rw, lab, fml, xp, fmt in [
        (PS['aprice'], 'Share price (AED)', f"=Summary!$C${SU['spot']}", SPOT, PX),
        (PS['apriceusd'], 'Share price (USD)',
         f"='Relative & Normalized'!$C${RN['spotusd']}", SPOT / PEG, PX),
        (PS['amkt'], 'Market capitalisation (USD 000)',
         f"='Relative & Normalized'!$C${RN['mktcap']}", MKTCAP, NUM0),
        (PS['aev'], 'Enterprise value (USD 000)',
         f"='Relative & Normalized'!$C${RN['evnow']}", EV_NOW, NUM0),
        (PS['aeveb'], 'Enterprise value / 2025 reported EBITDA',
         f"='Relative & Normalized'!$C${RN['eveb_ttm']}", OWN_EVEB_TTM, MULT),
        (PS['aeveb26'], 'Enterprise value / FY2026E EBITDA',
         f"='Relative & Normalized'!$C${RN['eveb_26']}", OWN_EVEB_26, MULT),
        (PS['ape'], 'Price / 2025 ordinary earnings',
         f"='Relative & Normalized'!$C${RN['pe_ttm']}", OWN_PE_TTM, MULT),
        (PS['apb'], 'Price / book at 31 March 2026',
         f"='Relative & Normalized'!$C${RN['pb']}", OWN_PB, MULT),
        (PS['ady'], 'Dividend yield on the 2026 distribution',
         f"='Relative & Normalized'!$C${RN['dy']}", OWN_DY, PCT)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=True)

# ============ 16 PEER & SECTOR ===============================================================
ws = sheet('Peer & Sector')
title(ws, 'Peer frame and sector context', 'No single clean comparable exists: the company '
      'is part contracted marine logistics and part spot tanker owner, so the frame is '
      'built from both ends and blended on the company\'s own disclosed exposure.', 10,
      awidth=34, cwidth=15)
hdr(ws, PR['hdr'], ['Company', 'Market', 'Business model', 'EV/EBITDA', 'Forward P/E',
                    'Trailing P/E', 'Price / book', 'Dividend yield', 'Source', 'As at'])
for j, p in enumerate(PEERS):
    rw = PR['p0'] + j
    put(ws, f'A{rw}', p['name'], fmt=None)
    put(ws, f'B{rw}', p['market'], fmt=None)
    put(ws, f'C{rw}', p['model'], fmt=None, wrap=True)
    put(ws, f'D{rw}', p['ev_ebitda'], BLUE, MULT)
    put(ws, f'E{rw}', p['pe_fwd'] if p['pe_fwd'] is not None else 'n/a', BLUE, MULT)
    put(ws, f'F{rw}', p['pe_ttm'] if p['pe_ttm'] is not None else 'n/a', BLUE, MULT)
    put(ws, f'G{rw}', p['pb'] if p['pb'] is not None else 'n/a', BLUE, MULT)
    put(ws, f'H{rw}', p['dy'] if p['dy'] is not None else 'n/a', BLUE, PCT)
    put(ws, f'I{rw}', p['src'], fmt=None, wrap=True)
    put(ws, f'J{rw}', p['asof'], fmt=None)
    ws.row_dimensions[rw].height = 28
ws.column_dimensions['C'].width = 30; ws.column_dimensions['I'].width = 34
ws.column_dimensions['A'].width = 30
band(ws, PR['mb'], 10)
put(ws, f"A{PR['mb']}", 'THE MULTIPLES USED IN THE RELATIVE LENS', bold=True, fmt=None)
for rw, lab, fml, xp, fmt in [
        (PR['mcon'], 'Contracted-shipping multiple', f"=D{PR['p0']}", MULT_CONTR, MULT),
        (PR['mspot'], 'Spot-tanker multiple — the two spot owners averaged',
         f"=(D{PR['p0']+1}+D{PR['p0']+2})/2", MULT_SPOT, MULT),
        (PR['mw'], 'Share of 2026 earnings exposed to spot rates, as disclosed',
         f"={a('spot_w')}", SPOT_W, PCT),
        (PR['mev'], 'Blended enterprise multiple',
         f"=(1-C{PR['mw']})*C{PR['mcon']}+C{PR['mw']}*C{PR['mspot']}", BLEND_EV, MULT),
        (PR['pecon'], 'Contracted-shipping forward price/earnings', f"=E{PR['p0']}",
         PEERS[0]['pe_fwd'], MULT),
        (PR['pespot'], 'Spot-tanker forward price/earnings', f"=E{PR['p0']+1}",
         PEERS[1]['pe_fwd'], MULT),
        (PR['pe'], 'Blended forward price/earnings — the multiple the relative lens '
         'applies to forward earnings',
         f"=(1-C{PR['mw']})*C{PR['pecon']}+C{PR['mw']}*C{PR['pespot']}", BLEND_PE, MULT),
        (PR['pecon_t'], 'Contracted-shipping trailing price/earnings', f"=F{PR['p0']}",
         PEERS[0]['pe_ttm'], MULT),
        (PR['pespot_t'], 'Spot-tanker trailing price/earnings', f"=F{PR['p0']+1}",
         PEERS[1]['pe_ttm'], MULT),
        (PR['pe_t'], 'Blended trailing price/earnings — the same blend on the trailing '
         'basis, so the company can be compared with the peers on either basis rather '
         'than one against the other',
         f"=(1-C{PR['mw']})*C{PR['pecon_t']}+C{PR['mw']}*C{PR['pespot_t']}", BLEND_PE_TTM,
         MULT)]:
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=fml.startswith('=Assumptions'))
band(ws, PR['ob'], 10)
put(ws, f"A{PR['ob']}", "THE COMPANY'S OWN MULTIPLES AT THE ANCHOR PRICE", bold=True,
    fmt=None)
for j, (lab, src, xp, fmt) in enumerate([
        ('Enterprise value / 2025 reported EBITDA', RN['eveb_ttm'], OWN_EVEB_TTM, MULT),
        ('Enterprise value / FY2026E EBITDA', RN['eveb_26'], OWN_EVEB_26, MULT),
        ('Price / 2025 ordinary earnings — trailing, against the blended trailing '
         'multiple above', RN['pe_ttm'], OWN_PE_TTM, MULT),
        ('Price / FY2026E ordinary earnings — forward, against the blended forward '
         'multiple above', RN['pe_fwd'], OWN_PE_FWD, MULT),
        ('Price / book at 31 March 2026', RN['pb'], OWN_PB, MULT),
        ('Dividend yield on the 2026 distribution', RN['dy'], OWN_DY, PCT)]):
    rw = PR['o0'] + j
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', f"='Relative & Normalized'!$C${src}", xp, fmt, green=True)
note(ws, f"A{PR['o0']+6}", 'The contracted peer is a long-term contracted gas shipping '
     'company whose earnings look nothing like a spot tanker owner\'s; the two spot owners '
     'carry the opposite exposure. Neither is a comparable on its own, which is why the '
     'multiple applied is a blend weighted by the share of earnings the company itself '
     'discloses as spot-exposed, and why the relative lens is a cross-check rather than an '
     'independent valuation.')

# ============ 2 SUMMARY (filled now the source rows are known) ==================
ws = wb['Summary']
_LSRC = {'dcf': (f"='Fundamental Valuation'!$C${FV['dcfbear']}", f"=DCF!$C${DF_['fvaed']}",
                 f"='Fundamental Valuation'!$C${FV['dcfbull']}"),
         'relative': (f"='Relative & Normalized'!$C${RN['bear']}",
                      f"='Relative & Normalized'!$C${RN['base']}",
                      f"='Relative & Normalized'!$D${RN['bear']}"),
         'normalized': (f"='Relative & Normalized'!$C${RN['nbear']}",
                        f"='Relative & Normalized'!$C${RN['nbase']}",
                        f"='Relative & Normalized'!$D${RN['nbear']}"),
         'book': (f"='Relative & Normalized'!$C${RN['bbear']}",
                  f"='Relative & Normalized'!$C${RN['bbase']}",
                  f"='Relative & Normalized'!$D${RN['bbear']}")}
_WKEY = {'dcf': 'w_dcf', 'relative': 'w_rel', 'normalized': 'w_norm', 'book': 'w_book'}
_SUKEY = {'dcf': 'dcf', 'relative': 'rel', 'normalized': 'norm', 'book': 'book'}
for k in ('dcf', 'relative', 'normalized', 'book'):
    rw = SU[_SUKEY[k]]
    put(ws, f'A{rw}', LENS_LABEL[k], fmt=None)
    for col, idx in (('B', 0), ('C', 1), ('D', 2)):
        putf(ws, f'{col}{rw}', _LSRC[k][idx], LB[k][idx], PX, green=True)
    # THE WEIGHT COLUMN IS THE RETIRED BLEND'S, kept so a reader can see what the
    # previous architecture said and what it cost. Nothing downstream reads column F.
    putf(ws, f'E{rw}', f"={a(_WKEY[k])}", LW[k], PCT, green=True)
    putf(ws, f'F{rw}', f'=C{rw}*E{rw}', LB[k][1] * LW[k], PX)
    putf(ws, f'G{rw}', f"=C{rw}/$C${SU['spot']}-1", LB[k][1] / SPOT - 1, PCT)
putf(ws, f"H{SU['dcf']}", f"=DCF!$C${DF_['tvshare']}", DC['tv_share'], PCT, green=True)
band(ws, SU['central'], 8)
put(ws, f"A{SU['central']}",
    'CENTRAL — THE CASH-FLOW LENS, NOT A BLEND', bold=True, fmt=None)
_LK = ['dcf', 'rel', 'norm', 'book']
for col, idx, xp in (('B', 0, CENTRAL_BEAR), ('D', 2, CENTRAL_BULL)):
    putf(ws, f"{col}{SU['central']}", f"={col}{SU['dcf']}", xp, PX, bold=True)
putf(ws, f"C{SU['central']}", f"=C{SU['dcf']}", CENTRAL, PX, bold=True)
putf(ws, f"F{SU['central']}", f"=SUM(F{SU['dcf']}:F{SU['book']})", RETIRED_BLEND, PX)
putf(ws, f"E{SU['central']}", f"=SUM(E{SU['dcf']}:E{SU['book']})", 1.0, PCT)
putf(ws, f"G{SU['central']}", f"=C{SU['central']}/$C${SU['spot']}-1", CENTRAL / SPOT - 1,
     PCT, bold=True)
band(ws, SU['cb'], 8)
put(ws, f"A{SU['cb']}", 'THE CONTESTED JUDGEMENT — PUBLISHED BOTH WAYS, NEVER AVERAGED',
    bold=True, fmt=None)
put(ws, f"A{SU['dcfa']}", 'Discounted cash flow (composite-index beta of '
    f"{V['beta_composite']:.3f})", fmt=None)
putf(ws, f"B{SU['dcfa']}", f"=B{SU['dcf']}", LN['dcf']['bear'], PX)
putf(ws, f"C{SU['dcfa']}", f"=DCF!$C${DF_['fvaeda']}", DA['fv_aed'], PX, green=True)
putf(ws, f"D{SU['dcfa']}", f"=D{SU['dcf']}", LN['dcf']['bull'], PX)
putf(ws, f"E{SU['dcfa']}", f"=E{SU['dcf']}", LW['dcf'], PCT)
putf(ws, f"F{SU['dcfa']}", f"=C{SU['dcfa']}*E{SU['dcfa']}", DA['fv_aed'] * LW['dcf'], PX)
putf(ws, f"G{SU['dcfa']}", f"=C{SU['dcfa']}/$C${SU['spot']}-1", DA['fv_aed'] / SPOT - 1, PCT)
putf(ws, f"H{SU['dcfa']}", f"=DCF!$C${DF_['tvsharea']}", DA['tv_share'], PCT, green=True)
put(ws, f"A{SU['centrala']}", 'Central on the composite-index beta — the cash-flow '
    'lens on the other regressor, not a blend', bold=True, fmt=None)
putf(ws, f"C{SU['centrala']}",
     f"=C{SU['dcfa']}", CENTRAL_A, PX, bold=True)
putf(ws, f"G{SU['centrala']}", f"=C{SU['centrala']}/$C${SU['spot']}-1", CENTRAL_A / SPOT - 1,
     PCT, bold=True)
band(ws, SU['centrala'], 8)
put(ws, f"A{SU['panel']}", 'Expert panel average', fmt=None)
putf(ws, f"C{SU['panel']}", f"='Fundamental Valuation'!$C${FV['epanel']}",
     D['panel_centre'], PX, green=True)
putf(ws, f"G{SU['panel']}", f"=C{SU['panel']}/$C${SU['spot']}-1",
     D['panel_centre'] / SPOT - 1, PCT)
put(ws, f"A{SU['spot']}", 'Market price (AED, anchor)', bold=True, fmt=None)
putf(ws, f"C{SU['spot']}", f"={a('spot')}", SPOT, PX, bold=True, green=True)
band(ws, SU['spot'], 8)
hdr(ws, SU['keyhdr'], ['Key figure', '', 'Value'])
_KEY = [('Shares outstanding (mn)', f"={a('shares')}", SH, NUM1),
        ('Market capitalisation (USD 000)', f"=DCF!$C${DF_['mktcap']}", MKTCAP, NUM0),
        ('Enterprise value at the anchor price (USD 000)',
         f"='Relative & Normalized'!$C${RN['evnow']}", EV_NOW, NUM0),
        ('Net debt at 31 March 2026, including the deferred consideration and the eleven '
         'vessels bought on 7 August 2026 (USD 000)',
         f"='Relative & Normalized'!$C${RN['netdebt']}", NETDEBT, NUM0),
        ('The eleven vessels bought on 7 August 2026, at the announced price (USD 000)',
         f"={a('acq_cost')}", ACQ_COST, NUM0),
        ('FY2025 revenue (USD 000)', f"='Income Statement'!D{IS['rev']}", H_REV[2], NUM0),
        ('FY2025 EBITDA as reported (USD 000)', f"='Income Statement'!D{IS['ebrep']}",
         HI['ebitda_reported'][2], NUM0),
        ('FY2025 profit attributable to shareholders (USD 000)',
         f"='Income Statement'!D{IS['npa']}", H_NPA[2], NUM0),
        ('Cost of equity — published-index beta (primary)', f"=DCF!$C${DF_['ke']}", KE,
         PCT2),
        ('Cost of equity — composite-index beta (alternative)', f"=DCF!$C${DF_['kea']}",
         KE_A, PCT2),
        ('Cost of equity at the lower 90% confidence bound on the primary beta',
         f"=DCF!$C${DF_['kecilo']}", KE_CI_LO, PCT2),
        ('Cost of equity at the upper 90% confidence bound on the primary beta',
         f"=DCF!$C${DF_['kecihi']}", KE_CI_HI, PCT2),
        ('Cost of debt ADOPTED — the instruments actually outstanding, weighted by '
         'balance', f"=DCF!$C${DF_['kd']}", KD, PCT2),
        ('Cost of debt — MEMORANDUM, the retired average of three constructions',
         f"=DCF!$C${DF_['kdbal']}", KD_RETIRED_AVG, PCT2),
        ('Cost of capital — explicit window', f"=DCF!$C${DF_['wacc']}", W_EXP, PCT2),
        ('Cost of capital — terminal', f"=DCF!$C${DF_['waccterm']}", W_TERM, PCT2),
        ('Terminal growth', f"=DCF!$C${DF_['g']}", G, PCT),
        ('Sum-of-the-parts cross-check (AED per share)',
         f"='SOTP Bridge'!$C${SB['bfv']}", SOTP_FV, PX)]
for j, (lab, fml, xp, fmt) in enumerate(_KEY):
    rw = SU['key0'] + j
    put(ws, f'A{rw}', lab, fmt=None)
    putf(ws, f'C{rw}', fml, xp, fmt, green=True)
note(ws, f"A{SU['key0']+len(_KEY)+1}", 'The terminal value share beside the discounted-cash-'
     'flow row is a live formula — the present value of the terminal value over the '
     'enterprise value of operations — and it is shown again in the equity bridge on the '
     'SOTP Bridge sheet. It is high, as it is for any long-lived asset owner whose fleet '
     'outlives the forecast window, and it is the reason the terminal cost of capital and '
     'the terminal growth rate carry more of this valuation than any single trading year.')
ANCH.update(summary_central=f"C{SU['central']}", summary_spot=f"C{SU['spot']}",
            dcf_fv=f"C{DF_['fvaed']}", dcf_fv_beta_alt=f"C{DF_['fvaeda']}",
            summary_central_beta_alt=f"C{SU['centrala']}")

# ============ 3 FUNDAMENTAL VALUATION (filled) ==================================
ws = wb['Fundamental Valuation']
hdr(ws, FV['hdr'], ['Lens / step', 'Basis', 'AED per share'])
_fv = [(FV['dcf'], 'Discounted cash flow (own regressed beta)',
        'links to the DCF sheet — five explicit years plus a capitalised terminal value',
        f"=DCF!$C${DF_['fvaed']}", DC['fv_aed'], True),
       (FV['dcfbear'], '  bear', f"beta {V['beta_ci_hi']:.3f} — the TOP of the "
        "regression's own 90% confidence interval, not a round number picked by hand — "
        'with the rate anchor at 0.85x and capital expenditure at 1.10x. A whole-model '
        're-run', LN['dcf']['bear'], LN['dcf']['bear'], False),
       (FV['dcfbull'], '  bull', f"beta {V['beta_ci_lo']:.3f} — the BOTTOM of the same "
        'interval — with the rate anchor at 1.15x and capital expenditure at 0.95x. Taking '
        'both bounds from the interval means the published range is the range the estimate '
        'itself supports. A whole-model re-run', LN['dcf']['bull'], LN['dcf']['bull'],
        False),
       (FV['rel'], 'Relative multiples',
        'blended enterprise and earnings multiples on 2026 earnings',
        f"='Relative & Normalized'!$C${RN['base']}", REL_BASE, True),
       (FV['norm'], 'Normalised earnings power',
        'the same multiples on the five-year average of the forecast',
        f"='Relative & Normalized'!$C${RN['nbase']}", NORM_BASE, True),
       (FV['book'], 'Book value and sustainable return',
        'the book the company already has, plus the present value of earning more than '
        'the cost of equity on it while that lasts, plus a fading remainder — a residual '
        'income build, not a justified multiple',
        f"='Relative & Normalized'!$C${RN['bbase']}", BOOK_BASE, True)]
for rw, lab, basis, val, xp, isf in _fv:
    put(ws, f'A{rw}', lab, fmt=None)
    put(ws, f'B{rw}', basis, fmt=None, wrap=True)
    ws.row_dimensions[rw].height = 28
    if isf:
        putf(ws, f'C{rw}', val, xp, PX, green=True)
    else:
        put(ws, f'C{rw}', val, BLUE, PX)
band(ws, FV['central'], 3)
put(ws, f"A{FV['central']}", 'Weighted central', bold=True, fmt=None)
putf(ws, f"C{FV['central']}", f"=Summary!$C${SU['central']}", CENTRAL, PX, bold=True,
     green=True)
band(ws, FV['cb'], 3)
put(ws, f"A{FV['cb']}", 'THE CONTESTED JUDGEMENT — TWO INDEX CONSTRUCTIONS, BOTH CARRIED '
    'THROUGH IN FULL', bold=True, fmt=None)
for rw, lab, basis, fml, xp, fmt in [
        (FV['beta'], 'Beta — weekly regression against the published index of its own '
         'exchange (primary)',
         f"{D['beta']['n']} weekly observations, R-squared {D['beta']['r2']:.1%}, standard "
         f"error {D['beta']['se']:.3f}, 90% confidence interval "
         f"{V['beta_ci_lo']:.3f} to {V['beta_ci_hi']:.3f} — the usability gate is passed, "
         'and this is the index the method asks for',
         f"={a('beta')}", V['beta'], BETA),
        (FV['ke'], 'Cost of equity on the published-index beta',
         'normalised risk-free rate plus beta times the equity risk premium',
         f"=DCF!$C${DF_['ke']}", KE, PCT2),
        (FV['wacc'], 'Cost of capital on the published-index beta',
         'explicit window, market-value weights', f"=DCF!$C${DF_['wacc']}", W_EXP, PCT2),
        (FV['fv'], 'Fair value per share — published-index beta (AED)',
         'the primary reading', f"=DCF!$C${DF_['fvaed']}", DC['fv_aed'], PX),
        (FV['cen'], 'Weighted central — published-index beta (AED)', 'all four lenses',
         f"=Summary!$C${SU['central']}", CENTRAL, PX),
        (FV['betaa'], 'Beta — the same regression against an equal-weight composite of the '
         'same exchange\'s names (alternative)',
         'the published index is weighted by size and is dominated by the same '
         'large-capitalisation group this company belongs to; the composite gives the '
         'exchange\'s smallest names the same say as its largest. Only the measure of the '
         'market changes',
         f"={a('beta_a')}", V['beta_composite'], BETA),
        (FV['kea'], 'Cost of equity on the composite-index beta',
         'the same construction, the same premium, the composite beta',
         f"=DCF!$C${DF_['kea']}", KE_A, PCT2),
        (FV['wacca'], 'Cost of capital on the composite-index beta',
         'explicit window, market-value weights', f"=DCF!$C${DF_['wacca']}", W_EXP_A, PCT2),
        (FV['fva'], 'Fair value per share — composite-index beta (AED)',
         'the alternative reading', f"=DCF!$C${DF_['fvaeda']}", DA['fv_aed'], PX),
        (FV['cena'], 'Weighted central — composite-index beta (AED)',
         'all four lenses, the discounted-cash-flow leg swapped',
         f"=Summary!$C${SU['centrala']}", CENTRAL_A, PX),
        (FV['cilo'], 'Beta — lower bound of the 90% confidence interval on the primary '
         'regression', 'the bull-case beta; the bear and bull cases take the two ends of '
         'this interval rather than round numbers picked by hand',
         f"={a('beta_ci_lo')}", V['beta_ci_lo'], BETA),
        (FV['cihi'], 'Beta — upper bound of the 90% confidence interval on the primary '
         'regression', 'the bear-case beta; the published range is therefore the range '
         'the estimate itself supports', f"={a('beta_ci_hi')}", V['beta_ci_hi'], BETA)]:
    put(ws, f'A{rw}', lab, fmt=None)
    put(ws, f'B{rw}', basis, fmt=None, wrap=True)
    ws.row_dimensions[rw].height = 26
    putf(ws, f'C{rw}', fml, xp, fmt, green=True)
note(ws, f"A{FV['note']}", 'These two readings are published side by side and are never '
     'averaged into a single number. They are the same regression of the same weekly '
     'returns; what differs is how the market itself is measured. The published index of '
     'the exchange is what the beta method asks for and is the primary reading. The '
     'equal-weight composite is what an earlier construction of this study used, and it is '
     'kept in view because a difference of this size is a property of index construction '
     'rather than a fact about the company. The regression passes its usability gate, but '
     'the history is only three years long and the 90% interval shown above spans more '
     'than half the point estimate, so the alternative is not a stress case — it is a '
     'second legitimate answer to the same question.')
band(ws, FV['eb'], 5)
put(ws, f"A{FV['eb']}", 'EXPERT PANEL — THREE METHODS, WORKED INDEPENDENTLY', bold=True,
    fmt=None)
hdr(ws, FV['ehdr'], ['Expert', 'Method', 'Base (AED per share)', 'Low', 'High'])
for j, k in enumerate(['e1', 'e2', 'e3']):
    e = EXP[k]
    rw = FV['e0'] + j
    put(ws, f'A{rw}', f'Expert {j+1}', fmt=None)
    put(ws, f'B{rw}', e['method_short'], fmt=None)
    put(ws, f'C{rw}', e['base'], BLUE, PX)
    put(ws, f'D{rw}', e['rng'][0], BLUE, PX)
    put(ws, f'E{rw}', e['rng'][1], BLUE, PX)
band(ws, FV['epanel'], 5)
put(ws, f"A{FV['epanel']}", 'Panel average', bold=True, fmt=None)
putf(ws, f"C{FV['epanel']}", f"=AVERAGE(C{FV['e0']}:C{FV['e0']+2})", D['panel_centre'], PX,
     bold=True)
ws.column_dimensions['B'].width = 56

# ============ the Assumptions sheet: derived cells become formulas ==============
# A driver sheet may carry only GENUINELY INDEPENDENT inputs. Anything on it that is itself
# derived from other inputs is written here as a live formula pointing at the build that
# produces it, so a reader can never mistake an output for an assumption and a change in
# the underlying disclosure carries all the way through.
ws = wb['Assumptions']
for _k, _fml, _xp, _fmt in [
        ('opex_day', f"=Segments!B{SG['opexd0']}", OPEX_DAY, NUM1),
        ('gas_rate', f"=Segments!B{SG['gasrate0']}", GAS_RATE, NUM0),
        ('dso_rep', f"='Balance Sheet'!D{BS['recv']}/'Income Statement'!D{IS['rev']}*365",
         DSO_REPORTED, NUM1),
        ('gu25', f"=Segments!D{SG['revh0'] + SEGREF['Tankers']}/Segments!B{SG['tcerev25']}",
         GROSSUP25, '0.00'),
        ('gu26', f"={a('grossup')}", GROSSUP, '0.00'),
        # THE RE-BASING, IN CELLS. The ratio is measured on reported 2025 revenue, which
        # carries the 2025 gross-up; the forecast revenue it is applied to is built at the
        # 2026 one. So the denominator is 2025 revenue with its tanker leg restated at the
        # gross-up the forecast uses, and the days are scaled by the ratio of the two.
        ('dso', f"=C{A['dso_rep']}*'Income Statement'!D{IS['rev']}/('Income Statement'!"
                f"D{IS['rev']}-Segments!D{SG['revh0'] + SEGREF['Tankers']}"
                f"+Segments!D{SG['revh0'] + SEGREF['Tankers']}/C{A['gu25']}"
                f"*C{A['gu26']})", DSO, NUM1),
        ('dio', f"='Balance Sheet'!D{BS['inv']}/'Income Statement'!D{IS['opcost']}*365",
         DIO, NUM1),
        ('dpo', f"='Balance Sheet'!D{BS['pay']}/'Income Statement'!D{IS['opcost']}*365",
         DPO, NUM1),
        ('nwc25', f"='Balance Sheet'!D{BS['nwc']}", NWC25, NUM0),
        # the same audited figure cannot be the record in two places: the balance sheet
        # carries it and this sheet points at it
        ('intang', f"='Balance Sheet'!D{BS['intang']}", INTANG, NUM0),
        ('gw', f"='Balance Sheet'!D{BS['gw']}", GW, NUM0)]:
    putf(ws, f"C{A[_k]}", _fml, _xp, _fmt, green=True)
# the charter table's two date columns display as dates, but hold real serial numbers, so
# nothing on this sheet is text sitting in an arithmetic chain
for _k in range(len(CHARTERS)):
    for _c in ('C', 'D'):
        ws[f"{_c}{A['ch'+str(_k)]}"].number_format = DATEFMT

put(ws, f"H{A['erp']}", 'No sovereign credit-default-swap entry exists for the United Arab '
    'Emirates in the country risk file, so the alternative rating-versus-swap premium basis '
    'cannot be built for this country; one basis is published rather than two.',
    fmt=None).font = SUB
put(ws, f"H{A['sofr']}", 'The Central Bank base rate of 3.65% was maintained at the 29 July '
    '2026 decision; the last change was a 25 basis point cut from 3.90% on 10 December '
    '2025.', fmt=None).font = SUB
put(ws, f"H{A['b25_mr']}", 'The handysize class is not broken out in the disclosure, so the '
    'medium-range rate stands in for it — SCALED by the disclosed relative move, because '
    'the company said handysize rates fell 21% while medium range rose 29%, and an '
    'unadjusted substitution therefore had the sign wrong. The gap is still flagged. The '
    '2024 quarterly rates for the medium-range class are not disclosed either, so its 2025 '
    'average stands in on both sides of the mid-cycle average.', fmt=None).font = SUB
put(ws, f"H{A['opex_day']}", 'GREEN, NOT BLUE. The running cost per vessel-day is not an '
    'assumption: it is solved on the Segments sheet so that the same construction '
    'reproduces the tanker earnings the company actually reported for 2025. The same is '
    'true of the gas-carrier day rate, the three days ratios and the opening net working '
    'capital immediately below.', fmt=None).font = SUB
put(ws, f'A{ASSUMPTIONS_LAST}', 'Blue cells on this sheet are independent inputs. Green '
    'cells are DERIVED — they are live formulas pointing at the build that produces them, '
    'and they are on this sheet only so that every driver the model reads can be found in '
    'one place. Nothing here is a pasted result.', fmt=None).font = SUB

# ============ save and verify against the committed study numbers ================
def close(x, y, tol):
    assert abs(float(x) - float(y)) <= tol, f'{x} vs {y}'


for i in range(5):
    close(REV_F[i], FC['revenue'][i], 1e-6)
    close(EB_F[i], FC['ebitda'][i], 1e-6)
    close(DNA_F[i], FC['dna'][i], 1e-6)
    close(EBIT_F[i], FC['ebit'][i], 1e-6)
    close(TAX_F[i], FC['tax'][i], 1e-6)
    close(NOPAT_F[i], FC['nopat'][i], 1e-6)
    close(NWC_F[i], FC['nwc'][i], 1e-6)
    close(DNWC_F[i], FC['dnwc'][i], 1e-6)
    close(FCFF_F[i], FC['fcff'][i], 1e-6)
    close(PPE_CLOSE[i], FC['ppe'][i], 1e-6)
    close(ND_CLOSE[i], FIN['net_debt'][i], 1e-6)
    close(GROSS_D[i], FIN['gross_debt'][i], 1e-6)
    close(INT_F[i], FIN['interest'][i], 1e-6)
    close(NPA_F[i], FIN['npa'][i], 1e-6)
    close(EQ_CLOSE[i], FBS[i]['equity_parent'], 1e-6)
    close(ROE_F[i], FBS[i]['roe'], 1e-12)
    close(ROIC_F[i], FBS[i]['roic'], 1e-12)
    close(IC_F[i], FBS[i]['invested_capital'], 1e-6)
    close(DC['df'][i], DCFB['df'][i], 1e-12)
    close(DC['pv'][i], DCFB['pv'][i], 1e-6)
    close(DA['pv'][i], DCFA['pv'][i], 1e-6)
    close(TNK_PATH['vlcc'][i], SN['market_cross_check']['vlcc_path'][i], 1e-6)
for k in ('pv_expl', 'ev_ops', 'ev', 'equity', 'fv_aed', 'tv', 'pv_tv'):
    kk = {'pv_expl': 'pv_explicit', 'ev_ops': 'ev_ops', 'ev': 'ev', 'equity': 'equity',
          'fv_aed': 'fv_aed', 'tv': 'tv', 'pv_tv': 'pv_tv'}[k]
    close(DC[k], DCFB[kk], 1e-6)
    close(DA[k], DCFA[kk], 1e-6)
close(DC['tv_share'], DCFB['tv_share'], 1e-12)
close(DA['tv_share'], DCFA['tv_share'], 1e-12)
close(W_EXP, WACC['wacc'], 1e-12); close(W_TERM, WACC['wacc_term'], 1e-12)
close(KE, WACC['ke'], 1e-12); close(KE_A, WACC['ke_beta1'], 1e-12)
close(KD, WACC['kd'], 1e-12); close(KD1, WACC['kd_method1'], 1e-12)
close(KD2, WACC['kd_method2'], 1e-12); close(KD3, WACC['kd_method3'], 1e-12)
close(WE, WACC['we'], 1e-12); close(MKTCAP, WACC['mktcap'], 1e-6)
close(REL_BASE, LN['relative']['base'], 1e-9)
close(REL_BEAR, LN['relative']['bear'], 1e-9)
close(REL_BULL, LN['relative']['bull'], 1e-9)
close(NORM_BASE, LN['normalized']['base'], 1e-9)
close(NORM_BEAR, LN['normalized']['bear'], 1e-9)
close(NORM_BULL, LN['normalized']['bull'], 1e-9)
close(BOOK_BASE, LN['book']['base'], 1e-9)
close(BOOK_BEAR, LN['book']['bear'], 1e-9)
close(BOOK_BULL, LN['book']['bull'], 1e-9)
close(ROE_SUST, BK['roe_sustainable'], 1e-12)
close(PB_FAIR, BK['pb_fair'], 1e-12)
close(BVPS0, BK['bvps_usd'], 1e-12)
close(VSB_RATIO, BK['vessel_value_to_book'], 1e-12)
# the residual-income build, line by line against the committed detail table
assert BK['method'] == 'residual income', BK['method']
close(RI_FADE, BK['fade'], 1e-12)
close(BOOK_EQ, BK['equity_value'], 1e-9)
close(RI['pv_tv'], BK['pv_terminal'], 1e-9)
for i in range(5):
    close(ROE_ORD[i], BK['roe_path'][i], 1e-12)
    close(RI['det'][i]['open'], BK['detail'][i]['opening_book'], 1e-9)
    close(RI['det'][i]['roe'], BK['detail'][i]['roe'], 1e-12)
    close(RI['det'][i]['ri'], BK['detail'][i]['residual_income'], 1e-9)
    close(RI['det'][i]['df'], BK['detail'][i]['discount_factor'], 1e-12)
    close(RI['det'][i]['pv'], BK['detail'][i]['pv'], 1e-9)
# the cost of capital now carries three weights and two hybrid costs
close(KH, WACC['kh'], 1e-12); close(KH_T, WACC['kh_term'], 1e-12)
close(WH, WACC['wh'], 1e-12); close(WD, WACC['wd'], 1e-12)
close(HYBRID_CAP, WACC['hybrid_cap'], 1e-6)
close(WE + WD + WH, 1.0, 1e-12)
close(W_EXP_A, DCFA['wacc'], 1e-12); close(W_TERM_A, DCFA['wacc_term'], 1e-12)
# earnings per share is struck AFTER the perpetual coupon
for i in range(5):
    close(ORD_F[i] / SH / 1000.0, FIN['eps'][i], 1e-12)
    close(NPA_F[i] / SH / 1000.0, FIN['eps_pre_coupon'][i], 1e-12)
    close(ORD_F[i], FIN['npa_ordinary'][i], 1e-6)
close(OWN_EVEB_26_BR, REL['own_ev_ebitda_26_bridge'], 1e-9)
close(EV_BRIDGE, REL['own_ev_bridge'], 1e-6)
# the tanker leg, vessel by vessel
close(OPEX_DAY, FLEET['opex_day'], 1e-9)
close(GAS_RATE, FLEET['gas_rate_day'], 1e-9)
close(VDAYS25, FLEET['vessel_days_25'], 1e-9)
close(TCEREV25, FLEET['tce_rev_25'], 1e-6)
for c in CLS:
    close(SPOT25[c], FLEET['spot_fy25'][c], 1e-9)
    close(SPOT_MID[c], FLEET['spot_mid'][c], 1e-9)
    close(SPOT_Q1[c], FLEET['spot_q1_26'][c], 1e-9)
    close(SPOT_Q2[c], FLEET['spot_q2_26'][c], 1e-9)
    close(BLEND_MID[c], FLEET['blend_mid'][c], 1e-9)
    close(TCE25[c], FLEET['blend_fy25'][c], 1e-9)
    close(TCE24[c], FLEET['blend_fy24'][c], 1e-9)
close(TNK_EB25, SEGH['Tankers']['ebitda'][2], 1e-6)
close(GAS_REV25, SEGH['Gas Carriers']['revenue'][2], 1e-6)
# the three days ratios and the opening working capital are DERIVED on the sheet from the
# audited columns, so the audited columns must be the very figures the model solved them on
close(HB['receivables'][2] + HB['due_from_related'][2], V['recv_fy25'] + V['dfr_fy25'], 1e-6)
close(HB['inventories'][2], V['inv_fy25'], 1e-6)
close(HB['payables'][2] + HB['due_to_related'][2], V['pay_fy25'] + V['dtr_c_fy25'], 1e-6)
close(H_OPCOST[2], OPCOST25, 1e-6)
close(DSO_REPORTED, CCC['dso'][2], 1e-12)
close(DSO_REPORTED, V['dso_days_reported'], 0.05)
# the re-based ratio, against the committed input that records it to one decimal
close(DSO, V['dso_days'], 0.05)
close(GROSSUP25, V['tnk_grossup_25'], 0.0005)
close(GROSSUP, V['tnk_grossup_26'], 1e-12)
# the fleet purchase: counts, price and the two places it lands
close(ACQ_GAS_SECONDHAND_N + ACQ_GAS_NEWBUILD_N, ACQ_GAS_N, 1e-12)
close(sum(ACQ_N.values()), ACQ_VLCC_N, 1e-12)
close(ACQ_CAPEX[0], ACQ_COST, 1e-12)
close(NETDEBT - NDCO - DEFERRED, ACQ_COST, 1e-6)
for i in range(5):
    close(GAS_VY_BASE[i] + GAS_ACQ_VY[i], GAS_VY[i], 1e-12)
    close(YSD['vlcc'][i] - (OWNED['vlcc'] * YRDAYS[i] - YCD['vlcc'][i]), YACD['vlcc'][i],
          1e-12)
for i in range(1, 5):                    # a full year of the acquired ships from 2027
    close(YACD['vlcc'][i], ACQ_VLCC_N * YRDAYS[i], 1e-12)
    close(GAS_ACQ_VY[i], float(ACQ_GAS_N), 1e-12)
close(KD_BALWTD, WACC['kd_balance_weighted'], 1e-12)
close(HS_REL, V['handysize_relative'], 1e-12)
for _c26 in (0, 1):                      # the two disclosed 2026 quarters, handysize
    close(B26['hs'][_c26], B26['mr'][_c26] * HS_REL, 1e-12)
close(CENTRAL, D['central'], 1e-9)
close(CENTRAL_A, D['central_beta_alt'], 1e-9)
# the beta framing block published in the study must be the framing the workbook builds
close(BF['primary']['beta'], V['beta'], 1e-12)
close(BF['primary']['ke'], KE, 1e-12)
close(BF['primary']['wacc'], W_EXP, 1e-12)
close(BF['primary']['fv'], DC['fv_aed'], 1e-9)
close(BF['primary']['central'], CENTRAL, 1e-9)
close(BF['alternative']['beta'], V['beta_composite'], 1e-12)
close(BF['alternative']['ke'], KE_A, 1e-12)
close(BF['alternative']['fv'], DA['fv_aed'], 1e-9)
close(BF['alternative']['central'], CENTRAL_A, 1e-9)
close(BF['ci90'][0], V['beta_ci_lo'], 1e-12)
close(BF['ci90'][1], V['beta_ci_hi'], 1e-12)
close(BF['blume'], V['beta_blume'], 1e-12)
close(KE_BLUME, WACC['ke_blume'], 1e-12)
# the sensitivity beta grid must span the two constructions and the interval around the
# adopted one — not round numbers chosen by hand
close(SN['betas'][0], V['beta_composite'], 1e-12)
close(SN['betas'][2], V['beta'], 1e-12)
close(SN['betas'][4], V['beta_ci_hi'], 1e-12)
assert SN['betas'] == sorted(SN['betas']), 'the sensitivity beta grid is not monotone'
_gi = SN['gs'].index(G)
# THE GRID MUST BE CENTRED ON THE VALUE IT BRACKETS. Every cell is a complete re-run of
# the model, so nothing forces those re-runs to discount at the rate the published
# valuation uses — and for one edition they did not: the perpetual tranche reached the base
# and terminal rates but not the scenario rate, and the grid's own centre cell printed 6.85
# against a published 6.03. The identity is asserted here, at the centre cell and at 1.00x
# on both single-driver rows, and the whole beta row is reproduced from its own beta so a
# grid that silently changed construction could not pass.
close(SN['grid_beta_g'][2][_gi], DC['fv_aed'], 1e-6)
close(SN['grid_beta_g'][0][_gi], DA['fv_aed'], 1e-6)
close(SN['anchor']['1.0'], DC['fv_aed'], 1e-6)
close(SN['capex']['1.0'], DC['fv_aed'], 1e-6)
for _bi, _b in enumerate(SN['betas']):
    _kes = RF_STAR + _b * V['erp_total']; _ket = V['rf_terminal'] + _b * V['erp_total']
    close(SN['grid_beta_g'][_bi][_gi],
          dcf_legs(WE * _kes + WD * KD_AT + WH * KH,
                   WE * _ket + WD * KD_T_AT + WH * KH_T)['fv_aed'], 1e-6)

close(CENTRAL_BEAR, LN['central']['bear'], 1e-9)
close(CENTRAL_BULL, LN['central']['bull'], 1e-9)
close(BLEND_EV, REL['blend_ev_ebitda'], 1e-12)
close(BLEND_PE, REL['blend_pe'], 1e-12)
close(OWN_EVEB_TTM, REL['own_ev_ebitda_ttm'], 1e-9)
close(OWN_EVEB_26, REL['own_ev_ebitda_26'], 1e-9)
close(OWN_PE_TTM, REL['own_pe_ttm'], 1e-9)
close(OWN_PB, REL['own_pb'], 1e-9)
close(OWN_DY, REL['own_dy'], 1e-12)
close(SOTP_EVOPS, SOTP['ev_ops'], 1e-6)
close(SOTP_EQ, SOTP['equity'], 1e-6)
close(SOTP_FV, SOTP['fv_aed'], 1e-9)
for j, g in enumerate(GROUPS):
    close(SOTP_EV[g], SOTP['legs'][j]['ev'], 1e-6)
    close(GRP_EB_F[g][0], SOTP['legs'][j]['ebitda_26'], 1e-6)
for i in range(5):
    for s in SEGS:
        close(SEG_REV_F[s][i], D['fcst_seg'][s]['rev'][i], 1e-6)
        close(SEG_EB_F[s][i], D['fcst_seg'][s]['ebitda'][i], 1e-6)

ANCH.update(
    fv=f"DCF!C{DF_['fvaed']}", fv_beta_alt=f"DCF!C{DF_['fvaeda']}",
    ke_ci_lo=f"DCF!C{DF_['kecilo']}", ke_ci_hi=f"DCF!C{DF_['kecihi']}",
    ke_blume=f"DCF!C{DF_['keblume']}",
    fv_usd=f"DCF!C{DF_['fvusd']}", pv_expl=f"DCF!C{DF_['pvex']}", tv=f"DCF!C{DF_['tv']}",
    ev=f"DCF!C{DF_['ev']}", tv_share=f"DCF!C{DF_['tvshare']}",
    wacc=f"DCF!C{DF_['wacc']}", wacc_term=f"DCF!C{DF_['waccterm']}",
    ke=f"DCF!C{DF_['ke']}", kd=f"DCF!C{DF_['kd']}",
    rev26=f"DCF!B{DF_['rev']}", ebitda26=f"DCF!B{DF_['ebitda']}",
    ebitda30=f"DCF!F{DF_['ebitda']}", nopat26=f"DCF!B{DF_['nopat']}",
    tax26=f"DCF!B{DF_['taxtot']}", fcff26=f"DCF!B{DF_['fcff']}",
    tankers26=f"Segments!B{SG['teb']}", tankers30=f"Segments!F{SG['teb']}",
    gas28=f"Segments!D{SG['gaseb']}",
    central=f"Summary!C{SU['central']}", central_beta_alt=f"Summary!C{SU['centrala']}",
    relative=f"'Relative & Normalized'!C{RN['base']}",
    normalized=f"'Relative & Normalized'!C{RN['nbase']}",
    book=f"'Relative & Normalized'!C{RN['bbase']}",
    book_bear=f"'Relative & Normalized'!C{RN['bbear']}",
    book_bull=f"'Relative & Normalized'!D{RN['bbear']}",
    roe_sust=f"'Relative & Normalized'!C{RN['broe']}",
    sotp=f"'SOTP Bridge'!C{SB['bfv']}",
    nd30=f"'Balance Sheet'!I{BS['nd']}", bvps30=f"'Balance Sheet'!I{BS['bvps']}",
    nwc26=f"'Balance Sheet'!B{BS['wcnwc']}", ppe30=f"'Balance Sheet'!F{BS['ppeclose']}",
    npa26=f"'Income Statement'!E{IS['npa']}",
    ordn26=f"'Income Statement'!E{IS['ordn']}")

ANCH.update(book_equity=f"'Relative & Normalized'!C{RN['beq']}",
            eveb_bridge=f"'Relative & Normalized'!C{RN['ebbr_26']}",
            wh=f"DCF!C{DF_['whyb']}", kh=f"DCF!C{DF_['kh']}",
            kh_term=f"DCF!C{DF_['khterm']}",
            eps26=f"'Income Statement'!E{IS['eps']}",
            eps26_pre=f"'Income Statement'!E{IS['epspre']}",
            tnk_spot_vlcc_q1=f"Segments!F{SG['sp0']+4}",
            tnk_tce26=f"Segments!B{SG['tcerev']}",
            tnk_tce30=f"Segments!F{SG['tcerev']}",
            tnk_opexday=f"Segments!B{SG['opexd0']}",
            gas_rate_solved=f"Segments!B{SG['gasrate0']}",
            tnk_chrev26=f"Segments!B{SG['chrevt']}",
            tnk_sprev26=f"Segments!B{SG['sprevt']}",
            tnk_acqdays26=f"Segments!B{SG['yacd0']+4}",
            gas26=f"Segments!B{SG['gaseb']}",
            gas_vy27=f"Assumptions!C{A['gas_vy']}",
            kd_balance_weighted=f"DCF!C{DF_['kdbal']}",
            dso=f"Assumptions!C{A['dso']}",
            dso_reported=f"Assumptions!C{A['dso_rep']}",
            ppe26=f"'Balance Sheet'!B{BS['ppeclose']}")

out = os.path.join(HERE, 'ADNOCLS_Valuation_Model_09082026_public.xlsx')
wb.save(out)
# The row plans go out with the expectations, so the gates address rows symbolically. A
# hand-kept copy of these numbers in three other files drifted the moment a row moved.
json.dump({'expected': EXPECT, 'anchors': ANCH,
           'rows': {'Segments': SG, 'Relative & Normalized': RN, 'DCF': DF_,
                    'Income Statement': IS, 'Balance Sheet': BS, 'Cash Flow': CF,
                    'SOTP Bridge': SB, 'Summary': SU, 'Fundamental Valuation': FV,
                    'Per-Share & Ratios': PS, 'Peer & Sector': PR, 'Assumptions': A,
                    'Monte Carlo': MC, 'Sensitivity': SE}},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nchk = sum(len(v) for v in EXPECT.values())
# THE COUNT IS TAKEN OFF THE DELIVERED FILE, NOT THE OBJECT IN MEMORY. A cell written as a
# date serial with a date format comes back from the saved file as a date, and counting the
# in-memory object instead reported thirty-six calendar cells as pasted figures — a number
# that did not match what a reader measuring the delivered workbook would find.
import openpyxl as _xl
_del = _xl.load_workbook(out)
nform = nlit = ndate = 0
per = {}
for _s in _del.worksheets:
    f_ = l_ = d_ = 0
    for row in _s.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith('='):
                f_ += 1
            elif isinstance(v, (_dt.datetime, _dt.date)):
                d_ += 1
            elif isinstance(v, (int, float)):
                l_ += 1
    per[_s.title] = (f_, l_, d_); nform += f_; nlit += l_; ndate += d_
print(f'wrote {out} | {len(wb.sheetnames)} sheets: {wb.sheetnames}')
for k, (f_, l_, d_) in per.items():
    print(f'  {k:24s} formulas {f_:5d}   pasted numeric {l_:5d}   dates {d_:3d}')
print(f'formulas: {nform} (of which {nchk} carry a checked expected value) | '
      f'pasted numeric cells: {nlit} | calendar cells: {ndate}')
