"""DU_Valuation_Model_09082026_public.xlsx — 16 sheets mirroring the house canonical
model (operating-company variant). Blue = inputs · black = formulas · green = cross-sheet
links.

The workbook is FORMULA-DRIVEN. Every quantity that is arithmetically derivable from an
input is written as a live Excel formula, not as a pasted number, so the reader can trace
each figure back to the drivers on the Assumptions sheet and change one to see the model
reprice. Only three classes of cell are pasted values:

  1. audited and disclosed historical figures (the primary record);
  2. the unit build's OUTPUT — forecast segment revenue for the four disclosed segments
     (Mobile and Fixed built as subscribers x ARPU, Wholesale and ICT grown at segment
     level), pasted per year with the subscriber/ARPU driver table shown beside it;
     everything downstream (contribution, opex stack, EBITDA, D&A, the DCF) is formula;
  3. engine outputs that are whole-model re-runs by construction: the Monte Carlo price
     map, the sensitivity grids, and the DCF scenario bear/bull bounds.

Every formula cell also carries the model's own value for that cell into
xlsx_expected.json, and recalc.py evaluates the workbook independently and asserts the two
agree. A formula that computes the right thing the wrong way therefore fails the gate.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
YH = ['FY2023', 'FY2024', 'FY2025']
YF = D['fcst']['years']
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, REL, NRM, BK = D['experts'], D['rel'], D['norm'], D['book']
SEG, S0, STK, BU = D['seg_fy25'], D['step0'], D['strike'], D['bottomup']
UC = D['unitcost']
IN = {k: v['value'] for k, v in D['inputs'].items()}
TRC = D['terminal_record']
TRI, TRO = TRC['inputs'], TRC['outputs']
SPOT, SH = M['spot'], M['shares_mn']
SEGS = ['mobile', 'fixed', 'wholesale', 'ict']
TAX = IN['tax_eff']
PAYOUT = F['payout']
NWC_PCT = D['nwc_pct']
NETCASH = HB['FY25']['net_cash']
LEASE = HB['FY25']['lease']
H3 = ['FY23', 'FY24', 'FY25']
CD = ['B', 'C', 'D', 'E', 'F']              # forecast columns on the DCF / Segments blocks
HC = ['B', 'C', 'D']                        # historical columns on the statements
FCOL = ['E', 'F', 'G', 'H', 'I']            # forecast columns on the statements
CFF = ['D', 'E', 'F', 'G', 'H']             # forecast columns on the cash-flow sheet
ALL = HC + FCOL

wb = Workbook()
EXPECT = {}
ANCH = {}

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula and record the model's own value for the same cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Emirates Integrated Telecommunications Company PJSC (DFM: DU)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the du valuation study. Every blue cell is an input;',
 'every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice: the cost of capital is built from',
 'the risk-free rate net of the sovereign spread, beta and the premium rather than pasted; the discount',
 'factors compound from the glide; and the income statement, balance sheet, cash flow, ratios and all four',
 'lenses chain off the same cells.', '',
 'THREE THINGS ARE PASTED VALUES, and it is worth knowing exactly which. First, audited and disclosed',
 'history — the primary record, not a calculation. Second, the unit build\'s output: forecast revenue for',
 'the four segments the company itself discloses (Mobile, Fixed, Wholesale, ICT and associated telecom',
 'services). Mobile and Fixed are built bottom-up as subscribers x ARPU — the driver table is shown on the',
 'Segments sheet — and Wholesale and ICT are grown at segment level because the company discloses no unit',
 'KPIs for them; only the build\'s OUTPUT (segment revenue per year) is pasted, and everything downstream of',
 'it is formula. Third, whole-model engine outputs, where each figure is a complete re-run of the entire',
 'valuation and so cannot be a single formula: the Monte Carlo price map, the sensitivity grids, and the',
 'DCF scenario bear/bull bounds. Everything else — including every lens base value, the relative/normalised/',
 'book bear and bull bounds, and the anchor-date roll — is a live formula. Changing a driver reprices the',
 'model but does NOT redraw the engine outputs.', '',
 'How revenue is built. Not as one growth rate. Mobile = average subscriber base x blended ARPU x 12; the',
 'subscribers-x-ARPU frame reproduces the audited FY2025 mobile segment to within 0.1%. Fixed = subscriber',
 'base x implied revenue per subscriber (a consumer-plus-enterprise blend, so an intensity metric, not a',
 'tariff). Wholesale and ICT are grown on their own paths — the war-hit roaming/transit recovery and the',
 'data-centre ramp respectively.', '',
 'How cost is built, and why no margin is an input. Direct cost is a COST PER UNIT, not a margin. Mobile',
 'carries three separate per-subscriber lines - interconnect, commission, and devices - each with its own',
 'escalator, because they are driven by physically different things: termination rates and messaging-app',
 'substitution, the cost of winning a customer, and handset volume. Fixed carries a per-subscriber capacity',
 'cost. Wholesale and ICT carry a cost rate on their own revenue because the company discloses no volume',
 'unit for either - that gap is flagged rather than filled. Every rate is anchored on the H1-2026 reviewed',
 'actual and held flat unless a named mechanism has a measured direction in the half-year pair. CONTRIBUTION',
 'MARGIN AND GROUP EBITDA MARGIN ARE THEREFORE OUTPUTS, computed on the Segments sheet, never typed in.', '',
 'The contested judgement — the required return. This is the one judgement that moves the answer more',
 'than any other, so it is computed BOTH WAYS on the Fundamental Valuation sheet and never averaged.',
 "Framing 1 takes du's own measured beta and capitalises the terminal at the Gordon rate. Framing 2",
 "refuses the re-rating that implies and holds du's CURRENT trailing EV/EBITDA into perpetuity instead.",
 'The gap between them is the study, and it is published, not resolved. (The post-2026 fiscal regime',
 'was the prior edition\'s contested judgement; it is no longer contested — du disclosed the 2027-2029',
 'extension itself on 24 July 2026 on the same structure, with the AED 1.8bn combined floor retained.',
 'A reversion after 2029 is still priced, as a named tail, on the same sheet.)', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges.', '',
 'Sourcing note, up front. FY2023, FY2024 and FY2025 all come from the company\'s own audited consolidated',
 'financial statements read from investors.du.ae, and H1-2026 from the KPMG-reviewed interims — every',
 'income-statement, balance-sheet and segment line is the filed figure. Segment revenue ties EXACTLY to',
 'consolidated revenue in every year (Note 38). du carries ZERO drawn borrowings in every year studied; the',
 'only debt-like item is IFRS-16 lease liabilities, and the bridge treats them as debt. Every input is',
 'annotated where it appears and listed with source and date in the companion bibliography document.', '',
 'Leases are debt, and charged once. du has no drawn borrowings; its only debt-like item is the IFRS-16',
 'lease liability. That liability is deducted in the enterprise-to-equity bridge and carries a debt',
 'weight in the cost of capital, so NO lease charge belongs in the cash-flow waterfall as well —',
 'perpetual renewal is already paid for in the terminal, where invested capital includes the',
 'right-of-use asset and terminal reinvestment maintains it. The right-of-use book is held flat because',
 'a new lease creates an asset and a liability together and is non-cash.', '',
 'Discount convention, stated. Each explicit year is discounted at its own forward cost of capital,',
 f"at full-year END-of-period factors, gliding {W['wacc_exp']*100:.2f}% -> {W['wacc_term']*100:.2f}% "
 'on the AED risk-free path (du has no debt whose cost could',
 'define the glide); the terminal value, a value dated at the end of year five, is discounted at the',
 'year-five factor. A mid-year convention would raise the answer and is not adopted.',
 'One date, one price of time: the bridge is dated 31-Dec-2025 and rolled to the 07-Aug-2026 anchor at',
 f"the cost of equity, net of the AED {IN['div_between']:.2f} of dividends whose EX-dates fall in that",
 'window (the 0.40 final, paid 28-Apr-2026, and the 0.26 interim, ex 31-Jul-2026).', '',
 f"Currency. AED million unless stated. Spot AED {SPOT:.2f} ({M['asof']} close). Sheets: READ FIRST · Summary ·",
 'Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income',
 'Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share &',
 'Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 112

# ============ 2 SUMMARY =======================================================
ws = sheet('Summary')
title(ws, 'Summary — valuation at a glance', 'All values link live to their source sheets', 7,
      awidth=44, cwidth=15)
hdr(ws, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Role', '', 'vs price'])
LENS_SRC = {'dcf': '=DCF!C63', 'relative': "='Relative & Normalized'!C11",
            'normalized': "='Relative & Normalized'!C27", 'book': "='Relative & Normalized'!C35"}
BEAR_SRC = {'relative': "='Relative & Normalized'!C12",
            'normalized': "='Relative & Normalized'!E27", 'book': "='Relative & Normalized'!E35"}
BULL_SRC = {'relative': "='Relative & Normalized'!D12",
            'normalized': "='Relative & Normalized'!F27", 'book': "='Relative & Normalized'!F35"}
LK = ['dcf', 'relative', 'normalized', 'book']
RETW = D['lens_record']['retired']['blend']
# The 'vs price' column divides by the market price, which sits at the FOOT of this block —
# below the rows that reference it — so its address is computed here from the layout and
# ASSERTED when the row is actually written. Retiring the blend added a row, and every
# absolute reference left behind silently re-pointed at the expert-panel median instead of
# the price: a formula naming a cell by address moves with the re-issue, and only an
# assertion notices when it does not.
SPOT_ROW = 5 + len(LK) + 6
# THE WEIGHT AND CONTRIBUTION COLUMNS WENT WITH THE BLEND, and a role column replaces
# them: which lens is the answer, which sit beside it, and which this class does not
# publish at all.
ROLE = {'dcf': 'THE CENTRAL — the class primary', 'relative': 'cross-check',
        'normalized': 'REMOVED — not a lens this class publishes',
        'book': 'a disclosed floor, never weighted'}
r = 5
for k in LK:
    l = LN[k]
    put(ws, f'A{r}', l['name'], fmt=None)
    if k in BEAR_SRC:
        putf(ws, f'B{r}', BEAR_SRC[k], l['bear'], PX, green=True)
    else:
        put(ws, f'B{r}', l['bear'], BLUE, PX)   # DCF bear/bull are whole-model scenario re-runs
    putf(ws, f'C{r}', LENS_SRC[k], l['base'], PX, green=True)
    if k in BULL_SRC:
        putf(ws, f'D{r}', BULL_SRC[k], l['bull'], PX, green=True)
    else:
        put(ws, f'D{r}', l['bull'], BLUE, PX)
    put(ws, f'E{r}', ROLE[k], fmt=None)
    putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', l['base'] / SPOT - 1, PCT)
    r += 1
band(ws, r, 7)
put(ws, f'A{r}', 'THE CENTRAL — the cash-flow lens, not an average', bold=True, fmt=None)
ANCH['summary_central'] = f'C{r}'
putf(ws, f'B{r}', '=B5', LN['dcf']['bear'], PX, bold=True)
putf(ws, f'C{r}', '=C5', D['central'], PX, bold=True)
putf(ws, f'D{r}', '=D5', LN['dcf']['bull'], PX, bold=True)
put(ws, f'E{r}', 'the class primary', fmt=None)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['central'] / SPOT - 1, PCT, bold=True)
r += 1
put(ws, f'A{r}', 'NOT AVERAGED — the retired blend, published unused', bold=True, fmt=None)
putf(ws, f'C{r}', '=' + '+'.join('C%d*%g' % (5 + i, RETW[k]) for i, k in enumerate(LK)),
     D['retired_blend_value'], PX)
put(ws, f'E{r}', 'retired 02-Sep-2026', fmt=None)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['retired_blend_value'] / SPOT - 1, PCT)
r += 1
put(ws, f'A{r}', 'Span across the lenses (min/max) — a spread between METHODS, not a range '
                 'around the answer', fmt=None)
putf(ws, f'B{r}', '=MIN(B5:B8)', min(LN[k]['bear'] for k in LK), PX)
putf(ws, f'D{r}', '=MAX(D5:D8)', max(LN[k]['bull'] for k in LK), PX)
r += 1
put(ws, f'A{r}', "Contested judgement, other way — no terminal re-rating (Framing 2)", fmt=None)
SUMMARY_ALT_ROW = r   # resolved after the Fundamental Valuation rows are known
ANCH['summary_alt'] = f'C{r}'
put(ws, f'C{r}', DCF['ps_mkt_term'], BLUE, PX)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', DCF['ps_mkt_term'] / SPOT - 1, PCT)
r += 1                                             # r = 12
put(ws, f'A{r}', 'Terminal value share of DCF enterprise value', fmt=None)
ANCH['summary_tv_share'] = f'C{r}'
putf(ws, f'C{r}', '=DCF!C29', DCF['tv_share'], PCT, green=True)
r += 1                                             # r = 13
put(ws, f'A{r}', 'Expert panel median', fmt=None)
ANCH['summary_panel'] = f'C{r}'
putf(ws, f'C{r}', "='Fundamental Valuation'!C27", D['panel_centre'], PX, green=True)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['panel_centre'] / SPOT - 1, PCT)
r += 1
band(ws, r, 7)
ANCH['summary_spot'] = f'C{r}'
assert r == SPOT_ROW, (f'market price landed on row {r}, not the row {SPOT_ROW} that every '
                       f'"vs price" formula on this sheet divides by')
put(ws, f'A{r}', 'Market price (anchor)', bold=True, fmt=None)
put(ws, f'C{r}', SPOT, BLUE, PX, bold=True)
r += 2
hdr(ws, r, ['Key figure', 'Value'])
KEY = [('Shares outstanding (mn)', 'SHARES', SH, NUM0),
       ('Market capitalisation (AED mn)', 'MKTCAP', M['mktcap'], NUM0),
       ('Lease liabilities, FY2025 (AED mn) — the only debt-like item', 'LEASE', LEASE, NUM0),
       ('Cash and term deposits, FY2025 (AED mn)', 'NETCASH', NETCASH, NUM0),
       ('FY2025 revenue (AED mn)', "='Income Statement'!D5", HI['FY25']['rev'], NUM0),
       ('FY2025 EBITDA (AED mn)', "='Income Statement'!D9", HI['FY25']['ebitda'], NUM0),
       ('FY2025 net profit (AED mn)', "='Income Statement'!D18", HI['FY25']['np'], NUM0),
       ('Cost of capital — explicit window', '=DCF!C47', W['wacc_exp'], PCT),
       ('Cost of capital — terminal', '=DCF!C54', W['wacc_term'], PCT),
       ('Terminal growth', '=DCF!C24', TRI['nominal_growth'], PCT)]
SUMMARY_KEY_START = r + 1

# ============ 3 FUNDAMENTAL VALUATION =========================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — the four lenses and the contested judgement', None, 6,
      awidth=52, cwidth=15)
hdr(ws, 4, ['Lens / step', 'Basis', 'AED per share'])
rows = [
    ('Discounted cash flow', 'links to the DCF sheet', "=DCF!C63", DCF['ps']),
    ('  bear', 'ARPU −5%, subscribers −250k, margins −3%, +100bp cost of capital, g 2%',
     LN['dcf']['bear'], None),
    ('  bull', 'ARPU +3%, subscribers +200k, margins +1.5%, −50bp cost of capital, g 3%',
     LN['dcf']['bull'], None),
    ('Relative multiples', f"{IN['pe_just']}x justified P/E on FY2026E earnings",
     "='Relative & Normalized'!C11", LN['relative']['base']),
    ('Normalised earnings power', f"{IN['pe_just']}x normalised earnings per share",
     "='Relative & Normalized'!C27", LN['normalized']['base']),
    ('Book value and sustainable return', 'justified price-to-book on sustainable return on equity',
     "='Relative & Normalized'!C35", LN['book']['base']),
]
r = 5
for a_, b_, c_, xp in rows:
    put(ws, f'A{r}', a_, fmt=None); put(ws, f'B{r}', b_, fmt=None)
    if isinstance(c_, str):
        putf(ws, f'C{r}', c_, xp, PX, green=True)
    else:
        put(ws, f'C{r}', c_, BLUE, PX)
    r += 1
r += 1
band(ws, r, 3)
put(ws, f'A{r}', 'THE CENTRAL — the cash-flow lens itself, not an average of the four',
    bold=True, fmt=None)
ANCH['fv_central'] = f'C{r}'
putf(ws, f'C{r}', f"=Summary!{ANCH['summary_central']}", D['central'], PX, bold=True,
     green=True)
r += 2
put(ws, f'A{r}', 'THE CONTESTED JUDGEMENT — THE REQUIRED RETURN, BOTH WAYS', bold=True,
    fmt=None); r += 1
_F1_ROW = r                                   # Framing 1 lands here; rows captured, never hardcoded
for lab, val, fmt, xp, fml in [
        ("Framing 1: du's own measured beta sets the cost of equity", None, PX, DCF['ps'],
         '=DCF!C63'),
        ('  the terminal that implies — exit multiple on terminal EBITDA',
         DCF['tv_implied_mult'], MULT, None, None),
        ("  against du's OWN current trailing EV/EBITDA", DCF['ev_ebitda_now'], MULT, None, None),
        ("Framing 2: no terminal re-rating — du's current multiple held in perpetuity",
         DCF['ps_mkt_term'], PX, None, None),
        ('The judgement is worth (Framing 1 less Framing 2, per share)', None, PX,
         DCF['ps'] - DCF['ps_mkt_term'], 'GAP'),
        ('Post-2029 fiscal tail: the pre-2024 royalty construction returns after the disclosed '
         '2027-2029 extension lapses', DCF['ps_framing_b'], PX, None, None)]:
    put(ws, f'A{r}', lab, fmt=None)
    if fml == 'GAP':
        putf(ws, f'C{r}', f'=C{_F1_ROW}-C{_F1_ROW + 3}', xp, fmt)
    elif fml is not None:
        putf(ws, f'C{r}', fml, xp, fmt, green='DCF' in fml)
    elif val is not None:
        put(ws, f'C{r}', val, BLUE, fmt)
    r += 1
FRAMING2_ROW = _F1_ROW + 3
ANCH['fv_framing1'] = f'C{_F1_ROW}'; ANCH['fv_framing2'] = f'C{FRAMING2_ROW}'
ANCH['fv_gap'] = f'C{_F1_ROW + 4}'; ANCH['fv_tail'] = f'C{_F1_ROW + 5}'
r += 1
put(ws, f'A{r}', 'EXPERT PANEL', bold=True, fmt=None); r += 1
hdr(ws, r, ['Expert', 'Method', 'Base (AED/share)', 'Low', 'High']); r += 1
_PANEL_FIRST = r
for k, nm in [('e1', 'Expert 1'), ('e2', 'Expert 2'), ('e3', 'Expert 3')]:
    e = EXP[k]
    put(ws, f'A{r}', nm, fmt=None); put(ws, f'B{r}', e['method_short'], fmt=None)
    put(ws, f'C{r}', e['base'], BLUE, PX); put(ws, f'D{r}', e['rng'][0], BLUE, PX)
    put(ws, f'E{r}', e['rng'][1], BLUE, PX); r += 1
band(ws, r, 5); put(ws, f'A{r}', 'Panel median', bold=True, fmt=None)
putf(ws, f'C{r}', f'=MEDIAN(C{_PANEL_FIRST}:C{_PANEL_FIRST + 2})', D['panel_centre'],
     PX, bold=True)
ANCH['fv_panel'] = f'C{r}'
r += 2
put(ws, f'A{r}', 'The risk-free tenor question, priced', fmt=None); r += 1
put(ws, f'A{r}', 'There is no liquid 10-year AED government point. The base model uses the '
    'Jan-2031 AED T-bond (4.48%); discounting instead at the peg-extrapolated 10-year proxy '
    '(4.69%) is an engine re-run worth the figure at right', fmt=None, wrap=True)
put(ws, f'C{r}', DCF['ps_rf_alt'], BLUE, PX)
ws.row_dimensions[r].height = 40

# ============ 4 ASSUMPTIONS ====================================================
ws = sheet('Assumptions')
title(ws, 'Assumptions — every input in the model', 'Blue cells are inputs. Change one and the '
      'model reprices: everything downstream is a formula.', 8, awidth=52, cwidth=13)
r = 4
A = {}          # key -> row on Assumptions

def block(name, items):
    global r
    band(ws, r, 8); put(ws, f'A{r}', name, bold=True, fmt=None); r += 1
    for key, lab, val, fmt in items:
        put(ws, f'A{r}', lab, fmt=None)
        if isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(ws, f'{get_column_letter(2+i)}{r}', v, BLUE, fmt)
        else:
            put(ws, f'C{r}', val, BLUE, fmt)
        A[key] = r
        r += 1
    r += 1

def a(key, i=None):
    """Absolute reference to an Assumptions cell; i selects a year column for list inputs."""
    col = get_column_letter(2 + i) if i is not None else 'C'
    return f"Assumptions!${col}${A[key]}"

hdr(ws, 3, ['Input', YF[0], YF[1], YF[2], YF[3], YF[4]])
block('Anchors', [
    ('spot', 'Spot price (AED)', SPOT, PX),
    ('shares', 'Shares outstanding (mn)', SH, NUM0),
    ('tax_eff', 'Combined federal royalty + income tax rate (Framing A, audited FY2025)',
     IN['tax_eff'], PCT),
    ('reg_share', 'Regulated revenue share (Framing B base, audited FY2023)', IN['reg_share'], PCT),
    ('royB_rev', 'Framing B — royalty rate on regulated revenue', IN['royB_rev_rate'], PCT),
    ('royB_prof', 'Framing B — royalty rate on regulated profit', IN['royB_prof_rate'], PCT)])
block('H1-2026 reviewed segment actuals — the base the FY2026 build chains off', [
    ('h1_mobile', 'Mobile revenue, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['h1_26_seg']['mobile'], NUM0),
    ('h1_fixed', 'Fixed revenue, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['h1_26_seg']['fixed'], NUM0),
    ('subs_m_q2', "Mobile subscribers at 30-Jun-2026 ('000, reviewed period end)",
     BU['subs_mobile']['Q2_2026'], NUM0),
    ('subs_f_q2', "Fixed subscribers at 30-Jun-2026 ('000, reviewed period end)",
     BU['subs_fixed']['Q2_2026'], NUM0),
    ('whl_fy25', 'Wholesale revenue FY2025 (AED mn, audited)', SEG['rev']['wholesale'], NUM0),
    ('ict_fy25', 'ICT revenue FY2025 (AED mn, audited)', SEG['rev']['ict'], NUM0),
    ('h1_whl', 'Wholesale revenue, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['h1_26_seg']['wholesale'], NUM0),
    ('h1_ict', 'ICT revenue, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['h1_26_seg']['ict'], NUM0)])
block('Unit build — mobile and fixed (subscribers x ARPU)', [
    ('subs_m', "Mobile subscribers, end of year ('000)", IN['subs_mobile_path'], NUM0),
    ('subs_f', "Fixed subscribers, end of year ('000)", IN['subs_fixed_path'], NUM0),
    ('arpu_m', 'Blended mobile ARPU (AED/month)', IN['arpu_mobile_path'], NUM1),
    ('arpu_f', 'Implied fixed revenue per subscriber (AED/month)', IN['arpu_fixed_path'], NUM0),
    ('g_whl', 'Wholesale revenue growth', IN['seg_g']['wholesale'], PCT),
    ('g_ict', 'ICT and associated telecom revenue growth', IN['seg_g']['ict'], PCT)])
block('Direct-cost stack — cost per unit, one escalator per driver class. '
      'CONTRIBUTION MARGIN IS NOT AN INPUT HERE: it is computed on the Segments sheet as '
      'what is left after each cost line is grown on its own physical driver.', [
    ('uc_inter_h1', 'Mobile interconnect cost, H1-2026 actual (AED/subscriber/month)',
     UC['hist']['H126']['mob_inter'], NUM2),
    ('esc_inter', 'Mobile interconnect escalator (termination rates, OTT substitution)',
     IN['esc_dc_inter'], PCT),
    ('uc_comm_h1', 'Mobile commission cost, H1-2026 actual (AED/subscriber/month)',
     UC['hist']['H126']['mob_comm'], NUM2),
    ('esc_comm', 'Mobile commission escalator (acquisition/retention cost)',
     IN['esc_dc_comm'], PCT),
    ('uc_dev_h1', 'Mobile devices and direct services, H1-2026 actual (AED/subscriber/month)',
     UC['hist']['H126']['mob_dev'], NUM2),
    ('esc_dev', 'Mobile devices escalator (held flat — lumpy, no trend read into it)',
     IN['esc_dc_dev'], PCT),
    ('uc_fixed_h1', 'Fixed capacity and direct cost, H1-2026 actual (AED/subscriber/month)',
     UC['hist']['H126']['fixed_cap'], NUM2),
    ('esc_fixed', 'Fixed capacity escalator (held flat against an observed decline)',
     IN['esc_dc_fixed'], PCT),
    ('dc_rate_whl', 'Wholesale direct cost / wholesale revenue (H1-2026 rate, held flat)',
     IN['dc_rate_wholesale'], PCT),
    ('dc_rate_ict', 'ICT direct cost / ICT revenue (H1-2026 rate, held flat)',
     IN['dc_rate_ict'], PCT),
    ('h1_dc_mobile', 'Mobile direct cost, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['seg_dc_h1']['H126']['mobile'], NUM0),
    ('h1_dc_fixed', 'Fixed direct cost, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['seg_dc_h1']['H126']['fixed'], NUM0),
    ('h1_dc_whl', 'Wholesale direct cost, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['seg_dc_h1']['H126']['wholesale'], NUM0),
    ('h1_dc_ict', 'ICT direct cost, six months to 30-Jun-2026 (AED mn, reviewed)',
     IN['seg_dc_h1']['H126']['ict'], NUM0),
    ('arpu_drift', 'Blended ARPU drift (mix-exhaustion sensitivity; zero in the base case)',
     IN['arpu_drift'], PCT)])
block('Operating expense stack — one escalator per cost class', [
    ('staff26', 'Staff cost, FY2026E level (AED mn)', IN['staff_fy26'], NUM0),
    ('esc_staff', 'Staff escalator (UAE wage inflation)', IN['esc_staff'], PCT),
    ('network_fy25', 'Network & maintenance, FY2025 base (AED mn, audited)',
     IN['opex_base_fy25']['network'], NUM0),
    ('esc_network', 'Network escalator (network scale)', IN['esc_network'], PCT),
    ('admin_fy25', 'Administrative expense, FY2025 base (AED mn, audited)',
     IN['opex_base_fy25']['admin'], NUM0),
    ('esc_admin', 'Administrative escalator (CPI)', IN['esc_admin'], PCT),
    ('other_fy25', 'Other operating expense, FY2025 base (AED mn, audited)',
     IN['opex_base_fy25']['other'], NUM0),
    ('esc_other', 'Other-expense escalator (CPI)', IN['esc_other'], PCT),
    ('marketing_pct', 'Marketing / revenue', IN['marketing_pct'], PCT),
    ('licence_pct', 'Telecom licence and related fees / revenue (regulatory revenue share)',
     IN['licence_pct'], PCT),
    ('ecl_pct', 'Expected credit losses / revenue', IN['ecl_pct'], PCT),
    ('other_inc', 'Other operating income (AED mn/yr)', IN['other_inc_path'], NUM0)])
block('Capital intensity and working capital', [
    ('capex_pct', 'Capital expenditure / revenue', IN['capex_pct'], PCT),
    ('tang_share', 'Tangible share of capex (audited FY2025)', IN['capex_tang_share'], PCT),
    ('dep_rate', 'PP&E depreciation rate on opening balance (audited FY2025)',
     IN['dep_rate_ppe'], PCT),
    ('amort_rate', 'Intangibles amortisation rate on opening balance (audited FY2025)',
     IN['amort_rate'], PCT),
    ('rou_dep', 'Right-of-use depreciation, matched by non-cash lease additions (AED mn)',
     IN['rou_dep_path'], NUM0),
    ('nwc_pct', 'Net working capital / revenue (audited FY2025 component days)', NWC_PCT, PCT)])
block('Cost of capital', [
    ('rf', 'Risk-free rate — Jan-2031 AED T-bond (longest liquid AED tenor)', IN['rf'], PCT),
    ('sov', 'UAE sovereign default spread (netted out; Aa2 rating basis)',
     IN['sov_spread_market_observed'], PCT),
    ('erp', 'Equity risk premium (UAE total, rating basis)', IN['erp_market_basis'], PCT),
    ('beta', 'Beta (DU weekly vs FTSE ADX General, 5y)', IN['beta'], '0.000'),
    ('kd', 'Marginal cost of debt (AED sovereign + GCC telecom spread)', IN['kd'], PCT),
    ('rf_path', 'AED risk-free path (defines the glide)', IN['rf_path'], PCT),
    ('rf_term', 'Terminal risk-free rate', IN['rf_term'], PCT),
    ('erp_term', 'Terminal equity risk premium', IN['erp_term'], PCT),
    ('kd_term', 'Terminal cost of debt', IN['kd_term'], PCT),
    ('wd_term', 'Terminal debt weight', IN['wd_term'], PCT),
    ('greal', 'Terminal REAL growth (stated, not derived)', IN['g_term_real'], PCT),
    ('pit', 'Terminal inflation — UAE house macro path', TRI['inflation'], PCT),
    ('life', 'Weighted asset life, DERIVED from notes 6 and 8 (gross cost of the '
     'depreciable owned base over the year\'s own charge)', IN['asset_life_years'], NUM1),
    ('g_term', 'Terminal growth = (1+inflation)(1+real growth) − 1', None, PCT)])
# derived, once, so every lens that needs a nominal terminal growth links to the same cell
putf(ws, f'C{A["g_term"]}', f'=(1+$C${A["pit"]})*(1+$C${A["greal"]})-1',
     TRI['nominal_growth'], PCT)
block('Balance-sheet and bridge anchors', [
    ('lease', 'Lease liabilities at FY2025 (AED mn, audited — the only debt-like item)',
     IN['lease_fy25'], NUM0),
    ('netcash', 'Cash and term deposits at FY2025 (AED mn, audited)', NETCASH, NUM0),
    ('investees', 'Equity-accounted investees at carrying value (AED mn)',
     IN['investees_bv'], NUM1),
    ('payout', 'Forecast dividend payout ratio (FY2024 actual 98%, FY2025 ~100%)', PAYOUT, PCT),
    ('dep_yield', 'Yield on cash and term deposits (audited FY2025 effective)',
     F['dep_yield'], PCT),
    ('lease_rate', 'Lease interest rate (audited FY2025 effective)', F['lease_rate'], PCT),
    ('div_between', 'Dividends gone ex between 31-Dec-2025 and the anchor (AED/share)',
     IN['div_between'], PX),
    ('anchor_days', 'Days from the 31-Dec-2025 valuation date to the 07-Aug-2026 anchor',
     IN['anchor_days'], NUM0)])
block('Lens inputs', [
    ('pe_just', 'Justified price/earnings (GCC telecom peer median)', IN['pe_just'], MULT),
    ('yield_peer', 'Peer benchmark dividend yield', IN['div_yield_peer'], PCT),
    ('roe_sust', 'Sustainable return on equity', IN['roe_sust'], PCT)])
# The four lens-weight rows that stood here went with the blend. They are REMOVED rather
# than zeroed: a weight of zero is still a weight, and a reader opening this sheet would
# reasonably read four zeros as a scheme somebody had switched off rather than one that no
# longer exists. The retired weights are published once, in the Summary's own retired row,
# where they are labelled as retired.

# now that the Assumptions addresses exist, finish the Summary key-figure block
ws = wb['Summary']
rr = SUMMARY_KEY_START
SHARES_ROW = SUMMARY_KEY_START            # the first key figure is the share count
for lab, fml, val, fmt in KEY:
    if fml == 'SHARES':
        fml = f'={a("shares")}'
    elif fml == 'MKTCAP':
        fml = f'=$C${SPOT_ROW}*C{SHARES_ROW}'
    elif fml == 'LEASE':
        fml = f'={a("lease")}'
    elif fml == 'NETCASH':
        fml = f'={a("netcash")}'
    put(ws, f'A{rr}', lab, fmt=None)
    putf(ws, f'C{rr}', fml, val, fmt, green=True)
    rr += 1
ANCH['summary_shares'] = f'C{SHARES_ROW}'
ANCH['summary_mktcap'] = f'C{SHARES_ROW + 1}'

# ============ 5 SOTP BRIDGE ====================================================
ws = sheet('SOTP Bridge')
title(ws, 'Enterprise value to equity — the bridge', 'du has no drawn borrowings and no minority '
      'interests: the bridge is leases out, cash and deposits in', 5, awidth=52, cwidth=16)
hdr(ws, 4, ['Step', 'AED mn', 'Per share (AED)'])
brows = [('Present value of the five forecast years', '=DCF!C27', DCF['pv_explicit']),
         ('Present value of the terminal value', '=DCF!C28', DCF['pv_tv']),
         ('Enterprise value', '=C5+C6', DCF['ev']),
         ('Less lease liabilities (the only debt-like item)', f'=-{a("lease")}', -LEASE),
         ('Plus cash and term deposits', f'={a("netcash")}', NETCASH),
         ('Plus equity-accounted investees at carrying value', f'={a("investees")}', 0.511),
         ('Equity value', '=C7+C8+C9+C10', DCF['eq_val'])]
r = 5
for lab, v, xp in brows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, NUM0, bold=(r in (7, 11)),
         green=v.startswith(('=DCF', '=Assumptions', '=-Assumptions')))
    putf(ws, f'D{r}', f'=C{r}/{a("shares")}', xp / SH, PX, bold=(r in (7, 11)))
    r += 1
band(ws, 11, 4)
r += 1
put(ws, f'A{r}', 'Terminal value as a share of enterprise value', fmt=None)
putf(ws, f'C{r}', '=DCF!C29', DCF['tv_share'], PCT, green=True)
r += 1
put(ws, f'A{r}', 'No minority interests: every subsidiary is wholly owned (the two H1-2026 Cayman '
    'SPVs are 100%-held); non-controlling interests do not appear on the audited balance sheet.',
    fmt=None).font = SUB

# ============ 6 SEGMENTS =======================================================
ws = sheet('Segments')
title(ws, 'Segments — the unit build and the margin stack', 'FY2025 disclosed revenue and '
      'contribution margins; forecast segment revenue is the unit build\'s OUTPUT (pasted, blue) '
      'and everything below it is formula', 9, awidth=38, cwidth=13)
hdr(ws, 4, ['Segment', 'FY2025 revenue', 'Share', 'FY2025 margin'] + YF)
r = 5
REV_TOT = r + len(SEGS)                          # 9
# THE UNIT BUILD IS LIVE IN THE SHEET (changed 17-Aug-2026). Forecast segment revenue was
# previously PASTED as "the unit build's output". It is not an unreadable grid — mobile and fixed
# are two multiplications off drivers that already sit on Assumptions, and wholesale and ICT are a
# growth step. All twenty cells are therefore formulas now, and changing a subscriber count or an
# ARPU reprices the model from the top.
_SEGROW = {}
for s in SEGS:
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    put(ws, f'B{r}', SEG['rev'][s], BLUE, NUM0)
    putf(ws, f'C{r}', f'=B{r}/$B${REV_TOT}', SEG['rev'][s] / HI['FY25']['rev'], PCT)
    put(ws, f'D{r}', SEG['margin'][s], BLUE, PCT)
    _SEGROW[s] = r
    for i in range(5):
        col = get_column_letter(5 + i)
        if s in ('mobile', 'fixed'):
            _sub = 'subs_m' if s == 'mobile' else 'subs_f'
            _arp = 'arpu_m' if s == 'mobile' else 'arpu_f'
            _q2 = 'subs_m_q2' if s == 'mobile' else 'subs_f_q2'
            _h1 = 'h1_mobile' if s == 'mobile' else 'h1_fixed'
            if i == 0:
                # FY2026 = reviewed H1 actual + unit-built H2 on the average of the Jun-2026
                # base and the year-end base
                f_ = (f'={a(_h1)}+({a(_q2)}+{a(_sub, 0)})/2*{a(_arp, 0)}*6/1000')
            else:
                f_ = (f'=({a(_sub, i-1)}+{a(_sub, i)})/2*{a(_arp, i)}'
                      f'*(1+{a("arpu_drift")})^{i}*12/1000')
        else:
            _g = 'g_whl' if s == 'wholesale' else 'g_ict'
            base = f"{a('whl_fy25') if s == 'wholesale' else a('ict_fy25')}" if i == 0 \
                else f'{get_column_letter(4+i)}{r}'
            f_ = f'={base}*(1+{a(_g, i)})'
        putf(ws, f'{col}{r}', f_, F['seg_rev'][s][i], NUM0)
    r += 1
_last = r - 1                                    # 8
band(ws, r, 9); put(ws, f'A{r}', 'Total revenue', bold=True, fmt=None)
putf(ws, f'B{r}', f'=SUM(B5:B{_last})', HI['FY25']['rev'], NUM0, bold=True)
putf(ws, f'C{r}', f'=SUM(C5:C{_last})', 1.0, PCT, bold=True)
for i in range(5):
    col = get_column_letter(5 + i)
    putf(ws, f'{col}{r}', f'=SUM({col}5:{col}{_last})', F['rev'][i], NUM0, bold=True)
r += 2
# ---- THE DIRECT-COST UNIT STACK (installed 17-Aug-2026) ---------------------------
# This block replaces a row of PASTED contribution margins. The margin used to be an input
# held at the audited FY2025 rate; it is now the OUTPUT of a costed unit build, so a change
# to any per-unit cost driver reprices the model from here down.
hdr(ws, r, ['Direct cost per unit — each line grown on its own physical driver'] + YF); r += 1
UROW = {}
for key, label, base_key, esc_key in (
        ('mi', 'Mobile interconnect (AED per subscriber per month)', 'uc_inter_h1', 'esc_inter'),
        ('mc', 'Mobile commission (AED per subscriber per month)', 'uc_comm_h1', 'esc_comm'),
        ('md', 'Mobile devices and direct services (AED per subscriber per month)',
         'uc_dev_h1', 'esc_dev')):
    put(ws, f'A{r}', label, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', f'={a(base_key)}*(1+{a(esc_key)})^{i}',
             F['unit_cost'][{'mi': 'mob_inter', 'mc': 'mob_comm', 'md': 'mob_dev'}[key]][i], NUM2)
    UROW[key] = r; r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Mobile — total direct cost per subscriber per month',
                    bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f"={CD[i]}{UROW['mi']}+{CD[i]}{UROW['mc']}+{CD[i]}{UROW['md']}",
         F['unit_cost']['mob_tot'][i], NUM2, bold=True)
UROW['mt'] = r; r += 1
put(ws, f'A{r}', 'Fixed capacity and direct cost (AED per subscriber per month)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f"={a('uc_fixed_h1')}*(1+{a('esc_fixed')})^{i}",
         F['unit_cost']['fixed_cap'][i], NUM2)
UROW['fc'] = r; r += 1
put(ws, f'A{r}', 'Wholesale direct cost / wholesale revenue (no disclosed unit — flagged)',
    fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f"={a('dc_rate_whl')}", F['unit_cost']['whl_rate'][i], PCT)
UROW['wr'] = r; r += 1
put(ws, f'A{r}', 'ICT direct cost / ICT revenue (no disclosed unit — flagged)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f"={a('dc_rate_ict')}", F['unit_cost']['ict_rate'][i], PCT)
UROW['ir'] = r; r += 2

hdr(ws, r, ['Direct cost (AED mn) — volume x cost per unit'] + YF); r += 1
first_dc = r
DCROW = {}
for j, s in enumerate(SEGS):
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    _rev_row = 5 + j
    for i in range(5):
        rcol = get_column_letter(5 + i)          # revenue table runs E:I
        if s in ('mobile', 'fixed'):
            _sub = 'subs_m' if s == 'mobile' else 'subs_f'
            _q2 = 'subs_m_q2' if s == 'mobile' else 'subs_f_q2'
            _h1dc = 'h1_dc_mobile' if s == 'mobile' else 'h1_dc_fixed'
            _urow = UROW['mt'] if s == 'mobile' else UROW['fc']
            if i == 0:
                f_ = (f'={a(_h1dc)}+({a(_q2)}+{a(_sub, 0)})/2*{CD[0]}{_urow}*6/1000')
            else:
                f_ = f'=({a(_sub, i-1)}+{a(_sub, i)})/2*{CD[i]}{_urow}*12/1000'
        else:
            _h1dc = 'h1_dc_whl' if s == 'wholesale' else 'h1_dc_ict'
            _h1r = 'h1_whl' if s == 'wholesale' else 'h1_ict'
            _urow = UROW['wr'] if s == 'wholesale' else UROW['ir']
            if i == 0:
                f_ = (f'={a(_h1dc)}+({rcol}{_rev_row}-{a(_h1r)})*{CD[0]}{_urow}')
            else:
                f_ = f'={rcol}{_rev_row}*{CD[i]}{_urow}'
        putf(ws, f'{CD[i]}{r}', f_, F['seg_dc'][s][i], NUM0)
    DCROW[s] = r; r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Total direct cost', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'=SUM({CD[i]}{first_dc}:{CD[i]}{r-1})', F['dc_tot'][i], NUM0,
         bold=True)
DC_TOT = r; r += 2

hdr(ws, r, ['Contribution = revenue less direct cost — AN OUTPUT, NOT AN INPUT'] + YF); r += 1
first_c = r
for j, s in enumerate(SEGS):
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', f'={get_column_letter(5+i)}{5+j}-{CD[i]}{DCROW[s]}',
             F['contrib'][s][i], NUM0)
    r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Total contribution', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'=SUM({CD[i]}{first_c}:{CD[i]}{r-1})', F['contrib_tot'][i], NUM0,
         bold=True)
CONTRIB_TOT = r; r += 2

hdr(ws, r, ['Contribution margin — computed, never assumed'] + YF); r += 1
for j, s in enumerate(SEGS):
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}',
             f'={CD[i]}{first_c+j}/{get_column_letter(5+i)}{5+j}',
             F['contrib_margin'][s][i], PCT)
    r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Group gross margin', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{CONTRIB_TOT}/{get_column_letter(5+i)}{REV_TOT}',
         F['gross_margin'][i], PCT, bold=True)
r += 2
hdr(ws, r, ['Operating expense stack — one escalator per cost class'] + YF); r += 1
OL = F['opex_lines']
opex_rows = [
    ('Staff (wage escalator)', lambda i: f'={a("staff26")}*(1+{a("esc_staff")})^{i}',
     OL['staff']),
    ('Network & maintenance (network-scale escalator)',
     lambda i: f'={a("network_fy25")}*(1+{a("esc_network")})^{i+1}', OL['network']),
    ('Marketing (% of revenue)',
     lambda i: f'={get_column_letter(5+i)}{REV_TOT}*{a("marketing_pct")}', OL['marketing']),
    ('Telecom licence and related fees (% of revenue)',
     lambda i: f'={get_column_letter(5+i)}{REV_TOT}*{a("licence_pct")}', OL['licence']),
    ('Administrative (CPI escalator)',
     lambda i: f'={a("admin_fy25")}*(1+{a("esc_admin")})^{i+1}', OL['admin']),
    ('Other operating expense (CPI escalator)',
     lambda i: f'={a("other_fy25")}*(1+{a("esc_other")})^{i+1}', OL['other']),
    ('Expected credit losses (% of revenue)',
     lambda i: f'={get_column_letter(5+i)}{REV_TOT}*{a("ecl_pct")}', OL['ecl']),
    ('Less other operating income', lambda i: f'=-{a("other_inc")}', [-IN['other_inc_path']] * 5),
]
first_o = r                                      # 19
for lab, fml, vals in opex_rows:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fml(i), vals[i], NUM0)
    r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Total operating expenses before D&A', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'=SUM({CD[i]}{first_o}:{CD[i]}{r-1})', F['opex'][i], NUM0, bold=True)
OPEX_TOT = r                                     # 27
r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Group EBITDA (contribution less the opex stack)', bold=True,
                    fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{CONTRIB_TOT}-{CD[i]}{OPEX_TOT}', F['ebitda'][i], NUM0,
         bold=True)
EBITDA_TOT = r                                   # 28
r += 1
put(ws, f'A{r}', 'Group EBITDA margin', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{EBITDA_TOT}/{get_column_letter(5+i)}${REV_TOT}',
         F['ebitda_margin'][i], PCT)
EBITDA_MGN = r
r += 2
hdr(ws, r, ['The unit build behind Mobile and Fixed (triangulation shown, not asserted)'] + YF)
r += 1
UNIT_ROWS = [
    ("Mobile subscribers, end of year ('000)", lambda i: f'={a("subs_m", i)}',
     IN['subs_mobile_path'], NUM0),
    ("Average mobile base ('000) — (prior end + end)/2",
     lambda i: (f'=({BU["subs_mobile"]["Q2_2026"]}+{a("subs_m",0)})/2' if i == 0
                else f'=({a("subs_m", i-1)}+{a("subs_m", i)})/2'),
     [(BU['subs_mobile']['Q2_2026'] + IN['subs_mobile_path'][0]) / 2]
     + [(IN['subs_mobile_path'][i - 1] + IN['subs_mobile_path'][i]) / 2 for i in range(1, 5)],
     NUM0),
    ('Blended mobile ARPU (AED/month)', lambda i: f'={a("arpu_m", i)}',
     IN['arpu_mobile_path'], NUM1),
]
for lab, fml, vals, fmt in UNIT_ROWS:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fml(i), vals[i], fmt)
    r += 1
AVG_ROW = r - 2
put(ws, f'A{r}', 'Unit-implied mobile revenue — average base x ARPU x 12 (FY2026 is H1 actual + '
    'unit-built H2, so the check applies from FY2027)', fmt=None)
unit_mob = [None] + [(IN['subs_mobile_path'][i - 1] + IN['subs_mobile_path'][i]) / 2
                     * IN['arpu_mobile_path'][i] * 12 / 1000 for i in range(1, 5)]
for i in range(5):
    if i == 0:
        put(ws, f'{CD[i]}{r}', '-', BLACK, NUM0)
    else:
        putf(ws, f'{CD[i]}{r}', f'={CD[i]}{AVG_ROW}*{CD[i]}{AVG_ROW+1}*12/1000', unit_mob[i],
             NUM0)
r += 1
put(ws, f'A{r}', 'Pasted mobile segment revenue (the build output above)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={get_column_letter(5+i)}5', F['seg_rev']['mobile'][i], NUM0,
         green=True)
r += 1
put(ws, f'A{r}', 'Check — unit-implied less pasted', fmt=None)
for i in range(5):
    if i == 0:
        put(ws, f'{CD[i]}{r}', '-', BLACK, NUM0)
    else:
        putf(ws, f'{CD[i]}{r}', f'={CD[i]}{r-2}-{CD[i]}{r-1}',
             unit_mob[i] - F['seg_rev']['mobile'][i], NUM1)
r += 1
put(ws, f'A{r}', 'FY2025 reconciliation: average base 9,310k x ARPU 63.3 x 12 = 7,072 vs the '
    'audited mobile segment 7,075 (−0.04%). Wholesale and ICT have no disclosed unit KPIs — '
    'segment-level growth is the finest sourced level, and that gap is flagged.',
    fmt=None).font = SUB
ANCH.update(seg_rev_tot=REV_TOT, seg_contrib_tot=CONTRIB_TOT, seg_ebitda_tot=EBITDA_TOT,
            seg_ebitda_mgn=EBITDA_MGN)

# ============ 7 RELATIVE & NORMALIZED ==========================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative multiples and normalised earnings power', None, 6, awidth=52, cwidth=15)
hdr(ws, 4, ['Relative lens', 'Value'])
rel_rows = [
    ('FY2026E earnings per share (AED)', "='Income Statement'!E19", F['eps'][0], PX),
    ('Justified price / earnings (GCC peer median)', f'={a("pe_just")}', IN['pe_just'], MULT),
    ('Implied value at 31-Dec-2025 (AED)', '=C5*C6', IN['pe_just'] * F['eps'][0], PX),
    ('Anchor accretion factor', '=DCF!C62', DCF['roll'], DF4),
    ('Less final FY2025 dividend paid 28-Apr-2026 (AED)', f'={a("div_between")}',
     IN['div_between'], PX),
    ('', None, None, None),
    ('Implied value per share at the anchor (AED)', '=C7*C8-C9', LN['relative']['base'], PX)]
r = 5
for lab, v, xp, fmt in rel_rows:
    put(ws, f'A{r}', lab, fmt=None)
    if v is not None:
        putf(ws, f'C{r}', v, xp, fmt, green=('DCF' in v or 'Income Statement' in v
                                             or v.startswith('=Assumptions')))
    r += 1
band(ws, 11, 3)
put(ws, 'A12', 'Bear at 12.0x (C) / bull at 18.5x (D), same construction', fmt=None)
putf(ws, 'C12', '=C5*12*C8-C9', LN['relative']['bear'], PX)
putf(ws, 'D12', '=C5*18.5*C8-C9', LN['relative']['bull'], PX)
r = 13
mktcap_f = f'({a("spot")}*{a("shares")})'
for lab, v, xp, fmt in [
        ('Trailing price / earnings', f"={a('spot')}/'Income Statement'!D19",
         REL['pe_trailing'], MULT),
        ('Trailing enterprise value / EBITDA',
         f"=({mktcap_f}+{a('lease')}-{a('netcash')})/'Income Statement'!D9",
         REL['ev_ebitda_trailing'], MULT),
        ('Dividend-yield cross: FY2026E dividend per share (AED)', "='Income Statement'!E20",
         F['dps'][0], PX),
        ('Peer benchmark dividend yield', f'={a("yield_peer")}', IN['div_yield_peer'], PCT),
        ('Implied value from the yield cross (AED)', '=C15/C16', D['yield_ps'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=('Income Statement' in v)); r += 1
r += 1
hdr(ws, r, ['Normalised earnings lens — mid-cycle margin (FY2028E) at current (FY2026E) scale',
            'Value'])
r += 1                                            # r = 20
for lab, v, xp, fmt in [
        ('Current-scale revenue (FY2026E, AED mn)', '=DCF!B5', NRM['rev'], NUM0),
        ('Mid-cycle EBITDA margin (FY2028E)', '=DCF!D7', NRM['margin'], PCT),
        ('Normalised EBITDA (AED mn)', '=C20*C21', NRM['ebitda'], NUM0),
        ('Normalised EBIT (AED mn) — less FY2026E depreciation and amortisation',
         "=C22+'Income Statement'!E11", NRM['ebit'], NUM0),
        ('Net finance income (FY2026E, AED mn)',
         "='Income Statement'!E13+'Income Statement'!E14",
         F['int_inc'][0] - F['int_exp'][0], NUM0),
        ('Normalised net profit (AED mn)', f'=(C23+C24)*(1-{a("tax_eff")})', NRM['np'], NUM0),
        ('Normalised earnings per share (AED)', f'=C25/{a("shares")}', NRM['eps'], PX),
        ('Implied value per share at the anchor (AED)',
         f'=C26*{a("pe_just")}*DCF!$C$62-{a("div_between")}', LN['normalized']['base'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=('DCF' in v or 'Income Statement' in v))
    r += 1
band(ws, r - 1, 3)                                 # implied value lands on row 27
put(ws, 'D27', 'bear 12.0x (E) / bull 18.5x (F):', fmt=None)
putf(ws, 'E27', f'=C26*12*DCF!$C$62-{a("div_between")}', LN['normalized']['bear'], PX)
putf(ws, 'F27', f'=C26*18.5*DCF!$C$62-{a("div_between")}', LN['normalized']['bull'], PX)
r += 1
hdr(ws, r, ['Book lens', 'Value']); r += 1         # r = 30
for lab, v, xp, fmt in [
        ('Book value per share (AED)', f"='Balance Sheet'!D14/{a('shares')}", BK['bvps'], PX),
        ('Sustainable return on equity', f'={a("roe_sust")}', BK['roe_sust'], PCT),
        ('Trailing return on equity',
         "='Income Statement'!D18/(('Balance Sheet'!C14+'Balance Sheet'!D14)/2)",
         BK['roe_trailing'], PCT),
        ('Perpetual (terminal) cost of equity — a steady-state multiple takes a steady-state rate',
         '=DCF!C50', BK['ke_term'], PCT),
        ('Justified price / book', f'=(C31-{a("g_term")})/(C33-{a("g_term")})', BK['pb_just'],
         MULT),
        ('Implied value per share at the anchor (AED)',
         f'=C30*C34*DCF!$C$62-{a("div_between")}', LN['book']['base'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt,
         green=('DCF' in v or 'Balance Sheet' in v or 'Income Statement' in v))
    r += 1
band(ws, r - 1, 3)                                 # implied value lands on row 35
put(ws, 'D35', 'bear / bull constructions (E / F):', fmt=None)
# ONE g in a justified price-to-book, taken from the registered terminal growth; the bear is the
# +100bp stress on the cost of equity, not a second, different growth rate (corrected 17-Aug-2026)
putf(ws, 'E35', f"=(({a('roe_sust')}-{a('g_term')})/((DCF!C40+DCF!C50)/2+0.01-{a('g_term')}))*C30"
     f"*DCF!$C$62-{a('div_between')}", LN['book']['bear'], PX)
putf(ws, 'F35', f"=(({a('roe_sust')}+0.02-{a('g_term')})/(DCF!C50-{a('g_term')}))*C30"
     f"*DCF!$C$62-{a('div_between')}", LN['book']['bull'], PX)

# ============ 8 DCF =============================================================
ws = sheet('DCF')
title(ws, 'Discounted cash flow — the full waterfall', 'Every line is a live formula: the cost of '
      'capital is built below, the glide is derived from the AED risk-free path, and the terminal '
      'value is capitalised at the terminal rate and discounted at the year-5 factor', 6,
      awidth=46, cwidth=15)
hdr(ws, 4, ['AED mn'] + YF)

def wf(r, lab, fmls, vals, fmt=NUM0, bd=False, green=False):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls(i), vals[i], fmt, bold=bd, green=green)
    if bd: band(ws, r, 6)

wf(5, 'Revenue', lambda i: f'=Segments!{get_column_letter(5+i)}{REV_TOT}', F['rev'], green=True)
wf(6, 'EBITDA', lambda i: f'=Segments!{CD[i]}{EBITDA_TOT}', F['ebitda'], green=True)
wf(7, 'EBITDA margin', lambda i: f'={CD[i]}6/{CD[i]}5', F['ebitda_margin'], PCT)
wf(8, 'Less depreciation and amortisation', lambda i: f"='Income Statement'!{FCOL[i]}11",
   [-x for x in F['dna']], green=True)
wf(9, 'EBIT', lambda i: f'={CD[i]}6+{CD[i]}8', F['ebit'], bd=True)
wf(10, f"NOPAT — EBIT x (1 - {TAX:.1%} combined royalty and tax)",
   lambda i: f'={CD[i]}9*(1-{a("tax_eff")})', F['nopat'])
wf(11, 'Add back depreciation and amortisation', lambda i: f'=-{CD[i]}8', F['dna'])
wf(12, 'Less capital expenditure', lambda i: f'=-{CD[i]}5*{a("capex_pct", i)}',
   [-x for x in F['capex']])
# CORRECTED 17-Aug-2026: this line is NO LONGER DEDUCTED. Leases are debt — the liability is
# netted in the bridge and carries a debt weight in the cost of capital, and perpetual renewal is
# paid for in the terminal (invested capital includes the right-of-use asset). Charging it here as
# well billed the same obligation twice. The figure is kept visible as a memo, outside the sum.
wf(13, 'Memo — lease replacement at right-of-use depreciation, NOT deducted (leases are debt)',
   lambda i: f'={a("rou_dep", i)}', DCF['rou_repl_retired'])
wf(14, 'Less change in working capital',
   lambda i: (f"=-({CD[i]}5*{a('nwc_pct')}-'Balance Sheet'!D11)" if i == 0
              else f'=-({CD[i]}5-{CD[i-1]}5)*{a("nwc_pct")}'), [-x for x in F['dnwc']])
wf(15, 'Free cash flow to the firm',
   lambda i: f'={CD[i]}10+{CD[i]}11+{CD[i]}12+{CD[i]}14', F['fcff'], bd=True)
wf(16, 'Forward cost of capital', lambda i: f'=$C$47-($C$47-$C$54)*{CD[i]}58', F['fwd_wacc'], PCT2)
wf(17, 'Discount factor', lambda i: (f'=1/(1+{CD[i]}16)' if i == 0
                                     else f'={CD[i-1]}17/(1+{CD[i]}16)'), F['df'], DF4)
wf(18, 'Present value of FCFF', lambda i: f'={CD[i]}15*{CD[i]}17', F['pv'], bd=True)

put(ws, 'A20', 'TERMINAL VALUE, BRIDGE AND THE ANCHOR ROLL', bold=True, fmt=None)
nopat_grown = F['nopat'][-1] * (1 + TRI['nominal_growth'])
# THE TERMINAL WATERFALL IS BUILT IN A BLOCK OF ITS OWN AT ROWS 70-77 AND SUMMARISED HERE.
# Rows 21-32 keep their positions deliberately: a dozen formulas on five other sheets name
# C24, C26, C27, C28, C29, C30, C32, C40, C50, C54 and C62 by address, and inserting four
# rows into this block to hold the waterfall would move every one of them silently [L-300].
tv_block = [
    ('Terminal-year operating profit after tax, grown one year (AED mn)',
     '=F10*(1+C24)', nopat_grown, NUM0),
    ('Terminal free cash flow — built line by line at rows 70-77 (AED mn)', '=C77',
     TRO['fcff'], NUM0),
    ('Terminal free cash flow as a share of terminal profit', '=C22/C21',
     TRO['fcff'] / TRI['nopat'], PCT),
    ('Terminal growth — (1+inflation)(1+real growth)−1, both at rows 70-71',
     '=(1+C70)*(1+C71)-1', TRI['nominal_growth'], PCT),
    ('Terminal cost of capital', '=C54', W['wacc_term'], PCT),
    ('Terminal value — terminal free cash flow C22 grown one year, capitalised (AED mn)',
     '=C22*(1+C24)/(C25-C24)', DCF['tv'], NUM0),
    ('Present value of the five forecast years (AED mn)', '=SUM(B18:F18)', DCF['pv_explicit'],
     NUM0),
    ('Present value of the terminal value (AED mn)', '=C26*F17', DCF['pv_tv'], NUM0),
    ('Terminal value as a share of enterprise value', '=C28/(C27+C28)', DCF['tv_share'], PCT),
    ('Enterprise value (AED mn)', '=C27+C28', DCF['ev'], NUM0),
    ('Equity value (AED mn)', "='SOTP Bridge'!C11", DCF['eq_val'], NUM0),
    ('Fair value per share at 31-Dec-2025 (AED)', f'=C31/{a("shares")}', DCF['ps_dec'], PX)]
r = 21
for lab, v, xp, fmt in tv_block:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt,
         green=('SOTP' in v or 'Summary Financials' in v or v.startswith('=Assumptions')))
    r += 1
band(ws, 32, 4)

put(ws, 'A34', 'COST OF CAPITAL — BUILT HERE, NOT ASSUMED', bold=True, fmt=None)
coc = [
    ('Risk-free rate — Jan-2031 AED T-bond (longest liquid AED tenor)', f'={a("rf")}',
     IN['rf'], PCT2),
    ('Less UAE sovereign default spread (removed to avoid double-counting)', f'={a("sov")}',
     IN['sov_spread_market_observed'], PCT2),
    ('Risk-free rate net of the sovereign spread', '=C35-C36', W['rf_star'], PCT2),
    ('Beta (DU weekly vs the FTSE ADX General, 5 years)', f'={a("beta")}',
     IN['beta'], '0.000'),
    ('Equity risk premium (UAE total, rating basis)', f'={a("erp")}', IN['erp_market_basis'], PCT2),
    ('Cost of equity, explicit window', '=C37+C38*C39', W['ke_exp'], PCT2),
    ('Marginal cost of debt (AED sovereign + GCC telecom spread)', f'={a("kd")}', IN['kd'],
     PCT2),
    ('Cost of debt after tax (interest shields both fiscal legs)',
     f'=C41*(1-{a("tax_eff")})', W['kd_at'], PCT2),
    ('Market capitalisation (AED mn)', f'={a("spot")}*{a("shares")}', SPOT * SH, NUM0),
    ('Lease liabilities — the only debt-like item (AED mn)', f'={a("lease")}', LEASE, NUM0),
    ('Debt weight (leases / (leases + market capitalisation))', '=C44/(C44+C43)',
     W['wd_exp'], PCT2),
    ('Equity weight', '=1-C45', W['we_exp'], PCT2),
    ('Cost of capital, explicit window', '=C46*C40+C45*C42', W['wacc_exp'], PCT2),
    ('Terminal risk-free rate', f'={a("rf_term")}', IN['rf_term'], PCT2),
    ('Terminal risk-free net of the sovereign spread', '=C48-C36',
     IN['rf_term'] - IN['sov_spread_market_observed'], PCT2),
    ('Terminal cost of equity', f'=C49+C38*{a("erp_term")}', W['ke_term'], PCT2),
    ('Terminal cost of debt', f'={a("kd_term")}', IN['kd_term'], PCT2),
    ('Terminal cost of debt after tax', f'=C51*(1-{a("tax_eff")})', W['kd_term_at'], PCT2),
    ('Terminal debt weight', f'={a("wd_term")}', IN['wd_term'], PCT2),
    ('Terminal cost of capital', '=(1-C53)*C50+C53*C52', W['wacc_term'], PCT2)]
r = 35
for lab, v, xp, fmt in coc:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=v.startswith('=Assumptions'))
    r += 1
band(ws, 47, 4); band(ws, 54, 4)
hdr(ws, 56, ['The glide — inherited from the AED risk-free path, not invented'] + YF)
put(ws, 'A57', 'AED risk-free path', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}57', f'={a("rf_path", i)}', IN['rf_path'][i], PCT2, green=True)
put(ws, 'A58', 'Glide fraction — cumulative progress of the easing', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}58', f'=($B$57-{CD[i]}57)/($B$57-$F$57)', F['glide_frac'][i], PCT2)
put(ws, 'A59', 'Note: row 16 above is the explicit-window cost of capital walked down to the '
    'terminal rate by the glide fraction on row 58. du carries no debt whose cost could define '
    'the glide, so the cost-of-money path itself (the AED risk-free path) does — the shape is '
    'inherited, not a second free parameter.', fmt=None).font = SUB
put(ws, 'A61', 'THE ANCHOR ROLL — one date, one price of time', bold=True, fmt=None)
put(ws, 'A62', 'Anchor accretion factor — (1 + cost of equity)^(days to anchor / 365)', fmt=None)
putf(ws, 'C62', f'=(1+C40)^({a("anchor_days")}/365)', DCF['roll'], DF4)
put(ws, 'A63', 'Fair value per share at the 07-Aug-2026 anchor (AED)', fmt=None)
putf(ws, 'C63', f'=C32*C62-{a("div_between")}', DCF['ps'], PX, bold=True)
band(ws, 63, 4)
put(ws, 'A66', 'THE TERMINAL, LINE BY LINE — capital maintained at replacement cost over '
    'the asset life the depreciation notes themselves imply', bold=True, fmt=None)
put(ws, 'A67', 'The retired construction grew terminal profit and deducted a reinvestment '
    'rate set by the growth rate over the return on capital, which is arithmetically the '
    'same as rebuilding the whole capital base every 1/g years — a fact about the dirham\'s '
    'peg to the dollar rather than about a mobile network. The asset life below is DERIVED '
    'from notes 6 and 8: the gross cost of the depreciable owned base over the year\'s own '
    'charge.', fmt=None).font = SUB
_tw = [('Terminal inflation — UAE house macro path', f'={a("pit")}', TRI['inflation'], PCT),
       ('Terminal REAL growth (stated)', f'={a("greal")}', IN['g_term_real'], PCT),
       ('Weighted asset life, derived from notes 6 and 8 (years)', f'={a("life")}',
        IN['asset_life_years'], NUM1),
       ('Terminal operating profit after tax (from row 21)', '=C21', TRI['nopat'], NUM0),
       ('Plus owned depreciation and amortisation, grown one year — the right-of-use '
        'charge is neither added back nor charged, which is a lease renewed at its own '
        'current cost', f'={TRI["dna_book"] / (1 + TRI["nominal_growth"]):.6f}*(1+C24)',
        TRI['dna_book'], NUM0),
       ('Less capital maintenance at replacement cost — that charge escalated over half '
        'the derived life', '=-C74*(1+C70)^(C72/2)', -TRO['maintenance'], NUM0),
       ('Less capital for real growth, and less inflation on working capital (a CREDIT '
        'here: this company collects before it pays)',
        f'=-C71*{TRI["incremental_capital_per_unit_growth"]:.6f}'
        f'-C70*{TRI["working_capital"]:.6f}',
        -TRO['growth_capex'] - TRO['wc_charge'], NUM0),
       ('Terminal free cash flow', '=SUM(C73:C76)', TRO['fcff'], NUM0)]
_r = 70
for lab, fml, xp, fmt in _tw:
    put(ws, f'A{_r}', lab, fmt=None)
    putf(ws, f'C{_r}', fml, xp, fmt, bold=(_r == 77),
         green=fml.startswith('=Assumptions'))
    _r += 1
put(ws, 'A78', 'Memo — the no-growth perpetuity at book depreciation, a diagnostic and not '
    'a bound', fmt=None)
putf(ws, 'C78', '=C21/C25', TRO['floor'], NUM0)

put(ws, 'A64', 'The bridge on row 32 is dated 31-Dec-2025 (the audited balance-sheet date it '
    'nets leases and cash at). Row 63 rolls it to the anchor at the cost of equity, net of the '
    'AED 0.40 final FY2025 dividend paid 28-Apr-2026. Every lens on every sheet is rolled the '
    'same way. The H1-2026 interim dividend of AED 0.26 (declared 23-Jul-2026, unpaid at the '
    'anchor) stays in the share.', fmt=None).font = SUB

# ============ 9 INCOME STATEMENT =================================================
ws = sheet('Income Statement')
title(ws, 'Income statement — three years historical, five years forecast', 'AED mn, '
      'consolidated. History is the audited record; every forecast line is a formula', 9,
      awidth=40, cwidth=12)
hdr(ws, 4, ['AED mn'] + YH + YF)

def isline(r, lab, hist, fc_f, fc_v, fmt=NUM0, bd=False, hist_f=None, hist_v=None, green=False):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i in range(3):
        if hist_f is not None:
            putf(ws, f'{HC[i]}{r}', hist_f(i), hist_v[i], fmt, bold=bd)
        elif hist is not None:
            put(ws, f'{HC[i]}{r}', hist[i], BLUE, fmt, bold=bd)
    if fc_f is not None:
        for i in range(5):
            putf(ws, f'{FCOL[i]}{r}', fc_f(i), fc_v[i], fmt, bold=bd, green=green)
    if bd: band(ws, r, 9)

# every figure below is READ from the committed numbers file — no financial numeral is typed
# into this builder (numeric-traceability rule; the prior edition violated it in 30 places)
DIRECT = IN['direct_costs_hist']
OPEX_H = IN['opex_before_dna_hist']
CONTRIB_H = {y: HI[y]['rev'] + DIRECT[y] for y in ('FY24', 'FY25')}
isline(5, 'Revenue', [HI[y]['rev'] for y in H3], lambda i: f'=DCF!{CD[i]}5', F['rev'], bd=True,
       green=True)
put(ws, 'A6', 'Direct costs (interconnect, commissions, devices)', fmt=None)
put(ws, 'B6', '-', BLACK, NUM0)
put(ws, 'C6', DIRECT['FY24'], BLUE, NUM0); put(ws, 'D6', DIRECT['FY25'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}6', f'=-(DCF!{CD[i]}5-Segments!{CD[i]}{CONTRIB_TOT})',
         -(F['rev'][i] - F['contrib_tot'][i]), NUM0, green=True)
put(ws, 'A7', 'Segment contribution', fmt=None)
put(ws, 'B7', '-', BLACK, NUM0)
for i, y in [(1, 'FY24'), (2, 'FY25')]:
    putf(ws, f'{HC[i]}7', f'={HC[i]}5+{HC[i]}6', CONTRIB_H[y], NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}7', f'={FCOL[i]}5+{FCOL[i]}6', F['contrib_tot'][i], NUM0)
put(ws, 'A8', 'Operating expenses before D&A', fmt=None)
put(ws, 'B8', '-', BLACK, NUM0)
put(ws, 'C8', OPEX_H['FY24'], BLUE, NUM0); put(ws, 'D8', OPEX_H['FY25'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}8', f'=-Segments!{CD[i]}{OPEX_TOT}', -F['opex'][i], NUM0, green=True)
put(ws, 'A9', 'EBITDA', bold=True, fmt=None)
put(ws, 'B9', HI['FY23']['ebitda'], BLUE, NUM0, bold=True)   # derived pre-IFRS-18, flagged
for i in (1, 2):
    putf(ws, f'{HC[i]}9', f'={HC[i]}7+{HC[i]}8', HI[H3[i]]['ebitda'], NUM0, bold=True)
for i in range(5):
    putf(ws, f'{FCOL[i]}9', f'=DCF!{CD[i]}6', F['ebitda'][i], NUM0, bold=True, green=True)
band(ws, 9, 9)
isline(10, 'EBITDA margin', None, lambda i: f'={FCOL[i]}9/{FCOL[i]}5', F['ebitda_margin'], PCT,
       hist_f=lambda i: f'={HC[i]}9/{HC[i]}5',
       hist_v=[HI[y]['ebitda'] / HI[y]['rev'] for y in H3])
isline(11, 'Depreciation and amortisation', [-abs(HI[y]['dna']) for y in H3],
       lambda i: (f"=-('Balance Sheet'!{FCOL[i]}6+'Balance Sheet'!{FCOL[i]}9"
                  f"+{a('rou_dep', i)})"), [-x for x in F['dna']])
isline(12, 'Operating profit (EBIT)', None, lambda i: f'={FCOL[i]}9+{FCOL[i]}11', F['ebit'],
       bd=True, hist_f=lambda i: f'={HC[i]}9+{HC[i]}11', hist_v=[HI[y]['ebit'] for y in H3])
isline(13, 'Interest income', [IN['int_inc_fy23'], IN['int_inc_fy24'], IN['int_inc_fy25']],
       lambda i: (f"={a('dep_yield')}*MAX('Balance Sheet'!{'D' if i == 0 else FCOL[i-1]}12,0)"),
       F['int_inc'])
isline(14, 'Interest expense (principally leases)',
       [-IN['int_exp_fy23'], -IN['int_exp_fy24'], -IN['int_exp_fy25']],
       lambda i: f"=-{a('lease_rate')}*'Balance Sheet'!{FCOL[i]}13", [-x for x in F['int_exp']])
isline(15, 'Share of equity-accounted investments',
       [IN['assoc_hist'][y] for y in H3], None, None)
put(ws, 'A16', 'Profit before federal royalty and income tax', bold=True, fmt=None)
for i, v in enumerate([HI[y]['pbt'] for y in H3]):
    put(ws, f'{HC[i]}16', v, BLUE, NUM0, bold=True)   # printed line (face); small net-impairment
for i in range(5):                                     # items keep it from being a 3-line sum
    putf(ws, f'{FCOL[i]}16', f'={FCOL[i]}12+{FCOL[i]}13+{FCOL[i]}14',
         F['ebit'][i] + F['int_inc'][i] - F['int_exp'][i], NUM0, bold=True)
band(ws, 16, 9)
isline(17, 'Federal royalty and income tax',
       [-(HI[y]['royalty'] + HI[y]['tax']) for y in H3],
       lambda i: f'=-{FCOL[i]}16*{a("tax_eff")}',
       [-(F['ebit'][i] + F['int_inc'][i] - F['int_exp'][i]) * TAX for i in range(5)])
isline(18, 'Net profit', None, lambda i: f'={FCOL[i]}16+{FCOL[i]}17', F['np'], bd=True,
       hist_f=lambda i: f'={HC[i]}16+{HC[i]}17', hist_v=[HI[y]['np'] for y in H3])
np_all = [HI[y]['np'] for y in H3] + F['np']
put(ws, 'A19', 'Earnings per share (AED)', fmt=None)
for i in range(8):
    putf(ws, f'{ALL[i]}19', f'={ALL[i]}18/{a("shares")}', np_all[i] / SH, PX)
put(ws, 'A20', 'Dividend per share (AED)', fmt=None)
for i, v in enumerate([IN['dps_fy23'], IN['dps_fy24'], IN['dps_fy25']]):
    put(ws, f'{HC[i]}20', v, BLUE, PX)
for i in range(5):
    putf(ws, f'{FCOL[i]}20', f'={FCOL[i]}19*{a("payout")}', F['dps'][i], PX)
put(ws, 'A22', 'Notes: FY2024-25 are the audited IFRS 18 face (EBITDA is printed); FY2023 '
    'predates IFRS 18 — its by-nature split is not shown ("-") and its EBITDA is derived from '
    'audited components and flagged. The historical profit-before-royalty line is the printed '
    'figure (small net-impairment items keep it from being a three-line sum; the forecast has '
    'none). FY2023 royalty is the old-regime charge (15% of regulated revenue + 30% of '
    'regulated profit); no corporate income tax existed before 2024.', fmt=None).font = SUB

# ============ 10 BALANCE SHEET ====================================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — condensed', 'AED mn, consolidated. Every FY2023-25 line is the '
      'audited closing figure. du carries ZERO drawn borrowings in every year shown', 9,
      awidth=40, cwidth=12)
hdr(ws, 4, ['AED mn'] + YH + YF)

DEP_H = [IN['dep_ppe_hist'][y] for y in H3]
AMO_H = [IN['amort_hist'][y] for y in H3]
put(ws, 'A5', 'Property, plant and equipment', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}5', HB[y]['ppe'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}5',
         f'={"D" if i == 0 else FCOL[i-1]}5-DCF!{CD[i]}12*{a("tang_share")}-{FCOL[i]}6',
         F['ppe'][i], NUM0)
put(ws, 'A6', '  of which depreciation charge (memo)', fmt=None)
for i in range(3):
    put(ws, f'{HC[i]}6', DEP_H[i], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}6', f'={a("dep_rate")}*{"D" if i == 0 else FCOL[i-1]}5',
         F['dep_ppe'][i], NUM0)
put(ws, 'A7', 'Right-of-use assets', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}7', HB[y]['rou'], BLUE, NUM0)
for i in range(5):
    # additions equal depreciation BY CONSTRUCTION (a new lease creates asset and liability
    # together and is non-cash), so the book is held flat — stated, not dressed as a driver link
    putf(ws, f'{FCOL[i]}7', f'=$D$7', F['rou'][i], NUM0)
put(ws, 'A8', 'Intangible assets (excl. goodwill)', fmt=None)
put(ws, 'B8', HB['FY23']['intang'], BLUE, NUM0)
put(ws, 'C8', HB['FY24']['intang'], BLUE, NUM0); put(ws, 'D8', HB['FY25']['intang'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}8',
         f'={"D" if i == 0 else FCOL[i-1]}8-DCF!{CD[i]}12*(1-{a("tang_share")})-{FCOL[i]}9',
         F['intang'][i], NUM0)
put(ws, 'A9', '  of which amortisation charge (memo)', fmt=None)
for i in range(3):
    put(ws, f'{HC[i]}9', AMO_H[i], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}9', f'={a("amort_rate")}*{"D" if i == 0 else FCOL[i-1]}8',
         F['amort'][i], NUM0)
put(ws, 'A10', 'Goodwill', fmt=None)
put(ws, 'B10', '-', BLACK, NUM0)
put(ws, 'C10', HB['FY24']['goodwill'], BLUE, NUM0); put(ws, 'D10', HB['FY25']['goodwill'], BLUE,
                                                        NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}10', '=$D$10', HB['FY25']['goodwill'], NUM0)
put(ws, 'A11', 'Net working capital (royalty accrual excluded)', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}11', HB[y]['nwc'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}11', f"='Income Statement'!{FCOL[i]}5*{a('nwc_pct')}", F['nwc'][i], NUM0)
put(ws, 'A12', 'Cash and term deposits', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}12', HB[y]['net_cash'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}12',
         f"={'D' if i == 0 else FCOL[i-1]}12+'Income Statement'!{FCOL[i]}18"
         f"-'Income Statement'!{FCOL[i]}11+DCF!{CD[i]}12+DCF!{CD[i]}14"
         f"-'Income Statement'!{FCOL[i]}18*{a('payout')}",
         F['net_cash'][i], NUM0)
put(ws, 'A13', 'Lease liabilities (the only debt-like item)', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}13', HB[y]['lease'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{FCOL[i]}13', '=$D$13', LEASE, NUM0)
put(ws, 'A14', 'Total equity', bold=True, fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{HC[i]}14', HB[y]['eq'], BLUE, NUM0, bold=True)
for i in range(5):
    putf(ws, f'{FCOL[i]}14',
         f"={'D' if i == 0 else FCOL[i-1]}14+'Income Statement'!{FCOL[i]}18*(1-{a('payout')})",
         F['equity'][i], NUM0, bold=True)
band(ws, 14, 9)
put(ws, 'A15', 'Net cash after lease liabilities', fmt=None)
nc_after = ([HB[y]['net_cash'] - HB[y]['lease'] for y in H3]
            + [F['net_cash'][i] - LEASE for i in range(5)])
for i in range(8):
    putf(ws, f'{ALL[i]}15', f'={ALL[i]}12-{ALL[i]}13', nc_after[i], NUM0)
put(ws, 'A16', 'Lease liabilities / EBITDA', fmt=None)
eb_all = [HI[y]['ebitda'] for y in H3] + F['ebitda']
lease_all = [HB[y]['lease'] for y in H3] + [LEASE] * 5
for i in range(8):
    putf(ws, f'{ALL[i]}16', f"={ALL[i]}13/'Income Statement'!{ALL[i]}9",
         lease_all[i] / eb_all[i], MULT)
put(ws, 'A18', 'Notes: a CONDENSED layout — receivables, payables, contract balances and the '
    'royalty accrual are netted inside the working-capital line, so the sheet does not foot to '
    'total assets. The FY2023 payables line still contained the royalty accrual (AED 2,033mn, '
    'separately disclosed from FY2024); it is excluded here in every year so the series is '
    'like-for-like. The lease book is held flat (replacement = depreciation); the FY2025 '
    'goodwill balance is carried unchanged. Cash and term deposits roll forward from profit '
    'less the ~98% payout — the H1-2026 interims (term deposits nil at 30-Jun-2026 after the '
    'AED 4.1bn royalty-plus-dividend outflow) show exactly this mechanic mid-year.',
    fmt=None).font = SUB

# ============ 11 CASH FLOW =========================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow — historical markers and the forecast waterfall', 'AED mn', 9,
      awidth=44, cwidth=12)
hdr(ws, 4, ['AED mn', 'FY2024', 'FY2025'] + YF)
put(ws, 'A5', 'EBITDA', fmt=None)
putf(ws, 'B5', "='Income Statement'!C9", HI['FY24']['ebitda'], NUM0, green=True)
putf(ws, 'C5', "='Income Statement'!D9", HI['FY25']['ebitda'], NUM0, green=True)
for i in range(5):
    putf(ws, f'{CFF[i]}5', f"='Income Statement'!{FCOL[i]}9", F['ebitda'][i], NUM0, green=True)
for r_, lab, v24, v25 in [
        (6, 'Federal royalty and income tax paid',
         IN['tax_paid_hist']['FY24'], IN['tax_paid_hist']['FY25']),
        (7, 'Net cash generated from operating activities',
         IN['ocf_hist']['FY24'], IN['ocf_hist']['FY25']),
        (8, 'Purchase of PP&E and intangibles',
         -IN['capex_cash_hist']['FY24'], -IN['capex_cash_hist']['FY25'])]:
    put(ws, f'A{r_}', lab, fmt=None)
    put(ws, f'B{r_}', v24, BLUE, NUM0)
    put(ws, f'C{r_}', v25, BLUE, NUM0)
put(ws, 'A9', 'NOPAT', fmt=None)
put(ws, 'B9', '-', BLACK, NUM0); put(ws, 'C9', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}9', f'=DCF!{CD[i]}10', F['nopat'][i], NUM0, green=True)
put(ws, 'A10', 'Add back depreciation and amortisation', fmt=None)
put(ws, 'B10', '-', BLACK, NUM0); put(ws, 'C10', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}10', f'=DCF!{CD[i]}11', F['dna'][i], NUM0, green=True)
put(ws, 'A11', 'Capital expenditure', fmt=None)
put(ws, 'B11', '-', BLACK, NUM0); put(ws, 'C11', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}11', f'=DCF!{CD[i]}12', -F['capex'][i], NUM0, green=True)
put(ws, 'A12', 'Memo — lease replacement, NOT deducted (leases are debt)', fmt=None)
put(ws, 'B12', '-', BLACK, NUM0); put(ws, 'C12', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}12', f'=DCF!{CD[i]}13', DCF['rou_repl_retired'][i], NUM0, green=True)
put(ws, 'A13', 'Change in working capital', fmt=None)
put(ws, 'B13', '-', BLACK, NUM0)
putf(ws, 'C13', "=-('Balance Sheet'!D11-'Balance Sheet'!C11)",
     -(HB['FY25']['nwc'] - HB['FY24']['nwc']), NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}13', f'=DCF!{CD[i]}14', -F['dnwc'][i], NUM0, green=True)
put(ws, 'A14', 'Free cash flow to the firm', bold=True, fmt=None)
put(ws, 'B14', '-', BLACK, NUM0); put(ws, 'C14', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}14', f'={CFF[i]}9+{CFF[i]}10+{CFF[i]}11+{CFF[i]}13',
         F['fcff'][i], NUM0, bold=True)
band(ws, 14, 9)
put(ws, 'A16', 'Dividends paid', fmt=None)
put(ws, 'B16', IN['div_paid_hist']['FY24'], BLUE, NUM0)
put(ws, 'C16', IN['div_paid_hist']['FY25'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}16', f"=-'Income Statement'!{FCOL[i]}18*{a('payout')}",
         -F['div'][i], NUM0)
put(ws, 'A17', 'Cash conversion — operating cash flow as a share of EBITDA, FY2025', fmt=None)
putf(ws, 'C17', '=C7/C5', IN['ocf_hist']['FY25'] / HI['FY25']['ebitda'], PCT)
put(ws, 'A18', 'du converts EBITDA to operating cash at over 70% AFTER the 43.6% fiscal take — '
    'the negative-working-capital model means growth releases cash. The constraint on equity '
    'cash flow is the ~100% payout plus the data-centre capex ramp, which is why term deposits '
    'fell to nil at 30-Jun-2026.', fmt=None).font = SUB

# ============ 12 SUMMARY FINANCIALS =================================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials — the eight-year picture', 'AED mn unless stated. Every cell on '
      'this sheet is a link or a ratio; nothing is typed twice', 9, awidth=40, cwidth=12)
hdr(ws, 4, ['AED mn'] + YH + YF)
rev_all = [HI[y]['rev'] for y in H3] + F['rev']
ebit_all = [HI[y]['ebit'] for y in H3] + F['ebit']

def sfline(r, lab, fml, vals, fmt=NUM0, skip=()):
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{r}', '-', BLACK, fmt)
        else:
            f_ = fml(i)
            putf(ws, f'{ALL[i]}{r}', f_, vals[i], fmt,
                 green=f_.startswith(("='I", "='B", '=DCF', "='C")))

IC_H = [None,
        HB['FY24']['ppe'] + HB['FY24']['rou'] + HB['FY24']['intang'] + HB['FY24']['goodwill']
        + HB['FY24']['nwc'],
        HB['FY25']['ppe'] + HB['FY25']['rou'] + HB['FY25']['intang'] + HB['FY25']['goodwill']
        + HB['FY25']['nwc']]
sfline(5, 'Revenue', lambda i: f"='Income Statement'!{ALL[i]}5", rev_all)
sfline(6, 'Revenue growth', lambda i: f'={ALL[i]}5/{ALL[i-1]}5-1',
       [None] + [rev_all[i] / rev_all[i - 1] - 1 for i in range(1, 8)], PCT, skip=(0,))
sfline(7, 'EBITDA', lambda i: f"='Income Statement'!{ALL[i]}9", eb_all)
sfline(8, 'EBITDA margin', lambda i: f'={ALL[i]}7/{ALL[i]}5',
       [eb_all[i] / rev_all[i] for i in range(8)], PCT)
sfline(9, 'EBIT', lambda i: f"='Income Statement'!{ALL[i]}12", ebit_all)
sfline(10, 'Net profit', lambda i: f"='Income Statement'!{ALL[i]}18", np_all)
sfline(11, 'Free cash flow to the firm', lambda i: f"='Cash Flow'!{CFF[i-3]}14",
       [None] * 3 + F['fcff'], skip=(0, 1, 2))
sfline(12, 'Net cash after lease liabilities', lambda i: f"='Balance Sheet'!{ALL[i]}15",
       nc_after)
sfline(13, 'Invested capital',
       lambda i: (f"='Balance Sheet'!{ALL[i]}5+'Balance Sheet'!{ALL[i]}7"
                  f"+'Balance Sheet'!{ALL[i]}8+'Balance Sheet'!{ALL[i]}10"
                  f"+'Balance Sheet'!{ALL[i]}11"),
       IC_H + F['ic'], skip=(0,))
sfline(14, 'Return on invested capital (average capital)',
       lambda i: f'=DCF!{CD[i-3]}10/(({ALL[i-1]}13+{ALL[i]}13)/2)',
       [None] * 3 + F['roic'], PCT, skip=(0, 1, 2))
put(ws, 'A16', 'Invested capital = PP&E + right-of-use assets + intangibles + goodwill + net '
    'working capital (negative for du). FY2023 is not shown because goodwill was not disclosed '
    'as a separate line that year.', fmt=None).font = SUB

# ============ 13 MONTE CARLO ==========================================================
ws = sheet('Monte Carlo')
title(ws, 'Probabilistic price map', 'A map of price dispersion. It carries no view on value and '
      'is never blended with the valuation. Each figure is an engine output, not a formula.', 8,
      awidth=40, cwidth=14)
hdr(ws, 4, ['Horizon', '5th', '25th', 'Median', '75th', '95th', 'P(above spot)'])
r = 5
for tag in ('1M', '3M'):
    h = STK['horizons'][tag]
    put(ws, f'A{r}', f"{'One month' if tag=='1M' else 'Three months'} — to {h['grade_date']}",
        fmt=None)
    for i, k in enumerate(('p5', 'p25', 'p50', 'p75', 'p95')):
        put(ws, f'{get_column_letter(2+i)}{r}', h['pct'][k], BLUE, PX)
    put(ws, f'G{r}', h['p_above'], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Level event', 'One month', 'Three months']); r += 1
for lab, k in [('Finishes 10% or more above spot', 'p_up10'),
               ('Finishes 10% or more below spot', 'p_dn10'),
               ('Touches 10% above spot at any point', 'touch_up10'),
               ('Touches 10% below spot at any point', 'touch_dn10')]:
    put(ws, f'A{r}', lab, fmt=None)
    put(ws, f'B{r}', STK['horizons']['1M'][k], BLUE, PCT)
    put(ws, f'C{r}', STK['horizons']['3M'][k], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Engine setting', 'Value']); r += 1
for lab, v, fmt, green in [('Simulated paths', 50000, NUM0, False),
                           ('Annualised volatility (3-month anchor)',
                            STK['horizons']['3M']['anchor_vol_ann'], PCT, False),
                           ('Spot price (AED)', f'=Summary!C{SPOT_ROW}', PX, True),
                           ('Anchor date', STK['anchor_date'], None, False)]:
    put(ws, f'A{r}', lab, fmt=None)
    if green:
        putf(ws, f'C{r}', v, SPOT, fmt, green=True)
    else:
        put(ws, f'C{r}', v, BLUE, fmt)
    r += 1

# ============ 14 SENSITIVITY ============================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — what the valuation needs the world to do', 'AED per share. Each cell is '
      'a complete re-run of the model, including the unit build, so these grids are engine '
      'outputs rather than formulas and do NOT redraw when a driver is changed.', 8,
      awidth=40, cwidth=13)
r = 4
put(ws, f'A{r}', 'Terminal cost of capital (rows) x terminal growth (columns)', bold=True,
    fmt=None)
r += 1
hdr(ws, r, [''] + [f'{g:.1%}' for g in SN['g_grid']]); r += 1
for i, wt in enumerate(SN['wt_grid']):
    put(ws, f'A{r}', f'{wt:.2%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid_wacc_g'][i][j], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Explicit-window cost of capital (columns) x terminal cost of capital (rows)',
    bold=True, fmt=None); r += 1
hdr(ws, r, [''] + [f'{x:.2%}' for x in SN['we_grid']]); r += 1
for i, wt in enumerate(SN['wt_grid']):
    put(ws, f'A{r}', f'terminal {wt:.2%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid_exp_term'][j][i], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Single-driver sensitivities — five engine re-runs per row; the parameter grid '
    'for each row is shown beside its name', bold=True, fmt=None); r += 1
hdr(ws, r, ['Driver (parameter grid)', '', '', '', '', '', '', 'Swing']); r += 1
for lab, grid, vals, gfmt in [
        ('Beta (regression CI ends, then priors)', SN['beta_grid'], SN['grid_beta'], '{:.2f}'),
        ('Combined fiscal take (top of grid = the pre-2024 construction\'s FY2023 take)',
         SN['tax_grid'], SN['grid_tax'], '{:.1%}'),
        ('Blended ARPU multiplier — applied to BOTH the mobile ARPU and the implied fixed '
         'revenue-per-subscriber paths', SN['arpu_grid'], SN['grid_arpu'], '{:.2f}x'),
        ("Subscriber path shift ('000) — applied in full to the mobile base and at 8% of that "
         'to the fixed base', SN['subs_grid'], SN['grid_subs'], '{:+.0f}'),
        ('Direct cost per unit, multiplicative', SN['mg_grid'], SN['grid_margin'], '{:.2f}x'),
        ('Blended ARPU drift per year (mix exhaustion)', SN['drift_grid'],
         SN['grid_drift'], '{:+.1%}'),
        ('Capex path multiplier', SN['capex_grid'], SN['grid_capex'], '{:.3f}x'),
        ('Working capital / revenue', SN['nwc_grid'], SN['grid_nwc'], '{:.1%}'),
        ('Asset life, years — derived from the depreciation notes',
         SN['life_grid'], SN['grid_life'], '{:.1f}'),
        ('Terminal growth (at base cost of capital)', SN['g_grid'], SN['grid_wacc_g'][2],
         '{:.1%}')]:
    put(ws, f'A{r}', f"{lab}  ({' / '.join(gfmt.format(g) for g in grid)})", fmt=None)
    vv = vals[:6]
    for j, v in enumerate(vv):
        put(ws, f'{get_column_letter(2+j)}{r}', v, BLUE, PX)
    last = get_column_letter(1 + len(vv))
    putf(ws, f'H{r}', f'=MAX(B{r}:{last}{r})-MIN(B{r}:{last}{r})', max(vv) - min(vv), PX)
    r += 1
ws.column_dimensions['H'].width = 13
ws.column_dimensions['A'].width = 56

# ============ 15 PER-SHARE & RATIOS =========================================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share and ratio analysis', 'The indicator set for an integrated telecom operator: '
      'ARPU, subscribers, margin, capital intensity, cash conversion and the dividend. Every '
      'ratio is a formula off the statements.', 9, awidth=44, cwidth=12)
hdr(ws, 4, ['Measure'] + YH + YF)
r = 5

def ratio(lab, fml, vals, fmt, skip=()):
    global r
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALL[i]}{r}', fml(i), vals[i], fmt)
    r += 1

eq_all = [HB[y]['eq'] for y in H3] + F['equity']
dps_all = [IN['dps_fy23'], IN['dps_fy24'], IN['dps_fy25']] + F['dps']
ratio('Earnings per share (AED)', lambda i: f"='Income Statement'!{ALL[i]}19",
      [x / SH for x in np_all], PX)
ratio('Dividend per share (AED)', lambda i: f"='Income Statement'!{ALL[i]}20", dps_all, PX)
ratio('Payout ratio', lambda i: f"='Income Statement'!{ALL[i]}20/'Income Statement'!{ALL[i]}19",
      [dps_all[i] / (np_all[i] / SH) for i in range(8)], PCT)
ratio('Dividend yield at the anchor spot', lambda i: f"='Income Statement'!{ALL[i]}20/{a('spot')}",
      [d / SPOT for d in dps_all], PCT)
ratio('Book value per share (AED)', lambda i: f"='Balance Sheet'!{ALL[i]}14/{a('shares')}",
      [x / SH for x in eq_all], PX)
ratio('Free cash flow to the firm per share (AED)',
      lambda i: f"='Summary Financials'!{ALL[i]}11/{a('shares')}",
      [None] * 3 + [x / SH for x in F['fcff']], PX, skip=(0, 1, 2))
ratio('EBITDA margin', lambda i: f"='Summary Financials'!{ALL[i]}8",
      [eb_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('EBIT margin', lambda i: f"='Income Statement'!{ALL[i]}12/'Income Statement'!{ALL[i]}5",
      [ebit_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Net margin', lambda i: f"='Income Statement'!{ALL[i]}18/'Income Statement'!{ALL[i]}5",
      [np_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Return on equity',
      lambda i: (f"='Income Statement'!{ALL[i]}18/(('Balance Sheet'!{ALL[i-1]}14+"
                 f"'Balance Sheet'!{ALL[i]}14)/2)"),
      [None] + [np_all[i] / ((eq_all[i - 1] + eq_all[i]) / 2) for i in range(1, 8)], PCT,
      skip=(0,))
ratio('Return on invested capital', lambda i: f"='Summary Financials'!{ALL[i]}14",
      [None] * 3 + F['roic'], PCT, skip=(0, 1, 2))
capex_h = [IN['capex_cash_hist'][y] for y in H3]
ratio('Capital expenditure / revenue (capital intensity)',
      lambda i: (f"=-'Cash Flow'!{CFF[i-3]}11/'Income Statement'!{ALL[i]}5" if i >= 3 else None),
      [None] * 3 + [F['capex'][i] / F['rev'][i] for i in range(5)], PCT, skip=(0, 1, 2))
ratio('Lease liabilities / EBITDA', lambda i: f"='Balance Sheet'!{ALL[i]}16",
      [lease_all[i] / eb_all[i] for i in range(8)], MULT)
ratio('Working capital / revenue',
      lambda i: f"='Balance Sheet'!{ALL[i]}11/'Income Statement'!{ALL[i]}5",
      [HB[y]['nwc'] / HI[y]['rev'] for y in H3] + [F['nwc'][i] / F['rev'][i] for i in range(5)],
      PCT)
r += 1
put(ws, f'A{r}', 'Operating KPIs (company-disclosed history, house forecast)', bold=True,
    fmt=None); r += 1
SMH, SFH, AH = (IN['subs_mobile_hist_display'], IN['subs_fixed_hist_display'],
                IN['arpu_hist_display'])
kpi_rows = [
    ("Mobile subscribers, end of year ('000)", [None, SMH['FY24'], SMH['FY25']],
     IN['subs_mobile_path'], NUM0, 'subs_m'),
    ("Fixed subscribers, end of year ('000)", [None, SFH['FY24'], SFH['FY25']],
     IN['subs_fixed_path'], NUM0, 'subs_f'),
    ('Blended mobile ARPU (AED/month)', [None, AH['FY24'], AH['FY25']],
     IN['arpu_mobile_path'], NUM1, 'arpu_m'),
]
for lab, hist, path, fmt, key in kpi_rows:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(3):
        v = hist[i]
        put(ws, f'{HC[i]}{r}', v if v is not None else '-', BLUE if v is not None else BLACK,
            fmt)
    for i in range(5):
        putf(ws, f'{FCOL[i]}{r}', f'={a(key, i)}', path[i], fmt, green=True)
    r += 1
put(ws, f'A{r+1}', 'Subscriber and ARPU history are the company\'s own year-end prints; FY2023 is '
    'not shown because the KPI series in the current disclosure pack starts at Q4-2024. The '
    'Q2-2026 actual mobile base is '
    f"{BU['subs_mobile']['Q2_2026']:,.0f}k after the war quarter (total base −"
    f"{BU['subs_mobile']['Q1_2026']-BU['subs_mobile']['Q2_2026']:,.0f}k on the quarter), which is "
    'the base the forecast recovers from.', fmt=None).font = SUB

# ============ 16 PEER & SECTOR ===============================================================
ws = sheet('Peer & Sector')
title(ws, 'Peer frame and sector context', 'Aggregator multiple reads, labelled cross-check — '
      'never a build source. du computed from its own audited figures', 6, awidth=34, cwidth=22)
hdr(ws, 4, ['Company', 'Market', 'Trailing P/E', 'Dividend yield', 'Relevance / caution'])
r = 5
for nm, mkt, pe, dy, note in [
    ('e& (EAND)', 'UAE (ADX)', '~20.7x', '~4.8%',
     'the other half of the duopoly; scale, international assets and a higher multiple'),
    ('stc', 'Saudi Arabia', '~18.9x', '~5.2%',
     'regional benchmark payer; the yield anchor for the cross-check'),
    ('Mobily', 'Saudi Arabia', '~15.5x', '~2.9%',
     'closest structural analogue — the #2 operator that closed the gap; the justified P/E'),
    ('Ooredoo', 'Qatar', '~12.5x', '~4.6%', 'multi-market; softer growth, lower multiple'),
    ('Zain', 'Kuwait', '~9x', '~6-7%', 'levered multi-market; the low end of the bracket'),
    ('Omantel', 'Oman', '~11.4x', '~6.7%', 'small-market incumbent; yield-heavy'),
    ('Deutsche Telekom / Vodafone', 'Developed markets', 'n/m', '3.4% / n/m',
     'the developed-market bracket: slower growth, more leverage, lower returns'),
]:
    put(ws, f'A{r}', nm, fmt=None); put(ws, f'B{r}', mkt, fmt=None)
    put(ws, f'C{r}', pe, BLUE, None); put(ws, f'D{r}', dy, BLUE, None)
    put(ws, f'E{r}', note, fmt=None, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1
ws.column_dimensions['E'].width = 52
r += 1
hdr(ws, r, ['du own multiples (computed live)', 'Value']); r += 1
for lab, v, xp, fmt in [
        ('Trailing price / earnings', "='Relative & Normalized'!C13", REL['pe_trailing'], MULT),
        ('Trailing enterprise value / EBITDA', "='Relative & Normalized'!C14",
         REL['ev_ebitda_trailing'], MULT),
        ('Trailing dividend yield (FY2025 DPS / spot)', f"='Income Statement'!D20/{a('spot')}",
         IN['dps_fy25'] / SPOT, PCT),
        ('Justified price / earnings applied', "='Relative & Normalized'!C6", IN['pe_just'],
         MULT)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', v, xp, fmt, green=True); r += 1
put(ws, f'A{r+1}', 'Peer EV/EBITDA was not reliably sourceable from public aggregators at the '
    'sweep date (net-debt figures missing); the relative lens therefore runs on P/E and '
    'dividend yield, both sourced, and that limitation is stated rather than papered over.',
    fmt=None).font = SUB

# print layout: wide sheets paginate label-from-value in portrait; landscape+fit-to-width
# keeps every row's label and values on one page in the rendered PDF
for _ws in wb.worksheets:
    _ws.page_setup.orientation = 'landscape'
    _ws.page_setup.fitToWidth = 1
    _ws.page_setup.fitToHeight = 0
    _ws.sheet_properties.pageSetUpPr.fitToPage = True
# post-pass: Summary's contested-judgement alternative links LIVE to the captured FV cell
_ws = wb['Summary']
putf(_ws, f'C{SUMMARY_ALT_ROW}', f"='Fundamental Valuation'!{ANCH['fv_framing2']}",
     DCF['ps_mkt_term'], PX, green=True)

out = os.path.join(HERE, 'DU_Valuation_Model_09082026_public.xlsx')
wb.save(out)
json.dump({'expected': EXPECT, 'anchors': ANCH},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nchk = sum(len(v) for v in EXPECT.values())
nform = nlit = 0
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
            elif isinstance(c.value, (int, float)):
                nlit += 1
print(f"wrote {out} | {len(wb.sheetnames)} sheets: {wb.sheetnames}")
print(f"formulas: {nform} (of which {nchk} carry a checked expected value) | numeric literals: {nlit}")
