"""RIYADHCABLE_Valuation_Model_18082026_public.xlsx — 16 sheets, formula-driven.
Blue = inputs · black = formulas · green = cross-sheet links.

Every quantity arithmetically derivable from a driver is a live Excel formula, so the
reader can change a blue cell on Assumptions and watch the model reprice. Only three
classes of cell are pasted values, and READ FIRST names them:
  1. audited/disclosed historical figures (the primary record) and the FY2025 disclosed
     base the ground-up build starts from;
  2. no flattened unit-build output is pasted — the cost stack is built live on Segments;
  3. whole-model re-runs: the Monte Carlo price map, the sensitivity grids and the DCF
     scenario bear/bull bounds, each a complete revaluation.

Every formula cell records the model's own value into xlsx_expected.json; recalc.py
evaluates the workbook independently and asserts agreement; driver_test.py perturbs each
Assumptions driver and asserts the headline moves in the right direction.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
REL, NRM, BKL, EXPP = D['rel'], D['norm'], D['book'], D['experts']
SEG, S0, STK, BT, TR = D['seg_fy25'], D['step0'], D['strike'], D['backtest'], D['terminal_recon']
UE = D['unit_econ']; SF = D['seg_fcst']
IN = {k: v['value'] for k, v in D['inputs'].items()}
SPOT, SH, TAX = M['spot'], M['shares_mn'], IN['tax_eff']
MKTCAP = SPOT * SH
NCI_SH = DCF['nci_share']
PAYOUT = F['payout']

# ---- segment arrays exactly as compute.py builds them (read, not recomputed) --------
# Cables & wires leg (the metal converter):
VOL0 = UE['vol0']; MATPU0 = UE['mat_pu0']; CONVPU0 = UE['conv_pu0']
vol, matpu, convpu = SF['vol'], SF['mat_pu'], SF['conv_pu']
cab_mat, cab_conv, cab_gp, cab_rev = SF['cab_mat'], SF['cab_conv'], SF['cab_gp'], SF['cab_rev']
gppu = SF['gp_pu']
gmt = [IN['spread_anchor'] + IN['margin_glide'][i] for i in range(5)]
# HV turnkey and Other legs:
hv_rev, hv_gp, oth_rev, oth_gp = SF['hv_rev'], SF['hv_gp'], SF['oth_rev'], SF['oth_gp']
CAB_MAT25 = UE['cab_mat25']; CAB_CONV25 = UE['cab_conv25']; CAB_REV25 = UE['cab_rev25']
CAB_GP25 = UE['cab_gp25']; HV_REV25 = UE['hv_rev25']; OTH_REV25 = UE['oth_rev25']
# Each leg is built on its OWN driver; the group is the sum and its gross margin is the blended
# OUTPUT. The cable leg is the metal converter: gross profit = volume x conversion-spread-per-unit,
# revenue = materials + conversion + gross profit; HV/Other grow at their own rate and margin.
BLUE = Font(color='0000FF'); GREEN = Font(color='1F6F3C'); BLACK = Font(color='1A1A1A')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='55625E')
FILL_T = PatternFill('solid', start_color='14322E'); FILL_H = PatternFill('solid', start_color='E8EFEC')
FILL_G = PatternFill('solid', start_color='F5F0E4')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
YH = ['FY2023', 'FY2024', 'FY2025']
YF = F['years']
FCUR = ['C', 'D', 'E', 'F', 'G']    # forecast columns on Segments/DCF blocks (FY2025 base in B)
HCOL = ['C', 'D', 'E']              # historical columns on the statements
FCOL = ['F', 'G', 'H', 'I', 'J']   # forecast columns on the statements

wb = Workbook()
EXPECT = {}
ANCH = {}


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=44, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


# =========================================================================
# ASSUMPTIONS — build first so every sheet can reference driver cells
# =========================================================================
wsa = sheet('Assumptions')
title(wsa, 'Assumptions — every blue cell is a driver', 'Change one and the model reprices', 10,
      awidth=54, cwidth=11)
AR = {}
_row = [4]


def a_scalar(label, value, fmt=PCT2):
    r = _row[0]
    put(wsa, f'A{r}', label, fmt=None)
    put(wsa, f'C{r}', value, BLUE, fmt)
    AR[label] = r; _row[0] += 1
    return r


def a_path(label, values, fmt=PCT2):
    r = _row[0]
    put(wsa, f'A{r}', label, fmt=None)
    for i, val in enumerate(values):
        put(wsa, f'{FCUR[i]}{r}', val, BLUE, fmt)
    AR[label] = r; _row[0] += 1
    return r


def a_section(label):
    band(wsa, _row[0], 10)
    put(wsa, f'A{_row[0]}', label, bold=True, fmt=None)
    _row[0] += 1


def AC(label):
    return f"Assumptions!$C${AR[label]}"


def AP(label, i):
    return f"Assumptions!{FCUR[i]}${AR[label]}"


a_section('Anchors & market')
a_scalar('Spot price (SAR)', SPOT, PX)
a_scalar('Shares outstanding (mn)', SH, NUM1)
a_scalar('Net financial debt at FY2025 (SAR mn, disclosed)', DCF['nd'], NUM0)
a_section('Cost of capital — explicit window')
a_scalar('Risk-free rate (10-year SAR sukuk)', IN['rf'])
a_scalar('Sovereign default spread (netted out)', IN['sov_spread'])
a_scalar('Equity risk premium (rating basis)', IN['erp'])
a_scalar('Beta (own-stock vs TASI)', IN['beta'], '0.000')
a_scalar('Cost of debt, marginal pre-tax', IN['kd'])
a_scalar('Effective zakat and income tax rate', TAX)
a_section('Cost of capital — terminal')
a_scalar('Terminal risk-free rate', IN['rf_term'])
a_scalar('Terminal equity risk premium', IN['erp_term'])
a_scalar('Terminal beta', IN['beta_term'], '0.000')
a_scalar('Terminal cost of debt', IN['kd_term'])
a_scalar('Terminal net-debt weight', IN['wd_term'])
a_scalar('Terminal growth', IN['g_term'])
a_section('Forecast drivers — Cables & wires leg (metal converter)')
a_path('Cable volume index growth', IN['vol_growth'])
a_path('Metal content price growth', IN['metal_growth'])
a_path('Conversion cost inflation', IN['conv_infl'])
a_scalar('Sustained gross margin (H1-2026 anchor)', IN['spread_anchor'])
a_path('Gross-margin glide (added to anchor)', IN['margin_glide'], '0.000')
a_section('Forecast drivers — HV turnkey & Other legs')
a_path('HV turnkey revenue growth', IN['hv_growth'])
a_scalar('HV turnkey segment margin', SF['hv_margin'])
a_path('Other segment revenue growth', IN['other_growth'])
a_scalar('Other segment margin', SF['oth_margin'])
a_section('Forecast drivers — group')
a_path('Operating expenses / revenue', IN['opex_pct'])
a_scalar('Depreciation and amortisation / revenue', IN['dna_pct'])
a_path('Capital expenditure / revenue', IN['capex_pct'])
a_scalar('Net working capital / revenue', IN['nwc_pct'])
a_path('Cost-of-debt path', IN['kd_path'])
a_scalar('Forecast dividend payout ratio', PAYOUT)
a_scalar('Yield on surplus cash', 0.04)
a_scalar('NCI share of forecast profit', NCI_SH)
a_section('Lens inputs')
a_scalar('Justified EV/EBITDA', IN['ev_ebitda_just'], MULT)
a_scalar('Justified price/earnings', IN['pe_just'], MULT)
a_scalar('Sustainable return on equity', IN['roe_sust'])
# The four lens-weight rows that stood here went with the blend. Removed rather than
# zeroed: a weight of zero is still a weight, and four zeros read as a scheme somebody
# switched off rather than one that no longer exists.
a_section('Anchor roll')
a_scalar('Days to the anchor', IN['anchor_days'], NUM0)
a_scalar('FY2025 dividend per share paid in window', IN['div_window'], PX)
a_section('FY2025 disclosed base (audited)')
a_scalar('FY2025 cables metal content (SAR mn)', CAB_MAT25, NUM0)
a_scalar('FY2025 cables conversion cost (SAR mn)', CAB_CONV25, NUM0)
a_scalar('FY2025 HV turnkey revenue (SAR mn)', HV_REV25, NUM0)
a_scalar('FY2025 Other segment revenue (SAR mn)', OTH_REV25, NUM0)
a_scalar('FY2025 net working capital (SAR mn)', F['nwc_fy25'], NUM0)
a_scalar('FY2025 PP&E (SAR mn)', F['ppe_fy25'], NUM0)
a_scalar('FY2025 gross borrowings incl. leases (SAR mn)', F['debt_fy25'], NUM0)
a_scalar('FY2025 equity attributable (SAR mn)', F['eqp_fy25'], NUM0)
a_scalar('FY2025 associates carrying value (SAR mn)', DCF['assoc'], NUM0)
a_scalar('FY2025 non-operating assets (SAR mn)', DCF['nonop'], NUM0)
a_scalar('FY2025 NCI carrying value (SAR mn)', DCF['nci'], NUM0)

# =========================================================================
# SEGMENTS — three disclosed legs, each on its own driver, summed to the group
# =========================================================================
wsg = sheet('Segments')
title(wsg, 'Segments — three legs, built to the group', 'Each disclosed Note 40 segment on its own '
      'driver; the cable leg is the metal converter; group gross margin is the blended OUTPUT', 8,
      awidth=48, cwidth=12)
hdr(wsg, 4, ['SAR mn / index', 'FY2025', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E'])
r = {}
_sr = [5]


def seg_head(name):
    band(wsg, _sr[0], 8); put(wsg, f'A{_sr[0]}', name, bold=True, fmt=None); _sr[0] += 1


def seg_row(key, name):
    r[key] = _sr[0]; put(wsg, f'A{_sr[0]}', name, fmt=None); _sr[0] += 1


seg_head('Cables & wires — the metal converter')
for k, nm in [('cvol', 'Cable volume index (FY2025=100)'), ('cmatpu', 'Metal content per unit'),
              ('cconvpu', 'Conversion cost per unit'), ('cgmt', 'Target gross margin (anchor + glide)'),
              ('cgppu', 'Gross profit per unit (conversion spread)'), ('cmat', 'Materials (metal leg)'),
              ('cconv', 'Conversion cost'), ('cgp', 'Gross profit = volume x spread'),
              ('crev', 'Cables revenue = materials + conversion + gross profit'),
              ('cgm', 'Cables gross margin (OUTPUT)')]:
    seg_row(k, nm)
seg_head('HV turnkey projects — own growth, disclosed margin')
for k, nm in [('hvrev', 'HV revenue (own growth path)'), ('hvmar', 'HV segment margin'),
              ('hvgp', 'HV gross profit = revenue x margin')]:
    seg_row(k, nm)
seg_head('Other (telephone cables & services) — own growth, disclosed margin')
for k, nm in [('orev', 'Other revenue (own growth path)'), ('omar', 'Other segment margin'),
              ('ogp', 'Other gross profit = revenue x margin')]:
    seg_row(k, nm)
seg_head('Group — the sum of the three legs')
for k, nm in [('rev', 'Revenue (group)'), ('gp', 'Gross profit (group)'), ('cogs', 'Cost of revenue (group)'),
              ('gm', 'Gross margin (OUTPUT = group gross profit / group revenue)'),
              ('opex', 'Operating expenses'), ('dna', 'Depreciation & amortisation'), ('ebit', 'EBIT'),
              ('ebitda', 'EBITDA'), ('ebmar', 'EBITDA margin')]:
    seg_row(k, nm)

# ---- Cables leg: FY2025 base (B) then the forecast ----
put(wsg, f'B{r["cvol"]}', VOL0, BLUE, NUM1)
putf(wsg, f'B{r["cmatpu"]}', f'={AC("FY2025 cables metal content (SAR mn)")}/{VOL0}', MATPU0, NUM2)
putf(wsg, f'B{r["cconvpu"]}', f'={AC("FY2025 cables conversion cost (SAR mn)")}/{VOL0}', CONVPU0, NUM2)
put(wsg, f'B{r["cgp"]}', CAB_GP25, BLUE, NUM0)  # disclosed FY2025 cables gross profit (Note 40)
putf(wsg, f'B{r["cgppu"]}', f'=B{r["cgp"]}/B{r["cvol"]}', CAB_GP25 / VOL0, NUM2)
putf(wsg, f'B{r["cmat"]}', f'=B{r["cvol"]}*B{r["cmatpu"]}', CAB_MAT25, NUM0)
putf(wsg, f'B{r["cconv"]}', f'=B{r["cvol"]}*B{r["cconvpu"]}', CAB_CONV25, NUM0)
putf(wsg, f'B{r["crev"]}', f'=B{r["cmat"]}+B{r["cconv"]}+B{r["cgp"]}', CAB_REV25, NUM0)
putf(wsg, f'B{r["cgm"]}', f'=B{r["cgp"]}/B{r["crev"]}', CAB_GP25 / CAB_REV25, PCT2)
for i in range(5):
    c = FCUR[i]; pc = 'B' if i == 0 else FCUR[i - 1]
    putf(wsg, f'{c}{r["cvol"]}', f'={pc}{r["cvol"]}*(1+{AP("Cable volume index growth", i)})', vol[i], NUM1)
    putf(wsg, f'{c}{r["cmatpu"]}', f'={pc}{r["cmatpu"]}*(1+{AP("Metal content price growth", i)})', matpu[i], NUM2)
    putf(wsg, f'{c}{r["cconvpu"]}', f'={pc}{r["cconvpu"]}*(1+{AP("Conversion cost inflation", i)})', convpu[i], NUM2)
    putf(wsg, f'{c}{r["cgmt"]}', f'={AC("Sustained gross margin (H1-2026 anchor)")}+{AP("Gross-margin glide (added to anchor)", i)}', gmt[i], PCT2)
    putf(wsg, f'{c}{r["cgppu"]}', f'={c}{r["cgmt"]}/(1-{c}{r["cgmt"]})*({c}{r["cmatpu"]}+{c}{r["cconvpu"]})', gppu[i], NUM2)
    putf(wsg, f'{c}{r["cmat"]}', f'={c}{r["cvol"]}*{c}{r["cmatpu"]}', cab_mat[i], NUM0)
    putf(wsg, f'{c}{r["cconv"]}', f'={c}{r["cvol"]}*{c}{r["cconvpu"]}', cab_conv[i], NUM0)
    putf(wsg, f'{c}{r["cgp"]}', f'={c}{r["cvol"]}*{c}{r["cgppu"]}', cab_gp[i], NUM0)
    putf(wsg, f'{c}{r["crev"]}', f'={c}{r["cmat"]}+{c}{r["cconv"]}+{c}{r["cgp"]}', cab_rev[i], NUM0)
    putf(wsg, f'{c}{r["cgm"]}', f'={c}{r["cgp"]}/{c}{r["crev"]}', cab_gp[i] / cab_rev[i], PCT2)
# ---- HV leg ----
putf(wsg, f'B{r["hvrev"]}', f'={AC("FY2025 HV turnkey revenue (SAR mn)")}', HV_REV25, NUM0)
putf(wsg, f'B{r["hvmar"]}', f'={AC("HV turnkey segment margin")}', SF['hv_margin'], PCT2)
putf(wsg, f'B{r["hvgp"]}', f'=B{r["hvrev"]}*B{r["hvmar"]}', HV_REV25 * SF['hv_margin'], NUM0)
for i in range(5):
    c = FCUR[i]; pc = 'B' if i == 0 else FCUR[i - 1]
    putf(wsg, f'{c}{r["hvrev"]}', f'={pc}{r["hvrev"]}*(1+{AP("HV turnkey revenue growth", i)})', hv_rev[i], NUM0)
    putf(wsg, f'{c}{r["hvmar"]}', f'={AC("HV turnkey segment margin")}', SF['hv_margin'], PCT2)
    putf(wsg, f'{c}{r["hvgp"]}', f'={c}{r["hvrev"]}*{c}{r["hvmar"]}', hv_gp[i], NUM0)
# ---- Other leg ----
putf(wsg, f'B{r["orev"]}', f'={AC("FY2025 Other segment revenue (SAR mn)")}', OTH_REV25, NUM0)
putf(wsg, f'B{r["omar"]}', f'={AC("Other segment margin")}', SF['oth_margin'], PCT2)
putf(wsg, f'B{r["ogp"]}', f'=B{r["orev"]}*B{r["omar"]}', OTH_REV25 * SF['oth_margin'], NUM0)
for i in range(5):
    c = FCUR[i]; pc = 'B' if i == 0 else FCUR[i - 1]
    putf(wsg, f'{c}{r["orev"]}', f'={pc}{r["orev"]}*(1+{AP("Other segment revenue growth", i)})', oth_rev[i], NUM0)
    putf(wsg, f'{c}{r["omar"]}', f'={AC("Other segment margin")}', SF['oth_margin'], PCT2)
    putf(wsg, f'{c}{r["ogp"]}', f'={c}{r["orev"]}*{c}{r["omar"]}', oth_gp[i], NUM0)
# ---- Group = sum of the legs ----
putf(wsg, f'B{r["rev"]}', f'=B{r["crev"]}+B{r["hvrev"]}+B{r["orev"]}', HI['FY25']['rev'], NUM0)
putf(wsg, f'B{r["gp"]}', f'=B{r["cgp"]}+B{r["hvgp"]}+B{r["ogp"]}', HI['FY25']['gp'], NUM0)
putf(wsg, f'B{r["cogs"]}', f'=B{r["rev"]}-B{r["gp"]}', HI['FY25']['rev'] - HI['FY25']['gp'], NUM0)
putf(wsg, f'B{r["gm"]}', f'=B{r["gp"]}/B{r["rev"]}', HI['FY25']['gp'] / HI['FY25']['rev'], PCT2)
for i in range(5):
    c = FCUR[i]
    putf(wsg, f'{c}{r["rev"]}', f'={c}{r["crev"]}+{c}{r["hvrev"]}+{c}{r["orev"]}', F['rev'][i], NUM0)
    putf(wsg, f'{c}{r["gp"]}', f'={c}{r["cgp"]}+{c}{r["hvgp"]}+{c}{r["ogp"]}', F['gp'][i], NUM0)
    putf(wsg, f'{c}{r["cogs"]}', f'={c}{r["rev"]}-{c}{r["gp"]}', F['rev'][i] - F['gp'][i], NUM0)
    putf(wsg, f'{c}{r["gm"]}', f'={c}{r["gp"]}/{c}{r["rev"]}', F['gm'][i], PCT2)
    putf(wsg, f'{c}{r["opex"]}', f'={AP("Operating expenses / revenue", i)}*{c}{r["rev"]}', F['opex'][i], NUM0)
    putf(wsg, f'{c}{r["dna"]}', f'={AC("Depreciation and amortisation / revenue")}*{c}{r["rev"]}', F['dna'][i], NUM0)
    putf(wsg, f'{c}{r["ebit"]}', f'={c}{r["gp"]}-{c}{r["opex"]}', F['ebit'][i], NUM0)
    putf(wsg, f'{c}{r["ebitda"]}', f'={c}{r["ebit"]}+{c}{r["dna"]}', F['ebitda'][i], NUM0)
    putf(wsg, f'{c}{r["ebmar"]}', f'={c}{r["ebitda"]}/{c}{r["rev"]}', F['ebitda_margin'][i], PCT2)
ANCH['seg'] = r

# =========================================================================
# DCF — cost of capital, glide, FCFF waterfall, terminal, bridge
# =========================================================================
wsd = sheet('DCF')
title(wsd, 'Discounted cash flow', 'Cost of capital built from drivers; glide from the cost-of-debt '
      'path; terminal ROIC-consistent', 8, awidth=46, cwidth=13)
# --- cost of capital block ---
rr = 4
put(wsd, f'A{rr}', 'Cost of capital', bold=True, fmt=None); band(wsd, rr, 8); rr += 1
row_rfstar = rr; putf(wsd, f'C{rr}', f'={AC("Risk-free rate (10-year SAR sukuk)")}-{AC("Sovereign default spread (netted out)")}',
                      W['rf_star'], PCT2); put(wsd, f'A{rr}', 'Normalised risk-free rate rf*', fmt=None); rr += 1
row_ke = rr; putf(wsd, f'C{rr}', f'=C{row_rfstar}+{AC("Beta (own-stock vs TASI)")}*{AC("Equity risk premium (rating basis)")}',
                  W['ke'], PCT2); put(wsd, f'A{rr}', 'Cost of equity Ke', fmt=None); rr += 1
row_kdat = rr; putf(wsd, f'C{rr}', f'={AC("Cost of debt, marginal pre-tax")}*(1-{AC("Effective zakat and income tax rate")})',
                    W['kd_at'], PCT2); put(wsd, f'A{rr}', 'Cost of debt after tax', fmt=None); rr += 1
row_mktcap = rr; putf(wsd, f'C{rr}', f'={AC("Spot price (SAR)")}*{AC("Shares outstanding (mn)")}', MKTCAP, NUM0)
put(wsd, f'A{rr}', 'Market capitalisation', fmt=None); rr += 1
row_wd = rr; putf(wsd, f'C{rr}', f'={AC("Net financial debt at FY2025 (SAR mn, disclosed)")}/({AC("Net financial debt at FY2025 (SAR mn, disclosed)")}+C{row_mktcap})',
                  W['wd'], PCT2); put(wsd, f'A{rr}', 'Net-debt weight', fmt=None); rr += 1
row_wacc = rr; putf(wsd, f'C{rr}', f'=(1-C{row_wd})*C{row_ke}+C{row_wd}*C{row_kdat}', W['wacc'], PCT2, bold=True)
put(wsd, f'A{rr}', 'WACC — explicit window', bold=True, fmt=None); ANCH['wacc'] = rr; rr += 1
row_keterm = rr; putf(wsd, f'C{rr}', f'={AC("Terminal risk-free rate")}+{AC("Terminal beta")}*{AC("Terminal equity risk premium")}',
                      W['ke_term'], PCT2); put(wsd, f'A{rr}', 'Terminal cost of equity', fmt=None); rr += 1
row_kdtat = rr; putf(wsd, f'C{rr}', f'={AC("Terminal cost of debt")}*(1-{AC("Effective zakat and income tax rate")})',
                     W['kd_term_at'], PCT2); put(wsd, f'A{rr}', 'Terminal cost of debt after tax', fmt=None); rr += 1
row_waccterm = rr; putf(wsd, f'C{rr}', f'=(1-{AC("Terminal net-debt weight")})*C{row_keterm}+{AC("Terminal net-debt weight")}*C{row_kdtat}',
                        W['wacc_term'], PCT2, bold=True); put(wsd, f'A{rr}', 'WACC — terminal', bold=True, fmt=None)
ANCH['wacc_term'] = rr; rr += 1
# --- glide + discount factors ---
rr += 1; put(wsd, f'A{rr}', 'Discount-rate glide & waterfall', bold=True, fmt=None); band(wsd, rr, 8); rr += 1
hdr(wsd, rr, ['SAR mn', '', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']); rr += 1
row_glide = rr; put(wsd, f'A{rr}', 'Glide fraction (from Kd path)', fmt=None)
gf = F['glide_frac']
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'=({AP("Cost-of-debt path", 0)}-{AP("Cost-of-debt path", i)})/({AP("Cost-of-debt path", 0)}-{AP("Cost-of-debt path", 4)})',
         gf[i], DF4)
rr += 1
row_fwd = rr; put(wsd, f'A{rr}', 'Forward WACC', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'=C{row_wacc}-(C{row_wacc}-C{row_waccterm})*{FCUR[i]}{row_glide}', F['fwd_wacc'][i], PCT2)
rr += 1
row_df = rr; put(wsd, f'A{rr}', 'Discount factor', fmt=None)
for i in range(5):
    if i == 0:
        putf(wsd, f'{FCUR[i]}{rr}', f'=1/(1+{FCUR[i]}{row_fwd})', F['df'][i], DF4)
    else:
        putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i-1]}{rr}/(1+{FCUR[i]}{row_fwd})', F['df'][i], DF4)
rr += 1
# --- FCFF waterfall (links to Segments) ---
row_ebit = rr; put(wsd, f'A{rr}', 'EBIT', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f"=Segments!{FCUR[i]}{ANCH['seg']['ebit']}", F['ebit'][i], NUM0, green=True)
rr += 1
row_nopat = rr; put(wsd, f'A{rr}', 'NOPAT = EBIT x (1 - tax)', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_ebit}*(1-{AC("Effective zakat and income tax rate")})', F['nopat'][i], NUM0)
rr += 1
row_dna = rr; put(wsd, f'A{rr}', '+ Depreciation & amortisation', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f"=Segments!{FCUR[i]}{ANCH['seg']['dna']}", F['dna'][i], NUM0, green=True)
rr += 1
row_rev = rr; put(wsd, f'A{rr}', 'Revenue (memo)', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f"=Segments!{FCUR[i]}{ANCH['seg']['rev']}", F['rev'][i], NUM0, green=True)
rr += 1
row_capex = rr; put(wsd, f'A{rr}', '- Capital expenditure', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={AP("Capital expenditure / revenue", i)}*{FCUR[i]}{row_rev}', F['capex'][i], NUM0)
rr += 1
row_nwc = rr; put(wsd, f'A{rr}', 'Net working capital', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={AC("Net working capital / revenue")}*{FCUR[i]}{row_rev}', F['nwc'][i], NUM0)
rr += 1
row_dnwc = rr; put(wsd, f'A{rr}', '- Change in working capital', fmt=None)
for i in range(5):
    if i == 0:
        putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_nwc}-{AC("FY2025 net working capital (SAR mn)")}', F['dnwc'][i], NUM0)
    else:
        putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_nwc}-{FCUR[i-1]}{row_nwc}', F['dnwc'][i], NUM0)
rr += 1
row_fcff = rr; put(wsd, f'A{rr}', 'Free cash flow to firm', bold=True, fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_nopat}+{FCUR[i]}{row_dna}-{FCUR[i]}{row_capex}-{FCUR[i]}{row_dnwc}', F['fcff'][i], NUM0, bold=True)
rr += 1
row_pv = rr; put(wsd, f'A{rr}', 'PV of FCFF', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_fcff}*{FCUR[i]}{row_df}', F['pv'][i], NUM0)
rr += 1
# --- PP&E, invested capital, ROIC, terminal ---
rr += 1; put(wsd, f'A{rr}', 'Invested capital & terminal', bold=True, fmt=None); band(wsd, rr, 8); rr += 1
row_ppe = rr; put(wsd, f'A{rr}', 'PP&E roll-forward', fmt=None)
for i in range(5):
    if i == 0:
        putf(wsd, f'{FCUR[i]}{rr}', f'={AC("FY2025 PP&E (SAR mn)")}+{FCUR[i]}{row_capex}-{FCUR[i]}{row_dna}', F['ppe'][i], NUM0)
    else:
        putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i-1]}{rr}+{FCUR[i]}{row_capex}-{FCUR[i]}{row_dna}', F['ppe'][i], NUM0)
rr += 1
row_ic = rr; put(wsd, f'A{rr}', 'Invested capital = NWC + PP&E', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_nwc}+{FCUR[i]}{row_ppe}', F['ic'][i], NUM0)
rr += 1
row_roic = rr; put(wsd, f'A{rr}', 'ROIC = NOPAT / invested capital', fmt=None)
for i in range(5):
    putf(wsd, f'{FCUR[i]}{rr}', f'={FCUR[i]}{row_nopat}/{FCUR[i]}{row_ic}', F['roic'][i], PCT2)
rr += 1
rr += 1
row_pvexpl = rr; putf(wsd, f'C{rr}', f'=SUM(C{row_pv}:G{row_pv})', DCF['pv_explicit'], NUM0)
put(wsd, f'A{rr}', 'PV of explicit FCFF (5 years)', fmt=None); rr += 1
row_roicterm = rr; putf(wsd, f'C{rr}', f'=G{row_nopat}*(1+{AC("Terminal growth")})/G{row_ic}', DCF['roic_term'], PCT2)
put(wsd, f'A{rr}', 'Terminal ROIC', fmt=None); ANCH['roic_term'] = rr; rr += 1
row_rrterm = rr; putf(wsd, f'C{rr}', f'={AC("Terminal growth")}/C{row_roicterm}', DCF['rr_term'], PCT2)
put(wsd, f'A{rr}', 'Terminal reinvestment rate = g / ROIC', fmt=None); rr += 1
row_nopatterm = rr; putf(wsd, f'C{rr}', f'=G{row_nopat}*(1+{AC("Terminal growth")})', DCF['tv'] * (W['wacc_term'] - IN['g_term']) / (1 - DCF['rr_term']), NUM0)
put(wsd, f'A{rr}', 'Terminal NOPAT (next year)', fmt=None); rr += 1
row_tv = rr; putf(wsd, f'C{rr}', f'=C{row_nopatterm}*(1-C{row_rrterm})/(C{row_waccterm}-{AC("Terminal growth")})', DCF['tv'], NUM0)
put(wsd, f'A{rr}', 'Terminal value', fmt=None); rr += 1
row_pvtv = rr; putf(wsd, f'C{rr}', f'=C{row_tv}*G{row_df}', DCF['pv_tv'], NUM0)
put(wsd, f'A{rr}', 'PV of terminal value', fmt=None); rr += 1
row_ev = rr; putf(wsd, f'C{rr}', f'=C{row_pvexpl}+C{row_pvtv}', DCF['ev'], NUM0, bold=True)
put(wsd, f'A{rr}', 'Enterprise value', bold=True, fmt=None); ANCH['ev'] = rr; rr += 1
row_tvshare = rr; putf(wsd, f'C{rr}', f'=C{row_pvtv}/C{row_ev}', DCF['tv_share'], PCT, bold=True)
put(wsd, f'A{rr}', 'Terminal value as % of EV', bold=True, fmt=None); ANCH['tv_share'] = rr; rr += 1
# --- bridge ---
rr += 1; put(wsd, f'A{rr}', 'Enterprise value to equity', bold=True, fmt=None); band(wsd, rr, 8); rr += 1
row_lessnd = rr; putf(wsd, f'C{rr}', f'=-{AC("Net financial debt at FY2025 (SAR mn, disclosed)")}', -DCF['nd'], NUM0)
put(wsd, f'A{rr}', 'Less: net financial debt', fmt=None); rr += 1
row_addassoc = rr; putf(wsd, f'C{rr}', f'={AC("FY2025 associates carrying value (SAR mn)")}+{AC("FY2025 non-operating assets (SAR mn)")}', DCF['assoc'] + DCF['nonop'], NUM0)
put(wsd, f'A{rr}', 'Add: associates + non-operating assets', fmt=None); rr += 1
row_lessnci = rr; putf(wsd, f'C{rr}', f'=-{AC("FY2025 NCI carrying value (SAR mn)")}', -DCF['nci'], NUM0)
put(wsd, f'A{rr}', 'Less: non-controlling interests', fmt=None); rr += 1
row_eqattr = rr; putf(wsd, f'C{rr}', f'=C{row_ev}+C{row_lessnd}+C{row_addassoc}+C{row_lessnci}', DCF['eq_attr'], NUM0, bold=True)
put(wsd, f'A{rr}', 'Equity attributable (31-Dec-2025)', bold=True, fmt=None); rr += 1
row_psdec = rr; putf(wsd, f'C{rr}', f'=C{row_eqattr}/{AC("Shares outstanding (mn)")}', DCF['ps_dec'], PX)
put(wsd, f'A{rr}', 'Value per share (31-Dec-2025)', fmt=None); rr += 1
row_roll = rr; putf(wsd, f'C{rr}', f'=(1+C{row_ke})^({AC("Days to the anchor")}/365)', DCF['roll'], DF4)
put(wsd, f'A{rr}', 'Anchor accretion factor', fmt=None); rr += 1
row_ps = rr; putf(wsd, f'C{rr}', f'=C{row_psdec}*C{row_roll}-{AC("FY2025 dividend per share paid in window")}', DCF['ps'], PX, bold=True)
put(wsd, f'A{rr}', 'Value per share at anchor (18-Aug-2026)', bold=True, fmt=None); ANCH['dcf_ps'] = rr; rr += 1
row_vsspot = rr; putf(wsd, f'C{rr}', f'=C{row_ps}/{AC("Spot price (SAR)")}-1', DCF['ps'] / SPOT - 1, PCT)
put(wsd, f'A{rr}', 'vs spot', fmt=None); rr += 1
# DCF scenario bounds — pasted whole-model re-runs
rr += 1; put(wsd, f'A{rr}', 'DCF scenarios (whole-model re-runs — pasted)', bold=True, fmt=None); band(wsd, rr, 8); rr += 1
put(wsd, f'A{rr}', 'Bear', fmt=None); put(wsd, f'C{rr}', DCF['bear'], BLUE, PX); row_bear = rr; rr += 1
put(wsd, f'A{rr}', 'Base (= value per share at anchor)', fmt=None); putf(wsd, f'C{rr}', f'=C{row_ps}', DCF['ps'], PX, green=True); rr += 1
put(wsd, f'A{rr}', 'Bull', fmt=None); put(wsd, f'C{rr}', DCF['bull'], BLUE, PX); row_bull = rr; rr += 1
ANCH['dcf_bear'], ANCH['dcf_bull'] = row_bear, row_bull

seg = ANCH['seg']
DR = f"='DCF'!C{row_roll}"                       # anchor roll factor cell (cross-sheet)


def anchor(expr_v):
    """wrap a raw value expression to the anchor: v*roll - div_window."""
    return f"({expr_v})*DCF!$C${row_roll}-{AC('FY2025 dividend per share paid in window')}"


# =========================================================================
# BALANCE SHEET — audited history (pasted) + the forecast net-debt / profit
# / equity roll-forward that the income statement and lenses consume
# =========================================================================
wsb = sheet('Balance Sheet')
title(wsb, 'Balance sheet — audited history + forecast roll-forward', 'SAR mn; FY2023-25 audited, '
      'FY2026E-30E rolled forward from the cash flow', 10, awidth=40, cwidth=11)
hdr(wsb, 4, ['SAR mn', ''] + YH + YF)
bl = {}
rr = 5
for key, nm in [('ppe', 'Property, plant & equipment'), ('inv', 'Inventory'),
                ('recv', 'Trade & other receivables'), ('cash', 'Cash & equivalents'),
                ('assets', 'Total assets'), ('debt', 'Gross borrowings incl. leases'),
                ('pay', 'Trade & other payables'), ('nwc', 'Net working capital'),
                ('nd', 'Net financial debt'), ('eqp', 'Equity attributable')]:
    bl[key] = rr; put(wsb, f'A{rr}', nm, fmt=None); rr += 1
# historical (pasted, audited)
for ci, y in zip(HCOL, ['FY23', 'FY24', 'FY25']):
    for key in ['ppe', 'inv', 'recv', 'cash', 'assets', 'debt', 'pay', 'nwc', 'nd']:
        val = HB[y].get(key)
        if val is not None:
            put(wsb, f'{ci}{bl[key]}', val, BLUE, NUM0)
    if 'eqp' in HB[y]:
        put(wsb, f'{ci}{bl["eqp"]}', HB[y]['eqp'], BLUE, NUM0)
# forecast roll-forward rows (helper block below the visible statement)
rf0 = rr + 1
put(wsb, f'A{rf0}', 'Forecast roll-forward (feeds the income statement & lenses)', bold=True, fmt=None)
band(wsb, rf0, 10)
rf = {}
for key, nm in [('ndstart', 'Net debt, start of year'), ('cash_s', 'Surplus cash'),
                ('int', 'Net interest'), ('ebit', 'EBIT (from DCF)'), ('pbt', 'Profit before zakat'),
                ('npa', 'Attributable profit'), ('div', 'Dividend'), ('ndend', 'Net debt, end of year'),
                ('eq', 'Equity attributable')]:
    rf[key] = rf0 + 1 + list(['ndstart', 'cash_s', 'int', 'ebit', 'pbt', 'npa', 'div', 'ndend', 'eq']).index(key)
    put(wsb, f'A{rf[key]}', nm, fmt=None)
rr = rf['eq'] + 1
DEBT = AC('FY2025 gross borrowings incl. leases (SAR mn)')
YLD = AC('Yield on surplus cash')
for i in range(5):
    c = FCUR[i]; pc = FCUR[i - 1]
    # start net debt = prior year end (year0 uses disclosed FY2025 net debt)
    if i == 0:
        putf(wsb, f'{c}{rf["ndstart"]}', f'={AC("Net financial debt at FY2025 (SAR mn, disclosed)")}', DCF['nd'], NUM0)
    else:
        putf(wsb, f'{c}{rf["ndstart"]}', f'={pc}{rf["ndend"]}', F['net_debt'][i - 1], NUM0)
    putf(wsb, f'{c}{rf["cash_s"]}', f'={DEBT}-{c}{rf["ndstart"]}', F['debt_fy25'] - (DCF['nd'] if i == 0 else F['net_debt'][i - 1]), NUM0)
    putf(wsb, f'{c}{rf["int"]}', f'={AP("Cost-of-debt path", i)}*{DEBT}-{YLD}*MAX({c}{rf["cash_s"]},0)', F['interest'][i], NUM0)
    putf(wsb, f'{c}{rf["ebit"]}', f"=DCF!{c}{row_ebit}", F['ebit'][i], NUM0, green=True)
    putf(wsb, f'{c}{rf["pbt"]}', f'={c}{rf["ebit"]}-{c}{rf["int"]}', F['ebit'][i] - F['interest'][i], NUM0)
    putf(wsb, f'{c}{rf["npa"]}', f'={c}{rf["pbt"]}*(1-{AC("Effective zakat and income tax rate")})*(1-{AC("NCI share of forecast profit")})', F['np_attr'][i], NUM0)
    putf(wsb, f'{c}{rf["div"]}', f'={AC("Forecast dividend payout ratio")}*{c}{rf["npa"]}', F['div'][i], NUM0)
    putf(wsb, f'{c}{rf["ndend"]}', f'={c}{rf["ndstart"]}-(DCF!{c}{row_fcff}-{c}{rf["int"]}*(1-{AC("Effective zakat and income tax rate")}))+{c}{rf["div"]}', F['net_debt'][i], NUM0)
    if i == 0:
        putf(wsb, f'{c}{rf["eq"]}', f'={AC("FY2025 equity attributable (SAR mn)")}+{c}{rf["npa"]}-{c}{rf["div"]}', F['equity'][i], NUM0)
    else:
        putf(wsb, f'{c}{rf["eq"]}', f'={pc}{rf["eq"]}+{c}{rf["npa"]}-{c}{rf["div"]}', F['equity'][i], NUM0)
# link the visible forecast net-debt / PP&E / equity rows to the roll-forward + DCF
for i in range(5):
    c = FCOL[i]; cf = FCUR[i]
    putf(wsb, f'{c}{bl["nd"]}', f'={cf}{rf["ndend"]}', F['net_debt'][i], NUM0, green=True)
    putf(wsb, f'{c}{bl["ppe"]}', f"=DCF!{cf}{row_ppe}", F['ppe'][i], NUM0, green=True)
    putf(wsb, f'{c}{bl["nwc"]}', f"=DCF!{cf}{row_nwc}", F['nwc'][i], NUM0, green=True)
    putf(wsb, f'{c}{bl["eqp"]}', f'={cf}{rf["eq"]}', F['equity'][i], NUM0, green=True)
    putf(wsb, f'{c}{bl["debt"]}', f'={DEBT}', F['debt_fy25'], NUM0, green=True)
ANCH['bl'], ANCH['rf'] = bl, rf

# =========================================================================
# INCOME STATEMENT — 3y audited history (pasted) + 5y forecast (formula)
# =========================================================================
wsi = sheet('Income Statement')
title(wsi, 'Income statement — 3y audited + 5y forecast', 'SAR mn', 10, awidth=38, cwidth=11)
hdr(wsi, 4, ['SAR mn', ''] + YH + YF)
il = {}
rows_is = [('rev', 'Revenue'), ('gp', 'Gross profit'), ('gm', 'Gross margin'),
           ('opex', 'Operating expenses'), ('ebitda', 'EBITDA'), ('dna', 'Depreciation & amortisation'),
           ('ebit', 'Operating profit (EBIT)'), ('fin', 'Net finance cost'), ('ebt', 'Profit before zakat'),
           ('zak', 'Zakat & income tax'), ('pat', 'Profit for the year'), ('npa', 'Attributable profit')]
rr = 5
for k, nm in rows_is:
    il[k] = rr; put(wsi, f'A{rr}', nm, fmt=None); rr += 1
# history (audited, pasted)
for ci, y in zip(HCOL, ['FY23', 'FY24', 'FY25']):
    h = HI[y]
    put(wsi, f'{ci}{il["rev"]}', h['rev'], BLUE, NUM0)
    put(wsi, f'{ci}{il["gp"]}', h['gp'], BLUE, NUM0)
    putf(wsi, f'{ci}{il["gm"]}', f'={ci}{il["gp"]}/{ci}{il["rev"]}', h['gp'] / h['rev'], PCT2)
    put(wsi, f'{ci}{il["ebit"]}', h['ebit'], BLUE, NUM0)
    put(wsi, f'{ci}{il["dna"]}', h['dna'], BLUE, NUM0)
    putf(wsi, f'{ci}{il["ebitda"]}', f'={ci}{il["ebit"]}+{ci}{il["dna"]}', h['ebitda'], NUM0)
    putf(wsi, f'{ci}{il["opex"]}', f'={ci}{il["gp"]}-{ci}{il["ebit"]}', h['gp'] - h['ebit'], NUM0)
    put(wsi, f'{ci}{il["fin"]}', h['fin'], BLUE, NUM0)
    putf(wsi, f'{ci}{il["ebt"]}', f'={ci}{il["ebit"]}+{ci}{il["fin"]}', h['ebt'], NUM0)
    put(wsi, f'{ci}{il["zak"]}', h['zakat'], BLUE, NUM0)
    put(wsi, f'{ci}{il["pat"]}', h['pat'], BLUE, NUM0)   # audited (pasted); FY2023 carries a small
    #                                                     undisclosed deferred/other tax residual so
    #                                                     it is the disclosed figure, not ebt+zakat
put(wsi, f'E{il["npa"]}', IN['npa_fy25'], BLUE, NUM0)  # FY25 attributable profit (audited, disclosed)
# forecast (formula, links to Segments + BS roll-forward)
rf = ANCH['rf']
for i in range(5):
    c = FCOL[i]; cf = FCUR[i]
    putf(wsi, f'{c}{il["rev"]}', f"=Segments!{cf}{seg['rev']}", F['rev'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["gp"]}', f"=Segments!{cf}{seg['gp']}", F['gp'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["gm"]}', f'={c}{il["gp"]}/{c}{il["rev"]}', F['gp'][i] / F['rev'][i], PCT2)
    putf(wsi, f'{c}{il["opex"]}', f"=Segments!{cf}{seg['opex']}", F['opex'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["dna"]}', f"=Segments!{cf}{seg['dna']}", F['dna'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["ebitda"]}', f"=Segments!{cf}{seg['ebitda']}", F['ebitda'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["ebit"]}', f"=Segments!{cf}{seg['ebit']}", F['ebit'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["fin"]}', f"=-'Balance Sheet'!{cf}{rf['int']}", -F['interest'][i], NUM0, green=True)
    putf(wsi, f'{c}{il["ebt"]}', f'={c}{il["ebit"]}+{c}{il["fin"]}', F['ebit'][i] - F['interest'][i], NUM0)
    putf(wsi, f'{c}{il["zak"]}', f'=-{c}{il["ebt"]}*{AC("Effective zakat and income tax rate")}', -(F['ebit'][i] - F['interest'][i]) * TAX, NUM0)
    putf(wsi, f'{c}{il["pat"]}', f'={c}{il["ebt"]}+{c}{il["zak"]}', (F['ebit'][i] - F['interest'][i]) * (1 - TAX), NUM0)
    putf(wsi, f'{c}{il["npa"]}', f"='Balance Sheet'!{cf}{rf['npa']}", F['np_attr'][i], NUM0, green=True)
ANCH['il'] = il

# =========================================================================
# CASH FLOW — the FCFF waterfall as a statement (forecast)
# =========================================================================
wsc = sheet('Cash Flow')
title(wsc, 'Cash flow — free cash flow to the firm', 'SAR mn; the DCF waterfall as a statement', 8,
      awidth=40, cwidth=12)
hdr(wsc, 4, ['SAR mn', ''] + YF)
cl = {}
rows_cf = [('nopat', 'NOPAT'), ('dna', '+ Depreciation & amortisation'), ('capex', '- Capital expenditure'),
           ('dnwc', '- Change in working capital'), ('fcff', 'Free cash flow to firm'),
           ('df', 'Discount factor'), ('pv', 'PV of FCFF')]
rr = 5
for k, nm in rows_cf:
    cl[k] = rr; put(wsc, f'A{rr}', nm, fmt=None); rr += 1
for i in range(5):
    c = FCUR[i]
    putf(wsc, f'{c}{cl["nopat"]}', f'=DCF!{c}{row_nopat}', F['nopat'][i], NUM0, green=True)
    putf(wsc, f'{c}{cl["dna"]}', f'=DCF!{c}{row_dna}', F['dna'][i], NUM0, green=True)
    putf(wsc, f'{c}{cl["capex"]}', f'=-DCF!{c}{row_capex}', -F['capex'][i], NUM0, green=True)
    putf(wsc, f'{c}{cl["dnwc"]}', f'=-DCF!{c}{row_dnwc}', -F['dnwc'][i], NUM0, green=True)
    putf(wsc, f'{c}{cl["fcff"]}', f'={c}{cl["nopat"]}+{c}{cl["dna"]}+{c}{cl["capex"]}+{c}{cl["dnwc"]}', F['fcff'][i], NUM0, bold=True)
    putf(wsc, f'{c}{cl["df"]}', f'=DCF!{c}{row_df}', F['df'][i], DF4, green=True)
    putf(wsc, f'{c}{cl["pv"]}', f'={c}{cl["fcff"]}*{c}{cl["df"]}', F['pv'][i], NUM0)
ANCH['cl'] = cl

ROLLC = f"DCF!$C${row_roll}"
DIVW = AC('FY2025 dividend per share paid in window')
NDc = AC('Net financial debt at FY2025 (SAR mn, disclosed)')
ASSOCc = AC('FY2025 associates carrying value (SAR mn)')
NONOPc = AC('FY2025 non-operating assets (SAR mn)')
NCIc = AC('FY2025 NCI carrying value (SAR mn)')
SHc = AC('Shares outstanding (mn)')


def ANC(raw):
    return f'({raw})*{ROLLC}-{DIVW}'


# =========================================================================
# RELATIVE & NORMALIZED — the relative, normalised-earnings and book lenses
# =========================================================================
wsr = sheet('Relative & Normalized')
title(wsr, 'Relative, normalised-earnings & book lenses', 'Each base value is a live formula; bear/bull '
      'vary the multiple', 7, awidth=46, cwidth=13)
rr = 4
put(wsr, f'A{rr}', 'Relative — EV/EBITDA on FY2027E', bold=True, fmt=None); band(wsr, rr, 7); rr += 1
put(wsr, f'A{rr}', 'FY2027E EBITDA', fmt=None); putf(wsr, f'C{rr}', f"=Segments!D{seg['ebitda']}", F['ebitda'][1], NUM0, green=True); row_eb27 = rr; rr += 1
put(wsr, f'A{rr}', 'Justified EV/EBITDA', fmt=None); putf(wsr, f'C{rr}', f'={AC("Justified EV/EBITDA")}', IN['ev_ebitda_just'], MULT, green=True); row_mult = rr; rr += 1
put(wsr, f'A{rr}', 'PV of interim FY2026-27E FCFF', fmt=None); putf(wsr, f'C{rr}', f'=DCF!C{row_pv}+DCF!D{row_pv}', F['pv'][0] + F['pv'][1], NUM0, green=True); row_pvint = rr; rr += 1


def rel_raw(multexpr):
    return (f'({multexpr}*C{row_eb27}*DCF!D{row_df}+C{row_pvint}-{NDc}+{ASSOCc}+{NONOPc}-{NCIc})/{SHc}')


row_relbase = rr; putf(wsr, f'C{rr}', f'={ANC(rel_raw(f"C{row_mult}"))}', LN['relative']['base'], PX, bold=True)
put(wsr, f'A{rr}', 'Relative lens — implied value (base)', bold=True, fmt=None); rr += 1
row_relbounds = rr; put(wsr, f'A{rr}', 'Bear (7.5x) / Bull (11.0x)', fmt=None)
putf(wsr, f'C{rr}', f'={ANC(rel_raw("7.5"))}', LN['relative']['bear'], PX)
putf(wsr, f'D{rr}', f'={ANC(rel_raw("11"))}', LN['relative']['bull'], PX); rr += 1
rr += 1
put(wsr, f'A{rr}', 'Normalised earnings power', bold=True, fmt=None); band(wsr, rr, 7); rr += 1
put(wsr, f'A{rr}', 'Mid-cycle EBITDA margin (FY2028E)', fmt=None); putf(wsr, f'C{rr}', f"=Segments!E{seg['ebmar']}", F['ebitda_margin'][2], PCT2, green=True); row_nmar = rr; rr += 1
put(wsr, f'A{rr}', 'FY2026E revenue', fmt=None); putf(wsr, f'C{rr}', f"=Segments!C{seg['rev']}", F['rev'][0], NUM0, green=True); row_nrev = rr; rr += 1
put(wsr, f'A{rr}', 'FY2026E net interest', fmt=None); putf(wsr, f'C{rr}', f"='Balance Sheet'!C{ANCH['rf']['int']}", F['interest'][0], NUM0, green=True); row_nint = rr; rr += 1
DNAP = AC('Depreciation and amortisation / revenue')
TAXc = AC('Effective zakat and income tax rate')
PEc = AC('Justified price/earnings')
ROEc = AC('Sustainable return on equity')
Gc = AC('Terminal growth')
KEc = f'DCF!C{row_ke}'
KETc = f'DCF!C{row_keterm}'
row_neps = rr
putf(wsr, f'C{rr}', f'=((C{row_nmar}*C{row_nrev}-{DNAP}*C{row_nrev})-C{row_nint})*(1-{TAXc})*(1-{AC("NCI share of forecast profit")})/{SHc}',
     NRM['eps'], PX); put(wsr, f'A{rr}', 'Normalised EPS', fmt=None); rr += 1
row_normbase = rr
putf(wsr, f'C{rr}', f'={ANC(f"{PEc}*C{row_neps}")}', LN['normalized']['base'], PX, bold=True)
put(wsr, f'A{rr}', 'Normalised lens — implied value (base)', bold=True, fmt=None); rr += 1
put(wsr, f'A{rr}', 'Bear (10x) / Bull (16x)', fmt=None)
putf(wsr, f'C{rr}', f'={ANC(f"10*C{row_neps}")}', LN['normalized']['bear'], PX)
putf(wsr, f'D{rr}', f'={ANC(f"16*C{row_neps}")}', LN['normalized']['bull'], PX); rr += 1
rr += 1
put(wsr, f'A{rr}', 'Book value & sustainable return', bold=True, fmt=None); band(wsr, rr, 7); rr += 1
row_bvps = rr
putf(wsr, f'C{rr}', f'={AC("FY2025 equity attributable (SAR mn)")}/{SHc}', BKL['bvps'], PX)
put(wsr, f'A{rr}', 'Book value per share', fmt=None); rr += 1
row_pbj = rr
putf(wsr, f'C{rr}', f'=({ROEc}-{Gc})/({KETc}-{Gc})', BKL['pb_just'], MULT)
put(wsr, f'A{rr}', 'Justified price / book', fmt=None); rr += 1
row_bookbase = rr
putf(wsr, f'C{rr}', f'={ANC(f"C{row_pbj}*C{row_bvps}")}', LN['book']['base'], PX, bold=True)
put(wsr, f'A{rr}', 'Book lens — implied value (base)', bold=True, fmt=None); rr += 1
put(wsr, f'A{rr}', 'Bear / Bull', fmt=None)
book_bear_raw = f"(({ROEc})-0.05-{Gc})/({KEc}-{Gc})*C{row_bvps}"
book_bull_raw = f"(({ROEc})+0.03-{Gc})/({KETc}-{Gc})*C{row_bvps}"
putf(wsr, f'C{rr}', f'={ANC(book_bear_raw)}', LN['book']['bear'], PX)
putf(wsr, f'D{rr}', f'={ANC(book_bull_raw)}', LN['book']['bull'], PX); rr += 1
ANCH['rel'] = dict(relbase=row_relbase, relbounds=row_relbounds, normbase=row_normbase, bookbase=row_bookbase, bvps=row_bvps)

# =========================================================================
# FUNDAMENTAL VALUATION — the four lenses, weighted central, contested judgement
# =========================================================================
wsf = sheet('Fundamental Valuation')
title(wsf, 'Fundamental valuation — four lenses, one field', 'Every base value links live to its '
      'source sheet', 7, awidth=48, cwidth=13)
hdr(wsf, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Role', '', 'vs price'])
LSRC = {'dcf': f'=DCF!C{row_ps}', 'relative': f"='Relative & Normalized'!C{ANCH['rel']['relbase']}",
        'normalized': f"='Relative & Normalized'!C{ANCH['rel']['normbase']}",
        'book': f"='Relative & Normalized'!C{ANCH['rel']['bookbase']}"}
BEARSRC = {'relative': f"='Relative & Normalized'!C{ANCH['rel']['relbounds']}",
           'normalized': f"='Relative & Normalized'!C{ANCH['rel']['normbase']+1}",
           'book': f"='Relative & Normalized'!C{ANCH['rel']['bookbase']+1}"}
BULLSRC = {'relative': f"='Relative & Normalized'!D{ANCH['rel']['relbounds']}",
           'normalized': f"='Relative & Normalized'!D{ANCH['rel']['normbase']+1}",
           'book': f"='Relative & Normalized'!D{ANCH['rel']['bookbase']+1}"}
LK = ['dcf', 'relative', 'normalized', 'book']
RETW = D['lens_record']['retired']['blend']
ROLE = {'dcf': 'THE CENTRAL — the class primary', 'relative': 'cross-check',
        'normalized': 'REMOVED — not a lens this class publishes',
        'book': 'a disclosed floor, never weighted'}
rr = 5
row_lens0 = rr
for k in ['dcf', 'relative', 'normalized', 'book']:
    put(wsf, f'A{rr}', LN[k]['name'], fmt=None)
    if k == 'dcf':
        putf(wsf, f'B{rr}', f'=DCF!C{ANCH["dcf_bear"]}', LN[k]['bear'], PX, green=True)
        putf(wsf, f'D{rr}', f'=DCF!C{ANCH["dcf_bull"]}', LN[k]['bull'], PX, green=True)
    else:
        putf(wsf, f'B{rr}', BEARSRC[k], LN[k]['bear'], PX, green=True)
        putf(wsf, f'D{rr}', BULLSRC[k], LN[k]['bull'], PX, green=True)
    putf(wsf, f'C{rr}', LSRC[k], LN[k]['base'], PX, green=True)
    put(wsf, f'E{rr}', ROLE[k], fmt=None)
    putf(wsf, f'G{rr}', f'=C{rr}/{AC("Spot price (SAR)")}-1', LN[k]['base'] / SPOT - 1, PCT)
    rr += 1
band(wsf, rr, 7)
row_cent = rr
# THE WEIGHT AND CONTRIBUTION COLUMNS WENT WITH THE BLEND, and a role column replaces them.
put(wsf, f'A{rr}', 'THE CENTRAL — the cash-flow lens, not an average', bold=True, fmt=None)
putf(wsf, f'B{rr}', f'=B{row_lens0}', LN['dcf']['bear'], PX, bold=True)
putf(wsf, f'C{rr}', f'=C{row_lens0}', D['central'], PX, bold=True)
putf(wsf, f'D{rr}', f'=D{row_lens0}', LN['dcf']['bull'], PX, bold=True)
put(wsf, f'E{rr}', 'the class primary', fmt=None)
putf(wsf, f'G{rr}', f'=C{rr}/{AC("Spot price (SAR)")}-1', D['central'] / SPOT - 1, PCT, bold=True)
rr += 1
put(wsf, f'A{rr}', 'NOT AVERAGED — the retired 45/20/20/15 blend, published unused',
    bold=True, fmt=None)
putf(wsf, f'C{rr}', '=' + '+'.join('C%d*%g' % (row_lens0 + i, RETW[k])
                                   for i, k in enumerate(LK)),
     D['retired_blend_value'], PX)
put(wsf, f'E{rr}', 'retired 04-Sep-2026', fmt=None)
putf(wsf, f'G{rr}', f'=C{rr}/{AC("Spot price (SAR)")}-1',
     D['retired_blend_value'] / SPOT - 1, PCT)
rr += 1
put(wsf, f'A{rr}', 'Span across the lenses (min/max) — a spread between METHODS, not a '
                   'range around the answer', fmt=None)
putf(wsf, f'B{rr}', f'=MIN(B{row_lens0}:B{row_lens0+3})', min(LN[k]['bear'] for k in LK), PX)
putf(wsf, f'D{rr}', f'=MAX(D{row_lens0}:D{row_lens0+3})', max(LN[k]['bull'] for k in LK), PX)
rr += 2
put(wsf, f'A{rr}', 'Expert panel median (whole-model — pasted)', fmt=None)
row_panel = rr                 # anchored, never reached by an offset from the central
put(wsf, f'C{rr}', D['panel_centre'], BLUE, PX); rr += 1
put(wsf, f'A{rr}', 'Market price (anchor)', fmt=None); putf(wsf, f'C{rr}', f'={AC("Spot price (SAR)")}', SPOT, PX, green=True); rr += 2
band(wsf, rr, 7); put(wsf, f'A{rr}', 'Central contested judgement — sustained gross margin, both ways', bold=True, fmt=None); rr += 1
put(wsf, f'A{rr}', 'H1-2026 anchor (15.26%) — DCF value', fmt=None); putf(wsf, f'C{rr}', f'=DCF!C{row_ps}', DCF['spread_base'], PX, green=True); rr += 1
put(wsf, f'A{rr}', 'FY2025-peak framing (16.0%) — DCF value (pasted re-run)', fmt=None); put(wsf, f'C{rr}', DCF['spread_bull'], BLUE, PX); rr += 1
put(wsf, f'A{rr}', 'Further-compression framing (14.5%) — DCF value (pasted re-run)', fmt=None); put(wsf, f'C{rr}', DCF['spread_bear'], BLUE, PX); rr += 1
# Anchored by NAME rather than reached by an offset. The Summary sheet used to find
# the panel median at cent+2; retiring the blend added two rows between them and the
# reference silently landed on an empty cell — L-067, and the recalculation gate
# caught it within the minute.
ANCH['fund'] = dict(cent=row_cent, lens0=row_lens0, panel=row_panel)

# =========================================================================
# SOTP BRIDGE — EV to equity, with terminal-value share visible (gate p)
# =========================================================================
wsp = sheet('SOTP Bridge')
title(wsp, 'Enterprise value to equity bridge', 'All rows link to the DCF sheet; terminal-value share '
      'shown to the reader', 5, awidth=48, cwidth=15)
rr = 4
bridge = [('Enterprise value', f'=DCF!C{row_ev}', DCF['ev']),
          ('  of which terminal value', f'=DCF!C{row_pvtv}', DCF['pv_tv']),
          ('  terminal value as % of EV', f'=DCF!C{row_tvshare}', DCF['tv_share']),
          ('Less: net financial debt', f'=DCF!C{row_lessnd}', -DCF['nd']),
          ('Add: associates + non-operating assets', f'=DCF!C{row_addassoc}', DCF['assoc'] + DCF['nonop']),
          ('Less: non-controlling interests', f'=DCF!C{row_lessnci}', -DCF['nci']),
          ('Equity attributable (31-Dec-2025)', f'=DCF!C{row_eqattr}', DCF['eq_attr']),
          ('Value per share at anchor (SAR)', f'=DCF!C{row_ps}', DCF['ps'])]
for nm, formula, exp in bridge:
    put(wsp, f'A{rr}', nm, fmt=None)
    fmt = PCT if '%' in nm else (PX if 'per share' in nm else NUM0)
    putf(wsp, f'C{rr}', formula, exp, fmt, green=True, bold=('Equity attributable' in nm or 'Value per share' in nm))
    rr += 1

# =========================================================================
# SUMMARY FINANCIALS — the forecast engine at a glance
# =========================================================================
wsm = sheet('Summary Financials')
title(wsm, 'Summary financials', 'SAR mn; forecast links to the model', 8, awidth=34, cwidth=12)
hdr(wsm, 4, ['SAR mn', ''] + YF)
sfrows = [('rev', 'Revenue', 'rev', row_rev), ('ebitda', 'EBITDA', None, None),
          ('nopat', 'NOPAT', 'nopat', row_nopat), ('fcff', 'FCFF', 'fcff', row_fcff),
          ('ic', 'Invested capital', 'ic', row_ic), ('roic', 'ROIC', 'roic', row_roic)]
rr = 5
for key, nm, fk, drow in sfrows:
    put(wsm, f'A{rr}', nm, fmt=None)
    for i in range(5):
        c = FCUR[i]
        if key == 'ebitda':
            putf(wsm, f'{c}{rr}', f"=Segments!{c}{seg['ebitda']}", F['ebitda'][i], NUM0, green=True)
        elif key == 'roic':
            putf(wsm, f'{c}{rr}', f'=DCF!{c}{drow}', F['roic'][i], PCT2, green=True)
        else:
            putf(wsm, f'{c}{rr}', f'=DCF!{c}{drow}', F[fk][i], NUM0, green=True)
    rr += 1

# =========================================================================
# PER-SHARE & RATIOS
# =========================================================================
wpr = sheet('Per-Share & Ratios')
title(wpr, 'Per-share & ratios', '', 8, awidth=36, cwidth=12)
hdr(wpr, 4, ['', 'FY2025'] + YF[:5])
il = ANCH['il']; bl = ANCH['bl']
rr = 5
row_eps = rr
put(wpr, f'A{rr}', 'EPS (SAR)', fmt=None)
putf(wpr, f'B{rr}', f"='Income Statement'!E{il['npa']}/{SHc}", IN['npa_fy25'] / SH, PX, green=True)
for i in range(5):
    putf(wpr, f'{FCUR[i]}{rr}', f"='Balance Sheet'!{FCUR[i]}{ANCH['rf']['npa']}/{SHc}", F['np_attr'][i] / SH, PX)
rr += 1
put(wpr, f'A{rr}', 'Book value per share (SAR)', fmt=None)
putf(wpr, f'B{rr}', f'={AC("FY2025 equity attributable (SAR mn)")}/{SHc}', F['eqp_fy25'] / SH, PX)
for i in range(5):
    putf(wpr, f'{FCUR[i]}{rr}', f"='Balance Sheet'!{FCOL[i]}{ANCH['bl']['eqp']}/{SHc}", F['equity'][i] / SH, PX)
rr += 1
put(wpr, f'A{rr}', 'ROE (%) — attributable profit / average equity', fmt=None)
putf(wpr, f'B{rr}', f"='Income Statement'!E{il['npa']}/AVERAGE('Balance Sheet'!D{bl['eqp']},'Balance Sheet'!E{bl['eqp']})", BKL['roe_trailing'], PCT, green=True)
rr += 1
put(wpr, f'A{rr}', 'Net debt / EBITDA (x)', fmt=None)
putf(wpr, f'B{rr}', f"={AC('Net financial debt at FY2025 (SAR mn, disclosed)')}/'Income Statement'!E{il['ebitda']}", DCF['nd'] / HI['FY25']['ebitda'], MULT, green=True)
for i in range(5):
    putf(wpr, f'{FCUR[i]}{rr}', f"='Balance Sheet'!{FCOL[i]}{ANCH['bl']['nd']}/Segments!{FCUR[i]}{seg['ebitda']}", F['net_debt'][i] / F['ebitda'][i], MULT)
rr += 1
row_evebitda = rr
put(wpr, f'A{rr}', 'Trailing EV/EBITDA (x)', fmt=None)
putf(wpr, f'B{rr}', f"=({AC('Spot price (SAR)')}*{SHc}+{AC('Net financial debt at FY2025 (SAR mn, disclosed)')})/'Income Statement'!E{il['ebitda']}", REL['ev_ebitda_trailing'], MULT, green=True)
rr += 1
put(wpr, f'A{rr}', 'Trailing P/E (x)', fmt=None)
putf(wpr, f'B{rr}', f"={AC('Spot price (SAR)')}/B{row_eps}", REL['pe_trailing'], MULT, green=True)
rr += 1
ANCH['per_share'] = dict(evebitda=row_evebitda)

# =========================================================================
# MONTE CARLO — the calibrated forward cone (pasted engine output)
# =========================================================================
wmc = sheet('Monte Carlo')
title(wmc, 'Monte Carlo — calibrated forward price cone', 'Engine output (50,000 paths, seed 42); '
      'pasted whole-model re-runs, not driver formulas', 6, awidth=34, cwidth=13)
hdr(wmc, 4, ['Percentile', '1-month', '3-month'])
h1, h3 = STK['horizons']['1M'], STK['horizons']['3M']
rr = 5
for p in ['p5', 'p25', 'p50', 'p75', 'p95']:
    put(wmc, f'A{rr}', p.upper(), fmt=None)
    put(wmc, f'B{rr}', h1['pct'][p], BLUE, PX); put(wmc, f'C{rr}', h3['pct'][p], BLUE, PX); rr += 1
put(wmc, f'A{rr}', 'Spot', fmt=None); put(wmc, f'B{rr}', STK['spot'], BLUE, PX); put(wmc, f'C{rr}', STK['spot'], BLUE, PX); rr += 1
put(wmc, f'A{rr}', 'P(above spot)', fmt=None); put(wmc, f'B{rr}', h1['p_above'], BLUE, PCT); put(wmc, f'C{rr}', h3['p_above'], BLUE, PCT); rr += 1
put(wmc, f'A{rr}', 'Annualised vol', fmt=None); put(wmc, f'B{rr}', h1['anchor_vol_ann'], BLUE, PCT); put(wmc, f'C{rr}', h3['anchor_vol_ann'], BLUE, PCT); rr += 2
put(wmc, f'A{rr}', f"3-month calibration: {S0['verdict']} vs the random walk; coverage "
    f"{S0['cov50']:.2f}/{S0['cov80']:.2f}/{S0['cov90']:.2f}; PIT mean {S0['pit_mean']:.2f}.", fmt=None)

# =========================================================================
# SENSITIVITY — pasted whole-model grids
# =========================================================================
wsn = sheet('Sensitivity')
title(wsn, 'Sensitivity — whole-model re-runs (pasted)', 'Each cell is a complete revaluation and does '
      'NOT redraw when a driver changes', 8, awidth=30, cwidth=11)
rr = 4
put(wsn, f'A{rr}', 'DCF value per share: terminal WACC (rows) x terminal growth (cols)', bold=True, fmt=None); rr += 1
hdr(wsn, rr, [''] + [f'g={g:.0%}' for g in SN['g_grid']]); rr += 1
for j, wt in enumerate(SN['wt_grid']):
    put(wsn, f'A{rr}', f'WACC_t={wt:.2%}', fmt=None)
    for i, val in enumerate(SN['grid_wacc_g'][j]):
        if val is None:
            put(wsn, f'{get_column_letter(2+i)}{rr}', 'n.m.', BLUE, None)
        else:
            put(wsn, f'{get_column_letter(2+i)}{rr}', val, BLUE, PX)
    rr += 1
rr += 1
for label, grid, axis in [('Beta', SN['grid_beta'], SN['beta_grid']),
                          ('Metal price x', SN['grid_metal'], SN['metal_grid']),
                          ('Sustained gross margin', SN['grid_spread'], SN['spread_grid']),
                          ('Volume x', SN['grid_vol'], SN['vol_grid']),
                          ('Net working capital / revenue', SN['grid_nwc'], SN['nwc_grid'])]:
    put(wsn, f'A{rr}', f'{label} sensitivity (DCF value/share)', bold=True, fmt=None); rr += 1
    for i, (a, val) in enumerate(zip(axis, grid)):
        put(wsn, f'A{rr}', str(a), fmt=None); put(wsn, f'B{rr}', val, BLUE, PX); rr += 1
    rr += 1

# =========================================================================
# PEER & SECTOR — context multiples (pasted)
# =========================================================================
wpe = sheet('Peer & Sector')
title(wpe, 'Peer & sector context', 'Cross-check multiples only — never a source for the subject '
      'historicals', 5, awidth=40, cwidth=15)
hdr(wpe, 4, ['Peer', 'Market', 'Fwd EV/EBITDA', 'Note'])
peers = [('Riyadh Cables (subject)', 'Tadawul', None, 'trailing; ~14.5x trailing P/E'),
         ('Prysmian', 'Milan', 8.5, 'developed-market major'),
         ('Nexans', 'Paris', 7.5, 'developed-market major'),
         ('Polycab India', 'NSE', 26.0, 'high-growth EM premium'),
         ('KEI Industries', 'NSE', 24.0, 'high-growth EM premium'),
         ('Ducab (private)', 'UAE', None, 'regional peer, not listed')]
rr = 5
for nm, mk, mult, note in peers:
    put(wpe, f'A{rr}', nm, fmt=None); put(wpe, f'B{rr}', mk, fmt=None)
    if nm.startswith('Riyadh Cables'):
        putf(wpe, f'C{rr}', f"='Per-Share & Ratios'!B{ANCH['per_share']['evebitda']}", REL['ev_ebitda_trailing'], MULT, green=True)
    elif mult is not None:
        put(wpe, f'C{rr}', mult, BLUE, MULT)
    put(wpe, f'D{rr}', note, fmt=None); rr += 1
put(wpe, f'A{rr+1}', f"Justified EV/EBITDA applied in the relative lens: {IN['ev_ebitda_just']:.1f}x "
    f"(mid-range, single-country discount).", fmt=None)

# =========================================================================
# SUMMARY — valuation at a glance (links to Fundamental Valuation)
# =========================================================================
wsu = sheet('Summary')
title(wsu, f"Testahil — Riyadh Cables Group Company (Tadawul: 4142)", 'Independent valuation study · '
      'educational · not investment advice', 7, awidth=44, cwidth=14)
rr = 4
hdr(wsu, rr, ['Lens', 'Bear', 'Base', 'Bull', 'Role', '', 'vs price']); rr += 1
row_su0 = rr
for k in ['dcf', 'relative', 'normalized', 'book']:
    li = ANCH['fund']['lens0'] + ['dcf', 'relative', 'normalized', 'book'].index(k)
    put(wsu, f'A{rr}', LN[k]['name'], fmt=None)
    putf(wsu, f'B{rr}', f"='Fundamental Valuation'!B{li}", LN[k]['bear'], PX, green=True)
    putf(wsu, f'C{rr}', f"='Fundamental Valuation'!C{li}", LN[k]['base'], PX, green=True)
    putf(wsu, f'D{rr}', f"='Fundamental Valuation'!D{li}", LN[k]['bull'], PX, green=True)
    put(wsu, f'E{rr}', ROLE[k], fmt=None)
    putf(wsu, f'G{rr}', f'=C{rr}/{AC("Spot price (SAR)")}-1', LN[k]['base'] / SPOT - 1, PCT)
    rr += 1
band(wsu, rr, 7)
put(wsu, f'A{rr}', 'THE CENTRAL — the cash-flow lens, not an average', bold=True, fmt=None)
putf(wsu, f'C{rr}', f"='Fundamental Valuation'!C{ANCH['fund']['cent']}", D['central'], PX, bold=True, green=True)
putf(wsu, f'G{rr}', f'=C{rr}/{AC("Spot price (SAR)")}-1', D['central'] / SPOT - 1, PCT, bold=True); rr += 1
put(wsu, f'A{rr}', 'Terminal value as % of DCF EV', fmt=None); putf(wsu, f'C{rr}', f'=DCF!C{row_tvshare}', DCF['tv_share'], PCT, green=True); rr += 1
put(wsu, f'A{rr}', 'Expert panel median', fmt=None); putf(wsu, f'C{rr}', f"='Fundamental Valuation'!C{ANCH['fund']['panel']}", D['panel_centre'], PX, green=True); rr += 1
band(wsu, rr, 7); put(wsu, f'A{rr}', 'Market price (anchor)', bold=True, fmt=None); putf(wsu, f'C{rr}', f'={AC("Spot price (SAR)")}', SPOT, PX, bold=True, green=True); rr += 1
ANCH['summary_mktcap'] = f'C{rr}'
put(wsu, f'A{rr}', 'Market capitalisation (SAR mn)', fmt=None); putf(wsu, f'C{rr}', f'={AC("Spot price (SAR)")}*{AC("Shares outstanding (mn)")}', MKTCAP, NUM0, green=True); rr += 1

# =========================================================================
# READ FIRST
# =========================================================================
wrd = sheet('READ FIRST')
title(wrd, 'Testahil — Riyadh Cables Group Company (Tadawul: 4142)', None, 9)
readfirst = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the Riyadh Cables valuation study. Every blue cell is an',
 'input; every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure derivable from a driver is a live formula, so you can change a blue cell',
 'on Assumptions and watch the model reprice: the cost of capital is built from the risk-free rate net of the',
 'sovereign spread, beta and the premium; the discount factors compound from the cost-of-debt glide; the DCF',
 'waterfall, the terminal block, the statement roll-forwards, all four lenses and every ratio chain off the',
 'same cells.', '',
 'THREE CLASSES OF CELL ARE PASTED VALUES, and it is worth knowing exactly which. First, audited and disclosed',
 'history — the primary record (FY2023-25 income statement and balance sheet, the FY2025 cost-stack base the',
 'ground-up build starts from), not a calculation. Second, there is NO flattened unit-build paste: the cable',
 'cost stack is built live on the Segments sheet from the disclosed FY2025 materials and conversion figures and',
 'the driver paths. Third, whole-model engine outputs, where each figure is a complete re-run of the valuation',
 'and so cannot be one formula: the Monte Carlo price map, the sensitivity grids, the DCF scenario bear/bull',
 'bounds, the two pasted framings of the contested gross-margin judgement, and the expert-panel median.',
 'Everything else — every lens base value, the relative/normalised/book bear and bull bounds, and the anchor',
 'roll — is a live formula. Changing a driver reprices the model but does NOT redraw the engine outputs.', '',
 'How revenue is built. Not as one growth rate. A cable maker is a metal converter: materials (copper and',
 'aluminium) are 94.9% of cost of revenue. The build prices a tonnage index as metal content — on its own',
 'commodity path — plus a conversion spread whose cost escalates on domestic inflation; gross margin is the',
 'OUTPUT of that stack, not an input. The central judgement is the margin the business sustains once the',
 'FY2024-25 metal tailwind has passed, anchored on the most recent reviewed actual (H1-2026, 15.26%) rather',
 'than the higher FY2025 full-year print (16.24%), and computed both ways.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges.', '',
 'Sourcing. FY2022-FY2025 come from the company\'s own audited consolidated financial statements (KPMG); the',
 'FY2026 near-term anchor is the Tadawul-filed reviewed H1-2026 interim. Every input is listed with source and',
 'date in the companion bibliography document. No aggregator or broker figure is a build source.', '',
 f"Currency. SAR million unless stated. Spot SAR {SPOT:.2f} ({M['asof']} close). Sheets: READ FIRST · Summary ·",
 'Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income',
 'Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share &',
 'Ratios · Peer & Sector.']
for i, ln in enumerate(readfirst, start=3):
    wrd.cell(row=i, column=1, value=ln).font = Font(size=10)
wrd.column_dimensions['A'].width = 116

# ---- reorder to the required 16-sheet order + save + expected ledger ----------
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge', 'Segments',
         'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet', 'Cash Flow',
         'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios', 'Peer & Sector']
wb._sheets.sort(key=lambda ws: ORDER.index(ws.title))
assert [ws.title for ws in wb._sheets] == ORDER, "sheet order wrong"

# Fit every sheet's columns onto one page width (landscape) so a printed/PDF reader never sees a
# value column orphaned from its labels; height is left to flow across pages as needed.
from openpyxl.worksheet.properties import PageSetupProperties
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = False

OUT_XLSX = os.path.join(HERE, 'RIYADHCABLE_Valuation_Model_18082026_public.xlsx')
wb.save(OUT_XLSX)
ANCH['seg_rev_tot'] = seg['rev']; ANCH['seg_ebitda'] = seg['ebitda']
json.dump({'expected': EXPECT, 'anchors': {k: (v if not isinstance(v, dict) else v) for k, v in ANCH.items() if k in ('summary_mktcap', 'ev', 'tv_share', 'dcf_ps')}},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1, default=str)
nform = sum(1 for _ in range(0))
from collections import Counter
cnt = Counter()
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                cnt['formula'] += 1
            elif isinstance(c.value, (int, float)):
                cnt['value'] += 1
nexp = sum(len(v) for v in EXPECT.values())
print(f"WROTE {os.path.basename(OUT_XLSX)} | 16 sheets | formulas {cnt['formula']} · numeric values "
      f"{cnt['value']} | expected-ledger cells {nexp}")
try:
    os.remove(os.path.join(HERE, '_wb_partial.xlsx'))
except OSError:
    pass
