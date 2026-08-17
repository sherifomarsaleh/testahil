"""ADNOC Drilling — build the delivered valuation workbook.

THE WORKBOOK CALCULATES. Everything arithmetically derivable from a driver is a
live formula: the cost of capital is built in the sheet from the risk-free rate
net of the sovereign spread, the beta and the premium; the cost of debt is built
from the five-year Treasury and the company's own facility margin and then taxed;
the weights come from net debt and market capitalisation; the discount factors
compound; the waterfall chains margin to EBITDA to EBIT to NOPAT to free cash
flow to present value; the terminal block derives its reinvestment rate from
growth over return on capital and its value from terminal NOPAT; the statements
roll forward; every ratio and per-share figure is a formula.

Only three classes of cell are pasted, and READ FIRST names them:
  1. audited and disclosed history — and where a line is both disclosed and
     derivable, the DISCLOSED figure is carried;
  2. the output of the unit build that would be unreadable flattened into a grid
     — here, the rig-count plan, which is a fleet schedule, not arithmetic;
  3. whole-model re-runs — the Monte Carlo map and the sensitivity grids, where
     each cell is a complete revaluation and which therefore do NOT redraw when a
     driver changes.

As it writes, this builder records the model's own value for every formula cell
into xlsx_expected.json. recalc.py then evaluates the DELIVERED file
independently and asserts that every one of those cells reproduces it, with none
left unchecked.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
ST = json.load(open(os.path.join(HERE, 'strike_result.json')))
EX = json.load(open(os.path.join(HERE, 'experts.json')))
TA = json.load(open(os.path.join(HERE, 'technicals.json')))
IN, W, M, H, U, UE = (D['inputs'], D['wacc'], D['market'], D['history'],
                      D['units_history'], D['unit_economics'])
CA, CB = D['cases']['A'], D['cases']['B']
REL, BOOK, NORM, SENS, FV = (D['relative'], D['book'], D['normalised'], D['sensitivity'],
                             D['fair_value'])
RA, RB = CA['rows'], CB['rows']
SH = M['shares_outstanding_k']
FX = IN['fx_aed_usd']['value']
TAXR = IN['tax_rate']['value']
FLEET = {c: {k: {int(y): v for y, v in plan.items()} for k, plan in D['fleet_plan'][c].items()}
         for c in ('A', 'B')}
CAPEX_PLAN = {c: {int(y): v for y, v in D['capex_plan'][c].items()} for c in ('A', 'B')}

OUT_XLSX = os.path.join(HERE, 'ADNOCDRILL_Valuation_Model_09082026.xlsx')

# ---------------------------------------------------------------- styling ----
INK, GOLD, SAGE, BAND = '1C3A36', 'C0A45F', '9FB0AC', 'F3EFE4'
F_TITLE = Font(name='Calibri', size=14, bold=True, color=INK)
F_H1 = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
F_H2 = Font(name='Calibri', size=10, bold=True, color=INK)
F_TXT = Font(name='Calibri', size=10, color='1A1A1A')
F_IN = Font(name='Calibri', size=10, color='0033CC')          # blue = input
F_FX = Font(name='Calibri', size=10, color='000000')          # black = formula
F_TOT = Font(name='Calibri', size=10, bold=True, color=INK)
FILL_H = PatternFill('solid', fgColor=INK)
FILL_B = PatternFill('solid', fgColor=BAND)
FILL_G = PatternFill('solid', fgColor='EFE7D2')
THIN = Side(style='thin', color='C8D0CE')
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

N0, N1, N2, N3 = '#,##0', '#,##0.0', '#,##0.00', '#,##0.000'
P1, P2, X2 = '0.0%', '0.00%', '0.00"x"'

wb = openpyxl.Workbook()
wb.remove(wb.active)
EXPECT = {}


def ws(name, widths):
    s = wb.create_sheet(name)
    for i, w_ in enumerate(widths, start=1):
        s.column_dimensions[get_column_letter(i)].width = w_
    EXPECT[name] = {}
    return s


def title(s, text, span):
    s['A1'] = text
    s['A1'].font = F_TITLE
    s.merge_cells(f'A1:{get_column_letter(span)}1')
    s.row_dimensions[1].height = 22


def hdr(s, row, vals, start=1):
    for j, v in enumerate(vals):
        c = s.cell(row=row, column=start + j, value=v)
        c.font = F_H1
        c.fill = FILL_H
        c.alignment = Alignment(horizontal='center' if j else 'left', wrap_text=True,
                                vertical='center')
        c.border = BOX
    s.row_dimensions[row].height = 30


def lab(s, row, text, col=1, bold=False, fill=None, wrap=False):
    c = s.cell(row=row, column=col, value=text)
    c.font = F_TOT if bold else F_TXT
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = WRAP
    return c


def put(s, coord, value, fmt=None, font=None, fill=None):
    """A PASTED cell — disclosed history, a fleet-schedule entry, or a re-run grid."""
    c = s[coord]
    c.value = value
    c.font = font or (F_IN if font is None and fill is FILL_G else F_FX)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    c.border = BOX
    return c


def inp(s, coord, value, fmt=None):
    """A blue INPUT cell on the Assumptions sheet."""
    c = s[coord]
    c.value = value
    c.font = F_IN
    c.fill = FILL_G
    if fmt:
        c.number_format = fmt
    c.border = BOX
    return c


def fml(s, coord, formula, expected, fmt=None, bold=False):
    """A LIVE FORMULA cell. `expected` is the model's own value for this cell and
    is recorded so the delivered file can be checked against the model."""
    c = s[coord]
    c.value = formula
    c.font = F_TOT if bold else F_FX
    if fmt:
        c.number_format = fmt
    c.border = BOX
    EXPECT[s.title][coord] = float(expected)
    return c


COLS = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']         # FY2023 .. FY2030E
HCOL = {2023: 'B', 2024: 'C', 2025: 'D'}
FCOL = {2026: 'E', 2027: 'F', 2028: 'G', 2029: 'H', 2030: 'I'}
YRS_H = [2023, 2024, 2025]
YRS_F = [2026, 2027, 2028, 2029, 2030]
YHDR = ['FY2023', 'FY2024', 'FY2025', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']

# ============================================================ ASSUMPTIONS ====
# Written first because every other sheet points at it.
A = ws('Assumptions', [58, 10, 16, 92])
title(A, 'Assumptions — every blue cell is an input; change one and the workbook reprices', 4)
hdr(A, 3, ['Driver', 'Unit', 'Value', 'Source and construction'])
AR = {}
r = 4


def arow(label, unit, value, source, fmt=N2, section=False):
    """A DISCLOSED or JUDGEMENT driver: a blue cell nothing in the workbook derives."""
    global r
    if section:
        c = lab(A, r, label, bold=True, fill=FILL_B)
        for cc in range(2, 5):
            A.cell(row=r, column=cc).fill = FILL_B
        r += 1
        return
    lab(A, r, label)
    A.cell(row=r, column=2, value=unit).font = F_TXT
    inp(A, f'C{r}', value, fmt)
    c = A.cell(row=r, column=4, value=source)
    c.font = F_TXT
    c.alignment = WRAP
    AR[label] = r
    r += 1


def afml(label, unit, formula, expected, source, fmt=N2):
    """A DERIVED driver: a live formula built out of the disclosed cells above it.

    Every rate, ratio and margin the model computes rather than reads is written
    this way. Perturb any disclosed cell the formula reaches and this cell moves,
    and so does everything downstream of it — which is what the driver test
    exists to prove, one input at a time.
    """
    global r
    lab(A, r, label)
    A.cell(row=r, column=2, value=unit).font = F_TXT
    fml(A, f'C{r}', formula, expected, fmt)
    c = A.cell(row=r, column=4, value=source)
    c.font = F_TXT
    c.alignment = WRAP
    AR[label] = r
    r += 1


def SRC(k):
    return IN[k]['source']


def AC(label):
    return f"Assumptions!$C${AR[label]}"


# A handful of Assumptions rows are formulas over cells on sheets written LATER —
# last-twelve-month EBITDA off the income statement, the sustainable return off
# the forecast. They are placed here in reading order and their formula text is
# filled in once those rows exist. Nothing is left as a pasted number: DEFERRED is
# drained at the end of the build and asserted empty.
DEFERRED = []


def adef(label, unit, expected, source, fmt=N2):
    """An Assumptions row whose formula is supplied later by resolve_deferred()."""
    global r
    lab(A, r, label)
    A.cell(row=r, column=2, value=unit).font = F_TXT
    DEFERRED.append((label, f'C{r}', float(expected), fmt))
    c = A.cell(row=r, column=4, value=source)
    c.font = F_TXT
    c.alignment = WRAP
    AR[label] = r
    r += 1


def resolve_deferred(label, formula):
    for i, (lbl, coord, expected, fmt) in enumerate(DEFERRED):
        if lbl == label:
            fml(A, coord, formula, expected, fmt)
            DEFERRED.pop(i)
            return
    raise KeyError(f'no deferred Assumptions row called {label!r}')


arow('MARKET AND SHARE COUNT', '', 0, '', section=True)
arow('Market price', 'AED/sh', IN['spot_aed']['value'], SRC('spot_aed'), N2)
arow('AED per USD', 'AED/USD', FX, SRC('fx_aed_usd'), N3)
arow('Shares issued', "'000", IN['shares_issued_k']['value'], SRC('shares_issued_k'), N0)
arow('Treasury shares held by the market maker', "'000", IN['treasury_shares_k']['value'],
     SRC('treasury_shares_k'), N0)

arow('COST OF CAPITAL', '', 0, '', section=True)
arow('US 10-year Treasury yield (observed)', '%', IN['ust10']['value'], SRC('ust10'), P2)
arow('US adjusted default spread', '%', IN['us_default_spread']['value'],
     SRC('us_default_spread'), P2)
arow('Equity beta', 'x', IN['beta_raw']['value'], SRC('beta_raw'), N3)
arow('Equity risk premium — rating basis', '%', IN['erp_rating']['value'], SRC('erp_rating'), P2)
arow('Equity risk premium — credit-default-swap basis', '%', IN['erp_cds']['value'],
     SRC('erp_cds'), P2)
arow('US 5-year Treasury yield', '%', IN['ust5']['value'], SRC('ust5'), P2)
arow('Borrowing margin over Term SOFR on the latest facility', '%',
     IN['facility_margin']['value'], SRC('facility_margin'), P2)
arow('Secured Overnight Financing Rate, spot', '%', IN['sofr_spot']['value'], SRC('sofr_spot'), P2)
arow('Abu Dhabi sovereign credit-default-swap spread', '%', IN['ad_cds']['value'], SRC('ad_cds'),
     P2)
arow('Corporate income tax rate', '%', TAXR, SRC('tax_rate'), P2)

FLEET_NICE = (('onshore_ad', 'Abu Dhabi onshore rigs, year end'),
              ('regional', 'Regional onshore rigs, year end'),
              ('jackup', 'Jack-up rigs, year end'),
              ('island', 'Island rigs, year end'),
              ('ids', 'Integrated-services rigs, year end'),
              ('discrete', 'Rigs given at least one discrete service, year end'))
FLEET_SRC = {
    'A': {'onshore_ad': 'Abu Dhabi onshore fleet held at the FY2025 disclosed 92 rigs and grown '
                        'two a year thereafter under the continued-expansion case',
          'regional': 'Opens at the 30 regional rigs disclosed at 30 June 2026 (8 from the SLDC '
                      'joint venture with SLB, 22 from MBPS) and adds two a year',
          'jackup': 'Held flat at the FY2025 disclosed 36; no jack-ups are on order',
          'island': 'FY2025 disclosed 12, plus the six island rigs the company has ordered for '
                    'gradual delivery between 2026 and 2028',
          'ids': "The company's own target of approximately 70 integrated-services rigs by the "
                 'end of 2026, extended at four a year and tapering',
          'discrete': 'The second rig population Oilfield Services serves and the company '
                      'discloses separately: 48 at end-2024, 58 at end-2025, 53 at 30 June 2026. '
                      'Held broadly flat, because it has not grown as the integrated fleet has'},
    'B': {'onshore_ad': 'Held flat at the FY2025 disclosed 92 rigs: the domestic fleet stops '
                        'growing once the production-capacity target is met',
          'regional': 'Held flat at the 30 regional rigs disclosed at 30 June 2026',
          'jackup': 'Held flat at the FY2025 disclosed 36',
          'island': 'The six ordered island rigs still arrive — they are already paid for — but '
                    'nothing follows them',
          'ids': 'Stops just above the stated end-2026 target of approximately 70',
          'discrete': 'Held flat at the disclosed count on the same reasoning as the expansion '
                      'case: this population is not growing'},
}
CASE_TAG = {'A': '', 'B': ' (plateau)'}
for case in ('A', 'B'):
    label = ('CONTINUED-EXPANSION CASE' if case == 'A' else 'CAPACITY-PLATEAU CASE')
    arow(f'FLEET PLAN — {label} (a delivery schedule, not arithmetic)', '', 0, '', section=True)
    for key, nice in FLEET_NICE:
        for y in YRS_F:
            v = FLEET[case][key][y]
            arow(f'{nice}{CASE_TAG[case]} — {y}', 'rigs', float(v), FLEET_SRC[case][key], N0)

arow('DISCLOSED BASE FOR THE UNIT RATES (reported figures, nothing derived)', '', 0, '',
     section=True)
arow('Onshore segment revenue — FY2025', "USD '000", H['2025']['seg_onshore'],
     SRC('seg_on_fy25'), N0)
arow('Offshore segment revenue — FY2025', "USD '000", H['2025']['seg_offshore'],
     SRC('seg_off_fy25'), N0)
arow('Oilfield Services segment revenue — FY2024', "USD '000", H['2024']['seg_ofs'],
     SRC('seg_ofs_fy24'), N0)
arow('Oilfield Services segment revenue — FY2025', "USD '000", H['2025']['seg_ofs'],
     SRC('seg_ofs_fy25'), N0)
arow('Offshore Island revenue — FY2023', "USD '000", IN['seg_island_fy23']['value'],
     SRC('seg_island_fy23'), N0)
arow('Offshore Jack-up revenue — FY2023', "USD '000", IN['seg_jackup_fy23']['value'],
     SRC('seg_jackup_fy23'), N0)
arow('Unconventional revenue — FY2025', "USD '000", IN['unconv_fy25']['value'],
     SRC('unconv_fy25'), N0)
arow('Unconventional booked to Oilfield Services — FY2024', "USD '000",
     IN['unconv_ofs_fy24']['value'], SRC('unconv_ofs_fy24'), N0)
arow('Unconventional booked to Oilfield Services — FY2025', "USD '000",
     IN['unconv_ofs_fy25']['value'], SRC('unconv_ofs_fy25'), N0)
for _k, _nice in (('rigs_onshore_fy24', 'Abu Dhabi onshore rigs — FY2024 year end'),
                  ('rigs_onshore_fy25', 'Abu Dhabi onshore rigs — FY2025 year end'),
                  ('rigs_jackup_fy22', 'Jack-up rigs — FY2022 year end'),
                  ('rigs_jackup_fy23', 'Jack-up rigs — FY2023 year end'),
                  ('rigs_jackup_fy24', 'Jack-up rigs — FY2024 year end'),
                  ('rigs_jackup_fy25', 'Jack-up rigs — FY2025 year end'),
                  ('rigs_island_fy23', 'Island rigs — FY2023 year end'),
                  ('rigs_island_fy24', 'Island rigs — FY2024 year end'),
                  ('rigs_island_fy25', 'Island rigs — FY2025 year end'),
                  ('ids_fy24', 'Integrated-services rigs — FY2024 year end'),
                  ('ids_fy25', 'Integrated-services rigs — FY2025 year end'),
                  ('discrete_fy24', 'Rigs given at least one discrete service — FY2024'),
                  ('discrete_fy25', 'Rigs given at least one discrete service — FY2025'),
                  ('rigs_regional_2q26', 'Regional onshore rigs — 30 June 2026'),
                  ('rigs_island_2q26', 'Island rigs — 30 June 2026'),
                  ('ids_2q26', 'Integrated-services rigs — 30 June 2026'),
                  ('discrete_2q26', 'Rigs given at least one discrete service — 30 June 2026'),
                  ):
    arow(_nice, 'rigs' if 'Wells' not in _nice else 'wells', IN[_k]['value'], SRC(_k), N0)

arow('UNIT REVENUE RATES (every one derived from the block above)', '', 0, '', section=True)
afml('Revenue per Abu Dhabi onshore rig-year', "USD '000",
     f'=({AC("Onshore segment revenue — FY2025")}'
     f'-({AC("Unconventional revenue — FY2025")}'
     f'-{AC("Unconventional booked to Oilfield Services — FY2025")}))'
     f'/(({AC("Abu Dhabi onshore rigs — FY2024 year end")}'
     f'+{AC("Abu Dhabi onshore rigs — FY2025 year end")})/2)',
     U['2025']['rev_per_onshore_rig'],
     'FY2025 reported Onshore segment revenue less the unconventional land-drilling revenue '
     'inside it, over the average of the FY2024 and FY2025 reported rig counts', N0)
arow('Revenue per regional onshore rig-year', "USD '000", IN['rev_per_rig_regional']['value'],
     SRC('rev_per_rig_regional'), N0)
afml('Island-to-jack-up revenue ratio', 'x',
     f'=({AC("Offshore Island revenue — FY2023")}/{AC("Island rigs — FY2023 year end")})'
     f'/({AC("Offshore Jack-up revenue — FY2023")}'
     f'/(({AC("Jack-up rigs — FY2022 year end")}+{AC("Jack-up rigs — FY2023 year end")})/2))',
     UE['island_to_jackup_ratio'],
     'FY2023 is the last year the island and jack-up businesses were reported separately, so it '
     'is the only year that can fix the ratio between them. Revenue per island rig-year over '
     'revenue per jack-up rig-year, both on that year', N3)
afml('Revenue per jack-up rig-year', "USD '000",
     f'={AC("Offshore segment revenue — FY2025")}'
     f'/((({AC("Jack-up rigs — FY2024 year end")}+{AC("Jack-up rigs — FY2025 year end")})/2)'
     f'+{AC("Island-to-jack-up revenue ratio")}'
     f'*(({AC("Island rigs — FY2024 year end")}+{AC("Island rigs — FY2025 year end")})/2))',
     UE['rev_per_jackup_fy25'],
     'FY2025 reported Offshore segment revenue divided across jack-up-equivalent rig-years, '
     'island rigs counted at the ratio above', N0)
afml('Revenue per island rig-year', "USD '000",
     f'={AC("Revenue per jack-up rig-year")}*{AC("Island-to-jack-up revenue ratio")}',
     UE['rev_per_island_fy25'], 'The jack-up rate at the FY2023 island-to-jack-up ratio', N0)
afml('Rigs served by Oilfield Services — FY2024', 'rigs',
     f'={AC("Integrated-services rigs — FY2024 year end")}'
     f'+{AC("Rigs given at least one discrete service — FY2024")}',
     UE['ofs_served']['2024'],
     'Oilfield Services serves TWO rig populations, and the company discloses both: the '
     'integrated-services fleet and, separately, the rigs given at least one discrete service. '
     'The first edition of this model recognised only the first', N0)
afml('Rigs served by Oilfield Services — FY2025', 'rigs',
     f'={AC("Integrated-services rigs — FY2025 year end")}'
     f'+{AC("Rigs given at least one discrete service — FY2025")}',
     UE['ofs_served']['2025'], 'The same two populations a year later', N0)
afml('Revenue per rig served — FY2024', "USD '000",
     f'=({AC("Oilfield Services segment revenue — FY2024")}'
     f'-{AC("Unconventional booked to Oilfield Services — FY2024")})'
     f'/{AC("Rigs served by Oilfield Services — FY2024")}',
     UE['ofs_rev_per_served']['2024'],
     'FY2024 reported Oilfield Services revenue less the unconventional revenue inside it, over '
     'the rigs served', N0)
afml('Revenue per rig served — FY2025', "USD '000",
     f'=({AC("Oilfield Services segment revenue — FY2025")}'
     f'-{AC("Unconventional booked to Oilfield Services — FY2025")})'
     f'/{AC("Rigs served by Oilfield Services — FY2025")}',
     UE['ofs_rev_per_served']['2025'], 'The same construction a year later', N0)
afml('Realised growth in revenue per rig served', '%',
     f'={AC("Revenue per rig served — FY2025")}/{AC("Revenue per rig served — FY2024")}-1',
     UE['ofs_intensity_realised'],
     'Measured, not assumed: what the company actually achieved between the two disclosed years. '
     'It fades to the contract escalator across the forecast, because a gain of this size in '
     'revenue per rig served is a mix and efficiency effect that cannot compound indefinitely',
     P1)

arow('ESCALATORS — ONE PER DRIVER CLASS', '', 0, '', section=True)
arow('Contract day-rate escalation', '%', IN['esc_dayrate']['value'], SRC('esc_dayrate'), P2)
arow('Wage escalation (domestic labour only)', '%', IN['esc_wages']['value'], SRC('esc_wages'), P2)
arow('Oilfield-services cost escalation', '%', IN['esc_oilfield']['value'], SRC('esc_oilfield'), P2)
arow('Fuel escalation (own commodity path)', '%', IN['esc_fuel']['value'], SRC('esc_fuel'), P2)
arow('General escalation', '%', IN['esc_general']['value'], SRC('esc_general'), P2)

arow('Unconventional EBITDA margin', '%', IN['unconv_ebitda_margin']['value'],
     SRC('unconv_ebitda_margin'), P2)

arow('FY2025 DIRECT COST — AS REPORTED IN THE COST NOTE', '', 0, '', section=True)
_dc_first = r
CS = UE['conventional_cost_stack_fy25']
CS_LABEL = {'repairs': 'Repairs and maintenance', 'staff': 'Staff costs',
            'hire': 'Hire of equipment', 'chemicals': 'Chemicals',
            'fuel': 'Fuel and lubricants', 'major_maintenance': 'Major maintenance charges',
            'other': 'Other direct cost'}
CS_KEY = {'repairs': 'dc_repairs_fy25', 'staff': 'dc_staff_fy25', 'hire': 'dc_hire_fy25',
          'chemicals': 'dc_chem_fy25', 'fuel': 'dc_fuel_fy25',
          'major_maintenance': 'dc_majmaint_fy25', 'other': 'dc_other_fy25'}
for k, nice in CS_LABEL.items():
    arow(f'{nice} — as reported', "USD '000", IN[CS_KEY[k]]['value'], SRC(CS_KEY[k]), N0)
_dc_last = r - 1
_DC_RANGE = f'Assumptions!$C${_dc_first}:$C${_dc_last}'
arow('General and administrative expenses — as reported', "USD '000", IN['gna_fy25']['value'],
     SRC('gna_fy25'), N0)
arow('Depreciation inside general and administrative expenses', "USD '000",
     IN['gna_dep_fy25']['value'], SRC('gna_dep_fy25'), N0)
afml('Direct cost carried by the unconventional programme', "USD '000",
     f'={AC("Unconventional revenue — FY2025")}*(1-{AC("Unconventional EBITDA margin")})',
     UE['unconventional_cost_fy25'],
     'FY2025 unconventional revenue at one minus the unconventional EBITDA margin — the cash '
     'cost the unconventional programme carried, which has to come out of the reported direct '
     'cost before the remainder can be escalated on conventional volume drivers', N0)

arow('FY2025 CONVENTIONAL COST STACK (reported cost less the unconventional share)', '', 0, '',
     section=True)
for k, nice in CS_LABEL.items():
    afml(f'{nice} — conventional base', "USD '000",
         f'={AC(f"{nice} — as reported")}'
         f'*(1-{AC("Direct cost carried by the unconventional programme")}'
         f'/SUM({_DC_RANGE}))',
         CS[k],
         'The reported line less its pro-rata share of the direct cost the unconventional '
         'programme carried', N0)
afml('General and administrative expenses excluding depreciation', "USD '000",
     f'={AC("General and administrative expenses — as reported")}'
     f'-{AC("Depreciation inside general and administrative expenses")}',
     IN['gna_fy25']['value'] - IN['gna_dep_fy25']['value'],
     'FY2025 reported general and administrative expenses less the depreciation and '
     'amortisation inside them, which is charged separately in the model', N0)
arow('Other income, FY2025 base', "USD '000", IN['othinc_fy25']['value'], SRC('othinc_fy25'), N0)
arow('Share of joint-venture results, FY2025 base', "USD '000", IN['jv_fy25']['value'],
     SRC('jv_fy25'), N0)

arow('UNCONVENTIONAL PROGRAMME', '', 0, '', section=True)
for y in YRS_F:
    arow(f'Unconventional revenue — {y}', "USD '000", FLEET['A']['unconv'][y],
         'Phased off the disclosed $0.86 billion of remaining Phase 1 contract value at end-2025 '
         'and the $206 million recognised in the first half of 2026', N0)
for y in YRS_F:
    arow(f'Unconventional revenue (plateau) — {y}', "USD '000", FLEET['B']['unconv'][y],
         'The same programme with no successor contract: Phase 1 runs off and nothing replaces '
         'it', N0)
afml('Unconventional share booked to Onshore', '%',
     f'=({AC("Unconventional revenue — FY2025")}'
     f'-{AC("Unconventional booked to Oilfield Services — FY2025")})'
     f'/{AC("Unconventional revenue — FY2025")}',
     (IN['unconv_fy25']['value'] - IN['unconv_ofs_fy25']['value']) / IN['unconv_fy25']['value'],
     'The FY2025 reported split, carried forward: the residual of total unconventional revenue '
     'over the part reported inside Oilfield Services is the part reported inside Onshore', P2)

arow('CAPITAL AND WORKING CAPITAL', '', 0, '', section=True)
for y in YRS_F:
    arow(f'Capital expenditure — {y}', "USD '000", CAPEX_PLAN['A'][y],
         'FY2026 is the midpoint of the guided $0.6-0.8 billion; later years step down toward '
         'the stated maintenance level of around $250 million as the ordered island rigs are '
         'delivered', N0)
for y in YRS_F:
    arow(f'Capital expenditure (plateau) — {y}', "USD '000", CAPEX_PLAN['B'][y],
         'The same FY2026 guidance, stepping down faster once the ordered rigs are delivered and '
         'nothing follows them', N0)
arow('Depreciation and amortisation — FY2025 as reported', "USD '000", IN['dna_fy25']['value'],
     SRC('dna_fy25'), N0)
arow('Property and equipment — FY2024 as reported', "USD '000", IN['ppe_fy24']['value'],
     SRC('ppe_fy24'), N0)
arow('Right-of-use assets — FY2024 as reported', "USD '000", IN['rou_fy24']['value'],
     SRC('rou_fy24'), N0)
arow('Intangible assets — FY2024 as reported', "USD '000", IN['intang_fy24']['value'],
     SRC('intang_fy24'), N0)
afml('Depreciation rate on opening fixed assets', '%',
     f'={AC("Depreciation and amortisation — FY2025 as reported")}'
     f'/({AC("Property and equipment — FY2024 as reported")}'
     f'+{AC("Right-of-use assets — FY2024 as reported")}'
     f'+{AC("Intangible assets — FY2024 as reported")})',
     UE['depreciation_rate'],
     'FY2025 reported depreciation and amortisation over the opening (FY2024) property, '
     'equipment, right-of-use assets and intangibles it was charged on', P2)

arow('WORKING CAPITAL — THE 30 JUNE 2026 BALANCE SHEET, WHICH IS THE ONLY ONE THAT '
     'CONSOLIDATES THE ACQUISITIONS', '', 0, '', section=True)
for _k, _nice in (('inv_1h26', 'Inventories — 30 June 2026'),
                  ('recv_1h26', 'Trade and other receivables — 30 June 2026'),
                  ('dfrp_1h26', 'Due from related parties — 30 June 2026'),
                  ('tp_1h26', 'Trade and other payables — 30 June 2026'),
                  ('dtrp_1h26', 'Due to related parties — 30 June 2026'),
                  ('rev_1h26', 'Revenue — first half of 2026')):
    arow(_nice, "USD '000", IN[_k]['value'], SRC(_k), N0)
afml('Working capital — 30 June 2026', "USD '000",
     f'={AC("Inventories — 30 June 2026")}'
     f'+{AC("Trade and other receivables — 30 June 2026")}'
     f'+{AC("Due from related parties — 30 June 2026")}'
     f'-{AC("Trade and other payables — 30 June 2026")}'
     f'-{AC("Due to related parties — 30 June 2026")}',
     UE['wc_1h26'], 'The same five lines the audited years are measured on', N0)
afml('Working capital / revenue', '%',
     f'={AC("Working capital — 30 June 2026")}/({AC("Revenue — first half of 2026")}*2)',
     UE['wc_pct_revenue'],
     'Working capital at 30 June 2026 over annualised first-half revenue. The first edition used '
     'the average of the three audited years, every one of which pre-dates the two regional '
     'acquisitions, and applied it to a revenue line that includes them. A mid-year balance sheet '
     'is not seasonally flattering here: over the first half of 2025 working capital RELEASED '
     'cash, so this company runs BELOW its year end at mid-year rather than above it', P2)
arow("FY2026 GUIDANCE — THE COMPANY GUIDES ALL THREE SEGMENTS, SO ALL THREE ARE RECONCILED",
     '', 0, '', section=True)
arow('FY2026 guided Onshore revenue', "USD '000", IN['g26_rev_onshore']['value'],
     SRC('g26_rev_onshore'), N0)
arow('FY2026 guided Offshore revenue', "USD '000", IN['g26_rev_offshore']['value'],
     SRC('g26_rev_offshore'), N0)
arow('FY2026 guided Oilfield Services revenue', "USD '000", IN['g26_rev_ofs']['value'],
     SRC('g26_rev_ofs'), N0)

arow('Dividend, FY2026 floor', "USD '000", IN['g26_dividend']['value'], SRC('g26_dividend'), N0)
arow('Dividend growth', '%', 0.05, 'The guided floor is stated as 5% above the prior year', P2)

arow('TERMINAL BLOCK', '', 0, '', section=True)
arow('Terminal growth — continued expansion', '%', IN['terminal_growth_A']['value'],
     SRC('terminal_growth_A'), P2)
arow('Terminal growth — capacity plateau', '%', IN['terminal_growth_B']['value'],
     SRC('terminal_growth_B'), P2)
arow('Terminal return on invested capital', '%', IN['terminal_roic']['value'],
     SRC('terminal_roic'), P2)

arow('THE 2026 BUSINESS COMBINATIONS (note 5 of the interim, line by line)', '', 0, '',
     section=True)
for _k, _nice in (('acq_ppe', 'Property and equipment acquired'),
                  ('acq_rou', 'Right-of-use assets acquired'),
                  ('acq_goodwill', 'Provisional goodwill'),
                  ('acq_inventories', 'Inventories acquired'),
                  ('acq_receivables', 'Trade and other receivables acquired'),
                  ('acq_payables', 'Trade and other payables assumed'),
                  ('acq_cash', 'Cash acquired'),
                  ('acq_borrowings', 'Term loans, overdraft and borrowings assumed'),
                  ('acq_leases', 'Lease liabilities assumed'),
                  ('acq_deferred_tax', 'Deferred tax liability assumed'),
                  ('acq_income_tax', 'Income tax payable assumed'),
                  ('acq_eosb', 'End-of-service benefits assumed'),
                  ('acq_contingent', 'Contingent consideration outstanding'),
                  ('acq_nci', 'Minority interests recognised'),
                  ('acq_consideration_2026', 'Cash consideration paid in 2026'),
                  ('acq_cash_returned', 'Consideration received back against the 2025 advance'),
                  ('advacq_fy25', 'The 2025 advance for acquisition of a subsidiary')):
    arow(_nice, "USD '000", IN[_k]['value'], SRC(_k), N0)

arow('THE 31 DECEMBER 2025 BALANCE SHEET THE FORECAST OPENS ON', '', 0, '', section=True)
for _k, _nice in (('debt_fy25', 'Borrowings — FY2025 audited'),
                  ('lease_fy25', 'Lease liabilities — FY2025 audited'),
                  ('eosb_fy25', "End-of-service benefits — FY2025 audited"),
                  ('dta_fy25', 'Deferred tax assets — FY2025 audited'),
                  ('advnc_fy25', 'Advances — FY2025 audited'),
                  ('jvinv_fy25', 'Investment in joint ventures — FY2025 audited'),
                  ('cash_fy25', 'Cash and cash equivalents — FY2025 audited'),
                  ('equity_fy25', 'Total equity — FY2025 audited'),
                  ('taxpay_fy25', 'Income tax payable — FY2025 audited'),
                  ('ppe_fy25', 'Property and equipment — FY2025 audited'),
                  ('rou_fy25', 'Right-of-use assets — FY2025 audited'),
                  ('intang_fy25', 'Intangible assets — FY2025 audited')):
    arow(_nice, "USD '000", IN[_k]['value'], SRC(_k), N0)

arow('BRIDGE FROM ENTERPRISE VALUE TO EQUITY (30 June 2026)', '', 0, '', section=True)
arow('Investment in joint ventures', "USD '000", IN['jvinv_1h26']['value'], SRC('jvinv_1h26'), N0)
arow('Cash and cash equivalents', "USD '000", IN['cash_1h26']['value'], SRC('cash_1h26'), N0)
arow('Borrowings', "USD '000", IN['debt_1h26']['value'], SRC('debt_1h26'), N0)
arow('Lease liabilities', "USD '000", IN['lease_1h26']['value'], SRC('lease_1h26'), N0)
arow('Financial liability over the acquired minorities', "USD '000", IN['finliab_1h26']['value'],
     SRC('finliab_1h26'),
     N0)
arow('Non-controlling interests (NOT deducted — see the note)', "USD '000",
     IN['nci_1h26']['value'],
     'Reported at 30 June 2026 and shown here for completeness, but NOT deducted in the bridge. '
     'The financial liability above is the present value of the exercise price over these same '
     'minorities, and the company has already charged a matching investment reserve of the same '
     'amount against owners equity. Deducting the liability and the minority both would charge '
     'the parent twice for one claim', N0)
arow('Net cash from operating activities — first half of 2026', "USD '000",
     IN['cfo_1h26']['value'], SRC('cfo_1h26'), N0)
arow('Capital expenditure — first half of 2026', "USD '000", IN['capex_1h26']['value'],
     SRC('capex_1h26'), N0)
arow('Days from 30 June 2026 to the price anchor', 'days', IN['days_jun26_to_anchor']['value'],
     SRC('days_jun26_to_anchor'), N0)

arow('OTHER LENSES', '', 0, '', section=True)
arow('Total equity — 30 June 2026', "USD '000", IN['equity_1h26']['value'], SRC('equity_1h26'), N0)
afml('Book equity attributable to owners', "USD '000",
     f'={AC("Total equity — 30 June 2026")}'
     f'-{AC("Non-controlling interests (NOT deducted — see the note)")}',
     BOOK['book_equity'], 'Total equity at 30 June 2026 less non-controlling interests', N0)
arow('EBITDA — first half of 2026', "USD '000", IN['ebitda_1h26']['value'], SRC('ebitda_1h26'), N0)
arow('EBITDA — first half of 2025', "USD '000", IN['ebitda_1h25']['value'], SRC('ebitda_1h25'), N0)
arow('Share of joint-venture results — first half of 2026', "USD '000", IN['jv_1h26']['value'],
     SRC('jv_1h26'), N0)
arow('Share of joint-venture results — first half of 2025', "USD '000", IN['jv_1h25']['value'],
     SRC('jv_1h25'), N0)
arow('Depreciation and amortisation — first half of 2026', "USD '000", IN['dna_1h26']['value'],
     SRC('dna_1h26'), N0)
arow('FY2026 guided revenue', "USD '000", IN['g26_revenue']['value'], SRC('g26_revenue'), N0)
arow('FY2026 guided EBITDA — bottom of the range', "USD '000", IN['g26_ebitda_lo']['value'],
     SRC('g26_ebitda_lo'), N0)
arow('FY2026 guided EBITDA — top of the range', "USD '000", IN['g26_ebitda_hi']['value'],
     SRC('g26_ebitda_hi'), N0)
afml('Normalised EBITDA margin', '%',
     f'=({AC("FY2026 guided EBITDA — bottom of the range")}'
     f'+{AC("FY2026 guided EBITDA — top of the range")})/2/{AC("FY2026 guided revenue")}',
     NORM['ebitda_margin'],
     "The company's own FY2026 guided group EBITDA margin — the midpoint of the guided EBITDA "
     'range over the guided revenue. Deliberately the GROUP margin and not the higher '
     'conventional-only margin guided separately, because the fleet priced in this lens includes '
     'the lower-margin oilfield-services book', P2)
afml('Normalised depreciation and amortisation', "USD '000",
     f'={AC("Depreciation and amortisation — first half of 2026")}*2',
     NORM['dna'],
     'The charge the fleet being priced actually carries: the first-half 2026 charge, annualised. '
     'The first edition set this halfway between maintenance capital expenditure and the 2030 '
     'charge — but the 2030 charge belongs to a larger fleet than this lens is allowed to price',
     N0)
adef('EBITDA — FY2025 audited', "USD '000", H['2025']['ebitda'],
     'Read live off the FY2025 column of the income statement in this workbook, which is itself '
     'built from the audited statements', N0)
adef('Sustainable return on equity', '%', BOOK['roe_sustainable'],
     'The return on average equity this workbook itself forecasts for FY2030, read live off the '
     'forecast statements. The first edition used the FY2024-25 realised average of 36.7%, which '
     'was earned on a smaller balance sheet than the one the model builds: putting a historical '
     'return into a perpetual formula made this lens richer than the cash-flow model four '
     'sections above it, for no reason other than that the past was measured on less capital',
     P1)
afml('Last-twelve-month EBITDA excluding joint ventures', "USD '000",
     f'={AC("EBITDA — FY2025 audited")}'
     f'+{AC("EBITDA — first half of 2026")}-{AC("EBITDA — first half of 2025")}'
     f'-({AC("Share of joint-venture results, FY2025 base")}'
     f'+{AC("Share of joint-venture results — first half of 2026")}'
     f'-{AC("Share of joint-venture results — first half of 2025")})',
     REL['applied_ebitda'],
     'Audited FY2025 plus the first half of 2026 less the first half of 2025, then less the share '
     'of joint-venture results on the same basis. Every peer multiple in the table is an '
     'enterprise value struck today over that peer\'s last twelve months, so the earnings it is '
     'applied to are on the same footing — and the joint ventures are stripped out here because '
     'their carrying value is added back on the bridge', N0)
arow('Peer median EV/EBITDA — MENA national-oil-company drillers', 'x', REL['median_mena'],
     'Median of ADES Holding and Arabian Drilling on their own latest reported EBITDA and net '
     'debt, at market prices on 07-Aug-2026', N2)
arow('Peer median EV/EBITDA — global land drillers', 'x', REL['median_land'],
     'Median of Helmerich & Payne, Nabors, Patterson-UTI and Precision Drilling', N2)
arow('Peer median EV/EBITDA — global offshore drillers', 'x', REL['median_offshore'],
     'Median of Valaris, Noble and Borr Drilling', N2)
arow('Peer median EV/EBITDA — diversified oilfield services', 'x', REL['median_ofs'],
     'Median of SLB, Halliburton, Baker Hughes and Weatherford', N2)

arow('LENS WEIGHTS', '', 0, '', section=True)
LW = FV['weights']
for k, nice in (('dcf_A', 'Weight — discounted cash flow, continued expansion'),
                ('dcf_B', 'Weight — discounted cash flow, capacity plateau'),
                ('relative', 'Weight — relative multiples'),
                ('book', 'Weight — book value and sustainable return'),
                ('normalised', 'Weight — normalised earnings power')):
    arow(nice, '%', LW[k],
         'The two discounted-cash-flow framings carry equal weight and half the total, because '
         'the study refuses to choose between them; the remaining three lenses split the rest',
         P1)


# ================================================================ SEGMENTS ===
S = ws('Segments', [52] + [13] * 8)
title(S, 'Segments — the unit build: rigs times revenue per rig, and the cost stack that '
         'serves them', 9)
hdr(S, 3, ['USD thousands unless stated'] + YHDR)

BASE = UE['base_units_fy25']
ESC_LBL = {'repairs': 'Oilfield-services cost escalation',
           'staff': 'Wage escalation (domestic labour only)',
           'hire': 'Oilfield-services cost escalation',
           'chemicals': 'Oilfield-services cost escalation',
           'fuel': 'Fuel escalation (own commodity path)',
           'major_maintenance': 'Oilfield-services cost escalation',
           'other': 'General escalation'}
UNIT_KEY = {'onshore_ad': ('Abu Dhabi onshore rigs, year end', 'avg_onshore_ad', 'onshore'),
            'regional': ('Regional onshore rigs, year end', 'avg_regional', 'regional'),
            'jackup': ('Jack-up rigs, year end', 'avg_jackup', 'jackup'),
            'island': ('Island rigs, year end', 'avg_island', 'island'),
            'ids': ('Integrated-services rigs, year end', 'avg_ids', 'ids'),
            'discrete': ('Rigs given at least one discrete service, year end', 'avg_discrete',
                         'discrete')}
UNIT_NICE = {'onshore_ad': 'Average Abu Dhabi onshore rigs',
             'regional': 'Average regional onshore rigs',
             'jackup': 'Average jack-up rigs', 'island': 'Average island rigs',
             'ids': 'Average integrated-services rigs',
             'discrete': 'Average rigs given at least one discrete service'}
RATE_LBL = {'onshore_ad': 'Revenue per Abu Dhabi onshore rig-year',
            'regional': 'Revenue per regional onshore rig-year',
            'jackup': 'Revenue per jack-up rig-year', 'island': 'Revenue per island rig-year',
            'ids': 'Revenue per integrated-services rig-year'}
REV_MK = {'onshore_ad': 'rev_onshore_ad', 'regional': 'rev_regional', 'jackup': 'rev_jackup',
          'island': 'rev_island', 'ids': 'rev_ids'}
OPEN_FLEET = {'onshore_ad': AC('Abu Dhabi onshore rigs — FY2025 year end'),
              'regional': AC('Regional onshore rigs — 30 June 2026'),
              'jackup': AC('Jack-up rigs — FY2025 year end'),
              'island': AC('Island rigs — FY2025 year end'),
              'ids': AC('Integrated-services rigs — FY2025 year end'),
              'discrete': AC('Rigs given at least one discrete service — FY2025')}
BASE_DRILL = U['2025']['onshore'] + U['2025']['offshore']


def segment_block(case, model_rows, start, with_history):
    """Emit a complete live unit build for one case and return its row map.

    Both cases are built the SAME way, from the same drivers, so the plateau
    case is a live model rather than a pasted answer: only the fleet schedule,
    the unconventional path and the capital-expenditure plan differ, and all
    three are blue cells on the Assumptions sheet.
    """
    tag = CASE_TAG[case]
    R = {}
    rr = start
    lab(S, rr, ('OPERATING UNITS (average deployed over the year)' if with_history else
                'CAPACITY-PLATEAU CASE — OPERATING UNITS'), bold=True, fill=FILL_B)
    rr += 1
    R['units'] = {}
    for fkey, (plan_lbl, mkey, hkey) in UNIT_KEY.items():
        R['units'][fkey] = rr
        lab(S, rr, UNIT_NICE[fkey])
        if with_history:
            for y in YRS_H:
                put(S, f'{HCOL[y]}{rr}', U[str(y)][hkey], N1)
        for i, y in enumerate(YRS_F):
            c = FCOL[y]
            cur = AC(f'{plan_lbl}{tag} — {y}')
            prev = (str(OPEN_FLEET[fkey]) if i == 0
                    else AC(f'{plan_lbl}{tag} — {YRS_F[i-1]}'))
            fml(S, f'{c}{rr}', f'=({prev}+{cur})/2', model_rows[i][mkey], N1)
        rr += 1
    R['index'] = rr
    lab(S, rr, 'Day-rate escalation index (2025 = 1.00)')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{rr}', 1.0, N3)
    for i, y in enumerate(YRS_F):
        fml(S, f'{FCOL[y]}{rr}', f'=(1+{AC("Contract day-rate escalation")})^{i+1}',
            model_rows[i]['day_rate_index'], N3)
    rr += 2

    R['ofs_served'] = rr
    lab(S, rr, 'Rigs served by Oilfield Services (integrated plus discrete)')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{rr}', U[str(y)]['ids'] + U[str(y)].get('discrete_avg', 0.0), N1)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={c}{R["units"]["ids"]}+{c}{R["units"]["discrete"]}',
            model_rows[i]['avg_ids'] + model_rows[i]['avg_discrete'], N1)
    rr += 1
    R['ofs_index'] = rr
    lab(S, rr, 'Oilfield-services revenue-per-rig-served index (2025 = 1.00)')
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        # The step fades linearly from what the company actually achieved between
        # the two disclosed years to the contract escalator, and the index is the
        # running product of those steps.
        fade = i / (len(YRS_F) - 1)
        step = (f'({AC("Realised growth in revenue per rig served")}*{1 - fade:.6f}'
                f'+{AC("Contract day-rate escalation")}*{fade:.6f})')
        prev = '1' if i == 0 else f'{FCOL[YRS_F[i-1]]}{rr}'
        fml(S, f'{c}{rr}', f'={prev}*(1+{step})',
            model_rows[i]['ofs_intensity'], N3)
    rr += 2

    lab(S, rr, 'REVENUE BY UNIT — BEFORE RECONCILIATION TO FY2026 GUIDANCE', bold=True,
        fill=FILL_B)
    rr += 1
    R['raw'] = {}
    for fkey in ('onshore_ad', 'regional', 'jackup', 'island', 'ids'):
        R['raw'][fkey] = rr
        lab(S, rr, {'onshore_ad': 'Onshore conventional — Abu Dhabi',
                    'regional': 'Onshore conventional — regional',
                    'jackup': 'Offshore — jack-up', 'island': 'Offshore — island',
                    'ids': 'Oilfield Services — rigs served'}[fkey])
        for i, y in enumerate(YRS_F):
            c = FCOL[y]
            if fkey == 'ids':
                f = (f'={c}{R["ofs_served"]}*{AC("Revenue per rig served — FY2025")}'
                     f'*{c}{R["ofs_index"]}')
                exp = model_rows[i]['rev_ids_raw']
            else:
                f = f'={c}{R["units"][fkey]}*{AC(RATE_LBL[fkey])}*{c}{R["index"]}'
                exp = model_rows[i][REV_MK[fkey] + '_raw']
            fml(S, f'{c}{rr}', f, exp, N0)
        rr += 1
    rr += 1
    lab(S, rr, 'RECONCILIATION TO THE FY2026 SEGMENT GUIDANCE', bold=True, fill=FILL_B)
    rr += 1
    R['cal'] = {}
    E = FCOL[2026]
    _CAL_SPEC = (
        ('onshore', 'Onshore — factor on the conventional unit build',
         f'=({AC("FY2026 guided Onshore revenue")}-{E}{{unc_on}})'
         f'/({E}{R["raw"]["onshore_ad"]}+{E}{R["raw"]["regional"]})', 'calib_onshore'),
        ('offshore', 'Offshore — factor on the conventional unit build',
         f'={AC("FY2026 guided Offshore revenue")}'
         f'/({E}{R["raw"]["jackup"]}+{E}{R["raw"]["island"]})', 'calib_offshore'),
        ('ofs', 'Oilfield Services — factor on the conventional unit build',
         f'=({AC("FY2026 guided Oilfield Services revenue")}-{E}{{unc_ofs}})'
         f'/{E}{R["raw"]["ids"]}', 'calib_ofs'))
    for key, nice, _f, _mk in _CAL_SPEC:
        R['cal'][key] = rr
        lab(S, rr, nice)
        rr += 1
    R['cal_spec'] = _CAL_SPEC
    rr += 1

    R['unc_on'], R['unc_ofs'] = rr, rr + 1
    lab(S, rr, 'Unconventional — Onshore land drilling')
    lab(S, rr + 1, 'Unconventional — Oilfield Services')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{R["unc_on"]}', U[str(y)]['unconv_onshore'], N0)
            put(S, f'{HCOL[y]}{R["unc_ofs"]}', U[str(y)]['unconv_ofs'], N0)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        unc = AC(f'Unconventional revenue{tag} — {y}')
        fml(S, f'{c}{R["unc_on"]}',
            f'={unc}*{AC("Unconventional share booked to Onshore")}',
            model_rows[i]['unconv_onshore'], N0)
        fml(S, f'{c}{R["unc_ofs"]}',
            f'={unc}*(1-{AC("Unconventional share booked to Onshore")})',
            model_rows[i]['unconv_ofs'], N0)
    # The three reconciliation factors, now that the unconventional rows they net
    # off exist. Written only in the FY2026 column and referenced absolutely from
    # every later year, so the correction is a level shift on the unit rate rather
    # than a plug that is reapplied annually.
    for key, nice, f, mk in R['cal_spec']:
        fml(S, f'{E}{R["cal"][key]}',
            f.replace('{unc_on}', str(R['unc_on'])).replace('{unc_ofs}', str(R['unc_ofs'])),
            model_rows[0][mk], N3)
    rr += 3

    lab(S, rr, 'REVENUE BY REPORTED SEGMENT', bold=True, fill=FILL_B)
    rr += 1
    R['rev'] = {}
    for fkey, calkey in (('onshore_ad', 'onshore'), ('regional', 'onshore'),
                         ('jackup', 'offshore'), ('island', 'offshore'), ('ids', 'ofs')):
        R['rev'][fkey] = rr
        lab(S, rr, {'onshore_ad': 'Onshore conventional — Abu Dhabi',
                    'regional': 'Onshore conventional — regional',
                    'jackup': 'Offshore — jack-up', 'island': 'Offshore — island',
                    'ids': 'Oilfield Services — rigs served'}[fkey] + ' (reconciled)')
        for i, y in enumerate(YRS_F):
            c = FCOL[y]
            fml(S, f'{c}{rr}', f'={c}{R["raw"][fkey]}*${E}${R["cal"][calkey]}',
                model_rows[i][REV_MK[fkey]], N0)
        rr += 1
    rr += 1
    R['seg_on'], R['seg_off'], R['seg_ofs'] = rr, rr + 1, rr + 2
    R['total'], R['conv'] = rr + 3, rr + 4
    lab(S, R['seg_on'], 'Onshore', bold=True)
    lab(S, R['seg_off'], 'Offshore', bold=True)
    lab(S, R['seg_ofs'], 'Oilfield Services', bold=True)
    lab(S, R['total'], 'Total revenue', bold=True)
    lab(S, R['conv'], 'of which conventional')
    if with_history:
        for y in YRS_H:
            c = HCOL[y]
            put(S, f'{c}{R["seg_on"]}', H[str(y)]['seg_onshore'], N0, font=F_TOT)
            put(S, f'{c}{R["seg_off"]}', H[str(y)]['seg_offshore'], N0, font=F_TOT)
            put(S, f'{c}{R["seg_ofs"]}', H[str(y)]['seg_ofs'], N0, font=F_TOT)
            fml(S, f'{c}{R["total"]}',
                f'={c}{R["seg_on"]}+{c}{R["seg_off"]}+{c}{R["seg_ofs"]}',
                H[str(y)]['revenue'], N0, bold=True)
            fml(S, f'{c}{R["conv"]}',
                f'={c}{R["total"]}-{c}{R["unc_on"]}-{c}{R["unc_ofs"]}',
                H[str(y)]['revenue'] - U[str(y)]['unconv'], N0)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{R["seg_on"]}',
            f'={c}{R["rev"]["onshore_ad"]}+{c}{R["rev"]["regional"]}+{c}{R["unc_on"]}',
            model_rows[i]['seg_onshore'], N0, bold=True)
        fml(S, f'{c}{R["seg_off"]}', f'={c}{R["rev"]["jackup"]}+{c}{R["rev"]["island"]}',
            model_rows[i]['seg_offshore'], N0, bold=True)
        fml(S, f'{c}{R["seg_ofs"]}', f'={c}{R["rev"]["ids"]}+{c}{R["unc_ofs"]}',
            model_rows[i]['seg_ofs'], N0, bold=True)
        fml(S, f'{c}{R["total"]}', f'={c}{R["seg_on"]}+{c}{R["seg_off"]}+{c}{R["seg_ofs"]}',
            model_rows[i]['revenue'], N0, bold=True)
        fml(S, f'{c}{R["conv"]}', f'={c}{R["total"]}-{c}{R["unc_on"]}-{c}{R["unc_ofs"]}',
            model_rows[i]['conv_revenue'], N0)
    rr = R['conv'] + 2

    lab(S, rr, 'COST STACK — EACH LINE ON ITS OWN VOLUME DRIVER AND ITS OWN ESCALATOR',
        bold=True, fill=FILL_B)
    rr += 1
    R['drv_rig'] = rr
    lab(S, rr, 'Total rig-years (all classes, including integrated services)')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{rr}',
                U[str(y)]['onshore'] + U[str(y)]['offshore'] + U[str(y)]['ids'], N1)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={c}{R["units"]["onshore_ad"]}+{c}{R["units"]["regional"]}'
                           f'+{c}{R["units"]["jackup"]}+{c}{R["units"]["island"]}'
                           f'+{c}{R["units"]["ids"]}', model_rows[i]['units']['rig_years'], N1)
    rr += 1
    R['drv_off'] = rr
    lab(S, rr, 'Offshore rig-years')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{rr}', U[str(y)]['offshore'], N1)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={c}{R["units"]["jackup"]}+{c}{R["units"]["island"]}',
            model_rows[i]['units']['offshore_rig_years'], N1)
    rr += 1
    R['drv_well'] = rr
    lab(S, rr, 'Wells drilled (drilling rig-years scaled off the FY2025 disclosed 836)')
    if with_history:
        for y in YRS_H:
            put(S, f'{HCOL[y]}{rr}', U[str(y)]['wells'], N0)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={IN["wells_fy25"]["value"]}*({c}{R["units"]["onshore_ad"]}'
                           f'+{c}{R["units"]["regional"]}+{c}{R["units"]["jackup"]}'
                           f'+{c}{R["units"]["island"]})/{BASE_DRILL}',
            model_rows[i]['units']['wells'], N0)
    rr += 1

    drv_row = {'rig_years': R['drv_rig'], 'offshore_rig_years': R['drv_off'],
               'wells': R['drv_well'], 'conv_revenue': R['conv']}
    R['cost'] = {}
    for k, nice in CS_LABEL.items():
        R['cost'][k] = rr
        drv = UE['cost_driver'][k]
        lab(S, rr, f'{nice}  ({drv.replace("_", " ")})')
        for i, y in enumerate(YRS_F):
            c = FCOL[y]
            fml(S, f'{c}{rr}', f'={AC(f"{nice} — conventional base")}*{c}{drv_row[drv]}'
                               f'/{BASE[drv]}*(1+{AC(ESC_LBL[k])})^{i+1}',
                model_rows[i]['cost_stack'][k], N0)
        rr += 1
    R['conv_cost'] = rr
    lab(S, rr, 'Conventional cash cost', bold=True)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}',
            f'=SUM({c}{min(R["cost"].values())}:{c}{max(R["cost"].values())})',
            model_rows[i]['conv_cash_cost'], N0, bold=True)
    rr += 1
    R['unc_cost'] = rr
    lab(S, rr, 'Unconventional cash cost')
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'=({c}{R["unc_on"]}+{c}{R["unc_ofs"]})'
                           f'*(1-{AC("Unconventional EBITDA margin")})',
            model_rows[i]['unconv_cash_cost'], N0)
    rr += 1
    R['gna'] = rr
    lab(S, rr, 'General and administrative expenses excluding depreciation')
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}',
            f'={AC("General and administrative expenses excluding depreciation")}'
            f'*{c}{R["conv"]}/{BASE["conv_revenue"]}*(1+{AC("General escalation")})^{i+1}',
            model_rows[i]['gna'], N0)
    rr += 1
    R['oi'] = rr
    lab(S, rr, 'Other income')
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={AC("Other income, FY2025 base")}'
                           f'*(1+{AC("General escalation")})^{i+1}',
            model_rows[i]['other_income'], N0)
    rr += 1
    R['jv'] = rr
    lab(S, rr, 'Share of joint-venture results')
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{rr}', f'={AC("Share of joint-venture results, FY2025 base")}'
                           f'*(1+{AC("General escalation")})^{i+1}',
            model_rows[i]['jv_share'], N0)
    rr += 2
    R['ebitda'], R['margin'], R['ebitda_xjv'] = rr, rr + 1, rr + 2
    lab(S, R['ebitda'], 'EBITDA', bold=True)
    lab(S, R['margin'], 'EBITDA margin')
    lab(S, R['ebitda_xjv'], 'EBITDA excluding joint-venture results', bold=True)
    if with_history:
        for y in YRS_H:
            c = HCOL[y]
            put(S, f'{c}{R["ebitda"]}', H[str(y)]['ebitda'], N0, font=F_TOT)
            fml(S, f'{c}{R["margin"]}', f'={c}{R["ebitda"]}/{c}{R["total"]}',
                H[str(y)]['ebitda_margin'], P1)
            put(S, f'{c}{R["ebitda_xjv"]}', H[str(y)]['ebitda_ex_jv'], N0, font=F_TOT)
    for i, y in enumerate(YRS_F):
        c = FCOL[y]
        fml(S, f'{c}{R["ebitda"]}',
            f'={c}{R["total"]}-{c}{R["conv_cost"]}-{c}{R["unc_cost"]}-{c}{R["gna"]}'
            f'+{c}{R["oi"]}+{c}{R["jv"]}', model_rows[i]['ebitda'], N0, bold=True)
        fml(S, f'{c}{R["margin"]}', f'={c}{R["ebitda"]}/{c}{R["total"]}',
            model_rows[i]['ebitda_margin'], P1)
        fml(S, f'{c}{R["ebitda_xjv"]}', f'={c}{R["ebitda"]}-{c}{R["jv"]}',
            model_rows[i]['ebitda_ex_jv'], N0, bold=True)
    R['last_row'] = R['ebitda_xjv']
    return R


SEG_A = segment_block('A', RA, 4, True)
SEG_B = segment_block('B', RB, SEG_A['last_row'] + 3, False)
R_REV_TOT, R_EBITDA, R_EBITDA_XJV = SEG_A['total'], SEG_A['ebitda'], SEG_A['ebitda_xjv']
R_CONV, R_JV = SEG_A['conv'], SEG_A['jv']

# ============================================================== BALANCE SHEET =
# Written before the DCF because the DCF reads depreciation off the fixed-asset roll.
BS = ws('Balance Sheet', [46] + [13] * 8)
title(BS, 'Balance sheet — three audited years and the five-year roll-forward', 9)
hdr(BS, 3, ["USD thousands"] + YHDR)
b = 4
lab(BS, b, 'FIXED-ASSET ROLL-FORWARD', bold=True, fill=FILL_B)
b += 1
R_FA_OPEN, R_FA_ACQ, R_CAPEX, R_DNA, R_FA_CLOSE = b, b + 1, b + 2, b + 3, b + 4
lab(BS, R_FA_OPEN, 'Opening property, equipment, right-of-use and intangibles')
lab(BS, R_FA_ACQ, 'Acquired with the 2026 business combinations')
lab(BS, R_CAPEX, 'Capital expenditure')
lab(BS, R_DNA, 'Depreciation and amortisation')
lab(BS, R_FA_CLOSE, 'Closing fixed assets', bold=True)
for y in YRS_H:
    c = HCOL[y]
    put(BS, f'{c}{R_FA_ACQ}', 0.0, N0)
    put(BS, f'{c}{R_CAPEX}', H[str(y)]['capex'], N0)
    put(BS, f'{c}{R_DNA}', H[str(y)]['dna'], N0)
    fml(BS, f'{c}{R_FA_CLOSE}',
        f'={H[str(y)]["ppe"]}+{H[str(y)]["rou"]}+{H[str(y)]["intangibles"]}',
        H[str(y)]['ppe'] + H[str(y)]['rou'] + H[str(y)]['intangibles'], N0, bold=True)
_FA_2025 = (f'({AC("Property and equipment — FY2025 audited")}'
            f'+{AC("Right-of-use assets — FY2025 audited")}'
            f'+{AC("Intangible assets — FY2025 audited")})')
_FA_ACQ = (f'({AC("Property and equipment acquired")}+{AC("Right-of-use assets acquired")}'
           f'+{AC("Provisional goodwill")})')
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    fml(BS, f'{c}{R_FA_OPEN}', f'={_FA_2025}' if i == 0 else f'={prev}{R_FA_CLOSE}',
        RA[i]['ppe_open'] - (UE['acquisition_entry'][0][1] if i == 0 else 0.0), N0)
    fml(BS, f'{c}{R_FA_ACQ}', f'={_FA_ACQ}' if i == 0 else '=0',
        UE['acquisition_entry'][0][1] if i == 0 else 0.0, N0)
    fml(BS, f'{c}{R_CAPEX}', f'={AC(f"Capital expenditure — {y}")}', RA[i]['capex'], N0)
    fml(BS, f'{c}{R_DNA}',
        f'=({c}{R_FA_OPEN}+{c}{R_FA_ACQ})*{AC("Depreciation rate on opening fixed assets")}',
        RA[i]['dna'], N0)
    fml(BS, f'{c}{R_FA_CLOSE}',
        f'={c}{R_FA_OPEN}+{c}{R_FA_ACQ}+{c}{R_CAPEX}-{c}{R_DNA}', RA[i]['ppe_close'],
        N0, bold=True)
b = R_FA_CLOSE + 2

lab(BS, b, 'ASSETS', bold=True, fill=FILL_B)
b += 1
R_BS = {}
BS_ASSETS = [('Fixed assets', 'fixed_assets'), ('Other non-current assets', 'other_non_current'),
             ('Investment in joint ventures', 'jv_investment'), ('Inventories', 'inventories'),
             ('Trade and other receivables', 'receivables'),
             ('Due from related parties', 'due_from_rp'), ('Cash and cash equivalents', 'cash')]
HIST_BS = {'fixed_assets': lambda h: h['ppe'] + h['rou'] + h['intangibles'],
           'other_non_current': lambda h: h['deferred_tax_asset'] + h['advances'],
           'jv_investment': lambda h: h['jv_investment'],
           'inventories': lambda h: h['inventories'], 'receivables': lambda h: h['receivables'],
           'due_from_rp': lambda h: h['due_from_rp'],
           'cash': lambda h: h['cash'] + h['assets_held_for_sale'],
           'debt': lambda h: h['debt'] + h['leases'], 'payables': lambda h: h['payables'],
           'due_to_rp': lambda h: h['due_to_rp'], 'eosb': lambda h: h['eosb'],
           'tax_payable': lambda h: h['tax_payable'],
           'acquisition_liabilities': lambda h: 0.0, 'nci': lambda h: 0.0}
WCM = {k: H['2025'][k] / abs(H['2025']['working_capital']) for k in
       ('inventories', 'receivables', 'due_from_rp', 'payables', 'due_to_rp')}
R_WC = None
# The FY2025 column is the balance sheet the forecast opens on, so each of its
# lines is written as a live reference to the single disclosed cell that holds it
# on the Assumptions sheet, rather than pasted a second time here.
FY25_LINK = {'other_non_current': f'={AC("Deferred tax assets — FY2025 audited")}'
                                  f'+{AC("Advances — FY2025 audited")}'
                                  f'+{AC("The 2025 advance for acquisition of a subsidiary")}',
             'jv_investment': f'={AC("Investment in joint ventures — FY2025 audited")}',
             'cash': f'={AC("Cash and cash equivalents — FY2025 audited")}',
             'debt': f'={AC("Borrowings — FY2025 audited")}'
                     f'+{AC("Lease liabilities — FY2025 audited")}',
             'eosb': f'={AC("End-of-service benefits — FY2025 audited")}',
             'tax_payable': f'={AC("Income tax payable — FY2025 audited")}'}
for nice, key in BS_ASSETS:
    R_BS[key] = b
    lab(BS, b, nice)
    for y in YRS_H:
        if y == 2025 and key in FY25_LINK:
            fml(BS, f'{HCOL[y]}{b}', FY25_LINK[key], HIST_BS[key](H[str(y)]), N0)
        else:
            put(BS, f'{HCOL[y]}{b}', HIST_BS[key](H[str(y)]), N0)
    b += 1
R_TA = b
lab(BS, R_TA, 'Total assets', bold=True)
b += 2
lab(BS, b, 'LIABILITIES AND EQUITY', bold=True, fill=FILL_B)
b += 1
for nice, key in [('Borrowings and lease liabilities', 'debt'),
                  ('Trade and other payables', 'payables'),
                  ('Due to related parties', 'due_to_rp'),
                  ("Employees' end of service benefits", 'eosb'),
                  ('Income tax payable', 'tax_payable'),
                  ('Liabilities assumed with the 2026 acquisitions', 'acquisition_liabilities')]:
    R_BS[key] = b
    lab(BS, b, nice)
    for y in YRS_H:
        if y == 2025 and key in FY25_LINK:
            fml(BS, f'{HCOL[y]}{b}', FY25_LINK[key], HIST_BS[key](H[str(y)]), N0)
        else:
            put(BS, f'{HCOL[y]}{b}', HIST_BS[key](H[str(y)]), N0)
    b += 1
R_TL, R_NCI, R_EQ, R_TLE = b, b + 1, b + 2, b + 3
lab(BS, R_TL, 'Total liabilities', bold=True)
lab(BS, R_NCI, 'Non-controlling interests')
lab(BS, R_EQ, 'Equity attributable to owners of the company', bold=True)
lab(BS, R_TLE, 'Total liabilities and equity', bold=True)
b = R_TLE + 1
R_WCAP, R_ND, R_NDE = R_TLE + 2, R_TLE + 3, R_TLE + 4
lab(BS, R_WCAP, 'Working capital')
lab(BS, R_ND, 'Net debt')
lab(BS, R_NDE, 'Net debt / EBITDA')

for y in YRS_H:
    c = HCOL[y]
    h = H[str(y)]
    fml(BS, f'{c}{R_TA}', f'=SUM({c}{R_BS["fixed_assets"]}:{c}{R_BS["cash"]})',
        sum(HIST_BS[k](h) for _, k in BS_ASSETS), N0, bold=True)
    fml(BS, f'{c}{R_TL}', f'=SUM({c}{R_BS["debt"]}:{c}{R_BS["tax_payable"]})',
        sum(HIST_BS[k](h) for k in ('debt', 'payables', 'due_to_rp', 'eosb', 'tax_payable')),
        N0, bold=True)
    put(BS, f'{c}{R_NCI}', 0.0, N0)
    if y == 2025:
        fml(BS, f'{c}{R_EQ}', f'={AC("Total equity — FY2025 audited")}', h['equity'], N0,
            bold=True)
    else:
        put(BS, f'{c}{R_EQ}', h['equity'], N0, font=F_TOT)
    fml(BS, f'{c}{R_TLE}', f'={c}{R_TL}+{c}{R_NCI}+{c}{R_EQ}',
        sum(HIST_BS[k](h) for k in ('debt', 'payables', 'due_to_rp', 'eosb', 'tax_payable'))
        + h['equity'], N0, bold=True)
    fml(BS, f'{c}{R_WCAP}', f'={c}{R_BS["inventories"]}+{c}{R_BS["receivables"]}'
                            f'+{c}{R_BS["due_from_rp"]}-{c}{R_BS["payables"]}'
                            f'-{c}{R_BS["due_to_rp"]}', h['working_capital'], N0)
    fml(BS, f'{c}{R_ND}', f'={c}{R_BS["debt"]}-{c}{R_BS["cash"]}',
        HIST_BS['debt'](h) - HIST_BS['cash'](h), N0)
    fml(BS, f'{c}{R_NDE}', f'={c}{R_ND}/Segments!{c}{R_EBITDA}',
        (HIST_BS['debt'](h) - HIST_BS['cash'](h)) / h['ebitda'], N2)

# Debt is held flat at the audited FY2025 gross balance and then STEPS UP by the
# borrowings and leases that came with the two 2026 acquisitions. The other
# non-current assets fall by the 2025 advance, which converts into those same
# acquired assets rather than staying on the balance sheet.
_DEBT_OPEN = f'({AC("Borrowings — FY2025 audited")}+{AC("Lease liabilities — FY2025 audited")})'
_ACQ_DEBT = (f'({AC("Term loans, overdraft and borrowings assumed")}'
             f'+{AC("Lease liabilities assumed")})')
_ACQ_LIAB = (f'({AC("Deferred tax liability assumed")}+{AC("Income tax payable assumed")}'
             f'+{AC("Contingent consideration outstanding")})')
_OTHER_NC = (f'(D{{ONC}}-{AC("The 2025 advance for acquisition of a subsidiary")})')
OTHER_NC = H['2025']['deferred_tax_asset'] + H['2025']['advances']
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    bsrow = RA[i]['balance_sheet']
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    fml(BS, f'{c}{R_WCAP}', f'=Segments!{c}{R_REV_TOT}*{AC("Working capital / revenue")}',
        RA[i]['working_capital'], N0)
    fml(BS, f'{c}{R_BS["fixed_assets"]}', f'={c}{R_FA_CLOSE}', bsrow['fixed_assets'], N0)
    fml(BS, f'{c}{R_BS["other_non_current"]}',
        f'={_OTHER_NC}'.replace('{ONC}', str(R_BS['other_non_current'])),
        bsrow['other_non_current'], N0)
    fml(BS, f'{c}{R_BS["jv_investment"]}', f'={AC("Investment in joint ventures — FY2025 audited")}',
        bsrow['jv_investment'], N0)
    for k in ('inventories', 'receivables', 'due_from_rp', 'payables', 'due_to_rp'):
        fml(BS, f'{c}{R_BS[k]}', f'={c}{R_WCAP}*{WCM[k]}', bsrow[k], N0)
    fml(BS, f'{c}{R_BS["debt"]}', f'={_DEBT_OPEN}+{_ACQ_DEBT}', bsrow['debt'], N0)
    fml(BS, f'{c}{R_BS["eosb"]}',
        f'=({AC("End-of-service benefits — FY2025 audited")}'
        f'+{AC("End-of-service benefits assumed")})'
        f'*(1+{AC("Wage escalation (domestic labour only)")})^{i+1}',
        bsrow['eosb'], N0)
    fml(BS, f'{c}{R_BS["acquisition_liabilities"]}', f'={_ACQ_LIAB}',
        bsrow['acquisition_liabilities'], N0)
    fml(BS, f'{c}{R_BS["tax_payable"]}',
        f"='Income Statement'!{c}{'TAXROW'}", bsrow['tax_payable'], N0)   # patched below
    fml(BS, f'{c}{R_BS["cash"]}', f"='Cash Flow'!{c}{'CASHROW'}", bsrow['cash'], N0)  # patched
    fml(BS, f'{c}{R_TA}', f'=SUM({c}{R_BS["fixed_assets"]}:{c}{R_BS["cash"]})',
        bsrow['total_assets'], N0, bold=True)
    fml(BS, f'{c}{R_TL}',
        f'=SUM({c}{R_BS["debt"]}:{c}{R_BS["acquisition_liabilities"]})',
        bsrow['total_liabilities'], N0, bold=True)
    fml(BS, f'{c}{R_NCI}', f'={AC("Minority interests recognised")}', bsrow['nci'], N0)
    fml(BS, f'{c}{R_EQ}', f'={c}{R_TA}-{c}{R_TL}-{c}{R_NCI}', bsrow['equity_residual'], N0,
        bold=True)
    fml(BS, f'{c}{R_TLE}', f'={c}{R_TL}+{c}{R_NCI}+{c}{R_EQ}', bsrow['total_assets'], N0,
        bold=True)
    fml(BS, f'{c}{R_ND}', f'={c}{R_BS["debt"]}-{c}{R_BS["cash"]}', RA[i]['net_debt'], N0)
    fml(BS, f'{c}{R_NDE}', f'={c}{R_ND}/Segments!{c}{R_EBITDA}',
        RA[i]['net_debt'] / RA[i]['ebitda'], N2)

# --- the plateau case's own fixed-asset roll and working capital --------------
# The plateau case carries a quarter of the weight in the summary, so it is built
# live from its own capital-expenditure plan rather than pasted in as an answer.
bp = R_NDE + 2
lab(BS, bp, 'CAPACITY-PLATEAU CASE — FIXED ASSETS AND WORKING CAPITAL', bold=True, fill=FILL_B)
bp += 1
(RB_FA_OPEN, RB_FA_ACQ, RB_CAPEX, RB_DNA, RB_FA_CLOSE, RB_WCAP, RB_DEBT,
 RB_INT, RB_FIN, RB_JV, RB_EBIT, RB_PBT, RB_TAX, RB_PAT, RB_EOSB, RB_TAXP,
 RB_CASH, RB_ND) = range(bp, bp + 18)
lab(BS, RB_FA_OPEN, 'Opening fixed assets')
lab(BS, RB_FA_ACQ, 'Acquired with the 2026 business combinations')
lab(BS, RB_CAPEX, 'Capital expenditure')
lab(BS, RB_DNA, 'Depreciation and amortisation')
lab(BS, RB_FA_CLOSE, 'Closing fixed assets', bold=True)
lab(BS, RB_WCAP, 'Working capital')
lab(BS, RB_DEBT, 'Borrowings and lease liabilities')
lab(BS, RB_INT, 'Finance cost')
lab(BS, RB_FIN, 'Finance income')
lab(BS, RB_JV, 'Share of joint-venture results')
lab(BS, RB_EBIT, 'EBIT')
lab(BS, RB_PBT, 'Profit before tax')
lab(BS, RB_TAX, 'Income tax')
lab(BS, RB_PAT, 'Profit after tax', bold=True)
lab(BS, RB_EOSB, "Employees' end of service benefits")
lab(BS, RB_TAXP, 'Income tax payable')
lab(BS, RB_CASH, 'Closing cash')
lab(BS, RB_ND, 'Net debt', bold=True)
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    fml(BS, f'{c}{RB_FA_OPEN}', f'={_FA_2025}' if i == 0 else f'={prev}{RB_FA_CLOSE}',
        RB[i]['ppe_open'] - (UE['acquisition_entry'][0][1] if i == 0 else 0.0), N0)
    fml(BS, f'{c}{RB_FA_ACQ}', f'={_FA_ACQ}' if i == 0 else '=0',
        UE['acquisition_entry'][0][1] if i == 0 else 0.0, N0)
    fml(BS, f'{c}{RB_CAPEX}', f'={AC(f"Capital expenditure (plateau) — {y}")}', RB[i]['capex'],
        N0)
    fml(BS, f'{c}{RB_DNA}',
        f'=({c}{RB_FA_OPEN}+{c}{RB_FA_ACQ})*{AC("Depreciation rate on opening fixed assets")}',
        RB[i]['dna'], N0)
    fml(BS, f'{c}{RB_FA_CLOSE}', f'={c}{RB_FA_OPEN}+{c}{RB_FA_ACQ}+{c}{RB_CAPEX}-{c}{RB_DNA}',
        RB[i]['ppe_close'], N0, bold=True)
    fml(BS, f'{c}{RB_WCAP}',
        f'=Segments!{c}{SEG_B["total"]}*{AC("Working capital / revenue")}',
        RB[i]['working_capital'], N0)
    fml(BS, f'{c}{RB_DEBT}', f'={_DEBT_OPEN}+{_ACQ_DEBT}', RB[i]['balance_sheet']['debt'], N0)
    # The plateau case gets its OWN profit and cash statement, because the terminal
    # capitalisation rate is decided by whether THIS case's 2030 firm holds net
    # cash — a question its own balance sheet has to answer, not case A's.
    _copen = (f'{AC("Cash and cash equivalents — FY2025 audited")}' if i == 0
              else f'{prev}{RB_CASH}')
    _acqc = (f'+{AC("Cash acquired")}-{AC("Cash consideration paid in 2026")}'
             f'+{AC("Consideration received back against the 2025 advance")}') if i == 0 else ''
    _wcprev = f'D{R_WCAP}' if i == 0 else f'{prev}{RB_WCAP}'
    _acqwc = (f'-({AC("Inventories acquired")}+{AC("Trade and other receivables acquired")}'
              f'-{AC("Trade and other payables assumed")})') if i == 0 else ''
    _eosb_prev = (f'({AC("End-of-service benefits — FY2025 audited")}'
                  f'+{AC("End-of-service benefits assumed")})') if i == 0 \
        else f'{prev}{RB_EOSB}'
    _taxp_prev = (f'{AC("Income tax payable — FY2025 audited")}' if i == 0
                  else f'{prev}{RB_TAXP}')
    fml(BS, f'{c}{RB_INT}',
        f'={c}{RB_DEBT}*({AC("US 5-year Treasury yield")}'
        f'+{AC("Borrowing margin over Term SOFR on the latest facility")})',
        RB[i]['interest'], N0)
    fml(BS, f'{c}{RB_FIN}', f'=({_copen}{_acqc})*{AC("Secured Overnight Financing Rate, spot")}',
        RB[i]['finance_income'], N0)
    fml(BS, f'{c}{RB_JV}', f'=Segments!{c}{SEG_B["jv"]}', RB[i]['jv_share'], N0)
    fml(BS, f'{c}{RB_EBIT}', f'=Segments!{c}{SEG_B["ebitda_xjv"]}-{c}{RB_DNA}', RB[i]['ebit'], N0)
    fml(BS, f'{c}{RB_PBT}', f'={c}{RB_EBIT}+{c}{RB_JV}-{c}{RB_INT}+{c}{RB_FIN}', RB[i]['pbt'], N0)
    fml(BS, f'{c}{RB_TAX}', f'={c}{RB_PBT}*{AC("Corporate income tax rate")}', RB[i]['tax'], N0)
    fml(BS, f'{c}{RB_PAT}', f'={c}{RB_PBT}-{c}{RB_TAX}', RB[i]['pat'], N0, bold=True)
    fml(BS, f'{c}{RB_EOSB}',
        f'=({AC("End-of-service benefits — FY2025 audited")}'
        f'+{AC("End-of-service benefits assumed")})'
        f'*(1+{AC("Wage escalation (domestic labour only)")})^{i+1}',
        RB[i]['balance_sheet']['eosb'], N0)
    fml(BS, f'{c}{RB_TAXP}',
        f'={c}{RB_TAX}*{H["2025"]["tax_payable"] / H["2025"]["tax"]}',
        RB[i]['balance_sheet']['tax_payable'], N0)
    fml(BS, f'{c}{RB_CASH}',
        f'={_copen}{_acqc}+{c}{RB_PAT}+{c}{RB_DNA}-({c}{RB_WCAP}-{_wcprev}{_acqwc})'
        f'+({c}{RB_EOSB}-{_eosb_prev})+({c}{RB_TAXP}-{_taxp_prev})'
        f'-{c}{RB_CAPEX}-{AC("Dividend, FY2026 floor")}*(1+{AC("Dividend growth")})^{i}',
        RB[i]['cash_close'], N0)
    fml(BS, f'{c}{RB_ND}', f'={c}{RB_DEBT}-{c}{RB_CASH}', RB[i]['net_debt'], N0, bold=True)

# ============================================================ INCOME STATEMENT
IS = ws('Income Statement', [46] + [13] * 8)
title(IS, 'Income statement — three audited years and the five-year forecast', 9)
hdr(IS, 3, ["USD thousands"] + YHDR)
i_ = 4
IROW = {}
IS_LINES = [('Revenue', 'revenue'), ('EBITDA', 'ebitda'), ('EBITDA margin', 'margin'),
            ('Depreciation and amortisation', 'dna'),
            ('Share of joint-venture results', 'jv'),
            ('EBIT (after depreciation, before joint ventures)', 'ebit'),
            ('Finance cost', 'interest'), ('Finance income', 'fin_income'),
            ('Profit before tax', 'pbt'), ('Income tax', 'tax'), ('Profit after tax', 'pat'),
            ('Net margin', 'net_margin')]
for nice, key in IS_LINES:
    IROW[key] = i_
    lab(IS, i_, nice, bold=key in ('revenue', 'ebitda', 'pat'))
    i_ += 1
for y in YRS_H:
    c, h = HCOL[y], H[str(y)]
    fml(IS, f'{c}{IROW["revenue"]}', f'=Segments!{c}{R_REV_TOT}', h['revenue'], N0, bold=True)
    fml(IS, f'{c}{IROW["ebitda"]}', f'=Segments!{c}{R_EBITDA}', h['ebitda'], N0, bold=True)
    fml(IS, f'{c}{IROW["margin"]}', f'={c}{IROW["ebitda"]}/{c}{IROW["revenue"]}',
        h['ebitda_margin'], P1)
    fml(IS, f'{c}{IROW["dna"]}', f'=-\'Balance Sheet\'!{c}{R_DNA}', -h['dna'], N0)
    put(IS, f'{c}{IROW["jv"]}', h['jv_share'], N0)
    fml(IS, f'{c}{IROW["ebit"]}', f'={c}{IROW["ebitda"]}-{c}{IROW["jv"]}+{c}{IROW["dna"]}',
        h['ebit'] - h['jv_share'], N0)
    put(IS, f'{c}{IROW["interest"]}', -h['finance_cost'], N0)
    put(IS, f'{c}{IROW["fin_income"]}', h['finance_income'], N0)
    fml(IS, f'{c}{IROW["pbt"]}', f'={c}{IROW["ebit"]}+{c}{IROW["jv"]}+{c}{IROW["interest"]}'
                                 f'+{c}{IROW["fin_income"]}', h['pbt'], N0)
    put(IS, f'{c}{IROW["tax"]}', -h['tax'], N0)
    fml(IS, f'{c}{IROW["pat"]}', f'={c}{IROW["pbt"]}+{c}{IROW["tax"]}', h['pat'], N0, bold=True)
    fml(IS, f'{c}{IROW["net_margin"]}', f'={c}{IROW["pat"]}/{c}{IROW["revenue"]}',
        h['net_margin'], P1)
for i, y in enumerate(YRS_F):
    c, x = FCOL[y], RA[i]
    fml(IS, f'{c}{IROW["revenue"]}', f'=Segments!{c}{R_REV_TOT}', x['revenue'], N0, bold=True)
    fml(IS, f'{c}{IROW["ebitda"]}', f'=Segments!{c}{R_EBITDA}', x['ebitda'], N0, bold=True)
    fml(IS, f'{c}{IROW["margin"]}', f'={c}{IROW["ebitda"]}/{c}{IROW["revenue"]}',
        x['ebitda_margin'], P1)
    fml(IS, f'{c}{IROW["dna"]}', f'=-\'Balance Sheet\'!{c}{R_DNA}', -x['dna'], N0)
    fml(IS, f'{c}{IROW["jv"]}', f'=Segments!{c}{R_JV}', x['jv_share'], N0)
    fml(IS, f'{c}{IROW["ebit"]}', f'={c}{IROW["ebitda"]}-{c}{IROW["jv"]}+{c}{IROW["dna"]}',
        x['ebit'], N0)
    fml(IS, f'{c}{IROW["interest"]}', f"=-'Balance Sheet'!{c}{R_BS['debt']}"
                                      f'*(DCFRATE)', -x['interest'], N0)   # patched below
    # Opening cash for 2026 is the audited FY2025 balance PLUS the net cash the two
    # business combinations brought in, because that is the balance the year is
    # actually earning interest on.
    prevcash = ((f"('Balance Sheet'!D{R_BS['cash']}+{AC('Cash acquired')}"
                 f"-{AC('Cash consideration paid in 2026')}"
                 f"+{AC('Consideration received back against the 2025 advance')})") if i == 0
                else f"'Balance Sheet'!{FCOL[YRS_F[i-1]]}{R_BS['cash']}")
    fml(IS, f'{c}{IROW["fin_income"]}',
        f'={prevcash}*{AC("Secured Overnight Financing Rate, spot")}', x['finance_income'], N0)
    fml(IS, f'{c}{IROW["pbt"]}', f'={c}{IROW["ebit"]}+{c}{IROW["jv"]}+{c}{IROW["interest"]}'
                                 f'+{c}{IROW["fin_income"]}', x['pbt'], N0)
    fml(IS, f'{c}{IROW["tax"]}', f'=-{c}{IROW["pbt"]}*{AC("Corporate income tax rate")}',
        -x['tax'], N0)
    fml(IS, f'{c}{IROW["pat"]}', f'={c}{IROW["pbt"]}+{c}{IROW["tax"]}', x['pat'], N0, bold=True)
    fml(IS, f'{c}{IROW["net_margin"]}', f'={c}{IROW["pat"]}/{c}{IROW["revenue"]}',
        x['net_margin'], P1)

# ================================================================ CASH FLOW ===
CF = ws('Cash Flow', [46] + [13] * 8)
title(CF, 'Cash flow — the free-cash-flow bridge and the movement in cash', 9)
hdr(CF, 3, ["USD thousands"] + YHDR)
f_ = 4
CROW = {}
for nice, key in [('Profit after tax', 'pat'), ('Depreciation and amortisation', 'dna'),
                  ('Increase in working capital', 'dwc'),
                  ('Movement in provisions and tax payable', 'prov'),
                  ('Net cash from operating activities', 'cfo'),
                  ('Capital expenditure', 'capex'),
                  ('Net cash on the 2026 business combinations', 'acq'),
                  ('Dividends paid', 'div'),
                  ('Net movement in cash', 'net'), ('Opening cash', 'open'),
                  ('Closing cash', 'close'),
                  ('Memo: free cash flow to the firm', 'fcff')]:
    CROW[key] = f_
    lab(CF, f_, nice, bold=key in ('cfo', 'close', 'fcff'))
    f_ += 1
for y in YRS_H:
    c, h = HCOL[y], H[str(y)]
    fml(CF, f'{c}{CROW["pat"]}', f"='Income Statement'!{c}{IROW['pat']}", h['pat'], N0)
    fml(CF, f'{c}{CROW["dna"]}', f"='Balance Sheet'!{c}{R_DNA}", h['dna'], N0)
    fml(CF, f'{c}{CROW["capex"]}', f"=-'Balance Sheet'!{c}{R_CAPEX}", -h['capex'], N0)
    put(CF, f'{c}{CROW["div"]}', -h['dividends'], N0)
    put(CF, f'{c}{CROW["cfo"]}', h['cfo'], N0, font=F_TOT)
for i, y in enumerate(YRS_F):
    c, x = FCOL[y], RA[i]
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    fml(CF, f'{c}{CROW["pat"]}', f"='Income Statement'!{c}{IROW['pat']}", x['pat'], N0)
    fml(CF, f'{c}{CROW["dna"]}', f"='Balance Sheet'!{c}{R_DNA}", x['dna'], N0)
    # The working capital that ARRIVED with the acquisitions was bought, not funded
    # out of operations, so it is taken out of the operating movement in 2026.
    # Leaving it in would charge the free cash flow twice for the same balance —
    # once inside the consideration and once again here.
    _acq_wc = (f'-({AC("Inventories acquired")}+{AC("Trade and other receivables acquired")}'
               f'-{AC("Trade and other payables assumed")})') if i == 0 else ''
    fml(CF, f'{c}{CROW["dwc"]}',
        f"=-('Balance Sheet'!{c}{R_WCAP}-'Balance Sheet'!{prev}{R_WCAP}{_acq_wc})",
        -x['delta_wc'], N0)
    # The end-of-service provision assumed with the acquisitions arrived on the
    # balance sheet with the businesses; it is not a movement operations generated.
    _acq_eosb = f'-{AC("End-of-service benefits assumed")}' if i == 0 else ''
    fml(CF, f'{c}{CROW["prov"]}',
        f"=('Balance Sheet'!{c}{R_BS['eosb']}-'Balance Sheet'!{prev}{R_BS['eosb']}{_acq_eosb})"
        f"+('Balance Sheet'!{c}{R_BS['tax_payable']}-'Balance Sheet'!{prev}{R_BS['tax_payable']})",
        x['cfo'] - x['pat'] - x['dna'] + x['delta_wc'], N0)
    fml(CF, f'{c}{CROW["cfo"]}', f'={c}{CROW["pat"]}+{c}{CROW["dna"]}+{c}{CROW["dwc"]}'
                                 f'+{c}{CROW["prov"]}', x['cfo'], N0, bold=True)
    fml(CF, f'{c}{CROW["capex"]}', f"=-'Balance Sheet'!{c}{R_CAPEX}", -x['capex'], N0)
    _acq_cash = (f'={AC("Cash acquired")}-{AC("Cash consideration paid in 2026")}'
                 f'+{AC("Consideration received back against the 2025 advance")}') if i == 0 \
        else '=0'
    fml(CF, f'{c}{CROW["acq"]}', _acq_cash,
        (IN['acq_cash']['value'] - IN['acq_consideration_2026']['value']
         + IN['acq_cash_returned']['value']) if i == 0 else 0.0, N0)
    fml(CF, f'{c}{CROW["div"]}',
        f'=-{AC("Dividend, FY2026 floor")}*(1+{AC("Dividend growth")})^{i}', -x['dividend'], N0)
    fml(CF, f'{c}{CROW["net"]}',
        f'={c}{CROW["cfo"]}+{c}{CROW["capex"]}+{c}{CROW["acq"]}+{c}{CROW["div"]}',
        x['cash_close'] - x['cash_open']
        + ((IN['acq_cash']['value'] - IN['acq_consideration_2026']['value']
            + IN['acq_cash_returned']['value']) if i == 0 else 0.0), N0)
    if i == 0:
        fml(CF, f'{c}{CROW["open"]}', f'={AC("Cash and cash equivalents — FY2025 audited")}',
            H['2025']['cash'], N0)
    else:
        fml(CF, f'{c}{CROW["open"]}', f'={prev}{CROW["close"]}', x['cash_open'], N0)
    fml(CF, f'{c}{CROW["close"]}', f'={c}{CROW["open"]}+{c}{CROW["net"]}', x['cash_close'], N0,
        bold=True)
    fml(CF, f'{c}{CROW["fcff"]}', f"=('Income Statement'!{c}{IROW['ebit']}"
                                  f'*(1-{AC("Corporate income tax rate")}))'
                                  f'+{c}{CROW["dna"]}+{c}{CROW["capex"]}+{c}{CROW["dwc"]}',
        x['fcff'], N0, bold=True)

# patch the two forward references left as placeholders
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    BS[f'{c}{R_BS["tax_payable"]}'].value = (
        f"=-'Income Statement'!{c}{IROW['tax']}"
        f"*{H['2025']['tax_payable'] / H['2025']['tax']}")
    BS[f'{c}{R_BS["cash"]}'].value = f"='Cash Flow'!{c}{CROW['close']}"

# ===================================================================== DCF ====
DC = ws('DCF', [50, 16, 16] + [13] * 6)
title(DC, 'Discounted cash flow — the cost of capital and the waterfall, both built here', 9)
d_ = 3
lab(DC, d_, 'COST OF CAPITAL — BUILT, NOT PASTED', bold=True, fill=FILL_B)
d_ += 1
DR = {}


def drow(key, nice, formula=None, value=None, expected=None, fmt=P2, bold=False, note=''):
    global d_
    DR[key] = d_
    lab(DC, d_, nice, bold=bold)
    if formula:
        fml(DC, f'C{d_}', formula, expected, fmt, bold=bold)
    else:
        put(DC, f'C{d_}', value, fmt, font=F_TOT if bold else F_FX)
    if note:
        cc = DC.cell(row=d_, column=4, value=note)
        cc.font = F_TXT
        cc.alignment = WRAP
    d_ += 1


drow('rf_obs', 'US 10-year Treasury yield (observed)',
     f'={AC("US 10-year Treasury yield (observed)")}', expected=W['rf_observed'])
drow('ds_us', 'less the US adjusted default spread',
     f'={AC("US adjusted default spread")}', expected=W['us_default_spread'])
drow('rf_star', 'Normalised risk-free rate', f'=C{DR["rf_obs"]}-C{DR["ds_us"]}',
     expected=W['rf_star'], bold=True)
drow('beta', 'Equity beta', f'={AC("Equity beta")}', expected=W['beta'], fmt=N3)
drow('erp_r', 'Equity risk premium — rating basis',
     f'={AC("Equity risk premium — rating basis")}', expected=W['erp_rating'])
drow('ke_r', 'Cost of equity — rating basis',
     f'=C{DR["rf_star"]}+C{DR["beta"]}*C{DR["erp_r"]}', expected=W['ke_rating'], bold=True)
drow('erp_c', 'Equity risk premium — credit-default-swap basis',
     f'={AC("Equity risk premium — credit-default-swap basis")}', expected=W['erp_cds'])
drow('ke_c', 'Cost of equity — credit-default-swap basis',
     f'=C{DR["rf_star"]}+C{DR["beta"]}*C{DR["erp_c"]}', expected=W['ke_cds'], bold=True)
drow('ust5', 'US 5-year Treasury yield', f'={AC("US 5-year Treasury yield")}',
     expected=IN['ust5']['value'])
drow('margin', 'Borrowing margin on the latest facility',
     f'={AC("Borrowing margin over Term SOFR on the latest facility")}',
     expected=IN['facility_margin']['value'])
drow('kd_pre', 'Marginal cost of debt, pre-tax (term-matched)',
     f'=C{DR["ust5"]}+C{DR["margin"]}', expected=W['kd_pretax'], bold=True)
drow('kd_spot', 'Cross-check: spot floating all-in cost',
     f'={AC("Secured Overnight Financing Rate, spot")}+C{DR["margin"]}',
     expected=W['kd_candidates']['spot_floating'])
drow('kd_trail', 'Cross-check: trailing effective rate on average gross debt',
     value=W['kd_candidates']['trailing_effective'])
drow('sov', 'Sovereign floor — 5-year Treasury plus the Abu Dhabi credit-default-swap spread',
     f'=C{DR["ust5"]}+{AC("Abu Dhabi sovereign credit-default-swap spread")}',
     expected=W['sovereign_floor'])
drow('kd_post', 'Cost of debt after tax',
     f'=C{DR["kd_pre"]}*(1-{AC("Corporate income tax rate")})', expected=W['kd_after_tax'])
drow('mcap', 'Market capitalisation',
     f'=({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
     f'*{AC("Market price")}/{AC("AED per USD")}', expected=M['market_cap_usd_k'], fmt=N0)
drow('nd', 'Net debt at 30 June 2026',
     f'={AC("Borrowings")}+{AC("Lease liabilities")}-{AC("Cash and cash equivalents")}',
     expected=W['net_debt'], fmt=N0)
drow('gd', 'Gross interest-bearing debt at 30 June 2026',
     f'={AC("Borrowings")}+{AC("Lease liabilities")}',
     expected=W['gross_debt'], fmt=N0,
     note='The weights are struck on GROSS debt. Interest is paid on the gross balance, and cash '
          'is bridged separately in the enterprise-value-to-equity step; netting it off here as '
          'well would take credit for the same cash twice')
drow('we', 'Weight of equity', f'=C{DR["mcap"]}/(C{DR["mcap"]}+C{DR["gd"]})',
     expected=W['weight_equity'], fmt=P1)
drow('wd', 'Weight of debt', f'=1-C{DR["we"]}', expected=W['weight_debt'], fmt=P1)
drow('wacc_r', 'Weighted average cost of capital — rating basis',
     f'=C{DR["we"]}*C{DR["ke_r"]}+C{DR["wd"]}*C{DR["kd_post"]}', expected=W['wacc_rating'],
     bold=True)
drow('wacc_c', 'Weighted average cost of capital — credit-default-swap basis',
     f'=C{DR["we"]}*C{DR["ke_c"]}+C{DR["wd"]}*C{DR["kd_post"]}', expected=W['wacc_cds'],
     bold=True)

# patch the income-statement interest formula now that the rate cell exists
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    IS[f'{c}{IROW["interest"]}'].value = (f"=-'Balance Sheet'!{c}{R_BS['debt']}"
                                          f"*DCF!$C${DR['kd_pre']}")

d_ += 1
lab(DC, d_, 'FREE-CASH-FLOW WATERFALL — CONTINUED-EXPANSION CASE', bold=True, fill=FILL_B)
d_ += 1
WHDR = d_
for j, y in enumerate(YRS_F):
    cc = DC.cell(row=WHDR, column=2 + j, value=f'FY{y}E')
    cc.font = F_H2
    cc.alignment = Alignment(horizontal='center')
d_ += 1
WCOL = {y: get_column_letter(2 + j) for j, y in enumerate(YRS_F)}
WR = {}
for key, nice, bold_ in [('revenue', 'Revenue', False), ('ebitda', 'EBITDA', True),
                         ('margin', 'EBITDA margin', False),
                         ('dna', 'less depreciation and amortisation', False),
                         ('ebit', 'EBIT', True), ('taxr', 'Tax rate', False),
                         ('nopat', 'NOPAT = EBIT x (1 - tax rate)', True),
                         ('addback', 'add back depreciation and amortisation', False),
                         ('capex', 'less capital expenditure', False),
                         ('dwc', 'less increase in working capital', False),
                         ('fcff', 'Free cash flow to the firm', True),
                         ('df', 'Discount factor', False),
                         ('pv', 'Present value of free cash flow to the firm', True)]:
    WR[key] = d_
    lab(DC, d_, nice, bold=bold_)
    d_ += 1
for i, y in enumerate(YRS_F):
    c, sc, x = WCOL[y], FCOL[y], RA[i]
    fml(DC, f'{c}{WR["revenue"]}', f'=Segments!{sc}{R_REV_TOT}', x['revenue'], N0)
    fml(DC, f'{c}{WR["ebitda"]}', f'=Segments!{sc}{R_EBITDA_XJV}', x['ebitda_ex_jv'], N0,
        bold=True)
    fml(DC, f'{c}{WR["margin"]}', f'={c}{WR["ebitda"]}/{c}{WR["revenue"]}',
        x['ebitda_ex_jv'] / x['revenue'], P1)
    fml(DC, f'{c}{WR["dna"]}', f"=-'Balance Sheet'!{sc}{R_DNA}", -x['dna'], N0)
    fml(DC, f'{c}{WR["ebit"]}', f'={c}{WR["ebitda"]}+{c}{WR["dna"]}', x['ebit'], N0, bold=True)
    fml(DC, f'{c}{WR["taxr"]}', f'={AC("Corporate income tax rate")}', TAXR, P1)
    fml(DC, f'{c}{WR["nopat"]}', f'={c}{WR["ebit"]}*(1-{c}{WR["taxr"]})', x['nopat'], N0,
        bold=True)
    fml(DC, f'{c}{WR["addback"]}', f"='Balance Sheet'!{sc}{R_DNA}", x['dna'], N0)
    fml(DC, f'{c}{WR["capex"]}', f"=-'Balance Sheet'!{sc}{R_CAPEX}", -x['capex'], N0)
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    _acq_wc = (f'-({AC("Inventories acquired")}+{AC("Trade and other receivables acquired")}'
               f'-{AC("Trade and other payables assumed")})') if i == 0 else ''
    fml(DC, f'{c}{WR["dwc"]}',
        f"=-('Balance Sheet'!{sc}{R_WCAP}-'Balance Sheet'!{prev}{R_WCAP}{_acq_wc})",
        -x['delta_wc'], N0)
    fml(DC, f'{c}{WR["fcff"]}', f'={c}{WR["nopat"]}+{c}{WR["addback"]}+{c}{WR["capex"]}'
                                f'+{c}{WR["dwc"]}', x['fcff'], N0, bold=True)
    fml(DC, f'{c}{WR["df"]}', f'=1/(1+$C${DR["wacc_r"]})^{i+1}', x['discount_factor'], N3)
    fml(DC, f'{c}{WR["pv"]}', f'={c}{WR["fcff"]}*{c}{WR["df"]}', x['pv_fcff'], N0, bold=True)

d_ += 1
lab(DC, d_, 'TERMINAL BLOCK AND ENTERPRISE VALUE', bold=True, fill=FILL_B)
d_ += 1
TR = {}


def trow(key, nice, formula, expected, fmt=N0, bold=False):
    global d_
    TR[key] = d_
    lab(DC, d_, nice, bold=bold)
    fml(DC, f'C{d_}', formula, expected, fmt, bold=bold)
    d_ += 1


last = WCOL[2030]
trow('g', 'Terminal growth rate', f'={AC("Terminal growth — continued expansion")}',
     CA['terminal_growth'], P2)
trow('roic', 'Terminal return on invested capital',
     f'={AC("Terminal return on invested capital")}', CA['terminal_roic'], P2)
trow('nopat_t', 'Terminal-year NOPAT (one year beyond the window)',
     f'={last}{WR["nopat"]}*(1+C{TR["g"]})', CA['terminal_nopat'], N0)
trow('reinv', 'Reinvestment rate = growth / return on capital',
     f'=C{TR["g"]}/C{TR["roic"]}', CA['reinvestment_rate'], P1)
trow('tnd', 'Net debt in the terminal year, off this model\'s own balance sheet',
     f"='Balance Sheet'!{FCOL[2030]}{R_ND}", CA['terminal_net_debt'], N0)
trow('trate', 'Terminal capitalisation rate', f'=$C${DR["wacc_r"]}', CA['terminal_rate'], P2,
     bold=True)
trow('tv', 'Terminal value',
     f'=C{TR["nopat_t"]}*(1-C{TR["reinv"]})/(C{TR["trate"]}-C{TR["g"]})',
     CA['terminal_value'], N0, bold=True)
trow('pv_tv', 'Present value of the terminal value',
     f'=C{TR["tv"]}*{last}{WR["df"]}', CA['pv_terminal'], N0)
trow('pv_exp', 'Present value of the explicit five years',
     f'=SUM({WCOL[2026]}{WR["pv"]}:{last}{WR["pv"]})', CA['pv_explicit'], N0)
trow('ev', 'Enterprise value at 31 December 2025', f'=C{TR["pv_exp"]}+C{TR["pv_tv"]}',
     CA['enterprise_value_dec25'], N0, bold=True)
trow('tvshare', 'Terminal value as a share of enterprise value',
     f'=C{TR["pv_tv"]}/C{TR["ev"]}', CA['tv_pct_of_ev'], P1, bold=True)
lab(DC, TR['tnd'], "This model's own 2030 balance sheet has the firm holding net cash. A "
    'self-audit of this study proposed capitalising the terminal block at the cost of equity on '
    'that basis, and it was accepted and priced before being implemented. It was then reversed. '
    'The capital-structure weights above are struck on GROSS debt, because interest is paid on '
    'the gross balance and cash is bridged separately — and a firm holding cash alongside '
    'undiminished gross borrowings has not de-levered. Reading the terminal structure off net '
    'debt after refusing to read today\'s off net debt is one quantity treated two ways. The '
    'terminal net debt is published here because it is worth seeing; it does not set the rate.',
    col=4, wrap=True)

d_ += 1
lab(DC, d_, 'CAPACITY-PLATEAU CASE — the same waterfall on the plateau fleet plan',
    bold=True, fill=FILL_B)
d_ += 1
for j, y in enumerate(YRS_F):
    cc = DC.cell(row=d_, column=2 + j, value=f'FY{y}E')
    cc.font = F_H2
    cc.alignment = Alignment(horizontal='center')
d_ += 1
BR = {}
for key, nice, bold_ in (('ebitda', 'EBITDA excluding joint-venture results', False),
                         ('dna', 'less depreciation and amortisation', False),
                         ('ebit', 'EBIT', False),
                         ('nopat', 'NOPAT = EBIT x (1 - tax rate)', False),
                         ('addback', 'add back depreciation and amortisation', False),
                         ('capex', 'less capital expenditure', False),
                         ('dwc', 'less increase in working capital', False),
                         ('fcff', 'Free cash flow to the firm — plateau case', True),
                         ('df', 'Discount factor', False),
                         ('pv', 'Present value', True)):
    BR[key] = d_
    lab(DC, d_, nice, bold=bold_)
    d_ += 1
for i, y in enumerate(YRS_F):
    c, sc = WCOL[y], FCOL[y]
    fml(DC, f'{c}{BR["ebitda"]}', f'=Segments!{sc}{SEG_B["ebitda_xjv"]}', RB[i]['ebitda_ex_jv'],
        N0)
    fml(DC, f'{c}{BR["dna"]}', f"=-'Balance Sheet'!{sc}{RB_DNA}", -RB[i]['dna'], N0)
    fml(DC, f'{c}{BR["ebit"]}', f'={c}{BR["ebitda"]}+{c}{BR["dna"]}', RB[i]['ebit'], N0)
    fml(DC, f'{c}{BR["nopat"]}', f'={c}{BR["ebit"]}*(1-{AC("Corporate income tax rate")})',
        RB[i]['nopat'], N0)
    fml(DC, f'{c}{BR["addback"]}', f"='Balance Sheet'!{sc}{RB_DNA}", RB[i]['dna'], N0)
    fml(DC, f'{c}{BR["capex"]}', f"=-'Balance Sheet'!{sc}{RB_CAPEX}", -RB[i]['capex'], N0)
    prev = 'D' if i == 0 else FCOL[YRS_F[i - 1]]
    prev_wc = (f"'Balance Sheet'!D{R_WCAP}" if i == 0
               else f"'Balance Sheet'!{prev}{RB_WCAP}")
    _acq_wc_b = (f'-({AC("Inventories acquired")}+{AC("Trade and other receivables acquired")}'
                 f'-{AC("Trade and other payables assumed")})') if i == 0 else ''
    fml(DC, f'{c}{BR["dwc"]}', f"=-('Balance Sheet'!{sc}{RB_WCAP}-{prev_wc}{_acq_wc_b})",
        -RB[i]['delta_wc'], N0)
    fml(DC, f'{c}{BR["fcff"]}', f'={c}{BR["nopat"]}+{c}{BR["addback"]}+{c}{BR["capex"]}'
                                f'+{c}{BR["dwc"]}', RB[i]['fcff'], N0, bold=True)
    fml(DC, f'{c}{BR["df"]}', f'=1/(1+$C${DR["wacc_r"]})^{i+1}', RB[i]['discount_factor'], N3)
    fml(DC, f'{c}{BR["pv"]}', f'={c}{BR["fcff"]}*{c}{BR["df"]}', RB[i]['pv_fcff'], N0,
        bold=True)
BT = {}
for key, nice, formula, expected, fmt in (
        ('g', 'Terminal growth rate — plateau', f'={AC("Terminal growth — capacity plateau")}',
         CB['terminal_growth'], P2),
        ('nopat_t', 'Terminal-year NOPAT', f'={RB[-1]["nopat"]}*(1+C{d_})', CB['terminal_nopat'],
         N0),
        ('tnd', 'Net debt in the terminal year — plateau', '', CB['terminal_net_debt'], N0),
        ('trate', 'Terminal capitalisation rate — plateau', '', CB['terminal_rate'], P2),
        ('tv', 'Terminal value', '', CB['terminal_value'], N0),
        ('pv_tv', 'Present value of the terminal value', '', CB['pv_terminal'], N0),
        ('pv_exp', 'Present value of the explicit five years',
         f'=SUM({WCOL[2026]}{BR["pv"]}:{last}{BR["pv"]})', CB['pv_explicit'], N0),
        ('ev', 'Enterprise value — plateau, at 31 December 2025', '',
         CB['enterprise_value_dec25'], N0),
        ('tvshare', 'Terminal value as a share of enterprise value', '', CB['tv_pct_of_ev'], P1)):
    BT[key] = d_
    lab(DC, d_, nice, bold=key in ('ev', 'tvshare'))
    d_ += 1
fml(DC, f'C{BT["g"]}', f'={AC("Terminal growth — capacity plateau")}', CB['terminal_growth'], P2)
fml(DC, f'C{BT["nopat_t"]}', f'={last}{BR["nopat"]}*(1+C{BT["g"]})', CB['terminal_nopat'], N0)
fml(DC, f'C{BT["tnd"]}', f"='Balance Sheet'!{FCOL[2030]}{RB_ND}", CB['terminal_net_debt'], N0)
fml(DC, f'C{BT["trate"]}', f'=$C${DR["wacc_r"]}', CB['terminal_rate'], P2)
fml(DC, f'C{BT["tv"]}',
    f'=C{BT["nopat_t"]}*(1-C{BT["g"]}/C{TR["roic"]})/(C{BT["trate"]}-C{BT["g"]})',
    CB['terminal_value'], N0)
fml(DC, f'C{BT["pv_tv"]}', f'=C{BT["tv"]}*{last}{BR["df"]}', CB['pv_terminal'], N0)
fml(DC, f'C{BT["pv_exp"]}', f'=SUM({WCOL[2026]}{BR["pv"]}:{last}{BR["pv"]})', CB['pv_explicit'],
    N0)
fml(DC, f'C{BT["ev"]}', f'=C{BT["pv_exp"]}+C{BT["pv_tv"]}', CB['enterprise_value_dec25'],
    N0, bold=True)
fml(DC, f'C{BT["tvshare"]}', f'=C{BT["pv_tv"]}/C{BT["ev"]}', CB['tv_pct_of_ev'], P1, bold=True)

# ============================================================== SOTP BRIDGE ===
BG = ws('SOTP Bridge', [58, 18, 18, 60])
title(BG, 'Enterprise value to equity — the bridge, struck on the 30 June 2026 capital '
          'structure', 4)
hdr(BG, 3, ['', 'Continued expansion', 'Capacity plateau', 'Note'])
g_ = 4
GR = {}


def grow_(key, nice, fa, fb, ea, eb, note='', fmt=N0, bold=False):
    global g_
    GR[key] = g_
    lab(BG, g_, nice, bold=bold)
    fml(BG, f'B{g_}', fa, ea, fmt, bold)
    fml(BG, f'C{g_}', fb, eb, fmt, bold)
    c = BG.cell(row=g_, column=4, value=note)
    c.font = F_TXT
    c.alignment = WRAP
    g_ += 1


grow_('ev25', 'Enterprise value at 31 December 2025', f'=DCF!$C${TR["ev"]}',
      f'=DCF!$C${BT["ev"]}', CA['enterprise_value_dec25'], CB['enterprise_value_dec25'],
      'Present value of the explicit five years plus the discounted terminal value. Discounting '
      'FY2026 cash flow by a full year places the answer at the START of FY2026, which is 31 '
      'December 2025 — six months before the balance sheet the deductions below come from',
      bold=True)
grow_('tvshare', 'of which terminal value', f'=DCF!$C${TR["tvshare"]}',
      f'=DCF!$C${BT["tvshare"]}', CA['tv_pct_of_ev'], CB['tv_pct_of_ev'],
      'The share of the answer that rests on the period beyond 2030 — read this before the '
      'valuation', fmt=P1, bold=True)
grow_('carry', 'carried forward half a year at the cost of capital',
      f'=B{GR["ev25"]}*((1+DCF!$C${DR["wacc_r"]})^0.5-1)',
      f'=C{GR["ev25"]}*((1+DCF!$C${DR["wacc_r"]})^0.5-1)',
      CA['enterprise_value_dec25'] * ((1 + W['wacc_rating']) ** 0.5 - 1),
      CB['enterprise_value_dec25'] * ((1 + W['wacc_rating']) ** 0.5 - 1),
      'An enterprise compounds at its own cost of capital. Closing the six-month gap rather than '
      'ignoring it is what puts both sides of this bridge on one date')
_stub = -UE['fcff_1h26'] * (1 + W['wacc_rating']) ** 0.25
grow_('stub', 'less the free cash flow the business actually generated over that half year',
      f'=-({AC("Net cash from operating activities — first half of 2026")}'
      f'-{AC("Capital expenditure — first half of 2026")})'
      f'*(1+DCF!$C${DR["wacc_r"]})^0.25',
      f'=-({AC("Net cash from operating activities — first half of 2026")}'
      f'-{AC("Capital expenditure — first half of 2026")})'
      f'*(1+DCF!$C${DR["wacc_r"]})^0.25',
      _stub, _stub,
      'Reported, not forecast. Cash the enterprise has already handed out is no longer inside it. '
      'Acquisition consideration is deliberately NOT deducted here: it bought assets that sit '
      'inside the same enterprise, so it nets out of an enterprise-value roll-forward')
grow_('ev', 'Enterprise value at 30 June 2026',
      f'=B{GR["ev25"]}+B{GR["carry"]}+B{GR["stub"]}',
      f'=C{GR["ev25"]}+C{GR["carry"]}+C{GR["stub"]}',
      CA['enterprise_value'], CB['enterprise_value'],
      'Now on the same date as every line beneath it', bold=True)
grow_('jv', 'add investment in joint ventures', f'={AC("Investment in joint ventures")}',
      f'={AC("Investment in joint ventures")}', IN['jvinv_1h26']['value'],
      IN['jvinv_1h26']['value'],
      'Enersol and Turnwell at carrying value; their earnings are excluded from the discounted '
      'EBITDA, so the stake is added here instead')
grow_('cash', 'add cash and cash equivalents', f'={AC("Cash and cash equivalents")}',
      f'={AC("Cash and cash equivalents")}', IN['cash_1h26']['value'], IN['cash_1h26']['value'],
      'Balance sheet at 30 June 2026')
grow_('debt', 'less borrowings', f'=-{AC("Borrowings")}', f'=-{AC("Borrowings")}',
      -IN['debt_1h26']['value'], -IN['debt_1h26']['value'],
      'Current and non-current term loans')
grow_('lease', 'less lease liabilities', f'=-{AC("Lease liabilities")}',
      f'=-{AC("Lease liabilities")}', -IN['lease_1h26']['value'], -IN['lease_1h26']['value'], '')
grow_('finliab', 'less the financial liability over the acquired minorities',
      f'=-{AC("Financial liability over the acquired minorities")}',
      f'=-{AC("Financial liability over the acquired minorities")}',
      -IN['finliab_1h26']['value'], -IN['finliab_1h26']['value'],
      'The present value of the price at which the parent may be required to buy the 30% of SLDC '
      'and the 20% of MBPS it does not own. The minority interests reported at 30 June 2026 are '
      'NOT deducted as well: that is the same claim under another name, and the company has '
      'already charged a matching investment reserve against owners equity for it')
grow_('eqjun', 'Equity value at 30 June 2026',
      f'=B{GR["ev"]}+SUM(B{GR["jv"]}:B{GR["finliab"]})',
      f'=C{GR["ev"]}+SUM(C{GR["jv"]}:C{GR["finliab"]})',
      CA['bridge']['equity_30jun26'], CB['bridge']['equity_30jun26'], '', bold=True)
grow_('accr', 'accreted to the price anchor at the cost of equity',
      f'=B{GR["eqjun"]}*((1+DCF!$C${DR["ke_r"]})'
      f'^({AC("Days from 30 June 2026 to the price anchor")}/365)-1)',
      f'=C{GR["eqjun"]}*((1+DCF!$C${DR["ke_r"]})'
      f'^({AC("Days from 30 June 2026 to the price anchor")}/365)-1)',
      CA['bridge']['accretion'], CB['bridge']['accretion'],
      'The remaining 38 days from the balance-sheet date to the share price this value is '
      'compared against')
grow_('eq', 'Equity value at the price anchor',
      f'=B{GR["eqjun"]}+B{GR["accr"]}', f'=C{GR["eqjun"]}+C{GR["accr"]}',
      CA['equity_value'], CB['equity_value'], '', bold=True)
grow_('sh', 'Shares outstanding (thousands)',
      f'={AC("Shares issued")}-{AC("Treasury shares held by the market maker")}',
      f'={AC("Shares issued")}-{AC("Treasury shares held by the market maker")}', SH, SH,
      'Issued shares less the shares the appointed market maker holds on the company\'s behalf')
grow_('ps_usd', 'Value per share (USD)', f'=B{GR["eq"]}/B{GR["sh"]}', f'=C{GR["eq"]}/C{GR["sh"]}',
      CA['value_per_share_usd'], CB['value_per_share_usd'], '', fmt=N3)
grow_('ps_aed', 'Value per share (AED)', f'=B{GR["ps_usd"]}*{AC("AED per USD")}',
      f'=C{GR["ps_usd"]}*{AC("AED per USD")}', CA['value_per_share_aed'],
      CB['value_per_share_aed'], 'Converted at the peg', fmt=N2, bold=True)

# ====================================================== RELATIVE & NORMALIZED =
RN = ws('Relative & Normalized', [56, 18, 18, 54])
title(RN, 'Relative multiples, normalised earnings power, and the book-value lens', 4)
n_ = 3
lab(RN, n_, 'RELATIVE MULTIPLES — segment-weighted, applied to the company\'s own guided EBITDA',
    bold=True, fill=FILL_B)
n_ += 1
NR = {}


def nrow(key, nice, formula, expected, fmt=N2, note='', bold=False):
    global n_
    NR[key] = n_
    lab(RN, n_, nice, bold=bold)
    fml(RN, f'C{n_}', formula, expected, fmt, bold)
    c = RN.cell(row=n_, column=4, value=note)
    c.font = F_TXT
    c.alignment = WRAP
    n_ += 1


SEGW = REL['segment_weights']
nrow('w_on', 'Onshore share of FY2025 segment EBITDA',
     f'={IN["seg_ebitda_on_fy25"]["value"]}/({IN["seg_ebitda_on_fy25"]["value"]}'
     f'+{IN["seg_ebitda_off_fy25"]["value"]}+{IN["seg_ebitda_ofs_fy25"]["value"]})',
     SEGW['onshore'], P1, 'From the FY2025 segment note')
nrow('w_off', 'Offshore share of FY2025 segment EBITDA',
     f'={IN["seg_ebitda_off_fy25"]["value"]}/({IN["seg_ebitda_on_fy25"]["value"]}'
     f'+{IN["seg_ebitda_off_fy25"]["value"]}+{IN["seg_ebitda_ofs_fy25"]["value"]})',
     SEGW['offshore'], P1)
nrow('w_ofs', 'Oilfield Services share of FY2025 segment EBITDA',
     f'=1-C{NR["w_on"]}-C{NR["w_off"]}', SEGW['ofs'], P1)
nrow('m_on', 'Multiple applied to Onshore',
     f'=({AC("Peer median EV/EBITDA — MENA national-oil-company drillers")}'
     f'+{AC("Peer median EV/EBITDA — global land drillers")})/2',
     (REL['median_mena'] + REL['median_land']) / 2, N2,
     'Average of the MENA national-oil-company drillers and the global land drillers')
nrow('m_off', 'Multiple applied to Offshore',
     f'=({AC("Peer median EV/EBITDA — MENA national-oil-company drillers")}'
     f'+{AC("Peer median EV/EBITDA — global offshore drillers")})/2',
     (REL['median_mena'] + REL['median_offshore']) / 2, N2)
nrow('m_ofs', 'Multiple applied to Oilfield Services',
     f'={AC("Peer median EV/EBITDA — diversified oilfield services")}', REL['median_ofs'], N2)
nrow('blend', 'Segment-weighted multiple',
     f'=C{NR["w_on"]}*C{NR["m_on"]}+C{NR["w_off"]}*C{NR["m_off"]}+C{NR["w_ofs"]}*C{NR["m_ofs"]}',
     REL['blended_multiple'], N2, bold=True)
nrow('ebitda', 'Last-twelve-month EBITDA excluding joint ventures',
     f'={AC("Last-twelve-month EBITDA excluding joint ventures")}',
     REL['applied_ebitda'], N0,
     'The multiple and the earnings it multiplies are on the same footing: every peer multiple '
     'above is an enterprise value struck today over that peer\'s last twelve months, and the '
     'joint-venture earnings are taken out here because the joint-venture carrying value is '
     'added back on the bridge below')
nrow('ev', 'Implied enterprise value', f'=C{NR["blend"]}*C{NR["ebitda"]}', REL['enterprise_value'],
     N0, bold=True)
nrow('eq', 'Implied equity value at 30 June 2026',
     f'=C{NR["ev"]}+{AC("Investment in joint ventures")}+{AC("Cash and cash equivalents")}'
     f'-{AC("Borrowings")}-{AC("Lease liabilities")}'
     f'-{AC("Financial liability over the acquired minorities")}',
     REL['bridge']['equity_30jun26'], N0,
     'The same bridge the cash-flow lenses use, and the minority is deducted once')
nrow('eqa', 'Accreted to the price anchor at the cost of equity',
     f'=C{NR["eq"]}*(1+DCF!$C${DR["ke_r"]})'
     f'^({AC("Days from 30 June 2026 to the price anchor")}/365)',
     REL['equity_value'], N0)
nrow('ps', 'Implied value per share (AED)',
     f'=C{NR["eqa"]}/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
     f'*{AC("AED per USD")}', REL['value_per_share_aed'], N2, bold=True)
nrow('own', 'The company\'s own enterprise value / last-twelve-month EBITDA',
     f'=(DCF!$C${DR["mcap"]}+DCF!$C${DR["nd"]})/C{NR["ebitda"]}', REL['implied_own_ev_ebitda'],
     X2, 'The gap between this and the segment-weighted peer multiple is the whole of the '
         'relative lens', bold=True)

n_ += 1
lab(RN, n_, 'NORMALISED EARNINGS POWER — the installed fleet, no growth credited',
    bold=True, fill=FILL_B)
n_ += 1
nrow('nrev', 'Normalised revenue on the fleet installed at 30 June 2026',
     f'={AC("Abu Dhabi onshore rigs — FY2025 year end")}'
     f'*{AC("Revenue per Abu Dhabi onshore rig-year")}'
     f'+{AC("Regional onshore rigs — 30 June 2026")}'
     f'*{AC("Revenue per regional onshore rig-year")}'
     f'+{AC("Jack-up rigs — FY2025 year end")}*{AC("Revenue per jack-up rig-year")}'
     f'+{AC("Island rigs — 30 June 2026")}*{AC("Revenue per island rig-year")}'
     f'+({AC("Integrated-services rigs — 30 June 2026")}'
     f'+{AC("Rigs given at least one discrete service — 30 June 2026")})'
     f'*{AC("Revenue per rig served — FY2025")}', NORM['revenue'], N0,
     'The rigs the company has ALREADY taken delivery of, at FY2025 realised rates. The first '
     'edition carried 14 island rigs and the 70 integrated rigs the company TARGETS for the end '
     'of 2026; the reported counts at 30 June 2026 are 13 and 61. A target is growth, and this '
     'lens is the one that credits none')
nrow('nebitda', 'Normalised EBITDA', f'=C{NR["nrev"]}*{AC("Normalised EBITDA margin")}',
     NORM['ebitda'], N0)
nrow('ndna', 'Normalised depreciation and amortisation',
     f'={AC("Normalised depreciation and amortisation")}', NORM['dna'], N0,
     'The charge the fleet being priced actually carries, annualised off the first half of 2026')
nrow('nebit', 'Normalised EBIT', f'=C{NR["nebitda"]}-C{NR["ndna"]}', NORM['ebit'], N0)
nrow('nnopat', 'Normalised NOPAT',
     f'=C{NR["nebit"]}*(1-{AC("Corporate income tax rate")})', NORM['nopat'], N0, bold=True)
nrow('ncap', 'Capitalisation rate — the cost of capital, with no growth deducted from it',
     f'=DCF!$C${DR["wacc_r"]}',
     NORM['capitalisation_rate'], P2,
     'The first edition divided by the cost of capital LESS a terminal growth rate while the '
     'text beside it said no growth was credited. A denominator of cost-of-capital-minus-growth '
     'IS the growth credit; there is no other reason for the growth rate to appear in it')
nrow('nev', 'Implied enterprise value', f'=C{NR["nnopat"]}/C{NR["ncap"]}',
     NORM['enterprise_value'], N0)
nrow('neq', 'Implied equity value at 30 June 2026',
     f'=C{NR["nev"]}+{AC("Investment in joint ventures")}+{AC("Cash and cash equivalents")}'
     f'-{AC("Borrowings")}-{AC("Lease liabilities")}'
     f'-{AC("Financial liability over the acquired minorities")}',
     NORM['bridge']['equity_30jun26'], N0)
nrow('neqa', 'Accreted to the price anchor at the cost of equity',
     f'=C{NR["neq"]}*(1+DCF!$C${DR["ke_r"]})'
     f'^({AC("Days from 30 June 2026 to the price anchor")}/365)',
     NORM['equity_value'], N0)
nrow('nps', 'Implied value per share (AED)',
     f'=C{NR["neqa"]}/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
     f'*{AC("AED per USD")}', NORM['value_per_share_aed'], N2, bold=True)

n_ += 1
lab(RN, n_, 'BOOK VALUE AND SUSTAINABLE RETURN', bold=True, fill=FILL_B)
n_ += 1
nrow('roe', 'Sustainable return on equity', f'={AC("Sustainable return on equity")}',
     BOOK['roe_sustainable'], P1)
nrow('ke', 'Cost of equity', f'=DCF!$C${DR["ke_r"]}', BOOK['cost_of_equity'], P2)
nrow('gb', 'Growth', f'={AC("Terminal growth — continued expansion")}', BOOK['growth'], P2)
nrow('pb', 'Justified price / book = (return on equity - growth) / (cost of equity - growth)',
     f'=(C{NR["roe"]}-C{NR["gb"]})/(C{NR["ke"]}-C{NR["gb"]})', BOOK['justified_pb'], X2,
     'Highly geared to the denominator: with a cost of equity this close to the growth rate, a '
     'small move in either coordinate moves the multiple a long way. Read the sensitivity, not '
     'the point estimate', bold=True)
nrow('bv', 'Book equity attributable to owners', f'={AC("Book equity attributable to owners")}',
     BOOK['book_equity'], N0)
nrow('bvps', 'Book value per share (AED)',
     f'=C{NR["bv"]}/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
     f'*{AC("AED per USD")}', BOOK['book_equity'] / SH * FX, N2)
nrow('bps', 'Implied value per share (AED)',
     f'=C{NR["pb"]}*C{NR["bvps"]}*(1+C{NR["ke"]})'
     f'^({AC("Days from 30 June 2026 to the price anchor")}/365)',
     BOOK['value_per_share_aed'], N2, bold=True)
nrow('cpb', 'The company\'s own price / book', f'={AC("Market price")}/C{NR["bvps"]}',
     BOOK['current_pb'], X2)

# ==================================================== FUNDAMENTAL VALUATION ===
FVs = ws('Fundamental Valuation', [50, 16, 14, 16, 16, 54])
title(FVs, 'Fundamental valuation — the five lenses and the weighted central', 6)
hdr(FVs, 3, ['Lens', 'Value (AED/sh)', 'Weight', 'Bear', 'Bull', 'What the lens is measuring'])
v_ = 4
LENS_SRC = [
    ('dcf_A', 'Discounted cash flow — continued expansion',
     f"='SOTP Bridge'!B{GR['ps_aed']}", 'Weight — discounted cash flow, continued expansion',
     'Free cash flow to the firm over five explicit years plus a terminal value, on a fleet '
     'plan that keeps building'),
    ('dcf_B', 'Discounted cash flow — capacity plateau',
     f"='SOTP Bridge'!C{GR['ps_aed']}", 'Weight — discounted cash flow, capacity plateau',
     'The same waterfall on a fleet that stops growing once the capacity target is met'),
    ('relative', 'Relative multiples',
     f"='Relative & Normalized'!C{NR['ps']}", 'Weight — relative multiples',
     'What the listed drilling and oilfield-services universe pays for a dollar of EBITDA, '
     'weighted to this company\'s own segment mix'),
    ('book', 'Book value and sustainable return',
     f"='Relative & Normalized'!C{NR['bps']}", 'Weight — book value and sustainable return',
     'The multiple of book a franchise earning this return on equity justifies'),
    ('normalised', 'Normalised earnings power',
     f"='Relative & Normalized'!C{NR['nps']}", 'Weight — normalised earnings power',
     'What the fleet already installed earns at the guided margin, with no growth credited'),
]
LV = {}
for key, nice, formula, wlabel, note in LENS_SRC:
    LV[key] = v_
    lab(FVs, v_, nice)
    fml(FVs, f'B{v_}', formula, FV['by_lens'][key], N2)
    fml(FVs, f'C{v_}', f'={AC(wlabel)}', FV['weights'][key], P1)
    put(FVs, f'D{v_}', FV['lens_range'][key]['bear'], N2)
    put(FVs, f'E{v_}', FV['lens_range'][key]['bull'], N2)
    c = FVs.cell(row=v_, column=6, value=note)
    c.font = F_TXT
    c.alignment = WRAP
    v_ += 1
R_CENTRAL = v_
lab(FVs, v_, 'Weighted central', bold=True)
# The central DIVIDES BY THE SUM OF THE WEIGHTS rather than assuming they add to
# one. At the delivered weights they do add to one, so this changes nothing here;
# it was adopted after the driver test showed that raising a single weight in
# isolation raised the central even for a lens sitting BELOW it, because the
# weights were not renormalising. A reader who re-weights the lenses should get a
# weighted average, not a scaled-up sum.
WSUM = '+'.join(f'C{LV[k]}' for k in LV)
fml(FVs, f'B{v_}', '=(' + '+'.join(f'B{LV[k]}*C{LV[k]}' for k in LV) + f')/({WSUM})',
    FV['central'], N2, bold=True)
fml(FVs, f'C{v_}', '=' + WSUM, 1.0, P1)
fml(FVs, f'D{v_}', '=(' + '+'.join(f'D{LV[k]}*C{LV[k]}' for k in LV) + f')/({WSUM})',
    FV['central_range']['bear'], N2)
fml(FVs, f'E{v_}', '=(' + '+'.join(f'E{LV[k]}*C{LV[k]}' for k in LV) + f')/({WSUM})',
    FV['central_range']['bull'], N2)
v_ += 2
R_SPOT = v_
lab(FVs, v_, 'Market price')
fml(FVs, f'B{v_}', f'={AC("Market price")}', M['spot_aed'], N2)
v_ += 1
lab(FVs, v_, 'Central against the market price')
fml(FVs, f'B{v_}', f'=B{R_CENTRAL}/B{R_SPOT}-1', FV['upside_central'], P1, bold=True)
v_ += 1
lab(FVs, v_, 'Terminal value as a share of enterprise value (continued expansion)')
fml(FVs, f'B{v_}', f'=DCF!$C${TR["tvshare"]}', CA['tv_pct_of_ev'], P1)
R_TVS_FV = v_

# ================================================================== SUMMARY ===
SM = ws('Summary', [50, 16, 14, 16, 16, 46])
title(SM, 'Summary valuation table', 6)
hdr(SM, 3, ['', 'AED per share', 'Weight', 'Bear', 'Bull', 'Note'])
s_ = 4
SMR = {}
for key, nice, note in (
        ('dcf_A', 'Discounted cash flow — continued expansion', ''),
        ('dcf_B', 'Discounted cash flow — capacity plateau', ''),
        ('relative', 'Relative multiples', ''),
        ('book', 'Book value and sustainable return', ''),
        ('normalised', 'Normalised earnings power', '')):
    SMR[key] = s_
    lab(SM, s_, nice)
    fml(SM, f'B{s_}', f"='Fundamental Valuation'!B{LV[key]}", FV['by_lens'][key], N2)
    fml(SM, f'C{s_}', f"='Fundamental Valuation'!C{LV[key]}", FV['weights'][key], P1)
    fml(SM, f'D{s_}', f"='Fundamental Valuation'!D{LV[key]}", FV['lens_range'][key]['bear'], N2)
    fml(SM, f'E{s_}', f"='Fundamental Valuation'!E{LV[key]}", FV['lens_range'][key]['bull'], N2)
    if key in ('dcf_A', 'dcf_B'):
        src = TR['tvshare'] if key == 'dcf_A' else BT['tvshare']
        fml(SM, f'F{s_}', f'=DCF!$C${src}',
            (CA if key == 'dcf_A' else CB)['tv_pct_of_ev'], P1)
        SM.cell(row=s_, column=6).comment = None
    else:
        c = SM.cell(row=s_, column=6, value=note)
        c.font = F_TXT
    s_ += 1
SM_CENTRAL = s_
lab(SM, s_, 'Weighted central fair value', bold=True)
fml(SM, f'B{s_}', "='Fundamental Valuation'!B" + str(R_CENTRAL), FV['central'], N2, bold=True)
fml(SM, f'D{s_}', "='Fundamental Valuation'!D" + str(R_CENTRAL), FV['central_range']['bear'], N2)
fml(SM, f'E{s_}', "='Fundamental Valuation'!E" + str(R_CENTRAL), FV['central_range']['bull'], N2)
s_ += 1
lab(SM, s_, 'Market price, 07-Aug-2026')
fml(SM, f'B{s_}', f'={AC("Market price")}', M['spot_aed'], N2)
s_ += 1
lab(SM, s_, 'Central against the market price')
fml(SM, f'B{s_}', f'=B{SM_CENTRAL}/B{s_-1}-1', FV['upside_central'], P1, bold=True)
s_ += 2
lab(SM, s_, 'MEMO', bold=True, fill=FILL_B)
s_ += 1
for nice, formula, expected, fmt in (
        ('Terminal value as a share of enterprise value — continued expansion',
         f'=DCF!$C${TR["tvshare"]}', CA['tv_pct_of_ev'], P1),
        ('Terminal value as a share of enterprise value — capacity plateau',
         f'=DCF!$C${BT["tvshare"]}', CB['tv_pct_of_ev'], P1),
        ('Weighted average cost of capital — rating basis', f'=DCF!$C${DR["wacc_r"]}',
         W['wacc_rating'], P2),
        ('Weighted average cost of capital — credit-default-swap basis',
         f'=DCF!$C${DR["wacc_c"]}', W['wacc_cds'], P2),
        ('Market capitalisation (USD thousands)', f'=DCF!$C${DR["mcap"]}', M['market_cap_usd_k'],
         N0),
        ('Enterprise value at the market price (USD thousands)',
         f'=DCF!$C${DR["mcap"]}+DCF!$C${DR["nd"]}', M['enterprise_value_usd_k'], N0),
        ('Enterprise value / FY2026 guided EBITDA at the market price',
         f"='Relative & Normalized'!C{NR['own']}", REL['implied_own_ev_ebitda'], X2),
        ('Segment-weighted peer enterprise value / EBITDA',
         f"='Relative & Normalized'!C{NR['blend']}", REL['blended_multiple'], X2)):
    lab(SM, s_, nice)
    fml(SM, f'B{s_}', formula, expected, fmt)
    s_ += 1

# ======================================================== SUMMARY FINANCIALS ==
SF = ws('Summary Financials', [46] + [13] * 8)
title(SF, 'Summary financials — the eight-year picture on one page', 9)
hdr(SF, 3, ["USD thousands unless stated"] + YHDR)
q_ = 4
for nice, src_sheet, src_row, fmt in (
        ('Revenue', 'Income Statement', IROW['revenue'], N0),
        ('EBITDA', 'Income Statement', IROW['ebitda'], N0),
        ('EBITDA margin', 'Income Statement', IROW['margin'], P1),
        ('Profit after tax', 'Income Statement', IROW['pat'], N0),
        ('Net margin', 'Income Statement', IROW['net_margin'], P1),
        ('Capital expenditure', 'Balance Sheet', R_CAPEX, N0),
        ('Net debt', 'Balance Sheet', R_ND, N0),
        ('Net debt / EBITDA', 'Balance Sheet', R_NDE, N2),
        ('Total equity', 'Balance Sheet', R_EQ, N0),
        ('Working capital', 'Balance Sheet', R_WCAP, N0)):
    lab(SF, q_, nice)
    for y in YRS_H + YRS_F:
        c = HCOL.get(y) or FCOL[y]
        e = EXPECT[src_sheet].get(f'{c}{src_row}')
        if e is None:
            cell = (BS if src_sheet == 'Balance Sheet' else IS)[f'{c}{src_row}']
            e = float(cell.value) if isinstance(cell.value, (int, float)) else 0.0
        fml(SF, f'{c}{q_}', f"='{src_sheet}'!{c}{src_row}", e, fmt)
    q_ += 1
q_ += 1
lab(SF, q_, 'Return on invested capital')
for i, y in enumerate(YRS_F):
    c = FCOL[y]
    fml(SF, f'{c}{q_}', f"=DCF!{WCOL[y]}{WR['nopat']}/('Balance Sheet'!{c}{R_EQ}"
                        f"+'Balance Sheet'!{c}{R_ND})", RA[i]['roic'], P1)

# ======================================================= PER-SHARE & RATIOS ===
PS = ws('Per-Share & Ratios', [46] + [13] * 8)
title(PS, 'Per-share figures and ratios — every cell a formula', 9)
hdr(PS, 3, ['AED per share unless stated'] + YHDR)
p_ = 4
lab(PS, p_, 'Earnings per share')
for y in YRS_H + YRS_F:
    c = HCOL.get(y) or FCOL[y]
    e = (H[str(y)]['pat'] if y in HCOL else RA[YRS_F.index(y)]['pat']) / SH * FX
    fml(PS, f'{c}{p_}', f"='Income Statement'!{c}{IROW['pat']}"
                        f'/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
                        f'*{AC("AED per USD")}', e, N3)
p_ += 1
lab(PS, p_, 'Dividend per share')
for y in YRS_F:
    c = FCOL[y]
    e = RA[YRS_F.index(y)]['dividend'] / SH * FX
    fml(PS, f'{c}{p_}', f"=-'Cash Flow'!{c}{CROW['div']}"
                        f'/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
                        f'*{AC("AED per USD")}', e, N3)
p_ += 1
lab(PS, p_, 'Book value per share')
for y in YRS_H + YRS_F:
    c = HCOL.get(y) or FCOL[y]
    e = (H[str(y)]['equity'] if y in HCOL
         else RA[YRS_F.index(y)]['balance_sheet']['equity_residual']) / SH * FX
    fml(PS, f'{c}{p_}', f"='Balance Sheet'!{c}{R_EQ}"
                        f'/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
                        f'*{AC("AED per USD")}', e, N2)
p_ += 1
lab(PS, p_, 'Free cash flow per share')
for y in YRS_F:
    c = FCOL[y]
    e = RA[YRS_F.index(y)]['fcff'] / SH * FX
    fml(PS, f'{c}{p_}', f"='Cash Flow'!{c}{CROW['fcff']}"
                        f'/({AC("Shares issued")}-{AC("Treasury shares held by the market maker")})'
                        f'*{AC("AED per USD")}', e, N3)
p_ += 2
lab(PS, p_, 'RATIOS AT THE CURRENT MARKET PRICE', bold=True, fill=FILL_B)
p_ += 1
lab(PS, p_, 'Price / earnings')
for y in YRS_H + YRS_F:
    c = HCOL.get(y) or FCOL[y]
    eps = (H[str(y)]['pat'] if y in HCOL else RA[YRS_F.index(y)]['pat']) / SH * FX
    fml(PS, f'{c}{p_}', f'={AC("Market price")}/{c}4', M['spot_aed'] / eps, X2)
p_ += 1
lab(PS, p_, 'Dividend yield')
for y in YRS_F:
    c = FCOL[y]
    dps = RA[YRS_F.index(y)]['dividend'] / SH * FX
    fml(PS, f'{c}{p_}', f'={c}5/{AC("Market price")}', dps / M['spot_aed'], P1)
p_ += 1
lab(PS, p_, 'Price / book')
for y in YRS_H + YRS_F:
    c = HCOL.get(y) or FCOL[y]
    bvps = (H[str(y)]['equity'] if y in HCOL
            else RA[YRS_F.index(y)]['balance_sheet']['equity_residual']) / SH * FX
    fml(PS, f'{c}{p_}', f'={AC("Market price")}/{c}6', M['spot_aed'] / bvps, X2)

# ============================================================ PEER & SECTOR ===
PR = ws('Peer & Sector', [30, 34, 15, 15, 15, 15, 15, 14])
title(PR, 'Peer and sector — third-party market data, used for cross-check and for the '
          'relative lens only', 8)
c = PR.cell(row=2, column=1,
            value='Every figure on this sheet belongs to another company. Nothing here enters '
                  'ADNOC Drilling\'s own income statement, balance sheet or cash flow — those '
                  'are built exclusively from its own audited filings. Prices are at the close '
                  'on 07-Aug-2026; balance-sheet and earnings figures are each peer\'s own '
                  'latest reported period.')
c.font = F_TXT
c.alignment = WRAP
PR.merge_cells('A2:H2')
PR.row_dimensions[2].height = 40
hdr(PR, 4, ['Group', 'Company', 'Market cap (USD mn)', 'Net debt (USD mn)',
            'Enterprise value (USD mn)', 'Last 12 months EBITDA (USD mn)', 'EV / EBITDA',
            'Latest period'])
pr = 5
for row in REL['peers']:
    lab(PR, pr, row['group'])
    lab(PR, pr, row['name'], col=2)
    put(PR, f'C{pr}', row['market_cap_usd_mn'], N0)
    put(PR, f'D{pr}', row['net_debt_usd_mn'], N0)
    fml(PR, f'E{pr}', f'=C{pr}+D{pr}', row['ev_usd_mn'], N0)
    put(PR, f'F{pr}', row['ltm_ebitda_usd_mn'], N0)
    fml(PR, f'G{pr}', f'=E{pr}/F{pr}', row['ev_ebitda'], X2)
    lab(PR, pr, row['latest_period'], col=8)
    pr += 1
pr += 1
lab(PR, pr, 'ADNOC Drilling at the market price', bold=True)
lab(PR, pr, 'ADNOC Drilling', col=2, bold=True)
fml(PR, f'C{pr}', f'=DCF!$C${DR["mcap"]}/1000', M['market_cap_usd_k'] / 1000, N0)
fml(PR, f'D{pr}', f'=DCF!$C${DR["nd"]}/1000', W['net_debt'] / 1000, N0)
fml(PR, f'E{pr}', f'=C{pr}+D{pr}', M['enterprise_value_usd_k'] / 1000, N0)
fml(PR, f'F{pr}', f'={AC("Last-twelve-month EBITDA excluding joint ventures")}/1000', REL['applied_ebitda'] / 1000,
    N0)
fml(PR, f'G{pr}', f'=E{pr}/F{pr}', REL['implied_own_ev_ebitda'], X2, bold=True)

# =============================================================== MONTE CARLO ==
MC = ws('Monte Carlo', [46, 18, 18, 60])
title(MC, 'Probability map — a whole-model re-run; these cells do NOT redraw when a driver '
          'changes', 4)
c = MC.cell(row=2, column=1,
            value='Each figure below is the outcome of 50,000 simulated price paths struck on '
                  '07-Aug-2026. They are pasted because every one of them is a complete '
                  're-simulation, not an arithmetic function of a cell on the Assumptions '
                  'sheet. Changing a valuation driver does not move them.')
c.font = F_TXT
c.alignment = WRAP
MC.merge_cells('A2:D2')
MC.row_dimensions[2].height = 40
hdr(MC, 4, ['', 'One month', 'Three months', 'Note'])
mrow = 5
lab(MC, mrow, 'Check date')
put(MC, f'B{mrow}', ST['horizons']['1M']['grade_date'])
put(MC, f'C{mrow}', ST['horizons']['3M']['grade_date'])
mrow += 1
for p in ('p5', 'p25', 'p50', 'p75', 'p95'):
    lab(MC, mrow, f'{p[1:]}th percentile (AED)')
    put(MC, f'B{mrow}', ST['horizons']['1M']['pct'][p], N2)
    put(MC, f'C{mrow}', ST['horizons']['3M']['pct'][p], N2)
    mrow += 1
for key, nice in (('p_above', 'Probability of finishing above the current price'),
                  ('p_up10', 'Probability of finishing 10% or more above'),
                  ('p_dn10', 'Probability of finishing 10% or more below'),
                  ('touch_up10', 'Probability of touching +10% at any point'),
                  ('touch_dn10', 'Probability of touching -10% at any point'),
                  ('touch_up20', 'Probability of touching +20% at any point'),
                  ('touch_dn20', 'Probability of touching -20% at any point')):
    lab(MC, mrow, nice)
    put(MC, f'B{mrow}', ST['horizons']['1M'][key], P1)
    put(MC, f'C{mrow}', ST['horizons']['3M'][key], P1)
    mrow += 1
lab(MC, mrow, 'Annualised volatility at the anchor')
put(MC, f'B{mrow}', ST['horizons']['1M']['anchor_vol_ann'], P1)
put(MC, f'C{mrow}', ST['horizons']['3M']['anchor_vol_ann'], P1)

# ================================================================ SENSITIVITY =
SN = ws('Sensitivity', [30] + [14] * 8)
title(SN, 'Sensitivity — whole-model re-runs; these grids do NOT redraw when a driver changes', 9)
c = SN.cell(row=2, column=1,
            value='Each cell below is a complete revaluation of the model at a different pair of '
                  'inputs, computed and pasted. They are the one place in this workbook where a '
                  'number is not a live formula, and that is deliberate: a grid of full model '
                  're-runs cannot be expressed as arithmetic over a single set of driver cells.')
c.font = F_TXT
c.alignment = WRAP
SN.merge_cells('A2:I2')
SN.row_dimensions[2].height = 40
sn = 4
lab(SN, sn, 'Discounted-cash-flow value (AED/share) — cost of capital against terminal growth',
    bold=True, fill=FILL_B)
sn += 1
for j, g in enumerate(SENS['g_grid']):
    cc = SN.cell(row=sn, column=2 + j, value=g)
    cc.number_format = P2
    cc.font = F_H2
    cc.alignment = Alignment(horizontal='center')
sn += 1
for i, wv in enumerate(SENS['wacc_grid']):
    cc = SN.cell(row=sn + i, column=1, value=wv)
    cc.number_format = P2
    cc.font = F_H2
    for j, _ in enumerate(SENS['g_grid']):
        put(SN, f'{get_column_letter(2+j)}{sn+i}', SENS['matrix'][i][j], N2)
sn += len(SENS['wacc_grid']) + 2
lab(SN, sn, 'EBITDA-margin shift (percentage points of revenue)', bold=True, fill=FILL_B)
sn += 1
for j, m in enumerate(SENS['margin_shift']):
    cc = SN.cell(row=sn, column=2 + j, value=m['shift'])
    cc.number_format = P1
    cc.font = F_H2
    put(SN, f'{get_column_letter(2+j)}{sn+1}', m['aed'], N2)
sn += 3
lab(SN, sn, 'Equity beta', bold=True, fill=FILL_B)
sn += 1
for j, bpt in enumerate(SENS['beta_grid']):
    cc = SN.cell(row=sn, column=2 + j, value=bpt['beta'])
    cc.number_format = N3
    cc.font = F_H2
    put(SN, f'{get_column_letter(2+j)}{sn+1}', bpt['aed'], N2)

# ================================================================ READ FIRST ==
RF = wb.create_sheet('READ FIRST', 0)
EXPECT['READ FIRST'] = {}
RF.column_dimensions['A'].width = 4
RF.column_dimensions['B'].width = 116
title(RF, 'ADNOC Drilling Company P.J.S.C. — valuation model, 9 August 2026', 2)
READ = [
    ('THIS WORKBOOK CALCULATES.', True),
    ('Everything arithmetically derivable from a driver is a live formula. Change a blue cell '
     'on the Assumptions sheet and the workbook reprices: the cost of capital is built in the '
     'sheet from the risk-free rate net of the sovereign default spread, the equity beta and '
     'the equity risk premium; the cost of debt is built from the five-year Treasury yield and '
     'the company\'s own borrowing margin and then taxed; the weights come from net debt and '
     'market capitalisation; the discount factors compound; the waterfall chains EBITDA to '
     'EBIT to NOPAT to free cash flow to present value; the terminal block derives its '
     'reinvestment rate from growth over return on capital; the statements roll forward; and '
     'every ratio and per-share figure is a formula.', False),
    ('THREE CLASSES OF CELL ARE PASTED, AND ONLY THREE.', True),
    ('1. Audited and disclosed history. The three years of income statement, balance sheet and '
     'cash flow come from the company\'s own signed consolidated financial statements. Where a '
     'line is both disclosed and derivable, the DISCLOSED figure is carried — the model does '
     'not overwrite an audited number with its own arithmetic.', False),
    ('2. The output of the unit build. The rig delivery schedule on the Assumptions sheet is a '
     'fleet plan, not arithmetic: it says how many rigs of each class are expected in service '
     'in each year. Everything downstream of it — average deployed rigs, revenue per class, '
     'the cost stack that serves them — is a formula.', False),
    ('3. Whole-model re-runs. The Monte Carlo sheet, the three grids on the Sensitivity sheet, '
     'and the Bear and Bull columns on the Summary and Fundamental Valuation sheets are pasted, '
     'because each cell in them is a complete revaluation at a different set of inputs. THESE '
     'CELLS DO NOT REDRAW WHEN A DRIVER CHANGES. Everything else does — including both '
     'discounted-cash-flow cases, which are each built live from their own fleet schedule, '
     'their own capital-expenditure plan and the same cost stack.', False),
    ('One sheet is other companies\' data. ', True),
    ('The Peer and Sector sheet carries thirteen other listed drillers and oilfield-service '
     'companies. Their prices, earnings and balance sheets are theirs, not ADNOC Drilling\'s, '
     'and are pasted as observations. The only formulas on that sheet are the enterprise values '
     'and multiples computed from them. Nothing on it enters ADNOC Drilling\'s own statements.',
     False),
    ('BLUE MEANS INPUT, BLACK MEANS FORMULA.', True),
    ('Every blue cell is on the Assumptions sheet, and every one of them carries its source '
     'and the construction behind it in the right-hand column.', False),
    ('WHERE THE NUMBERS COME FROM.', True),
    ('Every historical figure traces to ADNOC Drilling\'s own issued financial statements: the '
     'audited consolidated statements for 2023, 2024 and 2025 and the reviewed interim '
     'statements for the first quarter and first half of 2026, each downloaded from the '
     'company\'s investor-relations site. Operating units — rig counts, wells drilled, '
     'integrated-services rig counts, unconventional revenue — come from the company\'s own '
     'quarterly management discussion and analysis, which is the only published source that '
     'carries them. The Peer and Sector sheet is the one place third-party data appears; it is '
     'other companies\' data, used for cross-check and for the relative lens, and none of it '
     'enters ADNOC Drilling\'s own statements.', False),
    ('CURRENCY.', True),
    ('The company reports in US dollars, which is its functional and presentation currency. Its '
     'shares trade in dirhams on the Abu Dhabi Securities Exchange. The model runs in dollars '
     'and converts per-share figures at the peg of 3.6725 dirhams to the dollar.', False),
    ('WHAT THIS IS NOT.', True),
    ('This workbook contains no rating and no price target. It produces a fair-value range and '
     'a distribution of outcomes.', False),
]
rrow = 3
for text, is_head in READ:
    c = RF.cell(row=rrow, column=2, value=text)
    c.font = F_H2 if is_head else F_TXT
    c.alignment = WRAP
    RF.row_dimensions[rrow].height = 16 if is_head else max(30, 13 * (len(text) // 105 + 1))
    rrow += 1 if is_head else 2

# ------------------------------------------- the deferred Assumptions rows ---
# Two rows on the Assumptions sheet are formulas over cells on sheets written
# after it. They are filled in here, and DEFERRED is then asserted empty so that
# a row placed with adef() and never resolved cannot ship as a blank cell.
resolve_deferred('EBITDA — FY2025 audited', f"='Income Statement'!$D${IROW['ebitda']}")
_bk_eq_2030, _bk_eq_2029 = FCOL[2030], FCOL[2029]
resolve_deferred(
    'Sustainable return on equity',
    f"='Income Statement'!{_bk_eq_2030}{IROW['pat']}"
    f"/(('Balance Sheet'!{_bk_eq_2029}{R_EQ}+'Balance Sheet'!{_bk_eq_2030}{R_EQ})/2)")
assert not DEFERRED, [d[0] for d in DEFERRED]

# --------------------------------------------------------------- sheet order -
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios',
         'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]
assert [s.title for s in wb.worksheets] == ORDER, 'sheet order'
assert len(wb.worksheets) == 16, len(wb.worksheets)

wb.save(OUT_XLSX)
anchors = dict(central_row=SM_CENTRAL, lens_rows=LV, dcf=DR, waterfall=WR, terminal=TR,
               plateau=BT, bridge=GR, relnorm=NR, segments=dict(
                   revenue=R_REV_TOT, ebitda=R_EBITDA, ebitda_ex_jv=R_EBITDA_XJV,
                   conv=R_CONV),
               balance=dict(equity=R_EQ, nci=R_NCI, net_debt=R_ND, dna=R_DNA, capex=R_CAPEX,
                            wcap=R_WCAP, total_assets=R_TA, total_liabilities=R_TL),
               income=IROW, cash=CROW, fundamental_central=R_CENTRAL)
json.dump(dict(expected=EXPECT, anchors=anchors),
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nf = sum(len(v) for v in EXPECT.values())
print(f'wrote {OUT_XLSX}')
print(f'  16 sheets, {nf} formula cells recorded against the model')
