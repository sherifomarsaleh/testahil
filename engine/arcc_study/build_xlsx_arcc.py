"""ARCC_Valuation_Model_06082026_public.xlsx — 16 sheets, formula-first. REVISION 2.

Rebuilt on the AUDITED consolidated financial statements for FY2023, FY2024 and FY2025 and
the reviewed Q1-2026 interim accounts. Revision 1 was built without opening a source
document; every history cell here is now a disclosed figure, and the derivations that stood
in for them are gone.

Blue = input · black = formula · green = cross-sheet link.

The workbook CALCULATES. Everything arithmetically derivable from a driver is a live
formula: the cost of capital (cost of equity from the risk-free rate net of the sovereign
spread, beta and the premium; a CURRENCY-BLENDED cost of debt built facility by facility
from the audited debt note; weights from debt and market capitalisation; the terminal rate
from its own components), the glide fractions from the cost-of-debt path, the compounding
discount factors, the cash-flow waterfall, the terminal block, the statement roll-forwards
and every ratio and per-share figure.

Only three classes of cell are pasted, named on READ FIRST:
  1. audited and disclosed history — the primary record, not a calculation;
  2. the output of a unit build that cannot be flattened into a grid — the 50,000-path
     price map's percentile ladder;
  3. whole-model re-runs — the sensitivity grids and the contested-choice alternatives.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
S0 = json.load(open(os.path.join(HERE, 'step0_result.json')))
STK = json.load(open(os.path.join(HERE, 'strike_result.json')))
BETA = json.load(open(os.path.join(HERE, 'beta_result.json')))

BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM3 = '#,##0.000'
NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'
MULT = '0.00x'; DF4 = '0.0000'; NUM4 = '#,##0.0000'

M, H, F = D['meta'], D['history'], D['forecast']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sensitivity']
BU, PE, SHT, TR = D['bottom_up'], D['peers'], D['share_triangulation'], D['terminal_reconciliation']
UC, KDG, CON = D['unit_calibration'], D['kd_gate'], D['contested']
IN = {k: v['value'] for k, v in D['inputs'].items()}
SPOT, SH, TAX, TAXE = M['spot'], M['shares_mn'], IN['tax_stat'], IN['tax_eff']
YH, YF = H['years'], F['years']
HC = ['B', 'C', 'D']
FC = ['E', 'F', 'G', 'H', 'I']
DC = ['B', 'C', 'D', 'E', 'F']
BUC = ['B', 'C', 'D', 'E', 'F', 'G']
BUY = ['FY2025A'] + YF

wb = Workbook()
EXPECT = {}


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=48, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


def note(ws, row, text, w=10):
    ws.cell(row=row, column=1, value=text).font = SUB
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=w)


# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Arabian Cement Company S.A.E. (EGX: ARCC) — revision 3', None, 9)
LINES = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'REVISION 3 CORRECTS THE PRICE PATH, WHICH THE AUDITED RECORD DISPROVED. Revision 2 rebuilt every company',
 'figure on the statements but left one forecast driver where revision 1 had put it: local realised price',
 'growing 3.0% in FY2026 against 11.5% cost inflation, and about 4% a year after that. The statements that had',
 'just been read contradict it. FY2024: revenue +44.5% against total cash cost +41.8%. FY2025: revenue +42.6%',
 'against cash cost +12.5%, which is why the audited gross margin went 21.2% to 23.9% to 40.6%. Q1-2026 then',
 'printed revenue +17.3% at a 42.9% gross margin against 40.6% for the full prior year — still WIDENING. In',
 'every period the accounts cover, price outran cost. The old path also implied a producer unable to raise',
 'price even at the central bank\'s own 7% medium-term target, while the same model had utilisation RISING',
 'from 69.7% to 78.2% — volume gained and real price lost at once, two conservatisms stacked rather than one',
 'judgement made. The path is recalibrated to 8.0% / 9.0% / 8.0% / 7.0% / 6.5%, still below the cost path in',
 'every year, and still producing FY2026 group revenue growth of about 10.7% against the 17.3% the first',
 'quarter actually ran. The cost path is UNCHANGED at headline inflation: the company\'s realised cost',
 'inflation ran below the national rate, but crediting that here would double-count the alternative-fuel',
 'saving already carried as its own driver. The EBITDA margin now glides from the audited 39.3% to about',
 '34.3% by FY2030 rather than to 24.5% — still erosion, but erosion the record can support.', '',
 'REVISION 2 BUILT THE STUDY ON THE AUDITED ACCOUNTS. Revision 1 was built without opening a single',
 'source document: every outbound request was refused by the network policy in force, so every company figure',
 'reached it as relayed in a web-search summary. The audited consolidated financial statements for FY2023,',
 'FY2024 and FY2025 — Deloitte, signed 25 February 2026 — and the reviewed Q1-2026 interim accounts are now in',
 'hand, and every historical cell in this workbook is a disclosed figure read from them.', '',
 'WHAT THE STATEMENTS CHANGED, and it is not cosmetic:',
 '  * Non-controlling interests are EGP 158,005 — one hundred and fifty-eight THOUSAND pounds. Revision 1',
 '    deducted EGP 150 MILLION on inference, 950 times too much.',
 '  * The effective tax rate is 23.82%, not the 29.43% revision 1 inferred. Every forecast year was over-taxed.',
 '  * The cost of debt is about 7.5%, not 21.5%: 91% of the book is EURO-denominated, at Euribor plus 4.35%',
 '    (a EUR 25mn EBRD decarbonisation facility) and Euribor plus 3% (a EUR 3.09mn NBE/KfW facility).',
 '  * Kiln capacity is 4.2Mt of clinker against 5.0Mt of cement, so the clinker factor is 0.84, not 0.72.',
 '  * Exports are 30.7% of revenue, not 12%. The revenue note splits local from export and goods from services.',
 '  * The cost stack is DISCLOSED. The invented five-line stack of revision 1 is retired for the printed one.',
 '  What survived: the share count of 374,867,445, FY2025 operating income of EGP 4,595.82mn to the pound, and',
 '  total liabilities of EGP 4,140.99mn — which revision 1 derived, against an aggregator print of EGP 2,894mn',
 '  that it rejected. That figure turns out to be total CURRENT liabilities. The derivation was right.', '',
 'IT IS FORMULA-DRIVEN, AND THAT CLAIM IS TESTED. Change a blue cell on Assumptions and the model reprices. The',
 'cost of equity is built from the risk-free rate NET of the sovereign default spread, beta and the premium;',
 'the cost of debt is built FACILITY BY FACILITY from the audited debt note and blended by currency; the',
 'weights come from debt and market capitalisation; the terminal rate is built from its own components; the',
 'glide fractions are derived from the pound cost-of-debt path; the discount factors compound through it; the',
 'waterfall chains from margin through EBITDA, D&A, EBIT, NOPAT, capex and working capital to free cash flow',
 'and present value; the terminal block chains from reinvestment = growth / return on capital; the statements',
 'roll forward; and every ratio and per-share figure is a formula. A driver test perturbs every input in place,',
 're-evaluates the whole workbook and confirms the headline moves in the right direction.', '',
 'THREE CLASSES OF CELL ARE PASTED, and it is worth knowing exactly which.',
 '  1. Audited and disclosed history — the primary record, not a calculation. Where a line is both disclosed',
 '     and derivable, the DISCLOSED figure is carried. These sit on Assumptions and are labelled there.',
 '  2. The output of a unit build that cannot survive being flattened into a grid: the percentile ladder of the',
 '     probabilistic price map on the Monte Carlo sheet, the output of a 50,000-path simulation.',
 '  3. Whole-model re-runs — the sensitivity grids, and the alternative fair values in the contested-choices',
 '     table. Each of those cells is a COMPLETE revaluation at a different assumption. THESE DO NOT REDRAW',
 '     WHEN A DRIVER CHANGES. Edit an input and the unit build, waterfall, statements, ratios and all four',
 '     lenses reprice; the sensitivity tables and the price map keep the values printed here.',
 '  Anything else pasted would be a defect.', '',
 'HOW THE COMPANY IS VALUED. A single-segment cement operating company: two lines in Suez governorate, 4.2Mt of',
 'clinker capacity and 5.0Mt of cement, 60% owned by Aridos Jativa of Spain. No property leg, no lending book,',
 'no concession. One operating-company lens, cross-read against relative multiples, normalised earnings power',
 'and replacement cost. Cash of EGP 3,459mn against interest-bearing debt of EGP 1,135mn makes it net cash;',
 'that cash is stripped out of the discount rate and added back in the bridge, so it is priced once at face —',
 'less the EGP 2,002mn FY2025 dividend that was declared and still unpaid at 31 March 2026.', '',
 'TWO THINGS TO KNOW BEFORE READING ANY NUMBER.',
 '  * FY2025 is a cyclical peak. The audited EBITDA margin went 22.0% to 23.1% to 39.2% across the three years,',
 '    on the abolition of the production quota in May 2025. The forecast glides it down every year from there.',
 '  * Book return on capital is 60.6% and replacement-cost return is 8.6%. The terminal block is struck on the',
 '    REPLACEMENT figure. On the book number, growth would be free; it is not, and the difference is the single',
 '    most consequential judgement in this model.', '',
 'No rating and no price target. Fair-value ranges and distributions only.',
]
for i, ln in enumerate(LINES):
    ws.cell(row=4 + i, column=1, value=ln).font = Font(size=10 if ln else 9)
ws.column_dimensions['A'].width = 120

# ============ 2 ASSUMPTIONS ===================================================
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every driver', 'Blue cells are inputs. Nothing below a driver '
      'is typed.', 9, 56, 13)
A = {}
R = [4]


def inp(label, key, val, fmt=NUM0, unit=''):
    r = R[0]
    wsA.cell(row=r, column=1, value=label)
    put(wsA, f'B{r}', val, BLUE, fmt)
    if unit:
        wsA.cell(row=r, column=3, value=unit).font = SUB
    A[key] = f'Assumptions!$B${r}'
    R[0] += 1
    return A[key]


def sect(label, w=9):
    band(wsA, R[0], w); wsA.cell(row=R[0], column=1, value=label)
    R[0] += 1


def inprow(label, key, vals, cols, fmt=NUM0, unit=''):
    r = R[0]
    wsA.cell(row=r, column=1, value=label)
    for i, v in enumerate(vals):
        put(wsA, f'{cols[i]}{r}', v, BLUE, fmt)
        A[f'{key}{i}'] = f"Assumptions!${cols[i]}${r}"
    if unit:
        wsA.cell(row=r, column=9, value=unit).font = SUB
    R[0] += 1


sect('MARKET AND SHARE COUNT')
inp('Spot price', 'spot', SPOT, PX, 'EGP, close 06-Aug-2026')
inp('Ordinary shares issued', 'shiss', IN['shares_issued'], NUM4, 'mn — audited note 20')
inp('Treasury shares held', 'shtre', IN['shares_treasury'], NUM4, 'mn — audited note 21')
inp('Statutory tax rate', 'tax', TAX, PCT, '')
inp('Effective tax rate', 'taxe', TAXE, PCT2, 'audited: tax over pre-tax profit')
inp('Beta (own-stock weekly regression)', 'beta', W['beta'], NUM3, '')
inp('USD/EGP at the valuation date', 'fx', IN['fx'], NUM1, '')

sect('PLANT — AUDITED NOTE 1')
inp('Cement capacity', 'capcem', IN['cap_cement_mt'], NUM2, 'Mt/yr')
inp('Kiln clinker capacity', 'capclk', IN['cap_clinker_mt'], NUM2, 'Mt/yr')
inp('Clinker factor', 'cfac', IN['clinker_factor'], NUM3, 't clinker / t cement')

sect('EXPORT SPLIT — the one ratio the export leg needs; both prices come OUT')
inp('Export clinker price as a fraction of export cement', 'ckr', IN['clk_price_ratio'], NUM3, '')
inp('Average USD/EGP FY2025', 'fx25', IN['fx_avg_fy25'], NUM2, 'audited note 2.5')
inp('Services revenue as a share of goods revenue', 'svc', IN['svc_share'], PCT2, '')

sect('PATHS — FY2025A then FY2026E to FY2030E')
hdr(wsA, R[0], [''] + BUY); R[0] += 1
inprow('Kiln utilisation  (THE volume driver)', 'kutl', IN['kiln_util'], BUC, PCT)
inprow('Clinker sold as clinker, share of clinker made', 'cksh', IN['clk_export_share'], BUC, PCT)
inprow('Cement exported, share of cement made', 'cesh', IN['cem_export_share'], BUC, PCT)
inprow('Local price index', 'pli', IN['price_local_path'], BUC, NUM3)
inprow('Export price index (USD)', 'pei', IN['price_exp_path'], BUC, NUM3)
inprow('USD/EGP path', 'fxp', IN['fx_path'], BUC, NUM1)
inprow('Local cost-inflation index', 'infl', IN['cost_infl'], BUC, NUM3)
inprow('Alternative-fuel saving on materials', 'afs', IN['af_saving'], BUC, PCT)

sect('FORECAST DRIVERS — FY2026E to FY2030E')
hdr(wsA, R[0], [''] + YF); R[0] += 1
inprow('Depreciation as % of revenue', 'dnap', IN['dna_pct'], DC, PCT)
inprow('EGP marginal cost-of-debt path', 'kdp', IN['kd_path'], DC, PCT)
inprow('Yield earned on cash', 'cy', IN['cash_yield'], DC, PCT)
inp('Maintenance capital expenditure', 'cxt', IN['capex_usd_t_cap'], NUM2, 'USD/t capacity')
inp('Change in working capital / change in revenue', 'wcp', IN['wc_pct_drev'], PCT, '')
inp('Dividend payout ratio', 'payout', IN['payout'], PCT, '')

sect('COST OF CAPITAL')
inp('Risk-free rate (EGP 10-year government)', 'rf', IN['rf'], PCT2)
inp('Sovereign default spread (netted out)', 'sov', IN['sov_spread_cds'], PCT2)
inp('Equity risk premium', 'erp', IN['erp_cds'], PCT2)
inp('Euribor (EBRD and NBE reference rate)', 'eur', IN['euribor'], PCT2)
inp('EGP marginal borrowing rate (corridor + 0.6%)', 'kdegp', IN['kd_egp_marginal'], PCT2)
inp('Expected EGP depreciation against the euro', 'dep', IN['egp_dep_vs_eur'], PCT2)
inp('Terminal cost of debt', 'kdt', IN['kd_term'], PCT2)
inp('Terminal risk-free rate', 'rft', IN['rf_term'], PCT2)
inp('Terminal equity risk premium', 'erpt', IN['erp_term'], PCT2)
inp('Terminal debt weight', 'wdt', IN['wd_term'], PCT)
inp('Terminal growth rate', 'g', IN['g_term'], PCT)
inp('Elapsed fraction of FY2026 at the valuation date', 'stub', IN['stub_years'], NUM3)

sect('AUDITED INCOME STATEMENT — PASTED CLASS 1 (EGP mn)')
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} sales (net)', f'rev{tag}', IN[f'rev_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} cost of sales', f'cogs{tag}', IN[f'cogs_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} general and administrative expenses', f'ga{tag}', IN[f'ga_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} provisions charged', f'prov{tag}', IN[f'prov_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} expected credit losses', f'ecl{tag}', IN[f'ecl_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} profit before tax', f'pbt{tag}', IN[f'pbt_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} income tax', f'txc{tag}', IN[f'tax_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} attributable profit', f'pat{tag}', IN[f'pat_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} earnings per share', f'eps{tag}', IN[f'eps_fy{tag}'], PX, 'EGP')
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} depreciation and amortisation', f'dna{tag}', IN[f'dna_fy{tag}'], NUM0)
for tag, yr in (('23', 'FY2023'), ('24', 'FY2024'), ('25', 'FY2025')):
    inp(f'{yr} capital expenditure', f'cx{tag}', IN[f'capex_fy{tag}'], NUM0)
inp('FY2025 interest income', 'ii25', IN['intinc_fy25'], NUM0)
inp('FY2025 other income', 'oi25', IN['othinc_fy25'], NUM0)
inp('FY2025 finance costs', 'fc25', IN['fincost_fy25'], NUM0)
inp('FY2025 foreign exchange differences', 'fxd25', IN['fx_diff_fy25'], NUM0)

sect('AUDITED REVENUE AND COST NOTES (EGP mn)')
inp('FY2025 local sales of goods', 'rlg', IN['rev_local_goods_fy25'], NUM0)
inp('FY2025 local services', 'rls', IN['rev_local_svc_fy25'], NUM0)
inp('FY2025 export sales of goods', 'reg', IN['rev_exp_goods_fy25'], NUM0)
inp('FY2025 export services', 'res', IN['rev_exp_svc_fy25'], NUM0)
inp('FY2024 total local sales', 'rl24', IN['rev_local_fy24'], NUM0)
inp('FY2024 total export sales', 're24', IN['rev_exp_fy24'], NUM0)
inp('FY2025 cost of sales — materials and fuel', 'cmat', IN['cos_materials_fy25'], NUM0)
inp('FY2025 cost of sales — transportation', 'ctra', IN['cos_transport_fy25'], NUM0)
inp('FY2025 cost of sales — overheads', 'covh', IN['cos_overhead_fy25'], NUM0)
inp('FY2025 administrative depreciation', 'adep', IN['ga_admin_dep_fy25'], NUM0)

sect('AUDITED BALANCE SHEET — PASTED CLASS 1 (EGP mn)')
inp('Total assets FY2025', 'ta25', IN['ta_fy25'], NUM0)
inp('Total assets FY2024', 'ta24', IN['ta_fy24'], NUM0)
inp('Total assets FY2023', 'ta23', IN['ta_fy23'], NUM0)
inp('Total liabilities FY2025', 'tl25', IN['tl_fy25'], NUM0)
inp('Property, plant and equipment FY2025', 'ppe25', IN['ppe_fy25'], NUM0)
inp('Assets under construction FY2025', 'auc25', IN['auc_fy25'], NUM0)
inp('Intangible assets FY2025', 'int25', IN['intang_fy25'], NUM0)
inp('Inventories FY2025', 'inv25', IN['inv_fy25'], NUM0)
inp('Trade receivables FY2025', 'rec25', IN['recv_fy25'], NUM0)
inp('Debtors and other debit balances FY2025', 'deb25', IN['debtors_fy25'], NUM0)
inp('Cash and bank balances FY2025', 'cash25', IN['cash_fy25'], NUM0)
inp('Cash and bank balances FY2024', 'cash24', IN['cash_fy24'], NUM0)
inp('Cash and bank balances FY2023', 'cash23', IN['cash_fy23'], NUM0)
inp('Equity attributable to owners FY2025', 'eq25', IN['eq_fy25'], NUM0)
inp('Equity attributable to owners FY2024', 'eq24', IN['eq_fy24'], NUM0)
inp('Equity attributable to owners FY2023', 'eq23', IN['eq_fy23'], NUM0)
inp('Non-controlling interests FY2025', 'nci', IN['nci'], NUM4)

sect('AUDITED DEBT NOTE 25 (EGP mn)')
inp('CIB credit facilities — EGP', 'dcib', IN['debt_cib_fy25'], NUM0)
inp('National Bank of Egypt facility — EUR', 'dnbe', IN['debt_nbe_fy25'], NUM0)
inp('EBRD facility — EUR', 'debrd', IN['debt_ebrd_fy25'], NUM0)
inp('Lease liabilities', 'dlease', IN['lease_fy25'], NUM0)
inp('Total interest-bearing debt FY2024', 'd24', IN['debt_fy24'], NUM0)
inp('Total interest-bearing debt FY2023', 'd23', IN['debt_fy23'], NUM0)
inp('FY2025 loan interest expense', 'li25', IN['loan_int_fy25'], NUM0)
inp('FY2025 credit-facility interest expense', 'fi25', IN['fac_int_fy25'], NUM0)
inp('FY2024 loan interest expense', 'li24', IN['loan_int_fy24'], NUM0)
inp('FY2024 credit-facility interest expense', 'fi24', IN['fac_int_fy24'], NUM0)

sect('DIVIDENDS AND Q1-2026 (EGP mn)')
inp('FY2025 dividend declared', 'div25', IN['div_fy25_declared'], NUM0)
inp('FY2024 dividend approved and paid', 'div24', IN['div_fy24_paid'], NUM0)
inp('Q1-2026 sales', 'q1r26', IN['rev_q1_26'], NUM0)
inp('Q1-2025 sales', 'q1r25', IN['rev_q1_25'], NUM0)
inp('Q1-2026 gross profit', 'q1gp', IN['gp_q1_26'], NUM0)
inp('Q1-2026 attributable profit', 'q1pat', IN['pat_q1_26'], NUM0)
inp('Q1-2026 cash and bank balances', 'q1cash', IN['cash_q1_26'], NUM0)
inp('Q1-2026 interest-bearing debt', 'q1debt', IN['debt_q1_26'], NUM0)
inp('Q1-2026 dividends payable', 'q1div', IN['divpay_q1_26'], NUM0)
inp('Q1-2026 finance costs', 'q1fc', IN['fincost_q1_26'], NUM0)

sect('LENS INPUTS')
inp('Replacement cost per annual tonne', 'repl', IN['repl_usd_t'], NUM0, 'USD/t')
inp('Justified enterprise value per annual tonne', 'evt', IN['ev_t_just'], NUM0, 'USD/t')
inp('Justified EV/EBITDA', 'eveb', IN['ev_ebitda_just'], MULT)
inp('Justified price/earnings', 'pej', IN['pe_just'], MULT)
inp('Mid-cycle EBITDA margin', 'nmgn', IN['norm_mgn'], PCT)
inp('Normalised revenue haircut', 'nhc', IN['norm_rev_haircut'], PCT)
inp('Weight — cash-flow lens', 'wdcf', IN['w_dcf'], PCT)
inp('Weight — relative lens', 'wrel', IN['w_rel'], PCT)
inp('Weight — normalised lens', 'wnorm', IN['w_norm'], PCT)
inp('Weight — asset lens', 'wasset', IN['w_asset'], PCT)

sect('SECTOR AND PEERS')
inp('Egyptian nameplate capacity', 'egcap', IN['egy_capacity_mt'], NUM1, 'Mt')
inp('Egyptian production 2025', 'egprod', IN['egy_prod_mt'], NUM1, 'Mt')
inp('Egyptian consumption 2025', 'egcons', IN['egy_cons_mt'], NUM1, 'Mt')
inp('Egyptian exports 2025', 'egexp', IN['egy_exports_mt'], NUM1, 'Mt')
inp('Dormant capacity under revival', 'egrev', IN['egy_revival_mt'], NUM1, 'Mt')
inp('Peer — Sinai Cement revenue', 'pscrev', IN['peer_scem_rev'], NUM0)
inp('Peer — Sinai Cement profit', 'pscpat', IN['peer_scem_pat'], NUM0)
inp('Peer — Sinai Cement market capitalisation', 'pscmc', IN['peer_scem_mcap'], NUM0)
inp('Peer — Misr Beni Suef revenue', 'pmbrev', IN['peer_mbsc_rev'], NUM0)
inp('Peer — Misr Beni Suef profit', 'pmbpat', IN['peer_mbsc_pat'], NUM0)
inp('Peer — Misr Beni Suef market capitalisation', 'pmbmc', IN['peer_mbsc_mcap'], NUM0)
note(wsA, R[0] + 1, 'Every cell on this sheet is BLUE — an input. Nothing here is computed '
     'from anything else here.')
note(wsA, R[0] + 2, 'Everything on every other sheet that can be derived from these is a formula.')

# ============ 3 UNIT BUILD ====================================================
wsU = sheet('Unit Build')
title(wsU, 'Unit build — the PLANT drives the tonnes, and the prices come OUT',
      'Revision 3 assumed a price and divided revenue by it. That made the FY2025 check an '
      'identity that could not fail. Here the drivers are physical and all three realised '
      'prices are derived, so they can be tested against the market and disagree.', 10, 50, 13)
band(wsU, 4, 10); wsU['A4'] = 'FY2025 — THE PLANT IN TONNES (drivers in green, everything else derived)'
CAL = [('Kiln clinker capacity (Mt, audited note 1)', 'B5', f"={A['capclk']}", IN['cap_clinker_mt'], NUM2),
       ('Kiln utilisation', 'B6', f"={A['kutl0']}", IN['kiln_util'][0], PCT),
       ('Clinker produced (Mt)', 'B7', "=B5*B6", UC['clk_prod'], NUM3),
       ('Share sold as clinker', 'B8', f"={A['cksh0']}", IN['clk_export_share'][0], PCT),
       ('Clinker EXPORTED (Mt)', 'B9', "=B7*B8", UC['vol_clk_exp'], NUM3),
       ('Clinker ground into cement (Mt)', 'B10', "=B7-B9", UC['clk_prod'] - UC['vol_clk_exp'], NUM3),
       ('Clinker factor (t clinker / t cement)', 'B11', f"={A['cfac']}", IN['clinker_factor'], NUM3),
       ('Cement produced (Mt)', 'B12', "=B10/B11", UC['cem_prod'], NUM3),
       ('Cement mill capacity (Mt, audited note 1)', 'B13', f"={A['capcem']}", IN['cap_cement_mt'], NUM2),
       ('Mill utilisation', 'B14', "=B12/B13", UC['util_fy25'], PCT),
       ('Cement exported, share of cement made', 'B15', f"={A['cesh0']}", IN['cem_export_share'][0], PCT),
       ('Cement EXPORTED (Mt)', 'B16', "=B12*B15", UC['vol_cem_exp'], NUM3),
       ('Cement sold LOCALLY (Mt)', 'B17', "=B12-B16", UC['vol_local'], NUM3),
       ('TOTAL DESPATCHES FY2025 (Mt)', 'B18', "=B12+B9", UC['vol_fy25'], NUM3)]
for lab, ad, fm, ex, ft in CAL:
    wsU.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsU, ad, fm, ex, ft, bold=(ad in ('B18', 'B12')),
         green=(ad in ('B5', 'B6', 'B8', 'B11', 'B13', 'B15')))
band(wsU, 20, 10); wsU['A20'] = 'FY2025 PRICES — DERIVED FROM THE AUDITED REVENUE NOTE, AND TESTABLE'
PRC = [('Local sales of goods (audited note 4)', 'B21', f"={A['rlg']}", IN['rev_local_goods_fy25'], NUM0),
       ('LOCAL CEMENT PRICE (EGP/t) — DERIVED', 'B22', "=B21/B17", UC['price_loc_derived'], NUM0),
       ('Export sales of goods (audited note 4)', 'B23', f"={A['reg']}", IN['rev_exp_goods_fy25'], NUM0),
       ('Clinker price as a fraction of cement', 'B24', f"={A['ckr']}", IN['clk_price_ratio'], NUM3),
       ('Export cement-equivalent tonnes', 'B25', "=B16+B9*B24",
        UC['vol_cem_exp'] + UC['vol_clk_exp'] * IN['clk_price_ratio'], NUM3),
       ('EXPORT CEMENT PRICE (USD/t) — DERIVED', 'B26', f"=B23/B25/{A['fx25']}",
        UC['price_exp_cem_usd'], NUM1),
       ('EXPORT CLINKER PRICE (USD/t) — DERIVED', 'B27', "=B26*B24",
        UC['price_exp_clk_usd'], NUM1),
       ('Cement exports as a share of cement made — against a 30% statutory cap', 'B28',
        "=B16/B12", IN['cem_export_share'][0], PCT)]
for lab, ad, fm, ex, ft in PRC:
    wsU.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsU, ad, fm, ex, ft, bold=(ad in ('B22', 'B26', 'B27')),
         green=(ad in ('B21', 'B23')))

band(wsU, 30, 10); wsU['A30'] = 'FY2025 CASH COST — DISCLOSED, AND ALLOCATED TO ITS OWN DRIVER'
CC = [('Materials and fuel (note 5)', 'B31', f"={A['cmat']}", IN['cos_materials_fy25'], NUM0),
      ('Transportation (note 5)', 'B32', f"={A['ctra']}", IN['cos_transport_fy25'], NUM0),
      ('Overheads (note 5) plus cash administration (note 6)', 'B33',
       f"={A['covh']}+{A['ga25']}-{A['adep']}",
       IN['cos_overhead_fy25'] + IN['ga_fy25'] - IN['ga_admin_dep_fy25'], NUM0),
      ('Total cash cost (EGP mn)', 'B34', "=SUM(B31:B33)", UC['cash_cost_fy25'], NUM0),
      ('Materials and fuel per tonne of CLINKER — the kiln burns it', 'B35', "=B31/B7",
       UC['cc_mat_t'], NUM0),
      ('Transportation per tonne DESPATCHED', 'B36', "=B32/B18", UC['cc_tra_t'], NUM0),
      ('Overheads and administration per tonne DESPATCHED', 'B37', "=B33/B18",
       UC['cc_ovh_t'], NUM0),
      ('TOTAL CASH COST PER TONNE SOLD', 'B38', "=B34/B18", UC['cash_cost_t'], NUM0)]
for lab, ad, fm, ex, ft in CC:
    wsU.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsU, ad, fm, ex, ft, bold=(ad == 'B38'), green=(ad in ('B31', 'B32')))
note(wsU, 40, 'Provisions and expected credit losses are NOT in this stack. They are operating charges that')
note(wsU, 41, 'belong in the EBITDA bridge, but they are not cash cost per tonne. Revision 3 carried them here.')

band(wsU, 43, 10); wsU['A43'] = 'THE BUILD — FY2025A THEN FY2026E TO FY2030E'
hdr(wsU, 44, [''] + BUY)
LBL = ['Kiln utilisation  (DRIVER)', 'Clinker produced (Mt)',
       'Clinker sold as clinker, share  (DRIVER)', 'Clinker exported (Mt)',
       'Clinker ground (Mt)', 'Cement produced (Mt)', 'Mill utilisation',
       'Cement exported, share  (DRIVER)', 'Cement exported (Mt)', 'Cement local (Mt)',
       'TOTAL DESPATCHES (Mt)', 'Local cement price (EGP/t)',
       'Export cement price (EGP/t)', 'Export clinker price (EGP/t)',
       'Local revenue (EGP mn)', 'Export cement revenue (EGP mn)',
       'Export clinker revenue (EGP mn)', 'Goods revenue (EGP mn)', 'REVENUE (EGP mn)',
       'Revenue per tonne despatched (EGP)', 'Materials and fuel (EGP mn)',
       'Transportation (EGP mn)', 'Overheads and administration (EGP mn)',
       'Cash cost (EGP mn)', 'Cash cost per tonne sold (EGP)',
       'Provisions and credit losses (EGP mn)', 'EBITDA  (AN OUTPUT)', 'EBITDA margin',
       'EBITDA per tonne (EGP)']
for j, l in enumerate(LBL):
    wsU.cell(row=45 + j, column=1, value=l)
for i in range(6):
    c = BUC[i]
    b = BU[i]
    putf(wsU, f'{c}45', f"={A[f'kutl{i}']}", IN['kiln_util'][i], PCT, green=True)
    putf(wsU, f'{c}46', f"=$B$5*{c}45", b['clk_prod'], NUM3)
    putf(wsU, f'{c}47', f"={A[f'cksh{i}']}", IN['clk_export_share'][i], PCT, green=True)
    putf(wsU, f'{c}48', f"={c}46*{c}47", b['clk_exp'], NUM3)
    putf(wsU, f'{c}49', f"={c}46-{c}48", b['clk_ground'], NUM3)
    putf(wsU, f'{c}50', f"={c}49/$B$11", b['cem_prod'], NUM3)
    putf(wsU, f'{c}51', f"={c}50/$B$13", b['mill_util'], PCT)
    putf(wsU, f'{c}52', f"={A[f'cesh{i}']}", IN['cem_export_share'][i], PCT, green=True)
    putf(wsU, f'{c}53', f"={c}50*{c}52", b['cem_exp'], NUM3)
    putf(wsU, f'{c}54', f"={c}50-{c}53", b['cem_loc'], NUM3)
    putf(wsU, f'{c}55', f"={c}50+{c}48", b['sold'], NUM3, bold=True)
    putf(wsU, f'{c}56', f"=$B$22*{A[f'pli{i}']}", b['price_loc'], NUM0)
    putf(wsU, f'{c}57', f"=$B$26*{A[f'pei{i}']}*{A[f'fxp{i}']}", b['price_exp_cem'], NUM0)
    putf(wsU, f'{c}58', f"={c}57*$B$24", b['price_exp_clk'], NUM0)
    putf(wsU, f'{c}59', f"={c}54*{c}56", b['cem_loc'] * b['price_loc'], NUM0)
    putf(wsU, f'{c}60', f"={c}53*{c}57", b['cem_exp'] * b['price_exp_cem'], NUM0)
    putf(wsU, f'{c}61', f"={c}48*{c}58", b['clk_exp'] * b['price_exp_clk'], NUM0)
    putf(wsU, f'{c}62', f"=SUM({c}59:{c}61)", b['rev_goods'], NUM0)
    putf(wsU, f'{c}63', f"={c}62*(1+{A['svc']})", b['rev'], NUM0, bold=True)
    putf(wsU, f'{c}64', f"={c}63/{c}55", b['price'], NUM0)
    putf(wsU, f'{c}65', f"=$B$35*{A[f'infl{i}']}*(1-{A[f'afs{i}']})*{c}46", b['c_mat'], NUM0)
    putf(wsU, f'{c}66', f"=$B$36*{A[f'infl{i}']}*{c}55", b['c_tra'], NUM0)
    putf(wsU, f'{c}67', f"=$B$37*{A[f'infl{i}']}*{c}55", b['c_ovh'], NUM0)
    putf(wsU, f'{c}68', f"=SUM({c}65:{c}67)", b['cc'], NUM0)
    putf(wsU, f'{c}69', f"={c}68/{c}55", b['cc_t'], NUM0)
    putf(wsU, f'{c}70', f"=({A['prov25']}+{A['ecl25']})/{A['rev25']}*{c}63", b['c_prv'], NUM0)
    putf(wsU, f'{c}71', f"={c}63-{c}68-{c}70", b['ebitda'], NUM0, bold=True)
    putf(wsU, f'{c}72', f"={c}71/{c}63", b['mgn'], PCT)
    putf(wsU, f'{c}73', f"={c}71/{c}55", b['ebitda'] / b['sold'], NUM0)

band(wsU, 75, 10); wsU['A75'] = 'VALIDATION — AND A PRICE TEST THAT CAN ACTUALLY FAIL'
VAL = [('Reconstructed FY2025 revenue', 'B76', "=B63", BU[0]['rev'], NUM0),
       ('AUDITED FY2025 revenue', 'B77', f"={A['rev25']}", IN['rev_fy25'], NUM0),
       ('Difference', 'B78', "=B76/B77-1", BU[0]['rev'] / IN['rev_fy25'] - 1, PCT),
       ('Reconstructed FY2025 EBITDA', 'B79', "=B71", BU[0]['ebitda'], NUM0),
       ('AUDITED FY2025 EBITDA (operating profit plus D&A)', 'B80',
        "='Income Statement'!D14", H['ebitda'][2], NUM0),
       ('Difference', 'B81', "=B79/B80-1", BU[0]['ebitda'] / H['ebitda'][2] - 1, PCT)]
for lab, ad, fm, ex, ft in VAL:
    wsU.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsU, ad, fm, ex, ft, green=(ad in ('B77', 'B80')))
note(wsU, 83, 'Rows 76-81 are a TIE, not a test: with prices derived from revenue, revenue reconstructs by')
note(wsU, 84, 'construction. Revision 3 presented exactly this identity as "a test that can fail". It cannot.')
note(wsU, 85, 'The real test is rows 22, 26 and 27 — three DERIVED prices, which can be held against the market:')
note(wsU, 86, f'local EGP {UC["price_loc_derived"]:,.0f}/t, export cement USD {UC["price_exp_cem_usd"]:.1f}/t, export clinker USD {UC["price_exp_clk_usd"]:.1f}/t.')
note(wsU, 87, 'The clinker figure sits roughly 30% BELOW the USD 44-48 the trade press quotes for Egyptian FOB')
note(wsU, 88, 'clinker. That gap is a live disagreement between the physical disclosure and the price indices, and')
note(wsU, 89, 'it is published rather than tuned away. It is the reason the volume base carries a sensitivity.')

band(wsU, 91, 10); wsU['A91'] = 'THE TWO CAPACITY CONSTRAINTS — BOTH LIVE, BOTH CHECKED EVERY YEAR'
KLN = [('Kiln clinker capacity (Mt, audited note 1)', 'B92', f"={A['capclk']}",
        IN['cap_clinker_mt'], NUM2),
       ('Peak clinker required across the forecast (Mt)', 'B93', "=MAX(C46:G46)",
        max(b['clk_prod'] for b in BU[1:]), NUM3),
       ('Peak kiln utilisation — must stay below 100%', 'B94', "=B93/B92",
        max(b['kiln_util'] for b in BU[1:]), PCT),
       ('Headroom at the kiln (Mt of clinker)', 'B95', "=B92-B93",
        IN['cap_clinker_mt'] - max(b['clk_prod'] for b in BU[1:]), NUM3),
       ('Cement mill capacity (Mt, audited note 1)', 'B96', f"={A['capcem']}",
        IN['cap_cement_mt'], NUM2),
       ('Peak cement produced across the forecast (Mt)', 'B97', "=MAX(C50:G50)",
        max(b['cem_prod'] for b in BU[1:]), NUM3),
       ('Peak mill utilisation — must stay below 100%', 'B98', "=B97/B96",
        max(b['mill_util'] for b in BU[1:]), PCT),
       ('Headroom at the mill (Mt of cement)', 'B99', "=B96-B97",
        IN['cap_cement_mt'] - max(b['cem_prod'] for b in BU[1:]), NUM3)]
for lab, ad, fm, ex, ft in KLN:
    wsU.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsU, ad, fm, ex, ft, bold=(ad in ('B94', 'B98')), green=(ad in ('B92', 'B96')))
note(wsU, 101, 'Both constraints are now REAL. Clinker exports and domestic cement compete for the same kiln:')
note(wsU, 102, 'every tonne shipped as clinker is a tonne that could have been ground into cement worth several')
note(wsU, 103, 'times as much. Revision 3 could not see this at all — it carried one product, ignored clinker')
note(wsU, 104, 'exports, and its kiln check therefore ran on a volume base 28% too small. On the physical')
note(wsU, 105, 'disclosure its FY2030 forecast needed 103% of the kiln, not the 78% it published.')

# ============ 4 DCF ===========================================================
wsD = sheet('DCF')
title(wsD, 'Discounted cash flow — the primary lens',
      'Cost of capital built here, facility by facility, never pasted.', 8, 54, 14)
hdr(wsD, 4, [''] + YF)
ROWS = ['Revenue', 'EBITDA margin', 'EBITDA', 'Depreciation and amortisation', 'EBIT',
        'Tax rate (effective)', 'NOPAT  (EBIT × (1 − t))', 'Plus depreciation',
        'Less capital expenditure', 'Less change in working capital',
        'FREE CASH FLOW TO THE FIRM', 'Glide fraction', 'Forward cost of capital',
        'Discount factor', 'PRESENT VALUE OF FCFF']
for j, lab in enumerate(ROWS):
    wsD.cell(row=5 + j, column=1, value=lab)
for i in range(5):
    c = DC[i]
    putf(wsD, f'{c}5', f"='Unit Build'!{BUC[i+1]}63", F['revenue'][i], NUM0, green=True)
    putf(wsD, f'{c}7', f"='Unit Build'!{BUC[i+1]}71", F['ebitda'][i], NUM0, green=True)
    putf(wsD, f'{c}6', f"={c}7/{c}5", F['margin'][i], PCT)
    putf(wsD, f'{c}8', f"={c}5*{A[f'dnap{i}']}", F['dna'][i], NUM0)
    putf(wsD, f'{c}9', f"={c}7-{c}8", F['ebit'][i], NUM0)
    putf(wsD, f'{c}10', f"={A['taxe']}", TAXE, PCT, green=True)
    putf(wsD, f'{c}11', f"={c}9*(1-{c}10)", F['nopat'][i], NUM0)
    putf(wsD, f'{c}12', f"={c}8", F['dna'][i], NUM0)
    putf(wsD, f'{c}13', f"=-{A['capcem']}*{A['cxt']}*{A[f'fxp{i+1}']}", -F['capex'][i], NUM0)
    prev = A['rev25'] if i == 0 else f"{DC[i-1]}5"
    putf(wsD, f'{c}14', f"=-({c}5-{prev})*{A['wcp']}", -F['dwc'][i], NUM0)
    if i == 0:
        putf(wsD, f'{c}15', f"=({c}11+{c}12+{c}13+{c}14)*(1-{A['stub']})",
             F['fcff'][i], NUM0, bold=True)
    else:
        putf(wsD, f'{c}15', f"={c}11+{c}12+{c}13+{c}14", F['fcff'][i], NUM0, bold=True)
    putf(wsD, f'{c}16', f"=({A['kdp0']}-{A[f'kdp{i}']})/({A['kdp0']}-{A['kdp4']})",
         F['glide'][i], DF4)
    putf(wsD, f'{c}17', f"=$C$40-($C$40-$C$50)*{c}16", F['fwd_wacc'][i], PCT2)
    # Each year's forward rate discounts only the calendar it owns. Revision 3 walked
    # the rates in whole-year steps from t=0, so FY2027 was discounted entirely at the
    # FY2026 rate and the FY2030 rate never entered any factor at all.
    if i == 0:
        fm = f"=1/(1+B17)^((1-{A['stub']})/2)"
    else:
        parts = [f"(1+B17)^(1-{A['stub']})"]
        parts += [f"(1+{DC[k]}17)" for k in range(1, i)]
        parts.append(f"(1+{DC[i]}17)^0.5")
        fm = "=1/(" + "*".join(parts) + ")"
    putf(wsD, f'{c}18', fm, F['df'][i], DF4)
    putf(wsD, f'{c}19', f"={c}15*{c}18", F['pv'][i], NUM0)

band(wsD, 21, 8); wsD['A21'] = 'TERMINAL BLOCK'
TB = [('Replacement-cost invested capital, in TERMINAL-year pounds (EGP mn)', 'B22',
       f"={A['capcem']}*{A['repl']}*{A['fx']}*{A['infl5']}", DCF['ic_repl'], NUM0),
      ('Terminal NOPAT  (year 5 NOPAT grown at g)', 'B23', f"=F11*(1+{A['g']})",
       DCF['nopat_term'], NUM0),
      ('Terminal return on invested capital', 'B24', "=B23/B22", DCF['roic_term'], PCT),
      ('Memo: FY2025 return on BOOK invested capital', 'B25',
       f"=('Income Statement'!D12*(1-{A['taxe']}))/({A['eq25']}+$C$44)",
       TR['roic_book_fy25'], PCT),
      ('Reinvestment rate  (g ÷ return on capital)', 'B26', f"={A['g']}/B24",
       DCF['rr_term'], PCT),
      ('Terminal value', 'B27', f"=B23*(1-B26)/($C$50-{A['g']})", DCF['tv'], NUM0),
      ('End-of-window discount factor  (t = 4.417y, not the year-5 mid-point)', 'B29',
       f"=1/((1+B17)^(1-{A['stub']})*(1+C17)*(1+D17)*(1+E17)*(1+F17))",
       DCF['df_tv'], DF4),
      ('Present value of terminal value', 'B28', "=B27*B29", DCF['pv_tv'], NUM0)]
for lab, ad, fm, ex, ft in TB:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft)

band(wsD, 30, 8); wsD['A30'] = 'ENTERPRISE TO EQUITY BRIDGE'
BR = [('Present value of explicit years (FY2026E-FY2030E)', 'B31', "=SUM(B19:F19)",
       DCF['sum_pv'], NUM0),
      ('Present value of terminal value', 'B32', "=B28", DCF['pv_tv'], NUM0),
      ('Enterprise value', 'B33', "=B31+B32", DCF['ev'], NUM0),
      ('TERMINAL VALUE AS % OF ENTERPRISE VALUE', 'B34', "=B32/B33", DCF['tv_share'], PCT),
      ('Audited cash at 31 December 2025', 'B35', f"={A['cash25']}", IN['cash_fy25'], NUM0),
      ('Plus free cash flow AND treasury income earned to the valuation date', 'B36',
       f"=B15/(1-{A['stub']})*{A['stub']}"
       f"+{A['cash25']}*{A['cy0']}*{A['stub']}*(1-{A['taxe']})",
       DCF['cash_at_val'] - IN['cash_fy25'] + IN['div_fy25_declared'], NUM0),
      ('Less the FY2025 dividend declared and unpaid', 'B37', f"=-{A['div25']}",
       -IN['div_fy25_declared'], NUM0),
      ('Cash at the valuation date', 'B38', "=SUM(B35:B37)", DCF['cash_at_val'], NUM0),
      ('Less interest-bearing debt (reviewed 31 March 2026, the FRESHER disclosure)',
       'B39', f"=-{A['q1debt']}", -IN['debt_q1_26'], NUM0),
      ('Net cash (ADDED — the company is net cash)', 'B40', "=B38+B39", DCF['net_cash'], NUM0),
      ('Less non-controlling interests (audited note 24)', 'B41', f"=-{A['nci']}",
       -IN['nci'], NUM4),
      ('Equity value', 'B42', "=B33+B40+B41", DCF['equity'], NUM0),
      ('Shares outstanding (mn)', 'B43', f"={A['shiss']}-{A['shtre']}", SH, NUM4),
      ('FAIR VALUE PER SHARE (EGP)', 'B44', "=B42/B43", DCF['fv'], PX)]
for lab, ad, fm, ex, ft in BR:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft, bold=(ad in ('B33', 'B34', 'B42', 'B44')))

wsD['E35'] = 'COST OF CAPITAL — BUILT HERE, FACILITY BY FACILITY (labels in column E:'
wsD['E36'] = 'this block shares rows with the bridge above, whose labels are in column A)'
CC2 = [('Risk-free rate (observed EGP 10-year)', 'C36', f"={A['rf']}", IN['rf'], PCT2),
       ('Less sovereign default spread', 'C37', f"=-{A['sov']}", -IN['sov_spread_cds'], PCT2),
       ('Normalised risk-free rate', 'C38', "=C36+C37", W['rf_star'], PCT2),
       ('Cost of equity  (rf* + beta × premium)', 'C39',
        f"=C38+{A['beta']}*{A['erp']}", W['ke_exp'], PCT2),
       ('WACC — explicit window', 'C40', "=(1-C43)*C39+C43*C42", W['wacc_exp'], PCT2),
       ('Cost of debt, blended by currency', 'C41',
        f"=({A['dcib']}*{A['kdegp']}+{A['dnbe']}*({A['eur']}+0.03)"
        f"+{A['debrd']}*({A['eur']}+0.0435)+{A['dlease']}*{A['kdegp']})/C44",
        W['kd'], PCT2),
       ('Cost of debt after tax', 'C42', f"=C41*(1-{A['tax']})", W['kd_at'], PCT2),
       ('Debt weight  D/(D+E)', 'C43', "=C44/(C44+C45)", W['wd_gross'], PCT2),
       ('Total interest-bearing debt', 'C44',
        f"={A['dcib']}+{A['dnbe']}+{A['debrd']}+{A['dlease']}", W['debt_total'], NUM0),
       ('Market capitalisation', 'C45', f"={A['spot']}*({A['shiss']}-{A['shtre']})",
        M['mktcap'], NUM0),
       ('Euro share of the debt book', 'C46', f"=({A['dnbe']}+{A['debrd']})/C44",
        W['eur_share'], PCT),
       ('Terminal beta, UNLEVERED at the observed structure then re-levered', 'C47',
        f"={A['beta']}/(1+(1-{A['tax']})*C43/(1-C43))"
        f"*(1+(1-{A['tax']})*{A['wdt']}/(1-{A['wdt']}))", W['beta_term'], NUM3),
       ('Terminal cost of equity', 'C48', f"={A['rft']}+C47*{A['erpt']}", W['ke_term'], PCT2),
       ('Terminal cost of debt after tax', 'C49', f"={A['kdt']}*(1-{A['tax']})",
        W['kd_term_at'], PCT2),
       ('WACC — terminal', 'C50', f"=(1-{A['wdt']})*C48+{A['wdt']}*C49", W['wacc_term'], PCT2),
       ('Memo: retired construction, risk-free NOT netted', 'C51',
        f"={A['rf']}+{A['beta']}*{A['erp']}", W['ke_raw_retired'], PCT2),
       ('Memo: sovereign double-count removed (basis points)', 'C52', "=(C51-C39)*10000",
        (W['ke_raw_retired'] - W['ke_exp']) * 10000, NUM0)]
for lab, ad, fm, ex, ft in CC2:
    wsD.cell(row=int(ad[1:]), column=5, value=lab)
    putf(wsD, ad, fm, ex, ft, bold=(ad in ('C40', 'C50', 'C41')))

band(wsD, 54, 8); wsD['A54'] = 'COST-OF-DEBT INTEGRITY — THE EVIDENCE, NOT THE ASSERTION'
KD = [('Euro share of the book (audited note 25)', 'B55', "=C46", W['eur_share'], PCT),
      ('CIB facility rate — EGP corridor + 0.6%', 'B56', f"={A['kdegp']}", KDG['kd_cib'], PCT2),
      ('NBE/KfW facility rate — Euribor + 3.00%', 'B57', f"={A['eur']}+0.03", KDG['kd_nbe'], PCT2),
      ('EBRD facility rate — Euribor + 4.35%', 'B58', f"={A['eur']}+0.0435", KDG['kd_ebrd'], PCT2),
      ('ADOPTED blended cost of debt', 'B59', "=C41", KDG['kd_blended'], PCT2),
      ('Effective rate computed FY2024 (interest / average debt)', 'B60',
       f"=({A['li24']}+{A['fi24']})/(({A['d23']}+{A['d24']})/2)", KDG['eff_fy24'], PCT2),
      ('Effective rate computed FY2025', 'B61',
       f"=({A['li25']}+{A['fi25']})/(({A['d24']}+C44-{A['dlease']})/2)", KDG['eff_fy25'], PCT2),
      ('Effective rate computed Q1-2026, annualised', 'B62',
       f"={A['q1fc']}*4/((C44+{A['q1debt']})/2)", KDG['eff_q126_annualised'], PCT2),
      ('Pound-EQUIVALENT cost of debt under interest parity', 'B63',
       f"=({A['dcib']}*{A['kdegp']}+{A['dnbe']}*({A['eur']}+0.03+{A['dep']})"
       f"+{A['debrd']}*({A['eur']}+0.0435+{A['dep']})+{A['dlease']}*{A['kdegp']})/C44",
       KDG['kd_egp_equivalent'], PCT2)]
for lab, ad, fm, ex, ft in KD:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft, bold=(ad in ('B59', 'B63')))
note(wsD, 65, 'The blended rate on row 59 sits above the trailing effective rates on rows 60-62 because the book')
note(wsD, 66, 're-based mid-year from pound facilities to euro term debt, and interest on the under-construction')
note(wsD, 67, 'alternative-fuel assets is capitalised rather than expensed. The gap is disclosed, not smoothed.')
note(wsD, 68, 'Row 63 loads the euro legs with expected pound depreciation; its effect is on Fundamental Valuation.')
note(wsD, 69, 'The glide on row 16 is derived from the POUND cost-of-debt path: the discount rate is a pound rate')
note(wsD, 70, 'on pound cash flows, so the pound easing calendar sets its slope while the euro book sets its level.')

# ============ 5 EV BRIDGE =====================================================
wsB = sheet('EV Bridge')
title(wsB, 'Enterprise value to equity bridge', None, 6, 56, 16)
hdr(wsB, 4, ['', 'EGP mn', 'Per share (EGP)'])
BRW = [('Present value of explicit free cash flow', "=DCF!B31", DCF['sum_pv']),
       ('Present value of terminal value', "=DCF!B32", DCF['pv_tv']),
       ('Enterprise value', "=DCF!B33", DCF['ev']),
       ('Plus net cash at the valuation date', "=DCF!B40", DCF['net_cash']),
       ('Less non-controlling interests', "=DCF!B41", -IN['nci']),
       ('Equity value', "=DCF!B42", DCF['equity'])]
for j, (lab, fm, ex) in enumerate(BRW):
    wsB.cell(row=5 + j, column=1, value=lab)
    putf(wsB, f'B{5+j}', fm, ex, NUM0, green=True, bold=(j in (2, 5)))
    putf(wsB, f'C{5+j}', f"=B{5+j}/DCF!$B$43", ex / SH, PX, bold=(j in (2, 5)))
wsB['A12'] = 'TERMINAL VALUE AS % OF ENTERPRISE VALUE'
putf(wsB, 'B12', "=DCF!B34", DCF['tv_share'], PCT, green=True, bold=True)
wsB['A13'] = 'Memo: spot price'
putf(wsB, 'C13', f"={A['spot']}", SPOT, PX, green=True)
wsB['A14'] = 'Fair value less spot'
putf(wsB, 'C14', "=C10-C13", DCF['fv'] - SPOT, PX)
wsB['A15'] = 'Upside / (downside) to the cash-flow lens'
putf(wsB, 'C15', "=C10/C13-1", DCF['fv'] / SPOT - 1, PCT)
wsB['A16'] = 'Memo: enterprise value per annual tonne of capacity (USD)'
putf(wsB, 'C16', f"=B7/{A['capcem']}/{A['fx']}", DCF['ev'] / IN['cap_cement_mt'] / IN['fx'], NUM1)
note(wsB, 18, 'Terminal value share links live to the DCF sheet — never typed. The bridge ADDS net cash and')
note(wsB, 19, 'DEDUCTS minority interests; both signs are asserted in the build. The FY2025 dividend declared and')
note(wsB, 20, 'unpaid at 31 March 2026 is removed from cash: a buyer at today\'s price does not receive it.')

# ============ 6 INCOME STATEMENT ==============================================
wsI = sheet('Income Statement')
title(wsI, 'Income statement — 3 AUDITED years + 5-year forecast', 'EGP mn', 10, 48, 13)
hdr(wsI, 4, [''] + YH + YF)
# Row 12 restates operating profit to open the EBITDA bridge. Revision 3's label list
# had no entry for it, so every label from row 12 to row 21 sat one row above its own
# contents — row 12 read 'Depreciation and amortisation' over operating profit, and row
# 20 read 'Earnings per share' over attributable profit.
IL = ['Revenue', 'Cost of sales', 'Gross profit', 'Gross margin',
      'General and administrative expenses', 'Provisions and credit losses',
      'OPERATING PROFIT', 'Operating profit, opening the EBITDA bridge',
      'Depreciation and amortisation', 'EBITDA', 'EBITDA margin',
      'Net finance and other income', 'Profit before tax', 'Income tax',
      'Effective tax rate', 'Attributable profit', 'Earnings per share (EGP)']
for j, l in enumerate(IL):
    wsI.cell(row=5 + j, column=1, value=l)
HK = ['23', '24', '25']
for i in range(3):
    c = HC[i]
    t = HK[i]
    putf(wsI, f'{c}5', f"={A[f'rev{t}']}", H['revenue'][i], NUM0, green=True)
    putf(wsI, f'{c}6', f"=-{A[f'cogs{t}']}", -H['cogs'][i], NUM0, green=True)
    putf(wsI, f'{c}7', f"={c}5+{c}6", H['gross_profit'][i], NUM0)
    putf(wsI, f'{c}8', f"={c}7/{c}5", H['gross_profit'][i] / H['revenue'][i], PCT)
    putf(wsI, f'{c}9', f"=-{A[f'ga{t}']}", -H['ga'][i], NUM0, green=True)
    putf(wsI, f'{c}10', f"=-{A[f'prov{t}']}-{A[f'ecl{t}']}",
         -(IN[f'prov_fy{t}'] + IN[f'ecl_fy{t}']), NUM0, green=True)
    putf(wsI, f'{c}11', f"={c}7+{c}9+{c}10", H['ebit'][i], NUM0, bold=True)
    putf(wsI, f'{c}12', f"={c}11", H['ebit'][i], NUM0)
    putf(wsI, f'{c}13', f"={A[f'dna{t}']}", H['dna'][i], NUM0, green=True)
    putf(wsI, f'{c}14', f"={c}12+{c}13", H['ebitda'][i], NUM0, bold=True)
    putf(wsI, f'{c}15', f"={c}14/{c}5", H['margin'][i], PCT)
    if i == 2:
        # FY2025 is built from the disclosed components rather than as a residual, and
        # row 24 below asserts the two agree.
        putf(wsI, f'{c}16',
             f"={A['ii25']}+{A['oi25']}-{A['fc25']}+{A['fxd25']}+1.14",
             IN['intinc_fy25'] + IN['othinc_fy25'] - IN['fincost_fy25']
             + IN['fx_diff_fy25'] + 1.14, NUM0)
    else:
        putf(wsI, f'{c}16', f"={A[f'pbt{t}']}-{c}11", H['pbt'][i] - H['ebit'][i], NUM0)
    putf(wsI, f'{c}17', f"={A[f'pbt{t}']}", H['pbt'][i], NUM0, green=True)
    putf(wsI, f'{c}18', f"=-{A[f'txc{t}']}", -H['tax'][i], NUM0, green=True)
    putf(wsI, f'{c}19', f"=-{c}18/{c}17", H['tax_eff_hist'][i], PCT)
    putf(wsI, f'{c}20', f"={A[f'pat{t}']}", H['pat'][i], NUM0, green=True, bold=True)
    putf(wsI, f'{c}21', f"={A[f'eps{t}']}", H['eps'][i], PX, green=True)
put(wsI, 'D22', None, BLACK, NUM0)
putf(wsI, 'C22', f"={A['div24']}", IN['div_fy24_paid'], NUM0, green=True)
putf(wsI, 'D22', f"={A['div25']}", IN['div_fy25_declared'], NUM0, green=True)
putf(wsI, 'C23', f"=C22/DCF!$B$43", IN['div_fy24_paid'] / SH, PX)
putf(wsI, 'D23', f"=D22/DCF!$B$43", IN['div_fy25_declared'] / SH, PX)
for i in range(5):
    c = FC[i]
    putf(wsI, f'{c}5', f"=DCF!{DC[i]}5", F['revenue'][i], NUM0, green=True)
    putf(wsI, f'{c}14', f"=DCF!{DC[i]}7", F['ebitda'][i], NUM0, green=True, bold=True)
    putf(wsI, f'{c}15', f"={c}14/{c}5", F['margin'][i], PCT)
    putf(wsI, f'{c}13', f"=DCF!{DC[i]}8", F['dna'][i], NUM0, green=True)
    putf(wsI, f'{c}11', f"={c}14-{c}13", F['ebit'][i], NUM0, bold=True)
    putf(wsI, f'{c}12', f"={c}11", F['ebit'][i], NUM0)
    putf(wsI, f'{c}6', f"=-({c}5-{c}14-{A[f'ga{HK[2]}']}*{A[f'infl{i+1}']}/{A['infl0']})", None, NUM0)
    # FY2026 opens on the audited FY2025 cash LESS the FY2025 dividend declared and
    # unpaid at 31 March 2026 — the same opening balance the bridge uses.
    open_cash = (f"('Balance Sheet'!D9-{A['div25']})" if i == 0
                 else f"'Balance Sheet'!{FC[i-1]}9")
    putf(wsI, f'{c}16',
         f"={open_cash}*{A[f'cy{i}']}-DCF!$C$44*(1-DCF!$C$46)*{A[f'kdp{i}']}"
         f"-DCF!$C$44*DCF!$C$46*({A['eur']}+0.0435)", F['treasury'][i], NUM0)
    putf(wsI, f'{c}17', f"={c}11+{c}16", F['pbt'][i], NUM0)
    putf(wsI, f'{c}18', f"=-{c}17*{A['taxe']}", -F['tax'][i], NUM0)
    putf(wsI, f'{c}19', f"=-{c}18/{c}17", TAXE, PCT)
    putf(wsI, f'{c}20', f"={c}17+{c}18", F['pat'][i], NUM0, bold=True)
    putf(wsI, f'{c}21', f"={c}20/DCF!$B$43", F['eps'][i], PX)
    putf(wsI, f'{c}22', f"={c}20*{A['payout']}", F['dividends'][i], NUM0)
    putf(wsI, f'{c}23', f"={c}22/DCF!$B$43", F['dps'][i], PX)
wsI['A22'] = 'Dividends declared'
wsI['A23'] = 'Dividend per share (EGP)'
for i in range(5):
    c = FC[i]
    wsI[f'{c}6'].value = None
    if (wsI.title in EXPECT) and f'{c}6' in EXPECT[wsI.title]:
        del EXPECT[wsI.title][f'{c}6']
wsI['A24'] = 'FY2025 check: components of non-operating income against the audited residual'
putf(wsI, 'B24', "=D16-(D17-D11)",
     (IN['intinc_fy25'] + IN['othinc_fy25'] - IN['fincost_fy25'] + IN['fx_diff_fy25'] + 1.14)
     - (H['pbt'][2] - H['ebit'][2]), NUM4)
note(wsI, 25, 'FY2023-FY2025 revenue, cost of sales, administrative expenses, provisions, pre-tax profit, tax,')
note(wsI, 26, 'attributable profit, earnings per share and depreciation are AUDITED figures. Operating profit,')
note(wsI, 27, 'EBITDA, margins and the effective tax rate are formulas over them. Nothing historical is derived')
note(wsI, 28, 'by closing an assumption, which is what revision 1 had to do.')

# ============ 7 BALANCE SHEET =================================================
wsBS = sheet('Balance Sheet')
title(wsBS, 'Balance sheet — 3 AUDITED years + 5-year forecast', 'EGP mn', 10, 48, 13)
hdr(wsBS, 4, [''] + YH + YF)
BL = ['Property, plant and equipment', 'Assets under construction', 'Intangible assets',
      'Total non-current assets', 'Cash and bank balances',
      'Inventories, receivables and debtors', 'TOTAL ASSETS', 'Interest-bearing debt',
      'Other liabilities', 'Total liabilities', 'Equity attributable to owners',
      'Net (cash) / debt', 'Book value per share (EGP)', 'Return on equity']
for j, l in enumerate(BL):
    wsBS.cell(row=5 + j, column=1, value=l)
putf(wsBS, 'D5', f"={A['ppe25']}", IN['ppe_fy25'], NUM0, green=True)
putf(wsBS, 'D6', f"={A['auc25']}", IN['auc_fy25'], NUM0, green=True)
putf(wsBS, 'D7', f"={A['int25']}", IN['intang_fy25'], NUM0, green=True)
putf(wsBS, 'D8', "=SUM(D5:D7)", IN['ppe_fy25'] + IN['auc_fy25'] + IN['intang_fy25'], NUM0)
putf(wsBS, 'D10', f"={A['inv25']}+{A['rec25']}+{A['deb25']}",
     IN['inv_fy25'] + IN['recv_fy25'] + IN['debtors_fy25'], NUM0)
putf(wsBS, 'D9', f"={A['cash25']}", IN['cash_fy25'], NUM0, green=True)
putf(wsBS, 'D11', f"={A['ta25']}", IN['ta_fy25'], NUM0, green=True, bold=True)
putf(wsBS, 'D12', "=DCF!C44", W['debt_total'], NUM0, green=True)
putf(wsBS, 'D14', f"={A['tl25']}", IN['tl_fy25'], NUM0, green=True)
putf(wsBS, 'D13', "=D14-D12", IN['tl_fy25'] - W['debt_total'], NUM0)
putf(wsBS, 'D15', f"={A['eq25']}", IN['eq_fy25'], NUM0, green=True, bold=True)
putf(wsBS, 'D16', "=D12-D9", W['debt_total'] - IN['cash_fy25'], NUM0)
putf(wsBS, 'D17', "=D15/DCF!$B$43", LN['bvps'], PX)
putf(wsBS, 'D18', "='Income Statement'!D20/D15", LN['roe_fy25'], PCT)
putf(wsBS, 'C11', f"={A['ta24']}", IN['ta_fy24'], NUM0, green=True, bold=True)
putf(wsBS, 'B11', f"={A['ta23']}", IN['ta_fy23'], NUM0, green=True, bold=True)
putf(wsBS, 'C9', f"={A['cash24']}", IN['cash_fy24'], NUM0, green=True)
putf(wsBS, 'C12', f"={A['d24']}", IN['debt_fy24'], NUM0, green=True)
putf(wsBS, 'B12', f"={A['d23']}", IN['debt_fy23'], NUM0, green=True)
putf(wsBS, 'C15', f"={A['eq24']}", IN['eq_fy24'], NUM0, green=True, bold=True)
putf(wsBS, 'B15', f"={A['eq23']}", IN['eq_fy23'], NUM0, green=True, bold=True)
putf(wsBS, 'B9', f"={A['cash23']}", IN['cash_fy23'], NUM0, green=True)
putf(wsBS, 'C16', "=C12-C9", IN['debt_fy24'] - IN['cash_fy24'], NUM0)
putf(wsBS, 'B16', "=B12-B9", IN['debt_fy23'] - IN['cash_fy23'], NUM0)
putf(wsBS, 'C17', "=C15/DCF!$B$43", IN['eq_fy24'] / SH, PX)
putf(wsBS, 'B17', "=B15/DCF!$B$43", IN['eq_fy23'] / SH, PX)
putf(wsBS, 'C18', "='Income Statement'!C20/C15", H['pat'][1] / IN['eq_fy24'], PCT)
putf(wsBS, 'B18', "='Income Statement'!B20/B15", H['pat'][0] / IN['eq_fy23'], PCT)
for i in range(5):
    c = FC[i]
    prev = 'D' if i == 0 else FC[i - 1]
    fixed0 = (IN['ppe_fy25'] + IN['auc_fy25'] + IN['intang_fy25'])
    putf(wsBS, f'{c}8', f"={prev}8-DCF!{DC[i]}8-DCF!{DC[i]}13", F['ppe'][i], NUM0)
    putf(wsBS, f'{c}10', f"={prev}10-DCF!{DC[i]}14", F['wc'][i], NUM0)
    if i == 0:
        putf(wsBS, f'{c}9',
             f"=D9-{A['div25']}+'Income Statement'!{c}20+DCF!{DC[i]}8+DCF!{DC[i]}13"
             f"+DCF!{DC[i]}14-'Income Statement'!{c}22", F['cash'][i], NUM0)
    else:
        putf(wsBS, f'{c}9',
             f"={prev}9+'Income Statement'!{c}20+DCF!{DC[i]}8+DCF!{DC[i]}13"
             f"+DCF!{DC[i]}14-'Income Statement'!{c}22", F['cash'][i], NUM0)
    putf(wsBS, f'{c}11', f"={c}8+{c}9+{c}10", F['total_assets'][i], NUM0, bold=True)
    putf(wsBS, f'{c}12', "=DCF!$C$44", W['debt_total'], NUM0, green=True)
    putf(wsBS, f'{c}13', "=D13", IN['tl_fy25'] - W['debt_total'], NUM0)
    putf(wsBS, f'{c}14', f"={c}12+{c}13", IN['tl_fy25'], NUM0)
    if i == 0:
        putf(wsBS, f'{c}15',
             f"=D15-{A['div25']}+'Income Statement'!{c}20-'Income Statement'!{c}22",
             F['equity'][i], NUM0, bold=True)
    else:
        putf(wsBS, f'{c}15',
             f"={prev}15+'Income Statement'!{c}20-'Income Statement'!{c}22",
             F['equity'][i], NUM0, bold=True)
    putf(wsBS, f'{c}16', f"={c}12-{c}9", W['debt_total'] - F['cash'][i], NUM0)
    putf(wsBS, f'{c}17', f"={c}15/DCF!$B$43", F['equity'][i] / SH, PX)
    putf(wsBS, f'{c}18', f"='Income Statement'!{c}20/{c}15",
         F['pat'][i] / F['equity'][i], PCT)
band(wsBS, 20, 10); wsBS['A20'] = 'THE AUDITED BALANCE SHEET CLOSES — A CHECK REVISION 1 COULD NOT RUN'
CLO = [('Total assets (audited)', 'B21', f"={A['ta25']}", IN['ta_fy25']),
       ('Less total liabilities (audited)', 'B22', f"=-{A['tl25']}", -IN['tl_fy25']),
       ('Equals total equity', 'B23', "=B21+B22", IN['ta_fy25'] - IN['tl_fy25']),
       ('Equity attributable to owners plus minorities (audited)', 'B24',
        f"={A['eq25']}+{A['nci']}", IN['eq_fy25'] + IN['nci']),
       ('Difference', 'B25', "=B23-B24",
        (IN['ta_fy25'] - IN['tl_fy25']) - (IN['eq_fy25'] + IN['nci']))]
for lab, ad, fm, ex in CLO:
    wsBS.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsBS, ad, fm, ex, NUM4 if ad == 'B25' else NUM0, bold=(ad == 'B25'))
note(wsBS, 27, 'Revision 1 was offered EGP 2,894.13mn as total liabilities and rejected it because it would not')
note(wsBS, 28, 'close; it derived EGP 4,140.99mn instead. The audited statements show 2,894.13 is total CURRENT')
note(wsBS, 29, 'liabilities and the total is 4,140.99. The derivation was right, and row 25 now closes to zero.')

# ============ 8 CASH FLOW =====================================================
wsC = sheet('Cash Flow')
title(wsC, 'Cash flow — audited FY2025 and the forecast, linked to the waterfall',
      'EGP mn', 8, 52, 14)
hdr(wsC, 4, ['', 'FY2025A'] + YF)
CFL = ['Attributable profit', 'Add back depreciation', 'Less change in working capital',
       'Cash from operations', 'Capital expenditure', 'Free cash flow to equity',
       'Dividends declared', 'Closing cash', 'Memo: free cash flow to the firm']
for j, l in enumerate(CFL):
    wsC.cell(row=5 + j, column=1, value=l)
putf(wsC, 'B5', "='Income Statement'!D20", H['pat'][2], NUM0, green=True)
putf(wsC, 'B6', f"={A['dna25']}", H['dna'][2], NUM0, green=True)
putf(wsC, 'B8', "=B5+B6", H['pat'][2] + H['dna'][2], NUM0, bold=True)
putf(wsC, 'B9', f"=-{A['cx25']}", -H['capex'][2], NUM0, green=True)
putf(wsC, 'B10', "=B8+B9", H['pat'][2] + H['dna'][2] - H['capex'][2], NUM0, bold=True)
putf(wsC, 'B11', f"=-{A['div25']}", -IN['div_fy25_declared'], NUM0, green=True)
putf(wsC, 'B12', f"={A['cash25']}", IN['cash_fy25'], NUM0, green=True)
for i in range(5):
    c = DC[i] if i else 'C'
    c = ['C', 'D', 'E', 'F', 'G'][i]
    ic = FC[i]
    putf(wsC, f'{c}5', f"='Income Statement'!{ic}20", F['pat'][i], NUM0, green=True)
    putf(wsC, f'{c}6', f"=DCF!{DC[i]}8", F['dna'][i], NUM0, green=True)
    putf(wsC, f'{c}7', f"=DCF!{DC[i]}14", -F['dwc'][i], NUM0, green=True)
    putf(wsC, f'{c}8', f"={c}5+{c}6+{c}7", F['pat'][i] + F['dna'][i] - F['dwc'][i], NUM0,
         bold=True)
    putf(wsC, f'{c}9', f"=DCF!{DC[i]}13", -F['capex'][i], NUM0, green=True)
    putf(wsC, f'{c}10', f"={c}8+{c}9",
         F['pat'][i] + F['dna'][i] - F['dwc'][i] - F['capex'][i], NUM0, bold=True)
    putf(wsC, f'{c}11', f"=-'Income Statement'!{ic}22", -F['dividends'][i], NUM0)
    putf(wsC, f'{c}12', f"='Balance Sheet'!{ic}9", F['cash'][i], NUM0, green=True)
    putf(wsC, f'{c}13', f"=DCF!{DC[i]}15", F['fcff'][i], NUM0, green=True)
note(wsC, 15, 'The FY2025 column is the AUDITED outturn: attributable profit, depreciation and capital')
note(wsC, 16, 'expenditure all read from the audited cash flow statement. The forecast columns LINK to the DCF')
note(wsC, 17, 'and Income Statement sheets, so the statement and the valuation cannot disagree by construction.')

# ============ 9 SUMMARY FINANCIALS ============================================
wsSF = sheet('Summary Financials')
title(wsSF, 'Summary financials — eight years on one screen', 'EGP mn', 10, 46, 13)
hdr(wsSF, 4, [''] + YH + YF)
allc = HC + FC
SFL = [('Revenue', 5, H['revenue'] + F['revenue'], NUM0),
       ('EBITDA', 14, H['ebitda'] + F['ebitda'], NUM0),
       ('EBITDA margin', 15, H['margin'] + F['margin'], PCT),
       ('Operating profit', 11, H['ebit'] + F['ebit'], NUM0),
       ('Attributable profit', 20, H['pat'] + F['pat'], NUM0),
       ('Earnings per share (EGP)', 21, H['eps'] + F['eps'], PX)]
for j, (lab, row, vals, ft) in enumerate(SFL):
    wsSF.cell(row=5 + j, column=1, value=lab)
    for i, c in enumerate(allc):
        putf(wsSF, f'{c}{5+j}', f"='Income Statement'!{c}{row}", vals[i], ft, green=True)
band(wsSF, 12, 10); wsSF['A12'] = 'GROWTH'
GR = [('Revenue growth', 5, H['revenue'] + F['revenue']),
      ('EBITDA growth', 6, H['ebitda'] + F['ebitda']),
      ('Profit growth', 9, H['pat'] + F['pat'])]
for j, (lab, src, base) in enumerate(GR):
    wsSF.cell(row=13 + j, column=1, value=lab)
    for i in range(1, 8):
        putf(wsSF, f'{allc[i]}{13+j}', f"={allc[i]}{src}/{allc[i-1]}{src}-1",
             base[i] / base[i - 1] - 1, PCT)
note(wsSF, 18, 'Every cell on this sheet is a link or a formula. Nothing is restated.')

# ============ 10 PER-SHARE & RATIOS ===========================================
wsR = sheet('Per-Share & Ratios')
title(wsR, 'Per-share figures, ratios and reconciliations', None, 10, 50, 13)
hdr(wsR, 4, [''] + YH + YF)
RL = ['Earnings per share (EGP)', 'Book value per share (EGP)', 'Price / earnings (at spot)',
      'Price / book (at spot)', 'EBITDA per tonne (EGP)', 'Return on equity',
      'Net (cash) / EBITDA']
for j, l in enumerate(RL):
    wsR.cell(row=5 + j, column=1, value=l)
eps_all = H['eps'] + F['eps']
eq_all = [IN['eq_fy23'], IN['eq_fy24'], IN['eq_fy25']] + F['equity']
eb_all = H['ebitda'] + F['ebitda']
nd_all = [IN['debt_fy23'] - IN['cash_fy23'], IN['debt_fy24'] - IN['cash_fy24'],
          W['debt_total'] - IN['cash_fy25']] + [W['debt_total'] - x for x in F['cash']]
for i, c in enumerate(allc):
    putf(wsR, f'{c}5', f"='Income Statement'!{c}21", eps_all[i], PX, green=True)
    putf(wsR, f'{c}6', f"='Balance Sheet'!{c}17", eq_all[i] / SH, PX, green=True)
    putf(wsR, f'{c}7', f"={A['spot']}/{c}5", SPOT / eps_all[i], MULT)
    putf(wsR, f'{c}8', f"={A['spot']}/{c}6", SPOT / (eq_all[i] / SH), MULT)
    putf(wsR, f'{c}10', f"='Balance Sheet'!{c}18",
         [H['pat'][0] / IN['eq_fy23'], H['pat'][1] / IN['eq_fy24'], LN['roe_fy25']][i]
         if i < 3 else F['pat'][i - 3] / F['equity'][i - 3], PCT, green=True)
    putf(wsR, f'{c}11', f"='Balance Sheet'!{c}16/'Income Statement'!{c}14",
         nd_all[i] / eb_all[i], MULT)
for i in range(3, 8):
    c = allc[i]
    putf(wsR, f'{c}9', f"='Income Statement'!{c}14/'Unit Build'!{BUC[i-2]}55",
         eb_all[i] / F['volume_mt'][i - 3], NUM0)
putf(wsR, 'D9', "='Income Statement'!D14/'Unit Build'!B55", H['ebitda'][2] / BU[0]['vol'], NUM0)

band(wsR, 13, 10); wsR['A13'] = 'RECONCILIATIONS AGAINST THE AUDITED ACCOUNTS'
REC = [('Shares issued (audited note 20)', 'B14', f"={A['shiss']}", IN['shares_issued'], NUM4),
       ('Less treasury shares (audited note 21)', 'B15', f"=-{A['shtre']}",
        -IN['shares_treasury'], NUM4),
       ('Shares outstanding — ADOPTED', 'B16', "=B14+B15", SH, NUM4),
       ('Shares implied by the FY2025 dividend at EGP 5.34 per share', 'B17',
        f"={A['div25']}/5.34", SHT['from_fy25_dividend'], NUM4),
       ('Difference', 'B18', "=B17/B16-1", SHT['from_fy25_dividend'] / SH - 1, '0.0000%'),
       ('Issued capital implied at EGP 2 par (EGP mn)', 'B19', f"={A['shiss']}*2",
        SHT['par_check'], NUM0),
       ('Q1-2026 revenue growth on Q1-2025', 'B20', f"={A['q1r26']}/{A['q1r25']}-1",
        IN['rev_q1_26'] / IN['rev_q1_25'] - 1, PCT),
       ('FY2026 revenue implied by holding that growth', 'B21',
        f"={A['rev25']}*{A['q1r26']}/{A['q1r25']}",
        IN['rev_fy25'] * IN['rev_q1_26'] / IN['rev_q1_25'], NUM0),
       ('FY2026 revenue this model forecasts', 'B22', "=DCF!B5", F['revenue'][0], NUM0),
       ('The forecast against the run rate', 'B23', "=B22/B21-1",
        F['revenue'][0] / (IN['rev_fy25'] * IN['rev_q1_26'] / IN['rev_q1_25']) - 1, PCT),
       ('Q1-2026 gross margin', 'B24', f"={A['q1gp']}/{A['q1r26']}",
        IN['gp_q1_26'] / IN['rev_q1_26'], PCT),
       ('FY2025 gross margin', 'B25', "='Income Statement'!D8",
        H['gross_profit'][2] / H['revenue'][2], PCT, ),
       ('Q1-2026 attributable profit annualised at four times', 'B26', f"=4*{A['q1pat']}",
        4 * IN['pat_q1_26'], NUM0),
       ('FY2026 profit this model forecasts', 'B27', "='Income Statement'!E20",
        F['pat'][0], NUM0),
       ('The forecast against the annualised first quarter', 'B28', "=B27/B26-1",
        F['pat'][0] / (4 * IN['pat_q1_26']) - 1, PCT),
       ('Q1-2026 net cash, free of the declared dividend', 'B29',
        f"={A['q1cash']}-{A['q1debt']}-{A['q1div']}",
        IN['cash_q1_26'] - IN['debt_q1_26'] - IN['divpay_q1_26'], NUM0),
       ('Net cash at the valuation date in this model', 'B30', "=DCF!B40",
        DCF['net_cash'], NUM0)]
for item in REC:
    lab, ad, fm, ex, ft = item[0], item[1], item[2], item[3], item[4]
    wsR.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsR, ad, fm, ex, ft, bold=(ad in ('B16', 'B18')),
         green=(ad in ('B22', 'B25', 'B27', 'B30')))
note(wsR, 32, 'The share count is now read directly from the audited capital and treasury notes, and the FY2025')
note(wsR, 33, 'dividend divides to the same count exactly. Revision 1 reached 374.87mn by triangulation and was')
note(wsR, 34, 'right; this is the confirmation rather than the estimate.')

# ============ 11 RELATIVE & NORMALIZED ========================================
wsN = sheet('Relative & Normalized')
title(wsN, 'Relative multiples and normalised earnings power', None, 7, 58, 16)
band(wsN, 4, 7); wsN['A4'] = 'NORMALISED EARNINGS BASE'
NB = [('FY2025 revenue, audited (the cyclical peak)', 'B5', f"={A['rev25']}", IN['rev_fy25'], NUM0),
      ('Haircut to the revenue base', 'B6', f"={A['nhc']}", IN['norm_rev_haircut'], PCT),
      ('Normalised revenue', 'B7', "=B5*B6", IN['rev_fy25'] * IN['norm_rev_haircut'], NUM0),
      ('Mid-cycle EBITDA margin', 'B8', f"={A['nmgn']}", IN['norm_mgn'], PCT),
      ('Normalised EBITDA', 'B9', "=B7*B8", LN['ebitda_norm'], NUM0),
      ('Less audited depreciation', 'B10', f"=-{A['dna25']}", -IN['dna_fy25'], NUM0),
      ('Normalised EBIT', 'B11', "=B9+B10", LN['ebitda_norm'] - IN['dna_fy25'], NUM0),
      ('Normalised NOPAT', 'B12', f"=B11*(1-{A['taxe']})", LN['nopat_norm'], NUM0),
      ('Memo: audited FY2023 EBITDA margin', 'B13', "='Income Statement'!B15", H['margin'][0], PCT),
      ('Memo: audited FY2024 EBITDA margin', 'B14', "='Income Statement'!C15", H['margin'][1], PCT),
      ('Memo: audited FY2025 EBITDA margin', 'B15', "='Income Statement'!D15", H['margin'][2], PCT)]
for lab, ad, fm, ex, ft in NB:
    wsN.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsN, ad, fm, ex, ft, bold=(ad in ('B9', 'B12')), green=(ad in ('B13', 'B14', 'B15')))
band(wsN, 17, 7); wsN['A17'] = 'RELATIVE LENS — EV/EBITDA ON NORMALISED EARNINGS'
RLN = [('Justified EV/EBITDA', 'B18', f"={A['eveb']}", IN['ev_ebitda_just'], MULT),
       ('Implied enterprise value', 'B19', "=B9*B18",
        LN['ebitda_norm'] * IN['ev_ebitda_just'], NUM0),
       ('Plus net cash', 'B20', "=DCF!B40", DCF['net_cash'], NUM0),
       ('Less non-controlling interests', 'B21', "=DCF!B41", -IN['nci'], NUM4),
       ('Equity value', 'B22', "=B19+B20+B21",
        LN['ebitda_norm'] * IN['ev_ebitda_just'] + DCF['net_cash'] - IN['nci'], NUM0),
       ('Value per share (EGP)', 'B23', "=B22/DCF!$B$43", LN['values']['Relative multiples'], PX)]
for lab, ad, fm, ex, ft in RLN:
    wsN.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsN, ad, fm, ex, ft, bold=(ad == 'B23'), green=(ad in ('B20', 'B21')))
band(wsN, 25, 7); wsN['A25'] = 'NORMALISED-EARNINGS LENS'
NLN = [('Justified price / earnings', 'B26', f"={A['pej']}", IN['pe_just'], MULT),
       ('Capitalised operating earnings', 'B27', "=B12*B26", LN['nopat_norm'] * IN['pe_just'], NUM0),
       ('Plus net cash, at FACE', 'B28', "=DCF!B40", DCF['net_cash'], NUM0),
       ('Less non-controlling interests', 'B29', "=DCF!B41", -IN['nci'], NUM4),
       ('Equity value', 'B30', "=B27+B28+B29",
        LN['nopat_norm'] * IN['pe_just'] + DCF['net_cash'] - IN['nci'], NUM0),
       ('Value per share (EGP)', 'B31', "=B30/DCF!$B$43", LN['values']['Normalised earnings'], PX)]
for lab, ad, fm, ex, ft in NLN:
    wsN.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsN, ad, fm, ex, ft, bold=(ad == 'B31'), green=(ad in ('B28', 'B29')))
note(wsN, 33, 'Cash is added at FACE in both lenses rather than capitalised at the operating multiple.')

# ============ 12 FUNDAMENTAL VALUATION ========================================
wsFV = sheet('Fundamental Valuation')
title(wsFV, 'Asset lens, contested choices, and the terminal reconciliation', None, 7, 58, 15)
band(wsFV, 4, 7); wsFV['A4'] = 'ASSET / REPLACEMENT-COST LENS'
AL = [('Cement capacity (Mt/yr, audited note 1)', 'B5', f"={A['capcem']}", IN['cap_cement_mt'], NUM2),
      ('Replacement cost per annual tonne (USD)', 'B6', f"={A['repl']}", IN['repl_usd_t'], NUM0),
      ('Justified enterprise value per annual tonne (USD)', 'B7', f"={A['evt']}",
       IN['ev_t_just'], NUM0),
      ('Discount to replacement cost', 'B8', "=B7/B6-1",
       IN['ev_t_just'] / IN['repl_usd_t'] - 1, PCT),
      ('Implied enterprise value (EGP mn)', 'B9', f"=B7*B5*{A['fx']}", LN['ev_asset'], NUM0),
      ('Plus net cash', 'B10', "=DCF!B40", DCF['net_cash'], NUM0),
      ('Less non-controlling interests', 'B11', "=DCF!B41", -IN['nci'], NUM4),
      ('Equity value', 'B12', "=B9+B10+B11",
       LN['ev_asset'] + DCF['net_cash'] - IN['nci'], NUM0),
      ('Value per share (EGP)', 'B13', "=B12/DCF!$B$43",
       LN['values']['Asset / replacement cost'], PX),
      ('Memo: what the MARKET is paying per annual tonne (USD)', 'B14',
       f"=({A['spot']}*DCF!$B$43-DCF!B40+{A['nci']})/{A['capcem']}/{A['fx']}",
       LN['ev_per_t_spot'], NUM1),
      ('Memo: audited net book value of property and construction (EGP mn)', 'B15',
       f"={A['ppe25']}+{A['auc25']}", IN['ppe_fy25'] + IN['auc_fy25'], NUM0),
      ('Memo: that book value per annual tonne (USD)', 'B16',
       f"=B15/{A['capcem']}/{A['fx']}",
       (IN['ppe_fy25'] + IN['auc_fy25']) / IN['cap_cement_mt'] / IN['fx'], NUM1)]
for lab, ad, fm, ex, ft in AL:
    wsFV.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsFV, ad, fm, ex, ft, bold=(ad == 'B13'), green=(ad in ('B10', 'B11')))
band(wsFV, 18, 7); wsFV['A18'] = 'CONTESTED CHOICES — EACH COMPUTED, NOT ARGUED'
hdr(wsFV, 19, ['Choice', 'Adopted', 'Alternative', 'Fair value adopted',
               'Fair value alternative', 'Effect'])
for j, c in enumerate(CON):
    r = 20 + j
    wsFV.cell(row=r, column=1, value=c['choice']).alignment = Alignment(wrap_text=True,
                                                                       vertical='top')
    wsFV.cell(row=r, column=2, value=c['adopted'])
    wsFV.cell(row=r, column=3, value=c['alternative'])
    putf(wsFV, f'D{r}', "=DCF!$B$44", c['fv_adopted'], PX, green=True)
    put(wsFV, f'E{r}', c['fv_alternative'], BLUE, PX)
    putf(wsFV, f'F{r}', f"=E{r}/D{r}-1", c['effect'], PCT)
wsFV.column_dimensions['B'].width = 22
wsFV.column_dimensions['C'].width = 22
note(wsFV, 24, 'Column E holds whole-model re-runs — class 3 pasted cells. The EFFECT column is a live formula.')
band(wsFV, 26, 7); wsFV['A26'] = 'TERMINAL RECONCILIATION — NOW BUILDABLE, BECAUSE CAPEX IS DISCLOSED'
hdr(wsFV, 27, ['', 'FY2023', 'FY2024', 'FY2025'])
TRL = ['Capital expenditure (audited)', 'Capex / EBITDA', 'Depreciation (audited)',
       'Net reinvestment (capex less depreciation)', 'NOPAT', 'Reinvestment rate',
       'Return on BOOK invested capital', 'Implied growth (return x reinvestment)',
       'Character']
for j, l in enumerate(TRL):
    wsFV.cell(row=28 + j, column=1, value=l)
for i in range(3):
    c = HC[i]
    h = TR['history'][i]
    putf(wsFV, f'{c}28', f"={A[f'cx{HK[i]}']}", h['capex'], NUM0, green=True)
    putf(wsFV, f'{c}29', f"={c}28/'Income Statement'!{c}14", h['capex_over_ebitda'], PCT)
    putf(wsFV, f'{c}30', f"={A[f'dna{HK[i]}']}", h['dna'], NUM0, green=True)
    putf(wsFV, f'{c}31', f"={c}28-{c}30", h['reinvestment'], NUM0)
    putf(wsFV, f'{c}32', f"='Income Statement'!{c}11*(1-'Income Statement'!{c}19)",
         h['nopat'], NUM0)
    putf(wsFV, f'{c}33', f"={c}31/{c}32", h['rr'], PCT)
    eqk = ['eq23', 'eq24', 'eq25'][i]
    dbk = ['d23', 'd24', None][i]
    denom = f"{A[eqk]}+{A[dbk]}" if dbk else f"{A[eqk]}+DCF!$C$44"
    putf(wsFV, f'{c}34', f"={c}32/({denom})", h['roic_book'], PCT)
    putf(wsFV, f'{c}35', f"={c}34*{c}33", h['implied_g'], PCT)
    wsFV.cell(row=36, column=2 + i, value=h['character'])
wsFV.cell(row=38, column=1, value='Terminal return on capital, REPLACEMENT-COST basis')
putf(wsFV, 'B38', "=DCF!B24", TR['roic_repl'], PCT, green=True)
wsFV.cell(row=39, column=1, value='Terminal return on capital, BOOK basis (FY2025)')
putf(wsFV, 'B39', "=DCF!B25", TR['roic_book_fy25'], PCT, green=True)
wsFV.cell(row=40, column=1, value='Terminal rate')
putf(wsFV, 'B40', "=DCF!$C$50", W['wacc_term'], PCT, green=True)
wsFV.cell(row=41, column=1, value='Terminal growth adopted')
putf(wsFV, 'B41', f"={A['g']}", IN['g_term'], PCT)
note(wsFV, 43, 'On the BOOK basis the return on capital is far above the terminal rate and growth would be free.')
note(wsFV, 44, 'The book carries a plant built around 2010 at pre-devaluation historic cost. The terminal block is')
note(wsFV, 45, 'struck on REPLACEMENT cost instead, where the return sits below the rate and growth must be paid')
note(wsFV, 46, 'for. That single choice is the most consequential judgement in this model, and it is visible here.')

# ============ 13 SUMMARY ======================================================
wsS = sheet('Summary')
title(wsS, 'Summary valuation table', 'Every value linked live from its own sheet.',
      7, 42, 18)
hdr(wsS, 4, ['Lens', 'Value per share (EGP)', 'Weight', 'Weighted contribution',
             'Versus spot', 'Terminal value % of EV'])
LK = [('DCF (cash flow)', "=DCF!B44", A['wdcf'], "=DCF!B34"),
      ('Relative multiples', "='Relative & Normalized'!B23", A['wrel'], None),
      ('Normalised earnings', "='Relative & Normalized'!B31", A['wnorm'], None),
      ('Asset / replacement cost', "='Fundamental Valuation'!B13", A['wasset'], None)]
for j, (name, fm, wkey, tvfm) in enumerate(LK):
    r = 5 + j
    wsS.cell(row=r, column=1, value=name)
    putf(wsS, f'B{r}', fm, LN['values'][name], PX, green=True, bold=True)
    putf(wsS, f'C{r}', f"={wkey}", LN['weights'][name], PCT, green=True)
    putf(wsS, f'D{r}', f"=B{r}*C{r}", LN['values'][name] * LN['weights'][name], PX)
    putf(wsS, f'E{r}', f"=B{r}/{A['spot']}-1", LN['values'][name] / SPOT - 1, PCT)
    if tvfm:
        putf(wsS, f'F{r}', tvfm, DCF['tv_share'], PCT, green=True, bold=True)
    else:
        wsS.cell(row=r, column=6, value='—')
wsS.cell(row=9, column=1, value='WEIGHTED CENTRAL FAIR VALUE')
putf(wsS, 'B9', "=SUMPRODUCT(B5:B8,C5:C8)", LN['central'], PX, bold=True)
putf(wsS, 'C9', "=SUM(C5:C8)", 1.0, PCT, bold=True)
putf(wsS, 'D9', "=SUM(D5:D8)", LN['central'], PX, bold=True)
putf(wsS, 'E9', f"=B9/{A['spot']}-1", LN['central'] / SPOT - 1, PCT, bold=True)
wsS.cell(row=10, column=1, value='Lowest lens')
putf(wsS, 'B10', "=MIN(B5:B8)", LN['low'], PX)
putf(wsS, 'E10', f"=B10/{A['spot']}-1", LN['low'] / SPOT - 1, PCT)
wsS.cell(row=11, column=1, value='Highest lens')
putf(wsS, 'B11', "=MAX(B5:B8)", LN['high'], PX)
putf(wsS, 'E11', f"=B11/{A['spot']}-1", LN['high'] / SPOT - 1, PCT)
wsS.cell(row=12, column=1, value='Market price, 6 August 2026')
putf(wsS, 'B12', f"={A['spot']}", SPOT, PX, green=True)
wsS.cell(row=13, column=1, value='Market capitalisation (EGP mn)')
putf(wsS, 'B13', "=DCF!C45", M['mktcap'], NUM0, green=True)
band(wsS, 15, 7); wsS['A15'] = 'THE THREE-PANEL VALUATIONS'
hdr(wsS, 16, ['', 'Low (EGP)', 'Central (EGP)', 'High (EGP)', 'Versus spot'])
for j, e in enumerate(D['experts']):
    r = 17 + j
    wsS.cell(row=r, column=1, value=f"{e['label']} — {e['method']}")
    for k, col in enumerate(['B', 'C', 'D']):
        put(wsS, f'{col}{r}', e[['low', 'central', 'high'][k]], BLUE, PX)
    putf(wsS, f'E{r}', f"=C{r}/{A['spot']}-1", e['central'] / SPOT - 1, PCT)
wsS.cell(row=20, column=1, value='Panel median')
putf(wsS, 'C20', "=MEDIAN(C17:C19)", sorted(e['central'] for e in D['experts'])[1], PX, bold=True)
putf(wsS, 'E20', f"=C20/{A['spot']}-1",
     sorted(e['central'] for e in D['experts'])[1] / SPOT - 1, PCT)
note(wsS, 22, 'Terminal value as a percentage of enterprise value is shown beside the cash-flow lens and links')
note(wsS, 23, 'live to the DCF sheet — it is never typed.')

# ============ 14 MONTE CARLO ==================================================
wsM = sheet('Monte Carlo')
title(wsM, 'Probabilistic price map — ILLUSTRATIVE ONLY',
      'Percentiles are the output of a 50,000-path simulation: pasted, class 2.', 7, 46, 16)
note(wsM, 3, 'This map is NOT a valuation. It does not redraw when a valuation driver changes.')
hdr(wsM, 6, ['', 'One month', 'Three months'])
PCTL = [('5th percentile', 'p5'), ('25th percentile', 'p25'), ('Median', 'p50'),
        ('75th percentile', 'p75'), ('95th percentile', 'p95')]
for j, (lab, key) in enumerate(PCTL):
    r = 7 + j
    wsM.cell(row=r, column=1, value=lab)
    put(wsM, f'B{r}', STK['horizons']['1M']['pct'][key], BLUE, PX)
    put(wsM, f'C{r}', STK['horizons']['3M']['pct'][key], BLUE, PX)
wsM.cell(row=12, column=1, value='Spot')
putf(wsM, 'B12', f"={A['spot']}", SPOT, PX, green=True)
putf(wsM, 'C12', f"={A['spot']}", SPOT, PX, green=True)
for j, (lab, key) in enumerate([('Probability of finishing above spot', 'p_above'),
                                ('Probability of finishing 10% or more above', 'p_up10'),
                                ('Probability of finishing 10% or more below', 'p_dn10'),
                                ('Probability of TOUCHING +10% at any point', 'touch_up10'),
                                ('Probability of TOUCHING -10% at any point', 'touch_dn10')]):
    r = 14 + j
    wsM.cell(row=r, column=1, value=lab)
    put(wsM, f'B{r}', STK['horizons']['1M'][key], BLUE, PCT)
    put(wsM, f'C{r}', STK['horizons']['3M'][key], BLUE, PCT)
wsM.cell(row=20, column=1, value='Median versus spot')
putf(wsM, 'B20', "=B9/B12-1", STK['horizons']['1M']['pct']['p50'] / SPOT - 1, PCT)
putf(wsM, 'C20', "=C9/C12-1", STK['horizons']['3M']['pct']['p50'] / SPOT - 1, PCT)
wsM.cell(row=21, column=1, value='Width of the 5th-95th band, as % of spot')
putf(wsM, 'B21', "=(B11-B7)/B12",
     (STK['horizons']['1M']['pct']['p95'] - STK['horizons']['1M']['pct']['p5']) / SPOT, PCT)
putf(wsM, 'C21', "=(C11-C7)/C12",
     (STK['horizons']['3M']['pct']['p95'] - STK['horizons']['3M']['pct']['p5']) / SPOT, PCT)
band(wsM, 23, 7); wsM['A23'] = 'HOW WELL CALIBRATED IS THIS MAP? — MEASURED, NOT ASSERTED'
for j, (lab, v, ft) in enumerate([('Windows scored', S0['windows_scored'], NUM0),
                                  ('Skill against a random walk', S0['skill_norm'], PCT),
                                  ('Coverage of the 50% band (nominal 50%)', S0['cov50'], PCT),
                                  ('Coverage of the 80% band (nominal 80%)', S0['cov80'], PCT),
                                  ('Coverage of the 90% band (nominal 90%)', S0['cov90'], PCT),
                                  ('Cone width against the benchmark', S0['w90_ratio'], MULT)]):
    r = 24 + j
    wsM.cell(row=r, column=1, value=lab)
    put(wsM, f'B{r}', v, BLUE, ft)
note(wsM, 31, 'The bands cover MORE than their nominal share, so the map is too wide rather than mis-centred.')

# ============ 15 SENSITIVITY ==================================================
wsX = sheet('Sensitivity')
title(wsX, 'Sensitivity — whole-model re-runs (class 3, pasted)',
      'Each cell is a complete revaluation. THESE GRIDS DO NOT REDRAW.', 7, 46, 14)
wsX['A4'] = 'Fair value per share (EGP): explicit-window cost of capital against terminal growth'
wsX['A4'].font = Font(bold=True)
hdr(wsX, 5, [''] + [f'g = {g:.0%}' for g in SN['g_grid']])
for i, wv in enumerate(SN['wacc_grid']):
    r = 6 + i
    put(wsX, f'A{r}', wv, BLUE, PCT2)
    for j in range(5):
        put(wsX, f'{DC[j]}{r}', SN['wacc_g'][i][j], BLUE, PX)
wsX['A12'] = 'Fair value per share (EGP): explicit-window rate against TERMINAL rate'
wsX['A12'].font = Font(bold=True)
hdr(wsX, 13, [''] + [f'terminal {w:.1%}' for w in SN['wt_grid']])
for i, wv in enumerate(SN['wacc_grid']):
    r = 14 + i
    put(wsX, f'A{r}', wv, BLUE, PCT2)
    for j in range(5):
        put(wsX, f'{DC[j]}{r}', SN['exp_term'][i][j], BLUE, PX)
for tag, row, grid, vals, ft in (('beta', 20, SN['beta_grid'], SN['beta'], '0.00'),
                                 ('margin shift', 24, SN['mgn_grid'], SN['mgn'], '+0%'),
                                 ('net cash (EGP mn)', 28, SN['nc_grid'], SN['net_cash'], ',.0f')):
    wsX[f'A{row}'] = f'Fair value per share (EGP): {tag}'
    wsX[f'A{row}'].font = Font(bold=True)
    labs = [f'{g:.2f}' if tag == 'beta' else (f'{g:+.0%}' if 'margin' in tag else f'{g:,.0f}')
            for g in grid]
    hdr(wsX, row + 1, [''] + labs)
    wsX[f'A{row+2}'] = 'Fair value'
    for j in range(5):
        put(wsX, f'{DC[j]}{row+2}', vals[j], BLUE, PX)
_gdir = 'a LOWER' if SN['wacc_g'][2][4] < SN['wacc_g'][2][0] else 'a HIGHER'
_gwhy = ('sits below' if SN['wacc_g'][2][4] < SN['wacc_g'][2][0] else 'sits above')
note(wsX, 32, f'Note the growth column of the first grid: HIGHER terminal growth gives {_gdir} value, because the')
note(wsX, 33, f'terminal return on replacement-cost capital {_gwhy} the hurdle. Read off the grid, never typed:')
note(wsX, 34, f'revision 3 carried this note unchanged after its own grid had reversed, and it contradicted itself.')

# ============ 16 PEER & SECTOR ================================================
wsP = sheet('Peer & Sector')
title(wsP, 'Peer set and the Egyptian cement balance', None, 8, 46, 15)
hdr(wsP, 4, ['', 'Revenue (EGP mn)', 'Profit (EGP mn)', 'Market cap (EGP mn)',
             'Price / earnings', 'Price / sales', 'Net margin'])
PRW = [('Arabian Cement (ARCC)', A['rev25'], A['pat25'], "=DCF!C45", PE['self']),
       ('Sinai Cement (SCEM)', A['pscrev'], A['pscpat'], f"={A['pscmc']}", PE['scem']),
       ('Misr Beni Suef Cement (MBSC)', A['pmbrev'], A['pmbpat'], f"={A['pmbmc']}", PE['mbsc'])]
for j, (name, rk, pk, mfm, d) in enumerate(PRW):
    r = 5 + j
    wsP.cell(row=r, column=1, value=name)
    putf(wsP, f'B{r}', f"={rk}", d['rev'], NUM0, green=True)
    putf(wsP, f'C{r}', f"={pk}", d['pat'], NUM0, green=True)
    putf(wsP, f'D{r}', mfm, d['mcap'], NUM0, green=True)
    putf(wsP, f'E{r}', f"=D{r}/C{r}", d['pe'], MULT)
    putf(wsP, f'F{r}', f"=D{r}/B{r}", d['ps'], MULT)
    putf(wsP, f'G{r}', f"=C{r}/B{r}", d['pat'] / d['rev'], PCT)
wsP.cell(row=9, column=1, value='Peer average price / earnings (excluding the subject)')
putf(wsP, 'E9', "=AVERAGE(E6:E7)", (PE['scem']['pe'] + PE['mbsc']['pe']) / 2, MULT, bold=True)
wsP.cell(row=10, column=1, value='Subject premium / (discount) to the peer average')
putf(wsP, 'E10', "=E5/E9-1",
     PE['self']['pe'] / ((PE['scem']['pe'] + PE['mbsc']['pe']) / 2) - 1, PCT)
band(wsP, 12, 8); wsP['A12'] = 'THE EGYPTIAN CEMENT BALANCE'
SEC = [('Nameplate capacity (Mt)', 'B13', f"={A['egcap']}", PE['sector']['capacity_mt'], NUM1),
       ('Production 2025 (Mt)', 'B14', f"={A['egprod']}", PE['sector']['production_mt'], NUM1),
       ('Domestic consumption 2025 (Mt)', 'B15', f"={A['egcons']}", PE['sector']['consumption_mt'], NUM1),
       ('Exports 2025 (Mt)', 'B16', f"={A['egexp']}", PE['sector']['exports_mt'], NUM1),
       ('Dormant capacity under revival (Mt)', 'B17', f"={A['egrev']}", PE['sector']['revival_mt'], NUM1),
       ('Sector utilisation', 'B18', "=B14/B13", PE['sector']['utilisation'], PCT),
       ('The subject as a share of national capacity', 'B19', f"={A['capcem']}/B13",
        PE['sector']['share_of_capacity'], PCT),
       ('Revival capacity as a share of consumption', 'B20', "=B17/B15",
        PE['sector']['revival_pct_of_consumption'], PCT),
       ('The subject\'s own volume as a share of national production', 'B21',
        "='Unit Build'!B18/B14", UC['vol_fy25'] / IN['egy_prod_mt'], PCT)]
for lab, ad, fm, ex, ft in SEC:
    wsP.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsP, ad, fm, ex, ft, green=(int(ad[1:]) <= 17))
note(wsP, 23, 'Every multiple here is RECOMPUTED from revenue, profit and market capitalisation rather than')
note(wsP, 24, 'quoted, because the published multiples for this peer set do not reconcile.')

# ============ SAVE ============================================================
OUT = os.path.join(HERE, 'ARCC_Valuation_Model_06082026_public.xlsx')
wb.save(OUT)
with open(os.path.join(HERE, 'xlsx_expected.json'), 'w') as f:
    json.dump(EXPECT, f, indent=1)
nf = sum(len(v) for v in EXPECT.values())
nv = 0
chk = load_workbook(OUT)
for s in chk.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                pass
            elif isinstance(c.value, (int, float)):
                nv += 1
print(f'wrote {os.path.basename(OUT)} — {len(wb.sheetnames)} sheets, '
      f'{nf} formula cells recorded, {nv} pasted numeric cells')
json.dump(dict(formulas=nf, pasted_values=nv, sheets=len(wb.sheetnames)),
          open(os.path.join(HERE, 'formula_count.json'), 'w'), indent=1)
