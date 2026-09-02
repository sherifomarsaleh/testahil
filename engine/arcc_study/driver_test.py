"""Prove the workbook is a LIVE DRIVER model.

READ FIRST tells the reader that changing a blue cell on Assumptions reprices the model.
That is a claim about the DELIVERED file, so it is tested on the delivered file: each
driver is perturbed IN PLACE, the whole workbook is re-evaluated from scratch through the
independent evaluator, and the test asserts the headline moves in the asserted DIRECTION.
A dead-input sweep then bumps every remaining driver and requires it to move something.

THREE DIRECTIONS HERE ARE THE OPPOSITE OF THE TEXTBOOK ONE, AND NONE IS A BUG. Each was
decomposed BEFORE the sign was set, not after a test failed.

  * Higher terminal growth LOWERS the value. Terminal return on invested capital is 10.3%
    against a terminal rate of 16.5%, so growth has to be bought with reinvestment that
    earns less than it costs. The reinvestment rate is g / return on capital, so raising g
    raises what must be spent faster than it raises what is earned.

  * A WIDER sovereign spread RAISES the value. The spread is netted OUT of the local
    risk-free rate before the country equity premium is added, precisely so that Egypt's
    default risk is charged once rather than twice. Widening it therefore lowers the
    normalised risk-free rate and with it the cost of equity.

  * Depreciation runs BOTH ways, and which way depends on the year. A heavier charge is
    added back inside free cash flow, so in a mid-window year it is worth only its tax
    shield and the value RISES (FY2027 +1pp of revenue: enterprise value +35mn, terminal
    block untouched). In the TERMINAL BASE YEAR the same bump runs the other way, because
    capital expenditure here is set in dollars per tonne of capacity and does not follow
    the book charge: year-five NOPAT falls, and the terminal value falls with it (FY2030
    +1pp: enterprise value -502mn, of which -527mn is the terminal block). The first
    version of this test asserted the terminal mechanism against a mid-window bump and
    failed. The expectation was wrong, not the model.

  * Cash and debt WERE not clean one-way bridge levers, and in revision 4 they are. In the
    earlier revisions the effective tax rate was INFERRED from the FY2025 closure —
    disclosed operating profit plus modelled net finance income against disclosed profit
    after tax — so a balance-sheet change moved the imputed tax rate on every forecast
    year and the two legs very nearly cancelled. Revision 4 takes the effective rate from
    a disclosed figure, so EGP 1,000mn of cash on the reviewed 30-June-2026 balance sheet
    now moves the value by exactly 1,000 / shares outstanding and by nothing else.

WHICH BALANCE SHEET IS THE BRIDGE ON. Revision 4 moved the valuation date to 30 June 2026
and put the bridge on the reviewed balance sheet of that date. The FY2025 cash, minority
and declared-dividend rows are therefore HISTORY: they open the cash roll-forward and close
the audited year, and they no longer touch the headline. Revision 3's assertions said they
did, and those assertions passed for a month because this gate was pointed at the
superseded 06-08-2026 workbook rather than the delivered one — an empty answer wearing the
costume of a clean one [R-ENF-04]. It is pointed at the delivered file now, and each of
those five drivers is asserted against what it actually moves.
"""
import json, os
import openpyxl
import xlcalc

GDV = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  'study_numbers.json')))['growth_destroys_value']

HERE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(os.path.join(HERE, 'ARCC_Valuation_Model_02092026_public.xlsx'))
A = {}
for row in wb['Assumptions'].iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str):
        A.setdefault(c.value, c.row)


def row_of(label):
    if label not in A:
        raise KeyError(f'no Assumptions row labelled {label!r}')
    return A[label]


def read(overrides=None):
    bk = xlcalc.Book(wb, overrides)
    return dict(
        dcf=bk.cell_value('DCF', 'B44'),
        central=bk.cell_value('Summary', 'B9'),
        pv_expl=bk.cell_value('DCF', 'B31'),
        pv_tv=bk.cell_value('DCF', 'B32'),
        ev=bk.cell_value('DCF', 'B33'),
        netcash=bk.cell_value('DCF', 'B40'),
        cashval=bk.cell_value('DCF', 'B38'),
        wacc=bk.cell_value('DCF', 'C40'),
        kd=bk.cell_value('DCF', 'C41'),
        eurshare=bk.cell_value('DCF', 'C46'),
        wacc_term=bk.cell_value('DCF', 'C50'),
        beta_term=bk.cell_value('DCF', 'C47'),
        roic=bk.cell_value('DCF', 'B24'),
        roic_book=bk.cell_value('DCF', 'B25'),
        kd_eff25=bk.cell_value('DCF', 'B61'),
        kd_egp=bk.cell_value('DCF', 'B63'),
        vol25=bk.cell_value('Segments', 'B18'),
        util25=bk.cell_value('Segments', 'B14'),
        ccost25=bk.cell_value('Segments', 'B38'),
        cmat25=bk.cell_value('Segments', 'B35'),
        rev26=bk.cell_value('Segments', 'C63'),
        vol26=bk.cell_value('Segments', 'C55'),
        cc26=bk.cell_value('Segments', 'C69'),
        cmat26=bk.cell_value('Segments', 'C65'),
        ebitda25=bk.cell_value('Segments', 'B71'),
        ebitda26=bk.cell_value('Segments', 'C71'),
        resid_rev=bk.cell_value('Segments', 'B78'),
        gp23=bk.cell_value('Income Statement', 'B7'),
        gp25=bk.cell_value('Income Statement', 'D7'),
        ebit25=bk.cell_value('Income Statement', 'D11'),
        ebitda_is25=bk.cell_value('Income Statement', 'D14'),
        taxe25=bk.cell_value('Income Statement', 'D19'),
        pat26=bk.cell_value('Income Statement', 'E20'),
        dps25=bk.cell_value('Income Statement', 'D23'),
        cash30=bk.cell_value('Balance Sheet', 'I9'),
        bvps25=bk.cell_value('Balance Sheet', 'D17'),
        roe25=bk.cell_value('Balance Sheet', 'D18'),
        close=bk.cell_value('Balance Sheet', 'B25'),
        nd23=bk.cell_value('Balance Sheet', 'B16'),
        othliab=bk.cell_value('Balance Sheet', 'D13'),
        rel_lens=bk.cell_value('Relative & Normalized', 'B23'),
        norm_lens=bk.cell_value('Relative & Normalized', 'B31'),
        asset_lens=bk.cell_value('Fundamental Valuation', 'B13'),
        ev_per_t=bk.cell_value('Fundamental Valuation', 'B14'),
        bookval_t=bk.cell_value('Fundamental Valuation', 'B16'),
        rr23=bk.cell_value('Fundamental Valuation', 'B33'),
        rr25=bk.cell_value('Fundamental Valuation', 'D33'),
        roicb24=bk.cell_value('Fundamental Valuation', 'C34'),
        sh_out=bk.cell_value('Per-Share & Ratios', 'B16'),
        sh_div=bk.cell_value('Per-Share & Ratios', 'B17'),
        parchk=bk.cell_value('Per-Share & Ratios', 'B19'),
        runrate=bk.cell_value('Per-Share & Ratios', 'B21'),
        q1mgn=bk.cell_value('Per-Share & Ratios', 'B24'),
        q1ann=bk.cell_value('Per-Share & Ratios', 'B26'),
        q1nc=bk.cell_value('Per-Share & Ratios', 'B29'),
        pe_spot=bk.cell_value('Per-Share & Ratios', 'D7'),
        peer_pe=bk.cell_value('Peer & Sector', 'E9'),
        peer_ps_scem=bk.cell_value('Peer & Sector', 'F6'),
        peer_ps_mbsc=bk.cell_value('Peer & Sector', 'F7'),
        sector_util=bk.cell_value('Peer & Sector', 'B18'),
        share_cap=bk.cell_value('Peer & Sector', 'B19'),
        revival=bk.cell_value('Peer & Sector', 'B20'),
        share_prod=bk.cell_value('Peer & Sector', 'B21'),
        exports=bk.cell_value('Peer & Sector', 'B16'),
        # FY2025 revenue growth sits in the FY2025 column (D), not the FY2026 one.
        growth25=bk.cell_value('Summary Financials', 'D13'),
        # The sweep is only as strong as the span of what it watches. These reach the
        # corners of the workbook that no headline touches.
        # the corners of the physical build: both capacity constraints, all three
        # DERIVED prices, and the product split. None of these existed in revision 3.
        kiln_pk=bk.cell_value('Segments', 'B94'),
        mill_pk=bk.cell_value('Segments', 'B98'),
        mill26=bk.cell_value('Segments', 'C51'),
        p_loc=bk.cell_value('Segments', 'B22'),
        p_ecem=bk.cell_value('Segments', 'B26'),
        p_eclk=bk.cell_value('Segments', 'B27'),
        clk_prod=bk.cell_value('Segments', 'B7'),
        cem_prod=bk.cell_value('Segments', 'B12'),
        cap_chk=bk.cell_value('Segments', 'B28'),
        nonop25=bk.cell_value('Income Statement', 'D16'),
        nonop_chk=bk.cell_value('Income Statement', 'B24'),
        roe24=bk.cell_value('Balance Sheet', 'C18'),
        ta24=bk.cell_value('Balance Sheet', 'C11'),
        ta23=bk.cell_value('Balance Sheet', 'B11'),
        nca25=bk.cell_value('Balance Sheet', 'D8'),
        wc25=bk.cell_value('Balance Sheet', 'D10'),
        cash25=bk.cell_value('Balance Sheet', 'D9'),
        nd24=bk.cell_value('Balance Sheet', 'C16'),
        kd_eff24=bk.cell_value('DCF', 'B60'),
    )


base = read()
print('base: ' + ' · '.join(
    f'{k} {v:,.4f}' if isinstance(v, (int, float)) else f'{k} {v}'
    for k, v in base.items()))

CASES = [
    # ---- THE UNIT BUILD, now anchored on the audited notes ---------------------
    # The build now runs the OTHER way round: the plant sets the tonnes and the prices
    # fall out of the audited revenue. Every one of these directions is the reverse of
    # revision 3's, and that is the whole point of the rebuild.
    ('Kiln utilisation  (THE volume driver)', 'B', +0.02, 'clk_prod', +1,
     'running the kiln harder must make more clinker — the driver revision 3 did not have'),
    ('Kiln utilisation  (THE volume driver)', 'B', +0.02, 'vol25', +1,
     'and more clinker must mean more tonnes despatched'),
    ('Kiln utilisation  (THE volume driver)', 'B', +0.02, 'p_loc', -1,
     'prices are DERIVED, so more tonnes behind the same audited revenue is a LOWER '
     'realised price. This is the test revision 3 could not run: its price was an input '
     'and its FY2025 residual was an identity that could not fail'),
    ('Clinker sold as clinker, share of clinker made', 'B', +0.05, 'cem_prod', -1,
     'every tonne shipped as clinker is a tonne that is not ground into cement'),
    ('Clinker sold as clinker, share of clinker made', 'B', +0.05, 'p_loc', +1,
     'and fewer local cement tonnes behind the same audited local revenue is a HIGHER '
     'derived cement price'),
    ('Clinker factor', 'B', +0.05, 'cem_prod', -1,
     'more clinker per tonne of cement means less cement from the same kiln'),
    ('Export clinker price as a fraction of export cement', 'B', +0.10, 'p_ecem', -1,
     'lifting the clinker leg must lower the cement leg — the audited export revenue is '
     'fixed and the two prices split it between them'),
    ('Export clinker price as a fraction of export cement', 'B', +0.10, 'p_eclk', +1,
     'and the clinker price itself must rise'),
    ('Cement exported, share of cement sold', 'B', +0.03, 'p_loc', +1,
     'exporting more cement leaves fewer local tonnes behind the same local revenue'),
    ('Kiln utilisation  (THE volume driver)', 'C', +0.03, 'vol26', +1,
     'running the plant harder must make more cement'),
    ('Kiln utilisation  (THE volume driver)', 'C', +0.03, 'rev26', +1,
     'and more cement must raise revenue'),
    ('Kiln utilisation  (THE volume driver)', 'C', +0.03, 'kiln_pk', +1,
     'and the capacity check must SEE it — revision 3 ran a kiln test on a volume base '
     'that ignored clinker exports, so it could never bind'),
    ('Clinker sold as clinker, share of clinker made', 'C', +0.05, 'mill26', -1,
     'shipping more clinker leaves the mill with less to grind. Asserted on the FY2026 '
     'column, not the peak: the peak is a MAX across the window and sits in FY2030'),
    ('Local price index', 'C', +0.05, 'rev26', +1, 'a higher local price must raise revenue'),
    ('Export price index (USD)', 'C', +0.05, 'rev26', +1,
     'a higher export price must raise revenue'),
    ('USD/EGP path', 'C', +5.0, 'rev26', +1,
     'a weaker pound raises the pound value of export revenue'),
    ('Cement exported, share of cement sold', 'C', +0.05, 'rev26', -1,
     'export cement realises less per tonne than local, so a heavier export mix LOWERS '
     'revenue'),
    ('Local cost-inflation index', 'C', +0.10, 'cc26', +1,
     'inflating the pound cost lines must raise cost per tonne'),
    ('Local cost-inflation index', 'C', +0.10, 'ebitda26', -1, 'and must cut EBITDA'),
    ('Alternative-fuel saving on materials', 'C', +0.05, 'cmat26', -1,
     'the substitution programme must cut the materials and fuel bill'),
    ('Alternative-fuel saving on materials', 'C', +0.05, 'ebitda26', +1,
     'and the saving must reach EBITDA — this is the company-specific lever, and the EBRD '
     'facility funding it is on the audited balance sheet'),
    ('Services revenue as a share of goods revenue', 'B', +0.02, 'ebitda25', +1,
     'more transportation revenue on the same tonnes, in the BASE year. It does not move '
     'FY2026 revenue and must not: revision 4 calibrates price, cost AND services on the '
     'same reviewed half of 2026, so the forecast year is anchored on a filed actual '
     'rather than grown off the base-year share'),
    ('Cement capacity', 'B', +0.30, 'mill26', -1,
     'the mill is now a CONSTRAINT, not a volume driver: tonnes come off the kiln, so more '
     'mill capacity lowers mill utilisation and does not create cement. Revision 3 drove '
     'volume off cement capacity, which is why its kiln check could never bind'),
    ('Cement capacity', 'B', +0.30, 'asset_lens', +1,
     'and more capacity at the same value per tonne must raise the asset lens'),
    ('FY2025 cost of sales — materials and fuel', 'B', +200.0, 'ccost25', +1,
     'a bigger disclosed materials bill is a higher cost per tonne'),
    ('FY2025 cost of sales — transportation', 'B', +100.0, 'ebitda26', -1,
     'and a bigger transport bill carries into the forecast cost stack'),
    ('FY2025 cost of sales — overheads', 'B', +100.0, 'ebitda26', -1, 'as do overheads'),
    # ---- COST OF CAPITAL --------------------------------------------------------
    # The naive expectation here is that growth destroys value because terminal return on
    # capital (14.50%) sits below the terminal rate (15.75%). That is the WRONG test, and
    # revision 3 crossed the line that exposes it. Reinvestment is g/ROIC and ROIC is
    # N*(1+g)/IC, so the reinvestment charge collapses to a fixed g*IC and the block is
    #        TV(g) = [N*(1+g) - g*IC] / (W - g),   dTV/dg  prop.  N*(1+W) - IC*W
    # which contains no g at all: the direction is a CONSTANT of the model and the hurdle
    # is N/IC vs W/(1+W), not ROIC vs W — the two differ by (1+g)/(1+W) because ROIC is
    # measured on terminal-YEAR profit against a valuation-DATE capital base. Revision 2
    # sat at N/IC = 8.2% against the 13.61% hurdle and growth destroyed value; the
    # corrected price path lifts terminal NOPAT to N/IC = 13.81%, 21bp past it, so growth
    # now adds value. The magnitude is what matters and it is trivial: +0.10% of the DCF
    # per point of terminal growth, +0.5% across the whole 3%-7% range.
    ('Terminal growth rate', 'B', +0.01, 'dcf', (1 if GDV['analytic_adds_value'] else -1),
     f"the growth lever moves in the direction the terminal algebra requires. Its sign is "
     f"the constant N(1+W) - IC.W, so THE DIRECTION IS READ FROM THE MODEL rather than "
     f"typed: N/IC {GDV['n_over_ic']:.2%} against a hurdle of W/(1+W) {GDV['hurdle']:.2%}. "
     f"Revision 3 hard-typed it, and when the terminal capital was restated into "
     f"terminal-year pounds the sign reversed and the assertion went stale"),
    ('Beta (own-stock weekly regression)', 'B', +0.20, 'dcf', -1,
     'a higher beta must lower the valuation'),
    ('Beta (own-stock weekly regression)', 'B', +0.20, 'beta_term', +1,
     'and it must re-lever into the terminal beta'),
    ('Risk-free rate (EGP 10-year government)', 'B', +0.02, 'dcf', -1,
     'a higher risk-free rate must lower the valuation'),
    ('Sovereign default spread (netted out)', 'B', +0.01, 'dcf', +1,
     'the spread is netted OUT of the risk-free rate, so a wider one LOWERS the cost of equity'),
    ('Equity risk premium', 'B', +0.02, 'dcf', -1, 'a higher premium must lower the valuation'),
    ('Terminal risk-free rate', 'B', +0.02, 'dcf', -1,
     'a higher terminal risk-free rate must lower the valuation'),
    ('Terminal equity risk premium', 'B', +0.02, 'dcf', -1,
     'a higher terminal premium must lower the valuation'),
    ('Terminal debt weight', 'B', +0.10, 'beta_term', +1,
     'more terminal leverage must RAISE the re-levered beta'),
    ('Terminal cost of debt', 'B', +0.03, 'wacc_term', +1,
     'dearer terminal debt must raise the terminal rate'),
    ('Euribor (EBRD and NBE reference rate)', 'B', +0.02, 'kd', +1,
     '91% of the audited debt book is euro-denominated at Euribor-linked rates, so the '
     'reference rate carries almost all of the blended cost of debt'),
    ('Euribor (EBRD and NBE reference rate)', 'B', +0.02, 'wacc', +1,
     'and it must raise the blended cost of capital, even if barely — debt is 4.9% of it'),
    ('EGP marginal borrowing rate (corridor + 0.6%)', 'B', +0.05, 'kd', +1,
     'the pound facility is only 8.8% of the book, so this moves the blend by less'),
    ('Expected EGP depreciation against the euro', 'B', +0.02, 'kd_egp', +1,
     'loading the euro legs with faster pound depreciation must raise the '
     'pound-equivalent cost of debt'),
    ('EGP marginal cost-of-debt path', 'C', +0.02, 'dcf', -1,
     'a slower easing path flattens the glide, so the second year is discounted harder'),
    ('EBRD facility — EUR', 'B', +500.0, 'eurshare', +1,
     'a larger euro facility raises the euro share of the book'),
    ('EBRD facility — EUR', 'B', +500.0, 'kd', -1,
     'and because the euro rate is far below the pound rate, it LOWERS the blended cost'),
    ('Statutory tax rate', 'B', +0.02, 'wacc', -1,
     'a higher statutory rate deepens the tax shield on debt and lowers the blended rate'),
    ('Elapsed fraction of FY2026 at the valuation date', 'B', +0.10, 'dcf', +1,
     'a longer elapsed fraction means less of FY2026 left to discount and a shorter '
     'discount period on every year behind it, so the value rises. It does NOT move the '
     'cash at the valuation date any more: revision 4 reads that off the reviewed 30-June '
     '2026 balance sheet instead of rolling FY2025 forward, and a filed figure does not '
     'respond to a modelling assumption'),
    # ---- CAPITAL INTENSITY AND THE BRIDGE --------------------------------------
    ('Maintenance capital expenditure', 'B', +1.00, 'dcf', -1,
     'more capital spending leaves less free cash flow'),
    ('Depreciation as % of revenue', 'C', +0.01, 'dcf', +1,
     'a heavier charge in a MID-window year is worth only its tax shield: it is added back '
     'inside free cash flow and the terminal base year is untouched, so the value RISES'),
    ('Depreciation as % of revenue', 'F', +0.01, 'dcf', -1,
     'the same bump in the TERMINAL BASE YEAR runs the other way: year-five NOPAT falls and '
     'the terminal value falls with it, because capex is set in dollars per tonne and does '
     'not follow the book charge'),
    ('Depreciation as % of revenue', 'F', +0.01, 'pv_tv', -1,
     'and the loss is located in the terminal block, where the decomposition said it would be'),
    ('Change in working capital / change in revenue', 'B', +0.05, 'dcf', -1,
     'growth funded in working capital is growth the shareholder does not receive'),
    ('Yield earned on cash', 'C', +0.03, 'cash30', +1,
     'a better return on the cash pile leaves more cash at the end of the forecast'),
    ('Dividend payout ratio', 'B', +0.20, 'cash30', -1,
     'paying more out leaves less cash at the end of the forecast'),
    ('Cash and bank balances, reviewed 30 June 2026', 'B', +1000.0, 'dcf', +1,
     'THE bridge cash lever, and in revision 4 it is exactly clean: EGP 1,000mn of cash '
     'moves the value by 1,000 / shares outstanding and by nothing else, because the '
     'effective tax rate is now a disclosed figure rather than one inferred by closing a '
     'modelled finance income'),
    ('CIB credit facilities — EGP', 'B', +500.0, 'dcf', +1,
     'more debt RAISES the value, and that is a structural consequence of a correction '
     'made this revision rather than a defect. The bridge now deducts the FRESHER reviewed '
     '30-June-2026 debt, so a change to the FY2025 facility no longer moves net cash; it '
     'moves only the WACC weight, and after-tax debt is far cheaper than equity'),
    ('Interest-bearing debt, reviewed 30 June 2026', 'B', +500.0, 'dcf', -1,
     'the bridge deducts the debt on the LATEST reviewed balance sheet, so more of it is '
     'less equity, one for one'),
    ('Non-controlling interests, reviewed 30 June 2026', 'B', +0.02, 'dcf', -1,
     'minorities own part of the enterprise and must be deducted — on the same reviewed '
     'balance sheet the cash and the debt come off'),
    ('Cash and bank balances FY2025', 'B', +1000.0, 'cash30', +1,
     'the FY2025 balance is HISTORY in revision 4, not a bridge lever: it opens the cash '
     'roll-forward and so moves the closing FY2030 balance, while the bridge stands on the '
     'reviewed 30-June-2026 sheet above. Revision 3 asserted this against the headline and '
     'the assertion was never run, because the gate was pointed at the superseded workbook'),
    ('Non-controlling interests FY2025', 'B', +0.02, 'close', -1,
     'likewise history: it moves the balance-sheet closure check on the audited year and '
     'not the valuation, whose minority is the reviewed 30-June-2026 figure'),
    ('Ordinary shares issued', 'B', +20.0, 'dcf', -1,
     'the same equity across more shares must lower the value per share'),
    ('Treasury shares held', 'B', +5.0, 'dcf', +1,
     'treasury shares are NOT outstanding, so buying more raises the value of each remaining share'),
    ('FY2025 dividend declared', 'B', +500.0, 'cash30', -1,
     'a dividend declared out of FY2025 was PAID before the 30-June-2026 balance sheet the '
     'bridge now stands on, so it is already inside that cash figure and must not be '
     'deducted a second time. What it still moves is the cash roll-forward to FY2030'),
    # ---- THE AUDITED HISTORY ----------------------------------------------------
    ('FY2025 sales (net)', 'B', +500.0, 'gp25', +1, 'more audited revenue is more gross profit'),
    ('FY2025 cost of sales', 'B', +500.0, 'gp25', -1, 'more cost of sales is less gross profit'),
    ('FY2025 general and administrative expenses', 'B', +100.0, 'ebit25', -1,
     'more administrative expense is less operating profit'),
    ('FY2025 provisions charged', 'B', +50.0, 'ebit25', -1, 'as are more provisions'),
    ('FY2025 expected credit losses', 'B', +10.0, 'ebit25', -1, 'and more credit losses'),
    ('FY2025 depreciation and amortisation', 'B', +50.0, 'ebitda_is25', +1,
     'EBITDA is operating profit PLUS depreciation, so a bigger audited charge raises it'),
    ('FY2025 profit before tax', 'B', +200.0, 'taxe25', -1,
     'the same audited tax charge on more pre-tax profit is a LOWER effective rate'),
    ('FY2025 income tax', 'B', +200.0, 'taxe25', +1, 'and a bigger charge is a higher rate'),
    ('FY2023 sales (net)', 'B', +500.0, 'gp23', +1, 'the same mechanism in FY2023'),
    ('FY2023 cost of sales', 'B', +200.0, 'gp23', -1, 'and on its cost line'),
    ('FY2024 sales (net)', 'B', +500.0, 'growth25', -1,
     'a bigger FY2024 base makes the FY2025 growth rate smaller'),
    ('FY2024 cost of sales', 'B', +200.0, 'roicb24', -1,
     'more FY2024 cost is less operating profit and a lower return on FY2024 capital'),
    ('FY2024 general and administrative expenses', 'B', +50.0, 'roicb24', -1, 'as is more overhead'),
    ('FY2024 provisions charged', 'B', +50.0, 'roicb24', -1, 'and more provisions'),
    ('FY2024 expected credit losses', 'B', +10.0, 'roicb24', -1, 'and more credit losses'),
    ('FY2023 general and administrative expenses', 'B', +50.0, 'rr23', -1,
     'less FY2023 operating profit is less NOPAT, so the same net reinvestment is a LARGER '
     'share of it — and FY2023 reinvestment was NEGATIVE, so a larger share is more negative'),
    ('FY2023 provisions charged', 'B', +20.0, 'rr23', -1, 'the same mechanism'),
    ('FY2023 expected credit losses', 'B', +10.0, 'rr23', -1, 'and again'),
    ('FY2023 profit before tax', 'B', +100.0, 'rr23', +1,
     'a higher FY2023 pre-tax profit against the same tax charge is a LOWER effective rate, '
     'so more NOPAT, so a smaller (less negative) reinvestment share'),
    ('FY2023 income tax', 'B', +50.0, 'rr23', -1, 'and the mirror of it'),
    ('FY2024 profit before tax', 'B', +100.0, 'roicb24', +1,
     'a lower effective rate on the same operating profit is more NOPAT on the same capital'),
    ('FY2024 income tax', 'B', +100.0, 'roicb24', -1, 'and the mirror'),
    ('FY2023 attributable profit', 'B', +100.0, 'close', 0,
     'the audited profit line does not touch the balance-sheet closure check'),
    ('FY2025 attributable profit', 'B', +200.0, 'roe25', +1,
     'more audited profit on the same audited equity is a higher return'),
    ('FY2025 earnings per share', 'B', +1.00, 'pe_spot', -1,
     'the same price on higher published earnings per share is a lower multiple'),
    ('FY2023 earnings per share', 'B', +0.50, 'pe_spot', 0,
     'the FY2023 figure does not touch the FY2025 multiple'),
    ('FY2024 earnings per share', 'B', +0.50, 'pe_spot', 0, 'nor does FY2024'),
    ('FY2023 capital expenditure', 'B', +100.0, 'rr23', +1,
     'more capital spending against the same depreciation is more net reinvestment'),
    ('FY2025 capital expenditure', 'B', +100.0, 'rr25', +1, 'the same in FY2025'),
    ('FY2024 capital expenditure', 'B', +100.0, 'roicb24', 0,
     'capital expenditure does not enter the return on capital, only the reinvestment rate'),
    ('FY2023 depreciation and amortisation', 'B', +50.0, 'rr23', -1,
     'more depreciation against the same capex is LESS net reinvestment'),
    ('FY2024 depreciation and amortisation', 'B', +50.0, 'rr23', 0,
     'the FY2024 charge does not touch the FY2023 reinvestment rate'),
    # ---- THE AUDITED BALANCE SHEET ---------------------------------------------
    ('Total assets FY2025', 'B', +500.0, 'close', +1,
     'assets less liabilities must exceed equity if assets rise alone — the closure check '
     'exists precisely to catch that'),
    ('Total liabilities FY2025', 'B', +500.0, 'close', -1, 'and the mirror'),
    ('Total liabilities FY2025', 'B', +500.0, 'othliab', +1,
     'other liabilities are total liabilities less interest-bearing debt'),
    ('Equity attributable to owners FY2025', 'B', +500.0, 'bvps25', +1,
     'more audited equity over the same shares is a higher book value per share'),
    ('Equity attributable to owners FY2024', 'B', +500.0, 'roicb24', -1,
     'more FY2024 capital against the same profit is a lower return on it'),
    ('Equity attributable to owners FY2023', 'B', +500.0, 'rr23', 0,
     'FY2023 equity does not enter the FY2023 reinvestment rate'),
    ('Property, plant and equipment FY2025', 'B', +500.0, 'bookval_t', +1,
     'a bigger audited property base is a higher book value per annual tonne'),
    ('Assets under construction FY2025', 'B', +200.0, 'bookval_t', +1, 'as is more construction'),
    ('Cash and bank balances FY2023', 'B', +100.0, 'nd23', -1,
     'more FY2023 cash is less FY2023 net debt'),
    ('Total interest-bearing debt FY2023', 'B', +100.0, 'nd23', +1, 'and more debt is more'),
    ('Total interest-bearing debt FY2024', 'B', +100.0, 'kd_eff25', -1,
     'a bigger opening balance raises the average the FY2025 interest is divided by'),
    ('FY2025 loan interest expense', 'B', +10.0, 'kd_eff25', +1,
     'more interest on the same average balance is a higher computed effective rate'),
    ('FY2025 credit-facility interest expense', 'B', +10.0, 'kd_eff25', +1, 'the same'),
    ('Q1-2026 interest-bearing debt', 'B', +200.0, 'q1nc', -1,
     'more debt at the quarter end is less net cash there'),
    ('Q1-2026 cash and bank balances', 'B', +200.0, 'q1nc', +1, 'and more cash is more'),
    ('Q1-2026 dividends payable', 'B', +200.0, 'q1nc', -1,
     'a declared and unpaid dividend is an obligation against that cash'),
    ('Q1-2026 sales', 'B', +200.0, 'runrate', +1,
     'a stronger first quarter raises the run rate the forecast is checked against'),
    ('Q1-2025 sales', 'B', +200.0, 'runrate', -1,
     'a stronger prior-year quarter lowers the growth rate, and so the run rate'),
    ('Q1-2026 gross profit', 'B', +100.0, 'q1mgn', +1, 'a higher first-quarter gross margin'),
    ('Q1-2026 attributable profit', 'B', +100.0, 'q1ann', +1,
     'the annualised first quarter is four times it'),
    ('Q1-2026 finance costs', 'B', +5.0, 'dcf', 0,
     'the Q1 finance cost is a cost-of-debt CHECK, not an input to the valuation'),
    ('FY2024 dividend approved and paid', 'B', +100.0, 'dcf', 0,
     'the FY2024 distribution is history and is deliberately consumed nowhere downstream'),
    # ---- REVENUE AND SHARE NOTES ------------------------------------------------
    ('FY2025 local sales of goods', 'B', +200.0, 'p_loc', +1,
     'more disclosed local revenue at the same price is more tonnes'),
    ('FY2025 export sales of goods', 'B', +200.0, 'p_ecem', +1,
     'the same on the export leg: revenue sets the price, the plant sets the tonnes'),
    ('FY2025 local services', 'B', +50.0, 'resid_rev', 0,
     'the services line does not enter the volume build; it is carried as a share'),
    ('FY2025 export services', 'B', +50.0, 'ebitda26', 0, 'nor does the export services line'),
    ('FY2024 total local sales', 'B', +200.0, 'dcf', 0,
     'the FY2024 revenue split is disclosure for the reader and drives nothing'),
    ('FY2024 total export sales', 'B', +200.0, 'dcf', 0, 'the same'),
    ('FY2025 administrative depreciation', 'B', +10.0, 'ccost25', -1,
     'administrative depreciation is removed from the cash overhead line, so more of it is '
     'LESS cash cost'),
    # ---- LENSES AND SECTOR ------------------------------------------------------
    ('Justified EV/EBITDA', 'B', +1.0, 'rel_lens', +1, 'a higher multiple lifts the lens'),
    ('Justified price/earnings', 'B', +1.0, 'norm_lens', +1, 'a higher multiple lifts the lens'),
    ('Justified enterprise value per annual tonne', 'B', +10.0, 'asset_lens', +1,
     'a higher value per tonne lifts it'),
    ('Mid-cycle EBITDA margin', 'B', +0.02, 'rel_lens', +1, 'a richer mid-cycle margin lifts it'),
    ('Normalised revenue haircut', 'B', +0.05, 'rel_lens', +1,
     'a smaller haircut leaves a bigger normalised base'),
    ('Replacement cost per annual tonne', 'B', +20.0, 'roic', -1,
     'more invested capital against the same terminal profit lowers the return on it'),
    # The lens weights were drivers until this edition and are drivers no longer:
    # the cash-flow lens IS the central. They are asserted ABSENT at the foot of
    # this file rather than tested for a direction they can no longer have.
    ('Effective tax rate', 'B', +0.02, 'dcf', -1, 'a higher tax rate lowers every NOPAT'),
    ('Spot price', 'B', +5.0, 'ev_per_t', +1,
     'a higher share price is a higher enterprise value over the same capacity'),
    ('USD/EGP at the valuation date', 'B', +5.0, 'asset_lens', +1,
     'capacity is valued in dollars per tonne, so a weaker pound raises its pound value'),
    ('Egyptian nameplate capacity', 'B', +5.0, 'sector_util', -1,
     'the same production over more capacity is lower utilisation'),
    ('Egyptian production 2025', 'B', +5.0, 'sector_util', +1, 'more production is higher utilisation'),
    ('Egyptian consumption 2025', 'B', +5.0, 'revival', -1,
     'the restart programme is a smaller share of a larger market'),
    ('Dormant capacity under revival', 'B', +2.0, 'revival', +1,
     'more dormant capacity is a larger share of the same market'),
    ('Egyptian exports 2025', 'B', +2.0, 'exports', +1, 'the disclosed export total'),
    ('Peer — Sinai Cement profit', 'B', +500.0, 'peer_pe', -1,
     'more peer profit at the same peer market value is a lower peer multiple'),
    ('Peer — Sinai Cement market capitalisation', 'B', +2000.0, 'peer_pe', +1,
     'a dearer peer at the same profit is a higher peer multiple'),
    ('Peer — Sinai Cement revenue', 'B', +500.0, 'peer_ps_scem', -1,
     'more peer revenue at the same market value is a lower price to sales'),
    ('Peer — Misr Beni Suef profit', 'B', +500.0, 'peer_pe', -1, 'the same mechanism'),
    ('Peer — Misr Beni Suef market capitalisation', 'B', +2000.0, 'peer_pe', +1, 'the same'),
    ('Peer — Misr Beni Suef revenue', 'B', +500.0, 'peer_ps_mbsc', -1, 'and the same'),
]

fails, rows = [], []
for label, col, bump, key, sign, why in CASES:
    r = row_of(label)
    cur = wb['Assumptions'][f'{col}{r}'].value
    out = read({('Assumptions', f'{col}{r}'): cur + bump})
    delta = out[key] - base[key]
    rel = delta / abs(base[key]) if base[key] else 0.0
    if sign == 0:
        ok = abs(delta) < 1e-9
    else:
        ok = (delta * sign > 0) and abs(rel) > 1e-9
    rows.append(dict(driver=label, col=col, bump=bump, headline=key, base=base[key],
                     bumped=out[key], rel=rel,
                     direction=('up' if sign > 0 else ('down' if sign < 0 else 'unchanged')),
                     passed=bool(ok), why=why))
    print(f"  [{'OK ' if ok else 'BAD'}] {label} [{col}] {bump:+g} -> {key} "
          f"{base[key]:,.3f} -> {out[key]:,.3f} ({rel:+.3%})")
    if not ok:
        fails.append((label, key, delta, why))

# ---- dead-input sweep --------------------------------------------------------
print('\nDEAD-INPUT SWEEP — every remaining driver is bumped and must move something')
dead = []
seen = {c[0] for c in CASES}
for label, r in sorted(A.items(), key=lambda kv: kv[1]):
    if label in seen:
        continue
    for col in ('B', 'C', 'D', 'E', 'F', 'G'):
        cell = wb['Assumptions'][f'{col}{r}']
        if not isinstance(cell.value, (int, float)):
            continue
        out = read({('Assumptions', f'{col}{r}'): cell.value * 1.10 + 1e-6})
        if all(abs(out[k] - base[k]) < 1e-9 for k in base):
            dead.append(f'{label} [{col}{r}]')
        break
print('  inputs that changed nothing:', dead if dead else 'none — every driver reprices')

json.dump(dict(base=base, cases=rows, dead=dead, n_cases=len(CASES), n_failed=len(fails)),
          open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1, default=float)

assert not fails, f'{len(fails)} drivers failed: {fails}'

# THE RETIRED WEIGHTS ARE ASSERTED GONE. [R-LENS-03] retired the typed blend; this
# checks the workbook actually stopped carrying the knobs, because a retirement
# that leaves the levers on the sheet is a rule nobody can see was applied — and a
# weight cell that a reader can type into and that moves nothing is worse than no
# cell at all.
_labels = {ws.cell(row=r, column=1).value
           for ws in wb for r in range(1, ws.max_row + 1)}
_left = sorted(l for l in _labels if isinstance(l, str) and l.startswith('Weight \u2014 '))
assert not _left, ('the retired lens weights are still live inputs in the workbook: %s'
                   % _left)
print('  [OK ] the four retired lens weights carry no live input cell')
assert not dead, f'dead inputs: {dead}'
print(f'\nDRIVER TEST OK — {len(CASES)} driver assertions, every one in the asserted '
      f'direction; 0 dead inputs')
