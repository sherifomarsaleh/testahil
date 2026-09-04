"""Does the delivered workbook actually reprice when a driver changes?

Recalculation proves the formulas reproduce the model as written. It does NOT prove the model
is wired: a workbook whose headline is a constant would pass it. This test perturbs each input
IN PLACE on the Assumptions sheet, re-evaluates the WHOLE workbook from scratch, and asserts
the headline value per share moves in the asserted DIRECTION.

A dead-input sweep runs over every remaining numeric input on the Assumptions sheet: each is
nudged and the headline must move at all. An input that moves nothing is either genuinely
inert or a wiring failure, and the test names it either way.

If a directional expectation fails, the first hypothesis is that the EXPECTATION is wrong, not
the model. Decompose the mechanism before changing anything.
"""
import json, os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import openpyxl
import xlcalc

XLSX = os.path.join(HERE, 'EIPICO_Valuation_Model_09082026.xlsx')
wb = openpyxl.load_workbook(XLSX)
wa = wb['Assumptions']


def row_of(text):
    for row in wa.iter_rows(min_col=1, max_col=1):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().startswith(text):
                return c.row
    raise KeyError(text)


HEADLINE = ('SOTP Bridge', None)
for row in wb['SOTP Bridge'].iter_rows(min_col=1, max_col=1):
    for cc in row:
        if isinstance(cc.value, str) and cc.value.startswith('Value per share — Frame A'):
            HEADLINE = ('SOTP Bridge', f'C{cc.row}')


def headline(overrides=None):
    return float(xlcalc.Book(wb, overrides).cell_value(*HEADLINE))


HEADLINE_NAME = 'value per share, Frame A'
BASE = headline()
print(f'base headline value per share: {BASE:.4f}\n')

# (label, cell on Assumptions, multiplier, expected direction, why)
CASES = [
    # RE-DERIVED, NOT DELETED, AND THE RE-DERIVATION IS A FINDING. This asserted that
    # more domestic packs raise the VALUE, on the reasoning that revenue rises faster
    # than the cost base. The gross margin does behave that way — but at this company's
    # DISCLOSED working-capital day ratios, in an economy running at the house inflation
    # ladder, an incremental pound of domestic revenue absorbs roughly half of itself in
    # receivables and inventory in the year it arrives, against a gross margin near 37%.
    # Incremental domestic volume is therefore CASH-NEGATIVE in the near years and only
    # turns later, and at these discount rates the near years dominate. So the claim the
    # model actually makes is about REVENUE, and it is asserted on revenue; the value
    # direction is a finding about the company rather than a law about volume, and it is
    # recorded in the gap review rather than asserted away.
    ('Domestic pack volume growth', f'C{row_of("Domestic pack volume growth")}', 1.20, +1,
     'more packs, more revenue', 'revenue'),
    ('Export pack volume growth', f'C{row_of("Export pack volume growth")}', 1.20, +1,
     'same, on the hard-currency book'),
    ('Domestic price per pack growth', f'C{row_of("Domestic price per pack growth")}', 1.20, +1,
     'price straight to the margin'),
    ('Export price per pack growth', f'C{row_of("Export price per pack growth")}', 1.50, +1,
     'price straight to the margin'),
    ('Exchange rate, year one', f'C{row_of("Exchange rate (EGP per USD")}', 1.10, -1,
     'a WEAKER pound raises export revenue but ALSO raises the imported ingredient and the '
     'imported SHARE of the packaging cost, 68.4% of the disclosed cost stack against a 32% '
     'export share — the cost side dominates, so the net effect is negative'),
    ('Selling and marketing / revenue', f'C{row_of("Selling and marketing / revenue")}', 1.20,
     -1, 'a bigger cost share'),
    ('Provision charge, Frame A', f'C{row_of("FRAME A")}', 1.30, -1, 'a bigger cost share'),
    ('Capital expenditure / revenue', f'C{row_of("Capital expenditure / revenue")}', 1.30, -1,
     'cash out of free cash flow'),
    ('Depreciation rate', f'C{row_of("Depreciation rate on the property")}', 1.30, -1,
     'a bigger tax-effected charge; the add-back is not a full offset because the tax shield '
     'is worth less than the charge'),
    ('Inventory days', f'C{row_of("Inventory days")}', 1.20, -1, 'more cash tied up'),
    ('Receivable days', f'C{row_of("Receivable days")}', 1.20, -1, 'more cash tied up'),
    ('Payable days', f'C{row_of("Payable days")}', 1.20, +1, 'less cash tied up'),
    ('Ten-year government yield', f'C{row_of("Ten-year local-currency government yield")}',
     1.10, -1, 'a higher discount rate'),
    ('Sovereign swap spread', f'C{row_of("Sovereign credit-default-swap spread")}', 1.50, +1,
     'a LARGER spread is subtracted from the quoted yield, so the normalised risk-free rate '
     'FALLS and the discount rate with it — this is the whole point of the normalisation'),
    ('Country equity risk premium', f'C{row_of("Country equity risk premium (swap")}', 1.20,
     -1, 'a higher cost of equity'),
    ('Beta', f'C{row_of("Beta")}', 1.30, -1, 'a higher cost of equity'),
    ('Cost of local-currency debt', f'C{row_of("Cost of local-currency debt")}', 1.20, -1,
     'a higher cost of capital'),
    ('Terminal risk-free rate', f'C{row_of("Terminal risk-free rate")}', 1.20, -1,
     'a higher terminal discount rate against a 74% terminal weight'),
    ('Terminal growth', f'C{row_of("Terminal growth")}', 1.20, +1,
     'growth in the terminal numerator and a smaller denominator'),
    # Terminal return on invested capital is no longer an input — it is COMPUTED on the DCF
    # sheet from the model's own FY2030E rows, so there is no cell left to perturb. It is
    # exercised indirectly by every driver that moves FY2030E operating profit or invested
    # capital, and directly by the recalculation gate.
    ('Terminal debt weight', f'C{row_of("Terminal debt weight — DERIVED")}', 1.30, +1,
     'more weight on the cheaper after-tax debt lowers the terminal discount rate'),
    ('Normalised associate contribution',
     f'C{row_of("Normalised associate contribution")}', 1.30, +1,
     'a bigger addition in the bridge'),
    ('Associate earnings multiple', f'C{row_of("Associate earnings multiple")}', 1.30, +1,
     'a bigger addition in the bridge'),
    # The STATUTORY rate no longer taxes free cash flow — the cash-flow engine applies the
    # EFFECTIVE rate. All the statutory rate now does is size the interest tax shield inside
    # the after-tax cost of debt, so raising it LOWERS the discount rate and RAISES the value.
    # The earlier expectation (down) was written when this row taxed the cash flows; it was
    # the expectation that went stale, not the model.
    ('Tax rate, statutory — the interest shield only',
     f'C{row_of("Corporate income tax rate")}', 1.20, +1,
     'the statutory rate sizes the interest tax shield in the after-tax cost of debt and '
     'nothing else, so a higher rate lowers the discount rate'),
    ('Tax rate, effective — what the cash flows bear',
     f'C{row_of("Effective tax rate on forecast")}', 1.20, -1,
     'the rate the free-cash-flow engine actually applies: more tax, less cash'),
    # DECOMPOSED before the expectation was changed. Bringing construction into service
    # earlier does raise the explicit-year depreciation charge, but depreciation is added back
    # in free cash flow, so all that reaches the cash flow is the tax shield: +22.5 on the sum
    # of present values. And it shrinks the balance still parked at FY2030E, which cuts the
    # terminal depreciation catch-up from 182.7 to 84.2 and lifts the present value of the
    # terminal block by +21.5 net of the lower FY2030E operating profit. Both legs point the
    # same way. The old expectation (down) was written when the terminal block did not charge
    # the parked balance at all; it was the EXPECTATION that went stale, not the model.
    ('Transfers out of construction', f'C{row_of("Transfers out of construction")}', 1.30, +1,
     'the explicit-year tax shield on higher depreciation, plus a smaller never-depreciated '
     'balance left parked at FY2030E for the terminal block to charge'),
    ('Gross borrowings', f'C{row_of("Gross borrowings including leases")}', 1.20, -1,
     'more debt to subtract in the bridge'),
    ('Cash', f'C{row_of("Cash and bank balances")}', 1.50, +1, 'less net debt to subtract'),
    # RE-POINTED. The centre is now the cash-flow lens alone, and a FIRM-level free cash
    # flow does not consume the finance charge — it reaches equity through net debt. The
    # driver was asserted against the centre because the centre used to be a BLEND that
    # included two equity-level lenses. It still moves what it is for, and that is what
    # is now asserted.
    ('Finance cost charged to profit',
     f'C{row_of("Finance cost charged to profit")}', 1.20, -1,
     'lower attributable earnings feed the relative and normalised lenses; the free-cash-flow '
     'lens is unaffected because free cash flow to the firm is struck before financing, so the '
     'Frame A headline itself should NOT move — this case therefore asserts on the weighted '
     'central value, not on the bridge'),
    ('Active-ingredient company at cost',
     f'C{row_of("Active-ingredient company at carrying cost")}', 1.30, +1,
     'a bigger addition in the bridge'),
    ('Non-controlling interests in the bridge',
     f'C{row_of("Non-controlling interests deducted in the bridge")}', 2.00, -1,
     'a bigger deduction in the bridge'),
]

results, failures = [], []
CENTRAL = None
for row in wb['Fundamental Valuation'].iter_rows(min_col=1, max_col=1):
    for cc in row:
        if isinstance(cc.value, str) and cc.value.startswith('CENTRE — FRAME A'):
            CENTRAL = ('Fundamental Valuation', f'B{cc.row}')
assert CENTRAL is not None, (
    "no row on the Fundamental Valuation sheet is labelled 'CENTRE — FRAME A'. This "
    "file located the centre by its LABEL and the label changed when the weighted blend "
    "was retired; the failure that produced was a TypeError deep in a call rather than "
    "a message naming the cause, which is why the absence is now asserted here.")
CENTRAL_BASE = float(xlcalc.Book(wb).cell_value(*CENTRAL))

REV_CELL = None
for cc_row in wb['Segments'].iter_rows(min_col=1, max_col=1):
    for cc in cc_row:
        if isinstance(cc.value, str) and cc.value.startswith('REVENUE, consolidated'):
            REV_CELL = ('Segments', f'F{cc.row}')          # the last forecast year
assert REV_CELL is not None, ('no row on the Segments sheet labelled "REVENUE, '
                              'consolidated" to assert a volume driver against; a target '
                              'named by label needs the label to exist')
REV_BASE = float(xlcalc.Book(wb).cell_value(*REV_CELL))

NORM_VAL_ROW = None
for cc_row in wb['Relative & Normalized'].iter_rows(min_col=1, max_col=1):
    for cc in cc_row:
        if isinstance(cc.value, str) and cc.value.startswith('NORMALISED EARNINGS POWER'):
            NORM_VAL_ROW = cc.row
assert NORM_VAL_ROW is not None, ('no row labelled "NORMALISED EARNINGS POWER" on the '
                                  'Relative & Normalized sheet')
NORM_BASE = float(xlcalc.Book(wb).cell_value('Relative & Normalized', f'B{NORM_VAL_ROW}'))

for case in CASES:
    label, cell, mult, want, why = case[:5]
    target = case[5] if len(case) > 5 else 'value'
    base_v = wa[cell].value
    if target == 'revenue':
        got = float(xlcalc.Book(wb, {('Assumptions', cell): base_v * mult})
                    .cell_value(*REV_CELL))
        move = got - REV_BASE
        ok = move > 1e-6
        results.append(dict(driver=label, cell=cell, multiplier=mult,
                            base_input=float(base_v), headline=got, move=move,
                            expected='up (final-year revenue)', passed=bool(ok),
                            mechanism=why))
        print(f"{'PASS' if ok else 'FAIL'}  {label:38s} x{mult:<5.2f} "
              f"{REV_BASE:8.0f} -> {got:8.0f}  ({move:+8.0f})  expected up [revenue]")
        if not ok:
            failures.append(label)
        continue
    if label == 'Finance cost charged to profit':
        # RE-POINTED at a lens that actually consumes it. A firm-level free cash flow
        # does not: the finance charge reaches equity through net debt, not through the
        # enterprise value. It used to move the centre only because the centre was a
        # BLEND carrying two equity-level lenses at 35% between them, so retiring the
        # blend made this driver inert on the centre while leaving it live where it
        # belongs. Asserting it against the centre now would be asserting a mechanism
        # this model does not have.
        # THE ROW SPANS FIVE FORECAST YEARS and the case names one column. Bumping C
        # alone moves FY2026 while the normalised lens is built on FY2027 earnings, so
        # the driver read as dead when it was simply not the year being looked at. Every
        # year of the row is bumped, which is what "the finance cost rises" means.
        _row = cell[1:]
        _ov = {('Assumptions', f'{col}{_row}'): wa[f'{col}{_row}'].value * mult
               for col in ('C', 'D', 'E', 'F', 'G')
               if isinstance(wa[f'{col}{_row}'].value, (int, float))}
        got = float(xlcalc.Book(wb, _ov)
                    .cell_value('Relative & Normalized', f'B{NORM_VAL_ROW}'))
        move = got - NORM_BASE
        ok = move < -1e-6
        results.append(dict(driver=label, cell=cell, multiplier=mult,
                            base_input=float(base_v), headline=got, move=move,
                            expected='down (the normalised earnings figure)',
                            passed=bool(ok), mechanism=why))
        print(f"{'PASS' if ok else 'FAIL'}  {label:38s} x{mult:<5.2f} "
              f"{NORM_BASE:8.3f} -> {got:8.3f}  ({move:+8.3f})  expected down "
              f"[normalised earnings, the lens that consumes it]")
        if not ok:
            failures.append(label)
        continue
    got = headline({('Assumptions', cell): base_v * mult})
    move = got - BASE
    ok = (move > 1e-6) if want > 0 else (move < -1e-6)
    results.append(dict(driver=label, cell=cell, multiplier=mult, base_input=float(base_v),
                        headline=got, move=move, expected='up' if want > 0 else 'down',
                        passed=bool(ok), mechanism=why))
    print(f"{'PASS' if ok else 'FAIL'}  {label:38s} x{mult:<5.2f} "
          f"{BASE:8.3f} -> {got:8.3f}  ({move:+8.3f})  expected "
          f"{'up' if want > 0 else 'down'}")
    if not ok:
        failures.append(label)
        print(f'       mechanism asserted: {why}')

# ---- dead-input sweep over everything else -----------------------------------
# An input is not "dead" merely because it leaves the Frame A headline alone. The model
# has several headlines and each input belongs to one of them: the effective tax rate
# drives reported earnings, not the free-cash-flow discount; the Frame B provision row
# drives the Frame B value BY CONSTRUCTION and must not touch Frame A; the rating-basis
# spread and premium drive the published alternative cost of equity. So the sweep tests
# each input against the FULL headline set and only calls it dead if it moves none of them.
def cellref(sheet, startswith, col='B'):
    for row in wb[sheet].iter_rows(min_col=1, max_col=1):
        for cc in row:
            if isinstance(cc.value, str) and cc.value.strip().startswith(startswith):
                return (sheet, f'{col}{cc.row}')
    raise KeyError(f'{sheet}: {startswith}')


def rowref(sheet, startswith, col):
    for row in wb[sheet].iter_rows(min_col=1, max_col=1):
        for cc in row:
            if isinstance(cc.value, str) and cc.value.strip().startswith(startswith):
                return (sheet, f'{col}{cc.row}')
    raise KeyError(f'{sheet}: {startswith}')


HEADLINES = {
    'value per share, Frame A': HEADLINE,
    'value per share, Frame B': cellref('DCF', 'VALUE PER SHARE ON FRAME B'),
    'weighted centre, Frame A': cellref('Fundamental Valuation', 'CENTRE — FRAME A'),
    'weighted centre, Frame B': cellref('Fundamental Valuation', 'CENTRE — FRAME B'),
    'book value and sustainable return lens':
        cellref('Fundamental Valuation', 'Book value and sustainable'),
    'relative multiples lens': cellref('Fundamental Valuation', 'Relative multiples'),
    'FY2030E gross borrowings': rowref('Balance Sheet', 'Gross borrowings', 'I'),
    'FY2026E earnings per share': rowref('Income Statement', 'Earnings per share', 'E'),
    'FY2030E book value per share': rowref('Balance Sheet', 'Book value per share', 'I'),
    'FY2030E total assets': rowref('Balance Sheet', 'TOTAL ASSETS', 'I'),
    'FY2030E return on invested capital': rowref('Per-Share & Ratios',
                                                 'Return on invested', 'I'),
    'cost of equity, rating basis': ('Assumptions',
                                     f"C{row_of('Cost of equity, rating basis')}"),
}
BASES = {k: float(xlcalc.Book(wb).cell_value(*v)) for k, v in HEADLINES.items()}
print('\nheadline set used by the sweep:')
for k, v in BASES.items():
    print(f'   {k:38s} {v:14,.4f}')

print('\ndead-input sweep over every remaining numeric input on Assumptions:')
tested = {c for _, c, *_ in CASES}
dead, live, disclosure = [], [], []
for row in wa.iter_rows(min_col=3, max_col=7):
    for cell in row:
        if cell.coordinate in tested or not isinstance(cell.value, (int, float)):
            continue
        basis = wa[f'B{cell.row}'].value
        if isinstance(basis, str) and basis.strip() == 'calculated':
            continue                                   # a derived row, not an input
        if isinstance(basis, str) and basis.startswith('disclosure'):
            disclosure.append((wa[f'A{cell.row}'].value, cell.coordinate))
            continue
        base_v = cell.value
        if base_v == 0:
            continue
        ov = {('Assumptions', cell.coordinate): base_v * 1.15}
        bk = xlcalc.Book(wb, ov)
        moved = [k for k, v in HEADLINES.items()
                 if abs(float(bk.cell_value(*v)) - BASES[k]) > 1e-9]
        if moved:
            live.append((wa[f'A{cell.row}'].value, cell.coordinate, moved))
        else:
            dead.append((wa[f'A{cell.row}'].value, cell.coordinate))
print(f'  {len(live)} further inputs move at least one headline; {len(dead)} move none; '
      f'{len(disclosure)} are disclosure rows, excluded by design')
byhead = {}
for name, coord, moved in live:
    if HEADLINE_NAME not in moved:
        byhead.setdefault(', '.join(moved), []).append(f'{coord} {name}')
if byhead:
    print('  inputs that leave the Frame A headline alone, and what they DO move:')
    for heads, items in sorted(byhead.items()):
        print(f'    -> {heads}')
        for it in items:
            print(f'         {it}')
for name, coord in disclosure:
    print(f'  DISCLOSURE (not a driver)  {coord}  {name}')
for name, coord in dead:
    print(f'  DEAD  {coord}  {name}')

json.dump(dict(base=BASE, headline_bases=BASES, cases=results, failures=failures,
               dead_inputs=[[n, c] for n, c in dead],
               disclosure_rows=[[n, c] for n, c in disclosure],
               live_inputs=[[n, c, m] for n, c, m in live]),
          open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1)
assert not failures, f'{len(failures)} driver(s) did not move the headline as asserted: {failures}'
assert not dead, f'{len(dead)} dead input(s): {dead}'
print(f'\nDRIVER TEST PASSED — {len(CASES)} asserted directions all correct, '
      f'{len(live)} further inputs live against at least one headline, 0 dead '
      f'({len(disclosure)} disclosure rows excluded by design).')
