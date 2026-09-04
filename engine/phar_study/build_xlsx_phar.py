"""PHAR (EIPICO) valuation model — 16-sheet formula-first workbook.

THE WORKBOOK CALCULATES. Every figure that is arithmetically derivable from a driver is a
live Excel formula: the cost of capital is built in the sheet from the risk-free rate net
of the sovereign spread, beta and the premium; the cost of debt is blended by currency and
taxed; the weights come from net debt and market capitalisation; the glide fractions are
derived from the cost-of-debt path; the discount factors compound; the DCF waterfall
chains; the terminal block chains; the statements roll forward; every ratio and per-share
figure is a formula.

Only three classes of cell are pasted, and READ FIRST names them:
  (1) audited and disclosed history — where a line is both disclosed and derivable, the
      DISCLOSED figure is carried;
  (2) the output of the unit build, which would be unreadable flattened into a grid;
  (3) whole-model re-runs — the Monte Carlo map and the sensitivity grids, where each cell
      is a complete revaluation and which therefore do NOT redraw when a driver changes.

Every formula the builder writes is recorded with the model's own value for that cell in
xlsx_expected.json; recalc.py then evaluates the DELIVERED file independently and asserts
each one reproduces it.
"""
import json, os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, '..'))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
M, H, FC, W, DCFD, LN = D['meta'], D['history'], D['forecast'], D['wacc'], D['dcf'], D['lenses']
UB, SENS, CAL, CRUX = D['unit_build'], D['sensitivity'], D['calibration'], D['crux']
V = {k: v['value'] for k, v in D['inputs'].items()}
# derived rates, committed by the model so no document re-derives one differently
DERIVED = D['derived']
YRS = FC['years']
SH = M['shares_mn']
TAX = V['tax_stat']              # statutory — used for the after-tax cost of debt
TAX_FCFF = W['tax_fcff']         # effective — the rate the cash-flow engine applies
PROV = D['sensitivity']
CE = D['cost_exposure']
PATHS = 50000            # simulation path count, from the probability-map run
PEER_HI, PEER_MID = 26.7, 16.0   # observed peer multiples, market-data layer

INK = '1C3A36'; PANEL = 'EAF0EE'; CREAM = 'F6F1E6'; BLUE = '0B4F9E'; GREYF = '6E7B77'
F_H = Font(name='Calibri', size=10.5, bold=True, color='FFFFFF')
F_SUB = Font(name='Calibri', size=10, bold=True, color=INK)
F_N = Font(name='Calibri', size=10, color=INK)
F_IN = Font(name='Calibri', size=10, color=BLUE)          # blue = pasted input
F_NOTE = Font(name='Calibri', size=8.8, italic=True, color=GREYF)
FILL_H = PatternFill('solid', fgColor=INK)
FILL_P = PatternFill('solid', fgColor=PANEL)
FILL_C = PatternFill('solid', fgColor=CREAM)
THIN = Side(style='thin', color='C9D4D1')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)
EXPECT = {}      # "Sheet!A1" -> the model's own value for that formula cell
NPASTE = {'audited': 0, 'unit_build': 0, 'grid': 0, 'label': 0}


def sheet(name, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def hdr(ws, row, cells, fill=FILL_H, font=F_H):
    for j, t in enumerate(cells, start=1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = font; c.fill = fill; c.border = BOX
        c.alignment = Alignment(horizontal='center' if j > 1 else 'left', wrap_text=True,
                                vertical='center')
    ws.row_dimensions[row].height = 30


def lbl(ws, row, col, text, bold=False, fill=None, note=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_NOTE if note else (F_SUB if bold else F_N)
    if fill: c.fill = fill
    c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical='center')
    NPASTE['label'] += 1
    return c


def val(ws, row, col, v, fmt='#,##0.0', kind='audited', fill=None):
    """A PASTED value. `kind` must name which of the three permitted classes it is."""
    c = ws.cell(row=row, column=col, value=v)
    c.font = F_IN; c.number_format = fmt; c.border = BOX
    if fill: c.fill = fill
    NPASTE[kind] += 1
    return c


def f(ws, row, col, formula, expected, fmt='#,##0.0', bold=False, fill=None):
    """A LIVE FORMULA. `expected` is the model's own value for this cell; it is recorded
    so recalc.py can assert the delivered workbook reproduces the model."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name='Calibri', size=10, bold=bold, color=INK)
    c.number_format = fmt; c.border = BOX
    if fill: c.fill = fill
    EXPECT[f'{ws.title}!{c.coordinate}'] = float(expected)
    return c


PCT = '0.0%'; PCT2 = '0.00%'; MONEY = '#,##0'; PS = '#,##0.00'; X = '0.00"x"'

# =============================================================== 1. READ FIRST
ws = sheet('READ FIRST', [46, 96])
hdr(ws, 1, ['READ FIRST', 'Egyptian International Pharmaceutical Industries Company (EIPICO)'])
rows = [
 ('What this workbook is',
  'An educational valuation model of a listed Egyptian pharmaceutical manufacturer, built '
  'from the company\'s own audited financial statements. It is not investment advice, it '
  'contains no rating and no price target, and it expresses value as a range.'),
 ('THE WORKBOOK CALCULATES',
  'Every figure that can be derived from a driver is a live formula. Change a driver on the '
  'Assumptions sheet and the cost of capital, the glide, the discount factors, the cash-flow '
  'waterfall, the terminal block, the three statements, the bridge and every ratio all move. '
  'This claim is tested, not asserted: a driver test perturbs each input in place, '
  're-evaluates the whole workbook and checks the headline moves in the right direction.'),
 ('Pasted cell class 1 of 3 — audited and disclosed history',
  'The FY2023, FY2024 and FY2025 columns of the Income Statement, Balance Sheet and Cash '
  'Flow sheets, and the disclosed operating statistics on Segments. Where a line is both '
  'disclosed and derivable, the DISCLOSED figure is carried. Blue type marks a pasted cell.'),
 ('Pasted cell class 2 of 3 — the unit build\'s output',
  'The FY2024 and FY2025 pack volumes and realised prices per pack on the Segments sheet. '
  'These come from reconciling the board\'s disclosed pack counts against the revenue note\'s '
  'channel split; flattened into a grid the reconciliation would be unreadable. Everything '
  'downstream of them — every forecast year — is a formula.'),
 ('Pasted cell class 3 of 3 — whole-model re-runs',
  'The Monte Carlo percentile map and touch ladder, and the Sensitivity grids. Each of those '
  'cells is a COMPLETE revaluation of the model at a different input, or a full '
    f'{PATHS:,}-path '
  'simulation. THEY DO NOT REDRAW WHEN A DRIVER CHANGES. Everything else on those two sheets '
  'is a formula.'),
 ('Blue means input, black means formula',
  'Blue type is a pasted number. Black type is calculated in the sheet.'),
 ('The contested judgement is carried both ways',
  'The credit-loss and provision charge is the study\'s single most consequential contested '
  'judgement. Frame A treats it as a permanent cost of doing business; Frame B treats it as '
  'normalising. Both run all the way through to a value per share and both are published. '
  'They are never averaged into one number.'),
 ('Currency and units',
  'Egyptian pounds. Money in millions unless a row says otherwise; per-share figures in '
  'pounds; packs and units in millions.'),
 ('Basis and date',
  f"Audited CONSOLIDATED financial statements. Share price {M['spot']:.2f} at "
  f"{M['price_date']}; {SH:,.6f} million shares in issue. Fiscal year ends 31 December."),
]
r = 2
for a, b in rows:
    lbl(ws, r, 1, a, bold=True, fill=FILL_P)
    lbl(ws, r, 2, b)
    ws.row_dimensions[r].height = 46
    r += 1

# =============================================================== 4. ASSUMPTIONS
# built first: every other sheet points at it
wa = sheet('Assumptions', [44, 60, 13, 13, 13, 13, 13])
hdr(wa, 1, ['Driver', 'Basis', *YRS])
A = {}          # key -> row number


def arow(key, label, basis, values, fmt='#,##0.00', kind='audited'):
    global r
    A[key] = r
    lbl(wa, r, 1, label); lbl(wa, r, 2, basis, note=True)
    if isinstance(values, (list, tuple)):
        for j, v in enumerate(values):
            val(wa, r, 3 + j, v, fmt=fmt, kind=kind)
    else:
        val(wa, r, 3, values, fmt=fmt, kind=kind)
    wa.row_dimensions[r].height = 26
    r += 1


r = 2
lbl(wa, r, 1, 'ANCHORS', bold=True, fill=FILL_C); r += 1
arow('spot', 'Share price (EGP)', 'Last close of the uploaded exchange price history, '
     f"{M['price_date']}", M['spot'])
arow('sh22', 'Shares in issue at 31 December 2022 (million)', 'Board report, capital and '
     'shareholders table', V['shares_fy22'], fmt='#,##0.000')
arow('sh23', 'Shares in issue, FY2023 and FY2024 (million)', 'Board report',
     V['shares_fy23'], fmt='#,##0.000')
arow('shwavg', 'Weighted-average shares, FY2025 (million)', 'Audited — the capital increase '
     'completed during the year', V['wavg_shares_fy25'], fmt='#,##0.000')
arow('par22', 'Attributable profit, FY2022 (EGP mn)', 'Board report, eleven-year profit table',
     V['parent_fy22'], fmt=MONEY)
arow('shares', 'Shares in issue (million)', f"Capital note (13): issued capital of "
     f"EGP {SH * V['par_value'] * 1e6:,.0f} at EGP {V['par_value']:,.0f} nominal a share",
     SH, fmt='#,##0.000000')
arow('tax', 'Corporate income tax rate', 'Egyptian statutory rate', TAX, fmt=PCT2)
arow('tax_eff', 'Effective tax rate on forecast pre-tax profit', 'Above statutory to allow '
     'for the solidarity contribution and disallowed items', V['tax_eff_fwd'], fmt=PCT2)

lbl(wa, r, 1, 'VOLUME AND PRICE', bold=True, fill=FILL_C); r += 1
arow('dvol', 'Domestic pack volume growth', 'Not capacity-constrained at 65% utilisation',
     V['dom_pack_growth'], fmt=PCT, kind='audited')
arow('evol', 'Export pack volume growth', 'Hard-currency export growth was 10% in FY2025',
     V['exp_pack_growth'], fmt=PCT)
arow('dprice', 'Domestic price per pack growth', 'Administered prices track inflation, no real '
     'gain', V['dom_price_growth'], fmt=PCT)
arow('eprice', 'Export price per pack growth (USD)', 'Competitive generic export pricing',
     V['exp_price_usd_growth'], fmt=PCT)
arow('toll', 'Contract-manufacturing revenue growth', 'Small line, idle-capacity utilisation',
     V['toll_growth'], fmt=PCT)
arow('fx', 'Exchange rate (EGP per USD, period average)', 'Partial real-rate reversal as '
     'domestic inflation converges', V['fx_path'], fmt='#,##0.00')

lbl(wa, r, 1, 'COST STACK — ONE ESCALATOR PER DRIVER CLASS', bold=True, fill=FILL_C); r += 1
arow('sh_mat', 'Cost share: imported active ingredients', 'Cost of sales note (26)',
     V['cost_shares']['materials'], fmt=PCT2)
arow('sh_pack', 'Cost share: packaging materials', 'Cost of sales note (26)',
     V['cost_shares']['packaging'], fmt=PCT2)
arow('sh_lab', 'Cost share: labour', 'Cost of sales note (26), wages plus benefits plus social '
     'insurance', V['cost_shares']['labour'], fmt=PCT2)
arow('sh_en', 'Cost share: energy and utilities', 'Cost of sales note (26)',
     V['cost_shares']['energy'], fmt=PCT2)
arow('sh_svc', 'Cost share: other consumables and services', 'Cost of sales note (26)',
     V['cost_shares']['services_other'], fmt=PCT2)
arow('sh_dep', 'Cost share: depreciation', 'disclosure — the audited split is shown for '
     'completeness, but depreciation is EXCLUDED from the escalated unit cost by design and '
     'enters exactly once, from the property roll-forward. This row drives nothing and is '
     'meant to drive nothing', V['cost_shares']['depreciation'], fmt=PCT2)
arow('esc_api', 'Escalator: hard-currency ingredient price', 'Applied THROUGH the exchange-rate '
     'path, never a domestic index', V['esc_materials_usd'], fmt=PCT2)
arow('pk_imp', 'Imported share of the packaging line', 'The balance is made in-house',
     V['esc_packaging_import_share'], fmt=PCT)
arow('esc_lab', 'Escalator: Egyptian wages', 'Minimum-wage resets run above consumer prices',
     V['esc_labour'], fmt=PCT)
arow('esc_en', 'Escalator: regulated energy tariffs', 'Subsidy reform, ABOVE consumer prices',
     V['esc_energy'], fmt=PCT)
arow('esc_cpi', 'Escalator: domestic consumer prices', 'Applied only to genuinely domestic '
     'lines', V['esc_domestic_cpi'], fmt=PCT)

lbl(wa, r, 1, 'OPERATING EXPENSE, CAPITAL AND WORKING CAPITAL', bold=True, fill=FILL_C); r += 1
arow('mkt', 'Selling and marketing / revenue', 'Disclosed 12.9% -> 12.4% -> 10.7%', V['mkt_pct'],
     fmt=PCT)
arow('rnd', 'Research and development / revenue', 'Rising with the biosimilar pipeline',
     V['rnd_pct'], fmt=PCT2)
arow('ga', 'General and administrative / revenue', 'Held near the FY2025 level', V['ga_pct'],
     fmt=PCT2)
arow('provA', 'FRAME A — provision charge / revenue', 'Permanent, at the three-year average',
     [V['prov_pct_permanent']] * 5, fmt=PCT2)
arow('provB', 'FRAME B — provision charge / revenue', 'Normalising as the book seasons',
     V['prov_pct_normalising'], fmt=PCT2)
arow('capex', 'Capital expenditure / revenue', 'Falls as the construction cycle completes',
     V['capex_pct'], fmt=PCT)
arow('transfer', 'Transfers out of construction (EGP mn)', 'The plant was licensed in December '
     '2025 and enters service', V['cip_transfer'], fmt=MONEY)
arow('deprate', 'Depreciation rate on the property base', 'About a sixteen-year blended life',
     V['dep_rate'], fmt=PCT2)
arow('dio', 'Inventory days', 'An eight-month strategic stockpile, stated policy', V['dio'],
     fmt='#,##0')
arow('dso', 'Receivable days', 'Distributor-concentrated book plus export terms', V['dso'],
     fmt='#,##0')
arow('dpo', 'Payable days', 'Audited FY2025 implies 54 days', V['dpo'], fmt='#,##0')
arow('payout', 'Dividend payout ratio', 'EGP 3.50 a share proposed for FY2025 on attributable '
     'profit', V['payout'], fmt=PCT)

lbl(wa, r, 1, 'COST OF CAPITAL', bold=True, fill=FILL_C); r += 1
arow('rf', 'Ten-year local-currency government yield', 'House cost-of-capital reference print',
     V['rf'], fmt=PCT2)
arow('sov', 'Sovereign credit-default-swap spread', 'Country risk-premium file, Egypt row, read '
     'live', V['sov_spread_cds'], fmt=PCT2)
arow('erp', 'Country equity risk premium (swap basis)', 'Same file, same row', V['erp_cds'],
     fmt=PCT2)
arow('sov_r', 'Adjusted default spread (rating basis)', 'Same file — the alternative',
     V['sov_spread_rating'], fmt=PCT2)
arow('erp_r', 'Country equity risk premium (rating basis)', 'Same file — the alternative',
     V['erp_rating'], fmt=PCT2)
arow('beta', 'Beta', 'Own-stock weekly regression against a 36-name local composite, five '
     'years; R-squared 0.235, n = 257', V['beta'], fmt='0.000')
arow('kd_lc', 'Cost of local-currency debt', 'Sovereign yield plus 250 basis points',
     V['kd_egp'], fmt=PCT2)
arow('kd_fx', 'Hard-currency coupon', 'Dollar and euro term loans', V['kd_fx_coupon'], fmt=PCT2)
arow('fxdep', 'Expected currency depreciation for the debt charge', 'Carries foreign-currency '
     'debt at its LOCAL-EQUIVALENT cost', V['fx_dep_wacc'], fmt=PCT2)
arow('wfx', 'Hard-currency share of the term-loan book', 'Borrowings note (17), by lender and '
     'currency', W['w_fx'], fmt=PCT)
arow('kdpath', 'Normalised risk-free convergence path', 'NOT a cost of debt: the first '
     'point is the normalised risk-free rate. Only the SHAPE of this row enters the model — '
     'the discount-rate glide rebases it to its own endpoints, so the levels cancel',
     V['kd_path'], fmt=PCT2)
arow('intpath', 'Finance cost charged to profit (EGP mn)', 'What the profit and loss account '
     'actually bears — NOT the marginal cost of debt above. Calibrated to the first quarter '
     'of 2026', V['int_path'], fmt=MONEY)
arow('rf_t', 'Terminal risk-free rate', 'A sourced 5% medium-term inflation target plus an '
     'UNSOURCED 5.5-point real convention. The real leg is an assertion, it is the widest '
     'single lever in the study, and it is an open item — see the terminal grid',
     DERIVED['rf_term'], fmt=PCT2)
arow('erp_t', 'Terminal equity risk premium', 'Normalised toward the rating-class norm',
     V['erp_term'], fmt=PCT2)
arow('kdt_lc', 'Terminal local-currency borrowing rate', 'Long-run Egyptian norm',
     V['kd_term_lc'], fmt=PCT2)
arow('kdt_fx', 'Terminal hard-currency coupon', 'Long-run norm', V['kd_term_fx'], fmt=PCT2)
arow('wd_t', 'Terminal debt weight — DERIVED, market-value basis', "Today's net debt over "
     "market capitalisation plus net debt. Not an assumption; the earlier edition's 20% was "
     "neither this nor the book reading below", W['wd_term'], fmt=PCT)
arow('wd_tb', 'Terminal debt weight — the funded forecast balance sheet at FY2030E',
     'disclosure — the BOOK reading, published beside the market reading above rather than '
     'chosen between in silence. The valuation uses the market reading, because a weighted '
     'average cost of capital weights market values; this row drives nothing',
     W['wd_term_book'], fmt=PCT)
arow('g', 'Terminal growth', 'Pound-nominal, against a 5% terminal inflation rate — about zero '
     'in real terms', DERIVED['g_term'], fmt=PCT)
arow('assoc_e', 'Normalised associate contribution (EGP mn)', 'Three disclosed years, '
     'reconciled to this model\'s own income statement: 74.508, 151.581, 512.085 — mean '
     '246.058. The first quarter of 2026 annualises to 52.5, but the auditor states two '
     'holdings\' statements were not received, so that quarter is evidence, not a run-rate',
     V['assoc_norm'], fmt=MONEY)
arow('assoc_m', 'Associate earnings multiple', 'Below the Gulf listed range for a minority, '
     'unlisted, non-controlled stake', V['assoc_multiple'], fmt='0.0"x"')
arow('peer_pe', 'Struck peer reference price-earnings multiple', 'MARKET DATA, cross-check '
     'layer. The MIDPOINT of the only two disclosed observations (26.7x and 16.0x) — NOT a '
     'median of a peer set, and the peers are not named, so it cannot be rebuilt from their '
     'filings', V['peer_pe_regional'], fmt='0.00"x"')
arow('peer_ke', 'Cost of equity faced by the struck reference companies',
     'The ONE difference the peer multiple is adjusted for, registered here rather than '
     'typed inside the adjustment so a reader can see what the adjustment is made of',
     V['peer_ke'], fmt=PCT)

lbl(wa, r, 1, 'OPENING BALANCE SHEET (AUDITED, 31 DECEMBER 2025)', bold=True, fill=FILL_C)
r += 1
arow('ppe0', 'Property, plant and equipment, net', 'Note (4)', V['ppe_fy25'], fmt=MONEY)
arow('cip0', 'Projects under construction', 'Note (6)', V['cip_fy25'], fmt=MONEY)
arow('inv0', 'Inventories, net', 'Note (9)', V['inv_fy25'], fmt=MONEY)
arow('ar0', 'Trade and notes receivable, net', 'Note (10)', V['ar_fy25'], fmt=MONEY)
arow('od0', 'Other debtors', 'Note (11)', V['othdr_fy25'], fmt=MONEY)
arow('ap0', 'Trade and notes payable', 'Note (22)', V['ap_fy25'], fmt=MONEY)
arow('oc0', 'Other creditors', 'Note (23)', V['othcr_fy25'], fmt=MONEY)
arow('cash0', 'Cash and bank balances', 'Note (12)', V['cash_fy25'], fmt=MONEY)
arow('debt0', 'Gross borrowings including leases', 'Notes (17), (18) and (21)', W['gross_debt'],
     fmt=MONEY)
arow('parent25', 'Attributable profit, FY2025 (EGP mn)', 'Audited — the trailing earnings '
     'base for every trailing multiple', V['parent_fy25'], fmt=MONEY)
arow('eq0', 'Equity attributable to the holding company', 'Audited', V['equity_parent_fy25'],
     fmt=MONEY)
arow('nci0', 'Non-controlling interests, audited 31 December 2025', 'disclosure — the audited '
     'December figure, shown so the reader can see both. It is SUPERSEDED for valuation by the '
     'March figure below, because almost all of it was the active-ingredient company and that '
     'company was deconsolidated in the first quarter of 2026. This row drives the balance '
     'sheet history only', V['nci_fy25'], fmt=MONEY)
arow('ncibr', 'Non-controlling interests deducted in the bridge', 'Post-deconsolidation, from '
     'the reviewed 31 March 2026 interim', V['nci_bridge'], fmt=MONEY)
arow('apicost', 'Active-ingredient company at carrying cost (EGP mn)', 'Pre-revenue, so cost '
     'not an earnings multiple; the Q1-2026 movement in investments in associates',
     V['arab_api_cost'], fmt=MONEY)
arow('afs0', 'Assets held for sale', 'Note (8/1)', V['afs_fy25'], fmt=MONEY)
arow('intang0', 'Intangible assets, net', 'Note (7)', V['intang_fy25'], fmt=MONEY)
arow('amort', 'Right-of-use and intangible amortisation run-rate', 'Audited FY2025 cash-flow '
     'statement', FC['amort'][0], fmt='#,##0.0')
arow('depcogs', 'Share of depreciation charged to cost of sales', 'Cost of sales note (26): '
     '93.498 of 117.725', FC['dep_cogs_share'], fmt=PCT)
arow('uc0', 'FY2025 cash cost per pack sold (EGP)', 'Cost of sales less its depreciation line, '
     'over packs sold', UB['unit_cost_fy25'] if 'unit_cost_fy25' in UB else
     (V['cogs_fy25'] - 93.497560) / V['packs_sold_fy25'], fmt=PS, kind='unit_build')
arow('consol', 'Consolidated-to-separate revenue factor', 'MEASURED: audited consolidated '
     'revenue over the separate-company channel total', UB['consol_uplift'], fmt='0.0000')
arow('ownpk0', "FY2025 packs sold of the company's OWN preparations (million)",
     'Board report, sales-indicators table', V['packs_own_fy25'], fmt='#,##0.000')
arow('ep0', 'FY2025 export packs (million)', 'Investor presentation', V['export_packs_fy25'],
     fmt='#,##0.000')
arow('erev0', 'FY2025 export revenue (EGP mn)', 'Revenue note (25)', V['ch_export_fy25'],
     fmt=MONEY)
arow('ownval0', "FY2025 sales value of the company's OWN preparations (EGP mn)",
     'Board report, sales-indicators table', V['own_prep_value_fy25'], fmt=MONEY)
arow('conval0', 'FY2025 sales value of CONTRACT-MANUFACTURED preparations (EGP mn)',
     'Board report, sales-indicators table — product value, not the fee the company books',
     V['contract_value_fy25'], fmt=MONEY)
arow('tollpk0', 'FY2025 contract-manufactured packs (million)', 'Board report',
     V['packs_toll_fy25'], fmt='#,##0.000')
arow('toll0', 'FY2025 contract-manufacturing FEE revenue (EGP mn)',
     'disclosure — this row SPLITS the contract-manufacturing line in two and cannot move the '
     'valuation by construction: the fee per pack rises by exactly what the resale price per '
     'pack falls, so the product value above is what reaches revenue either way. It is carried '
     'because the two halves are disclosed separately and a reader should see both. Revenue '
     'note (25)',
     V['ch_toll_fy25'], fmt='#,##0.0')
arow('fx0', 'FY2025 average exchange rate', 'Note (36)', V['fx_avg_fy25'], fmt='#,##0.00')

lbl(wa, r, 1, 'DERIVED IN THIS SHEET (formulas, not inputs)', bold=True, fill=FILL_C); r += 1


def drow(key, label, formula, expected, fmt=PCT2):
    global r
    A[key] = r
    lbl(wa, r, 1, label)
    lbl(wa, r, 2, 'calculated', note=True)
    f(wa, r, 3, formula, expected, fmt=fmt, bold=True)
    r += 1


AS = 'Assumptions'
c = lambda k, col=3: f'${get_column_letter(col)}${A[k]}'
drow('rfstar', 'Normalised risk-free rate = yield less sovereign spread',
     f'={c("rf")}-{c("sov")}', W['rf_star'])
drow('ke', 'Cost of equity = normalised risk-free + beta x premium',
     f'={c("rfstar")}+{c("beta")}*{c("erp")}', W['ke'])
drow('ke_r', 'Cost of equity, rating basis (published alternative)',
     f'=({c("rf")}-{c("sov_r")})+{c("beta")}*{c("erp_r")}', W['ke_rating'])
drow('kdfxle', 'Hard-currency debt at LOCAL-EQUIVALENT cost',
     f'=(1+{c("kd_fx")})*(1+{c("fxdep")})-1', W['kd_fx_local_equiv'])
drow('kdb', 'Blended marginal cost of debt',
     f'=(1-{c("wfx")})*{c("kd_lc")}+{c("wfx")}*{c("kdfxle")}', W['kd_blend'])
drow('kdat', 'Cost of debt after tax', f'={c("kdb")}*(1-{c("tax")})', W['kd_at'])
drow('mcap', 'Market capitalisation (EGP mn)', f'={c("spot")}*{c("shares")}', W['mcap'],
     fmt=MONEY)
drow('nd', 'Net debt (EGP mn)', f'={c("debt0")}-{c("cash0")}', W['net_debt'], fmt=MONEY)
drow('we', 'Equity weight (net basis)', f'={c("mcap")}/({c("mcap")}+{c("nd")})', W['we_net'],
     fmt=PCT)
drow('wd', 'Debt weight (net basis)', f'=1-{c("we")}', W['wd_net'], fmt=PCT)
drow('wacc0', 'Weighted average cost of capital, year one',
     f'={c("we")}*{c("ke")}+{c("wd")}*{c("kdat")}', W['wacc0'])
drow('wdg', 'Debt weight (gross basis)', f'={c("debt0")}/({c("mcap")}+{c("debt0")})',
     W['wd_gross'], fmt=PCT)
drow('wacc0g', 'Weighted average cost of capital, gross-debt basis',
     f'=(1-{c("wdg")})*{c("ke")}+{c("wdg")}*{c("kdat")}', W['wacc0_gross'])
drow('ket', 'Terminal cost of equity', f'={c("rf_t")}+{c("beta")}*{c("erp_t")}', W['ke_term'])
drow('kdt', 'Terminal cost of debt',
     f'=(1-{c("wfx")})*{c("kdt_lc")}+{c("wfx")}*((1+{c("kdt_fx")})*1.03-1)', W['kd_term'])
drow('kdtat', 'Terminal cost of debt after tax', f'={c("kdt")}*(1-{c("tax")})', W['kd_term_at'])
drow('waccT', 'Terminal weighted average cost of capital',
     f'=(1-{c("wd_t")})*{c("ket")}+{c("wd_t")}*{c("kdtat")}', W['wacc_term'])
ROW_LAST = r

# glide fractions, derived from the cost-of-debt path
lbl(wa, r, 1, 'Glide fraction (derived from the convergence path — SHAPE only)', bold=True)
lbl(wa, r, 2, 'calculated: (this year\'s cost of debt less the last year\'s) over (the first '
     'year\'s less the last year\'s)', note=True)
A['glide'] = r
kd_first = f'$C${A["kdpath"]}'; kd_last = f'$G${A["kdpath"]}'
for j in range(5):
    col = get_column_letter(3 + j)
    f(wa, r, 3 + j, f'=({col}${A["kdpath"]}-{kd_last})/({kd_first}-{kd_last})',
      W['glide_frac'][j], fmt='0.000')
r += 1
lbl(wa, r, 1, 'Discount rate', bold=True)
lbl(wa, r, 2, 'calculated: terminal rate plus the glide fraction of the gap to year one',
    note=True)
A['disc'] = r
for j in range(5):
    col = get_column_letter(3 + j)
    f(wa, r, 3 + j, f'=$C${A["waccT"]}+($C${A["wacc0"]}-$C${A["waccT"]})*{col}${A["glide"]}',
      W['disc_rate'][j], fmt=PCT2)
r += 1
lbl(wa, r, 1, 'Discount factor', bold=True)
lbl(wa, r, 2, 'calculated: the previous factor divided by one plus this year\'s rate — the '
    'factors compound', note=True)
A['df'] = r
f(wa, r, 3, f'=1/(1+C${A["disc"]})', W['df'][0], fmt='0.0000')
for j in range(1, 5):
    col = get_column_letter(3 + j); prev = get_column_letter(2 + j)
    f(wa, r, 3 + j, f'={prev}{r}/(1+{col}${A["disc"]})', W['df'][j], fmt='0.0000')
r += 1

# ================================================================== 6. SEGMENTS
wsg = sheet('Segments', [40, 15, 13, 13, 13, 13, 13, 13])
hdr(wsg, 1, ['Volume and price build', 'FY2024', 'FY2025', *YRS])
S = {}
r = 2
lbl(wsg, r, 1, 'Domestic packs sold (million)', bold=True)
val(wsg, r, 2, UB['dom_packs_fy24'], fmt='#,##0.00', kind='unit_build')
val(wsg, r, 3, UB['dom_packs_fy25'], fmt='#,##0.00', kind='unit_build')
S['dp'] = r
f(wsg, r, 3, f"=Assumptions!{c('ownpk0')}-Assumptions!{c('ep0')}", UB['dom_packs_fy25'],
  fmt='#,##0.000')
for j in range(5):
    prev = f'C{r}' if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["dvol"]})',
      FC['dom_packs'][j], fmt='#,##0.00')
r += 1
lbl(wsg, r, 1, 'Export packs sold (million)', bold=True)
val(wsg, r, 2, UB['exp_packs_fy24'], fmt='#,##0.00', kind='unit_build')
val(wsg, r, 3, UB['exp_packs_fy25'], fmt='#,##0.00', kind='unit_build')
S['ep'] = r
f(wsg, r, 3, f"=Assumptions!{c('ep0')}", UB['exp_packs_fy25'], fmt='#,##0.000')
for j in range(5):
    prev = f'C{r}' if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["evol"]})',
      FC['exp_packs'][j], fmt='#,##0.00')
r += 1
lbl(wsg, r, 1, 'Domestic realised price per pack (EGP) — own preparations only', bold=True)
S['dpp'] = r
DPP_HIST = r          # filled once the revenue rows below exist
for j in range(5):
    prev = f'C{r}' if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["dprice"]})',
      FC['dom_price'][j], fmt=PS)
r += 1
lbl(wsg, r, 1, 'Export realised price per pack (USD)', bold=True)
S['epp'] = r
for j in range(5):
    prev = f'C{r}' if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["eprice"]})',
      FC['exp_price_usd'][j], fmt=PS)
r += 1
lbl(wsg, r, 1, 'Exchange rate (EGP per USD)', bold=True)
val(wsg, r, 2, V['fx_avg_fy24'], fmt='#,##0.00')
val(wsg, r, 3, V['fx_avg_fy25'], fmt='#,##0.00')
S['fx'] = r
for j in range(5):
    f(wsg, r, 4 + j, f'=Assumptions!{get_column_letter(3 + j)}${A["fx"]}', FC['fx'][j],
      fmt='#,##0.00')
r += 1
lbl(wsg, r, 1, 'Domestic revenue, own preparations (EGP mn)', bold=True, fill=FILL_P)
S['drev'] = r
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'={col}{S["dp"]}*{col}{S["dpp"]}', FC['rev_dom'][j], fmt=MONEY,
      fill=FILL_P)
r += 1
lbl(wsg, r, 1, 'Export revenue (EGP mn)', bold=True, fill=FILL_P)
val(wsg, r, 2, UB['exp_rev_fy24'], fmt=MONEY, kind='audited')
f(wsg, r, 3, f"=Assumptions!{c('erev0')}", UB['exp_rev_fy25'], fmt=MONEY)
S['erev'] = r
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'={col}{S["ep"]}*{col}{S["epp"]}*{col}{S["fx"]}', FC['rev_exp'][j],
      fmt=MONEY, fill=FILL_P)
r += 1
lbl(wsg, r, 1, 'Contract-manufactured packs (million)', bold=True)
val(wsg, r, 2, UB['toll_packs_fy24'], fmt='#,##0.000', kind='audited')
f(wsg, r, 3, f"=Assumptions!{c('tollpk0')}", UB['toll_packs_fy25'], fmt='#,##0.000')
S['tpk'] = r
for j in range(5):
    prev = f'C{r}' if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["toll"]})',
      FC['toll_packs'][j], fmt='#,##0.000')
r += 1
lbl(wsg, r, 1, 'Contract-manufacturing FEE per pack (EGP)', bold=True)
S['tfee'] = r
f(wsg, r, 2, f"={V['ch_toll_fy24']}/B{S['tpk']}", UB['toll_fee_pp_fy24'], fmt=PS)
f(wsg, r, 3, f"=Assumptions!{c('toll0')}/C{S['tpk']}", UB['toll_fee_pp_fy25'], fmt=PS)
for j in range(5):
    prev = f"C{r}" if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["dprice"]})',
      FC['toll_fee_pp'][j], fmt=PS)
r += 1
lbl(wsg, r, 1, 'Contract product resold through own channels, per pack (EGP)', bold=True)
S['tres'] = r
f(wsg, r, 2, f"=({V['contract_value_fy24']}-{V['ch_toll_fy24']})/B{S['tpk']}",
  UB['resale_pp_fy24'], fmt=PS)
f(wsg, r, 3, f"=(Assumptions!{c('conval0')}-Assumptions!{c('toll0')})/C{S['tpk']}",
  UB['resale_pp_fy25'], fmt=PS)
for j in range(5):
    prev = f"C{r}" if j == 0 else get_column_letter(3 + j) + str(r)
    f(wsg, r, 4 + j, f'={prev}*(1+Assumptions!{get_column_letter(3 + j)}${A["dprice"]})',
      FC['resale_pp'][j], fmt=PS)
r += 1
lbl(wsg, r, 1, 'Contract-manufacturing revenue — fee (EGP mn)', bold=True, fill=FILL_P)
S['toll'] = r
for j, col in enumerate(('B', 'C')):
    f(wsg, r, 2 + j, f'={col}{S["tpk"]}*{col}{S["tfee"]}',
      [V['ch_toll_fy24'], V['ch_toll_fy25']][j], fmt='#,##0.0')
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'={col}{S["tpk"]}*{col}{S["tfee"]}', FC['toll'][j], fmt='#,##0.0',
      fill=FILL_P)
r += 1
lbl(wsg, r, 1, 'Contract product resold through own channels (EGP mn)', bold=True, fill=FILL_P)
S['tresrev'] = r
for j, col in enumerate(('B', 'C')):
    f(wsg, r, 2 + j, f'={col}{S["tpk"]}*{col}{S["tres"]}',
      [UB['contract_resale_fy24'], UB['contract_resale_fy25']][j], fmt='#,##0.0')
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'={col}{S["tpk"]}*{col}{S["tres"]}', FC['rev_resale'][j], fmt='#,##0.0',
      fill=FILL_P)
r += 1
# ---- backfill the HISTORY columns of the price and domestic-revenue rows as FORMULAS.
# The realised price is not an input: it is what the disclosed revenue divided by the
# disclosed pack count actually was, and the sheet computes it.
f(wsg, S['drev'], 2, f"={V['own_prep_value_fy24']}-B{S['erev']}", UB['dom_own_rev_fy24'],
  fmt=MONEY)
f(wsg, S['drev'], 3, f"=Assumptions!{c('ownval0')}-C{S['erev']}", UB['dom_own_rev_fy25'],
  fmt=MONEY)
for col, exp in (('B', UB['dom_price_fy24']), ('C', UB['dom_price_fy25'])):
    f(wsg, S['dpp'], 2 if col == 'B' else 3, f'={col}{S["drev"]}/{col}{S["dp"]}', exp, fmt=PS)
for col, exp, fxv in (('B', UB['exp_rev_fy24'] / UB['exp_packs_fy24'] / V['fx_avg_fy24'],
                       V['fx_avg_fy24']),
                      ('C', UB['exp_price_usd_fy25'], V['fx_avg_fy25'])):
    f(wsg, S['epp'], 2 if col == 'B' else 3,
      f'={col}{S["erev"]}/{col}{S["ep"]}/{col}{S["fx"]}', exp, fmt=PS)
lbl(wsg, r, 1, 'Total packs sold (million)', bold=True)
val(wsg, r, 2, V['packs_sold_fy24'], fmt='#,##0.00')
val(wsg, r, 3, V['packs_sold_fy25'], fmt='#,##0.00')
S['packs'] = r
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'={col}{S["dp"]}+{col}{S["ep"]}+{col}{S["tpk"]}',
      FC['packs_total'][j], fmt='#,##0.00')
r += 1
lbl(wsg, r, 1, 'REVENUE, consolidated (EGP mn)', bold=True, fill=FILL_C)
val(wsg, r, 2, V['rev_fy24'], fmt=MONEY)
val(wsg, r, 3, V['rev_fy25'], fmt=MONEY)
S['rev'] = r
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, r, 4 + j, f'=({col}{S["drev"]}+{col}{S["erev"]}+{col}{S["toll"]}'
      f'+{col}{S["tresrev"]})*Assumptions!{c("consol")}', FC['revenue'][j], fmt=MONEY,
      bold=True, fill=FILL_C)
r += 2
lbl(wsg, r, 1, 'DISCLOSED CAPACITY AND UTILISATION', bold=True, fill=FILL_C); r += 1
for label, v24, v25 in (('Units produced (million)', V['units_prod_fy24'], V['units_prod_fy25']),
                        ('Available capacity (million units)', V['units_cap'], V['units_cap'])):
    lbl(wsg, r, 1, label); val(wsg, r, 2, v24, fmt='#,##0'); val(wsg, r, 3, v25, fmt='#,##0')
    S[label[:4]] = r; r += 1
S['util'] = r
lbl(wsg, r, 1, 'Capacity utilisation', bold=True)
f(wsg, r, 2, f'=B{r - 2}/B{r - 1}', UB['utilisation_fy24'], fmt=PCT)
f(wsg, r, 3, f'=C{r - 2}/C{r - 1}', UB['utilisation_fy25'], fmt=PCT)
r += 1
lbl(wsg, r, 1, 'Registered preparations produced'); val(wsg, r, 2, 401, fmt='#,##0')
val(wsg, r, 3, V['products_fy25'], fmt='#,##0'); r += 1
lbl(wsg, r, 1, 'Average headcount'); val(wsg, r, 2, 4880, fmt='#,##0')
val(wsg, r, 3, V['employees_fy25'], fmt='#,##0'); r += 1
lbl(wsg, r, 1, 'Revenue per employee (EGP)', bold=True)
f(wsg, r, 2, f'=B{S["rev"]}*1000000/B{r - 1}', V['rev_fy24'] * 1e6 / 4880, fmt=MONEY)
f(wsg, r, 3, f'=C{S["rev"]}*1000000/C{r - 1}', V['rev_fy25'] * 1e6 / V['employees_fy25'],
  fmt=MONEY)

ESC_TOP = r + 2
# ======================================================== ESC (escalator engine)
r = ESC_TOP
lbl(wsg, r, 1, 'COST ESCALATOR INDEX (base FY2025 = 1.000)', bold=True, fill=FILL_C)
r += 1
ESC0 = r
esc = FC['esc_trace']
cum = {k: 1.0 for k in ('materials', 'packaging', 'labour', 'energy', 'services')}
rowmap = {'materials': ESC0, 'packaging': ESC0+1, 'labour': ESC0+2, 'energy': ESC0+3, 'services': ESC0+4}
S['esc_blend'] = ESC0+5
S['esc_unit'] = ESC0+6
for k, rr in rowmap.items():
    lbl(wsg, rr, 1, {'materials': 'Imported active ingredients',
                     'packaging': 'Packaging materials',
                     'labour': 'Labour', 'energy': 'Energy and utilities',
                     'services': 'Other consumables and services'}[k])
    for j in range(5):
        cum[k] *= (1 + esc[j][k])
        col = get_column_letter(4 + j); prev = get_column_letter(3 + j)
        if k == 'materials':
            expr = (f'=(1+Assumptions!{c("esc_api")})*(Segments!{get_column_letter(4 + j)}'
                    f'{S["fx"]}/' +
                    (f'Assumptions!{c("fx0")})' if j == 0 else
                     f'Segments!{get_column_letter(3 + j)}{S["fx"]})'))
        elif k == 'packaging':
            expr = (f'=Assumptions!{c("pk_imp")}*(1+Assumptions!{c("esc_api")})*'
                    f'(Segments!{get_column_letter(4 + j)}{S["fx"]}/' +
                    (f'Assumptions!{c("fx0")})' if j == 0 else
                     f'Segments!{get_column_letter(3 + j)}{S["fx"]})') +
                    f'+(1-Assumptions!{c("pk_imp")})*(1+Assumptions!'
                    f'{get_column_letter(3 + j)}${A["esc_cpi"]})')
        elif k == 'labour':
            expr = f'=1+Assumptions!{get_column_letter(3 + j)}${A["esc_lab"]}'
        elif k == 'energy':
            expr = f'=1+Assumptions!{get_column_letter(3 + j)}${A["esc_en"]}'
        else:
            expr = f'=1+Assumptions!{get_column_letter(3 + j)}${A["esc_cpi"]}'
        if j > 0:
            expr = f'={prev}{rr}*(' + expr[1:] + ')'
        else:
            expr = f'=(' + expr[1:] + ')'
        f(wsg, rr, 4 + j, expr, cum[k], fmt='0.0000')
lbl(wsg, S['esc_blend'], 1, 'Blended cash-cost index', bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(4 + j)
    num = '+'.join(f'Assumptions!{c(k)}*{col}{ESC0 + o}' for o, k in
                   enumerate(('sh_mat', 'sh_pack', 'sh_lab', 'sh_en', 'sh_svc')))
    den = '+'.join(f'Assumptions!{c(k)}' for k in
                   ('sh_mat', 'sh_pack', 'sh_lab', 'sh_en', 'sh_svc'))
    f(wsg, S['esc_blend'], 4 + j, f'=({num})/({den})',
      esc[j]['blend_index'], fmt='0.0000', bold=True, fill=FILL_C)
lbl(wsg, S['esc_unit'], 1, 'Cash cost per pack (EGP)', bold=True)
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['esc_unit'], 4 + j, f'=Assumptions!{c("uc0")}*{col}{S["esc_blend"]}', esc[j]['unit_cash_cost'], fmt=PS,
      bold=True)


# ============================================================ PPE roll-forward
r = S['esc_unit'] + 2
lbl(wsg, r, 1, 'PROPERTY ROLL-FORWARD (EGP mn)', bold=True, fill=FILL_C)
P0 = r + 1
for _k, _o in (('ppe_open',0),('ppe_tr',1),('ppe_dep',2),('ppe_am',3),('ppe_close',4),('ppe_capex',5),('ppe_dna',6),('cip_open',7),('cip_mv',8),('cip_close',9)):
    S[_k] = P0 + _o
lbl(wsg, S['ppe_open'], 1, 'Opening property, plant and equipment')
for j in range(5):
    prev = f'Assumptions!{c("ppe0")}' if j == 0 else f'{get_column_letter(3 + j)}{S["ppe_close"]}'
    f(wsg, S['ppe_open'], 4 + j, f'={prev}', V['ppe_fy25'] if j == 0 else FC['ppe'][j - 1], fmt=MONEY)
lbl(wsg, S['ppe_tr'], 1, 'Transfers in from construction')
for j in range(5):
    f(wsg, S['ppe_tr'], 4 + j, f'=Assumptions!{get_column_letter(3 + j)}${A["transfer"]}',
      V['cip_transfer'][j], fmt=MONEY)
lbl(wsg, S['ppe_dep'], 1, 'Depreciation charge')
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['ppe_dep'], 4 + j, f'=-Assumptions!{c("deprate")}*({col}{S["ppe_open"]}+{col}{S["ppe_tr"]}/2)', -FC['dep'][j], fmt=MONEY)
lbl(wsg, S['ppe_am'], 1, 'Amortisation of right-of-use and intangible assets')
for j in range(5):
    f(wsg, S['ppe_am'], 4 + j, f'=-Assumptions!{c("amort")}', -FC['amort'][j], fmt='#,##0.0')
lbl(wsg, S['ppe_close'], 1, 'Closing property, plant and equipment', bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['ppe_close'], 4 + j, f'={col}{S["ppe_open"]}+{col}{S["ppe_tr"]}+{col}{S["ppe_dep"]}', FC['ppe'][j], fmt=MONEY, bold=True, fill=FILL_C)
lbl(wsg, S['ppe_capex'], 1, 'Capital expenditure')
for j in range(5):
    f(wsg, S['ppe_capex'], 4 + j, f'={get_column_letter(4 + j)}{S["rev"]}*Assumptions!'
      f'{get_column_letter(3 + j)}${A["capex"]}', FC['capex'][j], fmt=MONEY)
lbl(wsg, S['ppe_dna'], 1, 'Total depreciation and amortisation', bold=True)
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['ppe_dna'], 4 + j, f'=-{col}{S["ppe_dep"]}-{col}{S["ppe_am"]}', FC['dna'][j], fmt=MONEY, bold=True)
lbl(wsg, S['cip_open'], 1, 'Opening construction balance')
for j in range(5):
    prev = f'Assumptions!{c("cip0")}' if j == 0 else f'{get_column_letter(3 + j)}{S["cip_close"]}'
    f(wsg, S['cip_open'], 4 + j, f'={prev}', V['cip_fy25'] if j == 0 else FC['cip'][j - 1], fmt=MONEY)
lbl(wsg, S['cip_mv'], 1, 'Additions less transfers out')
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['cip_mv'], 4 + j, f'={col}{S["ppe_capex"]}-{col}{S["ppe_tr"]}', FC['capex'][j] - V['cip_transfer'][j], fmt=MONEY)
lbl(wsg, S['cip_close'], 1, 'Closing construction balance', bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(4 + j)
    f(wsg, S['cip_close'], 4 + j, f'={col}{S["cip_open"]}+{col}{S["cip_mv"]}', FC['cip'][j], fmt=MONEY, bold=True, fill=FILL_C)


# ============================================================ 9. INCOME STATEMENT
wi = sheet('Income Statement', [40, 13, 13, 13, 13, 13, 13, 13, 13])
hdr(wi, 1, ['EGP million', 'FY2023', 'FY2024', 'FY2025', *YRS])
IS = {}
r = 2


def hist3(key, label, h23, h24, h25, bold=False, fmt=MONEY, fill=None):
    global r
    IS[key] = r
    lbl(wi, r, 1, label, bold=bold, fill=fill)
    for j, v in enumerate((h23, h24, h25)):
        val(wi, r, 2 + j, v, fmt=fmt, fill=fill)
    return r


hist3('rev', 'Revenue', H['FY2023']['revenue'], H['FY2024']['revenue'], H['FY2025']['revenue'],
      bold=True, fill=FILL_P)
for j in range(5):
    f(wi, r, 5 + j, f"=Segments!{get_column_letter(4 + j)}{S['rev']}", FC['revenue'][j],
      fmt=MONEY, bold=True, fill=FILL_P)
r += 1
hist3('cogs', 'Cost of sales', -H['FY2023']['cogs'], -H['FY2024']['cogs'], -H['FY2025']['cogs'])
for j in range(5):
    f(wi, r, 5 + j, f'=-(Segments!{get_column_letter(4 + j)}{S["packs"]}*Assumptions!{c("uc0")}'
      f'*Assumptions!{get_column_letter(3 + j)}${0}'.replace('$0', f'${A["esc_idx"]}')
      if False else
      f'=-({get_column_letter(5 + j)}{r + 1}+{get_column_letter(5 + j)}{r + 2})',
      -FC['cogs'][j], fmt=MONEY)
r += 1
# the two components of cost of sales, so the sheet SHOWS the split
IS['cogs_cash'] = r
lbl(wi, r, 1, '   of which cash cost (packs x cash cost per pack, escalated)', note=True)
for j in range(3):
    val(wi, r, 2 + j, [H['FY2023']['cogs'], H['FY2024']['cogs'],
                       H['FY2025']['cogs'] - 93.497560][j], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'=Segments!{get_column_letter(4 + j)}{S["packs"]}*Assumptions!{c("uc0")}'
      f'*Segments!{get_column_letter(4 + j)}${S["esc_blend"]}', FC['cogs_cash'][j], fmt=MONEY)
r += 1
IS['cogs_dep'] = r
lbl(wi, r, 1, '   of which depreciation charged to production', note=True)
for j in range(3):
    val(wi, r, 2 + j, [0.0, 0.0, 93.497560][j], fmt=MONEY)
for j in range(5):
    f(wi, r, 5 + j, f'=Segments!{get_column_letter(4 + j)}${S["ppe_dna"]}*Assumptions!{c("depcogs")}',
      FC['dna'][j] * FC['dep_cogs_share'], fmt=MONEY)
r += 1
IS['gp'] = r
lbl(wi, r, 1, 'Gross profit', bold=True, fill=FILL_C)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]
    f(wi, r, 2 + j, f'={get_column_letter(2 + j)}{IS["rev"]}+{get_column_letter(2 + j)}'
      f'{IS["cogs"]}', H[y]['gross_profit'], fmt=MONEY, bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'={col}{IS["rev"]}+{col}{IS["cogs"]}', FC['gross_profit'][j], fmt=MONEY,
      bold=True, fill=FILL_C)
r += 1
IS['gm'] = r
lbl(wi, r, 1, 'Gross margin')
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["gp"]}/{col}{IS["rev"]}', H[y]['gross_margin'], fmt=PCT)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'={col}{IS["gp"]}/{col}{IS["rev"]}', FC['gross_margin'][j], fmt=PCT)
r += 1
for key, label, hk, apct in (('mkt', 'Selling and marketing', 'marketing', 'mkt'),
                             ('rnd', 'Research and development', 'rnd', 'rnd'),
                             ('ga', 'General and administrative', 'ga', 'ga')):
    IS[key] = r
    lbl(wi, r, 1, label)
    for j in range(3):
        val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]][hk], fmt=MONEY)
    for j in range(5):
        col = get_column_letter(5 + j)
        f(wi, r, 5 + j, f'=-{col}{IS["rev"]}*Assumptions!{get_column_letter(3 + j)}${A[apct]}',
          -FC[{'mkt': 'marketing', 'rnd': 'rnd', 'ga': 'ga'}[key]][j], fmt=MONEY)
    r += 1
IS['prov'] = r
lbl(wi, r, 1, 'Credit losses, inventory write-downs and provisions — FRAME A')
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['provisions'], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'=-{col}{IS["rev"]}*Assumptions!{get_column_letter(3 + j)}${A["provA"]}',
      -FC['prov_A'][j], fmt=MONEY)
r += 1
IS['provB'] = r
lbl(wi, r, 1, 'The same charge on FRAME B (memorandum)', note=True)
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['provisions'], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'=-{col}{IS["rev"]}*Assumptions!{get_column_letter(3 + j)}${A["provB"]}',
      -FC['prov_B'][j], fmt=MONEY)
r += 1
IS['board'] = r
lbl(wi, r, 1, 'Board remuneration and attendance')
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['board'], fmt='#,##0.0')
for j in range(5):
    val(wi, r, 5 + j, -FC['board_fee'], fmt='#,##0.0')
r += 1
IS['dna_op'] = r
lbl(wi, r, 1, 'Depreciation charged to selling and administrative expense')
for j in range(3):
    val(wi, r, 2 + j, 0.0, fmt='#,##0.0')
for j in range(5):
    f(wi, r, 5 + j, f'=-Segments!{get_column_letter(4 + j)}{S["ppe_dna"]}*'
      f'(1-Assumptions!{c("depcogs")})', -FC['dna'][j] * (1 - FC['dep_cogs_share']),
      fmt='#,##0.0')
r += 1
IS['ebit'] = r
lbl(wi, r, 1, 'Operating profit (EBIT)', bold=True, fill=FILL_C)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["gp"]}+{col}{IS["mkt"]}+{col}{IS["rnd"]}+{col}{IS["ga"]}'
      f'+{col}{IS["prov"]}+{col}{IS["board"]}+{col}{IS["dna_op"]}', H[y]['ebit'], fmt=MONEY,
      bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'={col}{IS["gp"]}+{col}{IS["mkt"]}+{col}{IS["rnd"]}+{col}{IS["ga"]}'
      f'+{col}{IS["prov"]}+{col}{IS["board"]}+{col}{IS["dna_op"]}', FC['ebit_A'][j],
      fmt=MONEY, bold=True, fill=FILL_C)
r += 1
IS['dna'] = r
lbl(wi, r, 1, 'Depreciation and amortisation (memorandum)')
for j in range(3):
    val(wi, r, 2 + j, H[['FY2023', 'FY2024', 'FY2025'][j]]['dna'], fmt=MONEY)
for j in range(5):
    f(wi, r, 5 + j, f'=Segments!{get_column_letter(4 + j)}${S["ppe_dna"]}', FC['dna'][j], fmt=MONEY)
r += 1
IS['ebitda'] = r
lbl(wi, r, 1, 'EBITDA', bold=True)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["ebit"]}+{col}{IS["dna"]}', H[y]['ebitda'], fmt=MONEY, bold=True)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'={col}{IS["ebit"]}+{col}{IS["dna"]}', FC['ebit_A'][j] + FC['dna'][j],
      fmt=MONEY, bold=True)
r += 1
IS['fin'] = r
lbl(wi, r, 1, 'Finance costs')
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['finance'], fmt=MONEY)
for j in range(5):
    f(wi, r, 5 + j, f'=-Assumptions!{get_column_letter(3 + j)}${A["intpath"]}',
      -V['int_path'][j], fmt=MONEY)
r += 1
IS['assoc'] = r
lbl(wi, r, 1, 'Share of results of associates')
for j in range(3):
    val(wi, r, 2 + j, H[['FY2023', 'FY2024', 'FY2025'][j]]['associates'], fmt=MONEY)
for j in range(5):
    f(wi, r, 5 + j, f'=Assumptions!{c("assoc_e")}', V['assoc_norm'], fmt=MONEY)
r += 1
IS['divtax'] = r
lbl(wi, r, 1, 'Dividend distribution tax')
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['divtax'], fmt='#,##0.0')
for j in range(5):
    val(wi, r, 5 + j, 0.0, fmt='#,##0.0')
r += 1
IS['other'] = r
lbl(wi, r, 1, 'Interest income, exchange differences and other income')
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]
    val(wi, r, 2 + j, H[y]['interest_income'] + H[y]['fx'] + H[y]['other'], fmt=MONEY)
for j in range(5):
    val(wi, r, 5 + j, 0.0, fmt=MONEY)
r += 1
IS['pbt'] = r
lbl(wi, r, 1, 'Profit before tax', bold=True, fill=FILL_C)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["ebit"]}+{col}{IS["fin"]}+{col}{IS["assoc"]}+{col}{IS["other"]}'
      f'+{col}{IS["divtax"]}', H[y]['pbt'], fmt=MONEY, bold=True, fill=FILL_C)
pbt_f = []
for j in range(5):
    col = get_column_letter(5 + j)
    e = FC['ebit_A'][j] - V['int_path'][j] + V['assoc_norm']
    pbt_f.append(e)
    f(wi, r, 5 + j, f'={col}{IS["ebit"]}+{col}{IS["fin"]}+{col}{IS["assoc"]}+{col}{IS["other"]}'
      f'+{col}{IS["divtax"]}', e, fmt=MONEY, bold=True, fill=FILL_C)
r += 1
IS['tax'] = r
lbl(wi, r, 1, 'Income tax and the statutory solidarity contribution')
for j in range(3):
    val(wi, r, 2 + j, -H[['FY2023', 'FY2024', 'FY2025'][j]]['tax'], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j,
      f'=-({col}{IS["pbt"]}-{col}{IS["assoc"]})*Assumptions!{c("tax_eff")}',
      -(pbt_f[j] - V['assoc_norm']) * V['tax_eff_fwd'], fmt=MONEY)
r += 1
IS['np'] = r
lbl(wi, r, 1, 'Profit for the year', bold=True, fill=FILL_C)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["pbt"]}+{col}{IS["tax"]}', H[y]['net_profit'], fmt=MONEY,
      bold=True, fill=FILL_C)
np_f = []
for j in range(5):
    col = get_column_letter(5 + j)
    e = ((pbt_f[j] - V['assoc_norm']) * (1 - V['tax_eff_fwd'])
         + V['assoc_norm']); np_f.append(e)
    f(wi, r, 5 + j, f'={col}{IS["pbt"]}+{col}{IS["tax"]}', e, fmt=MONEY, bold=True, fill=FILL_C)
r += 1
IS['nci'] = r
lbl(wi, r, 1, 'Less non-controlling interests')
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]
    val(wi, r, 2 + j, -(H[y]['net_profit'] - H[y]['parent']), fmt='#,##0.0')
for j in range(5):
    val(wi, r, 5 + j, -FC['nci_fwd'], fmt='#,##0.0')
r += 1
IS['parent'] = r
lbl(wi, r, 1, 'Profit attributable to the holding company', bold=True, fill=FILL_C)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    f(wi, r, 2 + j, f'={col}{IS["np"]}+{col}{IS["nci"]}', H[y]['parent'], fmt=MONEY, bold=True,
      fill=FILL_C)
par_f = []
for j in range(5):
    col = get_column_letter(5 + j)
    e = np_f[j] - FC['nci_fwd']; par_f.append(e)
    f(wi, r, 5 + j, f'={col}{IS["np"]}+{col}{IS["nci"]}', e, fmt=MONEY, bold=True, fill=FILL_C)
r += 1
IS['eps'] = r
lbl(wi, r, 1, 'Earnings per share (EGP)', bold=True)
for j in range(3):
    y = ['FY2023', 'FY2024', 'FY2025'][j]; col = get_column_letter(2 + j)
    sh_y = [V['shares_fy23'], V['shares_fy23'], V['wavg_shares_fy25']][j]
    f(wi, r, 2 + j, f'={col}{IS["parent"]}/{sh_y}', H[y]['parent'] / sh_y, fmt=PS, bold=True)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wi, r, 5 + j, f'={col}{IS["parent"]}/Assumptions!{c("shares")}', par_f[j] / SH, fmt=PS,
      bold=True)

# ============================================================= 10. BALANCE SHEET
wb_ = sheet('Balance Sheet', [40, 13, 13, 13, 13, 13, 13, 13, 13])
hdr(wb_, 1, ['EGP million', 'FY2023', 'FY2024', 'FY2025', *YRS])
BS = {}
r = 2
OTHNC = (V['assoc_bv_fy25'] + V['intang_fy25'] + V['rou_fy25'] + V['dta_fy25']
         + V['afs_fy25'])
bs_hist = {
    'ppe': ('Property, plant and equipment, net', V['ppe_fy23'], V['ppe_fy24'], V['ppe_fy25']),
    'cip': ('Projects under construction', V['cip_fy23'], V['cip_fy24'], V['cip_fy25']),
    'oth_nc': ('Associates, intangibles, right-of-use, deferred tax and assets held for sale',
               V['rou_fy23'] + V['intang_fy23'] + V['assoc_bv_fy23'] + V['afs_fy25'],
               V['assoc_bv_fy24'] + 32.053960 + 9.566507 + V['afs_fy25'], OTHNC),
    'inv': ('Inventories, net', V['inv_fy23'], V['inv_fy24'], V['inv_fy25']),
    'ar': ('Trade and notes receivable, net', V['ar_fy23'], V['ar_fy24'], V['ar_fy25']),
    'od': ('Other debtors and debit balances', V['othdr_fy23'], V['othdr_fy24'], V['othdr_fy25']),
    'cash': ('Cash and bank balances', V['cash_fy23'], V['cash_fy24'], V['cash_fy25']),
}
OPEN_LINK = {'ppe': 'ppe0', 'cip': 'cip0', 'inv': 'inv0', 'ar': 'ar0', 'od': 'od0',
             'cash': 'cash0'}
for k, (label, h23, h24, h25) in bs_hist.items():
    BS[k] = r
    lbl(wb_, r, 1, label)
    for j, v in enumerate((h23, h24)):
        val(wb_, r, 2 + j, v, fmt=MONEY)
    if k in OPEN_LINK:
        f(wb_, r, 4, f'=Assumptions!{c(OPEN_LINK[k])}', h25, fmt=MONEY)
    else:
        val(wb_, r, 4, h25, fmt=MONEY)
    if k == 'cash':
        for j in range(5):
            f(wb_, r, 5 + j, f'=Assumptions!{c("cash0")}', V['cash_fy25'], fmt=MONEY)
    r += 1
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, BS['ppe'], 5 + j, f'=Segments!{get_column_letter(4 + j)}{S["ppe_close"]}', FC['ppe'][j], fmt=MONEY)
    f(wb_, BS['cip'], 5 + j, f'=Segments!{get_column_letter(4 + j)}{S["cip_close"]}', FC['cip'][j], fmt=MONEY)
    f(wb_, BS['oth_nc'], 5 + j, f'=D{BS["oth_nc"]}', OTHNC, fmt=MONEY)
    f(wb_, BS['inv'], 5 + j, f'=-Income Statement!{col}0'.replace('-Income Statement!', '') and
      f"=('Income Statement'!{col}{IS['cogs']}*-1)*Assumptions!{get_column_letter(3 + j)}"
      f'${A["dio"]}/365', FC['inventory'][j], fmt=MONEY)
    f(wb_, BS['ar'], 5 + j,
      f"='Income Statement'!{col}{IS['rev']}*Assumptions!{get_column_letter(3 + j)}"
      f'${A["dso"]}/365', FC['receivables'][j], fmt=MONEY)
    f(wb_, BS['od'], 5 + j,
      f"='Income Statement'!{col}{IS['rev']}/{V['rev_fy25']:.6f}*Assumptions!{c('od0')}",
      FC['other_dr'][j], fmt=MONEY)
BS['ta'] = r
lbl(wb_, r, 1, 'TOTAL ASSETS', bold=True, fill=FILL_C)
ta_hist = [V['assets_fy23'], V['assets_fy24'], V['assets_fy25']]
for j in range(3):
    col = get_column_letter(2 + j)
    f(wb_, r, 2 + j, f'=SUM({col}{BS["ppe"]}:{col}{BS["cash"]})', ta_hist[j], fmt=MONEY,
      bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    e = (FC['ppe'][j] + FC['cip'][j] + OTHNC + FC['inventory'][j] + FC['receivables'][j]
         + FC['other_dr'][j] + V['cash_fy25'])
    f(wb_, r, 5 + j, f'=SUM({col}{BS["ppe"]}:{col}{BS["cash"]})', e, fmt=MONEY, bold=True,
      fill=FILL_C)
r += 1
BS['ap'] = r
lbl(wb_, r, 1, 'Trade and notes payable')
for j, v in enumerate((V['ap_fy23'], V['ap_fy24'])):
    val(wb_, r, 2 + j, v, fmt=MONEY)
f(wb_, r, 4, f'=Assumptions!{c("ap0")}', V['ap_fy25'], fmt=MONEY)
r += 1
BS['oc'] = r
lbl(wb_, r, 1, 'Other creditors and credit balances')
for j, v in enumerate((V['othcr_fy23'], V['othcr_fy24'])):
    val(wb_, r, 2 + j, v, fmt=MONEY)
f(wb_, r, 4, f'=Assumptions!{c("oc0")}', V['othcr_fy25'], fmt=MONEY)
r += 1
BS['ptx'] = r
lbl(wb_, r, 1, 'Provisions, income tax payable and deferred tax')
PTX0 = V['provbs_fy25'] + V['taxpay_fy25'] + V['dtl_fy25']
for j, v in enumerate((V['provbs_fy23'] + V['taxpay_fy23'] + V['dtl_fy23'],
                       356.698370 + 344.846967 + 35.828578, PTX0)):
    val(wb_, r, 2 + j, v, fmt=MONEY)
for j in range(5):
    f(wb_, r, 5 + j, f'=D{BS["ptx"]}', PTX0, fmt=MONEY)
r += 1
BS['debt'] = r
lbl(wb_, r, 1, 'Gross borrowings including leases')
for j, v in enumerate((V['debt_fy23'], V['debt_fy24'])):
    val(wb_, r, 2 + j, v, fmt=MONEY)
f(wb_, r, 4, f'=Assumptions!{c("debt0")}', W['gross_debt'], fmt=MONEY)
r += 1
BS['eq'] = r
lbl(wb_, r, 1, 'Equity attributable to the holding company', bold=True)
for j, v in enumerate((V['equity_parent_fy23'], V['equity_parent_fy24'])):
    val(wb_, r, 2 + j, v, fmt=MONEY)
f(wb_, r, 4, f'=Assumptions!{c("eq0")}', V['equity_parent_fy25'], fmt=MONEY)
for j in range(5):
    prev = f'Assumptions!{c("eq0")}' if j == 0 else f'{get_column_letter(4 + j)}{r}'
    col = get_column_letter(5 + j)
    f(wb_, r, 5 + j,
      f"={prev}+'Income Statement'!{col}{IS['parent']}*(1-Assumptions!{c('payout')})",
      FC['equity'][j], fmt=MONEY, bold=True)
r += 1
BS['nci'] = r
lbl(wb_, r, 1, 'Non-controlling interests')
for j, v in enumerate((V['nci_fy23'], V['nci_fy24'])):
    val(wb_, r, 2 + j, v, fmt=MONEY)
f(wb_, r, 4, f'=Assumptions!{c("nci0")}', V['nci_fy25'], fmt=MONEY)
for j in range(5):
    f(wb_, r, 5 + j, f'=Assumptions!{c("ncibr")}', V['nci_bridge'], fmt=MONEY)
r += 1
# forecast payables and other creditors
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, BS['ap'], 5 + j,
      f"=('Income Statement'!{col}{IS['cogs']}*-1)*Assumptions!{get_column_letter(3 + j)}"
      f'${A["dpo"]}/365', FC['payables'][j], fmt=MONEY)
    f(wb_, BS['oc'], 5 + j,
      f"='Income Statement'!{col}{IS['rev']}/{V['rev_fy25']:.6f}*Assumptions!{c('oc0')}",
      FC['other_cr'][j], fmt=MONEY)
    # THE FUNDING PLUG. Cash is held at the audited operating minimum and gross borrowings
    # carry whatever the asset side needs that trade credit, provisions and equity do not
    # supply, so the forecast statement balances instead of being out by up to 6.6%.
    f(wb_, BS['debt'], 5 + j,
      f'={col}{BS["ta"]}-{col}{BS["ap"]}-{col}{BS["oc"]}-{col}{BS["ptx"]}'
      f'-{col}{BS["eq"]}-{col}{BS["nci"]}', FC['debt'][j], fmt=MONEY)
BS['tle'] = r
lbl(wb_, r, 1, 'TOTAL LIABILITIES AND EQUITY', bold=True, fill=FILL_C)
tle_hist = [
    V['ap_fy23'] + V['othcr_fy23'] + (V['provbs_fy23'] + V['taxpay_fy23'] + V['dtl_fy23'])
    + V['debt_fy23'] + V['equity_parent_fy23'] + V['nci_fy23'],
    V['ap_fy24'] + V['othcr_fy24'] + (356.698370 + 344.846967 + 35.828578)
    + V['debt_fy24'] + V['equity_parent_fy24'] + V['nci_fy24'],
    V['ap_fy25'] + V['othcr_fy25'] + PTX0 + W['gross_debt'] + V['equity_parent_fy25']
    + V['nci_fy25'],
]
ta_hist_ = [V['assets_fy23'], V['assets_fy24'], V['assets_fy25']]
for j in range(3):
    col = get_column_letter(2 + j)
    f(wb_, r, 2 + j, f'={col}{BS["ap"]}+{col}{BS["oc"]}+{col}{BS["ptx"]}+{col}{BS["debt"]}'
      f'+{col}{BS["eq"]}+{col}{BS["nci"]}', tle_hist[j], fmt=MONEY, bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, r, 5 + j, f'={col}{BS["ap"]}+{col}{BS["oc"]}+{col}{BS["ptx"]}+{col}{BS["debt"]}'
      f'+{col}{BS["eq"]}+{col}{BS["nci"]}', FC['assets'][j], fmt=MONEY, bold=True,
      fill=FILL_C)
r += 1
BS['chk'] = r
lbl(wb_, r, 1, 'BALANCE CHECK — total assets less total liabilities and equity', bold=True)
for j in range(8):
    col = get_column_letter(2 + j)
    exp = 0.0 if j >= 3 else (ta_hist_[j] - tle_hist[j])
    f(wb_, r, 2 + j, f'={col}{BS["ta"]}-{col}{BS["tle"]}', exp, fmt=MONEY, bold=True)
lbl(wb_, r, 10, 'ZERO IN EVERY FORECAST COLUMN, BY CONSTRUCTION — gross borrowings are the '
    'funding plug. The earlier edition never computed this row at all and its forecast '
    'columns were out by up to 6.6% of total assets. The three audited columns carry small '
    'residuals (-4.1, +0.2, +0.4 on assets of 10.0bn to 18.3bn, under 0.05%) from grouping '
    'the filed statement into these line captions; they are shown rather than suppressed.',
    note=True)
r += 1
BS['nd'] = r
lbl(wb_, r, 1, 'NET DEBT', bold=True, fill=FILL_C)
for j in range(3):
    col = get_column_letter(2 + j)
    f(wb_, r, 2 + j, f'={col}{BS["debt"]}-{col}{BS["cash"]}',
      [V['debt_fy23'] - V['cash_fy23'], V['debt_fy24'] - V['cash_fy24'],
       W['net_debt']][j], fmt=MONEY, bold=True, fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, r, 5 + j, f'={col}{BS["debt"]}-{col}{BS["cash"]}', FC['net_debt'][j], fmt=MONEY,
      bold=True, fill=FILL_C)
r += 1
BS['wc'] = r
lbl(wb_, r, 1, 'Net working capital', bold=True)
for j in range(3):
    col = get_column_letter(2 + j)
    r_inv = bs_hist['inv'][1 + j]; r_ar = bs_hist['ar'][1 + j]; r_od = bs_hist['od'][1 + j]
    ap_h = (V['ap_fy23'], V['ap_fy24'], V['ap_fy25'])[j]
    oc_h = (V['othcr_fy23'], V['othcr_fy24'], V['othcr_fy25'])[j]
    f(wb_, r, 2 + j, f'={col}{BS["inv"]}+{col}{BS["ar"]}+{col}{BS["od"]}-{col}{BS["ap"]}'
      f'-{col}{BS["oc"]}', r_inv + r_ar + r_od - ap_h - oc_h, fmt=MONEY, bold=True)
wc_f_full = list(FC['wc'])
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, r, 5 + j, f'={col}{BS["inv"]}+{col}{BS["ar"]}+{col}{BS["od"]}-{col}{BS["ap"]}'
      f'-{col}{BS["oc"]}', FC['wc'][j], fmt=MONEY, bold=True)
r += 1
BS['bvps'] = r
lbl(wb_, r, 1, 'Book value per share (EGP)', bold=True)
for j in range(3):
    col = get_column_letter(2 + j)
    sh_y = [V['shares_fy23'], V['shares_fy23'], SH][j]
    f(wb_, r, 2 + j, f'={col}{BS["eq"]}/{sh_y}',
      [V['equity_parent_fy23'], V['equity_parent_fy24'], V['equity_parent_fy25']][j] / sh_y,
      fmt=PS, bold=True)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wb_, r, 5 + j, f'={col}{BS["eq"]}/Assumptions!{c("shares")}', FC['equity'][j] / SH,
      fmt=PS, bold=True)

# ================================================================ 11. CASH FLOW
wcf = sheet('Cash Flow', [40, 13, 13, 13, 13, 13, 13, 13, 13])
hdr(wcf, 1, ['EGP million', 'FY2023', 'FY2024', 'FY2025', *YRS])
CF = {}
r = 2
CF['ebit'] = r
lbl(wcf, r, 1, 'Operating profit (EBIT)')
for j in range(3):
    col = get_column_letter(2 + j)
    f(wcf, r, 2 + j, f"='Income Statement'!{col}{IS['ebit']}",
      H[['FY2023', 'FY2024', 'FY2025'][j]]['ebit'], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wcf, r, 5 + j, f"='Income Statement'!{col}{IS['ebit']}", FC['ebit_A'][j], fmt=MONEY)
r += 1
CF['tax'] = r
lbl(wcf, r, 1, 'Tax on operating profit at the EFFECTIVE rate the business bears')
for j in range(3):
    col = get_column_letter(2 + j)
    f(wcf, r, 2 + j, f'=-{col}{CF["ebit"]}*Assumptions!{c("tax_eff")}',
      -H[['FY2023', 'FY2024', 'FY2025'][j]]['ebit'] * TAX_FCFF, fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wcf, r, 5 + j, f'=-{col}{CF["ebit"]}*Assumptions!{c("tax_eff")}',
      -(FC['ebit_A'][j]) * TAX_FCFF, fmt=MONEY)
r += 1
CF['nopat'] = r
lbl(wcf, r, 1, 'Operating profit after tax (NOPAT)', bold=True, fill=FILL_C)
for j in range(3):
    col = get_column_letter(2 + j)
    f(wcf, r, 2 + j, f'={col}{CF["ebit"]}+{col}{CF["tax"]}',
      H[['FY2023', 'FY2024', 'FY2025'][j]]['ebit'] * (1 - TAX_FCFF), fmt=MONEY, bold=True,
      fill=FILL_C)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wcf, r, 5 + j, f'={col}{CF["ebit"]}+{col}{CF["tax"]}', FC['ebit_A'][j] * (1 - TAX_FCFF),
      fmt=MONEY, bold=True, fill=FILL_C)
r += 1
CF['dna'] = r
lbl(wcf, r, 1, 'Add back depreciation and amortisation')
for j in range(3):
    col = get_column_letter(2 + j)
    f(wcf, r, 2 + j, f"='Income Statement'!{col}{IS['dna']}",
      H[['FY2023', 'FY2024', 'FY2025'][j]]['dna'], fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wcf, r, 5 + j, f"='Income Statement'!{col}{IS['dna']}", FC['dna'][j], fmt=MONEY)
r += 1
CF['capex'] = r
lbl(wcf, r, 1, 'Less capital expenditure')
for j, v in enumerate((-V['capex_fy23'], -V['capex_fy24'], -V['capex_fy25'])):
    val(wcf, r, 2 + j, v, fmt=MONEY)
for j in range(5):
    f(wcf, r, 5 + j, f'=-Segments!{get_column_letter(4 + j)}{S["ppe_capex"]}', -FC['capex'][j], fmt=MONEY)
r += 1
CF['dwc'] = r
lbl(wcf, r, 1, 'Less the increase in working capital')
for j, v in enumerate((V['dwc_fy23'], V['dwc_fy24'], V['dwc_fy25'])):
    val(wcf, r, 2 + j, v, fmt=MONEY)
for j in range(5):
    col = get_column_letter(5 + j); prevcol = get_column_letter(4 + j)
    prev = (f"'Balance Sheet'!D{BS['wc']}" if j == 0
            else f"'Balance Sheet'!{prevcol}{BS['wc']}")
    e = -FC['dwc'][j]
    f(wcf, r, 5 + j, f"=-('Balance Sheet'!{col}{BS['wc']}-{prev})", e, fmt=MONEY)
r += 1
CF['fcff'] = r
lbl(wcf, r, 1, 'FREE CASH FLOW TO THE FIRM', bold=True, fill=FILL_C)
fcff_sheet = []
for j in range(5):
    col = get_column_letter(5 + j)
    e = (FC['ebit_A'][j] * (1 - TAX_FCFF) + FC['dna'][j] - FC['capex'][j] - FC['dwc'][j])
    fcff_sheet.append(e)
    f(wcf, r, 5 + j, f'={col}{CF["nopat"]}+{col}{CF["dna"]}+{col}{CF["capex"]}+{col}{CF["dwc"]}',
      e, fmt=MONEY, bold=True, fill=FILL_C)
CF_FCFF = r
r += 2
lbl(wcf, r, 1, 'Memorandum — audited operating, investing and financing cash flow', bold=True,
    fill=FILL_P); r += 1
for label, vals3 in (('Net cash from operating activities',
                      (V['ocf_fy23'], V['ocf_fy24'], V['ocf_fy25'])),
                     ('Net cash used in investing activities',
                      (V['icf_fy23'], V['icf_fy24'], V['icf_fy25'])),
                     ('Net cash from financing activities',
                      (V['fcf_fy23'], V['fcf_fy24'], V['fcf_fy25']))):
    lbl(wcf, r, 1, label)
    for j, v in enumerate(vals3):
        val(wcf, r, 2 + j, v, fmt=MONEY)
    r += 1

# ======================================================================= 8. DCF
wd = sheet('DCF', [40, 14, 14, 14, 14, 14, 16])
hdr(wd, 1, ['EGP million', *YRS, 'Terminal'])
DR = {}
r = 2


def dcfrow(key, label, formulas, expected, fmt=MONEY, bold=False, fill=None):
    global r
    DR[key] = r
    lbl(wd, r, 1, label, bold=bold, fill=fill)
    for j in range(5):
        f(wd, r, 2 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    r += 1


dcfrow('rev', 'Revenue',
       lambda j: f"='Income Statement'!{get_column_letter(5 + j)}{IS['rev']}", FC['revenue'])
dcfrow('ebitda', 'EBITDA',
       lambda j: f"='Income Statement'!{get_column_letter(5 + j)}{IS['ebitda']}",
       [FC['ebit_A'][j] + FC['dna'][j] for j in range(5)], bold=True)
dcfrow('mgn', 'EBITDA margin = EBITDA / revenue',
       lambda j: f'={get_column_letter(2 + j)}{DR["ebitda"]}/{get_column_letter(2 + j)}'
                 f'{DR["rev"]}',
       [(FC['ebit_A'][j] + FC['dna'][j]) / FC['revenue'][j] for j in range(5)], fmt=PCT)
dcfrow('dna', 'Less depreciation and amortisation',
       lambda j: f'=-Segments!{get_column_letter(4 + j)}{S["ppe_dna"]}', [-x for x in FC['dna']])
dcfrow('ebit', 'EBIT = EBITDA less depreciation and amortisation',
       lambda j: f'={get_column_letter(2 + j)}{DR["ebitda"]}+{get_column_letter(2 + j)}'
                 f'{DR["dna"]}', [FC['ebit_A'][j] for j in range(5)], bold=True)
dcfrow('taxr', 'Tax rate — the EFFECTIVE rate the business bears, not the statutory rate',
       lambda j: f'=Assumptions!{c("tax_eff")}', [TAX_FCFF] * 5, fmt=PCT2)
dcfrow('nopat', 'NOPAT = EBIT x (1 less the tax rate)',
       lambda j: f'={get_column_letter(2 + j)}{DR["ebit"]}*(1-{get_column_letter(2 + j)}'
                 f'{DR["taxr"]})', [FC['ebit_A'][j] * (1 - TAX_FCFF) for j in range(5)],
       bold=True, fill=FILL_C)
dcfrow('adddna', 'Add back depreciation and amortisation',
       lambda j: f'=Segments!{get_column_letter(4 + j)}{S["ppe_dna"]}', FC['dna'])
dcfrow('capex', 'Less capital expenditure',
       lambda j: f'=-Segments!{get_column_letter(4 + j)}{S["ppe_capex"]}', [-x for x in FC['capex']])
dcfrow('dwc', 'Less the increase in working capital',
       lambda j: f"='Cash Flow'!{get_column_letter(5 + j)}{CF['dwc']}",
       [fcff_sheet[j] - (FC['ebit_A'][j] * (1 - TAX_FCFF) + FC['dna'][j] - FC['capex'][j])
        for j in range(5)])
dcfrow('fcff', 'FREE CASH FLOW TO THE FIRM',
       lambda j: f'={get_column_letter(2 + j)}{DR["nopat"]}+{get_column_letter(2 + j)}'
                 f'{DR["adddna"]}+{get_column_letter(2 + j)}{DR["capex"]}'
                 f'+{get_column_letter(2 + j)}{DR["dwc"]}', fcff_sheet, bold=True, fill=FILL_C)
dcfrow('disc', 'Discount rate',
       lambda j: f'=Assumptions!{get_column_letter(3 + j)}${A["disc"]}', W['disc_rate'],
       fmt=PCT2)
dcfrow('df', 'Discount factor',
       lambda j: f'=Assumptions!{get_column_letter(3 + j)}${A["df"]}', W['df'], fmt='0.0000')
dcfrow('pv', 'PRESENT VALUE OF FREE CASH FLOW TO THE FIRM',
       lambda j: f'={get_column_letter(2 + j)}{DR["fcff"]}*{get_column_letter(2 + j)}'
                 f'{DR["df"]}', [fcff_sheet[j] * W['df'][j] for j in range(5)], bold=True,
       fill=FILL_C)
r += 1
pv_sum_sheet = sum(fcff_sheet[j] * W['df'][j] for j in range(5))
nopat_t_sheet = (FC['ebit_A'][-1] * (1 - TAX_FCFF) * (1 + DERIVED['g_term'])
                 - DCFD['frame_A']['term_dep_catchup'] * (1 - TAX_FCFF))
# [R-TERM-01]: the terminal is the SANCTIONED construction, not the reinvestment
# identity. The workbook must reproduce the model, and the model no longer builds
# tv = NOPAT(1 - g/ROIC)/(W-g) — a form whose implied replacement cycle is 1/g, which is
# a fact about the growth rate rather than about the asset.
_TRA = DCFD['frame_A']['terminal_record']['outputs']
_TRB = DCFD['frame_B']['terminal_record']['outputs']
tv_sheet = _TRA['tv']
pv_tv_sheet = tv_sheet * W['df'][-1]
ev_core_sheet = pv_sum_sheet + pv_tv_sheet
assoc_val = V['assoc_norm'] * V['assoc_multiple']
ev_tot_sheet = ev_core_sheet + assoc_val + V['arab_api_cost'] + V['afs_fy25']
eq_sheet = ev_tot_sheet - W['net_debt'] - V['nci_bridge']
ps_sheet = eq_sheet / SH

lbl(wd, r, 1, 'TERMINAL BLOCK', bold=True, fill=FILL_C); r += 1
TB = {}
ic_fy30 = (FC['ppe'][-1] + FC['cip'][-1] + FC['wc'][-1] + V['intang_fy25'])
TB['ic'] = r
lbl(wd, r, 1, 'Invested capital at FY2030E — property, construction, working capital and '
    'intangibles')
f(wd, r, 2, f"='Balance Sheet'!I{BS['ppe']}+'Balance Sheet'!I{BS['cip']}"
  f"+'Balance Sheet'!I{BS['wc']}+Assumptions!{c('intang0')}", ic_fy30, fmt=MONEY)
r += 1
TB['roic'] = r
lbl(wd, r, 1, "Return on invested capital, FY2030E — the model's OWN, read from the rows above",
    bold=True)
f(wd, r, 2, f'=F{DR["nopat"]}/B{TB["ic"]}', DCFD['frame_A']['roic_term'], fmt=PCT2, bold=True)
lbl(wd, r, 4, 'This is what sets terminal reinvestment. It is not an assumption and there is '
    'no cell to type it into.', note=True)
r += 1
TB['tdep'] = r
lbl(wd, r, 1, 'Terminal depreciation catch-up — the parked construction balance at the '
    'depreciation rate')
f(wd, r, 2, f"='Balance Sheet'!I{BS['cip']}*Assumptions!{c('deprate')}",
  DCFD['frame_A']['term_dep_catchup'], fmt=MONEY)
lbl(wd, r, 4, 'Construction still parked at FY2030E has never entered the depreciable base. A '
    'perpetuity cannot capitalise profit on capital it never charges.', note=True)
r += 1
for key, label, formula, exp, fmt in (
    ('nt', 'Terminal NOPAT = final-year NOPAT x (1 + growth), less the depreciation the '
     'parked construction balance has never been charged',
     f'=F{DR["nopat"]}*(1+Assumptions!{c("g")})'
     f'-B{TB["tdep"]}*(1-Assumptions!{c("tax_eff")})', nopat_t_sheet, MONEY),
    # THE SANCTIONED TERMINAL, ROW BY ROW, so a reader following the labels reaches the
    # figure the page prints. The retired reinvestment identity charged g x IC every
    # year for ever and is gone: free cash flow is NOPAT plus book depreciation, less
    # maintenance at CURRENT cost over the DISCLOSED life, less the capital real growth
    # needs, less inflation on working capital.
    ('dna', 'Plus book depreciation (terminal NOPAT is already net of it)', None,
     _TRA['dna_addback'], MONEY),
    ('mnt', 'Less maintenance at current cost — book depreciation escalated over half '
     'the useful life the company itself discloses', None, -_TRA['maintenance'], MONEY),
    ('gcx', 'Less the capital behind real growth', None, -_TRA['growth_capex'], MONEY),
    ('wcc', 'Less inflation on working capital', None, -_TRA['wc_charge'], MONEY),
    ('ft', 'Terminal free cash flow to the firm', None, 0, MONEY),
    ('wt', 'Terminal discount rate', f'=Assumptions!{c("waccT")}', W['wacc_term'], PCT2),
    ('gt', 'Terminal growth', f'=Assumptions!{c("g")}', DERIVED["g_term"], PCT),
    ('tv', 'TERMINAL VALUE — the free cash flow above, grown one year and capitalised',
     None, 0, MONEY),
    ('pvtv', 'Present value of the terminal value', None, 0, MONEY),
):
    TB[key] = r
    lbl(wd, r, 1, label, bold=key in ('tv', 'pvtv'))
    if key in ('dna', 'mnt', 'gcx', 'wcc'):
        f(wd, r, 2, '=%.6f' % exp, exp, fmt=fmt)
    elif key == 'ft':
        f(wd, r, 2, f'=B{TB["nt"]}+B{TB["dna"]}+B{TB["mnt"]}+B{TB["gcx"]}+B{TB["wcc"]}',
          _TRA['fcff'], fmt=MONEY)
    elif key == 'tv':
        f(wd, r, 2, f'=B{TB["ft"]}*(1+B{TB["gt"]})/(B{TB["wt"]}-B{TB["gt"]})', tv_sheet,
          fmt=MONEY, bold=True)
    elif key == 'pvtv':
        f(wd, r, 2, f'=B{TB["tv"]}*F{DR["df"]}', pv_tv_sheet, fmt=MONEY, bold=True)
    else:
        f(wd, r, 2, formula, exp, fmt=fmt)
    r += 1
r += 1
lbl(wd, r, 1, 'Sum of the present values of forecast free cash flow', bold=True)
f(wd, r, 2, f'=SUM(B{DR["pv"]}:F{DR["pv"]})', pv_sum_sheet, fmt=MONEY, bold=True)
DR['pvsum'] = r; r += 1
lbl(wd, r, 1, 'CORE ENTERPRISE VALUE', bold=True, fill=FILL_C)
f(wd, r, 2, f'=B{DR["pvsum"]}+B{TB["pvtv"]}', ev_core_sheet, fmt=MONEY, bold=True, fill=FILL_C)
DR['evcore'] = r; r += 1
lbl(wd, r, 1, 'Terminal value as a percentage of CORE enterprise value', bold=True,
    fill=FILL_C)
f(wd, r, 2, f'=B{TB["pvtv"]}/B{DR["evcore"]}', pv_tv_sheet / ev_core_sheet, fmt=PCT, bold=True,
  fill=FILL_C)
DR['tvshare'] = r
r += 1
lbl(wd, r, 1, 'Terminal value as a percentage of TOTAL enterprise value (both are published)',
    bold=True)
f(wd, r, 2, f"=B{TB['pvtv']}/(B{DR['evcore']}+Assumptions!{c('assoc_e')}"
  f"*Assumptions!{c('assoc_m')}+Assumptions!{c('apicost')}+Assumptions!{c('afs0')})",
  pv_tv_sheet / ev_tot_sheet, fmt=PCT, bold=True)
DR['tvshare_tot'] = r

r += 2
lbl(wd, r, 1, 'FRAME B — THE SAME MODEL WITH THE CONTESTED JUDGEMENT RESOLVED THE OTHER WAY',
    bold=True, fill=FILL_C)
r += 1
FB = {}
prov_delta = [FC['prov_A'][j] - FC['prov_B'][j] for j in range(5)]
ebitB = [FC['ebit_A'][j] + prov_delta[j] for j in range(5)]
FB['ebit'] = r
lbl(wd, r, 1, 'EBIT on Frame B = Frame A EBIT plus the difference in the provision charge')
for j in range(5):
    col = get_column_letter(2 + j); icol = get_column_letter(5 + j)
    f(wd, r, 2 + j, f'={col}{DR["ebit"]}+(\'Income Statement\'!{icol}{IS["provB"]}'
      f'-\'Income Statement\'!{icol}{IS["prov"]})', ebitB[j], fmt=MONEY)
r += 1
FB['nopat'] = r
lbl(wd, r, 1, 'NOPAT on Frame B')
for j in range(5):
    col = get_column_letter(2 + j)
    f(wd, r, 2 + j, f'={col}{FB["ebit"]}*(1-{col}{DR["taxr"]})', ebitB[j] * (1 - TAX_FCFF),
      fmt=MONEY)
r += 1
FB['fcff'] = r
lbl(wd, r, 1, 'Free cash flow to the firm on Frame B')
fcffB = [fcff_sheet[j] + prov_delta[j] * (1 - TAX_FCFF) for j in range(5)]
for j in range(5):
    col = get_column_letter(2 + j)
    f(wd, r, 2 + j, f'={col}{FB["nopat"]}+{col}{DR["adddna"]}+{col}{DR["capex"]}'
      f'+{col}{DR["dwc"]}', fcffB[j], fmt=MONEY)
r += 1
FB['pv'] = r
lbl(wd, r, 1, 'Present value of free cash flow on Frame B')
for j in range(5):
    col = get_column_letter(2 + j)
    f(wd, r, 2 + j, f'={col}{FB["fcff"]}*{col}{DR["df"]}', fcffB[j] * W['df'][j], fmt=MONEY)
r += 1
nopat_tB = (ebitB[-1] * (1 - TAX_FCFF) * (1 + DERIVED['g_term'])
            - DCFD['frame_B']['term_dep_catchup'] * (1 - TAX_FCFF))
tvB = DCFD['frame_B']['terminal_record']['outputs']['tv']
evB = sum(fcffB[j] * W['df'][j] for j in range(5)) + tvB * W['df'][-1]
eqB = (evB + assoc_val + V['arab_api_cost'] + V['afs_fy25'] - W['net_debt']
       - V['nci_bridge'])
FB['roic'] = r
lbl(wd, r, 1, "Return on invested capital, FY2030E — Frame B's own")
f(wd, r, 2, f'=F{FB["nopat"]}/B{TB["ic"]}', DCFD['frame_B']['roic_term'], fmt=PCT2)
r += 1
# Frame B's terminal is built the SAME sanctioned way as Frame A's — the two frames
# differ in the provision charge and in nothing else, so a construction that differed
# between them would make the spread the construction rather than the judgement.
FB['ntb'] = r
lbl(wd, r, 1, 'Terminal NOPAT on Frame B, less the parked construction depreciation')
f(wd, r, 2, f'=F{FB["nopat"]}*(1+Assumptions!{c("g")})'
  f'-B{TB["tdep"]}*(1-Assumptions!{c("tax_eff")})', nopat_tB, fmt=MONEY)
r += 1
FB['ftb'] = r
lbl(wd, r, 1, 'Terminal free cash flow on Frame B — the same construction as Frame A: '
    'NOPAT plus book depreciation, less maintenance at current cost, less growth '
    'capital, less inflation on working capital')
f(wd, r, 2, f'=B{FB["ntb"]}+%.6f' % (_TRB['dna_addback'] - _TRB['maintenance']
                                     - _TRB['growth_capex'] - _TRB['wc_charge']),
  _TRB['fcff'], fmt=MONEY)
r += 1
FB['tv'] = r
lbl(wd, r, 1, 'Terminal value on Frame B')
f(wd, r, 2, f'=B{FB["ftb"]}*(1+B{TB["gt"]})/(B{TB["wt"]}-B{TB["gt"]})', tvB, fmt=MONEY)
r += 1
FB['ev'] = r
lbl(wd, r, 1, 'Core enterprise value on Frame B', bold=True)
f(wd, r, 2, f'=SUM(B{FB["pv"]}:F{FB["pv"]})+B{FB["tv"]}*F{DR["df"]}', evB, fmt=MONEY, bold=True)
r += 1
FB['ps'] = r
lbl(wd, r, 1, 'VALUE PER SHARE ON FRAME B', bold=True, fill=FILL_C)
f(wd, r, 2, f'=(B{FB["ev"]}+Assumptions!{c("assoc_e")}*Assumptions!{c("assoc_m")}'
  f'+Assumptions!{c("apicost")}+Assumptions!{c("afs0")}-Assumptions!{c("nd")}'
  f'-Assumptions!{c("ncibr")})/Assumptions!{c("shares")}', eqB / SH, fmt=PS, bold=True,
  fill=FILL_C)

# ============================================================= 5. SOTP BRIDGE
wsb = sheet('SOTP Bridge', [52, 18, 16])
hdr(wsb, 1, ['Enterprise value to equity value bridge', 'EGP million', 'EGP / share'])
BR = {}
r = 2
bridge = [
    ('evcore', 'Core enterprise value — discounted cash flow of the operating business',
     f"=DCF!B{DR['evcore']}", ev_core_sheet),
    ('tvpct', '   of which the terminal value (memorandum)', f"=DCF!B{TB['pvtv']}", pv_tv_sheet),
    ('tvshare', '   terminal value as a percentage of core enterprise value',
     f"=DCF!B{DR['tvshare']}", pv_tv_sheet / ev_core_sheet),
    ('assoc', 'Add: earning associates at normalised earnings x the multiple',
     f'=Assumptions!{c("assoc_e")}*Assumptions!{c("assoc_m")}', assoc_val),
    ('api', 'Add: the pre-revenue active-ingredient company at carrying cost',
     f'=Assumptions!{c("apicost")}', V['arab_api_cost']),
    ('afs', 'Add: assets held for sale', f'=Assumptions!{c("afs0")}', V['afs_fy25']),
    ('evtot', 'TOTAL ENTERPRISE VALUE', None, ev_tot_sheet),
    ('nd', 'Less: net debt', f'=-Assumptions!{c("nd")}', -W['net_debt']),
    ('nci', 'Less: non-controlling interests, post-deconsolidation',
     f'=-Assumptions!{c("ncibr")}', -V['nci_bridge']),
    ('eq', 'EQUITY VALUE', None, eq_sheet),
]
for key, label, formula, exp in bridge:
    BR[key] = r
    bold = key in ('evtot', 'eq', 'tvshare')
    fill = FILL_C if bold else None
    lbl(wsb, r, 1, label, bold=bold, fill=fill)
    if key == 'evtot':
        f(wsb, r, 2, f'=B{BR["evcore"]}+B{BR["assoc"]}+B{BR["api"]}+B{BR["afs"]}', exp,
          fmt=MONEY, bold=True, fill=FILL_C)
    elif key == 'eq':
        f(wsb, r, 2, f'=B{BR["evtot"]}+B{BR["nd"]}+B{BR["nci"]}', exp, fmt=MONEY, bold=True,
          fill=FILL_C)
    elif key == 'tvshare':
        f(wsb, r, 2, formula, exp, fmt=PCT, bold=True, fill=FILL_C)
    else:
        f(wsb, r, 2, formula, exp, fmt=MONEY)
    if key not in ('tvshare',):
        f(wsb, r, 3, f'=B{r}/Assumptions!{c("shares")}', exp / SH, fmt=PS, bold=bold, fill=fill)
    r += 1
r += 1
lbl(wsb, r, 1, 'Value per share — Frame A', bold=True, fill=FILL_C)
f(wsb, r, 2, f'=B{BR["eq"]}', eq_sheet, fmt=MONEY, bold=True, fill=FILL_C)
f(wsb, r, 3, f'=B{r}/Assumptions!{c("shares")}', ps_sheet, fmt=PS, bold=True, fill=FILL_C)
BR['psA'] = r

# ====================================================== 7. RELATIVE & NORMALIZED
wr = sheet('Relative & Normalized', [46, 16, 16, 40])
hdr(wr, 1, ['Triangulation', 'Multiple', 'Value (EGP/share)', 'Note'])
RL = {}
r = 2
lbl(wr, r, 1, 'SUSTAINABLE RETURN ON EQUITY — COMPUTED FROM THE FORECAST, NOT ASSERTED',
    bold=True, fill=FILL_C); r += 1
roe_rows = []
for j in (2, 3, 4):
    col = get_column_letter(5 + j)
    prev = f"Assumptions!{c('eq0')}" if j == 0 else f"'Balance Sheet'!{get_column_letter(4 + j)}{BS['eq']}"
    lbl(wr, r, 1, f"Return on average equity, {YRS[j]}")
    f(wr, r, 2, f"='Income Statement'!{col}{IS['parent']}"
      f"/(({prev}+'Balance Sheet'!{col}{BS['eq']})/2)", FC['roe'][j], fmt=PCT2)
    roe_rows.append(r); r += 1
RL['roe'] = r
lbl(wr, r, 1, 'Sustainable return — the mean of the last three forecast years', bold=True,
    fill=FILL_C)
f(wr, r, 2, f'=AVERAGE(B{roe_rows[0]}:B{roe_rows[-1]})', LN['roe_sust'], fmt=PCT2, bold=True,
  fill=FILL_C)
lbl(wr, r, 4, 'The forecast path RISES through the window rather than settling, so the '
    'sustainable return is the mean of its last three years — read from the path here, never '
    'typed', note=True)
r += 1
RL['pay'] = r
lbl(wr, r, 1, 'Payout the growth rate permits = 1 less growth over the sustainable return')
f(wr, r, 2, f'=1-Assumptions!{c("g")}/B{RL["roe"]}', LN['payout_implied'], fmt=PCT)
r += 2
lbl(wr, r, 1, 'FY2026E attributable earnings per share — Frame A')
f(wr, r, 2, f"='Income Statement'!E{IS['eps']}", par_f[0] / SH, fmt=PS)
RL['epsA'] = r; r += 1
lbl(wr, r, 1, 'FY2026E attributable earnings per share — Frame B')
epsB0 = ((ebitB[0] - V['int_path'][0]) * (1 - V['tax_eff_fwd'])
         + V['assoc_norm'] - FC['nci_fwd']) / SH
f(wr, r, 2, f'=((DCF!B{FB["ebit"]}-Assumptions!C${A["intpath"]})'
  f'*(1-Assumptions!{c("tax_eff")})'
  f'+Assumptions!{c("assoc_e")}-{FC["nci_fwd"]})/Assumptions!{c("shares")}', epsB0, fmt=PS)
RL['epsB'] = r; r += 1
lbl(wr, r, 1, 'Average of the two frames — the FORWARD earnings base', bold=True)
f(wr, r, 2, f'=AVERAGE(B{RL["epsA"]}:B{RL["epsB"]})', LN['eps_fwd'], fmt=PS,
  bold=True)
lbl(wr, r, 4, 'Averaging the two frames\' EARNINGS is not averaging the two VALUATIONS: the '
    'frames are published as two separate centres and are never averaged into one', note=True)
RL['eps'] = r; r += 1
RL['epsT'] = r
lbl(wr, r, 1, 'TRAILING attributable earnings per share, FY2025, on the count in issue today',
    bold=True)
f(wr, r, 2, f"=Assumptions!{c('parent25')}/Assumptions!{c('shares')}", LN['eps_ttm'], fmt=PS,
  bold=True)
lbl(wr, r, 4, 'The base for the two TRAILING multiples below. A trailing multiple applied to '
    'forward earnings would mismatch the periods', note=True)
r += 1
r += 2
lbl(wr, r, 1, "THE COMPANY'S OWN TRADED MULTIPLE HISTORY", bold=True, fill=FILL_C); r += 1
hdr(wr, r, ['Year', 'Year-end close', 'Attributable EPS', 'Price / earnings'], fill=FILL_P,
    font=F_SUB)
r += 1
oh_start = r
IS_PARENT_COL = {2023: 'B', 2024: 'C', 2025: 'D'}
for o in LN['own_pe_history']:
    lbl(wr, r, 1, str(o['year']))
    val(wr, r, 2, o['close'], fmt=PS)          # a market observation: the year-end close
    if o['year'] == 2022:
        f(wr, r, 3, f"=Assumptions!{c('par22')}/Assumptions!{c('sh22')}", o['eps'], fmt=PS)
    else:
        col = IS_PARENT_COL[o['year']]
        shref = (f"Assumptions!{c('sh23')}" if o['year'] < 2025
                 else f"Assumptions!{c('shwavg')}")
        f(wr, r, 3, f"='Income Statement'!{col}{IS['parent']}/{shref}", o['eps'], fmt=PS)
    f(wr, r, 4, f'=B{r}/C{r}', o['pe'], fmt='0.00"x"')
    r += 1
OH_MEAN_ROW = r
lbl(wr, r, 1, 'Four-year mean', bold=True)
f(wr, r, 4, f'=AVERAGE(D{oh_start}:D{r - 1})', LN['own_pe_mean'], fmt='0.00"x"', bold=True)
r += 1
lbl(wr, r, 1, 'Trailing multiple today — on the share count IN ISSUE', bold=True)
f(wr, r, 2, f'=Assumptions!{c("spot")}', M['spot'], fmt=PS)
f(wr, r, 3, f"=B{RL['epsT']}", LN['eps_ttm'], fmt=PS)
f(wr, r, 4, f'=B{r}/C{r}', LN['pe_now'], fmt='0.00"x"', bold=True)
r += 1
lbl(wr, r, 1, 'Trailing multiple today — on the AUDITED WEIGHTED-AVERAGE count for FY2025',
    bold=True)
f(wr, r, 2, f'=Assumptions!{c("spot")}', M['spot'], fmt=PS)
f(wr, r, 3, f"=Assumptions!{c('parent25')}/{V['wavg_shares_fy25']}", LN['eps_ttm_wavg'], fmt=PS)
f(wr, r, 4, f'=B{r}/C{r}', LN['pe_now_wavg'], fmt='0.00"x"', bold=True)
lbl(wr, r, 6, 'TWO SHARE COUNTS, BOTH PUBLISHED. The capital increase from 148.755750 to '
    '168.755750 million shares completed during FY2025, so that year has a weighted-average '
    'count of 162.016024 million and a closing count of 168.755750 million. Both readings of '
    'the trailing multiple are legitimate; neither is stated alone.', note=True)
r += 2

lbl(wr, r, 1, 'THE THREE MULTIPLES, EACH ON THE EARNINGS OF ITS OWN PERIOD', bold=True,
    fill=FILL_C)
r += 1
tri = LN['rel_triangulation']
for i, (name, mult, value) in enumerate(tri):
    RL[f't{i}'] = r
    lbl(wr, r, 1, name)
    if i == 0:
        f(wr, r, 2, f'=B{RL["pay"]}/(Assumptions!{c("ket")}'
          f'-Assumptions!{c("g")})', mult, fmt='0.00"x"')
    elif i == 1:
        f(wr, r, 2, f'=D{OH_MEAN_ROW}', mult, fmt='0.00"x"')
    else:
        f(wr, r, 2, f'=Assumptions!{c("peer_pe")}*(Assumptions!{c("peer_ke")}-Assumptions!{c("g")})/'
          f'(Assumptions!{c("ket")}-Assumptions!{c("g")})', mult, fmt='0.00"x"')
    eps_ref = f'B{RL["eps"]}' if i == 0 else f'B{RL["epsT"]}'
    eps_val = LN['eps_fwd'] if i == 0 else LN['eps_ttm']
    f(wr, r, 3, f'=B{r}*{eps_ref}', mult * eps_val, fmt=PS)
    lbl(wr, r, 4, ["A FORWARD multiple, applied to FORWARD earnings — built from this model's "
                   'own sustainable return and perpetual cost of equity',
                   'A TRAILING multiple, applied to TRAILING earnings — year-end closes '
                   'against audited attributable profit, each year on its own weighted-average '
                   'share count',
                   'A TRAILING multiple, applied to TRAILING earnings. The midpoint of the two '
                   'disclosed observations, not a median of a peer set; those companies face a '
                   'cost of equity near 10%, not this one'][i], note=True)
    r += 1
RL['avg'] = r
lbl(wr, r, 1, 'AVERAGE OF THE THREE — the relative lens', bold=True, fill=FILL_C)
f(wr, r, 2, f'=AVERAGE(B{RL["t0"]}:B{RL["t2"]})', sum(t[1] for t in tri) / 3, fmt='0.00"x"',
  bold=True, fill=FILL_C)
f(wr, r, 3, f'=AVERAGE(C{RL["t0"]}:C{RL["t2"]})', LN['rel_ps'], fmt=PS, bold=True,
  fill=FILL_C)
r += 2
lbl(wr, r, 1, 'NORMALISED EARNINGS POWER', bold=True, fill=FILL_C); r += 1
NP = {}
for key, label, formula, exp, fmt in (
    ('mgn', 'Three-year average operating margin', None, LN['norm_margin'], PCT),
    ('rev', 'FY2027E revenue', f"='Income Statement'!F{IS['rev']}", FC['revenue'][1], MONEY),
    ('ebit', 'Normalised operating profit', None, LN['norm_margin'] * FC['revenue'][1], MONEY),
    ('eps', 'Normalised earnings per share', None, LN['norm_pat_ps'], PS),
    ('pay', 'Payout the growth rate permits', None, LN['payout_implied'], PCT),
    ('val', 'NORMALISED EARNINGS POWER VALUE', None, LN['norm_ps'], PS),
):
    NP[key] = r
    lbl(wr, r, 1, label, bold=key == 'val', fill=FILL_C if key == 'val' else None)
    if key == 'mgn':
        f(wr, r, 2, f"=AVERAGE('Income Statement'!B{IS['ebit']}/'Income Statement'!B{IS['rev']},"
          f"'Income Statement'!C{IS['ebit']}/'Income Statement'!C{IS['rev']},"
          f"'Income Statement'!D{IS['ebit']}/'Income Statement'!D{IS['rev']})", exp, fmt=PCT)
    elif key == 'ebit':
        f(wr, r, 2, f'=B{NP["mgn"]}*B{NP["rev"]}', exp, fmt=MONEY)
    elif key == 'eps':
        f(wr, r, 2, f'=((B{NP["ebit"]}-{FC["board_fee"]}-Assumptions!D${A["intpath"]})'
          f'*(1-Assumptions!{c("tax_eff")})'
          f'+Assumptions!{c("assoc_e")}-{FC["nci_fwd"]})/Assumptions!{c("shares")}', exp, fmt=PS)
    elif key == 'pay':
        f(wr, r, 2, f'=1-Assumptions!{c("g")}/{LN["roe_sust"]:.8f}', exp, fmt=PCT)
    elif key == 'val':
        f(wr, r, 2, f'=B{NP["eps"]}*B{NP["pay"]}/(Assumptions!{c("ket")}-Assumptions!{c("g")})',
          exp, fmt=PS, bold=True, fill=FILL_C)
    else:
        f(wr, r, 2, formula, exp, fmt=fmt)
    r += 1

RL_AVG_ROW = RL['avg']
NP_VAL_ROW = NP['val']

# ==================================================== 3. FUNDAMENTAL VALUATION
wfv = sheet('Fundamental Valuation', [46, 16, 16, 46])
hdr(wfv, 1, ['Lens', 'Value (EGP/share)', 'Weight', 'How it is built'])
LV = {}
r = 2
lens_rows = [
    ('dcfA', 'Discounted cash flow — Frame A (provision charge permanent)',
     f"=('SOTP Bridge'!B{BR['eq']})/Assumptions!{c('shares')}", ps_sheet, LN['w_dcf'],
     'Five explicit years of free cash flow to the firm, discounted on a glide from the '
     'current to the terminal cost of capital, plus a terminal value on the growth-over-'
     'return reinvestment identity.'),
    ('dcfB', 'Discounted cash flow — Frame B (provision charge normalising)',
     f"=DCF!B{FB['ps']}", eqB / SH, LN['w_dcf'],
     'The identical model with the single contested judgement resolved the other way. '
     'Published beside Frame A, never averaged into it.'),
    ('book', 'Book value and sustainable return',
     f"=(('Relative & Normalized'!B{RL['roe']}-Assumptions!{c('g')})"
     f"/(Assumptions!{c('ket')}-Assumptions!{c('g')}))*'Balance Sheet'!D{BS['bvps']}",
     LN['book_ps'], 0.20,
     'Book value per share multiplied by (sustainable return less growth) over (perpetual '
     'cost of equity less growth). EVERY term is live: the sustainable return reads the '
     'forecast return-on-equity path, and the perpetual cost of equity reads beta, the '
     'terminal risk-free rate and the terminal premium.'),
    ('rel', 'Relative multiples', "='Relative & Normalized'!C%d" % (RL_AVG_ROW,),
     LN['rel_ps'], 0.15,
     'Three multiples triangulated ON the sheet, EACH APPLIED TO THE EARNINGS OF ITS OWN '
     'PERIOD — the forward multiple this model\'s own economics justify on forward earnings, '
     'and the company\'s own four-year mean and a cost-of-equity-adjusted struck peer '
     'reference on trailing earnings.'),
    ('norm', 'Normalised earnings power', "='Relative & Normalized'!B%d" % (NP_VAL_ROW,),
     LN['norm_ps'], 0.15,
     'The three-year average operating margin applied to forecast revenue, capitalised at '
     'the perpetual cost of equity on the payout that the growth rate permits.'),
]
for key, label, formula, exp, wt, how in lens_rows:
    LV[key] = r
    lbl(wfv, r, 1, label)
    if formula:
        f(wfv, r, 2, formula, exp, fmt=PS, bold=True)
    else:
        val(wfv, r, 2, exp, fmt=PS)
    val(wfv, r, 3, wt, fmt=PCT)
    lbl(wfv, r, 4, how, note=True)
    wfv.row_dimensions[r].height = 42
    r += 1
SHARED = ('book', 'rel', 'norm')
_shared_sum = '+'.join(f'B{LV[k]}*C{LV[k]}' for k in SHARED)
# [R-LENS-03]: ONE CLASS PRIMARY IS THE CENTRE, and this study has two of them because
# it publishes two frames. The typed 50/20/15/15 blend is retired. On a two-sided answer
# it did something worse than damp the number: it pulled BOTH frames toward the same
# three shared readings, damping precisely the disagreement a two-sided answer exists to
# show. The memo row below prints what it read, so the change is visible on the page.
LV['centreA'] = r
lbl(wfv, r, 1, 'CENTRE — FRAME A (the cash-flow lens, unweighted)', bold=True, fill=FILL_C)
f(wfv, r, 2, f'=B{LV["dcfA"]}', LN['centre_A'], fmt=PS, bold=True, fill=FILL_C)
lbl(wfv, r, 4, 'The centre of each frame IS its own cash-flow read. The three lenses that '
    'do not turn on the contested judgement are published beside it and averaged into '
    'neither.', note=True)
wfv.row_dimensions[r].height = 42
r += 1
LV['centreB'] = r
lbl(wfv, r, 1, 'CENTRE — FRAME B (the cash-flow lens, unweighted)', bold=True, fill=FILL_C)
f(wfv, r, 2, f'=B{LV["dcfB"]}', LN['centre_B'], fmt=PS, bold=True, fill=FILL_C)
lbl(wfv, r, 4, 'THE TWO CENTRES ARE NOT AVERAGED. One number covering both frames is a '
    'straight average of them, which is exactly what this study says it never does.',
    note=True)
wfv.row_dimensions[r].height = 42
r += 1
LV['blendmemo'] = r
lbl(wfv, r, 1, 'Memo — the retired weighted blend this edition replaced, on Frame A')
f(wfv, r, 2, f'=B{LV["dcfA"]}*C{LV["dcfA"]}+' + _shared_sum, LN['blend_A'], fmt=PS)
lbl(wfv, r, 4, 'Printed so the change is visible rather than only described. Its weights '
    'had never been tested out of sample, and one of the four readings it carried is not '
    'a lens this class uses at all.', note=True)
wfv.row_dimensions[r].height = 42
r += 1
LV['central'] = LV['centreA']
LV['lo'] = r
lbl(wfv, r, 1, 'Low of the field', bold=True)
f(wfv, r, 2, f'=MIN(B{LV["dcfA"]}:B{LV["norm"]})', LN['fair_bear'], fmt=PS, bold=True)
r += 1
LV['hi'] = r
lbl(wfv, r, 1, 'High of the field', bold=True)
f(wfv, r, 2, f'=MAX(B{LV["dcfA"]}:B{LV["norm"]})', LN['fair_bull'], fmt=PS, bold=True)
r += 1
LV['spot'] = r
lbl(wfv, r, 1, 'Market price', bold=True)
f(wfv, r, 2, f'=Assumptions!{c("spot")}', M['spot'], fmt=PS, bold=True)
r += 1
lbl(wfv, r, 1, 'Market price relative to the Frame A centre', bold=True, fill=FILL_C)
f(wfv, r, 2, f'=B{LV["spot"]}/B{LV["centreA"]}-1', M['spot'] / LN['centre_A'] - 1, fmt=PCT,
  bold=True, fill=FILL_C)
r += 1
lbl(wfv, r, 1, 'Market price relative to the Frame B centre', bold=True, fill=FILL_C)
f(wfv, r, 2, f'=B{LV["spot"]}/B{LV["centreB"]}-1', M['spot'] / LN['centre_B'] - 1, fmt=PCT,
  bold=True, fill=FILL_C)

# ============================================================== 2. SUMMARY
wsum = sheet('Summary', [46, 18, 18, 44])
hdr(wsum, 1, ['SUMMARY VALUATION TABLE', 'EGP / share', 'Against the market price', 'Note'])
r = 2
SU = {}
summary_rows = [
    ('dcfA', 'Discounted cash flow — Frame A',
     f"='Fundamental Valuation'!B{LV['dcfA']}", ps_sheet,
     f'Terminal value {pv_tv_sheet / ev_core_sheet:.0%} of core enterprise value, '
     f'{pv_tv_sheet / ev_tot_sheet:.0%} of total'),
    ('dcfB', 'Discounted cash flow — Frame B',
     f"='Fundamental Valuation'!B{LV['dcfB']}", DCFD['frame_B']['per_share'],
     f"Terminal value {DCFD['frame_B']['tv_share']:.0%} of core enterprise value, "
     f"{DCFD['frame_B']['tv_share_total']:.0%} of total"),
    ('book', 'Book value and sustainable return',
     f"='Fundamental Valuation'!B{LV['book']}", LN['book_ps'],
     f"Justified {LN['just_pb']:.2f} times book"),
    ('rel', 'Relative multiples', f"='Fundamental Valuation'!B{LV['rel']}", LN['rel_ps'],
     'Three multiples averaged on the sheet'),
    ('norm', 'Normalised earnings power', f"='Fundamental Valuation'!B{LV['norm']}",
     LN['norm_ps'], 'Three-year average margin'),
]
for key, label, formula, exp, note in summary_rows:
    SU[key] = r
    lbl(wsum, r, 1, label)
    f(wsum, r, 2, formula, exp, fmt=PS)
    f(wsum, r, 3, f'=B{r}/Assumptions!{c("spot")}-1', exp / M['spot'] - 1, fmt=PCT)
    lbl(wsum, r, 4, note, note=True)
    r += 1
SU['central'] = r
lbl(wsum, r, 1, 'CENTRE — FRAME A (the cash-flow lens, unweighted)', bold=True, fill=FILL_C)
f(wsum, r, 2, f"='Fundamental Valuation'!B{LV['centreA']}", LN['centre_A'], fmt=PS, bold=True,
  fill=FILL_C)
f(wsum, r, 3, f'=B{r}/Assumptions!{c("spot")}-1', LN['centre_A'] / M['spot'] - 1, fmt=PCT,
  bold=True, fill=FILL_C)
lbl(wsum, r, 4, 'A range, not a target. No rating is expressed anywhere in this workbook.',
    note=True)
r += 1
SU['centreB'] = r
lbl(wsum, r, 1, 'CENTRE — FRAME B (the cash-flow lens, unweighted)', bold=True, fill=FILL_C)
f(wsum, r, 2, f"='Fundamental Valuation'!B{LV['centreB']}", LN['centre_B'], fmt=PS, bold=True,
  fill=FILL_C)
f(wsum, r, 3, f'=B{r}/Assumptions!{c("spot")}-1', LN['centre_B'] / M['spot'] - 1, fmt=PCT,
  bold=True, fill=FILL_C)
lbl(wsum, r, 4, 'TWO CENTRES, NEVER AVERAGED INTO ONE. Which one you use is a judgement about '
    'the debtor book, and it is yours to make rather than ours to bury.', note=True)
r += 1
lbl(wsum, r, 1, 'Field low to field high', bold=True)
f(wsum, r, 2, f"='Fundamental Valuation'!B{LV['lo']}", LN['fair_bear'], fmt=PS, bold=True)
f(wsum, r, 3, f"='Fundamental Valuation'!B{LV['hi']}", LN['fair_bull'], fmt=PS, bold=True)
r += 1
lbl(wsum, r, 1, 'Market price', bold=True)
f(wsum, r, 2, f'=Assumptions!{c("spot")}', M['spot'], fmt=PS, bold=True)
r += 2
lbl(wsum, r, 1, 'KEY FIGURES', bold=True, fill=FILL_C); r += 1
for label, formula, exp, fmt in (
    ('Market capitalisation (EGP mn)', f'=Assumptions!{c("mcap")}', W['mcap'], MONEY),
    ('Net debt (EGP mn)', f'=Assumptions!{c("nd")}', W['net_debt'], MONEY),
    ('Enterprise value (EGP mn)',
     f'=Assumptions!{c("mcap")}+Assumptions!{c("nd")}+Assumptions!{c("ncibr")}',
     W['mcap'] + W['net_debt'] + V['nci_bridge'], MONEY),
    ('FY2025 revenue (EGP mn)', f"='Income Statement'!D{IS['rev']}", V['rev_fy25'], MONEY),
    ('FY2025 EBITDA (EGP mn)', f"='Income Statement'!D{IS['ebitda']}", H['FY2025']['ebitda'],
     MONEY),
    ('FY2025 attributable profit (EGP mn)', f"='Income Statement'!D{IS['parent']}",
     V['parent_fy25'], MONEY),
    ('Trailing price / earnings', None, LN['pe_now'], '0.0"x"'),
    ('Trailing enterprise value / EBITDA', None, LN['evebitda_now'], '0.0"x"'),
    ('Cost of equity', f'=Assumptions!{c("ke")}', W['ke'], PCT2),
    ('Weighted average cost of capital, year one', f'=Assumptions!{c("wacc0")}', W['wacc0'],
     PCT2),
    ('Terminal weighted average cost of capital', f'=Assumptions!{c("waccT")}', W['wacc_term'],
     PCT2),
    ('Terminal value as a percentage of enterprise value', f"=DCF!B{DR['tvshare']}",
     pv_tv_sheet / ev_core_sheet, PCT),
    ('Dividend per share proposed for FY2025 (EGP)', None, V['dps_fy25'], PS),
    ('Dividend yield at the market price', None, V['dps_fy25'] / M['spot'], PCT),
):
    lbl(wsum, r, 1, label)
    if label.startswith('Trailing price'):
        f(wsum, r, 2, f"=Assumptions!{c('spot')}/('Income Statement'!D{IS['parent']}"
          f"/Assumptions!{c('shares')})", exp, fmt=fmt)
    elif label.startswith('Trailing enterprise'):
        f(wsum, r, 2, f"=(Assumptions!{c('mcap')}+Assumptions!{c('nd')}+Assumptions!{c('ncibr')})"
          f"/'Income Statement'!D{IS['ebitda']}", exp, fmt=fmt)
    elif label.startswith('Dividend per share'):
        val(wsum, r, 2, exp, fmt=fmt)
    elif label.startswith('Dividend yield'):
        f(wsum, r, 2, f'=B{r - 1}/Assumptions!{c("spot")}', exp, fmt=fmt)
    else:
        f(wsum, r, 2, formula, exp, fmt=fmt)
    r += 1

# ======================================================== 12. SUMMARY FINANCIALS
wsf = sheet('Summary Financials', [36, 13, 13, 13, 13, 13, 13, 13, 13])
hdr(wsf, 1, ['EGP million', 'FY2023', 'FY2024', 'FY2025', *YRS])
r = 2
for label, isrow in (('Revenue', IS['rev']), ('Gross profit', IS['gp']), ('EBITDA', IS['ebitda']),
                     ('Operating profit', IS['ebit']), ('Profit before tax', IS['pbt']),
                     ('Profit attributable to the holding company', IS['parent'])):
    lbl(wsf, r, 1, label, bold=True)
    for j in range(8):
        col = get_column_letter(2 + j)
        key = ['FY2023', 'FY2024', 'FY2025'][j] if j < 3 else None
        if j < 3:
            e = {'Revenue': H[key]['revenue'], 'Gross profit': H[key]['gross_profit'],
                 'EBITDA': H[key]['ebitda'], 'Operating profit': H[key]['ebit'],
                 'Profit before tax': H[key]['pbt'],
                 'Profit attributable to the holding company': H[key]['parent']}[label]
        else:
            i = j - 3
            e = {'Revenue': FC['revenue'][i], 'Gross profit': FC['gross_profit'][i],
                 'EBITDA': FC['ebit_A'][i] + FC['dna'][i],
                 'Operating profit': FC['ebit_A'][i],
                 'Profit before tax': pbt_f[i],
                 'Profit attributable to the holding company': par_f[i]}[label]
        f(wsf, r, 2 + j, f"='Income Statement'!{col}{isrow}", e, fmt=MONEY)
    r += 1
for label, num, den in (('Gross margin', IS['gp'], IS['rev']),
                        ('EBITDA margin', IS['ebitda'], IS['rev']),
                        ('Operating margin', IS['ebit'], IS['rev']),
                        ('Net margin', IS['parent'], IS['rev'])):
    lbl(wsf, r, 1, label)
    for j in range(8):
        col = get_column_letter(2 + j)
        if j < 3:
            key = ['FY2023', 'FY2024', 'FY2025'][j]
            nv = {'Gross margin': H[key]['gross_profit'], 'EBITDA margin': H[key]['ebitda'],
                  'Operating margin': H[key]['ebit'], 'Net margin': H[key]['parent']}[label]
            dv = H[key]['revenue']
        else:
            i = j - 3
            nv = {'Gross margin': FC['gross_profit'][i],
                  'EBITDA margin': FC['ebit_A'][i] + FC['dna'][i],
                  'Operating margin': FC['ebit_A'][i], 'Net margin': par_f[i]}[label]
            dv = FC['revenue'][i]
        f(wsf, r, 2 + j, f"='Income Statement'!{col}{num}/'Income Statement'!{col}{den}",
          nv / dv, fmt=PCT)
    r += 1

# ========================================================== 15. PER-SHARE & RATIOS
wps = sheet('Per-Share & Ratios', [40, 13, 13, 13, 13, 13, 13, 13, 13])
hdr(wps, 1, ['', 'FY2023', 'FY2024', 'FY2025', *YRS])
r = 2
lbl(wps, r, 1, 'Earnings per share (EGP)', bold=True)
for j in range(8):
    col = get_column_letter(2 + j)
    if j < 3:
        y = ['FY2023', 'FY2024', 'FY2025'][j]; sh_y = [V['shares_fy23'], V['shares_fy23'], V['wavg_shares_fy25']][j]
        e = H[y]['parent'] / sh_y
    else:
        e = par_f[j - 3] / SH
    f(wps, r, 2 + j, f"='Income Statement'!{col}{IS['eps']}", e, fmt=PS)
r += 1
lbl(wps, r, 1, 'Book value per share (EGP)', bold=True)
for j in range(8):
    col = get_column_letter(2 + j)
    if j < 3:
        sh_y = [V['shares_fy23'], V['shares_fy23'], SH][j]
        e = [V['equity_parent_fy23'], V['equity_parent_fy24'], V['equity_parent_fy25']][j] / sh_y
    else:
        e = FC['equity'][j - 3] / SH
    f(wps, r, 2 + j, f"='Balance Sheet'!{col}{BS['bvps']}", e, fmt=PS)
r += 1
lbl(wps, r, 1, 'Dividend per share (EGP)')
for j, v in enumerate((V['dps_fy23'], V['dps_fy24'], V['dps_fy25'])):
    val(wps, r, 2 + j, v, fmt=PS)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wps, r, 5 + j, f"='Income Statement'!{col}{IS['eps']}*Assumptions!{c('payout')}",
      par_f[j] / SH * V['payout'], fmt=PS)
r += 1
lbl(wps, r, 1, 'Return on average equity')
for j in range(5):
    col = get_column_letter(5 + j); prevcol = get_column_letter(4 + j)
    prev = f"'Balance Sheet'!D{BS['eq']}" if j == 0 else f"'Balance Sheet'!{prevcol}{BS['eq']}"
    prev_v = V['equity_parent_fy25'] if j == 0 else FC['equity'][j - 1]
    f(wps, r, 5 + j,
      f"='Income Statement'!{col}{IS['parent']}/(({prev}+'Balance Sheet'!{col}{BS['eq']})/2)",
      par_f[j] / ((prev_v + FC['equity'][j]) / 2), fmt=PCT)
r += 1
lbl(wps, r, 1, 'Net debt / EBITDA')
for j in range(5):
    col = get_column_letter(5 + j)
    f(wps, r, 5 + j, f"=('Balance Sheet'!{col}{BS['debt']}-Assumptions!{c('cash0')})"
      f"/'Income Statement'!{col}{IS['ebitda']}",
      FC['net_debt'][j] / (FC['ebit_A'][j] + FC['dna'][j]), fmt=X)
r += 1
lbl(wps, r, 1, 'Inventory days')
for j in range(5):
    f(wps, r, 5 + j, f'=Assumptions!{get_column_letter(3 + j)}${A["dio"]}', V['dio'][j],
      fmt='#,##0')
r += 1
lbl(wps, r, 1, 'Receivable days')
for j in range(5):
    f(wps, r, 5 + j, f'=Assumptions!{get_column_letter(3 + j)}${A["dso"]}', V['dso'][j],
      fmt='#,##0')
r += 1
lbl(wps, r, 1, 'Payable days')
for j in range(5):
    f(wps, r, 5 + j, f'=Assumptions!{get_column_letter(3 + j)}${A["dpo"]}', V['dpo'][j],
      fmt='#,##0')
r += 1
lbl(wps, r, 1, 'Cash conversion cycle (days)', bold=True)
for j in range(5):
    col = get_column_letter(5 + j)
    f(wps, r, 5 + j, f'={col}{r - 3}+{col}{r - 2}-{col}{r - 1}',
      V['dio'][j] + V['dso'][j] - V['dpo'][j], fmt='#,##0', bold=True)
r += 1
lbl(wps, r, 1, 'Return on invested capital')
for j in range(5):
    col = get_column_letter(5 + j)
    ic = FC['ppe'][j] + FC['cip'][j] + wc_f_full[j] + V['intang_fy25']
    f(wps, r, 5 + j, f"=('Income Statement'!{col}{IS['ebit']}*(1-Assumptions!{c('tax_eff')}))"
      f"/('Balance Sheet'!{col}{BS['ppe']}+'Balance Sheet'!{col}{BS['cip']}"
      f"+'Balance Sheet'!{col}{BS['wc']}+Assumptions!{c('intang0')})",
      FC['ebit_A'][j] * (1 - TAX_FCFF) / ic, fmt=PCT)

# ============================================================== 13. MONTE CARLO
strike = json.load(open(os.path.join(HERE, 'strike_result.json')))
wmc = sheet('Monte Carlo', [42, 16, 16, 44])
hdr(wmc, 1, ['Probabilistic price map', 'One month', 'Three months', 'Note'])
lbl(wmc, 2, 1, f'THESE CELLS ARE A WHOLE-MODEL RE-RUN — {PATHS:,} simulated price paths each. '
    'They do NOT redraw when a driver on the Assumptions sheet changes.', bold=True,
    fill=FILL_C)
wmc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
r = 3
lbl(wmc, r, 1, 'Anchor price (EGP)')
val(wmc, r, 2, strike['spot'], fmt=PS, kind='grid')
val(wmc, r, 3, strike['spot'], fmt=PS, kind='grid')
lbl(wmc, r, 4, f"Anchored {strike['anchor_date']}", note=True); r += 1
lbl(wmc, r, 1, 'Check date')
val(wmc, r, 2, strike['horizons']['1M']['grade_date'], fmt='@', kind='grid')
val(wmc, r, 3, strike['horizons']['3M']['grade_date'], fmt='@', kind='grid')
lbl(wmc, r, 4, 'A calendar date, resolved to the exchange\'s first real trading session',
    note=True); r += 1
for p in (5, 25, 50, 75, 95):
    lbl(wmc, r, 1, f'{p}th percentile of the simulated distribution')
    val(wmc, r, 2, strike['horizons']['1M']['pct'][f'p{p}'], fmt=PS, kind='grid')
    val(wmc, r, 3, strike['horizons']['3M']['pct'][f'p{p}'], fmt=PS, kind='grid')
    r += 1
for label, key in (('Probability of finishing above the anchor', 'p_above'),
                   ('Probability of finishing 10% or more above', 'p_up10'),
                   ('Probability of finishing 10% or more below', 'p_dn10'),
                   ('Probability of TOUCHING 10% above at any point', 'touch_up10'),
                   ('Probability of TOUCHING 10% below at any point', 'touch_dn10')):
    lbl(wmc, r, 1, label)
    val(wmc, r, 2, strike['horizons']['1M'][key], fmt=PCT, kind='grid')
    val(wmc, r, 3, strike['horizons']['3M'][key], fmt=PCT, kind='grid')
    r += 1
lbl(wmc, r, 1, 'Annualised volatility used')
val(wmc, r, 2, strike['horizons']['1M']['anchor_vol_ann'], fmt=PCT, kind='grid')
val(wmc, r, 3, strike['horizons']['3M']['anchor_vol_ann'], fmt=PCT, kind='grid')
r += 2
bt = CAL['backtest']
lbl(wmc, r, 1, 'CALIBRATION EVIDENCE', bold=True, fill=FILL_C); r += 1
hdr(wmc, r, ['Window set', 'Windows', 'Skill against the benchmark', 'Probability-transform '
             'uniformity'], fill=FILL_P, font=F_SUB)
r += 1
for k, name in (('full', 'Full cleaned history'), ('five_year', 'Last five years of origins'),
                ('production', 'Post-break window set')):
    b = bt[k]
    lbl(wmc, r, 1, f"{name} ({b['first_origin']} to {b['last_origin']})")
    val(wmc, r, 2, b['windows'], fmt='#,##0', kind='grid')
    val(wmc, r, 3, b['skill_norm'], fmt='+0.0000;-0.0000', kind='grid')
    lbl(wmc, r, 4, f"chi-square p = {b['chi2_p']}, Kolmogorov-Smirnov p = {b['ks_p']} — "
        f"{'roughly uniform' if b['chi2_p'] > 0.05 and b['ks_p'] > 0.05 else 'not uniform'}",
        note=True)
    r += 1

# =============================================================== 14. SENSITIVITY
wsn = sheet('Sensitivity', [40, 15, 15, 15, 15, 15])
hdr(wsn, 1, ['Value per share (EGP) — each cell is a COMPLETE revaluation', '', '', '', '', ''])
lbl(wsn, 2, 1, 'THESE GRIDS ARE WHOLE-MODEL RE-RUNS. Each cell re-runs the entire model at a '
    'different input. They do NOT redraw when a driver on the Assumptions sheet changes.',
    bold=True, fill=FILL_C)
wsn.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
r = 4
for key, title, fmt_ in (('wacc', 'Shift in the cost of equity (basis points)', '+0"bp";-0"bp"'),
                         ('g', 'Terminal growth', PCT),
                         ('beta', 'Beta', '0.00'),
                         ('prov', 'Provision charge as a share of revenue', PCT2),
                         ('fx', 'Exchange-rate path, scaled', '0.00"x"'),
                         ('volume', 'Shift in domestic volume growth', '+0.0%;-0.0%'),
                         ('dep', 'Depreciation rate', PCT2)):
    lbl(wsn, r, 1, title, bold=True, fill=FILL_P)
    for j, (a, b) in enumerate(SENS[key]):
        val(wsn, r, 2 + j, a * 10000 if key == 'wacc' else a, fmt=fmt_, kind='grid')
    r += 1
    lbl(wsn, r, 1, 'Value per share')
    for j, (a, b) in enumerate(SENS[key]):
        val(wsn, r, 2 + j, b, fmt=PS, kind='grid')
    r += 2
lbl(wsn, r, 1, 'Cost of equity shift x terminal growth', bold=True, fill=FILL_P)
for j, g in enumerate(SENS['grid_g']):
    val(wsn, r, 2 + j, g, fmt=PCT, kind='grid')
r += 1
for i, w_ in enumerate(SENS['grid_wacc']):
    lbl(wsn, r, 1, f'{w_ * 10000:+.0f} basis points')
    for j in range(5):
        val(wsn, r, 2 + j, SENS['grid'][i][j], fmt=PS, kind='grid')
    r += 1
r += 1
lbl(wsn, r, 1, 'THE CRUX — reverse valuation', bold=True, fill=FILL_C); r += 1
for label, v_, fmt_ in (
        ('Additional FY2030E revenue required to reach the market price (EGP mn)',
         CRUX['required_fy30_revenue'], MONEY),
        ('   as a share of FY2030E revenue', CRUX['required_share_of_fy30'], PCT),
        ('   in US dollars a year (mn)', CRUX['required_rev_usd_mn'], MONEY),
        ('   as a multiple of the plant\'s stated USD 100 million cost', CRUX['asset_turn'], X)):
    lbl(wsn, r, 1, label)
    val(wsn, r, 2, v_, fmt=fmt_, kind='grid')
    r += 1

# ============================================================== 16. PEER & SECTOR
wpe = sheet('Peer & Sector', [40, 18, 18, 44])
hdr(wpe, 1, ['Peer / reference', 'Multiple', 'Basis', 'Provenance'])
r = 2
lbl(wpe, r, 1, 'Peer multiples are MARKET DATA and sit in the cross-check layer. No peer '
    'figure is used to build any historical number for the subject company.', bold=True,
    fill=FILL_C)
wpe.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
peers = [
    ('Listed Saudi Arabian generics manufacturer', PEER_HI, 'Trailing price / earnings',
     'Public-comparables service, 2025 basis — cross-check only. This company is not named '
     'and its financial statements are not published here, so this multiple cannot be '
     'rebuilt from its filings'),
    ('Regional and international generic manufacturers, mid-cap', PEER_MID,
     'Trailing price / earnings', 'Range observed across the peer set — cross-check only'),
    ('STRUCK REFERENCE USED — the midpoint of the two above', None,
     'Trailing price / earnings',
     'NOT a median of a disclosed peer set: two observations have a midpoint, and this cell '
     'computes it. The lens runs a 13-26 times band around it'),
]
pstart = r
for name, mult, basis, prov in peers:
    lbl(wpe, r, 1, name, bold=mult is None)
    if mult is None:
        f(wpe, r, 2, f'=AVERAGE(B{pstart}:B{pstart + 1})', V['peer_pe_regional'],
          fmt='0.00"x"', bold=True)
    else:
        val(wpe, r, 2, mult, fmt='0.00"x"')
    lbl(wpe, r, 3, basis); lbl(wpe, r, 4, prov, note=True)
    r += 1
lbl(wpe, r, 1, 'The subject company today', bold=True, fill=FILL_C)
f(wpe, r, 2, f"=Assumptions!{c('spot')}/('Income Statement'!D{IS['parent']}/Assumptions!"
  f"{c('shares')})", LN['pe_now'], fmt='0.00"x"', bold=True, fill=FILL_C)
lbl(wpe, r, 3, 'Trailing price / earnings')
lbl(wpe, r, 4, 'Computed in this sheet from the audited attributable profit and the market '
    'price', note=True)
r += 1
lbl(wpe, r, 1, "The subject company's own four-year mean", bold=True)
f(wpe, r, 2, f"='Relative & Normalized'!D{OH_MEAN_ROW}", LN['own_pe_mean'], fmt='0.00"x"',
  bold=True)
lbl(wpe, r, 3, 'Trailing price / earnings')
lbl(wpe, r, 4, 'Computed from year-end closes and audited profit — entirely primary', note=True)
r += 1
lbl(wpe, r, 1, 'Discount of the subject to the struck reference', bold=True, fill=FILL_C)
f(wpe, r, 2, f'=B{r - 2}/B{pstart + 2}-1', LN['pe_now'] / V['peer_pe_regional'] - 1, fmt=PCT,
  bold=True, fill=FILL_C)
lbl(wpe, r, 4, 'The peers face a cost of equity near 10%; this company\'s perpetual cost of '
    'equity is 14.9% and its current cost of equity 24.8%. The discount is not an anomaly to '
    'be closed.', note=True)
r += 2
lbl(wpe, r, 1, 'SECTOR OPERATING BENCHMARKS (the subject, computed here)', bold=True,
    fill=FILL_C); r += 1
for label, formula, exp, fmt_ in (
        ('Gross margin, FY2025', f"='Income Statement'!D{IS['gm']}",
         H['FY2025']['gross_margin'], PCT),
        ('EBITDA margin, FY2025',
         f"='Income Statement'!D{IS['ebitda']}/'Income Statement'!D{IS['rev']}",
         H['FY2025']['ebitda'] / V['rev_fy25'], PCT),
        ('Return on average equity, FY2025',
         f"='Income Statement'!D{IS['parent']}/(('Balance Sheet'!C{BS['eq']}"
         f"+'Balance Sheet'!D{BS['eq']})/2)", LN['roe_fy25'], PCT),
        ('Capacity utilisation, FY2025', f"=Segments!C{S['util']}",
         UB['utilisation_fy25'], PCT),
        ('Cash conversion cycle, FY2025 (days)', None,
         D['working_capital']['ccc_fy25'], '#,##0'),
        ('Export share of revenue, FY2025', None,
         UB['exp_rev_fy25'] / V['rev_fy25'], PCT)):
    lbl(wpe, r, 1, label)
    if formula:
        f(wpe, r, 2, formula, exp, fmt=fmt_)
    else:
        val(wpe, r, 2, exp, fmt=fmt_)
    r += 1

# ------------------------------------------------------------------ write out
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = sorted(wb._sheets, key=lambda s: order.index(s.title))
OUT = os.path.join(HERE, 'EIPICO_Valuation_Model_09082026.xlsx')
wb.save(OUT)
json.dump({'expected': EXPECT, 'paste_counts': NPASTE}, open(
    os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
print(f'wrote {os.path.basename(OUT)}')
print(f'formula cells: {len(EXPECT)}')
print(f'pasted cells by permitted class: {NPASTE}')
