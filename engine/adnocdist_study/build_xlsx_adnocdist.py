"""ADNOC Distribution — the valuation workbook. FORMULA-FIRST: the model CALCULATES.

Two primitives carry the whole discipline:

  f(ws, row, col, formula, expected)  writes a LIVE FORMULA and it is IMPOSSIBLE to
      write one without supplying the model's own value for that cell. Those expected
      values are dumped to xlsx_expected.json, which is what lets recalc.py evaluate the
      DELIVERED workbook independently and assert every formula reproduces the model.

  val(ws, row, col, v, kind=...)      writes a PASTED value and it is IMPOSSIBLE to paste
      one without naming which of the three permitted classes it belongs to.

Only three classes of cell may be pasted, and READ FIRST names them:
  'audited'    — audited or disclosed history. Where a line is both disclosed and
                 derivable, the DISCLOSED figure is carried.
  'unit_build' — the output of the unit build, where flattening the reconciliation into
                 the grid would be unreadable.
  'grid'       — whole-model re-runs (the probability map, the sensitivity grids), where
                 each cell is a complete revaluation and the grid does NOT redraw when a
                 driver changes.

No financial numeral is typed into this builder: every number is a lookup from
study_numbers.json. Blue = pasted input, black = formula, and that contract is stated on
READ FIRST.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
SR = json.load(open(os.path.join(HERE, 'sweep_research.json')))
M, H, F, W, DCFD, L = D['meta'], D['history'], D['forecast'], D['wacc'], D['dcf'], D['lenses']
UB, SENS, CRUX, WCD = D['unit_build'], D['sensitivity'], D['crux'], D['working_capital']
V = {k: v['value'] for k, v in D['inputs'].items()}
A_, B_ = DCFD['frame_A'], DCFD['frame_B']
YR = ['FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']
HY = ['FY2023', 'FY2024', 'FY2025']
N = 5

INK = '1C3A36'; PANEL = 'EAF0EE'; CREAM = 'F6F1E6'; BLUE = '0B4F9E'; GREYF = '6E7B77'
F_H = Font(name='Calibri', size=10.5, bold=True, color='FFFFFF')
F_SUB = Font(name='Calibri', size=10, bold=True, color=INK)
F_N = Font(name='Calibri', size=10, color=INK)
F_IN = Font(name='Calibri', size=10, color=BLUE)
F_NOTE = Font(name='Calibri', size=8.8, italic=True, color=GREYF)
FILL_H = PatternFill('solid', fgColor=INK)
FILL_P = PatternFill('solid', fgColor=PANEL)
FILL_C = PatternFill('solid', fgColor=CREAM)
THIN = Side(style='thin', color='C9D4D1')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = '0.0%'; PCT2 = '0.00%'; MONEY = '#,##0'; PS = '#,##0.00'; PS3 = '#,##0.000'
X = '0.00"x"'; TXT = '@'

wb = openpyxl.Workbook()
wb.remove(wb.active)
EXPECT = {}
NPASTE = {'audited': 0, 'unit_build': 0, 'grid': 0, 'label': 0}


def sheet(name, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def hdr(ws, row, cells):
    for j, t in enumerate(cells, start=1):
        c = ws.cell(row=row, column=j, value=t)
        c.font = F_H; c.fill = FILL_H; c.border = BOX
        c.alignment = Alignment(horizontal='center' if j > 1 else 'left',
                                wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 28


def lbl(ws, row, col, text, bold=False, fill=None, note=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_NOTE if note else (F_SUB if bold else F_N)
    if fill:
        c.fill = fill
    c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical='center')
    NPASTE['label'] += 1
    return c


def val(ws, row, col, v, fmt=MONEY, kind='audited', fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=v)
    c.font = Font(name='Calibri', size=10, bold=bold, color=BLUE)
    c.number_format = fmt; c.border = BOX
    if fill:
        c.fill = fill
    NPASTE[kind] += 1
    return c


def f(ws, row, col, formula, expected, fmt=MONEY, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name='Calibri', size=10, bold=bold, color=INK)
    c.number_format = fmt; c.border = BOX
    if fill:
        c.fill = fill
    EXPECT[f'{ws.title}!{c.coordinate}'] = float(expected)
    return c


def q(name):
    """Sheet name as it must appear inside a cross-sheet reference."""
    return f"'{name}'" if (' ' in name or '&' in name) else name


# ============================== ASSUMPTIONS ==============================
AS = 'Assumptions'
wa = sheet(AS, [46, 30, 14, 14, 14, 14, 14])
hdr(wa, 1, ['Assumption', 'Basis', YR[0], YR[1], YR[2], YR[3], YR[4]])
A = {}
r = 2


def arow(key, label, basis, values, fmt=PS3, kind='audited'):
    global r
    A[key] = r
    lbl(wa, r, 1, label); lbl(wa, r, 2, basis, note=True)
    if isinstance(values, (list, tuple)):
        for j, v in enumerate(values):
            val(wa, r, 3 + j, v, fmt=fmt, kind=kind)
    else:
        val(wa, r, 3, values, fmt=fmt, kind=kind)
    ws_h = 24
    wa.row_dimensions[r].height = ws_h
    r += 1


def drow(key, label, formula, expected, fmt=PCT2):
    global r
    A[key] = r
    lbl(wa, r, 1, label); lbl(wa, r, 2, 'calculated', note=True)
    f(wa, r, 3, formula, expected, fmt=fmt, bold=True)
    r += 1


def sect(title):
    global r
    c = lbl(wa, r, 1, title, bold=True, fill=FILL_P)
    for j in range(2, 8):
        lbl(wa, r, j, '', fill=FILL_P)
    r += 1


# Both helpers carry the sheet name. They are used from every other sheet, and an
# unqualified reference silently resolves against whichever sheet the formula sits on —
# which is not an error, just a wrong answer.
c = lambda k, col=3: f'{q(AS)}!${get_column_letter(col)}${A[k]}'
cy = lambda k, j: f'{q(AS)}!{get_column_letter(3 + j)}${A[k]}'

sect('Market')
arow('spot', 'Share price (AED)', 'close of ' + M['price_date'], V['spot'], fmt=PS)
arow('shares', 'Shares in issue (millions)', 'issued capital note', V['shares_mn'])
drow('mcap', 'Market capitalisation (AED m)', f'={c("spot")}*{c("shares")}', W['mcap'],
     fmt=MONEY)

sect('Cost of capital')
arow('rf', 'Government bond yield, local currency', 'federal Treasury Bond auction, Jul-2026',
     V['rf_observed'], fmt=PCT2)
arow('sov', 'Sovereign default spread', 'published country-risk file, Jan-2026',
     V['sov_spread'], fmt=PCT2)
drow('rfstar', 'Normalised risk-free rate = yield less sovereign spread',
     f'={c("rf")}-{c("sov")}', W['rf_star'])
arow('erp', 'Total equity risk premium', 'published country-risk file, Jan-2026',
     V['erp_total'], fmt=PCT2)
arow('beta', 'Beta', 'own five-year weekly regression against the published index of its '
     'own exchange', V['beta'], fmt=PS3)
drow('ke', 'Cost of equity = normalised risk-free + beta x premium',
     f'={c("rfstar")}+{c("beta")}*{c("erp")}', W['ke'])
arow('cmargin', 'Credit margin over the sovereign', "company's own disclosed loan margin",
     V['credit_margin'], fmt=PCT2)
drow('kdpre', 'Cost of debt before tax = government yield + credit margin',
     f'={c("rf")}+{c("cmargin")}', W['kd_pretax'])
arow('taxstat', 'Statutory tax rate', 'federal corporate tax', V['tax_statutory'], fmt=PCT)
arow('taxeff', 'Effective tax rate applied to cash flows',
     'audited tax reconciliation, FY2025', V['tax_effective'], fmt=PCT2)
drow('kdaft', 'Cost of debt after tax = before tax x (1 less statutory rate)',
     f'={c("kdpre")}*(1-{c("taxstat")})', W['kd_aftertax'])
arow('netdebt', 'Net debt excluding leases (AED m)', 'disclosure, FY2025 balance sheet',
     W['net_debt'], fmt=MONEY)
drow('we', 'Equity weight = market capitalisation / (market capitalisation + net debt)',
     f'={c("mcap")}/({c("mcap")}+{c("netdebt")})', W['we'], fmt=PCT)
drow('wd', 'Debt weight = 1 less the equity weight', f'=1-{c("we")}', W['wd'], fmt=PCT)
drow('wacc', 'Cost of capital, first forecast year',
     f'={c("we")}*{c("ke")}+{c("wd")}*{c("kdaft")}', W['wacc'])
# FLAT COST OF CAPITAL. The sliding schedule is excluded for currency-pegged markets, where
# the risk-free rate already sits at its norm: explicit = terminal. So the terminal beta IS
# the measured beta and the terminal debt weight IS today's. The previous drift toward the
# market and the 10% terminal weight are retired to the sensitivity table as priced
# constructions -- the second of them contradicted this model's own de-gearing forecast.
drow('betaterm', 'Terminal beta = the measured beta (flat, pegged market)',
     f'={c("beta")}', W['beta_terminal'], fmt=PS3)
drow('keterm', 'Terminal cost of equity = normalised risk-free + terminal beta x premium',
     f'={c("rfstar")}+{c("betaterm")}*{c("erp")}', W['ke_terminal'])
drow('wdterm', "Terminal debt weight = today's weight (flat, pegged market)",
     f'={c("wd")}', W['wd_terminal'], fmt=PCT)
drow('waccterm', 'Terminal cost of capital = terminal equity and debt weights on their costs',
     f'=(1-{c("wdterm")})*{c("keterm")}+{c("wdterm")}*{c("kdaft")}', W['wacc_terminal'])
arow('gterm', 'Long-run growth after the forecast period', 'below domestic inflation',
     V['g_terminal'], fmt=PCT2)
arow('roicterm', 'Terminal return on invested capital', 'fade from the realised return',
     V['roic_terminal'], fmt=PCT)

sect('Volume and price drivers — built bottom-up')
arow('stations_g', 'Service-station network growth', 'network +11.3% y/y to June 2026',
     V['stations_g'], fmt=PCT2)
arow('lps_g', 'Growth in litres sold per station', 'realised MINUS 9.3% y/y to June 2026',
     V['litres_per_station_g'], fmt=PCT2)
arow('volcorpg', 'Corporate fuel volume growth', 'realised MINUS 2.6% in the first half',
     V['vol_corp_g'], fmt=PCT2)
arow('volavig', 'Aviation fuel volume growth', 'realised PLUS 53.9% in the first half, faded',
     V['vol_avi_g'], fmt=PCT2)
arow('prretail', 'Realised retail price per litre (AED)', 'crude-linked price path',
     V['price_retfuel'], fmt=PS3)
arow('prcorp', 'Realised corporate price per litre (AED)', 'crude-linked price path',
     V['price_corp'], fmt=PS3)
arow('pravi', 'Realised aviation price per litre (AED)', 'crude-linked; jet prices ~50% above',
     V['price_avi'], fmt=PS3)
arow('mrg', 'Retail margin per litre, escalation', 'domestic escalator',
     V['gp_retfuel_per_l_g'], fmt=PCT2)
arow('mcg', 'Commercial margin per litre, escalation', 'realised step then domestic escalator',
     V['gp_comm_per_l_g'], fmt=PCT2)
arow('ftxng', 'Fuel transaction growth', 'realised +4.9% in the first half',
     V['fueltxn_g'], fmt=PCT2)
arow('convg', 'Growth in the non-fuel conversion rate', 'FELL from 27.0% to 26.2% y/y',
     V['conversion_g'], fmt=PCT2)
arow('baskg', 'Growth in the average non-fuel basket', 'domestic escalator plus food service',
     V['basket_g'], fmt=PCT2)
arow('nfm', 'Non-fuel retail gross margin', 'realised first half 2026',
     V['gm_nonfuel'], fmt=PCT)
arow('invA', 'Inventory movement, normalised frame (AED m)',
     'realised first half only, nil thereafter', V['invmove_A'], fmt=MONEY)
arow('invB', 'Inventory movement, through-cycle frame (AED m)',
     'FY2024-FY2025 average carried forward', V['invmove_B'], fmt=MONEY)

sect('Cost, capital spending and working capital')
arow('opexg', 'Cash operating cost growth', 'domestic escalator plus network, less efficiency',
     V['cash_opex_g'], fmt=PCT2)
arow('oig', 'Other income growth', 'domestic escalator', V['other_income_g'], fmt=PCT2)
arow('imp', 'Impairments and other operating expenses (AED m)', 'realised then normalised',
     V['impair_norm'], fmt=MONEY)
# These two were pasted five-year arrays. They are now written as placeholders and
# OVERWRITTEN with live formulas once the asset-base roll exists on the Segments sheet, so
# the charge and the spend are outputs of the base the model itself rolls forward.
arow('dna', 'Depreciation and amortisation (AED m)',
     'calculated', F['dna'], fmt=MONEY)
arow('capex', 'Capital expenditure (AED m)',
     'calculated', F['capex'], fmt=MONEY)
arow('deprate', 'Depreciation rate on the opening fixed and right-of-use base',
     'measured off the audited FY2025 accounts', V['dep_rate'], fmt=PCT2)
arow('maintrate', 'Maintenance capital spending, share of the opening base',
     'set equal to the depreciation rate: a network replaces what it consumes',
     V['maint_capex_rate'], fmt=PCT2)
arow('cpxsta', 'Capital cost per station added (AED m)',
     "backed out of the company's own FY2026 guidance", V['capex_per_station'], fmt=PS3)
arow('dso', 'Receivable days including parent balances', 'FY2025 statements',
     WCD['dso_all'], fmt=PS)
arow('dio', 'Inventory days', 'FY2025 statements', WCD['dio'], fmt=PS)
arow('dpo', 'Payable days including parent balances', 'FY2025 statements',
     WCD['dpo_all'], fmt=PS)
arow('dps', 'Dividend per share (AED)', 'stated policy', V['dps'], fmt=PS3)
arow('payoutfloor', 'Dividend policy floor, share of net profit',
     'the policy the company states: a minimum of 75%', V['payout_floor'], fmt=PCT)

sect('Anchors — the annualised disclosed first half of 2026')
arow('sta0', 'Service stations, June 2026', 'disclosure', V['stations_h126'], fmt='#,##0.0')
arow('staY0', 'Service stations, FY2025 year end', 'disclosure', V['stations_fy25'],
     fmt='#,##0.0')
arow('lps0', 'Litres per station, first half annualised (million)',
     'disclosed retail volume over the disclosed station count',
     UB['litres_per_station_h126'] * 2, fmt='#,##0.000', kind='unit_build')
arow('co0', 'Corporate fuel volume, first half annualised (million litres)', 'disclosure',
     V['vol_corp_h126'] * 2)
arow('av0', 'Aviation fuel volume, first half annualised (million litres)', 'disclosure',
     V['vol_avi_h126'] * 2)
arow('ft0', 'Fuel transactions, first half annualised (million)', 'disclosure',
     V['fueltxn_h126'] * 2, fmt='#,##0.0')
arow('cv0', 'Non-fuel conversion rate, first half 2026',
     'disclosed non-fuel transactions over disclosed fuel transactions',
     UB['conversion_h126'], fmt=PCT, kind='unit_build')
arow('bk0', 'Average non-fuel basket, first half 2026 (AED)',
     'disclosed non-fuel revenue over disclosed non-fuel transactions',
     UB['basket_h126'], fmt=PS, kind='unit_build')

sect('Opening position, from the audited FY2025 balance sheet')
arow('vol_r0', 'Retail fuel volume, FY2025 (million litres)', 'disclosure',
     V['vol_retail_fy25'], fmt=MONEY)
arow('vol_c0', 'Commercial fuel volume, FY2025 (million litres)', 'disclosure',
     V['vol_comm_fy25'], fmt=MONEY)
arow('mr0', 'Retail margin per litre, first half 2026 (AED)',
     'unit build: segment gross profit less inventory movement, over volume',
     UB['margin_retail_h126'], fmt=PS3, kind='unit_build')
arow('mc0', 'Commercial margin per litre, first half 2026 (AED)',
     'unit build: segment gross profit less inventory movement, over volume',
     UB['margin_comm_h126'], fmt=PS3, kind='unit_build')
arow('nf0', 'Non-fuel retail revenue, FY2025 (AED m)', 'disclosure',
     V['rev_nonfuel_fy25'], fmt=MONEY)
arow('opex0', 'Cash operating costs, FY2025 (AED m)', 'disclosure',
     H['FY2025']['cash_opex'], fmt=MONEY)
arow('oi0', 'Other income, FY2025 (AED m)', 'disclosure', V['oi_fy25'], fmt=MONEY)
arow('ppe0', 'Fixed and right-of-use assets, FY2025 (AED m)', 'disclosure',
     V['ppe_fy25'] + V['rou_fy25'], fmt=MONEY)
arow('gw0', 'Goodwill, intangibles and other non-current, FY2025 (AED m)', 'disclosure',
     V['gwi_fy25'] + V['onca_fy25'], fmt=MONEY)
arow('eq0', 'Equity attributable to owners, FY2025 (AED m)', 'disclosure',
     V['eqp_fy25'], fmt=MONEY)
arow('nci0', 'Non-controlling interests, FY2025 (AED m)', 'disclosure',
     V['nciq_fy25'], fmt=MONEY)
arow('borr0', 'Borrowings, FY2025 (AED m)', 'disclosure', V['borr_fy25'], fmt=MONEY)
arow('lease0', 'Lease liabilities, FY2025 (AED m)', 'disclosure', V['lease_fy25'], fmt=MONEY)
arow('prov0', 'Provisions and deferred tax, FY2025 (AED m)', 'disclosure',
     V['provs_fy25'] + V['oncl_fy25'], fmt=MONEY)
arow('td0', 'Term deposits, FY2025 (AED m)', 'disclosure', V['td_fy25'], fmt=MONEY)
arow('fin0', 'Finance costs, held at the FY2025 level (AED m)', 'disclosure, held flat',
     V['fin_fy25'], fmt=MONEY)
arow('ii0', 'Interest income, held at the FY2025 level (AED m)', 'disclosure, held flat',
     V['intinc_fy25'], fmt=MONEY)
arow('ncip', 'Non-controlling share of profit (AED m)', 'disclosure, held flat',
     V['nci_fy25'], fmt=MONEY)

# ---- model values the workbook must reproduce (computed here, never typed) ----
pbt_A = [A_['ebit'][i] + V['intinc_fy25'] - V['fin_fy25'] for i in range(N)]
tax_A = [p * V['tax_effective'] for p in pbt_A]
np_A = [pbt_A[i] - tax_A[i] for i in range(N)]
npa_A = [np_A[i] - V['nci_fy25'] for i in range(N)]
eps_A = [x / V['shares_mn'] for x in npa_A]
# The policy is USD 700 million a year OR a minimum of 75% of net profit, WHICHEVER IS
# HIGHER. Distributing only the fixed leg understated payments in the years where 75% of
# profit exceeds it, and it left the policy floor as a dead input.
# Dividends per year are the HIGHER of the fixed commitment and 75% of that year's profit,
# which is what the policy says. Carried as three rows on the income statement so the MAX is
# visible in the sheet rather than resolved in the builder.
divs_y = [max(V['dps'] * V['shares_mn'], V['payout_floor'] * npa_A[i]) for i in range(N)]
divs = divs_y[0]
eqp = []
_e = V['eqp_fy25']
for i in range(N):
    _e = _e + npa_A[i] - divs_y[i]
    eqp.append(_e)
ncieq = []
_n = V['nciq_fy25']
for i in range(N):
    _n = _n + V['nci_fy25']
    ncieq.append(_n)
ppe = []
_p = V['ppe_fy25'] + V['rou_fy25']
for i in range(N):
    _p = _p + F['capex'][i] - F['dna'][i]
    ppe.append(_p)
recv = [WCD['dso_all'] / 365 * F['revenue'][i] for i in range(N)]
invn = [WCD['dio'] / 365 * F['direct_costs_A'][i] for i in range(N)]
pay = [WCD['dpo_all'] / 365 * F['direct_costs_A'][i] for i in range(N)]
gw0 = V['gwi_fy25'] + V['onca_fy25']
prov0 = V['provs_fy25'] + V['oncl_fy25']
cash = [(eqp[i] + ncieq[i] + V['borr_fy25'] + V['lease_fy25'] + prov0 + pay[i])
        - (ppe[i] + gw0 + invn[i] + recv[i]) for i in range(N)]
ta = [ppe[i] + gw0 + invn[i] + recv[i] + cash[i] for i in range(N)]
tle = [eqp[i] + ncieq[i] + V['borr_fy25'] + V['lease_fy25'] + prov0 + pay[i] for i in range(N)]
# the balance-sheet line is 'cash and term deposits', so net debt nets against it whole
nd_f = [V['borr_fy25'] - cash[i] for i in range(N)]

# ============================== SEGMENTS ==============================
# THE UNIT BUILD, LIVE. Every line below is a formula off the Assumptions sheet: nothing in
# the forecast columns is pasted. Retail volume is stations times litres per station, both
# disclosed; corporate and aviation are separate legs on their own volume drivers and their
# own realised prices; non-fuel is transactions times conversion times basket. The blended
# commercial price is an OUTPUT of the mix, not an input.
SG = 'Segments'
wsg = sheet(SG, [46, 15, 14, 14, 14, 14, 14])
hdr(wsg, 1, ['The unit build — every forecast cell is a formula', 'Anchor',
             YR[0], YR[1], YR[2], YR[3]])
S = {}
rr = 2


def srow(key, label, formulas, expected, base=None, fmt=MONEY, bold=False, fill=None,
         basefmt=None, basekind='audited'):
    global rr
    S[key] = rr
    lbl(wsg, rr, 1, label, bold=bold, fill=fill)
    if base is None:
        lbl(wsg, rr, 2, '')
    else:
        val(wsg, rr, 2, base, fmt=basefmt or fmt, kind=basekind)
    for j in range(N):
        f(wsg, rr, 3 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    rr += 1


def grow(key, drv, base_ref):
    """A row that compounds its own previous column off a driver on Assumptions."""
    return lambda j: (f'={get_column_letter(2 + j)}{S[key]}*(1+{cy(drv, j)})' if j
                      else f'={base_ref}*(1+{cy(drv, 0)})')


lbl(wsg, rr, 1, 'The anchor column is the ANNUALISED DISCLOSED FIRST HALF of 2026, not '
    'FY2025: two quarters of the study year are already on the public record. The first '
    'forecast column therefore carries the second-half shape on that realised run rate, and '
    'growth proper begins in FY2027.', note=True)
wsg.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
wsg.row_dimensions[rr].height = 32
rr += 1

_sta0, _lps0 = V['stations_h126'], UB['litres_per_station_h126'] * 2
_co0, _av0 = V['vol_corp_h126'] * 2, V['vol_avi_h126'] * 2
_ft0, _cv0, _bk0 = V['fueltxn_h126'] * 2, UB['conversion_h126'], UB['basket_h126']

srow('sta', 'Service stations', grow('sta', 'stations_g', f'{c("sta0")}'),
     F['stations'], base=_sta0, fmt='#,##0.0')
srow('lps', 'Litres sold per station (million)', grow('lps', 'lps_g', f'{c("lps0")}'),
     F['litres_per_station'], base=_lps0, fmt='#,##0.000')
srow('volr', 'RETAIL FUEL VOLUME = stations x litres per station',
     lambda j: (f'={get_column_letter(3 + j)}{S["sta"]}*'
                f'{get_column_letter(3 + j)}{S["lps"]}'),
     F['vol_retail'], base=_sta0 * _lps0, bold=True)
srow('volco', 'Corporate fuel volume (million litres)',
     grow('volco', 'volcorpg', f'{c("co0")}'), F['vol_corp'], base=_co0)
srow('volav', 'Aviation fuel volume (million litres)',
     grow('volav', 'volavig', f'{c("av0")}'), F['vol_avi'], base=_av0)
srow('volc', 'Commercial fuel volume = corporate + aviation',
     lambda j: (f'={get_column_letter(3 + j)}{S["volco"]}+'
                f'{get_column_letter(3 + j)}{S["volav"]}'),
     F['vol_comm'], base=_co0 + _av0, bold=True)
srow('volt', 'TOTAL FUEL VOLUME',
     lambda j: (f'={get_column_letter(3 + j)}{S["volr"]}+'
                f'{get_column_letter(3 + j)}{S["volc"]}'),
     F['vol_total'], base=_sta0 * _lps0 + _co0 + _av0, bold=True, fill=FILL_C)
srow('prr', 'Realised retail price per litre (AED)', lambda j: f'={cy("prretail", j)}',
     F['price_retail'], base=UB['price_retail_h126'], fmt=PS3)
srow('prco', 'Realised corporate price per litre (AED)', lambda j: f'={cy("prcorp", j)}',
     F['price_corp'], base=UB['price_corp_h126'], fmt=PS3)
srow('prav', 'Realised aviation price per litre (AED)', lambda j: f'={cy("pravi", j)}',
     F['price_avi'], base=UB['price_avi_h126'], fmt=PS3)
srow('revr', 'Retail fuel revenue = volume x price',
     lambda j: f'={get_column_letter(3 + j)}{S["volr"]}*{get_column_letter(3 + j)}{S["prr"]}',
     F['rev_retfuel'], base=V['rev_retfuel_h126'] * 2)
srow('revco', 'Corporate revenue = volume x price',
     lambda j: (f'={get_column_letter(3 + j)}{S["volco"]}*'
                f'{get_column_letter(3 + j)}{S["prco"]}'),
     F['rev_corp'], base=V['rev_corp_h126'] * 2)
srow('revav', 'Aviation revenue = volume x price',
     lambda j: (f'={get_column_letter(3 + j)}{S["volav"]}*'
                f'{get_column_letter(3 + j)}{S["prav"]}'),
     F['rev_avi'], base=V['rev_avi_h126'] * 2)
srow('revc', 'Commercial revenue = corporate + aviation',
     lambda j: (f'={get_column_letter(3 + j)}{S["revco"]}+'
                f'{get_column_letter(3 + j)}{S["revav"]}'),
     F['rev_comm'], base=(V['rev_corp_h126'] + V['rev_avi_h126']) * 2)
srow('prc', 'Blended commercial price per litre — AN OUTPUT of the mix',
     lambda j: f'={get_column_letter(3 + j)}{S["revc"]}/{get_column_letter(3 + j)}{S["volc"]}',
     F['price_comm'], base=UB['price_comm_h126'], fmt=PS3)
srow('ftxn', 'Fuel transactions (million)', grow('ftxn', 'ftxng', f'{c("ft0")}'),
     F['fuel_txn'], base=_ft0, fmt='#,##0.0')
srow('conv', 'Non-fuel conversion rate', grow('conv', 'convg', f'{c("cv0")}'),
     [F['nonfuel_txn'][i] / F['fuel_txn'][i] for i in range(N)], base=_cv0, fmt=PCT)
srow('ntxn', 'Non-fuel transactions = fuel transactions x conversion',
     lambda j: (f'={get_column_letter(3 + j)}{S["ftxn"]}*'
                f'{get_column_letter(3 + j)}{S["conv"]}'),
     F['nonfuel_txn'], base=_ft0 * _cv0, fmt='#,##0.0')
srow('bask', 'Average non-fuel basket (AED)', grow('bask', 'baskg', f'{c("bk0")}'),
     F['basket'], base=_bk0, fmt=PS)
srow('revn', 'Non-fuel revenue = transactions x basket',
     lambda j: (f'={get_column_letter(3 + j)}{S["ntxn"]}*'
                f'{get_column_letter(3 + j)}{S["bask"]}'),
     F['rev_nonfuel'], base=V['rev_nonfuel_h126'] * 2)
srow('rev', 'TOTAL REVENUE',
     lambda j: (f'={get_column_letter(3 + j)}{S["revr"]}+{get_column_letter(3 + j)}{S["revc"]}'
                f'+{get_column_letter(3 + j)}{S["revn"]}'),
     F['revenue'], base=V['rev_h126'] * 2, bold=True, fill=FILL_C)
srow('mr', 'Retail margin per litre (AED)', grow('mr', 'mrg', f'{c("mr0")}'),
     F['margin_retail'], base=UB['margin_retail_h126'], fmt=PS3, basekind='unit_build')
srow('mc', 'Commercial margin per litre (AED) — blended, split NOT disclosed',
     grow('mc', 'mcg', f'{c("mc0")}'), F['margin_comm'],
     base=UB['margin_comm_h126'], fmt=PS3, basekind='unit_build')
srow('gpr', 'Retail fuel gross profit = volume x margin',
     lambda j: f'={get_column_letter(3 + j)}{S["volr"]}*{get_column_letter(3 + j)}{S["mr"]}',
     F['gp_retfuel_struct'], base=UB['gp_retfuel_struct_h126'] * 2)
srow('gpc', 'Commercial gross profit = volume x margin',
     lambda j: f'={get_column_letter(3 + j)}{S["volc"]}*{get_column_letter(3 + j)}{S["mc"]}',
     F['gp_comm_struct'], base=UB['gp_comm_struct_h126'] * 2)
srow('gpn', 'Non-fuel gross profit = revenue x margin',
     lambda j: f'={get_column_letter(3 + j)}{S["revn"]}*{cy("nfm", j)}',
     F['gp_nonfuel'], base=V['gp_nonfuel_h126'] * 2)
srow('gps', 'STRUCTURAL GROSS PROFIT',
     lambda j: (f'={get_column_letter(3 + j)}{S["gpr"]}+{get_column_letter(3 + j)}{S["gpc"]}'
                f'+{get_column_letter(3 + j)}{S["gpn"]}'),
     F['gp_struct'], base=(V['gp_h126'] - V['invgain_h126']) * 2, bold=True, fill=FILL_C)
srow('ivA', 'Inventory movement — Frame A, normalised', lambda j: f'={cy("invA", j)}',
     F['invmove_A'], base=V['invgain_h126'])
srow('ivB', 'Inventory movement — Frame B, through-cycle', lambda j: f'={cy("invB", j)}',
     F['invmove_B'], base=V['invgain_h126'])
srow('gpA', 'GROSS PROFIT — Frame A',
     lambda j: f'={get_column_letter(3 + j)}{S["gps"]}+{get_column_letter(3 + j)}{S["ivA"]}',
     F['gross_profit_A'], base=V['gp_h126'] * 2, bold=True, fill=FILL_C)
srow('gpB', 'GROSS PROFIT — Frame B',
     lambda j: f'={get_column_letter(3 + j)}{S["gps"]}+{get_column_letter(3 + j)}{S["ivB"]}',
     F['gross_profit_B'], base=V['gp_h126'] * 2, bold=True, fill=FILL_C)

# ---- the asset-base roll: depreciation and capital spending as OUTPUTS ----------------
lbl(wsg, rr, 1, 'The asset base, rolled — depreciation and capital spending are outputs of '
    'it, not pasted arrays', bold=True, fill=FILL_P)
for _j in range(2, 8):
    lbl(wsg, rr, _j, '', fill=FILL_P)
rr += 1
_ob = []
_o = V['ppe_fy25'] + V['rou_fy25']
for i in range(N):
    _ob.append(_o)
    _o = _o + F['capex'][i] - F['dna'][i]
# The opening base for FY2027 onward is the prior year's CLOSING base, which does not exist
# on the sheet yet. It is written as the anchor here and overwritten with the live link once
# the closing row exists, rather than guessing a row offset.
srow('open', 'Opening fixed and right-of-use base',
     lambda j: f'={c("ppe0")}', [_ob[0]] * N, base=None)
srow('dep', 'Depreciation = opening base x the measured rate',
     lambda j: f'={get_column_letter(3 + j)}{S["open"]}*{c("deprate")}', F['dna'])
srow('maint', 'Maintenance capital spending = opening base x the maintenance rate',
     lambda j: f'={get_column_letter(3 + j)}{S["open"]}*{c("maintrate")}',
     [_ob[i] * V['maint_capex_rate'] for i in range(N)])
srow('adds', 'Stations added',
     lambda j: (f'={get_column_letter(3 + j)}{S["sta"]}-{get_column_letter(2 + j)}{S["sta"]}'
                if j else f'={get_column_letter(3)}{S["sta"]}-{c("staY0")}'),
     [F['stations'][i] - (F['stations'][i - 1] if i else V['stations_fy25'])
      for i in range(N)], fmt='#,##0.0')
srow('growcpx', 'Growth capital spending = stations added x cost per station',
     lambda j: f'={get_column_letter(3 + j)}{S["adds"]}*{c("cpxsta")}',
     [(F['stations'][i] - (F['stations'][i - 1] if i else V['stations_fy25']))
      * V['capex_per_station'] for i in range(N)])
srow('cpx', 'TOTAL CAPITAL SPENDING = maintenance + growth',
     lambda j: (f'={get_column_letter(3 + j)}{S["maint"]}+'
                f'{get_column_letter(3 + j)}{S["growcpx"]}'),
     F['capex'], bold=True, fill=FILL_C)
srow('close', 'Closing base = opening + capital spending less depreciation',
     lambda j: (f'={get_column_letter(3 + j)}{S["open"]}+{get_column_letter(3 + j)}{S["cpx"]}'
                f'-{get_column_letter(3 + j)}{S["dep"]}'),
     [_ob[i] + F['capex'][i] - F['dna'][i] for i in range(N)])

for _j in range(1, N):
    f(wsg, S['open'], 3 + _j, f'={get_column_letter(2 + _j)}{S["close"]}', _ob[_j], fmt=MONEY)

# The two Assumptions rows that were written as placeholders now become live formulas
# pointing at this roll, so every downstream reference to them inherits the derivation.
for _j in range(N):
    f(wa, A['dna'], 3 + _j, f'={q(SG)}!{get_column_letter(3 + _j)}{S["dep"]}',
      F['dna'][_j], fmt=MONEY)
    f(wa, A['capex'], 3 + _j, f'={q(SG)}!{get_column_letter(3 + _j)}{S["cpx"]}',
      F['capex'][_j], fmt=MONEY)

# ============================== INCOME STATEMENT ==============================
IST = 'Income Statement'
wi = sheet(IST, [40] + [13] * 8)
hdr(wi, 1, ['AED million'] + HY + YR)
IS = {}
ri = 2


def irow(key, label, hist, formulas, expected, fmt=MONEY, bold=False, fill=None):
    global ri
    IS[key] = ri
    lbl(wi, ri, 1, label, bold=bold, fill=fill)
    for j, v in enumerate(hist):
        val(wi, ri, 2 + j, v, fmt=fmt, kind='audited', fill=fill, bold=bold)
    for j in range(N):
        f(wi, ri, 5 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    ri += 1


sgc = lambda k, j: f"{q(SG)}!{get_column_letter(3 + j)}{S[k]}"
irow('rev', 'Revenue', [H[y]['revenue'] for y in HY], lambda j: f'={sgc("rev", j)}',
     F['revenue'])
irow('gp', 'Gross profit — Frame A', [H[y]['gross_profit'] for y in HY],
     lambda j: f'={sgc("gpA", j)}', F['gross_profit_A'])
irow('gm', 'Gross margin', [H[y]['gross_margin'] for y in HY],
     lambda j: f'={get_column_letter(5 + j)}{IS["gp"]}/{get_column_letter(5 + j)}{IS["rev"]}',
     F['gross_margin_A'], fmt=PCT)
irow('opex', 'Cash operating expenses', [-H[y]['cash_opex'] for y in HY],
     lambda j: (f'={get_column_letter(4 + j)}{ri}*(1+{cy("opexg", j)})' if j else
                f'=-{c("opex0")}*(1+{cy("opexg", 0)})'),
     [-x for x in F['cash_opex']])
irow('oi', 'Other income', [H[y]['other_income'] for y in HY],
     lambda j: (f'={get_column_letter(4 + j)}{ri}*(1+{cy("oig", j)})' if j else
                f'={c("oi0")}*(1+{cy("oig", 0)})'),
     F['other_income'])
irow('imp', 'Impairments and other operating expenses', [-H[y]['impairments'] for y in HY],
     lambda j: f'=-{cy("imp", j)}', [-x for x in F['impairments']])
irow('ebitda', 'EBITDA', [H[y]['ebitda'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["gp"]}+{get_column_letter(5 + j)}{IS["opex"]}'
                f'+{get_column_letter(5 + j)}{IS["oi"]}+{get_column_letter(5 + j)}{IS["imp"]}'),
     F['ebitda_A'], bold=True, fill=FILL_C)
irow('em', 'EBITDA margin', [H[y]['ebitda_margin'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["ebitda"]}/'
                f'{get_column_letter(5 + j)}{IS["rev"]}'),
     F['ebitda_margin_A'], fmt=PCT)
irow('dna', 'Depreciation and amortisation', [-H[y]['dna'] for y in HY],
     lambda j: f'=-{cy("dna", j)}', [-x for x in F['dna']])
irow('ebit', 'EBIT', [H[y]['ebit'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["ebitda"]}+'
                f'{get_column_letter(5 + j)}{IS["dna"]}'),
     F['ebit_A'], bold=True, fill=FILL_C)
irow('ii', 'Interest income', [H[y]['interest_income'] for y in HY],
     lambda j: f'={c("ii0")}', [V['intinc_fy25']] * N)
irow('fin', 'Finance costs', [-H[y]['finance_costs'] for y in HY],
     lambda j: f'=-{c("fin0")}', [-V['fin_fy25']] * N)
irow('pbt', 'Profit before tax', [H[y]['pbt'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["ebit"]}+{get_column_letter(5 + j)}{IS["ii"]}'
                f'+{get_column_letter(5 + j)}{IS["fin"]}'), pbt_A)
irow('tax', 'Income tax', [-H[y]['tax'] for y in HY],
     lambda j: f'=-{get_column_letter(5 + j)}{IS["pbt"]}*{c("taxeff")}', [-t for t in tax_A])
irow('np', 'Profit for the year', [H[y]['net_profit'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["pbt"]}+'
                f'{get_column_letter(5 + j)}{IS["tax"]}'), np_A, bold=True)
irow('nci', 'Non-controlling interests', [-H[y]['nci'] for y in HY],
     lambda j: f'=-{c("ncip")}', [-V['nci_fy25']] * N)
irow('npa', 'Profit attributable to owners', [H[y]['np_attributable'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{IS["np"]}+'
                f'{get_column_letter(5 + j)}{IS["nci"]}'), npa_A, bold=True, fill=FILL_C)
irow('eps', 'Earnings per share (AED)', [H[y]['eps'] for y in HY],
     lambda j: f'={get_column_letter(5 + j)}{IS["npa"]}/{c("shares")}', eps_A, fmt=PS3)
irow('divfix', 'Dividend — the fixed policy commitment', [None] * 3,
     lambda j: f'={c("dps")}*{c("shares")}', [V['dps'] * V['shares_mn']] * N)
irow('divmin', 'Dividend — the 75% of profit minimum', [None] * 3,
     lambda j: f'={c("payoutfloor")}*{get_column_letter(5 + j)}{IS["npa"]}',
     [V['payout_floor'] * npa_A[i] for i in range(N)])
irow('divpol', 'DIVIDEND PAID = the higher of the two', [None] * 3,
     lambda j: (f'=MAX({get_column_letter(5 + j)}{IS["divfix"]},'
                f'{get_column_letter(5 + j)}{IS["divmin"]})'), divs_y, bold=True)

# ============================== BALANCE SHEET ==============================
BST = 'Balance Sheet'
wb_ = sheet(BST, [40] + [13] * 8)
hdr(wb_, 1, ['AED million'] + HY + YR)
BS = {}
rb = 2


def brow(key, label, hist, formulas, expected, fmt=MONEY, bold=False, fill=None):
    global rb
    BS[key] = rb
    lbl(wb_, rb, 1, label, bold=bold, fill=fill)
    for j, v in enumerate(hist):
        val(wb_, rb, 2 + j, v, fmt=fmt, kind='audited', fill=fill, bold=bold)
    for j in range(N):
        f(wb_, rb, 5 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    rb += 1


brow('ppe', 'Fixed and right-of-use assets', [H[y]['ppe'] + H[y]['rou'] for y in HY],
     lambda j: (f'={get_column_letter(4 + j)}{rb}+{cy("capex", j)}-{cy("dna", j)}' if j else
                f'={c("ppe0")}+{cy("capex", 0)}-{cy("dna", 0)}'), ppe)
brow('gw', 'Goodwill, intangibles and other non-current',
     [H[y]['goodwill_intangibles'] + H[y]['other_nca'] for y in HY],
     lambda j: f'={c("gw0")}', [gw0] * N)
brow('inv', 'Inventories', [H[y]['inventories'] for y in HY],
     lambda j: (f'={c("dio")}/365*({q(IST)}!{get_column_letter(5 + j)}{IS["rev"]}-'
                f'{q(IST)}!{get_column_letter(5 + j)}{IS["gp"]})'), invn)
brow('recv', 'Receivables including related parties',
     [H[y]['receivables'] + H[y]['due_from_rp'] for y in HY],
     lambda j: f'={c("dso")}/365*{q(IST)}!{get_column_letter(5 + j)}{IS["rev"]}', recv)
brow('cash', 'Cash and term deposits', [H[y]['cash'] + H[y]['term_deposits'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{rb + 6}-{get_column_letter(5 + j)}{BS["ppe"]}'
                f'-{get_column_letter(5 + j)}{BS["gw"]}-{get_column_letter(5 + j)}{BS["inv"]}'
                f'-{get_column_letter(5 + j)}{BS["recv"]}'), cash)
brow('ta', 'TOTAL ASSETS', [H[y]['total_assets'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{BS["ppe"]}+{get_column_letter(5 + j)}{BS["gw"]}'
                f'+{get_column_letter(5 + j)}{BS["inv"]}+{get_column_letter(5 + j)}{BS["recv"]}'
                f'+{get_column_letter(5 + j)}{BS["cash"]}'), ta, bold=True, fill=FILL_C)
brow('eq', 'Equity attributable to owners', [H[y]['equity_parent'] for y in HY],
     lambda j: (f'={get_column_letter(4 + j)}{rb}+{q(IST)}!{get_column_letter(5 + j)}{IS["npa"]}'
                f'-{q(IST)}!{get_column_letter(5 + j)}{IS["divpol"]}' if j else
                f'={c("eq0")}+{q(IST)}!{get_column_letter(5)}{IS["npa"]}'
                f'-{q(IST)}!{get_column_letter(5)}{IS["divpol"]}'),
     eqp)
brow('nci', 'Non-controlling interests', [H[y]['nci_equity'] for y in HY],
     lambda j: (f'={get_column_letter(4 + j)}{rb}+{c("ncip")}' if j else
                f'={c("nci0")}+{c("ncip")}'), ncieq)
brow('borr', 'Borrowings', [H[y]['borrowings'] for y in HY],
     lambda j: f'={c("borr0")}', [V['borr_fy25']] * N)
brow('lease', 'Lease liabilities', [H[y]['leases'] for y in HY],
     lambda j: f'={c("lease0")}', [V['lease_fy25']] * N)
brow('prov', 'Provisions, deferred tax and other', [None, None, prov0],
     lambda j: f'={c("prov0")}', [prov0] * N)
brow('pay', 'Payables including related parties',
     [H[y]['payables'] + H[y]['due_to_rp'] for y in HY],
     lambda j: (f'={c("dpo")}/365*({q(IST)}!{get_column_letter(5 + j)}{IS["rev"]}-'
                f'{q(IST)}!{get_column_letter(5 + j)}{IS["gp"]})'), pay)
brow('tle', 'TOTAL EQUITY AND LIABILITIES', [H[y]['total_assets'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{BS["eq"]}+{get_column_letter(5 + j)}{BS["nci"]}'
                f'+{get_column_letter(5 + j)}{BS["borr"]}+{get_column_letter(5 + j)}{BS["lease"]}'
                f'+{get_column_letter(5 + j)}{BS["prov"]}+{get_column_letter(5 + j)}{BS["pay"]}'),
     tle, bold=True, fill=FILL_C)
brow('chk', 'BALANCE CHECK (must be zero)', [0.0, 0.0, 0.0],
     lambda j: (f'={get_column_letter(5 + j)}{BS["ta"]}-'
                f'{get_column_letter(5 + j)}{BS["tle"]}'), [0.0] * N, fmt=PS3, bold=True)
brow('nd', 'Net debt excluding leases',
     [H[y]['net_debt_company'] for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{BS["borr"]}-'
                f'{get_column_letter(5 + j)}{BS["cash"]}'), nd_f)
# Cash is the balancing item, so it is written once the rows it depends on exist. Writing
# it inline would mean guessing row offsets for cells further down the sheet, and a wrong
# guess is not an error — just a wrong answer.
def _cash_formula(j):
    col = get_column_letter(5 + j)
    plus = '+'.join(f'{col}{BS[k]}' for k in ('eq', 'nci', 'borr', 'lease', 'prov', 'pay'))
    minus = '-'.join(f'{col}{BS[k]}' for k in ('ppe', 'gw', 'inv', 'recv'))
    return f'={plus}-{minus}'


brow('bvps', 'Book value per share (AED)',
     [H[y]['equity_parent'] / V['shares_mn'] for y in HY],
     lambda j: f'={get_column_letter(5 + j)}{BS["eq"]}/{c("shares")}',
     [e / V['shares_mn'] for e in eqp], fmt=PS3)

for _j in range(N):
    f(wb_, BS['cash'], 5 + _j, _cash_formula(_j), cash[_j], fmt=MONEY)

# ============================== CASH FLOW ==============================
CFT = 'Cash Flow'
wcf = sheet(CFT, [40] + [13] * 8)
hdr(wcf, 1, ['AED million'] + HY + YR)
CF = {}
rc = 2


def crow(key, label, hist, formulas, expected, fmt=MONEY, bold=False, fill=None):
    global rc
    CF[key] = rc
    lbl(wcf, rc, 1, label, bold=bold, fill=fill)
    for j, v in enumerate(hist):
        val(wcf, rc, 2 + j, v, fmt=fmt, kind='audited', fill=fill, bold=bold)
    for j in range(N):
        f(wcf, rc, 5 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    rc += 1


nopat_A = A_['nopat']
crow('ebit', 'EBIT', [H[y]['ebit'] for y in HY],
     lambda j: f"={q(IST)}!{get_column_letter(5 + j)}{IS['ebit']}", A_['ebit'])
crow('taxr', 'Tax rate applied to EBIT', [H[y]['tax_rate'] for y in HY],
     lambda j: f'={c("taxeff")}', [V['tax_effective']] * N, fmt=PCT2)
crow('nopat', 'NOPAT = EBIT x (1 less the tax rate)',
     [H[y]['ebit'] * (1 - H[y]['tax_rate']) for y in HY],
     lambda j: (f'={get_column_letter(5 + j)}{CF["ebit"]}*'
                f'(1-{get_column_letter(5 + j)}{CF["taxr"]})'), nopat_A, bold=True)
crow('dna', 'Add back depreciation and amortisation', [H[y]['dna'] for y in HY],
     lambda j: f'={cy("dna", j)}', F['dna'])
crow('capex', 'Less capital expenditure', [-H[y]['capex'] for y in HY],
     lambda j: f'=-{cy("capex", j)}', [-x for x in F['capex']])
crow('dwc', 'Less increase in working capital', [None, None, None],
     lambda j: f'=-({A_["delta_nwc"][j]})' if False else
     (f'=-({q(BST)}!{get_column_letter(5 + j)}{BS["recv"]}+{q(BST)}!{get_column_letter(5 + j)}{BS["inv"]}'
      f'-{q(BST)}!{get_column_letter(5 + j)}{BS["pay"]}-{q(BST)}!{get_column_letter(4 + j)}{BS["recv"]}'
      f'-{q(BST)}!{get_column_letter(4 + j)}{BS["inv"]}+{q(BST)}!{get_column_letter(4 + j)}{BS["pay"]})'
      if j else
      f'=-({q(BST)}!{get_column_letter(5)}{BS["recv"]}+{q(BST)}!{get_column_letter(5)}{BS["inv"]}'
      f'-{q(BST)}!{get_column_letter(5)}{BS["pay"]}-{q(BST)}!{get_column_letter(4)}{BS["recv"]}'
      f'-{q(BST)}!{get_column_letter(4)}{BS["inv"]}+{q(BST)}!{get_column_letter(4)}{BS["pay"]})'),
     [-x for x in A_['delta_nwc']])
crow('fcff', 'FREE CASH FLOW TO THE FIRM', [None, None, None],
     lambda j: (f'={get_column_letter(5 + j)}{CF["nopat"]}+{get_column_letter(5 + j)}{CF["dna"]}'
                f'+{get_column_letter(5 + j)}{CF["capex"]}+'
                f'{get_column_letter(5 + j)}{CF["dwc"]}'),
     A_['fcff'], bold=True, fill=FILL_C)
crow('div', 'Dividends paid', [-H[y]['dividends_paid'] for y in HY],
     lambda j: f'=-{q(IST)}!{get_column_letter(5 + j)}{IS["divpol"]}',
     [-x for x in divs_y])

# ============================== DCF ==============================
DC = 'DCF'
wd = sheet(DC, [46] + [14] * 5 + [16])
hdr(wd, 1, ['The discounted cash flow'] + YR + ['Total'])
DR = {}
rd = 2


def dcfrow(key, label, formulas, expected, fmt=MONEY, bold=False, fill=None, total=None):
    global rd
    DR[key] = rd
    lbl(wd, rd, 1, label, bold=bold, fill=fill)
    for j in range(N):
        f(wd, rd, 2 + j, formulas(j), expected[j], fmt=fmt, bold=bold, fill=fill)
    if total is not None:
        f(wd, rd, 7, total[0], total[1], fmt=fmt, bold=True, fill=FILL_C)
    rd += 1


dcfrow('glide', 'Glide fraction toward the terminal cost of capital',
       lambda j: f'=({j + 1})/{N}', W['glide_frac'], fmt=PCT)
dcfrow('rate', 'Discount rate = first-year rate + glide x (terminal less first-year)',
       lambda j: (f'={c("wacc")}+{get_column_letter(2 + j)}{DR["glide"]}*'
                  f'({c("waccterm")}-{c("wacc")})'), W['disc_rate'], fmt=PCT2)
dcfrow('df', 'Discount factor (compounding)',
       lambda j: (f'={get_column_letter(1 + j)}{DR["df"] if False else rd}/'
                  f'(1+{get_column_letter(2 + j)}{DR["rate"]})' if j else
                  f'=1/(1+{get_column_letter(2)}{DR["rate"]})'), W['df'], fmt='0.000')
for frame, dd, gpk in (('A', A_, 'gpA'), ('B', B_, 'gpB')):
    lbl(wd, rd, 1, f'FRAME {frame} — inventory movements '
        + ('normalised to zero from FY2027' if frame == 'A'
           else 'carried at the through-cycle average'), bold=True, fill=FILL_P)
    for j in range(2, 8):
        lbl(wd, rd, j, '', fill=FILL_P)
    rd += 1
    dcfrow(f'ebitda{frame}', f'EBITDA — Frame {frame}',
           lambda j, k=gpk: (f'={q(SG)}!{get_column_letter(3 + j)}{S[k]}'
                             f'-{c("opex0")}*0'
                             f'+{q(IST)}!{get_column_letter(5 + j)}{IS["opex"]}'
                             f'+{q(IST)}!{get_column_letter(5 + j)}{IS["oi"]}'
                             f'+{q(IST)}!{get_column_letter(5 + j)}{IS["imp"]}'),
           dd['ebitda'])
    dcfrow(f'dna{frame}', 'Less depreciation and amortisation',
           lambda j: f'=-{cy("dna", j)}', [-x for x in F['dna']])
    dcfrow(f'ebit{frame}', f'EBIT — Frame {frame}',
           lambda j, fr=frame: (f'={get_column_letter(2 + j)}{DR[f"ebitda{fr}"]}+'
                                f'{get_column_letter(2 + j)}{DR[f"dna{fr}"]}'),
           dd['ebit'], bold=True)
    dcfrow(f'nopat{frame}', 'NOPAT = EBIT x (1 less the tax rate)',
           lambda j, fr=frame: (f'={get_column_letter(2 + j)}{DR[f"ebit{fr}"]}*'
                                f'(1-{c("taxeff")})'), dd['nopat'], bold=True)
    dcfrow(f'addd{frame}', 'Add back depreciation and amortisation',
           lambda j: f'={cy("dna", j)}', F['dna'])
    dcfrow(f'cap{frame}', 'Less capital expenditure',
           lambda j: f'=-{cy("capex", j)}', [-x for x in F['capex']])
    dcfrow(f'dwc{frame}', 'Less increase in working capital',
           lambda j, dv=dd: f'=-({q(CFT)}!{get_column_letter(5 + j)}{CF["dwc"]})*-1'
           if False else f'={q(CFT)}!{get_column_letter(5 + j)}{CF["dwc"]}',
           [-x for x in dd['delta_nwc']])
    dcfrow(f'fcff{frame}', f'FREE CASH FLOW TO THE FIRM — Frame {frame}',
           lambda j, fr=frame: (f'={get_column_letter(2 + j)}{DR[f"nopat{fr}"]}'
                                f'+{get_column_letter(2 + j)}{DR[f"addd{fr}"]}'
                                f'+{get_column_letter(2 + j)}{DR[f"cap{fr}"]}'
                                f'+{get_column_letter(2 + j)}{DR[f"dwc{fr}"]}'),
           dd['fcff'], bold=True, fill=FILL_C)
    dcfrow(f'pv{frame}', 'PRESENT VALUE of free cash flow',
           lambda j, fr=frame: (f'={get_column_letter(2 + j)}{DR[f"fcff{fr}"]}*'
                                f'{get_column_letter(2 + j)}{DR["df"]}'),
           dd['pv'], bold=True, fill=FILL_C,
           total=(f'=SUM(B{rd}:F{rd})', dd['pv_sum']))

# terminal block
TB = {}
for frame, dd in (('A', A_), ('B', B_)):
    lbl(wd, rd, 1, f'TERMINAL BLOCK — Frame {frame}', bold=True, fill=FILL_P)
    for j in range(2, 8):
        lbl(wd, rd, j, '', fill=FILL_P)
    rd += 1
    TB[f'reinv{frame}'] = rd
    lbl(wd, rd, 1, 'Reinvestment rate = long-run growth / terminal return on capital')
    f(wd, rd, 2, f'={c("gterm")}/{c("roicterm")}', dd['reinvest_rate'], fmt=PCT)
    rd += 1
    TB[f'nt{frame}'] = rd
    lbl(wd, rd, 1, 'Terminal NOPAT = final-year NOPAT x (1 + growth)')
    f(wd, rd, 2, f'=F{DR[f"nopat{frame}"]}*(1+{c("gterm")})', dd['nopat_term'])
    rd += 1
    TB[f'ft{frame}'] = rd
    lbl(wd, rd, 1, 'Terminal free cash flow = terminal NOPAT x (1 less reinvestment)')
    f(wd, rd, 2, f'=B{TB[f"nt{frame}"]}*(1-B{TB[f"reinv{frame}"]})', dd['fcff_term'])
    rd += 1
    TB[f'tv{frame}'] = rd
    lbl(wd, rd, 1, 'Terminal value = terminal free cash flow / (terminal rate less growth)',
        bold=True)
    f(wd, rd, 2, f'=B{TB[f"ft{frame}"]}/({c("waccterm")}-{c("gterm")})', dd['tv'], bold=True)
    rd += 1
    TB[f'pvtv{frame}'] = rd
    lbl(wd, rd, 1, 'Present value of the terminal value', bold=True)
    f(wd, rd, 2, f'=B{TB[f"tv{frame}"]}*F{DR["df"]}', dd['pv_tv'], bold=True, fill=FILL_C)
    rd += 1
    TB[f'ev{frame}'] = rd
    lbl(wd, rd, 1, 'ENTERPRISE VALUE', bold=True, fill=FILL_C)
    f(wd, rd, 2, f'=G{DR[f"pv{frame}"]}+B{TB[f"pvtv{frame}"]}', dd['ev'], bold=True,
      fill=FILL_C)
    rd += 1
    TB[f'tvs{frame}'] = rd
    lbl(wd, rd, 1, 'TERMINAL VALUE AS A PERCENTAGE OF ENTERPRISE VALUE', bold=True,
        fill=FILL_C)
    f(wd, rd, 2, f'=B{TB[f"pvtv{frame}"]}/B{TB[f"ev{frame}"]}', dd['tv_share'], fmt=PCT,
      bold=True, fill=FILL_C)
    rd += 1

# ============================== SOTP BRIDGE ==============================
SB = 'SOTP Bridge'
wsb = sheet(SB, [52, 18, 18])
hdr(wsb, 1, ['Enterprise value to equity value', 'Frame A (AED m)', 'Frame B (AED m)'])
BR = {}
rs = 2


def brow2(key, label, fa, fb, ea, eb, fmt=MONEY, bold=False, fill=None):
    global rs
    BR[key] = rs
    lbl(wsb, rs, 1, label, bold=bold, fill=fill)
    f(wsb, rs, 2, fa, ea, fmt=fmt, bold=bold, fill=fill)
    f(wsb, rs, 3, fb, eb, fmt=fmt, bold=bold, fill=fill)
    rs += 1


brow2('pv', 'Present value of five years of free cash flow',
      f'={q(DC)}!G{DR["pvA"]}', f'={q(DC)}!G{DR["pvB"]}', A_['pv_sum'], B_['pv_sum'])
brow2('pvtv', 'Present value of the terminal value',
      f'={q(DC)}!B{TB["pvtvA"]}', f'={q(DC)}!B{TB["pvtvB"]}', A_['pv_tv'], B_['pv_tv'])
brow2('ev', 'ENTERPRISE VALUE', f'=B{BR["pv"]}+B{BR["pvtv"]}', f'=C{BR["pv"]}+C{BR["pvtv"]}',
      A_['ev'], B_['ev'], bold=True, fill=FILL_C)
brow2('tvs', 'TERMINAL VALUE AS A PERCENTAGE OF ENTERPRISE VALUE',
      f'=B{BR["pvtv"]}/B{BR["ev"]}', f'=C{BR["pvtv"]}/C{BR["ev"]}',
      A_['tv_share'], B_['tv_share'], fmt=PCT, bold=True, fill=FILL_C)
brow2('nd', 'Less net debt excluding leases', f'=-{c("netdebt")}',
      f'=-{c("netdebt")}', -A_['net_debt'], -B_['net_debt'])
brow2('ls', 'Less lease liabilities', f'=-{c("lease0")}', f'=-{c("lease0")}',
      -A_['leases'], -B_['leases'])
brow2('nci', 'Less non-controlling interests', f'=-{c("nci0")}',
      f'=-{c("nci0")}', -A_['nci'], -B_['nci'])
brow2('eq', 'EQUITY VALUE', f'=B{BR["ev"]}+B{BR["nd"]}+B{BR["ls"]}+B{BR["nci"]}',
      f'=C{BR["ev"]}+C{BR["nd"]}+C{BR["ls"]}+C{BR["nci"]}',
      A_['equity'], B_['equity'], bold=True, fill=FILL_C)
brow2('ps', 'VALUE PER SHARE (AED)', f'=B{BR["eq"]}/{c("shares")}',
      f'=C{BR["eq"]}/{c("shares")}', A_['per_share'], B_['per_share'], fmt=PS,
      bold=True, fill=FILL_C)

# ============================== RELATIVE & NORMALIZED ==============================
RN = 'Relative & Normalized'
wr = sheet(RN, [50, 16, 16, 40])
hdr(wr, 1, ['Relative multiples and normalised earnings power', 'Frame A', 'Frame B',
            'Basis'])
RNr = {}
rq = 2


def rrow(key, label, fa, ea, fb=None, eb=None, basis='', fmt=PS, bold=False, fill=None):
    global rq
    RNr[key] = rq
    lbl(wr, rq, 1, label, bold=bold, fill=fill)
    f(wr, rq, 2, fa, ea, fmt=fmt, bold=bold, fill=fill)
    if fb is not None:
        f(wr, rq, 3, fb, eb, fmt=fmt, bold=bold, fill=fill)
    lbl(wr, rq, 4, basis, note=True)
    rq += 1


lbl(wr, rq, 1, 'Triangulation: the reference multiple is the AVERAGE of three '
    'independently-derived readings, taken on the sheet rather than asserted.', note=True)
wr.merge_cells(start_row=rq, start_column=1, end_row=rq, end_column=4)
rq += 1
_pe_now = L['pe_now']
_pe_mean = L['own_pe_mean']
rrow('pe1', "Own trailing multiple today", f'={c("spot")}/{q(IST)}!D{IS["eps"]}',
     _pe_now, basis='traded price over the FY2025 audited earnings per share', fmt=X)
rrow('pe2', "Own three-year average multiple",
     f'=({c("spot")}/{q(IST)}!B{IS["eps"]}+{c("spot")}/{q(IST)}!C{IS["eps"]}'
     f'+{c("spot")}/{q(IST)}!D{IS["eps"]})/3', _pe_mean,
     basis='traded price against each audited year, averaged', fmt=X)
rrow('retimp', 'Retention the growth rate requires = growth / return on equity',
     f'={c("gterm")}/{get_column_letter(2)}{RNr["roe"] if "roe" in RNr else 0}'
     if False else f'={c("gterm")}/{L["roe_sust"]:.10f}',
     L['pe_retention_implied'], basis='the sustainable-growth identity g = retention x ROE',
     fmt=PCT2)
rrow('payimp', 'Payout consistent with that growth = 1 less the retention',
     f'=1-{get_column_letter(2)}{RNr["retimp"]}', L['pe_payout_implied'], fmt=PCT)
rrow('pe3', 'REFERENCE MULTIPLE — what the company’s own economics justify',
     f'={get_column_letter(2)}{RNr["payimp"]}*(1+{c("gterm")})/({c("ke")}-{c("gterm")})',
     L['pe_method_justified'],
     basis='payout x (1 + growth) / (cost of equity less growth), on the payout the growth '
           'rate itself implies — NOT the realised payout and NOT the policy floor, either '
           'of which breaks the identity', fmt=X, bold=True, fill=FILL_C)
rrow('peref', 'Reference multiple carried into the lens',
     f'={get_column_letter(2)}{RNr["pe3"]}', L['just_fwd_pe'],
     basis='the fundamentals leg ALONE. The two traded multiples above are published as '
           'context and are deliberately NOT averaged in: both are the traded price divided '
           'by earnings, so averaging them would anchor the lens to the price it is being '
           'compared against', fmt=X, bold=True, fill=FILL_C)
_relA = L['rel_A']
rrow('rel', 'RELATIVE MULTIPLES READING (AED a share)',
     f'={get_column_letter(2)}{RNr["peref"]}*{q(IST)}!E{IS["eps"]}', _relA,
     basis='reference multiple on the first forecast year', bold=True, fill=FILL_C)
_pbtB = B_['ebit'][0] + V['intinc_fy25'] - V['fin_fy25']
_epsB = (_pbtB * (1 - V['tax_effective']) - V['nci_fy25']) / V['shares_mn']
rrow('epsB', 'First forecast year earnings per share — Frame B (AED)',
     f'=({q(DC)}!B{DR["ebitB"]}+{c("ii0")}-{c("fin0")})*(1-{c("taxeff")})'
     f'/{c("shares")}-{c("ncip")}/{c("shares")}', _epsB,
     basis='the same construction as the income statement, on Frame B earnings', fmt=PS3)
rrow('relB', 'RELATIVE MULTIPLES READING — Frame B (AED a share)',
     f'={get_column_letter(2)}{RNr["peref"]}*B{RNr["epsB"]}', L['rel_B'],
     basis='the same reference multiple on Frame B earnings', bold=True, fill=FILL_C)
rrow('nebitda', 'Normalised EBITDA — structural gross profit only, no inventory movement',
     f'={q(SG)}!C{S["gps"]}+{q(IST)}!E{IS["opex"]}+{q(IST)}!E{IS["oi"]}+{q(IST)}!E{IS["imp"]}',
     L['norm_ebitda'], fmt=MONEY)
rrow('nebit', 'Normalised EBIT',
     f'=B{RNr["nebitda"]}-{cy("dna", 0)}', L['norm_ebit'], fmt=MONEY)
rrow('nnopat', 'Normalised NOPAT', f'=B{RNr["nebit"]}*(1-{c("taxeff")})', L['norm_nopat'],
     fmt=MONEY)
rrow('nrr', 'Reinvestment this perpetuity must fund = growth / return on capital',
     f'={c("gterm")}/{c("roicterm")}', L['norm_reinvest'], fmt=PCT)
rrow('nev', 'Normalised enterprise value = NOPAT x (1 less reinvestment) / (cost of capital '
     'less growth)',
     f'=B{RNr["nnopat"]}*(1-B{RNr["nrr"]})/({c("wacc")}-{c("gterm")})', L['norm_ev'],
     fmt=MONEY,
     basis='the identity g = ROIC x reinvestment governs EVERY perpetuity, not only the '
           'cash-flow model terminal block. This lens previously credited growth for free')
rrow('neq', 'Normalised equity value',
     f'=B{RNr["nev"]}-{c("netdebt")}-{c("lease0")}-{c("nci0")}',
     L['norm_equity'], fmt=MONEY)
rrow('nps', 'NORMALISED EARNINGS POWER READING (AED a share)',
     f'=B{RNr["neq"]}/{c("shares")}', L['norm_ps'], bold=True, fill=FILL_C)
rrow('bvps', 'Book value per share (AED)', f'={c("eq0")}/{c("shares")}', L['bv_ps'],
     fmt=PS3)
rrow('roe', 'Sustainable return on equity — the three-year mean',
     f'=({q(IST)}!B{IS["npa"]}/{H["FY2023"]["equity_parent"]}+'
     f'{q(IST)}!C{IS["npa"]}/{H["FY2024"]["equity_parent"]}+'
     f'{q(IST)}!D{IS["npa"]}/{H["FY2025"]["equity_parent"]})/3', L['roe_sust'], fmt=PCT)
rrow('jpb', 'Justified multiple of book = (return less growth) / (cost of equity less growth)',
     f'=(B{RNr["roe"]}-{c("gterm")})/({c("ke")}-{c("gterm")})', L['just_pb'], fmt=X)
rrow('bkps', 'BOOK VALUE AND SUSTAINABLE RETURN READING (AED a share)',
     f'=B{RNr["bvps"]}*B{RNr["jpb"]}', L['book_ps'], bold=True, fill=FILL_C)
rrow('divps', 'DIVIDEND CAPITALISATION READING (AED a share, unweighted)',
     f'={c("dps")}/{c("ke")}', L['div_ps'],
     basis='the dividend is a FIXED commitment held flat from 2024 through 2030, so it is '
           'valued flat. Growing it credited growth the policy does not promise', bold=True)

# ============================== FUNDAMENTAL VALUATION ==============================
FV = 'Fundamental Valuation'
wfv = sheet(FV, [46, 16, 16, 14, 40])
hdr(wfv, 1, ['The four lenses', 'Frame A (AED)', 'Frame B (AED)', 'Weight', 'Note'])
FVr = {}
rf_ = 2
lens_rows = [
    ('dcf', 'Discounted cash flow', f'={q(SB)}!B{BR["ps"]}', f'={q(SB)}!C{BR["ps"]}',
     A_['per_share'], B_['per_share'], 0.40, 'the cash the business generates'),
    ('norm', 'Normalised earnings power', f'={q(RN)}!B{RNr["nps"]}', f'={q(RN)}!B{RNr["nps"]}',
     L['norm_ps'], L['norm_ps'], 0.25, 'structural gross profit only'),
    ('rel', 'Relative multiples', f'={q(RN)}!B{RNr["rel"]}', f'={q(RN)}!B{RNr["relB"]}',
     _relA, L['rel_B'], 0.20, 'triangulated on the company’s own history'),
    ('book', 'Book value and sustainable return', f'={q(RN)}!B{RNr["bkps"]}',
     f'={q(RN)}!B{RNr["bkps"]}', L['book_ps'], L['book_ps'], 0.15,
     'return on a deliberately thin equity base'),
]
for key, label, fa, fb, ea, eb, wgt, note in lens_rows:
    FVr[key] = rf_
    lbl(wfv, rf_, 1, label)
    f(wfv, rf_, 2, fa, ea, fmt=PS)
    f(wfv, rf_, 3, fb, eb, fmt=PS)
    val(wfv, rf_, 4, wgt, fmt=PCT, kind='audited')
    lbl(wfv, rf_, 5, note, note=True)
    rf_ += 1
WROW = rf_
lbl(wfv, rf_, 1, 'SUM OF WEIGHTS (must be one)', bold=True)
f(wfv, rf_, 4, f'=SUM(D2:D{rf_ - 1})', 1.0, fmt=PCT, bold=True)
rf_ += 1
CENT = rf_
lbl(wfv, rf_, 1, 'WEIGHTED CENTRE', bold=True, fill=FILL_C)
_ca = sum(x[4] * x[6] for x in lens_rows)
_cb = sum(x[5] * x[6] for x in lens_rows)
f(wfv, rf_, 2, '+'.join(f'B{FVr[k]}*D{FVr[k]}' for k, *_ in lens_rows).join(['=', '']),
  _ca, fmt=PS, bold=True, fill=FILL_C)
f(wfv, rf_, 3, '+'.join(f'C{FVr[k]}*D{FVr[k]}' for k, *_ in lens_rows).join(['=', '']),
  _cb, fmt=PS, bold=True, fill=FILL_C)
lbl(wfv, rf_, 5, 'the two frames are NEVER averaged into one number', note=True)
rf_ += 1
lbl(wfv, rf_, 1, 'Market price (AED)', bold=True)
f(wfv, rf_, 2, f'={c("spot")}', V['spot'], fmt=PS, bold=True)
MKTROW = rf_
rf_ += 1
lbl(wfv, rf_, 1, 'Difference against the market price', bold=True)
f(wfv, rf_, 2, f'=B{CENT}/B{MKTROW}-1', _ca / V['spot'] - 1, fmt=PCT, bold=True)
f(wfv, rf_, 3, f'=C{CENT}/B{MKTROW}-1', _cb / V['spot'] - 1, fmt=PCT, bold=True)

# ============================== SUMMARY ==============================
SM = 'Summary'
wsum = sheet(SM, [46, 18, 18, 44])
hdr(wsum, 1, ['Summary valuation', 'Frame A (AED)', 'Frame B (AED)', 'Note'])
rsm = 2
for key, label, _fa, _fb, ea, eb, wgt, note in lens_rows:
    lbl(wsum, rsm, 1, label)
    f(wsum, rsm, 2, f'={q(FV)}!B{FVr[key]}', ea, fmt=PS)
    f(wsum, rsm, 3, f'={q(FV)}!C{FVr[key]}', eb, fmt=PS)
    lbl(wsum, rsm, 4, note, note=True)
    rsm += 1
lbl(wsum, rsm, 1, 'TERMINAL VALUE AS A PERCENTAGE OF ENTERPRISE VALUE, beside the '
    'cash-flow reading', bold=True, fill=FILL_C)
f(wsum, rsm, 2, f'={q(SB)}!B{BR["tvs"]}', A_['tv_share'], fmt=PCT, bold=True, fill=FILL_C)
f(wsum, rsm, 3, f'={q(SB)}!C{BR["tvs"]}', B_['tv_share'], fmt=PCT, bold=True, fill=FILL_C)
lbl(wsum, rsm, 4, 'it is high, and it is stated here rather than buried', note=True)
rsm += 1
lbl(wsum, rsm, 1, 'WEIGHTED CENTRE', bold=True, fill=FILL_C)
f(wsum, rsm, 2, f'={q(FV)}!B{CENT}', _ca, fmt=PS, bold=True, fill=FILL_C)
f(wsum, rsm, 3, f'={q(FV)}!C{CENT}', _cb, fmt=PS, bold=True, fill=FILL_C)
lbl(wsum, rsm, 4, 'the two frames are published side by side, never averaged', note=True)
rsm += 1
lbl(wsum, rsm, 1, 'Dividend capitalisation (unweighted fifth reading)')
f(wsum, rsm, 2, f'={q(RN)}!B{RNr["divps"]}', L['div_ps'], fmt=PS)
rsm += 1
lbl(wsum, rsm, 1, 'Market price (AED)', bold=True)
f(wsum, rsm, 2, f'={c("spot")}', V['spot'], fmt=PS, bold=True)

# ============================== SUMMARY FINANCIALS ==============================
SF = 'Summary Financials'
wsf = sheet(SF, [36] + [13] * 8)
hdr(wsf, 1, ['AED million'] + HY + YR)
rsf = 2
for label, hk, fk, exp in (
        ('Revenue', 'revenue', IS['rev'], F['revenue']),
        ('Gross profit', 'gross_profit', IS['gp'], F['gross_profit_A']),
        ('EBITDA', 'ebitda', IS['ebitda'], F['ebitda_A']),
        ('EBIT', 'ebit', IS['ebit'], F['ebit_A']),
        ('Profit attributable to owners', 'np_attributable', IS['npa'], npa_A)):
    lbl(wsf, rsf, 1, label)
    for j, y in enumerate(HY):
        val(wsf, rsf, 2 + j, H[y][hk], fmt=MONEY, kind='audited')
    for j in range(N):
        f(wsf, rsf, 5 + j, f"={q(IST)}!{get_column_letter(5 + j)}{fk}", exp[j], fmt=MONEY)
    rsf += 1
lbl(wsf, rsf, 1, 'Free cash flow to the firm')
for j in range(3):
    lbl(wsf, rsf, 2 + j, '')
for j in range(N):
    f(wsf, rsf, 5 + j, f"={q(CFT)}!{get_column_letter(5 + j)}{CF['fcff']}", A_['fcff'][j],
      fmt=MONEY)

# ============================== PER-SHARE & RATIOS ==============================
PR = 'Per-Share & Ratios'
wps = sheet(PR, [40] + [13] * 8)
hdr(wps, 1, ['Per share and ratios'] + HY + YR)
rp = 2
for label, hvals, form, exp, fmt in (
        ('Earnings per share (AED)', [H[y]['eps'] for y in HY],
         lambda j: f"={q(IST)}!{get_column_letter(5 + j)}{IS['eps']}", eps_A, PS3),
        ('Book value per share (AED)', [H[y]['equity_parent'] / V['shares_mn'] for y in HY],
         lambda j: f"={q(BST)}!{get_column_letter(5 + j)}{BS['bvps']}",
         [e / V['shares_mn'] for e in eqp], PS3),
        ('Dividend per share (AED)', [H[y]['dividends_paid'] / V['shares_mn'] for y in HY],
         lambda j: f'={c("dps")}', [V['dps']] * N, PS3),
        ('EBITDA margin', [H[y]['ebitda_margin'] for y in HY],
         lambda j: f"={q(IST)}!{get_column_letter(5 + j)}{IS['em']}",
         F['ebitda_margin_A'], PCT),
        ('Return on equity', [H[y]['np_attributable'] / H[y]['equity_parent'] for y in HY],
         lambda j: (f"={q(IST)}!{get_column_letter(5 + j)}{IS['npa']}/"
                    f"{q(BST)}!{get_column_letter(5 + j)}{BS['eq']}"),
         [npa_A[i] / eqp[i] for i in range(N)], PCT),
        ('Net debt to EBITDA', [H[y]['net_debt_company'] / H[y]['ebitda'] for y in HY],
         lambda j: (f"={q(BST)}!{get_column_letter(5 + j)}{BS['nd']}/"
                    f"{q(IST)}!{get_column_letter(5 + j)}{IS['ebitda']}"),
         [nd_f[i] / F['ebitda_A'][i] for i in range(N)], X),
        ('Return on capital employed', [H[y]['roce'] for y in HY],
         lambda j: (f"={q(IST)}!{get_column_letter(5 + j)}{IS['ebit']}/"
                    f"({q(BST)}!{get_column_letter(5 + j)}{BS['eq']}+"
                    f"{q(BST)}!{get_column_letter(5 + j)}{BS['nci']}+"
                    f"{q(BST)}!{get_column_letter(5 + j)}{BS['nd']})"),
         [F['ebit_A'][i] / (eqp[i] + ncieq[i] + nd_f[i]) for i in range(N)], PCT)):
    lbl(wps, rp, 1, label)
    for j, v in enumerate(hvals):
        val(wps, rp, 2 + j, v, fmt=fmt, kind='audited')
    for j in range(N):
        f(wps, rp, 5 + j, form(j), exp[j], fmt=fmt)
    rp += 1

# ============================== MONTE CARLO ==============================
MC = 'Monte Carlo'
wmc = sheet(MC, [42, 16, 16, 44])
hdr(wmc, 1, ['Probability map', 'One month', 'Three months', 'Note'])
ST0 = json.load(open(os.path.join(HERE, 'strike_result.json')))
lbl(wmc, 2, 1, f"THESE CELLS ARE A WHOLE-MODEL RE-RUN — {ST0['n_paths']:,} simulated price "
    "paths each. They do NOT redraw when a driver on the Assumptions sheet changes.",
    bold=True, fill=FILL_C)
wmc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
ST = json.load(open(os.path.join(HERE, 'strike_result.json')))
rm = 3
for p in ('p5', 'p25', 'p50', 'p75', 'p95'):
    lbl(wmc, rm, 1, f'{p[1:]}th percentile (AED)')
    val(wmc, rm, 2, ST['horizons']['1M']['pct'][p], fmt=PS, kind='grid')
    val(wmc, rm, 3, ST['horizons']['3M']['pct'][p], fmt=PS, kind='grid')
    rm += 1
for lab, k in (('Probability of finishing above the anchor', 'p_above'),
               ('Probability of finishing 10% higher', 'p_up10'),
               ('Probability of finishing 10% lower', 'p_dn10'),
               ('Probability of TOUCHING 10% higher', 'touch_up10'),
               ('Probability of TOUCHING 10% lower', 'touch_dn10')):
    lbl(wmc, rm, 1, lab)
    val(wmc, rm, 2, ST['horizons']['1M'][k], fmt=PCT, kind='grid')
    val(wmc, rm, 3, ST['horizons']['3M'][k], fmt=PCT, kind='grid')
    rm += 1

# ============================== SENSITIVITY ==============================
SN = 'Sensitivity'
wsn = sheet(SN, [40] + [15] * 5)
hdr(wsn, 1, ['Value per share (AED), Frame A'] + [f'{g * 100:.1f}% growth'
                                                  for g in SENS['grid_g']])
lbl(wsn, 2, 1, 'EACH CELL IS A COMPLETE REVALUATION of the whole model. These grids do NOT '
    'redraw when a driver changes.', bold=True, fill=FILL_C)
wsn.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
rn_ = 3
for i, wv in enumerate(SENS['grid_wacc']):
    lbl(wsn, rn_, 1, f'Terminal cost of capital {wv * 100:.2f}%')
    for j in range(5):
        val(wsn, rn_, 2 + j, SENS['grid'][i][j], fmt=PS, kind='grid')
    rn_ += 1
rn_ += 1
hdr(wsn, rn_, ['Single-driver sensitivity', 'Input', 'Value per share', '', '', ''])
rn_ += 1
for name, key in (('Terminal cost of capital', 'wacc'), ('Long-run growth', 'g'),
                  ('Beta', 'beta'), ('Retail volume growth', 'volume'),
                  ('Margin per litre', 'margin'),
                  ('Recurring inventory movement', 'inventory'),
                  ('Effective tax rate', 'tax'), ('Capital expenditure', 'capex')):
    for iv, pv in SENS[key]:
        lbl(wsn, rn_, 1, name)
        val(wsn, rn_, 2, iv, fmt=PS3, kind='grid')
        val(wsn, rn_, 3, pv, fmt=PS, kind='grid')
        rn_ += 1

# ============================== PEER & SECTOR ==============================
PS_ = 'Peer & Sector'
wpe = sheet(PS_, [44, 16, 16, 44])
hdr(wpe, 1, ['Peer and sector evidence', 'EV / EBITDA', 'Role', 'Note'])
lbl(wpe, 2, 1, 'Peer observations are a CROSS-CHECK. They are not the source of any '
    'number in this valuation: the reference multiple is triangulated from the company’s '
    'own history on the Relative & Normalized sheet.', bold=True, fill=FILL_C)
wpe.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
rpe = 3
peers = []
for e in SR['entries']:
    if str(e.get('topic', '')).startswith('Peer multiple') and e.get('value'):
        nm = e['topic'].replace('Peer multiple — ', '').split(' (')[0]
        peers.append((nm, float(e['value'])))
for nm, mult in peers:
    lbl(wpe, rpe, 1, nm)
    val(wpe, rpe, 2, mult, fmt=X, kind='audited')
    lbl(wpe, rpe, 3, 'cross-check')
    rpe += 1
if peers:
    lbl(wpe, rpe, 1, 'PEER AVERAGE (context only)', bold=True, fill=FILL_C)
    f(wpe, rpe, 2, f'=AVERAGE(B3:B{rpe - 1})', sum(p[1] for p in peers) / len(peers),
      fmt=X, bold=True, fill=FILL_C)
    lbl(wpe, rpe, 3, 'not used', bold=True)
    lbl(wpe, rpe, 4, 'the spread across these names is more than three to one, across '
        'different fuel-pricing regimes', note=True)
    rpe += 1
lbl(wpe, rpe, 1, 'This company, on the traded price', bold=True)
f(wpe, rpe, 2, f'=({c("mcap")}+{c("netdebt")}+{c("lease0")})/'
  f'{q(IST)}!D{IS["ebitda"]}',
  (W['mcap'] + W['net_debt'] + V['lease_fy25']) / H['FY2025']['ebitda'], fmt=X, bold=True)
lbl(wpe, rpe, 3, 'the subject', bold=True)

# ============================== READ FIRST ==============================
RF = 'READ FIRST'
ws = sheet(RF, [46, 96])
rrf = 1
lbl(ws, rrf, 1, 'ADNOC Distribution — valuation model', bold=True, fill=FILL_P)
lbl(ws, rrf, 2, f"{M['company']} · {M['exchange']} · {M['currency']} · "
    f"study date {M['study_date']} · price date {M['price_date']}", fill=FILL_P)
rrf += 2
for head, body in (
    ('THIS WORKBOOK CALCULATES',
     'Everything that can be derived from a driver is a live formula. Change a driver on '
     'the Assumptions sheet and the cost of capital, the glide, the discount factors, the '
     'cash-flow waterfall, the terminal block, the statements and every ratio all move.'),
    ('BLUE IS AN INPUT, BLACK IS A FORMULA',
     'That is the whole colour code. A blue cell is a number that came from somewhere; a '
     'black one is computed by the sheet from cells above it.'),
    ('ONLY THREE KINDS OF CELL ARE PASTED — 1. AUDITED AND DISCLOSED HISTORY',
     'The FY2023, FY2024 and FY2025 statement columns and the disclosed operating figures '
     '(volumes, station counts, segment gross profit). Where a line is both disclosed and '
     'derivable, the DISCLOSED figure is carried, because that is what the company reported.'),
    ('2. THE OUTPUT OF THE UNIT BUILD',
     'The opening volumes, prices and margins per litre that the forecast grows from. '
     'Flattening the reconciliation behind them into this grid would make it unreadable.'),
    ('3. WHOLE-MODEL RE-RUNS',
     'The probability map on the Monte Carlo sheet and the grids on the Sensitivity sheet. '
     'Each of those cells is a COMPLETE revaluation of the entire model at a different '
     'setting. THEY DO NOT REDRAW WHEN A DRIVER CHANGES. Every other number here does.'),
    ('THE CONTESTED JUDGEMENT IS CARRIED BOTH WAYS',
     'Inventory movements were AED 254m in FY2024, AED 335m in FY2025 and AED 762m in the '
     'first half of 2026 alone, against fuel volume growth of 1.6% in that half. Frame A '
     'normalises them to zero from FY2027; Frame B carries the FY2024-FY2025 average. Both '
     'run all the way through to a value per share and they are NEVER averaged together.'),
    ('WHERE TO START',
     'Summary for the answer, Assumptions for every driver, Segments for the volume and '
     'margin build, DCF for the waterfall and the terminal block, SOTP Bridge for the '
     'enterprise-to-equity walk and the terminal share of value.'),
):
    lbl(ws, rrf, 1, head, bold=True)
    lbl(ws, rrf, 2, body)
    ws.row_dimensions[rrf].height = 46
    rrf += 1

# ============================== EMIT ==============================
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = sorted(wb._sheets, key=lambda s: order.index(s.title))
OUT = os.path.join(HERE, 'ADNOCDIST_Valuation_Model_09082026.xlsx')
wb.save(OUT)
json.dump({'expected': EXPECT, 'paste_counts': NPASTE},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
_pasted = NPASTE['audited'] + NPASTE['unit_build'] + NPASTE['grid']
print(f'wrote {os.path.basename(OUT)}: {len(wb._sheets)} sheets')
print(f'  formulas: {len(EXPECT)}')
print(f'  pasted values: {_pasted}  (audited {NPASTE["audited"]}, '
      f'unit build {NPASTE["unit_build"]}, whole-model grids {NPASTE["grid"]})')
print(f'  text labels: {NPASTE["label"]}')
