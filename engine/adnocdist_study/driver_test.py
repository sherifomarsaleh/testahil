"""Does the workbook actually reprice when a driver moves?

Recalculation proves the formulas reproduce the model as written. It does NOT prove the
model is wired: a workbook whose headline is a constant would pass that check perfectly.
This one perturbs each input IN PLACE — nothing is written to disk, the override layer in
xlcalc does the work — re-evaluates the WHOLE workbook from an empty cache, and asserts the
headline moves in the direction claimed.

STANDING RULE, and it has earned its place twice below: if a directional expectation fails,
THE FIRST HYPOTHESIS IS THAT THE EXPECTATION IS WRONG, NOT THE MODEL. Decompose the
mechanism and report what actually happens before changing anything.

Then a dead-input sweep: every remaining numeric input is nudged and checked against TWELVE
headlines, not one, because an input is not dead merely because it leaves the headline value
per share alone. A driver of the balance sheet, or of the cost of equity, is still live.
"""
import json, os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import openpyxl
import xlcalc

XLSX = os.path.join(HERE, 'ADNOCDIST_Valuation_Model_09082026.xlsx')
wb = openpyxl.load_workbook(XLSX)
wa = wb['Assumptions']


def row_of(text):
    for row in wa.iter_rows(min_col=1, max_col=1):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().startswith(text):
                return c.row
    raise KeyError(text)


def find(sheet, text, col=1):
    ws = wb[sheet]
    for row in ws.iter_rows(min_col=col, max_col=col):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().startswith(text):
                return c.row
    raise KeyError(f'{sheet}: {text!r}')


def cellref(sheet, text, col_letter='B'):
    return (sheet, f'{col_letter}{find(sheet, text)}')


def rowref(sheet, text, col_letter):
    return (sheet, f'{col_letter}{find(sheet, text)}')


HEADLINE = cellref('SOTP Bridge', 'VALUE PER SHARE', 'B')


def headline(overrides=None):
    return float(xlcalc.Book(wb, overrides).cell_value(*HEADLINE))


BASE = headline()
print(f'base value per share, Frame A: {BASE:.4f}')

R = row_of
CASES = [
    ('Service-station network growth', f'C{R("Service-station network growth")}', 1.50, +1,
     'more sites selling fuel and running shops'),
    ('Litres per station', f'C{R("Growth in litres sold per station")}', 1.50, -1,
     'EXPECTATION CORRECTED. This driver is a GROWTH RATE and it is NEGATIVE — throughput per '
     'site is falling. Multiplying a negative rate by 1.5 accelerates the decline, so volume '
     'and value fall. The first expectation read the driver as a level rather than a rate.'),
    ('Corporate fuel volume growth', f'D{R("Corporate fuel volume growth")}', 1.50, +1,
     'more litres on the business-to-business leg at the same margin per litre'),
    ('Aviation fuel volume growth', f'D{R("Aviation fuel volume growth")}', 1.50, +1,
     'same mechanism on the jet-fuel leg, now modelled separately because it moved the '
     'opposite way to corporate'),
    ('Retail margin per litre, escalation', f'C{R("Retail margin per litre, escalation")}',
     1.50, +1, 'the margin per litre IS the economics of a regulated fuel retailer'),
    ('Commercial margin per litre, escalation',
     f'C{R("Commercial margin per litre, escalation")}', 1.20, +1, 'same, commercial legs'),
    ('Fuel transaction growth', f'C{R("Fuel transaction growth")}', 1.50, +1,
     'transactions drive the non-fuel leg through the conversion rate'),
    ('Non-fuel conversion rate', f'C{R("Growth in the non-fuel conversion rate")}', 1.50, -1,
     'EXPECTATION CORRECTED, same reason as litres per station: the conversion GROWTH rate is '
     'negative, so scaling it up deepens the deterioration.'),
    ('Non-fuel basket', f'C{R("Growth in the average non-fuel basket")}', 1.50, +1,
     'spend per non-fuel visit'),
    ('Non-fuel retail gross margin', f'C{R("Non-fuel retail gross margin")}', 1.10, +1,
     'more gross profit on the same non-fuel revenue'),
    ('Realised retail price per litre', f'C{R("Realised retail price per litre")}', 1.10, +1,
     'A PASS-THROUGH, and the direction is not the obvious one. It lifts revenue and direct '
     'cost by the same amount and leaves gross profit — volume x margin per litre — '
     'untouched, so it reaches value ONLY through working capital. Receivable plus inventory '
     'days less payable days is about MINUS 31, so a higher price makes an already negative '
     'balance more negative and RELEASES cash. The move is tiny, which is what a '
     'pass-through should look like.'),
    ('Realised corporate price per litre', f'C{R("Realised corporate price per litre")}',
     1.10, +1, 'same pass-through mechanism on the corporate leg'),
    ('Realised aviation price per litre', f'C{R("Realised aviation price per litre")}',
     1.10, +1, 'same pass-through mechanism on the aviation leg'),
    ('Cash operating cost growth', f'C{R("Cash operating cost growth")}', 1.50, -1,
     'costs compound against a margin that does not'),
    ('Impairments and other operating expenses', f'C{R("Impairments and other")}', 1.50, -1,
     'straight charge against operating profit'),
    ('Depreciation rate on the opening base', f'C{R("Depreciation rate on the opening")}',
     1.50, -1,
     'NOT the obvious direction, and the mechanism is worth stating. Work the waterfall '
     'through: free cash flow = EBITDA x (1 - tax) + depreciation x tax - capital spending '
     '- working capital. Depreciation does not cancel — it leaves a TAX SHIELD worth '
     'depreciation x the tax rate. That shield outweighs the lower terminal NOPAT here. '
     'EXPECTATION CORRECTED AND THE MECHANISM CHANGED. When depreciation was a pasted array '
     'the shield dominated and value rose. Now that depreciation is DERIVED from an asset '
     'base the model rolls forward, a higher rate also shrinks that base faster, so later '
     'depreciation and the terminal NOPAT both fall — and the terminal effect now dominates. '
     'Making the charge endogenous flipped the sign, which is itself worth knowing.'),
    ('Maintenance capital spending rate', f'C{R("Maintenance capital spending, share")}',
     1.50, -1, 'cash out of the door, and it is not in the terminal reinvestment rate'),
    ('Capital cost per station added', f'C{R("Capital cost per station added")}', 1.50, -1,
     'growth spending is now per station added rather than a pasted total'),
    ('Inventory movement, normalised frame', f'C{R("Inventory movement, normalised")}',
     2.00, +1, 'straight addition to gross profit in the frame the headline uses'),
    ('Receivable days', f'C{R("Receivable days")}', 1.20, -1, 'slower collection absorbs cash'),
    ('Inventory days', f'C{R("Inventory days")}', 1.20, -1, 'more stock absorbs cash'),
    ('Payable days', f'C{R("Payable days")}', 1.20, +1,
     'paying later releases cash — why the working-capital line is a source every year'),
    ('Government bond yield', f'C{R("Government bond yield")}', 1.20, -1,
     'lifts the cost of equity and the cost of debt together'),
    ('Sovereign default spread', f'C{R("Sovereign default spread")}', 1.50, +1,
     'THE SIGN LOOKS WRONG AND IS NOT. The spread is SUBTRACTED from the quoted yield to '
     'normalise it, so a larger spread means a lower normalised risk-free rate. That is the '
     'point of normalising: country risk is priced once, inside the equity premium. What '
     'CHANGED after audit is the SIZE of the strip — 4 basis points, the spread the bond '
     'actually carries over comparable US Treasuries, not a 42-basis-point ratings lookup '
     'that pushed the result below the risk-free rate of the currency it is pegged to.'),
    ('Total equity risk premium', f'C{R("Total equity risk premium")}', 1.20, -1,
     'raises the cost of equity in the explicit years and the terminal alike, because the '
     'cost of capital is now flat'),
    ('Beta', f'C{R("Beta")}', 1.50, -1,
     'raises the cost of equity. With a flat pegged-market cost of capital the terminal beta '
     'IS the measured beta, so the beta ladder now reproduces the model at every point'),
    ('Credit margin over the sovereign', f'C{R("Credit margin over the sovereign")}',
     1.50, -1, 'raises the after-tax cost of debt; the debt weight is small but real'),
    ('Long-run growth', f'C{R("Long-run growth")}', 1.50, +1,
     'the terminal is a perpetuity and growth is in its denominator'),
    ('Terminal return on invested capital', f'C{R("Terminal return on invested")}', 1.50, +1,
     'a higher return needs less reinvestment to fund the same growth — and it now moves the '
     'normalised lens too, which charges the same reinvestment'),
    ('Statutory tax rate', f'C{R("Statutory tax rate")}', 1.50, +1,
     'THE SIGN LOOKS WRONG AND IS NOT. The statutory rate no longer taxes the cash flows — '
     'the effective rate does. All it still does is size the interest tax shield inside the '
     'after-tax cost of debt, so raising it LOWERS the discount rate.'),
    ('Effective tax rate', f'C{R("Effective tax rate")}', 1.20, -1,
     'this is the rate that taxes the cash flows'),
    # The dividend policy floor is deliberately NOT a headline case. The headline is an
    # FCFF valuation and dividends are a FINANCING flow, so it cannot move it — expecting it
    # to was wrong. It is a live input all the same: it drives the equity roll-forward, cash
    # and net debt, and the dead-input sweep below confirms that against the balance-sheet
    # headlines rather than this one.
    ('Share price', f'C{R("Share price")}', 1.20, -1,
     'a higher price raises the market capitalisation and therefore the equity weight, and '
     'equity is the dearer money'),
    ('Net debt excluding leases', f'C{R("Net debt excluding leases")}', 1.50, -1,
     'straight deduction in the bridge, and it lifts the debt weight'),
    ('Lease liabilities, FY2025', f'C{R("Lease liabilities, FY2025")}', 1.50, -1,
     'deducted in the bridge on the same basis as debt'),
    ('Non-controlling interests, FY2025', f'C{R("Non-controlling interests, FY2025")}',
     1.50, -1, 'deducted in the bridge'),
    ('Shares in issue', f'C{R("Shares in issue")}', 1.20, -1,
     'the same equity value spread over more shares'),
]

failures, results = [], []
for label, cell, mult, want, why in CASES:
    base_v = wa[cell].value
    got = headline({('Assumptions', cell): base_v * mult})
    move = got - BASE
    ok = (move > 1e-9) if want > 0 else (move < -1e-9)
    results.append(dict(driver=label, cell=cell, multiplier=mult,
                        expected='up' if want > 0 else 'down',
                        base=BASE, moved_to=got, change=move, ok=ok, why=why))
    print(f"  {'ok ' if ok else 'FAIL'} {label:44s} x{mult:<5} "
          f"{'up' if want > 0 else 'down':4s} -> {got:8.4f} ({move:+.4f})")
    if not ok:
        failures.append(f'{label}: expected {"up" if want > 0 else "down"}, moved {move:+.6f}')

# ---- dead-input sweep, against twelve headlines --------------------------------
HEADLINES = {
    'value per share, Frame A': HEADLINE,
    'value per share, Frame B': cellref('SOTP Bridge', 'VALUE PER SHARE', 'C'),
    'weighted centre, Frame A': cellref('Fundamental Valuation', 'WEIGHTED CENTRE', 'B'),
    'weighted centre, Frame B': cellref('Fundamental Valuation', 'WEIGHTED CENTRE', 'C'),
    'book and sustainable return lens': cellref('Relative & Normalized',
                                                'BOOK VALUE AND SUSTAINABLE', 'B'),
    'relative multiples lens': cellref('Relative & Normalized', 'RELATIVE MULTIPLES READING',
                                       'B'),
    'normalised earnings power lens': cellref('Relative & Normalized',
                                              'NORMALISED EARNINGS POWER', 'B'),
    'dividend capitalisation lens': cellref('Relative & Normalized',
                                            'DIVIDEND CAPITALISATION', 'B'),
    'first forecast year earnings per share': rowref('Income Statement',
                                                     'Earnings per share', 'E'),
    'final year total assets': rowref('Balance Sheet', 'TOTAL ASSETS', 'I'),
    'final year net debt': rowref('Balance Sheet', 'Net debt excluding leases', 'I'),
    'cost of equity': ('Assumptions', f'C{row_of("Cost of equity")}'),
}
BASES = {k: float(xlcalc.Book(wb).cell_value(*v)) for k, v in HEADLINES.items()}
tested = {c for _, c, _, _, _ in CASES}
live, dead, disclosure = [], [], []
for row in wa.iter_rows(min_col=3, max_col=7):
    for cell in row:
        if cell.value is None or isinstance(cell.value, str):
            continue
        if not isinstance(cell.value, (int, float)) or cell.value == 0:
            continue
        coord = cell.coordinate
        if coord in tested:
            continue
        basis = wa.cell(row=cell.row, column=2).value or ''
        basis = str(basis).strip()
        if basis == 'calculated':
            continue   # a derived row, not an input
        label = wa.cell(row=cell.row, column=1).value
        moved = []
        for name, ref in HEADLINES.items():
            v = float(xlcalc.Book(wb, {('Assumptions', coord): cell.value * 1.15}
                                 ).cell_value(*ref))
            if abs(v - BASES[name]) > 1e-9:
                moved.append(name)
        if moved:
            live.append((coord, label, moved))
        elif basis.startswith('disclosure'):
            # A figure carried because the reader needs to see it, not because the model
            # consumes it. Recorded and named rather than quietly counted as live.
            disclosure.append((coord, label, basis))
        else:
            dead.append((coord, label, basis))

print(f'\ndead-input sweep: {len(live)} live, {len(dead)} dead, '
      f'{len(disclosure)} excluded by design')
for coord, label, basis in dead:
    print(f'   DEAD {coord} {label!r} (basis: {basis})')
for coord, label, basis in disclosure:
    print(f'   excluded by design: {coord} {label!r} — carried for the reader, '
          f'not consumed by the model')

json.dump(dict(base=BASE, cases=results, failures=failures,
               live_inputs=len(live), dead_inputs=[d[:2] for d in dead],
               excluded_by_design=[d[:2] for d in disclosure],
               headlines=list(HEADLINES)),
          open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1)

assert not failures, f'{len(failures)} driver(s) did not move the headline as asserted: {failures}'
assert not dead, f'{len(dead)} dead input(s): {dead}'
print(f'\n{len(CASES)} directions correct, {len(live)} live inputs, 0 dead, '
      f'{len(disclosure)} excluded by design')
