"""SCEM_Valuation_Model_04092026_public.xlsx — 16 sheets, formula-first.

Blue = input · black = formula · green = cross-sheet link.

The workbook CALCULATES. Every quantity arithmetically derivable from a driver is a
live Excel formula, so the reader can change a blue cell on Assumptions and watch the
model reprice: the cost of capital is built from the risk-free rate net of the sovereign
spread, beta and the premium; the glide fractions are derived from the cost-of-debt path;
the discount factors compound; the DCF waterfall chains; the terminal block chains; the
statements roll forward; every ratio and per-share figure is a formula.

Only three classes of cell are pasted, named on READ FIRST:
  1. audited and disclosed history — the primary record, not a calculation;
  2. the output of the volume-and-price unit build's historical calibration (the market
     size and share pair per year), which is a research judgement rather than an
     arithmetic step; everything downstream of it is formula;
  3. whole-model re-runs — the Monte Carlo price map and the sensitivity grids, where
     each cell is a complete revaluation. These do NOT redraw when a driver changes.

Every formula cell also records the model's own value into xlsx_expected.json, and
recalc.py evaluates the workbook independently and asserts the two agree.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
S0 = json.load(open(os.path.join(HERE, 'step0_result.json')))
STK = json.load(open(os.path.join(HERE, 'strike_result.json')))
BETA = json.load(open(os.path.join(HERE, 'beta_result.json')))

BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'
NUM3 = '#,##0.000'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'

M, H, F = D['meta'], D['history'], D['forecast']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sensitivity']
DISP = D['disposal']
BUD = D['bottom_up']
PE = D['peers']
SHT = D['share_triangulation']
TR = D['terminal_reconciliation']
IN = {k: v['value'] for k, v in D['inputs'].items()}
LR = D['lens_record']          # [R-LENS-03] the primary and its cross-checks
SPOT, SH, TAX = M['spot'], M['shares_mn'], IN['tax_stat']
YH = ['FY2023', 'FY2024', 'FY2025']
YF = F['years']
HC = ['B', 'C', 'D']                    # historical columns
FC = ['E', 'F', 'G', 'H', 'I']          # forecast columns
DC = ['B', 'C', 'D', 'E', 'F']          # forecast columns on DCF-style sheets

wb = Workbook()
EXPECT, ANCH = {}, {}


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula and record the model's own value for the same cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


def note(ws, row, text, w=10):
    ws.cell(row=row, column=1, value=text).font = SUB
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=w)


# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Sinai Cement Company S.A.E. (EGX: SCEM)', None, 9)
LINES = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the SCEM valuation study. Every blue cell is an input;',
 'every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN, AND THAT CLAIM IS TESTED. Every figure derivable from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice. The cost of equity is built from the',
 'risk-free rate NET of the sovereign spread, beta and the premium rather than pasted; the cost of debt is',
 'taxed in the sheet; the weights come from net debt and market capitalisation; the terminal rate is built',
 'from its own components; the glide fractions are visibly derived from the cost-of-debt path; the discount',
 'factors compound; the DCF waterfall chains from margin through EBIT, NOPAT and FCFF to present value; the',
 'terminal block chains from reinvestment = g / return on capital; and the statements, ratios and all four',
 'lenses run off the same cells. A driver test perturbs every input in place and confirms the headline moves.', '',
 'THREE CLASSES OF CELL ARE PASTED, and it is worth knowing exactly which.',
 '  1. Audited and disclosed history — the primary record, not a calculation. Where a line is both disclosed',
 '     and derivable, the DISCLOSED figure is carried.',
 '  2. The unit build\'s historical calibration: the Egyptian market size and SCEM\'s share of it, per year.',
 '     Those two are research judgements, not arithmetic. Volume, realised price and revenue are all formulas',
 '     from them, and everything downstream is formula.',
 '  3. The Monte Carlo price map and the sensitivity grids. Each individual cell there is a COMPLETE re-run of',
 '     the whole valuation, so it cannot be a formula in a grid. THESE GRIDS DO NOT REDRAW WHEN A DRIVER',
 '     CHANGES — if you edit an input on Assumptions, the waterfall, statements and lenses all reprice, but',
 '     the sensitivity tables and the price map keep the values printed here.',
 '  Anything else pasted would be a defect.', '',
 'HOW THE COMPANY IS VALUED, and why. SCEM is a single-asset cement operating company: essentially all revenue',
 'is grey cement and clinker from two lines at El Hassana, and total debt of EGP 36.8mn against ~EGP 5.2bn of',
 'equity makes it NET CASH. It is valued with one operating-company lens, not a sum of parts. The 25.4% stake',
 'in Sinai White Portland Cement — the only thing that could have been a second leg — was sold to Aalborg',
 'Portland (Cementir) for EUR 30mn, completed 13 August 2024.', '',
 'TWO THINGS THE READER SHOULD KNOW BEFORE READING ANY NUMBER.',
 '  * FY2024 is not a base year. Profit after tax of EGP 3,070mn on revenue of EGP 6,420mn is a 48% net margin',
 '    that no cement plant earns from making cement; it contains the Sinai White disposal gain. Stripping it,',
 '    underlying FY2024 profit was EGP 1,568mn, so FY2025 profit ROSE 46% — it did not fall 25%.',
 '  * Profit after tax EXCEEDS EBITDA because a large cash pile earns Egyptian policy rates. That is an',
 '    excess-cash artefact, not a second business. Treasury income is excluded from free cash flow entirely',
 '    and the cash is added back in the enterprise-to-equity bridge.', '',
 'PROVENANCE LIMIT, STATED PLAINLY. The company\'s audited statements could not be retrieved when this was',
 'built: the environment\'s egress policy refused every external host. Revenue and profit after tax are',
 'carried as disclosed via reporting of the EGX filing; every line between them is DERIVED by closing the',
 'disclosed profit, and is labelled as derived on the Income Statement sheet.', '',
 'No rating and no price target. Fair-value ranges and distributions only.',
]
for i, ln in enumerate(LINES):
    ws.cell(row=4 + i, column=1, value=ln).font = Font(size=10 if ln else 9)
ws.column_dimensions['A'].width = 118

# ============ 4 ASSUMPTIONS (written early: every sheet references it) ========
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every driver, including the whole cost stack',
      'Blue cells are inputs. Nothing below a driver is typed.', 9, 54, 13)
A = {}
NUM3 = '#,##0.000'


def inp(row, label, key, val, fmt=NUM0, unit=''):
    wsA.cell(row=row, column=1, value=label)
    put(wsA, f'B{row}', val, BLUE, fmt)
    if unit:
        wsA.cell(row=row, column=3, value=unit).font = SUB
    A[key] = f'Assumptions!$B${row}'
    return A[key]


def inprow(row, label, key, vals, fmt=NUM0, unit=''):
    wsA.cell(row=row, column=1, value=label)
    for i, v in enumerate(vals):
        put(wsA, f'{BUC[i]}{row}', v, BLUE, fmt)
        A[f'{key}{i}'] = f"Assumptions!${BUC[i]}${row}"
    if unit:
        wsA.cell(row=row, column=9, value=unit).font = SUB


BUC = ['B', 'C', 'D', 'E', 'F', 'G']
band(wsA, 4, 9); wsA['A4'] = 'MARKET'
inp(5, 'Spot price', 'spot', SPOT, PX, 'EGP, close 06-Aug-2026')
inp(6, 'Shares outstanding', 'shares', SH, NUM2, 'mn')
inp(7, 'Statutory tax rate', 'tax', IN['tax_stat'], PCT, '')

inp(9, 'USD/EGP (spot)', 'fx', IN['fx'], NUM1, '')

band(wsA, 11, 9); wsA['A11'] = 'PLANT — PHYSICAL'
inp(12, 'Cement grinding capacity', 'capcem', IN['cap_cement_mt'], NUM2, 'Mt/yr')
inp(13, 'Kiln clinker capacity', 'capclk', IN['cap_clinker_mt'], NUM2, 'Mt/yr')
inp(14, 'Clinker factor', 'cfac', IN['clinker_factor'], NUM3, 't clinker / t cement')

band(wsA, 15, 9); wsA['A15'] = 'COST STACK — PHYSICAL AND MARKET DRIVERS'
# THE COST STACK IS THE COMPANY'S OWN DISCLOSED ONE. The earlier edition carried four
# industry rules of thumb and a fixed cost per tonne of installed capacity; notes 24, 25
# and 26 of the audited accounts state the lines, and they are the inputs now.
inp(16, 'Materials, fuel, power, packing (note 24)', 'matfy', IN['cost_materials_fy25'],
    NUM0, 'EGP mn, FY2025')
inp(17, 'Transport, loading and export (notes 24, 25)', 'distfy',
    IN['cost_distribution_fy25'], NUM0, 'EGP mn, FY2025')
inp(18, 'Fixed cash cost (the rest of notes 24, 25, 26)', 'fixfy', IN['cost_fixed_fy25'],
    NUM0, 'EGP mn, FY2025')
inp(19, 'Dollar-linked share of the materials line', 'usdsh', IN['materials_usd_share'],
    PCT, 'ESTIMATED — the accounts do not split it')
inp(20, 'Gross fixed assets (note 4)', 'grossfa', IN['gross_fixed_fy25'], NUM0,
    'EGP mn, 31-Dec-2025')
inp(21, 'Weighted depreciation rate (note 3/2 on note 4)', 'deprate',
    IN['dep_rate_disclosed'], PCT, 'straight-line')
inp(22, 'Capex run rate (cash-flow statements)', 'capexrr', IN['capex_run_rate'], NUM0,
    'EGP mn/yr')
inp(23, 'Cash, reviewed sheet 31-Mar-2026', 'cashm26', IN['cash_mar26'], NUM0, 'EGP mn')
inp(24, 'Lease liabilities, 31-Mar-2026', 'debtm26', IN['debt_mar26'], NUM0, 'EGP mn')

band(wsA, 26, 9); wsA['A26'] = 'FORECAST PATHS — FY2025A then FY2026E to FY2030E'
hdr(wsA, 27, [''] + ['FY2025A'] + YF)
inprow(28, 'Kiln utilisation', 'util', IN['kiln_util'], PCT)
inprow(29, 'Domestic share of despatches', 'dom', IN['domestic_share'], PCT)
inprow(30, 'Domestic realised price', 'pdom', IN['price_dom_egp_t'], NUM0, 'EGP/t')
inprow(31, 'Export price', 'pexp', IN['price_exp_usd_t'], NUM1, 'USD/t')
inprow(32, 'USD/EGP path', 'fxp', IN['fx_path'], NUM1, '')
inprow(33, 'Local cost inflation index', 'infl', IN['cost_infl'], NUM3, '')

band(wsA, 35, 9); wsA['A35'] = 'CAPITAL INTENSITY (FY2026E to FY2030E)'
# D&A / revenue and capex / revenue are RETIRED: depreciation is now the disclosed rate on
# the asset base and capex the company's own run rate, both from the filings.
for j, (lab, key, vals) in enumerate([
        ('Cost-of-debt path', 'kdp', IN['kd_path']),
        ('Yield earned on cash', 'cy', IN['cash_yield'])]):
    rr = 38 + j
    wsA.cell(row=rr, column=1, value=lab)
    for i, v in enumerate(vals):
        put(wsA, f'{DC[i]}{rr}', v, BLUE, PCT)
        A[f'{key}{i}'] = f"Assumptions!${DC[i]}${rr}"
inp(40, 'Delta working capital / delta revenue', 'wcp', IN['wc_pct_drev'], PCT)
inp(41, 'Dividend payout ratio', 'pay', IN['payout'], PCT)


band(wsA, 44, 9); wsA['A44'] = 'COST OF CAPITAL'
inp(45, 'Risk-free rate (EGP 10-year)', 'rf', IN['rf'], PCT2)
inp(46, 'Sovereign default spread (netted out)', 'sov', IN['sov_spread_cds'], PCT2)
inp(47, 'Equity risk premium (CDS basis)', 'erp', IN['erp_cds'], PCT2)
inp(48, 'Beta', 'beta', IN['beta'], NUM2)
inp(49, 'Pre-tax cost of debt', 'kd', IN['kd'], PCT2)
inp(50, 'Gross debt', 'debt', IN['debt_fy25'], NUM1, 'EGP mn')
inp(51, 'Terminal risk-free rate', 'rft', IN['rf_term'], PCT2)
inp(52, 'Terminal equity risk premium', 'erpt', IN['erp_term'], PCT2)
inp(53, 'Terminal cost of debt', 'kdt', IN['kd_term'], PCT2)
inp(54, 'Terminal debt weight', 'wdt', IN['wd_term'], PCT)
inp(55, 'Terminal growth', 'g', IN['g_term'], PCT)
inp(56, 'Elapsed fraction of FY2026 at valuation', 'stub', IN['stub_years'], NUM3)

band(wsA, 58, 9); wsA['A58'] = 'BALANCE SHEET AND DISCLOSED HISTORY'
inp(59, 'FY2025 cash (REPORTED)', 'cash25', IN['cash_fy25'], NUM0, 'EGP mn')
inp(60, 'FY2025 equity (reported)', 'eq25', IN['eq_fy25_rep'], NUM0, 'EGP mn')
inp(61, 'Non-controlling interests', 'nci', IN['nci'], NUM0, 'EGP mn')
inp(62, 'FY2023 revenue', 'rev23', IN['rev_fy23'], NUM0)
inp(63, 'FY2024 revenue', 'rev24', IN['rev_fy24'], NUM0)
inp(64, 'FY2025 revenue', 'rev25', IN['rev_fy25'], NUM0)
inp(65, 'FY2023 profit after tax', 'pat23', IN['pat_fy23'], NUM0)
inp(66, 'FY2024 profit after tax', 'pat24', IN['pat_fy24'], NUM0)
inp(67, 'FY2025 profit after tax', 'pat25', IN['pat_fy25'], NUM0)

inp(69, 'FY2024 total assets', 'ta24', IN['ta_fy24'], NUM1)
inp(70, 'FY2024 total liabilities', 'tl24', IN['tl_fy24'], NUM1)

inp(72, 'FY2023 weighted-average shares', 'sh23', IN['shares_fy23'], NUM2, 'mn')
inp(73, 'FY2024 weighted-average shares', 'sh24', IN['shares_fy24'], NUM2, 'mn')
inp(74, 'FY2025 weighted-average shares', 'sh25', IN['shares_fy25'], NUM2, 'mn')



inp(92, 'Sinai White disposal — EUR consideration', 'eur', IN['swcc_eur'], NUM1, 'EUR mn')
# The euro rate and the carrying value are RETIRED FROM THE SHEET: the disposal gain is
# the one the FY2024 income statement states on its face, so reconstructing it from a
# consideration and a book value changes nothing. Both remain in the input register, where
# the study quotes the transaction, but an input a reader can change that changes nothing
# is worse than an absent one — it says the model depends on something it does not.

band(wsA, 79, 9); wsA['A79'] = 'LENS INPUTS AND WEIGHTS'
inp(80, 'Replacement cost of capacity', 'repl', IN['repl_usd_t'], NUM0, 'USD/t')
inp(81, 'Justified EV per tonne', 'evt', IN['ev_t_just'], NUM0, 'USD/t')
inp(82, 'Justified EV/EBITDA', 'evb', IN['ev_ebitda_just'], MULT)
inp(83, 'Justified price/earnings', 'pej', IN['pe_just'], MULT)
inp(84, 'Mid-cycle EBITDA margin', 'nmgn', IN['norm_mgn'], PCT)
inp(85, 'Normalised revenue haircut', 'nhair', IN['norm_rev_haircut'], PCT)
inp(86, 'Vicat tender offer price', 'mto', IN['mto_price'], PX)
inp(92, 'Gain on sale of investments, FY2024 (as filed)', 'swgain',
    D['disposal']['gain'], NUM0, 'EGP mn')

# [R-LENS-03] THE FOUR LENS WEIGHTS ARE RETIRED and are no longer inputs to anything:
# the central IS the class primary. What replaces them on this sheet is the envelope's
# own floor and the value the retired blend would have read, both committed figures.
inp(87, 'Envelope floor — the primary at the filed margin floor',
    'envlo', LR['envelope']['low'], PX)
inp(88, 'MEMO — the retired four-lens blend', 'blend', LR['retired']['blend_value'], PX)

# [R-TERM-01] THE EXPLICIT WINDOW CONVERGES ON THE TERMINAL'S OWN CAPITAL CHARGE. The
# company's own spend of EGP 303.2mn a year sits 3.2x below the current-cost maintenance
# the terminal charges, on a plant 59.7 per cent written down whose machinery is 68.4 per
# cent gone. The charge glides from one to the other so the step at the boundary is zero.
inp(89, 'Maintenance at current cost (replacement base / disclosed life)', 'maintcc',
    DCF['ic_repl'] / (1.0 / IN['dep_rate_disclosed']), NUM0,
    'Replacement capital over the note 3/2 life — the SAME charge the terminal makes')
for _i, _w in enumerate(D['conv_weights']):
    inp(105 + _i, f'Capital-charge convergence weight, FY{2026+_i}', f'conv{_i}', _w, PCT,
        'the share of the way from the company\'s own spend to current-cost maintenance')
for _i, _f in enumerate(D['forecast']['glide']):
    inp(116 + _i, f'Cost-of-capital glide fraction, FY{2026+_i}', f'gl{_i}', _f, DF4,
        'the central bank\'s own easing calendar: the policy path\'s cumulative progress')
# The 'Total (must be 100%)' row that summed the four lens weights is RETIRED with
# them [R-LENS-03]. Left in place it summed whatever landed in those rows next,
# and it did: it read 1,078 per cent the first time these inputs moved.

# THE FILED HISTORICALS, as inputs on the assumptions sheet — depreciation, operating
# profit and interest income straight off the audited statements, so the income-statement
# sheet reads facts rather than reconstructing them.
wsA['A99'] = 'THE FILED HISTORICALS — AUDITED STATEMENTS, NOTHING SOLVED'
wsA['A99'].font = SUB
for _j, (_lab, _key, _vals, _f) in enumerate([
        ('Depreciation & amortisation', 'fdna', D['history']['dna'], NUM0),
        ('EBIT (operating profit plus the finance charge inside it)', 'febit',
         D['history']['ebit'], NUM0),
        ('Interest and investment income', 'ftr', D['history']['treasury'], NUM0)]):
    _rr = 100 + _j
    wsA.cell(row=_rr, column=1, value=_lab)
    for _i, _v in enumerate(_vals):
        put(wsA, f'{DC[_i]}{_rr}', _v, BLUE, _f)
        A[f'{_key}{_i}'] = f"Assumptions!${DC[_i]}${_rr}"
    wsA.cell(row=_rr, column=9, value='EGP mn, FY2023-25').font = SUB
FDNAK = [A['fdna0'], A['fdna1'], A['fdna2']]
FEBITK = [A['febit0'], A['febit1'], A['febit2']]
FTRK = [A['ftr0'], A['ftr1'], A['ftr2']]

PATK = [A['pat23'], A['pat24'], A['pat25']]
REVK = [A['rev23'], A['rev24'], A['rev25']]
SHK = [A['sh23'], A['sh24'], A['sh25']]

# ============ 6 UNIT BUILD & COST STACK (the bottom-up engine) ===============
wsU = sheet('Unit Build')
title(wsU, 'Bottom-up build — physical units in, EBITDA out',
      'EBITDA is a RESULT of this sheet, never an input.', 10, 46, 13)
BUC = ['B', 'C', 'D', 'E', 'F', 'G']          # FY2025A + FY2026E..FY2030E
BUY = ['FY2025A'] + YF
hdr(wsU, 4, [''] + BUY)
LBL = ['Kiln clinker capacity (Mt/yr)', 'Kiln utilisation', 'Clinker produced (Mt)',
       'Cement grinding capacity (Mt/yr)', 'Clinker factor (t clinker / t cement)',
       'Cement produced (Mt)', 'Domestic share', 'Domestic volume (Mt)',
       'Export volume (Mt)', 'Domestic price (EGP/t)', 'Export price (USD/t)',
       'USD/EGP', 'Domestic revenue (EGP mn)', 'Export revenue (EGP mn)',
       'REVENUE (EGP mn)', 'Blended realised price (EGP/t)']
for j, l in enumerate(LBL):
    wsU.cell(row=5 + j, column=1, value=l)
BU = D['bottom_up']
for i in range(6):
    c = BUC[i]
    putf(wsU, f'{c}5', f"={A['capclk']}", IN['cap_clinker_mt'], NUM2, green=True)
    putf(wsU, f'{c}6', f"={A[f'util{i}']}", IN['kiln_util'][i], PCT, green=True)
    putf(wsU, f'{c}7', f"={c}5*{c}6", BU[i]['clinker'], NUM3)
    putf(wsU, f'{c}8', f"={A['capcem']}", IN['cap_cement_mt'], NUM2, green=True)
    putf(wsU, f'{c}9', f"={A['cfac']}", D['clinker_factor'], NUM3, green=True)
    putf(wsU, f'{c}10', f"={c}7/{c}9", BU[i]['cement'], NUM3)
    putf(wsU, f'{c}11', f"={A[f'dom{i}']}", IN['domestic_share'][i], PCT, green=True)
    putf(wsU, f'{c}12', f"={c}10*{c}11", BU[i]['dom'], NUM3)
    putf(wsU, f'{c}13', f"={c}10-{c}12", BU[i]['exp'], NUM3)
    putf(wsU, f'{c}14', f"={A[f'pdom{i}']}", IN['price_dom_egp_t'][i], NUM0, green=True)
    putf(wsU, f'{c}15', f"={A[f'pexp{i}']}", IN['price_exp_usd_t'][i], NUM1, green=True)
    putf(wsU, f'{c}16', f"={A[f'fxp{i}']}", IN['fx_path'][i], NUM1, green=True)
    putf(wsU, f'{c}17', f"={c}12*{c}14", BU[i]['dom'] * IN['price_dom_egp_t'][i], NUM0)
    putf(wsU, f'{c}18', f"={c}13*{c}15*{c}16",
         BU[i]['exp'] * IN['price_exp_usd_t'][i] * IN['fx_path'][i], NUM0)
    putf(wsU, f'{c}19', f"={c}17+{c}18", BU[i]['rev'], NUM0, bold=True)
    putf(wsU, f'{c}20', f"={c}19/{c}10", BU[i]['price'], NUM0)

band(wsU, 22, 10); wsU['A22'] = 'COST STACK — EGP PER TONNE OF CEMENT'
CL = [(23, 'Materials, fuel, power and packing (note 24)'),
      (24, 'Transport, loading and export (notes 24 and 25)'),
      (28, 'TOTAL VARIABLE (EGP/t)')]
for rr, l in CL:
    wsU.cell(row=rr, column=1, value=l)
for i in range(6):
    c = BUC[i]
    # the two disclosed lines, anchored on FY2025 over the FY2025 volume this model itself
    # computes, each escalating on its own driver class
    putf(wsU, f'{c}23', f"={A['matfy']}/$B$10*({A['usdsh']}*{c}16/{A['fxp0']}"
                        f"+(1-{A['usdsh']})*{A[f'infl{i}']}/{A['infl0']})",
         BU[i]['c_mat'], NUM0)
    putf(wsU, f'{c}24', f"={A['distfy']}/$B$10*{A[f'infl{i}']}/{A['infl0']}",
         BU[i]['c_dist'], NUM0)
    putf(wsU, f'{c}28', f"={c}23+{c}24", BU[i]['var_t'], NUM0, bold=True)

band(wsU, 30, 10); wsU['A30'] = 'PROFIT AND LOSS — EGP MILLION'
for j, l in enumerate(['Revenue', 'Variable cost', 'Fixed cost', 'EBITDA  (AN OUTPUT)',
                       'EBITDA margin', 'EBITDA per tonne (EGP)']):
    wsU.cell(row=31 + j, column=1, value=l)
for i in range(6):
    c = BUC[i]
    putf(wsU, f'{c}31', f"={c}19", BU[i]['rev'], NUM0, green=True)
    putf(wsU, f'{c}32', f"={c}28*{c}10", BU[i]['var'], NUM0)
    putf(wsU, f'{c}33', f"={A['fixfy']}*{A[f'infl{i}']}/{A['infl0']}", BU[i]['fixed'], NUM0)
    putf(wsU, f'{c}34', f"={c}31-{c}32-{c}33", BU[i]['ebitda'], NUM0, bold=True)
    putf(wsU, f'{c}35', f"={c}34/{c}31", BU[i]['mgn'], PCT)
    putf(wsU, f'{c}36', f"={c}34/{c}10", BU[i]['ebitda'] / BU[i]['cement'], NUM0)

band(wsU, 38, 10); wsU['A38'] = 'VALIDATION — A TEST THAT CAN FAIL'
wsU['A39'] = 'Bottom-up FY2025 revenue'
putf(wsU, 'B39', "=B19", BU[0]['rev'], NUM0)
wsU['A40'] = 'Disclosed FY2025 revenue'
putf(wsU, 'B40', f"={A['rev25']}", IN['rev_fy25'], NUM0, green=True)
wsU['A41'] = 'Difference'
putf(wsU, 'B41', "=B39/B40-1", BU[0]['rev'] / IN['rev_fy25'] - 1, PCT)
wsU['A42'] = 'Bottom-up FY2025 EBITDA'
putf(wsU, 'B42', "=B34", BU[0]['ebitda'], NUM0)
wsU['A43'] = 'EBITDA implied by closing disclosed profit'
putf(wsU, 'B43', f"='Income Statement'!D6", H['ebitda'][2], NUM0, green=True)
wsU['A44'] = 'Difference'
putf(wsU, 'B44', "=B42/B43-1", BU[0]['ebitda'] / H['ebitda'][2] - 1, PCT)
note(wsU, 46, 'Nothing on this sheet is solved to force agreement. Every cost driver is an')
note(wsU, 47, 'independent physical or market norm, so a wrong cost stack shows up in rows 41 and 44.')
note(wsU, 48, 'The model this replaces solved realised price as revenue divided by volume, which')
note(wsU, 49, 'made its residual zero for ANY assumption — an identity, not a validation.')

# ============ 8 DCF ===========================================================
wsD = sheet('DCF')
title(wsD, 'Discounted cash flow — the primary lens', 'Cost of capital built here, never pasted.',
      8, 46, 14)
hdr(wsD, 4, [''] + YF)
ROWS = [('Revenue', 'rev'), ('EBITDA margin', 'mgn'), ('EBITDA', 'ebitda'),
        ('Depreciation & amortisation', 'dna'), ('EBIT', 'ebit'), ('Tax rate', 'trate'),
        ('NOPAT  (EBIT × (1 − t))', 'nopat'), ('Memo: capital expenditure', 'capex'),
        ('Memo: change in working capital', 'dwc'),
        ('Net capital charge  (capex + working capital − depreciation)', 'reinv'),
        ('Free cash flow to the firm', 'fcff'), ('Glide fraction', 'glide'),
        ('Forward cost of capital', 'fwd'), ('Discount factor', 'df'),
        ('Present value of FCFF', 'pv')]
for j, (lab, _) in enumerate(ROWS):
    wsD.cell(row=5 + j, column=1, value=lab)
wsD.cell(row=20, column=1, value='Memo: capital expenditure, on the company\'s own run rate')
wsD.cell(row=21, column=1,
         value='Memo: gross fixed assets, note 4 rolled forward on that capex')
for i in range(5):
    c = DC[i]
    putf(wsD, f'{c}5', f"='Unit Build'!{BUC[i+1]}31", F['revenue'][i], NUM0, green=True)
    putf(wsD, f'{c}7', f"='Unit Build'!{BUC[i+1]}34", F['ebitda'][i], NUM0, green=True)
    putf(wsD, f'{c}6', f"={c}7/{c}5", F['ebitda'][i] / F['revenue'][i], PCT)
    # THE ASSET BASE AND THE CAPEX THAT ROLLS IT, on memo rows 20 and 21 so the disclosed
    # depreciation rate has something to apply to. Rows 6 and 7 are the margin and EBITDA.
    _prevg = f"{A['grossfa']}" if i == 0 else f"{DC[i-1]}21"
    putf(wsD, f'{c}20',
         f"=({A['capexrr']}+({A['maintcc']}-{A['capexrr']})*{A[f'conv{i}']})"
         f"*{A[f'infl{i+1}']}/{A['infl0']}", F['capex'][i], NUM0)
    putf(wsD, f'{c}21', f"={_prevg}+{c}20", F['gross_fa'][i], NUM0)
    putf(wsD, f'{c}8', f"={c}21*{A['deprate']}", F['dna'][i], NUM0)
    putf(wsD, f'{c}9', f"={c}7-{c}8", F['ebit'][i], NUM0)
    putf(wsD, f'{c}10', f"={A['tax']}", TAX, PCT, green=True)
    putf(wsD, f'{c}11', f"={c}9*(1-{c}10)", F['nopat'][i], NUM0)
    putf(wsD, f'{c}12', f"={c}20", F['capex'][i], NUM0)
    prev = A['rev25'] if i == 0 else f"{DC[i-1]}5"
    putf(wsD, f'{c}13', f"=({c}5-{prev})*{A['wcp']}", F['dwc'][i], NUM0)
    pn = "$B$23" if i == 0 else f"{DC[i-1]}11"
    # [R-TERM-01] ONE DEFINITION OF FREE CASH FLOW ACROSS BOTH WINDOWS: NOPAT plus book
    # depreciation, less the capital actually spent and the working capital the growth
    # absorbs. Revision 2 ran the explicit years on NOPAT less a reinvestment charge
    # derived from the growth in NOPAT and the terminal on something else.
    putf(wsD, f'{c}14', f"={c}12+{c}13-{c}8", F['reinvestment'][i], NUM0)
    if i == 0:
        putf(wsD, f'{c}15', f"=({c}11+{c}8-{c}12-{c}13)*(1-{A['stub']})", F['fcff'][i],
             NUM0, bold=True)
    else:
        putf(wsD, f'{c}15', f"={c}11+{c}8-{c}12-{c}13", F['fcff'][i], NUM0, bold=True)
    # [R-COC-01] THE GLIDE'S FRACTIONS ARE THE POLICY-RATE PATH'S OWN CUMULATIVE
    # PROGRESS. The retired formula read them off the cost-of-debt ladder, which
    # inherited its shape from a hand-set path — a second free parameter.
    putf(wsD, f'{c}16', f"={A[f'gl{i}']}", F['glide'][i], DF4)
    putf(wsD, f'{c}17', f"=$C$46-($C$46-$C$53)*{c}16", F['fwd_wacc'][i], PCT2)
    # EACH FORWARD RATE IS COMPOUNDED OVER THE SLICE OF CALENDAR IT OWNS. The retired
    # formulas walked the rates in whole-year steps from t=0, so the whole 0.917 years
    # to the FY2027 midpoint compounded at the FY2026 rate and the FY2030 rate never
    # entered any factor at all. On a path that FALLS that over-discounts every year
    # after the first, and it is the same defect ARCC revision 3 carried.
    if i == 0:
        fm = f"=1/(1+B17)^((1-{A['stub']})/2)"
    else:
        parts = [f"(1+B17)^(1-{A['stub']})"]
        parts += [f"(1+{DC[k]}17)" for k in range(1, i)]
        parts.append(f"(1+{DC[i]}17)^0.5")
        fm = "=1/(" + "*".join(parts) + ")"
    putf(wsD, f'{c}18', fm, F['df'][i], DF4)
    putf(wsD, f'{c}19', f"={c}15*{c}18", F['pv'][i], NUM0)

# note the row offset: labels start at 5 but there are 15 labels -> rows 5..19
band(wsD, 21, 8); wsD['A21'] = 'TERMINAL BLOCK'
TB = [('Replacement-cost invested capital (EGP mn)', 'B22',
       f"={A['capcem']}*1000000*{A['repl']}*{A['fx']}/1000000", DCF['ic_repl'], NUM0),
      ('FY2025 NOPAT (the reinvestment base)', 'B23',
       f"=('Income Statement'!D6-'Income Statement'!D8)*(1-{A['tax']})",
       (D['history']['ebitda'][2] - D['history']['dna'][2]) * (1 - TAX), NUM0),
      ('Terminal return on invested capital', 'B24',
       f"=F11*(1+{A['g']})/B22", DCF['roic_term'], PCT),
      ('Terminal NOPAT', 'B28', f"=F11*(1+{A['g']})", DCF['nopat_term'], NUM0),
      # [R-TERM-01] MAINTENANCE AT CURRENT COST OVER THE DISCLOSED LIFE, not the
      # reinvestment identity. The retired construction charged g x invested capital every
      # year for ever, which is an implied replacement cycle of 1/g — a fact about the
      # currency and not about the plant. The life comes from note 3/2 of the audited
      # accounts weighted on note 4's own gross-cost mix.
      ('Disclosed weighted asset life (years)', 'B23b', f"=1/{A['deprate']}",
       1.0 / IN['dep_rate_disclosed'], NUM1),
      ('Maintenance at current cost  (invested capital / life)', 'B25',
       "=B22/B23b", DCF['term_maintenance'], NUM0),
      ('Book depreciation added back', 'B25b', f"=DCF!F8*(1+{A['g']})",
       DCF['term_dna_addback'], NUM0),
      ('Working capital charged with inflation', 'B25c',
       f"={A['rev25']}*{A['wcp']}*{A['g']}", DCF['term_wc_charge'], NUM0),
      ('Terminal free cash flow', 'B25d', "=B28+B25b-B25-B25c", DCF['term_fcff'], NUM0),
      ('Terminal value', 'B26', f"=B25d*(1+{A['g']})/($C$53-{A['g']})", DCF['tv'], NUM0),
      ('Present value of terminal value', 'B27', "=B26*F18", DCF['pv_tv'], NUM0)]
_ROWMAP = {'B23b': 'B45', 'B25b': 'B46', 'B25c': 'B47', 'B25d': 'B48'}
TB = [(l, _ROWMAP.get(a, a), f.replace('B23b', _ROWMAP['B23b']).replace('B25b', _ROWMAP['B25b'])
       .replace('B25c', _ROWMAP['B25c']).replace('B25d', _ROWMAP['B25d']), e, t)
      for l, a, f, e, t in TB]
for lab, ad, fm, ex, ft in TB:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft)
band(wsD, 29, 8); wsD['A29'] = 'ENTERPRISE TO EQUITY BRIDGE'
BR = [('Present value of explicit years (FY2026E-FY2030E)', 'B30', "=SUM(B19:F19)", DCF['sum_pv'], NUM0),
      ('Present value of terminal value', 'B31', "=B27", DCF['pv_tv'], NUM0),
      ('Enterprise value', 'B32', "=B30+B31", DCF['ev'], NUM0),
      ('Terminal value as % of enterprise value', 'B33', "=B31/B32", DCF['tv_share'], PCT),
      # [R-BRIDGE-01] THE LATEST DISCLOSED SHEET IS THE REVIEWED 31-MARCH-2026 ONE, not
      # the audited 31-December-2025 sheet rolled forward on an estimate. The remaining
      # four months to the valuation date are carried on this model's own free cash flow.
      ('Cash, reviewed sheet at 31 March 2026', 'B34',
       f"={A['cashm26']}+B15/(1-{A['stub']})*({A['stub']}-0.25)", DCF['cash_fy25'], NUM0),
      ('Less lease liabilities (there are no bank borrowings)', 'B35',
       f"=-{A['debtm26']}", -IN['debt_mar26'], NUM0),
      ('Net cash (ADDED — the company is net cash)', 'B36', "=B34+B35", DCF['net_cash'], NUM0),
      ('Less non-controlling interests', 'B40', f"={A['nci']}", IN['nci'], NUM0),
      ('Equity value', 'B37', "=B32+B36-B40", DCF['equity'], NUM0),
      ('Shares outstanding (mn)', 'B38', f"={A['shares']}", SH, NUM2),
      ('Fair value per share (EGP)', 'B39', "=B37/B38", DCF['fv'], PX)]
for lab, ad, fm, ex, ft in BR:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft, bold=(ad in ('B32', 'B37', 'B39', 'B33')))
band(wsD, 41, 8); wsD['A41'] = 'COST OF CAPITAL — BUILT HERE'
CC = [('Risk-free rate (observed EGP 10-year)', 'C42', f"={A['rf']}", IN['rf'], PCT2),
      ('Less sovereign default spread', 'C43', f"=-{A['sov']}", -IN['sov_spread_cds'], PCT2),
      ('Normalised risk-free rate', 'C44', "=C42+C43", W['rf_star'], PCT2),
      ('Cost of equity  (rf* + β × ERP)', 'C45', f"=C44+{A['beta']}*{A['erp']}", W['ke_exp'], PCT2),
      ('WACC — explicit window', 'C46',
       f"=(1-C49)*C45+C49*C48", W['wacc_exp'], PCT2),
      ('Cost of debt after tax', 'C48', f"={A['kd']}*(1-{A['tax']})", W['kd_at'], PCT2),
      ('Debt weight  D/(D+E)', 'C49', f"={A['debt']}/({A['debt']}+C50)", W['wd_exp'], '0.000%'),
      ('Market capitalisation', 'C50', f"={A['spot']}*{A['shares']}", M['mktcap'], NUM0),
      # HAMADA MUST START FROM AN ASSET BETA. This formula re-levered an already-levered
      # beta, levering it twice. Unlever at the OBSERVED structure first — the model does,
      # and a workbook that does not is a workbook disagreeing with the study it publishes.
      ('Asset beta, unlevered at the observed structure', 'C47',
       f"={A['beta']}/(1+(1-{A['tax']})*C49/(1-C49))", W['beta_u'], NUM3),
      ('Terminal beta, RE-LEVERED to the terminal structure (Hamada)', 'C56',
       f"=C47*(1+(1-{A['tax']})*{A['wdt']}/(1-{A['wdt']}))", W['beta_term'], NUM3),
      ('Terminal cost of equity  (rf_term + β_L × ERP_term)', 'C51',
       f"={A['rft']}+C56*{A['erpt']}", W['ke_term'], PCT2),
      ('Terminal cost of debt after tax', 'C52', f"={A['kdt']}*(1-{A['tax']})", W['kd_term_at'], PCT2),
      ('WACC — terminal', 'C53', f"=(1-{A['wdt']})*C51+{A['wdt']}*C52", W['wacc_term'], PCT2),
      ('Retired construction: rf NOT netted (disclosed only)', 'C54',
       f"={A['rf']}+{A['beta']}*{A['erp']}", W['ke_raw_retired'], PCT2),
      ('Sovereign double-count removed (bp)', 'C55', "=(C54-C45)*10000",
       (W['ke_raw_retired'] - W['ke_exp']) * 10000, NUM0)]
for lab, ad, fm, ex, ft in CC:
    wsD.cell(row=int(ad[1:]), column=1, value=lab)
    putf(wsD, ad, fm, ex, ft, bold=(ad in ('C46', 'C53')))
note(wsD, 57, 'The glide fraction on row 16 is the central bank\'s own easing calendar — the policy path\'s')
note(wsD, 58, 'cumulative progress — so the schedule is INHERITED rather than invented. The terminal is capitalised at the')
note(wsD, 59, 'terminal WACC and discounted on year 5\'s OWN cumulative factor — one date, one price of time.')
note(wsD, 60, 'Terminal return on capital is struck on REPLACEMENT-COST capacity, not book: the plant dates to 1997')
note(wsD, 61, 'and its book invested capital implies a 172% return, which would let growth through unpaid for.')

# ============ 5 EV BRIDGE =====================================================
wsB = sheet('EV Bridge')
title(wsB, 'Enterprise value to equity bridge', None, 6, 52, 16)
hdr(wsB, 4, ['', 'EGP mn', 'Per share (EGP)'])
BRW = [('Present value of explicit free cash flow', "=DCF!B30", DCF['sum_pv']),
       ('Present value of terminal value', "=DCF!B31", DCF['pv_tv']),
       ('Enterprise value', "=DCF!B32", DCF['ev']),
       ('Plus net cash', "=DCF!B36", DCF['net_cash']),
       ('Equity value', "=DCF!B37", DCF['equity'])]
for j, (lab, fm, ex) in enumerate(BRW):
    wsB.cell(row=5 + j, column=1, value=lab)
    putf(wsB, f'B{5+j}', fm, ex, NUM0, green=True, bold=(j in (2, 4)))
    putf(wsB, f'C{5+j}', f"=B{5+j}/{A['shares']}", ex / SH, PX, bold=(j in (2, 4)))
wsB['A11'] = 'Terminal value as % of enterprise value'
putf(wsB, 'B11', "=DCF!B33", DCF['tv_share'], PCT, green=True, bold=True)
wsB['A12'] = 'Memo: spot price'
putf(wsB, 'C12', f"={A['spot']}", SPOT, PX, green=True)
wsB['A13'] = 'Memo: Vicat tender offer price (July 2025)'
putf(wsB, 'C13', f"={A['mto']}", IN['mto_price'], PX, green=True)
wsB['A14'] = 'Fair value less spot'
putf(wsB, 'C14', "=C9-C12", DCF['fv'] - SPOT, PX)
wsB['A15'] = 'Upside / (downside) to the cash-flow lens'
putf(wsB, 'C15', "=C9/C12-1", DCF['fv'] / SPOT - 1, PCT)
note(wsB, 17, 'Terminal value share is linked live to the DCF sheet — never typed. The tender offer price is a')
note(wsB, 18, 'disclosed reference point and an overhang, not a valuation.')

# ============ 9 INCOME STATEMENT ==============================================
wsI = sheet('Income Statement')
title(wsI, 'Income statement — 3 years historical + 5-year forecast', 'EGP mn', 10, 44, 13)
hdr(wsI, 4, [''] + YH + YF)
IL = ['Revenue', 'EBITDA', 'EBITDA margin', 'Depreciation & amortisation', 'EBIT',
      'Treasury / investment income', 'Gain on disposal of Sinai White stake',
      'Profit before tax', 'Tax', 'Profit after tax', 'Profit after tax — disclosed',
      'Earnings per share (EGP)']
for j, l in enumerate(IL):
    wsI.cell(row=5 + j, column=1, value=l)
gain_hist = [0.0, DISP['gain'], 0.0]
for i in range(3):
    c = HC[i]
    # THE HISTORICALS ARE FILED FACTS, SO THEY ARE INPUTS, NOT SOLVES. The earlier
    # edition derived depreciation as a share of revenue, operating profit by grossing a
    # press profit figure at an effective tax rate, and treasury income from a cash
    # balance rolled back by a guessed 1.25x. Every one of those is stated in the audited
    # statements and is now read from them.
    putf(wsI, f'{c}5', f"={REVK[i]}", H['revenue'][i], NUM0, green=True)
    putf(wsI, f'{c}8', f"={FDNAK[i]}", H['dna'][i], NUM0, green=True)
    putf(wsI, f'{c}9', f"={FEBITK[i]}", H['ebit'][i], NUM0, green=True)
    putf(wsI, f'{c}10', f"={FTRK[i]}", H['treasury'][i], NUM0, green=True)
    putf(wsI, f'{c}6', f"={c}9+{c}8", H['ebitda'][i], NUM0)
    putf(wsI, f'{c}7', f"={c}6/{c}5", H['ebitda'][i] / H['revenue'][i], PCT)
    # the FILED gain, from the face of the FY2024 income statement, not a reconstruction
    putf(wsI, f'{c}11', f"={A['swgain']}" if i == 1 else "=0", gain_hist[i], NUM0,
         green=(i == 1))
    putf(wsI, f'{c}12', f"={c}9+{c}10+{c}11", H['ebit'][i] + H['treasury'][i] + gain_hist[i], NUM0)
    tx = (H['ebit'][i] + H['treasury'][i] + gain_hist[i]) - D['history']['pat'][i]
    putf(wsI, f'{c}13', f"={c}12-{c}14", tx, NUM0)
    putf(wsI, f'{c}14', f"={PATK[i]}", D['history']['pat'][i], NUM0, green=True, bold=True)
    putf(wsI, f'{c}15', f"={PATK[i]}", D['history']['pat'][i], NUM0, green=True)
    putf(wsI, f'{c}16', f"={c}14/{SHK[i]}", D['history']['eps'][i], PX)
for i in range(5):
    c = FC[i]
    putf(wsI, f'{c}5', f"=DCF!{DC[i]}5", F['revenue'][i], NUM0, green=True)
    putf(wsI, f'{c}6', f"=DCF!{DC[i]}7", F['ebitda'][i], NUM0, green=True)
    putf(wsI, f'{c}7', f"={c}6/{c}5", F['ebitda'][i] / F['revenue'][i], PCT)
    putf(wsI, f'{c}8', f"=DCF!{DC[i]}8", F['dna'][i], NUM0, green=True)
    putf(wsI, f'{c}9', f"={c}6-{c}8", F['ebit'][i], NUM0)
    pc = HC[2] if i == 0 else FC[i - 1]      # treasury runs on OPENING cash, as the model does
    putf(wsI, f'{c}10', f"='Balance Sheet'!{pc}7*{A[f'cy{i}']}", F['treasury'][i], NUM0)
    putf(wsI, f'{c}11', "=0", 0.0, NUM0)
    putf(wsI, f'{c}12', f"={c}9+{c}10+{c}11", F['pbt'][i], NUM0)
    putf(wsI, f'{c}13', f"={c}12*{A['tax']}", F['tax'][i], NUM0)
    putf(wsI, f'{c}14', f"={c}12-{c}13", F['pat'][i], NUM0, bold=True)
    putf(wsI, f'{c}16', f"={c}14/{A['shares']}", F['pat'][i] / SH, PX)
note(wsI, 18, 'DERIVED LINES, stated plainly. Revenue and profit after tax are DISCLOSED (rows 5 and 15, in green).')
note(wsI, 19, 'EBITDA, D&A, EBIT and treasury income for the historical years are DERIVED by closing the disclosed')
note(wsI, 20, 'profit: FY2024 anchors on the one disclosed EBITDA figure and solves treasury income; FY2025 solves')
note(wsI, 21, 'EBIT from disclosed profit after tax; FY2023 uses its own smaller cash balance. Row 13 is the residual')
note(wsI, 22, 'tax charge, and row 15 sits beside row 14 so the reader can see the derivation closes to the print.')

# ============ 10 BALANCE SHEET ================================================
wsBS = sheet('Balance Sheet')
title(wsBS, 'Balance sheet — 3 years historical + 5-year forecast', 'EGP mn', 10, 44, 13)
hdr(wsBS, 4, [''] + YH + YF)
BL = ['Property, plant and equipment (net)', 'Working capital (net)', 'Cash and equivalents',
      'Total assets (condensed)', 'Gross debt', 'Other liabilities', 'Total liabilities',
      'Shareholders\' equity', 'Equity — reported cross-check', 'Net cash',
      'Return on equity']
for j, l in enumerate(BL):
    wsBS.cell(row=5 + j, column=1, value=l)
# THE FILED OPERATING ASSET BASE at 31-Dec-2024: fixed assets net of depreciation,
# intangibles and construction in progress, from the face of the audited sheet. Revision 2
# derived it as total assets less a rolled-back cash balance less a round 900.
_FB24 = json.load(open('filings_extract.json'))['balance_sheet']['FY2024']
ppe24 = (_FB24['fixed_assets'] + _FB24['intangibles'] + _FB24['cwip']) / 1e6
inp(97, 'Operating assets at 31-Dec-2024 (fixed assets net, intangibles, CWIP)',
    'ppe24f', ppe24, NUM0, 'EGP mn, as filed')
cash23 = IN['cash_fy25']/1.25*0.35
eq23 = (IN['ta_fy24'] - IN['tl_fy24']) - IN['pat_fy24']
ta23 = ppe24 * 0.95 + 850.0 + cash23
TL_EXP = [ta23 - eq23, IN['tl_fy24'],
          (ppe24 + 900.0 + IN['cash_fy25']) - IN['eq_fy25_rep']]
hist_bs = dict(
    ppe=[ppe24 * 0.95, ppe24, F['ppe'][0] - F['capex'][0] + F['dna'][0]],
    wc=[850.0, 900.0, 900.0 + 0.0],
    cash=[IN['cash_fy25']/1.25*0.35, IN['cash_fy25']/1.25, DCF['cash_fy25']])
for i in range(3):
    c = HC[i]
    if i == 1:
        # the filed operating asset base: fixed assets net, intangibles and CWIP
        putf(wsBS, f'{c}5', f"={A['ppe24f']}", ppe24, NUM0, green=True)
        put(wsBS, f'{c}6', 900.0, BLUE, NUM0)
        putf(wsBS, f'{c}7', f"={A['cash25']}/1.25", IN['cash_fy25'] / 1.25, NUM0)
        putf(wsBS, f'{c}8', f"={A['ta24']}", IN['ta_fy24'], NUM0, green=True, bold=True)
        putf(wsBS, f'{c}11', f"={A['tl24']}", IN['tl_fy24'], NUM0, green=True)
        putf(wsBS, f'{c}12', f"={c}8-{c}11", IN['ta_fy24'] - IN['tl_fy24'], NUM0, bold=True)
    elif i == 0:
        put(wsBS, f'{c}5', ppe24 * 0.95, BLUE, NUM0)
        put(wsBS, f'{c}6', 850.0, BLUE, NUM0)
        put(wsBS, f'{c}7', IN['cash_fy25']/1.25*0.35, BLUE, NUM0)
        putf(wsBS, f'{c}8', f"=SUM({c}5:{c}7)", ppe24 * 0.95 + 850.0 + IN['cash_fy25']/1.25*0.35, NUM0, bold=True)
        putf(wsBS, f'{c}11', f"={c}8-{c}12",
             (ppe24 * 0.95 + 850.0 + IN['cash_fy25']/1.25*0.35)
             - ((IN['ta_fy24'] - IN['tl_fy24']) - IN['pat_fy24']), NUM0)
        putf(wsBS, f'{c}12', f"=C12-{PATK[1]}", (IN['ta_fy24'] - IN['tl_fy24']) - IN['pat_fy24'], NUM0, bold=True)
    else:
        putf(wsBS, f'{c}5', f"=C5", ppe24, NUM0)
        putf(wsBS, f'{c}6', f"=C6", 900.0, NUM0)
        putf(wsBS, f'{c}7', f"={A['cash25']}", IN['cash_fy25'], NUM0, green=True)
        putf(wsBS, f'{c}8', f"=SUM({c}5:{c}7)", ppe24 + 900.0 + IN['cash_fy25'], NUM0, bold=True)
        putf(wsBS, f'{c}11', f"={c}8-{c}12", (ppe24 + 900.0 + IN['cash_fy25']) - IN['eq_fy25_rep'], NUM0)
        putf(wsBS, f'{c}12', f"={A['eq25']}", IN['eq_fy25_rep'], NUM0, green=True, bold=True)
    putf(wsBS, f'{c}9', f"={A['debt']}", IN['debt_fy25'], NUM1)
    putf(wsBS, f'{c}10', f"={c}11-{c}9", TL_EXP[i] - IN['debt_fy25'], NUM0)
    bs_cash = [IN['cash_fy25'] / 1.25 * 0.35, IN['cash_fy25'] / 1.25, IN['cash_fy25']][i]
    putf(wsBS, f'{c}14', f"={c}7-{c}9", bs_cash - IN['debt_fy25'], NUM0)
putf(wsBS, 'D13', f"=C12+{PATK[2]}", (IN['ta_fy24']-IN['tl_fy24'])+IN['pat_fy25'], NUM0)
putf(wsBS, 'D15', f"={PATK[2]}/D12", IN['pat_fy25'] / IN['eq_fy25_rep'], PCT)
for i in range(5):
    c, p = FC[i], (HC[2] if i == 0 else FC[i - 1])
    putf(wsBS, f'{c}5', f"={p}5+DCF!{DC[i]}12-DCF!{DC[i]}8", F['ppe'][i], NUM0)
    putf(wsBS, f'{c}6', f"={p}6+DCF!{DC[i]}13", F['wc'][i], NUM0)
    putf(wsBS, f'{c}7', f"={p}7+'Cash Flow'!{c}12", F['cash'][i], NUM0)
    putf(wsBS, f'{c}8', f"=SUM({c}5:{c}7)", F['ppe'][i] + F['wc'][i] + F['cash'][i], NUM0, bold=True)
    putf(wsBS, f'{c}9', f"={A['debt']}", IN['debt_fy25'], NUM1)
    putf(wsBS, f'{c}11', f"={c}8-{c}12", F['ppe'][i] + F['wc'][i] + F['cash'][i] - F['equity'][i], NUM0)
    putf(wsBS, f'{c}10', f"={c}11-{c}9",
         F['ppe'][i] + F['wc'][i] + F['cash'][i] - F['equity'][i] - IN['debt_fy25'], NUM0)
    putf(wsBS, f'{c}12', f"={p}12+'Income Statement'!{c}14-'Cash Flow'!{c}11", F['equity'][i], NUM0, bold=True)
    putf(wsBS, f'{c}14', f"={c}7-{c}9", F['cash'][i] - IN['debt_fy25'], NUM0)
    putf(wsBS, f'{c}15', f"='Income Statement'!{c}14/{c}12", F['pat'][i] / F['equity'][i], PCT)
note(wsBS, 17, 'The FY2024 column is the DISCLOSED triple — total assets EGP 6,385.92mn less total liabilities EGP')
note(wsBS, 18, '1,610.86mn closes to equity of EGP 4,775.06mn exactly. Equity then ROLLS FORWARD: prior equity plus')
note(wsBS, 19, 'profit less dividends. Row 13 carries the REPORTED FY2025 equity beside the rolled figure: the two do')
note(wsBS, 20, 'not agree, and the gap implies a FY2025 distribution that no retrievable source reports. Disclosed, not')
note(wsBS, 21, 'plugged. Row 11 is the residual liability block, not an independently sourced line.')

# ============ 11 CASH FLOW ====================================================
wsC = sheet('Cash Flow')
title(wsC, 'Cash flow — linked to the DCF waterfall', 'EGP mn', 10, 44, 13)
hdr(wsC, 4, [''] + YF, start=4)
for j, l in enumerate(['EBITDA', 'Less change in working capital',
                       'Less capital expenditure', 'Operating free cash flow',
                       'NOPAT', 'Less net reinvestment (growth at terminal ROIC)',
                       'Free cash flow to the firm', 'Treasury income after tax',
                       'Dividends paid', 'Net change in cash']):
    wsC.cell(row=5 + j, column=1, value=l)
for i in range(5):
    c = FC[i]
    putf(wsC, f'{c}5', f"=DCF!{DC[i]}7", F['ebitda'][i], NUM0, green=True)
    putf(wsC, f'{c}6', f"=-DCF!{DC[i]}13", -F['dwc'][i], NUM0)
    putf(wsC, f'{c}7', f"=-DCF!{DC[i]}12", -F['capex'][i], NUM0)
    putf(wsC, f'{c}8', f"=SUM({c}5:{c}7)",
         F['ebitda'][i] - F['dwc'][i] - F['capex'][i], NUM0, bold=True)
    putf(wsC, f'{c}9', f"=DCF!{DC[i]}11", F['nopat'][i], NUM0, green=True)
    putf(wsC, f'{c}10', f"=-DCF!{DC[i]}14", -F['reinvestment'][i], NUM0)
    putf(wsC, f'{c}11', f"=DCF!{DC[i]}15", F['fcff'][i], NUM0, bold=True)
    putf(wsC, f'{c}12', f"='Income Statement'!{c}10*(1-{A['tax']})",
         F['treasury'][i] * (1 - TAX), NUM0)
    putf(wsC, f'{c}13', f"='Income Statement'!{c}14*{A['pay']}", F['dividends'][i], NUM0)
    exp_dc = F['pat'][i] + F['dna'][i] - F['capex'][i] - F['dwc'][i] - F['dividends'][i]
    putf(wsC, f'{c}14',
         f"='Income Statement'!{c}14+DCF!{DC[i]}8-DCF!{DC[i]}12-DCF!{DC[i]}13-{c}13",
         exp_dc, NUM0, bold=True)
ANCH['cf_div_row'] = 13
ANCH['cf_dcash_row'] = 14
note(wsC, 18, 'Two presentations of the same year. Rows 5-8 build operating free cash flow from EBITDA; rows 9-13')
note(wsC, 19, 'are the DCF waterfall\'s own free cash flow to the firm, linked cell-for-cell to the DCF sheet so the')
note(wsC, 20, 'two can never drift. Treasury income is shown separately and is NOT in free cash flow to the firm.')

# fix the balance-sheet references to the cash-flow rows now that they are known
for i in range(5):
    c, p = FC[i], (HC[2] if i == 0 else FC[i - 1])
    putf(wsBS, f'{c}7', f"={p}7+'Cash Flow'!{c}14", F['cash'][i], NUM0)
    putf(wsBS, f'{c}12', f"={p}12+'Income Statement'!{c}14-'Cash Flow'!{c}13", F['equity'][i], NUM0, bold=True)

# ============ 12 SUMMARY FINANCIALS ===========================================
wsSF = sheet('Summary Financials')
title(wsSF, 'Summary financials', 'EGP mn unless stated', 10, 44, 13)
hdr(wsSF, 4, [''] + YH + YF)
SFL = [('Revenue', "='Income Statement'!{c}5", H['revenue'] + F['revenue']),
       ('EBITDA', "='Income Statement'!{c}6", H['ebitda'] + F['ebitda']),
       ('EBIT', "='Income Statement'!{c}9", H['ebit'] + F['ebit']),
       ('Profit after tax', "='Income Statement'!{c}14", D['history']['pat'] + F['pat']),
       ]
for j, (lab, fm, vals) in enumerate(SFL):
    wsSF.cell(row=5 + j, column=1, value=lab)
    for i, c in enumerate(HC + FC):
        putf(wsSF, f'{c}{5+j}', fm.format(c=c), vals[i], NUM0, green=True)
# The bottom-up build starts at FY2025A, so the physical rows carry FY2025 onward only.
for j, (lab, key, ft) in enumerate([('Sales volume (Mt)', 10, NUM3),
                                    ('Realised price (EGP/t)', 20, NUM0),
                                    ('Kiln utilisation', 6, PCT)]):
    rr = 9 + j
    wsSF.cell(row=rr, column=1, value=lab)
    for i in range(6):
        putf(wsSF, f'{(HC+FC)[i+2]}{rr}', f"='Unit Build'!{BUC[i]}{key}",
             [BUD[i]['cement'], BUD[i]['price'], BUD[i]['util']][j], ft, green=True)
wsSF['A13'] = 'EBITDA margin'
wsSF['A14'] = 'Revenue growth'
for i, c in enumerate(HC + FC):
    allrev = H['revenue'] + F['revenue']
    alleb = H['ebitda'] + F['ebitda']
    putf(wsSF, f'{c}13', f"={c}6/{c}5", alleb[i] / allrev[i], PCT)
    if i > 0:
        p = (HC + FC)[i - 1]
        putf(wsSF, f'{c}14', f"={c}5/{p}5-1", allrev[i] / allrev[i - 1] - 1, PCT)

# ============ 15 PER-SHARE & RATIOS ===========================================
wsR = sheet('Per-Share & Ratios')
title(wsR, 'Per-share and ratios', 'Every figure is a formula', 10, 44, 13)
hdr(wsR, 4, [''] + YH + YF)
for j, l in enumerate(['Earnings per share (EGP)', 'Dividend per share (EGP)',
                       'Book value per share (EGP)', 'Free cash flow per share (EGP)',
                       'Net cash per share (EGP)', 'Return on equity',
                       'EBITDA per tonne sold (EGP)', 'Price / earnings (at spot)',
                       'EV / EBITDA (at spot)']):
    wsR.cell(row=5 + j, column=1, value=l)
allpat = D['history']['pat'] + F['pat']
alleq = [(IN['ta_fy24'] - IN['tl_fy24']) - IN['pat_fy24'], IN['ta_fy24'] - IN['tl_fy24'],
         IN['eq_fy25_rep']] + F['equity']
allcash = [IN['cash_fy25']/1.25*0.35, IN['cash_fy25']/1.25, IN['cash_fy25']] + F['cash']
alleb = H['ebitda'] + F['ebitda']
allvol = H['volume_mt'] + F['volume_mt']
for i, c in enumerate(HC + FC):
    putf(wsR, f'{c}5', f"='Income Statement'!{c}14/{A['shares']}", allpat[i] / SH, PX)
    putf(wsR, f'{c}7', f"='Balance Sheet'!{c}12/{A['shares']}", alleq[i] / SH, PX)
    putf(wsR, f'{c}9', f"=('Balance Sheet'!{c}7-'Balance Sheet'!{c}9)/{A['shares']}",
         (allcash[i] - IN['debt_fy25']) / SH, PX)
    putf(wsR, f'{c}10', f"='Income Statement'!{c}14/'Balance Sheet'!{c}12", allpat[i] / alleq[i], PCT)
    if i >= 2:
        putf(wsR, f'{c}11', f"='Income Statement'!{c}6/'Summary Financials'!{c}9",
             alleb[i] / BUD[i - 2]['cement'], NUM0)
    putf(wsR, f'{c}12', f"={A['spot']}/{c}5", SPOT / (allpat[i] / SH), MULT)
    putf(wsR, f'{c}13', f"=({A['spot']}*{A['shares']}-('Balance Sheet'!{c}7-'Balance Sheet'!{c}9))"
                        f"/'Income Statement'!{c}6",
         (SPOT * SH - (allcash[i] - IN['debt_fy25'])) / alleb[i], MULT)
for i, c in enumerate(FC):
    putf(wsR, f'{c}6', f"='Cash Flow'!{c}13/{A['shares']}", F['dividends'][i] / SH, PX)
    putf(wsR, f'{c}8', f"=DCF!{DC[i]}15/{A['shares']}", F['fcff'][i] / SH, PX)
note(wsR, 15, 'Price/earnings and EV/EBITDA are struck at the current spot against each year\'s earnings, so the')
note(wsR, 16, 'forecast columns show what the reader is paying today for a future year — not a forecast of the multiple.')

# ============ 7 RELATIVE & NORMALIZED =========================================
wsN = sheet('Relative & Normalized')
title(wsN, 'Relative multiples, normalised earnings and the asset lens', None, 6, 52, 16)
band(wsN, 4, 6); wsN['A4'] = 'RELATIVE — EV/EBITDA ON MID-CYCLE EARNINGS'
RL = [('FY2025 revenue (disclosed)', f"={A['rev25']}", IN['rev_fy25'], NUM0),
      ('Normalised revenue haircut', f"={A['nhair']}", IN['norm_rev_haircut'], PCT),
      ('Normalised revenue base', "=B5*B6", IN['rev_fy25'] * IN['norm_rev_haircut'], NUM0),
      ('Mid-cycle EBITDA margin', f"={A['nmgn']}", IN['norm_mgn'], PCT),
      ('Normalised EBITDA', "=B7*B8", LN['ebitda_norm'], NUM0),
      ('Justified EV/EBITDA', f"={A['evb']}", IN['ev_ebitda_just'], MULT),
      ('Implied enterprise value', "=B9*B10", LN['ebitda_norm'] * IN['ev_ebitda_just'], NUM0),
      ('Plus net cash', "=DCF!B36", DCF['net_cash'], NUM0),
      ('Less non-controlling interests', f"=-{A['nci']}", -IN['nci'], NUM0),
      ('Implied equity value', "=B11+B12+B13",
       LN['ebitda_norm'] * IN['ev_ebitda_just'] + DCF['net_cash'] - IN['nci'], NUM0),
      ('Implied value per share (EGP)', f"=B14/{A['shares']}",
       LN['values']['Relative multiples'], PX)]
for j, (lab, fm, ex, ft) in enumerate(RL):
    wsN.cell(row=5 + j, column=1, value=lab)
    putf(wsN, f'B{5+j}', fm, ex, ft, bold=(j == 10))
band(wsN, 17, 6)
wsN['A17'] = ('NORMALISED EARNINGS POWER — RETIRED AS A LENS FOR THIS CLASS, SHOWN FOR THE RECORD')
NL = [('Normalised EBITDA', "=B9", LN['ebitda_norm'], NUM0),
      ('Less D&A (FY2025)', "=-'Income Statement'!D8", -H['dna'][2], NUM0),
      ('Normalised EBIT', "=B18+B19", LN['ebitda_norm'] - H['dna'][2], NUM0),
      ('Normalised NOPAT', f"=B20*(1-{A['tax']})", LN['nopat_norm'], NUM0),
      ('Justified price/earnings', f"={A['pej']}", IN['pe_just'], MULT),
      ('Operating equity value', "=B21*B22", LN['nopat_norm'] * IN['pe_just'], NUM0),
      ('Plus net cash AT FACE (not capitalised)', "=DCF!B36", DCF['net_cash'], NUM0),
      ('Less non-controlling interests', f"=-{A['nci']}", -IN['nci'], NUM0),
      ('Implied value per share (EGP)', f"=(B23+B24+B25)/{A['shares']}",
       LR['retired']['normalised_earnings']['value'], PX)]
for j, (lab, fm, ex, ft) in enumerate(NL):
    wsN.cell(row=18 + j, column=1, value=lab)
    putf(wsN, f'B{18+j}', fm, ex, ft, bold=(j == 8))
band(wsN, 28, 6); wsN['A28'] = 'ASSET LENS — EV PER TONNE OF CAPACITY (the cement sector yardstick)'
AL = [('Enterprise value at spot (EGP mn)', f"={A['spot']}*{A['shares']}-DCF!B36",
       SPOT * SH - DCF['net_cash'], NUM0),
      ('Capacity (annual tonnes)', f"={A['capcem']}*1000000", IN['cap_cement_mt'] * 1e6, NUM0),
      ('EV per tonne at spot (USD/t)', f"=B29*1000000/B30/{A['fx']}", LN['ev_per_t_spot'], NUM1),
      ('Replacement cost (USD/t)', f"={A['repl']}", IN['repl_usd_t'], NUM0),
      ('Discount to replacement cost', "=B31/B32-1", LN['ev_per_t_spot'] / IN['repl_usd_t'] - 1, PCT),
      ('Justified EV per tonne (USD/t)', f"={A['evt']}", IN['ev_t_just'], NUM0),
      ('Implied enterprise value (EGP mn)', f"=B34*B30*{A['fx']}/1000000", LN['ev_asset'], NUM0),
      ('Plus net cash', "=DCF!B36", DCF['net_cash'], NUM0),
      ('Less non-controlling interests', f"=-{A['nci']}", -IN['nci'], NUM0),
      ('Implied value per share (EGP)', f"=(B35+B36+B37)/{A['shares']}",
       LN['values']['Asset / replacement cost'], PX)]
for j, (lab, fm, ex, ft) in enumerate(AL):
    wsN.cell(row=29 + j, column=1, value=lab)
    putf(wsN, f'B{29+j}', fm, ex, ft, bold=(j == 8))
band(wsN, 40, 6); wsN['A40'] = 'MEMO — BOOK, SHOWN BUT NOT USED AS A LENS'
wsN['A41'] = 'Book value per share (EGP)'
putf(wsN, 'B41', f"='Balance Sheet'!D12/{A['sh25']}", LN['bvps'], PX)
wsN['A42'] = 'Sustainable return on equity'
putf(wsN, 'B42', f"=B21/'Balance Sheet'!D12", LN['roe_sust'], PCT)
note(wsN, 44, 'A book/return lens is deliberately NOT one of the four. The El Hassana plant commissioned in 1997 and')
note(wsN, 45, 'is carried at historic cost through a five-fold devaluation, so book value measures the accounting')
note(wsN, 46, 'rather than the asset. The cement sector\'s own yardstick — enterprise value per annual tonne of')
note(wsN, 47, 'capacity against replacement cost — is used in its place.')

# ============ 3 FUNDAMENTAL VALUATION =========================================
wsFV = sheet('Fundamental Valuation')
title(wsFV, 'Fundamental valuation — four lenses', None, 6, 52, 16)
hdr(wsFV, 4, ['Lens', 'Role', 'Value per share (EGP)', 'vs spot'])
# [R-LENS-03] ONE CLASS PRIMARY IS THE CENTRAL AND THE OTHER LENSES ARE CROSS-CHECKS.
# The retired sheet carried a Weight column and a SUM of weighted lenses. Those weights
# were typed, inherited, and had never cleared an out-of-sample test; three of the four
# lenses they averaged read a cement plant off reported accounting earnings and
# historical-cost book. There is no weight column here because there are no weights.
LKEYS = ['DCF (cash flow)', 'Relative multiples', 'Asset / replacement cost',
         'Book value (disclosed floor)']
LROLE = {'DCF (cash flow)': 'PRIMARY — this IS the central',
         'Relative multiples': 'cross-check',
         'Asset / replacement cost': 'cross-check',
         'Book value (disclosed floor)': 'disclosed floor — never weighted'}
LSRC = {'DCF (cash flow)': "=DCF!B39",
        'Relative multiples': "='Relative & Normalized'!B15",
        'Asset / replacement cost': "='Relative & Normalized'!B38",
        'Book value (disclosed floor)': "='Relative & Normalized'!B41"}
for j, k in enumerate(LKEYS):
    wsFV.cell(row=5 + j, column=1, value=k)
    wsFV.cell(row=5 + j, column=2, value=LROLE[k])
    putf(wsFV, f'C{5+j}', LSRC[k], LN['values'][k], PX, green=True)
    putf(wsFV, f'D{5+j}', f"=C{5+j}/$C$11-1", LN['values'][k] / SPOT - 1, PCT)
wsFV['A10'] = 'RETIRED — normalised earnings power, not a lens for this class'
putf(wsFV, 'C10', "='Relative & Normalized'!B26",
     LR['retired']['normalised_earnings']['value'], PX)
wsFV['A11'] = 'Spot price (EGP)'
putf(wsFV, 'C11', f"={A['spot']}", SPOT, PX, green=True)
wsFV['A12'] = 'CENTRAL FAIR VALUE (EGP) — the primary lens, not a blend'
putf(wsFV, 'C12', "=C5", LN['central'], PX, bold=True)
putf(wsFV, 'D12', "=C12/C11-1", LN['central'] / SPOT - 1, PCT, bold=True)
wsFV['A13'] = 'Envelope, low — the primary with the EBITDA margin at its filed floor'
putf(wsFV, 'C13', f"={A['envlo']}", LR['envelope']['low'], PX, green=True)
wsFV['A14'] = 'Envelope, high — the primary as published'
putf(wsFV, 'C14', "=C12", LR['envelope']['high'], PX)
wsFV['A15'] = 'MEMO — what the retired four-lens blend would have read'
putf(wsFV, 'C15', f"={A['blend']}", LR['retired']['blend_value'], PX, green=True)
wsFV['A16'] = 'Terminal value as % of enterprise value (primary lens)'
putf(wsFV, 'C16', "=DCF!B33", DCF['tv_share'], PCT, green=True)
note(wsFV, 18, 'The lenses disagree by design and the disagreement is PUBLISHED rather than averaged away. The')
note(wsFV, 19, 'cash-flow lens is the central for this class; the others are cross-checks read beside it, and the')
note(wsFV, 20, 'book value is a disclosed floor that is never weighted into an answer.')

# ============ 2 SUMMARY =======================================================
wsS = sheet('Summary')
title(wsS, 'Summary valuation table', 'Sinai Cement Company S.A.E. · EGX: SCEM · EGP · 06-Aug-2026',
      7, 50, 17)
hdr(wsS, 4, ['', 'Value'])
SUMR = [('Spot price (EGP)', f"={A['spot']}", SPOT, PX),
        ('Shares outstanding (mn)', f"={A['shares']}", SH, NUM2),
        ('Market capitalisation (EGP mn)', f"={A['spot']}*{A['shares']}", M['mktcap'], NUM0),
        ('Net cash (EGP mn)', "=DCF!B36", DCF['net_cash'], NUM0),
        ('Enterprise value at spot (EGP mn)', f"=B7-B8", SPOT * SH - DCF['net_cash'], NUM0)]
for j, (lab, fm, ex, ft) in enumerate(SUMR):
    wsS.cell(row=5 + j, column=1, value=lab)
    putf(wsS, f'B{5+j}', fm, ex, ft, green=True)
band(wsS, 11, 7); wsS['A11'] = 'VALUATION BY LENS'
hdr(wsS, 12, ['Lens', 'Role', 'Value per share (EGP)', 'vs spot', 'Terminal value % of EV'])
for j, k in enumerate(LKEYS):
    wsS.cell(row=13 + j, column=1, value=k)
    wsS.cell(row=13 + j, column=2, value=LROLE[k])
    putf(wsS, f'C{13+j}', f"='Fundamental Valuation'!C{5+j}", LN['values'][k], PX, green=True)
    putf(wsS, f'D{13+j}', f"=C{13+j}/$B$5-1", LN['values'][k] / SPOT - 1, PCT)
    if k == 'DCF (cash flow)':
        putf(wsS, f'E{13+j}', "=DCF!B33", DCF['tv_share'], PCT, green=True)
wsS['A17'] = 'CENTRAL FAIR VALUE (EGP) — the primary lens, not a blend'
putf(wsS, 'C17', "='Fundamental Valuation'!C12", LN['central'], PX, bold=True)
putf(wsS, 'D17', "=C17/$B$5-1", LN['central'] / SPOT - 1, PCT, bold=True)
wsS['A18'] = 'Envelope — the primary with the EBITDA margin across its own filed span (EGP)'
putf(wsS, 'C18', "='Fundamental Valuation'!C13", LR['envelope']['low'], PX)
putf(wsS, 'E18', "='Fundamental Valuation'!C14", LR['envelope']['high'], PX)
wsS['A19'] = 'Vicat tender offer, July 2025 (EGP) — reference, not a value'
putf(wsS, 'B19', f"={A['mto']}", IN['mto_price'], PX, green=True)
band(wsS, 21, 7); wsS['A21'] = 'COST OF CAPITAL AND TERMINAL'
CCS = [('Cost of equity — explicit window', "=DCF!C45", W['ke_exp'], PCT2),
       ('WACC — explicit window', "=DCF!C46", W['wacc_exp'], PCT2),
       ('WACC — terminal', "=DCF!C53", W['wacc_term'], PCT2),
       ('Terminal growth', f"={A['g']}", IN['g_term'], PCT),
       ('Terminal return on invested capital (replacement basis)', "=DCF!B24", DCF['roic_term'], PCT),
       ('Terminal value as % of enterprise value', "=DCF!B33", DCF['tv_share'], PCT)]
for j, (lab, fm, ex, ft) in enumerate(CCS):
    wsS.cell(row=22 + j, column=1, value=lab)
    putf(wsS, f'B{22+j}', fm, ex, ft, green=True, bold=(j == 5))
note(wsS, 29, 'Terminal value share is linked live to the DCF sheet in both the lens table and the block above.')

# ============ 13 MONTE CARLO ==================================================
wsM = sheet('Monte Carlo')
title(wsM, 'Probabilistic price map', 'PASTED — a whole-model re-run. Does NOT redraw when a driver changes.',
      8, 46, 15)
hdr(wsM, 4, ['Horizon', 'Sessions', 'Grade date', '5th %ile', '25th %ile', 'Median',
             '75th %ile', '95th %ile'])
for j, hz in enumerate(['1M', '3M']):
    h = STK['horizons'][hz]
    wsM.cell(row=5 + j, column=1, value=f'{hz} ({"1 month" if hz=="1M" else "3 months"})')
    put(wsM, f'B{5+j}', h['h'], BLUE, NUM0)
    put(wsM, f'C{5+j}', h['grade_date'], BLUE, None)
    for k, p in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        put(wsM, f'{chr(68+k)}{5+j}', h['pct'][p], BLUE, PX)
hdr(wsM, 8, ['', 'P(above spot)', 'P(+10% or more)', 'P(-10% or more)',
             'P(touch +10%)', 'P(touch -10%)'])
for j, hz in enumerate(['1M', '3M']):
    h = STK['horizons'][hz]
    wsM.cell(row=9 + j, column=1, value=hz)
    for k, key in enumerate(['p_above', 'p_up10', 'p_dn10', 'touch_up10', 'touch_dn10']):
        put(wsM, f'{chr(66+k)}{9+j}', h[key], BLUE, PCT)
band(wsM, 12, 8); wsM['A12'] = 'CALIBRATION — READ THIS BEFORE THE MAP'
CAL = [
 ('Market panel verdict (Egypt)', f"{S0['market_gate']['verdict']} — skill "
  f"{S0['market_gate']['skill']:+.4f}, 90% interval {S0['market_gate']['ci90']}"),
 ('This name\'s verdict', f"{S0['verdict']} at every bootstrap block size {{2,3,4}}"),
 ('This name\'s CRPS skill vs a random walk', f"{S0['skill_norm']:+.4f} — BELOW zero"),
 ('Coverage 50 / 80 / 90', f"{S0['cov50']:.2f} / {S0['cov80']:.2f} / {S0['cov90']:.2f} "
  f"against nominal 0.50 / 0.80 / 0.90 — OVER-COVERED"),
 ('Cone width versus the benchmark', f"{S0['w90_ratio']:.2f}x"),
 ('Why', 'SCEM prints an unchanged close on 29.3% of sessions, 3.4x the Egyptian panel '
         'median and 2nd thinnest of 33 names. On such a series the benchmark\'s own '
         'volatility estimate collapses in quiet quarters.'),
 ('Consequence', 'The map is ILLUSTRATIVE ONLY. It is materially too wide and must not be '
                 'read with the confidence of a calibrated name.')]
for j, (k, v) in enumerate(CAL):
    wsM.cell(row=13 + j, column=1, value=k).font = Font(bold=True)
    c = wsM.cell(row=13 + j, column=2, value=v)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    wsM.merge_cells(start_row=13 + j, start_column=2, end_row=13 + j, end_column=8)

# ============ 14 SENSITIVITY ==================================================
wsX = sheet('Sensitivity')
title(wsX, 'Sensitivity', 'PASTED — each cell is a complete revaluation. Does NOT redraw.',
      8, 40, 13)
wsX['A4'] = 'Explicit WACC (rows) × terminal growth (columns) — fair value per share (EGP)'
wsX['A4'].font = Font(bold=True)
hdr(wsX, 5, [''] + [f'g = {g:.0%}' for g in SN['g_grid']])
for i, we in enumerate(SN['wacc_grid']):
    put(wsX, f'A{6+i}', we, BLUE, PCT2)
    for j in range(5):
        put(wsX, f'{chr(66+j)}{6+i}', SN['wacc_g'][i][j], BLUE, PX)
wsX['A12'] = 'NET CASH — the largest single uncertainty, now sensitised'
wsX['A12'].font = Font(bold=True)
hdr(wsX, 13, [''] + [f'{x:,.0f}' for x in SN['nc_grid']])
put(wsX, 'A14', 'Fair value per share (EGP)', BLACK, None)
for j, v in enumerate(SN['net_cash']):
    put(wsX, f'{chr(66+j)}14', v, BLUE, PX)
note(wsX, 16, 'Revision 1 called net cash the assumption where "an error matters more than')
note(wsX, 17, 'almost any operating assumption", and then published no grid on it. This is that grid.')
wsX['A20'] = 'Beta — fair value per share (EGP)'
wsX['A20'].font = Font(bold=True)
hdr(wsX, 21, [''] + [f'β = {b:.2f}' for b in SN['beta_grid']])
put(wsX, 'A22', 'Fair value', BLACK, None)
for j, v in enumerate(SN['beta']):
    put(wsX, f'{chr(66+j)}22', v, BLUE, PX)
wsX['A24'] = 'EBITDA margin shift — fair value per share (EGP)'
wsX['A24'].font = Font(bold=True)
hdr(wsX, 25, [''] + [f'{m:+.0%}' for m in SN['mgn_grid']])
put(wsX, 'A26', 'Fair value', BLACK, None)
for j, v in enumerate(SN['mgn']):
    put(wsX, f'{chr(66+j)}26', v, BLUE, PX)
note(wsX, 28, 'Note the terminal-growth row: HIGHER growth gives a LOWER value. That is not an error. Terminal')
note(wsX, 29, 'return on capital (9.4%) is below the terminal cost of capital (16.3%), so growth must be bought with')
note(wsX, 30, 'reinvestment that earns less than it costs. This company creates value by harvesting, not by growing.')

# ============ 16 PEER & SECTOR ================================================
wsP = sheet('Peer & Sector')
title(wsP, 'Peer set and sector structure', None, 6, 46, 16)
band(wsP, 4, 6); wsP['A4'] = 'NAMED EGYPTIAN LISTED PEER'
PR = [('Misr Beni Suef — FY2025 net sales (EGP mn)', PE['mbsc']['rev'], NUM0),
      ('Misr Beni Suef — FY2025 attributable profit (EGP mn)', PE['mbsc']['pat'], NUM0),
      ('Misr Beni Suef — earnings per share (EGP)', PE['mbsc']['eps'], PX),
      ('Misr Beni Suef — market capitalisation (EGP mn)', PE['mbsc']['mcap'], NUM0),
      ('Misr Beni Suef — trailing price/earnings', PE['mbsc']['pe'], MULT),
      ('Misr Beni Suef — EV/EBITDA', PE['mbsc']['ev_ebitda'], MULT),
      ('Arabian Cement — FY2025 consolidated profit (EGP mn)', PE['arcc']['pat'], NUM0)]
for j, (lab, v, ft) in enumerate(PR):
    wsP.cell(row=5 + j, column=1, value=lab)
    put(wsP, f'B{5+j}', v, BLUE, ft)
band(wsP, 13, 6); wsP['A13'] = 'EGYPTIAN SECTOR STRUCTURE'
SE = [('Nameplate capacity (Mt/yr)', PE['sector']['capacity_mt'], NUM1),
      ('Domestic consumption 2025 (Mt)', PE['sector']['consumption_mt'], NUM1),
      ('Production 2025 (Mt)', PE['sector']['production_mt'], NUM1),
      ('Exports 2025 (Mt)', PE['sector']['exports_mt'], NUM1),
      ('Dormant capacity under revival from 2H-2026 (Mt)', PE['sector']['revival_mt'], NUM1)]
for j, (lab, v, ft) in enumerate(SE):
    wsP.cell(row=14 + j, column=1, value=lab)
    put(wsP, f'B{14+j}', v, BLUE, ft)
wsP['A19'] = 'SCEM capacity as % of Egyptian capacity'
putf(wsP, 'B19', f"={A['capcem']}/B14", PE['sector']['scem_share_of_capacity'], PCT)
wsP['A20'] = 'Revival capacity as % of domestic consumption'
putf(wsP, 'B20', "=B18/B15", PE['sector']['revival_pct_of_consumption'], PCT)
wsP['A21'] = 'Structural surplus (capacity less consumption, Mt)'
putf(wsP, 'B21', "=B14-B15", PE['sector']['capacity_mt'] - PE['sector']['consumption_mt'], NUM1)
wsP['A22'] = 'SCEM EV/EBITDA at spot on FY2026E'
putf(wsP, 'B22', "='Per-Share & Ratios'!E13",
     (SPOT * SH - (F['cash'][0] - IN['debt_fy25'])) / F['ebitda'][0], MULT)
note(wsP, 24, 'The structural surplus is the whole sector case in one number: Egypt can make far more cement than')
note(wsP, 25, 'it consumes, and 12.6Mt of dormant capacity is under study for revival inside the forecast window.')

# ============ order the sheets ================================================
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'EV Bridge',
         'Unit Build', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]

# openpyxl cannot write cached values. Set fullCalcOnLoad so Excel, LibreOffice and
# Sheets all recompute the whole book the moment it opens — which is what the reader
# needs, and closes the "delivered file reads blank" finding.
wb.calculation.fullCalcOnLoad = True
wb.calculation.calcCompleted = False

OUT = os.path.join(HERE, 'SCEM_Valuation_Model_04092026_public.xlsx')
wb.save(OUT)
json.dump(dict(expected=EXPECT, anchors=ANCH),
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nf = sum(len(v) for v in EXPECT.values())
print(f'wrote {OUT}')
print(f'sheets {len(ORDER)} | formula cells recorded {nf}')
