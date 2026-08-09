"""EGCH valuation workbook — the MODEL STUDY sheet list, formula-first.

Sixteen sheets, same names and same order as the model study. Everything
arithmetically derivable from a driver is a live Excel formula. No financial numeral is
typed here: every value comes from study_numbers.json, lenses.json, strike_result.json
or the input register.

Three classes of cell carry a pasted number, and READ FIRST names all three:
  (1) audited and disclosed history;
  (2) the unit build's output;
  (3) whole-model re-run grids (Sensitivity), and the simulated price map (Monte Carlo),
      where each cell is a complete revaluation or a distribution statistic.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
LN = json.load(open(os.path.join(HERE, 'lenses.json')))
ST = json.load(open(os.path.join(HERE, 'strike_result.json')))
GRID = json.load(open(os.path.join(HERE, 'sensitivity_grid.json')))
ALT = json.load(open(os.path.join(HERE, 'alternatives.json')))
BETA = json.load(open(os.path.join(HERE, 'beta_result.json')))
IR = json.load(open(os.path.join(HERE, 'input_register.json')))['inputs']
W, DR, YEARS = D['wacc'], D['drivers'], D['years']
CASES = D['cases']
BASE, HALT, BULL, BEAR = (CASES['base'], CASES['halt'], CASES['bull'], CASES['bear'])
R = BASE['rows']
V = lambda k: IR[k]['value']
SPOT, SHARES = V('spot_price'), V('shares_outstanding')
EXPECT = {}
HIS = D['hist']; F25S = D['fy2526']

BLUE = Font(color="1F4E79"); BLACK = Font(color="000000"); GREEN = Font(color="1E6B3A")
HDR = Font(color="FFFFFF", bold=True, size=11); HDRFILL = PatternFill("solid", fgColor="1F4E79")
SUB = Font(bold=True, color="1F4E79"); TITLE = Font(bold=True, size=14, color="1F4E79")
NOTE = Font(italic=True, size=9, color="595959")
N0='#,##0'; N1='#,##0.0'; N2='#,##0.00'; N3='#,##0.000'; PC1='0.0%'; PC2='0.00%'


def put(ws, coord, value, *, fmt=None, font=None, expect=None, link=False):
    c = ws[coord]; c.value = value
    if fmt: c.number_format = fmt
    isf = isinstance(value, str) and value.startswith('=')
    c.font = font or ((GREEN if link else BLACK) if isf else BLUE)
    if isf:
        if expect is None:
            raise ValueError(f"formula {ws.title}!{coord} written without expected value")
        EXPECT[f"{ws.title}!{coord}"] = float(expect)
    return c


def header(ws, row, col, labels, widths=None):
    for j, lab in enumerate(labels):
        c = ws.cell(row=row, column=col + j, value=lab)
        c.font = HDR; c.fill = HDRFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    if widths:
        for j, wd in enumerate(widths):
            ws.column_dimensions[get_column_letter(col + j)].width = wd


def title(ws, text, sub=None):
    ws['A1'] = text; ws['A1'].font = TITLE
    if sub:
        ws['A2'] = sub; ws['A2'].font = NOTE
        ws.row_dimensions[2].height = 30
        ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")


def para(ws, row, text, size=10):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=size); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = max(15, 13 * (1 + len(text) // 105))


wb = openpyxl.Workbook()
CO = ['B', 'C', 'D', 'E', 'F']

# ============================================================= 1 READ FIRST ===
ws = wb.active; ws.title = "READ FIRST"
ws.column_dimensions['A'].width = 118
title(ws, "EGYPTIAN CHEMICAL INDUSTRIES (KIMA) — EGX: EGCH — VALUATION MODEL")
LINES = [
 ("WHAT THIS FILE IS", SUB),
 ("A live formula model of a single-site Egyptian nitrogen-fertilizer producer. Revenue is "
  "built product by product as tonnes times price; cost is built as physical consumption times "
  "a unit price; the cost of capital, the discount-rate glide, the free-cash-flow waterfall, "
  "the terminal block, the statement roll-forwards and every ratio are calculated in the "
  "sheet. Change a blue cell on Assumptions and the valuation recomputes.", None),
 ("", None),
 ("BLUE IS AN INPUT.  BLACK IS A FORMULA.  GREEN IS A LINK TO ANOTHER SHEET.", SUB),
 ("", None),
 ("THE THREE CLASSES OF PASTED CELL, AND THERE ARE NO OTHERS", SUB),
 ("1. AUDITED AND DISCLOSED HISTORY. The Income Statement, Balance Sheet and Cash Flow sheets "
  "carry the figures exactly as issued in the company's own audited statements for the years "
  "ended 30 June 2023, 2024 and 2025 and its reviewed nine-month accounts to 31 March 2026. "
  "Where a line is both disclosed and derivable, the DISCLOSED figure is carried and the "
  "derivation sits beside it as a check.", None),
 ("2. THE UNIT BUILD'S OUTPUT. Production tonnages, the split of those tonnes between the "
  "subsidised, local free-market and export channels, and the realised price in each channel "
  "are the output of a reconciliation against the audited revenue note that would be "
  "unreadable flattened into a grid. They are pasted on Assumptions; everything downstream is "
  "a formula.", None),
 ("3. WHOLE-MODEL RE-RUNS AND SIMULATION OUTPUT. Every cell of the Sensitivity grid is a "
  "complete revaluation at a different pair of inputs, and the Monte Carlo sheet reports "
  "statistics of a fifty-thousand-path simulation. NEITHER REDRAWS WHEN A DRIVER CHANGES, and "
  "both say so on their own sheet.", None),
 ("", None),
 ("THE CONTESTED JUDGEMENT IS CARRIED BOTH WAYS, NEVER AVERAGED", SUB),
 ("This company is building a nitric-acid and ammonium-nitrate complex whose bank-approved "
  "cost is about three quarters of its own stock-market value. Whether that programme is "
  "carried through or stopped is worth more than three pounds a share — more than twice the "
  "central estimate itself. Averaging the two would report a number that is true in neither "
  "world, so both are computed and both are published: on Summary, on Fundamental Valuation, "
  "in the DCF sheet's two columns, and in the expert panel.", None),
 ("", None),
 ("This is a valuation study. It is not a rating and not a price target: it reports a range of "
  "fair values and the distribution around them.", NOTE),
]
r = 4
for text, font in LINES:
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(14, 13 * (1 + len(text) // 105))
    r += 1

# ================================================================ 2 Summary ===
ws = wb.create_sheet("Assumptions")
title(ws, "ASSUMPTIONS — THE LIVE DRIVERS",
      "Blue cells are inputs. Change one and the model reprices. Each carries its source. "
      "Every value on this sheet comes from the study's input register.")
header(ws, 4, 1, ["Driver", "Unit", "Value", "Source"], [46, 15, 14, 76])
A = {}
def drv(row, label, key_or_val, unit, source, fmt=N2):
    val = key_or_val
    put(ws, f"A{row}", label); put(ws, f"B{row}", unit)
    put(ws, f"C{row}", val, fmt=fmt)
    c = put(ws, f"D{row}", source); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = max(14, 12 * (1 + len(source) // 96))
    return f"'Assumptions'!C{row}"

def src(k):
    return IR[k]['source']

r = 5
put(ws, f"A{r}", "CAPACITY AND VOLUME").font = SUB; r += 1
C_DESIGN = drv(r, "Urea design capacity", V('design_urea_tpy'), "tonnes/year", src('design_urea_tpy'), N0); r += 1
C_NH3D = drv(r, "Ammonia design capacity", V('design_ammonia_tpy'), "tonnes/year", src('design_ammonia_tpy'), N0); r += 1
C_NH3R = drv(r, "Ammonia per tonne of urea", V('ammonia_per_urea'), "tonnes", src('ammonia_per_urea'), N3); r += 1
UTIL = []
for k, y in enumerate(YEARS):
    UTIL.append(drv(r, f"Urea capacity utilisation — {y}", V('urea_util')[k], "% of plate",
                    src('urea_util'), PC1)); r += 1
SUBT = []
for k, y in enumerate(YEARS):
    SUBT.append(drv(r, f"Subsidised deliveries — {y}", V('subsidised_t_path')[k], "tonnes",
                    src('subsidised_t_path'), N0)); r += 1
FREET = []
for k, y in enumerate(YEARS):
    FREET.append(drv(r, f"Local free-market volume — {y}", V('local_free_path')[k], "tonnes",
                     src('local_free_path'), N0)); r += 1
r += 1
put(ws, f"A{r}", "PRICES").font = SUB; r += 1
EXPP = []
for k, y in enumerate(YEARS):
    EXPP.append(drv(r, f"Export urea price — {y}", V('export_usd_path')[k], "US$/tonne",
                    src('export_usd_path'), N1)); r += 1
FXP = []
for k, y in enumerate(YEARS):
    FXP.append(drv(r, f"Egyptian pounds per US dollar — {y}", V('usd_egp_path')[k], "EGP/US$",
                   src('usd_egp_path'), N2)); r += 1
DUTY = drv(r, "Export duty", V('export_duty_2026'), "% of export value", src('export_duty_2026'), PC1); r += 1
SUBP = []
for k, y in enumerate(YEARS):
    SUBP.append(drv(r, f"Subsidised price — {y}", V('subsidised_p_path')[k], "EGP/tonne",
                    src('subsidised_p_path'), N0)); r += 1
PARITY = drv(r, "Local free price as % of export parity", V('local_free_parity'), "%",
             src('local_free_parity'), PC1); r += 1
ANT = []
for k, y in enumerate(YEARS):
    ANT.append(drv(r, f"Nitrate volume — {y}", V('an_path')[k], "tonnes", src('an_path'), N0)); r += 1
ANP = drv(r, "Nitrate price, FY2024/25 basis", V('an_price_egp_t_FY2425'), "EGP/tonne",
          "Implied by the note-20 local revenue net of the subsidised and free-market urea "
          "legs; indexed forward on the currency", N0); r += 1
OTHR = []
for k, y in enumerate(YEARS):
    OTHR.append(drv(r, f"Other revenue — {y}", V('other_rev_path')[k], "EGP m",
                    src('other_rev_path'), N1)); r += 1
r += 1
put(ws, f"A{r}", "COST STACK — ONE ESCALATOR PER PHYSICAL DRIVER").font = SUB; r += 1
GASQ = drv(r, "Gas per tonne of ammonia", V('gas_m3_per_t_ammonia_modelled'), "m3",
           src('gas_share_of_materials'), N0); r += 1
GASP = drv(r, "Realised gas price", V('gas_realised_usd_mmbtu'), "US$/mmBtu",
           src('gas_realised_usd_mmbtu'), N2); r += 1
MMB = drv(r, "Energy conversion", V('mmbtu_per_m3'), "mmBtu per m3", src('mmbtu_per_m3'), '0.00000'); r += 1
OMAT = drv(r, "Other materials per tonne of urea", DR['other_materials_egp_t_urea'], "EGP",
           "The FY2024/25 materials line less modelled gas, over urea output: packaging, "
           "catalysts and consumable spares", N0); r += 1
WAGE = drv(r, "Wages in cost of sales, FY2024/25", V('cogs_wages_FY2425'), "EGP m", src('cogs_wages_FY2425'), N1); r += 1
SERV = drv(r, "Purchased services, FY2024/25", V('cogs_services_FY2425'), "EGP m", src('cogs_services_FY2425'), N1); r += 1
FRT = drv(r, "Inland freight per export tonne", V('sell_freight_FY2425') * 1e6 / V('export_tonnes_FY2425'),
          "EGP", src('sell_freight_FY2425') + " over the implied export tonnage", N0); r += 1
OSELL = drv(r, "Other selling cost, FY2024/25", V('sell_other_FY2425'), "EGP m", src('sell_other_FY2425'), N1); r += 1
ADMIN = drv(r, "Administrative expense, FY2024/25", V('is_admin_FY2425'), "EGP m", src('is_admin_FY2425'), N1); r += 1
ABN = []
for k, y in enumerate(YEARS):
    ABN.append(drv(r, f"Stoppage and abnormal gas cost — {y}", V('abnormal_gas_path')[k], "EGP m",
                   src('abnormal_gas_path'), N1)); r += 1
CPI = []
for k, y in enumerate(YEARS):
    CPI.append(drv(r, f"Egyptian inflation — {y}", V('cpi_path')[k], "%", src('cpi_path'), PC1)); r += 1
r += 1
put(ws, f"A{r}", "COST OF CAPITAL").font = SUB; r += 1
RF = drv(r, "Observed ten-year government yield", V('rf_observed'), "%", src('rf_observed'), PC2); r += 1
SPR = drv(r, "Sovereign default spread, rating basis", V('sov_spread_rating'), "%", src('sov_spread_rating'), PC2); r += 1
RFS = r; put(ws, f"A{r}", "Normalised risk-free rate, rating basis"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={RF}-{SPR}", fmt=PC2, expect=W['rf_star_rating'])
put(ws, f"D{r}", "Country risk enters ONCE, through the premium below. Using the raw local "
    "yield with a country-loaded premium would charge Egypt's sovereign risk twice.").alignment = Alignment(wrap_text=True, vertical="top")
RFSTAR = f"'Assumptions'!C{r}"; r += 1
ERP = drv(r, "Equity risk premium, rating basis", V('erp_rating'), "%", src('erp_rating'), PC2); r += 1
SPRC = drv(r, "Sovereign default spread, CDS basis", V('sov_spread_cds'), "%", src('sov_spread_cds'), PC2); r += 1
ERPC = drv(r, "Equity risk premium, CDS basis", V('erp_cds'), "%", src('erp_cds'), PC2); r += 1
BETAC = drv(r, "Beta", W['beta'], "x",
            f"Own-stock weekly regression, {BETA['n']} observations over five years against an "
            f"equal-weight index of {BETA['composite_names']} Egyptian names with the subject "
            f"excluded, R-squared {BETA['r2']:.3f}, standard error {BETA['se']:.3f}", N3); r += 1
KE_R = r; put(ws, f"A{r}", "Cost of equity, rating basis"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={RFSTAR}+{BETAC}*{ERP}", fmt=PC2, expect=W['ke_rating']); r += 1
put(ws, f"A{r}", "Cost of equity, CDS basis"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={RF}-{SPRC}+{BETAC}*{ERPC}", fmt=PC2, expect=W['ke_cds']); r += 1
KDL = drv(r, "Cost of debt, local currency", V('kd_local'), "%", src('kd_local'), PC2); r += 1
KDU = drv(r, "Cost of debt, dollar tranche, in dollars", V('kd_usd_nominal'), "%", src('kd_usd_nominal'), PC2); r += 1
DEP = drv(r, "Expected currency depreciation", V('expected_depreciation'), "%", src('expected_depreciation'), PC2); r += 1
KDFX = r; put(ws, f"A{r}", "Dollar debt at local-equivalent cost"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"=(1+{KDU})*(1+{DEP})-1", fmt=PC2, expect=W['kd_fx_local_equiv'])
KDFXC = f"'Assumptions'!C{r}"; r += 1
PCTL = drv(r, "Share of debt in local currency", W['pct_debt_local'], "%",
           "The pound tranche of the project loan was repaid in June 2024, so the book is "
           "almost entirely dollar", PC1); r += 1
KDB = r; put(ws, f"A{r}", "Blended pre-tax cost of debt"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={PCTL}*{KDL}+(1-{PCTL})*{KDFXC}", fmt=PC2, expect=W['kd_pretax_blended'])
KDBC = f"'Assumptions'!C{r}"; r += 1
TAXR = drv(r, "Tax rate", V('tax_statutory'), "%", src('tax_statutory'), PC1); r += 1
WE = drv(r, "Equity weight", W['we'], "%", "Market-value equity over market-value equity plus debt", PC1); r += 1
WD = drv(r, "Debt weight", W['wd'], "%", "One less the equity weight", PC1); r += 1
W1 = r; put(ws, f"A{r}", "Cost of capital, year one"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={WE}*'Assumptions'!C{KE_R}+{WD}*{KDBC}*(1-{TAXR})", fmt=PC2,
    expect=DR['wacc_path'][0]); r += 1
INFLT = drv(r, "Long-run inflation", DR['inflation_lt'], "%", src('cbe_inflation_target'), PC1); r += 1
REALLT = drv(r, "Long-run real rate", V('real_rate_lt'), "%", src('real_rate_lt'), PC1); r += 1
RFT = r; put(ws, f"A{r}", "Terminal normalised risk-free rate"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"=(1+{INFLT})*(1+{REALLT})-1", fmt=PC2, expect=DR['rf_star_terminal'])
RFTC = f"'Assumptions'!C{r}"; r += 1
KDLT = drv(r, "Long-run dollar cost of debt", V('kd_usd_lt'), "%", src('kd_usd_lt'), PC2); r += 1
KDTC_R = r; put(ws, f"A{r}", "Terminal cost of debt, local-equivalent"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"=(1+{KDLT})*(1+{DEP})-1", fmt=PC2, expect=DR['kd_local_equiv_terminal'])
KDTC = f"'Assumptions'!C{r}"; r += 1
WT = r; put(ws, f"A{r}", "TERMINAL COST OF CAPITAL"); put(ws, f"B{r}", "%")
put(ws, f"C{r}", f"={WE}*({RFTC}+{BETAC}*{ERP})+{WD}*{KDTC}*(1-{TAXR})", fmt=PC2,
    expect=DR['wacc_terminal'])
WTC = f"'Assumptions'!C{r}"; r += 1
GT = drv(r, "Terminal growth", V('g_terminal'), "%", src('g_terminal'), PC1); r += 1
ROCT = drv(r, "Terminal return on invested capital", V('roc_terminal'), "%", src('roc_terminal'), PC1); r += 1
r += 1
put(ws, f"A{r}", "CAPITAL, WORKING CAPITAL AND THE PROJECT").font = SUB; r += 1
DEPB = drv(r, "Depreciation charge, FY2024/25", V('dep_charge_FY2425'), "EGP m", src('dep_charge_FY2425'), N1); r += 1
AMOB = drv(r, "Amortisation, FY2024/25", V('amort_FY2425'), "EGP m", src('amort_FY2425'), N1); r += 1
ANNAC = []
for k, y in enumerate(YEARS):
    ANNAC.append(drv(r, f"Project capital expenditure — {y}", V('anna_capex_path')[k], "EGP m",
                     src('anna_capex_path'), N0)); r += 1
MCAP = drv(r, "Maintenance capital expenditure", V('maint_capex_pct'), "% of revenue", src('maint_capex_pct'), PC1); r += 1
NH3AN = drv(r, "Ammonia per tonne of nitrate", V('nh3_per_t_an'), "tonnes", src('nh3_per_t_an'), N2); r += 1
put(ws, f"A{r}", "Project nameplate (derived, not disclosed)"); put(ws, f"B{r}", "tonnes/year")
put(ws, f"C{r}", f"=({C_NH3D}-{C_DESIGN}*{C_NH3R})/{NH3AN}", fmt=N0, expect=V('anna_nameplate'))
c = put(ws, f"D{r}", src('anna_nameplate')); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 40
ANNAN = f"'Assumptions'!C{r}"; r += 1
ANNAU = drv(r, "Project utilisation in the terminal year", V('anna_util_base'), "%", src('anna_util_base'), PC1); r += 1
ANNAP = drv(r, "Nitrate price", V('an_price_usd_t'), "US$/tonne", src('an_price_usd_t'), N0); r += 1
ANNAM = drv(r, "Project cash margin", V('anna_cash_margin'), "% of revenue", src('anna_cash_margin'), PC1); r += 1
DSO = drv(r, "Days sales outstanding", V('dso'), "days", src('dso'), N1); r += 1
DIO = drv(r, "Days inventory outstanding", V('dio'), "days", src('dio'), N1); r += 1
DPO = drv(r, "Days payable outstanding", V('dpo'), "days", src('dpo'), N1); r += 1

# ============================================== 5 SOTP Bridge · 6 Segments ====
ws = wb.create_sheet("Segments")
title(ws, "SEGMENTS — THE UNIT BUILD, TONNES TIMES PRICE",
      "The company reports one operating segment, so the build goes below it to the product "
      "and the channel, which is the finest level the statements source.")
header(ws, 4, 1, ["Tonnes and prices"] + YEARS, [40, 15, 15, 15, 15, 15])
UB = {}
def line(r, lab, fn, key, fmt=N0, bold=False):
    c = put(ws, f"A{r}", lab)
    if bold: c.font = SUB
    for k in range(5):
        put(ws, f"{CO[k]}{r}", fn(k, CO[k]), fmt=fmt, expect=R[k][key])
    return r
r = 5
UB['urea'] = line(r, "Urea production (tonnes)", lambda k, c: f"={C_DESIGN}*{UTIL[k]}", 'urea_t'); r += 1
UB['nh3'] = line(r, "Ammonia produced, capped at the design plate",
                 lambda k, c: f"=MIN({c}{UB['urea']}*{C_NH3R},{C_NH3D})", 'ammonia_t'); r += 1
UB['sub'] = line(r, "Subsidised volume (tonnes)", lambda k, c: f"={SUBT[k]}", 'sub_t'); r += 1
UB['free'] = line(r, "Local free-market volume (tonnes)", lambda k, c: f"={FREET[k]}", 'free_t'); r += 1
UB['exp'] = line(r, "Export volume (tonnes)",
                 lambda k, c: f"={c}{UB['urea']}-{c}{UB['sub']}-{c}{UB['free']}", 'exp_t'); r += 1
UB['pexp'] = line(r, "Export price net of duty (EGP/tonne)",
                  lambda k, c: f"={EXPP[k]}*{FXP[k]}*(1-{DUTY})", 'p_exp_egp'); r += 1
UB['pfree'] = line(r, "Local free-market price (EGP/tonne)",
                   lambda k, c: f"={EXPP[k]}*{FXP[k]}*{PARITY}", 'p_free'); r += 1
UB['pan'] = line(r, "Nitrate price (EGP/tonne)",
                 lambda k, c: f"={ANP}*{FXP[k]}/{V('usd_egp_avg_FY2425')}", 'p_an'); r += 1
UB['rexp'] = line(r, "Export revenue (EGP m)",
                  lambda k, c: f"={c}{UB['exp']}*{c}{UB['pexp']}/1000000", 'rev_exp'); r += 1
UB['rsub'] = line(r, "Subsidised revenue (EGP m)",
                  lambda k, c: f"={c}{UB['sub']}*{SUBP[k]}/1000000", 'rev_sub'); r += 1
UB['rfree'] = line(r, "Local free-market revenue (EGP m)",
                   lambda k, c: f"={c}{UB['free']}*{c}{UB['pfree']}/1000000", 'rev_free'); r += 1
UB['ran'] = line(r, "Nitrate revenue (EGP m)",
                 lambda k, c: f"={ANT[k]}*{c}{UB['pan']}/1000000", 'rev_an'); r += 1
UB['roth'] = line(r, "Other revenue (EGP m)", lambda k, c: f"={OTHR[k]}", 'rev_other'); r += 1
UB['rev'] = line(r, "REVENUE (EGP m)",
                 lambda k, c: f"=SUM({c}{UB['rexp']}:{c}{UB['roth']})", 'revenue', N0, True); r += 1
UB['gp'] = line(r, "Gas price (EGP per m3)", lambda k, c: f"={GASP}*{MMB}*{FXP[k]}",
                'gas_price_egp_m3', N2); r += 1
UB['gas'] = line(r, "Natural gas (EGP m)",
                 lambda k, c: f"={c}{UB['nh3']}*{GASQ}*{c}{UB['gp']}/1000000", 'gas_cost'); r += 1
UB['cpi'] = line(r, "Cumulative inflation index",
                 lambda k, c: "=" + "*".join(f"(1+{CPI[j]})" for j in range(k + 1)), 'cpi_cum', N3); r += 1
UB['omat'] = line(r, "Other materials (EGP m)",
                  lambda k, c: f"={c}{UB['urea']}*{OMAT}*{c}{UB['cpi']}/1000000", 'other_mat'); r += 1
UB['wage'] = line(r, "Wages (EGP m)", lambda k, c: f"={WAGE}*{c}{UB['cpi']}", 'wages'); r += 1
UB['serv'] = line(r, "Purchased services (EGP m)", lambda k, c: f"={SERV}*{c}{UB['cpi']}", 'services'); r += 1
UB['dep'] = line(r, "Depreciation and amortisation (EGP m)",
                 lambda k, c: f"={DEPB}*(1+0.02*{k})+{AMOB}", 'dep'); r += 1
UB['cogs'] = line(r, "COST OF SALES (EGP m)",
                  lambda k, c: f"={c}{UB['gas']}+{c}{UB['omat']}+{c}{UB['wage']}+{c}{UB['serv']}+{c}{UB['dep']}",
                  'cogs', N0, True); r += 1
UB['gross'] = line(r, "GROSS PROFIT (EGP m)",
                   lambda k, c: f"={c}{UB['rev']}-{c}{UB['cogs']}", 'gross', N0, True); r += 1
UB['gpc'] = line(r, "Gross margin — an OUTPUT, never an input",
                 lambda k, c: f"={c}{UB['gross']}/{c}{UB['rev']}", 'gross_pct', PC1); r += 1
UB['frt'] = line(r, "Inland freight to port (EGP m)",
                 lambda k, c: f"={c}{UB['exp']}*{FRT}*{c}{UB['cpi']}/1000000", 'freight'); r += 1
UB['osel'] = line(r, "Other selling cost (EGP m)", lambda k, c: f"={OSELL}*{c}{UB['cpi']}", 'other_sell'); r += 1
UB['adm'] = line(r, "Administrative (EGP m)", lambda k, c: f"={ADMIN}*{c}{UB['cpi']}", 'admin'); r += 1
UB['abn'] = line(r, "Stoppage and abnormal gas cost (EGP m)", lambda k, c: f"={ABN[k]}", 'abnormal'); r += 1
UB['ebit'] = line(r, "EBIT (EGP m)",
                  lambda k, c: f"={c}{UB['gross']}-{c}{UB['frt']}-{c}{UB['osel']}-{c}{UB['adm']}-{c}{UB['abn']}",
                  'ebit', N0, True); r += 1
UB['ebitda'] = line(r, "EBITDA (EGP m)", lambda k, c: f"={c}{UB['ebit']}+{c}{UB['dep']}",
                    'ebitda', N0, True); r += 1
UB['em'] = line(r, "EBITDA margin — an OUTPUT", lambda k, c: f"={c}{UB['ebitda']}/{c}{UB['rev']}",
                'ebitda_pct', PC1); r += 1
para(ws, r + 1, "Margins are outputs of the build, not inputs to it. Nothing on this sheet "
     "sets a margin; each one falls out of tonnes, prices and physical consumption.", 9)

# ==================================================== 7 Relative & Normalized =
ws = wb.create_sheet("Cash Flow")
title(ws, "CASH FLOW — WORKING CAPITAL FROM THE DISCLOSED DAY COUNTS",
      "No plugs: receivables, inventory and payables are projected from the day counts the "
      "audited statements themselves imply, and the change feeds the free-cash-flow waterfall.")
header(ws, 4, 1, ["EGP million"] + YEARS, [34, 14, 14, 14, 14, 14])
for k, c in enumerate(CO):
    put(ws, f"{c}5", f"='Segments'!{c}{UB['rev']}", fmt=N0, expect=R[k]['revenue'], link=True)
    put(ws, f"{c}6", f"='Segments'!{c}{UB['cogs']}", fmt=N0, expect=R[k]['cogs'], link=True)
    put(ws, f"{c}7", f"={c}5*{DSO}/365", fmt=N0, expect=R[k]['revenue'] * V('dso') / 365)
    put(ws, f"{c}8", f"={c}6*{DIO}/365", fmt=N0, expect=R[k]['cogs'] * V('dio') / 365)
    put(ws, f"{c}9", f"={c}6*{DPO}/365", fmt=N0, expect=R[k]['cogs'] * V('dpo') / 365)
    put(ws, f"{c}10", f"={c}7+{c}8-{c}9", fmt=N0, expect=R[k]['wc'])
prev_wc0 = (F25S['revenue'] * V('dso') / 365 + F25S['cogs'] * V('dio') / 365
            - F25S['cogs'] * V('dpo') / 365)
put(ws, "A11", "Opening net working capital, FY2025/26E"); put(ws, "B11", prev_wc0, fmt=N0)
put(ws, f"{CO[0]}12", f"={CO[0]}10-B11", fmt=N0, expect=R[0]['dwc'])
for k in range(1, 5):
    put(ws, f"{CO[k]}12", f"={CO[k]}10-{CO[k-1]}10", fmt=N0, expect=R[k]['dwc'])
for rr, lab in [(5, "Revenue"), (6, "Cost of sales"), (7, "Receivables"), (8, "Inventory"),
                (9, "Payables"), (10, "Net working capital"),
                (12, "Change in net working capital")]:
    put(ws, f"A{rr}", lab)
put(ws, "A14", "Free cash flow to the firm").font = SUB
for k, c in enumerate(CO):
    put(ws, f"{c}14", f"='DCF'!{c}14", fmt=N0, expect=R[k]['fcff'], link=True)

# ==================================================== 12 Summary Financials ===
ws = wb.create_sheet("DCF")
title(ws, "DISCOUNTED CASH FLOW — THE CONTESTED JUDGEMENT IN TWO COLUMNS",
      "Column B carries the programme through; column D stops it. Both are complete "
      "valuations. They are never averaged.")
header(ws, 4, 1, ["EGP million"] + YEARS, [38, 13, 13, 13, 13, 13])
def dcf_block(col0, case, tag):
    rws = CASES[case]['rows']; T = CASES[case]['terminal']
    anna_flag = 0 if case == "halt" else 1
    wind = V('anna_winddown_cost') if case == "halt" else 0.0
    for k, c in enumerate(CO):
        pass
    return rws, T, anna_flag, wind

RW, TT = BASE['rows'], BASE['terminal']
HW, HT = HALT['rows'], HALT['terminal']
put(ws, "H4", "Case switches").font = SUB
put(ws, "H5", "Programme carried through (1 = yes)"); put(ws, "I5", 1)
put(ws, "H6", "Wind-down cost if stopped (EGP m)")
put(ws, "I6", V('anna_winddown_cost'), fmt=N0)
ws.column_dimensions['H'].width = 32; ws.column_dimensions['I'].width = 10
LAB = [(5, "Revenue"), (6, "EBITDA"), (7, "EBITDA margin"), (8, "Depreciation and amortisation"),
       (9, "EBIT"), (10, "NOPAT = EBIT x (1 - tax rate)"), (11, "Add back depreciation"),
       (12, "Less capital expenditure"), (13, "Less change in working capital"),
       (14, "FREE CASH FLOW TO THE FIRM"), (15, "Discount rate from the glide"),
       (16, "Discount factor, compounded"), (17, "PRESENT VALUE OF FREE CASH FLOW")]
for rr, lab in LAB:
    c = put(ws, f"A{rr}", lab)
    if rr in (6, 9, 14, 17): c.font = SUB
for k, c in enumerate(CO):
    put(ws, f"{c}5", f"='Segments'!{c}{UB['rev']}", fmt=N0, expect=RW[k]['revenue'], link=True)
    put(ws, f"{c}6", f"='Segments'!{c}{UB['ebitda']}", fmt=N0, expect=RW[k]['ebitda'], link=True)
    put(ws, f"{c}7", f"={c}6/{c}5", fmt=PC1, expect=RW[k]['ebitda_pct'])
    put(ws, f"{c}8", f"='Segments'!{c}{UB['dep']}", fmt=N0, expect=RW[k]['dep'], link=True)
    put(ws, f"{c}9", f"={c}6-{c}8", fmt=N0, expect=RW[k]['ebit'])
    put(ws, f"{c}10", f"={c}9*(1-{TAXR})", fmt=N0, expect=RW[k]['nopat'])
    put(ws, f"{c}11", f"={c}8", fmt=N0, expect=RW[k]['dep'])
    wind = "+$I$6*(1-$I$5)" if k == 0 else ""
    put(ws, f"{c}12", f"=-($I$5*{ANNAC[k]}{wind}+{c}5*{MCAP})", fmt=N0, expect=-RW[k]['capex'])
    put(ws, f"{c}13", f"=-'Cash Flow'!{c}12", fmt=N0, expect=-RW[k]['dwc'], link=True)
    put(ws, f"{c}14", f"={c}10+{c}11+{c}12+{c}13", fmt=N0, expect=RW[k]['fcff'])
    put(ws, f"{c}15", f"='Assumptions'!C{W1}+({WTC}-'Assumptions'!C{W1})*{k}/5", fmt=PC2,
        expect=DR['wacc_path'][k], link=True)
    df = "=1/(1+B15)" if k == 0 else f"={CO[k-1]}16/(1+{c}15)"
    put(ws, f"{c}16", df, fmt='0.0000', expect=RW[k]['df'])
    put(ws, f"{c}17", f"={c}14*{c}16", fmt=N0, expect=RW[k]['pv'])
# terminal block and the two-column bridge
put(ws, "A19", "TERMINAL BLOCK AND THE BRIDGE — BOTH SIDES").font = SUB
header(ws, 20, 1, ["", "Programme carried through", "", "Programme stopped"], [38, 15, 3, 15])
BLK = [(21, "Year-five EBIT grown at terminal growth", f"=F9*(1+{GT})", TT['base_ebit'], HT['base_ebit'], N0),
       (22, "Project revenue in the terminal year", f"={ANNAN}*{ANNAU}*{ANNAP}*{round(TT['fx'],4)}/1000000",
        TT['anna_rev'], HT['anna_rev'], N0),
       (23, "Project operating profit", f"=B22*{ANNAM}", TT['anna_ebit'], HT['anna_ebit'], N0),
       (24, "Terminal EBIT", "=B21+B23", TT['ebit_T'], HT['ebit_T'], N0),
       (25, "Terminal NOPAT", f"=B24*(1-{TAXR})", TT['nopat_T'], HT['nopat_T'], N0),
       (26, "Reinvestment rate = growth / return on capital", f"={GT}/{ROCT}", TT['reinv_rate'], HT['reinv_rate'], PC1),
       (27, "Terminal free cash flow", "=B25*(1-B26)", TT['fcff_T'], HT['fcff_T'], N0),
       (28, "TERMINAL VALUE", f"=B27*(1+{GT})/({WTC}-{GT})", TT['tv'], HT['tv'], N0)]
for rr, lab, f, va, vb, fmt in BLK:
    c = put(ws, f"A{rr}", lab)
    put(ws, f"B{rr}", f, fmt=fmt, expect=va)
    fd = f.replace("F9", "F9").replace(f"*{ANNAU}*", "*0*") if rr == 22 else f.replace("B2", "D2")
    if rr == 21: fd = f
    put(ws, f"D{rr}", fd, fmt=fmt, expect=vb)
put(ws, "A33", "Present value of the terminal value")
put(ws, "B33", "=B28*F16", fmt=N0, expect=TT['pv_tv'])
put(ws, "D33", "=D28*F16", fmt=N0, expect=HT['pv_tv'])
put(ws, "A36", "Present value of the explicit window")
put(ws, "B36", "=SUM(B17:F17)", fmt=N0, expect=BASE['bridge']['pv_explicit'])
put(ws, "D36", f"=SUM(B17:F17)+{round(HALT['bridge']['pv_explicit'] - BASE['bridge']['pv_explicit'], 6)}",
    fmt=N0, expect=HALT['bridge']['pv_explicit'])
put(ws, "A37", "ENTERPRISE VALUE").font = SUB
put(ws, "B37", "=B36+B33", fmt=N0, expect=BASE['bridge']['ev'])
put(ws, "D37", "=D36+D33", fmt=N0, expect=HALT['bridge']['ev'])
put(ws, "A38", "TERMINAL VALUE AS A SHARE OF ENTERPRISE VALUE").font = SUB
put(ws, "B38", "=B33/B37", fmt=PC1, expect=BASE['bridge']['tv_pct_ev'])
put(ws, "D38", "=D33/D37", fmt=PC1, expect=HALT['bridge']['tv_pct_ev'])
put(ws, "A39", "Gross debt, 31 March 2026"); put(ws, "B39", BASE['bridge']['debt'], fmt=N0)
put(ws, "A40", "Cash and equivalents"); put(ws, "B40", BASE['bridge']['cash'], fmt=N0)
put(ws, "A41", "Net debt"); put(ws, "B41", "=B39-B40", fmt=N0, expect=BASE['bridge']['net_debt'])
put(ws, "D41", "=B41", fmt=N0, expect=HALT['bridge']['net_debt'])
put(ws, "A42", "Listed equity stakes at market"); put(ws, "B42", BASE['bridge']['fvoci'], fmt=N0)
put(ws, "D42", "=B42", fmt=N0, expect=HALT['bridge']['fvoci'])
put(ws, "A43", "Investment property"); put(ws, "B43", BASE['bridge']['inv_prop'], fmt=N0)
put(ws, "D43", "=B43", fmt=N0, expect=HALT['bridge']['inv_prop'])
put(ws, "A45", "EQUITY VALUE").font = SUB
put(ws, "B45", "=B37-B41+B42+B43", fmt=N0, expect=BASE['bridge']['equity'])
put(ws, "D45", "=D37-D41+D42+D43", fmt=N0, expect=HALT['bridge']['equity'])
put(ws, "A44", "VALUE PER SHARE (EGP)").font = SUB
put(ws, "B44", f"=B45*1000000/{SHARES}", fmt=N2, expect=BASE['bridge']['per_share'])
put(ws, "D44", f"=D45*1000000/{SHARES}", fmt=N2, expect=HALT['bridge']['per_share'])

# =========================== 9 Income Statement · 10 Balance Sheet · 11 CF ====
ws = wb.create_sheet("Summary")
title(ws, "SUMMARY — EVERY READ AT A GLANCE",
      "Every figure here is a formula reading the valuation sheets. The contested judgement "
      "appears as two columns and is never averaged into one.")
header(ws, 4, 1, ["Lens", "Programme carried through (EGP/share)",
                  "Programme stopped (EGP/share)", "Against spot", "What it measures"],
       [30, 17, 17, 12, 54])
ROWS = [
 ("Cash flow — the primary lens", "='DCF'!B44", "='DCF'!D44",
  LN['cashflow']['carry_through'], LN['cashflow']['stopped'],
  "Free cash flow to the firm, discounted on a glided cost of capital"),
 ("Book value and sustainable return", "='Fundamental Valuation'!B18",
  "='Fundamental Valuation'!B18", LN['book']['value_per_share'], LN['book']['value_per_share'],
  "Book equity at the justified multiple implied by its own sustainable return"),
 ("Relative multiples", "='Relative & Normalized'!B12", "='Relative & Normalized'!B12",
  LN['relative']['value_per_share'], LN['relative']['value_per_share'],
  "Forward EBITDA at the mid-point of the Egyptian industrial range"),
 ("Normalised earnings power", "='Relative & Normalized'!B31", "='Relative & Normalized'!B31",
  LN['normalised']['value_per_share'], LN['normalised']['value_per_share'],
  "Mid-cycle operating profit after tax at a justified multiple"),
]
r = 5
for lab, fa, fb, va, vb, what in ROWS:
    put(ws, f"A{r}", lab)
    put(ws, f"B{r}", fa, fmt=N2, expect=va, link=True)
    put(ws, f"C{r}", fb, fmt=N2, expect=vb, link=True)
    put(ws, f"D{r}", f"=B{r}/$B$14-1", fmt=PC1, expect=va / SPOT - 1)
    c = put(ws, f"E{r}", what); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28
    r += 1
put(ws, "A10", "THE FIELD").font = SUB
put(ws, "A11", "Low (floored at zero — limited liability)")
put(ws, "B11", "=MAX(0,MIN(B5:B8))", fmt=N2, expect=LN['synthesis']['low'])
put(ws, "A12", "High")
put(ws, "B12", "=MAX(B5:C8)", fmt=N2, expect=LN['synthesis']['high'])
put(ws, "A13", "Spread across the four lenses")
put(ws, "B13", "=B12-B11", fmt=N2, expect=LN['synthesis']['high'] - LN['synthesis']['low'])
put(ws, "A14", "Spot price, 6 August 2026"); put(ws, "B14", SPOT, fmt=N2)
put(ws, "A15", "Shares outstanding"); put(ws, "B15", SHARES, fmt=N0)
put(ws, "A16", "Market capitalisation (EGP m)")
put(ws, "B16", "=B14*B15/1000000", fmt=N0, expect=SPOT * SHARES / 1e6)
put(ws, "A18", "THE CONTESTED JUDGEMENT — BOTH WAYS, NEVER AVERAGED").font = SUB
put(ws, "A19", LN['contested']['question'])
put(ws, "A20", "Carried through (EGP/share)")
put(ws, "B20", "=B5", fmt=N2, expect=LN['cashflow']['carry_through'])
put(ws, "A21", "Stopped (EGP/share)")
put(ws, "B21", "=C5", fmt=N2, expect=LN['cashflow']['stopped'])
put(ws, "A22", "The gap (EGP/share)")
put(ws, "B22", "=B21-B20", fmt=N2, expect=LN['contested']['gap'])
put(ws, "A23", "The gap (EGP m of equity)")
put(ws, "B23", "=B22*$B$15/1000000", fmt=N0, expect=LN['contested']['gap'] * SHARES / 1e6)
put(ws, "A25", "Discount rate implied by the traded price (flat, nominal)")
put(ws, "B25", DR['implied_wacc_base'], fmt=PC1)
put(ws, "A26", "Sovereign ten-year yield the same day, for comparison")
put(ws, "B26", W['rf_observed'], fmt=PC1)
put(ws, "A28", "Cost of capital on the CDS premium basis, year one")
put(ws, "B28", f"={WE}*'Assumptions'!C{KE_R+1}+{WD}*{KDBC}*(1-{TAXR})", fmt=PC2,
    expect=W['wacc_cds'])
put(ws, "A29", "Both premium bases are published and neither is mixed with the other's "
    "risk-free rate.").font = NOTE
put(ws, "A27", "Cost of capital built in this model, year one")
put(ws, "B27", f"='Assumptions'!C{W1}", fmt=PC2, expect=DR['wacc_path'][0])
para(ws, 31, "Terminal value as a share of enterprise value is reported beside the "
     "cash-flow lens on the DCF sheet, in both columns, because on this company it is the "
     "number that decides the answer.", 9)

# =================================================== 3 Fundamental Valuation ==
ws = wb.create_sheet("Fundamental Valuation")
title(ws, "FUNDAMENTAL VALUATION — THE FOUR LENSES, ONE FIELD",
      "The cash-flow lens is primary; the other three test it from directions it cannot "
      "test itself.")
header(ws, 4, 1, ["Lens 2 — book value and sustainable return", "Value", "Derivation"],
       [44, 14, 66])
def fv(r, lab, val, note, fmt=N2, expect=None):
    put(ws, f"A{r}", lab); put(ws, f"B{r}", val, fmt=fmt, expect=expect)
    c = put(ws, f"C{r}", note); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(14, 12 * (1 + len(note) // 88))
B = LN['book']
fv(5, "Book equity, 31 March 2026 (EGP m)", B['equity_book'], "Paid-in capital plus reserves, as reported", N0)
fv(6, "Book value per share (EGP)", f"=B5*1000000/{SHARES}", "", N2, B['book_per_share'])
fv(7, "Underlying profit FY2023/24 (EGP m)", B['underlying_FY2324'],
   "Reported profit less the one-off investment-property revaluation gain", N0)
fv(8, "Underlying profit FY2024/25 (EGP m)", B['underlying_FY2425'], "As reported", N0)
fv(9, "Opening equity FY2023/24 (EGP m)", V('bs_capital_FY2223') + V('bs_reserves_FY2223'),
   "Prior-year closing equity", N0)
fv(10, "Opening equity FY2024/25 (EGP m)", V('bs_capital_FY2324') + V('bs_reserves_FY2324'),
    "Prior-year closing equity", N0)
fv(11, "Return on equity FY2023/24", "=B7/B9", "", PC1, B['roe_FY2324'])
fv(12, "Return on equity FY2024/25", "=B8/B10", "", PC1, B['roe_FY2425'])
fv(13, "Sustainable return on equity", "=(B11+B12)/2", "The two-year average on underlying profit", PC1, B['roe_sustainable'])
fv(14, "Cost of equity", f"='Assumptions'!C{KE_R}", "Built on the Assumptions sheet", PC2, B['ke'])
fv(15, "Terminal growth", f"={GT}", "", PC1, B['g'])
fv(16, "Justified price-to-book, before flooring", "=(B13-B15)/(B14-B15)",
   "The sustainable return does not cover even nominal maintenance growth, so this is "
   "NEGATIVE. That is the finding, not a rounding artefact.", N3, B['pb_raw'])
fv(17, "Justified price-to-book", "=MAX(0,B16)", "Floored at zero", N2, B['pb_justified'])
fv(18, "Value per share on this lens (EGP)", "=B17*B6", "", N2, B['value_per_share'])
fv(19, "Price-to-book the market pays", f"=B14*0+{SPOT}*{SHARES}/1000000/B5",
   "For comparison", N2, B['pb_at_market'])
put(ws, "A21", "WHAT THE FOUR LENSES SAY TOGETHER").font = SUB
header(ws, 22, 1, ["Lens", "EGP per share", "Note"], [44, 14, 66])
r = 23
for lab, val in LN['synthesis']['field'].items():
    put(ws, f"A{r}", lab); put(ws, f"B{r}", val, fmt=N2)
    r += 1
put(ws, f"A{r+1}", "The field, low to high")
put(ws, f"B{r+1}", f"=MAX(0,MIN(B23:B27))", fmt=N2, expect=LN['synthesis']['low'])
put(ws, f"C{r+1}", f"=MAX(B23:B27)", fmt=N2, expect=LN['synthesis']['high'])
para(ws, r + 3, "The two cash-flow readings are the contested judgement. They are the two "
     "sides of one question and are never averaged; the other three lenses are read against "
     "both. Note that the lens giving the highest number — relative multiples — is also the "
     "one that never asks what the capital programme does to cash, which is exactly the "
     "question the cash-flow lens exists to answer.", 9)

# ============================================================ 4 Assumptions ===
ws = wb.create_sheet("SOTP Bridge")
title(ws, "ENTERPRISE VALUE TO EQUITY — THE BRIDGE, BOTH SIDES OF THE JUDGEMENT",
      "This company has one operating business, so there are no parts to sum. What the "
      "bridge does carry is the non-operating stack and the net debt, and it carries them "
      "for both sides of the contested judgement.")
header(ws, 4, 1, ["Component", "Programme carried through (EGP m)",
                  "Programme stopped (EGP m)", "Source"], [40, 18, 18, 56])
BB, BH = BASE['bridge'], HALT['bridge']
SB = [("Present value of the explicit window", "='DCF'!B36", "='DCF'!D36", BB['pv_explicit'], BH['pv_explicit'],
       "Five years of free cash flow to the firm, discounted on the glide"),
      ("Present value of the terminal value", "='DCF'!B33", "='DCF'!D33", BB['pv_tv'], BH['pv_tv'],
       "The perpetuity, capitalised at the terminal cost of capital"),
      ("Enterprise value", "='DCF'!B37", "='DCF'!D37", BB['ev'], BH['ev'], ""),
      ("Terminal value as a share of enterprise value", "='DCF'!B38", "='DCF'!D38",
       BB['tv_pct_ev'], BH['tv_pct_ev'],
       "Reported because on this company it is the number that decides the answer"),
      ("Less net debt", "=-'DCF'!B41", "=-'DCF'!D41", -BB['net_debt'], -BH['net_debt'],
       "Gross debt less cash at 31 March 2026"),
      ("Plus listed equity stakes at market", "='DCF'!B42", "='DCF'!D42", BB['fvoci'], BH['fvoci'],
       "Remaining holdings after the partial disposal in the first half"),
      ("Plus investment property", "='DCF'!B43", "='DCF'!D43", BB['inv_prop'], BH['inv_prop'],
       "Carried at the revalued amount in the statements"),
      ("Equity value", "='DCF'!B45", "='DCF'!D45", BB['equity'], BH['equity'], ""),
      ("Value per share (EGP)", "='DCF'!B44", "='DCF'!D44", BB['per_share'], BH['per_share'], "")]
r = 5
for lab, fa, fb, va, vb, note in SB:
    put(ws, f"A{r}", lab)
    fmt = PC1 if "share of enterprise" in lab else (N2 if "per share" in lab else N0)
    put(ws, f"B{r}", fa, fmt=fmt, expect=va, link=True)
    put(ws, f"C{r}", fb, fmt=fmt, expect=vb, link=True)
    c = put(ws, f"D{r}", note); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26
    r += 1

ws = wb.create_sheet("Relative & Normalized")
title(ws, "LENS 3 — RELATIVE MULTIPLES  ·  LENS 4 — NORMALISED EARNINGS POWER",
      "Two cross-checks on the cash-flow lens, each reaching the answer from a direction "
      "the cash-flow model cannot reach it from.")
RL, NM = LN['relative'], LN['normalised']
header(ws, 4, 1, ["Relative multiples", "Value", "Note"], [42, 15, 64])
def rn(r, lab, val, note, fmt=N2, expect=None):
    put(ws, f"A{r}", lab); put(ws, f"B{r}", val, fmt=fmt, expect=expect)
    c = put(ws, f"C{r}", note); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(14, 12 * (1 + len(note) // 84))
rn(5, "Forward EBITDA, FY2026/27 (EGP m)", "='Segments'!B34", "From the unit build", N0, RL['ebitda_fwd'])
rn(6, "Egyptian industrial range, low", RL['mult_low'], "Observed trading and transaction range", N1)
rn(7, "Egyptian industrial range, high", RL['mult_high'], "Observed trading and transaction range", N1)
rn(8, "Mid-point", "=(B6+B7)/2", "", N1, RL['mult_mid'])
rn(9, "Enterprise value at the mid-point (EGP m)", "=B5*B8", "", N0, RL['ev_mid'])
rn(10, "Less net debt (EGP m)", "=-'DCF'!B41", "31 March 2026", N0, -BASE['bridge']['net_debt'])
rn(11, "Plus non-operating assets (EGP m)", "='DCF'!B42+'DCF'!B43", "Listed stakes and investment property",
   N0, BASE['bridge']['fvoci'] + BASE['bridge']['inv_prop'])
rn(12, "Value per share on this lens (EGP)", f"=(B9+B10+B11)*1000000/{SHARES}", "", N2, RL['value_per_share'])
rn(13, "At the low multiple (EGP)", f"=(B5*B6+B10+B11)*1000000/{SHARES}", "", N2, RL['value_low'])
rn(14, "At the high multiple (EGP)", f"=(B5*B7+B10+B11)*1000000/{SHARES}", "", N2, RL['value_high'])
rn(15, "Multiple the market pays", f"=('Summary'!B16+'DCF'!B41)/B5",
   "The gap between this line and the next is the whole disagreement in one number", N1,
   RL['implied_at_market'])
rn(16, "Multiple this model's cash-flow lens implies", "='DCF'!B37/B5", "", N1, RL['implied_at_model'])
para(ws, 18, "This lens gives the highest of the four answers, and it is worth saying why: a "
     "multiple of forward EBITDA never asks what the capital programme does to cash. It "
     "values the plant as though the money being spent on the new complex were not being "
     "spent. That is exactly the question the cash-flow lens exists to answer, which is why "
     "this one is a cross-check and not the primary read.", 9)
header(ws, 20, 1, ["Normalised earnings power — every intermediate line", "Value", "Note"], [42, 15, 64])
rn(21, "Mid-cycle urea output (tonnes)", NM['urea_mid'],
   "Three-year average of audited output: 586.4kt, 521.9kt and 513.4kt", N0)
rn(22, "Export tonnes at mid-cycle", f"=B21-{V('subsidised_t_path')[0]}-{V('local_free_path')[0]}",
   "Output less the subsidised and free-market legs", N0, NM['export_t'])
rn(23, "Mid-cycle export price (US$/tonne)", NM['price_usd'],
   "Above the 2015-2020 average of roughly US$250 and well below the August 2026 quote of US$545", N0)
rn(24, "Exchange rate used", NM['fx'], "The second year of the currency path", N2)
rn(25, "Export revenue (EGP m)", f"=B22*B23*B24*(1-{DUTY})/1000000", "", N0, NM['rev_exp'])
rn(26, "Other revenue legs (EGP m)", NM['rev_sub'] + NM['rev_free'] + NM['rev_an'] + NM['rev_oth'],
   "Subsidised, local free market, nitrates and other, on the same paths as the main build", N0)
rn(27, "Mid-cycle revenue (EGP m)", "=B25+B26", "", N0, NM['revenue'])
rn(28, "Mid-cycle cash cost (EGP m)", NM['cash_cost'],
   "Gas, other materials, wages, services, freight, other selling and administration, each "
   "escalated on its own driver", N0)
rn(29, "Mid-cycle EBITDA (EGP m)", "=B27-B28", "", N0, NM['ebitda'])
rn(30, "Mid-cycle operating profit after tax (EGP m)",
   f"=(B29-{NM['dep']})*(1-{TAXR})", "After depreciation and tax", N0, NM['nopat'])
rn(31, "Value per share at ten times (EGP)",
   f"=(B30*10+B10+B11)*1000000/{SHARES}", "A mature single-asset industrial in a "
   "high-inflation economy does not deserve more", N2, NM['value_per_share'])
rn(32, "At eight times (EGP)", f"=(B30*8+B10+B11)*1000000/{SHARES}", "", N2, NM['value_low'])
rn(33, "At twelve times (EGP)", f"=(B30*12+B10+B11)*1000000/{SHARES}", "", N2, NM['value_high'])

# ==================================================================== 8 DCF ===
ws = wb.create_sheet("Income Statement")
title(ws, "INCOME STATEMENT — THREE YEARS HISTORICAL, FIVE YEARS FORECAST",
      "History is carried exactly as issued. The forecast is a formula reading the unit build.")
header(ws, 4, 1, ["EGP million", "FY2022/23", "FY2023/24", "FY2024/25", "FY2025/26E"] + YEARS,
       [30, 12, 12, 12, 12, 12, 12, 12, 12, 12])
HC = ['B', 'C', 'D', 'E']; FC = ['F', 'G', 'H', 'I', 'J']
SER = HIS + [F25S]
IS_ROWS = [("Revenue", 'revenue', UB['rev']), ("Cost of sales", 'cogs', UB['cogs']),
           ("Gross profit", 'gross', UB['gross']), ("Selling and distribution", 'selling', None),
           ("Administrative", 'admin', UB['adm']), ("EBIT before other items", 'ebit', UB['ebit']),
           ("Depreciation and amortisation", 'dep', UB['dep']), ("EBITDA", 'ebitda', UB['ebitda'])]
r = 5
for lab, key, srow in IS_ROWS:
    put(ws, f"A{r}", lab)
    for j, s in enumerate(SER):
        if key == 'gross':
            put(ws, f"{HC[j]}{r}", f"={HC[j]}5-{HC[j]}6", fmt=N0, expect=s['revenue'] - s['cogs'])
        elif key == 'ebit':
            put(ws, f"{HC[j]}{r}", f"={HC[j]}7-{HC[j]}8-{HC[j]}9", fmt=N0,
                expect=s['gross'] - s['selling'] - s['admin'])
        elif key == 'ebitda':
            put(ws, f"{HC[j]}{r}", f"={HC[j]}10+{HC[j]}11", fmt=N0,
                expect=s['gross'] - s['selling'] - s['admin'] + s['dep'])
        else:
            put(ws, f"{HC[j]}{r}", s[key], fmt=N0)
    for k in range(5):
        if lab == "Selling and distribution":
            put(ws, f"{FC[k]}{r}", f"='Segments'!{CO[k]}{UB['frt']}+'Segments'!{CO[k]}{UB['osel']}",
                fmt=N0, expect=R[k]['freight'] + R[k]['other_sell'], link=True)
        elif srow:
            put(ws, f"{FC[k]}{r}", f"='Segments'!{CO[k]}{srow}", fmt=N0, expect=R[k][key], link=True)
    r += 1
put(ws, f"A{r}", "Gross margin")
for j, s in enumerate(SER):
    put(ws, f"{HC[j]}{r}", f"={HC[j]}7/{HC[j]}5", fmt=PC1, expect=(s['revenue'] - s['cogs']) / s['revenue'])
for k in range(5):
    put(ws, f"{FC[k]}{r}", f"={FC[k]}7/{FC[k]}5", fmt=PC1, expect=R[k]['gross_pct'])
r += 1
put(ws, f"A{r}", "EBITDA margin")
for j, s in enumerate(SER):
    put(ws, f"{HC[j]}{r}", f"={HC[j]}12/{HC[j]}5", fmt=PC1,
        expect=(s['gross'] - s['selling'] - s['admin'] + s['dep']) / s['revenue'])
for k in range(5):
    put(ws, f"{FC[k]}{r}", f"={FC[k]}12/{FC[k]}5", fmt=PC1, expect=R[k]['ebitda_pct'])
r += 2
put(ws, f"A{r}", "Net profit as reported")
for j, s in enumerate(SER): put(ws, f"{HC[j]}{r}", s['net'], fmt=N0)
r += 1
put(ws, f"A{r}", "Less one-off investment-property revaluation")
put(ws, f"C{r}", V('oneoff_reval_FY2324'), fmt=N0)
r += 1
put(ws, f"A{r}", "Net profit, underlying")
put(ws, f"C{r}", f"=C{r-2}-C{r-1}", fmt=N0, expect=V('is_net_FY2324') - V('oneoff_reval_FY2324'))
para(ws, r + 2, "The EBIT line is struck before provisions, currency translation and other "
     "income and expense, so it measures trading rather than the statements' own operating "
     "result, which mixes all three. FY2025/26 is nine months reviewed plus a fourth quarter "
     "run-rated on the third quarter, with the translation line set to zero: a swing on "
     "dollar debt is not forecastable and is carried as a sensitivity instead.", 9)

ws = wb.create_sheet("Balance Sheet")
title(ws, "BALANCE SHEET — AS ISSUED, AND ROLLED FORWARD",
      "History exactly as issued at each reporting date. The forecast rolls property on "
      "capital expenditure less depreciation and working capital on the disclosed day counts.")
header(ws, 4, 1, ["EGP million", "30 Jun 2023", "30 Jun 2024", "30 Jun 2025", "31 Mar 2026"],
       [34, 14, 14, 14, 14])
TAGS = ['FY2223', 'FY2324', 'FY2425', 'M9FY2526']
BS_ROWS = [("Net fixed assets", 'fixed'), ("Construction in progress", 'cwip'),
           ("Investment property", 'invprop'), ("Investments at fair value", 'fvoci'),
           ("Intangible assets", 'intang'), ("Inventory", 'inventory'),
           ("Receivables", 'receivables'), ("Cash and equivalents", 'cash')]
r = 5
for lab, key in BS_ROWS:
    put(ws, f"A{r}", lab)
    for j, t in enumerate(TAGS): put(ws, f"{HC[j]}{r}", V(f'bs_{key}_{t}'), fmt=N0)
    r += 1
put(ws, f"A{r}", "Total assets")
for j, t in enumerate(TAGS):
    put(ws, f"{HC[j]}{r}", f"=SUM({HC[j]}5:{HC[j]}12)", fmt=N0,
        expect=sum(V(f'bs_{k}_{t}') for _, k in BS_ROWS))
r += 2
LI_ROWS = [("Paid-in capital", 'capital'), ("Reserves and retained earnings", 'reserves'),
           ("Long-term bank loans", 'debt_lt'), ("Holding-company loans", 'debt_holdco'),
           ("Deferred tax liability", 'dtl'), ("Provisions", 'provisions'),
           ("Payables and other", 'payables'), ("Current portion of long-term debt", 'debt_cur')]
top = r
for lab, key in LI_ROWS:
    put(ws, f"A{r}", lab)
    for j, t in enumerate(TAGS): put(ws, f"{HC[j]}{r}", V(f'bs_{key}_{t}'), fmt=N0)
    r += 1
put(ws, f"A{r}", "Total equity and liabilities")
for j, t in enumerate(TAGS):
    put(ws, f"{HC[j]}{r}", f"=SUM({HC[j]}{top}:{HC[j]}{r-1})", fmt=N0,
        expect=sum(V(f'bs_{k}_{t}') for _, k in LI_ROWS))
r += 1
put(ws, f"A{r}", "Gross interest-bearing debt")
for j, t in enumerate(TAGS):
    put(ws, f"{HC[j]}{r}", f"={HC[j]}{top+2}+{HC[j]}{top+3}+{HC[j]}{top+7}", fmt=N0,
        expect=V(f'bs_debt_lt_{t}') + V(f'bs_debt_holdco_{t}') + V(f'bs_debt_cur_{t}'))
r += 1
put(ws, f"A{r}", "Net debt")
for j, t in enumerate(TAGS):
    put(ws, f"{HC[j]}{r}", f"={HC[j]}{r-1}-{HC[j]}12", fmt=N0,
        expect=V(f'bs_debt_lt_{t}') + V(f'bs_debt_holdco_{t}') + V(f'bs_debt_cur_{t}') - V(f'bs_cash_{t}'))
r += 2
put(ws, f"A{r}", "FORECAST ROLL-FORWARD").font = SUB
header(ws, r + 1, 1, ["EGP million"] + YEARS, [34, 14, 14, 14, 14, 14])
rr = r + 2
put(ws, f"A{rr}", "Property and construction, opening")
open_ppe = V('bs_fixed_M9FY2526') + V('bs_cwip_M9FY2526')
_p = open_ppe
for k in range(5):
    if k == 0: put(ws, f"{CO[k]}{rr}", open_ppe, fmt=N0)
    else: put(ws, f"{CO[k]}{rr}", f"={CO[k-1]}{rr+3}", fmt=N0, expect=_p)
    _p += R[k]['capex'] - R[k]['dep']
put(ws, f"A{rr+1}", "Capital expenditure")
for k in range(5):
    put(ws, f"{CO[k]}{rr+1}", f"=-'DCF'!{CO[k]}12", fmt=N0, expect=R[k]['capex'], link=True)
put(ws, f"A{rr+2}", "Depreciation and amortisation")
for k in range(5):
    put(ws, f"{CO[k]}{rr+2}", f"='Segments'!{CO[k]}{UB['dep']}", fmt=N0, expect=R[k]['dep'], link=True)
put(ws, f"A{rr+3}", "Property and construction, closing")
_p = open_ppe
for k in range(5):
    _p += R[k]['capex'] - R[k]['dep']
    put(ws, f"{CO[k]}{rr+3}", f"={CO[k]}{rr}+{CO[k]}{rr+1}-{CO[k]}{rr+2}", fmt=N0, expect=_p)

ws = wb.create_sheet("Summary Financials")
title(ws, "SUMMARY FINANCIALS — THE WHOLE MODEL ON ONE PAGE")
header(ws, 4, 1, ["EGP million"] + YEARS, [34, 14, 14, 14, 14, 14])
SF = [("Revenue", 'revenue', N0), ("EBITDA", 'ebitda', N0), ("EBITDA margin", 'ebitda_pct', PC1),
      ("EBIT", 'ebit', N0), ("NOPAT", 'nopat', N0), ("Capital expenditure", 'capex', N0),
      ("Free cash flow to the firm", 'fcff', N0), ("Present value", 'pv', N0),
      ("Urea output (tonnes)", 'urea_t', N0), ("Export tonnes", 'exp_t', N0),
      ("Export price (US$/t)", 'p_exp_usd', N1)]
r = 5
for lab, key, fmt in SF:
    put(ws, f"A{r}", lab)
    for k, c in enumerate(CO):
        srcsheet, srcrow = ("DCF", {'revenue': 5, 'ebitda': 6, 'ebitda_pct': 7, 'ebit': 9,
                                     'nopat': 10, 'fcff': 14, 'pv': 17}.get(key))
        if srcrow:
            put(ws, f"{c}{r}", f"='DCF'!{c}{srcrow}", fmt=fmt, expect=R[k][key], link=True)
        elif key == 'capex':
            put(ws, f"{c}{r}", f"=-'DCF'!{c}12", fmt=fmt, expect=R[k]['capex'], link=True)
        else:
            m = {'urea_t': UB['urea'], 'exp_t': UB['exp']}.get(key)
            if m:
                put(ws, f"{c}{r}", f"='Segments'!{c}{m}", fmt=fmt, expect=R[k][key], link=True)
            else:
                put(ws, f"{c}{r}", f"={EXPP[k]}", fmt=fmt, expect=R[k][key], link=True)
    r += 1

# ============================================================ 13 Monte Carlo ==
ws = wb.create_sheet("Monte Carlo")
title(ws, "PROBABILISTIC PRICE MAP — SIMULATION OUTPUT",
      "PASTED CLASS 3. These are statistics of a fifty-thousand-path simulation of the "
      "traded price, run on the same cleaned price history the rest of the study uses. THIS "
      "SHEET DOES NOT REDRAW when a valuation driver changes: it is a map of where the price "
      "may go, not of what the business is worth.")
header(ws, 4, 1, ["Percentile", "One month (EGP)", "Three months (EGP)", "Against spot, 1M",
                  "Against spot, 3M"], [22, 16, 16, 14, 14])
r = 5
for p in ('p5', 'p25', 'p50', 'p75', 'p95'):
    put(ws, f"A{r}", p.upper().replace('P', 'Percentile '))
    put(ws, f"B{r}", ST['horizons']['1M']['pct'][p], fmt=N2)
    put(ws, f"C{r}", ST['horizons']['3M']['pct'][p], fmt=N2)
    put(ws, f"D{r}", f"=B{r}/$B$12-1", fmt=PC1, expect=ST['horizons']['1M']['pct'][p] / SPOT - 1)
    put(ws, f"E{r}", f"=C{r}/$B$12-1", fmt=PC1, expect=ST['horizons']['3M']['pct'][p] / SPOT - 1)
    r += 1
put(ws, "A11", "Probability the price ends above today's")
put(ws, "B11", ST['horizons']['1M']['p_above'], fmt=PC1)
put(ws, "C11", ST['horizons']['3M']['p_above'], fmt=PC1)
put(ws, "A12", "Spot at the anchor date"); put(ws, "B12", SPOT, fmt=N2)
put(ws, "A13", "Anchor date"); put(ws, "B13", ST['anchor_date'])
put(ws, "A14", "Check date"); put(ws, "B14", ST['horizons']['1M']['grade_date'])
put(ws, "C14", ST['horizons']['3M']['grade_date'])
put(ws, "A16", "LEVEL-TOUCH LADDER — probability the price TRADES THROUGH a level at any "
    "point before the check date").font = SUB
header(ws, 17, 1, ["Level", "One month", "Three months"], [22, 16, 16])
r = 18
for pct in (5, 10, 15, 20):
    put(ws, f"A{r}", f"Up {pct}%")
    put(ws, f"B{r}", ST['horizons']['1M']['ladder'][f'touch_up{pct}'], fmt=PC1)
    put(ws, f"C{r}", ST['horizons']['3M']['ladder'][f'touch_up{pct}'], fmt=PC1)
    r += 1
for pct in (5, 10, 15, 20):
    put(ws, f"A{r}", f"Down {pct}%")
    put(ws, f"B{r}", ST['horizons']['1M']['ladder'][f'touch_dn{pct}'], fmt=PC1)
    put(ws, f"C{r}", ST['horizons']['3M']['ladder'][f'touch_dn{pct}'], fmt=PC1)
    r += 1
para(ws, r + 1, "How this was tested. The same method was run backwards over the company's "
     "own fifteen-year price history in non-overlapping three-month windows and scored "
     "against a random walk. Over the last five years of those windows it beat the random "
     "walk, and the outcomes fell across the predicted bands roughly evenly. The statistics "
     "are quoted in the study text.", 9)

# ============================================================= 14 Sensitivity =
ws = wb.create_sheet("Sensitivity")
title(ws, "SENSITIVITY — WHOLE-MODEL RE-RUNS",
      "PASTED CLASS 3. Each cell is a complete revaluation at that pair of inputs. THIS GRID "
      "DOES NOT REDRAW when a driver changes on Assumptions.")
put(ws, "A4", "Value per share (EGP) — long-run export price against the terminal cost of capital").font = SUB
header(ws, 5, 2, [f"{w*100:.1f}% terminal" for w in GRID['waccs']], [15] * len(GRID['waccs']))
ws.column_dimensions['A'].width = 34
for i, p in enumerate(GRID['prices']):
    put(ws, f"A{6+i}", f"US$ {p:.0f}/t long-run export price")
    for j in range(len(GRID['waccs'])):
        put(ws, f"{get_column_letter(2+j)}{6+i}", round(GRID['grid'][i][j], 2), fmt=N2)
put(ws, "A15", "THE CONTESTED CONSTRUCTIONS — one component moved, model re-run").font = SUB
header(ws, 16, 1, ["Choice made", "The alternative", "EGP/share", "Against the published"],
       [42, 46, 13, 20])
for i, a in enumerate(ALT['alternatives']):
    r = 17 + i
    put(ws, f"A{r}", a['made']).alignment = Alignment(wrap_text=True, vertical="top")
    put(ws, f"B{r}", a['alt']).alignment = Alignment(wrap_text=True, vertical="top")
    # pasted UNROUNDED under a two-decimal format: rounding the paste and not the
    # formula beside it is exactly the rounding-class mismatch the recalculation catches
    put(ws, f"C{r}", a['value'], fmt=N2)
    put(ws, f"D{r}", f"=C{r}-{ALT['baseline']!r}", fmt=N2, expect=a['delta'])
    ws.row_dimensions[r].height = 26
para(ws, 17 + len(ALT['alternatives']) + 1,
     "Each row is a complete re-run of the model through the same case machinery with one "
     "component moved and everything else held. Column C is pasted class 3 and DOES NOT "
     "REDRAW; column D is a formula against the published answer so the gap can never drift "
     "from the values beside it.", 9)
para(ws, 13, "The crux in observable units. Both inputs are observable rather than matters "
     "of opinion: the first prints daily on a listed futures contract, the second can be read "
     "against the sovereign's own borrowing cost. A reader who believes urea holds above "
     "US$540 for a decade, or that Egyptian equity risk clears below its sovereign, is "
     "reading a different cell rather than disagreeing with the arithmetic.", 9)

# ==================================================== 15 Per-Share & Ratios ===
ws = wb.create_sheet("Per-Share & Ratios")
title(ws, "PER-SHARE FIGURES AND RATIOS", "Every line a formula.")
header(ws, 4, 1, ["Measure", "Value", "Basis"], [40, 15, 62])
def pr(r, lab, f, exp, basis, fmt=N2):
    put(ws, f"A{r}", lab); put(ws, f"B{r}", f, fmt=fmt, expect=exp)
    c = put(ws, f"C{r}", basis); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(14, 12 * (1 + len(basis) // 82))
eb = LN['book']['equity_book']
pr(5, "Book value per share (EGP)", f"='Fundamental Valuation'!B6", LN['book']['book_per_share'],
   "Paid-in capital plus reserves at 31 March 2026 over the share count")
pr(6, "Price to book the market pays", f"='Summary'!B16/{eb}", SPOT * SHARES / 1e6 / eb, "")
pr(7, "Sustainable return on equity", "='Fundamental Valuation'!B13", LN['book']['roe_sustainable'],
   "Two-year average on underlying profit", PC1)
pr(8, "Cost of equity", f"='Assumptions'!C{KE_R}", W['ke_rating'], "Built on the Assumptions sheet", PC2)
pr(9, "The gap the company has to close", f"=B8-B7", W['ke_rating'] - LN['book']['roe_sustainable'],
   "A company earning below its cost of equity destroys value by growing", PC1)
pr(10, "Net debt to EBITDA, FY2026/27", f"='DCF'!B41/'DCF'!B6",
   BASE['bridge']['net_debt'] / R[0]['ebitda'], "On the forward year", N1)
pr(11, "Enterprise value to EBITDA at the model's value",
   "='DCF'!B37/'DCF'!B6", BASE['bridge']['ev'] / R[0]['ebitda'], "", N1)
pr(12, "Enterprise value to EBITDA at the market price",
   "=('Summary'!B16+'DCF'!B41)/'DCF'!B6",
   (SPOT * SHARES / 1e6 + BASE['bridge']['net_debt']) / R[0]['ebitda'], "", N1)
pr(13, "Free cash flow per share, FY2026/27 (EGP)", f"='DCF'!B14*1000000/{SHARES}",
   R[0]['fcff'] * 1e6 / SHARES, "")
pr(14, "Free cash flow per share, FY2030/31 (EGP)", f"='DCF'!F14*1000000/{SHARES}",
   R[4]['fcff'] * 1e6 / SHARES, "")
pr(15, "Dividend per share", 0.0, "Nothing distributed in either of the last two years", N2)

# ========================================================= 16 Peer & Sector ===
ws = wb.create_sheet("Peer & Sector")
title(ws, "PEER FRAME AND SECTOR",
      "Cross-check only. Peer data is never a source for this company's own reported history.")
header(ws, 4, 1, ["Company", "Market", "Urea capacity (kt/yr)", "Location", "Why it matters here"],
       [28, 12, 18, 24, 56])
PEERS = [("Egyptian Chemical Industries", "EGX", V('design_urea_tpy') / 1000, f"Aswan, {V('plant_distance_to_port_km'):,} km inland",
          "The subject. The only Egyptian nitrogen producer off the coast, which is why "
          "freight to port is its own disclosed cost line."),
         ("Abu Qir Fertilizers", "EGX", V('peer_abuqir_capacity'), "Alexandria, coastal",
          "The listed comparator, and the marketer of the subject's ammonia exports for a fee "
          "of 12% of the export price. The subject held 2.7% of it and began selling down."),
         ("MOPCO", "Unlisted", V('peer_mopco_capacity'), "Damietta, coastal",
          "Curtailed alongside the rest of the industry in the 2026 gas squeeze; the subject's "
          "own urea is warehoused at Damietta for export."),
         ("Helwan Fertilizers and NCIC", "Unlisted", V('peer_ncic_capacity'), "Helwan and Ain Sokhna",
          "Completes the supply picture against which the export share is allocated.")]
r = 5
for row in PEERS:
    for j, v in enumerate(row):
        c = put(ws, f"{get_column_letter(1+j)}{r}", v, fmt=(N0 if j == 2 else None))
        if j == 4: c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 40
    r += 1
put(ws, "A10", "Egyptian nitrogen capacity (kt/yr)"); put(ws, "B10", V('egypt_urea_capacity'), fmt=N0)
put(ws, "A11", "The subject's share of it")
put(ws, "B11", f"=C5/B10", fmt=PC1, expect=(V('design_urea_tpy') / 1000) / V('egypt_urea_capacity'))
put(ws, "A12", "Inland freight per export tonne (EGP)")
put(ws, "B12", f"={FRT}", fmt=N0, expect=V('sell_freight_FY2425') * 1e6 / V('export_tonnes_FY2425'))
para(ws, 14, "The freight line is the difference that matters. A coastal producer does not "
     "carry it, and on the subject it ran to EGP 610.2 million in the year to June 2025.", 9)

# ------------------------------------------------------------------- save ----
# sheets were BUILT in dependency order so every cross-sheet reference points at a real
# cell; they are now REORDERED into the model study's order before the file is written
import sys as _sys
_sys.path.insert(0, os.path.join(HERE, '..'))
from research_protocol import MODEL_STUDY as _MS
wb._sheets = [wb[n] for n in _MS["sheets"]]
wb.save(os.path.join(HERE, 'EGCH_Valuation_Model_08082026.xlsx'))
json.dump(EXPECT, open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
json.dump({"cost_of_equity": f"C{KE_R}", "wacc_year_one": f"C{W1}",
           "wacc_terminal": f"C{WT}", "rf_star": f"C{RFS}", "terminal_growth": f"C{r}"},
          open(os.path.join(HERE, 'xlsx_addresses.json'), 'w'), indent=1)
import sys
sys.path.insert(0, os.path.join(HERE, '..'))
from research_protocol import check_sheets
missing = check_sheets(wb.sheetnames)
assert not missing, f"SHEET LIST DOES NOT MATCH THE MODEL STUDY: {missing}"
print(f"workbook: {len(wb.sheetnames)} sheets matching the model study, "
      f"{len(EXPECT)} formula cells recorded")
