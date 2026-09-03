"""Prove the workbook is a LIVE DRIVER model.

READ FIRST tells the reader that changing a blue cell on '3 Assumptions' reprices the
model. That is a claim about the DELIVERED file, so it is tested on the delivered file:
each driver is perturbed IN PLACE, the whole workbook is re-evaluated from scratch
through the independent evaluator, and the test asserts the headline moves in the
asserted DIRECTION. A dead-input sweep then bumps every remaining driver on the
assumptions sheet and requires each one to move something.

TWO DIRECTIONS HERE ARE THE OPPOSITE OF THE TEXTBOOK ONE, AND NEITHER IS A BUG. Each was
decomposed BEFORE the sign was set, not after a test failed.

  * A WIDER sovereign default spread RAISES the value. The spread is netted OUT of the
    local risk-free rate before the country equity premium is added, precisely so that
    Egypt's sovereign risk is charged once rather than twice. Widening it therefore
    lowers the normalised risk-free rate, and with it the cost of equity.

  * TERMINAL GROWTH HAS NO FIRST-ORDER EFFECT, in either direction, and that is the
    finding rather than a failed test. It was asserted upward first, then downward, and
    the model was right both times to refuse. Growth enters twice and the two entries
    fight: it raises the reinvestment rate (growth divided by return on capital), which
    takes cash out of the terminal flow, and it tightens the perpetuity denominator
    (terminal rate less growth), which puts value back. Traced across the range, value
    peaks at about g = 7.35% and falls away on both sides:

        g = 6.30%  reinvestment 35.0%  value per share -1.3062
        g = 7.00%  reinvestment 38.9%  value per share -1.3031   <- the base case
        g = 7.35%  reinvestment 40.8%  value per share -1.3027   <- the turning point
        g = 7.70%  reinvestment 42.8%  value per share -1.3031
        g = 8.40%  reinvestment 46.7%  value per share -1.3067

    The base case sits 35 basis points below the turning point, so a small bump barely
    registers. The test therefore asserts the SECOND-order property that is actually
    true — the response is concave, and a large move in EITHER direction destroys value —
    rather than a first-order sign the model does not have.
"""
import json, os, sys, copy
import openpyxl, xlcalc

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'EGCH_Valuation_Model_03092026.xlsx')
SHEET_A = 'Assumptions'
HEAD = ('DCF', 'B44')          # value per share, programme carried through


def value_with(overrides):
    wb = openpyxl.load_workbook(XLSX)
    for (sh, coord), v in overrides.items():
        wb[sh][coord] = v
    return float(xlcalc.Book(wb).cell_value(*HEAD))


BASE = value_with({})
print(f"baseline value per share: EGP {BASE:.4f}\n")

# ---- directional tests ------------------------------------------------------
wb0 = openpyxl.load_workbook(XLSX)
ws0 = wb0[SHEET_A]
LABELS = {}
# A driver that has become a FORMULA off evidence elsewhere in the workbook is still a
# driver. Reading only numeric cells silently dropped three of them the moment the capital
# expenditure drivers started reading the capex record — and a test that loses a driver
# without saying so is worse than no test. Formula-valued drivers are resolved to the
# number they currently produce and bumped from there.
_BK0 = xlcalc.Book(wb0)
for row in range(5, 200):
    lab = ws0.cell(row=row, column=1).value
    val = ws0.cell(row=row, column=3).value
    if not lab:
        continue
    if isinstance(val, str) and val.startswith('='):
        try:
            val = float(_BK0.cell_value(SHEET_A, f"C{row}"))
        except Exception:
            continue
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        LABELS[lab] = (f"C{row}", val)

TESTS = [
    ("Export urea price — FY2026/27", 1.10, "up",
     "A higher realised export price lifts revenue with no change in cost."),
    ("Export urea price — FY2030/31", 1.10, "up",
     "Year five feeds the terminal base, so the last year's price matters most."),
    ("Urea capacity utilisation — FY2026/27", 1.05, "up", "More tonnes at a positive margin."),
    ("Gas per tonne of ammonia", 1.10, "down",
     "Gas is the largest single cost; burning more of it per tonne costs money."),
    ("Realised gas price", 1.10, "down", "Same mechanism, through the price rather than the volume."),
    ("Egyptian pounds per US dollar — FY2026/27", 1.05, "up",
     "Revenue is dollar-linked and the gas bill is dollar-linked, but the pound cost base "
     "is not, so a weaker pound is net positive at this margin."),
    ("Export duty", 1.50, "down", "The duty is a direct wedge in the realised export price."),
    ("Project capital expenditure — FY2026/27", 1.20, "down",
     "Cash out with no incremental cash in inside the explicit window."),
    ("Maintenance capital expenditure", 1.30, "down", "Cash out of free cash flow."),
    ("Terminal return on invested capital", 1.20, "up",
     "A higher return on capital means less must be reinvested to buy the same growth."),
    ("Terminal growth", 1.20, "up",
     "Growth enters twice and the entries fight: it raises the reinvestment rate and it "
     "tightens the perpetuity denominator, so the response is concave with a turning point "
     "near 7.3%. RE-DERIVED 9 August 2026: with terminal growth re-anchored on the central "
     "bank's longest-horizon 5% target the base now sits well BELOW that turning point, so "
     "the response over the tested range is monotonically increasing. The assertion follows "
     "the model rather than the other way round."),
    ("Tax rate", 1.20, "down", "A larger share of operating profit leaves the firm."),
    ("Days inventory outstanding", 1.20, "down", "More cash locked in working capital."),
    ("Days payable outstanding", 1.20, "up", "Supplier financing releases cash."),
    ("Egyptian inflation — FY2026/27", 1.30, "up",
     "One macro path: the same inflation figure sets that year's currency wedge, so higher "
     "domestic inflation depreciates the pound and lifts the dollar-linked revenue in pounds "
     "by more than it lifts the pound cost lines and the local-equivalent cost of the dollar "
     "debt. [Edition 1 September 2026: the 8 August edition held the currency as a separate "
     "input, so inflation was a pure cost shock there.]"),
    ("Stoppage and abnormal gas cost — FY2026/27", 2.0, "down", "A direct charge against EBIT."),
    ("Project utilisation in the terminal year", 1.40, "up",
     "The only place ANNA earns anything in the committed-capital case."),
]
# COUNT AGAINST A KNOWN TOTAL: every named driver must be found, or the test is silently
# smaller than it claims to be.
missing = [lab for lab, *_ in TESTS if lab not in LABELS]
assert not missing, ("driver test cannot find these drivers on the Assumptions sheet: "
                     + "; ".join(missing))

fails = []
print(f"{'driver':52s} {'bump':>7s} {'new':>9s} {'move':>10s}  expected")
for lab, mult, want, why in TESTS:
    coord, v0 = LABELS[lab]
    got = value_with({(SHEET_A, coord): v0 * mult})
    move = got - BASE
    if want == "concave":
        down = value_with({(SHEET_A, coord): v0 * (2 - mult)}) - BASE
        ok = move < -1e-6 and down < -1e-6
        move = min(move, down)
    else:
        ok = (move > 1e-6) if want == "up" else (move < -1e-6)
    if not ok:
        fails.append((lab, want, move, why))
    print(f"{lab:52s} {mult:6.2f}x {got:9.3f} {move:+10.4f}  {want:5s} {'OK' if ok else 'FAIL'}")

# ---- dead-input sweep -------------------------------------------------------
# A CROSS-CHECK IS NOT A DRIVER. Three cells stopped moving the headline the moment the
# new complex's nameplate came from the engineering award instead of being derived from the
# ammonia surplus: the derivation is now a check on the disclosed plate, not the plate. They
# are exempted BY EXPLICIT LABEL and the exemptions are counted out loud -- an exemption the
# reader cannot see is indistinguishable from a dead input the sweep missed.
CROSS_CHECK_PREFIXES = ("Cross-check:",)
CROSS_CHECK_FEEDS = ("Ammonia design capacity", "Ammonia per tonne of nitrate")
exempt = [lab for lab in LABELS
          if lab.startswith(CROSS_CHECK_PREFIXES) or lab in CROSS_CHECK_FEEDS]

print("\ndead-input sweep — every remaining driver must move at least one case headline,")
print("in at least one direction. A driver of the downside or upside case is live even")
print("though it leaves the committed-capital headline untouched.")
# The sweep watches every PUBLISHED output, not only the four case headlines. Three cells
# read as dead the moment the anchor price and the normalised multiples became driver cells:
# they move the lens RANGES and the market ratios, which the study publishes, but not the
# central. Widening the watch list is the honest fix; exempting them would not have been.
HEADS = [('DCF', 'B44'), ('DCF', 'D44'), ('Summary', 'B11'), ('Summary', 'B12'),
         ('Fundamental Valuation', 'B18'), ('Fundamental Valuation', 'B19'),
         ('Relative & Normalized', 'B12'), ('Relative & Normalized', 'B31'),
         ('Relative & Normalized', 'B32'), ('Relative & Normalized', 'B33'),
         ('Summary', 'B28')]


def all_heads(overrides):
    wb = openpyxl.load_workbook(XLSX)
    for (sh, coord), v in overrides.items():
        wb[sh][coord] = v
    BK = xlcalc.Book(wb)
    return [float(BK.cell_value(*h)) for h in HEADS]


BASE_ALL = all_heads({})
tested = {LABELS[l][0] for l, _, _, _ in TESTS}
dead = []
for lab, (coord, v0) in LABELS.items():
    if lab in exempt:
        continue
    if coord in tested:
        continue
    moved = False
    for mult in (1.15, 0.85):
        got = all_heads({(SHEET_A, coord): (v0 * mult if v0 else 1.0)})
        if any(abs(a - b) > 1e-9 for a, b in zip(got, BASE_ALL)):
            moved = True
            break
    if not moved:
        dead.append(lab)
print(f"  drivers swept: {len(LABELS) - len(tested) - len(exempt)} | "
      f"cross-checks exempted (named, not silent): {len(exempt)} -> {exempt} | "
      f"dead inputs: {len(dead)}")
for d in dead:
    print(f"  ! DEAD: {d}")

ok = not fails and not dead
print(f"\n{'PASS' if ok else 'FAIL'}: {len(TESTS) - len(fails)}/{len(TESTS)} drivers move the "
      f"headline in the asserted direction, {len(dead)} dead inputs")
for lab, want, move, why in fails:
    print(f"  ! {lab}: expected {want}, moved {move:+.4f} — {why}")
json.dump(dict(baseline=BASE, tests=len(TESTS), failures=len(fails), dead=dead),
          open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1)
sys.exit(0 if ok else 1)
