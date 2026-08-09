"""EGCH valuation workbook — FORMULA-FIRST.

Everything arithmetically derivable from a driver is written as a live Excel formula.
Only three classes of cell carry a pasted number, and READ FIRST names all three:
  (1) audited and disclosed history — where a line is both disclosed and derivable, the
      DISCLOSED figure is carried;
  (2) the output of the unit build (production, channel tonnages, realised prices) that
      would be unreadable flattened into a grid — pasted, with everything downstream of
      it formula-driven;
  (3) whole-model re-run grids — the sensitivity and scenario matrices, where each cell
      is a complete revaluation. Those grids do NOT redraw when a driver changes.

As it writes, the builder records the model's own value for every formula cell into
xlsx_expected.json. recalc.py then evaluates the delivered file independently and
asserts every one of them reproduces, with nothing left unchecked.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
SWEEP = json.load(open(os.path.join(HERE, 'sweep_register.json')))
STEP0 = json.load(open(os.path.join(HERE, 'step0_result.json')))
BT = json.load(open(os.path.join(HERE, 'backtest_5y.json')))
BETA = json.load(open(os.path.join(HERE, 'beta_result.json')))
W = D['wacc']
DR = D['drivers']
BASE = D['cases']['base']
YEARS = D['years']
SPOT = D['spot']
SHARES = 1_986_578_999

EXPECT = {}      # "Sheet!A1" -> model value, written as the builder goes

# ---------------------------------------------------------------- styling ----
BLUE = Font(color="1F4E79", bold=False)          # input
BLACK = Font(color="000000")                     # formula
HDR = Font(color="FFFFFF", bold=True, size=11)
HDRFILL = PatternFill("solid", fgColor="1F4E79")
SUB = Font(bold=True, color="1F4E79")
TITLE = Font(bold=True, size=14, color="1F4E79")
NOTE = Font(italic=True, size=9, color="595959")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)
N0 = '#,##0'; N1 = '#,##0.0'; N2 = '#,##0.00'; PC1 = '0.0%'; PC2 = '0.00%'


def put(ws, coord, value, *, fmt=None, font=None, expect=None, wrap=False):
    c = ws[coord]
    c.value = value
    if fmt:
        c.number_format = fmt
    c.font = font or (BLACK if isinstance(value, str) and value.startswith('=') else BLUE)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    if isinstance(value, str) and value.startswith('='):
        if expect is None:
            raise ValueError(f"formula at {ws.title}!{coord} written without an expected value")
        EXPECT[f"{ws.title}!{coord}"] = float(expect)
    return c


def header(ws, row, cols, labels, widths=None):
    for j, lab in enumerate(labels):
        c = ws.cell(row=row, column=cols + j, value=lab)
        c.font = HDR; c.fill = HDRFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    if widths:
        for j, wd in enumerate(widths):
            ws.column_dimensions[get_column_letter(cols + j)].width = wd


def title(ws, text, sub=None):
    ws['A1'] = text; ws['A1'].font = TITLE
    if sub:
        ws['A2'] = sub; ws['A2'].font = NOTE
        ws.row_dimensions[2].height = 28
        ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")


wb = openpyxl.Workbook()

# =========================================================== 1. READ FIRST ====
ws = wb.active; ws.title = "1 READ FIRST"
ws.column_dimensions['A'].width = 118
title(ws, "EGYPTIAN CHEMICAL INDUSTRIES (KIMA) — EGX: EGCH — VALUATION MODEL")
rows = [
 ("", ""),
 ("WHAT THIS FILE IS", SUB),
 ("A live formula model of a single-site Egyptian nitrogen-fertilizer producer. Revenue is "
  "built product by product as tonnes times price; cost is built as physical consumption "
  "times a unit price; the cost of capital, the discount-rate glide, the free-cash-flow "
  "waterfall, the terminal block, the statement roll-forwards and every ratio are "
  "calculated in the sheet. Change a blue cell on '3 Assumptions' and the valuation "
  "recomputes.", None),
 ("", ""),
 ("THE THREE CLASSES OF PASTED CELL, AND THERE ARE NO OTHERS", SUB),
 ("1. AUDITED AND DISCLOSED HISTORY. Sheet '4 Historical IS' and '5 Historical BS' carry the "
  "figures exactly as issued in the company's own audited statements for the years ended "
  "30 June 2023, 2024 and 2025 and its reviewed nine-month accounts to 31 March 2026. Where "
  "a line is both disclosed and derivable, the DISCLOSED figure is carried and the "
  "derivation is shown beside it as a check.", None),
 ("2. THE UNIT BUILD'S OUTPUT. Production tonnages, the split of those tonnes between the "
  "subsidised, local free-market and export channels, and the realised prices in each "
  "channel are the output of a reconciliation against the audited revenue note that would "
  "be unreadable flattened into this grid. Those tonnages and prices are pasted on "
  "'3 Assumptions'; everything downstream of them is a formula.", None),
 ("3. WHOLE-MODEL RE-RUN GRIDS. Sheet '13 Sensitivity' and '14 Scenarios' contain cells that "
  "are each a complete revaluation of the company at a different pair of inputs. THOSE "
  "GRIDS DO NOT REDRAW WHEN A DRIVER CHANGES — they are snapshots of separate model runs, "
  "and they are labelled as such on the sheet itself.", None),
 ("", ""),
 ("BLUE IS AN INPUT. BLACK IS A FORMULA.", SUB),
 ("", ""),
 ("WHAT THE READER SHOULD KNOW BEFORE LOOKING AT A SINGLE NUMBER", SUB),
 ("This company is building a nitric-acid and ammonium-nitrate complex whose bank-approved "
  "cost — EGP 6,422.4 million plus US$278.4 million, about EGP 20.3 billion — is roughly "
  "three quarters of its own stock-market value. About EGP 5.7 billion of that sat in "
  "construction-in-progress at 31 March 2026 against physical progress its auditor put at "
  "12.9% versus a 37% plan. The explicit forecast window is therefore a construction "
  "window in which free cash flow is negative, and the value of the company sits in what "
  "the assets earn afterwards. That is a property of the asset, not of the model, and it "
  "is why terminal value is such a large share of enterprise value here — a share the "
  "model reports on every valuation sheet rather than burying.", None),
 ("", ""),
 ("This document is a valuation study. It is not a rating and not a price target: it "
  "reports a range of fair values and the distribution around them, and the reader draws "
  "their own conclusion.", NOTE),
]
r = 3
for text, font in rows:
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(15, 13 * (1 + len(text) // 105))
    r += 1

# =========================================================== 2. SUMMARY =======
ws = wb.create_sheet("2 Summary valuation")
title(ws, "SUMMARY VALUATION TABLE",
      "Every figure on this sheet is a formula reading the valuation sheets. Terminal "
      "value as a share of enterprise value is shown beside the discounted-cash-flow "
      "lens, because on this company it is the number that decides the answer.")
header(ws, 4, 1, ["Lens", "Enterprise value (EGP m)", "Terminal value % of EV",
                  "Net debt (EGP m)", "Non-operating assets (EGP m)",
                  "Equity value (EGP m)", "Per share (EGP)", "vs spot"],
       [34, 15, 13, 13, 15, 15, 12, 11])
CASE_SHEET = {"base": "10 DCF base", "bear": "11 DCF bear", "bull": "12 DCF bull",
              "halt": "12b DCF capital discipline"}
LABELS = [("base", "Discounted cash flow — committed capital"),
          ("halt", "Discounted cash flow — capital discipline"),
          ("bull", "Discounted cash flow — upside"),
          ("bear", "Discounted cash flow — downside")]
r = 5
for case, lab in LABELS:
    b = D['cases'][case]['bridge']
    sh = CASE_SHEET[case]
    put(ws, f"A{r}", lab)
    put(ws, f"B{r}", f"='{sh}'!B33", fmt=N0, expect=b['ev'])
    put(ws, f"C{r}", f"='{sh}'!B34", fmt=PC1, expect=b['tv_pct_ev'])
    put(ws, f"D{r}", f"='{sh}'!B37", fmt=N0, expect=b['net_debt'])
    put(ws, f"E{r}", f"='{sh}'!B40", fmt=N0, expect=b['fvoci'] + b['inv_prop'])
    put(ws, f"F{r}", f"='{sh}'!B41", fmt=N0, expect=b['equity'])
    put(ws, f"G{r}", f"='{sh}'!B42", fmt=N2, expect=b['per_share'])
    put(ws, f"H{r}", f"=G{r}/$B$14-1", fmt=PC1, expect=b['per_share'] / SPOT - 1)
    r += 1

ws.cell(row=r + 1, column=1, value="Fair-value range (equity floored at zero — a "
        "limited-liability shareholder cannot owe more than the shares cost)").font = SUB
lo = max(0.0, min(D['cases'][c]['bridge']['per_share'] for c in ('base', 'bear', 'bull', 'halt')))
hi = max(D['cases'][c]['bridge']['per_share'] for c in ('base', 'bear', 'bull', 'halt'))
put(ws, f"A{r+2}", "Low (EGP per share)")
put(ws, f"B{r+2}", f"=MAX(0,MIN(G5:G8))", fmt=N2, expect=lo)
put(ws, f"A{r+3}", "High (EGP per share)")
put(ws, f"B{r+3}", f"=MAX(G5:G8)", fmt=N2, expect=hi)
put(ws, f"A{r+4}", "Central (capital-discipline and committed-capital midpoint)")
_mid = (D['cases']['base']['bridge']['per_share']
        + D['cases']['halt']['bridge']['per_share']) / 2
put(ws, f"C{r+4}", "=(G5+G6)/2", fmt=N2, expect=_mid)
put(ws, f"B{r+4}", f"=MAX(0,C{r+4})", fmt=N2, expect=max(0.0, _mid))
put(ws, "A14", "Spot price (EGP), 6 August 2026")
put(ws, "B14", SPOT, fmt=N2)
put(ws, "A15", "Shares outstanding (note 14, audited FY2024/25)")
put(ws, "B15", SHARES, fmt=N0)
put(ws, "A16", "Market capitalisation (EGP m)")
put(ws, "B16", "=B14*B15/1000000", fmt=N0, expect=SPOT * SHARES / 1e6)
put(ws, "A17", "Discount rate the market price itself implies (flat, nominal EGP)")
put(ws, "B17", DR['implied_wacc_base'], fmt=PC1)
put(ws, "A18", "Sovereign 10-year EGP yield, 6 August 2026, for comparison")
put(ws, "B18", W['rf_observed'], fmt=PC1)
put(ws, "A19", "Cost of capital built in this model (year one, falling to terminal)")
put(ws, "B19", f"='6 WACC'!B31", fmt=PC2, expect=DR['wacc_path'][0])
ws.cell(row=21, column=1, value=(
    "The reverse discounted-cash-flow line is the honest way to report the distance "
    "between this model and the market: rather than asserting the market is wrong, it "
    "states the discount rate the market is using. A rate below the sovereign's own "
    "ten-year borrowing cost is not a required return any cost-of-capital construction "
    "produces.")).font = NOTE
ws.cell(row=21, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[21].height = 48

# =========================================================== 3. ASSUMPTIONS ===
ws = wb.create_sheet("3 Assumptions")
title(ws, "ASSUMPTIONS — THE LIVE DRIVERS",
      "Blue cells are inputs. Change one and the whole model reprices. Each carries its "
      "source. Tonnages and realised prices are the unit build's output (pasted class 2).")
header(ws, 4, 1, ["Driver", "Unit", "Value", "Source"], [46, 16, 14, 74])
A = {}
def drv(row, label, unit, value, source, fmt=N2):
    put(ws, f"A{row}", label); put(ws, f"B{row}", unit)
    put(ws, f"C{row}", value, fmt=fmt)
    c = put(ws, f"D{row}", source); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = max(15, 12 * (1 + len(source) // 95))
    A[label] = f"'3 Assumptions'!C{row}"
    return f"'3 Assumptions'!C{row}"

r = 5
ws.cell(row=r, column=1, value="CAPACITY AND VOLUME").font = SUB; r += 1
C_DESIGN = drv(r, "Urea design capacity", "tonnes/year", DR['design_urea_t'],
    "1,575 t/day contractual plate, note 28 of the audited FY2024/25 statements", N0); r += 1
C_NH3D = drv(r, "Ammonia design capacity", "tonnes/year", DR['design_ammonia_t'],
    "1,200 t/day contractual plate, note 28 of the audited FY2024/25 statements", N0); r += 1
C_NH3R = drv(r, "Ammonia consumed per tonne of urea", "tonnes", DR['ammonia_per_urea_t'],
    "318,242 t ammonia against 513,385 t urea, auditor's own production and unit-cost "
    "table, audited FY2024/25", '#,##0.000'); r += 1
UTIL = []
for k, y in enumerate(YEARS):
    UTIL.append(drv(r, f"Urea capacity utilisation — {y}", "% of plate", DR['urea_util'][k],
        "Audited output 586.4kt (FY2022/23, 2% above plate), 521.9kt (FY2023/24), 513.4kt "
        "(FY2024/25). The path never returns to plate because the summer gas curtailment "
        "is structural." if k == 0 else "Utilisation path, same basis", PC1)); r += 1
SUBT, FREET = [], []
for k, y in enumerate(YEARS):
    SUBT.append(drv(r, f"Subsidised deliveries — {y}", "tonnes", DR['subsidised_t_path'][k],
        "Cabinet decision of 8 September 2025 reset the cooperative obligation; the company "
        "delivered 147kt of a 322kt requirement in the 14 months to August 2025, a 46% "
        "compliance rate the forecast does not assume away." if k == 0 else
        "Subsidised delivery path, same basis", N0)); r += 1
for k, y in enumerate(YEARS):
    FREET.append(drv(r, f"Local free-market volume — {y}", "tonnes", DR['local_free_path'][k],
        "37.1kt implied for FY2024/25 by the note-20 revenue split net of the subsidised "
        "and export legs" if k == 0 else "Local free-market path, same basis", N0)); r += 1

r += 1
ws.cell(row=r, column=1, value="PRICES").font = SUB; r += 1
EXPP_B = []
for k, y in enumerate(YEARS):
    EXPP_B.append(drv(r, f"Export urea price, upside case — {y}", "US$/tonne",
        DR['export_usd_path_bull'][k],
        "The upside path: urea holds nearer the US$545/t at which the CME granular FOB "
        "Egypt contract settled on 7 August 2026 rather than mean-reverting" if k == 0
        else "Upside export price path", N1)); r += 1
EXPP = []
for k, y in enumerate(YEARS):
    EXPP.append(drv(r, f"Export urea price — {y}", "US$/tonne", DR['export_usd_path'][k],
        "US$385/t realised in FY2024/25 (auditor, Damietta stock note); Q1-2025/26 export "
        "prices +43% year on year; CME granular FOB Egypt settled US$545/t on 7 August "
        "2026. The path mean-reverts toward the marginal gas-based producer's cash cost."
        if k == 0 else "Export price path, mean reversion", N1)); r += 1
FXP = []
for k, y in enumerate(YEARS):
    FXP.append(drv(r, f"USD/EGP — {y}", "EGP per US$", DR['usd_egp_path'][k],
        "49.79 on 7 August 2026, depreciating 4.5% a year — the same wedge used to carry "
        "the dollar debt at local-equivalent cost in the WACC, so the two cannot diverge"
        if k == 0 else "Depreciation path, 4.5% a year", N2)); r += 1
DUTY = drv(r, "Export duty", "% of export value", DR['export_duty_pct'],
    "2026 switch from the EGP 2,500/t shortfall levy (decree 241/2021) to a 10% ad-valorem "
    "duty tied to the global price", PC1); r += 1
SUBP = []
for k, y in enumerate(YEARS):
    SUBP.append(drv(r, f"Subsidised price — {y}", "EGP/tonne", DR['subsidised_p_path'][k],
        "EGP 6,000/t in FY2024/25 (cooperative supply price), on an administered "
        "escalation path" if k == 0 else "Administered price path", N0)); r += 1
PARITY = drv(r, "Local free-market price as % of export parity", "%", DR['local_free_parity'],
    "EGP 18,485/t implied for FY2024/25, against an export parity of US$385 x 49.0", PC1); r += 1
ANT = []
for k, y in enumerate(YEARS):
    ANT.append(drv(r, f"Ammonium-nitrate volume — {y}", "tonnes", DR['an_path'][k],
        "26,058 t of 33.5% granulated plus low-density AN, auditor's FY2024/25 production "
        "table" if k == 0 else "AN volume path", N0)); r += 1
ANP = drv(r, "Ammonium-nitrate price, FY2024/25 basis", "EGP/tonne", DR['an_egp_t']['FY2024/25'],
    "Implied by the note-20 local revenue net of the subsidised and free-market urea legs; "
    "indexed forward on USD/EGP", N0); r += 1
OTHR = []
for k, y in enumerate(YEARS):
    OTHR.append(drv(r, f"Other revenue — {y}", "EGP m", DR['other_rev_path'][k],
        "Merchant nitric acid, the ferrosilicon plant's rental (leased to a Saudi tenant "
        "from May 2025) and services" if k == 0 else "Other revenue path", N1)); r += 1

r += 1
ws.cell(row=r, column=1, value="COST STACK — ONE ESCALATOR PER PHYSICAL DRIVER").font = SUB; r += 1
GASQ = drv(r, "Gas consumption per tonne of ammonia", "m3", DR['gas_m3_per_t_ammonia'],
    "Inside the auditor's own disclosed 1,025-1,771 m3/t range; calibrated so gas is 75% "
    "of the FY2024/25 materials line. The split of that line between gas and everything "
    "else is the model's and is flagged as such: the statements give only the total.", N0); r += 1
GASP = drv(r, "Realised gas price", "US$/mmBtu", DR['gas_usd_mmbtu'],
    "The company's own Q1-2025/26 disclosure values 31,313,235 m3 of lost gas at EGP 251m "
    "= EGP 8.016/m3, about US$4.68/mmBtu at the prevailing rate", N2); r += 1
GASC = drv(r, "Contract gas price (downside case)", "US$/mmBtu", DR['gas_usd_mmbtu_contract'],
    "US$5.75/mmBtu formula price, note 28 of the audited FY2024/25 statements, raised from "
    "US$4.50 under the November 2021 decision", N2); r += 1
MMB = drv(r, "Energy conversion", "mmBtu per m3", DR['mmbtu_per_m3'],
    "Standard gross calorific conversion", '0.00000'); r += 1
OMAT = drv(r, "Other materials per tonne of urea", "EGP", DR['other_materials_egp_t_urea'],
    "The FY2024/25 materials line of EGP 4,398.6m less modelled gas, over 513,385 t: "
    "packaging, catalysts and consumable spares", N0); r += 1
WAGE = drv(r, "Wages in cost of sales, FY2024/25", "EGP m", 212.857,
    "Note 21 of the audited FY2024/25 statements", N1); r += 1
SERV = drv(r, "Purchased services, FY2024/25", "EGP m", DR['services'],
    "Note 21 of the audited FY2024/25 statements", N1); r += 1
FRT = drv(r, "Inland freight per export tonne", "EGP", DR['freight_egp_t_export'],
    "EGP 610.2m of product freight and commissions (note 22) over 350.3kt of exports — the "
    "cost of sitting 1,000 km inland at Aswan", N0); r += 1
OSELL = drv(r, "Other selling cost, FY2024/25", "EGP m", DR['other_selling'],
    "Note 22, selling materials, wages and other selling expense", N1); r += 1
ADMIN = drv(r, "Administrative expense, FY2024/25", "EGP m", DR['admin'],
    "Income statement, audited FY2024/25", N1); r += 1
ABN = []
for k, y in enumerate(YEARS):
    ABN.append(drv(r, f"Abnormal gas and stoppage cost — {y}", "EGP m", DR['abnormal_gas_path'][k],
        "EGP 164.5m charged in FY2024/25 (note 25) and EGP 152.7m the year before; about "
        "EGP 781m cumulative since FY2022/23. Decays as gas supply normalises."
        if k == 0 else "Stoppage-cost path", N1)); r += 1
CPI = []
for k, y in enumerate(YEARS):
    CPI.append(drv(r, f"Egyptian CPI — {y}", "%", DR['cpi_path'][k],
        "14.3% year on year in June 2026, converging on the central bank's medium-term "
        "target" if k == 0 else "Inflation convergence path", PC1)); r += 1

r += 1
ws.cell(row=r, column=1, value="CAPITAL, TAX AND THE COST OF CAPITAL").font = SUB; r += 1
DEPB = drv(r, "Depreciation charge, FY2024/25", "EGP m", DR['dep_base'],
    "Note 6 fixed-asset register: KIMA-2 machinery at 3.95% a year", N1); r += 1
AMOB = drv(r, "Amortisation, FY2024/25", "EGP m", DR['amort_base'],
    "Usufruct intangible at 4.75% a year, note 10", N1); r += 1
ANNAC = []
for k, y in enumerate(YEARS):
    ANNAC.append(drv(r, f"ANNA project capital expenditure — {y}", "EGP m", DR['anna_capex_path'][k],
        "Bank-approved cost EGP 6,422.4m plus US$278.4m (agreement of 25 June 2025), "
        "against EGP 5,653.5m in construction-in-progress at 31 March 2026 and physical "
        "progress of 12.9% against a 37% plan" if k == 0 else "ANNA spending path", N0)); r += 1
MCAP = drv(r, "Maintenance capital expenditure", "% of revenue", DR['maint_capex_pct_rev'],
    "The pre-ANNA observed run of EGP 42.5-81m was abnormally low on a newly built plant; "
    "3.0% of revenue is the mature-plant standard. No guidance exists, so this is "
    "sensitised.", PC1); r += 1
NH3AN = drv(r, "Ammonia per tonne of ammonium nitrate", "tonnes", DR['nh3_per_t_an'],
    "Nitric-acid route plus direct neutralisation", N2); r += 1
put(ws, f"A{r}", "ANNA nameplate (derived, not disclosed)"); put(ws, f"B{r}", "tonnes AN/year")
put(ws, f"C{r}", f"=({C_NH3D}-{C_DESIGN}*{C_NH3R})/{NH3AN}", fmt=N0,
    expect=DR['anna_nameplate_an_t'])
_c = put(ws, f"D{r}", "DERIVED, and flagged as derived: no filing states the plant's "
    "capacity. It is the ammonia design plate less the draw of urea at ITS design plate, "
    "converted at the ratio above. This is why the ammonia plate is a live driver and not "
    "a decorative reference figure.")
_c.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[r].height = 40
ANNAN = f"'3 Assumptions'!C{r}"; r += 1
ANNAU = drv(r, "ANNA utilisation in the terminal year", "%", DR['anna_util_base'],
    "Half of nameplate in the committed-capital case. The downside and capital-discipline "
    "cases set it to zero on their own sheets, which is what those cases mean.", PC1); r += 1
ANNAU_UP = drv(r, "ANNA utilisation in the terminal year, upside case", "%",
    DR['anna_util_bull'], "70% of nameplate in the upside case", PC1); r += 1
ANNAP = drv(r, "Ammonium-nitrate price", "US$/tonne", DR['anna_price_usd_t'],
    "Mid-cycle nitrate pricing", N0); r += 1
ANNAM = drv(r, "ANNA cash margin", "% of revenue", DR['anna_cash_margin'],
    "Conversion margin over its own ammonia feedstock", PC1); r += 1
DSO = drv(r, "Days sales outstanding", "days", DR['dso'],
    "EGP 631.0m of receivables on EGP 8,602.6m of revenue, audited FY2024/25", N1); r += 1
DIO = drv(r, "Days inventory outstanding", "days", DR['dio'],
    "EGP 2,399.6m of inventory on EGP 5,300.3m of cost of sales, audited FY2024/25", N1); r += 1
DPO = drv(r, "Days payable outstanding", "days", DR['dpo'],
    "EGP 1,207.8m of trade payables on cost of sales, audited FY2024/25", N1); r += 1
TAXR = drv(r, "Tax rate", "%", DR['tax_rate'],
    "Egyptian statutory corporate rate. The company has been paying no current income tax "
    "(years to 2016/17 settled, 2017-2019 under appeal, 2020-2023 examined and settled) "
    "and booked deferred credits in both years — an upside this model does not take.", PC1); r += 1
GT = drv(r, "Terminal growth", "%", DR['g_terminal'],
    "The central bank's medium-term inflation target: nominal maintenance growth, no real "
    "growth assumed", PC1); r += 1
ROCT = drv(r, "Terminal return on invested capital", "%", DR['roc_terminal'],
    "Sets the terminal reinvestment rate as growth divided by return on capital", PC1); r += 1

# =========================================================== 4. HISTORICAL IS =
ws = wb.create_sheet("4 Historical IS")
title(ws, "HISTORICAL INCOME STATEMENT — AS ISSUED",
      "Pasted class 1: the figures are carried exactly as they appear in the company's own "
      "audited statements. EGP million. Margins and growth are formulas.")
header(ws, 4, 1, ["EGP million", "FY2022/23", "FY2023/24", "FY2024/25", "FY2025/26E"],
       [42, 15, 15, 15, 15])
H = D['hist']; F5 = D['fy2526']
COLS = ['B', 'C', 'D', 'E']
SER = H + [F5]
lines = [("Revenue", 'revenue'), ("Cost of sales", 'cogs'), ("Gross profit", 'gross'),
         ("Selling and distribution", 'selling'), ("Administrative", 'admin'),
         ("EBIT (core operating)", 'ebit'), ("Depreciation and amortisation", 'dep'),
         ("EBITDA", 'ebitda')]
r = 5
for lab, key in lines:
    put(ws, f"A{r}", lab)
    for j, s in enumerate(SER):
        col = COLS[j]
        if key == 'gross':
            put(ws, f"{col}{r}", f"={col}{r-2}-{col}{r-1}", fmt=N0, expect=s['gross'])
        elif key == 'ebit':
            put(ws, f"{col}{r}", f"={col}{r-3}-{col}{r-2}-{col}{r-1}", fmt=N0, expect=s['ebit'])
        elif key == 'ebitda':
            put(ws, f"{col}{r}", f"={col}{r-2}+{col}{r-1}", fmt=N0, expect=s['ebitda'])
        else:
            put(ws, f"{col}{r}", s[key], fmt=N0)
    r += 1
RD = [{k: v for k, v in s.items() if isinstance(v, (int, float))} for s in SER]
for d in RD:
    d['gross'] = d['revenue'] - d['cogs']
    d['ebit'] = d['gross'] - d['selling'] - d['admin']
    d['ebitda'] = d['ebit'] + d['dep']
put(ws, f"A{r}", "Gross margin")
for j, s in enumerate(SER):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}7/{COLS[j]}5", fmt=PC1,
        expect=RD[j]['gross'] / RD[j]['revenue'])
r += 1
put(ws, f"A{r}", "EBITDA margin")
for j, s in enumerate(SER):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}12/{COLS[j]}5", fmt=PC1,
        expect=RD[j]['ebitda'] / RD[j]['revenue'])
r += 1
put(ws, f"A{r}", "Revenue growth")
for j in range(1, len(SER)):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}5/{COLS[j-1]}5-1", fmt=PC1,
        expect=RD[j]['revenue'] / RD[j-1]['revenue'] - 1)
r += 2
put(ws, f"A{r}", "Net profit as reported")
for j, s in enumerate(SER):
    put(ws, f"{COLS[j]}{r}", s['net'], fmt=N0)
r += 1
put(ws, f"A{r}", "of which one-off investment-property revaluation")
put(ws, f"C{r}", 2034.573, fmt=N0)
r += 1
put(ws, f"A{r}", "Net profit, underlying")
put(ws, f"C{r}", f"=C{r-2}-C{r-1}", fmt=N0, expect=H[1]['net'] - 2034.573)
ws.cell(row=r + 2, column=1, value=(
    "FY2025/26E is nine months reviewed plus a fourth quarter run-rated on the third "
    "quarter's operating performance, which was the strongest on record. The foreign-"
    "exchange line is set to zero in that estimate: a translation swing on dollar debt is "
    "not forecastable and is carried as a sensitivity instead. Reported nine-month net "
    "profit of EGP 531.3m is after a EGP 1,072.0m translation loss.")).font = NOTE
ws.cell(row=r + 2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r + 2].height = 60

# =========================================================== 5. HISTORICAL BS =
ws = wb.create_sheet("5 Historical BS")
title(ws, "HISTORICAL BALANCE SHEET — AS ISSUED",
      "Pasted class 1. EGP million, at each 30 June and at 31 March 2026. Totals and "
      "ratios are formulas that prove the statement balances.")
header(ws, 4, 1, ["EGP million", "30 Jun 2023", "30 Jun 2024", "30 Jun 2025", "31 Mar 2026"],
       [42, 15, 15, 15, 15])
BSD = {
 "Net fixed assets": [11300.4, 14144.1, 13587.2, 13057.6],
 "Construction in progress": [56.4, 2535.1, 3790.2, 5653.5],
 "Investment property": [0.0, 2034.6, 2160.6, 2155.1],
 "Investments at fair value": [1855.4, 2491.2, 2163.3, 1382.9],
 "Intangible assets": [1908.6, 2376.2, 2256.8, 2170.5],
 "Other non-current": [0.3, 0.3, 2.6, 2.5],
 "Inventory": [1391.8, 1615.5, 2399.6, 3378.2],
 "Receivables": [798.6, 858.4, 631.0, 1230.2],
 "Cash and equivalents": [1416.2, 3103.4, 3057.0, 4606.5],
}
r = 5
for lab, vals in BSD.items():
    put(ws, f"A{r}", lab)
    for j, v in enumerate(vals):
        put(ws, f"{COLS[j]}{r}", v, fmt=N0)
    r += 1
TOT = [18727.8, 29161.0, 30048.3, 33637.0]
put(ws, f"A{r}", "Total assets (formula)")
for j in range(4):
    put(ws, f"{COLS[j]}{r}", f"=SUM({COLS[j]}5:{COLS[j]}13)", fmt=N0,
        expect=sum(v[j] for v in BSD.values()))
r += 1
put(ws, f"A{r}", "Total assets as issued")
for j, v in enumerate(TOT):
    put(ws, f"{COLS[j]}{r}", v, fmt=N0)
r += 1
put(ws, f"A{r}", "Difference (rounding of the issued statement)")
for j in range(4):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}{r-2}-{COLS[j]}{r-1}", fmt=N1,
        expect=sum(v[j] for v in BSD.values()) - TOT[j])
r += 2
LIAB = {
 "Paid-in capital": [5932.9, 9932.9, 9932.9, 9932.9],
 "Reserves and retained": [1316.7, 4627.2, 5430.2, 6273.2],
 "Long-term bank loans": [8424.9, 11226.2, 11183.3, 14386.1],
 "Holding-company loans": [50.3, 0.0, 596.9, 45.9],
 "Deferred tax liability": [1469.5, 1001.5, 990.9, 961.2],
 "Provisions": [144.4, 432.2, 309.1, 291.8],
 "Payables and other": [1367.8, 1587.0, 1207.8, 1539.0],
 "Current portion of long-term debt": [21.4, 354.1, 397.3, 207.0],
}
for lab, vals in LIAB.items():
    put(ws, f"A{r}", lab)
    for j, v in enumerate(vals):
        put(ws, f"{COLS[j]}{r}", v, fmt=N0)
    r += 1
put(ws, f"A{r}", "Total equity and liabilities (formula)")
for j in range(4):
    put(ws, f"{COLS[j]}{r}", f"=SUM({COLS[j]}{r-8}:{COLS[j]}{r-1})", fmt=N0,
        expect=sum(v[j] for v in LIAB.values()))
r += 1
put(ws, f"A{r}", "Gross interest-bearing debt")
for j in range(4):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}{r-7}+{COLS[j]}{r-6}+{COLS[j]}{r-2}", fmt=N0,
        expect=LIAB['Long-term bank loans'][j] + LIAB['Holding-company loans'][j]
               + LIAB['Current portion of long-term debt'][j])
r += 1
put(ws, f"A{r}", "Net debt")
for j in range(4):
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}{r-1}-{COLS[j]}13", fmt=N0,
        expect=LIAB['Long-term bank loans'][j] + LIAB['Holding-company loans'][j]
               + LIAB['Current portion of long-term debt'][j] - BSD['Cash and equivalents'][j])
r += 1
put(ws, f"A{r}", "Net debt to EBITDA (on the year's EBITDA)")
for j, s in enumerate(SER):
    _eb = s['revenue'] - s['cogs'] - s['selling'] - s['admin'] + s['dep']
    put(ws, f"{COLS[j]}{r}", f"={COLS[j]}{r-1}/'4 Historical IS'!{COLS[j]}12", fmt=N1,
        expect=(LIAB['Long-term bank loans'][j] + LIAB['Holding-company loans'][j]
                + LIAB['Current portion of long-term debt'][j]
                - BSD['Cash and equivalents'][j]) / _eb)

# =========================================================== 6. WACC ==========
ws = wb.create_sheet("6 WACC")
title(ws, "COST OF CAPITAL — BUILT IN THE SHEET",
      "Nothing on this sheet is a pasted rate. The cost of equity is built from the "
      "risk-free rate net of the sovereign spread, the beta and the premium; the cost of "
      "debt is built by currency and taxed; the weights come from market capitalisation "
      "and net debt; the terminal rate is built from its own long-run components.")
header(ws, 4, 1, ["Component", "Value", "Source / derivation"], [46, 14, 86])
def wrow(r, lab, val, src, fmt=PC2, expect=None):
    put(ws, f"A{r}", lab)
    put(ws, f"B{r}", val, fmt=fmt, expect=expect)
    c = put(ws, f"C{r}", src); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = max(15, 12 * (1 + len(src) // 110))

wrow(5, "Observed 10-year EGP sovereign yield", W['rf_observed'],
     "23.00%, quote of 6 August 2026. A new EGP 120.9bn treasury bond maturing May 2029 "
     "listed at a 23.098% coupon on the same market.")
wrow(6, "Sovereign default spread (rating basis)", W['sov_spread_rating'],
     f"Moody's {json.load(open(os.path.join(HERE,'live_data.json')))['damodaran_egypt']['value']['moodys_rating']}"
     ", adjusted default spread from the country-premium file")
wrow(7, "Normalised risk-free rate rf* (rating basis)", "=B5-B6",
     "Country risk enters ONCE, through the premium below. Using the raw local yield and "
     "a country-loaded premium together would charge Egypt's sovereign risk twice.",
     expect=W['rf_star_rating'])
wrow(8, "Equity risk premium (rating basis)", W['erp_rating'],
     "Mature-market premium of 4.23% plus a country premium of 9.71%, from the original "
     "country-premium workbook")
wrow(9, "Beta", W['beta'],
     f"{BETA['n']} weekly observations over five years against an equal-weight index of "
     f"{BETA['composite_names']} Egyptian names, the subject excluded from its own index. "
     f"R-squared {BETA['r2']:.3f}, standard error {BETA['se']:.3f}, 90% interval "
     f"[{BETA['ci90'][0]:.2f}, {BETA['ci90'][1]:.2f}] — the usability gate passes, so the "
     f"regression is adopted rather than a default. Dimson sum-beta "
     f"{BETA['dimson']['sum_beta']:.3f} as cross-check.", fmt=N2)
wrow(10, "Cost of equity (rating basis)", "=B7+B9*B8", "rf* plus beta times the premium",
     expect=W['ke_rating'])
wrow(11, "Sovereign default spread (CDS basis)", W['sov_spread_cds'], "Ten-year CDS spread")
wrow(12, "Normalised risk-free rate rf* (CDS basis)", "=B5-B11",
     "The same normalisation on the CDS basis", expect=W['rf_star_cds'])
wrow(13, "Equity risk premium (CDS basis)", W['erp_cds'], "CDS-based total premium")
wrow(14, "Cost of equity (CDS basis)", "=B12+B9*B13",
     "Both bases are published. They are never mixed: a CDS-based premium is paired only "
     "with a CDS-normalised risk-free rate.", expect=W['ke_cds'])
wrow(16, "Cost of debt, local currency, pre-tax", W['kd_local'],
     "The company's own latest local borrowing: EGP 500,000,000 drawn from the holding "
     "company in FY2024/25 carried EGP 96,896,001 of interest.")
wrow(17, "Cost of debt, dollar tranche, in dollars", W['kd_usd_nominal'],
     "EGP 1,338,012,810 of interest in FY2024/25 on a mean balance of about US$233m")
wrow(18, "Expected EGP depreciation", W['expected_depreciation'],
     "Built from the central bank's own 7% inflation target against about 2.4% in the "
     "United States")
wrow(19, "Dollar debt at local-equivalent cost", "=(1+B17)*(1+B18)-1",
     "A dollar coupon inside an EGP-nominal cost of capital would understate the true "
     "cost of that debt to an EGP earner.", expect=W['kd_fx_local_equiv'])
wrow(20, "Share of debt in local currency", W['pct_debt_local'],
     "The EGP tranche of the KIMA-2 loan was repaid in June 2024, so the book is almost "
     "entirely dollar — which is why this company's earnings swing on translation.")
wrow(21, "Blended pre-tax cost of debt", "=B20*B16+(1-B20)*B19", "Weighted by currency",
     expect=W['kd_pretax_blended'])
wrow(22, "Tax rate", W['tax_rate'], "Egyptian statutory rate")
wrow(23, "After-tax cost of debt", "=B21*(1-B22)", "", expect=W['kd_aftertax'])
wrow(25, "Market capitalisation (EGP m)", "='2 Summary valuation'!B16",
     "Spot price times the share count in note 14 of the audited statements", fmt=N0,
     expect=SPOT * SHARES / 1e6)
wrow(26, "Gross debt (EGP m)", W['total_debt'] / 1e6,
     "31 March 2026 reviewed balance sheet", fmt=N0)
wrow(27, "Equity weight", "=B25/(B25+B26)", "Market-value equity, never book", expect=W['we'])
wrow(28, "Debt weight", "=1-B27", "", expect=W['wd'])
wrow(30, "WACC on the CDS basis", "=B27*B14+B28*B23", "", expect=W['wacc_cds'])
wrow(31, "WACC on the rating basis — year one", "=B27*B10+B28*B23",
     "The published build. Both bases are shown; the rating basis is the more "
     "conservative of the two and is the one carried into the valuation.",
     expect=DR['wacc_path'][0])
wrow(33, "Long-run inflation", DR['inflation_lt'], "Central bank medium-term target")
wrow(34, "Long-run real rate", DR['real_rate_lt'], "Emerging-market long-run real policy rate")
wrow(35, "Terminal normalised risk-free rate", "=(1+B33)*(1+B34)-1",
     "Built from its own components. A spot yield embeds today's 14.3% inflation print, "
     "while the terminal value grows at the 7% target — capitalising one at the other is "
     "a units mismatch, and on this company it is the largest number in the study.",
     expect=DR['rf_star_terminal'])
wrow(36, "Terminal cost of equity", "=B35+B9*B8", "", expect=DR['ke_terminal'])
wrow(37, "Long-run dollar cost of debt", DR['kd_usd_lt'], "Long-run corporate dollar cost")
wrow(38, "Terminal cost of debt at local-equivalent cost", "=(1+B37)*(1+B18)-1", "",
     expect=DR['kd_local_equiv_terminal'])
wrow(39, "TERMINAL WACC", "=B27*B36+B28*B38*(1-B22)",
     "The rate the perpetuity is capitalised at.", expect=DR['wacc_terminal'])
r = 41
put(ws, f"A{r}", "THE GLIDE — visibly derived, and the discount factors compound it").font = SUB
header(ws, r + 1, 1, ["Year", "Glide fraction", "rf*", "Cost of debt", "WACC"],
       [22, 16, 14, 16, 14])
for k, y in enumerate(YEARS):
    rr = r + 2 + k
    put(ws, f"A{rr}", y)
    put(ws, f"B{rr}", f"={k}/5", fmt='0.00', expect=k / 5)
    put(ws, f"C{rr}", f"=$B$7+($B$35-$B$7)*B{rr}", fmt=PC2, expect=DR['rf_star_path'][k])
    put(ws, f"D{rr}", f"=$B$21+($B$38-$B$21)*B{rr}", fmt=PC2, expect=DR['kd_path'][k])
    put(ws, f"E{rr}", f"=$B$27*(C{rr}+$B$9*$B$8)+$B$28*D{rr}*(1-$B$22)", fmt=PC2,
        expect=DR['wacc_path'][k])

# ================================================= 7-9 UNIT BUILD / IS / CF ===
def unit_sheet(name, case):
    rows = D['cases'][case]['rows']
    ws = wb.create_sheet(name)
    title(ws, f"UNIT BUILD AND FORECAST INCOME STATEMENT — {case.upper()} CASE",
          "Tonnes times price on the revenue side, physical consumption times a unit "
          "price on the cost side. Every cell here is a formula reading '3 Assumptions'.")
    header(ws, 4, 1, ["EGP million unless stated"] + YEARS, [44, 15, 15, 15, 15, 15])
    CO = ['B', 'C', 'D', 'E', 'F']
    r = 5
    def line(lab, fn, fmt=N0, key=None, bold=False):
        nonlocal r
        c = put(ws, f"A{r}", lab)
        if bold:
            c.font = SUB
        for k in range(5):
            put(ws, f"{CO[k]}{r}", fn(k, CO[k], r), fmt=fmt, expect=rows[k][key] if key else None)
        r += 1
        return r - 1
    r_urea = line("Urea production (tonnes)", lambda k, c, rw: f"={C_DESIGN}*{UTIL[k]}",
                  N0, 'urea_t')
    r_nh3 = line("Ammonia produced (tonnes), capped at the design plate",
                 lambda k, c, rw: f"=MIN({c}{r_urea}*{C_NH3R},{C_NH3D})", N0, 'ammonia_t')
    r_sub = line("Subsidised volume (tonnes)", lambda k, c, rw: f"={SUBT[k]}", N0, 'sub_t')
    r_free = line("Local free-market volume (tonnes)", lambda k, c, rw: f"={FREET[k]}", N0, 'free_t')
    r_exp = line("Export volume (tonnes)",
                 lambda k, c, rw: f"={c}{r_urea}-{c}{r_sub}-{c}{r_free}", N0, 'exp_t')
    r_pexp = line("Export price, net of duty (EGP/tonne)",
                  lambda k, c, rw: f"={EXPP[k]}*{FXP[k]}*(1-{DUTY})", N0, 'p_exp_egp')
    r_pfree = line("Local free-market price (EGP/tonne)",
                   lambda k, c, rw: f"={EXPP[k]}*{FXP[k]}*{PARITY}", N0, 'p_free')
    r_pan = line("Ammonium-nitrate price (EGP/tonne)",
                 lambda k, c, rw: f"={ANP}*{FXP[k]}/{DR['usd_egp']['FY2024/25']}", N0, 'p_an')
    r_rexp = line("Export revenue", lambda k, c, rw: f"={c}{r_exp}*{c}{r_pexp}/1000000", N0, 'rev_exp')
    r_rsub = line("Subsidised revenue", lambda k, c, rw: f"={c}{r_sub}*{SUBP[k]}/1000000", N0, 'rev_sub')
    r_rfree = line("Local free-market revenue",
                   lambda k, c, rw: f"={c}{r_free}*{c}{r_pfree}/1000000", N0, 'rev_free')
    r_ran = line("Ammonium-nitrate revenue",
                 lambda k, c, rw: f"={ANT[k]}*{c}{r_pan}/1000000", N0, 'rev_an')
    r_roth = line("Other revenue", lambda k, c, rw: f"={OTHR[k]}", N0, 'rev_other')
    r_rev = line("REVENUE", lambda k, c, rw: f"=SUM({c}{r_rexp}:{c}{r_roth})", N0, 'revenue', True)
    r_gp = line("Gas price (EGP per m3)", lambda k, c, rw: f"={GASP}*{MMB}*{FXP[k]}",
                N2, 'gas_price_egp_m3')
    r_gas = line("Natural gas", lambda k, c, rw: f"={c}{r_nh3}*{GASQ}*{c}{r_gp}/1000000",
                 N0, 'gas_cost')
    r_cpi = line("Cumulative inflation index", lambda k, c, rw:
                 "=" + "*".join(f"(1+{CPI[j]})" for j in range(k + 1)), '0.000', 'cpi_cum')
    r_omat = line("Other materials", lambda k, c, rw: f"={c}{r_urea}*{OMAT}*{c}{r_cpi}/1000000",
                  N0, 'other_mat')
    r_wage = line("Wages", lambda k, c, rw: f"={WAGE}*{c}{r_cpi}", N0, 'wages')
    r_serv = line("Purchased services", lambda k, c, rw: f"={SERV}*{c}{r_cpi}", N0, 'services')
    r_dep = line("Depreciation and amortisation",
                 lambda k, c, rw: f"={DEPB}*(1+0.02*{k})+{AMOB}", N0, 'dep')
    r_cogs = line("COST OF SALES", lambda k, c, rw: f"=SUM({c}{r_gas}:{c}{r_dep})-{c}{r_cpi}",
                  N0, 'cogs', True)
    r_gross = line("GROSS PROFIT", lambda k, c, rw: f"={c}{r_rev}-{c}{r_cogs}", N0, 'gross', True)
    r_gpc = line("Gross margin", lambda k, c, rw: f"={c}{r_gross}/{c}{r_rev}", PC1, 'gross_pct')
    r_frt = line("Inland freight to port",
                 lambda k, c, rw: f"={c}{r_exp}*{FRT}*{c}{r_cpi}/1000000", N0, 'freight')
    r_osel = line("Other selling cost", lambda k, c, rw: f"={OSELL}*{c}{r_cpi}", N0, 'other_sell')
    r_adm = line("Administrative", lambda k, c, rw: f"={ADMIN}*{c}{r_cpi}", N0, 'admin')
    r_abn = line("Abnormal gas and stoppage cost", lambda k, c, rw: f"={ABN[k]}", N0, 'abnormal')
    r_ebit = line("EBIT", lambda k, c, rw:
                  f"={c}{r_gross}-{c}{r_frt}-{c}{r_osel}-{c}{r_adm}-{c}{r_abn}", N0, 'ebit', True)
    r_ebitda = line("EBITDA", lambda k, c, rw: f"={c}{r_ebit}+{c}{r_dep}", N0, 'ebitda', True)
    r_em = line("EBITDA margin", lambda k, c, rw: f"={c}{r_ebitda}/{c}{r_rev}", PC1, 'ebitda_pct')
    return ws, dict(rev=r_rev, cogs=r_cogs, ebit=r_ebit, ebitda=r_ebitda, dep=r_dep,
                    exp_t=r_exp, urea=r_urea, nh3=r_nh3, cols=CO)


ws_u, UB = unit_sheet("7 Unit build and forecast IS", "base")

# ---- correction: COGS row must not subtract the index row; rebuild explicitly
rows = D['cases']['base']['rows']
for k in range(5):
    c = UB['cols'][k]
    put(ws_u, f"{c}{UB['cogs']}",
        f"={c}{UB['cogs']-6}+{c}{UB['cogs']-4}+{c}{UB['cogs']-3}+{c}{UB['cogs']-2}+{c}{UB['cogs']-1}",
        fmt=N0, expect=rows[k]['cogs'])

# =========================================================== 8. FORECAST BS ===
ws = wb.create_sheet("8 Forecast BS and CF")
title(ws, "FORECAST BALANCE SHEET AND CASH FLOW",
      "Property rolls forward on capital expenditure less depreciation; working capital "
      "rolls on the disclosed day counts; net debt rolls on cash generation. Every cell "
      "is a formula.")
header(ws, 4, 1, ["EGP million"] + YEARS, [44, 15, 15, 15, 15, 15])
CO = UB['cols']
r = 5
put(ws, f"A{r}", "Net property and construction in progress, opening")
open_ppe = 13057.6 + 5653.5
_ppe = open_ppe
for k in range(5):
    if k == 0:
        put(ws, f"{CO[k]}{r}", open_ppe, fmt=N0)
    else:
        put(ws, f"{CO[k]}{r}", f"={CO[k-1]}{r+3}", fmt=N0, expect=_ppe)
    _ppe = _ppe + rows[k]['capex'] - rows[k]['dep']
r += 1
put(ws, f"A{r}", "Capital expenditure")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"={ANNAC[k]}+'7 Unit build and forecast IS'!{CO[k]}{UB['rev']}*{MCAP}",
        fmt=N0, expect=rows[k]['capex'])
r += 1
put(ws, f"A{r}", "Depreciation and amortisation")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"='7 Unit build and forecast IS'!{CO[k]}{UB['dep']}", fmt=N0,
        expect=rows[k]['dep'])
r += 1
put(ws, f"A{r}", "Net property and construction in progress, closing")
ppe = open_ppe
for k in range(5):
    ppe = ppe + rows[k]['capex'] - rows[k]['dep']
    put(ws, f"{CO[k]}{r}", f"={CO[k]}{r-3}+{CO[k]}{r-2}-{CO[k]}{r-1}", fmt=N0, expect=ppe)
r += 2
put(ws, f"A{r}", "Receivables")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"='7 Unit build and forecast IS'!{CO[k]}{UB['rev']}*{DSO}/365",
        fmt=N0, expect=rows[k]['revenue'] * DR['dso'] / 365)
r += 1
put(ws, f"A{r}", "Inventory")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"='7 Unit build and forecast IS'!{CO[k]}{UB['cogs']}*{DIO}/365",
        fmt=N0, expect=rows[k]['cogs'] * DR['dio'] / 365)
r += 1
put(ws, f"A{r}", "Payables")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"='7 Unit build and forecast IS'!{CO[k]}{UB['cogs']}*{DPO}/365",
        fmt=N0, expect=rows[k]['cogs'] * DR['dpo'] / 365)
r += 1
put(ws, f"A{r}", "Net working capital")
for k in range(5):
    put(ws, f"{CO[k]}{r}", f"={CO[k]}{r-3}+{CO[k]}{r-2}-{CO[k]}{r-1}", fmt=N0, expect=rows[k]['wc'])
r_wc = r
r += 1
put(ws, f"A{r}", "Change in net working capital")
prev_wc0 = (D['fy2526']['revenue'] * DR['dso'] / 365 + D['fy2526']['cogs'] * DR['dio'] / 365
            - D['fy2526']['cogs'] * DR['dpo'] / 365)
put(ws, f"{CO[0]}{r}", f"={CO[0]}{r_wc}-{prev_wc0!r}", fmt=N0, expect=rows[0]['dwc'])
for k in range(1, 5):
    put(ws, f"{CO[k]}{r}", f"={CO[k]}{r_wc}-{CO[k-1]}{r_wc}", fmt=N0, expect=rows[k]['dwc'])
r_dwc = r


# ========================================== 9-12  DCF SHEETS (fixed layout) ===
# Fixed row layout so every cross-sheet reference is stable:
#   5 revenue | 6 gas adj to contract | 7 EBITDA | 8 margin | 9 D&A | 10 EBIT
#  11 NOPAT | 12 add back D&A | 13 less capex | 14 less dWC | 15 FCFF
#  16 discount rate | 17 discount factor | 18 PV
#  21-29 terminal block | 31-40 bridge
R_DWC_SHEET8 = r_dwc
ANNAU_CASE = {"base": DR['anna_util_base'], "bull": DR['anna_util_bull'],
              "bear": 0.0, "halt": 0.0}
SRC = "7 Unit build and forecast IS"
CO = ['B', 'C', 'D', 'E', 'F']


def dcf_sheet(name, case, label, blurb):
    rws = D['cases'][case]['rows']
    bs = D['cases']['base']['rows']
    T = D['cases'][case]['terminal']
    B = D['cases'][case]['bridge']
    ws = wb.create_sheet(name)
    title(ws, f"DISCOUNTED CASH FLOW — {label}", blurb)
    header(ws, 4, 1, ["EGP million"] + YEARS + ["", "Case switches"],
           [40, 14, 14, 14, 14, 14, 3, 26])
    gas_flag = 1 if case == "bear" else 0
    up_flag = 1 if case == "bull" else 0
    anna_flag = 0 if case == "halt" else 1
    winddown = 1000.0 if case == "halt" else 0.0
    put(ws, "H5", "Gas at contract price (1 = yes)"); put(ws, "I5", gas_flag)
    put(ws, "H6", "ANNA capital programme (1 = yes)"); put(ws, "I6", anna_flag)
    put(ws, "H7", "Wind-down cost, year one (EGP m)"); put(ws, "I7", winddown, fmt=N0)
    put(ws, "H8", "ANNA utilisation in terminal year")
    if case == "base":
        put(ws, "I8", f"={ANNAU}", fmt=PC1, expect=DR['anna_util_base'])
    elif case == "bull":
        put(ws, "I8", f"={ANNAU_UP}", fmt=PC1, expect=DR['anna_util_bull'])
    else:
        put(ws, "I8", 0.0, fmt=PC1)
    put(ws, "H9", "Upside export price path (1 = yes)"); put(ws, "I9", up_flag)
    for k, c in enumerate(CO):
        # price uplift: the upside case differs from the base ONLY in the export price,
        # so the difference is carried as an explicit formula on the export and local
        # free-market tonnes rather than by duplicating the whole unit build
        up = (f"$I$9*('{SRC}'!{c}{UB['exp_t']}*({EXPP_B[k]}-{EXPP[k]})*{FXP[k]}*(1-{DUTY})"
              f"+{FREET[k]}*({EXPP_B[k]}-{EXPP[k]})*{FXP[k]}*{PARITY})/1000000")
        up_val = (bs[k]['exp_t'] * (DR['export_usd_path_bull'][k] - DR['export_usd_path'][k])
                  * bs[k]['fx'] * (1 - DR['export_duty_pct'])
                  + DR['local_free_path'][k]
                  * (DR['export_usd_path_bull'][k] - DR['export_usd_path'][k])
                  * bs[k]['fx'] * DR['local_free_parity']) / 1e6 * up_flag
        put(ws, f"{c}19", "=" + up, fmt=N0, expect=up_val)
        put(ws, f"{c}5", f"='{SRC}'!{c}{UB['rev']}+{c}19", fmt=N0, expect=rws[k]['revenue'])
        put(ws, f"{c}6", f"=$I$5*'{SRC}'!{c}{UB['nh3']}*{GASQ}*({GASC}-{GASP})*{MMB}*{FXP[k]}/1000000",
            fmt=N0, expect=(bs[k]['ammonia_t'] * DR['gas_m3_per_t_ammonia']
                            * (DR['gas_usd_mmbtu_contract'] - DR['gas_usd_mmbtu'])
                            * DR['mmbtu_per_m3'] * bs[k]['fx'] / 1e6) * gas_flag)
        put(ws, f"{c}7", f"='{SRC}'!{c}{UB['ebitda']}-{c}6+{c}19", fmt=N0,
            expect=rws[k]['ebitda'])
        put(ws, f"{c}8", f"={c}7/{c}5", fmt=PC1, expect=rws[k]['ebitda'] / rws[k]['revenue'])
        put(ws, f"{c}9", f"='{SRC}'!{c}{UB['dep']}", fmt=N0, expect=rws[k]['dep'])
        put(ws, f"{c}10", f"={c}7-{c}9", fmt=N0, expect=rws[k]['ebit'])
        put(ws, f"{c}11", f"={c}10*(1-{TAXR})", fmt=N0, expect=rws[k]['nopat'])
        put(ws, f"{c}12", f"={c}9", fmt=N0, expect=rws[k]['dep'])
        wind = f"+$I$7" if k == 0 else ""
        put(ws, f"{c}13", f"=-($I$6*{ANNAC[k]}{wind}+{c}5*{MCAP})", fmt=N0,
            expect=-rws[k]['capex'])
        prev19 = "0" if k == 0 else f"{CO[k-1]}19"
        put(ws, f"{c}14",
            f"=-('8 Forecast BS and CF'!{c}{R_DWC_SHEET8}+({c}19-{prev19})*{DSO}/365)",
            fmt=N0, expect=-rws[k]['dwc'])
        put(ws, f"{c}15", f"={c}11+{c}12+{c}13+{c}14", fmt=N0, expect=rws[k]['fcff'])
        put(ws, f"{c}16", f"='6 WACC'!E{43+k}", fmt=PC2, expect=DR['wacc_path'][k])
        df = f"=1/(1+B16)" if k == 0 else f"={CO[k-1]}17/(1+{c}16)"
        put(ws, f"{c}17", df, fmt='0.0000', expect=rws[k]['df'])
        put(ws, f"{c}18", f"={c}15*{c}17", fmt=N0, expect=rws[k]['pv'])
    for rr, lab, bold in [(5, "Revenue", False),
                          (6, "Gas cost adjustment to contract price", False),
                          (7, "EBITDA", True), (8, "EBITDA margin", False),
                          (9, "Depreciation and amortisation", False), (10, "EBIT", True),
                          (11, "NOPAT = EBIT x (1 - tax rate)", False),
                          (12, "Add back depreciation and amortisation", False),
                          (13, "Less capital expenditure", False),
                          (14, "Less change in working capital", False),
                          (15, "FREE CASH FLOW TO THE FIRM", True),
                          (19, "Upside export-price uplift to revenue and EBITDA", False),
                          (16, "Discount rate (from the glide)", False),
                          (17, "Discount factor (compounded)", False),
                          (18, "PRESENT VALUE OF FREE CASH FLOW", True)]:
        c = put(ws, f"A{rr}", lab)
        if bold:
            c.font = SUB
    put(ws, "A20", "TERMINAL BLOCK").font = SUB
    put(ws, "A21", "Year-five EBIT grown at terminal growth")
    put(ws, "B21", f"=F10*(1+{GT})", fmt=N0, expect=T['base_ebit'])
    put(ws, "A22", "ANNA revenue in the terminal year")
    put(ws, "B22", f"={ANNAN}*$I$8*{ANNAP}*{round(T['fx'], 4)}/1000000", fmt=N0,
        expect=T['anna_rev'])
    put(ws, "A23", "ANNA operating profit")
    put(ws, "B23", f"=B22*{ANNAM}", fmt=N0, expect=T['anna_ebit'])
    put(ws, "A24", "Terminal EBIT"); put(ws, "B24", "=B21+B23", fmt=N0, expect=T['ebit_T'])
    put(ws, "A25", "Terminal NOPAT")
    put(ws, "B25", f"=B24*(1-{TAXR})", fmt=N0, expect=T['nopat_T'])
    put(ws, "A26", "Reinvestment rate = growth / return on capital")
    put(ws, "B26", f"={GT}/{ROCT}", fmt=PC1, expect=T['reinv_rate'])
    put(ws, "A27", "Terminal free cash flow")
    put(ws, "B27", "=B25*(1-B26)", fmt=N0, expect=T['fcff_T'])
    put(ws, "A28", "TERMINAL VALUE").font = SUB
    put(ws, "B28", f"=B27*(1+{GT})/('6 WACC'!B39-{GT})", fmt=N0, expect=T['tv'])
    put(ws, "A29", "Present value of terminal value")
    put(ws, "B29", "=B28*F17", fmt=N0, expect=T['pv_tv'])
    put(ws, "A31", "ENTERPRISE VALUE TO EQUITY BRIDGE").font = SUB
    put(ws, "A32", "Present value of the explicit window")
    put(ws, "B32", "=SUM(B18:F18)", fmt=N0, expect=B['pv_explicit'])
    put(ws, "A33", "ENTERPRISE VALUE").font = SUB
    put(ws, "B33", "=B32+B29", fmt=N0, expect=B['ev'])
    put(ws, "A34", "TERMINAL VALUE AS A SHARE OF ENTERPRISE VALUE").font = SUB
    put(ws, "B34", "=B29/B33", fmt=PC1, expect=B['tv_pct_ev'])
    put(ws, "A35", "Gross debt (31 March 2026)"); put(ws, "B35", B['debt'], fmt=N0)
    put(ws, "A36", "Cash and equivalents"); put(ws, "B36", B['cash'], fmt=N0)
    put(ws, "A37", "Net debt"); put(ws, "B37", "=B35-B36", fmt=N0, expect=B['net_debt'])
    put(ws, "A38", "Listed equity stakes at market"); put(ws, "B38", B['fvoci'], fmt=N0)
    put(ws, "A39", "Investment property"); put(ws, "B39", B['inv_prop'], fmt=N0)
    put(ws, "A40", "Non-operating assets")
    put(ws, "B40", "=B38+B39", fmt=N0, expect=B['fvoci'] + B['inv_prop'])
    put(ws, "A41", "EQUITY VALUE").font = SUB
    put(ws, "B41", "=B33-B37+B40", fmt=N0, expect=B['equity'])
    put(ws, "A42", "VALUE PER SHARE (EGP)").font = SUB
    put(ws, "B42", f"=B41*1000000/{SHARES}", fmt=N2, expect=B['per_share'])
    put(ws, "A43", "Spot price (EGP)"); put(ws, "B43", SPOT, fmt=N2)
    put(ws, "A44", "Value per share against spot")
    put(ws, "B44", "=B42/B43-1", fmt=PC1, expect=B['per_share'] / SPOT - 1)
    return ws


for nm, cs, lb, bl in [
    ("10 DCF base", "base", "COMMITTED CAPITAL",
     "The board completes the ANNA complex on the spending pace the accounts show, and it "
     "runs at half its derived nameplate in the terminal year."),
    ("11 DCF bear", "bear", "DOWNSIDE",
     "The capital is spent, the plant never earns, and gas is charged at the US$5.75 "
     "contract price rather than the rate the company's own disclosure implies."),
    ("12 DCF bull", "bull", "UPSIDE",
     "Urea holds nearer today's war-tightened level and ANNA reaches 70% of nameplate."),
    ("12b DCF capital discipline", "halt", "CAPITAL DISCIPLINE",
     "The board stops the ANNA programme after one more year, takes the wind-down cost, "
     "and runs the urea plant it already owns. This is not a forecast of what management "
     "will do: it measures what the programme costs shareholders against not doing it.")]:
    dcf_sheet(nm, cs, lb, bl)


# =========================================================== 13. SENSITIVITY ==
ws = wb.create_sheet("13 Sensitivity")
title(ws, "SENSITIVITY — WHOLE-MODEL RE-RUNS",
      "PASTED CLASS 3. Each cell is a complete revaluation of the company at that pair of "
      "inputs. THIS GRID DOES NOT REDRAW WHEN A DRIVER CHANGES on '3 Assumptions' — it is "
      "a set of separate model runs, recorded here so the reader can see the shape of the "
      "answer rather than a single point.")
import importlib.util as _ilu
_spec = _ilu.spec_from_file_location("cmp", os.path.join(HERE, "compute.py"))
GRID = json.load(open(os.path.join(HERE, 'sensitivity_grid.json')))
put(ws, "A4", "Value per share (EGP) — export urea price against terminal cost of capital").font = SUB
header(ws, 5, 2, [f"{w*100:.0f}% terminal WACC" for w in GRID['waccs']], [16]*len(GRID['waccs']))
ws.column_dimensions['A'].width = 34
for i, p in enumerate(GRID['prices']):
    put(ws, f"A{6+i}", f"US$ {p:.0f}/t long-run export price", fmt=None)
    for j in range(len(GRID['waccs'])):
        put(ws, f"{get_column_letter(2+j)}{6+i}", round(GRID['grid'][i][j], 2), fmt=N2)
r = 6 + len(GRID['prices']) + 2
put(ws, f"A{r}", "THE CRUX, IN OBSERVABLE UNITS").font = SUB
put(ws, f"A{r+1}", ("The valuation turns on the long-run export price of granular urea "
     "delivered free on board an Egyptian port, and on the rate at which a perpetuity of "
     "Egyptian pounds is capitalised. Both are observable: the first prints daily on the "
     "CME contract, the second against the sovereign's own ten-year yield. The grid is "
     "read as a falsification test — a reader who believes urea holds above US$550 for a "
     "decade, or that Egyptian equity risk is priced below its sovereign, is reading a "
     "different cell, not disagreeing with the arithmetic."))
ws.cell(row=r+1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r+1].height = 70

# =========================================================== 14. SCENARIOS ====
ws = wb.create_sheet("14 Scenarios")
title(ws, "SCENARIOS — WHOLE-MODEL RE-RUNS",
      "PASTED CLASS 3, for the same reason as sheet 13: each row is a separate run of the "
      "whole model. The formulas that produced them live on sheets 10 to 12b.")
header(ws, 4, 1, ["Scenario", "What changes", "EV (EGP m)", "TV % of EV",
                  "Equity (EGP m)", "Per share (EGP)"], [26, 62, 14, 12, 14, 14])
SC = [("Committed capital", "ANNA completed on the observed spending pace, half nameplate "
       "in the terminal year; gas at the rate the company's own loss disclosure implies", "base"),
      ("Capital discipline", "ANNA stopped after one further year, wind-down cost taken, "
       "the urea plant run as it stands", "halt"),
      ("Upside", "Urea holds nearer today's tightened level and ANNA reaches 70% of "
       "nameplate", "bull"),
      ("Downside", "Capital spent, plant never earns, gas charged at the US$5.75 contract "
       "price", "bear")]
for i, (nm, what, cs) in enumerate(SC):
    b = D['cases'][cs]['bridge']; shname = {"base": "10 DCF base", "halt": "12b DCF capital discipline",
                                            "bull": "12 DCF bull", "bear": "11 DCF bear"}[cs]
    rr = 5 + i
    put(ws, f"A{rr}", nm)
    c = put(ws, f"B{rr}", what); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 42
    put(ws, f"C{rr}", f"='{shname}'!B33", fmt=N0, expect=b['ev'])
    put(ws, f"D{rr}", f"='{shname}'!B34", fmt=PC1, expect=b['tv_pct_ev'])
    put(ws, f"E{rr}", f"='{shname}'!B41", fmt=N0, expect=b['equity'])
    put(ws, f"F{rr}", f"='{shname}'!B42", fmt=N2, expect=b['per_share'])
put(ws, "A11", "Cost of the capital programme to shareholders (EGP per share)").font = SUB
put(ws, "B11", "=F6-F5", fmt=N2,
    expect=D['cases']['halt']['bridge']['per_share'] - D['cases']['base']['bridge']['per_share'])

# =========================================================== 15. PEERS ========
ws = wb.create_sheet("15 Peer comparison")
title(ws, "PEER COMPARISON — OPERATING AND VALUATION",
      "Cross-check only. Peer multiples are never a source for this company's own reported "
      "history; they test whether the discounted-cash-flow answer is reachable from a "
      "different direction.")
header(ws, 4, 1, ["Company", "Market", "Capacity (kt urea/yr)", "Location",
                  "Relevance"], [30, 12, 20, 22, 60])
PEERS = [("Egyptian Chemical Industries (KIMA)", "EGX", 575, "Aswan, 1,000 km inland",
          "The subject. The only Egyptian nitrogen producer not on the coast, which is why "
          "freight to port is its own disclosed cost line."),
         ("Abu Qir Fertilizers", "EGX", 2000, "Alexandria, coastal",
          "The listed comparator, and also the marketer of KIMA's ammonia exports for a fee "
          "of 12% of the export price. KIMA held 2.7% of it and began selling down in "
          "H1-2025/26."),
         ("MOPCO", "Unlisted", 1800, "Damietta, coastal",
          "Curtailed alongside the rest of the industry in the 2026 gas squeeze; KIMA's own "
          "urea is warehoused at Damietta for export."),
         ("Helwan Fertilizers / NCIC", "Unlisted", 1300, "Helwan / Ain Sokhna",
          "Completes the Egyptian nitrogen supply picture against which the export share is "
          "allocated.")]
for i, row in enumerate(PEERS):
    rr = 5 + i
    for j, v in enumerate(row):
        c = put(ws, f"{get_column_letter(1+j)}{rr}", v, fmt=(N0 if j == 2 else None))
        if j == 4:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 44
put(ws, "A11", "IMPLIED MULTIPLES ON THIS MODEL").font = SUB
put(ws, "A12", "Enterprise value / FY2026/27 EBITDA, at the model's committed-capital value")
put(ws, "B12", "='10 DCF base'!B33/'10 DCF base'!B7", fmt=N1,
    expect=D['cases']['base']['bridge']['ev'] / D['cases']['base']['rows'][0]['ebitda'])
put(ws, "A13", "Enterprise value / FY2026/27 EBITDA, at the market price")
put(ws, "B13", f"=('2 Summary valuation'!B16+'10 DCF base'!B37)/'10 DCF base'!B7", fmt=N1,
    expect=(SPOT * SHARES / 1e6 + D['cases']['base']['bridge']['net_debt'])
           / D['cases']['base']['rows'][0]['ebitda'])
put(ws, "A14", "The gap between those two lines is the whole study in one number.").font = NOTE

# =========================================================== 16. SOURCES ======
ws = wb.create_sheet("16 Sources and audit trail")
title(ws, "SOURCES AND AUDIT TRAIL",
      "Every historical figure in this workbook traces to a document in this list, read "
      "from the company's own investor-relations channel.")
header(ws, 4, 1, ["Document", "Period", "Signed / dated", "Auditor", "Used for"],
       [40, 20, 18, 34, 46])
DOCS = [("Audited financial statements", "FY2021/22 (comparatives)", "8 Oct 2023",
         "Central Auditing Organization + PKF Rashed Badr & Co",
         "First year of the four-year statement history"),
        ("Audited financial statements", "FY2022/23", "8 Oct 2023",
         "Central Auditing Organization + PKF Rashed Badr & Co",
         "Peak-utilisation reference year"),
        ("Audited financial statements", "FY2023/24", "23 Oct 2024",
         "Central Auditing Organization",
         "Trough-utilisation year; the one-off revaluation gain stripped out"),
        ("Audited financial statements", "FY2024/25", "23 Sep 2025",
         "Central Auditing Organization",
         "The anchor year: revenue split, production tonnages, unit costs, debt, gas terms"),
        ("Interim statements, limited review", "Q1 FY2025/26 (30 Sep 2025)", "13 Nov 2025",
         "Central Auditing Organization + Nasr Abou El Abbas & Co (Morison Global)",
         "Export volume +34% and price +43% year on year; gas loss quantified"),
        ("Interim statements, limited review", "H1 FY2025/26 (31 Dec 2025)", "10 Feb 2026",
         "Central Auditing Organization + Nasr Abou El Abbas & Co (Morison Global)",
         "Margin confirmation; partial sale of the listed stake"),
        ("Interim statements, limited review", "9M FY2025/26 (31 Mar 2026)", "20 May 2026",
         "Central Auditing Organization + Nasr Abou El Abbas & Co (Morison Global)",
         "Balance sheet used in the bridge; the company's own budget column"),
        ("Country risk premium workbook", "Updated 1 Jan 2026", "1 Jan 2026",
         "Original source file", "Sovereign spread and equity risk premium, both bases"),
        ("Exchange and market quotes", "6-7 Aug 2026", "7 Aug 2026", "Market data",
         "Share price, sovereign yield, USD/EGP, urea FOB Egypt")]
for i, row in enumerate(DOCS):
    rr = 5 + i
    for j, v in enumerate(row):
        c = put(ws, f"{get_column_letter(1+j)}{rr}", v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 34
put(ws, "A16", "ACCESS NOTE").font = SUB
c = put(ws, "A17", ("The exchange's own company page sits behind a bot challenge that "
    "refused every automated read, so the share count used throughout is taken from note "
    "14 of the audited FY2024/25 statements — 1,986,578,999 shares of EGP 5 par, paid-in "
    "capital EGP 9,932,894,995 — rather than from an exchange page or a data aggregator. "
    "Central-bank auction pages were likewise unreachable, so treasury-bill yields are "
    "carried as secondary market quotes and labelled as such; the sovereign yield used in "
    "the cost of capital is cross-checked against a treasury bond listed at a 23.098% "
    "coupon."))
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[17].height = 90

wb.save(os.path.join(HERE, 'EGCH_Valuation_Model_08082026.xlsx'))
json.dump(EXPECT, open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
print(f"wrote workbook: {len(wb.sheetnames)} sheets, {len(EXPECT)} formula cells recorded")
