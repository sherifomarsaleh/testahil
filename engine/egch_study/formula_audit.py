"""EGCH — is every formula in the delivered workbook structurally honest?

recalc.py proves each formula REPRODUCES the model. It cannot see whether the formula is
honest: `=SUM(B17:F17)+7259.375005` reproduces the model perfectly and is still a plug.
qc_checks.py scans BUILDER SOURCE for typed numerals and never opens the workbook, so a
constant emitted into a formula string was invisible to both. An external audit found the
plug on the study's single most valuation-critical switch. This gate closes that hole.

A formula may contain: cell references, operators, functions, and the small set of
structural constants below. Any other embedded number is a plug and fails.
"""
import json, os, re, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
os.chdir(HERE)
XLSX = 'EGCH_Valuation_Model_08082026.xlsx'

# Structural, not financial: unit conversions, day counts, percentage bases, and the
# small integers that index a model (years, halves, quarters).
ALLOWED = {'0', '1', '2', '3', '4', '5', '6', '10', '12', '100', '365', '1000',
           '1000000', '0.5', '1.5', '2.5'}
NUM = re.compile(r'(?<![A-Za-z0-9_.!$:])(\d+(?:\.\d+)?)(?![A-Za-z0-9_.!(])')

wb = openpyxl.load_workbook(XLSX)
bad, nform = [], 0
for sh in wb.sheetnames:
    for row in wb[sh].iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith('=')):
                continue
            nform += 1
            for m in NUM.finditer(v):
                tok = m.group(1)
                if tok in ALLOWED:
                    continue
                bad.append((sh, cell.coordinate, tok, v[:78]))

print(f"formula audit: {nform} formulas across {len(wb.sheetnames)} sheets")
print(f"  embedded constants that are not structural: {len(bad)}")
for sh, co, tok, f in bad[:25]:
    print(f"   ! {sh}!{co}  constant {tok}  in  {f}")
if bad:
    sys.exit(f"FAIL: {len(bad)} formula(s) carry an embedded constant. A number inside a "
             f"formula is a plug: it cannot be traced, sourced or moved by a driver.")
print("PASS: every formula is built from cell references and structural constants only")
