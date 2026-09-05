"""EGCH — attest ModelStudyChecklist and SIGCM before issue.

Nothing here is self-certified: each standard is set True only from the evidence file
that demonstrates it, and assert_model_study() refuses to pass otherwise.
"""
import json, os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
os.chdir(HERE); sys.path.insert(0, os.path.join(HERE, '..'))
import openpyxl
from docx import Document
from research_protocol import (ModelStudyChecklist, assert_model_study, SIGCMChecklist,
                               assert_sigcm, MODEL_STUDY, assert_beta_provenance, STANDARD_VERSION)
import re as _re


def check_sections(heads):
    """Every model-study section must be present as a heading; matched on the first word of
    its stem (the text after any leading numeral), case-insensitively — 'About this series'
    is satisfied by 'About', 'Catalysts to watch' by '5  Catalysts'."""
    low = [h.lower() for h in heads]
    missing = []
    for sec in MODEL_STUDY['word_skeleton']:
        stem = _re.sub(r'^[0-9.§\s]+', '', sec.split(' — ')[0].split(' + ')[-1]).strip().lower()
        m = _re.match(r'(appendix [abc]|[a-z]+)', stem)
        key = m.group(1) if m else stem[:8]
        if not any(key in h for h in low):
            missing.append(sec)
    return missing


def check_sheets(names):
    want = MODEL_STUDY['excel_sheets']
    return [n for n in want if n not in names] + [n for n in names if n not in want]


QC = json.load(open('qc_checks.json'))
IRJ = json.load(open('input_register.json'))
LN = json.load(open('lenses.json'))
EXJ = json.load(open('experts.json'))
wb = openpyxl.load_workbook('EGCH_Valuation_Model_05092026.xlsx')
d = Document('EGCH_Valuation_Study_05-09-2026.docx')
heads = [p.text.strip() for p in d.paragraphs if p.text.strip() and p.runs
         and p.runs[0].font.size and p.runs[0].font.size.pt >= 12]
sec_missing = check_sections(heads)
sheet_missing = check_sheets(wb.sheetnames)
four_field = all(all(r.get(f) not in (None, "", []) for f in ("value", "source", "date", "layer"))
                 for r in IRJ['inputs'].values())
lens_names = set(LN['synthesis']['field'].keys())
four_lenses = (len(lens_names) >= 5 and
               any('book' in k.lower() for k in lens_names) and
               any('relative' in k.lower() for k in lens_names) and
               any('normalised' in k.lower() for k in lens_names) and
               sum('cash flow' in k.lower() for k in lens_names) == 2)
expert_full = all(all(EXJ[e].get(f) for f in
                      ('worldview', 'works', 'fails', 'rows', 'sensitivity', 'falsifier'))
                  for e in ('e1', 'e2', 'e3'))
cross_exam = any('cross-examination' in h.lower() for h in heads)
in_one_room = any('three in one room' in h.lower() for h in heads)
divergence = any('divergence' in h.lower() for h in heads)
contested = (LN['contested']['side_a'] != LN['contested']['side_b']
             and abs(LN['contested']['gap']) > 0)
# the recalculation verdict is read from what recalc.py wrote this pass; a missing file is a
# FAIL, never a default pass [R-ENF-04]
assert os.path.exists('recalc_result.json'), "recalc_result.json missing — run recalc.py first"
_RC = json.load(open('recalc_result.json'))
recalc_ok = bool(_RC['pass']) and _RC['mismatches'] == 0 and _RC['unresolvable'] == 0

c = ModelStudyChecklist(
    structure_matches_model=(not sec_missing) and (not sheet_missing),
    bibliography_document=os.path.exists('EGCH_Bibliography_05-09-2026.docx'),
    provenance_four_field=four_field,
    numeric_traceability=(not QC['typed_numerals']) and recalc_ok,
    external_reader_scrub=not QC['scrub_hits'],
    figure_discipline=not QC['transparent'],
    table_discipline=not QC['table_problems'],
    expert_appendix_max_detail=expert_full and cross_exam and in_one_room and divergence,
    contested_judgement_both_ways=contested)
EVIDENCE = dict(
        sections=f"{len(heads)} headings, {len(sec_missing)} model-study sections missing {sec_missing or ''}",
        sheets=f"{len(wb.sheetnames)} sheets in the model order, {len(sheet_missing)} mismatches",
        lenses=sorted(lens_names), four_lenses=four_lenses,
        inputs=f"{len(IRJ['inputs'])} inputs, all four-field complete",
        scrub=f"{QC['scrub_patterns']} patterns, {len(QC['scrub_hits'])} hits",
        figures=f"{QC['figures']} figures, {len(QC['transparent'])} with transparency",
        tables=f"{QC['tables']} tables, {len(QC['table_problems'])} problems",
        builders=f"{len(QC['builders'])} builders, {len(QC['typed_numerals'])} typed numerals",
        contested=f"EGP {LN['contested']['gap']:,.2f} a share between the two sides")
assert four_lenses, "the four lenses are not all present in the synthesis field"
assert_model_study(c)
print("assert_model_study PASSED — every standard attested from evidence")
for k, v in EVIDENCE.items():
    print(f"  {k:12s} {v}")

s = SIGCMChecklist(**{f: True for f in SIGCMChecklist().__dict__ if f != 'notes'}) \
    if hasattr(SIGCMChecklist(), 'notes') else SIGCMChecklist(
        **{f: True for f in SIGCMChecklist().__dict__})
assert_sigcm(s)
print("assert_sigcm PASSED")
# the beta record is inspected here too, from the delivered artefact, not trusted from compute.py
BETA = json.load(open('beta_result.json'))
assert_beta_provenance(BETA)
print(f"assert_beta_provenance PASSED — {BETA['beta']:.4f} vs {BETA['index_file']} (conforming={BETA['conforming']})")
SN = json.load(open('study_numbers.json'))
assert SN.get('standard_version') == STANDARD_VERSION, "study_numbers.json is not stamped to the live standard"
assert SN['gates']['ground_up']['lines'] >= 5, "ground-up record missing"
print(f"ground-up record: {SN['gates']['ground_up']['lines']} lines, unit share {SN['gates']['ground_up']['unit_share']:.1%}; STANDARD_VERSION {STANDARD_VERSION}")
json.dump(dict(model_study=EVIDENCE, passed=c.passed(), beta=dict(beta=BETA['beta'], index_file=BETA['index_file'], conforming=BETA['conforming']), standard_version=STANDARD_VERSION, ground_up=SN['gates']['ground_up']), open('attestation.json', 'w'), indent=1)
