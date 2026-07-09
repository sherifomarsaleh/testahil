"""Part 2: Segments · DCF · Income Statement · Balance Sheet · Cash Flow (all formula-linked)."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook('GBCO_Valuation_Model_08072026_public.xlsx')
A = json.load(open('_asm_rows.json'))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
NUM = '#,##0.0;(#,##0.0);"-"'; NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'; PX = '0.00'

def sheet(name):
    ws = wb.create_sheet(name); ws.title = name; return ws
def title(ws, text, sub=None, width=10):
    ws['A1'] = text; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, width+1): ws.cell(row=1, column=c).fill = FILL_T
    if sub: ws['A2'] = sub; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 40
    for c in range(2, width+1): ws.column_dimensions[get_column_letter(c)].width = 11.5
def put(ws, addr, v, font=BLACK, fmt=NUM, bold=False, fill=None):
    c = ws[addr]; c.value = v
    c.font = Font(color=(font.color if font else '000000'), bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
def ac(label, col):  # Assumptions driver cell for forecast column j (col letter on that sheet)
    return f"Assumptions!${col}${A[label]}"

YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
# columns: B,C,D historical · E..I forecast on Segments/IS/BS ; Assumptions drivers live in C..G
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']

# ================= SEGMENTS ==================================================
ws = sheet('Segments')
title(ws, 'Segment view — the units × ASP build', 'Volumes and ASP per line of business, FY23–FY30E. Drivers disclosed in company releases; forecast links to Assumptions.', 10)
hdrs = [''] + YH + YF
for j, h in enumerate(hdrs):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
SR = {}
def srow(r, label, hist, ffml=None, fmt=NUM, font_hist=BLUE):
    SR[label] = r
    put(ws, f'A{r}', label)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{r}', v, font_hist, fmt)
    if ffml:
        for j, col in enumerate(FCOLS):
            f = ffml(j, col)
            if f is not None: put(ws, f'{col}{r}', f, BLACK, fmt)
    return r + 1
r = 6
r = srow(r, 'PC volume (units)', [26994, 42043, 56548],
         lambda j, c: f"={get_column_letter(ord(c[0])-65)}{SR['PC volume (units)']}*(1+{ac('PC volume growth', ACOLS[j])})", NUM0)
r = srow(r, 'PC volume (units)', [26994, 42043, 56548],
         lambda j, c: f"={chr(ord(c)-1)}{SR['PC volume (units)']}*(1+{ac('PC volume growth', ACOLS[j])})", NUM0)
r = srow(r, 'PC revenue', [16544.3, 36533.4, 52827.3],
         lambda j, c: f"={c}{SR['PC volume (units)']}*{c}{r}", NUM0)  # temp, fixed after ASP row
PC_REV = SR['PC revenue']
r = srow(r, 'PC ASP (EGP mn/unit)',
         [f'=B{PC_REV}/B{PC_REV-1}', f'=C{PC_REV}/C{PC_REV-1}', f'=D{PC_REV}/D{PC_REV-1}'],
         lambda j, c: f"={chr(ord(c)-1)}{r}*(1+{ac('PC ASP growth', ACOLS[j])})", PX, BLACK)
ASP = SR['PC ASP (EGP mn/unit)']
for j, c in enumerate(FCOLS):  # rewire PC revenue forecast = vol × ASP
    put(ws, f'{c}{PC_REV}', f"={c}{PC_REV-1}*{c}{ASP}", BLACK, NUM0)
r = srow(r, 'CV&CE volume (units)', [2273, 2096, 3404],
         lambda j, c: f"={chr(ord(c)-1)}{SR['CV&CE volume (units)']}*(1+{ac('CV&CE volume growth', ACOLS[j])})", NUM0)
r = srow(r, 'CV&CE revenue', [2323.0, 3984.5, 5956.8],
         lambda j, c: f"={chr(ord(c)-1)}{SR['CV&CE revenue']}*(1+{ac('CV&CE volume growth', ACOLS[j])})*(1+{ac('CV&CE ASP growth', ACOLS[j])})", NUM0)
r = srow(r, 'Light-Mobility volume (units)', [13610, 20189, 33906],
         lambda j, c: f"={chr(ord(c)-1)}{SR['Light-Mobility volume (units)']}*(1+{ac('Light-Mobility volume growth', ACOLS[j])})", NUM0)
r = srow(r, 'Light-Mobility revenue', [854.2, 1378.2, 2203.8],
         lambda j, c: f"={chr(ord(c)-1)}{SR['Light-Mobility revenue']}*(1+{ac('Light-Mobility volume growth', ACOLS[j])})*(1+{ac('Light-Mobility ASP growth', ACOLS[j])})", NUM0)
r = srow(r, 'Trading revenue (tires + parts)', [2506.8, 3815.5, 4242.8],
         lambda j, c: f"={chr(ord(c)-1)}{SR['Trading revenue (tires + parts)']}*(1+{ac('Trading revenue growth', ACOLS[j])})", NUM0)
r = srow(r, 'Other Auto / after-sales & adj.', [7625.2, 1353.4, 1127.6], lambda j, c: '=0', NUM0)
r = srow(r, 'GB Auto total revenue', [23854.0, 47065.0, 66358.3],
         lambda j, c: f"=SUM({c}{SR['PC revenue']},{c}{SR['CV&CE revenue']},{c}{SR['Light-Mobility revenue']},{c}{SR['Trading revenue (tires + parts)']},{c}{SR['Other Auto / after-sales & adj.']})", NUM0)
AUTOR = SR['GB Auto total revenue']
r = srow(r, 'GB Capital revenue', [4950.9, 7383.6, 14743.0],
         lambda j, c: f"={chr(ord(c)-1)}{SR['GB Capital revenue']}*(1+{ac('GB Capital revenue growth', ACOLS[j])})", NUM0)
CAPR = SR['GB Capital revenue']
r = srow(r, 'Intercompany eliminations', [-487.7, -479.0, -871.5],
         lambda j, c: f"=-({c}{AUTOR}+{c}{CAPR})*{ac('Intercompany eliminations (% of gross revenue)', ACOLS[j])}", NUM0)
ELIM = SR['Intercompany eliminations']
r = srow(r, 'Group revenue', [28317.2, 53969.5, 80229.8],
         lambda j, c: f"={c}{AUTOR}+{c}{CAPR}+{c}{ELIM}", NUM0, BLACK)
for col in 'BCD':
    ws[f'{col}{SR["Group revenue"]}'] = f'={col}{AUTOR}+{col}{CAPR}+{col}{ELIM}'
    ws[f'{col}{SR["Group revenue"]}'].font = BLACK
GRPR = SR['Group revenue']
r += 1
r = srow(r, 'Auto gross profit', [5813.1, 9057.4, 9837.1],
         lambda j, c: f"={c}{AUTOR}*{ac('Auto gross margin', ACOLS[j])}", NUM0)
AGP = SR['Auto gross profit']
r = srow(r, 'Auto gross margin', [f'=B{AGP}/B{AUTOR}', f'=C{AGP}/C{AUTOR}', f'=D{AGP}/D{AUTOR}'],
         lambda j, c: f"={c}{AGP}/{c}{AUTOR}", PCT, BLACK)
r = srow(r, 'Auto EBIT (operating profit)', [3460.9, 5564.9, 5830.9],
         lambda j, c: f"={c}{AGP}-{c}{AUTOR}*{ac('Auto GS&A (% of revenue)', ACOLS[j])}+{c}{AUTOR}*{ac('Auto other operating income (% rev)', ACOLS[j])}+{c}{AUTOR}*{ac('Auto provisions (% rev)', ACOLS[j])}", NUM0)
AEBIT = SR['Auto EBIT (operating profit)']
r = srow(r, 'Auto D&A', [374.3, 525.7, 683.3],
         lambda j, c: f"={c}{AUTOR}*{ac('Auto D&A (% of revenue)', ACOLS[j])}", NUM0)
ADNA = SR['Auto D&A']
r = srow(r, 'Auto EBITDA', [3794.6, 5880.5, 6363.3],
         lambda j, c: f"={c}{AEBIT}+{c}{ADNA}", NUM0)
AEBITDA = SR['Auto EBITDA']
r = srow(r, 'Auto EBITDA margin', [f'=B{AEBITDA}/B{AUTOR}', f'=C{AEBITDA}/C{AUTOR}', f'=D{AEBITDA}/D{AUTOR}'],
         lambda j, c: f"={c}{AEBITDA}/{c}{AUTOR}", PCT, BLACK)
r = srow(r, 'GB Capital operating profit', [243.7, 380.1, 788.5], lambda j, c: None, NUM0)
r = srow(r, 'GB Capital net profit (after NCI)', [1207.6, 1091.5, 1365.9], lambda j, c: None, NUM0)
r = srow(r, 'GB Capital on-book loan portfolio', [8980.5, 13183.4, 19495.2], lambda j, c: None, NUM0)
put(ws, f'A{r+1}', 'Source: GB Corp 4Q23 / 4Q24 / 4Q25 earnings releases (Tables 1–14). Other Auto = after-sales & regional lines folded into the four LoBs from FY24; forecast conservatively set to zero.', SUB, None)
json.dump(SR, open('_seg_rows.json', 'w'))

# ================= DCF (Auto operating leg) =================================
ws = sheet('DCF')
title(ws, 'DCF — GB Auto operating leg, explicit 5-year FCFF',
      'Revenue → EBITDA → D&A → EBIT → NOPAT → +D&A → −Capex → −ΔWC → FCFF → discount factor → PV. Links to Segments / Balance Sheet / Assumptions.', 8)
for j, y in enumerate(YF):
    put(ws, f'{get_column_letter(2+j)}4', y, BLACK, None, bold=True, fill=FILL_H)
DC = {}
def drow(r, label, fml, fmt=NUM0, bold=False):
    DC[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j in range(5):
        c = get_column_letter(2+j)
        f = fml(j, c)
        put(ws, f'{c}{r}', f, GREEN if 'Segments!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = drow(r, 'Auto revenue', lambda j, c: f"=Segments!{FCOLS[j]}{AUTOR}")
r = drow(r, 'EBITDA', lambda j, c: f"=Segments!{FCOLS[j]}{AEBITDA}")
r = drow(r, 'D&A', lambda j, c: f"=-Segments!{FCOLS[j]}{ADNA}")
r = drow(r, 'EBIT', lambda j, c: f"=Segments!{FCOLS[j]}{AEBIT}")
r = drow(r, 'NOPAT = EBIT × (1 − tax)', lambda j, c: f"={c}{DC['EBIT']}*(1-Assumptions!$B$7)")
r = drow(r, '+ D&A', lambda j, c: f"=Segments!{FCOLS[j]}{ADNA}")
r = drow(r, '− Capex', lambda j, c: f"=-{ac('Auto capex (EGP mn)', ACOLS[j])}")
r = drow(r, '− Increase in net working capital', lambda j, c: f"='Balance Sheet'!{FCOLS[j]}40*-1")  # placeholder row 40 fixed later
DWC = DC['− Increase in net working capital']
r = drow(r, 'FCFF', lambda j, c: f"=SUM({c}{DC['NOPAT = EBIT × (1 − tax)']}:{c}{DC['− Increase in net working capital']})", bold=True)
r = drow(r, 'Discount factor', lambda j, c: f"=1/(1+Assumptions!$B$16)^{j+1}", '0.000')
r = drow(r, 'PV of FCFF', lambda j, c: f"={c}{DC['FCFF']}*{c}{DC['Discount factor']}", bold=True)
r += 1
def dline(r, label, fml, fmt=NUM0, bold=False):
    put(ws, f'A{r}', label, BLACK, None, bold=bold); put(ws, f'B{r}', fml, BLACK, fmt, bold=bold); return r + 1
r = dline(r, 'Σ PV of explicit FCFF (FY26–30E)', f"=SUM(B{DC['PV of FCFF']}:F{DC['PV of FCFF']})", bold=True); SPV = r-1
r = dline(r, 'Terminal value (Gordon)', f"=F{DC['FCFF']}*(1+Assumptions!$B$17)/(Assumptions!$B$16-Assumptions!$B$17)"); TVR = r-1
r = dline(r, 'PV of terminal value', f"=B{TVR}*F{DC['Discount factor']}"); PVT = r-1
r = dline(r, 'Enterprise value — Auto leg', f"=B{SPV}+B{PVT}", bold=True); EVR = r-1
r = dline(r, '% terminal of EV (device A-7)', f"=B{PVT}/B{EVR}", PCT, bold=True)
r = dline(r, 'less: Auto net debt', f"=-Assumptions!$B$19")
r = dline(r, 'less: Auto non-controlling interests', f"=-Assumptions!$B$20")
r = dline(r, 'Auto equity value', f"=B{EVR}-Assumptions!$B$19-Assumptions!$B$20", bold=True); AEQ = r-1
r += 1
r = dline(r, 'EV sensitivity per +1pp Auto GPM (helper for Sensitivity)',
          f"=0.01*(1-Assumptions!$B$7)*(SUMPRODUCT(B{DC['Auto revenue']}:F{DC['Auto revenue']},B{DC['Discount factor']}:F{DC['Discount factor']})"
          f"+F{DC['Auto revenue']}*(1+Assumptions!$B$17)/(Assumptions!$B$16-Assumptions!$B$17)*F{DC['Discount factor']})")
EVPP = r-1
put(ws, f'A{r+1}', 'WACC is built on the Assumptions sheet as Ke = rf + β × ERP (house rule §3.5-G) blended with after-tax Kd.', SUB, None)
json.dump(dict(DC=DC, SPV=SPV, TVR=TVR, PVT=PVT, EVR=EVR, AEQ=AEQ, EVPP=EVPP), open('_dcf_rows.json', 'w'))
wb.save('GBCO_Valuation_Model_08072026_public.xlsx')
print('part2a ok — segments+dcf; AUTOR', AUTOR, 'AEBITDA', AEBITDA)
