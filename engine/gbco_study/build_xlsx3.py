"""Part 3: Income Statement · Balance Sheet · Cash Flow (clean-surplus construction)."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook('GBCO_Valuation_Model_08072026_public.xlsx')
A = json.load(open('_asm_rows.json')); SR = json.load(open('_seg_rows.json'))
DCJ = json.load(open('_dcf_rows.json'))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'
YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']
AUTOR, CAPR, GRPR = SR['GB Auto total revenue'], SR['GB Capital revenue'], SR['Group revenue']
AGP, AEBIT, ADNA = SR['Auto gross profit'], SR['Auto EBIT (operating profit)'], SR['Auto D&A']

def sheet(n): ws = wb.create_sheet(n); ws.title = n; return ws
def title(ws, t, s=None, w=10):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w+1): ws.cell(row=1, column=c).fill = FILL_T
    if s: ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 42
    for c in range(2, w+1): ws.column_dimensions[get_column_letter(c)].width = 11.5
def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None):
    c = ws[ad]; c.value = v; c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
def ac(label, j): return f"Assumptions!${ACOLS[j]}${A[label]}"

# ================= INCOME STATEMENT =========================================
ws = sheet('Income Statement')
title(ws, 'Income statement (EGP mn, consolidated)',
      'FY23–FY25 disclosed (blue); FY26E–FY30E formulas linking to Segments and Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
IS = {}
def irow(r, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    IS[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{r}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{r}', f, GREEN if 'Segments!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = irow(r, 'GB Auto revenue', [23854.0, 47065.0, 66358.3], lambda j, c: f"=Segments!{c}{AUTOR}")
r = irow(r, 'GB Capital revenue', [4950.9, 7383.6, 14743.0], lambda j, c: f"=Segments!{c}{CAPR}")
r = irow(r, 'Intercompany eliminations', [-487.7, -479.0, -871.5], lambda j, c: f"=Segments!{c}{SR['Intercompany eliminations']}")
r = irow(r, 'Total revenue', [28317.2, 53969.5, 80229.8], lambda j, c: f"=SUM({c}{IS['GB Auto revenue']}:{c}{IS['Intercompany eliminations']})", bold=True)
REV = IS['Total revenue']
r = irow(r, 'Gross profit', [6884.3, 10514.6, 12431.1],
         lambda j, c: f"=Segments!{c}{AGP}+Segments!{c}{CAPR}*{ac('GB Capital gross margin', j)}+Segments!{c}{SR['Intercompany eliminations']}*0.2")
GP = IS['Gross profit']
r = irow(r, 'Gross profit margin', [f'=B{GP}/B{REV}', f'=C{GP}/C{REV}', f'=D{GP}/D{REV}'],
         lambda j, c: f"={c}{GP}/{c}{REV}", PCT, BLACK)
r = irow(r, 'Selling, marketing & administration', [-3430.8, -4843.6, -6547.3],
         lambda j, c: f"=-{c}{REV}*{ac('Group opex S&M+Admin (% of group rev)', j)}")
r = irow(r, 'Other income (expenses)', [518.8, 505.6, 912.2],
         lambda j, c: f"={c}{REV}*{ac('Group other income (% of group rev)', j)}")
r = irow(r, 'Provisions (net)', [-268.9, -355.7, -165.1],
         lambda j, c: f"={c}{REV}*{ac('Group provisions (% of group rev)', j)}")
r = irow(r, 'Operating profit', [3703.4, 5820.8, 6631.0],
         lambda j, c: f"=SUM({c}{GP},{c}{IS['Selling, marketing & administration']}:{c}{IS['Provisions (net)']})", bold=True)
OP = IS['Operating profit']
r = irow(r, 'Investment gains from associates', [1066.1, 867.6, 974.8], lambda j, c: f"={ac('Associates income (EGP mn)', j)}")
r = irow(r, 'EBIT', [4769.5, 6688.5, 7605.7], lambda j, c: f"={c}{OP}+{c}{IS['Investment gains from associates']}", bold=True)
EBIT = IS['EBIT']
r = irow(r, 'Group D&A', [820, 1030, 1243],
         lambda j, c: f"=Segments!{c}{ADNA}+{ac('GB Capital D&A (EGP mn)', j)}")
r = irow(r, 'EBITDA (EBIT + D&A)', [f'=B{EBIT}+B{IS["Group D&A"]}', f'=C{EBIT}+C{IS["Group D&A"]}', f'=D{EBIT}+D{IS["Group D&A"]}'],
         lambda j, c: f"={c}{EBIT}+{c}{IS['Group D&A']}", NUM0, BLACK)
r = irow(r, 'Foreign exchange gains (losses)', [-1499.2, -291.6, -37.1], lambda j, c: '=0')
r = irow(r, 'Net finance cost', [-965.8, -2397.8, -3702.1], lambda j, c: f"={ac('Net finance cost (EGP mn)', j)}")
r = irow(r, 'Earnings before tax', [2304.4, 3999.1, 3866.5],
         lambda j, c: f"={c}{EBIT}+{c}{IS['Foreign exchange gains (losses)']}+{c}{IS['Net finance cost']}", bold=True)
EBT = IS['Earnings before tax']
r = irow(r, 'Income taxes', [-493.2, -939.2, -1086.1], lambda j, c: f"=-{c}{EBT}*Assumptions!$B$7")
r = irow(r, 'Net profit before minority interest', [1811.2, 3059.9, 2780.5],
         lambda j, c: f"={c}{EBT}+{c}{IS['Income taxes']}", bold=True)
NPBM = IS['Net profit before minority interest']
r = irow(r, 'Minority interest', [79.5, -131.8, 99.6], lambda j, c: f"={c}{NPBM}*{ac('Minority interest (% of NP before MI)', j)}")
r = irow(r, 'Net profit (attributable)', [1890.8, 2928.1, 2880.0],
         lambda j, c: f"={c}{NPBM}+{c}{IS['Minority interest']}", bold=True)
NP = IS['Net profit (attributable)']
r = irow(r, 'Net profit margin', [f'=B{NP}/B{REV}', f'=C{NP}/C{REV}', f'=D{NP}/D{REV}'],
         lambda j, c: f"={c}{NP}/{c}{REV}", PCT, BLACK)
put(ws, f'A{r+1}', 'Source: GB Corp 4Q23 / 4Q24 / 4Q25 earnings releases, Table 1 & Table 11 (segment IS). FY23 operating profit shown after provisions for cross-year consistency with the FY25 layout.', SUB, None)
json.dump(IS, open('_is_rows.json', 'w'))

# ================= BALANCE SHEET ============================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet (EGP mn, consolidated)',
      'FY23–FY25 grouped from the disclosed segment balance sheets to a house layout (blue). Forecast rolls forward; the check row is zero by clean-surplus construction.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
BS = {}
def brow(r, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    BS[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{r}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{r}', f, GREEN if ('Cash Flow' in str(f) or 'Income' in str(f) or 'Segments' in str(f)) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = brow(r, 'PP&E, intangibles, ROU & inv. property', [6937.4, 10360.6, 13389.1],
         lambda j, c: f"={chr(ord(c)-1)}{BS['PP&E, intangibles, ROU & inv. property']}+{ac('Auto capex (EGP mn)', j)}+{ac('Rental-fleet & other capex (EGP mn)', j)}-'Income Statement'!{c}{IS['Group D&A']}")
r = brow(r, 'Investments in associates (MNT-Halan, Bedaya, Kaf)', [10732.4, 11743.6, 13689.5],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Investments in associates (MNT-Halan, Bedaya, Kaf)']}+{ac('Associates income (EGP mn)', j)}")
r = brow(r, 'GB Capital loan book (on balance sheet)', [7681.4, 11483.0, 17518.3],
         lambda j, c: f"={chr(ord(c)-1)}{BS['GB Capital loan book (on balance sheet)']}*(1+{ac('GB Capital loan-book growth', j)})")
r = brow(r, 'Inventories', [6366.1, 21134.3, 24649.7],
         lambda j, c: f"=Segments!{c}{AUTOR}*{ac('Auto inventory (% of Auto rev)', j)}")
r = brow(r, 'Trade receivables — Auto', [1743.5, 3708.7, 5316.9],
         lambda j, c: f"=Segments!{c}{AUTOR}*{ac('Auto receivables (% of Auto rev)', j)}")
r = brow(r, 'Advances, debtors & other current', [1039.1, 2942.2, 4670.6],
         lambda j, c: f"=Segments!{c}{AUTOR}*{ac('Auto advances & debtors (% rev)', j)}")
r = brow(r, 'Cash & cash equivalents', [4504.2, 7420.9, 9523.6],
         lambda j, c: f"='Cash Flow'!{c}22")
CASH = BS['Cash & cash equivalents']
r = brow(r, 'Other assets (DTA, held-for-sale, misc.)', [3581.4, 3931.9, 2601.3],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Other assets (DTA, held-for-sale, misc.)']}")
r = brow(r, 'TOTAL ASSETS', [42585.4, 72725.2, 91359.0],
         lambda j, c: f"=SUM({c}{BS['PP&E, intangibles, ROU & inv. property']}:{c}{BS['Other assets (DTA, held-for-sale, misc.)']})", NUM0, BLACK, True)
TA = BS['TOTAL ASSETS']
for col in 'BCD':
    ws[f'{col}{TA}'] = f"=SUM({col}{BS['PP&E, intangibles, ROU & inv. property']}:{col}{BS['Other assets (DTA, held-for-sale, misc.)']})"
    ws[f'{col}{TA}'].font = Font(bold=True)
r += 1
r = brow(r, 'Equity attributable to shareholders', [19838.8, 25438.5, 28788.7],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Equity attributable to shareholders']}+'Income Statement'!{c}{NP}-'Cash Flow'!{c}18*-1")
r = brow(r, 'Non-controlling interests', [1363.0, 1978.4, 1801.4],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Non-controlling interests']}-'Income Statement'!{c}{IS['Minority interest']}")
r = brow(r, 'Total equity', [21201.8, 27417.0, 30590.2],
         lambda j, c: f"={c}{BS['Equity attributable to shareholders']}+{c}{BS['Non-controlling interests']}", NUM0, BLACK, True)
r = brow(r, 'Borrowings (loans, overdrafts & bonds)', [12517.7, 22608.7, 38041.4],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Borrowings (loans, overdrafts & bonds)']}+{ac('Increase in borrowings, net (EGP mn)', j)}")
r = brow(r, 'Trade & notes payables', [7398.7, 19717.6, 18436.5],
         lambda j, c: f"=Segments!{c}{AUTOR}*{ac('Auto payables (% of Auto rev)', j)}+2716.3")
r = brow(r, 'Lease obligations', [371.3, 1123.8, 1554.3], lambda j, c: f"={chr(ord(c)-1)}{BS['Lease obligations']}")
r = brow(r, 'Provisions', [347.7, 709.9, 794.9], lambda j, c: f"={chr(ord(c)-1)}{BS['Provisions']}")
r = brow(r, 'Other liabilities', [415.2, 746.2, 1178.2], lambda j, c: f"={chr(ord(c)-1)}{BS['Other liabilities']}")
r = brow(r, 'Deferred tax liabilities', [333.1, 402.0, 763.5], lambda j, c: f"={chr(ord(c)-1)}{BS['Deferred tax liabilities']}")
r = brow(r, 'TOTAL EQUITY & LIABILITIES', [None, None, None],
         lambda j, c: f"=SUM({c}{BS['Total equity']}:{c}{BS['Deferred tax liabilities']})-{c}{BS['Equity attributable to shareholders']}-{c}{BS['Non-controlling interests']}+{c}{BS['Equity attributable to shareholders']}+{c}{BS['Non-controlling interests']}", NUM0, BLACK, True)
TLE = BS['TOTAL EQUITY & LIABILITIES']
for col in list('BCD') + FCOLS:
    ws[f'{col}{TLE}'] = f"={col}{BS['Total equity']}+SUM({col}{BS['Borrowings (loans, overdrafts & bonds)']}:{col}{BS['Deferred tax liabilities']})"
    ws[f'{col}{TLE}'].font = Font(bold=True); ws[f'{col}{TLE}'].number_format = NUM0
r += 1
r = brow(r, 'Balance check (assets − L&E)', [None]*3,
         lambda j, c: f"={c}{TA}-{c}{TLE}", NUM0, BLACK, True)
CHK = BS['Balance check (assets − L&E)']
for col in 'BCD':
    ws[f'{col}{CHK}'] = f"={col}{TA}-{col}{TLE}"; ws[f'{col}{CHK}'].number_format = NUM0
r = brow(r, 'Group net debt (borrowings − cash)', [None]*3,
         lambda j, c: f"={c}{BS['Borrowings (loans, overdrafts & bonds)']}-{c}{CASH}", NUM0, BLACK)
for col in 'BCD':
    ws[f'{col}{BS["Group net debt (borrowings − cash)"]}'] = f"={col}{BS['Borrowings (loans, overdrafts & bonds)']}-{col}{CASH}"
r += 1
r = brow(r, 'Net Auto working capital (inv + rec + adv − pay)', [None]*3,
         lambda j, c: f"={c}{BS['Inventories']}+{c}{BS['Trade receivables — Auto']}+{c}{BS['Advances, debtors & other current']}-({c}{BS['Trade & notes payables']}-2716.3)", NUM0, BLACK)
NWC = BS['Net Auto working capital (inv + rec + adv − pay)']
for col in 'BCD':
    ws[f'{col}{NWC}'] = f"={col}{BS['Inventories']}+{col}{BS['Trade receivables — Auto']}+{col}{BS['Advances, debtors & other current']}-({col}{BS['Trade & notes payables']}-2716.3)"
    ws[f'{col}{NWC}'].number_format = NUM0
r = brow(r, 'Increase in net Auto working capital', [None]*3,
         lambda j, c: f"={c}{NWC}-{chr(ord(c)-1)}{NWC}", NUM0, BLACK)
DNWC = BS['Increase in net Auto working capital']
put(ws, f'A{r+1}', 'Historic mapping note: lines grouped from the disclosed 4Q23/4Q24/4Q25 segment balance sheets (Table 12/13). "GB Capital loan book (on BS)" = notes receivable + GB Capital trade receivables net of eliminations; the disclosed portfolio metric (EGP 19.5bn FY25) differs by provisions and presentation. Payables forecast holds the non-Auto payables block (EGP 2,716mn) flat.', SUB, None)
json.dump(dict(BS=BS, NWC=NWC, DNWC=DNWC, CASH=CASH), open('_bs_rows.json', 'w'))

# fix the DCF ΔWC link now that DNWC row is known
dws = wb['DCF']; DC = DCJ['DC']
for j, c in enumerate(['B', 'C', 'D', 'E', 'F']):
    dws[f'{c}{DC["− Increase in net working capital"]}'] = f"=-'Balance Sheet'!{FCOLS[j]}{DNWC}"
    dws[f'{c}{DC["− Increase in net working capital"]}'].font = GREEN

# ================= CASH FLOW =================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow (EGP mn, forecast)',
      'Derived from the Income Statement and Balance Sheet. Closing cash ties to the Balance Sheet; every balance-sheet movement is captured, so the check row is zero.', 10)
for j, h in enumerate([''] + YF):
    put(ws, f'{get_column_letter(4+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
CF = {}
def crow(r, label, ffml, fmt=NUM0, bold=False):
    CF[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, c in enumerate(FCOLS):
        f = ffml(j, c)
        if f is not None:
            put(ws, f'{c}{r}', f, GREEN if '!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = crow(r, 'Net profit before minority interest', lambda j, c: f"='Income Statement'!{c}{NPBM}")
r = crow(r, '+ D&A', lambda j, c: f"='Income Statement'!{c}{IS['Group D&A']}")
r = crow(r, '− Associates income (non-cash)', lambda j, c: f"=-'Income Statement'!{c}{IS['Investment gains from associates']}")
r = crow(r, '− Increase in net Auto working capital', lambda j, c: f"=-'Balance Sheet'!{c}{DNWC}")
r = crow(r, '− Increase in GB Capital loan book', lambda j, c: f"=-('Balance Sheet'!{c}{BS['GB Capital loan book (on balance sheet)']}-'Balance Sheet'!{chr(ord(c)-1)}{BS['GB Capital loan book (on balance sheet)']})")
r = crow(r, 'Operating cash flow', lambda j, c: f"=SUM({c}{CF['Net profit before minority interest']}:{c}{CF['− Increase in GB Capital loan book']})", bold=True)
r = crow(r, '− Capex (Auto + rental fleet)', lambda j, c: f"=-{ac('Auto capex (EGP mn)', j)}-{ac('Rental-fleet & other capex (EGP mn)', j)}")
r = crow(r, 'Free cash flow', lambda j, c: f"={c}{CF['Operating cash flow']}+{c}{CF['− Capex (Auto + rental fleet)']}", bold=True)
r = crow(r, '+ Increase in borrowings (net)', lambda j, c: f"={ac('Increase in borrowings, net (EGP mn)', j)}")
r = crow(r, '− Dividends paid', lambda j, c: f"=-'Income Statement'!{c}{NP}*{ac('Dividend payout (of attributable NP)', j)}")
DIVR = CF['− Dividends paid']
r = crow(r, 'Net change in cash', lambda j, c: f"={c}{CF['Free cash flow']}+{c}{CF['+ Increase in borrowings (net)']}+{c}{DIVR}", bold=True)
r = crow(r, 'Opening cash', lambda j, c: (f"='Balance Sheet'!D{CASH}" if j == 0 else f"={chr(ord(c)-1)}{CF['Closing cash'] if 'Closing cash' in CF else r+1}"))
OPEN = CF['Opening cash']
r = crow(r, 'Closing cash', lambda j, c: f"={c}{OPEN}+{c}{CF['Net change in cash']}", bold=True)
CLOSE = CF['Closing cash']
for j, c in enumerate(FCOLS[1:], start=1):
    ws[f'{c}{OPEN}'] = f"={chr(ord(c)-1)}{CLOSE}"
# BS cash + equity dividend links now point at the right rows
bws = wb['Balance Sheet']
for j, c in enumerate(FCOLS):
    bws[f'{c}{CASH}'] = f"='Cash Flow'!{c}{CLOSE}"; bws[f'{c}{CASH}'].font = GREEN
    bws[f'{c}{BS["Equity attributable to shareholders"]}'] = (
        f"={chr(ord(c)-1)}{BS['Equity attributable to shareholders']}+'Income Statement'!{c}{NP}+'Cash Flow'!{c}{DIVR}")
    bws[f'{c}{BS["Equity attributable to shareholders"]}'].font = GREEN
put(ws, f'A{r+1}', 'Historical group cash-flow statements are published per segment (GB Auto CF in each release); the forecast is the consolidated clean-surplus build.', SUB, None)
json.dump(dict(CF=CF, CLOSE=CLOSE, DIVR=DIVR), open('_cf_rows.json', 'w'))
wb.save('GBCO_Valuation_Model_08072026_public.xlsx')
print('part3 ok — IS/BS/CF; NP row', NP, 'DNWC', DNWC, 'CLOSE', CLOSE)
