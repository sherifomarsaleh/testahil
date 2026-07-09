"""Part 4: remaining sheets + sheet ordering."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook('GBCO_Valuation_Model_08072026_public.xlsx')
D = json.load(open('study_numbers.json'))
A = json.load(open('_asm_rows.json')); SR = json.load(open('_seg_rows.json'))
IS = json.load(open('_is_rows.json')); BSJ = json.load(open('_bs_rows.json'))
CFJ = json.load(open('_cf_rows.json')); DCJ = json.load(open('_dcf_rows.json'))
BS = BSJ['BS']; CF = CFJ['CF']; DC = DCJ['DC']
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'; PX = '0.00'; MULT = '0.00x'
YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']; FCOLS = ['E', 'F', 'G', 'H', 'I']
YH = ['FY23', 'FY24', 'FY25']

def sheet(n): ws = wb.create_sheet(n); ws.title = n; return ws
def title(ws, t, s=None, w=9):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w+1): ws.cell(row=1, column=c).fill = FILL_T
    if s: ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 44
    for c in range(2, w+1): ws.column_dimensions[get_column_letter(c)].width = 12
def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None):
    c = ws[ad]; c.value = v; c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill

# ================= SOTP ======================================================
ws = sheet('SOTP')
title(ws, 'Sum-of-the-parts — split the legs (primary lens, §3.5-F5)',
      'Auto = FCFF DCF · captive lender = adjusted book × multiple · fintech associate = carrying value ≈ Jun-26 round. Links to DCF / Assumptions.', 7)
rows = [
 ('Leg', 'Basis', 'Value (EGP mn)', True),
 ('GB Auto operating leg', 'FCFF DCF (§1.2): EV − net debt − NCI', f"=DCF!B{DCJ['AEQ']}", False),
 ('GB Capital lending leg', 'Adjusted operating book 9,500 × 1.0×', "=Assumptions!B21*Assumptions!B22", False),
 ('Associate — MNT-Halan (41.61%, confirmed 9-Jun-2026)', 'USD 1.4bn Jun-26 round × stake × FX, × mark', "=Assumptions!B79*Assumptions!B24", False),
 ('Other associates (Bedaya, Kaf)', 'Residual carrying value × mark', "=Assumptions!B80*Assumptions!B24", False),
 ('Σ Equity value (pre-discount)', '', '=SUM(C6:C9)', True),
 ('  per share (EGP), pre-discount', '', '=C10/Assumptions!B6', True),
 ('less: complexity / conglomerate discount', 'applied', '=-C10*Assumptions!B25', False),
 ('SOTP equity value', '', '=C10+C12', True),
 ('SOTP fair value per share (EGP)', '', '=C13/Assumptions!B6', True),
 ('Upside / (downside) vs spot', '', '=C14/Assumptions!B5-1', True),
 ('memo: MNT-Halan share of SOTP value', 'C8 ÷ Σ legs (discount pro-rata)', '=C8/C10', False),
 ('memo: MNT-Halan per share, post-discount (EGP)', '', '=C8*(1-Assumptions!B25)/Assumptions!B6', False),
]
r = 5
for a, b, c, bold in rows:
    put(ws, f'A{r}', a, BLACK, None, bold=bold)
    put(ws, f'B{r}', b, SUB if b else BLACK, None)
    if c: put(ws, f'C{r}', c, GREEN if isinstance(c, str) and ('DCF!' in c or 'Assumptions!' in c) else BLACK,
              PCT if r in (15, 16) else (PX if r in (11, 14, 17) else NUM0), bold=bold)
    r += 1
ws.column_dimensions['B'].width = 56; ws.column_dimensions['C'].width = 16
r += 1  # blank spacer row (row 18)
put(ws, f'A{r}',
    "The market-implied read is now a genuine puzzle, not a sanity check: mkt cap − lender − associates at the "
    "confirmed 41.61% stake implies a NEGATIVE value for the Auto leg (≈ −EGP 3.6bn) — impossible for a real, "
    "profitable business. MNT-Halan alone is ≈82% of GB Corp's market cap. Either the market discounts this "
    "private mark far more steeply than this study's 10%, or GB Corp is meaningfully mispriced (see the study's §7).",
    SUB, None)
r += 1  # no blank spacer here — the stake-sensitivity header sits immediately below (row 20)
put(ws, f'A{r}', 'MNT-Halan stake sensitivity — now largely resolved: 41.61% is a confirmed, dated figure (9-Jun-2026)', BLACK, None, True, FILL_H)
r += 1
put(ws, f'A{r}', 'GB stake in MNT-Halan', bold=True); put(ws, f'B{r}', 'MNT-Halan value (EGP mn)', bold=True)
put(ws, f'C{r}', 'SOTP fair value/sh (EGP)', bold=True)
r += 1
stakes_notes = [
 (0.20, 'prior placeholder (wrong)'),
 (0.25, None), (0.30, None),
 (0.3545, 'implied by pro-rata defense of the 2024 raise only'),
 (0.40, None),
 (0.4258, 'superseded — was the pre-transaction figure (28-Jul-2024)'),
 (0.4161, 'CONFIRMED current (GB Corp PR, 9-Jun-2026, post Al Ahly round) — base case'),
]
for st, note in stakes_notes:
    put(ws, f'A{r}', st, BLUE, PCT)
    put(ws, f'B{r}', f"=A{r}*Assumptions!$B$76*Assumptions!$B$78", GREEN, NUM0)
    put(ws, f'C{r}', f"=((DCF!$B$25)+B{r}+Assumptions!$B$80+Assumptions!$B$21*Assumptions!$B$22)*(1-Assumptions!$B$25)/Assumptions!$B$6", GREEN, PX)
    if note: put(ws, f'D{r}', note, SUB)
    r += 1
put(ws, f'A{r}',
    "Every 5pp of stake ≈ EGP 3.3/share of SOTP fair value — by far the single largest swing factor in this study; "
    "confirm the FY25 annual-report associates footnote before relying on any point estimate.", SUB, None)

# ================= Relative & Normalized ====================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative & normalized-earnings lenses', 'Links to Assumptions and the Income Statement.', 7)
r = 5
put(ws, f'A{r}', 'Lens', BLACK, None, True, FILL_H); put(ws, f'B{r}', 'Workings', BLACK, None, True, FILL_H); put(ws, f'C{r}', 'EGP/share', BLACK, None, True, FILL_H); r += 1
put(ws, f'A{r}', 'Relative P/E'); put(ws, f'B{r}', 'FY26E group NP (conservative input) × justified P/E, /sh', SUB, None)
put(ws, f'C{r}', '=Assumptions!B27*Assumptions!B28/Assumptions!B6', GREEN, PX); r += 1
put(ws, f'A{r}', '  implied FY26E EPS (EGP)'); put(ws, f'C{r}', '=Assumptions!B27/Assumptions!B6', GREEN, PX); r += 1
put(ws, f'A{r}', '  IS-build FY26E EPS, for reference'); put(ws, f'C{r}', f"='Income Statement'!E{IS['Net profit (attributable)']}/Assumptions!B6", GREEN, PX); r += 1
put(ws, f'A{r}', 'Normalized earnings power'); put(ws, f'B{r}', 'Mid-cycle group PAT × through-cycle P/E, /sh', SUB, None)
put(ws, f'C{r}', '=Assumptions!B29*Assumptions!B30/Assumptions!B6', GREEN, PX); r += 1
put(ws, f'A{r}', '  normalized EPS (EGP)'); put(ws, f'C{r}', '=Assumptions!B29/Assumptions!B6', GREEN, PX); r += 1
put(ws, f'A{r+1}', 'P/B cross-check: spot / FY25 BVPS'); put(ws, f'C{r+1}', f"=Assumptions!B5/('Balance Sheet'!D{BS['Equity attributable to shareholders']}/Assumptions!B6)", BLACK, MULT)
ws.column_dimensions['B'].width = 52

# ================= Summary ===================================================
ws = sheet('Summary')
title(ws, 'Valuation summary — GB Corp (EGX: GBCO)',
      'Four-lens fair value vs spot. Base links live; bear/bull are scenario outputs (study §1.5).', 7)
L = D['lenses']
r = 5
put(ws, f'A{r}', '', BLACK, None); 
for j, h in enumerate(['Bear', 'Base', 'Bull', 'Weight']):
    put(ws, f'{get_column_letter(2+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
lens_rows = [
 ('Sum-of-the-parts (split legs)', L['sotp']['bear'], '=SOTP!C14', L['sotp']['bull'], '=Assumptions!B32'),
 ('Pre-discount NAV (DCF-anchored)', L['prediscount']['bear'], '=SOTP!C11', L['prediscount']['bull'], '=Assumptions!B33'),
 ('Relative multiples', L['relative']['bear'], "='Relative & Normalized'!C6", L['relative']['bull'], '=Assumptions!B34'),
 ('Normalized earnings', L['normalized']['bear'], "='Relative & Normalized'!C9", L['normalized']['bull'], '=Assumptions!B35'),
]
first = r
for nm, be, ba, bu, w in lens_rows:
    put(ws, f'A{r}', nm)
    put(ws, f'B{r}', round(be, 1), BLUE, PX)
    put(ws, f'C{r}', ba, GREEN, PX)
    put(ws, f'D{r}', round(bu, 1), BLUE, PX)
    put(ws, f'E{r}', w, GREEN, PCT)
    r += 1
put(ws, f'A{r}', 'Weighted central', BLACK, None, True)
for col in 'BCD':
    put(ws, f'{col}{r}', f"=SUMPRODUCT({col}{first}:{col}{r-1},$E${first}:$E${r-1})", BLACK, PX, True)
WC = r; r += 2
put(ws, f'A{r}', 'Spot price (EGP)'); put(ws, f'B{r}', '=Assumptions!B5', GREEN, PX); r += 1
put(ws, f'A{r}', 'Upside to central base'); put(ws, f'B{r}', f'=C{WC}/Assumptions!B5-1', BLACK, PCT); r += 2
put(ws, f'A{r}', 'Read: roughly fairly valued — the price already pays for the recovery; the swing is Auto cash conversion (working capital), the Egyptian nominal rate path, and the complexity discount.', SUB, None)

# ================= Fundamental Valuation ====================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — football-field data', 'Bear/base/bull per lens (links to Summary).', 7)
r = 5
for j, h in enumerate(['Lens', 'Bear', 'Base', 'Bull']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for i, nm in enumerate(['Sum-of-the-parts (split legs)', 'Pre-discount NAV (DCF-anchored)', 'Relative multiples', 'Normalized earnings', 'Weighted central']):
    put(ws, f'A{r}', nm, BLACK, None, i == 4)
    for j, col in enumerate('BCD'):
        put(ws, f'{col}{r}', f'=Summary!{col}{first + i}', GREEN, PX, i == 4)
    r += 1
put(ws, f'A{r}', 'Spot'); put(ws, f'B{r}', '=Assumptions!B5', GREEN, PX)

# ================= Summary Financials =======================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials (EGP mn)', 'Every cell links to the statement sheets.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
r = 6
links = [
 ('Revenue', 'Income Statement', IS['Total revenue']),
 ('EBITDA', 'Income Statement', IS['EBITDA (EBIT + D&A)']),
 ('EBIT', 'Income Statement', IS['EBIT']),
 ('Net profit (attributable)', 'Income Statement', IS['Net profit (attributable)']),
 ('Total assets', 'Balance Sheet', BS['TOTAL ASSETS']),
 ('Total equity', 'Balance Sheet', BS['Total equity']),
 ('Group net debt', 'Balance Sheet', BS['Group net debt (borrowings − cash)']),
]
for nm, sh, rr in links:
    put(ws, f'A{r}', nm)
    for j, col in enumerate(['B', 'C', 'D'] + FCOLS):
        put(ws, f'{col}{r}', f"='{sh}'!{col}{rr}", GREEN, NUM0)
    r += 1
put(ws, f'A{r}', 'Free cash flow (forecast)')
for j, col in enumerate(FCOLS):
    put(ws, f'{col}{r}', f"='Cash Flow'!{col}{CF['Free cash flow']}", GREEN, NUM0)

# ================= Monte Carlo ==============================================
ws = sheet('Monte Carlo')
title(ws, 'Monte Carlo — engine outputs (YZ-HAR v2)',
      '50,000 paths · 16 factors · seed 42 · computed by the Testahil MC engine (values, not a sheet simulation). No KVOL — the retired multiplier is replaced by the HAR width.', 8)
pr = D['mc']['prob_read']; q20, q60 = D['mc']['q20'], D['mc']['q60']
r = 5
put(ws, f'A{r}', 'The probability read (T+60)', BLACK, None, True, FILL_G); r += 1
prr = [
 ('P(price above spot)', pr['p_above'], PCT),
 ('P(+10%) vs P(−10%) — odds', f"{pr['p_up10']*100:.0f}% vs {pr['p_dn10']*100:.0f}% · {pr['odds']:.1f}:1", '@'),
 ('Median level (EGP) and % move', f"{pr['median']:.2f} ({pr['med_move']*100:+.1f}%)", '@'),
 ('50% band (25th–75th)', f"{pr['band50'][0]:.1f} – {pr['band50'][1]:.1f}  ({pr['band50_pct'][0]*100:+.0f}% / {pr['band50_pct'][1]*100:+.0f}%)", '@'),
 ('Touch(+10%) / touch(−10%)', f"{pr['touch_up10']*100:.0f}% / {pr['touch_dn10']*100:.0f}%", '@'),
]
for nm, v, fmt in prr:
    put(ws, f'A{r}', nm); put(ws, f'B{r}', v, BLACK, fmt); r += 1
r += 1
put(ws, f'A{r}', 'Percentile map (EGP/share)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for tag, q in [('T+20 sessions', q20), ('T+60 sessions', q60)]:
    put(ws, f'A{r}', tag)
    for j, p in enumerate(['5', '25', '50', '75', '95']):
        put(ws, f'{get_column_letter(2+j)}{r}', round(q[p], 1), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Engine inputs (from Assumptions)', BLACK, None, True, FILL_H); r += 1
for nm, ref in [('Anchor volatility (HAR, annualized)', '=Assumptions!B37'),
                ('Secular drift (daily)', '=Assumptions!B38'),
                ('Net factor drift / quarter', '=Assumptions!B39')]:
    put(ws, f'A{r}', nm); put(ws, f'B{r}', ref, GREEN, PCT); r += 1
r += 1
put(ws, f'A{r}', 'Level-touch ladder (probability of touching by horizon)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Level (EGP)', 'T+20', 'T+60']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for L_, tv in D['mc']['touch'].items():
    put(ws, f'A{r}', float(L_), BLACK, PX)
    put(ws, f'B{r}', tv['t20'], BLACK, PCT); put(ws, f'C{r}', tv['t60'], BLACK, PCT); r += 1

# ================= Sensitivity ==============================================
ws = sheet('Sensitivity')
title(ws, 'SOTP sensitivity — Auto margin × complexity discount',
      'Fair value (EGP/share, post-discount). Live off Assumptions and the DCF helper (EV per +1pp GPM).', 8)
gm = D['sens']['grid_margin']; gd = D['sens']['grid_disc']
put(ws, 'A5', 'GPM shift \\ discount', BLACK, None, True, FILL_H)
for j, dd in enumerate(gd):
    put(ws, f'{get_column_letter(2+j)}5', dd, BLACK, PCT, True, FILL_H)
for i, mm in enumerate(gm):
    rr = 6 + i
    put(ws, f'A{rr}', mm, BLUE, '+0.0%;-0.0%;"base"')
    for j, dd in enumerate(gd):
        col = get_column_letter(2+j)
        put(ws, f'{col}{rr}',
            f"=((DCF!$B${DCJ['EVR']}+$A{rr}*100*DCF!$B${DCJ['EVPP']}-Assumptions!$B$19-Assumptions!$B$20"
            f"+Assumptions!$B$21*Assumptions!$B$22+Assumptions!$B$23*Assumptions!$B$24)*(1-{col}$5))/Assumptions!$B$6",
            BLACK, PX)
put(ws, 'A12', 'Bold-face cells near spot (31.25) show how much Auto-margin recovery the price already carries.', SUB, None)

# ================= Per-Share & Ratios =======================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share & ratios — the standing dashboard (device A-5)', 'Links to statements / Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
r = 6
allc = ['B', 'C', 'D'] + FCOLS
def prow(r, label, fml, fmt=PX, cols=allc):
    put(ws, f'A{r}', label)
    for col in cols:
        put(ws, f'{col}{r}', fml(col), BLACK, fmt)
    return r + 1
NPR = IS['Net profit (attributable)']; REVR = IS['Total revenue']; EBITDAR = IS['EBITDA (EBIT + D&A)']
EQR = BS['Equity attributable to shareholders']; NDR = BS['Group net debt (borrowings − cash)']
r = prow(r, 'EPS (EGP)', lambda c: f"='Income Statement'!{c}{NPR}/Assumptions!$B$6")
r = prow(r, 'DPS (EGP)', lambda c: f"='Income Statement'!{c}{NPR}*{'0.13' if c in 'BCD' else 'Assumptions!' + ['C','D','E','F','G'][allc.index(c)-3] + '$' + str(A['Dividend payout (of attributable NP)'])}/Assumptions!$B$6")
r = prow(r, 'Book value / share (EGP)', lambda c: f"='Balance Sheet'!{c}{EQR}/Assumptions!$B$6")
r = prow(r, 'P/E at spot (×)', lambda c: f"=Assumptions!$B$5/('Income Statement'!{c}{NPR}/Assumptions!$B$6)")
ws[f'A{r-1}'].value = 'P/E at spot (×)'
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'P/B at spot (×)', lambda c: f"=Assumptions!$B$5/('Balance Sheet'!{c}{EQR}/Assumptions!$B$6)")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Dividend yield at spot', lambda c: f"={c}7/Assumptions!$B$5", PCT)
r = prow(r, 'EBITDA margin', lambda c: f"='Income Statement'!{c}{EBITDAR}/'Income Statement'!{c}{REVR}", PCT)
r = prow(r, 'Net margin', lambda c: f"='Income Statement'!{c}{NPR}/'Income Statement'!{c}{REVR}", PCT)
r = prow(r, 'ROAE (attributable)', lambda c: f"='Income Statement'!{c}{NPR}/'Balance Sheet'!{c}{EQR}", PCT)
r = prow(r, 'Group net debt / EBITDA (×)', lambda c: f"='Balance Sheet'!{c}{NDR}/'Income Statement'!{c}{EBITDAR}")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Revenue YoY', lambda c: (f"='Income Statement'!{c}{REVR}/'Income Statement'!{chr(ord(c)-1)}{REVR}-1" if c != 'B' else None), PCT, allc)
put(ws, f'B{r-1}', None, BLACK, None)
r = prow(r, 'EPS YoY', lambda c: (f"=('Income Statement'!{c}{NPR}/'Income Statement'!{chr(ord(c)-1)}{NPR})-1" if c != 'B' else None), PCT, allc)
put(ws, f'B{r-1}', None, BLACK, None)
put(ws, f'A{r+1}', 'Group net-debt/EBITDA blends the lender funding book; the Auto-leg discipline metric is Auto ND/EBITDA (2.39× FY25, 2.14× 1Q26, disclosed).', SUB, None)

# ================= Peer & Sector ============================================
ws = sheet('Peer & Sector')
title(ws, 'Peer set & sector', 'No clean single comparable — GB Corp is an auto operator plus an NBFI plus a fintech associate. Split by leg.', 7)
r = 5
for j, h in enumerate(['Leg', 'Closest peers', 'Typical valuation']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
peers = [
 ('Auto assembly & distribution', 'Al Watania (KSA autos), Saudi Co. for Vehicles, Türk Traktör, Astra Industrial; EGX consumer durables', 'EV/EBITDA ~4–6× (EM distribution); P/E ~8–12×'),
 ('NBFI / leasing & consumer finance', 'EFG Hermes (NBFI arm), Contact Financial (CNFN), Raya Holding finance leg', 'P/B ~1.0–1.8× on mid-teens ROE; P/E ~6–10×'),
 ('Fintech associate (MNT-Halan)', 'Fawry (FWRY), e-Finance (EFIH), MNT-Halan private rounds', 'Private marks; EV/Rev 2–5×; last round USD 1.4bn (Jun-26)'),
]
for a, b, c in peers:
    put(ws, f'A{r}', a); put(ws, f'B{r}', b, BLACK, None); put(ws, f'C{r}', c, BLACK, None); r += 1
ws.column_dimensions['B'].width = 58; ws.column_dimensions['C'].width = 46
r += 1
put(ws, f'A{r}', 'Sector context: Egypt PC registrations ~210k in 2025 (+40% y/y), GB share ~21%; CBE easing cycle (19.00% deposit rate) supports affordability and GB Capital margins; BYD entering Egypt; Iraq/Jordan drag from the regional conflict; Sadat CKD facility inaugurated Jun-2026.', SUB, None)
put(ws, f'A{r+1}', 'Analyst context is deliberately not used as a model input (house rule: no rating, no target).', SUB, None)

# ================= sheet order ==============================================
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP', 'Segments',
         'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet', 'Cash Flow',
         'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in order]
wb.save('GBCO_Valuation_Model_08072026_public.xlsx')
print('part4 ok — sheets:', wb.sheetnames)
