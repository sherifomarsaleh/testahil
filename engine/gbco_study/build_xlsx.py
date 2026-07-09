"""GBCO_Valuation_Model_08072026_public.xlsx — 16 sheets mirroring the TMPV canonical model.
Blue = inputs · black = formulas · green = cross-sheet links. All inputs live on Assumptions."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = json.load(open('study_numbers.json'))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
BOLD = Font(bold=True); TITLE = Font(bold=True, size=13, color='F6F1E6')
SUB = Font(size=9, color='6E7B77'); HDR = Font(bold=True, color='1C3A36')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM = '#,##0.0;(#,##0.0);"-"'; NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'
MULT = '0.00x'; PX = '0.00'

wb = Workbook()

def sheet(name):
    ws = wb.create_sheet(name) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = name
    return ws

def title(ws, text, sub=None, width=10):
    ws['A1'] = text; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, width + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if sub:
        ws['A2'] = sub; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 42
    for c in range(2, width + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12.5

def put(ws, addr, val, font=BLACK, fmt=None, bold=False, fill=None):
    c = ws[addr]; c.value = val
    c.font = Font(color=font.color, bold=bold) if font else (BOLD if bold else BLACK)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

# ============ READ FIRST =====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — GB Corp S.A.E. (EGX: GBCO)', width=9)
lines = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the GBCO valuation study. Every blue cell is an input; every',
 'black cell is a formula; green cells link across sheets. All inputs live on the Assumptions sheet — change one',
 '(the complexity discount, the Auto gross margin, a growth rate, the GB Capital multiple) and the whole model reprices.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs shown',
 'as ranges. The preparer is not licensed by any securities regulator and may hold a position in the security.', '',
 'Entity note. GB Corp S.A.E. (formerly GB Auto / Ghabbour Auto; renamed March 2023) is an operating company with a',
 'captive finance arm: GB Auto (passenger cars, CV&CE, trading, light mobility; Egypt · Iraq · Jordan) plus GB Capital',
 '(leasing, factoring, consumer finance, SME lending) and associate stakes in MNT-Halan, Bedaya and Kaf.',
 'House lens: split the legs — Auto = FCFF DCF; captive lender = adjusted book × a return-justified multiple;',
 'the fintech associate = balance-sheet carrying value cross-checked to the June-2026 USD 1.4bn funding round.', '',
 'Currency. EGP million unless stated. Spot EGP 31.25 (7 Jul 2026 close). Historical financials are company disclosure',
 '(FY23–FY25 earnings releases; 1Q26 release 14 May 2026). Some consolidated balance-sheet lines are grouped from the',
 'disclosed segment balance sheets to a house layout — flagged on the Balance Sheet.', '',
 'Sheets: Summary · Fundamental Valuation · Assumptions · SOTP · Segments · Relative & Normalized · DCF ·',
 'Income Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share & Ratios · Peer & Sector.']
for i, ln in enumerate(lines, start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 118

# ============ ASSUMPTIONS (built early so links resolve) =====================
wa = sheet('Assumptions')
title(wa, 'Assumptions — the single input layer', 'All blue cells are inputs. Every other sheet links here.', 9)
r = 4
def hdr(ws_, row, text):
    put(ws_, f'A{row}', text, bold=True, fill=FILL_H); return row + 1
def inp(ws_, row, label, val, fmt=NUM, note=None):
    put(ws_, f'A{row}', label)
    put(ws_, f'B{row}', val, BLUE, fmt)
    if note: put(ws_, f'C{row}', note, SUB)
    return row + 1
r = hdr(wa, r, 'ANCHORS')
r = inp(wa, r, 'Spot price (EGP/share)', 31.25, PX)
r = inp(wa, r, 'Shares outstanding (mn)', 1085.5, NUM0)
r = inp(wa, r, 'Tax rate', 0.28, PCT)
r = hdr(wa, r, 'COST OF CAPITAL — bottom-up, sourced (house rule §3.5-G, rebuilt 09-07-2026; full detail row 82+)')
r = inp(wa, r, 'Risk-free rate (Egypt 10Y local govt bond, 3-Jul-2026)', 0.2255, PCT,
        'SOURCE: investing.com, "Egypt 10-Year Bond Yield Historical Data" — market quote, per '
        "Damodaran's own stated method (use the local bond yield directly when a liquid local market exists).")
r = inp(wa, r, 'Equity beta (see row 86 for why =1.0)', 1.0, '0.00',
        'CORRECTED 09-07-2026: was 0.95 (unsourced "house band" guess). A genuine GBCO-vs-EGX30 regression was '
        'attempted (n=5 annual: beta=-0.15, R2=0.008, unusable; higher-frequency EGX30 data inaccessible via '
        'available tools) and rejected. Beta=1.0 per standing instruction when no reliable regression is '
        'obtainable — see wacc_builder.py.')
r = inp(wa, r, 'Equity risk premium — Egypt, CDS-based (primary; rating-based alt. at row 84)', 0.0941, PCT,
        "Same Damodaran original file (ctryprem.html), Egypt row, CDS column — uses Egypt's actual sovereign CDS "
        'spread (3.41%), more current than the rating-based figure since CDS prices continuously. CORRECTED '
        '09-07-2026: was 0.07 (flat unsourced house number), then briefly 0.1487 (WRONG — a secondary '
        "source's misquote).")
put(wa, f'A{r}', 'Cost of equity Ke = rf + β × ERP'); put(wa, f'B{r}', '=B9+B10*B11', BLACK, PCT); KE=f'B{r}'
put(wa, f'C{r}', 'Primary Ke, used in the base-case WACC below.', SUB); r += 1
r = inp(wa, r, 'Pre-tax cost of debt (EGP; ~100% of debt is EGP-denominated, row 88)', 0.207, PCT,
        'CORRECTED 09-07-2026: was 0.21 (flat unsourced guess). SOURCE: CBE weighted-average EGP bank lending '
        "rate, <12mo tenor, Feb-2026 (CEIC/TradingEconomics quoting CBE); cross-checked against CBE's own "
        'overnight lending-rate ceiling 20.0% (held since the Apr-2026 policy pause). Currency mix: 5 '
        'separately disclosed GB Corp/GB Capital financing facilities found (Drive Finance EGP5bn syndication, '
        'Ghabbour Egypt EGP1.2bn Sadat facility, GB Lease EGP4.16bn securitization, Drive Finance EGP2.4bn '
        'bond, GB Capital Securitization EGP28.8bn book) are ALL EGP-denominated; zero USD facilities found '
        'in any search — treated as ~100% local currency.')
put(wa, f'A{r}', 'After-tax Kd'); put(wa, f'B{r}', '=B13*(1-B7)', BLACK, PCT); r+=1
r = inp(wa, r, 'Debt weight D/(D+E) — sourced, row 89', '=38041.4/(31.25*1085.5+38041.4)', PCT,
        'CORRECTED 09-07-2026: was 0.35 (assumed, not computed). Market cap = spot(31.25) x shares(1,085.5mn) '
        '= 33,922; total debt = FY25 disclosed consolidated borrowings (38,041.4). D/(D+E) = '
        '38,041.4/(33,922+38,041.4).')
wa[f'B{r-1}'].font = BLACK  # this is a formula, not an input — override inp()'s default blue
put(wa, f'A{r}', 'WACC'); put(wa, f'B{r}', f'=(1-B15)*B12+B15*B14', BLACK, PCT); WACC='Assumptions!$B$16'; r+=1
r = inp(wa, r, 'Terminal growth (nominal EGP)', 0.115, PCT)
TG='Assumptions!$B$17'
r = hdr(wa, r, 'SOTP — the three legs (split-the-legs, §3.5-F5)')
r = inp(wa, r, 'GB Auto net debt (31 Dec 2025)', 15210.0, NUM0, 'disclosed; 1Q26 ND/EBITDA 2.14x')
r = inp(wa, r, 'GB Auto non-controlling interests', 800.4, NUM0)
r = inp(wa, r, 'GB Capital adjusted operating equity', 9500.0, NUM0, 'from disclosed adjusted-ROAE basis (NP 1,366 / 15.1%)')
r = inp(wa, r, 'GB Capital multiple (× adjusted book)', 1.0, MULT, 'return-justified ~1× book')
r = inp(wa, r, 'Associates carrying value (MNT-Halan + Bedaya + Kaf)', '=B79+B80', NUM0,
        'FY25 carrying 13,689.5 — split in the MNT-Halan detail block (rows 75–80)')
wa[f'B{r-1}'].font = BLACK  # this is a formula, not an input — override inp()'s default blue
r = inp(wa, r, 'Associates mark (× carrying)', 1.0, MULT)
r = inp(wa, r, 'Complexity / conglomerate discount', 0.10, PCT, 'sensitized 0–20%')
r = hdr(wa, r, 'RELATIVE & NORMALIZED')
r = inp(wa, r, 'FY26E group net profit for the relative lens', 3300.0, NUM0, 'aligned to the IS build (≈ EGP 3.3bn FY26E)')
r = inp(wa, r, 'Justified P/E (base)', 9.5, MULT)
r = inp(wa, r, 'Mid-cycle group PAT (normalized)', 4200.0, NUM0)
r = inp(wa, r, 'Justified through-cycle P/E', 8.5, MULT)
r = hdr(wa, r, 'SYNTHESIS WEIGHTS')
r = inp(wa, r, 'SOTP weight', 0.40, PCT)
r = inp(wa, r, 'Pre-discount NAV weight', 0.15, PCT)
r = inp(wa, r, 'Relative weight', 0.20, PCT)
r = inp(wa, r, 'Normalized-earnings weight', 0.25, PCT)
r = hdr(wa, r, 'MONTE CARLO (YZ-HAR v2 — no KVOL; engine outputs on the Monte Carlo sheet)')
r = inp(wa, r, 'Anchor volatility (HAR forecast, annualized)', round(D['engine']['anchor_vol'], 4), PCT)
r = inp(wa, r, 'Secular drift (daily log-return, expanding window)', round(D['engine']['drift_daily'], 6), '0.0000%')
r = inp(wa, r, 'Net factor drift per quarter (16-factor stack)', round(D['engine']['factor_drift_q'], 4), PCT)
r = inp(wa, r, 'Paths / seed', '50,000 / 42', '@')
r = hdr(wa, r, 'FORECAST DRIVERS (FY26E–FY30E) — units × ASP build; drivers disclosed')
yrs = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
put(wa, f'A{r}', 'Driver \\ year', bold=True)
for j, y in enumerate(yrs):
    put(wa, f'{get_column_letter(3+j)}{r}', y, bold=True, fill=FILL_H)
r += 1
DRV = {}
def drv(row, label, vals, fmt=PCT):
    put(wa, f'A{row}', label)
    for j, v in enumerate(vals):
        put(wa, f'{get_column_letter(3+j)}{row}', v, BLUE, fmt)
    DRV[label] = row
    return row + 1
r = drv(r, 'PC volume growth', [0.12, 0.14, 0.10, 0.08, 0.06])
r = drv(r, 'PC ASP growth', [0.06, 0.07, 0.07, 0.06, 0.06])
r = drv(r, 'CV&CE volume growth', [0.25, 0.18, 0.12, 0.10, 0.08])
r = drv(r, 'CV&CE ASP growth', [0.05]*5)
r = drv(r, 'Light-Mobility volume growth', [0.30, 0.20, 0.15, 0.12, 0.10])
r = drv(r, 'Light-Mobility ASP growth', [0.05]*5)
r = drv(r, 'Trading revenue growth', [0.18, 0.15, 0.12, 0.10, 0.10])
r = drv(r, 'GB Capital revenue growth', [0.45, 0.30, 0.25, 0.20, 0.18])
r = drv(r, 'GB Capital loan-book growth', [0.35, 0.28, 0.24, 0.20, 0.18])
r = drv(r, 'Auto gross margin', [0.138, 0.142, 0.145, 0.145, 0.145])
r = drv(r, 'Auto GS&A (% of revenue)', [0.073, 0.072, 0.071, 0.070, 0.070])
r = drv(r, 'Auto other operating income (% rev)', [0.012]*5)
r = drv(r, 'Auto provisions (% rev)', [-0.003]*5)
r = drv(r, 'Auto D&A (% of revenue)', [0.011]*5)
r = drv(r, 'Auto capex (EGP mn)', [3000, 2400, 2500, 2600, 2800], NUM0)
r = drv(r, 'Rental-fleet & other capex (EGP mn)', [700, 800, 900, 1000, 1100], NUM0)
r = drv(r, 'GB Capital D&A (EGP mn)', [560, 640, 730, 830, 940], NUM0)
r = drv(r, 'Auto inventory (% of Auto rev)', [0.36, 0.338, 0.32, 0.308, 0.296])
r = drv(r, 'Auto receivables (% of Auto rev)', [0.08]*5)
r = drv(r, 'Auto advances & debtors (% rev)', [0.07, 0.069, 0.0675, 0.066, 0.0645])
r = drv(r, 'Auto payables (% of Auto rev)', [0.245, 0.237, 0.2325, 0.229, 0.2255])
r = drv(r, 'Group opex S&M+Admin (% of group rev)', [0.080, 0.079, 0.078, 0.0775, 0.077])
r = drv(r, 'Group other income (% of group rev)', [0.011]*5)
r = drv(r, 'Group provisions (% of group rev)', [-0.003]*5)
r = drv(r, 'GB Capital gross margin', [0.188, 0.19, 0.192, 0.194, 0.196])
r = drv(r, 'Associates income (EGP mn)', [1250, 1430, 1630, 1830, 2030], NUM0)
r = drv(r, 'Net finance cost (EGP mn)', [-4100, -3800, -3500, -3300, -3100], NUM0)
r = drv(r, 'Minority interest (% of NP before MI)', [-0.02]*5)
r = drv(r, 'Increase in borrowings, net (EGP mn)', [5500, 5200, 5600, 5800, 6100], NUM0)
r = drv(r, 'Dividend payout (of attributable NP)', [0.14, 0.15, 0.16, 0.18, 0.20])
r = drv(r, 'Intercompany eliminations (% of gross revenue)', [0.011]*5)
wa.column_dimensions['C'].width = 11

# ===== MNT-HALAN DETAIL (rows 75-80) — added 09-07-2026 to reflect the confirmed stake =====
r = hdr(wa, 75, 'MNT-HALAN DETAIL (feeds the associates line, row 23)')
r = inp(wa, r, 'MNT-Halan round valuation (USD mn, Jun-26 first close)', 1400.0, NUM0)
r = inp(wa, r, 'GB stake — CONFIRMED current (GB Corp PR, 9-Jun-2026, post Al Ahly Capital round)', 0.4161, PCT,
        "UPDATED 09-07-2026: GB Corp's own press release, 9 June 2026 (\"MNT-Halan ... Closes Capital Increase "
        'Round Led by Al Ahly Capital Holding\"): "GB Corp\'s ownership stake in MNT-Halan will be adjusted to '
        '41.61%, compared to 42.58% prior to the transaction." This is a current, dated, confirmed figure — not '
        "an estimate. Applying it to the round still implies MNT-Halan alone ≈ 82% of GB Corp's market cap; "
        'flagged as a genuine valuation puzzle (steep implied private-mark discount, or undervaluation) rather '
        'than a sourcing gap.')
r = inp(wa, r, 'EGP/USD (study date)', 47.5, PX)
put(wa, f'A{r}', 'MNT-Halan implied value (EGP mn)'); put(wa, f'B{r}', '=B76*B77*B78', BLACK, NUM0); r += 1
r = inp(wa, r, 'Other associates (Bedaya, Kaf) — residual carrying', 390.0, NUM0)

# ===== WACC BUILD — FULL DETAIL & SOURCING (rows 82-90) — reference only, feeds rows 9-17 =====
r = hdr(wa, 82, 'WACC BUILD — FULL DETAIL & SOURCING (feeds rows 9-17 above; this block is reference only)')
put(wa, f'A{r}', 'rf source')
put(wa, f'C{r}', 'investing.com, "Egypt 10-Year Bond Yield Historical Data" — market quote, 3-Jul-2026. Per '
                  "Damodaran's own stated method: use the local bond yield directly when a liquid local market exists.", SUB)
r += 1
r = inp(wa, r, 'ERP — rating-based (Damodaran "standard practice"), alternative to primary', 0.1394, PCT,
        'CORRECTED 09-07-2026: an interim figure of 14.87% (from a secondary source claiming to summarize '
        '"Damodaran\'s July 2026 table") was WRONG — verified against Damodaran\'s own ORIGINAL file '
        '(ctryprem.html), Egypt row, "Last updated January 2026": country risk premium 9.71% + mature-market ERP '
        '4.23% = 13.94%. Logged as Fundamental Driver Ledger S2.')
put(wa, f'A{r}', 'Ke, alternative (rating-based ERP)'); put(wa, f'B{r}', '=B9+B10*B84', BLACK, PCT); r += 1
put(wa, f'A{r}', 'WACC, alternative (rating-based ERP)'); put(wa, f'B{r}', '=(1-B15)*B85+B15*B14', BLACK, PCT); r += 1
put(wa, f'A{r}', 'Beta — why 1.0')
put(wa, f'C{r}', 'CORRECTED 09-07-2026: was 0.95 ("house band" guess). A genuine GBCO-vs-EGX30 regression was '
                  'attempted: n=5 annual observations gave beta=-0.15, R2=0.008 (statistically unusable — '
                  'SE(beta) exceeds the estimate itself). Higher-frequency (monthly) EGX30 data proved '
                  "inaccessible via any available tool (EGX's own site blocks automated access; other sources "
                  'are JS-rendered, no bulk export). Beta=1.0 applied per standing instruction (house rule '
                  '§3.5-G) when no reliable regression is obtainable — see wacc_builder.py / '
                  'RegressionBetaAttempt gate.', SUB)
r += 1
put(wa, f'A{r}', 'Kd source & currency-mix evidence')
put(wa, f'C{r}', 'Pre-tax rate: CBE weighted-average EGP bank lending rate, <12mo tenor, Feb-2026 '
                  "(CEIC/TradingEconomics quoting CBE); cross-checked against CBE's own overnight lending-rate "
                  'ceiling 20.0% (held since the Apr-2026 policy pause). Currency mix: 5 separately disclosed '
                  'GB Corp/GB Capital financing facilities found (Drive Finance EGP5bn syndication, Ghabbour '
                  'Egypt EGP1.2bn Sadat facility, GB Lease EGP4.16bn securitization, Drive Finance EGP2.4bn bond, '
                  'GB Capital Securitization EGP28.8bn book) are ALL EGP-denominated; zero USD facilities found '
                  'in any search — treated as ~100% local currency (no FX tranche blended in).', SUB)
r += 1
put(wa, f'A{r}', 'Weights source')
put(wa, f'C{r}', 'Market cap = spot(31.25) x shares(1,085.5mn) = 33,922; total debt = FY25 disclosed consolidated '
                  'borrowings (38,041.4). Sourced, not assumed (prior draft used a flat 65/35 house default).', SUB)
r += 1
put(wa, f'A{r}', 'Full reference & cache')
put(wa, f'C{r}', 'See Cost_of_Capital_Reference.md (Egypt row) and wacc_builder.py in the project repo for the '
                  'standing method applied to every future study across every market.', SUB)

ROWSA = {k: v for k, v in DRV.items()}
json.dump(ROWSA, open('_asm_rows.json', 'w'))
wb.save('GBCO_Valuation_Model_08072026_public.xlsx')
print('part1 ok; assumptions rows:', len(ROWSA))
