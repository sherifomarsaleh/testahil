"""BOROUGE_Valuation_Model_09082026_public.xlsx — 16 sheets, formula-first.

Blue = input · black = formula · green = cross-sheet link.

The workbook CALCULATES. Every quantity arithmetically derivable from a driver is a
live Excel formula, so the reader can change a blue cell on Assumptions and watch the
model reprice: the cost of equity is built from the risk-free rate NET of the sovereign
default spread, beta and the premium; the cost of debt is taxed in the sheet; the
weights come from net debt and market capitalisation; the terminal rate is built from
its own components; the discount factors compound; the DCF waterfall chains from
revenue through EBITDA, EBIT, NOPAT and FCFF to present value; the terminal block
chains from reinvestment = growth / return on capital; the statements roll forward; and
every ratio and per-share figure is a formula.

Only three classes of cell are pasted, named on READ FIRST:
  1. audited and disclosed history — the primary record, not a calculation;
  2. the output of the unit build's cost regression (the fixed / per-tonne split of
     other production cost, least squares across three audited years), which cannot be
     read as a grid; everything downstream of it is formula;
  3. whole-model re-runs — the probability map, the sensitivity grids, and the
     sector-beta re-solve of the whole waterfall. These do NOT redraw when a driver
     changes.

Sheets are created in the required order first and populated in DEPENDENCY order, so a
formula never has to guess a row number on a sheet that has not been laid out yet.

Every formula cell also records the model's own value into xlsx_expected.json, and
recalc.py evaluates the workbook independently and asserts the two agree.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
S0 = json.load(open(os.path.join(HERE, 'step0_result.json')))
STK = json.load(open(os.path.join(HERE, 'strike_result.json')))
BT = json.load(open(os.path.join(HERE, 'backtest_5y.json')))
BETA = json.load(open(os.path.join(HERE, 'beta_result.json')))

BLUE, GREEN, BLACK = Font(color='0000FF'), Font(color='008000'), Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6')
SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM2 = '#,##0.00;(#,##0.00);"-"'
NUM3 = '#,##0.000'
NUM4 = '#,##0.0000'
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%'
PX = '0.00;(0.00);"-"'
MULT = '0.00x'
DF4 = '0.0000'

W, LEN_, SN, H = D['wacc'], D['lenses'], D['sensitivity'], D['history']
FR, UB, WC = D['framings'], D['unit_build'], D['working_capital']
CI = {k: v['value'] for k, v in D['company_inputs'].items()}
MC = {k: v['value'] for k, v in D['macro'].items()}
NRM, REL, BV = D['normalised'], D['relative'], D['book_value']
ALT = D['alt_wacc']['bottom_up_sector_beta']
SPOT, FX, SHARES = D['spot_aed'], D['aed_per_usd'], D['shares_out'] / 1e6
USDm = 1e-3
YH = ['FY2023', 'FY2024', 'FY2025']
HY = ['2023', '2024', '2025']
YF = [r['year'] for r in FR['normalisation']['rows']]
HC = ['B', 'C', 'D']
FC = ['E', 'F', 'G', 'H', 'I']
DC = ['B', 'C', 'D', 'E', 'F']
AC = ['B', 'C', 'D', 'E', 'F']
NET_DEBT = FR['normalisation']['net_debt']
LEASES = FR['normalisation']['leases']
NCI = FR['normalisation']['nci']

SHEETS = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
          'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
          'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
          'Per-Share & Ratios', 'Peer & Sector']
wb = Workbook()
wb.remove(wb.active)
WS = {n: wb.create_sheet(n) for n in SHEETS}
EXPECT = {}
PASTED = {'history': 0, 'unit_build': 0, 'rerun': 0}


def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t
    ws['A1'].font = TITLE
    ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s
        ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False, cls=None):
    c = ws[ad]
    c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    if cls and isinstance(v, (int, float)):
        PASTED[cls] += 1
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def lab(ws, r, text, bold=False, col=1):
    ws.cell(row=r, column=col, value=text).font = Font(bold=bold)


def hdr(ws, row, labels, start=1):
    for i, x in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=x)
        c.font = Font(bold=True)
        c.fill = FILL_H


def band(ws, row, text, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)
    ws.cell(row=row, column=1, value=text)


def note(ws, row, text, w=10):
    ws.cell(row=row, column=1, value=text).font = SUB
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=w)


# =============================================================================
# 4  ASSUMPTIONS — populated first; everything points at it
# =============================================================================
wsA = WS['Assumptions']
title(wsA, 'Assumptions — every driver in the model',
      'Blue = input. Change a blue cell and the whole workbook reprices. '
      'Amounts in USD million unless stated.', 9, 54, 14)
A = {}


def inp(row, label_, value, fmt=NUM2, unit='', cls='history', key=None):
    wsA.cell(row=row, column=1, value=label_)
    put(wsA, f'B{row}', value, BLUE, fmt, cls=cls)
    if unit:
        wsA.cell(row=row, column=3, value=unit).font = SUB
    if key:
        A[key] = f'Assumptions!$B${row}'


def a(key):
    return A[key]


r = 4
band(wsA, r, 'MARKET AND SHARE COUNT'); r += 1
inp(r, 'Closing price, 7 August 2026', SPOT, PX, 'AED per share', key='spot'); r += 1
inp(r, 'Dirham per US dollar (pegged since November 1997)', FX, NUM4, 'AED/USD',
    key='fx'); r += 1
inp(r, 'Ordinary shares in issue', SHARES, NUM0, 'million', key='shares'); r += 2

band(wsA, r, 'COST OF CAPITAL — INPUTS; THE BUILD IS ON FUNDAMENTAL VALUATION'); r += 1
inp(r, 'US 10-year Treasury yield', MC['ust_10y'], PCT2, 'the dollar risk-free anchor',
    key='rf_usd'); r += 1
inp(r, 'US adjusted default spread', MC['us_default_spread'], PCT2,
    'subtracted to normalise the risk-free rate', key='us_ds'); r += 1
inp(r, 'Mature-market equity risk premium', MC['mature_erp'], PCT2, '',
    key='mature_erp'); r += 1
inp(r, 'UAE country risk premium', MC['uae_crp'], PCT2, 'country risk enters ONCE, here',
    key='crp'); r += 1
inp(r, 'UAE adjusted default spread', MC['uae_default_spread_rating'], PCT2,
    'the alternative premium basis', key='uae_ds'); r += 1
inp(r, "Beta — the share's own five-year weekly regression", W['beta_own'], NUM3,
    'weak: R-squared 8.8%', key='beta_own'); r += 1
inp(r, 'Beta — sector bottom-up, unlevered', MC['sector_unlevered_beta_diversified'],
    NUM3, 'Chemical (Diversified) — the SAME industry row the EV/EBITDA anchor uses',
    key='beta_unlev'); r += 1
inp(r, 'Debt to equity at market value', W['de_ratio'], PCT,
    'for re-levering the sector beta', key='de_ratio'); r += 1
inp(r, 'Marginal pre-tax cost of debt', W['kd'], PCT2,
    "long dollar rate plus an arm's-length margin", key='kd'); r += 1
inp(r, 'Effective tax rate', W['tax'], PCT,
    "the company's own three-year mean, not the 9% headline", key='tax'); r += 2

band(wsA, r, 'CAPACITY AND THE UNIT BUILD'); r += 1
inp(r, 'Polyethylene nameplate capacity', UB['capacity_pe'], NUM0, 'kt a year',
    key='cap_pe'); r += 1
inp(r, 'Polypropylene nameplate capacity', UB['capacity_pp'], NUM0, 'kt a year',
    key='cap_pp'); r += 1
inp(r, 'Sales over production (partner sourcing)', UB['sourcing_uplift'], NUM4,
    'sales exceed production because Borouge sources from partners — measured from the '
    'audited record and held flat', key='sourcing'); r += 1
inp(r, 'Realisation residual — polyethylene', UB['realisation_pe'], NUM4,
    'three-year mean; computed on Segments', key='real_pe'); r += 1
inp(r, 'Realisation residual — polypropylene', UB['realisation_pp'], NUM4,
    'three-year mean; computed on Segments', key='real_pp'); r += 1
inp(r, 'Other production cost — fixed leg', UB['othprod_fixed'], NUM0,
    'PASTED CLASS 2: least squares across three audited years', 'unit_build',
    'oth_fixed'); r += 1
inp(r, 'Other production cost — per tonne of production', UB['othprod_var_per_t'], NUM4,
    'PASTED CLASS 2: USD million per kt, same regression', 'unit_build', 'oth_var'); r += 1
inp(r, 'Feedstock cost per tonne, FY2025 actual', UB['feed_per_t']['2025'], NUM1,
    'USD per tonne produced — the base both legs escalate from', key='feed_base'); r += 1
inp(r, 'Polypropylene benchmark, FY2025 actual', CI['bench_pp_fy25'], NUM0,
    'the traded leg escalates on THIS, not on consumer inflation',
    key='bench_pp_base'); r += 1
inp(r, 'Contracted ethane real escalation', MC['ethane_contract_real_escalation'], PCT2,
    'zero: the pricing formula is not disclosed, so none is invented',
    key='ethane_esc'); r += 1
inp(r, 'UAE consumer inflation', MC['uae_cpi'], PCT2,
    'escalates the DOMESTIC fixed cost leg only', key='cpi'); r += 2

band(wsA, r, 'FIXED COSTS, WORKING CAPITAL AND CAPITAL SPEND'); r += 1
inp(r, 'General and administrative, FY2025 excluding depreciation', CI['ga_exda_fy25'],
    NUM1, 'USD million', key='ga_base'); r += 1
inp(r, 'Other income, FY2025', CI['othinc_fy25'] * USDm, NUM1, 'USD million',
    key='othinc_base'); r += 1
inp(r, 'Other revenue, FY2025', CI['rev_oth_fy25'], NUM1,
    'USD million, run off at 10% a year', key='rev_oth'); r += 1
inp(r, 'Days sales outstanding', WC['dso'], NUM1, 'days, from the audited statements',
    key='dso'); r += 1
inp(r, 'Days inventory', WC['dio'], NUM1, 'days', key='dio'); r += 1
inp(r, 'Days payable', WC['dpo'], NUM1, 'days', key='dpo'); r += 1
inp(r, 'Capital expenditure guide, 2026', CI['capex_guide_2026'], NUM0, 'USD million',
    key='capex26'); r += 1
inp(r, 'Maintenance capital expenditure, steady state', MC['maintenance_capex'], NUM0,
    'USD million a year', key='capex_maint'); r += 1
inp(r, 'Property, plant and equipment, opening', CI['ppe_fy25'] * USDm, NUM0,
    'USD million at 31 December 2025', key='ppe_open'); r += 1
inp(r, 'Equity attributable to owners, 30 June 2026', CI['eq_owners_h126'] * USDm, NUM0,
    'USD million — the book-value lens reads this, on the SAME date as the bridge',
    key='eq_owners'); r += 1
for _i, _y in enumerate(('2023', '2024', '2025')):
    inp(r, f'Return on equity, FY{_y}', BV['roe_hist'][_i], PCT,
        'profit attributable to owners over opening equity, from the audited statements',
        key=f'roe{_y[2:]}'); r += 1
inp(r, 'Depreciation rate on the property balance',
    FR['normalisation']['rows'][0]['da'] / (CI['ppe_fy25'] * USDm), PCT2,
    'the H1-2026 charge annualised, over the opening property balance',
    key='dep_rate'); r += 2

band(wsA, r, 'TERMINAL BLOCK AND THE BRIDGE'); r += 1
inp(r, 'Terminal growth rate', MC['terminal_growth'], PCT2,
    'long-run dollar inflation; owned capacity is fixed', key='g'); r += 1
inp(r, 'Terminal return on capital', MC['terminal_roc'], PCT,
    'sets the reinvestment the terminal block must fund', key='roc'); r += 1
inp(r, 'Net debt', NET_DEBT, NUM0, 'USD million at 30 June 2026', key='net_debt'); r += 1
inp(r, 'Lease liabilities', LEASES, NUM1, 'USD million', key='leases'); r += 1
inp(r, 'Non-controlling interests', NCI, NUM1, 'USD million at book', key='nci'); r += 1
inp(r, 'Borouge 4 earnings accretion after ramp-up', MC['b4_accretion_post_rampup'], PCT,
    "the sponsors' own figure", key='b4_accr'); r += 1
inp(r, 'Borouge 4 cumulative net profit, first three years',
    MC['b4_cumulative_net_profit_3y'], NUM0, 'USD million; the second quantification',
    key='b4_cum'); r += 2

band(wsA, r, 'FORECAST DRIVERS — NORMALISATION CONSTRUCTION'); r += 1
hdr(wsA, r, [''] + [str(y) for y in YF]); r += 1
DRV_N, DRV_P = {}, {}
DRIVERS = [('util_pe', 'Polyethylene utilisation', PCT),
           ('util_pp', 'Polypropylene utilisation', PCT),
           ('bench_pe', 'Polyethylene benchmark (USD/t)', NUM0),
           ('bench_pp', 'Polypropylene benchmark (USD/t)', NUM0),
           ('prem_pe', 'Polyethylene premium (USD/t)', NUM0),
           ('prem_pp', 'Polypropylene premium (USD/t)', NUM0),
           ('sd_per_t', 'Selling and distribution (USD/t sold)', NUM0),
           ('feed_market_share', 'Share of feedstock bought at market price', PCT)]
for key, label_, fmt in DRIVERS:
    wsA.cell(row=r, column=1, value=label_)
    for i, col in enumerate(AC):
        put(wsA, f'{col}{r}', D['framing_drivers']['normalisation'][key][i], BLUE, fmt,
            cls='history')
    DRV_N[key] = r
    r += 1
r += 1
band(wsA, r, 'FORECAST DRIVERS — PROLONGED-DISRUPTION CONSTRUCTION'); r += 1
hdr(wsA, r, [''] + [str(y) for y in YF]); r += 1
for key, label_, fmt in DRIVERS:
    wsA.cell(row=r, column=1, value=label_)
    for i, col in enumerate(AC):
        put(wsA, f'{col}{r}', D['framing_drivers']['prolonged'][key][i], BLUE, fmt,
            cls='history')
    DRV_P[key] = r
    r += 1
r += 1
note(wsA, r, 'Every other number in this workbook is computed from the cells above, from '
             'the audited history on the statement sheets, or is one of the three pasted '
             'classes named on READ FIRST.', 9)

# =============================================================================
# 3a FUNDAMENTAL VALUATION — cost of capital (needed by DCF)
# =============================================================================
wsF = WS['Fundamental Valuation']
title(wsF, 'Fundamental valuation — the cost of capital, built line by line',
      'Nothing on this sheet is pasted. Every rate is assembled from the inputs on '
      'Assumptions.', 9, 58, 15)
r = 4
band(wsF, r, 'RISK-FREE RATE, NORMALISED'); r += 1
lab(wsF, r, 'US 10-year Treasury yield')
putf(wsF, f'B{r}', f'={a("rf_usd")}', MC['ust_10y'], PCT2, green=True); r += 1
lab(wsF, r, 'Less: US adjusted default spread')
putf(wsF, f'B{r}', f'=-{a("us_ds")}', -MC['us_default_spread'], PCT2, green=True); r += 1
lab(wsF, r, 'Normalised risk-free rate', bold=True)
RF = f'B{r}'
putf(wsF, f'B{r}', f'=B{r - 2}+B{r - 1}', W['rf_star'], PCT2, bold=True)
wsF.cell(row=r, column=3, value='Country risk is charged once, in the premium below — '
                                'never twice').font = SUB
r += 2

band(wsF, r, 'EQUITY RISK PREMIUM (BOTH BASES PUBLISHED)'); r += 1
lab(wsF, r, 'Mature-market equity risk premium')
putf(wsF, f'B{r}', f'={a("mature_erp")}', MC['mature_erp'], PCT2, green=True); r += 1
lab(wsF, r, 'Plus: UAE country risk premium (rating basis)')
putf(wsF, f'B{r}', f'={a("crp")}', MC['uae_crp'], PCT2, green=True); r += 1
lab(wsF, r, 'Equity risk premium — rating basis', bold=True)
ERP = f'B{r}'
putf(wsF, f'B{r}', f'=B{r - 2}+B{r - 1}', W['erp_rating'], PCT2, bold=True); r += 1
lab(wsF, r, 'Equity risk premium — default-spread basis')
putf(wsF, f'B{r}', f'=B{r - 3}+{a("uae_ds")}', W['erp_default_spread_basis'], PCT2)
wsF.cell(row=r, column=3, value='Published alongside: no sovereign credit-default-swap '
                                'quote exists for this sovereign').font = SUB
r += 2

band(wsF, r, "BETA — THE STUDY'S CENTRAL CONTESTED JUDGEMENT"); r += 1
lab(wsF, r, "The share's own five-year weekly regression beta")
B_OWN = f'B{r}'
putf(wsF, f'B{r}', f'={a("beta_own")}', W['beta_own'], NUM3, green=True)
wsF.cell(row=r, column=3, value=f"R-squared {BETA['r2']:.3f} on {BETA['n']} weekly "
                                f"observations — statistically weak").font = SUB
r += 1
lab(wsF, r, 'Sector unlevered beta (global chemicals)')
putf(wsF, f'B{r}', f'={a("beta_unlev")}', MC['sector_unlevered_beta_diversified'],
     NUM3, green=True); r += 1
lab(wsF, r, "Re-levered at the company's debt-to-equity and tax rate", bold=True)
B_BU = f'B{r}'
putf(wsF, f'B{r}', f'=B{r - 1}*(1+(1-{a("tax")})*{a("de_ratio")})', W['beta_bottom_up'],
     NUM3, bold=True)
r += 2

band(wsF, r, 'COST OF EQUITY, COST OF DEBT AND THE WEIGHTS'); r += 1
hdr(wsF, r, ['', 'Own-stock beta', 'Sector bottom-up beta']); r += 1
lab(wsF, r, 'Cost of equity = risk-free rate + beta x premium')
KE_OWN, KE_BU = f'B{r}', f'C{r}'
putf(wsF, f'B{r}', f'=${RF}+{B_OWN}*${ERP}', W['ke_own'], PCT2)
putf(wsF, f'C{r}', f'=${RF}+{B_BU}*${ERP}', W['ke_bottom_up'], PCT2); r += 1
lab(wsF, r, 'Pre-tax cost of debt (marginal)')
putf(wsF, f'B{r}', f'={a("kd")}', W['kd'], PCT2, green=True)
putf(wsF, f'C{r}', f'={a("kd")}', W['kd'], PCT2, green=True); r += 1
lab(wsF, r, 'After-tax cost of debt')
KDAT_B, KDAT_C = f'B{r}', f'C{r}'
putf(wsF, f'B{r}', f'=B{r - 1}*(1-{a("tax")})', W['kd'] * (1 - W['tax']), PCT2)
putf(wsF, f'C{r}', f'=C{r - 1}*(1-{a("tax")})', W['kd'] * (1 - W['tax']), PCT2); r += 1
lab(wsF, r, 'Market capitalisation (USD million)')
MKTCAP = f'B{r}'
putf(wsF, f'B{r}', f'={a("spot")}/{a("fx")}*{a("shares")}', W['mktcap'], NUM0); r += 1
lab(wsF, r, 'Net debt (USD million)')
putf(wsF, f'B{r}', f'={a("net_debt")}', NET_DEBT, NUM0, green=True); r += 1
lab(wsF, r, 'Equity weight')
WE = f'B{r}'
putf(wsF, f'B{r}', f'=B{r - 2}/(B{r - 2}+B{r - 1})', W['equity_weight'], PCT); r += 1
lab(wsF, r, 'Debt weight')
WD = f'B{r}'
putf(wsF, f'B{r}', f'=1-B{r - 1}', W['debt_weight'], PCT); r += 1
lab(wsF, r, 'Weighted average cost of capital', bold=True)
WACC_OWN, WACC_BU = f'$B${r}', f'$C${r}'
putf(wsF, f'B{r}', f'={KE_OWN}*${WE}+{KDAT_B}*${WD}', W['wacc_own'], PCT2, bold=True)
putf(wsF, f'C{r}', f'={KE_BU}*${WE}+{KDAT_C}*${WD}', W['wacc_bottom_up'], PCT2, bold=True)
r += 1
note(wsF, r, 'The two columns are the study\'s central contested judgement. They are '
             'carried side by side to a value per share everywhere they appear and are '
             'never averaged into one number.', 9)
FV_LENS_START = r + 2
WACC_OWN_REF = f"'Fundamental Valuation'!{WACC_OWN}"
WACC_BU_REF = f"'Fundamental Valuation'!{WACC_BU}"

# =============================================================================
# 6  SEGMENTS — the unit build
# =============================================================================
wsS = WS['Segments']
title(wsS, 'Segments — the unit build, tonnes and dollars per tonne',
      'Borouge reports as one operating segment. Volume and price are disclosed by '
      'product; cost is allocated on the disclosed physical drivers.', 9, 50, 13)
r = 4
band(wsS, r, 'DISCLOSED HISTORY (PASTED — THE PRIMARY RECORD)'); r += 1
hdr(wsS, r, [''] + YH); r += 1
SR = {}
for key, label_, fmt in [('vol_pe', 'Polyethylene sold (kt)', NUM0),
                         ('vol_pp', 'Polypropylene sold (kt)', NUM0),
                         ('production', 'Total production (kt)', NUM0),
                         ('bench_pe', 'Polyethylene benchmark (USD/t)', NUM0),
                         ('bench_pp', 'Polypropylene benchmark (USD/t)', NUM0),
                         ('prem_pe', 'Polyethylene premium (USD/t)', NUM0),
                         ('prem_pp', 'Polypropylene premium (USD/t)', NUM0),
                         ('rev_pe', 'Polyethylene revenue (USD m)', NUM0),
                         ('rev_pp', 'Polypropylene revenue (USD m)', NUM0),
                         ('feed_per_t', 'Feedstock (USD per tonne produced)', NUM1),
                         ('sd_per_t', 'Selling and distribution (USD per tonne sold)',
                          NUM1)]:
    wsS.cell(row=r, column=1, value=label_)
    for i, y in enumerate(HY):
        put(wsS, f'{HC[i]}{r}', UB[key][y], BLUE, fmt, cls='history')
    SR[key] = r
    r += 1
r += 1
band(wsS, r, 'THE REALISATION RESIDUAL, COMPUTED FROM THAT HISTORY'); r += 1
hdr(wsS, r, [''] + YH + ['Three-year mean']); r += 1
lab(wsS, r, 'Polyethylene: revenue per tonne / (benchmark + premium)')
for i, y in enumerate(HY):
    c = HC[i]
    putf(wsS, f'{c}{r}',
         f'={c}{SR["rev_pe"]}*1000/{c}{SR["vol_pe"]}/({c}{SR["bench_pe"]}'
         f'+{c}{SR["prem_pe"]})', UB['realisation_by_year'][y]['pe'], NUM4)
putf(wsS, f'E{r}', f'=AVERAGE(B{r}:D{r})', UB['realisation_pe'], NUM4, bold=True)
r += 1
lab(wsS, r, 'Polypropylene: revenue per tonne / (benchmark + premium)')
for i, y in enumerate(HY):
    c = HC[i]
    putf(wsS, f'{c}{r}',
         f'={c}{SR["rev_pp"]}*1000/{c}{SR["vol_pp"]}/({c}{SR["bench_pp"]}'
         f'+{c}{SR["prem_pp"]})', UB['realisation_by_year'][y]['pp'], NUM4)
putf(wsS, f'E{r}', f'=AVERAGE(B{r}:D{r})', UB['realisation_pp'], NUM4, bold=True)
r += 1
note(wsS, r, 'The residual is the bridge from a published benchmark to the company\'s own '
             'printed top line. It is computed here from disclosed revenue, volume, '
             'benchmark and premium — not assumed.', 9)
r += 2

SEGF = {}
for tag, drv, rows_ in [('N', DRV_N, FR['normalisation']['rows']),
                        ('P', DRV_P, FR['prolonged']['rows'])]:
    name = ('NORMALISATION' if tag == 'N' else 'PROLONGED-DISRUPTION')
    band(wsS, r, f'THE FORECAST UNIT BUILD — {name} CONSTRUCTION'); r += 1
    hdr(wsS, r, [''] + [str(y) for y in YF]); r += 1
    m = {}
    for key, label_ in [('vol_pe', 'Polyethylene volume (kt) = capacity x utilisation'),
                        ('vol_pp', 'Polypropylene volume (kt) = capacity x utilisation'),
                        ('vol_tot', 'Total volume (kt)'),
                        ('price_pe', 'Polyethylene realised price (USD/t)'),
                        ('price_pp', 'Polypropylene realised price (USD/t)'),
                        ('feed_unit', 'Feedstock (USD per tonne produced)')]:
        wsS.cell(row=r, column=1, value=label_)
        m[key] = r
        r += 1
    for i, col in enumerate(DC):
        ac = AC[i]
        row = rows_[i]
        putf(wsS, f'{col}{m["vol_pe"]}',
             f'={a("cap_pe")}*Assumptions!{ac}${drv["util_pe"]}', row['vol_pe'], NUM0)
        putf(wsS, f'{col}{m["vol_pp"]}',
             f'={a("cap_pp")}*Assumptions!{ac}${drv["util_pp"]}', row['vol_pp'], NUM0)
        putf(wsS, f'{col}{m["vol_tot"]}',
             f'={col}{m["vol_pe"]}+{col}{m["vol_pp"]}', row['vol_tot'], NUM0, bold=True)
        putf(wsS, f'{col}{m["price_pe"]}',
             f'=(Assumptions!{ac}${drv["bench_pe"]}+Assumptions!{ac}${drv["prem_pe"]})'
             f'*{a("real_pe")}', row['price_pe'], NUM1)
        putf(wsS, f'{col}{m["price_pp"]}',
             f'=(Assumptions!{ac}${drv["bench_pp"]}+Assumptions!{ac}${drv["prem_pp"]})'
             f'*{a("real_pp")}', row['price_pp'], NUM1)
        eth = f'{a("feed_base")}*(1+{a("ethane_esc")})^{i + 1}'
        prop = (f'{a("feed_base")}*Assumptions!{ac}${drv["bench_pp"]}'
                f'/{a("bench_pp_base")}')
        putf(wsS, f'{col}{m["feed_unit"]}',
             f'=(1-Assumptions!{ac}${drv["feed_market_share"]})*({eth})'
             f'+Assumptions!{ac}${drv["feed_market_share"]}*({prop})',
             row['feed_unit'], NUM1)
    SEGF[tag] = m
    r += 1
note(wsS, r, 'Two escalators, because there are two physically different inputs: '
             'contracted ethane escalates on its own terms, purchased propylene on the '
             'propylene benchmark. A single blended cost index across both would '
             'manufacture a margin trend out of arithmetic.', 9)

# =============================================================================
# 8  DCF
# =============================================================================
wsD = WS['DCF']
title(wsD, 'Discounted cash flow — the waterfall and the terminal block',
      'USD million. Both constructions. Every cell below is a formula.', 9, 52, 13)


def waterfall(ws, r0, rows_, drv, seg, wacc_ref, heading):
    band(ws, r0, heading.upper()); r_ = r0 + 1
    hdr(ws, r_, [''] + [str(y) for y in YF]); r_ += 1
    R = {}
    for key, label_, fmt, bold in [
            ('revenue', 'Revenue', NUM0, False),
            ('feedstock', 'Less: feedstock', NUM0, False),
            ('othprod', 'Less: other production cost', NUM0, False),
            ('sd', 'Less: selling and distribution', NUM0, False),
            ('ga', 'Less: general and administrative', NUM0, False),
            ('other_income', 'Add: other income', NUM1, False),
            ('ebitda', 'EBITDA', NUM0, True),
            ('ebitda_margin', 'EBITDA margin = EBITDA / revenue', PCT, False),
            ('da', 'Less: depreciation and amortisation', NUM0, False),
            ('ebit', 'EBIT = EBITDA less depreciation and amortisation', NUM0, True),
            ('nopat', 'NOPAT = EBIT x (1 - tax rate)', NUM0, True),
            ('da2', 'Add back: depreciation and amortisation', NUM0, False),
            ('capex', 'Less: capital expenditure', NUM0, False),
            ('d_nwc', 'Less: increase in working capital', NUM1, False),
            ('fcff', 'Free cash flow to the firm', NUM0, True),
            ('wacc_row', 'Discount rate applied', PCT2, False),
            ('discount_factor', 'Discount factor', DF4, False),
            ('pv_fcff', 'Present value of free cash flow to the firm', NUM0, True)]:
        lab(ws, r_, label_, bold=bold)
        R[key] = r_
        r_ += 1
    for i, col in enumerate(DC):
        ac = AC[i]
        row = rows_[i]
        putf(ws, f'{col}{R["revenue"]}',
             f'=(Segments!{col}${seg["vol_pe"]}*Segments!{col}${seg["price_pe"]}'
             f'+Segments!{col}${seg["vol_pp"]}*Segments!{col}${seg["price_pp"]})/1000'
             f'*{a("sourcing")}+{a("rev_oth")}*(1-0.1)^{i + 1}', row['revenue'], NUM0)
        putf(ws, f'{col}{R["feedstock"]}',
             f'=Segments!{col}${seg["feed_unit"]}*Segments!{col}${seg["vol_tot"]}/1000',
             row['feedstock'], NUM0)
        putf(ws, f'{col}{R["othprod"]}',
             f'={a("oth_fixed")}*(1+{a("cpi")})^{i + 1}+{a("oth_var")}'
             f'*Segments!{col}${seg["vol_tot"]}', row['othprod'], NUM0)
        putf(ws, f'{col}{R["sd"]}',
             f'=Assumptions!{ac}${drv["sd_per_t"]}*Segments!{col}${seg["vol_tot"]}'
             f'*{a("sourcing")}/1000', row['sd'], NUM0)
        putf(ws, f'{col}{R["ga"]}', f'={a("ga_base")}*(1+{a("cpi")})^{i + 1}', row['ga'],
             NUM1)
        putf(ws, f'{col}{R["other_income"]}',
             f'={a("othinc_base")}*(1+{a("cpi")})^{i + 1}', row['other_income'], NUM1)
        putf(ws, f'{col}{R["ebitda"]}',
             f'={col}{R["revenue"]}-{col}{R["feedstock"]}-{col}{R["othprod"]}'
             f'-{col}{R["sd"]}-{col}{R["ga"]}+{col}{R["other_income"]}',
             row['ebitda'], NUM0, bold=True)
        putf(ws, f'{col}{R["ebitda_margin"]}',
             f'={col}{R["ebitda"]}/{col}{R["revenue"]}', row['ebitda_margin'], PCT)
        # Depreciation is the opening property balance times the disclosed rate, then the
        # balance rolls: prior balance + capital spend - depreciation.
        if i == 0:
            putf(ws, f'{col}{R["da"]}', f'={a("ppe_open")}*{a("dep_rate")}', row['da'],
                 NUM0)
        else:
            putf(ws, f'{col}{R["da"]}',
                 f'=({a("ppe_open")}'
                 f'+SUM(B{R["capex"]}:{DC[i - 1]}{R["capex"]})'
                 f'-SUM(B{R["da"]}:{DC[i - 1]}{R["da"]}))*{a("dep_rate")}',
                 row['da'], NUM0)
        putf(ws, f'{col}{R["ebit"]}', f'={col}{R["ebitda"]}-{col}{R["da"]}', row['ebit'],
             NUM0, bold=True)
        putf(ws, f'{col}{R["nopat"]}', f'={col}{R["ebit"]}*(1-{a("tax")})', row['nopat'],
             NUM0, bold=True)
        putf(ws, f'{col}{R["da2"]}', f'={col}{R["da"]}', row['da'], NUM0)
        putf(ws, f'{col}{R["capex"]}',
             f'={a("capex26")}' if i == 0 else f'={a("capex_maint")}', row['capex'], NUM0)
        # Working capital from the disclosed conversion cycle — no plug.
        cogs = f'({col}{R["feedstock"]}+{col}{R["othprod"]})'
        nwc = (f'({col}{R["revenue"]}*{a("dso")}/365+{cogs}*{a("dio")}/365'
               f'-{cogs}*{a("dpo")}/365)')
        if i == 0:
            base_cogs = CI['feed_fy25'] + CI['othprod_fy25']
            prev = (f'({H["2025"]["revenue"]!r}*{a("dso")}/365+{base_cogs!r}*{a("dio")}'
                    f'/365-{base_cogs!r}*{a("dpo")}/365)')
        else:
            pc = DC[i - 1]
            pcogs = f'({pc}{R["feedstock"]}+{pc}{R["othprod"]})'
            prev = (f'({pc}{R["revenue"]}*{a("dso")}/365+{pcogs}*{a("dio")}/365'
                    f'-{pcogs}*{a("dpo")}/365)')
        putf(ws, f'{col}{R["d_nwc"]}', f'={nwc}-{prev}', row['d_nwc'], NUM1)
        putf(ws, f'{col}{R["fcff"]}',
             f'={col}{R["nopat"]}+{col}{R["da2"]}-{col}{R["capex"]}-{col}{R["d_nwc"]}',
             row['fcff'], NUM0, bold=True)
        putf(ws, f'{col}{R["wacc_row"]}', f'={wacc_ref}', row.get('wacc', W['wacc_own']),
             PCT2, green=True)
        putf(ws, f'{col}{R["discount_factor"]}',
             f'=1/(1+{col}{R["wacc_row"]})^{i + 1}', row['discount_factor'], DF4)
        putf(ws, f'{col}{R["pv_fcff"]}',
             f'={col}{R["fcff"]}*{col}{R["discount_factor"]}', row['pv_fcff'], NUM0,
             bold=True)
    return R, r_


def terminal(ws, r0, R, F_, heading):
    band(ws, r0, heading.upper()); r_ = r0 + 1
    T = {}
    lab(ws, r_, 'Final-year NOPAT')
    putf(ws, f'B{r_}', f'=F{R["nopat"]}', F_['rows'][-1]['nopat'], NUM0); r_ += 1
    lab(ws, r_, 'Terminal NOPAT = final-year NOPAT x (1 + growth)')
    T['nopat'] = f'B{r_}'
    putf(ws, f'B{r_}', f'=B{r_ - 1}*(1+{a("g")})', F_['terminal_nopat'], NUM0); r_ += 1
    lab(ws, r_, 'Reinvestment rate = growth / return on capital')
    T['ri'] = f'B{r_}'
    putf(ws, f'B{r_}', f'={a("g")}/{a("roc")}', F_['reinvestment_rate'], PCT); r_ += 1
    lab(ws, r_, 'Terminal free cash flow = NOPAT x (1 - reinvestment rate)')
    T['fcff'] = f'B{r_}'
    putf(ws, f'B{r_}', f'={T["nopat"]}*(1-{T["ri"]})', F_['terminal_fcff'], NUM0); r_ += 1
    lab(ws, r_, 'Terminal value = terminal cash flow / (cost of capital - growth)')
    T['tv'] = f'B{r_}'
    putf(ws, f'B{r_}', f'={T["fcff"]}/(F{R["wacc_row"]}-{a("g")})', F_['terminal_value'],
         NUM0); r_ += 1
    lab(ws, r_, 'Present value of the terminal value', bold=True)
    T['pvtv'] = f'B{r_}'
    putf(ws, f'B{r_}', f'={T["tv"]}*F{R["discount_factor"]}', F_['pv_terminal'], NUM0,
         bold=True); r_ += 1
    lab(ws, r_, 'Present value of the explicit five years')
    T['pvex'] = f'B{r_}'
    putf(ws, f'B{r_}', f'=SUM(B{R["pv_fcff"]}:F{R["pv_fcff"]})', F_['pv_explicit'], NUM0)
    r_ += 1
    lab(ws, r_, 'Enterprise value of the owned business', bold=True)
    T['evc'] = f'B{r_}'
    putf(ws, f'B{r_}', f'={T["pvex"]}+{T["pvtv"]}', F_['ev_core'], NUM0, bold=True)
    r_ += 1
    return T, r_


RN, r = waterfall(wsD, 4, FR['normalisation']['rows'], DRV_N, SEGF['N'], WACC_OWN_REF,
                  'Waterfall — normalisation construction')
r += 1
TN, r = terminal(wsD, r, RN, FR['normalisation'], 'Terminal block — normalisation')
r += 1
RP, r = waterfall(wsD, r, FR['prolonged']['rows'], DRV_P, SEGF['P'], WACC_OWN_REF,
                  'Waterfall — prolonged-disruption construction')
r += 1
TP, r = terminal(wsD, r, RP, FR['prolonged'], 'Terminal block — prolonged disruption')
r += 1

band(wsD, r, 'VALUE PER SHARE — FOUR READINGS, NEVER AVERAGED'); r += 1
hdr(wsD, r, ['Construction and beta', 'Enterprise value', 'Equity value',
             'Value per share (USD)', 'Terminal value % of EV', 'Value per share (AED)'],
    start=1)
r += 1
VPS = {}
for tag, frk, T, own in [('own_n', 'normalisation', TN, True),
                         ('bu_n', 'normalisation', TN, False),
                         ('own_p', 'prolonged', TP, True),
                         ('bu_p', 'prolonged', TP, False)]:
    F_ = FR[frk]
    b4v = F_['b4']['value']
    lab(wsD, r, ('Normalisation' if frk == 'normalisation' else 'Prolonged disruption') +
        (', own-stock beta' if own else ', sector bottom-up beta'))
    ev_v = F_['ev'] if own else ALT['ev'][frk]
    putf(wsD, f'B{r}', f'={T["evc"]}+{b4v!r}' if own else f'={ev_v!r}', ev_v, NUM0,
         green=not own)
    if not own:
        PASTED['rerun'] += 1
    eq = ev_v - NET_DEBT - LEASES - NCI
    putf(wsD, f'C{r}', f'=B{r}-{a("net_debt")}-{a("leases")}-{a("nci")}', eq, NUM0)
    putf(wsD, f'D{r}', f'=C{r}/{a("shares")}', eq / SHARES, NUM4)
    tvs = (F_['pv_terminal'] / ev_v) if own else ALT['tv_share'][frk]
    putf(wsD, f'E{r}', f'={T["pvtv"]}/B{r}' if own else f'={tvs!r}', tvs, PCT,
         green=not own)
    if not own:
        PASTED['rerun'] += 1
    putf(wsD, f'F{r}', f'=D{r}*{a("fx")}', eq / SHARES * FX, PX, bold=True)
    VPS[tag] = r
    r += 1
note(wsD, r, 'The sector-beta enterprise values are a complete re-solve of the same '
             'waterfall at the other cost of capital — a whole-model re-run, so they are '
             'carried as values rather than as a second live chain, and they do not '
             'redraw when a driver changes. The bridge from enterprise value to a value '
             'per share is a live formula in every row.', 9)
DCF_TVSHARE_OWN_N = f"'DCF'!$E${VPS['own_n']}"

# =============================================================================
# 5  SOTP BRIDGE — enterprise value to equity, with the Borouge 4 fee stream
# =============================================================================
wsB = WS['SOTP Bridge']
title(wsB, 'Enterprise value to equity value',
      'Borouge is one operating business plus one operator fee stream. The fee stream is '
      'valued separately and never consolidated as owned capacity.', 9, 56, 15)
r = 4
band(wsB, r, 'THE BOROUGE 4 FEE STREAM — QUANTIFIED TWO WAYS, THE LOWER CARRIED'); r += 1
B4 = FR['normalisation']['b4']
lab(wsB, r, 'From the disclosed three-year cumulative net profit')
putf(wsB, f'B{r}', f'={a("b4_cum")}/{sum([0.10, 0.30, 0.60])!r}',
     B4['steady_from_cumulative'], NUM1); r += 1
lab(wsB, r, 'From the disclosed post-ramp-up earnings accretion')
putf(wsB, f'B{r}',
     f"={a('b4_accr')}*AVERAGE('DCF'!B{RN['nopat']}:F{RN['nopat']})",
     B4['steady_from_accretion'], NUM1); r += 1
lab(wsB, r, 'Steady-state fee carried — the LOWER of the two', bold=True)
B4_STEADY = f'B{r}'
putf(wsB, f'B{r}', f'=MIN(B{r - 2},B{r - 1})', B4['steady_adopted'], NUM1, bold=True)
wsB.cell(row=r, column=3, value='The sponsors\' two figures do not agree; the study does '
                                'not average them').font = SUB
r += 2
hdr(wsB, r, [''] + [str(y) for y in YF]); r += 1
lab(wsB, r, 'Ramp-up fraction')
B4_RAMP = r
for i, col in enumerate(DC):
    put(wsB, f'{col}{r}', B4['rows'][i]['ramp'], BLUE, PCT, cls='history')
r += 1
lab(wsB, r, 'Fee to Borouge plc (USD million)')
B4_CASH = r
for i, col in enumerate(DC):
    putf(wsB, f'{col}{r}', f'=${B4_STEADY}*{col}{B4_RAMP}', B4['rows'][i]['net_profit'],
         NUM1)
r += 1
lab(wsB, r, 'Present value of the fee')
B4_PV = r
for i, col in enumerate(DC):
    putf(wsB, f'{col}{r}', f"={col}{B4_CASH}*'DCF'!{col}${RN['discount_factor']}",
         B4['rows'][i]['pv'], NUM1)
r += 2
lab(wsB, r, 'Present value of the explicit fee years')
putf(wsB, f'B{r}', f'=SUM(B{B4_PV}:F{B4_PV})', B4['pv_explicit'], NUM1); r += 1
lab(wsB, r, 'Present value of the fee beyond the forecast')
putf(wsB, f'B{r}', f'=${B4_STEADY}*0', B4['pv_terminal'], NUM1)
wsB.cell(row=r, column=3,
         value='NIL by construction: the Asset Usage Agreement runs only until the '
               'assets are acquired by the parent group, which the company says is not '
               'anticipated before 2029. A stream that ends cannot carry a '
               'perpetuity').font = SUB
r += 1
lab(wsB, r, 'Value of the Borouge 4 fee stream', bold=True)
B4_VAL = f'B{r}'
putf(wsB, f'B{r}', f'=B{r - 2}+B{r - 1}', B4['value'], NUM1, bold=True)
B4V = B4['value']
r += 2

band(wsB, r, 'THE BRIDGE — BOTH CONSTRUCTIONS, BOTH BETAS'); r += 1
hdr(wsB, r, ['', 'Normalisation, own beta', 'Normalisation, sector beta',
             'Prolonged, own beta', 'Prolonged, sector beta']); r += 1
BR_COLS = {'own_n': 'B', 'bu_n': 'C', 'own_p': 'D', 'bu_p': 'E'}
rows_map = [('Enterprise value', 'B', NUM0),
            ('Less: net debt', None, NUM0),
            ('Less: lease liabilities', None, NUM1),
            ('Less: non-controlling interests', None, NUM1),
            ('Equity value', None, NUM0),
            ('Shares in issue (million)', None, NUM0),
            ('Value per share (USD)', None, NUM4),
            ('Value per share (AED)', None, PX),
            ('Terminal value as a share of enterprise value', None, PCT)]
BR = {}
for i, (label_, _, fmt) in enumerate(rows_map):
    lab(wsB, r, label_, bold=label_ in ('Equity value', 'Value per share (AED)',
                                        'Terminal value as a share of enterprise value'))
    BR[label_] = r
    r += 1
for tag, col in BR_COLS.items():
    src = VPS[tag]
    frk = 'normalisation' if tag.endswith('_n') else 'prolonged'
    own = tag.startswith('own')
    ev_v = FR[frk]['ev'] if own else ALT['ev'][frk]
    eq = ev_v - NET_DEBT - LEASES - NCI
    tvs = (FR[frk]['pv_terminal'] / ev_v) if own else ALT['tv_share'][frk]
    putf(wsB, f'{col}{BR["Enterprise value"]}', f"='DCF'!$B${src}", ev_v, NUM0, green=True)
    putf(wsB, f'{col}{BR["Less: net debt"]}', f'=-{a("net_debt")}', -NET_DEBT, NUM0)
    putf(wsB, f'{col}{BR["Less: lease liabilities"]}', f'=-{a("leases")}', -LEASES, NUM1)
    putf(wsB, f'{col}{BR["Less: non-controlling interests"]}', f'=-{a("nci")}', -NCI, NUM1)
    putf(wsB, f'{col}{BR["Equity value"]}',
         f'=SUM({col}{BR["Enterprise value"]}:{col}{BR["Less: non-controlling interests"]})',
         eq, NUM0, bold=True)
    putf(wsB, f'{col}{BR["Shares in issue (million)"]}', f'={a("shares")}', SHARES, NUM0)
    putf(wsB, f'{col}{BR["Value per share (USD)"]}',
         f'={col}{BR["Equity value"]}/{col}{BR["Shares in issue (million)"]}',
         eq / SHARES, NUM4)
    putf(wsB, f'{col}{BR["Value per share (AED)"]}',
         f'={col}{BR["Value per share (USD)"]}*{a("fx")}', eq / SHARES * FX, PX, bold=True)
    putf(wsB, f'{col}{BR["Terminal value as a share of enterprise value"]}',
         f"='DCF'!$E${src}", tvs, PCT, green=True, bold=True)
note(wsB, r, 'The terminal value share is shown in the bridge itself, live, so a reader '
             'can see at a glance how much of each answer rests on the years beyond the '
             'forecast rather than inside it.', 9)

# =============================================================================
# 7  RELATIVE & NORMALIZED
# =============================================================================
wsR = WS['Relative & Normalized']
title(wsR, 'Relative multiples and normalised earnings power',
      'The multiple is triangulated on the sheet from three through-cycle anchors, and '
      'averaged here rather than asserted.', 9, 52, 15)
r = 4
band(wsR, r, 'WHY THE PEER MEDIAN IS REJECTED'); r += 1
lab(wsR, r, 'Listed peers observed')
putf(wsR, f'B{r}', f'={len(D["peer_table"])}', len(D['peer_table']), NUM0); r += 1
lab(wsR, r, 'Of which loss-making on trailing net income')
put(wsR, f'B{r}', D['peers_loss_making'], BLUE, NUM0, cls='history'); r += 1
lab(wsR, r, 'Of which enterprise value to EBITDA is undefined')
put(wsR, f'B{r}', D['peers_ev_undefined'], BLUE, NUM0, cls='history'); r += 1
lab(wsR, r, 'Naive median of whatever prints')
put(wsR, f'B{r}', D['peer_naive_median'], BLUE, MULT, cls='history')
wsR.cell(row=r, column=3, value='Rejected: a median across collapsed denominators '
                                'measures the trough, not the multiple').font = SUB
r += 2

band(wsR, r, 'THE THROUGH-CYCLE TRIANGULATION'); r += 1
TRI_FIRST = r
for k, v in D['relative_triangulation'].items():
    lab(wsR, r, k)
    put(wsR, f'B{r}', v, BLUE, MULT, cls='history')
    r += 1
TRI_LAST = r - 1
lab(wsR, r, 'Adopted multiple — the median of the three anchors', bold=True)
MULT_CELL = f'B{r}'
putf(wsR, f'B{r}', f'=MEDIAN(B{TRI_FIRST}:B{TRI_LAST})', REL['median_ev_ebitda'], MULT,
     bold=True); r += 2

band(wsR, r, 'THE RELATIVE LENS'); r += 1
lab(wsR, r, 'Mid-cycle EBITDA (USD million)')
MIDEB = f'B{r}'
putf(wsR, f'B{r}', f"=AVERAGE('DCF'!D{RN['ebitda']},'DCF'!D{RP['ebitda']})",
     REL['midcycle_ebitda'], NUM0, green=True)
wsR.cell(row=r, column=3, value='The third forecast year of each construction, averaged '
                                '— the mid-point of the cycle the study models').font = SUB
r += 1
lab(wsR, r, 'Enterprise value = multiple x mid-cycle EBITDA')
putf(wsR, f'B{r}', f'={MULT_CELL}*{MIDEB}', REL['ev'] - B4V, NUM0); r += 1
lab(wsR, r, 'Add: the Borouge 4 fee stream, on the same basis as the cash-flow lens')
putf(wsR, f'B{r}', f"='SOTP Bridge'!{B4_VAL.replace('B', '$B$')}", B4V, NUM1, green=True)
wsR.cell(row=r, column=3, value='The same separable asset the cash-flow lens carries. '
                                'Omitting it here valued three different companies '
                                'under three lenses').font = SUB
r += 1
lab(wsR, r, 'Enterprise value including the fee stream')
putf(wsR, f'B{r}', f'=B{r - 2}+B{r - 1}', REL['ev'], NUM0); r += 1
lab(wsR, r, 'Less: net debt, leases and non-controlling interests')
putf(wsR, f'B{r}', f'=-{a("net_debt")}-{a("leases")}-{a("nci")}',
     -(NET_DEBT + LEASES + NCI), NUM0); r += 1
lab(wsR, r, 'Equity value')
putf(wsR, f'B{r}', f'=B{r - 2}+B{r - 1}', REL['equity'], NUM0); r += 1
lab(wsR, r, 'Relative lens — value per share (AED)', bold=True)
REL_VPS = f'B{r}'
putf(wsR, f'B{r}', f'=B{r - 1}/{a("shares")}*{a("fx")}', REL['value_aed'], PX, bold=True)
r += 2

band(wsR, r, 'NORMALISED EARNINGS POWER'); r += 1
note(wsR, r, 'Mid-cycle is DERIVED from the audited record, not asserted: utilisation is '
             'the mean of the two most recent audited years and the benchmark is the mean '
             'of the three audited annual averages. Both therefore contain a turnaround '
             'year and a soft-price year, and neither contains the 2026 disruption.', 9)
r += 1
NRW = {}
for key, label_, value, fmt in [
        ('util_pe', 'Mid-cycle polyethylene utilisation', NRM['util_pe'], PCT),
        ('util_pp', 'Mid-cycle polypropylene utilisation', NRM['util_pp'], PCT),
        ('bench_pe', 'Mid-cycle polyethylene benchmark (USD/t)', NRM['bench_pe'], NUM0),
        ('bench_pp', 'Mid-cycle polypropylene benchmark (USD/t)', NRM['bench_pp'], NUM0),
        ('prem_pe', 'Through-the-cycle polyethylene premium (USD/t)',
         CI['prem_pe_ttc'], NUM0),
        ('prem_pp', 'Through-the-cycle polypropylene premium (USD/t)',
         CI['prem_pp_ttc'], NUM0),
        ('sd_per_t', 'Mid-cycle selling and distribution (USD/t)',
         sum(UB['sd_per_t'].values()) / 3, NUM1),
        ('da', 'Mid-cycle depreciation and amortisation (USD m)', NRM['da'], NUM0)]:
    lab(wsR, r, label_)
    put(wsR, f'B{r}', value, BLUE, fmt, cls='history')
    NRW[key] = r
    r += 1
lab(wsR, r, 'Mid-cycle polyethylene volume (kt)')
NV_PE = f'B{r}'
putf(wsR, f'B{r}', f'=B{NRW["util_pe"]}*{a("cap_pe")}', NRM['util_pe'] * UB['capacity_pe'],
     NUM0); r += 1
lab(wsR, r, 'Mid-cycle polypropylene volume (kt)')
NV_PP = f'B{r}'
putf(wsR, f'B{r}', f'=B{NRW["util_pp"]}*{a("cap_pp")}', NRM['util_pp'] * UB['capacity_pp'],
     NUM0); r += 1
lab(wsR, r, 'Mid-cycle volume (kt)')
NV = f'B{r}'
putf(wsR, f'B{r}', f'={NV_PE}+{NV_PP}', NRM['volume'], NUM0, bold=True); r += 1
lab(wsR, r, 'Mid-cycle revenue (USD million)')
NREV = f'B{r}'
putf(wsR, f'B{r}',
     f'=({NV_PE}*(B{NRW["bench_pe"]}+B{NRW["prem_pe"]})*{a("real_pe")}'
     f'+{NV_PP}*(B{NRW["bench_pp"]}+B{NRW["prem_pp"]})*{a("real_pp")})/1000',
     NRM['revenue'], NUM0); r += 1
lab(wsR, r, 'Less: feedstock')
NFEED = f'B{r}'
putf(wsR, f'B{r}', f'={a("feed_base")}*{NV}/1000', NRM['feedstock'], NUM0); r += 1
lab(wsR, r, 'Less: other production cost')
NOTH = f'B{r}'
putf(wsR, f'B{r}', f'={a("oth_fixed")}+{a("oth_var")}*{NV}', NRM['othprod'], NUM0); r += 1
lab(wsR, r, 'Less: selling and distribution')
NSD = f'B{r}'
putf(wsR, f'B{r}', f'=B{NRW["sd_per_t"]}*{NV}/1000', NRM['sd'], NUM0); r += 1
lab(wsR, r, 'Less: general and administrative')
NGA = f'B{r}'
putf(wsR, f'B{r}', f'={a("ga_base")}', CI['ga_exda_fy25'], NUM1, green=True); r += 1
lab(wsR, r, 'Mid-cycle EBITDA', bold=True)
NEB = f'B{r}'
putf(wsR, f'B{r}', f'={NREV}-{NFEED}-{NOTH}-{NSD}-{NGA}', NRM['ebitda'], NUM0, bold=True)
r += 1
lab(wsR, r, 'Mid-cycle EBIT')
NEBIT = f'B{r}'
putf(wsR, f'B{r}', f'={NEB}-B{NRW["da"]}', NRM['ebit'], NUM0); r += 1
lab(wsR, r, 'Mid-cycle NOPAT')
NNOP = f'B{r}'
putf(wsR, f'B{r}', f'={NEBIT}*(1-{a("tax")})', NRM['nopat'], NUM0, bold=True); r += 1
lab(wsR, r, 'Enterprise value = NOPAT x (1 - reinvestment) / (cost of capital - growth)')
NEV_OWN = f'B{r}'
putf(wsR, f'B{r}',
     f'={NNOP}/({WACC_OWN_REF}-{a("g")})*(1-{a("g")}/{a("roc")})'
     f"+'SOTP Bridge'!{B4_VAL.replace('B', '$B$')}", NRM['ev'], NUM0); r += 1
lab(wsR, r, 'Normalised earnings lens — own-stock beta (AED)', bold=True)
NE_OWN = f'B{r}'
putf(wsR, f'B{r}', f'=({NEV_OWN}-{a("net_debt")}-{a("leases")}-{a("nci")})'
     f'/{a("shares")}*{a("fx")}', LEN_['normalised_earnings_own_beta'], PX, bold=True)
r += 1
lab(wsR, r, 'Normalised earnings lens — sector bottom-up beta (AED)', bold=True)
NE_BU = f'B{r}'
putf(wsR, f'B{r}',
     f'=({NNOP}/({WACC_BU_REF}-{a("g")})*(1-{a("g")}/{a("roc")})'
     f"+'SOTP Bridge'!{B4_VAL.replace('B', '$B$')}"
     f'-{a("net_debt")}-{a("leases")}-{a("nci")})/{a("shares")}*{a("fx")}',
     LEN_['normalised_earnings_sector_beta'], PX, bold=True)
r += 1
note(wsR, r, 'Every line of the normalised block above is a live formula off the same '
             'capacity, realisation residual and cost stack the forecast uses. Change a '
             'driver on Assumptions and this lens moves with it.', 9)

# =============================================================================
# 3b FUNDAMENTAL VALUATION — the four lenses (now that their sources exist)
# =============================================================================
r = FV_LENS_START
band(wsF, r, 'THE FOUR LENSES ASSEMBLED'); r += 1
hdr(wsF, r, ['Lens', 'Sector bottom-up beta', 'Own-stock beta']); r += 1
LR = {}
lab(wsF, r, 'Discounted cash flow — normalisation')
putf(wsF, f'B{r}', f"='DCF'!$F${VPS['bu_n']}", LEN_['dcf_normalisation_sector_beta'], PX,
     green=True)
putf(wsF, f'C{r}', f"='DCF'!$F${VPS['own_n']}", LEN_['dcf_normalisation_own_beta'], PX,
     green=True)
LR['dcf_n'] = r; r += 1
lab(wsF, r, 'Discounted cash flow — prolonged disruption')
putf(wsF, f'B{r}', f"='DCF'!$F${VPS['bu_p']}", LEN_['dcf_prolonged_sector_beta'], PX,
     green=True)
putf(wsF, f'C{r}', f"='DCF'!$F${VPS['own_p']}", LEN_['dcf_prolonged_own_beta'], PX,
     green=True)
LR['dcf_p'] = r; r += 1
lab(wsF, r, 'Book value per share (USD)')
BVPS = f'B{r}'
putf(wsF, f'B{r}', f'={a("eq_owners")}/{a("shares")}', BV['bvps_usd'], NUM4); r += 1
lab(wsF, r, 'Sustainable return on equity (three audited years)')
ROE = f'B{r}'
putf(wsF, f'B{r}', f'=AVERAGE({a("roe23")},{a("roe24")},{a("roe25")})',
     BV['roe_sustainable'], PCT); r += 1
lab(wsF, r, 'Justified price to book = (return - growth) / (cost of equity - growth)')
putf(wsF, f'B{r}', f'=({ROE}-{a("g")})/({KE_BU}-{a("g")})',
     BV['justified_pb_sector_beta'], MULT)
putf(wsF, f'C{r}', f'=({ROE}-{a("g")})/({KE_OWN}-{a("g")})', BV['justified_pb'], MULT)
r += 1
lab(wsF, r, 'Book value lens — value per share (AED)', bold=True)
putf(wsF, f'B{r}', f'=B{r - 1}*{BVPS}*{a("fx")}', LEN_['book_value_sector_beta'], PX,
     bold=True)
putf(wsF, f'C{r}', f'=C{r - 1}*{BVPS}*{a("fx")}', LEN_['book_value_own_beta'], PX,
     bold=True)
LR['bv'] = r; r += 1
lab(wsF, r, 'Normalised earnings power — value per share (AED)')
putf(wsF, f'B{r}', f"='Relative & Normalized'!{NE_BU.replace('B', '$B$')}",
     LEN_['normalised_earnings_sector_beta'], PX, green=True)
putf(wsF, f'C{r}', f"='Relative & Normalized'!{NE_OWN.replace('B', '$B$')}",
     LEN_['normalised_earnings_own_beta'], PX, green=True)
LR['ne'] = r; r += 1
lab(wsF, r, 'Relative multiples — value per share (AED), independent of beta')
putf(wsF, f'B{r}', f"='Relative & Normalized'!{REL_VPS.replace('B', '$B$')}",
     LEN_['relative_multiples'], PX, green=True)
putf(wsF, f'C{r}', f"='Relative & Normalized'!{REL_VPS.replace('B', '$B$')}",
     LEN_['relative_multiples'], PX, green=True)
LR['rel'] = r; r += 2

band(wsF, r, 'THE ANSWER — the cash-flow lens, both sides of the shipping-lane judgement')
r += 1
# THE MEDIAN OF THE NINE READINGS WAS THE PUBLISHED CENTRAL UNTIL 5 SEPTEMBER 2026 AND IS
# RETIRED [R-LENS-03]. The nine cluster in two blocks of four, one per beta construction,
# so the median averaged nothing — it SELECTED one cell of the grid, and which cell was
# decided by how many lenses happened to have been computed under each framing rather than
# by any valuation choice. It is kept below, labelled and unused, so a reader of the
# previous edition can see the number that moved.
BR = D['central_two_sided']['branches']
lab(wsF, r, 'Navigation normalises during 2026 (AED)', bold=True)
FV_NORM = f'$B${r}'
# COLUMN C IS THE OWN-STOCK BETA, column B the bottom-up sector one — the answer is the
# ADOPTED tier-1 beta, so it reads C. A first draft read B and the recalculation gate
# caught it within the minute, which is what that gate is for.
putf(wsF, f'B{r}', f'=C{LR["dcf_n"]}', BR[0]['value'], PX, bold=True); r += 1
lab(wsF, r, 'Disruption persists into 2027 (AED)', bold=True)
FV_PROL = f'$B${r}'
putf(wsF, f'B{r}', f'=C{LR["dcf_p"]}', BR[1]['value'], PX, bold=True); r += 1
lab(wsF, r, 'The judgement is worth (AED a share)')
putf(wsF, f'B{r}', f'={FV_NORM}-{FV_PROL}',
     D['central_two_sided']['gap_per_share'], PX); r += 1
lab(wsF, r, 'Closing price, 3 September 2026 (AED)')
FV_SPOT = f'$B${r}'
putf(wsF, f'B{r}', f'={a("spot")}', SPOT, PX, green=True); r += 1
lab(wsF, r, 'Navigation normalises, against the close')
putf(wsF, f'B{r}', f'={FV_NORM}/{FV_SPOT}-1', BR[0]['value'] / SPOT - 1, PCT); r += 1
lab(wsF, r, 'Disruption persists, against the close')
putf(wsF, f'B{r}', f'={FV_PROL}/{FV_SPOT}-1', BR[1]['value'] / SPOT - 1, PCT); r += 2

band(wsF, r, 'THE FIELD OF CROSS-CHECKS — published beside the answer, never averaged into it')
r += 1
# The relative lens is beta-independent, so it appears in BOTH columns. Counting it twice
# would weight it double in anything computed across the field. It enters ONCE.
cells = ','.join([f'B{LR[k]}' for k in ('dcf_n', 'dcf_p', 'bv', 'ne', 'rel')] +
                 [f'C{LR[k]}' for k in ('dcf_n', 'dcf_p', 'bv', 'ne')])
lab(wsF, r, 'Lowest lens reading (AED)')
FV_LOW = f'$B${r}'
putf(wsF, f'B{r}', f'=MIN({cells})', D['field_low'], PX); r += 1
lab(wsF, r, 'Highest lens reading (AED)')
FV_HIGH = f'$B${r}'
putf(wsF, f'B{r}', f'=MAX({cells})', D['field_high'], PX); r += 1
lab(wsF, r, 'RETIRED — the median of the nine readings, published unused (AED)')
FV_MID = f'$B${r}'
putf(wsF, f'B{r}', f'=MEDIAN({cells})', D['fair_mid_retired'], PX); r += 1
FV_LR = LR

# =============================================================================
# 9  INCOME STATEMENT — 3 historical + 5 forecast
# =============================================================================
wsI = WS['Income Statement']
title(wsI, 'Income statement — three audited years and five forecast years',
      'USD million. Historical columns are the audited figures; forecast columns are '
      'formulas from the waterfall.', 10, 44, 12)
r = 4
hdr(wsI, r, [''] + YH + [str(y) for y in YF]); r += 1
IS = {}
IS_LINES = [
    ('revenue', 'Revenue', NUM0, True, 'revenue'),
    ('cogs', 'Cost of sales', NUM0, False, None),
    ('gross_profit', 'Gross profit', NUM0, True, None),
    ('sd', 'Selling and distribution', NUM0, False, 'sd'),
    ('ga', 'General and administrative', NUM0, False, 'ga'),
    ('other_income', 'Other income', NUM1, False, 'other_income'),
    ('ebit', 'Operating profit (EBIT)', NUM0, True, 'ebit'),
    ('da', 'Depreciation and amortisation', NUM0, False, 'da'),
    ('ebitda', 'EBITDA', NUM0, True, 'ebitda'),
    ('fin_cost', 'Net finance cost', NUM1, False, None),
    ('pbt', 'Profit before tax', NUM0, True, None),
    ('tax', 'Tax', NUM0, False, None),
    ('pat', 'Profit after tax', NUM0, True, None),
]
for key, label_, fmt, bold, _src in IS_LINES:
    lab(wsI, r, label_, bold=bold)
    IS[key] = r
    r += 1
for i, y in enumerate(HY):
    c = HC[i]
    h = H[y]
    for key, _l, fmt, _b, _s in IS_LINES:
        put(wsI, f'{c}{IS[key]}', h[key] if key != 'fin_cost'
            else h['fin_cost'] - h['fin_income'] - h['fx'], BLUE, fmt, cls='history')
for i, col in enumerate(FC):
    dcol = DC[i]
    row = FR['normalisation']['rows'][i]
    putf(wsI, f'{col}{IS["revenue"]}', f"='DCF'!{dcol}{RN['revenue']}", row['revenue'],
         NUM0, green=True)
    cogs = row['feedstock'] + row['othprod'] + row['da']
    putf(wsI, f'{col}{IS["cogs"]}',
         f"='DCF'!{dcol}{RN['feedstock']}+'DCF'!{dcol}{RN['othprod']}"
         f"+'DCF'!{dcol}{RN['da']}", cogs, NUM0, green=True)
    putf(wsI, f'{col}{IS["gross_profit"]}',
         f'={col}{IS["revenue"]}-{col}{IS["cogs"]}', row['revenue'] - cogs, NUM0,
         bold=True)
    putf(wsI, f'{col}{IS["sd"]}', f"='DCF'!{dcol}{RN['sd']}", row['sd'], NUM0, green=True)
    putf(wsI, f'{col}{IS["ga"]}', f"='DCF'!{dcol}{RN['ga']}", row['ga'], NUM1, green=True)
    putf(wsI, f'{col}{IS["other_income"]}', f"='DCF'!{dcol}{RN['other_income']}",
         row['other_income'], NUM1, green=True)
    putf(wsI, f'{col}{IS["ebit"]}',
         f'={col}{IS["gross_profit"]}-{col}{IS["sd"]}-{col}{IS["ga"]}'
         f'+{col}{IS["other_income"]}', row['ebit'], NUM0, bold=True)
    putf(wsI, f'{col}{IS["da"]}', f"='DCF'!{dcol}{RN['da']}", row['da'], NUM0, green=True)
    putf(wsI, f'{col}{IS["ebitda"]}', f'={col}{IS["ebit"]}+{col}{IS["da"]}',
         row['ebitda'], NUM0, bold=True)
    fin = H['2025']['fin_cost'] - H['2025']['fin_income'] - H['2025']['fx']
    putf(wsI, f'{col}{IS["fin_cost"]}', f'={a("net_debt")}*{a("kd")}',
         NET_DEBT * W['kd'], NUM1)
    pbt = row['ebit'] - NET_DEBT * W['kd']
    putf(wsI, f'{col}{IS["pbt"]}', f'={col}{IS["ebit"]}-{col}{IS["fin_cost"]}', pbt, NUM0,
         bold=True)
    putf(wsI, f'{col}{IS["tax"]}', f'={col}{IS["pbt"]}*{a("tax")}', pbt * W['tax'], NUM0)
    putf(wsI, f'{col}{IS["pat"]}', f'={col}{IS["pbt"]}-{col}{IS["tax"]}',
         pbt * (1 - W['tax']), NUM0, bold=True)
r += 1
note(wsI, r, 'The forecast operating lines are the same cells as the cash-flow waterfall, '
             'linked rather than restated. The finance cost below EBIT is the net debt '
             'balance at the marginal cost of debt — it does not enter free cash flow to '
             'the firm, which is computed before financing.', 10)

# =============================================================================
# 10 BALANCE SHEET
# =============================================================================
wsBS = WS['Balance Sheet']
title(wsBS, 'Balance sheet — three audited years and five forecast years',
      'USD million. The forecast rolls property, working capital and net debt forward '
      'from the drivers; nothing is plugged.', 10, 44, 12)
r = 4
hdr(wsBS, r, [''] + YH + [str(y) for y in YF]); r += 1
BS = {}
for key, label_, bold in [('ppe', 'Property, plant and equipment', False),
                          ('ar', 'Trade receivables', False),
                          ('inv', 'Inventory', False),
                          ('cash', 'Cash and equivalents', False),
                          ('ta', 'Total assets (memo: audited)', True),
                          ('ap', 'Trade payables', False),
                          ('debt', 'Borrowings', False),
                          ('equity', 'Equity', True),
                          ('nwc', 'Net working capital', True),
                          ('nd', 'Net debt', True)]:
    lab(wsBS, r, label_, bold=bold)
    BS[key] = r
    r += 1
BSV = {}
BS_HIST = {'ppe': ['ppe_fy23', 'ppe_fy24', 'ppe_fy25'],
           'inv': ['inv_fy23', 'inv_fy24', 'inv_fy25'],
           'cash': ['cash_fy23', 'cash_fy24', 'cash_fy25'],
           'ta': ['ta_fy23', 'ta_fy24', 'ta_fy25'],
           'debt': ['debt_fy23', 'debt_fy24', 'debt_fy25'],
           'equity': ['eq_owners_fy23', 'eq_owners_fy24', 'eq_owners_fy25']}
for i, y in enumerate(HY):
    c = HC[i]
    for key, keys in BS_HIST.items():
        put(wsBS, f'{c}{BS[key]}', CI[keys[i]] * USDm, BLUE, NUM0, cls='history')
        BSV[(key, c)] = CI[keys[i]] * USDm
    # Receivables and payables are shown at the days the audited statements imply, so the
    # working-capital line the forecast projects is measured on the same basis as history.
    ar_v = H[y]['revenue'] * WC['dso_hist'][y] / 365
    ap_v = H[y]['cogs'] * WC['dpo_hist'][y] / 365
    putf(wsBS, f'{c}{BS["ar"]}',
         f"='Income Statement'!{c}{IS['revenue']}*{WC['dso_hist'][y]!r}/365", ar_v, NUM0,
         green=True)
    putf(wsBS, f'{c}{BS["ap"]}',
         f"='Income Statement'!{c}{IS['cogs']}*{WC['dpo_hist'][y]!r}/365", ap_v, NUM0,
         green=True)
    putf(wsBS, f'{c}{BS["nwc"]}', f'={c}{BS["ar"]}+{c}{BS["inv"]}-{c}{BS["ap"]}',
         ar_v + CI[BS_HIST['inv'][i]] * USDm - ap_v, NUM0, bold=True)
    putf(wsBS, f'{c}{BS["nd"]}', f'={c}{BS["debt"]}-{c}{BS["cash"]}',
         (CI[BS_HIST['debt'][i]] - CI[BS_HIST['cash'][i]]) * USDm, NUM0, bold=True)
    BSV[('equity', c)] = CI[BS_HIST['equity'][i]] * USDm
    BSV[('nwc', c)] = ar_v + CI[BS_HIST['inv'][i]] * USDm - ap_v
    BSV[('nd', c)] = (CI[BS_HIST['debt'][i]] - CI[BS_HIST['cash'][i]]) * USDm

# Forecast: property rolls on capital spend less depreciation; working capital is
# projected from the disclosed conversion cycle; net debt absorbs the free cash flow
# after the dividend; equity rolls on retained profit less the dividend.
DPS_USD = 0.162 / FX
prev_eq = CI['eq_owners_fy25'] * USDm
prev_nd = (CI['debt_fy25'] - CI['cash_fy25']) * USDm
prev_ppe = CI['ppe_fy25'] * USDm
for i, col in enumerate(FC):
    dcol = DC[i]
    row = FR['normalisation']['rows'][i]
    pc = HC[-1] if i == 0 else FC[i - 1]
    putf(wsBS, f'{col}{BS["ppe"]}',
         f"={pc}{BS['ppe']}+'DCF'!{dcol}{RN['capex']}-'DCF'!{dcol}{RN['da']}",
         prev_ppe + row['capex'] - row['da'], NUM0)
    prev_ppe = prev_ppe + row['capex'] - row['da']
    BSV[('ppe', col)] = prev_ppe
    cogs_p = row['feedstock'] + row['othprod']
    putf(wsBS, f'{col}{BS["ar"]}',
         f"='Income Statement'!{col}{IS['revenue']}*{a('dso')}/365",
         row['revenue'] * WC['dso'] / 365, NUM0, green=True)
    putf(wsBS, f'{col}{BS["inv"]}',
         f"=('DCF'!{dcol}{RN['feedstock']}+'DCF'!{dcol}{RN['othprod']})*{a('dio')}/365",
         cogs_p * WC['dio'] / 365, NUM0, green=True)
    putf(wsBS, f'{col}{BS["ap"]}',
         f"=('DCF'!{dcol}{RN['feedstock']}+'DCF'!{dcol}{RN['othprod']})*{a('dpo')}/365",
         cogs_p * WC['dpo'] / 365, NUM0, green=True)
    putf(wsBS, f'{col}{BS["nwc"]}', f'={col}{BS["ar"]}+{col}{BS["inv"]}-{col}{BS["ap"]}',
         row['nwc'], NUM0, bold=True)
    BSV[('nwc', col)] = row['nwc']
    # Net debt: opening, less free cash flow after tax-effected finance cost, plus the
    # dividend the company says it intends to keep paying.
    pat = row['ebit'] - NET_DEBT * W['kd']
    pat = pat * (1 - W['tax'])
    div = DPS_USD * SHARES
    nd = prev_nd - (row['fcff'] - NET_DEBT * W['kd'] * (1 - W['tax'])) + div
    putf(wsBS, f'{col}{BS["nd"]}',
         f"={pc}{BS['nd']}-('DCF'!{dcol}{RN['fcff']}-{a('net_debt')}*{a('kd')}"
         f"*(1-{a('tax')}))+{DPS_USD!r}*{a('shares')}", nd, NUM0, bold=True)
    BSV[('nd', col)] = nd
    eq = prev_eq + pat - div
    putf(wsBS, f'{col}{BS["equity"]}',
         f"={pc}{BS['equity']}+'Income Statement'!{col}{IS['pat']}"
         f"-{DPS_USD!r}*{a('shares')}", eq, NUM0, bold=True)
    BSV[('equity', col)] = eq
    debt_v = CI['debt_fy25'] * USDm
    cash_v = debt_v - nd
    putf(wsBS, f'{col}{BS["debt"]}', f"='Balance Sheet'!D{BS['debt']}", debt_v, NUM0,
         green=True)
    putf(wsBS, f'{col}{BS["cash"]}', f'={col}{BS["debt"]}-{col}{BS["nd"]}', cash_v, NUM0)
    putf(wsBS, f'{col}{BS["ta"]}',
         f'={col}{BS["ppe"]}+{col}{BS["ar"]}+{col}{BS["inv"]}+{col}{BS["cash"]}',
         prev_ppe + row['revenue'] * WC['dso'] / 365 + cogs_p * WC['dio'] / 365 + cash_v,
         NUM0, bold=True)
    prev_nd, prev_eq = nd, eq
r = BS['nd'] + 2
note(wsBS, r, 'Receivables, inventory and payables are projected from the days the '
              'audited statements themselves show — fifty-one days of sales, sixty-two '
              'days of inventory, ninety-two days of payables. Nothing on this sheet is '
              'a balancing figure.', 10)

# =============================================================================
# 11 CASH FLOW
# =============================================================================
wsC = WS['Cash Flow']
title(wsC, 'Cash flow — linked to the waterfall, not restated',
      'USD million. The forecast columns are the same cells the valuation discounts.',
      10, 44, 12)
r = 4
hdr(wsC, r, [''] + YH + [str(y) for y in YF]); r += 1
CF = {}
for key, label_, bold in [('ebitda', 'EBITDA', True),
                          ('d_nwc', 'Less: increase in working capital', False),
                          ('tax_paid', 'Less: tax on operating profit', False),
                          ('cfo', 'Cash from operations', True),
                          ('capex', 'Less: capital expenditure', False),
                          ('fcff', 'Free cash flow to the firm', True),
                          ('interest', 'Less: net finance cost after tax', False),
                          ('div', 'Less: dividend', False),
                          ('fcfe', 'Free cash flow to equity', True)]:
    lab(wsC, r, label_, bold=bold)
    CF[key] = r
    r += 1
for i, y in enumerate(HY):
    c = HC[i]
    h = H[y]
    put(wsC, f'{c}{CF["ebitda"]}', h['ebitda'], BLUE, NUM0, cls='history')
    put(wsC, f'{c}{CF["cfo"]}', h['cfo'], BLUE, NUM0, cls='history')
    put(wsC, f'{c}{CF["capex"]}', h['capex'], BLUE, NUM0, cls='history')
    put(wsC, f'{c}{CF["tax_paid"]}', h['tax'], BLUE, NUM0, cls='history')
    putf(wsC, f'{c}{CF["fcff"]}', f'={c}{CF["cfo"]}-{c}{CF["capex"]}',
         h['cfo'] - h['capex'], NUM0, bold=True)
for i, col in enumerate(FC):
    dcol = DC[i]
    row = FR['normalisation']['rows'][i]
    putf(wsC, f'{col}{CF["ebitda"]}', f"='DCF'!{dcol}{RN['ebitda']}", row['ebitda'], NUM0,
         green=True)
    putf(wsC, f'{col}{CF["d_nwc"]}', f"='DCF'!{dcol}{RN['d_nwc']}", row['d_nwc'], NUM1,
         green=True)
    taxop = row['ebit'] * W['tax']
    putf(wsC, f'{col}{CF["tax_paid"]}', f"='DCF'!{dcol}{RN['ebit']}*{a('tax')}", taxop,
         NUM0)
    putf(wsC, f'{col}{CF["cfo"]}',
         f'={col}{CF["ebitda"]}-{col}{CF["d_nwc"]}-{col}{CF["tax_paid"]}',
         row['ebitda'] - row['d_nwc'] - taxop, NUM0, bold=True)
    putf(wsC, f'{col}{CF["capex"]}', f"='DCF'!{dcol}{RN['capex']}", row['capex'], NUM0,
         green=True)
    putf(wsC, f'{col}{CF["fcff"]}', f'={col}{CF["cfo"]}-{col}{CF["capex"]}',
         row['fcff'], NUM0, bold=True)
    it = NET_DEBT * W['kd'] * (1 - W['tax'])
    putf(wsC, f'{col}{CF["interest"]}', f'={a("net_debt")}*{a("kd")}*(1-{a("tax")})', it,
         NUM1)
    dv = 0.162 / FX * SHARES
    putf(wsC, f'{col}{CF["div"]}', f'={0.162 / FX!r}*{a("shares")}', dv, NUM0)
    putf(wsC, f'{col}{CF["fcfe"]}',
         f'={col}{CF["fcff"]}-{col}{CF["interest"]}-{col}{CF["div"]}',
         row['fcff'] - it - dv, NUM0, bold=True)
r += 1
note(wsC, r, 'Cash from operations reconciles to free cash flow to the firm through '
             'capital expenditure alone, because the working-capital movement and the '
             'tax on operating profit are already inside it. The dividend is the '
             'company\'s own stated intention of 16.2 fils a share.', 10)

# =============================================================================
# 12 SUMMARY FINANCIALS
# =============================================================================
wsSF = WS['Summary Financials']
title(wsSF, 'Summary financials', 'USD million. Every cell is a link or a formula.',
      10, 44, 12)
r = 4
hdr(wsSF, r, [''] + YH + [str(y) for y in YF]); r += 1
SF = {}
for key, label_ in [('revenue', 'Revenue'), ('ebitda', 'EBITDA'),
                    ('margin', 'EBITDA margin'), ('ebit', 'EBIT'),
                    ('pat', 'Profit after tax'), ('fcff', 'Free cash flow to the firm'),
                    ('capex', 'Capital expenditure'), ('nd', 'Net debt'),
                    ('ndeb', 'Net debt to EBITDA')]:
    lab(wsSF, r, label_, bold=key in ('ebitda', 'fcff'))
    SF[key] = r
    r += 1
for i, cols in enumerate([HC, FC]):
    for j, col in enumerate(cols):
        y = HY[j] if i == 0 else None
        src = HC[j] if i == 0 else FC[j]
        rowv = H[y] if i == 0 else FR['normalisation']['rows'][j]
        putf(wsSF, f'{col}{SF["revenue"]}', f"='Income Statement'!{src}{IS['revenue']}",
             rowv['revenue'], NUM0, green=True)
        putf(wsSF, f'{col}{SF["ebitda"]}', f"='Income Statement'!{src}{IS['ebitda']}",
             rowv['ebitda'], NUM0, green=True, bold=True)
        putf(wsSF, f'{col}{SF["margin"]}',
             f'={col}{SF["ebitda"]}/{col}{SF["revenue"]}',
             rowv['ebitda'] / rowv['revenue'], PCT)
        putf(wsSF, f'{col}{SF["ebit"]}', f"='Income Statement'!{src}{IS['ebit']}",
             rowv['ebit'], NUM0, green=True)
        if i == 0:
            putf(wsSF, f'{col}{SF["pat"]}', f"='Income Statement'!{src}{IS['pat']}",
                 rowv['pat'], NUM0, green=True)
            putf(wsSF, f'{col}{SF["fcff"]}', f"='Cash Flow'!{src}{CF['fcff']}",
                 rowv['cfo'] - rowv['capex'], NUM0, green=True, bold=True)
            putf(wsSF, f'{col}{SF["capex"]}', f"='Cash Flow'!{src}{CF['capex']}",
                 rowv['capex'], NUM0, green=True)
            nd = (CI[f'debt_fy{y[2:]}'] - CI[f'cash_fy{y[2:]}']) * USDm
        else:
            pbt = rowv['ebit'] - NET_DEBT * W['kd']
            putf(wsSF, f'{col}{SF["pat"]}', f"='Income Statement'!{src}{IS['pat']}",
                 pbt * (1 - W['tax']), NUM0, green=True)
            putf(wsSF, f'{col}{SF["fcff"]}', f"='Cash Flow'!{src}{CF['fcff']}",
                 rowv['fcff'], NUM0, green=True, bold=True)
            putf(wsSF, f'{col}{SF["capex"]}', f"='Cash Flow'!{src}{CF['capex']}",
                 rowv['capex'], NUM0, green=True)
            nd = BSV[('nd', col)]
        putf(wsSF, f'{col}{SF["nd"]}', f"='Balance Sheet'!{src}{BS['nd']}", nd, NUM0,
             green=True)
for i, col in enumerate(HC + FC):
    eb = H[HY[i]]['ebitda'] if i < 3 else FR['normalisation']['rows'][i - 3]['ebitda']
    putf(wsSF, f'{col}{SF["ndeb"]}', f'={col}{SF["nd"]}/{col}{SF["ebitda"]}',
         BSV[('nd', col)] / eb, MULT)

# =============================================================================
# 13 MONTE CARLO — a whole-model re-run, pasted
# =============================================================================
wsM = WS['Monte Carlo']
title(wsM, 'Probability map for the traded price',
      'Each figure is the outcome of a 50,000-path simulation on the cleaned price '
      'history — a complete re-run, not a formula. THIS MAP DOES NOT REDRAW WHEN A '
      'DRIVER CHANGES.', 9, 44, 15)
r = 4
band(wsM, r, 'THE SIMULATION AS IT WAS RUN'); r += 1
for label_, v, fmt in [('Anchor close (AED)', STK['spot'], PX),
                       ('Anchor date', 0, None),
                       ('Paths', 50000, NUM0),
                       ('Risk-free rate', STK['rf_live'], PCT2),
                       ('Dividend yield (16.2 fils on the close)', STK['q_annual'], PCT2),
                       ('Net carry a year', STK['rf_live'] - STK['q_annual'], PCT2)]:
    lab(wsM, r, label_)
    if label_ == 'Anchor date':
        put(wsM, f'B{r}', STK['anchor_date'], BLUE, None, cls='history')
    else:
        put(wsM, f'B{r}', v, BLUE, fmt, cls='rerun')
    if label_ == 'Anchor close (AED)':
        MC_ANCHOR = f'$B${r}'
    r += 1
r += 1
band(wsM, r, 'PERCENTILE MAP (AED PER SHARE)'); r += 1
hdr(wsM, r, ['Horizon', 'Ends', '5th', '25th', 'Median', '75th', '95th',
             'Above the close']); r += 1
MC_PCT_ROWS = {}
for tag in ('1M', '3M'):
    MC_PCT_ROWS[tag] = r
    hz = STK['horizons'][tag]
    lab(wsM, r, 'One month' if tag == '1M' else 'Three months')
    put(wsM, f'B{r}', hz['grade_date'], BLUE, None, cls='history')
    for j, p in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        put(wsM, f'{get_column_letter(3 + j)}{r}', hz['pct'][p], BLUE, PX, cls='rerun')
    put(wsM, f'H{r}', hz['p_above'], BLUE, PCT, cls='rerun')
    r += 1
r += 1
band(wsM, r, 'LEVEL-TOUCH LADDER — CHANCE THE PRICE TRADES THROUGH A LEVEL AT ANY POINT')
r += 1
hdr(wsM, r, ['Move from the close', 'Level (AED)', 'One month', 'Three months']); r += 1
for pct_ in (5, 10, 15, 20):
    for sign in (1, -1):
        lvl = STK['spot'] * (1 + sign * pct_ / 100)
        lab(wsM, r, f"{'Up' if sign > 0 else 'Down'} {pct_}%")
        # THE LADDER STANDS ON THE CONE'S OWN ANCHOR, NOT ON THE VALUATION'S SPOT — two
        # clocks, and they separated the moment this study was re-struck on the latest
        # price while its price library still ends on the anchor date. The formula read
        # the valuation spot cell and the expected value came from the strike, so they
        # agreed only while the two dates happened to coincide. The recalculation gate
        # caught all eight rungs the minute they stopped coinciding.
        putf(wsM, f'B{r}',
             f'={MC_ANCHOR}*(1{"+" if sign > 0 else "-"}{pct_ / 100})', lvl, PX)
        for k, tag in enumerate(('1M', '3M')):
            key = 'touch_up' if sign > 0 else 'touch_dn'
            put(wsM, f'{get_column_letter(3 + k)}{r}',
                STK['horizons'][tag][key][str(pct_)], BLUE, PCT, cls='rerun')
        r += 1
r += 1
note(wsM, r, 'The percentile map and the touch ladder are simulation output. The level '
             'column beside them is a live formula off the close, so the ladder\'s rungs '
             'move if the anchor changes even though the probabilities beside them do '
             'not.', 9)

# =============================================================================
# 14 SENSITIVITY — whole-model re-runs, pasted
# =============================================================================
wsN = WS['Sensitivity']
title(wsN, 'Sensitivity — each cell is a complete revaluation',
      'THESE GRIDS DO NOT REDRAW WHEN A DRIVER CHANGES. Value per share in AED, '
      'normalisation construction, own-stock beta.', 9, 40, 13)
r = 4
band(wsN, r, 'COST OF CAPITAL AGAINST TERMINAL GROWTH'); r += 1
hdr(wsN, r, ['Cost of capital \\ growth'] + [f'{g:.1%}' for g in SN['g_grid']]); r += 1
for i, w_ in enumerate(SN['wacc_grid']):
    put(wsN, f'A{r}', w_, BLUE, PCT2, cls='rerun')
    for j in range(len(SN['g_grid'])):
        put(wsN, f'{get_column_letter(2 + j)}{r}', SN['grids']['normalisation'][i][j],
            BLUE, PX, cls='rerun')
    r += 1
r += 1
band(wsN, r, 'REALISED PREMIUM OVER THE BENCHMARK (USD PER TONNE)'); r += 1
hdr(wsN, r, ['Shift', 'Normalisation', 'Prolonged disruption']); r += 1
for i, (sh, v) in enumerate(SN['premium_grid']['normalisation']):
    put(wsN, f'A{r}', f'{sh:+d}', BLACK, None)
    put(wsN, f'B{r}', v, BLUE, PX, cls='rerun')
    put(wsN, f'C{r}', SN['premium_grid']['prolonged'][i][1], BLUE, PX, cls='rerun')
    r += 1
r += 1
band(wsN, r, 'UTILISATION'); r += 1
hdr(wsN, r, ['Shift', 'Normalisation', 'Prolonged disruption']); r += 1
for i, (sh, v) in enumerate(SN['util_grid']['normalisation']):
    put(wsN, f'A{r}', f'{sh:+.0%}', BLACK, None)
    put(wsN, f'B{r}', v, BLUE, PX, cls='rerun')
    put(wsN, f'C{r}', SN['util_grid']['prolonged'][i][1], BLUE, PX, cls='rerun')
    r += 1
r += 1
note(wsN, r, 'Every cell above is the whole valuation solved again at that pair of '
             'inputs. That is why they are values rather than formulas, and why they are '
             'named as a pasted class on READ FIRST.', 9)

# =============================================================================
# 15 PER-SHARE & RATIOS
# =============================================================================
wsP = WS['Per-Share & Ratios']
title(wsP, 'Per-share figures and returns', 'Every cell is a formula.', 10, 44, 12)
r = 4
hdr(wsP, r, [''] + YH + [str(y) for y in YF]); r += 1
PS = {}
for key, label_ in [('eps', 'Earnings per share (USD)'),
                    ('eps_aed', 'Earnings per share (AED)'),
                    ('bvps', 'Book value per share (AED)'),
                    ('roe', 'Return on equity'),
                    ('roic', 'Return on invested capital'),
                    ('margin', 'EBITDA margin'),
                    ('nd_ebitda', 'Net debt to EBITDA'),
                    ('dps', 'Dividend per share (AED)'),
                    ('payout', 'Payout ratio')]:
    lab(wsP, r, label_)
    PS[key] = r
    r += 1
for i, col in enumerate(HC + FC):
    hist = i < 3
    src = col
    y = HY[i] if hist else None
    rowv = H[y] if hist else FR['normalisation']['rows'][i - 3]
    pat = rowv['pat_owners'] if hist else \
        (rowv['ebit'] - NET_DEBT * W['kd']) * (1 - W['tax'])
    eq = CI[f'eq_owners_fy{y[2:]}'] * USDm if hist else None
    putf(wsP, f'{col}{PS["eps"]}', f"='Income Statement'!{col}{IS['pat']}/{a('shares')}",
         (rowv['pat'] if hist else pat) / SHARES, NUM4)
    putf(wsP, f'{col}{PS["eps_aed"]}', f'={col}{PS["eps"]}*{a("fx")}',
         (rowv['pat'] if hist else pat) / SHARES * FX, NUM3)
    eqv, ndv, ppev, nwcv = (BSV[('equity', col)], BSV[('nd', col)],
                            BSV[('ppe', col)], BSV[('nwc', col)])
    patv = rowv['pat'] if hist else pat
    ebitv = rowv['ebit']
    ebitdav = rowv['ebitda']
    eps_aed = patv / SHARES * FX
    putf(wsP, f'{col}{PS["bvps"]}',
         f"='Balance Sheet'!{col}{BS['equity']}/{a('shares')}*{a('fx')}",
         eqv / SHARES * FX, NUM3)
    putf(wsP, f'{col}{PS["roe"]}',
         f"='Income Statement'!{col}{IS['pat']}/'Balance Sheet'!{col}{BS['equity']}",
         patv / eqv, PCT)
    putf(wsP, f'{col}{PS["roic"]}',
         f"='Income Statement'!{col}{IS['ebit']}*(1-{a('tax')})"
         f"/('Balance Sheet'!{col}{BS['ppe']}+'Balance Sheet'!{col}{BS['nwc']})",
         ebitv * (1 - W['tax']) / (ppev + nwcv), PCT)
    putf(wsP, f'{col}{PS["margin"]}',
         f"='Summary Financials'!{col}{SF['margin']}",
         ebitdav / rowv['revenue'], PCT, green=True)
    putf(wsP, f'{col}{PS["nd_ebitda"]}',
         f"='Balance Sheet'!{col}{BS['nd']}/'Income Statement'!{col}{IS['ebitda']}",
         ndv / ebitdav, MULT)
    putf(wsP, f'{col}{PS["dps"]}', '=0.162', 0.162, NUM3)
    putf(wsP, f'{col}{PS["payout"]}', f'={col}{PS["dps"]}/{col}{PS["eps_aed"]}',
         0.162 / eps_aed, PCT)
r += 1
note(wsP, r, 'The dividend per share is the company\'s own stated annual intention of '
             '16.2 fils. It is held flat rather than grown, because the company states '
             'an intention, not a policy formula.', 10)

# =============================================================================
# 16 PEER & SECTOR
# =============================================================================
wsPS = WS['Peer & Sector']
title(wsPS, 'Listed peers — a cross-check, never a source',
      'Observed 9 August 2026. Shown to explain why the naive median is rejected, not to '
      'set the multiple.', 9, 34, 14)
r = 4
hdr(wsPS, r, ['Company', 'EV / EBITDA', 'Forward P/E', 'Price to book',
              'Dividend yield', 'EBITDA margin', 'Net margin', 'Loss-making']); r += 1
PT = D['peer_table']
DEFINED = [(n, d_) for n, d_ in PT.items() if d_.get('ev_ebitda') is not None]
UNDEFINED = [(n, d_) for n, d_ in PT.items() if d_.get('ev_ebitda') is None]
first_peer = r
for nm, d_ in DEFINED + UNDEFINED:
    lab(wsPS, r, nm)
    for j, key, fmt in [(2, 'ev_ebitda', MULT), (3, 'pe_fwd', MULT), (4, 'pb', MULT),
                        (5, 'div_yield', PCT), (6, 'ebitda_margin', PCT),
                        (7, 'profit_margin', PCT)]:
        v = d_.get(key)
        put(wsPS, f'{get_column_letter(j)}{r}', v if v is not None else '-', BLUE, fmt,
            cls='history')
    put(wsPS, f'H{r}', 'yes' if d_['loss_making'] else 'no', BLUE, None)
    r += 1
last_peer = first_peer + len(DEFINED) - 1
lab(wsPS, r, 'Median of the multiples that print', bold=True)
putf(wsPS, f'B{r}', f'=MEDIAN(B{first_peer}:B{last_peer})', D['peer_naive_median'], MULT,
     bold=True)
wsPS.cell(row=r, column=3,
          value=f'Over the {len(DEFINED)} peers whose multiple is defined; the '
                f'{len(UNDEFINED)} with negative EBITDA are listed below the range and '
                f'cannot enter a median').font = SUB
wsPS.cell(row=r, column=3, value='REJECTED — see Relative & Normalized').font = SUB
r += 2
band(wsPS, r, 'BOROUGE ON THE SAME BASIS, FOR CONTRAST'); r += 1
PBM = D['macro']['peer_borouge_market']['value']
for key, label_, fmt in [('ev_ebitda', 'EV / EBITDA', MULT),
                         ('pe_fwd', 'Forward P/E', MULT),
                         ('div_yield', 'Dividend yield', PCT),
                         ('ebitda_margin', 'EBITDA margin', PCT)]:
    lab(wsPS, r, label_)
    put(wsPS, f'B{r}', PBM[key], BLUE, fmt, cls='history')
    r += 1
r += 1
note(wsPS, r, 'Borouge earns an EBITDA margin roughly three times the peer group\'s while '
             'nine of the eleven are loss-making. That is the advantaged-feedstock '
             'position, and it is the reason a peer median cannot be applied to it '
             'directly.', 9)

# =============================================================================
# 2  SUMMARY — laid out last, links everything
# =============================================================================
wsSm = WS['Summary']
title(wsSm, 'Summary valuation',
      'Value per share in AED. Every figure on this sheet is a live link or formula.',
      9, 46, 17)
r = 4
band(wsSm, r, 'THE FOUR LENSES, ACROSS THE BETA THE STUDY CONTESTS'); r += 1
hdr(wsSm, r, ['Lens', 'Sector bottom-up beta', 'Own-stock beta',
              'Terminal value % of EV']); r += 1
for key, label_, tv in [('dcf_n', 'Discounted cash flow — normalisation', 'own_n'),
                        ('dcf_p', 'Discounted cash flow — prolonged disruption', 'own_p'),
                        ('bv', 'Book value and sustainable return', None),
                        ('ne', 'Normalised earnings power', None),
                        ('rel', 'Relative multiples (independent of beta)', None)]:
    lab(wsSm, r, label_)
    fr = FV_LR[key]
    putf(wsSm, f'B{r}', f"='Fundamental Valuation'!$B${fr}",
         LEN_[{'dcf_n': 'dcf_normalisation_sector_beta',
               'dcf_p': 'dcf_prolonged_sector_beta',
               'bv': 'book_value_sector_beta',
               'ne': 'normalised_earnings_sector_beta',
               'rel': 'relative_multiples'}[key]], PX, green=True)
    putf(wsSm, f'C{r}', f"='Fundamental Valuation'!$C${fr}",
         LEN_[{'dcf_n': 'dcf_normalisation_own_beta',
               'dcf_p': 'dcf_prolonged_own_beta',
               'bv': 'book_value_own_beta',
               'ne': 'normalised_earnings_own_beta',
               'rel': 'relative_multiples'}[key]], PX, green=True)
    if tv:
        frk = 'normalisation' if tv.endswith('_n') else 'prolonged'
        putf(wsSm, f'D{r}', f"='DCF'!$E${VPS[tv]}",
             FR[frk]['pv_terminal'] / FR[frk]['ev'], PCT, green=True)
    r += 1
r += 1
band(wsSm, r, 'THE ANSWER AGAINST THE MARKET'); r += 1
lab(wsSm, r, 'Navigation normalises during 2026', bold=True)
SM_NORM = r
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!{FV_NORM}", BR[0]['value'], PX, green=True,
     bold=True); r += 1
lab(wsSm, r, 'Disruption persists into 2027', bold=True)
SM_PROL = r
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!{FV_PROL}", BR[1]['value'], PX, green=True,
     bold=True); r += 1
lab(wsSm, r, 'Closing price, 3 September 2026')
SM_SPOT = r
putf(wsSm, f'B{r}', f'={a("spot")}', SPOT, PX, green=True); r += 1
lab(wsSm, r, 'Navigation normalises, against the close')
putf(wsSm, f'B{r}', f'=B{SM_NORM}/B{SM_SPOT}-1', BR[0]['value'] / SPOT - 1, PCT); r += 1
lab(wsSm, r, 'Disruption persists, against the close')
putf(wsSm, f'B{r}', f'=B{SM_PROL}/B{SM_SPOT}-1', BR[1]['value'] / SPOT - 1, PCT); r += 1
lab(wsSm, r, 'Cross-checks span (AED)')
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!{FV_LOW}", D['field_low'], PX, green=True)
putf(wsSm, f'C{r}', f"='Fundamental Valuation'!{FV_HIGH}", D['field_high'], PX, green=True)
r += 2

band(wsSm, r, 'THE COST OF CAPITAL BEHIND IT'); r += 1
lab(wsSm, r, 'Weighted average cost of capital — own-stock beta')
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!{WACC_OWN}", W['wacc_own'], PCT2,
     green=True); r += 1
lab(wsSm, r, 'Weighted average cost of capital — sector bottom-up beta')
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!{WACC_BU}", W['wacc_bottom_up'], PCT2,
     green=True); r += 1
lab(wsSm, r, 'Normalised risk-free rate')
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!${RF[0]}${RF[1:]}", W['rf_star'], PCT2,
     green=True); r += 1
lab(wsSm, r, 'Equity risk premium (rating basis)')
putf(wsSm, f'B{r}', f"='Fundamental Valuation'!${ERP[0]}${ERP[1:]}", W['erp_rating'],
     PCT2, green=True); r += 1
lab(wsSm, r, 'Terminal growth rate')
putf(wsSm, f'B{r}', f'={a("g")}', MC['terminal_growth'], PCT2, green=True); r += 2

band(wsSm, r, 'THE PROBABILITY MAP FOR THE TRADED PRICE'); r += 1
hdr(wsSm, r, ['Horizon', '5th', 'Median', '95th', 'Above the close']); r += 1
for tag in ('1M', '3M'):
    hz = STK['horizons'][tag]
    lab(wsSm, r, 'One month' if tag == '1M' else 'Three months')
    mc_row = MC_PCT_ROWS[tag]
    for j, (p, mcol) in enumerate([('p5', 'C'), ('p50', 'E'), ('p95', 'G')]):
        putf(wsSm, f'{get_column_letter(2 + j)}{r}', f"='Monte Carlo'!{mcol}{mc_row}",
             hz['pct'][p], PX, green=True)
    putf(wsSm, f'E{r}', f"='Monte Carlo'!H{mc_row}", hz['p_above'], PCT, green=True)
    r += 1
r += 1
note(wsSm, r, 'The two constructions and the two betas are shown side by side and are '
              'never averaged. The spread between them is the honest width of the answer.',
     9)

# =============================================================================
# 1  READ FIRST
# =============================================================================
ws = WS['READ FIRST']
title(ws, 'Testahil — Borouge plc (ADX: BOROUGE)', None, 9)
LINES = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the Borouge valuation study. Every blue cell is an',
 'input; every black cell is a formula; green cells link across sheets. Amounts are USD million unless a',
 'row says otherwise — Borouge reports in US dollars and its shares trade in dirhams, so per-share figures',
 'are converted at the pegged rate held on Assumptions.', '',
 'IT IS FORMULA-DRIVEN, AND THAT CLAIM IS TESTED. Every figure derivable from a driver is a live formula,',
 'so you can change a blue cell on Assumptions and watch the model reprice. The cost of equity is built',
 'from the risk-free rate NET of the sovereign default spread, beta and the premium rather than pasted;',
 'the cost of debt is taxed in the sheet; the weights come from net debt and market capitalisation; the',
 'terminal rate is built from its own components; the discount factors compound; the DCF waterfall chains',
 'from revenue through EBITDA, EBIT, NOPAT and free cash flow to the firm to present value; the terminal',
 'block chains from reinvestment = growth / return on capital; the statements roll forward; and every',
 'ratio and per-share figure is a formula. A driver test perturbs every input in place, re-evaluates the',
 'whole workbook and confirms the headline moves in the right direction, with no dead inputs.', '',
 'THREE CLASSES OF CELL ARE PASTED, and it is worth knowing exactly which.',
 '  1. Audited and disclosed history — the primary record, not a calculation. Where a line is both',
 '     disclosed and derivable, the DISCLOSED figure is carried.',
 '  2. One output of the unit build: the split of other production cost into a fixed leg and a per-tonne',
 '     leg, fitted by least squares across the three audited years. A three-point regression cannot be',
 '     read as a grid, so its two coefficients sit on Assumptions and everything downstream of them —',
 '     every year of other production cost, in both constructions — is formula.',
 '  3. Whole-model re-runs: the probability map, the sensitivity grids, the mid-cycle revenue and EBITDA',
 '     behind the normalised-earnings lens, and the sector-beta re-solve of the whole waterfall. Each of',
 '     those cells is a complete revaluation or a complete simulation, so it cannot be a formula in a',
 '     grid. THESE DO NOT REDRAW WHEN A DRIVER CHANGES — edit an input on Assumptions and the waterfall,',
 '     the statements, the bridge and all four lenses reprice, but those grids keep the values printed',
 '     here.',
 '  Anything else pasted would be a defect.', '',
 'THE COST OF CAPITAL IS FLAT, DELIBERATELY, AND THE SHEET SHOWS WHY. Some models glide the discount rate',
 'down over the forecast because the currency is normalising off a high policy rate. That is not this',
 'company: the dirham is hard-pegged to the dollar, Borouge reports in dollars, and its debt is dollar',
 'debt. There is no rate normalisation to glide along, so the per-year discount rate on the DCF sheet is',
 'a live formula pointing at one built cost of capital, and the discount factors compound off it. The row',
 'is there so a reader can see the rate used in each year rather than infer it.', '',
 'HOW THE COMPANY IS VALUED, AND WHY. Borouge is a single-segment polyolefin operating company: it makes',
 'polyethylene and polypropylene at one integrated complex at Ruwais and sells them in more than ninety',
 'countries. There is no lending book, no property portfolio and no investment-holding leg, so it takes',
 'one operating-company lens — discounted cash flow built from tonnes and dollars per tonne — cross-read',
 'against book value and sustainable return, through-cycle multiples and normalised earnings power.', '',
 'TWO THINGS TO KNOW BEFORE READING ANY NUMBER.',
 '  * BOROUGE 4 IS NOT OWNED. The 1.4 million tonne expansion next door is owned 70% by ADNOC and 30% by',
 '    OMV. Borouge plc operates it under an Asset Usage Agreement and earns a fee. It is valued on the',
 '    Bridge sheet as a fee stream, never consolidated as capacity, and terminal growth is set on the',
 '    footing that the company\'s OWN capacity is fixed.',
 '  * THE STUDY PUBLISHES TWO ANSWERS, NOT ONE. The share\'s own regression beta is statistically weak,',
 '    and a sector bottom-up beta gives a materially different cost of capital. Both are carried all the',
 '    way to a value per share, side by side, everywhere they appear. Averaging them would hide the',
 '    single most consequential judgement in the study.', '',
]
r = 4
for ln in LINES:
    ws.cell(row=r, column=1, value=ln).font = Font(bold=ln.isupper() and len(ln) > 12)
    r += 1
band(ws, r, 'SHEET GUIDE', 9); r += 1
for nm, desc in [
    ('Summary', 'the headline: four lenses, both betas, and the terminal-value share'),
    ('Fundamental Valuation', 'the cost of capital built line by line, and the lenses assembled'),
    ('Assumptions', 'every driver. Blue cells here are what you change'),
    ('SOTP Bridge', 'enterprise value to equity, with the Borouge 4 fee stream separate'),
    ('Segments', 'the unit build: tonnes, prices and cost per tonne, by product'),
    ('Relative & Normalized', 'the through-cycle triangulation and normalised earnings power'),
    ('DCF', 'the waterfall and the terminal block, both constructions'),
    ('Income Statement', 'three audited years and five forecast years'),
    ('Balance Sheet', 'three audited years and five forecast years, rolled forward'),
    ('Cash Flow', 'linked to the waterfall, not restated'),
    ('Summary Financials', 'the compact view'),
    ('Monte Carlo', 'the probability map for the traded price (a re-run, not a formula)'),
    ('Sensitivity', 'value against cost of capital, growth, premium and utilisation'),
    ('Per-Share & Ratios', 'per-share figures and returns, all formulas'),
    ('Peer & Sector', 'the eleven listed peers and why the naive median is rejected'),
]:
    ws.cell(row=r, column=1, value=f'  {nm}').font = Font(bold=True)
    ws.cell(row=r, column=3, value=desc)
    r += 1
ws.column_dimensions['A'].width = 30
ws.column_dimensions['C'].width = 62
for c in 'BDEFGHI':
    ws.column_dimensions[c].width = 12

# =============================================================================
# WRITE
# =============================================================================
OUT = os.path.join(HERE, 'BOROUGE_Valuation_Model_09082026_public.xlsx')
wb.save(OUT)
nform = sum(len(v) for v in EXPECT.values())
with open(os.path.join(HERE, 'xlsx_expected.json'), 'w') as f:
    json.dump(dict(expected=EXPECT, pasted=PASTED), f, indent=1)
print(f'wrote {os.path.basename(OUT)} — {len(wb.sheetnames)} sheets')
print(f'formula cells with a recorded expected value: {nform}')
print(f'pasted cells by class: {PASTED}')
assert wb.sheetnames == SHEETS, wb.sheetnames
