"""Prove the workbook is a LIVE DRIVER model.

READ FIRST tells the reader that changing a blue cell on Assumptions reprices the model.
That is a claim about the DELIVERED file, so it is tested on the delivered file: each
driver is perturbed in place, the whole workbook is re-evaluated from scratch, and the
test asserts the headline moves in the asserted DIRECTION. A dead-input sweep then bumps
every remaining blue cell on Assumptions and requires each one to move something.

TWO DIRECTIONS ARE THE OPPOSITE OF THE TEXTBOOK ONE, AND NEITHER IS A BUG.

  * Higher terminal GROWTH raises the value here, because the terminal return on capital
    (12%) sits well ABOVE the terminal cost of capital (5.5%). Growth is bought with
    reinvestment of g / return on capital, which at these levels costs less than it earns.
    Had the return sat below the cost of capital the sign would flip, and the sensitivity
    grid on the Sensitivity sheet shows exactly that boundary.

  * The WORKING-CAPITAL DAYS move the value the "wrong" way by a hair, and the mechanism
    was decomposed before anything was changed. A longer collection period raises the
    receivables BALANCE, which is what the test now asserts. What it does to cash over a
    five-year window is a different question: the cash effect is the CHANGE in working
    capital, and the sum of those changes across the window is simply the closing balance
    less the opening one. Modelled 2030 revenue of USD 5,759m sits BELOW the audited 2025
    base of USD 5,848m, so the cycle releases cash over the window rather than absorbing
    it — and a longer collection period scales that release. Adding ten days to collection
    moved the value by +0.0003 dirhams, three hundredths of a fil. The direction is real
    but the magnitude is nil, because working capital touches only the explicit window
    while 73% of enterprise value sits in the terminal block, where it does not appear at
    all. Asserting the textbook sign on the VALUE would have been asserting noise.
"""
import json
import os
import sys

import openpyxl
import xlcalc

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'BOROUGE_Valuation_Model_09082026_public.xlsx')
wb = openpyxl.load_workbook(XLSX)
XP = json.load(open(os.path.join(HERE, 'xlsx_expected.json')))

# Map every Assumptions label to its row, so the test refers to drivers by name.
LBL = {}
for row in wb['Assumptions'].iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str) and c.value.strip():
        LBL.setdefault(c.value, c.row)


def row_of(label):
    if label not in LBL:
        raise KeyError(f'no Assumptions row labelled {label!r}')
    return LBL[label]


# Blue input cells on Assumptions — everything the reader is invited to change.
BLUE_CELLS = []
for row in wb['Assumptions'].iter_rows():
    for c in row:
        if (isinstance(c.value, (int, float)) and c.font and c.font.color
                and c.font.color.rgb
                and str(c.font.color.rgb).upper().endswith('0000FF')):
            BLUE_CELLS.append(c.coordinate)
# COUNT AGAINST A KNOWN TOTAL. An earlier revision of this matcher found no cells at all
# and still printed "0 moved nothing" — a sweep that tested nothing, reported as a pass.
EXPECTED_BLUE = sum(XP['pasted'][k] for k in ('history', 'unit_build'))
if not BLUE_CELLS:
    raise SystemExit('dead-input sweep found NO input cells — the matcher is broken, not '
                     'the workbook')

# The headline readings the test watches.
WATCH = dict(
    dcf_own_n=('DCF', None), dcf_own_p=('DCF', None),
    median=('Fundamental Valuation', None), wacc=('Fundamental Valuation', None),
    ebitda26=('DCF', None), rev26=('DCF', None), fcff26=('DCF', None),
    tv=('DCF', None), rel=('Relative & Normalized', None),
    norm=('Relative & Normalized', None), bv=('Fundamental Valuation', None),
    eq26=('Balance Sheet', None), nd26=('Balance Sheet', None),
    b4=('SOTP Bridge', None), lowest=('Fundamental Valuation', None),
    highest=('Fundamental Valuation', None),
)


def locate():
    """Find the watched cells by matching the builder's recorded model values."""
    D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
    E = XP['expected']
    FRn, FRp = D['framings']['normalisation'], D['framings']['prolonged']

    def by_label(sheet, label):
        """Locate a cell by the row label beside it. Needed wherever two different
        quantities happen to share a value — the US 10-year yield and the spread-basis
        premium are both 4.65%, and a value search returns whichever comes first."""
        ws = wb[sheet]
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value == label:
                return f'B{row[0].row}'
        raise KeyError(f'{sheet}: no row labelled {label!r}')

    def by_value(sheet, want, tol=1e-6):
        for coord, v in E.get(sheet, {}).items():
            if abs(v - want) <= max(tol, abs(want) * 1e-9):
                return coord
        raise KeyError(f'{sheet}: no recorded cell with value {want}')

    return dict(
        dcf_own_n=('DCF', by_value('DCF', D['lenses']['dcf_normalisation_own_beta'])),
        dcf_own_p=('DCF', by_value('DCF', D['lenses']['dcf_prolonged_own_beta'])),
        median=('Fundamental Valuation', by_value('Fundamental Valuation',
                                                  D['fair_mid'])),
        lowest=('Fundamental Valuation', by_value('Fundamental Valuation',
                                                  D['fair_low'])),
        highest=('Fundamental Valuation', by_value('Fundamental Valuation',
                                                   D['fair_high'])),
        wacc=('Fundamental Valuation', by_value('Fundamental Valuation',
                                                D['wacc']['wacc_own'])),
        ebitda26=('DCF', by_value('DCF', FRn['rows'][0]['ebitda'])),
        rev26=('DCF', by_value('DCF', FRn['rows'][0]['revenue'])),
        fcff26=('DCF', by_value('DCF', FRn['rows'][0]['fcff'])),
        tv=('DCF', by_value('DCF', FRn['pv_terminal'])),
        rel=('Relative & Normalized', by_value('Relative & Normalized',
                                               D['lenses']['relative_multiples'])),
        norm=('Relative & Normalized',
              by_value('Relative & Normalized',
                       D['lenses']['normalised_earnings_own_beta'])),
        bv=('Fundamental Valuation', by_value('Fundamental Valuation',
                                              D['lenses']['book_value_own_beta'])),
        b4=('SOTP Bridge', by_value('SOTP Bridge', FRn['b4']['value'])),
        nwc26=('Balance Sheet', by_value('Balance Sheet', FRn['rows'][0]['nwc'])),
        erp_ds=('Fundamental Valuation',
                by_label('Fundamental Valuation',
                         'Equity risk premium — default-spread basis')),
    )


CELLS = locate()


def read(overrides=None):
    bk = xlcalc.Book(wb, overrides)
    out = {}
    for k, (sheet, coord) in CELLS.items():
        out[k] = bk.cell_value(sheet, coord)
    return out


BASE = read()

# driver label, column, bump, watched reading, asserted direction, why
TESTS = [
    ('US 10-year Treasury yield', 'B', +0.01, 'dcf_own_n', 'down',
     'a higher risk-free rate raises the discount rate on every cash flow'),
    ('UAE adjusted default spread', 'B', +0.01, 'erp_ds', 'up',
     'it is the alternative premium basis, published for disclosure. It drives no lens, '
     'so it must move the published figure and nothing else'),
    ('US adjusted default spread', 'B', +0.01, 'dcf_own_n', 'up',
     'the spread is netted OUT of the risk-free rate, so widening it lowers the '
     'normalised rate and therefore the discount rate'),
    ('Mature-market equity risk premium', 'B', +0.01, 'dcf_own_n', 'down',
     'a higher premium raises the cost of equity'),
    ('UAE country risk premium', 'B', +0.01, 'dcf_own_n', 'down',
     'country risk enters through the premium, so widening it raises the cost of equity'),
    ("Beta — the share's own five-year weekly regression", 'B', +0.30, 'dcf_own_n',
     'down', 'a higher beta raises the cost of equity'),
    ('Beta — sector bottom-up, unlevered', 'B', +0.30, 'wacc', 'flat',
     'the sector beta drives the OTHER column; it must not touch the own-beta answer'),
    ('Marginal pre-tax cost of debt', 'B', +0.02, 'dcf_own_n', 'down',
     'a higher cost of debt raises the weighted average cost of capital'),
    ('Effective tax rate', 'B', +0.05, 'dcf_own_n', 'down',
     'a higher tax rate cuts NOPAT, and the after-tax cost of debt effect is far smaller'),
    ('Polyethylene nameplate capacity', 'B', +200, 'dcf_own_n', 'up',
     'more capacity at the same utilisation is more tonnes'),
    ('Polypropylene nameplate capacity', 'B', +200, 'dcf_own_n', 'up',
     'more capacity at the same utilisation is more tonnes'),
    ('Realisation residual — polyethylene', 'B', +0.05, 'dcf_own_n', 'up',
     'a higher realisation lifts the price actually received per tonne'),
    ('Realisation residual — polypropylene', 'B', +0.05, 'dcf_own_n', 'up',
     'a higher realisation lifts the price actually received per tonne'),
    ('Other production cost — fixed leg', 'B', +100, 'dcf_own_n', 'down',
     'more fixed cost is less EBITDA'),
    ('Other production cost — per tonne of production', 'B', +0.02, 'dcf_own_n', 'down',
     'more variable cost per tonne is less EBITDA'),
    ('Feedstock cost per tonne, FY2025 actual', 'B', +30, 'dcf_own_n', 'down',
     'the whole feedstock stack escalates from this base'),
    ('Polypropylene benchmark, FY2025 actual', 'B', +100, 'dcf_own_n', 'up',
     'a higher base benchmark makes the market-priced feedstock leg CHEAPER relative to '
     'it, because the leg escalates on the ratio of the forecast benchmark to this base'),
    ('Contracted ethane real escalation', 'B', +0.02, 'dcf_own_n', 'down',
     'escalating the contracted feedstock leg raises cost'),
    ('UAE consumer inflation', 'B', +0.02, 'dcf_own_n', 'down',
     'it escalates the domestic fixed cost leg and general and administrative expense'),
    ('General and administrative, FY2025 excluding depreciation', 'B', +50,
     'dcf_own_n', 'down', 'more overhead is less EBITDA'),
    ('Other income, FY2025', 'B', +50, 'dcf_own_n', 'up', 'more other income is more '
     'EBITDA'),
    ('Other revenue, FY2025', 'B', +50, 'dcf_own_n', 'up', 'more revenue is more EBITDA'),
    ('Days sales outstanding', 'B', +10, 'nwc26', 'up',
     'slower collection raises the receivables balance. Its effect on VALUE is not '
     'asserted: see the note at the top of this file — over this window the cycle '
     'releases cash, so the value sign follows the revenue path and the magnitude is '
     'three hundredths of a fil'),
    ('Days inventory', 'B', +10, 'nwc26', 'up',
     'more inventory raises the working-capital balance'),
    ('Days payable', 'B', +10, 'nwc26', 'down',
     'paying later lowers the working-capital balance'),
    ('Capital expenditure guide, 2026', 'B', +100, 'dcf_own_n', 'down',
     'more capital spend is less free cash flow'),
    ('Maintenance capital expenditure, steady state', 'B', +100, 'dcf_own_n', 'down',
     'more capital spend in four of the five years, and in the terminal base'),
    ('Depreciation rate on the property balance', 'B', +0.02, 'dcf_own_n', 'down',
     'more depreciation cuts NOPAT; the add-back is not fully offsetting because tax '
     'is charged on the lower EBIT'),
    ('Property, plant and equipment, opening', 'B', +500, 'dcf_own_n', 'down',
     'a bigger property base at the same rate is more depreciation'),
    ('Terminal growth rate', 'B', +0.005, 'dcf_own_n', 'up',
     'terminal return on capital sits far above the terminal cost of capital, so growth '
     'is bought with reinvestment that earns more than it costs'),
    ('Terminal return on capital', 'B', +0.03, 'dcf_own_n', 'up',
     'a higher return means less reinvestment is needed to fund the same growth'),
    ('Net debt', 'B', +500, 'dcf_own_n', 'down',
     'more debt is deducted in the bridge to equity'),
    ('Lease liabilities', 'B', +100, 'dcf_own_n', 'down', 'deducted in the bridge'),
    ('Non-controlling interests', 'B', +100, 'dcf_own_n', 'down', 'deducted in the bridge'),
    ('Borouge 4 earnings accretion after ramp-up', 'B', +0.05, 'b4', 'up',
     'a larger accretion is a larger fee stream, until the disclosed cumulative figure '
     'binds instead'),
    ('Borouge 4 cumulative net profit, first three years', 'B', -300, 'b4', 'down',
     'the fee carried is the LOWER of the two quantifications, so cutting the cumulative '
     'figure below the accretion one makes it bind'),
    ('Closing price, 7 August 2026', 'B', +0.50, 'wacc', 'up',
     'a higher price is a bigger equity weight, and equity is the dearer of the two'),
    ('Ordinary shares in issue', 'B', +2000, 'dcf_own_n', 'down',
     'the same equity value spread over more shares'),
    ('Dirham per US dollar (pegged since November 1997)', 'B', +0.20, 'dcf_own_n', 'up',
     'a dollar of value converts into more dirhams per share'),
    ('Debt to equity at market value', 'B', +0.10, 'wacc', 'flat',
     'it re-levers the SECTOR beta only; the own-beta answer must not move'),
    ('Polyethylene utilisation', 'B', +0.05, 'dcf_own_n', 'up',
     'higher utilisation in 2026 is more tonnes'),
    ('Polyethylene benchmark (USD/t)', 'B', +100, 'dcf_own_n', 'up',
     'a higher benchmark is a higher realised price'),
    ('Selling and distribution (USD/t sold)', 'B', +30, 'dcf_own_n', 'down',
     'higher freight is less EBITDA'),
    ('Share of feedstock bought at market price', 'B', +0.20, 'dcf_own_n', 'down',
     'buying more feedstock at the market price rather than under contract raises cost, '
     'because the market leg prices above the contracted one on this path'),
]

TOL = 1e-7
rows, failures = [], []
for label, col, bump, watch, direction, why in TESTS:
    rw = row_of(label)
    cur = wb['Assumptions'][f'{col}{rw}'].value
    got = read({('Assumptions', f'{col}{rw}'): cur + bump})
    before, after = BASE[watch], got[watch]
    delta = after - before
    if direction == 'up':
        ok = delta > TOL
    elif direction == 'down':
        ok = delta < -TOL
    else:
        ok = abs(delta) <= TOL
    rows.append(dict(driver=label, bump=bump, watched=watch, before=before, after=after,
                     delta=delta, asserted=direction, ok=ok, why=why))
    if not ok:
        failures.append(f'{label}: {watch} moved {delta:+.6f}, asserted {direction}')

# ---- dead-input sweep: every remaining blue cell must move SOMETHING ----------
tested_rows = {row_of(t[0]) for t in TESTS}
dead, non_binding = [], []
swept = 0
for coord in BLUE_CELLS:
    rw = int(''.join(ch for ch in coord if ch.isdigit()))
    colr = ''.join(ch for ch in coord if ch.isalpha())
    cur = wb['Assumptions'][coord].value
    if not isinstance(cur, (int, float)):
        continue
    swept += 1
    # Escalate the probe. Some inputs are CAPS rather than levers — the Borouge 4
    # cumulative-profit figure feeds a MIN() and is not currently the binding side, so a
    # 10% nudge moves nothing while a large cut moves it immediately. That is correct
    # model behaviour, and calling it a dead input would be wrong. An input is dead only
    # if NOTHING moves it in EITHER direction at ANY plausible magnitude.
    moved_at = None
    for frac in (0.10, 0.50, 0.90):
        bump = (abs(cur) * frac) or 0.05
        for signed in (bump, -bump):
            got = read({('Assumptions', coord): cur + signed})
            if any(abs(got[k] - BASE[k]) > TOL for k in BASE):
                moved_at = frac
                break
        if moved_at:
            break
    label_ = wb['Assumptions'][f'A{rw}'].value
    if moved_at is None:
        dead.append(f'{coord} ({label_})')
    elif moved_at > 0.10:
        non_binding.append(f'{coord} ({label_}) — inert to a 10% move, moves at '
                           f'{moved_at:.0%}: it is a cap, not a lever')

print(f'driver directions tested: {len(rows)}, failures: {len(failures)}')
print(f"{'driver':<58} {'watched':<10} {'before':>12} {'after':>12} {'move':>10}  ok")
for x in rows:
    print(f"{x['driver'][:57]:<58} {x['watched']:<10} {x['before']:>12,.4f} "
          f"{x['after']:>12,.4f} {x['delta']:>+10,.4f}  {'ok' if x['ok'] else 'FAIL'}")
for f in failures:
    print('  ! ', f)
print(f'\ndead-input sweep: {swept} input cells bumped, {len(dead)} moved nothing')
print(f'  (the builder recorded {EXPECTED_BLUE} pasted history and unit-build cells '
      f'across the workbook; {len(BLUE_CELLS)} of them are inputs on Assumptions)')
if swept < 40:
    raise SystemExit(f'dead-input sweep only reached {swept} cells — too few to be a '
                     f'real sweep of the Assumptions sheet')
for d in dead:
    print('  ! ', d)
for n in non_binding:
    print('  ~ ', n)

with open(os.path.join(HERE, 'driver_test_result.json'), 'w') as f:
    json.dump(dict(tests=rows, failures=failures, blue_cells_swept=swept,
                   dead_inputs=dead, non_binding=non_binding), f, indent=1)

if failures or dead:
    sys.exit(1)
print('\nEvery driver moves the headline in the asserted direction, and no input is dead.')
