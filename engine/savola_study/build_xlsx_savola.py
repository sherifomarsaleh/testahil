"""SAVOLA_Valuation_Model_19082026_public.xlsx — 16 sheets mirroring the house canonical
model (operating-company variant). SECOND EDITION (19-Aug-2026, critique response).
Blue = inputs · black = formulas · green = cross-sheet links.

The workbook is FORMULA-DRIVEN. Every quantity that is arithmetically derivable from an
input is written as a live Excel formula, not as a pasted number. Only three classes of
cell are pasted values:

  1. audited and disclosed historical figures (the primary record);
  2. the unit build's disclosed BASES (FY2025 category volumes/revenue/GP and the H1-2026
     actuals) — the build itself is LIVE: category volumes compound off growth drivers,
     revenue = volume x revenue-per-tonne, gross profit = volume x GP-per-tonne, Panda
     revenue = stores x sales-per-store, and every margin is an OUTPUT;
  3. engine outputs that are whole-model re-runs by construction: the Monte Carlo price
     map, the sensitivity grids, the DCF scenario bear/bull bounds, Framing B, and the
     CDS-basis fair value.

Every formula cell also carries the model's own value into xlsx_expected.json, and
recalc.py evaluates the workbook independently and asserts the two agree.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'
NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, SEG, H1 = D['experts'], D['segments_fy25'], D['h1_2026']
STK, S0, BT = D['strike'], D['step0'], D['backtest']
IN = {k: v['value'] for k, v in D['inputs'].items()}
TRC = D['terminal_record']
TRI, TRO = TRC['inputs'], TRC['outputs']
SPOT, SH, SHW = M['spot'], M['shares_mn'], M['shares_val_mn']
YF = [str(y) + 'E' for y in F['years']]
YH = ['FY2023', 'FY2024', 'FY2025']
T = W['tax']
CD = ['B', 'C', 'D', 'E', 'F']

wb = Workbook()
EXPECT = {}
ANCH = {}

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Savola Group Company (Tadawul: 2050)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the Savola Group valuation study. Every blue cell is',
 'an input; every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice: the cost of capital is built from',
 'the risk-free rate net of the sovereign spread, beta and the premium rather than pasted; the discount',
 'factors compound; and the segment build, income statement, balance sheet, cash flow, ratios and all four',
 'lenses chain off the same cells.', '',
 'THIS IS THE SECOND EDITION (19-Aug-2026). It restrikes the 18-Aug first edition after a four-critique',
 'external audit: the settled 18-Aug close (25.40) replaces an intraday print; the cash-flow waterfall now',
 'charges the FULL lease additions; the terminal return on capital is COMPUTED from the model\'s own year',
 'five rather than input; the risk-free rate is the published SAR sovereign curve; the capital-structure',
 'weights use the 30-Jun-2026 reviewed balance sheet; the Tiryaki receivable moved from working capital to',
 'the bridge; the per-share divisor is the Q2-2026 ex-treasury count; the relative lens is trailing-on-',
 'trailing with Al Othaim n/m; and the bridge holds its three anchor-dated legs outside the roll.', '',
 'THREE THINGS ARE PASTED VALUES, and it is worth knowing exactly which. First, audited and disclosed',
 'history — the primary record, not a calculation: the FY2023-FY2025 statements, the FY2025 category',
 'volumes, revenues and gross profits, and the H1-2026 actuals. Second, nothing else of the build: the',
 'forecast itself is LIVE — category volumes compound off growth drivers, revenue is volume x revenue per',
 'tonne, gross profit is volume x gross profit per tonne, Panda revenue is stores x sales per store. Where',
 'a margin IS an input it is said plainly: Herfy and Al Kabeer disclose only revenue and margin, so their',
 'margins are inputs at the finest disclosed level (flagged); Panda\'s margin is the identity of its two',
 'inputs; Food Processing — 47% of FY2026E EBITDA — is fully unit-built and its margins are outputs.',
 'Third, whole-model engine outputs, where each figure is a complete re-run of the entire valuation and so',
 'cannot be a single formula: the Monte Carlo price map, the sensitivity grids, the DCF scenario bear/bull',
 'bounds, Framing B, the judgement variants, and the CDS-basis fair value. Changing a driver reprices the',
 'model but does NOT redraw those engine outputs.', '',
 'How revenue is built. Not as one growth rate. Edible oil, sugar and pasta are volume x price: thousand',
 'tonnes compound on their own growth paths (anchored on the disclosed H1-2026 actuals: oil +16.1%, sugar',
 '+7.2%, pasta +3.1% as the company states it) and carry their own revenue-per-tonne paths. Nuts & spices',
 'carries a revenue path with the July-2026 Mehbaj acquisition folded in small (SR 11.4mn consideration',
 'per the Q2-2026 interims, note 19). Panda is stores x sales-per-store: the store path follows the',
 'company\'s own 20-plus-per-year guidance (the +8/yr run-rate alternative is priced as a variant), and',
 'sales density carries the measured H1-2026 decline — published as a range, −7.1% to −6.0%, because the',
 'June-2025 store count is not disclosed — fading in Framing A and persisting in Framing B.',
 'Herfy and Al Kabeer are revenue glides anchored on H1-2026 actuals.', '',
 'How cost is built. Each category carries its own gross profit per tonne — anchored on the H1-2026',
 'actuals, with the second half held BELOW the first on the company\'s own replacement-cost warning and',
 'the July-2026 vegetable-oil price rise — and its own operating-cost ratio measured from FY2025. Segment',
 'EBITDA and the group margin are OUTPUTS computed on the Segments sheet.', '',
 'Leases are debt, and charged in full. The lease liability is deducted in the bridge and carries its',
 'measured 5.86% effective rate inside the cost of capital; the cash-flow waterfall charges the FULL lease',
 'additions — right-of-use depreciation (renewals) PLUS the lease-book growth the store programme creates.',
 'Leases fund assets, so lease-funded growth is reinvestment like owned capex; the first edition charged',
 'renewals only, and correcting that is this edition\'s largest single change (−2.3 on the DCF lens).', '',
 'The contested judgement — Panda\'s expansion. The model is computed BOTH WAYS and never averaged:',
 'Framing A lets sales density stabilise as the store-refresh programme and e-commerce mature; Framing B',
 'holds the measured H1-2026 density erosion forever. Both fair values sit side by side on the Fundamental',
 'Valuation sheet, in the study\'s summary table, and in an expert\'s range.', '',
 'Discount convention, stated. The five explicit years are discounted at the explicit-window cost of',
 'capital at full-year END-of-period factors; the terminal value, dated at the end of year five, is',
 'discounted at the year-five factor and uses terminal weights re-based toward equity. The bridge\'s',
 'Dec-2025-dated legs are rolled to the 18-Aug-2026 anchor at the cost of equity, net of the SAR 1.70',
 'dividend whose ex-date (07-May-2026) falls inside that window; the three anchor-dated legs (Kinan',
 'capitalized on H1-2026 earnings, Herfy\'s NCI at its 18-Aug price, the Jul-2026 Mehbaj consideration)',
 'sit OUTSIDE the roll at their own dates.', '',
 'Sourcing note, up front. FY2023, FY2024 and FY2025 come from the company\'s own audited consolidated',
 'financial statements read from savola.com/investors; Q1-2026 AND Q2-2026 from the reviewed interim',
 'statements (the Q2 interims carry the Mehbaj consideration, the 30-Jun balance sheet and the ex-treasury',
 'share divisor used here); the H1-2026 actuals from the company\'s own earnings release and investor',
 'presentation (06-Aug-2026). Every input is listed with source and date in the companion bibliography.', '',
 f'Currency. SAR million unless stated. Spot SAR {SPOT:.2f} (settled 18-Aug-2026 close). Sheets: READ FIRST ·',
 'Summary · Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF ·',
 'Income Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity ·',
 'Per-Share & Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 112

# ============ 4 ASSUMPTIONS (built early so refs exist) ========================
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every input in the model', 'Blue cells are inputs. Change one and '
      'the model reprices: everything downstream is a formula.', 8, awidth=56, cwidth=12)
rA = 4
A = {}

def block(name, items):
    global rA
    band(wsA, rA, 8); put(wsA, f'A{rA}', name, bold=True, fmt=None); rA += 1
    for key, lab, val, fmt in items:
        put(wsA, f'A{rA}', lab, fmt=None)
        if isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(wsA, f'{get_column_letter(2+i)}{rA}', v, BLUE, fmt)
        else:
            put(wsA, f'C{rA}', val, BLUE, fmt)
        A[key] = rA
        rA += 1
    rA += 1

def a(key, i=None):
    col = get_column_letter(2 + i) if i is not None else 'C'
    return f"Assumptions!${col}${A[key]}"

hdr(wsA, 3, ['Input', YF[0], YF[1], YF[2], YF[3], YF[4]])
block('Anchors', [
    ('spot', 'Spot price (SAR, SETTLED 18-Aug-2026 close)', SPOT, PX),
    ('shares', 'Shares issued (mn)', SH, NUM0),
    ('sharesw', 'Shares outstanding ex-treasury (mn; Q2-2026 interims EPS-note divisor)',
     SHW, NUM1),
    ('tax', 'Combined zakat + income tax rate on core profit', IN['tax_rate'], PCT),
    ('payout', 'Dividend payout (stated policy 50-60% of net profit; midpoint)',
     IN['payout'], PCT),
    ('anchor_days', 'Days from the 31-Dec-2025 valuation date to the 18-Aug-2026 anchor',
     IN['anchor_days'], NUM0),
    ('div_between', 'Dividend gone ex between valuation date and anchor (SAR/share)',
     IN['div_between'], PX)])
block('FY2025 disclosed bases (audited statements + company presentations)', [
    ('rev25', 'Group revenue FY2025', IN['rev_fy25'], NUM0),
    ('gp25', 'Group gross profit FY2025', HI['FY25']['gp'], NUM0),
    ('oilv25', 'Oil volume FY2025 (k MT)', IN['oil_vol_fy25'], NUM0),
    ('oilr25', 'Oil revenue FY2025', IN['oil_rev_fy25'], NUM0),
    ('sugv25', 'Sugar volume FY2025 (k MT)', IN['sug_vol_fy25'], NUM0),
    ('sugr25', 'Sugar revenue FY2025', IN['sug_rev_fy25'], NUM0),
    ('pasv25', 'Pasta volume FY2025 (k MT, DISCLOSED — FY2025 presentation p17)',
     IN['pas_vol_fy25'], NUM0),
    ('pasr25', 'Pasta revenue FY2025', IN['pas_rev_fy25'], NUM0),
    ('nutsr25', 'Nuts & spices revenue FY2025 (residual to the audited segment)',
     SEG['categories']['nuts']['rev'], NUM0),
    ('retrev25', 'Panda (Retail) segment revenue FY2025', IN['ret_segrev_fy25'], NUM0),
    ('fsvrev25', 'Food Services (Herfy) segment revenue FY2025', IN['fsv_segrev_fy25'], NUM0),
    ('frzrev25', 'Frozen Food (Al Kabeer) segment revenue FY2025', IN['frz_segrev_fy25'], NUM0),
    ('invsegrev', 'Investments segment revenue (held flat)', IN['invseg_rev'], NUM1),
    ('st24', 'Panda stores at Dec-2024', IN['stores_end24'], NUM0),
    ('st25', 'Panda stores at Dec-2025', IN['stores_end25'], NUM0)])
block('Unit build — Food Processing (volume x price x GP/tonne; margins are outputs)', [
    ('oilvg', 'Oil volume growth', IN['oil_vol_g'], PCT),
    ('oilpg', 'Oil revenue-per-tonne growth', IN['oil_rpt_g'], PCT),
    ('oilgpt', 'Oil gross profit per tonne (SAR/t)', IN['oil_gpt_path'], NUM0),
    ('sugvg', 'Sugar volume growth', IN['sug_vol_g'], PCT),
    ('sugpg', 'Sugar revenue-per-tonne growth', IN['sug_rpt_g'], PCT),
    ('suggpt', 'Sugar gross profit per tonne (SAR/t)', IN['sug_gpt_path'], NUM0),
    ('pasvg', 'Pasta volume growth', IN['pas_vol_g'], PCT),
    ('paspg', 'Pasta revenue-per-tonne growth', IN['pas_rpt_g'], PCT),
    ('pasgpt', 'Pasta gross profit per tonne (SAR/t)', IN['pas_gpt_path'], NUM0),
    ('nutsrev', 'Nuts & spices revenue path (Mehbaj folded in, flagged)',
     IN['nuts_rev_path'], NUM0),
    ('nutsgm', 'Nuts & spices gross margin path', IN['nuts_gm_path'], PCT),
    ('opx_oil', 'Oil operating cost / revenue (measured FY2025)',
     IN['fp_opex_ratio']['oil'], PCT2),
    ('opx_sug', 'Sugar operating cost / revenue (measured FY2025)',
     IN['fp_opex_ratio']['sugar'], PCT2),
    ('opx_pas', 'Pasta operating cost / revenue (measured FY2025)',
     IN['fp_opex_ratio']['pasta'], PCT2),
    ('opx_nuts', 'Nuts operating cost / revenue (measured FY2025)',
     IN['fp_opex_ratio']['nuts'], PCT2)])
block('Panda (Retail) build — stores x sales per store; margin is an output', [
    ('stores', 'Store count, end of year (company guidance 20+/yr; +8/yr run-rate '
     'variant priced on Sensitivity)', IN['stores_path'], NUM0),
    ('spsg', 'Sales-per-average-store growth (Framing A; opening measured as a range '
     '-7.1% to -6.0% over the undisclosed Jun-2025 count)', IN['sps_g_A'], PCT),
    ('pgm', 'Panda gross margin (H1-2026 actual, held)', IN['panda_gm'], PCT),
    ('popex', 'Panda store-opex / revenue (measured H1-2026)', IN['panda_opex_ratio'], PCT),
    ('pstep', 'Framing A scale gain on the opex ratio from FY2028 (per year)',
     IN['panda_scale_step'], PCT2)])
block('Food Services and Frozen Food', [
    ('herg', 'Herfy revenue growth', IN['herfy_rev_g'], PCT),
    ('hermg', 'Herfy EBITDA margin (H1-2026 actual 18.7%, held FLAT — a margin INPUT '
     'at the finest disclosed level, flagged)', IN['herfy_ebitda_mgn'], PCT),
    ('frzg', 'Al Kabeer revenue growth', IN['frz_rev_g'], PCT),
    ('frzmg', 'Al Kabeer EBITDA margin (H1-2026 actual, held)', IN['frz_ebitda_mgn'], PCT)])
block('Group items', [
    ('elim', 'Eliminations / Food-Processing segment revenue (measured FY2025)',
     IN['elim_ratio'], PCT2),
    ('unalloc', 'Unallocated corporate costs (SAR mn)', IN['unalloc_path'], NUM0)])
block('Capital intensity, D&A and leases', [
    ('capex', 'Capital expenditure (SAR mn, incl. intangibles)', IN['capex_path'], NUM0),
    ('deprate', 'Owned-PP&E depreciation rate on opening balance (measured FY2025)',
     IN['dep_rate_own'], PCT2),
    ('intam', 'Intangibles amortisation = intangible capex (book held flat)',
     IN['int_amort'], NUM0),
    ('roud25', 'Right-of-use depreciation FY2025 (audited components)', IN['rou_dna_fy25'],
     NUM1),
    ('roug', 'Right-of-use / lease growth (store-driven)', IN['rou_growth'], PCT),
    ('leaserate', 'Lease effective interest rate (measured FY2025)', IN['lease_rate'], PCT2)])
block('Working capital — component days measured FY2025', [
    ('dio', 'Days inventory outstanding', IN['dio_fy25'], NUM1),
    ('dso', 'Days sales outstanding', IN['dso_fy25'], NUM1),
    ('dpo', 'Days payable outstanding', IN['dpo_fy25'], NUM1),
    ('prepr', 'Prepayments & other receivables / revenue (EX the Tiryaki receivable)',
     IN['prepay_ratio'], PCT2),
    ('accrr', 'Accrued & other liabilities / revenue', IN['accrued_ratio'], PCT2),
    ('contr', 'Contract liabilities / revenue', IN['contr' 'act_ratio'], PCT2)])
block('Profit walk', [
    ('ncis', "Non-controlling interests' share of core profit (measured FY2025)",
     IN['nci_share'], PCT),
    ('kinh1', "Kinan share of results, H1-2026 actual (SAR mn)",
     IN['kinan_profit_share_h126'], NUM1),
    ('king', 'Kinan contribution growth', IN['kinan_g'], PCT),
    ('kindiv', 'Kinan cash dividend received FY2025 (grows with Kinan)', IN['kinan_div'],
     NUM1),
    ('sar1y', 'Yield on surplus cash (observed 1Y SAR sovereign)', IN['sar_1y_obs'], PCT2)])
block('Cost of capital — v2 method, rating basis primary', [
    ('rf', 'Risk-free rate: PUBLISHED SAR sovereign curve, FTSE SAGBI 7-10y YTM '
     '(31-Jul-2026 factsheet)', IN['rf_observed'], PCT2),
    ('sov', 'Saudi sovereign default spread (rating basis, July-2026 Damodaran, netted '
     'out of rf)', IN['sov_spread_rating'], PCT2),
    ('erp', 'Equity risk premium (Saudi total, rating basis, July-2026 Damodaran)',
     IN['erp_rating'], PCT2),
    ('sovcds', 'Saudi sovereign CDS spread (CDS basis — JANUARY-2026 vintage, flagged)',
     IN['sov_spread_cds'], PCT2),
    ('erpcds', 'Equity risk premium (CDS basis — January-2026 vintage, flagged)',
     IN['erp_cds'], PCT2),
    ('beta', 'Beta (SAVOLA weekly vs TASI, 5y, Dimson)', IN['beta'], '0.000'),
    ('kdsar', 'Marginal SAR cost of debt (SAIBOR + murabaha spread)', IN['kd_sar'], PCT2),
    ('kdeg', 'EGP tranche cost, SAR-equivalent (parity construction)', IN['kd_eg_localeq'],
     PCT2),
    ('kdot', 'Other tranches (AED/DZD) cost', IN['kd_other'], PCT2),
    ('lsa', 'Loans — Saudi Riyal tranche (constructed from the audited currency note; '
     'the note discloses currencies, not countries)', IN['loans_geo_sa'], NUM0),
    ('leg', 'Loans — Egyptian Pound tranche (constructed)', IN['loans_geo_eg'], NUM0),
    ('lot', 'Loans — other currencies, AED/DZD/USD (constructed)',
     IN['loans_geo_other'], NUM0),
    ('ljun', 'Loans and borrowings at 30-Jun-2026 (WACC weight leg, reviewed interims)',
     IN['loans_jun26'], NUM0),
    ('zjun', 'Lease liabilities at 30-Jun-2026 (WACC weight leg, reviewed interims)',
     IN['leases_jun26'], NUM0),
    ('twe', 'Terminal equity weight', IN['tw_e'], PCT),
    ('twl', 'Terminal loans weight', IN['tw_loans'], PCT),
    ('greal', 'Terminal REAL growth (stated, not derived)', IN['g_term_real'], PCT),
    ('pit', 'Terminal inflation — Saudi house macro path', TRI['inflation'], PCT),
    ('life', 'Weighted asset life, DERIVED from note 6 (gross cost of the depreciable '
     'base over the year\'s own depreciation charge)', IN['asset_life_years'], NUM1),
    ('lifevar', 'Downside variant — twice the directly measured average age (accumulated '
     'depreciation over the same charge)', DCF['life_variant_years'], NUM1),
    ('inccap', 'Invested capital per unit of real growth, at terminal revenue',
     TRI['incremental_capital_per_unit_growth'], NUM0)])
# Nominal terminal growth is DERIVED once, here, so the book lens and the dividend model
# link to the same cell the terminal does and the three cannot drift apart.
put(wsA, f'A{rA}', 'Terminal growth = (1 + inflation) x (1 + real growth) − 1', fmt=None)
putf(wsA, f'C{rA}', f"=(1+$C${A['pit']})*(1+$C${A['greal']})-1", TRI['nominal_growth'], PCT)
A['g'] = rA
rA += 2
block('Bridge anchors (audited 31-Dec-2025)', [
    ('cash', 'Cash and cash equivalents', IN['cash_fy25'], NUM0),
    ('loans', 'Loans and borrowings (all current)', IN['loans_fy25'], NUM0),
    ('lease', 'Lease liabilities', IN['leases_fy25'], NUM0),
    ('eb', 'Employee benefits liabilities', IN['eb_fy25'], NUM0),
    ('restor', 'Provision against asset restoration', IN['restor_fy25'], NUM0),
    ('othnl', 'Other net liabilities (tax/zakat accruals + DTL − DTA − other assets)',
     IN['other_net_liab'], NUM1),
    ('invnc', 'Non-current investments (MOF sukuk + Almarai FVOCI)', IN['inv_nc_fy25'], NUM1),
    ('invc', 'Current investments (T-bills, FVTPL)', IN['inv_c_fy25'], NUM1),
    ('invprop', 'Investment property', IN['invprop_fy25'], NUM1),
    ('herfyp', 'Herfy share price (its own Tadawul listing, 18-Aug-2026)', IN['herfy_price'],
     PX),
    ('herfysh', 'Herfy shares outstanding (mn)', IN['herfy_shares_mn'], NUM2),
    ('ncib', 'Non-controlling interests, book (audited)', IN['nci_book_fy25'], NUM1),
    ('nciherfy', 'of which the 51% Herfy NCI at book', IN['nci_herfy_book'], NUM1),
    ('ppe25', 'Property, plant and equipment (owned)', IN['ppe_fy25'], NUM0),
    ('rou25', 'Right-of-use assets', IN['rou_fy25'], NUM0),
    ('intang25', 'Intangible assets and goodwill', IN['intang_fy25'], NUM0),
    ('kinbv', 'Kinan carrying value (equity method)', IN['kinan_carry'], NUM1),
    ('eq25', 'Equity attributable to owners', IN['equity_att_fy25'], NUM0),
    ('eqjun', 'Equity attributable to owners, 30-Jun-2026 (reviewed; book-lens base)',
     IN['equity_att_jun26'], NUM0),
    ('tiryaki', 'Tiryaki sale-proceeds receivable (on the 31-Dec-2025 balance sheet; '
     'settled in Tiryaki shares H1-2026)', IN['tiryaki_recv'], NUM1),
    ('mehbaj', 'Al Mehbaj consideration (Q2-2026 interims note 19: 5.4 paid + 6.0 '
     'deferred)', IN['mehbaj_total'], NUM1),
    ('hliabnc', "Herfy non-current liabilities (note 20; Expert-1 carve-out)",
     IN['herfy_liab_nc'], NUM1),
    ('hcurstk', "Herfy current lease portion, CONSTRUCTED estimate (flagged)",
     IN['herfy_cur_stack_est'], NUM1),
    ('hcash', "Herfy cash, CONSTRUCTED estimate (flagged)", IN['herfy_cash_est'], NUM1)])
block('Lens inputs', [
    ('pe_alm', 'Peer P/E — Almarai (settled 18-Aug close)', IN['peer_pe']['ALMARAI'], MULT),
    ('pe_bin', 'Peer P/E — BinDawood (settled 18-Aug close)', IN['peer_pe']['BINDAWOOD'],
     MULT),
    ('pe_nad', 'Peer P/E — NADEC (18-Aug)', IN['peer_pe']['NADEC'], MULT),
    ('pe_wil', 'Peer P/E — Wilmar, international analogue (18-Aug)',
     IN['peer_pe']['WILMAR'], MULT),
    ('pedisc', 'Conglomerate / EM-mix discount on the peer-mix multiple', IN['pe_discount'],
     PCT),
    ('normm', 'Normalised mid-cycle operating EBITDA margin', IN['norm_ebitda_mgn'], PCT2),
    ('rec25', 'Recurring net profit FY2025 (company bridge)', IN['recurring_np_fy25'], NUM1),
    ('rec26h1', 'Recurring net income H1-2026 (company net-income analysis)',
     IN['h1_recurring_h126'], NUM1),
    ('rec25h1', 'Recurring net income H1-2025 (same table, comparative column)',
     IN['h1_recurring_h125'], NUM1),
    ('w_dcf', 'Weight — discounted cash flow', D['weights']['dcf'], PCT),
    ('w_rel', 'Weight — relative', D['weights']['relative'], PCT),
    ('w_norm', 'Weight — normalised', D['weights']['normalized'], PCT),
    ('w_book', 'Weight — book', D['weights']['book'], PCT)])
# Al Othaim is displayed n/m on Peer & Sector (H1-2026 loss announced 11-Aug-2026);
# the retail leg of the mix is BinDawood alone, and the processing-leg weight is
# COMPUTED on the Relative & Normalized sheet from the model's own segment EBITDA.

# ============ 6 SEGMENTS ======================================================
ws = sheet('Segments')
title(ws, 'Segments — the unit build; every margin is an output',
      'Volumes compound on their drivers; revenue = volume x price; gross profit = volume x '
      'GP/tonne; Panda = stores x sales per store', 7, awidth=48, cwidth=13)
hdr(ws, 4, ['Food Processing unit build'] + YF)
r = 5
RW = {}

def rowf(key, label, fmls, exps, fmt=NUM0, bold_=False):
    global r
    put(ws, f'A{r}', label, fmt=None, bold=bold_)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i], fmt, bold=bold_)
    RW[key] = r
    r += 1

OIL, SUG, PAS, NUTS = F['oil'], F['sugar'], F['pasta'], F['nuts']
rowf('oilv', 'Oil volume (k MT)',
     [f"={a('oilv25')}*(1+{a('oilvg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('oilvg',i)})" for i in range(1, 5)], OIL['vol'])
rowf('oilrpt', 'Oil revenue per tonne (SAR/t)',
     [f"={a('oilr25')}/{a('oilv25')}*1000*(1+{a('oilpg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('oilpg',i)})" for i in range(1, 5)],
     [OIL['rev'][i] / OIL['vol'][i] * 1000 for i in range(5)], NUM0)
rowf('oilrev', 'Oil revenue',
     [f"={CD[i]}{RW['oilv']}*{CD[i]}{RW['oilrpt']}/1000" for i in range(5)], OIL['rev'])
rowf('oilgp', 'Oil gross profit (= volume x GP/tonne)',
     [f"={CD[i]}{RW['oilv']}*{a('oilgpt',i)}/1000" for i in range(5)], OIL['gp'])
rowf('oileb', 'Oil EBITDA (= GP − measured opex ratio x revenue)',
     [f"={CD[i]}{RW['oilgp']}-{CD[i]}{RW['oilrev']}*{a('opx_oil')}" for i in range(5)],
     OIL['eb'])
rowf('sugv', 'Sugar volume (k MT)',
     [f"={a('sugv25')}*(1+{a('sugvg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('sugvg',i)})" for i in range(1, 5)], SUG['vol'])
rowf('sugrpt', 'Sugar revenue per tonne (SAR/t)',
     [f"={a('sugr25')}/{a('sugv25')}*1000*(1+{a('sugpg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('sugpg',i)})" for i in range(1, 5)],
     [SUG['rev'][i] / SUG['vol'][i] * 1000 for i in range(5)], NUM0)
rowf('sugrev', 'Sugar revenue',
     [f"={CD[i]}{RW['sugv']}*{CD[i]}{RW['sugrpt']}/1000" for i in range(5)], SUG['rev'])
rowf('suggp', 'Sugar gross profit',
     [f"={CD[i]}{RW['sugv']}*{a('suggpt',i)}/1000" for i in range(5)], SUG['gp'])
rowf('sugeb', 'Sugar EBITDA',
     [f"={CD[i]}{RW['suggp']}-{CD[i]}{RW['sugrev']}*{a('opx_sug')}" for i in range(5)],
     SUG['eb'])
rowf('pasv', 'Pasta volume (k MT)',
     [f"={a('pasv25')}*(1+{a('pasvg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('pasvg',i)})" for i in range(1, 5)], PAS['vol'])
rowf('pasrpt', 'Pasta revenue per tonne (SAR/t)',
     [f"={a('pasr25')}/{a('pasv25')}*1000*(1+{a('paspg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('paspg',i)})" for i in range(1, 5)],
     [PAS['rev'][i] / PAS['vol'][i] * 1000 for i in range(5)], NUM0)
rowf('pasrev', 'Pasta revenue',
     [f"={CD[i]}{RW['pasv']}*{CD[i]}{RW['pasrpt']}/1000" for i in range(5)], PAS['rev'])
rowf('pasgp', 'Pasta gross profit',
     [f"={CD[i]}{RW['pasv']}*{a('pasgpt',i)}/1000" for i in range(5)], PAS['gp'])
rowf('paseb', 'Pasta EBITDA',
     [f"={CD[i]}{RW['pasgp']}-{CD[i]}{RW['pasrev']}*{a('opx_pas')}" for i in range(5)],
     PAS['eb'])
rowf('nutsrev', 'Nuts & spices revenue (driver path)',
     [f"={a('nutsrev',i)}" for i in range(5)], NUTS['rev'])
rowf('nutsgp', 'Nuts & spices gross profit',
     [f"={CD[i]}{RW['nutsrev']}*{a('nutsgm',i)}" for i in range(5)], NUTS['gp'])
rowf('nutseb', 'Nuts & spices EBITDA',
     [f"={CD[i]}{RW['nutsgp']}-{CD[i]}{RW['nutsrev']}*{a('opx_nuts')}" for i in range(5)],
     NUTS['eb'])
band(ws, r, 6)
rowf('fprev', 'Food Processing revenue',
     [f"={CD[i]}{RW['oilrev']}+{CD[i]}{RW['sugrev']}+{CD[i]}{RW['pasrev']}"
      f"+{CD[i]}{RW['nutsrev']}" for i in range(5)], F['fp_rev'], NUM0, True)
rowf('fpgp', 'Food Processing gross profit',
     [f"={CD[i]}{RW['oilgp']}+{CD[i]}{RW['suggp']}+{CD[i]}{RW['pasgp']}"
      f"+{CD[i]}{RW['nutsgp']}" for i in range(5)], F['fp_gp'], NUM0, True)
rowf('fpeb', 'Food Processing EBITDA',
     [f"={CD[i]}{RW['oileb']}+{CD[i]}{RW['sugeb']}+{CD[i]}{RW['paseb']}"
      f"+{CD[i]}{RW['nutseb']}" for i in range(5)], F['fp_eb'], NUM0, True)
r += 1
hdr(ws, r, ['Panda (Retail) build'] + YF); r += 1
PAN = F['panda']
rowf('avgst', 'Average stores in year',
     [f"=({a('st25')}+{a('stores',0)})/2"] +
     [f"=({a('stores',i-1)}+{a('stores',i)})/2" for i in range(1, 5)],
     [(IN['stores_end25'] + IN['stores_path'][0]) / 2] +
     [(IN['stores_path'][i-1] + IN['stores_path'][i]) / 2 for i in range(1, 5)], NUM1)
rowf('sps', 'Sales per average store (SAR mn)',
     [f"={a('retrev25')}/(({a('st24')}+{a('st25')})/2)*(1+{a('spsg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('spsg',i)})" for i in range(1, 5)],
     [PAN['rev'][i] / ((IN['stores_end25'] if i == 0 else IN['stores_path'][i-1])
                       + IN['stores_path'][i]) * 2 for i in range(5)], NUM2)
rowf('panrev', 'Panda revenue (= stores x sales/store)',
     [f"={CD[i]}{RW['avgst']}*{CD[i]}{RW['sps']}" for i in range(5)], PAN['rev'])
rowf('pangp', 'Panda gross profit (margin held at the H1-2026 actual)',
     [f"={CD[i]}{RW['panrev']}*{a('pgm')}" for i in range(5)], PAN['gp'])
rowf('panopx', 'Panda store-opex ratio (Framing A scale gains from FY2028)',
     [f"={a('popex')}", f"={a('popex')}", f"={a('popex')}-{a('pstep')}",
      f"={a('popex')}-2*{a('pstep')}", f"={a('popex')}-3*{a('pstep')}"],
     [IN['panda_opex_ratio'], IN['panda_opex_ratio'],
      IN['panda_opex_ratio'] - IN['panda_scale_step'],
      IN['panda_opex_ratio'] - 2 * IN['panda_scale_step'],
      IN['panda_opex_ratio'] - 3 * IN['panda_scale_step']], PCT2)
rowf('paneb', 'Panda EBITDA (an output)',
     [f"={CD[i]}{RW['pangp']}-{CD[i]}{RW['panrev']}*{CD[i]}{RW['panopx']}"
      for i in range(5)], PAN['eb'], NUM0, True)
r += 1
hdr(ws, r, ['Food Services, Frozen Food, eliminations, group'] + YF); r += 1
rowf('herrev', 'Herfy revenue',
     [f"={a('fsvrev25')}*(1+{a('herg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('herg',i)})" for i in range(1, 5)], F['herfy']['rev'])
rowf('hereb', 'Herfy EBITDA',
     [f"={CD[i]}{RW['herrev']}*{a('hermg')}" for i in range(5)], F['herfy']['eb'])
rowf('frzrev', 'Al Kabeer revenue',
     [f"={a('frzrev25')}*(1+{a('frzg',0)})"] +
     [f"={CD[i-1]}{r}*(1+{a('frzg',i)})" for i in range(1, 5)], F['frozen']['rev'])
rowf('frzeb', 'Al Kabeer EBITDA',
     [f"={CD[i]}{RW['frzrev']}*{a('frzmg')}" for i in range(5)], F['frozen']['eb'])
rowf('elim', 'Eliminations (proportional to Food Processing)',
     [f"={CD[i]}{RW['fprev']}*{a('elim')}" for i in range(5)], F['elim'], NUM0)
band(ws, r, 6)
rowf('grev', 'Group revenue',
     [f"={CD[i]}{RW['fprev']}+{CD[i]}{RW['panrev']}+{CD[i]}{RW['herrev']}"
      f"+{CD[i]}{RW['frzrev']}+{CD[i]}{RW['elim']}+{a('invsegrev')}" for i in range(5)],
     F['rev'], NUM0, True)
rowf('unal', 'Unallocated corporate costs',
     [f"=-{a('unalloc',i)}" for i in range(5)], [-u for u in IN['unalloc_path']], NUM0)
band(ws, r, 6)
rowf('gebitda', 'Group operating EBITDA (output)',
     [f"={CD[i]}{RW['fpeb']}+{CD[i]}{RW['paneb']}+{CD[i]}{RW['hereb']}"
      f"+{CD[i]}{RW['frzeb']}+{CD[i]}{RW['unal']}" for i in range(5)],
     F['ebitda'], NUM0, True)
rowf('gmgn', 'Group EBITDA margin (output, never an input)',
     [f"={CD[i]}{RW['gebitda']}/{CD[i]}{RW['grev']}" for i in range(5)],
     F['ebitda_margin'], PCT)
r += 1
hdr(ws, r, ['FY2025 gross-profit footing (audited)', 'SAR mn']); r += 1
FOOT25 = r
for lab, fml, xp in [
        ('Oil gross profit FY2025 (disclosed)', None, IN['oil_gp_fy25']),
        ('Sugar gross profit FY2025 (disclosed)', None, IN['sug_gp_fy25']),
        ('Pasta gross profit FY2025 = disclosed revenue x disclosed margin',
         f"={a('pasr25')}*0.215", SEG['categories']['pasta']['gp']),
        ('Nuts & spices gross profit FY2025 (residual)', None,
         SEG['categories']['nuts']['gp'])]:
    put(ws, f'A{r}', lab, fmt=None)
    if fml:
        putf(ws, f'B{r}', fml, xp, NUM1)
    else:
        put(ws, f'B{r}', xp, BLUE, NUM1)
    r += 1
put(ws, f'A{r}', 'Sum of the four categories', bold=True, fmt=None)
putf(ws, f'B{r}', f"=SUM(B{FOOT25}:B{r-1})", SEG['fp']['gp'], NUM1, bold=True)
r += 1
put(ws, f'A{r}', 'Audited Food-Processing segment gross profit (revenue less cost of '
    'revenues, note 33)', fmt=None)
put(ws, f'B{r}', SEG['fp']['gp'], BLUE, NUM1)
r += 1
put(ws, f'A{r}', 'The pasta line ties to the deck\'s own per-tonne label: 117.2mn / 263k '
    'tonnes = SAR 445/tonne (FY2024: 102 / 232 = 440).', fmt=None).font = SUB
r += 1
put(ws, f'A{r}', 'MARGIN STATUS BY SEGMENT — stated plainly: Food Processing (47% of '
    'FY2026E EBITDA) is unit-built and its margins are OUTPUTS; Herfy and Al Kabeer '
    'margins are INPUTS at the finest level their disclosures allow (flagged above); '
    'Panda\'s margin is the identity of its gross-margin and opex-ratio inputs. The '
    'GROUP margin is an output of the mix.', fmt=None, wrap=True).font = SUB
ws.row_dimensions[r].height = 40
r += 1
put(ws, f'A{r}', 'FY2025 bases shown on Assumptions; H1-2026 actuals in the study text. '
    'The four category revenues foot to the audited Food-Processing segment revenue '
    '(13,279.9) and the four gross profits to the audited segment gross profit (1,689.2), '
    'as shown in the footing block above.', fmt=None).font = SUB

# ============ 10 BALANCE SHEET (built before DCF so links exist) ===============
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — audited FY2025 and the forecast roll-forward',
      'The sheet FOOTS: the check row at the bottom is assets less liabilities less equity '
      'and evaluates to zero in every year', 8, awidth=44, cwidth=13)
FCOL = ['E', 'F', 'G', 'H', 'I']
hdr(ws, 4, ['SAR mn', '', '', 'FY2025', *YF])
r = 5
BS = {}

def bsrow(key, label, hist, fmls, exps, fmt=NUM0, bold_=False):
    global r
    put(ws, f'A{r}', label, fmt=None, bold=bold_)
    if hist is not None:
        put(ws, f'D{r}', hist, BLUE, fmt, bold=bold_)
    for i in range(5):
        if fmls is not None:
            putf(ws, f'{FCOL[i]}{r}', fmls[i], exps[i], fmt, bold=bold_)
    BS[key] = r
    r += 1

bsrow('ppe', 'Property, plant and equipment (opening + capex − depreciation)',
      IN['ppe_fy25'],
      [f"=D{r}+{a('capex',0)}-{a('intam')}-'Balance Sheet'!D{r}*{a('deprate')}"] +
      [f"={FCOL[i-1]}{r}+{a('capex',i)}-{a('intam')}-{FCOL[i-1]}{r}*{a('deprate')}"
       for i in range(1, 5)], F['ppe'])
bsrow('rou', 'Right-of-use assets (grow with the lease book)', IN['rou_fy25'],
      [f"=D{r}+{a('lease')}*{a('roug',0)}"] +
      [f"={FCOL[i-1]}{r}+'Balance Sheet'!{FCOL[i-1]}{BS['ppe']+13}*{a('roug',i)}"
       for i in range(1, 5)], F['rou'])
bsrow('intang', 'Intangible assets and goodwill (held: capex = amortisation)',
      IN['intang_fy25'], [f"={a('intang25')}"] * 5, [IN['intang_fy25']] * 5)
bsrow('invprop', 'Investment property (held)', IN['invprop_fy25'],
      [f"={a('invprop')}"] * 5, [IN['invprop_fy25']] * 5)
bsrow('kinan', 'Kinan (equity method: + share of results − dividends)', IN['kinan_carry'],
      [f"=D{r}+'Income Statement'!{FCOL[0]}14-{a('kindiv')}*(1+{a('king')})^0"] +
      [f"={FCOL[i-1]}{r}+'Income Statement'!{FCOL[i]}14-{a('kindiv')}*(1+{a('king')})^{i}"
       for i in range(1, 5)], F['kinan_bv'])
bsrow('invnc', 'Non-current investments (held)', IN['inv_nc_fy25'],
      [f"={a('invnc')}"] * 5, [IN['inv_nc_fy25']] * 5)
bsrow('invc', 'Current investments (held)', IN['inv_c_fy25'],
      [f"={a('invc')}"] * 5, [IN['inv_c_fy25']] * 5)
bsrow('tiryaki', 'Tiryaki stake (the 274.6 sale-proceeds receivable reclassified from '
      'working capital; settled in shares H1-2026)', IN['tiryaki_recv'],
      [f"={a('tiryaki')}"] * 5, [IN['tiryaki_recv']] * 5)
# working capital components
_cogsr = 1 - HI['FY25']['gp'] / IN['rev_fy25']
_revp = F['rev']
_cogsp = [rv * _cogsr for rv in _revp]
_invp = [c * IN['dio_fy25'] / 365 for c in _cogsp]
_trp = [rv * IN['dso_fy25'] / 365 for rv in _revp]
_prep = [rv * IN['prepay_ratio'] for rv in _revp]
_tpp = [c * IN['dpo_fy25'] / 365 for c in _cogsp]
_accp = [rv * IN['accrued_ratio'] for rv in _revp]
_conp = [rv * IN['contract_ratio'] for rv in _revp]
put(ws, f'A{r}', 'Cost of revenues ratio (measured FY2025, drives inventory and payables)',
    fmt=None)
putf(ws, f'D{r}', f"=1-{a('gp25')}/{a('rev25')}", _cogsr, PCT2)
COGSR = r; r += 1
bsrow('inv', 'Inventories (days on cost of revenues)', IN['inventories_fy25'],
      [f"=Segments!{CD[i]}{RW['grev']}*'Balance Sheet'!$D${COGSR}*{a('dio')}/365"
       for i in range(5)], _invp)
bsrow('tr', 'Trade receivables (days on revenue)', IN['tr_fy25'],
      [f"=Segments!{CD[i]}{RW['grev']}*{a('dso')}/365" for i in range(5)], _trp)
bsrow('prep', 'Prepayments and other receivables, EX the Tiryaki receivable '
      '(audited 1,346.5 less 274.6; % of revenue in the forecast)',
      IN['prepay_fy25'] - IN['tiryaki_recv'],
      [f"=Segments!{CD[i]}{RW['grev']}*{a('prepr')}" for i in range(5)], _prep)
bsrow('tp', 'Trade payables (days on cost of revenues)', -IN['tp_fy25'],
      [f"=-Segments!{CD[i]}{RW['grev']}*'Balance Sheet'!$D${COGSR}*{a('dpo')}/365"
       for i in range(5)], [-x for x in _tpp])
bsrow('acc', 'Accrued and other liabilities (% of revenue)', -IN['accrued_fy25'],
      [f"=-Segments!{CD[i]}{RW['grev']}*{a('accrr')}" for i in range(5)],
      [-x for x in _accp])
bsrow('con', 'Contract liabilities (% of revenue)', -IN['contract_fy25'],
      [f"=-Segments!{CD[i]}{RW['grev']}*{a('contr')}" for i in range(5)],
      [-x for x in _conp])
_nwc0 = HB['FY25']['nwc']
bsrow('nwc', 'Net working capital', _nwc0,
      [f"=SUM({FCOL[i]}{BS['inv']}:{FCOL[i]}{BS['con']})" for i in range(5)], F['nwc'],
      NUM0, True)
bsrow('cash', 'Cash and cash equivalents (from the cash-flow walk)', IN['cash_fy25'],
      [f"='Cash Flow'!{CD[i]}17" for i in range(5)], F['cash'])
band(ws, r, 9)
bsrow('ta', 'Total assets (condensed)', None,
      [f"={FCOL[i]}{BS['ppe']}+{FCOL[i]}{BS['rou']}+{FCOL[i]}{BS['intang']}"
       f"+{FCOL[i]}{BS['invprop']}+{FCOL[i]}{BS['kinan']}+{FCOL[i]}{BS['invnc']}"
       f"+{FCOL[i]}{BS['invc']}+{FCOL[i]}{BS['tiryaki']}+{FCOL[i]}{BS['nwc']}"
       f"+{FCOL[i]}{BS['cash']}" for i in range(5)],
      [F['ppe'][i] + F['rou'][i] + IN['intang_fy25'] + IN['invprop_fy25'] + F['kinan_bv'][i]
       + IN['inv_nc_fy25'] + IN['inv_c_fy25'] + IN['tiryaki_recv'] + F['nwc'][i]
       + F['cash'][i] for i in range(5)], NUM0, True)
bsrow('loans', 'Loans and borrowings (held at the FY2025 level)', IN['loans_fy25'],
      [f"={a('loans')}"] * 5, [IN['loans_fy25']] * 5)
bsrow('leases', 'Lease liabilities (grow with the store network)', IN['leases_fy25'],
      [f"={a('lease')}*(1+{a('roug',0)})"] +
      [f"={FCOL[i-1]}{r}*(1+{a('roug',i)})" for i in range(1, 5)], F['leases'])
bsrow('eb', 'Employee benefits liabilities (held)', IN['eb_fy25'],
      [f"={a('eb')}"] * 5, [IN['eb_fy25']] * 5)
bsrow('restor', 'Asset-restoration provision (held)', IN['restor_fy25'],
      [f"={a('restor')}"] * 5, [IN['restor_fy25']] * 5)
bsrow('othnl', 'Other net liabilities (tax/zakat accruals + DTL − DTA − other, held)',
      IN['other_net_liab'], [f"={a('othnl')}"] * 5, [IN['other_net_liab']] * 5)
bsrow('eq', 'Equity attributable to owners (+ profit − dividends)', IN['equity_att_fy25'],
      [f"=D{r}+'Income Statement'!{FCOL[0]}16-'Cash Flow'!{CD[0]}13*-1"] +
      [f"={FCOL[i-1]}{r}+'Income Statement'!{FCOL[i]}16-'Cash Flow'!{CD[i]}13*-1"
       for i in range(1, 5)], F['equity_att'])
bsrow('nci', 'Non-controlling interests (+ NCI profit − NCI dividends)',
      IN['nci_book_fy25'],
      [f"=D{r}+'Income Statement'!{FCOL[0]}15-'Cash Flow'!{CD[0]}14*-1"] +
      [f"={FCOL[i-1]}{r}+'Income Statement'!{FCOL[i]}15-'Cash Flow'!{CD[i]}14*-1"
       for i in range(1, 5)], F['nci'])
band(ws, r, 9)
bsrow('foot', 'FOOT CHECK: assets − liabilities − equity (must be zero)', None,
      [f"={FCOL[i]}{BS['ta']}-{FCOL[i]}{BS['loans']}-{FCOL[i]}{BS['leases']}"
       f"-{FCOL[i]}{BS['eb']}-{FCOL[i]}{BS['restor']}-{FCOL[i]}{BS['othnl']}"
       f"-{FCOL[i]}{BS['eq']}-{FCOL[i]}{BS['nci']}" for i in range(5)],
      [0.0] * 5, NUM2, True)
bsrow('nd', 'Net debt (loans − cash − current investments; company definition)',
      IN['loans_fy25'] - IN['cash_fy25'] - IN['inv_c_fy25'],
      [f"={FCOL[i]}{BS['loans']}-{FCOL[i]}{BS['cash']}-{FCOL[i]}{BS['invc']}"
       for i in range(5)], F['netdebt'])
# fix the rou row's second+ formulas to reference the leases row (now known)
for i in range(1, 5):
    ws[f"{FCOL[i]}{BS['rou']}"] = (f"={FCOL[i-1]}{BS['rou']}"
                                   f"+'Balance Sheet'!{FCOL[i-1]}{BS['leases']}*{a('roug',i)}")
# single-source rule: the FY2025 column links to Assumptions instead of duplicating it,
# so a change to an audited base on Assumptions flows through the whole roll-forward
for key, akey, xp in [('ppe', 'ppe25', IN['ppe_fy25']), ('rou', 'rou25', IN['rou_fy25']),
                      ('intang', 'intang25', IN['intang_fy25']),
                      ('invprop', 'invprop', IN['invprop_fy25']),
                      ('kinan', 'kinbv', IN['kinan_carry']),
                      ('invnc', 'invnc', IN['inv_nc_fy25']),
                      ('invc', 'invc', IN['inv_c_fy25']),
                      ('tiryaki', 'tiryaki', IN['tiryaki_recv']),
                      ('cash', 'cash', IN['cash_fy25']), ('loans', 'loans', IN['loans_fy25']),
                      ('leases', 'lease', IN['leases_fy25']), ('eb', 'eb', IN['eb_fy25']),
                      ('restor', 'restor', IN['restor_fy25']),
                      ('othnl', 'othnl', IN['other_net_liab']),
                      ('eq', 'eq25', IN['equity_att_fy25']),
                      ('nci', 'ncib', IN['nci_book_fy25'])]:
    putf(ws, f"D{BS[key]}", f"={a(akey)}", xp, NUM0, green=True)
putf(ws, f"D{BS['nd']}", f"={a('loans')}-{a('cash')}-{a('invc')}",
     IN['loans_fy25'] - IN['cash_fy25'] - IN['inv_c_fy25'], NUM0, green=True)

# ============ 9 INCOME STATEMENT ==============================================
ws = sheet('Income Statement')
title(ws, 'Income statement — three audited years and the five-year forecast',
      'FY2023 is the continuing basis of the FY2024 statements and still includes Turkiye '
      '(disposed 2025); FY2024-25 are the FY2025 audited basis (ex-Turkiye)', 9,
      awidth=46, cwidth=13)
hdr(ws, 4, ['SAR mn', *YH, *YF])
HC = ['B', 'C', 'D']
H23, H24, H25 = HI['FY23'], HI['FY24'], HI['FY25']
IS = dict(rev=5, cogs=6, gp=7, opx=8, ebitda=9, dna=10, ebit=11, nf=12, core=13,
          kinan=14, nci=15, np=16, eps=17)
_gp_h = [H23['gp'], H24['gp'], H25['gp']]
_opx_h = [-(H['sda'] + H['adm'] - H['oth']) for H in (H23, H24, H25)]
_eb_h = [H23['ebitda'], H24['ebitda'], H25['ebitda']]

def isput(row, label, hist_vals, hist_fmls, fc_fmls, fc_exps, fmt=NUM0, bold_=False):
    put(ws, f'A{row}', label, fmt=None, bold=bold_)
    for j in range(3):
        if hist_fmls and hist_fmls[j] is not None:
            putf(ws, f'{HC[j]}{row}', hist_fmls[j],
                 hist_vals[j] if hist_vals else None, fmt, bold=bold_)
        elif hist_vals and hist_vals[j] is not None:
            put(ws, f'{HC[j]}{row}', hist_vals[j], BLUE, fmt, bold=bold_)
    for i in range(5):
        if fc_fmls is not None:
            putf(ws, f'{FCOL[i]}{row}', fc_fmls[i],
                 fc_exps[i] if fc_exps else None, fmt, bold=bold_)

isput(IS['rev'], 'Revenues', [H23['rev'], H24['rev'], H25['rev']], None,
      [f"=Segments!{CD[i]}{RW['grev']}" for i in range(5)], F['rev'], NUM0, True)
isput(IS['cogs'], 'Cost of revenues (audited; forecast modelled at category level)',
      [-H23['cogs'], -H24['cogs'], -H25['cogs']], None, None, None)
isput(IS['gp'], 'Gross profit (audited columns)', _gp_h,
      [f"={HC[j]}{IS['rev']}+{HC[j]}{IS['cogs']}" for j in range(3)], None, None, NUM0,
      True)
isput(IS['opx'], 'Operating costs before D&A (selling, admin, net other; forecast = '
      'revenue less EBITDA)', _opx_h, None,
      [f"=-({FCOL[i]}{IS['rev']}-{FCOL[i]}{IS['ebitda']})" for i in range(5)],
      [-(F['rev'][i] - F['ebitda'][i]) for i in range(5)])
isput(IS['ebitda'], 'Operating EBITDA (before associates and impairments)', _eb_h,
      [f"={HC[j]}{IS['gp']}+{HC[j]}{IS['opx']}+{HC[j]}{IS['dna']}" for j in range(3)],
      [f"=Segments!{CD[i]}{RW['gebitda']}" for i in range(5)], F['ebitda'], NUM0, True)
isput(IS['dna'], 'Depreciation and amortisation',
      [H23['dna'], H24['dna'], H25['dna']], None,
      [f"='Balance Sheet'!{'D' if i == 0 else FCOL[i-1]}{BS['ppe']}*{a('deprate')}"
       f"+{a('intam')}+'Cash Flow'!{CD[i]}19" for i in range(5)], F['dna'])
isput(IS['ebit'], 'EBIT', _eb_h and [None, None, None],
      [f"={HC[j]}{IS['ebitda']}-{HC[j]}{IS['dna']}" for j in range(3)],
      [f"={FCOL[i]}{IS['ebitda']}-{FCOL[i]}{IS['dna']}" for i in range(5)], F['ebit'],
      NUM0, True)
for j, v in enumerate([_eb_h[j] - [H23['dna'], H24['dna'], H25['dna']][j]
                       for j in range(3)]):
    EXPECT.setdefault('Income Statement', {})[f'{HC[j]}{IS["ebit"]}'] = v
for j, v in enumerate(_gp_h):
    EXPECT.setdefault('Income Statement', {})[f'{HC[j]}{IS["gp"]}'] = v
for j, v in enumerate(_eb_h):
    EXPECT.setdefault('Income Statement', {})[f'{HC[j]}{IS["ebitda"]}'] = v
isput(IS['nf'], 'Net finance cost (on prior-year net debt and lease book)',
      [None, None, -275.525], None,
      ["=0"] * 5, F['netfin'])          # placeholder — repointed after the DCF rows exist
isput(IS['core'], 'Core profit after zakat and tax (= (EBIT + net finance) x (1 − rate))',
      None, None,
      [f"=({FCOL[i]}{IS['ebit']}+{FCOL[i]}{IS['nf']})*(1-{a('tax')})" for i in range(5)],
      [(F['ebit'][i] + F['netfin'][i]) * (1 - T) for i in range(5)])
isput(IS['kinan'], 'Share of results of Kinan (net of zakat and tax)',
      [None, None, H25['assoc']], None,
      [f"={a('kinh1')}*2*(1+{a('king')})^{i}" for i in range(5)], F['kinan'], NUM1)
isput(IS['nci'], 'Profit attributable to non-controlling interests',
      [None, None, -66.037], None,
      [f"={FCOL[i]}{IS['core']}*{a('ncis')}" for i in range(5)], F['np_nci'], NUM1)
isput(IS['np'], 'Profit attributable to owners',
      [H23['np_att'], H24['np_att'], H25['np_att']], None,
      [f"={FCOL[i]}{IS['core']}*(1-{a('ncis')})+{FCOL[i]}{IS['kinan']}" for i in range(5)],
      F['np'], NUM0, True)
isput(IS['eps'], 'Earnings per share (SAR)', [None, None, H25['eps']], None,
      [f"={FCOL[i]}{IS['np']}/{a('sharesw')}" for i in range(5)], F['eps'], PX)
put(ws, f"A{IS['eps'] + 1}", 'FY2024 attributable profit includes the SAR 11,554.7mn '
    'Almarai distribution gain (audited FY2024 statements, segment note). FY2025 recurring '
    'net profit is 539.1 per the company\'s own bridge, whose items are a 300.0 zakat-and-'
    'other-accrual reversal, a 32.3 Turkiye gain and a 40.2 put-option gain (the audited '
    'FS note 29 zakat credit is 217.4; the announcement quotes 247.3 gross of related '
    'expenses — three disclosed figures on three bases). The forecast is the recurring '
    'construction.', fmt=None).font = SUB
IS_ROWS = IS

# ============ 11 CASH FLOW =====================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow — the walk that closes the balance sheet',
      'Lease principal equals right-of-use depreciation; new store leases are non-cash '
      'additions', 7, awidth=52, cwidth=13)
hdr(ws, 4, ['SAR mn', *YF])
r = 5
CF = {}

def cfrow(key, label, fmls, exps, fmt=NUM0, bold_=False):
    global r
    put(ws, f'A{r}', label, fmt=None, bold=bold_)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i] if exps else None, fmt, bold=bold_)
    CF[key] = r
    r += 1

cfrow('core', 'Core profit after zakat and tax',
      [f"='Income Statement'!{FCOL[i]}{IS['core']}" for i in range(5)],
      [(F['ebit'][i] + F['netfin'][i]) * (1 - T) for i in range(5)], NUM0)
cfrow('dna', 'Add back depreciation and amortisation',
      [f"='Income Statement'!{FCOL[i]}{IS['dna']}" for i in range(5)], F['dna'])
cfrow('dwc', 'Working-capital investment (increase is an outflow)',
      [f"=-('Balance Sheet'!{FCOL[0]}{BS['nwc']}-'Balance Sheet'!D{BS['nwc']})"] +
      [f"=-('Balance Sheet'!{FCOL[i]}{BS['nwc']}-'Balance Sheet'!{FCOL[i-1]}{BS['nwc']})"
       for i in range(1, 5)], [-x for x in F['dwc']])
cfrow('kdiv', 'Dividend received from Kinan',
      [f"={a('kindiv')}*(1+{a('king')})^{i}" for i in range(5)], F['kinan_div'], NUM1)
band(ws, r, 6)
cfrow('cfo', 'Net cash from operating activities',
      [f"={CD[i]}{CF['core']}+{CD[i]}{CF['dna']}+{CD[i]}{CF['dwc']}+{CD[i]}{CF['kdiv']}"
       for i in range(5)], F['cfo'], NUM0, True)
cfrow('capex', 'Capital expenditure (property, plant, equipment and intangibles)',
      [f"=-{a('capex',i)}" for i in range(5)], [-c for c in IN['capex_path']])
band(ws, r, 6)
cfrow('cfi', 'Net cash used in investing activities',
      [f"={CD[i]}{CF['capex']}" for i in range(5)], [-c for c in IN['capex_path']], NUM0,
      True)
cfrow('leasep', 'Payment of lease liabilities — principal (= right-of-use depreciation)',
      [f"=-'Cash Flow'!{CD[i]}19" for i in range(5)], [-x for x in F['dna_rou']])
cfrow('div', 'Dividends paid to owners (policy payout on the current year\'s profit)',
      [f"=-'Income Statement'!{FCOL[i]}{IS['np']}*{a('payout')}" for i in range(5)],
      [-x for x in F['div']])
cfrow('divnci', 'Dividends paid to non-controlling interests',
      [f"=-'Income Statement'!{FCOL[i]}{IS['nci']}*{a('payout')}" for i in range(5)],
      [-x for x in F['div_nci']])
band(ws, r, 6)
cfrow('cff', 'Net cash used in financing activities',
      [f"={CD[i]}{CF['leasep']}+{CD[i]}{CF['div']}+{CD[i]}{CF['divnci']}" for i in range(5)],
      [-(F['dna_rou'][i] + F['div'][i] + F['div_nci'][i]) for i in range(5)], NUM0, True)
cfrow('net', 'Net change in cash',
      [f"={CD[i]}{CF['cfo']}+{CD[i]}{CF['cfi']}+{CD[i]}{CF['cff']}" for i in range(5)],
      [F['cash'][i] - (IN['cash_fy25'] if i == 0 else F['cash'][i-1]) for i in range(5)])
cfrow('close', 'Cash and cash equivalents at year end',
      [f"={a('cash')}+{CD[0]}{CF['net']}"] +
      [f"={CD[i-1]}{r}+{CD[i]}{CF['net']}" for i in range(1, 5)], F['cash'], NUM0, True)
r += 1
cfrow('roud', 'Memo: right-of-use depreciation (grows with the store network)',
      [f"={a('roud25')}*(1+{a('roug',0)})"] +
      [f"={CD[i-1]}{r}*(1+{a('roug',i)})" for i in range(1, 5)], F['dna_rou'], NUM1)
put(ws, f'A{r}', 'The dividend row applies the stated 50-60% policy (midpoint) to the '
    'current year\'s attributable profit; Savola\'s actual FY2025 dividend of SAR 510mn was '
    'paid during H1-2026.', fmt=None).font = SUB
# cash walk row index used by the Balance Sheet: CF['close'] must be row 17
assert CF['close'] == 17, CF['close']
assert CF['roud'] == 19, CF['roud']

# ============ 8 DCF ============================================================
ws = sheet('DCF')
title(ws, 'DCF — the waterfall, the cost of capital built in-sheet, and the terminal',
      'NOPAT = EBIT x (1 − combined zakat/tax); lease replacement equals right-of-use '
      'depreciation; leases carry a debt weight', 7, awidth=52, cwidth=13)
hdr(ws, 4, ['SAR mn', *YF])
r = 5
DC = {}

def dcrow(key, label, fmls, exps, fmt=NUM0, bold_=False, green=False):
    global r
    put(ws, f'A{r}', label, fmt=None, bold=bold_)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i] if exps else None, fmt, bold=bold_,
             green=green)
    DC[key] = r
    r += 1

dcrow('rev', 'Revenue', [f"=Segments!{CD[i]}{RW['grev']}" for i in range(5)], F['rev'],
      NUM0, False, True)
dcrow('ebitda', 'EBITDA', [f"=Segments!{CD[i]}{RW['gebitda']}" for i in range(5)],
      F['ebitda'], NUM0, False, True)
dcrow('mgn', 'EBITDA margin', [f"={CD[i]}{DC['ebitda']}/{CD[i]}{DC['rev']}"
      for i in range(5)], F['ebitda_margin'], PCT)
dcrow('dna', 'Less depreciation and amortisation',
      [f"=-'Income Statement'!{FCOL[i]}{IS['dna']}" for i in range(5)],
      [-x for x in F['dna']])
dcrow('ebit', 'EBIT', [f"={CD[i]}{DC['ebitda']}+{CD[i]}{DC['dna']}" for i in range(5)],
      F['ebit'], NUM0, True)
dcrow('nopat', 'NOPAT = EBIT x (1 − combined zakat/tax rate)',
      [f"={CD[i]}{DC['ebit']}*(1-{a('tax')})" for i in range(5)], F['nopat'])
dcrow('adddna', 'Add back depreciation and amortisation',
      [f"=-{CD[i]}{DC['dna']}" for i in range(5)], F['dna'])
dcrow('capex', 'Less capital expenditure', [f"=-{a('capex',i)}" for i in range(5)],
      [-c for c in IN['capex_path']])
dcrow('leaser', 'Less lease renewals (= right-of-use depreciation)',
      [f"=-'Cash Flow'!{CD[i]}{CF['roud']}" for i in range(5)], [-x for x in F['dna_rou']])
dcrow('leaseg', 'Less lease-book growth (new-store leases; full additions charged)',
      [f"=-('Balance Sheet'!{FCOL[0]}{BS['leases']}-{a('lease')})"] +
      [f"=-('Balance Sheet'!{FCOL[i]}{BS['leases']}-'Balance Sheet'!{FCOL[i-1]}{BS['leases']})"
       for i in range(1, 5)], [-x for x in F['dlease']])
dcrow('dwc', 'Less increase in net working capital',
      [f"='Cash Flow'!{CD[i]}{CF['dwc']}" for i in range(5)], [-x for x in F['dwc']])
band(ws, r, 6)
dcrow('fcff', 'Free cash flow to the firm',
      [f"={CD[i]}{DC['nopat']}+{CD[i]}{DC['adddna']}+{CD[i]}{DC['capex']}"
       f"+{CD[i]}{DC['leaser']}+{CD[i]}{DC['leaseg']}+{CD[i]}{DC['dwc']}"
       for i in range(5)], F['fcff'], NUM0, True)
dcrow('df', 'Discount factor (compounding at the explicit cost of capital)',
      ["=1/(1+$C$47)"] + [f"={CD[i-1]}{r}/(1+$C$47)" for i in range(1, 5)], DCF['dfs'], DF4)
dcrow('pv', 'Present value of free cash flow',
      [f"={CD[i]}{DC['fcff']}*{CD[i]}{DC['df']}" for i in range(5)],
      [F['fcff'][i] * DCF['dfs'][i] for i in range(5)], NUM0)
r += 1
put(ws, f'A{r}', 'TERMINAL BLOCK', bold=True, fmt=None); r += 1
TB = {}
for key, label, fml, xp, fmt in [
        ('greal', 'Terminal REAL growth (stated)', None, IN['g_term_real'], PCT),
        ('pit', 'Terminal inflation — Saudi house macro path', None, TRI['inflation'], PCT),
        ('g', 'Terminal growth = (1+inflation)(1+real growth) − 1', None,
         TRI['nominal_growth'], PCT),
        ('nopatT', 'Terminal operating profit after tax (year five grown one year)',
         None, None, NUM0),
        ('dnaT', 'Plus owned and intangible depreciation, grown one year (the right-of-use '
         'charge cancels out of this model\'s own free cash flow and is not added back '
         'here)', None, TRI['dna_book'], NUM0),
        ('maintT', 'Less capital maintenance at replacement cost — that charge escalated '
         'over half the derived asset life', None, -TRO['maintenance'], NUM0),
        ('gcapT', 'Less capital for real growth (real growth x capital per unit of growth)',
         None, -TRO['growth_capex'], NUM0),
        ('wcT', 'Less inflation on the working capital and lease book the group carries',
         None, -TRO['wc_charge'], NUM0),
        ('fcffT', 'Terminal free cash flow', None, TRO['fcff'], NUM0),
        ('tv', 'Terminal value at end of year five', None, DCF['tv'], NUM0),
        ('floorT', 'Memo — the no-growth perpetuity at book depreciation (a diagnostic, '
         'not a bound)', None, TRO['floor'], NUM0),
        ('pvtv', 'Present value of the terminal value', None, DCF['pv_tv'], NUM0),
        ('pvexp', 'Present value of the five explicit years', None, DCF['pv_explicit'],
         NUM0),
        ('tvshare', 'Terminal value share of enterprise value', None, DCF['tv_share'], PCT),
        ('ev', 'Enterprise value (operating)', None, DCF['ev'], NUM0)]:
    TB[key] = r
    put(ws, f'A{r}', label, fmt=None, bold=key in ('ev', 'fcffT'))
    r += 1
putf(ws, f"C{TB['greal']}", f"={a('greal')}", IN['g_term_real'], PCT)
putf(ws, f"C{TB['pit']}", f"={a('pit')}", TRI['inflation'], PCT)
putf(ws, f"C{TB['g']}", f"=(1+C{TB['pit']})*(1+C{TB['greal']})-1", TRI['nominal_growth'], PCT)
putf(ws, f"C{TB['nopatT']}", f"=F{DC['nopat']}*(1+C{TB['g']})", TRI['nopat'], NUM0)
putf(ws, f"C{TB['dnaT']}", f"={DCF['dna_oi_last']}*(1+C{TB['g']})", TRI['dna_book'], NUM0)
putf(ws, f"C{TB['maintT']}", f"=-C{TB['dnaT']}*(1+C{TB['pit']})^({a('life')}/2)",
     -TRO['maintenance'], NUM0)
putf(ws, f"C{TB['gcapT']}", f"=-C{TB['greal']}*{a('inccap')}", -TRO['growth_capex'], NUM0)
putf(ws, f"C{TB['wcT']}", f"=-C{TB['pit']}*{DCF['wc_lease_last']:.6f}*(1+C{TB['g']})",
     -TRO['wc_charge'], NUM0)
putf(ws, f"C{TB['fcffT']}", f"=SUM(C{TB['nopatT']}:C{TB['wcT']})", TRO['fcff'], NUM0,
     bold=True)
# Placeholders: both are RE-POINTED at the real terminal-cost-of-capital row once the
# cost-of-capital block below has been placed. The expected values recorded here are the
# ones the re-pointed formulas must reproduce, so a re-point that lands on the wrong row
# fails the recalculation gate rather than passing quietly.
putf(ws, f"C{TB['tv']}", f"=C{TB['fcffT']}*(1+C{TB['g']})/($C$54-C{TB['g']})", DCF['tv'],
     NUM0)
putf(ws, f"C{TB['floorT']}", f"=C{TB['nopatT']}/$C$54", TRO['floor'], NUM0)
putf(ws, f"C{TB['pvtv']}", f"=C{TB['tv']}*F{DC['df']}", DCF['pv_tv'], NUM0)
putf(ws, f"C{TB['pvexp']}", f"=SUM(B{DC['pv']}:F{DC['pv']})", DCF['pv_explicit'], NUM0)
putf(ws, f"C{TB['tvshare']}", f"=C{TB['pvtv']}/C{TB['ev']}", DCF['tv_share'], PCT)
putf(ws, f"C{TB['ev']}", f"=C{TB['pvexp']}+C{TB['pvtv']}", DCF['ev'], NUM0, bold=True)
r += 1
put(ws, f'A{r}', 'Downside variant — the asset life read at twice the directly measured '
    'average age (engine re-run, published beside the base, never averaged)', fmt=None)
put(ws, f'C{r}', DCF['ps_life_variant'], BLUE, PX)
r += 1
put(ws, f'A{r}', 'COST OF CAPITAL — BUILT, NOT PASTED (v2: rf net of the sovereign spread; '
    'both ERP bases shown)', bold=True, fmt=None); r += 1
CC = {}
cc_rows = [
    ('rfobs', 'Observed 10Y SAR sovereign yield (FTSE SAGBI 7-10y)', f"={a('rf')}", W['rf_observed'], PCT2),
    ('sov', 'Less sovereign default spread (rating basis)', f"=-{a('sov')}",
     -W['sov_spread_rating'], PCT2),
    ('rfstar', 'Normalized risk-free rate rf*', None, W['rf_star_rating'], PCT2),
    ('beta', 'Beta (own stock vs TASI, five-year weekly)', f"={a('beta')}", W['beta'],
     '0.000'),
    ('erp', 'Equity risk premium (rating basis)', f"={a('erp')}", W['erp_rating'], PCT2),
    ('ke', 'Cost of equity = rf* + beta x ERP', None, W['ke_rating'], PCT2),
    ('kdl', 'Cost of loans (blend of SAR / EGP-parity / other tranches)', None,
     W['kd_loans'], PCT2),
    ('kdz', 'Cost of leases (measured effective rate)', f"={a('leaserate')}", W['kd_lease'],
     PCT2),
    ('mcap', 'Market capitalisation = settled spot x shares issued', None, M['mktcap'],
     NUM0),
    ('vtot', 'Capital base = equity (anchor) + loans + leases (both at 30-Jun-2026)',
     None, M['mktcap'] + IN['loans_jun26'] + IN['leases_jun26'], NUM0),
    ('we', 'Equity weight', None, W['we'], PCT2),
    ('wl', 'Loans weight', None, W['wl'], PCT2),
    ('wz', 'Lease weight', None, W['wlease'], PCT2),
    ('wacc', 'Cost of capital — explicit window', None, W['wacc_exp'], PCT2),
    ('kecds', 'Cost of equity, CDS basis = (rf − CDS) + beta x ERP(CDS)', None,
     W['ke_cds'], PCT2),
    ('wacccds', 'Cost of capital — explicit window, CDS basis', None, W['wacc_exp_cds'],
     PCT2),
    ('twe', 'Terminal weights: equity / loans / leases', f"={a('twe')}", IN['tw_e'], PCT),
    ('twl', ' ', f"={a('twl')}", IN['tw_loans'], PCT),
    ('twz', ' ', None, W['tw_lease'], PCT),
    ('wacct', 'Cost of capital — terminal', None, W['wacc_term'], PCT2),
]
for key, label, fml, xp, fmt in cc_rows:
    put(ws, f'A{r}', label, fmt=None)
    if fml is not None:
        putf(ws, f'C{r}', fml, xp, fmt)
    CC[key] = r
    r += 1
putf(ws, f"C{CC['rfstar']}", f"=C{CC['rfobs']}+C{CC['sov']}", W['rf_star_rating'], PCT2)
putf(ws, f"C{CC['ke']}", f"=C{CC['rfstar']}+C{CC['beta']}*C{CC['erp']}", W['ke_rating'],
     PCT2)
putf(ws, f"C{CC['kdl']}",
     f"=({a('lsa')}*{a('kdsar')}+{a('leg')}*{a('kdeg')}+{a('lot')}*{a('kdot')})/{a('loans')}",
     W['kd_loans'], PCT2)
putf(ws, f"C{CC['mcap']}", f"={a('spot')}*{a('shares')}", M['mktcap'], NUM0)
putf(ws, f"C{CC['vtot']}", f"=C{CC['mcap']}+{a('ljun')}+{a('zjun')}",
     M['mktcap'] + IN['loans_jun26'] + IN['leases_jun26'], NUM0)
putf(ws, f"C{CC['we']}", f"=C{CC['mcap']}/C{CC['vtot']}", W['we'], PCT2)
putf(ws, f"C{CC['wl']}", f"={a('ljun')}/C{CC['vtot']}", W['wl'], PCT2)
putf(ws, f"C{CC['wz']}", f"={a('zjun')}/C{CC['vtot']}", W['wlease'], PCT2)
putf(ws, f"C{CC['wacc']}",
     f"=C{CC['we']}*C{CC['ke']}+C{CC['wl']}*C{CC['kdl']}*(1-{a('tax')})"
     f"+C{CC['wz']}*C{CC['kdz']}*(1-{a('tax')})", W['wacc_exp'], PCT2, bold=True)
putf(ws, f"C{CC['kecds']}",
     f"={a('rf')}-{a('sovcds')}+C{CC['beta']}*{a('erpcds')}", W['ke_cds'], PCT2)
putf(ws, f"C{CC['wacccds']}",
     f"=C{CC['we']}*C{CC['kecds']}+C{CC['wl']}*C{CC['kdl']}*(1-{a('tax')})"
     f"+C{CC['wz']}*C{CC['kdz']}*(1-{a('tax')})", W['wacc_exp_cds'], PCT2)
putf(ws, f"C{CC['twz']}", f"=1-{a('twe')}-{a('twl')}", W['tw_lease'], PCT)
putf(ws, f"C{CC['wacct']}",
     f"={a('twe')}*C{CC['ke']}+{a('twl')}*C{CC['kdl']}*(1-{a('tax')})"
     f"+C{CC['twz']}*C{CC['kdz']}*(1-{a('tax')})", W['wacc_term'], PCT2, bold=True)
# now that the cost-of-capital rows are placed, point the discount factors, the
# terminal value and the Income Statement's finance line at the real rows
for i in range(5):
    cell = f'{CD[i]}{DC["df"]}'
    ws[cell] = (f"=1/(1+$C${CC['wacc']})" if i == 0
                else f"={CD[i-1]}{DC['df']}/(1+$C${CC['wacc']})")
ws[f"C{TB['tv']}"] = (f"=C{TB['fcffT']}*(1+C{TB['g']})"
                      f"/($C${CC['wacct']}-C{TB['g']})")
ws[f"C{TB['floorT']}"] = f"=C{TB['nopatT']}/$C${CC['wacct']}"
wsIS = wb['Income Statement']
for i in range(5):
    prevcol = 'D' if i == 0 else FCOL[i - 1]
    wsIS[f"{FCOL[i]}{IS['nf']}"] = (
        f"=MAX('Balance Sheet'!{prevcol}{BS['nd']},0)*-DCF!$C${CC['kdl']}"
        f"-'Balance Sheet'!{prevcol}{BS['leases']}*{a('leaserate')}"
        f"+MAX(-'Balance Sheet'!{prevcol}{BS['nd']},0)*{a('sar1y')}")
r += 1
put(ws, f'A{r}', 'FROM ENTERPRISE VALUE TO THE ANCHOR', bold=True, fmt=None); r += 1
AN = {}
for key, label, fml, xp, fmt, bold_ in [
        ('eq', 'Equity value AT THE ANCHOR (from the bridge: Dec-2025 legs rolled at the '
         'cost of equity; the anchor-dated legs — Kinan, Herfy NCI, Mehbaj — unrolled)',
         "='SOTP Bridge'!C18", DCF['eq_val'], NUM0, False),
        ('psdec', 'Fair value per share on the 31-Dec-2025 basis (anchor legs at their '
         'own dates)', None, DCF['ps_dec'], PX, False),
        ('roll', 'Anchor accretion factor on the Dec legs = (1 + cost of equity)^(days/365)',
         None, DCF['roll'], DF4, False),
        ('ps', 'Fair value per share at the 18-Aug-2026 anchor (ex the 1.70 dividend)',
         None, DCF['ps'], PX, True)]:
    put(ws, f'A{r}', label, fmt=None, bold=bold_)
    if fml:
        putf(ws, f'C{r}', fml, xp, fmt, bold=bold_, green=True)
    AN[key] = r
    r += 1
putf(ws, f"C{AN['roll']}", f"=(1+C{CC['ke']})^({a('anchor_days')}/365)", DCF['roll'], DF4)
putf(ws, f"C{AN['ps']}", f"=C{AN['eq']}/{a('sharesw')}-{a('div_between')}", DCF['ps'], PX,
     bold=True)
put(ws, f'A{r}', f"CDS-basis fair value (engine re-run at the CDS-basis cost of capital; "
    f"the CDS legs are the January-2026 vintage, flagged): SAR {DCF['ps_cds']:.2f}",
    fmt=None).font = SUB

# ============ 5 SOTP BRIDGE ====================================================
ws = sheet('SOTP Bridge')
title(ws, 'Enterprise value to equity — the bridge',
      'Dec-2025 legs roll to the anchor at the cost of equity; the three anchor-dated '
      'legs (Kinan capitalized, Herfy NCI at market, Mehbaj) sit OUTSIDE the roll', 5,
      awidth=56, cwidth=15)
hdr(ws, 4, ['Step', 'SAR mn', 'Per share (SAR)'])
kin_cap = DCF['kinan_capitalized']
brows = [
    ('pvexp', 'Present value of the five forecast years', f"=DCF!C{TB['pvexp']}",
     DCF['pv_explicit'], True),
    ('pvtv', 'Present value of the terminal value', f"=DCF!C{TB['pvtv']}", DCF['pv_tv'],
     True),
    ('ev', 'Enterprise value (operating)', '=C5+C6', DCF['ev'], False),
    ('sukuk', 'Government sukuk and other non-current investments', f"={a('invnc')}",
     IN['inv_nc_fy25'], True),
    ('invc', 'Current investments (T-bills, Almarai FVTPL)', f"={a('invc')}",
     IN['inv_c_fy25'], True),
    ('tiryaki', 'Tiryaki sale-proceeds receivable (on the audited 31-Dec-2025 balance '
     'sheet; carved out of working capital)', f"={a('tiryaki')}", IN['tiryaki_recv'], True),
    ('invprop', 'Investment property (its inter-segment rent is OUTSIDE group EBITDA, '
     'so the asset belongs here)', f"={a('invprop')}", IN['invprop_fy25'], True),
    ('cash', 'Cash and cash equivalents', f"={a('cash')}", IN['cash_fy25'], True),
    ('loans', 'Less loans and borrowings', f"=-{a('loans')}", -IN['loans_fy25'], True),
    ('lease', 'Less lease liabilities', f"=-{a('lease')}", -IN['leases_fy25'], True),
    ('eb', 'Less employee benefits liabilities', f"=-{a('eb')}", -IN['eb_fy25'], True),
    ('restor', 'Less asset-restoration provision', f"=-{a('restor')}", -IN['restor_fy25'],
     True),
    ('othnl', 'Less other net liabilities (tax/zakat accruals, deferred tax, net)',
     f"=-{a('othnl')}", -IN['other_net_liab'], True),
    ('onci', 'Less other non-controlling interests at book',
     f"=-({a('ncib')}-{a('nciherfy')})", -DCF['nci_other_book'], False),
]
r = 5
BR = {}
for key, lab, v, xp, grn in brows:
    put(ws, f'A{r}', lab, fmt=None, bold=key in ('ev',))
    putf(ws, f'C{r}', v, xp, NUM0, bold=key in ('ev',), green=grn)
    BR[key] = r
    r += 1
_dec_sum = (DCF['ev'] + IN['inv_nc_fy25'] + IN['inv_c_fy25'] + IN['tiryaki_recv']
            + IN['invprop_fy25'] + IN['cash_fy25'] - IN['loans_fy25'] - IN['leases_fy25']
            - IN['eb_fy25'] - IN['restor_fy25'] - IN['other_net_liab']
            - DCF['nci_other_book'])
band(ws, r, 4)
put(ws, f'A{r}', 'Dec-2025-dated legs, subtotal', bold=True, fmt=None)
putf(ws, f'C{r}', f"=SUM(C{BR['ev']}:C{BR['onci']})", _dec_sum, NUM0, bold=True)
BR['dec'] = r; r += 1
put(ws, f'A{r}', 'x accretion to 18-Aug-2026 at the cost of equity (Dec legs only)',
    fmt=None)
putf(ws, f'C{r}', f"=(1+DCF!C{CC['ke']})^({a('anchor_days')}/365)", DCF['roll'], DF4,
     green=True)
BR['roll'] = r; r += 1
put(ws, f'A{r}', 'Dec legs rolled to the anchor', fmt=None)
putf(ws, f'C{r}', f"=C{BR['dec']}*C{BR['roll']}", _dec_sum * DCF['roll'], NUM0)
BR['decroll'] = r; r += 1
put(ws, f'A{r}', 'ANCHOR-DATED LEGS (held outside the roll)', bold=True, fmt=None); r += 1
put(ws, f'A{r}', 'Kinan at capitalized earnings (annualized H1-2026 share / cost of '
    'equity — an H1-2026 run-rate value)', fmt=None)
putf(ws, f'C{r}', f"={a('kinh1')}*2/DCF!C{CC['ke']}", kin_cap, NUM0)
BR['kinan'] = r; r += 1
put(ws, f'A{r}', "Less Herfy's 51% NCI at Herfy's own settled 18-Aug-2026 price",
    fmt=None)
putf(ws, f'C{r}', f"=-0.51*{a('herfyp')}*{a('herfysh')}", -DCF['nci_herfy_mkt'], NUM0)
BR['herfy'] = r; r += 1
put(ws, f'A{r}', 'Less Al Mehbaj consideration (Jul-2026 acquisition whose revenue the '
    'forecast carries)', fmt=None)
putf(ws, f'C{r}', f"=-{a('mehbaj')}", -IN['mehbaj_total'], NUM1)
BR['mehbaj'] = r; r += 1
band(ws, r, 4)
put(ws, f'A{r}', 'Equity value attributable to owners AT THE ANCHOR', bold=True, fmt=None)
putf(ws, f'C{r}', f"=C{BR['decroll']}+C{BR['kinan']}+C{BR['herfy']}+C{BR['mehbaj']}",
     DCF['eq_val'], NUM0, bold=True)
putf(ws, f'D{r}', f"=C{r}/{a('sharesw')}-{a('div_between')}", DCF['ps'], PX, bold=True)
BR['eq'] = r
r += 1
put(ws, f'A{r}', '31-Dec-2025-basis view (Dec legs unrolled; anchor legs at their own '
    'dates)', fmt=None)
putf(ws, f'C{r}', f"=C{BR['dec']}+C{BR['kinan']}+C{BR['herfy']}+C{BR['mehbaj']}",
     DCF['eq_dec'], NUM0)
putf(ws, f'D{r}', f"=C{r}/{a('sharesw')}", DCF['ps_dec'], PX)
BR['eqdec'] = r
# point the DCF anchor block at the bridge's real rows
wb['DCF'][f"C{AN['eq']}"] = f"='SOTP Bridge'!C{BR['eq']}"
wb['DCF'][f"C{AN['psdec']}"] = f"='SOTP Bridge'!D{BR['eqdec']}"
EXPECT.setdefault('DCF', {})[f"C{AN['psdec']}"] = DCF['ps_dec']
r += 2
put(ws, f'A{r}', 'Terminal value as a share of enterprise value', fmt=None)
putf(ws, f'C{r}', f"=DCF!C{TB['tvshare']}", DCF['tv_share'], PCT, green=True)
r += 1
put(ws, f'A{r}', 'Kinan alternatives: carrying value 435.5 (floor) · share of net assets '
    '600.9 (disclosed) · capitalized earnings (used). First edition note superseded: the '
    'Tiryaki receivable IS on the 31-Dec-2025 balance sheet (notes 14/22) and now sits in '
    'the bridge; only its share settlement happened in H1-2026.', fmt=None).font = SUB

# ============ 7 RELATIVE & NORMALIZED =========================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative, normalised and book lenses',
      'The peer-mix multiple is computed from quoted peers, discounted for the '
      'conglomerate/Egypt mix, and applied to model earnings', 6, awidth=52, cwidth=14)
REL = LN['relative']
r = 4
put(ws, f'A{r}', 'PEER-MIX MULTIPLE', bold=True, fmt=None); r += 1
RL = {}
for key, label, fml, xp in [
        ('fp', 'Processing-leg P/E = median (NADEC, Wilmar)',
         f"=MEDIAN({a('pe_nad')},{a('pe_wil')})", REL['pe_fp_leg']),
        ('ret', 'Retail-leg P/E = BinDawood (Al Othaim n/m after its 11-Aug-2026 H1 loss)',
         f"={a('pe_bin')}", REL['pe_ret_leg']),
        ('mixw', 'Processing-leg weight = FP EBITDA / (FP + Panda EBITDA), FY2026E — '
         'COMPUTED from the segment sheet', None, REL['pe_mix_w_fp']),
        ('mix', 'Mix-weighted peer P/E', None, REL['pe_mix']),
        ('appl', 'Applied multiple after the disclosed discount', None, REL['pe'])]:
    put(ws, f'A{r}', label, fmt=None)
    if fml:
        putf(ws, f'C{r}', fml, xp, MULT)
    RL[key] = r
    r += 1
putf(ws, f"C{RL['mixw']}",
     f"=Segments!B{RW['fpeb']}/(Segments!B{RW['fpeb']}+Segments!B{RW['paneb']})",
     REL['pe_mix_w_fp'], PCT2)
putf(ws, f"C{RL['mix']}", f"=C{RL['mixw']}*C{RL['fp']}+(1-C{RL['mixw']})*C{RL['ret']}",
     REL['pe_mix'], MULT)
putf(ws, f"C{RL['appl']}", f"=C{RL['mix']}*(1-{a('pedisc')})", REL['pe'], MULT)
r += 1
put(ws, f'A{r}', 'RELATIVE LENS — trailing multiple on TRAILING earnings', bold=True,
    fmt=None); r += 1
put(ws, f'A{r}', 'Trailing recurring net income to 30-Jun-2026 = FY2025 recurring − '
    'H1-2025 recurring + H1-2026 recurring (all three company-disclosed)', fmt=None)
putf(ws, f'C{r}', f"={a('rec25')}-{a('rec25h1')}+{a('rec26h1')}", REL['ttm_recurring'],
     NUM1)
TTM_ROW = r; r += 1
put(ws, f'A{r}', 'Trailing recurring EPS (on the ex-treasury divisor)', fmt=None)
putf(ws, f'C{r}', f"=C{TTM_ROW}/{a('sharesw')}", REL['ttm_eps'], PX)
EPS_ROW = r; r += 1
put(ws, f'A{r}', 'Relative lens value = applied multiple x trailing recurring EPS '
    '(anchor-dated by construction: the multiples are 18-Aug quotes)', bold=True,
    fmt=None)
putf(ws, f'C{r}', f"=C{RL['appl']}*C{EPS_ROW}", REL['base'], PX, bold=True)
REL_ROW = r; r += 1
put(ws, f'A{r}', '  bear (30% discount) / bull (10% discount)', fmt=None)
putf(ws, f'C{r}', f"=C{RL['mix']}*0.7*C{EPS_ROW}", REL['bear'], PX)
putf(ws, f'D{r}', f"=C{RL['mix']}*0.9*C{EPS_ROW}", REL['bull'], PX)
REL_BB = r; r += 1
put(ws, f'A{r}', '  forward variant (the first edition\'s construction: applied multiple '
    'x FY2026E EPS — a trailing multiple on forward earnings imports the peers\' growth '
    'twice; shown, not used)', fmt=None)
putf(ws, f'C{r}', f"=C{RL['appl']}*'Income Statement'!{FCOL[0]}{IS['eps']}",
     REL['forward_variant'], PX)
r += 2
put(ws, f'A{r}', 'NORMALISED EARNINGS POWER', bold=True, fmt=None); r += 1
NRM = LN['normalized']
put(ws, f'A{r}', 'FY2026E revenue (the trailing-multiple year; FY2027E under a trailing '
    'multiple would double-count growth)', fmt=None)
putf(ws, f'C{r}', f"=Segments!B{RW['grev']}", F['rev'][0], NUM0, green=True)
NR1 = r; r += 1
put(ws, f'A{r}', 'Normalised EBITDA at the mid-cycle margin', fmt=None)
putf(ws, f'C{r}', f"=C{NR1}*{a('normm')}", F['rev'][0] * IN['norm_ebitda_mgn'], NUM0)
NR2 = r; r += 1
put(ws, f'A{r}', 'Normalised net profit (same D&A, finance, tax and minority frame)',
    fmt=None)
_np_norm = ((F['rev'][0] * IN['norm_ebitda_mgn'] - F['dna'][0] + F['netfin'][0])
            * (1 - T) * (1 - IN['nci_share']) + F['kinan'][0])
putf(ws, f'C{r}',
     f"=(C{NR2}-'Income Statement'!E{IS['dna']}+'Income Statement'!E{IS['nf']})"
     f"*(1-{a('tax')})*(1-{a('ncis')})+'Income Statement'!E{IS['kinan']}",
     _np_norm, NUM0)
NR3 = r; r += 1
put(ws, f'A{r}', 'Normalised EPS', fmt=None)
putf(ws, f'C{r}', f"=C{NR3}/{a('sharesw')}", NRM['eps_norm'], PX)
NR4 = r; r += 1
put(ws, f'A{r}', 'Normalised lens value = applied multiple x normalised EPS', bold=True,
    fmt=None)
putf(ws, f'C{r}', f"=C{RL['appl']}*C{NR4}", NRM['base'], PX, bold=True)
put(ws, f'E{r}', f"=C{RL['mix']}*0.7*C{NR4}", GREEN, PX)
EXPECT.setdefault(ws.title, {})[f'E{r}'] = NRM['bear']
put(ws, f'F{r}', f"=C{RL['mix']}*0.9*C{NR4}", GREEN, PX)
EXPECT.setdefault(ws.title, {})[f'F{r}'] = NRM['bull']
NRM_ROW = r
r += 2
put(ws, f'A{r}', 'BOOK VALUE AND SUSTAINABLE RETURN', bold=True, fmt=None); r += 1
BKL = LN['book']
put(ws, f'A{r}', 'Book value per share (30-Jun-2026 reviewed equity, ex the 1.70 '
    'dividend paid)', fmt=None)
putf(ws, f'C{r}', f"={a('eqjun')}/{a('sharesw')}", BKL['bvps'], PX)
BK1 = r; r += 1
put(ws, f'A{r}', 'Sustainable return on equity = the model\'s OWN FY2026E attributable '
    'profit (recurring construction) / FY2025 opening equity — one FY2026 base across '
    'lenses', fmt=None)
putf(ws, f'C{r}', f"='Income Statement'!E{IS['np']}/{a('eq25')}", BKL['roe'], PCT2)
BK2 = r; r += 1
put(ws, f'A{r}', 'Justified price-to-book = (ROE − g) / (Ke − g)', fmt=None)
putf(ws, f'C{r}', f"=(C{BK2}-{a('g')})/(DCF!C{CC['ke']}-{a('g')})", BKL['pb'], NUM2)
BK3 = r; r += 1
put(ws, f'A{r}', 'Book lens value = justified P/B x book value per share', bold=True,
    fmt=None)
putf(ws, f'C{r}', f"=C{BK3}*C{BK1}", BKL['base'], PX, bold=True)
put(ws, f'E{r}', f"=MAX(C{BK3}-0.15,0.5)*C{BK1}", GREEN, PX)
EXPECT.setdefault(ws.title, {})[f'E{r}'] = BKL['bear']
put(ws, f'F{r}', f"=(C{BK3}+0.15)*C{BK1}", GREEN, PX)
EXPECT.setdefault(ws.title, {})[f'F{r}'] = BKL['bull']
BOOK_ROW = r
r += 2
put(ws, f'A{r}', 'DIVIDEND-DISCOUNT CROSSWALK — the multiple our own cost of equity '
    'supports, computed three ways (none is typed; the study quotes the range)',
    bold=True, fmt=None); r += 1
DDM = REL['ddm']
put(ws, f'A{r}', f"Two-stage (model earnings growth {DDM['g_stage1']:.1%} for five years, "
    "then terminal growth)", fmt=None)
put(ws, f'C{r}', DDM['two_stage'], BLUE, MULT); r += 1
put(ws, f'A{r}', 'Gordon forward form = payout / (Ke − g)', fmt=None)
putf(ws, f'C{r}', f"={a('payout')}/(DCF!C{CC['ke']}-{a('g')})", DDM['gordon_fwd'], MULT)
r += 1
put(ws, f'A{r}', "Implied by Expert 2's dividend model (its value / FY2026E EPS)",
    fmt=None)
put(ws, f'C{r}', DDM['e2_implied'], BLUE, MULT); r += 1
put(ws, f'A{r}', f"The market pays the peer set {REL['pe_mix']:.1f}x against the "
    f"{DDM['gordon_fwd']:.1f}x-{DDM['two_stage']:.1f}x this study's cost of equity "
    'supports. That gap (the price of capital) is why the DCF sits below the '
    'peer-anchored lenses; it is examined in the study rather than averaged away.',
    fmt=None, wrap=True).font = SUB
ws.row_dimensions[r].height = 40
ANCH['rel_row'] = f'C{REL_ROW}'; ANCH['norm_row'] = f'C{NRM_ROW}'
ANCH['book_row'] = f'C{BOOK_ROW}'

# ============ 2 SUMMARY =======================================================
ws = sheet('Summary')
title(ws, 'Summary — valuation at a glance', 'All values link live to their source sheets',
      7, awidth=46, cwidth=14)
hdr(ws, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Weight', 'Contribution', 'vs spot'])
r = 5
lens_meta = [
    ('Discounted cash flow (Framing A)', 'dcf', f"=DCF!C{AN['ps']}", None, None, 'w_dcf'),
    ('Relative multiples', 'relative', f"='Relative & Normalized'!C{REL_ROW}",
     f"='Relative & Normalized'!C{REL_BB}", f"='Relative & Normalized'!D{REL_BB}", 'w_rel'),
    ('Normalised earnings power', 'normalized', f"='Relative & Normalized'!C{NRM_ROW}",
     f"='Relative & Normalized'!E{NRM_ROW}", f"='Relative & Normalized'!F{NRM_ROW}",
     'w_norm'),
    ('Book value and sustainable return', 'book', f"='Relative & Normalized'!C{BOOK_ROW}",
     f"='Relative & Normalized'!E{BOOK_ROW}", f"='Relative & Normalized'!F{BOOK_ROW}",
     'w_book'),
]
SPOT_ROW = 14
for label, k, base_f, bear_f, bull_f, wkey in lens_meta:
    l = LN[k]
    put(ws, f'A{r}', label, fmt=None)
    if bear_f:
        putf(ws, f'B{r}', bear_f, l['bear'], PX, green=True)
    else:
        put(ws, f'B{r}', l['bear'], BLUE, PX)
    putf(ws, f'C{r}', base_f, l['base'], PX, green=True)
    if bull_f:
        putf(ws, f'D{r}', bull_f, l['bull'], PX, green=True)
    else:
        put(ws, f'D{r}', l['bull'], BLUE, PX)
    putf(ws, f'E{r}', f"={a(wkey)}", D['weights'][k], PCT, green=True)
    putf(ws, f'F{r}', f'=C{r}*E{r}', l['base'] * D['weights'][k], PX)
    putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', l['base'] / SPOT - 1, PCT)
    r += 1
band(ws, r, 7)
LK = ['dcf', 'relative', 'normalized', 'book']
_WB = sum(LN[k]['bear'] * D['weights'][k] for k in LK)
_WU = sum(LN[k]['bull'] * D['weights'][k] for k in LK)
put(ws, f'A{r}', 'Weighted central', bold=True, fmt=None)
putf(ws, f'B{r}', '=B5*E5+B6*E6+B7*E7+B8*E8', _WB, PX, bold=True)
putf(ws, f'C{r}', '=SUM(F5:F8)', D['central'], PX, bold=True)
putf(ws, f'D{r}', '=D5*E5+D6*E6+D7*E7+D8*E8', _WU, PX, bold=True)
putf(ws, f'E{r}', '=SUM(E5:E8)', 1.0, PCT, bold=True)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['central'] / SPOT - 1, PCT, bold=True)
CENTRAL_ROW = r
assert CENTRAL_ROW == 9
r += 1
put(ws, f'A{r}', 'Span across lenses (min/max — NOT weighted)', fmt=None)
putf(ws, f'B{r}', '=MIN(B5:B8)', min(LN[k]['bear'] for k in LK), PX)
putf(ws, f'D{r}', '=MAX(D5:D8)', max(LN[k]['bull'] for k in LK), PX)
r += 1
put(ws, f'A{r}', 'Contested judgement, other way — Framing B (−6% then −3% forever)',
    fmt=None)
put(ws, f'C{r}', DCF['framingB'], BLUE, PX)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', DCF['framingB'] / SPOT - 1, PCT)
FRB_ROW = r
assert FRB_ROW == 11
r += 1
put(ws, f'A{r}', 'Terminal value share of DCF enterprise value', fmt=None)
putf(ws, f'C{r}', f"=DCF!C{TB['tvshare']}", DCF['tv_share'], PCT, green=True)
assert r == 12
r += 1
put(ws, f'A{r}', 'Expert panel median', fmt=None)
putf(ws, f'C{r}', f"='Fundamental Valuation'!C24", D['panel_median'], PX, green=True)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['panel_median'] / SPOT - 1, PCT)
r += 1
band(ws, r, 7)
put(ws, f'A{r}', 'Market price (settled 18-Aug-2026 close)', bold=True, fmt=None)
putf(ws, f'C{r}', f"={a('spot')}", SPOT, PX, bold=True)
assert r == SPOT_ROW
r += 2
hdr(ws, r, ['Key figure', '', 'Value']); r += 1
KEYS = [
    ('Shares issued (mn)', f"={a('shares')}", SH, NUM0),
    ('Market capitalisation (SAR mn)', f"=$C${SPOT_ROW}*{a('shares')}", M['mktcap'], NUM0),
    ('FY2025 revenue (SAR mn, audited)', f"='Income Statement'!D{IS['rev']}",
     HI['FY25']['rev'], NUM0),
    ('FY2025 operating EBITDA (SAR mn)', f"='Income Statement'!D{IS['ebitda']}",
     HI['FY25']['ebitda'], NUM0),
    ('FY2026E revenue (SAR mn)', f"='Income Statement'!E{IS['rev']}", F['rev'][0], NUM0),
    ('FY2026E EPS (SAR)', f"='Income Statement'!E{IS['eps']}", F['eps'][0], PX),
    ('Cost of capital — explicit window', f"=DCF!C{CC['wacc']}", W['wacc_exp'], PCT2),
    ('Cost of capital — terminal', f"=DCF!C{CC['wacct']}", W['wacc_term'], PCT2),
    ('Terminal growth', f"={a('g')}", TRI['nominal_growth'], PCT),
    ('Terminal asset life, derived from note 6 (years)', f"={a('life')}",
     IN['asset_life_years'], NUM2),
    ('Terminal free cash flow as a share of terminal profit',
     f"=DCF!C{TB['fcffT']}/DCF!C{TB['nopatT']}", TRO['fcff'] / TRI['nopat'], PCT),
    ('Net debt, 30-Jun-2026 (company definition, SAR mn)', None, H1['netdebt'], NUM0),
    ('Weighted central — first edition, 18-Aug-2026 (superseded by this edition)',
     None, D['edition1']['central'], PX),
]
MKTCAP_ROW = r + 1
for lab, fml, xp, fmt in KEYS:
    put(ws, f'A{r}', lab, fmt=None)
    if fml:
        putf(ws, f'C{r}', fml, xp, fmt, green=True)
    else:
        put(ws, f'C{r}', xp, BLUE, fmt)
    r += 1
ANCH['summary_mktcap'] = f'C{MKTCAP_ROW}'

# ============ 3 FUNDAMENTAL VALUATION =========================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — four lenses, the contested judgement, the panel', None,
      6, awidth=56, cwidth=14)
hdr(ws, 4, ['Lens / step', 'Basis', 'SAR per share'])
rows = [
    ('Discounted cash flow (Framing A)', 'five-year FCFF (full lease additions charged) '
     '+ computed-return terminal', f"=DCF!C{AN['ps']}", DCF['ps']),
    ('  bear', 'Framing B density (−6% then −3%) · oil GP/t −40 · sugar −15 · half '
     'volume growth · flat store opex', LN['dcf']['bear'], None),
    ('  bull', 'oil GP/t +25 · sugar +10 · volumes x1.2 · Panda gross margin +40bp',
     LN['dcf']['bull'], None),
    ('Relative multiples', 'trailing peer-mix P/E x trailing recurring EPS, '
     'conglomerate discount applied', f"='Relative & Normalized'!C{REL_ROW}",
     LN['relative']['base']),
    ('Normalised earnings power', 'mid-cycle margin on FY2026E revenue',
     f"='Relative & Normalized'!C{NRM_ROW}", LN['normalized']['base']),
    ('Book value and sustainable return', 'justified P/B on 30-Jun-2026 book; ROE = '
     'model FY2026E', f"='Relative & Normalized'!C{BOOK_ROW}", LN['book']['base']),
]
r = 5
for a_, b_, c_, xp in rows:
    put(ws, f'A{r}', a_, fmt=None); put(ws, f'B{r}', b_, fmt=None)
    if isinstance(c_, str):
        putf(ws, f'C{r}', c_, xp, PX, green=True)
    else:
        put(ws, f'C{r}', c_, BLUE, PX)
    r += 1
r += 1
band(ws, r, 3); put(ws, f'A{r}', 'Weighted central', bold=True, fmt=None)
putf(ws, f'C{r}', f'=Summary!C{CENTRAL_ROW}', D['central'], PX, bold=True, green=True)
r += 2
put(ws, f'A{r}', 'THE CONTESTED JUDGEMENT — PANDA\'S EXPANSION, BOTH WAYS', bold=True,
    fmt=None); r += 1
F1 = r
put(ws, f'A{r}', 'Framing A: density stabilises as CXR and e-commerce mature', fmt=None)
putf(ws, f'C{r}', f"=DCF!C{AN['ps']}", DCF['framingA'], PX, green=True); r += 1
put(ws, f'A{r}', 'Framing B: the density erosion never fades — −6% in FY2026E, then −3% '
    'every year forever (engine re-run)', fmt=None)
put(ws, f'C{r}', DCF['framingB'], BLUE, PX); r += 1
put(ws, f'A{r}', 'The judgement is worth (Framing A less Framing B, per share)', fmt=None)
putf(ws, f'C{r}', f'=C{F1}-C{F1+1}', DCF['framing_gap'], PX); r += 1
put(ws, f'A{r}', 'CDS-basis fair value (same model, CDS sovereign basis, Jan-2026 '
    'vintage flagged; engine re-run)', fmt=None)
put(ws, f'C{r}', DCF['ps_cds'], BLUE, PX); r += 1
ANCH['fv_framing1'] = f'C{F1}'; ANCH['fv_framing2'] = f'C{F1+1}'
ANCH['fv_gap'] = f'C{F1+2}'
r += 1
put(ws, f'A{r}', 'JUDGEMENT VARIANTS — every lever value published (engine re-runs)',
    bold=True, fmt=None); r += 1
for lab, val in [
        ('Store path at the H1-2026 run-rate (+8/yr instead of guidance; lease growth '
         'scaled to the 0.4x cadence)', DCF['stores_runrate']),
        ('Density opening at −7.1% (the June-2025 = 213 assumption end of the range)',
         DCF['sps_open_71']),
        ('Density opening at −6.0% (the interpolated end; the base input)',
         DCF['sps_open_59']),
        ('Asset life read at twice the directly measured average age of the base — the '
         'heavier of the two readings of note 6', DCF['ps_life_variant'])]:
    put(ws, f'A{r}', lab, fmt=None, wrap=True)
    put(ws, f'C{r}', val, BLUE, PX)
    ws.row_dimensions[r].height = 28
    r += 1
r += 1
put(ws, f'A{r}', 'EXPERT PANEL', bold=True, fmt=None); r += 1
hdr(ws, r, ['Expert', 'Method', 'Base (SAR/share)', 'Low', 'High']); r += 1
PF = r
for k, nm in [('e1', 'Expert 1'), ('e2', 'Expert 2'), ('e3', 'Expert 3')]:
    e = EXP[k]
    put(ws, f'A{r}', nm, fmt=None); put(ws, f'B{r}', e['method_short'], fmt=None)
    put(ws, f'C{r}', e['base'], BLUE, PX); put(ws, f'D{r}', e['rng'][0], BLUE, PX)
    put(ws, f'E{r}', e['rng'][1], BLUE, PX); r += 1
band(ws, r, 5); put(ws, f'A{r}', 'Panel median', bold=True, fmt=None)
putf(ws, f'C{r}', f'=MEDIAN(C{PF}:C{PF+2})', D['panel_median'], PX, bold=True)
ANCH['fv_panel'] = f'C{r}'
# point Summary's expert-panel row at the real panel-median row
wb['Summary'][f'C13'] = f"='Fundamental Valuation'!C{r}"
r += 2
put(ws, f'A{r}', 'The risk-free rate, priced both ways', fmt=None); r += 1
put(ws, f'A{r}', 'The base is the PUBLISHED SAR sovereign curve: FTSE SAGBI 7-10y '
    'yield 5.52% (31-Jul-2026 factsheet; iBoxx SAR sukuk 5.44% at 6.07y corroborates). '
    'Discounting at −50bp / +50bp instead is an engine re-run worth the figures at right',
    fmt=None, wrap=True)
put(ws, f'C{r}', SN['rf_alts']['5.02%'], BLUE, PX)
put(ws, f'D{r}', SN['rf_alts']['6.02%'], BLUE, PX)
ws.row_dimensions[r].height = 40

# ============ 12 SUMMARY FINANCIALS ===========================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials and returns', None, 9, awidth=44, cwidth=13)
hdr(ws, 4, ['SAR mn', 'FY2025', *YF])
r = 5
SF = {}
sf_rows = [
    ('rev', 'Revenue', HI['FY25']['rev'], [f"='Income Statement'!{FCOL[i]}{IS['rev']}"
     for i in range(5)], F['rev'], NUM0),
    ('ebitda', 'Operating EBITDA', HI['FY25']['ebitda'],
     [f"='Income Statement'!{FCOL[i]}{IS['ebitda']}" for i in range(5)], F['ebitda'], NUM0),
    ('mgn', 'EBITDA margin', HI['FY25']['ebitda'] / HI['FY25']['rev'],
     ["=1"] * 5, None, PCT),   # placeholder — rebuilt explicitly right below
    ('np', 'Profit attributable to owners', HI['FY25']['np_att'],
     [f"='Income Statement'!{FCOL[i]}{IS['np']}" for i in range(5)], F['np'], NUM0),
    ('eps', 'EPS (SAR)', HI['FY25']['eps'],
     [f"='Income Statement'!{FCOL[i]}{IS['eps']}" for i in range(5)], F['eps'], PX),
    ('capex', 'Capital expenditure', 858.5, [f"={a('capex',i)}" for i in range(5)],
     IN['capex_path'], NUM0),
    ('fcff', 'Free cash flow to the firm', None, [f"=DCF!{CD[i]}{DC['fcff']}"
     for i in range(5)], F['fcff'], NUM0),
    ('nd', 'Net debt (company definition)',
     IN['loans_fy25'] - IN['cash_fy25'] - IN['inv_c_fy25'],
     [f"='Balance Sheet'!{FCOL[i]}{BS['nd']}" for i in range(5)], F['netdebt'], NUM0),
    ('eq', 'Equity attributable to owners', IN['equity_att_fy25'],
     [f"='Balance Sheet'!{FCOL[i]}{BS['eq']}" for i in range(5)], F['equity_att'], NUM0),
]
for key, lab, hist, fmls, exps, fmt in sf_rows:
    put(ws, f'A{r}', lab, fmt=None)
    if hist is not None:
        put(ws, f'B{r}', hist, BLUE, fmt)
    for i in range(5):
        putf(ws, f'{CD[i+1]}{r}' if False else f'{get_column_letter(3+i)}{r}', fmls[i],
             exps[i] if exps else None, fmt)
    SF[key] = r
    r += 1
# fix margin row formulas (written against wrong cols above): rebuild explicitly
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f'{col}{SF["mgn"]}'] = f"={col}{SF['ebitda']}/{col}{SF['rev']}"
    EXPECT['Summary Financials'][f'{col}{SF["mgn"]}'] = F['ebitda_margin'][i]
put(ws, f'B{SF["mgn"]}', HI['FY25']['ebitda'] / HI['FY25']['rev'], BLUE, PCT)
r += 1
put(ws, f'A{r}', 'Operating invested capital (equity + minorities + debt + leases − '
    'cash − investments − Kinan − Tiryaki − investment property; the terminal return '
    'runs on this row)', fmt=None)
IC_ROW = r
for i in range(5):
    col = get_column_letter(3 + i)
    putf(ws, f'{col}{r}',
         f"='Balance Sheet'!{FCOL[i]}{BS['eq']}+'Balance Sheet'!{FCOL[i]}{BS['nci']}"
         f"+{a('loans')}+'Balance Sheet'!{FCOL[i]}{BS['leases']}"
         f"-'Balance Sheet'!{FCOL[i]}{BS['cash']}-{a('invc')}-{a('invnc')}"
         f"-'Balance Sheet'!{FCOL[i]}{BS['kinan']}-{a('tiryaki')}-{a('invprop')}",
         F['ic_path'][i], NUM0)
r += 1
put(ws, f'A{r}', 'Return on invested capital (NOPAT / opening invested capital)', fmt=None)
for i in range(5):
    col = get_column_letter(3 + i)
    if i == 0:
        f_ = (f"=DCF!B{DC['nopat']}/({a('eq25')}+{a('ncib')}+{a('loans')}+{a('lease')}"
              f"-{a('cash')}-{a('invc')}-{a('invnc')}-{a('kinbv')}-{a('tiryaki')}"
              f"-{a('invprop')})")
    else:
        f_ = f"=DCF!{CD[i]}{DC['nopat']}/{get_column_letter(2+i)}{IC_ROW}"
    putf(ws, f'{col}{r}', f_, F['roic_path'][i], PCT2)
r += 1
# The DCF sheet no longer carries a terminal-return cell to re-point: the sanctioned
# terminal does not use one. The return path stays on THIS sheet, where it is a fact about
# the forecast rather than a driver of the valuation, and the row above computes it from
# the invested-capital row rather than from anything typed.
put(ws, f'A{r}', 'Return on equity (attributable profit / opening equity)', fmt=None)
for i in range(5):
    col = get_column_letter(3 + i)
    prev = IN['equity_att_fy25'] if i == 0 else F['equity_att'][i-1]
    if i == 0:
        f_ = f"='Income Statement'!E{IS['np']}/{a('eq25')}"
    else:
        f_ = (f"='Income Statement'!{FCOL[i]}{IS['np']}"
              f"/'Balance Sheet'!{FCOL[i-1]}{BS['eq']}")
    putf(ws, f'{col}{r}', f_, F['np'][i] / prev, PCT2)

# ============ 13 MONTE CARLO ==================================================
ws = sheet('Monte Carlo')
title(ws, 'Monte Carlo price map — engine output (pasted by construction)',
      'Each figure is a 50,000-path simulation of the whole price process; the grid does '
      'not redraw when a driver changes', 7, awidth=52, cwidth=13)
r = 4
put(ws, f'A{r}', f"Anchor: SAR {STK['spot']:.2f} (settled close) on {STK['anchor_date']} "
    f"· annualized volatility ~{STK['horizons']['1M']['anchor_vol_ann']:.0%} · 50,000 "
    "simulated paths on a fixed seed (42), carry-anchored heavy-tailed process",
    fmt=None); r += 2
hdr(ws, r, ['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95', 'P(above spot)']); r += 1
for short, label in [('1M', 'One month (to {})'.format(STK['horizons']['1M']['grade_date'])),
                     ('3M', 'Three months (to {})'.format(STK['horizons']['3M']['grade_date']))]:
    hz = STK['horizons'][short]
    put(ws, f'A{r}', label, fmt=None)
    for j, p in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        put(ws, f'{get_column_letter(2+j)}{r}', hz['pct'][p], BLUE, PX)
    put(ws, f'G{r}', hz['p_above'], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Touch probability within 3M', '+5%', '+10%', '+15%', '−5%', '−10%', '−15%'])
r += 1
hz3 = STK['horizons']['3M']
put(ws, f'A{r}', 'Probability the level trades at least once', fmt=None)
for j, k in enumerate(['touch_up5', 'touch_up10', 'touch_up15', 'touch_dn5', 'touch_dn10',
                       'touch_dn15']):
    put(ws, f'{get_column_letter(2+j)}{r}', hz3[k], BLUE, PCT)
r += 2
b5 = BT['five_year']
put(ws, f'A{r}', 'Backtest of the same price engine on Savola (plain language):', bold=True,
    fmt=None); r += 1
for ln in [
    f"Over the last five years ({b5['first_origin']} to {b5['last_origin']}), "
    f"{b5['windows']} non-overlapping three-month forecasts scored "
    f"{b5['skill_norm']*100:+.1f}% against a carry-anchored random walk.",
    f"Realized outcomes fell inside the 50/80/90% bands {b5['cov50']:.0%}/{b5['cov80']:.0%}/"
    f"{b5['cov90']:.0%} of the time; the forecast percentiles were roughly uniform "
    f"(chi-square p = {b5['chi2_p']:.2f}, KS p = {b5['ks_p']:.2f}).",
    f"Across the full cleaned history ({BT['full']['windows']} windows) the score was "
    f"{BT['full']['skill_norm']*100:+.1f}%; the engine is statistically indistinguishable "
    "from the benchmark on this name — the map is honest about dispersion, not a source "
    "of edge."]:
    put(ws, f'A{r}', ln, fmt=None, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

# ============ 14 SENSITIVITY ==================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — engine re-runs (pasted by construction)',
      'Each cell is a complete revaluation of the whole model; the grids do not redraw '
      'when a driver changes', 7, awidth=30, cwidth=13)
r = 4
put(ws, f'A{r}', 'Fair value per share at the anchor: cost of capital x terminal growth',
    bold=True, fmt=None); r += 1
hdr(ws, r, ['WACC \\ g'] + [f'{g:.1%}' for g in SN['g_grid']]); r += 1
for i, wv in enumerate(SN['wacc_grid']):
    put(ws, f'A{r}', f'{wv:.2%}', fmt=None)
    for j in range(len(SN['g_grid'])):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid'][i][j], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Beta (the 90% confidence interval of the measured beta spans 0.73-1.44)',
    bold=True, fmt=None); r += 1
hdr(ws, r, ['Beta', 'Fair value (SAR)']); r += 1
for b_, v in SN['beta_grid'].items():
    put(ws, f'A{r}', float(b_), BLUE, '0.000')
    put(ws, f'B{r}', v, BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Risk-free construction', bold=True, fmt=None); r += 1
hdr(ws, r, ['10Y SAR level', 'Fair value (SAR)']); r += 1
for k, v in SN['rf_alts'].items():
    put(ws, f'A{r}', k, fmt=None)
    put(ws, f'B{r}', v, BLUE, PX)
    r += 1

# ============ 15 PER-SHARE & RATIOS ===========================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share figures, ratios and operating KPIs', None, 8, awidth=44, cwidth=13)
hdr(ws, 4, ['Per share (SAR)', *YF])
r = 5
pr_rows = [
    ('EPS', [f"='Income Statement'!{FCOL[i]}{IS['eps']}" for i in range(5)], F['eps'], PX),
    ('Dividend per share (policy payout)',
     [f"='Income Statement'!{FCOL[i]}{IS['np']}*{a('payout')}/{a('sharesw')}"
      for i in range(5)], [F['div'][i] / SHW for i in range(5)], PX),
    ('Book value per share',
     [f"='Balance Sheet'!{FCOL[i]}{BS['eq']}/{a('sharesw')}" for i in range(5)],
     [F['equity_att'][i] / SHW for i in range(5)], PX),
    ('Free cash flow to the firm per share',
     [f"=DCF!{CD[i]}{DC['fcff']}/{a('sharesw')}" for i in range(5)],
     [F['fcff'][i] / SHW for i in range(5)], PX),
]
for lab, fmls, exps, fmt in pr_rows:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i], fmt)
    r += 1
r += 1
hdr(ws, r, ['Ratios', *YF]); r += 1
rat_rows = [
    ('Net debt / EBITDA (loans basis)',
     [f"='Balance Sheet'!{FCOL[i]}{BS['nd']}/Segments!{CD[i]}{RW['gebitda']}"
      for i in range(5)], [F['netdebt'][i] / F['ebitda'][i] for i in range(5)], MULT),
    ('Payout ratio (policy)', [f"={a('payout')}" for _ in range(5)],
     [IN['payout']] * 5, PCT),
    ('Dividend yield at spot',
     [f"='Income Statement'!{FCOL[i]}{IS['np']}*{a('payout')}/{a('sharesw')}/{a('spot')}"
      for i in range(5)], [F['div'][i] / SHW / SPOT for i in range(5)], PCT),
]
for lab, fmls, exps, fmt in rat_rows:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i], fmt)
    r += 1
r += 1
hdr(ws, r, ['Operating KPIs (from the unit build)', *YF]); r += 1
kpi_rows = [
    ('Oil volume (k MT)', [f"=Segments!{CD[i]}{RW['oilv']}" for i in range(5)], OIL['vol'],
     NUM0),
    ('Sugar volume (k MT)', [f"=Segments!{CD[i]}{RW['sugv']}" for i in range(5)],
     SUG['vol'], NUM0),
    ('Pasta volume (k MT)', [f"=Segments!{CD[i]}{RW['pasv']}" for i in range(5)],
     PAS['vol'], NUM0),
    ('Panda stores (year end)', [f"={a('stores',i)}" for i in range(5)],
     IN['stores_path'], NUM0),
    ('Panda sales per average store (SAR mn)',
     [f"=Segments!{CD[i]}{RW['sps']}" for i in range(5)],
     [PAN['rev'][i] / ((IN['stores_end25'] if i == 0 else IN['stores_path'][i-1])
                       + IN['stores_path'][i]) * 2 for i in range(5)], NUM2),
    ('Group EBITDA margin (output)', [f"=Segments!{CD[i]}{RW['gmgn']}" for i in range(5)],
     F['ebitda_margin'], PCT),
]
for lab, fmls, exps, fmt in kpi_rows:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls[i], exps[i], fmt, green=True)
    r += 1

# ============ 16 PEER & SECTOR ================================================
ws = sheet('Peer & Sector')
title(ws, 'Peer frame — market data, cross-check only',
      'Settled closes of 18-Aug-2026; the peer-mix multiple used by the lenses is computed '
      'on the Relative & Normalized sheet from these cells', 6, awidth=44, cwidth=14)
hdr(ws, 4, ['Company', 'Exchange', 'P/E (TTM)', 'Note'])
r = 5
for nm, exch, key, note in [
        ('Almarai', 'Tadawul 2280', 'pe_alm', 'GCC food & beverage leader'),
        ('Al Othaim Markets', 'Tadawul 4001', None,
         'n/m — H1-2026 attributable loss announced 11-Aug-2026 (TTM earnings ~79mn); '
         'excluded like Herfy'),
        ('BinDawood Holding', 'Tadawul 4161', 'pe_bin', 'Saudi grocery retail'),
        ('NADEC', 'Tadawul 6010', 'pe_nad', 'Saudi agri-food'),
        ('Wilmar International', 'SGX F34', 'pe_wil', 'international agri-food analogue'),
        ('Herfy Food Services', 'Tadawul 6002', None, 'loss-making; consolidated at 49%')]:
    put(ws, f'A{r}', nm, fmt=None); put(ws, f'B{r}', exch, fmt=None)
    if key:
        putf(ws, f'C{r}', f"={a(key)}", IN['peer_pe'][{'pe_alm': 'ALMARAI',
             'pe_bin': 'BINDAWOOD', 'pe_nad': 'NADEC',
             'pe_wil': 'WILMAR'}[key]], MULT, green=True)
    else:
        put(ws, f'C{r}', 'n/m', fmt=None)
    put(ws, f'D{r}', note, fmt=None, wrap=True)
    r += 1
r += 1
put(ws, f'A{r}', 'Savola at the model fair value (Framing A)', bold=True, fmt=None); r += 1
sav_rows = [
    ('Implied P/E on FY2026E EPS', f"=DCF!C{AN['ps']}/'Income Statement'!E{IS['eps']}",
     DCF['ps'] / F['eps'][0], MULT),
    ('Implied EV / FY2026E EBITDA (lease-inclusive)',
     f"=(DCF!C{TB['ev']})/Segments!B{RW['gebitda']}", DCF['ev'] / F['ebitda'][0], MULT),
    ('P/E at spot on FY2026E EPS', f"={a('spot')}/'Income Statement'!E{IS['eps']}",
     SPOT / F['eps'][0], MULT),
]
for lab, fml, xp, fmt in sav_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, xp, fmt)
    r += 1
put(ws, f'A{r}', 'Herfy market value feeds the bridge directly: the 51% outside '
    'shareholding is deducted at Herfy\'s own market price rather than book.',
    fmt=None).font = SUB

# ============ anchors for recalc / driver tests ================================
ANCH.update(dict(
    dcf_ps=f"DCF!C{AN['ps']}", dcf_ev=f"DCF!C{TB['ev']}",
    dcf_wacc=f"DCF!C{CC['wacc']}", dcf_wacct=f"DCF!C{CC['wacct']}",
    dcf_ke=f"DCF!C{CC['ke']}", dcf_fcff_b=f"DCF!B{DC['fcff']}",
    dcf_ebitda_b=f"DCF!B{DC['ebitda']}", dcf_tvshare=f"DCF!C{TB['tvshare']}",
    bridge_eq=f"SOTP Bridge!C{BR['eq']}",
    seg_grev_b=f"Segments!B{RW['grev']}", seg_geb_f=f"Segments!F{RW['gebitda']}",
    seg_geb_b=f"Segments!B{RW['gebitda']}",
    bs_foot_i=f"Balance Sheet!I{BS['foot']}", bs_cash_i=f"Balance Sheet!I{BS['cash']}",
    bs_nd_b=f"Balance Sheet!E{BS['nd']}",
    cf_cash_b=f"Cash Flow!B{CF['close']}",
    is_ebitda_d=f"Income Statement!D{IS['ebitda']}",
    is_np_i=f"Income Statement!I{IS['np']}", is_eps_e=f"Income Statement!E{IS['eps']}",
    dcf_pvexp=f"DCF!C{TB['pvexp']}", bs_cash_i2=f"Balance Sheet!I{BS['cash']}",
    dcf_maintT=f"DCF!C{TB['maintT']}", dcf_fcffT=f"DCF!C{TB['fcffT']}",
    dcf_tv=f"DCF!C{TB['tv']}", dcf_floorT=f"DCF!C{TB['floorT']}",
    dcf_wacccds=f"DCF!C{CC['wacccds']}", sf_roic_c=f"Summary Financials!C{IC_ROW + 1}",
))

# ============ order sheets ====================================================
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]

# print setup: landscape, fit to one page wide — the portrait default clips labels from
# their values in the rendered PDF (the documented model-PDF failure mode)
for wsx in wb.worksheets:
    wsx.page_setup.orientation = 'landscape'
    wsx.page_setup.fitToWidth = 1
    wsx.page_setup.fitToHeight = 0
    wsx.sheet_properties.pageSetUpPr.fitToPage = True

XLSX = os.path.join(HERE, 'SAVOLA_Valuation_Model_19082026_public.xlsx')
wb.save(XLSX)
n_form = sum(len(v) for v in EXPECT.values())
json.dump(dict(expected=EXPECT, anchors=ANCH),
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
# count formulas vs pasted literals on the delivered file
import openpyxl
wb2 = openpyxl.load_workbook(XLSX)
nf = nl = 0
for wsx in wb2.worksheets:
    for row in wsx.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nf += 1
            elif isinstance(c.value, (int, float)):
                nl += 1
print(f'workbook written: {XLSX}')
print(f'formulas {nf} | numeric literals {nl} | expected-value records {n_form}')
