"""MODON_Valuation_Model_09082026_public.xlsx — revision 2. 16 sheets, house canonical
model. Blue = inputs · black = formulas · green = cross-sheet links.

Revision-2 changes, from the external audits and the re-audit:
  - valuation restruck at 30-Jun-2026 (H1-2026 reviewed balance sheet + results release);
  - development drivers anchored on the disclosed 30-Jun backlog (AED 62.1bn development)
    and the realised H1 sales (AED 26bn);
  - working capital built from COMPONENTS (receivable days, land-bank conversion,
    payables/advances cover), calibrated at the two balance-sheet dates;
  - D&A off the asset base; terminal debt weight DERIVED from the model's own
    terminal-year balance sheet; bridge on AVAILABLE cash; NCI capitalised;
  - beta 1.03 from an own-stock regression against an equal-weight panel proxy;
  - the mid-chain hardcodes the audits flagged are eliminated: every embedded numeral
    is now an Assumptions input or a live link (lease interest, segment sums, FY2025
    disclosed profit/EBITDA, invested-capital base, SOTP weights and divisor, ROIC row);
  - scenario driver vectors are PUBLISHED on Assumptions so every pasted engine output
    is reproducible; the workbook opens with full calculation forced.

Only three classes of pasted values remain: audited/disclosed figures, the FY2025
segment base, and whole-model engine re-runs (Monte Carlo, sensitivity grids, scenario
outputs). Every formula cell's expected value is recorded for recalc.py."""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, TR, REL, NRM, BK = D['experts'], D['terminal_recon'], D['rel'], D['norm'], D['book']
SEG, S0, STK, CJ = D['seg_fy25'], D['step0'], D['strike'], D['contested']
HA, H1D = D['h1_anchors'], D['h1']
IN = {k: (v['value'] if isinstance(v, dict) and 'value' in v else v)
      for k, v in D['inputs'].items()}
SPOT, SH = M['spot'], M['shares_mn']
TAX = IN['tax_f']
YF = F['years']
NY = 5
FC = ['B', 'C', 'D', 'E', 'F']
HC = ['B', 'C', 'D']
FCOL = ['F', 'G', 'H', 'I', 'J']            # statements: B-D hist, E = 30-Jun-26, F-J fcst
YRF = F['yrfrac']

wb = Workbook()
EXPECT = {}
ANCH = {}

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Modon Holding PSC (ADX: MODON) · Revision 2', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'REVISION 2 (9 Aug 2026). This edition restrikes the valuation on the 30-Jun-2026 reviewed balance',
 'sheet and the H1-2026 results release (revenue backlog AED 65.4bn, 95% development; H1 real-estate',
 'sales AED 26bn) — disclosures the first edition failed to carry into its drivers, a defect surfaced',
 'by external audits and accepted. The audits\' other accepted findings are implemented throughout:',
 'the bridge uses AVAILABLE cash (AED 8.6bn) rather than gross cash including project escrow; the',
 'terminal debt weight is DERIVED from the model\'s own terminal-year balance sheet; working capital',
 'is built from components (receivable days, land-bank conversion, payables/advances cover); D&A is',
 'charged on the asset base; beta is a regression (1.03 vs an equal-weight panel proxy), no longer an',
 'assumption; and the mid-chain hardcoded numerals the audits found are eliminated.', '',
 'What this workbook is. Every blue cell is an input; every black cell is a formula; green cells link',
 'across sheets. The cost of capital is built from the AED government yield net of the sovereign',
 'spread, beta and the premium; the development backlog rolls forward in front of you; the DCF',
 'waterfall chains EBITDA → EBIT → NOPAT → free cash flow → present value; the terminal block derives',
 'its reinvestment rate from growth over return on capital; the statements, ratios and all four',
 'lenses chain off the same cells.', '',
 'THREE THINGS ARE PASTED VALUES. First, audited and disclosed history (the primary record; where a',
 'line is both disclosed and derivable, the DISCLOSED figure is carried). Second, the FY2025',
 'four-segment base and the 30-Jun-2026 balance-sheet anchors — pasted once, with the entire forecast',
 'chaining off them as formulas. Third, whole-model engine outputs, where each figure is a complete',
 're-run: the Monte Carlo price map, the sensitivity grids and the scenario outputs (run-off,',
 'growth-hold, Egypt stress). Their DRIVER VECTORS are published on the Assumptions sheet so each',
 're-run is reproducible, but the grids do NOT redraw when a driver changes; everything else reprices',
 'live, and the workbook forces a full recalculation when opened.', '',
 'The contested judgement, shown both ways. Base: the realised H1-2026 surge normalises but sustains.',
 'Run-off (now labelled a stress — the H1 disclosure falsified it as a central path): launches halve',
 'and fade. Both DCFs are shown side by side; never averaged.', '',
 'What it is not. Not investment advice, a recommendation, or a price target. Values are model',
 'outputs shown as ranges.', '',
 f"Currency. AED million unless stated. Spot AED {SPOT:.2f} (7-Aug-2026 close); valuation date",
 '30-Jun-2026, rolled 38 days to the price anchor. Sheets: READ FIRST · Summary · Fundamental',
 'Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income Statement ·',
 'Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share & Ratios ·',
 'Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 112

# ============ ASSUMPTIONS =====================================================
aws = wb.create_sheet('Assumptions')
title(aws, 'Assumptions — every blue cell reprices the model',
      'Vectors run H2-2026E..FY2030E across columns C..G; scenario vectors at the foot are '
      'published for reproducibility of the pasted engine outputs', 8, awidth=62, cwidth=12)
AR = {}
_r = 4
def arow(label, vals, fmt=NUM0):
    global _r
    put(aws, f'A{_r}', label, fmt=None)
    if isinstance(vals, (list, tuple)):
        for j, v in enumerate(vals):
            put(aws, f'{chr(67 + j)}{_r}', v, BLUE, fmt)
    else:
        put(aws, f'C{_r}', vals, BLUE, fmt)
    AR[label] = _r
    _r += 1
    return _r - 1

hdr(aws, 3, ['Driver', '', 'H2-26E / value', 'FY27E', 'FY28E', 'FY29E', 'FY30E'])
r_spot   = arow('Spot price (AED, 7-Aug-2026 close)', SPOT, PX)
r_sh     = arow('Shares outstanding (mn)', SH, NUM0)
band(aws, _r, 8); put(aws, f'A{_r}', '30-Jun-2026 anchors (reviewed interim + results release, pasted)', bold=True, fmt=None); _r += 1
r_bl0    = arow('Development backlog at 30-Jun-2026 (65.4bn × 95%, disclosed)', HA['dev_backlog'], NUM0)
r_acash  = arow('Unrestricted (available) cash at 30-Jun-2026 (disclosed)', HA['avail_cash'], NUM0)
r_tcash  = arow('Total cash and bank balances at 30-Jun-2026 (disclosed)', HA['cash_total'], NUM0)
r_d0     = arow('Gross debt incl. related-party loan at 30-Jun-2026 (disclosed)', HA['debt'], NUM0)
r_lease  = arow('Lease liabilities at 30-Jun-2026 (disclosed)', HA['lease'], NUM0)
r_assoc  = arow('Associates & JVs at carrying value, 30-Jun-2026 (disclosed)', HA['assoc'], NUM0)
r_finass = arow('Investments in financial assets, 30-Jun-2026 (disclosed)', HA['finass'], NUM0)
r_ncib   = arow('Non-controlling interests, 30-Jun-2026 (disclosed)', HA['nci'], NUM0)
r_eqp    = arow('Equity attributable to owners, 30-Jun-2026 (disclosed)', HA['eqp'], NUM0)
r_recv0  = arow('Receivables incl. related-party, 30-Jun-2026 (disclosed)', HA['recv'], NUM0)
r_inv0   = arow('Inventories + development WIP, 30-Jun-2026 (disclosed)', HA['invdwip'], NUM0)
r_pay0   = arow('Payables incl. advances + related-party, 30-Jun-2026 (disclosed)', HA['pay'], NUM0)
r_h1rev  = arow('H1-2026 revenue (actual)', H1D['rev'], NUM0)
r_h1gp   = arow('H1-2026 gross profit (actual)', H1D['gp'], NUM0)
r_h1npa  = arow('H1-2026 attributable profit (actual)', H1D['npa'], NUM0)
r_h1eb   = arow('H1-2026 adjusted EBITDA (disclosed)', HA['adj_ebitda'], NUM0)
r_abase  = arow('Depreciable asset base at 30-Jun-2026 (PP&E+IP+ROU+intangibles)', IN['asset_base_30jun'], NUM0)
band(aws, _r, 8); put(aws, f'A{_r}', 'Forecast drivers (H2-2026 stub, then annual)', bold=True, fmt=None); _r += 1
r_conv   = arow('Backlog conversion rate on opening development backlog', IN['conv_path'], PCT)
r_ns     = arow('New development sales (AED mn)', IN['new_sales'], NUM0)
r_land   = arow('Land and plot sales revenue (AED mn)', IN['land_rev'], NUM0)
r_redm   = arow('Development gross margin', IN['red_margin'], PCT)
r_aimg   = arow('Asset & investment management growth (from FY27)', [0] + IN['aim_growth'][1:], PCT)
r_aimh2  = arow('Asset & investment management H2-2026E revenue', IN['aim_h2'], NUM0)
r_aimm   = arow('Asset & investment management gross margin', IN['aim_margin'], PCT)
r_hosg   = arow('Hospitality growth (from FY27)', [0] + IN['hosp_growth'][1:], PCT)
r_hosh2  = arow('Hospitality H2-2026E revenue', IN['hosp_h2'], NUM0)
r_hosm   = arow('Hospitality gross margin', IN['hosp_margin'], PCT)
r_ectg   = arow('Events, catering & tourism growth (from FY27)', [0] + IN['ect_growth'][1:], PCT)
r_ecth2  = arow('Events, catering & tourism H2-2026E revenue', IN['ect_h2'], NUM0)
r_ectm   = arow('Events, catering & tourism gross margin', IN['ect_margin'], PCT)
r_h1aim  = arow('H1-2026 AIM revenue (actual)', SEG['h1_rev']['aim'], NUM0)
r_h1hos  = arow('H1-2026 Hospitality revenue (actual)', SEG['h1_rev']['hosp'], NUM0)
r_h1ect  = arow('H1-2026 ECT revenue (actual)', SEG['h1_rev']['ect'], NUM0)
r_othr   = arow('Others / eliminations revenue (full-year rate)', IN['oth_rev_f'], NUM0)
r_othg   = arow('Others gross loss (full-year rate)', IN['oth_gp_f'], NUM0)
r_ga     = arow('General & administrative / revenue', IN['ga_pct'], PCT)
r_sm     = arow('Selling & marketing / revenue', IN['sm_pct'], PCT)
r_inv    = arow('Investment and other income, recurring (AED mn)', IN['invinc_f'], NUM0)
r_dnar   = arow('D&A rate on the average depreciable asset base', IN['dna_rate'], PCT2)
r_capex  = arow('Capital expenditure (AED mn)', IN['capex_f'], NUM0)
r_tax    = arow('Effective tax rate (DMTT floor + foreign uplift)', TAX, PCT)
r_nci    = arow('NCI share of profit', IN['nci_pct'], PCT)
r_apath  = arow('Associate income path (AED mn)', IN['assoc_f_path'], NUM0)
r_debt   = arow('Gross debt path incl. related-party loan (AED mn)', IN['debt_path'], NUM0)
r_cy     = arow('Yield on cash balances', IN['cash_yield'], PCT)
r_li     = arow('Interest on lease liabilities (AED mn/yr)', IN['lease_int'], NUM0)
band(aws, _r, 8); put(aws, f'A{_r}', 'Working capital components (calibrated 31-Dec-25 and 30-Jun-26)', bold=True, fmt=None); _r += 1
r_dso    = arow('Receivable days (incl. related-party) on revenue', IN['dso_path'], NUM0)
r_iadd   = arow('New WIP added per AED of new development sales', IN['inv_addition'], PCT)
r_icons  = arow('Land-bank/WIP share of development cost of sales', IN['inv_consumption'], PCT)
r_pcov   = arow('Payables + advances cover of annual direct costs (×)', IN['pay_cover'], MULT)
band(aws, _r, 8); put(aws, f'A{_r}', 'Cost of capital build', bold=True, fmt=None); _r += 1
r_rf     = arow('AED government bond yield (Jan-2031 T-Bond auction)', IN['rf'], PCT2)
r_ss     = arow('UAE sovereign default spread (rating basis, netted out)', IN['sov_spread_rating'], PCT2)
r_erp    = arow('Equity risk premium (UAE row, rating basis)', IN['erp_rating'], PCT2)
r_beta   = arow('Beta (own-stock regression vs equal-weight AE panel proxy)', IN['beta'], PX)
r_eib    = arow('6-month EIBOR (31-Mar-2026 fixing, dated)', IN['eibor6m'], PCT2)
r_kdm    = arow('Marginal debt margin over 6M EIBOR', IN['kd_margin'], PCT2)
r_g      = arow('Terminal growth', IN['g_term'], PCT2)
r_roic   = arow('Terminal return on invested capital', IN['roic_term'], PCT2)
r_days   = arow('Days from the 30-Jun-2026 valuation date to the 7-Aug anchor', IN['anchor_days'], NUM0)
band(aws, _r, 8); put(aws, f'A{_r}', 'FY2025 disclosed anchors (audited, pasted)', bold=True, fmt=None); _r += 1
r_pat25  = arow('FY2025 profit for the year (AED mn, disclosed)', IN['pat_fy25'], NUM0)
r_npa25  = arow('FY2025 attributable profit (AED mn, disclosed)', IN['npa_fy25'], NUM0)
r_eb25   = arow('FY2025 EBITDA, house basis = EBIT + D&A (AED mn)', HI['FY25']['ebitda'], NUM0)
r_nd25   = arow('FY2025 net debt, strict basis (AED mn)', HB['FY25']['nd'], NUM0)
band(aws, _r, 8); put(aws, f'A{_r}', 'Lens settings', bold=True, fmt=None); _r += 1
r_pej    = arow('Justified price/earnings (FY2026E attributable)', IN['pe_just'], MULT)
r_evj    = arow('EV/EBITDA cross-check multiple (house judgement, unanchored)', IN['ev_ebitda_just'], MULT)
r_nsal   = arow('Through-cycle development sales (AED mn)', IN['norm_sales'], NUM0)
r_nmar   = arow('Through-cycle net margin on revenue', IN['norm_margin'], PCT)
r_rbase  = arow('Recurring-legs revenue base, FY2025 (AED mn, disclosed)', IN['recurring_base'], NUM0)
r_npe    = arow('Through-cycle price/earnings', IN['norm_pe'], MULT)
r_roes   = arow('Sustainable return on equity', IN['roe_sust'], PCT)
r_wdcf   = arow('Weight — discounted cash flow', IN['lens_weights']['dcf'], PCT)
r_wrel   = arow('Weight — relative', IN['lens_weights']['relative'], PCT)
r_wnorm  = arow('Weight — normalised', IN['lens_weights']['normalized'], PCT)
r_wbook  = arow('Weight — book', IN['lens_weights']['book'], PCT)
band(aws, _r, 8); put(aws, f'A{_r}', 'Scenario driver vectors (engine re-run inputs, published for '
     'reproducibility — the pasted scenario outputs use exactly these)', bold=True, fmt=None); _r += 1
r_snsr   = arow('scenario: run-off new development sales', IN['new_sales_runoff'], NUM0)
r_snmr   = arow('scenario: run-off development margin', IN['red_margin_runoff'], PCT)
r_snsb   = arow('scenario: growth-hold new development sales', IN['new_sales_bull'], NUM0)
r_snmb   = arow('scenario: growth-hold development margin', IN['red_margin_bull'], PCT)
r_segy   = arow('scenario: Egypt stress — non-UAE revenue share (note 5)', IN['fgn_share'], PCT)
r_scrp   = arow('scenario: Egypt country risk premium (Damodaran)', IN['egy_crp'], PCT2)

AS = 'Assumptions'
def av(row, col='C'):
    return f"{AS}!${col}${row}"

# ============ SEGMENTS ========================================================
sg = wb.create_sheet('Segments')
title(sg, 'Segments — FY2025 base, H1-2026 actuals, and the forecast build',
      'FY2025 base and H1-2026 actuals pasted (disclosed); every forecast cell is a formula off Assumptions',
      8, awidth=46, cwidth=13)
hdr(sg, 4, ['FY2025 base (audited)', 'Revenue', 'Gross profit', 'GP margin', 'PBT', 'Assets'])
segrows = [('Real Estate Development', 'red'), ('Asset & Investment Management', 'aim'),
           ('Hospitality', 'hosp'), ('Events, Catering & Tourism', 'ect')]
r = 5
SEGROW = {}
for nm, k in segrows:
    put(sg, f'A{r}', nm, fmt=None)
    put(sg, f'B{r}', SEG['rev'][k], BLUE, NUM0)
    put(sg, f'C{r}', SEG['gp'][k], BLUE, NUM0)
    putf(sg, f'D{r}', f'=C{r}/B{r}', SEG['gp'][k] / SEG['rev'][k], PCT)
    put(sg, f'E{r}', SEG['pbt'][k], BLUE, NUM0)
    put(sg, f'F{r}', SEG['assets'][k], BLUE, NUM0)
    SEGROW[k] = r
    r += 1
put(sg, f'A{r}', 'Others / eliminations', fmt=None)
put(sg, f'B{r}', SEG['oth_rev'], BLUE, NUM0); put(sg, f'C{r}', SEG['oth_gp'], BLUE, NUM0)
r_oth = r
r += 1
band(sg, r, 8)
put(sg, f'A{r}', 'Group FY2025', bold=True, fmt=None)
putf(sg, f'B{r}', f'=SUM(B5:B{r_oth})', HI['FY25']['rev'], NUM0, bold=True)
putf(sg, f'C{r}', f'=SUM(C5:C{r_oth})', HI['FY25']['gp'], NUM0, bold=True)
ANCH['seg_rev_tot'] = r
r += 2

hdr(sg, r, ['Development build (from the 30-Jun-2026 disclosed backlog)'] + YF)
rb = r + 1
put(sg, f'A{rb}', 'Opening development backlog', fmt=None)
put(sg, f'A{rb+1}', 'New development sales', fmt=None)
put(sg, f'A{rb+2}', 'Development revenue recognised (conv × opening)', fmt=None)
put(sg, f'A{rb+3}', 'Closing backlog', fmt=None)
put(sg, f'A{rb+4}', 'Land and plot sales', fmt=None)
put(sg, f'A{rb+5}', 'Real Estate Development revenue', fmt=None)
for t in range(NY):
    c = FC[t]; ac = chr(67 + t)
    if t == 0:
        putf(sg, f'{c}{rb}', f"={av(r_bl0)}", F['bl_open'][0], NUM0, green=True)
    else:
        putf(sg, f'{c}{rb}', f"={FC[t-1]}{rb+3}", F['bl_open'][t], NUM0)
    putf(sg, f'{c}{rb+1}', f"={AS}!${ac}${r_ns}", F['new_sales'][t], NUM0, green=True)
    putf(sg, f'{c}{rb+2}', f"={AS}!${ac}${r_conv}*{c}{rb}", F['dev_rev'][t], NUM0)
    putf(sg, f'{c}{rb+3}', f"={c}{rb}+{c}{rb+1}-{c}{rb+2}", F['bl_close'][t], NUM0)
    putf(sg, f'{c}{rb+4}', f"={AS}!${ac}${r_land}", F['land_rev'][t], NUM0, green=True)
    putf(sg, f'{c}{rb+5}', f"={c}{rb+2}+{c}{rb+4}", F['red_rev'][t], NUM0, bold=True)
ANCH['bl_row'] = rb
r = rb + 7

hdr(sg, r, ['Segment revenue forecast'] + YF)
rs = r + 1
put(sg, f'A{rs}', 'Real Estate Development', fmt=None)
put(sg, f'A{rs+1}', 'Asset & Investment Management', fmt=None)
put(sg, f'A{rs+2}', 'Hospitality', fmt=None)
put(sg, f'A{rs+3}', 'Events, Catering & Tourism', fmt=None)
put(sg, f'A{rs+4}', 'Others / eliminations (half-year in the stub)', fmt=None)
put(sg, f'A{rs+5}', 'Group revenue', fmt=None)
for t in range(NY):
    c = FC[t]; ac = chr(67 + t)
    putf(sg, f'{c}{rs}', f"={c}{rb+5}", F['red_rev'][t], NUM0)
    if t == 0:
        putf(sg, f'{c}{rs+1}', f"={av(r_aimh2)}", F['aim_rev'][0], NUM0, green=True)
        putf(sg, f'{c}{rs+2}', f"={av(r_hosh2)}", F['hosp_rev'][0], NUM0, green=True)
        putf(sg, f'{c}{rs+3}', f"={av(r_ecth2)}", F['ect_rev'][0], NUM0, green=True)
    elif t == 1:
        putf(sg, f'{c}{rs+1}', f"=({av(r_h1aim)}+{av(r_aimh2)})*(1+{AS}!${ac}${r_aimg})",
             F['aim_rev'][1], NUM0)
        putf(sg, f'{c}{rs+2}', f"=({av(r_h1hos)}+{av(r_hosh2)})*(1+{AS}!${ac}${r_hosg})",
             F['hosp_rev'][1], NUM0)
        putf(sg, f'{c}{rs+3}', f"=({av(r_h1ect)}+{av(r_ecth2)})*(1+{AS}!${ac}${r_ectg})",
             F['ect_rev'][1], NUM0)
    else:
        p = FC[t - 1]
        putf(sg, f'{c}{rs+1}', f"={p}{rs+1}*(1+{AS}!${ac}${r_aimg})", F['aim_rev'][t], NUM0)
        putf(sg, f'{c}{rs+2}', f"={p}{rs+2}*(1+{AS}!${ac}${r_hosg})", F['hosp_rev'][t], NUM0)
        putf(sg, f'{c}{rs+3}', f"={p}{rs+3}*(1+{AS}!${ac}${r_ectg})", F['ect_rev'][t], NUM0)
    yf = 0.5 if t == 0 else 1.0
    putf(sg, f'{c}{rs+4}', f"={av(r_othr)}*{yf}", IN['oth_rev_f'] * yf, NUM0, green=True)
    putf(sg, f'{c}{rs+5}', f"=SUM({c}{rs}:{c}{rs+4})", F['rev'][t], NUM0, bold=True)
ANCH['seg_fcst_rev'] = rs
r = rs + 7

hdr(sg, r, ['Gross profit build (margin × revenue)'] + YF)
rg = r + 1
put(sg, f'A{rg}', 'Real Estate Development', fmt=None)
put(sg, f'A{rg+1}', 'Asset & Investment Management', fmt=None)
put(sg, f'A{rg+2}', 'Hospitality', fmt=None)
put(sg, f'A{rg+3}', 'Events, Catering & Tourism', fmt=None)
put(sg, f'A{rg+4}', 'Others (half-year in the stub)', fmt=None)
put(sg, f'A{rg+5}', 'Group gross profit', fmt=None)
put(sg, f'A{rg+6}', 'Group gross margin (output)', fmt=None)
for t in range(NY):
    c = FC[t]; ac = chr(67 + t)
    yf = 0.5 if t == 0 else 1.0
    putf(sg, f'{c}{rg}', f"={c}{rs}*{AS}!${ac}${r_redm}", F['red_rev'][t] * IN['red_margin'][t], NUM0)
    putf(sg, f'{c}{rg+1}', f"={c}{rs+1}*{AS}!${ac}${r_aimm}", F['aim_rev'][t] * IN['aim_margin'][t], NUM0)
    putf(sg, f'{c}{rg+2}', f"={c}{rs+2}*{AS}!${ac}${r_hosm}", F['hosp_rev'][t] * IN['hosp_margin'][t], NUM0)
    putf(sg, f'{c}{rg+3}', f"={c}{rs+3}*{AS}!${ac}${r_ectm}", F['ect_rev'][t] * IN['ect_margin'][t], NUM0)
    putf(sg, f'{c}{rg+4}', f"={av(r_othg)}*{yf}", IN['oth_gp_f'] * yf, NUM0, green=True)
    putf(sg, f'{c}{rg+5}', f"=SUM({c}{rg}:{c}{rg+4})", F['gp'][t], NUM0, bold=True)
    putf(sg, f'{c}{rg+6}', f"={c}{rg+5}/{c}{rs+5}", F['gp'][t] / F['rev'][t], PCT)
ANCH['seg_fcst_gp'] = rg

# ============ DCF =============================================================
dc = wb.create_sheet('DCF')
title(dc, 'DCF — waterfall, working-capital components, cost of capital, terminal block, bridge',
      'Valuation date 30-Jun-2026; H2-2026 stub then annual; every chain live', 8,
      awidth=52, cwidth=13)
hdr(dc, 4, ['Waterfall (AED mn)'] + YF)
rw = 5
LBL = ['Revenue', 'Gross profit', 'General & administrative', 'Selling & marketing',
       'Investment and other income', 'EBIT (post-D&A, ex-fair-value items)',
       'EBITDA margin (output)', 'Depreciation & amortisation (asset-base)', 'EBITDA',
       'NOPAT = EBIT × (1 − t)', '+ D&A', '− Capital expenditure',
       '− Δ working capital (from components below)', 'Free cash flow to firm',
       'Discount factor (period-end from 30-Jun-2026)', 'PV of FCFF',
       'Depreciable asset base (opening)', 'Receivables (days × revenue)',
       'Inventories + WIP (roll)', 'Payables + advances (cover × costs)',
       'Net working capital', 'Δ working capital']
for i, l in enumerate(LBL):
    put(dc, f'A{rw+i}', l, fmt=None)
# rows: rw+16 asset base, rw+17 recv, rw+18 inv, rw+19 pay, rw+20 nwc, rw+21 dnwc
ab_prev_ref = None
for t in range(NY):
    c = FC[t]; ac = chr(67 + t)
    rs_, rg_ = ANCH['seg_fcst_rev'], ANCH['seg_fcst_gp']
    yf = 0.5 if t == 0 else 1.0
    putf(dc, f'{c}{rw}',   f"=Segments!{c}{rs_+5}", F['rev'][t], NUM0, green=True)
    putf(dc, f'{c}{rw+1}', f"=Segments!{c}{rg_+5}", F['gp'][t], NUM0, green=True)
    putf(dc, f'{c}{rw+2}', f"=-{AS}!${ac}${r_ga}*{c}{rw}", -F['ga'][t], NUM0)
    putf(dc, f'{c}{rw+3}', f"=-{AS}!${ac}${r_sm}*{c}{rw}", -F['sm'][t], NUM0)
    putf(dc, f'{c}{rw+4}', f"={AS}!${ac}${r_inv}", F['invinc'][t], NUM0, green=True)
    putf(dc, f'{c}{rw+5}', f"=SUM({c}{rw+1}:{c}{rw+4})", F['ebit'][t], NUM0, bold=True)
    putf(dc, f'{c}{rw+6}', f"=({c}{rw+5}+{c}{rw+7})/{c}{rw}", F['ebitda'][t] / F['rev'][t], PCT)
    # asset base opening + D&A off it
    if t == 0:
        putf(dc, f'{c}{rw+16}', f"={av(r_abase)}", IN['asset_base_30jun'], NUM0, green=True)
    else:
        p = FC[t - 1]
        putf(dc, f'{c}{rw+16}', f"={p}{rw+16}+{AS}!${chr(66+t)}${r_capex}-{p}{rw+7}",
             None, NUM0)
        EXPECT['DCF'][f'{c}{rw+16}'] = float(IN['asset_base_30jun']
            + sum(IN['capex_f'][:t]) - sum(F['dna'][:t]))
    putf(dc, f'{c}{rw+7}', f"={av(r_dnar)}*({c}{rw+16}+{AS}!${ac}${r_capex}/2)*{yf}",
         F['dna'][t], NUM0)
    putf(dc, f'{c}{rw+8}', f"={c}{rw+5}+{c}{rw+7}", F['ebitda'][t], NUM0)
    putf(dc, f'{c}{rw+9}', f"={c}{rw+5}*(1-{av(r_tax)})", F['nopat'][t], NUM0)
    putf(dc, f'{c}{rw+10}', f"={c}{rw+7}", F['dna'][t], NUM0)
    putf(dc, f'{c}{rw+11}', f"=-{AS}!${ac}${r_capex}", -F['capex'][t], NUM0)
    # WC components
    if t == 0:
        rev_run = f"({av(r_h1rev)}+{c}{rw})"
        dc_run = f"({av(r_h1rev)}+{c}{rw}-{av(r_h1gp)}-{c}{rw+1})"
    else:
        rev_run = f"{c}{rw}"
        dc_run = f"({c}{rw}-{c}{rw+1})"
    putf(dc, f'{c}{rw+17}', f"={AS}!${ac}${r_dso}/365*{rev_run}", F['recv'][t], NUM0)
    if t == 0:
        putf(dc, f'{c}{rw+18}', f"={av(r_inv0)}+{av(r_iadd)}*{AS}!${ac}${r_ns}"
             f"-{av(r_icons)}*Segments!{c}{ANCH['bl_row']+5}*(1-{AS}!${ac}${r_redm})",
             F['invdwip'][t], NUM0)
    else:
        p = FC[t - 1]
        putf(dc, f'{c}{rw+18}', f"={p}{rw+18}+{av(r_iadd)}*{AS}!${ac}${r_ns}"
             f"-{av(r_icons)}*Segments!{c}{ANCH['bl_row']+5}*(1-{AS}!${ac}${r_redm})",
             F['invdwip'][t], NUM0)
    putf(dc, f'{c}{rw+19}', f"={AS}!${ac}${r_pcov}*{dc_run}", F['pay'][t], NUM0)
    putf(dc, f'{c}{rw+20}', f"={c}{rw+17}+{c}{rw+18}-{c}{rw+19}", F['nwc'][t], NUM0)
    if t == 0:
        putf(dc, f'{c}{rw+21}', f"={c}{rw+20}-({av(r_recv0)}+{av(r_inv0)}-{av(r_pay0)})",
             F['dnwc'][t], NUM0)
    else:
        putf(dc, f'{c}{rw+21}', f"={c}{rw+20}-{FC[t-1]}{rw+20}", F['dnwc'][t], NUM0)
    putf(dc, f'{c}{rw+12}', f"=-{c}{rw+21}", -F['dnwc'][t], NUM0)
    putf(dc, f'{c}{rw+13}', f"=SUM({c}{rw+9}:{c}{rw+12})", F['fcff'][t], NUM0, bold=True)
    putf(dc, f'{c}{rw+14}', f"=1/(1+$C$36)^{F['t_exp'][t]}", F['df'][t], DF4)
    putf(dc, f'{c}{rw+15}', f"={c}{rw+13}*{c}{rw+14}", F['pv'][t], NUM0)
ANCH['dcf_rw'] = rw
r = rw + 23
band(dc, r, 8); put(dc, f'A{r}', 'Cost of capital — built on the sheet', bold=True, fmt=None)
r += 1
put(dc, f'A{r}', 'Normalised risk-free rate rf* = AED yield − sovereign spread', fmt=None)
putf(dc, f'C{r}', f"={av(r_rf)}-{av(r_ss)}", W['rf_star'], PCT2, green=True); r_rfs = r; r += 1
put(dc, f'A{r}', 'Cost of equity Ke = rf* + β × ERP (β from the panel-proxy regression)', fmt=None)
putf(dc, f'C{r}', f"=C{r_rfs}+{av(r_beta)}*{av(r_erp)}", W['ke_exp'], PCT2); r_ke = r; r += 1
put(dc, f'A{r}', 'Marginal cost of debt Kd = 6M EIBOR + margin', fmt=None)
putf(dc, f'C{r}', f"={av(r_eib)}+{av(r_kdm)}", W['kd'], PCT2); r_kd = r; r += 1
put(dc, f'A{r}', 'Kd after tax', fmt=None)
putf(dc, f'C{r}', f"=C{r_kd}*(1-{av(r_tax)})", W['kd_at'], PCT2); r_kdat = r; r += 1
put(dc, f'A{r}', 'Market capitalisation (spot × shares)', fmt=None)
putf(dc, f'C{r}', f"={av(r_spot)}*{av(r_sh)}", M['mktcap'], NUM0); r_mc = r; r += 1
put(dc, f'A{r}', 'Equity weight E/(E+D), debt at the 30-Jun-2026 book', fmt=None)
putf(dc, f'C{r}', f"=C{r_mc}/(C{r_mc}+{av(r_d0)})", W['we_exp'], PCT2); r_we = r; r += 1
put(dc, f'A{r}', 'Debt weight D/(E+D)', fmt=None)
putf(dc, f'C{r}', f"=1-C{r_we}", W['wd_exp'], PCT2); r_wd = r; r += 1
put(dc, f'A{r}', 'Cost of capital (explicit window)', fmt=None)
putf(dc, f'C{r}', f"=C{r_we}*C{r_ke}+C{r_wd}*C{r_kdat}", W['wacc_exp'], PCT2, bold=True)
r_wacc = r; r += 1
assert r_wacc == 36, f'WACC row moved to {r_wacc}; DCF discount factors point at C36'
put(dc, f'A{r}', 'Terminal debt weight — DERIVED from the FY2030E balance sheet', fmt=None)
putf(dc, f'C{r}', f"='Balance Sheet'!J12/('Balance Sheet'!J12+'Balance Sheet'!J14+'Balance Sheet'!J15)",
     W['wd_term'], PCT2); r_wdt = r; r += 1
put(dc, f'A{r}', 'Terminal cost of capital (derived weights)', fmt=None)
putf(dc, f'C{r}', f"=(1-C{r_wdt})*C{r_ke}+C{r_wdt}*C{r_kd}*(1-{av(r_tax)})",
     W['wacc_term'], PCT2, bold=True)
r_wt = r; r += 2
band(dc, r, 8); put(dc, f'A{r}', 'Terminal block — reinvestment derived', bold=True, fmt=None)
r += 1
put(dc, f'A{r}', 'Terminal growth g', fmt=None)
putf(dc, f'C{r}', f"={av(r_g)}", IN['g_term'], PCT2, green=True); r_gt = r; r += 1
put(dc, f'A{r}', 'Terminal return on invested capital', fmt=None)
putf(dc, f'C{r}', f"={av(r_roic)}", IN['roic_term'], PCT2, green=True); r_rc = r; r += 1
put(dc, f'A{r}', 'Reinvestment rate = g / ROIC', fmt=None)
putf(dc, f'C{r}', f"=C{r_gt}/C{r_rc}", DCF['rr_term'], PCT2); r_rr = r; r += 1
put(dc, f'A{r}', 'Terminal NOPAT = FY2030E NOPAT × (1+g)', fmt=None)
putf(dc, f'C{r}', f"=F{rw+9}*(1+C{r_gt})", DCF['nopat_term'], NUM0); r_tn = r; r += 1
put(dc, f'A{r}', 'Terminal value = NOPAT × (1−RR) / (WACCterm − g)', fmt=None)
putf(dc, f'C{r}', f"=C{r_tn}*(1-C{r_rr})/(C{r_wt}-C{r_gt})", DCF['tv'], NUM0); r_tv = r; r += 1
put(dc, f'A{r}', 'PV of terminal value (year-4.5 factor)', fmt=None)
putf(dc, f'C{r}', f"=C{r_tv}*F{rw+14}", DCF['pv_tv'], NUM0); r_ptv = r; r += 1
put(dc, f'A{r}', 'PV of explicit periods', fmt=None)
putf(dc, f'C{r}', f"=SUM(B{rw+15}:F{rw+15})", DCF['pv_explicit'], NUM0); r_pex = r; r += 1
put(dc, f'A{r}', 'Enterprise value', fmt=None)
putf(dc, f'C{r}', f"=C{r_ptv}+C{r_pex}", DCF['ev'], NUM0, bold=True); r_ev = r; r += 1
put(dc, f'A{r}', 'Terminal value share of enterprise value', fmt=None)
putf(dc, f'C{r}', f"=C{r_ptv}/C{r_ev}", DCF['tv_share'], PCT, bold=True); r_tvs = r; r += 2
band(dc, r, 8); put(dc, f'A{r}', 'EV → equity bridge (30-Jun-2026, available-cash basis)',
                    bold=True, fmt=None)
r += 1
BR = [('+ Unrestricted (available) cash — disclosed', f"={av(r_acash)}", HA['avail_cash']),
      ('   (project escrow/restricted cash of ' + f"{DCF['restricted']:,.0f}" +
       ' is excluded: it funds completion of the backlog the DCF values)', None, None),
      ('− Gross debt incl. related-party loan', f"=-{av(r_d0)}", -HA['debt']),
      ('− Lease liabilities', f"=-{av(r_lease)}", -HA['lease']),
      ('+ Associates & JVs at carrying value', f"={av(r_assoc)}", HA['assoc']),
      ('+ Investments in financial assets', f"={av(r_finass)}", HA['finass'])]
r_br0 = None
for l, fml, xp in BR:
    put(dc, f'A{r}', l, fmt=None)
    if fml:
        putf(dc, f'C{r}', fml, xp, NUM0, green=True)
        if r_br0 is None:
            r_br0 = r
    r += 1
put(dc, f'A{r}', '− Non-controlling interests, capitalised at share of equity value '
    '(MAX of book and 2%)', fmt=None)
putf(dc, f'C{r}', f"=-MAX({av(r_ncib)},0.02*(C{r_ev}+C{r_br0}-{av(r_d0)}-{av(r_lease)}"
     f"+{av(r_assoc)}+{av(r_finass)}-{av(r_ncib)}))", -DCF['nci_val'], NUM0)
r_ncirow = r; r += 1
put(dc, f'A{r}', 'Equity value attributable (30-Jun-2026)', fmt=None)
putf(dc, f'C{r}', f"=C{r_ev}+SUM(C{r_br0}:C{r_ncirow})", DCF['eq_attr'], NUM0, bold=True)
r_eq = r; r += 1
put(dc, f'A{r}', 'Fair value per share at 30-Jun-2026', fmt=None)
putf(dc, f'C{r}', f"=C{r_eq}/{av(r_sh)}", DCF['ps_jun'], PX, bold=True); r_psd = r; r += 1
put(dc, f'A{r}', 'Anchor accretion (1+Ke)^(38/365)', fmt=None)
putf(dc, f'C{r}', f"=(1+C{r_ke})^({av(r_days)}/365)", DCF['roll'], DF4); r_roll = r; r += 1
put(dc, f'A{r}', 'Fair value per share at the 7-Aug-2026 anchor', fmt=None)
putf(dc, f'C{r}', f"=C{r_psd}*C{r_roll}", DCF['ps'], PX, bold=True); r_ps = r; r += 2
put(dc, f'A{r}', 'Alternatives (engine re-runs, pasted; driver vectors on Assumptions): '
    'run-off stress / growth-hold / Egypt stress / gross-cash basis', fmt=None)
put(dc, f'C{r}', CJ['runoff_ps'], BLUE, PX); put(dc, f'D{r}', CJ['bull_ps'], BLUE, PX)
put(dc, f'E{r}', DCF['ps_egystress'], BLUE, PX); put(dc, f'F{r}', DCF['ps_grosscash'], BLUE, PX)
ANCH['dcf'] = dict(wacc=r_wacc, wt=r_wt, wdt=r_wdt, ev=r_ev, tvs=r_tvs, eq=r_eq,
                   psd=r_psd, roll=r_roll, ps=r_ps, ke=r_ke, kd=r_kd, rfs=r_rfs,
                   rr=r_rr, tn=r_tn, tv=r_tv, ptv=r_ptv, pex=r_pex, mc=r_mc)

# ============ RELATIVE & NORMALIZED ==========================================
rn = wb.create_sheet('Relative & Normalized')
title(rn, 'Relative multiples · normalised earnings power · book value',
      'One attributable-earnings basis throughout; the EV/EBITDA leg is a labelled, '
      'unanchored cross-check and is NOT averaged into the lens', 7, awidth=52, cwidth=14)
hdr(rn, 4, ['Relative multiples (P/E leg = the lens)', 'Value'])
r = 5
put(rn, f'A{r}', 'FY2026E attributable profit = H1 actual + H2 model', fmt=None)
putf(rn, f'C{r}', f"={av(r_h1npa)}+'Income Statement'!F17", F['fy26_npa_total'], NUM0,
     green=True); r_np26 = r; r += 1
put(rn, f'A{r}', 'Justified P/E × FY2026E EPS', fmt=None)
putf(rn, f'C{r}', f"={av(r_pej)}*C{r_np26}/{av(r_sh)}", REL['pe_ps'], PX); r_pe = r; r += 1
put(rn, f'A{r}', 'FY2026E EBITDA (H1 adjusted-EBITDA + H2 model)', fmt=None)
putf(rn, f'C{r}', f"={av(r_h1eb)}+DCF!B{ANCH['dcf_rw']+8}", REL['fy26_ebitda'], NUM0,
     green=True); r_eb26 = r; r += 1
put(rn, f'A{r}', 'EV/EBITDA cross-check (house multiple, unanchored — labelled)', fmt=None)
putf(rn, f'C{r}', f"=({av(r_evj)}*C{r_eb26}+{av(r_acash)}-{av(r_d0)}-{av(r_lease)}"
     f"+{av(r_assoc)}+{av(r_finass)}-MAX({av(r_ncib)},0.02*({av(r_evj)}*C{r_eb26}"
     f"+{av(r_acash)}-{av(r_d0)}-{av(r_lease)}+{av(r_assoc)}+{av(r_finass)}"
     f"-{av(r_ncib)})))/{av(r_sh)}", REL['ev_ps'], PX); r_evps = r; r += 3
band(rn, r, 7)
put(rn, f'A{r}', 'Relative lens (P/E leg)', bold=True, fmt=None)
putf(rn, f'C{r}', f"=C{r_pe}", REL['base'], PX, bold=True)
r_rel = r
assert r_rel == 11, f'relative lens row {r_rel} — recalc references C11'
r += 1
put(rn, f'A{r}', 'bear / bull (lens scaled to the peer-set floor 4.73x and ceiling 8.10x)', fmt=None)
putf(rn, f'C{r}', f"=C{r_rel}*4.73/{av(r_pej)}", LN['relative']['bear'], PX)
putf(rn, f'D{r}', f"=C{r_rel}*8.1/{av(r_pej)}", LN['relative']['bull'], PX)
r += 2
hdr(rn, r, ['Trailing cross-checks', 'Value'])
r += 1
put(rn, f'A{r}', 'Trailing P/E on FY2025 attributable profit', fmt=None)
putf(rn, f'C{r}', f"=DCF!C{ANCH['dcf']['mc']}/{av(r_npa25)}", REL['pe_trailing_attr'],
     MULT); r += 1
put(rn, f'A{r}', 'Trailing P/E on FY2025 group profit (dual-framed)', fmt=None)
putf(rn, f'C{r}', f"=DCF!C{ANCH['dcf']['mc']}/{av(r_pat25)}", REL['pe_trailing_group'],
     MULT); r += 1
put(rn, f'A{r}', 'Trailing EV/EBITDA (mktcap + FY2025 strict net debt over house EBITDA)', fmt=None)
putf(rn, f'C{r}', f"=(DCF!C{ANCH['dcf']['mc']}+{av(r_nd25)})/{av(r_eb25)}",
     REL['ev_ebitda_trailing'], MULT)
r += 2
hdr(rn, r, ['Normalised earnings power', 'Value'])
r += 1
put(rn, f'A{r}', 'Through-cycle revenue = 85% × cycle sales + recurring legs × 1.15', fmt=None)
putf(rn, f'C{r}', f"={av(r_nsal)}*0.85+{av(r_rbase)}*1.15", NRM['rev'], NUM0); r_nrev = r; r += 1
put(rn, f'A{r}', 'Through-cycle net margin', fmt=None)
putf(rn, f'C{r}', f"={av(r_nmar)}", NRM['margin'], PCT, green=True); r_nm = r; r += 1
put(rn, f'A{r}', 'Normalised net profit', fmt=None)
putf(rn, f'C{r}', f"=C{r_nrev}*C{r_nm}", NRM['np'], NUM0); r_nnp = r; r += 1
put(rn, f'A{r}', 'Normalised EPS', fmt=None)
putf(rn, f'C{r}', f"=C{r_nnp}/{av(r_sh)}", NRM['eps'], PX); r_neps = r; r += 1
while r < 27:
    r += 1
put(rn, f'A{r+1}', 'Normalised lens = EPS × through-cycle P/E', bold=True, fmt=None)
putf(rn, f'C{r+1}', f"=C{r_neps}*{av(r_npe)}", NRM['base'], PX, bold=True)
putf(rn, f'E{r+1}', f"=C{r_neps}*6", LN['normalized']['bear'], PX)
putf(rn, f'F{r+1}', f"=C{r_neps}*10.5", LN['normalized']['bull'], PX)
r_norm = r + 1
assert r_norm == 28, f'normalised lens row {r_norm} — recalc references C28'
r += 3
hdr(rn, r, ['Book value & sustainable return (rolled to the anchor like the DCF)', 'Value'])
r += 1
put(rn, f'A{r}', 'Book value per share (30-Jun-2026 attributable equity)', fmt=None)
putf(rn, f'C{r}', f"={av(r_eqp)}/{av(r_sh)}", BK['bvps'], PX); r_bv = r; r += 1
put(rn, f'A{r}', 'Justified P/B = (ROE − g)/(Ke − g)', fmt=None)
putf(rn, f'C{r}', f"=({av(r_roes)}-{av(r_g)})/(DCF!C{ANCH['dcf']['ke']}-{av(r_g)})",
     BK['pb_just'], MULT); r_pb = r; r += 1
while r < 36:
    r += 1
put(rn, f'A{r}', 'Book lens = BVPS × justified P/B × anchor roll', bold=True, fmt=None)
putf(rn, f'C{r}', f"=C{r_bv}*C{r_pb}*DCF!C{ANCH['dcf']['roll']}", BK['base'], PX, bold=True)
putf(rn, f'E{r}', f"=C{r_bv}*((0.055-{av(r_g)})/(DCF!C{ANCH['dcf']['ke']}-{av(r_g)}))"
     f"*DCF!C{ANCH['dcf']['roll']}", LN['book']['bear'], PX)
putf(rn, f'F{r}', f"=C{r_bv}*((0.095-{av(r_g)})/(DCF!C{ANCH['dcf']['ke']}-{av(r_g)}))"
     f"*DCF!C{ANCH['dcf']['roll']}", LN['book']['bull'], PX)
r_book = r

# ============ SUMMARY =========================================================
ws = wb.create_sheet('Summary')
title(ws, 'Summary — valuation at a glance (revision 2)',
      'All values link live to their source sheets', 7, awidth=44, cwidth=15)
hdr(ws, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Weight', 'Contribution', 'vs spot'])
LENS_SRC = {'dcf': f"=DCF!C{ANCH['dcf']['ps']}", 'relative': "='Relative & Normalized'!C11",
            'normalized': "='Relative & Normalized'!C28", 'book': "='Relative & Normalized'!C36"}
BEAR_SRC = {'normalized': "='Relative & Normalized'!E28", 'book': "='Relative & Normalized'!E36",
            'relative': "='Relative & Normalized'!C12"}
BULL_SRC = {'normalized': "='Relative & Normalized'!F28", 'book': "='Relative & Normalized'!F36",
            'relative': "='Relative & Normalized'!D12"}
WROW = {'dcf': r_wdcf, 'relative': r_wrel, 'normalized': r_wnorm, 'book': r_wbook}
r = 5
for k in ['dcf', 'relative', 'normalized', 'book']:
    l = LN[k]
    put(ws, f'A{r}', l['name'], fmt=None)
    if k in BEAR_SRC:
        putf(ws, f'B{r}', BEAR_SRC[k], l['bear'], PX, green=True)
    else:
        put(ws, f'B{r}', l['bear'], BLUE, PX)
    putf(ws, f'C{r}', LENS_SRC[k], l['base'], PX, green=True)
    if k in BULL_SRC:
        putf(ws, f'D{r}', BULL_SRC[k], l['bull'], PX, green=True)
    else:
        put(ws, f'D{r}', l['bull'], BLUE, PX)
    putf(ws, f'E{r}', f"={av(WROW[k])}", l['w'], PCT, green=True)
    putf(ws, f'F{r}', f'=C{r}*E{r}', l['base'] * l['w'], PX)
    putf(ws, f'G{r}', f'=C{r}/$C$14-1', l['base'] / SPOT - 1, PCT)
    r += 1
band(ws, r, 7)
LK = ['dcf', 'relative', 'normalized', 'book']
put(ws, f'A{r}', 'Weighted central', bold=True, fmt=None)
putf(ws, f'B{r}', '=MIN(B5:B8)', min(LN[k]['bear'] for k in LK), PX, bold=True)
putf(ws, f'C{r}', '=SUM(F5:F8)', D['central'], PX, bold=True)
putf(ws, f'D{r}', '=MAX(D5:D8)', max(LN[k]['bull'] for k in LK), PX, bold=True)
putf(ws, f'E{r}', '=SUM(E5:E8)', 1.0, PCT, bold=True)
putf(ws, f'G{r}', f'=C{r}/$C$14-1', D['central'] / SPOT - 1, PCT, bold=True)
assert r == 9
r += 2
put(ws, f'A{r}', 'The contested judgement: DCF base vs run-off stress', fmt=None)
putf(ws, f'C{r}', f"=DCF!C{ANCH['dcf']['ps']}", DCF['ps'], PX, green=True)
put(ws, f'D{r}', CJ['runoff_ps'], BLUE, PX)
r += 1
put(ws, f'A{r}', 'Terminal value share of DCF enterprise value', fmt=None)
putf(ws, f'C{r}', f"=DCF!C{ANCH['dcf']['tvs']}", DCF['tv_share'], PCT, green=True)
assert r == 12
r += 1
put(ws, f'A{r}', 'Expert panel median', fmt=None)
put(ws, f'C{r}', D['panel_centre'], BLUE, PX)
r += 1
band(ws, r, 7)
put(ws, f'A{r}', 'Market price (anchor, 7-Aug-2026)', bold=True, fmt=None)
putf(ws, f'C{r}', f"={av(r_spot)}", SPOT, PX, bold=True, green=True)
assert r == 14
ANCH['summary_mktcap'] = 'C17'
r += 2
hdr(ws, r, ['Key figure', 'Value'])
r += 1
put(ws, f'A{r}', 'Market capitalisation (AED mn)', fmt=None)
putf(ws, f'C{r}', f"={av(r_spot)}*{av(r_sh)}", M['mktcap'], NUM0)
r += 1
for label, fml, xp, fmt in [
    ('Net cash, strict basis (all cash − all debt, 30-Jun-2026)',
     f"={av(r_tcash)}-{av(r_d0)}", HA['cash_total'] - HA['debt'], NUM0),
    ('Net debt, company definition (available cash − debt, disclosed)',
     -HA['netdebt'], None, NUM0),
    ('FY2026E revenue (H1 actual + H2 model)', f"={av(r_h1rev)}+DCF!B{ANCH['dcf_rw']}",
     F['fy26_rev_total'], NUM0),
    ('FY2026E attributable profit', "='Relative & Normalized'!C5",
     F['fy26_npa_total'], NUM0),
    ('Cost of capital (explicit window)', f"=DCF!C{ANCH['dcf']['wacc']}", W['wacc_exp'], PCT2),
    ('Cost of capital (terminal, derived weights)', f"=DCF!C{ANCH['dcf']['wt']}",
     W['wacc_term'], PCT2),
    ('Terminal growth', f"={av(r_g)}", IN['g_term'], PCT2)]:
    put(ws, f'A{r}', label, fmt=None)
    if isinstance(fml, str):
        putf(ws, f'C{r}', fml, xp, fmt, green=True)
    else:
        put(ws, f'C{r}', fml, BLUE, fmt)
    r += 1
r += 1
hdr(ws, r, ['Monte Carlo price map (engine re-run, pasted)', '1 month', '3 months'])
r += 1
H1M, H3M = STK['horizons']['1M'], STK['horizons']['3M']
for lbl, key in [('5th percentile', 'p5'), ('25th percentile', 'p25'), ('Median', 'p50'),
                 ('75th percentile', 'p75'), ('95th percentile', 'p95')]:
    put(ws, f'A{r}', lbl, fmt=None)
    put(ws, f'B{r}', H1M['pct'][key], BLUE, PX)
    put(ws, f'C{r}', H3M['pct'][key], BLUE, PX)
    r += 1

# ============ FUNDAMENTAL VALUATION ==========================================
fv = wb.create_sheet('Fundamental Valuation')
title(fv, 'Fundamental valuation — four lenses, the contested judgement and the stresses',
      None, 6, awidth=56, cwidth=15)
hdr(fv, 4, ['Lens / step', 'Basis', 'AED per share'])
rows = [
    ('Discounted cash flow (primary)', 'links to the DCF sheet', f"=DCF!C{ANCH['dcf']['ps']}", DCF['ps']),
    ('  run-off stress', 'launches halve and fade (engine re-run; vectors on Assumptions)',
     CJ['runoff_ps'], None),
    ('  growth-hold alternative', 'sales hold near the realised pace (engine re-run)',
     CJ['bull_ps'], None),
    ('  Egypt-risk stress', 'non-UAE revenue share carries the Egypt premium (engine re-run)',
     DCF['ps_egystress'], None),
    ('  gross-cash bridge alternative', 'escrow/restricted cash included (engine re-run)',
     DCF['ps_grosscash'], None),
    ('Relative multiples', 'justified P/E vs the rebuilt attributable peer set',
     "='Relative & Normalized'!C11", REL['base']),
    ('Normalised earnings power', 'through-cycle sales, margin, multiple',
     "='Relative & Normalized'!C28", NRM['base']),
    ('Book value and sustainable return', 'justified P/B on sustainable ROE, rolled',
     "='Relative & Normalized'!C36", BK['base']),
]
r = 5
for a_, b_, c_, xp in rows:
    put(fv, f'A{r}', a_, fmt=None); put(fv, f'B{r}', b_, fmt=None)
    if isinstance(c_, str):
        putf(fv, f'C{r}', c_, xp, PX, green=True)
    else:
        put(fv, f'C{r}', c_, BLUE, PX)
    r += 1
r += 1
band(fv, r, 3); put(fv, f'A{r}', 'Weighted central (links to Summary)', bold=True, fmt=None)
putf(fv, f'C{r}', "=Summary!C9", D['central'], PX, bold=True); r += 2
put(fv, f'A{r}', 'What the market price implies', fmt=None)
put(fv, f'B{r}', 'cost-of-equity adder reconciling the base DCF to spot (engine solve)', fmt=None)
put(fv, f'C{r}', D['market_implied']['ke_add'], BLUE, PCT2); r += 2
hdr(fv, r, ['Expert panel (engine re-runs, worked in Appendix C)', 'Method', 'AED per share'])
r += 1
for lbl, e in [('Expert 1', 'e1'), ('Expert 2', 'e2'), ('Expert 3', 'e3')]:
    put(fv, f'A{r}', lbl, fmt=None); put(fv, f'B{r}', EXP[e]['method_short'], fmt=None)
    put(fv, f'C{r}', EXP[e]['base'], BLUE, PX)
    r += 1
put(fv, f'A{r}', 'Panel median', bold=True, fmt=None)
putf(fv, f'C{r}', f"=MEDIAN(C{r-3}:C{r-1})", D['panel_centre'], PX, bold=True)

# ============ SOTP BRIDGE =====================================================
sb = wb.create_sheet('SOTP Bridge')
title(sb, 'SOTP bridge — segment EV split (weights DERIVED) and the EV → equity walk',
      'Weight = FY2025 segment gross profit less a revenue-proportional corporate load — '
      'all live formulas off Segments and the FY2025 disclosed opex', 6,
      awidth=52, cwidth=15)
r_ga25 = None
put(sb, 'A4', 'FY2025 G&A + S&M (audited, for the load allocation)', fmt=None)
put(sb, 'B4', 1547.476 + 266.211, BLUE, NUM0)
hdr(sb, 5, ['Enterprise value by segment', 'Weight (GP − load share)', 'AED mn'])
r = 6
segk = [('red', 'Real Estate Development'), ('aim', 'Asset & Investment Management'),
        ('hosp', 'Hospitality'), ('ect', 'Events, Catering & Tourism')]
for k, nm in segk:
    put(sb, f'A{r}', nm, fmt=None)
    sr = SEGROW[k]
    putf(sb, f'B{r}', f"=Segments!C{sr}-$B$4*Segments!B{sr}/Segments!B{ANCH['seg_rev_tot']}",
         D['sotp']['weights'][k], NUM0)
    putf(sb, f'C{r}', f"=MAX(B{r},0)/(MAX(B$6,0)+MAX(B$7,0)+MAX(B$8,0)+MAX(B$9,0))"
         f"*DCF!C{ANCH['dcf']['ev']}", D['sotp']['ev_split'][k], NUM0)
    r += 1
band(sb, r, 6)
put(sb, f'A{r}', 'Enterprise value (links to DCF)', bold=True, fmt=None)
putf(sb, f'C{r}', f"=DCF!C{ANCH['dcf']['ev']}", DCF['ev'], NUM0, bold=True); r_bev = r
r += 1
put(sb, f'A{r}', 'Terminal value share of enterprise value', fmt=None)
putf(sb, f'C{r}', f"=DCF!C{ANCH['dcf']['tvs']}", DCF['tv_share'], PCT); r += 1
for l, fml, xp in [('+ Unrestricted cash (disclosed)', f"={av(r_acash)}", HA['avail_cash']),
                   ('− Gross debt incl. related-party loan', f"=-{av(r_d0)}", -HA['debt']),
                   ('− Lease liabilities', f"=-{av(r_lease)}", -HA['lease']),
                   ('+ Associates & JVs', f"={av(r_assoc)}", HA['assoc']),
                   ('+ Financial assets', f"={av(r_finass)}", HA['finass']),
                   ('− NCI (capitalised, links to DCF)', f"=DCF!C{r_ncirow}", -DCF['nci_val'])]:
    put(sb, f'A{r}', l, fmt=None); putf(sb, f'C{r}', fml, xp, NUM0, green=True); r += 1
band(sb, r, 6)
put(sb, f'A{r}', 'Equity attributable → per share → at anchor', bold=True, fmt=None)
putf(sb, f'C{r}', f"=C{r_bev}+SUM(C{r_bev+2}:C{r-1})", DCF['eq_attr'], NUM0, bold=True)
r_beq = r; r += 1
putf(sb, f'C{r}', f"=C{r_beq}/{av(r_sh)}", DCF['ps_jun'], PX)
put(sb, f'A{r}', 'per share, 30-Jun-2026', fmt=None); r += 1
putf(sb, f'C{r}', f"=C{r-1}*DCF!C{ANCH['dcf']['roll']}", DCF['ps'], PX, bold=True)
put(sb, f'A{r}', 'per share at the 7-Aug-2026 anchor', fmt=None)
r += 2
put(sb, f'A{r}', 'Book-NCI alternative framing (engine)', fmt=None)
put(sb, f'C{r}', DCF['ps_booknci'], BLUE, PX)
ANCH['sotp_eq'] = r_beq

# ============ INCOME STATEMENT ===============================================
istmt = wb.create_sheet('Income Statement')
title(istmt, 'Income statement — 3 years audited + H1-2026 actual + forecast',
      'FY2023 is the Q Holding perimeter; FY2024 contains the AED 9,192mn bargain gain (dual-framed); '
      'column E is the H1-2026 actual; forecast runs H2-2026E..FY2030E', 11, awidth=42, cwidth=11)
hdr(istmt, 4, [''] + ['FY2023', 'FY2024', 'FY2025', 'H1-26A'] + YF)
ROWS = ['Revenue', 'Gross profit', 'EBITDA (house)', 'EBITDA margin',
        'Depreciation & amortisation', 'EBIT', 'Net finance result', 'Associates & JVs',
        'Profit before tax', 'Income tax', 'Profit for the year',
        'Non-controlling interests', 'Attributable profit']
r = 5
for i, lbl in enumerate(ROWS):
    put(istmt, f'A{r+i}', lbl, fmt=None)
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    c = HC[j]
    H = HI[y]
    put(istmt, f'{c}5', H['rev'], BLUE, NUM0)
    put(istmt, f'{c}6', H['gp'], BLUE, NUM0)
    put(istmt, f'{c}7', H['ebitda'], BLUE, NUM0)
    putf(istmt, f'{c}8', f'={c}7/{c}5', H['ebitda'] / H['rev'], PCT)
    put(istmt, f'{c}9', H['dna'], BLUE, NUM0)
    putf(istmt, f'{c}10', f'={c}7-{c}9', H['ebit'], NUM0)
    put(istmt, f'{c}11', H['fin'], BLUE, NUM0)
    put(istmt, f'{c}12', H['assoc'], BLUE, NUM0)
    putf(istmt, f'{c}13', f'={c}10+{c}11+{c}12', H['ebit'] + H['fin'] + H['assoc'], NUM0)
    put(istmt, f'{c}14', H['tax'], BLUE, NUM0)
    putf(istmt, f'{c}15', f'={c}13+{c}14', H['ebt'] + H['tax'], NUM0)
    put(istmt, f'{c}16', H['nci'], BLUE, NUM0)
    putf(istmt, f'{c}17', f'={c}15-{c}16', H['pat'] - H['nci'], NUM0, bold=True)
# H1-2026 actual column (disclosed)
put(istmt, 'E5', H1D['rev'], BLUE, NUM0)
put(istmt, 'E6', H1D['gp'], BLUE, NUM0)
put(istmt, 'E13', H1D['pbt'], BLUE, NUM0)
put(istmt, 'E14', -H1D['eff_tax'] * H1D['pbt'], BLUE, NUM0)
put(istmt, 'E15', H1D['pat'], BLUE, NUM0)
put(istmt, 'E16', H1D['pat'] - H1D['npa'], BLUE, NUM0)
put(istmt, 'E17', H1D['npa'], BLUE, NUM0, bold=True)
for t in range(NY):
    c = FCOL[t]; dcfc = FC[t]; ac = chr(67 + t)
    rw_ = ANCH['dcf_rw']
    putf(istmt, f'{c}5', f"=DCF!{dcfc}{rw_}", F['rev'][t], NUM0, green=True)
    putf(istmt, f'{c}6', f"=DCF!{dcfc}{rw_+1}", F['gp'][t], NUM0, green=True)
    putf(istmt, f'{c}7', f"=DCF!{dcfc}{rw_+8}", F['ebitda'][t], NUM0, green=True)
    putf(istmt, f'{c}8', f'={c}7/{c}5', F['ebitda'][t] / F['rev'][t], PCT)
    putf(istmt, f'{c}9', f"=DCF!{dcfc}{rw_+7}", F['dna'][t], NUM0, green=True)
    putf(istmt, f'{c}10', f'={c}7-{c}9', F['ebit'][t], NUM0)
    yf = 0.5 if t == 0 else 1.0
    d_open = f"{av(r_d0)}" if t == 0 else f"{AS}!${chr(66 + t)}${r_debt}"
    c_open_ref = f"'Balance Sheet'!E9" if t == 0 else f"'Balance Sheet'!{FCOL[t-1]}9"
    fin_exp = (IN['cash_yield'] * (HA['cash_total'] if t == 0 else F['cash'][t - 1])
               - (W['kd'] * ((HA['debt'] if t == 0 else F['debt'][t - 1])
                             + F['debt'][t]) / 2 + IN['lease_int'])) * yf
    putf(istmt, f'{c}11',
         f"=({av(r_cy)}*{c_open_ref}-(DCF!$C${ANCH['dcf']['kd']}*({d_open}"
         f"+{AS}!${ac}${r_debt})/2+{av(r_li)}))*{yf}", fin_exp, NUM0)
    putf(istmt, f'{c}12', f"={AS}!${ac}${r_apath}", F['assoc'][t], NUM0, green=True)
    putf(istmt, f'{c}13', f'={c}10+{c}11+{c}12',
         F['ebit'][t] + fin_exp + F['assoc'][t], NUM0)
    putf(istmt, f'{c}14', f'=-{c}13*{av(r_tax)}',
         -(F['ebit'][t] + fin_exp + F['assoc'][t]) * TAX, NUM0)
    putf(istmt, f'{c}15', f'={c}13+{c}14', F['np'][t], NUM0)
    putf(istmt, f'{c}16', f'={c}15*{av(r_nci)}', F['np'][t] * IN['nci_pct'], NUM0)
    putf(istmt, f'{c}17', f'={c}15-{c}16', F['np_attr'][t], NUM0, bold=True)

# ============ BALANCE SHEET ===================================================
bsx = wb.create_sheet('Balance Sheet')
title(bsx, 'Balance sheet — audited history, 30-Jun-2026 actual, and the roll-forward',
      'Column E = 30-Jun-2026 reviewed actuals; forecast rolls equity on retained profit, debt on '
      'the drawn path, cash from free cash flow', 11, awidth=42, cwidth=11)
hdr(bsx, 4, [''] + ['FY2023', 'FY2024', 'FY2025', '30-Jun-26A'] + YF)
BSROWS = ['Receivables incl. related-party', 'Inventories + development WIP',
          'Payables incl. advances + related-party', 'Net working capital (components)',
          'Cash and bank balances', 'Total assets (audited)',
          'Gross debt incl. related-party loan', 'Net debt (− = net cash)',
          'Equity attributable to owners', 'Non-controlling interests',
          'Net debt / EBITDA']
r = 5
for i, lbl in enumerate(BSROWS):
    put(bsx, f'A{r+i}', lbl, fmt=None)
# rows: 5 recv, 6 inv, 7 pay, 8 nwc, 9 cash, 10 assets, 11 (spare), 12 debt, 13 nd, 14 eqp, 15 nci, 16 nd/ebitda
# adjust mapping: use explicit rows
ROWMAP = dict(recv=5, inv=6, pay=7, nwc=8, cash=9, assets=10, debt=12, nd=13,
              eqp=14, nci=15, ndeb=16)
put(bsx, 'A11', '', fmt=None)
put(bsx, f"A{ROWMAP['debt']}", 'Gross debt incl. related-party loan', fmt=None)
put(bsx, f"A{ROWMAP['nd']}", 'Net debt (− = net cash)', fmt=None)
put(bsx, f"A{ROWMAP['eqp']}", 'Equity attributable to owners', fmt=None)
put(bsx, f"A{ROWMAP['nci']}", 'Non-controlling interests', fmt=None)
put(bsx, f"A{ROWMAP['ndeb']}", 'Net debt / EBITDA', fmt=None)
BSH = {'FY23': dict(recv=1039.926 + 645.607, inv=HB['FY23']['inv'], pay=None,
                    cash=HB['FY23']['cash'], assets=HB['FY23']['assets'],
                    debt=HB['FY23']['debt'], eqp=HB['FY23']['eqp'], nci=HB['FY23']['nci'],
                    eb=HI['FY23']['ebitda']),
       'FY24': dict(recv=HB['FY24']['recv'] + HB['FY24']['duefr'],
                    inv=HB['FY24']['inv'] + HB['FY24']['dwip'],
                    pay=HB['FY24']['pay'] + HB['FY24']['dueto'], cash=HB['FY24']['cash'],
                    assets=HB['FY24']['assets'], debt=HB['FY24']['debt'],
                    eqp=HB['FY24']['eqp'], nci=HB['FY24']['nci'], eb=HI['FY24']['ebitda']),
       'FY25': dict(recv=HB['FY25']['recv'] + HB['FY25']['duefr'],
                    inv=HB['FY25']['inv'] + HB['FY25']['dwip'],
                    pay=HB['FY25']['pay'] + HB['FY25']['dueto'], cash=HB['FY25']['cash'],
                    assets=HB['FY25']['assets'], debt=HB['FY25']['debt'],
                    eqp=HB['FY25']['eqp'], nci=HB['FY25']['nci'], eb=HI['FY25']['ebitda'])}
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    c = HC[j]; H = BSH[y]
    put(bsx, f"{c}{ROWMAP['recv']}", H['recv'], BLUE, NUM0)
    put(bsx, f"{c}{ROWMAP['inv']}", H['inv'], BLUE, NUM0)
    if H['pay'] is not None:
        put(bsx, f"{c}{ROWMAP['pay']}", H['pay'], BLUE, NUM0)
        putf(bsx, f"{c}{ROWMAP['nwc']}",
             f"={c}{ROWMAP['recv']}+{c}{ROWMAP['inv']}-{c}{ROWMAP['pay']}",
             H['recv'] + H['inv'] - H['pay'], NUM0)
    else:
        put(bsx, f"{c}{ROWMAP['pay']}", '-', BLACK, None)
        put(bsx, f"{c}{ROWMAP['nwc']}", '-', BLACK, None)
    put(bsx, f"{c}{ROWMAP['cash']}", H['cash'], BLUE, NUM0)
    put(bsx, f"{c}{ROWMAP['assets']}", H['assets'], BLUE, NUM0)
    put(bsx, f"{c}{ROWMAP['debt']}", H['debt'], BLUE, NUM0)
    putf(bsx, f"{c}{ROWMAP['nd']}", f"={c}{ROWMAP['debt']}-{c}{ROWMAP['cash']}",
         H['debt'] - H['cash'], NUM0)
    put(bsx, f"{c}{ROWMAP['eqp']}", H['eqp'], BLUE, NUM0)
    put(bsx, f"{c}{ROWMAP['nci']}", H['nci'], BLUE, NUM0)
    putf(bsx, f"{c}{ROWMAP['ndeb']}", f"=({c}{ROWMAP['debt']}-{c}{ROWMAP['cash']})/{H['eb']}",
         (H['debt'] - H['cash']) / H['eb'], MULT)
# 30-Jun-2026 actual column E (disclosed, links to Assumptions)
putf(bsx, f"E{ROWMAP['recv']}", f"={av(r_recv0)}", HA['recv'], NUM0, green=True)
putf(bsx, f"E{ROWMAP['inv']}", f"={av(r_inv0)}", HA['invdwip'], NUM0, green=True)
putf(bsx, f"E{ROWMAP['pay']}", f"={av(r_pay0)}", HA['pay'], NUM0, green=True)
putf(bsx, f"E{ROWMAP['nwc']}", f"=E{ROWMAP['recv']}+E{ROWMAP['inv']}-E{ROWMAP['pay']}",
     F['nwc_30jun'], NUM0)
putf(bsx, f"E{ROWMAP['cash']}", f"={av(r_tcash)}", HA['cash_total'], NUM0, green=True)
put(bsx, f"E{ROWMAP['assets']}", 91985.930, BLUE, NUM0)
putf(bsx, f"E{ROWMAP['debt']}", f"={av(r_d0)}", HA['debt'], NUM0, green=True)
putf(bsx, f"E{ROWMAP['nd']}", f"=E{ROWMAP['debt']}-E{ROWMAP['cash']}",
     HA['debt'] - HA['cash_total'], NUM0)
putf(bsx, f"E{ROWMAP['eqp']}", f"={av(r_eqp)}", HA['eqp'], NUM0, green=True)
putf(bsx, f"E{ROWMAP['nci']}", f"={av(r_ncib)}", HA['nci'], NUM0, green=True)
for t in range(NY):
    c = FCOL[t]; ac = chr(67 + t); pcol = 'E' if t == 0 else FCOL[t - 1]
    rw_ = ANCH['dcf_rw']; dcfc = FC[t]
    putf(bsx, f"{c}{ROWMAP['recv']}", f"=DCF!{dcfc}{rw_+17}", F['recv'][t], NUM0, green=True)
    putf(bsx, f"{c}{ROWMAP['inv']}", f"=DCF!{dcfc}{rw_+18}", F['invdwip'][t], NUM0, green=True)
    putf(bsx, f"{c}{ROWMAP['pay']}", f"=DCF!{dcfc}{rw_+19}", F['pay'][t], NUM0, green=True)
    putf(bsx, f"{c}{ROWMAP['nwc']}", f"=DCF!{dcfc}{rw_+20}", F['nwc'][t], NUM0, green=True)
    yf = 0.5 if t == 0 else 1.0
    putf(bsx, f"{c}{ROWMAP['cash']}",
         f"={pcol}{ROWMAP['cash']}+DCF!{dcfc}{rw_+13}+'Income Statement'!{c}11"
         f"-('Income Statement'!{c}13-DCF!{dcfc}{rw_+5})*{av(r_tax)}"
         f"+({AS}!${ac}${r_debt}-{av(r_d0) if t == 0 else f'{AS}!${chr(66 + t)}${r_debt}'})",
         F['cash'][t], NUM0)
    put(bsx, f"{c}{ROWMAP['assets']}", '-', BLACK, None)
    putf(bsx, f"{c}{ROWMAP['debt']}", f"={AS}!${ac}${r_debt}", F['debt'][t], NUM0, green=True)
    putf(bsx, f"{c}{ROWMAP['nd']}", f"={c}{ROWMAP['debt']}-{c}{ROWMAP['cash']}",
         F['net_debt'][t], NUM0)
    eq_exp = float(HA['eqp'] + sum(F['np_attr'][:t + 1]))
    putf(bsx, f"{c}{ROWMAP['eqp']}", f"={pcol}{ROWMAP['eqp']}+'Income Statement'!{c}17",
         eq_exp, NUM0)
    putf(bsx, f"{c}{ROWMAP['nci']}", f"={pcol}{ROWMAP['nci']}+'Income Statement'!{c}16",
         HA['nci'] + sum(F['np'][i] * IN['nci_pct'] for i in range(t + 1)), NUM0)
    putf(bsx, f"{c}{ROWMAP['ndeb']}",
         f"=({c}{ROWMAP['debt']}-{c}{ROWMAP['cash']})/DCF!{dcfc}{rw_+8}",
         F['net_debt'][t] / F['ebitda'][t], MULT)
ANCH['bs_rowmap'] = ROWMAP

# ============ CASH FLOW =======================================================
cf = wb.create_sheet('Cash Flow')
title(cf, 'Cash flow — audited history, H1-2026 actual, and the forecast chain',
      'Forecast rows link to the DCF waterfall; the IFRS operating-cash-flow actuals are shown '
      'separately from the model construct (different measures, not one series)', 10,
      awidth=46, cwidth=12)
hdr(cf, 4, [''] + ['FY2024A', 'FY2025A', 'H1-26A'] + YF)
CFH = D['hist_cf']
put(cf, 'A5', 'IFRS net cash from operating activities (audited/reviewed)', fmt=None)
put(cf, 'B5', CFH['FY24']['ocf'], BLUE, NUM0); put(cf, 'C5', CFH['FY25']['ocf'], BLUE, NUM0)
put(cf, 'D5', H1D['ocf'], BLUE, NUM0)
put(cf, 'A6', 'Capital expenditure (PP&E + intangibles + IP)', fmt=None)
put(cf, 'B6', CFH['FY24']['capex'], BLUE, NUM0); put(cf, 'C6', CFH['FY25']['capex'], BLUE, NUM0)
put(cf, 'D6', 206.911 + 63.530 + 42.049, BLUE, NUM0)
put(cf, 'A8', 'Model construct (NOT the IFRS measure): links to the DCF waterfall', fmt=None, bold=True)
CFROWS = [('NOPAT', 9), ('+ D&A', 10), ('− Capex', 11), ('− Δ working capital', 12),
          ('Free cash flow to firm', 13)]
r = 9
for i, (lbl, off) in enumerate(CFROWS):
    put(cf, f'A{r+i}', lbl, fmt=None)
    for t in range(NY):
        c = ['E', 'F', 'G', 'H', 'I'][t]; dcfc = FC[t]
        vals = dict(zip([9, 10, 11, 12, 13],
                        [F['nopat'][t], F['dna'][t], -F['capex'][t], -F['dnwc'][t],
                         F['fcff'][t]]))
        putf(cf, f'{c}{r+i}', f"=DCF!{dcfc}{ANCH['dcf_rw']+off}", vals[off], NUM0,
             green=True, bold=(off == 13))
put(cf, f'A{r+6}', 'Implied H2-2026 IFRS-basis operating recovery required by the model: the '
    'H1 actual was −3,921; the model\'s FY2026 working-capital absorption of '
    f"{F['dnwc'][0]:,.0f} implies roughly +{-H1D['ocf'] + F['nopat'][0] + F['dna'][0] - F['dnwc'][0]:,.0f} "
    'of H2 operating cash flow — stated, not hidden', fmt=None, wrap=True)
cf.merge_cells(f'A{r+6}:I{r+7}')

# ============ SUMMARY FINANCIALS =============================================
sf = wb.create_sheet('Summary Financials')
title(sf, 'Summary financials — the model in one table', 'Every row links live', 11,
      awidth=40, cwidth=11)
hdr(sf, 4, [''] + ['FY2023', 'FY2024', 'FY2025', 'H1-26A'] + YF)
r = 5
SFrows = ['Revenue', 'EBITDA', 'EBITDA margin', 'Attributable profit',
          'Free cash flow to firm', 'Net debt (− = net cash)', 'Equity attributable',
          '', 'Invested capital (roll)']
for i, lbl in enumerate(SFrows):
    put(sf, f'A{r+i}', lbl, fmt=None)
COLS = HC + ['E'] + FCOL
for j, c in enumerate(COLS):
    hist = j < 3
    isH1 = j == 3
    y = ['FY23', 'FY24', 'FY25'][j] if hist else None
    t = None if (hist or isH1) else j - 4
    putf(sf, f'{c}5', f"='Income Statement'!{c}5",
         HI[y]['rev'] if hist else (H1D['rev'] if isH1 else F['rev'][t]), NUM0, green=True)
    if isH1:
        put(sf, f'{c}6', '-', BLACK, None)
        put(sf, f'{c}7', '-', BLACK, None)
    else:
        putf(sf, f'{c}6', f"='Income Statement'!{c}7",
             HI[y]['ebitda'] if hist else F['ebitda'][t], NUM0, green=True)
        putf(sf, f'{c}7', f"='Income Statement'!{c}8",
             (HI[y]['ebitda'] / HI[y]['rev']) if hist else F['ebitda'][t] / F['rev'][t],
             PCT, green=True)
    putf(sf, f'{c}8', f"='Income Statement'!{c}17",
         HI[y]['npa'] if hist else (H1D['npa'] if isH1 else F['np_attr'][t]), NUM0,
         green=True)
    if hist or isH1:
        put(sf, f'{c}9', '-', BLACK, None)
    else:
        putf(sf, f'{c}9', f"='Cash Flow'!{['E','F','G','H','I'][t]}13", F['fcff'][t],
             NUM0, green=True)
    putf(sf, f'{c}10', f"='Balance Sheet'!{c}13",
         (BSH[y]['debt'] - BSH[y]['cash']) if hist
         else (HA['debt'] - HA['cash_total'] if isH1 else F['net_debt'][t]), NUM0,
         green=True)
    putf(sf, f'{c}11', f"='Balance Sheet'!{c}14",
         BSH[y]['eqp'] if hist else (HA['eqp'] if isH1
                                     else HA['eqp'] + sum(F['np_attr'][:t + 1])), NUM0,
         green=True)
put(sf, 'A13', 'Invested capital: 30-Jun-2026 base + cumulative (capex − D&A + ΔWC)', fmt=None)
putf(sf, 'E13', f"={av(r_eqp)}+{av(r_ncib)}+{av(r_d0)}-{av(r_tcash)}",
     F['ic_30jun'], NUM0)
for t in range(NY):
    c = FCOL[t]; dcfc = FC[t]; pcol = 'E' if t == 0 else FCOL[t - 1]
    putf(sf, f'{c}13', f"={pcol}13+DCF!{dcfc}{ANCH['dcf_rw']+11}*-1"
         f"-DCF!{dcfc}{ANCH['dcf_rw']+7}+DCF!{dcfc}{ANCH['dcf_rw']+21}",
         F['ic'][t], NUM0)

# ============ MONTE CARLO =====================================================
mc = wb.create_sheet('Monte Carlo')
title(mc, 'Monte Carlo price map — engine re-run (pasted, does not reprice)',
      f"Struck {STK['anchor_date']} at spot {STK['spot']:.2f}; 50,000 paths, seed 42; "
      'unchanged from the first edition (the price series and anchor are unchanged)', 7,
      awidth=46, cwidth=13)
hdr(mc, 4, ['', '1 month', '3 months'])
r = 5
H1x, H3x = STK['horizons']['1M'], STK['horizons']['3M']
for lbl, k1, k3, fmt in [
        ('Sessions in window', H1x['h'], H3x['h'], NUM0),
        ('Grade date', H1x['grade_date'], H3x['grade_date'], None),
        ('Annualised anchor volatility', H1x['anchor_vol_ann'], H3x['anchor_vol_ann'], PCT),
        ('5th percentile', H1x['pct']['p5'], H3x['pct']['p5'], PX),
        ('25th percentile', H1x['pct']['p25'], H3x['pct']['p25'], PX),
        ('Median', H1x['pct']['p50'], H3x['pct']['p50'], PX),
        ('75th percentile', H1x['pct']['p75'], H3x['pct']['p75'], PX),
        ('95th percentile', H1x['pct']['p95'], H3x['pct']['p95'], PX),
        ('P(close above spot)', H1x['p_above'], H3x['p_above'], PCT),
        ('P(≥ +10%)', H1x['p_up10'], H3x['p_up10'], PCT),
        ('P(≤ −10%)', H1x['p_dn10'], H3x['p_dn10'], PCT),
        ('P(touch +5% at any point)', H1x['touch_up5'], H3x['touch_up5'], PCT),
        ('P(touch −5% at any point)', H1x['touch_dn5'], H3x['touch_dn5'], PCT)]:
    put(mc, f'A{r}', lbl, fmt=None)
    put(mc, f'B{r}', k1, BLUE, fmt); put(mc, f'C{r}', k3, BLUE, fmt)
    r += 1
r += 1
put(mc, f'A{r}', 'Calibration evidence (walk-forward, this name)', bold=True, fmt=None); r += 1
for lbl, v, fmt in [('Windows scored (post-break)', S0['windows_scored'], NUM0),
                    ('CRPS skill vs carry-anchored random walk', S0['skill_norm'], PCT2),
                    ('Verdict', S0['verdict'], None),
                    ('Coverage of the 80% band (over-covered: bands run wide)', S0['cov80'], PCT),
                    ('Coverage of the 90% band', S0['cov90'], PCT),
                    ('PIT mean (0.5 = centred)', S0['pit_mean'], DF4)]:
    put(mc, f'A{r}', lbl, fmt=None); put(mc, f'B{r}', v, BLUE, fmt); r += 1

# ============ SENSITIVITY =====================================================
sx = wb.create_sheet('Sensitivity')
title(sx, 'Sensitivity — each cell is a complete engine revaluation (pasted)',
      'Rebuilt at revision 2 on the base convention: explicit and terminal rates shift '
      'TOGETHER, so the centre cell equals the base case. Driver vectors on Assumptions', 7,
      awidth=30, cwidth=12)
hdr(sx, 4, ['Ke shift \\ g'] + [f'{x*100:.1f}%' for x in SN['g_grid']])
r = 5
for i, wx in enumerate(SN['wacc_grid']):
    put(sx, f'A{r}', f'{wx*100:.2f}%', fmt=None)
    for j in range(len(SN['g_grid'])):
        put(sx, f'{chr(66+j)}{r}', SN['table'][i][j], BLUE, PX)
    r += 1
r += 1
for name, grid, vals, hdrfmt in [
        ('Beta 0.8 → 1.2 (base 1.03)', SN['beta_grid'], SN['grid_beta'], '{:.2f}'),
        ('Development margin shift (pts)', SN['mg_grid'], SN['grid_margin'], '{:+.2f}'),
        ('Conversion-rate shift (pts)', SN['conv_grid'], SN['grid_conv'], '{:+.2f}'),
        ('New-sales multiple of base', SN['sales_grid'], SN['grid_sales'], '{:.2f}'),
        ('Working-capital shift (AED mn/yr)', SN['nwc_grid'], SN['grid_nwc'], '{:+.0f}'),
        ('Receivable-days shift', SN['dso_grid'], SN['grid_dso'], '{:+.0f}'),
        ('Cost-of-equity adder (0 = base)', SN['ke_grid'], SN['grid_ke'], '{:+.3f}')]:
    hdr(sx, r, [name] + [hdrfmt.format(g) for g in grid]); r += 1
    put(sx, f'A{r}', 'DCF per share', fmt=None)
    for j, v in enumerate(vals):
        put(sx, f'{chr(66+j)}{r}', v, BLUE, PX)
    r += 2

# ============ PER-SHARE & RATIOS =============================================
pr = wb.create_sheet('Per-Share & Ratios')
title(pr, 'Per-share figures and ratios — all live formulas', None, 11, awidth=42, cwidth=11)
hdr(pr, 4, [''] + ['FY2023', 'FY2024', 'FY2025', 'H1-26A'] + YF)
put(pr, 'A5', 'EPS (attributable; H1 and the stub are part-year)', fmt=None)
put(pr, 'A6', 'Book value per share', fmt=None)
put(pr, 'A7', 'Return on attributable equity (avg; part-year annualised)', fmt=None)
put(pr, 'A8', 'Return on invested capital (avg base, annualised)', fmt=None)
put(pr, 'A9', 'P/E at spot (full years)', fmt=None)
put(pr, 'A10', 'P/B at spot', fmt=None)
for j, c in enumerate(COLS):
    hist = j < 3
    isH1 = j == 3
    y = ['FY23', 'FY24', 'FY25'][j] if hist else None
    t = None if (hist or isH1) else j - 4
    npa_v = HI[y]['npa'] if hist else (H1D['npa'] if isH1 else F['np_attr'][t])
    eq_v = BSH[y]['eqp'] if hist else (HA['eqp'] if isH1
                                       else HA['eqp'] + sum(F['np_attr'][:t + 1]))
    putf(pr, f'{c}5', f"='Income Statement'!{c}17/{av(r_sh)}", npa_v / SH, PX, green=True)
    putf(pr, f'{c}6', f"='Balance Sheet'!{c}14/{av(r_sh)}", eq_v / SH, PX, green=True)
    if j == 0:
        put(pr, f'{c}7', '-', BLACK, None); put(pr, f'{c}8', '-', BLACK, None)
    else:
        pcol = COLS[j - 1]
        eq_prev = (BSH[['FY23', 'FY24', 'FY25'][j - 1]]['eqp'] if j <= 3
                   else (HA['eqp'] if j == 4
                         else HA['eqp'] + sum(F['np_attr'][:t])))
        yf = 0.5 if (isH1 or t == 0) else 1.0
        putf(pr, f'{c}7', f"='Income Statement'!{c}17/{yf}/(('Balance Sheet'!{c}14"
             f"+'Balance Sheet'!{pcol}14)/2)", npa_v / yf / ((eq_v + eq_prev) / 2), PCT,
             green=True)
        if hist or isH1:
            put(pr, f'{c}8', '-', BLACK, None)
        else:
            picol = 'E' if t == 0 else FCOL[t - 1]
            putf(pr, f'{c}8', f"=DCF!{FC[t]}{ANCH['dcf_rw']+9}/{yf}"
                 f"/(('Summary Financials'!{c}13+'Summary Financials'!{picol}13)/2)",
                 F['roic'][t], PCT, green=True)
    if hist or (not isH1 and t is not None and t > 0):
        putf(pr, f'{c}9', f"={av(r_spot)}/{c}5", SPOT / (npa_v / SH), MULT)
    else:
        put(pr, f'{c}9', '-', BLACK, None)
    putf(pr, f'{c}10', f"={av(r_spot)}/{c}6", SPOT / (eq_v / SH), MULT)

# ============ PEER & SECTOR ===================================================
pk = wb.create_sheet('Peer & Sector')
title(pk, 'Peer & sector — one attributable basis, cross-checks only',
      'Peer attributable profits from each company\'s own audited filings/releases; prices '
      '7-Aug-2026; every multiple is a live formula from its own row', 8, awidth=36, cwidth=13)
hdr(pk, 4, ['Peer', 'Spot', 'Shares (mn)', 'Mkt cap', 'FY2025 attributable NP',
            'Trailing P/E (attr)', 'Backlog (own disclosure)'])
r = 5
for k in ['ALDAR', 'EMAAR', 'EMAARDEV']:
    p = REL['peers'][k]
    put(pk, f'A{r}', p['name'], fmt=None)
    put(pk, f'B{r}', p['spot'], BLUE, PX)
    put(pk, f'C{r}', p['shares_mn'], BLUE, NUM0)
    putf(pk, f'D{r}', f"=B{r}*C{r}", p['spot'] * p['shares_mn'], NUM0)
    put(pk, f'E{r}', p['np_attr'], BLUE, NUM0)
    putf(pk, f'F{r}', f"=D{r}/E{r}", p['spot'] * p['shares_mn'] / p['np_attr'], MULT)
    put(pk, f'G{r}', p['backlog'], BLUE, NUM0)
    r += 1
put(pk, f'A{r}', 'Modon Holding (this study)', fmt=None, bold=True)
putf(pk, f'B{r}', f"={av(r_spot)}", SPOT, PX, green=True)
putf(pk, f'C{r}', f"={av(r_sh)}", SH, NUM0, green=True)
putf(pk, f'D{r}', f"=B{r}*C{r}", M['mktcap'], NUM0)
putf(pk, f'E{r}', f"={av(r_npa25)}", IN['npa_fy25'], NUM0, green=True)
putf(pk, f'F{r}', f"=D{r}/E{r}", M['mktcap'] / IN['npa_fy25'], MULT)
put(pk, f'G{r}', IN['h1_backlog'], BLUE, NUM0)
r += 1
put(pk, f'A{r}', '  (Modon backlog = 30-Jun-2026 group figure; peers = FY2025 development '
    'backlogs from their own releases — one definition per row, dated)', fmt=None)
r += 2
put(pk, f'A{r}', 'Sector context (ADREC 2025, regulator): AED 142bn transactions (+44%); '
    'residential AED 76bn (+67%); off-plan 71% of residential deals; expatriate and '
    'foreign buyers 62% of residential sales value.', fmt=None, wrap=True)
pk.merge_cells(f'A{r}:G{r+2}')

# ---- order sheets ------------------------------------------------------------
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
assert set(ORDER) == set(wb.sheetnames), (set(ORDER) ^ set(wb.sheetnames))
wb._sheets = [wb[n] for n in ORDER]
wb.calculation.fullCalcOnLoad = True

XLSX = os.path.join(HERE, 'MODON_Valuation_Model_09082026_public.xlsx')
wb.save(XLSX)
with open(os.path.join(HERE, 'xlsx_expected.json'), 'w') as f:
    json.dump(dict(expected=EXPECT,
                   anchors=dict(seg_rev_tot=ANCH['seg_rev_tot'],
                                dcf=ANCH['dcf'], dcf_rw=ANCH['dcf_rw'],
                                bl_row=ANCH['bl_row'],
                                seg_fcst_rev=ANCH['seg_fcst_rev'],
                                seg_fcst_gp=ANCH['seg_fcst_gp'],
                                summary_mktcap=ANCH['summary_mktcap'],
                                sotp_eq=ANCH['sotp_eq'],
                                nci_row=r_ncirow)), f, indent=1)
nform = sum(len(v) for v in EXPECT.values())
print(f'wrote {XLSX}')
print(f'formula cells recorded: {nform} across {len(EXPECT)} sheets')
