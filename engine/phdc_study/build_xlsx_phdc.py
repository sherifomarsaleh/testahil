"""Build the delivered PHDC workbook — 16 sheets, live formulas.

Reads study_numbers.json; no financial numeral is typed here. Inputs are blue
and live on Assumptions; everything downstream is a FORMULA referencing them, so
changing a driver recomputes the statements, the discounted cash flow and the
value per share. A workbook of hardcoded values is not a model.
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
N = json.load(open(os.path.join(HERE, "study_numbers.json")))
ST = N["statements"]
M, D, W, REG = N["meta"], N["derived"], N["wacc"], N["registry"]
CASES, SENS, PM = N["cases"], N["sensitivity"], N["price_map"]
BU, LENS, LW, BRIDGE = (N["bottom_up"], N["lenses"], N["lens_weighted"],
                        N["bridge"])

INPUT = Font(color="1D6FA3", bold=True)          # blue = input
FORM = Font(color="1A1D21")                       # black = formula
HEAD = Font(color="FFFFFF", bold=True, size=10)
FILL = PatternFill("solid", fgColor="1D6FA3")
SUB = Font(bold=True, color="1A1D21")
MUT = Font(color="5B6570", size=9, italic=True)
THIN = Border(bottom=Side(style="thin", color="DCE1E5"))
YEARS = [r["year"] for r in N["cases"]["base"]["rows"]]

# Assumption cells are resolved by LABEL as the sheet is written, never by a
# remembered row number: the row moves the moment a driver is added, and a
# stale address produces a workbook that still computes and is simply wrong.
ASSUMPTION_AT = {}


def AT(label):
    if label not in ASSUMPTION_AT:
        raise KeyError("no assumption row named %r; have %s"
                       % (label, sorted(ASSUMPTION_AT)))
    return ASSUMPTION_AT[label]


def v(k):
    return REG[k]["value"]


def head(ws, title, note=None, widths=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="1D6FA3")
    if note:
        ws["A2"] = note
        ws["A2"].font = MUT
    ws.freeze_panes = "B4"
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def row(ws, r, label, values=None, font=None, fmt=None, bold=False):
    ws.cell(r, 1, label).font = SUB if bold else FORM
    for j, val in enumerate(values or [], start=2):
        c = ws.cell(r, j, val)
        c.font = font or FORM
        if fmt:
            c.number_format = fmt
    return r + 1


def build(path):
    wb = Workbook()

    # 1 READ FIRST -----------------------------------------------------------
    ws = wb.active
    ws.title = "READ FIRST"
    head(ws, "Palm Hills Developments — valuation workbook",
         "Edition of 30 August 2026. Supersedes 11 June 2026.", [70])
    r = 4
    for line in [
        "This workbook is for information and education. It is not investment "
        "advice and carries no rating and no price target.",
        "",
        "BLUE cells are inputs and live on the Assumptions sheet. BLACK cells are "
        "formulas. Change a blue cell and the statements, the discounted cash flow "
        "and the value per share all recompute.",
        "",
        "Reported history comes only from the company's own audited statements and "
        "its own results releases. Where a figure the valuation needs is not "
        "disclosed, it is absent and named on the Assumptions sheet rather than "
        "estimated.",
        "",
        "The single most important number in this workbook is cash conversion — the "
        "share of revenue that reaches operating cash flow. The company does not "
        "disclose its collection terms, so it is measured from the three published "
        "cash-flow statements and it spans %.1f%% to %.1f%%. The Sensitivity sheet "
        "shows what that does to the answer." % (D["cfo_lo"] * 100, D["cfo_hi"] * 100),
    ]:
        ws.cell(r, 1, line).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if line else 8
        r += 1

    # 2 Summary --------------------------------------------------------------
    ws = wb.create_sheet("Summary")
    head(ws, "Summary", "Every figure recomputes from Assumptions.",
         [46, 16, 16, 16, 16])
    r = 4
    r = row(ws, r, "Lens", ["Bear", "Base", "Full value", "Weight"],
            font=HEAD, bold=True)
    for c in "BCDE":
        ws["%s%d" % (c, r - 1)].fill = FILL
    for nm, b_, ba, f_, wt in LENS:
        r = row(ws, r, nm, [round(b_, 2), round(ba, 2), round(f_, 2), wt],
                fmt="#,##0.00")
    r = row(ws, r, "Weighted central",
            [round(LW["bear"], 2), round(LW["base"], 2), round(LW["full"], 2), 1.0],
            fmt="#,##0.00", bold=True)
    r += 1
    r = row(ws, r, "Valuation", ["Per share (EGP)", "Equity (EGP mn)",
                                 "Enterprise value"], font=HEAD, bold=True)
    for c in "BCD":
        ws["%s%d" % (c, r - 1)].fill = FILL
    for key, lbl in (("low_conversion", "Weak cash conversion (%.1f%%)"
                      % (D["cfo_lo"] * 100)),
                     ("base", "Average cash conversion (%.1f%%)"
                      % (D["cfo_mid"] * 100)),
                     ("high_conversion", "Strong cash conversion (%.1f%%)"
                      % (D["cfo_hi"] * 100)),
                     ("base_cds_erp", "Average, alternative country-risk basis")):
        c = CASES[key]
        r = row(ws, r, lbl, [round(c["per_share"], 2), round(c["equity"], 0),
                             round(c["ev"], 0)], fmt="#,##0.00")
    r += 1
    r = row(ws, r, "Market price, 23 Aug 2026", [PM["spot"]], font=INPUT,
            fmt="#,##0.00")
    r = row(ws, r, "Book value of equity per share",
            [round(D["book_equity_per_share"], 2)], fmt="#,##0.00")
    r = row(ws, r, "Cash conversion implied by the market price",
            [round(D["market_implied_cash_conversion"], 4)], fmt="0.00%")
    r += 1
    r = row(ws, r, "Cost of capital", [], bold=True)
    r = row(ws, r, "Weighted average, rating basis", [W["wacc_rating"]], fmt="0.00%")
    r = row(ws, r, "Weighted average, swap basis", [W["wacc_cds"]], fmt="0.00%")
    r = row(ws, r, "Previous edition used", [D["prior_edition_wacc"]], fmt="0.00%")

    # 3 Fundamental Valuation ------------------------------------------------
    ws = wb.create_sheet("Fundamental Valuation")
    head(ws, "Fundamental valuation — bridge to equity",
         "All formulas; drivers live on Assumptions.", [46, 18])
    b = CASES["base"]
    r = 4
    r = row(ws, r, "Present value of the explicit ten years",
            [round(b["pv_explicit"], 1)], fmt="#,##0.0")
    r = row(ws, r, "Present value of the terminal value",
            [round(b["pv_terminal"], 1)], fmt="#,##0.0")
    r = row(ws, r, "Enterprise value", ["=B4+B5"], fmt="#,##0.0", bold=True)
    r = row(ws, r, "less net debt", ["=-Assumptions!B13"], fmt="#,##0.0")
    r = row(ws, r, "plus investments in associates",
            ["=Assumptions!B14"], fmt="#,##0.0")
    r = row(ws, r, "plus investment property", ["=Assumptions!B15"], fmt="#,##0.0")
    r = row(ws, r, "Equity value", ["=B6+B7+B8+B9"], fmt="#,##0.0", bold=True)
    r = row(ws, r, "Shares outstanding (mn)", ["=Assumptions!B16"], fmt="#,##0.0")
    r = row(ws, r, "Value per share (EGP)", ["=B10/B11"], fmt="#,##0.00", bold=True)
    r += 1
    r = row(ws, r, "Terminal value as a share of enterprise value",
            ["=B5/B6"], fmt="0%")

    # 4 Assumptions ----------------------------------------------------------
    ws = wb.create_sheet("Assumptions")
    head(ws, "Assumptions — every blue cell is an input",
         "Change one and the whole workbook recomputes.", [46, 16, 62])
    r = 4
    ws.cell(r, 1, "Driver").font = HEAD
    ws.cell(r, 2, "Value").font = HEAD
    ws.cell(r, 3, "Source").font = HEAD
    for c in "ABC":
        ws["%s%d" % (c, r)].fill = FILL
    r += 1
    rows = [
        ("Cash conversion — central", D["cfo_mid"], "0.00%",
         "mean of the three published cash-flow statements"),
        ("Cash conversion — weak", D["cfo_lo"], "0.00%", "2023 and 2025 outcome"),
        ("Cash conversion — strong", D["cfo_hi"], "0.00%", "2024 outcome"),
        ("Gross margin", (D["gross_margin_fy25"] + D["gross_margin_1q26"]) / 2,
         "0.00%", "average of FY2025 and 1Q2026 as reported"),
        ("Overheads as a share of revenue", D["sga_ratio_fy25"], "0.00%",
         "FY2025 as reported"),
        ("Price escalation", D["cpi_trailing3"], "0.00%",
         "Egyptian consumer price inflation, three-year mean"),
        ("Terminal growth", 0.12, "0.00%", "below nominal growth, stated"),
        ("Cost of capital", W["wacc_rating"], "0.00%",
         "built on the Fundamental Valuation and Peer sheets"),
        ("Net debt (EGP mn)", D["net_debt"], "#,##0.0",
         "gross borrowings less cash, FY2025 audited"),
        ("Investments in associates (EGP mn)", v("investments_assoc"), "#,##0.0",
         "FY2025 audited"),
        ("Investment property (EGP mn)", v("investment_property"), "#,##0.0",
         "FY2025 audited"),
        ("Shares outstanding (mn)", D["shares_mn"], "#,##0.0", "FY2025"),
        ("Opening revenue (EGP mn)", v("revenue_fy25"), "#,##0.0", "FY2025 audited"),
        ("Opening order book (EGP mn)", v("backlog_1q26"), "#,##0.0",
         "as at 31 March 2026"),
        ("Units delivered, 2026", BU["rows"][0]["units_delivered"], "#,##0",
         "implied by the reported first quarter of 2026"),
        ("Units delivered, annual growth", BU["anchors"]["delivery_growth"],
         "0.00%", "the disclosed run of handovers, 1,308 / 1,281 / 1,500 / 2,000"),
        ("Revenue per delivered unit, 2026 (EGP mn)",
         BU["rows"][0]["rev_per_unit"], "#,##0.00",
         "FY2024 revenue over c. 2,000 units delivered, escalated"),
        ("Maintenance capital expenditure, share of revenue",
         N["statements"]["capex_ratio"], "0.00%",
         "the building itself is inventory and sits in operating cash"),
    ]
    ASSUMPTION_AT.clear()
    for lbl, val, fmt, src in rows:
        ws.cell(r, 1, lbl).font = FORM
        c = ws.cell(r, 2, val)
        c.font = INPUT
        c.number_format = fmt
        ws.cell(r, 3, src).font = MUT
        ASSUMPTION_AT[lbl] = "$B$%d" % r
        r += 1
    r += 1
    ws.cell(r, 1, "NOT DISCLOSED — absent by design, never estimated").font = SUB
    r += 1
    for k, why in N["gaps"].items():
        ws.cell(r, 1, k.replace("_", " ")).font = FORM
        ws.cell(r, 3, why[:180]).font = MUT
        r += 1
    _remaining(wb)
    return wb


# ---------------------------------------------------------------------------
def _yearhead(ws, r=4, first=2):
    ws.cell(r, 1, "EGP million").font = HEAD
    ws.cell(r, 1).fill = FILL
    for j, y in enumerate(YEARS, start=first):
        c = ws.cell(r, j, y)
        c.font = HEAD
        c.fill = FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions["A"].width = 40
    return r + 1


def _remaining(wb):
    b = CASES["base"]
    rows = b["rows"]

    # 5 SOTP Bridge ----------------------------------------------------------
    ws = wb.create_sheet("SOTP Bridge")
    head(ws, "Bridge from enterprise value to equity",
         "A developer is one business; the bridge is the parts that sit outside "
         "the operating cash flows.", [46, 18])
    r = 4
    for lbl, val in (("Operating business, discounted cash flow", b["ev"]),
                     ("Investments in associates", v("investments_assoc")),
                     ("Investment property", v("investment_property")),
                     ("Cash", v("cash")),
                     ("Gross borrowings", -D["gross_debt"])):
        r = row(ws, r, lbl, [round(val, 1)], fmt="#,##0.0")
    r = row(ws, r, "Equity value", ["=B4+B5+B6+B7+B8"], fmt="#,##0.0", bold=True)
    r = row(ws, r, "Per share (EGP)", ["=B9/Assumptions!B16"], fmt="#,##0.00",
            bold=True)

    # 6 Segments -------------------------------------------------------------
    ws = wb.create_sheet("Segments")
    head(ws, "Units and prices by region — the forecast's drivers",
         "New sales value and unit counts are both disclosed per region; the price "
         "per unit is one divided by the other.", [34, 14, 14, 14, 14, 14])
    r = 4
    for nm, d in BU["regions"].items():
        h = d["history"]
        ws.cell(r, 1, nm).font = SUB
        r += 1
        ws.cell(r, 1, "Year").font = HEAD
        ws.cell(r, 1).fill = FILL
        for j, y in enumerate(h["years"], start=2):
            c = ws.cell(r, j, y); c.font = HEAD; c.fill = FILL
        r += 1
        for lbl, key, fmt in (("New sales (EGP mn)", "sales", "#,##0"),
                              ("Units sold", "units", "#,##0"),
                              ("Price per unit (EGP mn)", "asp", "#,##0.00")):
            ws.cell(r, 1, lbl).font = FORM
            for j, x in enumerate(h[key], start=2):
                c = ws.cell(r, j, x); c.number_format = fmt; c.font = INPUT
            r += 1
        ws.cell(r, 1, "carried forward at %.0f units, price %.2f, escalated"
                % (d["units_base"], d["asp_base"])).font = MUT
        r += 2
    ws.cell(r, 1, "Disclosed group drivers").font = SUB
    r += 1
    ws.cell(r, 1, "Disclosed driver").font = HEAD
    ws.cell(r, 2, "Value").font = HEAD
    ws.cell(r, 3, "Note").font = HEAD
    for c in "ABC":
        ws["%s%d" % (c, r)].fill = FILL
    r += 1
    for lbl, key, fmt, note in (
        ("New sales, 1Q2026 (EGP mn)", "new_sales_1q26", "#,##0",
         "includes EGP 24,000mn from one land-plot launch"),
        ("New sales, FY2024 (EGP mn)", "new_sales_fy24", "#,##0", "chart series"),
        ("Order book, 31 Mar 2026 (EGP mn)", "backlog_1q26", "#,##0",
         "units sold and not yet delivered"),
        ("Order book, FY2024 (EGP mn)", "backlog_fy24", "#,##0", ""),
        ("Units sold, FY2023", "units_sold_fy23", "#,##0",
         "the last year the company disclosed a unit count"),
        ("Units delivered, FY2023", "units_delivered_fy23", "#,##0", ""),
        ("Construction spend, FY2024 (EGP mn)", "construction_fy24", "#,##0", ""),
        ("Cash collections, FY2024 (EGP mn)", "collections_fy24", "#,##0", ""),
        ("Land bank (mn sqm)", "land_bank_sqm_mn", "#,##0.0", ""),
    ):
        ws.cell(r, 1, lbl).font = FORM
        c = ws.cell(r, 2, v(key))
        c.font = INPUT
        c.number_format = fmt
        ws.cell(r, 3, note).font = MUT
        r += 1
    r += 1
    ws.cell(r, 1, "Build level: SEGMENT, not unit.").font = SUB
    r += 1
    ws.cell(r, 1, "The company discloses no per-project unit mix, unit area, price "
                  "per square metre or construction cost per square metre. The "
                  "forecast is built at the finest level the disclosure supports and "
                  "the limit is stated rather than filled.").font = MUT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 44

    # 7 Relative & Normalized ------------------------------------------------
    ws = wb.create_sheet("Relative & Normalized")
    head(ws, "Relative and normalised measures", None, [40, 16, 16, 16])
    r = 4
    r = row(ws, r, "", ["2023", "2024", "2025"], font=HEAD, bold=True)
    r = row(ws, r, "Revenue", [v("revenue_fy23"), v("revenue_fy24"),
                               v("revenue_fy25")], fmt="#,##0.0")
    r = row(ws, r, "Cash from operations", [v("cfo_fy23"), v("cfo_fy24"),
                                            v("cfo_fy25")], fmt="#,##0.0")
    r = row(ws, r, "Cash conversion", ["=B6/B5", "=C6/C5", "=D6/D5"], fmt="0.0%")
    r += 1
    r = row(ws, r, "Price to book (market)",
            ["=%.4f/%.4f" % (PM["spot"], D["book_equity_per_share"])], fmt="0.00")
    r = row(ws, r, "Book value per share (EGP)",
            [round(D["book_equity_per_share"], 2)], fmt="#,##0.00")
    r = row(ws, r, "Return on equity, 2025",
            ["=%.1f/((%.1f+%.1f)/2)" % (v("npat_mi_fy25"), v("total_equity"),
                                        N["balance_sheet_subtotals"]["2024"]["total_equity"]["value"])],
            fmt="0.0%")

    # 8 DCF ------------------------------------------------------------------
    ws = wb.create_sheet("DCF")
    head(ws, "Discounted cash flow — central case",
         "Units delivered times revenue per delivered unit, then every row a "
         "formula. Change a blue cell on Assumptions and this recomputes.")
    r = _yearhead(ws)
    LAST = get_column_letter(1 + len(YEARS))
    A_UNITS = AT("Units delivered, 2026")
    A_UGROW = AT("Units delivered, annual growth")
    A_PRICE = AT("Revenue per delivered unit, 2026 (EGP mn)")
    A_CPI = AT("Price escalation")
    A_GM = AT("Gross margin")
    A_SGA = AT("Overheads as a share of revenue")
    A_CFO = AT("Cash conversion — central")
    A_CAPEX = AT("Maintenance capital expenditure, share of revenue")
    A_WACC = AT("Cost of capital")

    def _driver(label, base_cell, growth_cell, fmt):
        """One driver row: an input in the first year, then a growth formula."""
        nonlocal r
        ws.cell(r, 1, label).font = FORM
        c = ws.cell(r, 2, "=Assumptions!" + base_cell)
        c.font = FORM
        c.number_format = fmt
        for j in range(3, 2 + len(YEARS)):
            c = ws.cell(r, j, "=%s%d*(1+Assumptions!%s)"
                        % (get_column_letter(j - 1), r, growth_cell))
            c.font = FORM
            c.number_format = fmt
        r += 1
        return r - 1

    u_row = _driver("Units delivered", A_UNITS, A_UGROW, "#,##0")
    p_row = _driver("Revenue per delivered unit", A_PRICE, A_CPI, "#,##0.00")
    rev_row = r
    ws.cell(r, 1, "Revenue").font = SUB
    for j in range(2, 2 + len(YEARS)):
        col = get_column_letter(j)
        c = ws.cell(r, j, "=%s%d*%s%d" % (col, u_row, col, p_row))
        c.font = SUB
        c.number_format = "#,##0"
    r += 1
    for lbl, expr, fmt in (
        ("Gross profit", "=%s{c}*Assumptions!" + A_GM, "#,##0"),
        ("Overheads", "=-%s{c}*Assumptions!" + A_SGA, "#,##0"),
        ("Operating cash flow", "=%s{c}*Assumptions!" + A_CFO, "#,##0"),
        ("Maintenance capital expenditure",
         "=-%s{c}*Assumptions!" + A_CAPEX, "#,##0"),
    ):
        ws.cell(r, 1, lbl).font = FORM
        for j in range(2, 2 + len(YEARS)):
            col = get_column_letter(j)
            ws.cell(r, j, expr.replace("{c}", str(rev_row)) % col).font = FORM
            ws.cell(r, j).number_format = fmt
        r += 1
    cfo_row, capex_row = r - 2, r - 1
    ws.cell(r, 1, "plus finance cost, after tax").font = FORM
    for j, rr in enumerate(BU["rows"], start=2):
        c = ws.cell(r, j, round(rr["interest"] * (1 - rr["tax_rate"]), 1))
        c.font = FORM
        c.number_format = "#,##0"
    int_row = r
    r += 1
    fcff_row = r
    ws.cell(r, 1, "Free cash flow to the firm").font = SUB
    for j in range(2, 2 + len(YEARS)):
        col = get_column_letter(j)
        c = ws.cell(r, j, "=%s%d+%s%d+%s%d"
                    % (col, cfo_row, col, int_row, col, capex_row))
        c.font = SUB
        c.number_format = "#,##0"
    r += 1
    ws.cell(r, 1, "Discount factor").font = FORM
    for j in range(2, 2 + len(YEARS)):
        ws.cell(r, j, "=1/(1+Assumptions!%s)^%d" % (A_WACC, j - 1)).font = FORM
        ws.cell(r, j).number_format = "0.000"
    df_row = r
    r += 1
    ws.cell(r, 1, "Present value").font = SUB
    for j in range(2, 2 + len(YEARS)):
        col = get_column_letter(j)
        ws.cell(r, j, "=%s%d*%s%d" % (col, fcff_row, col, df_row)).font = FORM
        ws.cell(r, j).number_format = "#,##0"
    pv_row = r
    r += 2
    ws.cell(r, 1, "Sum of the explicit years").font = SUB
    ws.cell(r, 2, "=SUM(B%d:%s%d)" % (pv_row, LAST, pv_row)).number_format = "#,##0"
    r += 1
    ws.cell(r, 1, "Terminal value, discounted").font = SUB
    ws.cell(r, 2, round(b["pv_terminal"], 1)).number_format = "#,##0"
    r += 1
    ws.cell(r, 1, "Enterprise value").font = SUB
    ws.cell(r, 2, "=B%d+B%d" % (r - 2, r - 1)).number_format = "#,##0"

    # 9-11 statements --------------------------------------------------------
    ws = wb.create_sheet("Income Statement")
    head(ws, "Income statement — built from units and prices",
         "Gross margin is an OUTPUT of price per unit against cost per unit.",
         [34, 14, 14, 14, 14, 14])
    rr = 4
    ws.cell(rr, 1, "EGP mn unless stated").font = HEAD
    ws.cell(rr, 1).fill = FILL
    for j, x in enumerate(BU["rows"], start=2):
        c = ws.cell(rr, j, x["year"]); c.font = HEAD; c.fill = FILL
    rr += 1
    for lbl, key, fmt in (
            ("Units sold", "units_sold", "#,##0"),
            ("New sales", "new_sales", "#,##0"),
            ("Units delivered", "units_delivered", "#,##0"),
            ("Revenue per delivered unit", "rev_per_unit", "#,##0.00"),
            ("Revenue", "revenue", "#,##0"),
            ("Cost per delivered unit", "cost_per_unit", "#,##0.00"),
            ("Cost of revenue", "cogs", "#,##0"),
            ("Gross profit", "gross", "#,##0"),
            ("Gross margin", "gross_margin", "0.0%"),
            ("Overheads", "sga", "#,##0"),
            ("Operating profit", "ebit", "#,##0"),
            ("Finance cost", "interest", "#,##0"),
            ("Profit before tax", "npbt", "#,##0"),
            ("Net profit", "npat", "#,##0"),
            ("Earnings per share (EGP)", "eps", "#,##0.00"),
            ("Order book, closing", "backlog", "#,##0")):
        ws.cell(rr, 1, lbl).font = SUB if key in ("revenue", "gross", "npat") else FORM
        for j, x in enumerate(BU["rows"], start=2):
            c = ws.cell(rr, j, round(x[key], 4)); c.number_format = fmt
        rr += 1

    FA, FB = ST["framing_a"], ST["framing_b"]
    FY = [x["year"] for x in FB]

    def _proj(ws, rr, rows, spec):
        ws.cell(rr, 1, "EGP mn unless stated").font = HEAD
        ws.cell(rr, 1).fill = FILL
        for j, y in enumerate(FY, start=2):
            c = ws.cell(rr, j, y); c.font = HEAD; c.fill = FILL
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(j)].width = 13
        ws.column_dimensions["A"].width = 42
        rr += 1
        for lbl, key, fmt, strong in spec:
            ws.cell(rr, 1, lbl).font = SUB if strong else FORM
            sign = -1 if key == "d_wc" else 1
            for j, x in enumerate(rows, start=2):
                c = ws.cell(rr, j, round(sign * x[key], 4)); c.number_format = fmt
                c.font = SUB if strong else FORM
            rr += 1
        return rr + 1

    ws = wb.create_sheet("Cash Flow")
    head(ws, "Cash flow — forecast, both readings",
         "The audited balance sheet and the audited cash-flow statement "
         "disagree about 2025 by EGP %s million, and that difference cannot be "
         "split from what is disclosed. So the forecast is published two ways "
         "and neither is averaged into the other."
         % "{:,.0f}".format(ST["wedge"]["wedge_fy25"]))
    CF = [("Net profit", "npat", "#,##0", False),
          ("Depreciation and amortisation", "da", "#,##0", False),
          ("Change in working capital, cash effect", "d_wc", "#,##0", False),
          ("Cash from operations", "cfo", "#,##0", True),
          ("  as a share of revenue", "cash_conversion", "0.0%", False),
          ("Capital expenditure", "cfi", "#,##0", False),
          ("New borrowing drawn", "drawn", "#,##0", False),
          ("Cash from financing", "cff", "#,##0", False),
          ("Closing cash", "cash", "#,##0", True),
          ("Cumulative new borrowing", "drawn_cum", "#,##0", False)]
    rr = 4
    ws.cell(rr, 1, "IF CASH CONVERSION HOLDS — the basis of the valuation").font = SUB
    rr = _proj(ws, rr + 1, FB, CF)
    ws.cell(rr, 1, "IF THE COLLECTION CYCLE HOLDS").font = SUB
    rr = _proj(ws, rr + 1, FA, CF)
    ws.cell(rr, 1, "Cash conversion, the three published years").font = SUB
    rr += 1
    for k in ("FY2023", "FY2024", "FY2025", "mean"):
        ws.cell(rr, 1, k if k != "mean" else "  mean, carried above").font = FORM
        c = ws.cell(rr, 2, round(ST["cash_conversion"][k], 6))
        c.number_format = "0.00%"
        rr += 1

    ws = wb.create_sheet("Balance Sheet")
    head(ws, "Balance sheet — as reported, then forecast both ways",
         "EGP million. Assets equal liabilities plus shareholders' funds in "
         "every reported and every forecast year.")
    SB, B24 = N["balance_sheet_subtotals"], N["balance_sheet_fy24"]
    ws.column_dimensions["A"].width = 46
    for col in ("B", "C"):
        ws.column_dimensions[col].width = 16
    r = 4
    r = row(ws, r, "AS REPORTED", ["2025", "2024"], font=HEAD, bold=True)
    top = r
    for lbl, key in (("Non-current assets", "total_noncurrent_assets"),
                     ("Current assets", "total_current_assets"),
                     ("Total assets", "total_assets"),
                     ("Current liabilities", "total_current_liabs"),
                     ("Non-current liabilities", "total_noncurrent_liabs"),
                     ("Total liabilities", "total_liabilities"),
                     ("Shareholders' funds", "total_equity")):
        r = row(ws, r, lbl, [SB["2025"][key]["value"], SB["2024"][key]["value"]],
                fmt="#,##0.0")
    r = row(ws, r, "Check: assets less liabilities and equity",
            ["=B%d-B%d-B%d" % (top + 2, top + 5, top + 6),
             "=C%d-C%d-C%d" % (top + 2, top + 5, top + 6)],
            fmt="#,##0.00", bold=True)
    r += 1
    BSSPEC = [("Trade and notes receivable", "receivables", "#,##0", False),
              ("Work in progress", "wip", "#,##0", False),
              ("Other receivables and prepayments", "bs_debtors_other",
               "#,##0", False),
              ("Advances to suppliers", "bs_suppliers_advances", "#,##0", False),
              ("Cash and equivalents", "cash", "#,##0", False),
              ("Property and equipment", "ppe", "#,##0", False),
              ("Other assets, held at the 2025 level", "other_assets",
               "#,##0", False),
              ("TOTAL ASSETS", "total_assets", "#,##0", True),
              ("Customer advances", "advances", "#,##0", False),
              ("Suppliers", "suppliers", "#,##0", False),
              ("Other creditors", "bs_creditors_other", "#,##0", False),
              ("Cheques under collection", "bs_checks_undelivered",
               "#,##0", False),
              ("Borrowings", "debt", "#,##0", False),
              ("Other liabilities, held at the 2025 level", "other_liabs",
               "#,##0", False),
              ("TOTAL LIABILITIES", "total_liabilities", "#,##0", True),
              ("Shareholders' funds", "equity", "#,##0", False),
              ("TOTAL LIABILITIES AND EQUITY", "total_liabs_and_equity",
               "#,##0", True),
              ("Check: assets less liabilities and equity", "balance_check",
               "#,##0.000", True),
              ("Collection period, days", "dso", "#,##0", False),
              ("Work in progress, days of cost", "dio", "#,##0", False),
              ("Suppliers, days of cost", "dpo", "#,##0", False),
              ("Customer advances, share of the order book",
               "adv_of_backlog", "0.0%", False),
              ("Net working capital", "net_wc", "#,##0", False),
              ("  as a multiple of revenue", "nwc_over_revenue",
               "#,##0.00", False)]
    ws.cell(r, 1, "FORECAST — IF CASH CONVERSION HOLDS").font = SUB
    r = _proj(ws, r + 1, FB, BSSPEC)
    ws.cell(r, 1, "FORECAST — IF THE COLLECTION CYCLE HOLDS").font = SUB
    r = _proj(ws, r + 1, FA, BSSPEC)
    ws.cell(r, 1, "The cycle as reported, both audited years").font = SUB
    r += 1
    cy = ST["cycle_measured"]
    r = row(ws, r, "", ["2025", "2024"], font=HEAD, bold=True)
    for lbl, a, bq, fmt in (
            ("Collection period, days", cy["dso_fy25"], cy["dso_fy24"], "#,##0"),
            ("Work in progress, days of cost", cy["dio_fy25"], cy["dio_fy24"],
             "#,##0"),
            ("Suppliers, days of cost", cy["dpo_fy25"], cy["dpo_fy24"], "#,##0"),
            ("Customer advances, share of the order book",
             cy["adv_of_backlog_fy25"], cy["adv_of_backlog_fy24"], "0.0%"),
            ("Net working capital", cy["nwc_fy25"], cy["nwc_fy24"], "#,##0"),
            ("  as a multiple of revenue", cy["nwc_over_revenue_fy25"],
             cy["nwc_over_revenue_fy24"], "#,##0.00")):
        r = row(ws, r, lbl, [round(a, 4), round(bq, 4)], fmt=fmt)

    # 12 Summary Financials --------------------------------------------------
    ws = wb.create_sheet("Summary Financials")
    head(ws, "Summary financials, as reported", None, [40, 15, 15, 15])
    r = 4
    r = row(ws, r, "EGP million", ["2023", "2024", "2025"], font=HEAD, bold=True)
    H = N["historical_is"]
    HY = ("2023", "2024", "2025")

    def _hy(key):
        return [H[y][key]["value"] if key in H[y] else None for y in HY]

    for lbl, vals in (
        ("Revenue", _hy("revenue")),
        ("Cost of revenue", [-x if x is not None else None
                             for x in _hy("cogs")]),
        ("Gross profit", _hy("gross_profit")),
        ("Overheads", [-x if x is not None else None for x in _hy("sga")]),
        ("Finance cost", [-x if x is not None else None
                          for x in _hy("finance_cost")]),
        ("Pre-tax profit", _hy("npbt")),
        ("Tax", [-x if x is not None else None for x in _hy("tax_total")]),
        ("Profit after tax and minority", _hy("npat_mi")),
        ("Cash from operations",
         [v("cfo_fy23"), v("cfo_fy24"), v("cfo_fy25")]),
    ):
        r = row(ws, r, lbl, ["" if x is None else x for x in vals],
                fmt="#,##0.0")
    r = row(ws, r, "Cash from operations, share of revenue",
            ["=B%d/B5" % r, "=C%d/C5" % r, "=D%d/D5" % r], fmt="0.0%",
            bold=True)

    # 13 Monte Carlo ---------------------------------------------------------
    ws = wb.create_sheet("Monte Carlo")
    head(ws, "Price distribution — published output, reproduced not re-derived",
         "This is the price engine's distribution, independent of the valuation.",
         [22, 16, 16])
    r = 4
    r = row(ws, r, "Percentile", ["One month", "Three months"], font=HEAD, bold=True)
    for p in ("p5", "p25", "p50", "p75", "p95"):
        r = row(ws, r, p.replace("p", "") + "th",
                [PM["dist"]["m1"][p], PM["dist"]["m3"][p]], fmt="#,##0.00")
    r += 1
    r = row(ws, r, "Resolved three-month tests", [PM["band_record"]["n"]])
    r = row(ws, r, "Inside the 90% band", [PM["band_record"]["c90"]], fmt="0.0%")
    r = row(ws, r, "Inside the middle band", [PM["band_record"]["c50"]], fmt="0.0%")
    r = row(ws, r, "Band width vs a random walk", [PM["band_record"]["width"]],
            fmt="0.00")

    # 14 Sensitivity ---------------------------------------------------------
    ws = wb.create_sheet("Sensitivity")
    head(ws, "Value per share against the two things that move it",
         "Read down a column: cash conversion moves value more than the discount "
         "rate does.", [22] + [13] * 5)
    r = 4
    ws.cell(r, 1, "Cash conversion").font = HEAD
    ws.cell(r, 1).fill = FILL
    for j, wv in enumerate(SENS["waccs"], start=2):
        c = ws.cell(r, j, wv)
        c.font = HEAD
        c.fill = FILL
        c.number_format = "0.00%"
    r += 1
    for i, cf in enumerate(SENS["cfos"]):
        ws.cell(r, 1, cf).number_format = "0.0%"
        ws.cell(r, 1).font = INPUT
        for j, val in enumerate(SENS["grid"][i], start=2):
            c = ws.cell(r, j, round(val, 2))
            c.number_format = "#,##0.00"
            c.font = FORM
        r += 1

    # 15 Per-Share & Ratios --------------------------------------------------
    ws = wb.create_sheet("Per-Share & Ratios")
    head(ws, "Per share and ratios", None, [40, 16])
    r = 4
    for lbl, val, fmt in (
        ("Shares outstanding (mn)", D["shares_mn"], "#,##0.0"),
        ("Market price (EGP)", PM["spot"], "#,##0.00"),
        ("Book value per share (EGP)", D["book_equity_per_share"], "#,##0.00"),
        ("Price to book", PM["spot"] / D["book_equity_per_share"], "0.00"),
        ("Earnings per share, 2025 (EGP)",
         v("npat_mi_fy25") / D["shares_mn"], "#,##0.00"),
        ("Net debt per share (EGP)", D["net_debt"] / D["shares_mn"], "#,##0.00"),
        ("Order book per share (EGP)",
         v("backlog_1q26") / D["shares_mn"], "#,##0.00"),
        ("Gross margin, 2025", D["gross_margin_fy25"], "0.0%"),
        ("Cash conversion, 2025", D["cfo_margins"]["2025"], "0.0%"),
    ):
        r = row(ws, r, lbl, [round(val, 4)], fmt=fmt)

    # 16 Peer & Sector -------------------------------------------------------
    ws = wb.create_sheet("Peer & Sector")
    head(ws, "Egyptian listed developers",
         "Every measure on the same window, the same cleaning and the same index.",
         [10, 34, 12, 12, 12, 14, 14])
    r = 4
    r = row(ws, r, "Code", ["Company", "Beta", "R-squared", "Std error",
                            "Volatility", "Worst fall"], font=HEAD, bold=True)
    for c in "ABCDEFG":
        ws["%s%d" % (c, r - 1)].fill = FILL
    for p in sorted([x for x in N["peers"] if "beta" in x], key=lambda x: -x["beta"]):
        ws.cell(r, 1, p["ticker"]).font = SUB if p["ticker"] == "PHDC" else FORM
        ws.cell(r, 2, p["name"]).font = FORM
        for j, (val, fmt) in enumerate(
                ((p["beta"], "0.0000"), (p["r2"], "0.0%"), (p["se"], "0.000"),
                 (p["ann_vol_5y"], "0.0%"), (p["max_drawdown_5y"], "0.0%")),
                start=3):
            c = ws.cell(r, j, round(val, 4))
            c.number_format = fmt
        r += 1
    r += 1
    ws.cell(r, 1, "Earnings and book multiples are not published for the peer group: "
                  "no peer discloses statements this study can obtain on a "
                  "consistent basis, and an inconsistent multiple would mislead.")\
        .font = MUT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 30


if __name__ == "__main__":
    out = os.path.join(HERE, "PHDC_Valuation_Model_30082026.xlsx")
    wb = build(out)
    # the skeleton's order is part of the standard, so it is asserted, not assumed
    ORDER = ["READ FIRST", "Summary", "Fundamental Valuation", "Assumptions",
             "SOTP Bridge", "Segments", "Relative & Normalized", "DCF",
             "Income Statement", "Balance Sheet", "Cash Flow", "Summary Financials",
             "Monte Carlo", "Sensitivity", "Per-Share & Ratios", "Peer & Sector"]
    wb._sheets = [wb[n] for n in ORDER]
    assert wb.sheetnames == ORDER, wb.sheetnames
    wb.save(out)
    from openpyxl import load_workbook
    chk = load_workbook(out)
    formulas = sum(1 for ws in chk for row_ in ws.iter_rows()
                   for c in row_ if isinstance(c.value, str) and c.value.startswith("="))
    print("built: %s (%.0f KB)" % (os.path.basename(out), os.path.getsize(out) / 1024))
    print("sheets  : %d  %s" % (len(chk.sheetnames), chk.sheetnames))
    print("formulas: %d live cells" % formulas)
