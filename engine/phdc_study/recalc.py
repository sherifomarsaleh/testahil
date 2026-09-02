"""Independently recalculate the delivered workbook and compare it to the study.

Numeric traceability: a clean-looking workbook is not evidence. This evaluates
the formula cells that matter from the workbook's OWN inputs, without touching
the Python model, and checks the answers against study_numbers.json. Anything
that will not parse is reported as a FAILURE, never skipped.
"""
import json, os, re, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

HERE = os.path.dirname(os.path.abspath(__file__))
N = json.load(open(os.path.join(HERE, "study_numbers.json")))
XL = os.path.join(HERE, "PHDC_Valuation_Model_02092026.xlsx")


def evaluate(wb, sheet, ref, depth=0):
    """Resolve one cell, following formula references. Raises on anything it
    cannot parse — an unparseable cell is a failure, not a skip."""
    if depth > 24:
        raise ValueError("reference cycle at %s!%s" % (sheet, ref))
    val = wb[sheet][ref].value
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if not (isinstance(val, str) and val.startswith("=")):
        raise ValueError("cell %s!%s is neither number nor formula: %r"
                         % (sheet, ref, val))
    expr = val[1:]

    # SUM over a contiguous range is resolved, not skipped: the rule is that an
    # unparseable cell is a FAILURE, and the honest way to satisfy it is to
    # parse the thing rather than to keep raising on a function the workbook
    # legitimately uses.
    def _sum(m):
        sh = (m.group(1) or m.group(2) or sheet).strip()
        c1, r1, c2, r2 = m.group(3), int(m.group(4)), m.group(5), int(m.group(6))
        a, b = column_index_from_string(c1), column_index_from_string(c2)
        total = 0.0
        for cc in range(min(a, b), max(a, b) + 1):
            for rr in range(min(r1, r2), max(r1, r2) + 1):
                total += evaluate(wb, sh, "%s%d" % (get_column_letter(cc), rr),
                                  depth + 1)
        return repr(total)

    expr = re.sub(
        r"SUM\(\s*(?:(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_\-]*))!)?"
        r"\$?([A-Z]+)\$?(\d+)\s*:\s*\$?([A-Z]+)\$?(\d+)\s*\)",
        _sum, expr, flags=re.I)
    tok = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_\-]*))!\$?([A-Z]+)\$?(\d+)"
                     r"|\$?([A-Z]+)\$?(\d+)")

    def sub(m):
        if m.group(3):
            sh = (m.group(1) or m.group(2)).strip()
            return repr(evaluate(wb, sh, "%s%s" % (m.group(3), m.group(4)), depth + 1))
        return repr(evaluate(wb, sheet, "%s%s" % (m.group(5), m.group(6)), depth + 1))

    py = tok.sub(sub, expr).replace("^", "**")
    if re.search(r"SUM\(", py, re.I):
        raise ValueError("SUM not resolved at %s!%s: %s" % (sheet, ref, expr))
    return float(eval(py, {"__builtins__": {}}, {}))


def main():
    wb = load_workbook(XL)
    checks, fails = [], 0

    def chk(name, got, want, tol):
        nonlocal fails
        ok = abs(got - want) <= tol
        checks.append({"check": name, "workbook": round(got, 4),
                       "study": round(want, 4), "ok": ok})
        if not ok:
            fails += 1

    base = N["cases"]["base"]
    D = N["derived"]

    ev = evaluate(wb, "Fundamental Valuation", "B6")
    chk("enterprise value", ev, base["ev"], 1.0)
    eqg = evaluate(wb, "Fundamental Valuation", "B10")
    chk("equity value before minority interests", eqg, base["equity_before_nci"], 1.0)
    nci = evaluate(wb, "Fundamental Valuation", "B11")
    chk("minority interests at their share of value", nci, -base["nci_deduction"], 1.0)
    eq = evaluate(wb, "Fundamental Valuation", "B12")
    chk("equity value attributable to shareholders", eq, base["equity"], 1.0)
    ps = evaluate(wb, "Fundamental Valuation", "B14")
    chk("value per share", ps, base["per_share"], 0.01)
    tv = evaluate(wb, "Fundamental Valuation", "B16")
    chk("terminal share of enterprise value", tv, base["terminal_share"], 0.001)

    sotp = evaluate(wb, "SOTP Bridge", "B11")
    chk("bridge sheet agrees with the valuation sheet on value per share",
        sotp, base["per_share"], 0.02)

    # The discounted-cash-flow sheet, resolved BY LABEL. Row numbers are not
    # remembered anywhere in this file: adding one driver shifts every row
    # below it, and a check reading a stale row still computes and is simply
    # wrong — it read "revenue per delivered unit" as revenue and reported a
    # mismatch of four orders of magnitude.
    def at(sheet, label, after=0):
        ws = wb[sheet]
        for r in range(after + 1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == label:
                return r
        raise ValueError("%r not found on %s after row %d"
                         % (label, sheet, after))

    ws = wb["DCF"]
    for lbl, key, tol in (("Units delivered", "units_delivered", 0.01),
                          ("Revenue per delivered unit", "rev_per_unit", 0.01),
                          ("Revenue", "revenue", 1.0)):
        rr = at("DCF", lbl)
        for j, r in zip(range(2, 2 + len(base["rows"])), N["bottom_up"]["rows"]):
            col = ws.cell(1, j).column_letter
            chk("DCF %s %d" % (lbl.lower(), r["year"]),
                evaluate(wb, "DCF", "%s%d" % (col, rr)), r[key], tol)
    fr = at("DCF", "Free cash flow to the firm")
    pr = at("DCF", "Present value")
    for j, w in zip(range(2, 2 + len(base["rows"])),
                    N["statements"]["dcf_b"]["waterfall"]):
        col = ws.cell(1, j).column_letter
        chk("DCF free cash flow %d" % w["year"],
            evaluate(wb, "DCF", "%s%d" % (col, fr)), w["fcff"], 1.0)
        chk("DCF present value %d" % w["year"],
            evaluate(wb, "DCF", "%s%d" % (col, pr)), w["pv"], 1.0)
    chk("DCF sheet's explicit-period total",
        evaluate(wb, "DCF", "B%d" % at("DCF", "Sum of the explicit years")),
        N["statements"]["dcf_b"]["pv_explicit"], 1.0)
    chk("DCF sheet's enterprise value",
        evaluate(wb, "DCF", "B%d" % at("DCF", "Enterprise value")),
        base["ev"], 1.0)

    ck = at("Balance Sheet", "Check: assets less liabilities and equity")
    chk("balance sheet balances, 2025",
        evaluate(wb, "Balance Sheet", "B%d" % ck), 0.0, 0.2)
    chk("balance sheet balances, 2024",
        evaluate(wb, "Balance Sheet", "C%d" % ck), 0.0, 0.2)

    # the projected statements: read the workbook's own cells back by label,
    # never by a remembered row number, and check every forecast year
    ST = N["statements"]

    _find = at

    for tag, key, first in (("cash conversion holds", "framing_b", 0),
                            ("the cycle holds", "framing_a", None)):
        rows = ST[key]
        hdr = _find("Balance Sheet",
                    "FORECAST — IF CASH CONVERSION HOLDS" if first == 0
                    else "FORECAST — IF THE COLLECTION CYCLE HOLDS")
        ta = _find("Balance Sheet", "TOTAL ASSETS", hdr)
        tl = _find("Balance Sheet", "TOTAL LIABILITIES", hdr)
        eqr = _find("Balance Sheet", "Shareholders' funds", tl)
        for j, r in enumerate(rows, start=2):
            col = wb["Balance Sheet"].cell(1, j).column_letter
            a = evaluate(wb, "Balance Sheet", "%s%d" % (col, ta))
            l = evaluate(wb, "Balance Sheet", "%s%d" % (col, tl))
            e = evaluate(wb, "Balance Sheet", "%s%d" % (col, eqr))
            chk("forecast sheet balances, %s, %d" % (tag, r["year"]),
                a - l - e, 0.0, 0.01)
            chk("forecast total assets, %s, %d" % (tag, r["year"]),
                a, r["total_assets"], 1.0)
        cfh = _find("Cash Flow",
                    "IF CASH CONVERSION HOLDS — the basis of the valuation"
                    if first == 0 else "IF THE COLLECTION CYCLE HOLDS")
        cfo = _find("Cash Flow", "Cash from operations", cfh)
        cash = _find("Cash Flow", "Closing cash", cfo)
        for j, r in enumerate(rows, start=2):
            col = wb["Cash Flow"].cell(1, j).column_letter
            chk("forecast operating cash, %s, %d" % (tag, r["year"]),
                evaluate(wb, "Cash Flow", "%s%d" % (col, cfo)), r["cfo"], 1.0)
            chk("closing cash never negative, %s, %d" % (tag, r["year"]),
                1.0 if evaluate(wb, "Cash Flow",
                                "%s%d" % (col, cash)) >= 0 else 0.0, 1.0, 0.0)

    # the waterfall the study discounts must be the workbook's own free cash flow
    for j, w in enumerate(ST["dcf_b"]["waterfall"], start=2):
        chk("free cash flow discounted, %d" % w["year"],
            w["cfo"] + w["interest_addback"] - w["capex"], w["fcff"], 0.01)
    chk("present values sum to the explicit-period figure",
        sum(x["pv"] for x in ST["dcf_b"]["waterfall"]),
        ST["dcf_b"]["pv_explicit"], 0.01)
    chk("the study's discounted cash flow is the one in the statements",
        ST["dcf_b"]["per_share"], base["per_share"], 0.01)

    for k, want in (("B5", D["cfo_mid"]), ("B12", N["wacc"]["wacc_rating"]),
                    ("B13", D["net_debt_bridge"]), ("B16", D["nci_value_share"]),
                    ("B17", D["shares_mn"])):
        chk("assumption %s" % k, evaluate(wb, "Assumptions", k), want, 0.01)

    json.dump(checks, open(os.path.join(HERE, "recalc_result.json"), "w"), indent=1)
    for c in checks:
        print("  %-52s workbook %12.4f   study %12.4f   %s"
              % (c["check"][:52], c["workbook"], c["study"],
                 "ok" if c["ok"] else "MISMATCH"))
    print("\n%d checks, %d mismatches" % (len(checks), fails))
    if fails:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
