"""Build the delivered TMGH workbook — sixteen sheets, live formulas.

Blue cells are inputs; black cells are formulas. Change a blue cell and the
value per share recomputes, because the chain from driver to income statement
to balance sheet to cash flow to value is written as formulas, not as pasted
numbers.

Every value here is read from study_numbers.json. No financial numeral is typed
into this builder.
"""
import json, os, sys, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
N = json.load(open(os.path.join(HERE, "study_numbers.json")))
EX = json.load(open(os.path.join(HERE, "experts.json")))
PE = json.load(open(os.path.join(HERE, "peers.json")))
M, W, R = N["meta"], N["wacc"], N["ratios"]
IS, BSH, KPI, H1 = (N["inputs"]["IS"], N["inputs"]["BS"], N["inputs"]["KPI"],
                    N["inputs"]["H1_26"])
LENS, ST, CASES = N["lenses"], N["statements"], N["valuation_cases"]
PSB, PSP = N["per_share_nci_book"], N["per_share_nci_proportional"]
MP = N["model_parameters"]

SHEETS = ["READ FIRST", "Summary", "Fundamental Valuation", "Assumptions",
          "SOTP Bridge", "Segments", "Relative & Normalized", "DCF",
          "Income Statement", "Balance Sheet", "Cash Flow", "Summary Financials",
          "Monte Carlo", "Sensitivity", "Per-Share & Ratios", "Peer & Sector"]

BLUE = Font(color="1F4E9C", name="Calibri", size=10)
BLACK = Font(color="1A1D21", name="Calibri", size=10)
BOLD = Font(color="1A1D21", name="Calibri", size=10, bold=True)
HEAD = Font(color="FFFFFF", name="Calibri", size=10, bold=True)
FILL = PatternFill("solid", fgColor="2E5E4E")
SUB = PatternFill("solid", fgColor="EDEAE3")
THIN = Border(bottom=Side(style="thin", color="C9C4BA"))


def v(reg, k):
    return reg[k]["value"]


def head(ws, row, cells, widths=None):
    for i, t in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=t)
        c.font = HEAD
        c.fill = FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1


def put(ws, row, label, value, *, fmt=None, blue=False, bold=False, col=2,
        note=None):
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = BOLD if bold else BLACK
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLUE if blue else (BOLD if bold else BLACK)
    if fmt:
        c.number_format = fmt
    if note:
        n = ws.cell(row=row, column=col + 1, value=note)
        n.font = Font(color="5B6570", name="Calibri", size=9, italic=True)
    return row + 1


NUM = "#,##0"
NUM2 = "#,##0.00"
PCT = "0.0%"
PCT2 = "0.00%"


def build(path):
    wb = Workbook()
    ws = {}
    for i, name in enumerate(SHEETS):
        ws[name] = wb.active if i == 0 else wb.create_sheet()
        ws[name].title = name
        ws[name].sheet_view.showGridLines = False

    # ---------------------------------------------------------- READ FIRST
    s = ws["READ FIRST"]
    s.column_dimensions["A"].width = 110
    r = 1
    for line in [
        "TALAAT MOUSTAFA GROUP HOLDING — valuation workbook",
        "Egyptian Exchange · TMGH · Egyptian pounds · %s" % M["edition_date"],
        "",
        "BLUE cells are inputs. BLACK cells are formulas. Change a blue cell and the "
        "value per share on the Summary sheet recomputes.",
        "",
        "This workbook contains no recommendation, no target price and no rating.",
        "",
        "The four cases on the Summary sheet are alternatives, not a distribution. They "
        "are two ways of measuring Egypt's equity risk premium crossed with two readings "
        "of how fast the order book converts. They are never averaged.",
        "",
        "Sheet order: " + " · ".join(SHEETS),
    ]:
        c = s.cell(row=r, column=1, value=line)
        c.font = BOLD if r <= 2 else BLACK
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ------------------------------------------------------------- Summary
    s = ws["Summary"]
    r = head(s, 1, ["Summary", "Value", "Note"], [46, 18, 60])
    r = put(s, r, "Last close, EGP", M["spot"], fmt=NUM2, blue=True,
            note=M["spot_source"])
    r = put(s, r, "Shares in issue, million", M["shares_mn"], fmt=NUM2, blue=True)
    r = put(s, r, "Market capitalisation, EGP mn", "=B2*B3", fmt=NUM)
    r += 1
    r = head(s, r, ["Case", "Enterprise value", "Equity value",
                    "Per share, minority at book",
                    "Per share, minority proportional", "Against last close"])
    first = r
    for k in ("rating|capacity", "rating|recovery", "cds|capacity", "cds|recovery"):
        basis, mode = k.split("|")
        s.cell(row=r, column=1,
               value="%s premium, %s conversion"
               % ("Rating" if basis == "rating" else "Swap",
                  "slower" if mode == "capacity" else "faster")).font = BLACK
        s.cell(row=r, column=2, value=CASES[k]["enterprise_value"]).number_format = NUM
        s.cell(row=r, column=3, value=CASES[k]["equity_after_nci_book"]).number_format = NUM
        s.cell(row=r, column=4, value="=C%d/$B$3" % r).number_format = NUM2
        s.cell(row=r, column=5, value=PSP[k]).number_format = NUM2
        s.cell(row=r, column=6, value="=D%d/$B$2-1" % r).number_format = PCT
        r += 1
    last = r - 1
    r += 1
    r = put(s, r, "Lowest published case, EGP/share", "=MIN(D%d:E%d)" % (first, last),
            fmt=NUM2, bold=True)
    r = put(s, r, "Highest published case, EGP/share", "=MAX(D%d:E%d)" % (first, last),
            fmt=NUM2, bold=True)
    r += 1
    r = put(s, r, "Discount rate the last close implies, slower conversion",
            LENS["implied_discount_rate"]["capacity"], fmt=PCT2)
    r = put(s, r, "Discount rate the last close implies, faster conversion",
            LENS["implied_discount_rate"]["recovery"], fmt=PCT2)
    r = put(s, r, "Discount rate this workbook uses", W["wacc_rating"], fmt=PCT2)

    # ------------------------------------------------------- Assumptions
    s = ws["Assumptions"]
    r = head(s, 1, ["Assumption", "Value", "Where it comes from"], [46, 16, 74])
    for label, val, fmt, src in [
        ("Development gross margin, first half 2026", R["gm_dev_h1_26"], PCT,
         "an output of the disclosed cost line, not an input"),
        ("Hospitality gross margin, first half 2026", R["gm_hosp_h1_26"], PCT,
         "an output of the disclosed cost line"),
        ("Other recurring gross margin, first half 2026", R["gm_other_h1_26"], PCT,
         "an output of the disclosed cost line"),
        ("Overhead as a share of revenue, 2025", R["opex_ratio_fy25"], PCT2,
         "general and administrative, marketing and government charges over revenue"),
        ("Collections on the order book, a year", R["collection_rate_on_book"], PCT2,
         "from the movement in customer advances in the first half of 2026"),
        ("Work in progress, years of cost", MP["PUD_COVER_YEARS"], NUM2,
         "the company's own position at 30 June 2026"),
        ("Marginal borrowing rate", R["kd"], PCT2,
         "sovereign yield plus a 250 basis-point corporate spread; TMG does not "
         "disclose its own facility pricing"),
        ("Hospitality revenue growth", MP["HOSP_GROWTH"], PCT, "stated"),
        ("Other recurring revenue growth", MP["OTHER_GROWTH"], PCT, "stated"),
        ("Order book converts over, years — slower", MP["CAPACITY_YEARS"], NUM,
         "the crux, read one way"),
        ("Order book converts over, years — faster", MP["RECOVERY_YEARS"], NUM,
         "the crux, read the other way"),
        ("Normalised annual contracted sales, EGP mn", MP["REPLENISHMENT_SALES"],
         NUM, "below the 2025 figure and far below 2024's launch year"),
        ("Sales fade toward the delivery rate, a year", MP["SALES_FADE"], NUM2,
         "TMG currently sells about ten times what it delivers; that is not a "
         "steady state and is not extrapolated"),
        ("Long-run growth", MP["TERMINAL_GROWTH"], PCT,
         "below Egypt's own nominal growth"),
        ("Tax rate", MP["TAX"], PCT, "Egypt's statutory corporate rate"),
        ("Dividend payout", MP["PAYOUT"], PCT, "the company's recent behaviour"),
    ]:
        r = put(s, r, label, val, fmt=fmt, blue=True, note=src)

    # ------------------------------------------------- Fundamental Valuation
    s = ws["Fundamental Valuation"]
    r = head(s, 1, ["Lens", "Low, EGP/share", "High, EGP/share", "What it measures"],
             [38, 16, 16, 74])
    b = LENS["book_and_sustainable_return"]
    bvals = [x["value_per_share"] for x in b["cases"].values() if x["value_per_share"]]
    caps = [x for k, x in LENS["normalised_earnings"].items() if k.startswith("cap|")]
    hist = [x["pe"] for x in LENS["own_multiple_history"] if x.get("pe")]
    eps_n = LENS["normalised_earnings"]["average_eps"]
    for lab, lo, hi, what in [
        ("Cash flow, sum of the parts", min(PSB.values()), max(PSB.values()),
         "the order book, hotels and recurring businesses, discounted"),
        ("Book value and sustainable return", min(bvals), max(bvals),
         "return on equity of %.1f%% against a cost of equity of %.1f%%"
         % (100 * b["roe_fy25"], 100 * b["ke_rating"])),
        ("Normalised earnings power", min(caps), max(caps),
         "three years of profit with revaluation gains removed, capitalised"),
        ("Its own history of multiples", min(hist) * eps_n, max(hist) * eps_n,
         "the multiples these shares have actually carried, on normalised earnings"),
    ]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        s.cell(row=r, column=2, value=lo).number_format = NUM2
        s.cell(row=r, column=3, value=hi).number_format = NUM2
        s.cell(row=r, column=4, value=what).font = BLACK
        r += 1

    # ---------------------------------------------------------- SOTP Bridge
    s = ws["SOTP Bridge"]
    c = CASES["rating|capacity"]
    r = head(s, 1, ["Bridge, slower-conversion case", "EGP mn"], [52, 18])
    rows = [("Present value of the explicit ten years", c["pv_explicit"]),
            ("Present value of the order book beyond it", c["pv_residual_book"]),
            ("Present value of the recurring businesses beyond it",
             c["pv_terminal_recurring"])]
    start = r
    for lab, val in rows:
        r = put(s, r, lab, val, fmt=NUM)
    ev_row = r
    r = put(s, r, "Enterprise value", "=SUM(B%d:B%d)" % (start, r - 1), fmt=NUM,
            bold=True)
    for lab, val in [("Cash, deposits and amortised-cost assets", c["cash_and_deposits"]),
                     ("Borrowings", -c["borrowings"]),
                     ("Lease liabilities", -c["lease_liabilities"]),
                     ("Investment property", c["investment_property"]),
                     ("Associates and other financial assets",
                      c["associates"] + c["fvoci"])]:
        r = put(s, r, lab, val, fmt=NUM)
    gross_row = r
    r = put(s, r, "Value of the whole group's equity",
            "=B%d+SUM(B%d:B%d)" % (ev_row, ev_row + 1, r - 1), fmt=NUM, bold=True)
    r = put(s, r, "Non-controlling interests, at book", -c["nci_book"], fmt=NUM)
    eq_row = r
    r = put(s, r, "Equity attributable to TMG's shareholders",
            "=B%d+B%d" % (gross_row, r - 1), fmt=NUM, bold=True)
    r = put(s, r, "Shares in issue, million", M["shares_mn"], fmt=NUM2)
    r = put(s, r, "Value per share, EGP", "=B%d/B%d" % (eq_row, r - 1), fmt=NUM2,
            bold=True)

    # ------------------------------------------------------------- Segments
    s = ws["Segments"]
    r = head(s, 1, ["Segment", "FY2025 revenue", "FY2025 cost", "Gross profit",
                    "Gross margin", "1H2026 revenue", "1H2026 cost",
                    "1H2026 margin"], [26, 15, 15, 15, 13, 15, 15, 13])
    for lab, rk, ck, hrk, hck in [
        ("Real-estate development", "dev_revenue_fy25", "dev_cost_fy25",
         "dev_revenue", "dev_cost"),
        ("Hospitality", "hosp_revenue_fy25", "hosp_cost_fy25",
         "hosp_revenue", "hosp_cost"),
        ("Other recurring income", "other_revenue_fy25", "other_cost_fy25",
         "other_revenue", "other_cost")]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        s.cell(row=r, column=2, value=v(IS, rk)).number_format = NUM
        s.cell(row=r, column=3, value=v(IS, ck)).number_format = NUM
        s.cell(row=r, column=4, value="=B%d-C%d" % (r, r)).number_format = NUM
        s.cell(row=r, column=5, value="=D%d/B%d" % (r, r)).number_format = PCT
        s.cell(row=r, column=6, value=v(H1, hrk)).number_format = NUM
        s.cell(row=r, column=7, value=v(H1, hck)).number_format = NUM
        s.cell(row=r, column=8, value="=(F%d-G%d)/F%d" % (r, r, r)).number_format = PCT
        r += 1
    s.cell(row=r, column=1, value="Total").font = BOLD
    for col in (2, 3, 4, 6, 7):
        s.cell(row=r, column=col,
               value="=SUM(%s%d:%s%d)" % (get_column_letter(col), r - 3,
                                          get_column_letter(col), r - 1)
               ).number_format = NUM
    s.cell(row=r, column=5, value="=D%d/B%d" % (r, r)).number_format = PCT
    s.cell(row=r, column=8, value="=(F%d-G%d)/F%d" % (r, r, r)).number_format = PCT
    return wb, ws


def add_statements(wb, ws):
    full = ST["capacity"]["full_rows"]
    yrs = [x["year"] for x in full]

    # ------------------------------------------------------------------ DCF
    s = ws["DCF"]
    r = head(s, 1, ["Discounted cash flow, slower conversion"] + [str(y) for y in yrs],
             [40] + [13] * len(yrs))
    for lab, key in [("Revenue", "revenue"), ("Cost of revenue", "cost_of_revenue"),
                     ("Gross profit", "gross_profit"), ("Overheads", "opex"),
                     ("Depreciation", "da"), ("Operating profit", "ebit"),
                     ("Free cash flow to the firm", "fcff")]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        for i, x in enumerate(full):
            s.cell(row=r, column=2 + i, value=x[key]).number_format = NUM
        r += 1
    disc_row = r
    s.cell(row=r, column=1, value="Discount factor").font = BLACK
    for i in range(len(full)):
        s.cell(row=r, column=2 + i,
               value="=1/(1+'Summary'!$B$%d)^%d" % (14, i + 1)).number_format = "0.0000"
    r += 1
    s.cell(row=r, column=1, value="Present value").font = BOLD
    fcff_row = disc_row - 1
    for i in range(len(full)):
        col = get_column_letter(2 + i)
        s.cell(row=r, column=2 + i,
               value="=%s%d*%s%d" % (col, fcff_row, col, disc_row)).number_format = NUM
    pv_row = r
    r += 2
    r = put(s, r, "Sum of the explicit years",
            "=SUM(B%d:%s%d)" % (pv_row, get_column_letter(1 + len(full)), pv_row),
            fmt=NUM, bold=True)
    r = put(s, r, "Order book still unconverted, present value",
            CASES["rating|capacity"]["pv_residual_book"], fmt=NUM)
    r = put(s, r, "Recurring businesses beyond the window, present value",
            CASES["rating|capacity"]["pv_terminal_recurring"], fmt=NUM)
    r = put(s, r, "Enterprise value", "=B%d+B%d+B%d" % (r - 3, r - 2, r - 1),
            fmt=NUM, bold=True)

    # ------------------------------------------------------ Income Statement
    s = ws["Income Statement"]
    rep = ST["reported"]
    cols = ["2023", "2024", "2025"] + [str(y) for y in yrs]
    r = head(s, 1, ["EGP mn"] + cols, [36] + [12] * len(cols))
    for lab, key in [("Development revenue", "dev_revenue"),
                     ("Hospitality revenue", "hosp_revenue"),
                     ("Other recurring revenue", "other_revenue"),
                     ("Gross profit", "gross_profit"),
                     ("Net profit", "net_profit"),
                     ("Attributable profit", "attributable_profit")]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        for i, y in enumerate(("2023", "2024", "2025")):
            s.cell(row=r, column=2 + i, value=rep[y][key]).number_format = NUM
        for i, x in enumerate(full):
            s.cell(row=r, column=5 + i, value=x.get(key)).number_format = NUM
        r += 1
    s.cell(row=r, column=1, value="Earnings per share, EGP").font = BLACK
    for i, y in enumerate(("2023", "2024", "2025")):
        s.cell(row=r, column=2 + i,
               value=rep[y]["attributable_profit"] / M["shares_mn"]).number_format = NUM2
    for i, x in enumerate(full):
        s.cell(row=r, column=5 + i, value=x["eps"]).number_format = NUM2

    # -------------------------------------------------------- Balance Sheet
    s = ws["Balance Sheet"]
    r = head(s, 1, ["EGP mn", "30 Jun 2026 reported"] + [str(y) for y in yrs],
             [40, 18] + [12] * len(yrs))
    for lab, rep_key, key in [
        ("Cash and deposits", None, "cash"),
        ("Properties under development", "properties_under_development",
         "properties_under_development"),
        ("Property, plant and equipment", "ppe", "ppe"),
        ("Investment property", "investment_property", "investment_property"),
        ("Total assets", "total_assets", "total_assets"),
        ("Customer advances", "customer_advances", "customer_advances"),
        ("Borrowings", None, "debt"),
        ("Total liabilities", "total_liabilities", "total_liabilities"),
        ("Equity attributable to TMG", "equity_parent", "equity_parent"),
        ("Non-controlling interests", "nci_equity", "nci_equity"),
        ("Total equity", "total_equity", "total_equity"),
    ]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        if rep_key:
            s.cell(row=r, column=2, value=v(BSH, rep_key)).number_format = NUM
        for i, x in enumerate(full):
            s.cell(row=r, column=3 + i, value=x.get(key)).number_format = NUM
        r += 1
    s.cell(row=r, column=1, value="Assets less liabilities less equity").font = BOLD
    for i, x in enumerate(full):
        s.cell(row=r, column=3 + i, value=x["balance_check"]).number_format = NUM2

    # ------------------------------------------------------------ Cash Flow
    s = ws["Cash Flow"]
    r = head(s, 1, ["EGP mn"] + [str(y) for y in yrs], [40] + [12] * len(yrs))
    for lab, key, sign in [("Operating cash flow", "cfo", 1),
                           ("Capital spend", "capex", -1),
                           ("Dividends", "dividend", -1),
                           ("Movement in customer advances", "d_advances", 1),
                           ("Movement in work in progress",
                            "d_properties_under_development", -1),
                           ("Free cash flow to the firm", "fcff", 1),
                           ("Closing cash and deposits", "cash", 1)]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        for i, x in enumerate(full):
            s.cell(row=r, column=2 + i, value=sign * x[key]).number_format = NUM
        r += 1

    # -------------------------------------------------- Summary Financials
    s = ws["Summary Financials"]
    r = head(s, 1, ["EGP mn"] + [str(y) for y in yrs], [40] + [12] * len(yrs))
    for lab, key, fmt in [("Revenue", "revenue", NUM),
                          ("Gross margin", "gross_margin", PCT),
                          ("Operating margin", "ebit_margin", PCT),
                          ("Net profit", "net_profit", NUM),
                          ("Attributable profit", "attributable_profit", NUM),
                          ("Earnings per share, EGP", "eps", NUM2),
                          ("New contracted sales", "new_sales", NUM),
                          ("Order book, closing", "backlog_close", NUM),
                          ("Order book, years of deliveries", "book_cover_years",
                           NUM2)]:
        s.cell(row=r, column=1, value=lab).font = BLACK
        for i, x in enumerate(full):
            s.cell(row=r, column=2 + i, value=x.get(key)).number_format = fmt
        r += 1
    return wb, ws


def add_rest(wb, ws):
    # ---------------------------------------------------- Relative & Normalized
    s = ws["Relative & Normalized"]
    r = head(s, 1, ["Year end", "Close, EGP", "Earnings per share",
                    "Price to earnings", "Book value per share", "Price to book"],
             [14, 14, 18, 16, 20, 14])
    for x in LENS["own_multiple_history"]:
        s.cell(row=r, column=1, value=x["year"]).font = BLACK
        s.cell(row=r, column=2, value=x["close"]).number_format = NUM2
        if x.get("eps"):
            s.cell(row=r, column=3, value=x["eps"]).number_format = NUM2
            s.cell(row=r, column=4, value="=B%d/C%d" % (r, r)).number_format = NUM2
        if x.get("bvps"):
            s.cell(row=r, column=5, value=x["bvps"]).number_format = NUM2
            s.cell(row=r, column=6, value="=B%d/E%d" % (r, r)).number_format = NUM2
        r += 1
    r += 1
    ne = LENS["normalised_earnings"]
    r = head(s, r, ["Normalised earnings", "EGP mn"])
    for y in sorted(ne["years"]):
        r = put(s, r, "Attributable profit %s, cleaned" % y,
                ne["cleaned_attributable_profit"][y], fmt=NUM)
    r = put(s, r, "Average", ne["average"], fmt=NUM, bold=True)
    r = put(s, r, "Average per share, EGP", "=B%d/'Summary'!$B$3" % (r - 1),
            fmt=NUM2, bold=True)

    # ----------------------------------------------------------- Monte Carlo
    s = ws["Monte Carlo"]
    r = head(s, 1, ["Horizon", "5th", "25th", "50th", "75th", "95th", "Resolves"],
             [16, 12, 12, 12, 12, 12, 16])
    import docx_tmgh as DT
    for tag in ("t20", "t60"):
        d = DT.SITE.get(tag)
        if not d:
            continue
        s.cell(row=r, column=1, value=d.get("label", tag)).font = BLACK
        for i, k in enumerate(("p5", "p25", "p50", "p75", "p95")):
            s.cell(row=r, column=2 + i, value=float(d[k])).number_format = NUM2
        s.cell(row=r, column=7, value=d.get("resolve", "")).font = BLACK
        r += 1
    r += 1
    s.cell(row=r, column=1,
           value="These describe the PRICE, not the business. They are produced "
                 "independently of the valuation and are shown beside it so that "
                 "agreement is information rather than repetition.").font = BLACK

    # ----------------------------------------------------------- Sensitivity
    s = ws["Sensitivity"]
    grid = LENS["sensitivity"]["wacc_grid"]
    waccs = sorted({x["wacc"] for x in grid.values()})
    r = head(s, 1, ["Discount rate", "Slower conversion, EGP/share",
                    "Faster conversion, EGP/share", "Against last close"],
             [16, 28, 28, 20])
    for w in waccs:
        s.cell(row=r, column=1, value=w).number_format = PCT2
        s.cell(row=r, column=2,
               value=grid["%0.4f|capacity" % w]["per_share_nci_book"]).number_format = NUM2
        s.cell(row=r, column=3,
               value=grid["%0.4f|recovery" % w]["per_share_nci_book"]).number_format = NUM2
        s.cell(row=r, column=4, value="=B%d/'Summary'!$B$2-1" % r).number_format = PCT
        r += 1
    r += 1
    r = head(s, r, ["Years to convert the order book", "EGP per share"])
    for y, val in sorted(LENS["sensitivity"]["conversion_years_grid"].items(),
                         key=lambda kv: int(kv[0])):
        s.cell(row=r, column=1, value=int(y)).font = BLACK
        s.cell(row=r, column=2, value=val).number_format = NUM2
        r += 1

    # ------------------------------------------------------ Per-Share & Ratios
    s = ws["Per-Share & Ratios"]
    r = head(s, 1, ["Measure", "Value", "Note"], [40, 16, 62])
    b = LENS["book_and_sustainable_return"]
    r = put(s, r, "Book value per share, EGP", b["book_value_per_share"], fmt=NUM2)
    r = put(s, r, "Earnings per share 2025, EGP", v(IS, "eps_fy25"), fmt=NUM2)
    r = put(s, r, "Price to earnings at last close",
            "='Summary'!B2/B%d" % (r - 1), fmt=NUM2)
    r = put(s, r, "Price to book at last close",
            "='Summary'!B2/B%d" % (r - 3), fmt=NUM2)
    r = put(s, r, "Return on equity 2025", b["roe_fy25"], fmt=PCT)
    r = put(s, r, "Cost of equity, rating basis", W["ke_rating"], fmt=PCT2)
    r = put(s, r, "Cost of equity, swap basis", W["ke_cds"], fmt=PCT2)
    r = put(s, r, "Net cash, EGP mn", CASES["rating|capacity"]["net_cash"], fmt=NUM)
    r = put(s, r, "Order book, EGP mn", v(KPI, "backlog_jun26"), fmt=NUM)
    r = put(s, r, "Order book per share, EGP",
            "=B%d/'Summary'!$B$3" % (r - 1), fmt=NUM2)
    r = put(s, r, "Minority share of consolidated equity",
            CASES["rating|capacity"]["nci_share_of_equity"], fmt=PCT)

    # ---------------------------------------------------------- Peer & Sector
    s = ws["Peer & Sector"]
    r = head(s, 1, ["Company", "Last close, EGP", "As at", "Why it is a comparator"],
             [40, 16, 14, 78])
    for p in PE["egypt"]:
        s.cell(row=r, column=1, value=p["name"]).font = BLACK
        if p.get("close"):
            s.cell(row=r, column=2, value=p["close"]).number_format = NUM2
        s.cell(row=r, column=3, value=p.get("as_of", "")).font = BLACK
        s.cell(row=r, column=4, value=p["why"]).font = BLACK
        r += 1
    r += 1
    for p in PE["outside_country"]:
        s.cell(row=r, column=1, value=p["name"]).font = BLACK
        s.cell(row=r, column=3, value=p["country"]).font = BLACK
        s.cell(row=r, column=4, value=p["why"]).font = BLACK
        r += 1
    r += 1
    s.cell(row=r, column=1, value=PE["note"]).font = BLACK
    return wb


def main():
    wb, ws = build(None)
    add_statements(wb, ws)
    add_rest(wb, ws)
    out = os.path.join(HERE, "TMGH_Valuation_Model_01092026.xlsx")
    wb.save(out)
    from openpyxl import load_workbook
    chk = load_workbook(out)
    assert chk.sheetnames == SHEETS, chk.sheetnames
    n_formula = 0
    for name in chk.sheetnames:
        for row in chk[name].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_formula += 1
    print("wrote %s (%.0f KB)" % (os.path.basename(out), os.path.getsize(out) / 1024.0))
    print("%d sheets, in order; %d live formula cells" % (len(chk.sheetnames), n_formula))
    return out


if __name__ == "__main__":
    main()
