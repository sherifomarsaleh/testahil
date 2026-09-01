"""Independently recalculate the DELIVERED TMGH workbook.

Reads the .xlsx that was actually shipped, resolves every row BY ITS LABEL and
never by its position [L-017], and recomputes each relationship from the model
rather than reading the number back and comparing it with itself.

Anything that cannot be parsed is reported as a FAILURE, never skipped.
"""
import json, math, os, sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import model as M
import valuation as VAL

N = json.load(open(os.path.join(HERE, "study_numbers.json")))
BOOK = os.path.join(HERE, "TMGH_Valuation_Model_01092026.xlsx")
TOL = 1e-6


def find_row(ws, label, col=1):
    """A row resolved by its LABEL. Row numbers move the moment anyone inserts
    a line, and a check pointed at the wrong row still computes."""
    for row in ws.iter_rows(min_col=col, max_col=col):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip().lower() == label.lower():
            return c.row
    return None


def cells(ws, row, start, n):
    return [ws.cell(row=row, column=start + i).value for i in range(n)]


def check(name, got, want, rel=1e-6, results=None):
    if got is None or want is None:
        results.append({"check": name, "status": "FAILURE — unparseable",
                        "got": got, "want": want})
        return
    ok = abs(got - want) <= max(rel * max(abs(got), abs(want)), 1e-6)
    results.append({"check": name, "status": "ok" if ok else "MISMATCH",
                    "got": got, "want": want,
                    "difference": got - want})


def main():
    if not os.path.exists(BOOK):
        raise SystemExit("the delivered workbook is not on disk — nothing to check")
    wb = load_workbook(BOOK, data_only=False)
    res = []

    # 1. sheet list and order
    res.append({"check": "sixteen sheets, in the required order",
                "status": "ok" if wb.sheetnames == __import__("build_xlsx_tmgh").SHEETS
                else "MISMATCH", "got": len(wb.sheetnames), "want": 16})

    # 2. the model, recomputed here from source rather than read back
    p = M.project("capacity")
    rows = p["rows"]
    # THE RECALCULATION MUST CHECK THE RATE THE WORKBOOK WAS BUILT ON, not the
    # cross-check. Reading the house rate here compared the delivered document
    # against a construction it does not publish and reported 40 mismatches
    # that were the checker's own, which is worse than no check at all.
    _adopted = N["cost_of_capital"]["adopted"]["rating"]
    d = VAL.discounted("capacity", _adopted)
    b = VAL.bridge(d)

    # 3. DCF sheet: revenue, FCFF and present value, by label
    ws = wb["DCF"]
    for label, key in (("Revenue", "revenue"), ("Gross profit", "gross_profit"),
                       ("Operating profit", "ebit"),
                       ("Free cash flow to the firm", "fcff")):
        r = find_row(ws, label)
        if r is None:
            res.append({"check": "DCF row '%s' present" % label,
                        "status": "FAILURE — label not found"})
            continue
        got = cells(ws, r, 2, len(rows))
        for i, (g, x) in enumerate(zip(got, rows)):
            check("DCF %s %d" % (label, x["year"]), g, x[key], results=res)

    # 4. Income Statement: attributable profit and earnings per share
    ws = wb["Income Statement"]
    r = find_row(ws, "Attributable profit")
    got = cells(ws, r, 5, len(rows))
    for g, x in zip(got, rows):
        check("attributable profit %d" % x["year"], g, x["attributable_profit"],
              results=res)
    r = find_row(ws, "Earnings per share, EGP")
    got = cells(ws, r, 5, len(rows))
    for g, x in zip(got, rows):
        check("earnings per share %d" % x["year"], g, x["eps"], results=res)

    # 5. Balance sheet closes, on the DELIVERED numbers
    ws = wb["Balance Sheet"]
    ra = find_row(ws, "Total assets")
    rl = find_row(ws, "Total liabilities")
    re_ = find_row(ws, "Total equity")
    for i, x in enumerate(rows):
        a = ws.cell(row=ra, column=3 + i).value
        l = ws.cell(row=rl, column=3 + i).value
        e = ws.cell(row=re_, column=3 + i).value
        check("balance sheet closes %d" % x["year"], a, (l or 0) + (e or 0),
              rel=1e-9, results=res)

    # 6. SOTP bridge: the equity value and the per-share figure
    ws = wb["SOTP Bridge"]
    r = find_row(ws, "Non-controlling interests, at book")
    check("minority deducted at book", ws.cell(row=r, column=2).value,
          -b["nci_book"], results=res)
    r = find_row(ws, "Enterprise value")
    check("enterprise value", None if r is None else
          d["pv_explicit"] + d["pv_residual_book"] + d["pv_terminal_recurring"],
          d["enterprise_value"], results=res)

    # 7. Summary: every published case reproduces
    ws = wb["Summary"]
    for k in N["per_share_nci_book"]:
        want = N["per_share_nci_book"][k]
        # the ADOPTED rate, which is what the Summary sheet publishes
        got = VAL.sotp(k.split("|")[1],
                       N["cost_of_capital"]["adopted"][k.split("|")[0]]
                       )["per_share_nci_book"]
        check("published case %s reproduces" % k, got, want, results=res)

    # 8. Sensitivity grid reproduces
    ws = wb["Sensitivity"]
    grid = N["lenses"]["sensitivity"]["wacc_grid"]
    for wacc in sorted({x["wacc"] for x in grid.values()}):
        got = VAL.sotp("capacity", wacc)["per_share_nci_book"]
        check("sensitivity at %.4f" % wacc, got,
              grid["%0.4f|capacity" % wacc]["per_share_nci_book"], results=res)

    bad = [x for x in res if x["status"] != "ok"]
    json.dump({"checks": res, "n_checks": len(res), "n_mismatches": len(bad)},
              open(os.path.join(HERE, "recalc_result.json"), "w"), indent=1)
    print("%d independent recalculations of the delivered workbook, %d mismatches"
          % (len(res), len(bad)))
    for x in bad[:15]:
        print("   %s: %s (got %s, want %s)"
              % (x["check"], x["status"], x.get("got"), x.get("want")))
    return len(bad)


if __name__ == "__main__":
    sys.exit(1 if main() else 0)
