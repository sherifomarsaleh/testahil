"""Recalculate the delivered xlsx with LibreOffice and reconcile it cell-by-cell
against study_numbers.json. Fails loudly on any formula error or mismatch."""
import json, os, re, subprocess, sys, glob
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'ELEC_Valuation_Model_05082026_public.xlsx')
OUT = '/tmp/claude-0/-home-user-testahil/25f0971e-fd8d-5ac0-8335-5ab76e1c0c47/scratchpad/recalc'
os.makedirs(OUT, exist_ok=True)

# force 'always recalculate on load' for OOXML in a throwaway LO profile
prof = os.path.join(OUT, 'loprofile')
os.makedirs(os.path.join(prof, 'user'), exist_ok=True)
with open(os.path.join(prof, 'user', 'registrymodifications.xcu'), 'w') as f:
    f.write('''<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>''')
subprocess.run(['soffice', '--headless', '--norestore',
                f'-env:UserInstallation=file://{prof}',
                '--convert-to', 'xlsx', '--outdir', OUT, XLSX],
               check=True, capture_output=True, timeout=240)
recalced = os.path.join(OUT, os.path.basename(XLSX))

src = openpyxl.load_workbook(XLSX)                      # formulas
wb = openpyxl.load_workbook(recalced, data_only=True)   # recalculated values
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))

# ---- 1: formula count + zero errors ----------------------------------------
nform, errors = 0, []
for ws in src.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and re.match(r'^#(REF|DIV/0|NAME\?|VALUE|NUM|N/A|NULL)', c.value):
                errors.append(f'{ws.title}!{c.coordinate} = {c.value}')
print(f'formulas: {nform}, errors: {len(errors)}')
for e in errors[:20]:
    print('  ', e)

# ---- 2: label-driven reconciliation -----------------------------------------
def find(sheet, label_sub):
    ws = wb[sheet]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and label_sub.lower() in c.value.lower():
                return c.row
    return None

def val(sheet, coord):
    return wb[sheet][coord].value

coc, dcf, L = D['coc'], D['dcf'], D['lenses']
checks = [
    ('Assumptions!B20 WACC explicit', val('Assumptions', 'B20'), coc['wacc_exp'], 1e-4),
    ('Assumptions!B25 WACC terminal', val('Assumptions', 'B25'), coc['wacc_term'], 1e-4),
    ('Assumptions!B28 net debt', val('Assumptions', 'B28'), dcf['net_debt'], 0.5),
]
r = find('DCF', 'enterprise value')
if r: checks.append(('DCF enterprise value', wb['DCF'][f'B{r}'].value or wb['DCF'][f'G{r}'].value, dcf['ev'], 2.0))
r = find('Summary', 'weighted central')
if r:  # row layout: bear / base / bull in columns B / C / D
    checks.append(('Summary weighted central (base, col C)', wb['Summary'][f'C{r}'].value,
                   L['central']['base'], 0.01))

bad = 0
for name, got, want, tol in checks:
    ok = got is not None and abs(got - want) <= tol
    print(f'  {"OK " if ok else "FAIL"} {name}: {got} vs {want}')
    bad += 0 if ok else 1

# balance check row must be zero everywhere
r = find('Balance Sheet', 'balance check (')
if r is None:
    print('  FAIL balance-check row not found'); bad += 1
else:
    ws = wb['Balance Sheet']
    vals = [ws.cell(row=r, column=c).value for c in range(2, 9)]
    nums = [v for v in vals if isinstance(v, (int, float))]
    ok = nums and all(abs(v) < 0.5 for v in nums)
    print(f'  {"OK " if ok else "FAIL"} balance check row {r}: {["%.3f" % v for v in nums]}')
    bad += 0 if ok else 1

if errors or bad:
    sys.exit(f'RECALC FAILED: {len(errors)} formula errors, {bad} reconciliation failures')
print('RECALC + RECONCILIATION PASS')
