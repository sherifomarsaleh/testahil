"""ELEC_Valuation_Model_05082026_public.xlsx — 16 sheets mirroring the house canonical
model (operating-co variant). Blue = inputs · black = formulas · green = cross-sheet
links. All inputs live on Assumptions; engine outputs (Monte Carlo, grids) are values."""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'
YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']
coc = D['coc']; dcf = D['dcf']; L = D['lenses']; H = D['hist_is']; B = D['hist_bs']

wb = Workbook()

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 46
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill

# ============ READ FIRST =====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Electro Cable Egypt Co. S.A.E. (EGX: ELEC)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the ELEC valuation study. Every blue cell is an input; every',
 'black cell is a formula; green cells link across sheets. All inputs live on the Assumptions sheet — change one',
 '(volumes, copper, the EGP path, conversion EBITDA per tonne, the working-capital intensity, the net-debt',
 'anchor) and the whole model reprices. The forecast is a bottom-up TONNAGE build: revenue = volume x (LME x',
 'EGP x 1.387 fabrication uplift); EBITDA = volume x conversion EBITDA/t. Margins are outputs, not inputs.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges. The preparer is not licensed by any securities regulator and may hold a position.', '',
 'Sourcing note, up front. ELEC\'s audited statements were unreachable through available channels. Headline',
 'revenue/profit/assets are multiply-sourced from bourse-disclosure reporting; several lines (interest, capex,',
 'the working-capital split, the FY25 net-debt anchor) are DERIVED — each is annotated where it appears and',
 'listed with source and date in the companion Source Register document. FY2024, the one fully-triangulated',
 'year, closes to the reported net profit within 0.8% using the derived lines.', '',
 'Discount convention. Each explicit year is discounted at its own forward WACC, gliding 21.5% -> 15.0% on the',
 'same easing calendar as the interest forecast; the terminal value is capitalised at the terminal WACC and',
 'discounted at the year-5 cumulative factor. One date, one price of time.', '',
 'Currency. EGP million unless stated. Spot EGP 2.19 (5 Aug 2026 close). Sheets: Summary · Fundamental',
 'Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income Statement ·',
 'Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share & Ratios · Peer & Sector.']
 , start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 118

# ============ ASSUMPTIONS ====================================================
wa = sheet('Assumptions')
title(wa, 'Assumptions — the single input layer', 'All blue cells are inputs. Every other sheet links here.', 9)
def hdr(ws_, row, text):
    put(ws_, f'A{row}', text, bold=True, fill=FILL_H); return row + 1
def inp(ws_, row, label, val, fmt=NUM0, note=None):
    put(ws_, f'A{row}', label); put(ws_, f'B{row}', val, BLUE, fmt)
    if note: put(ws_, f'C{row}', note, SUB, None)
    return row + 1
r = hdr(wa, 4, 'ANCHORS')
r = inp(wa, r, 'Spot price (EGP/share)', 2.19, PX, '5-Aug-2026 close, uploaded EGX history')             # B5
r = inp(wa, r, 'Shares outstanding (mn)', 3313.540373, NUM0, 'Mubasher paid-in capital 662,708,074.60 / par 0.20; EPS arithmetic confirms')  # B6
r = inp(wa, r, 'Tax rate', 0.225, PCT, 'PwC Tax Summaries Egypt — 22.5% unchanged')                       # B7
r = hdr(wa, r, 'COST OF CAPITAL — sovereign double-count removed; sliding schedule below')                # 8
r = inp(wa, r, 'Egypt 10Y local yield (observed)', 0.2231, PCT, 'investing.com print 21-Jul-2026; May-26 window avg 21.3% corroborates')  # B9
r = inp(wa, r, 'Sovereign default spread (CDS-implied)', 0.0340, PCT, 'Damodaran Jan-2026 original file; market 5Y CDS ~330bp May-26')  # B10
put(wa, f'A{r}', 'rf* for the equity build'); put(wa, f'B{r}', '=B9-B10', BLACK, PCT); r += 1              # B11
r = inp(wa, r, 'ERP — Egypt, CDS-based (primary)', 0.0941, PCT, "Damodaran ORIGINAL ctryprem, Egypt row, Jan-2026 (rating basis 13.94% in sensitivity)")  # B12
r = inp(wa, r, 'Beta (own weekly regression, 5yr)', 0.964, PX, 'vs 30-name equal-weight EGX composite: R2 0.222, n=257, SE 0.113, CI90 [0.78,1.15]')  # B13
put(wa, f'A{r}', 'Ke = rf* + beta × ERP'); put(wa, f'B{r}', '=B11+B13*B12', BLACK, PCT); r += 1            # B14
r = inp(wa, r, 'Kd pre-tax (marginal EGP)', 0.220, PCT, 'Corridor + credit margin; effective-rate checks 23.5% (FY24) / 22.1% (FY25, triangulated debt path) inside 150bp')  # B15
put(wa, f'A{r}', 'Kd after tax'); put(wa, f'B{r}', '=B15*(1-B7)', BLACK, PCT); r += 1                      # B16
put(wa, f'A{r}', 'Market cap (spot × shares)'); put(wa, f'B{r}', '=B5*B6', BLACK, NUM0); r += 1            # B17
r = inp(wa, r, 'Total debt for weights (FY25, triangulated)', 10465.0, NUM0, 'Balance-sheet residual: disclosed assets 16,460 − rolled equity 4,100 − non-debt liabilities ~1,890; = 96% of the disclosed 10.9bn facilities')  # B18
put(wa, f'A{r}', 'Equity weight E/(D+E)'); put(wa, f'B{r}', '=B17/(B17+B18)', BLACK, PCT); r += 1          # B19
put(wa, f'A{r}', 'WACC — explicit window', bold=True); put(wa, f'B{r}', '=B19*B14+(1-B19)*B16', BLACK, PCT, True); r += 1  # B20
r = inp(wa, r, 'Terminal rf (norm-built)', 0.105, PCT, "CBE's Q4-2028 inflation target 5% + 5.5pp real-rate convention")  # B21
r = inp(wa, r, 'Terminal ERP (normalised)', 0.070, PCT, 'Below the crisis-era 9.41%, toward the B-rating norm')  # B22
put(wa, f'A{r}', 'Terminal Ke'); put(wa, f'B{r}', '=B21+B13*B22', BLACK, PCT); r += 1                      # B23
r = inp(wa, r, 'Terminal Kd', 0.150, PCT, 'Egyptian long-run corporate norm 14-16%, midpoint')             # B24
put(wa, f'A{r}', 'WACC — terminal (normalized structure)', bold=True); put(wa, f'B{r}', 0.15, BLACK, PCT, True); r += 1  # B25 — real formula written once wd_term row is known
r = inp(wa, r, 'Terminal growth g', 0.05, PCT, 'Standing center 5%; grid 3-7% on Sensitivity')             # B26
r = hdr(wa, r, 'BRIDGE')                                                                                   # 27
r = inp(wa, r, 'Net debt (FY25, triangulated)', 9805.0, NUM0, 'Roll-forward from FY24 audited comparatives: drawn debt 10,465 − cash ~665; two methods agree ~9,800, facilities-as-drawn 10,235 is the upper check; range 9,120-10,360 (~±EGP 0.19/sh) — see study §1.6')  # B28
r = inp(wa, r, 'Non-controlling interests', 0.0, NUM0, 'Consolidated-vs-attributable gap ~1.1mn FY25 — immaterial')  # B29
r = hdr(wa, r, 'LENS INPUTS')                                                                              # 30
r = inp(wa, r, 'EV/EBITDA multiple (base)', 5.5, MULT, 'SWDY ~6.0x; discount for leverage/concentration')  # B31
r = inp(wa, r, 'Normalized P/E (base)', 6.5, MULT, 'SWDY 10.4x trailing; deep discount')                   # B32
r = inp(wa, r, 'Sustainable ROE (book lens)', 0.14, PCT, 'Normalized NP ~700 on ~4.6-5bn book')            # B33
r = inp(wa, r, 'Equity FY25e (book, derived)', 4100.0, NUM0, 'FY24 equity 3,600 + FY25 NP 500.3, no dividends')  # B34
r = inp(wa, r, 'Weight: DCF', 0.40, PCT); r = inp(wa, r, 'Weight: Relative', 0.20, PCT)                    # B35,B36
r = inp(wa, r, 'Weight: Normalized', 0.20, PCT); r = inp(wa, r, 'Weight: Book', 0.20, PCT)                 # B37,B38
r = hdr(wa, r, 'FORECAST DRIVERS (FY26E-FY30E)')                                                           # 39
put(wa, f'A{r}', 'Driver \\ year', bold=True)
for j, y in enumerate(YF):
    put(wa, f'{ACOLS[j]}{r}', y, BLACK, None, True, FILL_H)
r += 1                                                                                                     # grid starts 41
DRV = {}
def drv(row, label, vals, fmt=PCT, note=None):
    put(wa, f'A{row}', label)
    for j, v in enumerate(vals):
        put(wa, f'{ACOLS[j]}{row}', v, BLUE, fmt)
    if note: put(wa, f'H{row}', note, SUB, None)
    DRV[label] = row
    return row + 1
r = drv(r, 'Volume (kt)', [10.4, 11.9, 13.4, 14.8, 16.0], '0.0',
        note='Q1-26 implied ~2.4kt/qtr sets FY26E (~42% util.); recovery to 64% by FY30E on the grid-capex cycle')  # 41
r = drv(r, 'LME copper (USD/t)', [12600.0] * 5, '#,##0',
        note='Flat at the 2026 consensus avg — no house commodity view; bull/bear move it')                          # 42
r = drv(r, 'EGP/USD (avg)', [50.4, 52.0, 53.5, 55.0, 56.5], '0.0',
        note='~3%/yr crawl; inflation differential narrows as CBE targets bite')                                     # 43
r = drv(r, 'Conversion EBITDA per tonne (k EGP/t)', [40.0, 90.0, 115.0, 128.0, 135.0], '0',
        note='Hist: 111 (FY23), 146 (FY24), 182 (FY25, copper-gain inflated), 11 (Q1-26). FY30E 13.7% of price = pre-windfall 2022 norm')  # 44
r = drv(r, 'Capex (EGP mn)', [225.0, 243.0, 262.0, 283.0, 306.0], '#,##0',
        note='Maintenance ~EGP 9k per tonne of 25kt capacity, escalated ~8%/yr')                                     # 45
r = drv(r, 'D&A (% of revenue)', [0.013] * 5)                                                                        # 46
r = drv(r, 'Net working capital (% of revenue)', [1.12, 1.06, 1.00, 0.94, 0.88],
        note='From ~113% FY25e; NOT full reversion to FY24 76%')                                                     # 47
r = drv(r, 'Forward Kd path', [0.220, 0.200, 0.185, 0.168, 0.155], note='CBE easing resumption; sets the WACC glide shape')  # 48
put(wa, f'A{r}', 'Glide fraction (from the Kd path)')
for j, c in enumerate(ACOLS):
    put(wa, f'{c}{r}', f"=($C${DRV['Forward Kd path']}-{c}{DRV['Forward Kd path']})/($C${DRV['Forward Kd path']}-$G${DRV['Forward Kd path']})", BLACK, PCT)
GLIDE = r; r += 1                                                                                          # 47
put(wa, f'A{r}', 'Forward WACC (that year)')
for j, c in enumerate(ACOLS):
    put(wa, f'{c}{r}', f"=$B$20-($B$20-$B$25)*{c}{GLIDE}", BLACK, PCT)
FWD = r; r += 1                                                                                            # 48
put(wa, f'A{r}', 'Cumulative discount factor')
for j, c in enumerate(ACOLS):
    put(wa, f'{c}{r}', (f"=1/(1+{c}{FWD})" if j == 0 else f"={ACOLS[j-1]}{r}/(1+{c}{FWD})"), BLACK, '0.000')
DFR = r; r += 1                                                                                            # 49
r = hdr(wa, r, 'BASE-YEAR ANCHORS (FY2025)')                                                               # 50
r = inp(wa, r, 'Revenue FY25', 10819.0, NUM0, 'Arab Finance/Zawya FY2025 results'); R_REV25 = r - 1
r = inp(wa, r, 'Net working capital FY25 (derived)', 12245.0, NUM0, '~113% of FY25 revenue — construction in study §1.6'); R_NWC25 = r - 1
r = inp(wa, r, 'Fabrication uplift k (price/t ÷ copper cost/t)', 1.387, '0.000', 'Copper ~72% of cable price (industry norm); FY24 back-solves to 24.0kt = 96% of capacity — the validation'); R_K = r - 1
r = inp(wa, r, 'Terminal debt weight D/(D+E)', 0.40, PCT, 'NORMALIZED structure — not today’s ~60% distress weight (circular into perpetuity); conservative direction'); R_WD = r - 1
r = inp(wa, r, 'Capacity (kt/yr)', 25.0, '0.0', 'Company profile via IATF page — single-sourced, possibly parent-only; utilization indicative'); R_CAP = r - 1
# WACC-terminal formula can only be written once the wd_term row is known
wa['B25'] = f"=(1-$B${R_WD})*B23+$B${R_WD}*B24*(1-B7)"
wa['A25'] = f'WACC — terminal (normalized structure, see B{R_WD})'
wa.column_dimensions['C'].width = 11; wa.column_dimensions['H'].width = 60

# ============ SEGMENTS =======================================================
ws = sheet('Segments')
title(ws, 'Segment view — the revenue and margin build',
      'Single-segment build (power ~2/3, telecom ~1/3 of revenue — memo estimate; no segment disclosure). Forecast links to Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
SR = {}
def srow(rr, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    SR[label] = rr
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{rr}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{rr}', f, GREEN if 'Assumptions!' in str(f) else BLACK, fmt, bold=bold)
    return rr + 1
r = 6
TGX = D['tonnage']; HVX = TGX['hist_vol']; HEX = TGX['hist_ebitda_per_t']
r = srow(r, 'Volume (kt) — implied hist / driver fwd',
         [round(HVX['FY23']['vol_kt'], 1), round(HVX['FY24']['vol_kt'], 1), round(HVX['FY25']['vol_kt'], 1)],
         lambda j, c: f"=Assumptions!{ACOLS[j]}{DRV['Volume (kt)']}", '0.0', BLACK)
VOL = SR['Volume (kt) — implied hist / driver fwd']
r = srow(r, 'Utilization (÷ 25 kt capacity)', [f'=B{VOL}/Assumptions!$B${R_CAP}', f'=C{VOL}/Assumptions!$B${R_CAP}', f'=D{VOL}/Assumptions!$B${R_CAP}'],
         lambda j, c: f"={c}{VOL}/Assumptions!$B${R_CAP}", PCT, BLACK)
r = srow(r, 'LME copper (USD/t, avg)',
         [TGX['copper_hist']['FY23'], TGX['copper_hist']['FY24'], TGX['copper_hist']['FY25']],
         lambda j, c: f"=Assumptions!{ACOLS[j]}{DRV['LME copper (USD/t)']}", '#,##0', BLUE)
CU = SR['LME copper (USD/t, avg)']
r = srow(r, 'EGP/USD (avg)', [TGX['egp_hist']['FY23'], TGX['egp_hist']['FY24'], TGX['egp_hist']['FY25']],
         lambda j, c: f"=Assumptions!{ACOLS[j]}{DRV['EGP/USD (avg)']}", '0.0', BLUE)
EGPR = SR['EGP/USD (avg)']
r = srow(r, 'Price per tonne (EGP, = Cu × EGP × k)',
         [f'=B{CU}*B{EGPR}*Assumptions!$B${R_K}', f'=C{CU}*C{EGPR}*Assumptions!$B${R_K}', f'=D{CU}*D{EGPR}*Assumptions!$B${R_K}'],
         lambda j, c: f"={c}{CU}*{c}{EGPR}*Assumptions!$B${R_K}", '#,##0', BLACK)
PPT = SR['Price per tonne (EGP, = Cu × EGP × k)']
r = srow(r, 'Revenue (= volume × price/t)', [H['FY23']['rev'], H['FY24']['rev'], H['FY25']['rev']],
         lambda j, c: f"={c}{VOL}*{c}{PPT}/1000", NUM0, BLUE, True)
REV = SR['Revenue (= volume × price/t)']
r = srow(r, '  memo: power cables (~2/3, est.)', [None]*3, lambda j, c: f"={c}{REV}*2/3")
r = srow(r, '  memo: telecom & other (~1/3, est.)', [None]*3, lambda j, c: f"={c}{REV}/3")
r = srow(r, 'Conversion EBITDA per tonne (k EGP/t)',
         [round(HEX['FY23']), round(HEX['FY24']), round(HEX['FY25'])],
         lambda j, c: f"=Assumptions!{ACOLS[j]}{DRV['Conversion EBITDA per tonne (k EGP/t)']}", '0', BLACK)
EPT = SR['Conversion EBITDA per tonne (k EGP/t)']
r = srow(r, 'EBITDA (= volume × EBITDA/t)', [H['FY23']['ebitda'], H['FY24']['ebitda'], H['FY25']['ebitda']],
         lambda j, c: f"={c}{VOL}*{c}{EPT}", NUM0, BLUE, True)
EBITDA = SR['EBITDA (= volume × EBITDA/t)']
r = srow(r, 'EBITDA margin — OUTPUT', [f'=B{EBITDA}/B{REV}', f'=C{EBITDA}/C{REV}', f'=D{EBITDA}/D{REV}'],
         lambda j, c: f"={c}{EBITDA}/{c}{REV}", PCT, BLACK)
r = srow(r, 'D&A', [H['FY23']['dna'], H['FY24']['dna'], H['FY25']['dna']],
         lambda j, c: f"={c}{REV}*Assumptions!{ACOLS[j]}{DRV['D&A (% of revenue)']}")
DNA = SR['D&A']
r = srow(r, 'EBIT', [H['FY23']['ebit'], H['FY24']['ebit'], H['FY25']['ebit']],
         lambda j, c: f"={c}{EBITDA}-{c}{DNA}", NUM0, BLUE, True)
EBIT = SR['EBIT']
put(ws, f'A{r+1}', 'BOTTOM-UP TONNAGE BUILD. Historical volumes are IMPLIED (revenue ÷ [LME × EGP × k]; k = 1.387 from the '
    'industry copper-share norm; FY24 back-solves to 24.0kt = 96% of stated capacity — the validation). Historical '
    'conversion EBITDA/t: 111 (FY23) / 146 (FY24) / 182 (FY25, copper-gain inflated) / 11 (Q1-26 annualized ~9.5kt, '
    '38% utilization). Q1-26 disclosed: GM 5.7%, operating profit ~0. Forecast volumes/copper/EGP/EBITDA-per-t are '
    'Assumptions drivers; margins are outputs. Capacity single-sourced, possibly parent-only — utilization indicative.', SUB, None)
put(ws, f'A{r+2}', 'FY23 note: at the 2023 parallel rate (~38 vs official 30.7) implied volume is 19.4kt (78%) — the '
    'honest range for that year.', SUB, None)

# ============ DCF ============================================================
ws = sheet('DCF')
title(ws, 'DCF — explicit 5-year FCFF on the forward-WACC schedule',
      'Revenue -> EBITDA -> D&A -> EBIT -> NOPAT -> +D&A -> -Capex -> -dWC -> FCFF -> forward WACC -> cumulative DF -> PV. Terminal is ROIC-consistent.', 8)
for j, y in enumerate(YF):
    put(ws, f'{get_column_letter(2+j)}4', y, BLACK, None, True, FILL_H)
DC = {}
def drow(rr, label, fml, fmt=NUM0, bold=False):
    DC[label] = rr
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    for j in range(5):
        c = get_column_letter(2 + j)
        f = fml(j, c)
        put(ws, f'{c}{rr}', f, GREEN if ('Segments!' in str(f) or 'Assumptions!' in str(f)) else BLACK, fmt, bold=bold)
    return rr + 1
r = 6
r = drow(r, 'Revenue', lambda j, c: f"=Segments!{FCOLS[j]}{REV}")
r = drow(r, 'EBITDA', lambda j, c: f"=Segments!{FCOLS[j]}{EBITDA}")
r = drow(r, '− D&A', lambda j, c: f"=-Segments!{FCOLS[j]}{DNA}")
r = drow(r, 'EBIT', lambda j, c: f"=Segments!{FCOLS[j]}{EBIT}")
r = drow(r, 'NOPAT = EBIT × (1 − tax)', lambda j, c: f"={c}{DC['EBIT']}*(1-Assumptions!$B$7)")
r = drow(r, '+ D&A', lambda j, c: f"=Segments!{FCOLS[j]}{DNA}")
r = drow(r, '− Capex', lambda j, c: f"=-Assumptions!{ACOLS[j]}{DRV['Capex (EGP mn)']}")
r = drow(r, 'Net working capital', lambda j, c: f"={c}{DC['Revenue']}*Assumptions!{ACOLS[j]}{DRV['Net working capital (% of revenue)']}")
NWCR = DC['Net working capital']
r = drow(r, '− Δ working capital', lambda j, c: (f"=-(B{NWCR}-Assumptions!$B${R_NWC25})" if j == 0
                                                 else f"=-({c}{NWCR}-{get_column_letter(2+j-1)}{NWCR})"))
r = drow(r, 'Free cash flow to firm',
         lambda j, c: f"={c}{DC['NOPAT = EBIT × (1 − tax)']}+{c}{DC['+ D&A']}+{c}{DC['− Capex']}+{c}{DC['− Δ working capital']}", NUM0, True)
r = drow(r, 'Forward WACC (that year)', lambda j, c: f"=Assumptions!{ACOLS[j]}{FWD}", PCT)
r = drow(r, 'Cumulative discount factor', lambda j, c: f"=Assumptions!{ACOLS[j]}{DFR}", '0.000')
r = drow(r, 'PV of FCFF', lambda j, c: f"={c}{DC['Free cash flow to firm']}*{c}{DC['Cumulative discount factor']}", NUM0, True)
r += 1
def dline(rr, label, fml, fmt=NUM0, bold=False, font=BLACK):
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    put(ws, f'B{rr}', fml, font, fmt, bold=bold)
    return rr + 1
r = dline(r, 'Σ PV of explicit FCFF (FY26–30E)', f"=SUM(B{DC['PV of FCFF']}:F{DC['PV of FCFF']})", NUM0, True); SPV = r - 1
r = dline(r, 'Terminal ROIC (NOPAT ÷ [NWC + 5% of revenue])',
          f"=F{DC['NOPAT = EBIT × (1 − tax)']}/(F{NWCR}+0.05*F{DC['Revenue']})", PCT); ROIC = r - 1
r = dline(r, 'Terminal reinvestment rate = g ÷ ROIC', f"=Assumptions!$B$26/B{ROIC}", PCT); RR_ = r - 1
r = dline(r, 'Terminal FCFF = NOPAT × (1+g) × (1 − RR)',
          f"=F{DC['NOPAT = EBIT × (1 − tax)']}*(1+Assumptions!$B$26)*(1-B{RR_})"); FT = r - 1
r = dline(r, 'Terminal value (at terminal WACC)', f"=B{FT}/(Assumptions!$B$25-Assumptions!$B$26)"); TVR = r - 1
r = dline(r, 'PV of terminal value (year-5 factor)', f"=B{TVR}*F{DC['Cumulative discount factor']}"); PVT = r - 1
r = dline(r, 'Enterprise value', f"=B{SPV}+B{PVT}", NUM0, True); EVR = r - 1
r = dline(r, 'Terminal value as % of EV', f"=B{PVT}/B{EVR}", PCT)
r = dline(r, 'less: net debt', "=-Assumptions!$B$28", NUM0, False, GREEN)
r = dline(r, 'less: non-controlling interests', "=-Assumptions!$B$29", NUM0, False, GREEN)
r = dline(r, 'Equity value — INTRINSIC (may be negative)', f"=B{EVR}-Assumptions!$B$28-Assumptions!$B$29", NUM0, True); AEQ = r - 1
r = dline(r, 'Equity value — floored at zero (limited liability)', f"=MAX(B{AEQ},0)", NUM0, True); AEQF = r - 1
r = dline(r, 'Intrinsic per share (EGP, unfloored)', f"=B{AEQ}/Assumptions!$B$6", PX)
r = dline(r, 'Fair value per share (EGP) — floored', f"=MAX(B{AEQF}/Assumptions!$B$6,0.01)", PX, True); DPS_ = r - 1
r = dline(r, 'Upside / (downside) vs spot', f"=B{DPS_}/Assumptions!$B$5-1", PCT, True)
put(ws, f'A{r+1}', 'The glide: each year at its own forward WACC (22.6% -> 14.1%), shape from the forward Kd path; '
    'the terminal is capitalised at the terminal WACC and discounted at the identical year-5 cumulative factor.', SUB, None)

# ============ INCOME STATEMENT ==============================================
ws = sheet('Income Statement')
title(ws, 'Income statement (EGP mn, consolidated)',
      'FY23–FY25 as sourced/derived (blue; (d)-flagged in the study); FY26E–FY30E formulas. Interest on OPENING net debt at that year’s forward Kd; tax on positive EBT net of loss carryforward.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
IS = {}
def irow(rr, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    IS[label] = rr
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{rr}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{rr}', f, GREEN if ('!' in str(f)) else BLACK, fmt, bold=bold)
    return rr + 1
r = 6
r = irow(r, 'Revenue', [H['FY23']['rev'], H['FY24']['rev'], H['FY25']['rev']],
         lambda j, c: f"=Segments!{c}{REV}", NUM0, BLUE, True)
r = irow(r, 'EBITDA', [H['FY23']['ebitda'], H['FY24']['ebitda'], H['FY25']['ebitda']],
         lambda j, c: f"=Segments!{c}{EBITDA}")
r = irow(r, 'D&A', [-H['FY23']['dna'], -H['FY24']['dna'], -H['FY25']['dna']],
         lambda j, c: f"=-Segments!{c}{DNA}")
r = irow(r, 'EBIT', [H['FY23']['ebit'], H['FY24']['ebit'], H['FY25']['ebit']],
         lambda j, c: f"=Segments!{c}{EBIT}", NUM0, BLUE, True)
r = irow(r, 'Net finance cost', [H['FY23']['fin'], H['FY24']['fin'], H['FY25']['fin']],
         lambda j, c: (f"=-Assumptions!{ACOLS[j]}{DRV['Forward Kd path']}*Assumptions!$B$28" if j == 0
                       else f"=-Assumptions!{ACOLS[j]}{DRV['Forward Kd path']}*'Balance Sheet'!{['D','E','F','G'][j-1]}11"))
FIN = IS['Net finance cost']
r = irow(r, 'Earnings before tax', [H['FY23']['ebt'], H['FY24']['ebt'], H['FY25']['ebt']],
         lambda j, c: f"={c}{IS['EBIT']}+{c}{FIN}", NUM0, BLACK, True)
EBTR = IS['Earnings before tax']
r = irow(r, 'Cumulative EBT since FY26E (memo)', [None]*3,
         lambda j, c: f"=SUM($E${EBTR}:{c}{EBTR})")
CUM = IS['Cumulative EBT since FY26E (memo)']
r = irow(r, 'Income tax (22.5% of positive cumulative EBT, incremental)',
         [-(H['FY23']['ebt']-H['FY23']['np']), -(H['FY24']['ebt']-H['FY24']['np']), -(H['FY25']['ebt']-H['FY25']['np'])],
         lambda j, c: (f"=-Assumptions!$B$7*MAX(0,{c}{CUM})" if j == 0
                       else f"=-(Assumptions!$B$7*MAX(0,{c}{CUM})-Assumptions!$B$7*MAX(0,{get_column_letter(1+2+j-1)}{CUM}))"))
TAXR = IS['Income tax (22.5% of positive cumulative EBT, incremental)']
r = irow(r, 'Net profit (attributable)', [H['FY23']['np'], H['FY24']['np'], H['FY25']['np']],
         lambda j, c: f"={c}{EBTR}+{c}{TAXR}", NUM0, BLUE, True)
NP = IS['Net profit (attributable)']
r = irow(r, 'EPS (EGP)', [f'=B{NP}/Assumptions!$B$6', f'=C{NP}/Assumptions!$B$6', f'=D{NP}/Assumptions!$B$6'],
         lambda j, c: f"={c}{NP}/Assumptions!$B$6", PX, BLACK)
put(ws, f'A{r+1}', 'Historical net finance cost is (d)-derived (FY24 via the disclosed 2.0x coverage; FY23/FY25 back-solved '
    'to the reported NP). Loss carryforward per Egyptian tax law (5yr) modelled via cumulative-EBT tax.', SUB, None)

# ============ BALANCE SHEET =================================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — condensed invested-capital layout (EGP mn)',
      'IC = NWC + PP&E + other assets; financed by net debt + equity + flat non-debt liabilities. Check row is zero by construction.', 10)
for j, h in enumerate(['', 'FY24', 'FY25e'] + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
BSCOLS = ['D', 'E', 'F', 'G', 'H']   # forecast cols on this sheet start at D
BS = {}
def brow(rr, label, h24, h25, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    BS[label] = rr
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    if h24 is not None: put(ws, f'B{rr}', h24, hfont, fmt, bold=bold)
    if h25 is not None: put(ws, f'C{rr}', h25, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(BSCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{rr}', f, GREEN if '!' in str(f) else BLACK, fmt, bold=bold)
    return rr + 1
r = 6
r = brow(r, 'Net working capital', 10500.0, f'=Assumptions!B{R_NWC25}',
         lambda j, c: f"=DCF!{get_column_letter(2+j)}{NWCR}")
r = brow(r, 'PP&E (roll: + capex - D&A)', 650.0, 680.0, None)
PPE = BS['PP&E (roll: + capex - D&A)']
for j, c in enumerate(BSCOLS):
    prev = 'C' if j == 0 else BSCOLS[j - 1]
    ws[f'{c}{PPE}'] = (f"={prev}{PPE}-DCF!{get_column_letter(2+j)}{DC['− Capex']}"
                       f"-Segments!{FCOLS[j]}{DNA}")
    ws[f'{c}{PPE}'].font = GREEN; ws[f'{c}{PPE}'].number_format = NUM0
r = brow(r, 'Other assets (held flat)', 3000.0, 2980.0, lambda j, c: f"={'C' if j==0 else BSCOLS[j-1]}{r-1}")
OTH = BS['Other assets (held flat)']
for j, c in enumerate(BSCOLS):
    ws[f'{c}{OTH}'] = f"={'C' if j==0 else BSCOLS[j-1]}{OTH}"; ws[f'{c}{OTH}'].font = BLACK; ws[f'{c}{OTH}'].number_format = NUM0
r = brow(r, 'Invested capital (net of payables)', None, None,
         lambda j, c: f"=SUM({c}{BS['Net working capital']}:{c}{OTH})", NUM0, BLACK, True)
IC = BS['Invested capital (net of payables)']
for col, in zip(['B', 'C']):
    ws[f'{col}{IC}'] = f"=SUM({col}{BS['Net working capital']}:{col}{OTH})"
    ws[f'{col}{IC}'].font = Font(bold=True); ws[f'{col}{IC}'].number_format = NUM0
r += 1
r = brow(r, 'Net debt (schedule: opening − equity FCF)', 9805.0, '=Assumptions!B28',
         lambda j, c: f"={'Assumptions!$B$28' if j==0 else BSCOLS[j-1]+str(r-1)}-'Cash Flow'!{get_column_letter(3+j)}12")
ND = BS['Net debt (schedule: opening − equity FCF)']
for j, c in enumerate(BSCOLS):
    prev = "Assumptions!$B$28" if j == 0 else f"{BSCOLS[j-1]}{ND}"
    ws[f'{c}{ND}'] = f"={prev}-'Cash Flow'!{get_column_letter(3+j)}12"
    ws[f'{c}{ND}'].font = GREEN; ws[f'{c}{ND}'].number_format = NUM0
r = brow(r, 'Equity (roll: + net profit, no dividends)', '=Assumptions!B34-500.31', '=Assumptions!B34',
         lambda j, c: f"={'C' if j==0 else BSCOLS[j-1]}{r-1}+'Income Statement'!{FCOLS[j]}{NP}")
EQR2 = BS['Equity (roll: + net profit, no dividends)']
for j, c in enumerate(BSCOLS):
    prev = 'C' if j == 0 else BSCOLS[j - 1]
    ws[f'{c}{EQR2}'] = f"={prev}{EQR2}+'Income Statement'!{FCOLS[j]}{NP}"
    ws[f'{c}{EQR2}'].font = GREEN; ws[f'{c}{EQR2}'].number_format = NUM0
r = brow(r, 'Non-debt liabilities & other (flat plug at FY25e)', None, None, None)
PLUG = BS['Non-debt liabilities & other (flat plug at FY25e)']
ws[f'B{PLUG}'] = f"=B{IC}-B{ND}-B{EQR2}"; ws[f'B{PLUG}'].number_format = NUM0
ws[f'C{PLUG}'] = f"=C{IC}-C{ND}-C{EQR2}"; ws[f'C{PLUG}'].number_format = NUM0
for j, c in enumerate(BSCOLS):
    ws[f'{c}{PLUG}'] = f"=$C{PLUG}"; ws[f'{c}{PLUG}'].number_format = NUM0
r = brow(r, 'Total financing', None, None,
         lambda j, c: f"={c}{ND}+{c}{EQR2}+{c}{PLUG}", NUM0, BLACK, True)
TF = BS['Total financing']
for col in ['B', 'C']:
    ws[f'{col}{TF}'] = f"={col}{ND}+{col}{EQR2}+{col}{PLUG}"; ws[f'{col}{TF}'].font = Font(bold=True); ws[f'{col}{TF}'].number_format = NUM0
r = brow(r, 'Balance check (IC − financing)', None, None,
         lambda j, c: f"={c}{IC}-{c}{TF}", NUM0, BLACK, True)
CHK = BS['Balance check (IC − financing)']
for col in ['B', 'C']:
    ws[f'{col}{CHK}'] = f"={col}{IC}-{col}{TF}"; ws[f'{col}{CHK}'].number_format = NUM0
put(ws, f'A{r+1}', 'FY24 column: SWS-sourced anchors (debt 9,000/cash 828/equity 3,600); FY25e column: derived (study §1.6). '
    'The forecast check row is zero by construction: PP&E rolls with capex − D&A, NWC with its driver, net debt with the cash-flow sheet, '
    'equity with net profit, and the non-debt-liabilities plug is held flat.', SUB, None)

# ============ CASH FLOW =====================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow (EGP mn, forecast)',
      'Equity free cash flow feeds the net-debt schedule (surplus repays debt; deficit is drawn). FCFF memo ties to the DCF.', 9)
for j, h in enumerate([''] + YF):
    put(ws, f'{get_column_letter(2+j)}4', h, BLACK, None, True, FILL_H)
CF = {}
CFCOLS = ['C', 'D', 'E', 'F', 'G']
def crow(rr, label, ffml, fmt=NUM0, bold=False):
    CF[label] = rr
    put(ws, f'A{rr}', label, BLACK, None, bold=bold)
    for j, c in enumerate(CFCOLS):
        f = ffml(j, c)
        if f is not None:
            put(ws, f'{c}{rr}', f, GREEN if '!' in str(f) else BLACK, fmt, bold=bold)
    return rr + 1
r = 6
r = crow(r, 'Net profit', lambda j, c: f"='Income Statement'!{FCOLS[j]}{NP}")
r = crow(r, '+ D&A', lambda j, c: f"=Segments!{FCOLS[j]}{DNA}")
r = crow(r, '− Δ working capital', lambda j, c: f"=DCF!{get_column_letter(2+j)}{DC['− Δ working capital']}")
r = crow(r, '− Capex', lambda j, c: f"=DCF!{get_column_letter(2+j)}{DC['− Capex']}")
r = crow(r, 'Operating − investing cash flow', lambda j, c: f"=SUM({c}{CF['Net profit']}:{c}{CF['− Capex']})", NUM0, True)
r = crow(r, 'Dividends', lambda j, c: 0)
r = crow(r, 'Equity free cash flow (repays / draws net debt)',
         lambda j, c: f"={c}{CF['Operating − investing cash flow']}+{c}{CF['Dividends']}", NUM0, True)   # row 12
r += 1
r = crow(r, 'memo: FCFF (ties to DCF row)', lambda j, c: f"=DCF!{get_column_letter(2+j)}{DC['Free cash flow to firm']}")
r = crow(r, 'memo: net debt, closing', lambda j, c: f"='Balance Sheet'!{BSCOLS[j]}{ND}")
put(ws, f'A{r+1}', 'No dividends modelled — the company has never paid one and the balance sheet argues against a first.', SUB, None)

# ============ SOTP Bridge ====================================================
ws = sheet('SOTP Bridge')
title(ws, 'EV → equity bridge', 'Single-business company: the bridge is the DCF EV less net claims. Links to DCF / Assumptions.', 7)
rows_ = [
 ('Core enterprise value (DCF)', f"=DCF!B{EVR}", NUM0, True),
 ('+ Surplus / non-core assets', 0.0, NUM0, False),
 ('− Net debt (FY25: triangulated — study §1.6)', "=-Assumptions!B28", NUM0, False),
 ('− Non-controlling interests', "=-Assumptions!B29", NUM0, False),
 ('Equity value — intrinsic (may be negative)', f"=B5+B6+B7+B8", NUM0, True),
 ('Equity value — floored (limited liability)', f"=MAX(B9,0)", NUM0, True),
 ('per share (EGP) — floored', f"=MAX(B10/Assumptions!B6,0.01)", PX, True),
 ('Upside / (downside) vs spot', f"=B11/Assumptions!B5-1", PCT, True),
]
r = 5
for a, b, fmt, bold in rows_:
    put(ws, f'A{r}', a, BLACK, None, bold=bold)
    put(ws, f'B{r}', b, GREEN if isinstance(b, str) else BLUE if b else BLACK, fmt, bold=bold)
    r += 1
r += 1
put(ws, f'A{r}', 'Market-implied read (the same bridge run backwards)', BLACK, None, True, FILL_H); r += 1
put(ws, f'A{r}', 'Market cap at spot'); put(ws, f'B{r}', '=Assumptions!B17', GREEN, NUM0); r += 1
put(ws, f'A{r}', '+ Net debt = market-implied EV'); put(ws, f'B{r}', f'=B{r-1}+Assumptions!B28', BLACK, NUM0); r += 1
put(ws, f'A{r}', 'Market-implied EV / FY27E EBITDA'); put(ws, f'B{r}', f"=B{r-1}/Segments!F{EBITDA}", BLACK, MULT); r += 1
put(ws, f'A{r}', 'Market-implied EV / FY25 EBITDA (windfall-vintage)'); put(ws, f'B{r}', f"=B{r-2}/Segments!D{EBITDA}", BLACK, MULT); r += 1
put(ws, f'A{r+1}', 'The market pays a peer-level multiple on windfall-vintage EBITDA — the study’s central observation.', SUB, None)

# ============ Relative & Normalized =========================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative & normalized-earnings lenses', 'Links to Assumptions / Segments / Balance Sheet.', 7)
r = 5
put(ws, f'A{r}', 'Lens', BLACK, None, True, FILL_H); put(ws, f'B{r}', 'Workings', BLACK, None, True, FILL_H)
put(ws, f'C{r}', 'EGP/share', BLACK, None, True, FILL_H); r += 1
put(ws, f'A{r}', 'Relative EV/EBITDA'); put(ws, f'B{r}', 'FY27E EBITDA × multiple − net debt end-FY26E, /sh', SUB, None)
put(ws, f'C{r}', f"=MAX((Assumptions!B31*Segments!F{EBITDA}-'Balance Sheet'!D{ND})/Assumptions!B6,0.05)", GREEN, PX); RELC = r; r += 1
put(ws, f'A{r}', 'Normalized earnings power'); put(ws, f'B{r}', 'FY28E EBIT − 15% money on 6,000 mid-cycle ND, × (1−t), × P/E, /sh', SUB, None)
put(ws, f'C{r}', f"=(Segments!G{EBIT}-Assumptions!B24*6000)*(1-Assumptions!B7)*Assumptions!B32/Assumptions!B6", GREEN, PX); NORMC = r; r += 1
put(ws, f'A{r}', '  normalized EPS (EGP)'); put(ws, f'C{r}', f"=(Segments!G{EBIT}-Assumptions!B24*6000)*(1-Assumptions!B7)/Assumptions!B6", GREEN, PX); r += 1
put(ws, f'A{r}', 'Book lens: justified P/B × BVPS'); put(ws, f'B{r}', '(ROE−g)/(Ke_term−g) × FY25e book/sh', SUB, None)
put(ws, f'C{r}', f"=MAX((Assumptions!B33-Assumptions!B26)/(Assumptions!B23-Assumptions!B26),0.1)*Assumptions!B34/Assumptions!B6", GREEN, PX); BOOKC = r; r += 1
put(ws, f'A{r+1}', 'P/B at spot (memo)'); put(ws, f'C{r+1}', "=Assumptions!B5/(Assumptions!B34/Assumptions!B6)", BLACK, MULT)
ws.column_dimensions['B'].width = 56

# ============ Summary ========================================================
ws = sheet('Summary')
title(ws, 'Valuation summary — Electro Cable Egypt (EGX: ELEC)',
      'Four-lens fair value vs spot. Base cells link live; bear/bull are engine scenario outputs (study §1.5).', 7)
r = 5
for j, h in enumerate(['Bear', 'Base', 'Bull', 'Weight']):
    put(ws, f'{get_column_letter(2+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
lens_rows = [
 ('FCFF DCF (glide schedule)', L['dcf']['bear'], f"=DCF!B{DPS_}", L['dcf']['bull'], '=Assumptions!B35'),
 ('Relative (EV/EBITDA)', L['relative']['bear'], f"='Relative & Normalized'!C{RELC}", L['relative']['bull'], '=Assumptions!B36'),
 ('Normalized earnings', L['normalized']['bear'], f"='Relative & Normalized'!C{NORMC}", L['normalized']['bull'], '=Assumptions!B37'),
 ('Book / replacement', L['book']['bear'], f"='Relative & Normalized'!C{BOOKC}", L['book']['bull'], '=Assumptions!B38'),
]
first = r
for nm, be, ba, bu, w in lens_rows:
    put(ws, f'A{r}', nm)
    put(ws, f'B{r}', round(be, 2), BLUE, PX)
    put(ws, f'C{r}', ba, GREEN, PX)
    put(ws, f'D{r}', round(bu, 2), BLUE, PX)
    put(ws, f'E{r}', w, GREEN, PCT)
    r += 1
put(ws, f'A{r}', 'Weighted central', BLACK, None, True)
for col in 'BCD':
    put(ws, f'{col}{r}', f"=SUMPRODUCT({col}{first}:{col}{r-1},$E${first}:$E${r-1})", BLACK, PX, True)
WC = r; r += 2
put(ws, f'A{r}', 'Spot price (EGP)'); put(ws, f'B{r}', '=Assumptions!B5', GREEN, PX); r += 1
put(ws, f'A{r}', 'Upside / (downside) to central base'); put(ws, f'B{r}', f'=C{WC}/Assumptions!B5-1', BLACK, PCT); r += 2
put(ws, f'A{r}', 'Read: meaningfully overvalued — the price capitalises devaluation-era earnings; the swing factors are '
    'working-capital collection, margin normalisation, and the easing path.', SUB, None)

# ============ Fundamental Valuation =========================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — football-field data', 'Bear/base/bull per lens (links to Summary).', 7)
r = 5
for j, h in enumerate(['Lens', 'Bear', 'Base', 'Bull']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for i, nm in enumerate(['FCFF DCF (glide schedule)', 'Relative (EV/EBITDA)', 'Normalized earnings',
                        'Book / replacement', 'Weighted central']):
    put(ws, f'A{r}', nm, BLACK, None, i == 4)
    for j, col in enumerate('BCD'):
        put(ws, f'{col}{r}', f'=Summary!{col}{first + i}', GREEN, PX, i == 4)
    r += 1
put(ws, f'A{r}', 'Spot'); put(ws, f'B{r}', '=Assumptions!B5', GREEN, PX)

# ============ Summary Financials ============================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials (EGP mn)', 'Every forecast cell links to the statement sheets; history as sourced/derived.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
r = 6
for nm, hvals, sh_, rr in [
    ('Revenue', [H['FY23']['rev'], H['FY24']['rev'], H['FY25']['rev']], 'Income Statement', IS['Revenue']),
    ('EBITDA', [H['FY23']['ebitda'], H['FY24']['ebitda'], H['FY25']['ebitda']], 'Income Statement', IS['EBITDA']),
    ('EBIT', [H['FY23']['ebit'], H['FY24']['ebit'], H['FY25']['ebit']], 'Income Statement', IS['EBIT']),
    ('Net profit (attributable)', [H['FY23']['np'], H['FY24']['np'], H['FY25']['np']], 'Income Statement', NP),
]:
    put(ws, f'A{r}', nm)
    for j, col in enumerate(['B', 'C', 'D']):
        put(ws, f'{col}{r}', hvals[j], BLUE, NUM0)
    for j, col in enumerate(FCOLS):
        put(ws, f'{col}{r}', f"='{sh_}'!{col}{rr}", GREEN, NUM0)
    r += 1
put(ws, f'A{r}', 'Net debt (closing)')
put(ws, f'D{r}', '=Assumptions!B28', GREEN, NUM0)
for j, col in enumerate(FCOLS):
    put(ws, f'{col}{r}', f"='Balance Sheet'!{BSCOLS[j]}{ND}", GREEN, NUM0)
r += 1
put(ws, f'A{r}', 'Equity (attributable)')
put(ws, f'D{r}', '=Assumptions!B34', GREEN, NUM0)
for j, col in enumerate(FCOLS):
    put(ws, f'{col}{r}', f"='Balance Sheet'!{BSCOLS[j]}{EQR2}", GREEN, NUM0)
r += 1
put(ws, f'A{r}', 'FCFF (forecast)')
for j, col in enumerate(FCOLS):
    put(ws, f'{col}{r}', f"=DCF!{get_column_letter(2+j)}{DC['Free cash flow to firm']}", GREEN, NUM0)

# ============ Monte Carlo ====================================================
ws = sheet('Monte Carlo')
title(ws, 'Monte Carlo — engine outputs (carry-anchored YZ-HAR-t)',
      '50,000 paths · Student-t(6) · width calibrated on the 30-name Egyptian panel · drift = local carry (19.50%) · values computed by the engine, not a sheet simulation.', 8)
pr = D['mc']['prob_read']; p1 = D['mc']['pct1']; p3 = D['mc']['pct3']
r = 5
put(ws, f'A{r}', 'The probability read (3 months)', BLACK, None, True, FILL_G); r += 1
for nm, v, fmt in [
    ('P(price above spot)', pr['p_above'], PCT),
    ('P(+10%) vs P(−10%) — odds', f"{pr['p_up10']*100:.0f}% vs {pr['p_dn10']*100:.0f}% · {pr['odds']:.1f}:1", '@'),
    ('Median level (EGP) and % move', f"{pr['median']:.2f} ({pr['med_move']*100:+.1f}%)", '@'),
    ('50% band (25th–75th)', f"{pr['band50'][0]:.2f} – {pr['band50'][1]:.2f}", '@'),
    ('Touch(+10%) / touch(−10%)', f"{pr['touch_up10']*100:.0f}% / {pr['touch_dn10']*100:.0f}%", '@')]:
    put(ws, f'A{r}', nm); put(ws, f'B{r}', v, BLACK, fmt); r += 1
r += 1
put(ws, f'A{r}', 'Percentile map (EGP/share)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for tag, q in [('1 month', p1), ('3 months', p3)]:
    put(ws, f'A{r}', tag)
    for j, p in enumerate(['5', '25', '50', '75', '95']):
        put(ws, f'{get_column_letter(2+j)}{r}', round(q[p], 2), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Level-touch ladder (probability of touching by horizon)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Level (EGP)', '1 month', '3 months']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for Lv, tv in D['mc']['touch'].items():
    put(ws, f'A{r}', float(Lv), BLACK, PX)
    put(ws, f'B{r}', tv['t1'], BLACK, PCT); put(ws, f'C{r}', tv['t3'], BLACK, PCT); r += 1
r += 1
put(ws, f'A{r}', 'Probability zones (3 months)', BLACK, None, True, FILL_H); r += 1
for (zn, rge), z in zip([('Deep downside', '< 1.80'), ('Lower band', '1.80–2.05'), ('Around spot', '2.05–2.35'),
                          ('Upper band', '2.35–2.70'), ('Strong upside', '> 2.70')], D['mc']['zones']):
    put(ws, f'A{r}', zn); put(ws, f'B{r}', rge, BLACK, '@'); put(ws, f'C{r}', z, BLACK, PCT); r += 1

# ============ Sensitivity ====================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — engine grids (values) + the live base cell',
      'Grids are engine outputs under the full glide re-computation; the base cell reprices live off Assumptions.', 8)
r = 5
put(ws, f'A{r}', 'Live base DCF (links to the DCF sheet)'); put(ws, f'B{r}', f"=DCF!B{DPS_}", GREEN, PX); r += 2
put(ws, f'A{r}', 'Terminal WACC × terminal g (EGP/share)', BLACK, None, True, FILL_H); r += 1
S = D['sens_wg']
put(ws, f'A{r}', 'Terminal WACC \\ g', BLACK, None, True, FILL_H)
for j, g in enumerate(S['g_grid']):
    put(ws, f'{get_column_letter(2+j)}{r}', g, BLACK, PCT, True, FILL_H)
r += 1
for i, w in enumerate(S['wacc_grid']):
    put(ws, f'A{r}', w, BLACK, PCT)
    for j in range(len(S['g_grid'])):
        put(ws, f'{get_column_letter(2+j)}{r}', round(S['table'][i][j], 2), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Explicit-window WACC × terminal WACC (EGP/share)', BLACK, None, True, FILL_H); r += 1
S2 = D['sens_expl']
put(ws, f'A{r}', 'Explicit \\ terminal', BLACK, None, True, FILL_H)
for j, w in enumerate(S2['term_grid']):
    put(ws, f'{get_column_letter(2+j)}{r}', w, BLACK, PCT, True, FILL_H)
r += 1
for i, w in enumerate(S2['expl_grid']):
    put(ws, f'A{r}', w, BLACK, PCT)
    for j in range(len(S2['term_grid'])):
        put(ws, f'{get_column_letter(2+j)}{r}', round(S2['table'][i][j], 2), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'EBITDA-margin shift × terminal NWC intensity (EGP/share)', BLACK, None, True, FILL_H); r += 1
S3 = D['sens_mn']
put(ws, f'A{r}', 'Margin shift \\ NWC %', BLACK, None, True, FILL_H)
for j, n in enumerate(S3['nwc_grid']):
    put(ws, f'{get_column_letter(2+j)}{r}', n, BLACK, PCT, True, FILL_H)
r += 1
for i, m in enumerate(S3['margin_grid']):
    put(ws, f'A{r}', m, BLUE, '+0.0%;-0.0%;"base"')
    for j in range(len(S3['nwc_grid'])):
        put(ws, f'{get_column_letter(2+j)}{r}', round(S3['table'][i][j], 2), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Beta grid (Ke / WACCs / DCF)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Beta', 'Ke', 'WACC explicit', 'WACC terminal', 'DCF (EGP/sh)']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for b in D['sens_beta']:
    put(ws, f'A{r}', b['beta'], BLACK, PX)
    put(ws, f'B{r}', b['ke'], BLACK, PCT); put(ws, f'C{r}', b['wacc_exp'], BLACK, PCT)
    put(ws, f'D{r}', b['wacc_term'], BLACK, PCT); put(ws, f'E{r}', round(b['dcf'], 2), BLACK, PX)
    r += 1

# ============ Per-Share & Ratios ============================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share & ratios — the standing dashboard', 'Links to statements / Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
allc = ['B', 'C', 'D'] + FCOLS
r = 6
def prow(rr, label, fml, fmt=PX, cols=allc):
    put(ws, f'A{rr}', label)
    for col in cols:
        f = fml(col)
        if f is not None:
            put(ws, f'{col}{rr}', f, BLACK, fmt)
    return rr + 1
r = prow(r, 'EPS (EGP)', lambda c: f"='Income Statement'!{c}{NP}/Assumptions!$B$6")
r = prow(r, 'P/E at spot (×)', lambda c: f"=IF('Income Statement'!{c}{NP}>0,Assumptions!$B$5/('Income Statement'!{c}{NP}/Assumptions!$B$6),\"n.m.\")")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'EBITDA margin', lambda c: f"='Income Statement'!{c}{IS['EBITDA']}/'Income Statement'!{c}{IS['Revenue']}", PCT)
r = prow(r, 'Net margin', lambda c: f"='Income Statement'!{c}{NP}/'Income Statement'!{c}{IS['Revenue']}", PCT)
r = prow(r, 'Revenue YoY', lambda c: (f"='Income Statement'!{c}{IS['Revenue']}/'Income Statement'!{chr(ord(c)-1)}{IS['Revenue']}-1" if c != 'B' else None), PCT)
r = prow(r, 'Book value / share (EGP)', lambda c: (f"='Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{EQR2}/Assumptions!$B$6" if c in FCOLS else (f"=Assumptions!$B$34/Assumptions!$B$6" if c == 'D' else None)))
r = prow(r, 'P/B at spot (×)', lambda c: (f"=Assumptions!$B$5/('Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{EQR2}/Assumptions!$B$6)" if c in FCOLS else (f"=Assumptions!$B$5/(Assumptions!$B$34/Assumptions!$B$6)" if c == 'D' else None)))
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Net debt / EBITDA (×)', lambda c: (f"='Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{ND}/'Income Statement'!{c}{IS['EBITDA']}" if c in FCOLS else (f"=Assumptions!$B$28/'Income Statement'!D{IS['EBITDA']}" if c == 'D' else None)))
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Net debt / equity', lambda c: (f"='Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{ND}/'Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{EQR2}" if c in FCOLS else (f"=Assumptions!$B$28/Assumptions!$B$34" if c == 'D' else None)), PCT)
r = prow(r, 'ROE (attributable)', lambda c: (f"='Income Statement'!{c}{NP}/'Balance Sheet'!{BSCOLS[FCOLS.index(c)]}{EQR2}" if c in FCOLS else None), PCT)
put(ws, f'A{r+1}', 'Historical per-share ratios beyond EPS are omitted where the year-end balance sheet was never published (flagged gap).', SUB, None)

# ============ Peer & Sector =================================================
ws = sheet('Peer & Sector')
title(ws, 'Peer set & sector', 'One giant (Elsewedy) above a tier of mid-caps competing on price and local content in tenders.', 7)
r = 5
for j, h in enumerate(['Name', 'Mkt cap', 'P/E', 'EV/EBITDA', 'D/E', 'Note']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for row_ in [
 ('El Sewedy Electric (EGX: SWDY)', 'EGP 196bn', '10.4x', '~6.0x', 'moderate', 'FY25 rev EGP 281bn, NP 17.3bn; W&C segment +66% — the demand proof'),
 ('Riyadh Cables (Tadawul: 4142)', 'SAR 18.0bn', '18.0x', '15.0x', '0.27x', 'Same industry, clean balance sheet, different market'),
 ('Electro Cable Egypt (ELEC)', 'EGP 7.3bn', '14.6x trailing', '~5.3x FY25e', '~2.5x', 'Loss-making 1Q26; ~5.5x on FY24 windfall EPS'),
 ('Giza Cables (private)', '—', '—', '—', '—', 'No public financials'),
]:
    for j, v in enumerate(row_):
        put(ws, f'{get_column_letter(1+j)}{r}', v, BLACK, None)
    r += 1
ws.column_dimensions['F'].width = 56
r += 1
put(ws, f'A{r}', 'Sector: EETC EGP 45bn FY25/26 transmission plan + EU €690mn package + Egypt–Saudi interconnector (95% complete) '
    'drive demand; LME-linked pass-through pricing keeps margins thin; copper +51% y/y strains working capital at 20%+ funding.', SUB, None)
put(ws, f'A{r+1}', 'Analyst context: none exists — ELEC is uncovered; no rating and no target here either (house rule).', SUB, None)

# ============ sheet order ====================================================
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge', 'Segments',
         'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet', 'Cash Flow',
         'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in order]
wb.save(os.path.join(HERE, 'ELEC_Valuation_Model_05082026_public.xlsx'))
print('xlsx saved:', wb.sheetnames)
