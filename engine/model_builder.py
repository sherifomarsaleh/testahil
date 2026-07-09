"""
Testahil — Excel model builder (canonical 16-sheet skeleton).

Emulates the TMPV workbook structure: a single blue-cell Assumptions layer that every downstream
sheet links back to, so changing one input reprices the whole model. The DCF sheet is fully wired
(Revenue -> ... -> PV of FCFF in the required row order, terminal value, WACC from Ke = rf + ERP*beta,
TV as % of EV). The Monte Carlo sheet is populated from an MCResult. Income Statement / Balance
Sheet / Cash Flow carry the 3-year-historical + 5-year-forecast column frame for the analyst to
fill; formula values are recalculated by LibreOffice in study_runner.

build_model(path, ticker, spot, shares, mc_result, assumptions=None) -> path
"""
from __future__ import annotations
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")          # hardcoded inputs (change for scenarios)
GREEN = Font(color="008000")         # cross-sheet links
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14, color="0F766E")
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="0F766E")
INPUT_FILL = PatternFill("solid", fgColor="FFF7E6")
THIN = Side(style="thin", color="DDDDDD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEETS = [
    "READ FIRST", "Summary", "Fundamental Valuation", "Assumptions", "Primary Lens",
    "Segments", "Relative & Normalized", "DCF", "Income Statement", "Balance Sheet",
    "Cash Flow", "Summary Financials", "Monte Carlo", "Sensitivity", "Per-Share & Ratios",
    "Peer & Sector",
]

DEFAULT_ASSUMPTIONS = {
    "rf": 0.14, "erp": 0.06, "beta": 1.0, "tax": 0.225, "terminal_g": 0.05,
    "rev_base": 1000.0, "g1": 0.15, "g2": 0.12, "g3": 0.10, "g4": 0.08, "g5": 0.06,
    "ebitda_margin": 0.30, "da_pct": 0.04, "capex_pct": 0.05, "wc_pct": 0.03, "net_debt": 0.0,
    "cap_rate": 0.09, "nav_discount": 0.25,
}


def _put(ws, cell, value, font=None, fill=None, fmt=None, align=None):
    c = ws[cell]; c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align)
    return c


def _header(ws, text):
    _put(ws, "A1", text, TITLE)
    ws["A2"] = "Testahil · Independent Valuation Study — Educational Analysis · Not investment advice"
    ws["A2"].font = Font(italic=True, size=9, color="777777")


def build_model(path, ticker, spot, shares, mc_result, assumptions: Optional[dict] = None):
    a = {**DEFAULT_ASSUMPTIONS, **(assumptions or {})}
    wb = openpyxl.Workbook()
    ws = {}
    for i, name in enumerate(SHEETS):
        ws[name] = wb.active if i == 0 else wb.create_sheet(name)
        ws[name].title = name
        ws[name].sheet_view.showGridLines = False
        ws[name].column_dimensions["A"].width = 34
        for col in "BCDEFG":
            ws[name].column_dimensions[col].width = 13

    # ---------------- READ FIRST ----------------
    w = ws["READ FIRST"]; _header(w, f"{ticker} — Valuation Model")
    notes = [
        "", "This workbook publishes fair-value ranges and probability distributions — never a rating,",
        "target price, or buy/sell call.", "",
        "Assumptions is the single blue-cell input layer. Every downstream sheet links back to it,",
        "so changing one input reprices the whole model.", "",
        "Blue = input.  Green = link to another sheet.  Black = formula.",
        "The Monte Carlo sheet holds simulation outputs (50,000 paths, seed 42).",
    ]
    for i, t in enumerate(notes, start=4):
        w[f"A{i}"] = t

    # ---------------- Assumptions (the blue layer) ----------------
    w = ws["Assumptions"]; _header(w, "Assumptions — single input layer (blue = edit here)")
    rows = [
        ("Spot price", spot, "0.00"), ("Shares outstanding (mn)", shares, "#,##0.0"),
        ("Risk-free rate (rf)", a["rf"], "0.0%"), ("Equity risk premium (ERP)", a["erp"], "0.0%"),
        ("Beta", a["beta"], "0.00"), ("Tax rate", a["tax"], "0.0%"),
        ("Terminal growth (g)", a["terminal_g"], "0.0%"),
        ("Revenue base (mn)", a["rev_base"], "#,##0"),
        ("Revenue growth Y1", a["g1"], "0.0%"), ("Revenue growth Y2", a["g2"], "0.0%"),
        ("Revenue growth Y3", a["g3"], "0.0%"), ("Revenue growth Y4", a["g4"], "0.0%"),
        ("Revenue growth Y5", a["g5"], "0.0%"), ("EBITDA margin", a["ebitda_margin"], "0.0%"),
        ("D&A (% revenue)", a["da_pct"], "0.0%"), ("Capex (% revenue)", a["capex_pct"], "0.0%"),
        ("Δ working capital (% revenue)", a["wc_pct"], "0.0%"),
        ("Net debt (mn)", a["net_debt"], "#,##0"),
        ("Recurring cap rate", a["cap_rate"], "0.0%"), ("NAV discount", a["nav_discount"], "0.0%"),
    ]
    _put(w, "A4", "Input", BOLD); _put(w, "B4", "Value", BOLD)
    r0 = 5
    ref = {}
    for i, (label, val, fmt) in enumerate(rows):
        r = r0 + i
        _put(w, f"A{r}", label)
        _put(w, f"B{r}", val, BLUE, INPUT_FILL, fmt)
        ref[label] = f"Assumptions!B{r}"
    # named refs used by DCF
    A = lambda lbl: ref[lbl]
    # Ke computed on Assumptions
    r_ke = r0 + len(rows) + 1
    _put(w, f"A{r_ke}", "Cost of equity Ke = rf + beta*ERP", BOLD)
    _put(w, f"B{r_ke}", f"={A('Risk-free rate (rf)')}+{A('Beta')}*{A('Equity risk premium (ERP)')}", BOLD, None, "0.0%")
    ke_ref = f"Assumptions!B{r_ke}"

    # ---------------- DCF (fully wired, required row order) ----------------
    w = ws["DCF"]; _header(w, "Discounted cash flow (FCFF) — full waterfall")
    yr_hdr = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    _put(w, "A4", "All figures $mn unless stated", Font(italic=True, size=9, color="777777"))
    hr = 5
    _put(w, f"A{hr}", "Line", HDR, HDR_FILL)
    for j, h in enumerate(yr_hdr):
        _put(w, f"{get_column_letter(2+j)}{hr}", h, HDR, HDR_FILL, align="center")
    def row(label): 
        row.i += 1; _put(w, f"A{row.i}", label); return row.i
    row.i = hr
    r_rev = row("Revenue")
    r_ebitda = row("EBITDA")
    r_da = row("D&A")
    r_ebit = row("EBIT")
    r_nopat = row("NOPAT = EBIT x (1 - tax)")
    r_pda = row("(+) D&A")
    r_capex = row("(−) Capex")
    r_wc = row("(−) Δ working capital")
    r_fcff = row("Free cash flow to firm (FCFF)")
    r_df = row("Discount factor")
    r_pv = row("PV of FCFF")
    gcells = [f"{A('Revenue growth Y%d' % k)}" for k in range(1, 6)]
    for j in range(5):
        col = get_column_letter(2 + j); prev = get_column_letter(1 + j)
        # Revenue
        if j == 0:
            w[f"{col}{r_rev}"] = f"={A('Revenue base (mn)')}*(1+{gcells[0]})"
        else:
            w[f"{col}{r_rev}"] = f"={prev}{r_rev}*(1+{gcells[j]})"
        w[f"{col}{r_ebitda}"] = f"={col}{r_rev}*{A('EBITDA margin')}"
        w[f"{col}{r_da}"] = f"={col}{r_rev}*{A('D&A (% revenue)')}"
        w[f"{col}{r_ebit}"] = f"={col}{r_ebitda}-{col}{r_da}"
        w[f"{col}{r_nopat}"] = f"={col}{r_ebit}*(1-{A('Tax rate')})"
        w[f"{col}{r_pda}"] = f"={col}{r_da}"
        w[f"{col}{r_capex}"] = f"=-{col}{r_rev}*{A('Capex (% revenue)')}"
        w[f"{col}{r_wc}"] = f"=-{col}{r_rev}*{A('Δ working capital (% revenue)')}"
        w[f"{col}{r_fcff}"] = f"={col}{r_nopat}+{col}{r_pda}+{col}{r_capex}+{col}{r_wc}"
        w[f"{col}{r_df}"] = f"=1/(1+{ke_ref})^{j+1}"
        w[f"{col}{r_pv}"] = f"={col}{r_fcff}*{col}{r_df}"
    for rr in (r_rev, r_ebitda, r_da, r_ebit, r_nopat, r_pda, r_capex, r_wc, r_fcff, r_pv):
        for j in range(5):
            w[f"{get_column_letter(2+j)}{rr}"].number_format = "#,##0.0"
        for j in range(5):
            w[f"{get_column_letter(2+j)}{r_df}"].number_format = "0.000"
    # valuation block
    vb = r_pv + 2
    _put(w, f"A{vb}", "WACC (= Ke, equity-financed simplification)", BOLD); _put(w, f"B{vb}", f"={ke_ref}", BOLD, None, "0.0%")
    _put(w, f"A{vb+1}", "Terminal growth g", BOLD); _put(w, f"B{vb+1}", f"={A('Terminal growth (g)')}", BOLD, None, "0.0%")
    _put(w, f"A{vb+2}", "Terminal value (Gordon on Y5 FCFF)")
    w[f"B{vb+2}"] = f"=F{r_fcff}*(1+B{vb+1})/(B{vb}-B{vb+1})"; w[f"B{vb+2}"].number_format = "#,##0.0"
    _put(w, f"A{vb+3}", "PV of terminal value")
    w[f"B{vb+3}"] = f"=B{vb+2}/(1+B{vb})^5"; w[f"B{vb+3}"].number_format = "#,##0.0"
    _put(w, f"A{vb+4}", "Sum PV of explicit FCFF")
    w[f"B{vb+4}"] = f"=SUM(B{r_pv}:F{r_pv})"; w[f"B{vb+4}"].number_format = "#,##0.0"
    _put(w, f"A{vb+5}", "Enterprise value (EV)", BOLD)
    w[f"B{vb+5}"] = f"=B{vb+3}+B{vb+4}"; w[f"B{vb+5}"].number_format = "#,##0.0"; w[f"B{vb+5}"].font = BOLD
    _put(w, f"A{vb+6}", "TV as % of EV", BOLD)
    w[f"B{vb+6}"] = f"=B{vb+3}/B{vb+5}"; w[f"B{vb+6}"].number_format = "0.0%"; w[f"B{vb+6}"].font = BOLD
    _put(w, f"A{vb+7}", "(−) Net debt")
    w[f"B{vb+7}"] = f"=-{A('Net debt (mn)')}"; w[f"B{vb+7}"].number_format = "#,##0.0"
    _put(w, f"A{vb+8}", "Equity value")
    w[f"B{vb+8}"] = f"=B{vb+5}+B{vb+7}"; w[f"B{vb+8}"].number_format = "#,##0.0"
    _put(w, f"A{vb+9}", "Fair value per share", BOLD)
    w[f"B{vb+9}"] = f"=B{vb+8}/{A('Shares outstanding (mn)')}"; w[f"B{vb+9}"].number_format = "0.00"; w[f"B{vb+9}"].font = BOLD
    dcf_ps_ref = f"DCF!B{vb+9}"

    # ---------------- Monte Carlo (from MCResult) ----------------
    w = ws["Monte Carlo"]; _header(w, "Monte Carlo — 50,000 paths, seed 42 (simulation outputs)")
    pt = mc_result.percentile_table()
    _put(w, "A4", "Diffusion these paths use is spot-anchored; the §1 value gap is kept out of the drift.",
         Font(italic=True, size=9, color="777777"))
    _put(w, "A6", "Percentile", HDR, HDR_FILL)
    for j, k in enumerate((5, 25, 50, 75, 95)):
        _put(w, f"{get_column_letter(2+j)}6", f"P{k}", HDR, HDR_FILL, align="center")
    rr = 7
    for name, rowd in pt.items():
        _put(w, f"A{rr}", name, BOLD)
        for j, k in enumerate((5, 25, 50, 75, 95)):
            _put(w, f"{get_column_letter(2+j)}{rr}", round(rowd[k], 2), None, None, "0.00")
        rr += 1
    # probability read
    col = mc_result.paths[:, -1]; spot_v = mc_result.anchor
    import numpy as np
    pr = rr + 1
    _put(w, f"A{pr}", "Probability read (T+60)", BOLD)
    reads = [
        ("P(price > spot)", float((col > spot_v).mean()), "0.0%"),
        ("P(+10%)", float((col > spot_v*1.1).mean()), "0.0%"),
        ("P(−10%)", float((col < spot_v*0.9).mean()), "0.0%"),
        ("Median level", float(np.percentile(col, 50)), "0.00"),
        ("Median % move", float(np.percentile(col, 50)/spot_v - 1), "0.0%"),
        ("Touch(+10%)", mc_result.touch_probability(spot_v*1.1, mc_result.horizon), "0.0%"),
        ("Touch(−10%)", mc_result.touch_probability(spot_v*0.9, mc_result.horizon), "0.0%"),
    ]
    for i, (lbl, val, fmt) in enumerate(reads):
        _put(w, f"A{pr+1+i}", lbl); _put(w, f"B{pr+1+i}", round(val, 4), None, None, fmt)

    # ---------------- Sensitivity (WACC x terminal-g grid of DCF per share) ----------------
    w = ws["Sensitivity"]; _header(w, "Sensitivity — fair value per share (WACC × terminal g)")
    _put(w, "A5", "WACC ↓  /  g →", BOLD)
    gs = [-0.02, -0.01, 0.0, 0.01, 0.02]
    ws_off = [-0.02, -0.01, 0.0, 0.01, 0.02]
    for j, dg in enumerate(gs):
        _put(w, f"{get_column_letter(2+j)}5", f"g{dg:+.0%}", BOLD, None, None, "center")
    for i, dw in enumerate(ws_off):
        r = 6 + i
        _put(w, f"A{r}", f"WACC {dw:+.0%}", BOLD)
        for j, dg in enumerate(gs):
            col = get_column_letter(2+j)
            wexpr = f"({ke_ref}+({dw}))"; gexpr = f"({A('Terminal growth (g)')}+({dg}))"
            tv = f"DCF!F{r_fcff}*(1+{gexpr})/({wexpr}-{gexpr})"
            ev = f"(SUM(DCF!B{r_pv}:F{r_pv})+({tv})/(1+{wexpr})^5)"
            eq = f"({ev}-{A('Net debt (mn)')})"
            w[f"{col}{r}"] = f"={eq}/{A('Shares outstanding (mn)')}"
            w[f"{col}{r}"].number_format = "0.00"

    # ---------------- Summary ----------------
    w = ws["Summary"]; _header(w, f"{ticker} — Summary")
    _put(w, "A5", "Spot"); _put(w, "B5", f"={A('Spot price')}", GREEN, None, "0.00")
    _put(w, "A6", "DCF fair value / share"); _put(w, "B6", f"={dcf_ps_ref}", GREEN, None, "0.00")
    _put(w, "A7", "MC median (T+60)"); _put(w, "B7", f"='Monte Carlo'!D8", GREEN, None, "0.00")
    _put(w, "A8", "Fair-value range is set in Fundamental Valuation (five-lens football field).",
         Font(italic=True, size=9, color="777777"))

    # ---------------- statement / other scaffolds ----------------
    def statement_frame(name, lines):
        w = ws[name]; _header(w, name)
        cols = ["FY-3", "FY-2", "FY-1", "F+1", "F+2", "F+3"]
        _put(w, "A5", "Line ($mn)", HDR, HDR_FILL)
        # only 6 value columns fit B..G
        for j, c in enumerate(cols):
            _put(w, f"{get_column_letter(2+j)}5", c, HDR, HDR_FILL, align="center")
        for i, l in enumerate(lines):
            _put(w, f"A{6+i}", l)
        _put(w, f"A{7+len(lines)}", "3 years historical + 5-year forecast — analyst fills from primary filings.",
             Font(italic=True, size=9, color="777777"))
    statement_frame("Income Statement", ["Revenue", "COGS", "Gross profit", "Opex", "EBITDA",
                                         "D&A", "EBIT", "Net interest", "Pre-tax profit", "Tax",
                                         "Net profit", "EPS"])
    statement_frame("Balance Sheet", ["Cash & equivalents", "Receivables", "Inventory",
                                      "PP&E", "Investment property", "Total assets", "Debt",
                                      "Payables", "Total liabilities", "Equity", "BVPS"])
    statement_frame("Cash Flow", ["CFO", "Capex", "FCF", "CFI", "Dividends paid", "CFF",
                                  "Net change in cash"])
    statement_frame("Summary Financials", ["Revenue", "EBITDA", "Net profit", "EPS", "DPS",
                                           "FCF", "Net debt"])

    for name, title in [("Fundamental Valuation", "Fundamental valuation — five-lens football field"),
                        ("Primary Lens", "Primary lens (SOTP / RNAV / DCF / DDM — per instrument)"),
                        ("Segments", "Segment build"),
                        ("Relative & Normalized", "Relative multiples & normalized earnings power"),
                        ("Per-Share & Ratios", "Per-share & ratio dashboard (fixed panel)"),
                        ("Peer & Sector", "Peer set & sector structure")]:
        w = ws[name]; _header(w, title)
    # Per-Share & Ratios fixed labels
    w = ws["Per-Share & Ratios"]
    panel = ["EPS", "DPS", "BVPS", "P/E", "P/B", "FCF yield", "Dividend yield", "EV/EBITDA",
             "ROAIC", "ROAE", "EBITDA margin", "Net-debt / EBITDA", "Effective tax"]
    for i, m in enumerate(panel):
        _put(w, f"A{5+i}", m)

    wb.save(path)
    return path
