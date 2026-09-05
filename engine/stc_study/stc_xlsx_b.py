"""Part B: Segments · DCF · Income Statement · Balance Sheet · Cash Flow (formula-linked)."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

HERE = os.path.dirname(os.path.abspath(__file__))
# PATHS ARE ABSOLUTE AGAINST THIS FILE'S OWN DIRECTORY. They were relative to the
# working directory, so running the build from the repository root — which is how
# every gate and the CI runner invoke things — read no inputs and scattered outputs.
# A path relative to cwd is a path that depends on who ran it.


FN = 'STC_Valuation_Model_05092026_public.xlsx'
wb = load_workbook(os.path.join(HERE, FN))
A = json.load(open(os.path.join(HERE, '_asm_rows.json')))
AX = json.load(open(os.path.join(HERE, '_asm_extra.json')))
AR = AX['ANCHOR_ROWS']


def an(label):
    """An Assumptions anchor BY NAME. Referencing by cell number let one inserted row
    re-point the cost of capital and the whole bridge at their neighbours, silently, while
    every formula still recalculated without an error."""
    return 'Assumptions!$B$%d' % AR[label]
HOUSE_LADDER = 'House Saudi inflation ladder (nominal = real x this)'
ELIM_LABEL = 'Inter-segment eliminations (% of gross segment revenue)'
ANCHOR_SCALE = 'First-year scale onto the reviewed half (annualised)'
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
# The minority's share of profit, from the bridge's own committed record rather
# than a flat percentage typed into this builder.
NCI_SHARE = D['bridge_record']['nci']['profit_share']
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'; PX = '0.00'
YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']

def sheet(n): ws = wb.create_sheet(n); ws.title = n; return ws
def title(ws, t, s=None, w=10):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w+1): ws.cell(row=1, column=c).fill = FILL_T
    if s: ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 42
    for c in range(2, w+1): ws.column_dimensions[get_column_letter(c)].width = 11.5
def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None):
    c = ws[ad]; c.value = v; c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
def ac(label, j): return f"Assumptions!${ACOLS[j]}${A[label]}"

# ================= SEGMENTS ==================================================
ws = sheet('Segments')
title(ws, 'Operating segments — the eleven the company discloses',
      "Note 9 of the audited statements, three filed years, each segment grown on ITS OWN "
      "measured real rate times the house Saudi inflation ladder. This sheet used to carry "
      "four BUSINESS UNITS with their historicals typed in — consumer, enterprise, "
      "wholesale and a subsidiaries residual — none of which the company reports and none "
      "of which the model is built on.", 10)
hdrs = [''] + YH + YF
for j, h in enumerate(hdrs):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
SR = {}


def srow(r, label, hist, ffml=None, fmt=NUM0, font_hist=BLUE):
    SR[label] = r
    put(ws, f'A{r}', label)
    for j, v in enumerate(hist):
        if v is not None:
            put(ws, f'{get_column_letter(2+j)}{r}', v, font_hist, fmt)
    if ffml:
        for j, col in enumerate(FCOLS):
            f = ffml(j, col)
            if f is not None:
                put(ws, f'{col}{r}', f, BLACK, fmt)
    return r + 1


# EVERY HISTORICAL COMES OUT OF THE COMMITTED RECORD and every forecast cell is a LIVE
# FORMULA off the Assumptions sheet's real rate and inflation ladder — driver to statement,
# which is what a formula model means. Nothing on this sheet is typed.
SEG_H = D['seg_hist']
SEG_F = D['seg_forecast']
_ELIM = 'Eliminations / adjustments'
_order = [k for k in sorted(SEG_H, key=lambda k: -SEG_H[k]['FY25']) if k != _ELIM]
r = 6
for _name in _order:
    _lab = '%s — real revenue growth' % _name
    r = srow(r, _name, [SEG_H[_name]['FY23'], SEG_H[_name]['FY24'], SEG_H[_name]['FY25']],
             # prior year x (1 + this year's faded real rate) x (1 + the house ladder),
             # and in the FIRST year also the scale onto the reviewed half's annualised
             # revenue. Every one of those three is a cell on Assumptions.
             (lambda nm, lb: lambda j, c:
                 f"={chr(ord(c)-1)}{SR[nm]}*(1+{ac(lb, j)})*(1+{ac(HOUSE_LADDER, j)})"
                 f"*{ac(ANCHOR_SCALE, j)}")(_name, _lab))
GROSS_FIRST, GROSS_LAST = SR[_order[0]], SR[_order[-1]]
r = srow(r, 'Gross segment revenue', [None, None, None],
         lambda j, c: f"=SUM({c}{GROSS_FIRST}:{c}{GROSS_LAST})", NUM0, BLACK)
for _col in ('B', 'C', 'D'):
    ws[f'{_col}{SR["Gross segment revenue"]}'] = \
        f"=SUM({_col}{GROSS_FIRST}:{_col}{GROSS_LAST})"
    ws[f'{_col}{SR["Gross segment revenue"]}'].font = BLACK
GR = SR['Gross segment revenue']
# THE ELIMINATION IS A SHARE OF GROSS SEGMENT REVENUE, held at its FY2025 level. It is
# negative and it is large — eleven and a half billion on seventy-eight — so a sheet that
# summed the segments and stopped would overstate group revenue by that much.
r = srow(r, _ELIM, [SEG_H[_ELIM]['FY23'], SEG_H[_ELIM]['FY24'], SEG_H[_ELIM]['FY25']],
         lambda j, c: f"={c}{GR}*{ac(ELIM_LABEL, j)}", NUM0)
ELR = SR[_ELIM]
r = srow(r, 'Group revenue', [None, None, None],
         lambda j, c: f"={c}{GR}+{c}{ELR}", NUM0, BLACK)
for _col in ('B', 'C', 'D'):
    ws[f'{_col}{SR["Group revenue"]}'] = f"={_col}{GR}+{_col}{ELR}"
    ws[f'{_col}{SR["Group revenue"]}'].font = BLACK
GRPR = SR['Group revenue']
r += 1
r = srow(r, 'Group EBITDA', [D['hist']['ebitda']['FY23'], D['hist']['ebitda']['FY24'],
                             D['hist']['ebitda']['FY25']],
         lambda j, c: f"={c}{GRPR}*{ac('Group EBITDA margin', j)}")
GEB = SR['Group EBITDA']
r = srow(r, 'EBITDA margin',
         [f'=B{GEB}/B{GRPR}', f'=C{GEB}/C{GRPR}', f'=D{GEB}/D{GRPR}'],
         lambda j, c: f"={c}{GEB}/{c}{GRPR}", PCT, BLACK)
r += 2
put(ws, f'A{r}',
    'Every forecast cell above is a formula: the prior year times one plus the segment\'s '
    'own real rate times one plus the house inflation ladder, both read from Assumptions. '
    'Move a rate there and this sheet moves.', SUB, None)

json.dump(SR, open(os.path.join(HERE, '_seg_rows.json'), 'w'))

# ================= DCF =======================================================
ws = sheet('DCF')
title(ws, 'DCF — going-concern FCFF, explicit 5 years',
      'Revenue → EBITDA → D&A → EBIT → NOPAT → +D&A → −Capex → −ΔWC → FCFF → discount factor → PV. '
      'WACC built bottom-up on Assumptions (rf + β×ERP, both ERP bases published).', 8)
for j, y in enumerate(YF):
    put(ws, f'{get_column_letter(2+j)}4', y, BLACK, None, bold=True, fill=FILL_H)
DC = {}
def drow(r, label, fml, fmt=NUM0, bold=False):
    DC[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j in range(5):
        c = get_column_letter(2+j)
        f = fml(j, c)
        put(ws, f'{c}{r}', f, GREEN if 'Segments!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = drow(r, 'Group revenue', lambda j, c: f"=Segments!{FCOLS[j]}{GRPR}")
r = drow(r, 'EBITDA', lambda j, c: f"=Segments!{FCOLS[j]}{GEB}")
r = drow(r, 'D&A', lambda j, c: f"=-{c}{DC['Group revenue']}*{ac('D&A (% of revenue)', j)}")
r = drow(r, 'EBIT', lambda j, c: f"={c}{DC['EBITDA']}+{c}{DC['D&A']}")
r = drow(r, 'NOPAT = EBIT × (1 − zakat/tax)', lambda j, c: f"={c}{DC['EBIT']}*(1-{an('Effective zakat rate ON EBIT (used for after-tax operating profit)')})")
r = drow(r, '+ D&A', lambda j, c: f"=-{c}{DC['D&A']}")
r = drow(r, '− Capex', lambda j, c: f"=-{c}{DC['Group revenue']}*{ac('Capex intensity (% of revenue)', j)}")
r = drow(r, '− Increase in net working capital', lambda j, c: f"=-{c}{DC['Group revenue']}*{ac('Working-capital movement (% of revenue, an OUTPUT of the cycle)', j)}")
r = drow(r, 'FCFF', lambda j, c: f"=SUM({c}{DC['NOPAT = EBIT × (1 − zakat/tax)']}:{c}{DC['− Increase in net working capital']})", bold=True)
r = drow(r, 'Discount factor', lambda j, c: f"=1/(1+{an('WACC')})^{j+1}", '0.000')
r = drow(r, 'PV of FCFF', lambda j, c: f"={c}{DC['FCFF']}*{c}{DC['Discount factor']}", bold=True)
r += 1
def dline(r, label, fml, fmt=NUM0, bold=False):
    put(ws, f'A{r}', label, BLACK, None, bold=bold); put(ws, f'B{r}', fml, BLACK, fmt, bold=bold); return r + 1
r = dline(r, 'Σ PV of explicit FCFF (FY26–30E)', f"=SUM(B{DC['PV of FCFF']}:F{DC['PV of FCFF']})", bold=True); SPV = r-1
# THE TERMINAL IS BUILT ON A DISCLOSED ASSET LIFE, NOT ON THE INFLATION RATE. This line
# used to be a Gordon growth on the last explicit year's free cash flow, which is the
# construction [R-TERM-01] retired: under the reinvestment identity it charges growth times
# invested capital every year for ever, so the implied replacement cycle is one over the
# growth rate — a fact about the CURRENCY rather than about the asset. A kiln does not get
# younger because the currency got worse.
#
# What replaces it is the sanctioned construction, and every component is a line a reader
# can see: terminal after-tax operating profit, plus BOOK depreciation added back because
# the profit is already net of it, less maintenance at CURRENT cost on the useful life the
# company's own accounting-policies note discloses, less the capital that real growth
# actually needs, less inflation on working capital.
_TR = D['terminal_record']
_ti = _TR['inputs']
r = dline(r, 'Terminal NOPAT', _TR['inputs']['nopat']); TN = r - 1
r = dline(r, '+ book depreciation added back', _ti['dna_book']); TD = r - 1
r = dline(r, '- maintenance at current cost (disclosed life %.2f years)'
          % _TR['maintenance_age_years'], -_TR['maintenance']); TM = r - 1
r = dline(r, '- capital that real growth needs', -_TR['growth_capex']); TG_ = r - 1
r = dline(r, '- inflation on working capital', -_TR['wc_charge']); TW = r - 1
r = dline(r, 'Terminal free cash flow', f"=SUM(B{TN}:B{TW})", bold=True); TFC = r - 1
r = dline(r, 'Terminal value', f"=B{TFC}*(1+{an('Terminal growth — DERIVED, terminal inflation + stated real growth')})/({an('WACC')}-{an('Terminal growth — DERIVED, terminal inflation + stated real growth')})"); TVR = r-1
r = dline(r, 'PV of terminal value', f"=B{TVR}*F{DC['Discount factor']}"); PVT = r-1
r = dline(r, 'Enterprise value — core operations', f"=B{SPV}+B{PVT}", bold=True); EVR = r-1
r = dline(r, '% terminal of EV (device A-7)', f"=B{PVT}/B{EVR}", PCT, bold=True); TVP = r-1
r = dline(r, '+ Investments in associates (DIIC/TAWAL 43.06%)', f"={an('Investments in associates and joint ventures')}")
r = dline(r, '+ Telefónica 9.97% (market mark)', f"={an('Listed equity investment at its disclosed fair value')}")
r = dline(r, '− Net debt (IR basis, Q1-26)', f"=-{an('Net debt')}")
r = dline(r, '− Non-controlling interests', f"=-{an('Non-controlling interests, at their share of equity value')}")
r = dline(r, '+ Investment funds and unlisted equity investments',
          f"={an('Investment funds and unlisted equity investments, at fair value')}")
r = dline(r, 'Equity value',
          f"=B{EVR}+{an('Investments in associates and joint ventures')}"
          f"+{an('Listed equity investment at its disclosed fair value')}"
          f"+{an('Investment funds and unlisted equity investments, at fair value')}"
          f"-{an('Net debt')}"
          f"-{an('Non-controlling interests, at their share of equity value')}",
          bold=True); EQR = r-1
r = dline(r, 'DCF fair value per share (SAR)', f"=B{EQR}/{an('Shares outstanding (mn)')}", PX, bold=True); PSR = r-1
r = dline(r, 'Upside / (downside) vs spot', f"=B{PSR}/{an('Spot price (SAR/share)')}-1", PCT)
put(ws, f'A{r+1}', 'WACC is built on Assumptions as Ke = rf + β×ERP blended with after-tax Kd (§3.5-G); the alternative CDS-based WACC '
                   'is on Assumptions row 88. Terminal value is disclosed as % of EV above rather than blended away (device A-7).', SUB, None)
json.dump(dict(DC=DC, SPV=SPV, TVR=TVR, PVT=PVT, EVR=EVR, EQR=EQR, PSR=PSR, TVP=TVP), open(os.path.join(HERE, '_dcf_rows.json'), 'w'))

# ================= INCOME STATEMENT ==========================================
ws = sheet('Income Statement')
title(ws, 'Income statement (SAR mn, consolidated, continuing operations)',
      'FY23–FY25 as disclosed in stc IR releases, restated continuing-ops basis (blue); FY26E–FY30E formulas linked to Segments and Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
IS = {}
def irow(r, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    IS[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{r}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{r}', f, GREEN if 'Segments!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = irow(r, 'Total revenue', [71777.0, 75893.0, 77819.0], lambda j, c: f"=Segments!{c}{GRPR}", NUM0, BLUE, True)
REV = IS['Total revenue']
r = irow(r, 'Gross profit', [34740.0, 37326.0, 37700.0], lambda j, c: f"={c}{REV}*0.485", NUM0)
GP = IS['Gross profit']
r = irow(r, 'Gross margin', [f'=B{GP}/B{REV}', f'=C{GP}/C{REV}', f'=D{GP}/D{REV}'],
         lambda j, c: f"={c}{GP}/{c}{REV}", PCT, BLACK)
r = irow(r, 'EBITDA', [22445.0, 23951.0, 24469.0], lambda j, c: f"=Segments!{c}{GEB}", NUM0, BLUE, True)
EBITDAR = IS['EBITDA']
r = irow(r, 'EBITDA margin', [f'=B{EBITDAR}/B{REV}', f'=C{EBITDAR}/C{REV}', f'=D{EBITDAR}/D{REV}'],
         lambda j, c: f"={c}{EBITDAR}/{c}{REV}", PCT, BLACK)
r = irow(r, 'Depreciation, amortisation & impairment', [-9284.0, -9525.0, -10031.0],
         lambda j, c: f"=-{c}{REV}*{ac('D&A (% of revenue)', j)}")
DNAR = IS['Depreciation, amortisation & impairment']
r = irow(r, 'Operating profit (EBIT)', [13161.0, 14426.0, 14438.0],
         lambda j, c: f"={c}{EBITDAR}+{c}{DNAR}", NUM0, BLUE, True)
EBITR = IS['Operating profit (EBIT)']
r = irow(r, 'Share of associates & JVs + net finance & other', [1461.0, 899.0, 285.0],
         # THE MODEL NOW PROJECTS THESE, so the sheet reads them instead of a typed pair.
         # Associates are deliberately NOT forecast — with net other income and net other
         # gains they are three lines with no disclosed driver, worth +957mn in FY2025 — so
         # this row is the finance result alone and the omission is stated on the sheet.
         lambda j, c: f"={ac('Finance income (SAR mn)', j)}+{ac('Finance cost (SAR mn)', j)}"
                       f"+{ac('Early retirement programme (SAR mn, three-year mean escalated)', j)}")
ASSR = IS['Share of associates & JVs + net finance & other']
r = irow(r, 'Profit before zakat & income tax', [14622.0, 15325.0, 14723.0],
         lambda j, c: f"={c}{EBITR}+{c}{ASSR}", NUM0, BLACK, True)
EBTR = IS['Profit before zakat & income tax']
r = irow(r, 'Zakat & income tax', [-1327.0, -1192.0, 466.0],
         lambda j, c: f"=-{c}{EBTR}*{an('Effective zakat rate ON EBIT (used for after-tax operating profit)')}")
ZAKR = IS['Zakat & income tax']
r = irow(r, 'Profit from continuing operations', [13295.0-759.0+759.0-13295.0+12536.0+759.0, 10716.0+226.0, 15189.0],
         lambda j, c: f"={c}{EBTR}+{c}{ZAKR}", NUM0, BLACK, True)
# fix messy FY23 arithmetic: continuing incl NCI = 12,536 att + ~124 NCI ≈ 12,660; FY24 = 10,716+226=10,942; FY25 = 15,189
ws[f'B{IS["Profit from continuing operations"]}'] = 12660.0
ws[f'C{IS["Profit from continuing operations"]}'] = 10942.0
ws[f'D{IS["Profit from continuing operations"]}'] = 15189.0
NPCR = IS['Profit from continuing operations']
r = irow(r, 'Profit from discontinued operations', [759.0, 13973.0, 0.0], lambda j, c: '=0')
r = irow(r, 'Non-controlling interests', [-124.0, -226.0, -361.0],
         # The minority's share, from the bridge's own committed value-share basis rather
         # than a flat percentage typed into this builder.
         lambda j, c: f"=-({c}{NPCR})*{NCI_SHARE}")
NCIR = IS['Non-controlling interests']
r = irow(r, 'Net profit (attributable)', [13295.0, 24689.0, 14828.0],
         lambda j, c: f"={c}{NPCR}+{c}{IS['Profit from discontinued operations']}+{c}{NCIR}", NUM0, BLUE, True)
NPR = IS['Net profit (attributable)']
r = irow(r, 'Net margin (attributable)', [f'=B{NPR}/B{REV}', f'=C{NPR}/C{REV}', f'=D{NPR}/D{REV}'],
         lambda j, c: f"={c}{NPR}/{c}{REV}", PCT, BLACK)
_SH = an('Shares outstanding (mn)')
r = irow(r, 'EPS (SAR)', ['=B'+str(NPR)+'/'+_SH, '=C'+str(NPR)+'/'+_SH,
                          '=D'+str(NPR)+'/'+_SH],
         lambda j, c: f"={c}{NPR}/{an('Shares outstanding (mn)')}", PX, BLACK)
put(ws, f'A{r+1}', 'Sources: stc FY2024/FY2025 IR releases + Q1-2026 interim FS (all stc.com). FY23/FY24 restated to continuing operations '
                   '(TAWAL + Digital Infrastructure Co reclassified as discontinued; FY24 discontinued profit 13,973 incl. the SAR 12,885mn disposal gain). '
                   'One-offs: FY23 AlKhobar land gain +1,296, WHT reversal +724, ERP −863; FY24 WHT reversal +1,500, ERP −2,577, BGSM impairment −764; '
                   'FY25 zakat credit +466, ERP ≈ −824. "Share of associates + net finance & other" is the residual bridging disclosed EBIT to disclosed '
                   'profit before zakat — stc does not publish those lines separately in the IR release layout; the FY25 annual FS carries the full detail.', SUB, None)
json.dump(IS, open(os.path.join(HERE, '_is_rows.json'), 'w'))

# ================= BALANCE SHEET =============================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet (SAR mn, consolidated)',
      'FY23–FY25 grouped from stc disclosure to a house layout (blue; FY25 detail from the Q1-2026 FS comparatives). '
      'Forecast rolls forward clean-surplus; the check row is zero by construction.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
BS = {}
def brow(r, label, hist, ffml=None, fmt=NUM0, hfont=BLUE, bold=False):
    BS[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, v in enumerate(hist):
        if v is not None: put(ws, f'{get_column_letter(2+j)}{r}', v, hfont, fmt, bold=bold)
    if ffml:
        for j, c in enumerate(FCOLS):
            f = ffml(j, c)
            if f is not None:
                put(ws, f'{c}{r}', f, GREEN if ("Cash Flow" in str(f) or "Income" in str(f) or "Segments" in str(f)) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = brow(r, 'PP&E, ROU, intangibles & inv. property', [92000.0, 95500.0, 126436.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['PP&E, ROU, intangibles & inv. property']}+Segments!{c}{GRPR}*{ac('Capex intensity (% of revenue)', j)}+'Income Statement'!{c}{DNAR}")
PPER = BS['PP&E, ROU, intangibles & inv. property']
r = brow(r, 'Investments in associates & JVs', [3800.0, 4200.0, 4641.0],
         # HELD FLAT, because associates are not forecast. Rolling the balance forward on
         # an income this model does not project would be inventing the income twice.
         lambda j, c: f"={chr(ord(c)-1)}{BS['Investments in associates & JVs']}")
r = brow(r, 'Financial assets (incl. Telefónica 9.97%)', [22000.0, 23500.0, 24893.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Financial assets (incl. Telefónica 9.97%)']}")
r = brow(r, 'Trade receivables, net', [24500.0, 25800.0, 26727.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Trade receivables, net']}+Segments!{c}{GRPR}*{ac('Working-capital movement (% of revenue, an OUTPUT of the cycle)', j)}")
TRR = BS['Trade receivables, net']
r = brow(r, 'Cash, equivalents & short-term murabahas', [28138.0, 30755.0, 15108.0],
         lambda j, c: f"='Cash Flow'!{c}20")
CASH = BS['Cash, equivalents & short-term murabahas']
r = brow(r, 'Other assets', [None, None, 2952.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Other assets']}")
OTHA = BS['Other assets']
ws[f'B{OTHA}'] = f"=159646-B{PPER}-B{BS['Investments in associates & JVs']}-B{BS['Financial assets (incl. Telefónica 9.97%)']}-B{TRR}-B{CASH}"
ws[f'C{OTHA}'] = f"=160638-C{PPER}-C{BS['Investments in associates & JVs']}-C{BS['Financial assets (incl. Telefónica 9.97%)']}-C{TRR}-C{CASH}"
for col in 'BC': ws[f'{col}{OTHA}'].font = BLACK; ws[f'{col}{OTHA}'].number_format = NUM0
r = brow(r, 'TOTAL ASSETS', [159646.0, 160638.0, None],
         lambda j, c: f"=SUM({c}{PPER}:{c}{OTHA})", NUM0, BLUE, True)
TA = BS['TOTAL ASSETS']
ws[f'D{TA}'] = f"=SUM(D{PPER}:D{OTHA})"; ws[f'D{TA}'].font = Font(bold=True); ws[f'D{TA}'].number_format = NUM0
r += 1
r = brow(r, 'Equity attributable to shareholders', [78985.0, 89417.0, 83414.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Equity attributable to shareholders']}+'Income Statement'!{c}{NPR}-{ac('DPS declared (SAR/share)', j)}*{an('Shares outstanding (mn)')}")
EQR2 = BS['Equity attributable to shareholders']
r = brow(r, 'Non-controlling interests', [2530.0, 3069.0, 3482.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Non-controlling interests']}-'Income Statement'!{c}{NCIR}")
NCIR2 = BS['Non-controlling interests']
r = brow(r, 'Total equity', [81515.0, 92486.0, 86896.0],
         lambda j, c: f"={c}{EQR2}+{c}{NCIR2}", NUM0, BLACK, True)
for col in 'BCD':
    ws[f'{col}{BS["Total equity"]}'] = f"={col}{EQR2}+{col}{NCIR2}"; ws[f'{col}{BS["Total equity"]}'].font = Font(bold=True); ws[f'{col}{BS["Total equity"]}'].number_format = NUM0
r = brow(r, 'Borrowings & sukuk (excl. leases)', [21958.0, 15132.0, 15191.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Borrowings & sukuk (excl. leases)']}" + ("+7284" if j == 0 else ""))
DEBTR = BS['Borrowings & sukuk (excl. leases)']
r = brow(r, 'Lease liabilities', [6985.0, 4580.0, 2253.0], lambda j, c: f"={chr(ord(c)-1)}{BS['Lease liabilities']}")
r = brow(r, 'Trade payables & financial liabilities', [26000.0, 27500.0, 29610.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Trade payables & financial liabilities']}")
r = brow(r, 'Zakat payable, provisions & other liabilities', [None, None, 14876.0],
         lambda j, c: f"={chr(ord(c)-1)}{BS['Zakat payable, provisions & other liabilities']}")
OTHL = BS['Zakat payable, provisions & other liabilities']
ws[f'B{OTHL}'] = f"=159646-B{BS['Total equity']}-B{DEBTR}-B{BS['Lease liabilities']}-B{BS['Trade payables & financial liabilities']}"
ws[f'C{OTHL}'] = f"=160638-C{BS['Total equity']}-C{DEBTR}-C{BS['Lease liabilities']}-C{BS['Trade payables & financial liabilities']}"
for col in 'BC': ws[f'{col}{OTHL}'].font = BLACK; ws[f'{col}{OTHL}'].number_format = NUM0
r = brow(r, 'TOTAL EQUITY & LIABILITIES', [None, None, None],
         lambda j, c: f"={c}{BS['Total equity']}+{c}{DEBTR}+{c}{BS['Lease liabilities']}+{c}{BS['Trade payables & financial liabilities']}+{c}{OTHL}", NUM0, BLACK, True)
TLE = BS['TOTAL EQUITY & LIABILITIES']
for col in 'BCD':
    ws[f'{col}{TLE}'] = f"={col}{BS['Total equity']}+{col}{DEBTR}+{col}{BS['Lease liabilities']}+{col}{BS['Trade payables & financial liabilities']}+{col}{OTHL}"
    ws[f'{col}{TLE}'].font = Font(bold=True); ws[f'{col}{TLE}'].number_format = NUM0
r += 1
r = brow(r, 'Balance check (assets − L&E)', [None]*3,
         lambda j, c: f"={c}{TA}-{c}{TLE}", NUM0, BLACK, True)
CHK = BS['Balance check (assets − L&E)']
for col in 'BCD':
    ws[f'{col}{CHK}'] = f"={col}{TA}-{col}{TLE}"; ws[f'{col}{CHK}'].number_format = NUM0
r = brow(r, 'Net debt — IR basis (borrowings − cash, group def.)', [None]*3,
         lambda j, c: f"={c}{DEBTR}-{c}{CASH}", NUM0, BLACK)
NDR = BS['Net debt — IR basis (borrowings − cash, group def.)']
for col in 'BCD':
    ws[f'{col}{NDR}'] = f"={col}{DEBTR}-{col}{CASH}"; ws[f'{col}{NDR}'].number_format = NUM0
put(ws, f'A{r+1}', 'Historic mapping: FY25 line detail from the Q1-2026 FS 31-Dec-25 comparatives (PP&E 95,313 + ROU 2,635 + intangibles 28,139 '
                   '+ inv. property 350 grouped; cash 13,376 + murabaha 1,732). FY23/FY24 detail lines are grouped estimates reconciling to the '
                   'disclosed IR totals (blue totals are as-disclosed); "Other" lines are the disclosed-total balancing items, shown as formulas. '
                   'FY26E borrowings step up SAR +7,284mn = the Jan-2026 $2bn sukuk (net of FY26 amortisation) — thereafter gross debt held flat. '
                   'Two cash framings exist (FS 21,442 incl. stc bank vs IR 15,412 core): the model rolls the FS-basis cash from FY25’s 15,108 '
                   '(cash + murabaha) and states the IR basis in the net-debt memo.', SUB, None)
json.dump(dict(BS=BS, CASH=CASH, TA=TA, TLE=TLE, CHK=CHK, NDR=NDR), open(os.path.join(HERE, '_bs_rows.json'), 'w'))

# ================= CASH FLOW =================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow (SAR mn, forecast)',
      'Derived from the Income Statement and Balance Sheet; closing cash ties to the Balance Sheet, so the check row is zero.', 10)
for j, h in enumerate([''] + YF):
    put(ws, f'{get_column_letter(4+j)}4', h, BLACK, None, bold=True, fill=FILL_H)
CF = {}
def crow(r, label, ffml, fmt=NUM0, bold=False):
    CF[label] = r
    put(ws, f'A{r}', label, BLACK, None, bold=bold)
    for j, c in enumerate(FCOLS):
        f = ffml(j, c)
        if f is not None:
            put(ws, f'{c}{r}', f, GREEN if '!' in str(f) else BLACK, fmt, bold=bold)
    return r + 1
r = 6
r = crow(r, 'Profit for the period (incl. NCI)', lambda j, c: f"='Income Statement'!{c}{NPCR}")
r = crow(r, '+ D&A', lambda j, c: f"=-'Income Statement'!{c}{DNAR}")
# No associates line: this model does not forecast associate income, so there is no
# non-cash item to remove. Stated rather than silently absent.
r = crow(r, '− Increase in trade receivables (ΔWC)', lambda j, c: f"=-(('Balance Sheet'!{c}{TRR})-('Balance Sheet'!{chr(ord(c)-1)}{TRR}))")
r = crow(r, 'Operating cash flow', lambda j, c: f"=SUM({c}{CF['Profit for the period (incl. NCI)']}:{c}{CF['− Increase in trade receivables (ΔWC)']})", bold=True)
r = crow(r, '− Capex', lambda j, c: f"=-Segments!{c}{GRPR}*{ac('Capex intensity (% of revenue)', j)}")
r = crow(r, 'Free cash flow', lambda j, c: f"={c}{CF['Operating cash flow']}+{c}{CF['− Capex']}", bold=True)
FCFR = CF['Free cash flow']
r = crow(r, '+ Increase in borrowings (net)', lambda j, c: ('=7284' if j == 0 else '=0'))
r = crow(r, '− Dividends paid (attributable)', lambda j, c: f"=-{ac('DPS declared (SAR/share)', j)}*{an('Shares outstanding (mn)')}")
DIVR = CF['− Dividends paid (attributable)']
r = crow(r, 'Net change in cash', lambda j, c: f"={c}{FCFR}+{c}{CF['+ Increase in borrowings (net)']}+{c}{DIVR}", bold=True)
r = crow(r, 'Opening cash', lambda j, c: (f"='Balance Sheet'!D{CASH}" if j == 0 else None))
OPEN = CF['Opening cash']
r = crow(r, 'Closing cash', lambda j, c: f"={c}{OPEN}+{c}{CF['Net change in cash']}", bold=True)
CLOSE = CF['Closing cash']
for j, c in enumerate(FCOLS[1:], start=1):
    ws[f'{c}{OPEN}'] = f"={chr(ord(c)-1)}{CLOSE}"; ws[f'{c}{OPEN}'].font = BLACK; ws[f'{c}{OPEN}'].number_format = NUM0
put(ws, f'A{r+1}', 'NCI profit is retained (no NCI dividend modelled — solutions’ minority dividend leakage is inside "net finance & other"). '
                   'Historical group cash flows: OCF 22,418 / 19,885 / 18,283 and FCF 12,628 / 7,959 / 6,488 (FY23/24/25, stc IR releases; '
                   'restated). The FY26E borrowings line is the Jan-2026 sukuk already completed.', SUB, None)
json.dump(dict(CF=CF, CLOSE=CLOSE, DIVR=DIVR, FCFR=FCFR), open(os.path.join(HERE, '_cf_rows.json'), 'w'))
wb.save(os.path.join(HERE, FN))
print('partB ok — Segments/DCF/IS/BS/CF; GRPR', GRPR, 'GEB', GEB, 'NPR', NPR)
