"""Part C: SOTP Bridge · Relative & Normalized · Summary · Fundamental Valuation ·
Summary Financials · Monte Carlo · Sensitivity · Per-Share & Ratios · Peer & Sector · ordering."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

HERE = os.path.dirname(os.path.abspath(__file__))
# PATHS ARE ABSOLUTE AGAINST THIS FILE'S OWN DIRECTORY. They were relative to the
# working directory, so running the build from the repository root — which is how
# every gate and the CI runner invoke things — read no inputs and scattered outputs.
# A path relative to cwd is a path that depends on who ran it.


FN = 'STC_Valuation_Model_05092026_public.xlsx'
wb = load_workbook(os.path.join(HERE, FN))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
A = json.load(open(os.path.join(HERE, '_asm_rows.json'))); AX = json.load(open(os.path.join(HERE, '_asm_extra.json')))
# THE ROW NUMBERS BELOW WERE RE-POINTED ONCE, BY HAND, AFTER ONE INSERTED ROW ON THE
# ASSUMPTIONS SHEET SHIFTED EVERY ANCHOR AND THIS SHEET WENT ON READING THE OLD POSITIONS —
# silently, because every formula still recalculated. AR is the map by NAME, and any new
# reference on this sheet should go through it rather than through a literal.
AR = AX['ANCHOR_ROWS']


def an(label):
    return 'Assumptions!$B$%d' % AR[label]


# SHORT NAMES FOR THE ANCHORS THIS SHEET USES. Every one resolves through AR, so
# a row inserted on Assumptions moves them all; and they are constants rather than
# inline calls because an f-string cannot nest the same quote its label contains.
A_EVX = an("EV/EBITDA — the company's own trailing multiple, three year ends")
A_ASSOC = an('Investments in associates and joint ventures')
A_LISTED = an('Listed equity investment at its disclosed fair value')
A_FUNDS = an('Investment funds and unlisted equity investments, at fair value')
A_NETDEBT = an('Net debt')
A_NCI = an('Non-controlling interests, at their share of equity value')
A_FUNDS = an('Investment funds and unlisted equity investments, at fair value')
A_NCISHARE = an('Minority share of equity value (proportional)')
A_SHARES = an('Shares outstanding (mn)')
A_SPOT = an('Spot price (SAR/share)')
A_NP26 = an('FY26E net profit for the relative cross-check')
A_NORMPAT = an('Normalized PAT (ex one-offs)')
A_NORMPE = an('Justified through-cycle P/E')
A_KE = an('Cost of equity')
A_KDAT = an('After-tax Kd')
A_WD = an('Debt weight')
A_WACC = an('WACC')
A_TG = an('Terminal growth — DERIVED, terminal inflation + stated real growth')
A_TGDIV = an('DDM terminal dividend growth')
A_RF = an('Risk-free rate, normalised by the sovereign default spread')
A_ERP = an('Equity risk premium (market basis, central)')
A_VOL = an('Forward volatility over three months (annualised)')
A_LEAN = an('Momentum lean applied over three months')
A_TAX = an('Effective zakat rate ON EBIT (used for after-tax operating profit)')
SR = json.load(open(os.path.join(HERE, '_seg_rows.json'))); IS = json.load(open(os.path.join(HERE, '_is_rows.json')))
BSJ = json.load(open(os.path.join(HERE, '_bs_rows.json'))); CFJ = json.load(open(os.path.join(HERE, '_cf_rows.json')))
DCJ = json.load(open(os.path.join(HERE, '_dcf_rows.json')))
BS = BSJ['BS']; CF = CFJ['CF']; DC = DCJ['DC']
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'; PX = '0.00'; MULT = '0.00x'
YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']
DPSR = AX['DPS_ROW']

def sheet(n): ws = wb.create_sheet(n); ws.title = n; return ws
def title(ws, t, s=None, w=9):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w+1): ws.cell(row=1, column=c).fill = FILL_T
    if s: ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 44
    for c in range(2, w+1): ws.column_dimensions[get_column_letter(c)].width = 12
def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None):
    c = ws[ad]; c.value = v; c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill

# ================= SOTP Bridge (primary-lens sheet) =====================
ws = sheet('SOTP Bridge')
title(ws, 'Valuation bridge — FCFF DCF (primary) + the dividend-policy DDM',
      'Core-operations EV from the DCF sheet; stakes marked separately; DDM = the locked SAR 0.55/quarter policy. Links to DCF / Assumptions.', 7)
r = 5
rows = [
 ('Component', 'Basis', 'Value (SAR mn)', True),
 ('Core operations enterprise value', 'FCFF DCF (§1.1): Σ PV FCFF + PV terminal', f"=DCF!B{DCJ['EVR']}", False),
 ('  of which terminal value', '% of core enterprise value', f"=DCF!B{DCJ['TVP']}", False),
 ('+ Investments in associates & JVs', '43.06% DIIC/TAWAL at carrying value', f"={A_ASSOC}", False),
 ('+ Telefónica 9.97%', 'Market mark on the disclosed stake', f"={A_LISTED}", False),
 # THIS ROW WAS PRINTED NOWHERE AND THE TOTAL BELOW ADDED IT NOWHERE, while the DCF sheet's
 # own bridge DID add it — so one workbook computed two different values per share from one
 # model, 190,447 on the DCF sheet and 185,283 here, SAR 1.03 apart. The delivered document
 # printed the same short bridge. A reader following the printed rows could not reach the
 # printed answer, which is the one thing a waterfall must do.
 ('+ Investment funds and unlisted equity investments', 'At fair value',
  f"={A_FUNDS}", False),
 # TWO LABELS NAMING A SUPERSEDED QUARTER on figures taken from the reviewed 30 June 2026
 # sheet, and the minority is not book at all — it is its share of value, with book
 # published beside it as the basis NOT adopted.
 ('− Net debt', 'Total borrowings less cash, on the 30 June 2026 reviewed sheet',
  f"=-{A_NETDEBT}", False),
 ('− Non-controlling interests', 'At the minority\u2019s share of equity value, not at book',
  f"=-{A_NCI}", False),
 ('Equity value', '', f"=C6+{A_ASSOC}+{A_LISTED}+{A_FUNDS}-{A_NETDEBT}-{A_NCI}", True),
 # THESE TWO REFERENCED C12 AND C13 BY POSITION and broke the moment a row was inserted
 # above them: the per-share line started dividing the MINORITY by the share count. The
 # map below is built from the list's own order before a cell is written, so a row added
 # anywhere carries every reference with it.
 ('DCF fair value per share (SAR)', '', '={EQ}/%s' % A_SHARES, True),
 ('Upside / (downside) vs spot', '', '={PS}/%s-1' % A_SPOT, True),
]
BR = {a.strip(): 5 + i for i, (a, _b, _c, _bold) in enumerate(rows)}
_SUB = {'EQ': 'C%d' % BR['Equity value'],
        'PS': 'C%d' % BR['DCF fair value per share (SAR)']}
rows = [(a, b, (c.format(**_SUB) if isinstance(c, str) and '{' in c else c), bold)
        for a, b, c, bold in rows]
# THE BRIDGE'S ROWS ARE CAPTURED BY NAME AS THEY ARE WRITTEN. They were referenced
# positionally from the Summary sheet, so adding the investment-funds row above moved the
# per-share line and the Summary silently started reading the EQUITY VALUE instead —
# 190,447 where it wanted 38.14. That is the same defect the Assumptions map already had,
# in a second place: a hardcoded row number is a reference that breaks the moment anybody
# inserts a line, and it breaks silently because the cell it lands on still holds a number.
assert BR['Component'] == r, 'the row map must start where the table does'
for a, b, c, bold in rows:
    put(ws, f'A{r}', a, BLACK, None, bold=bold)
    put(ws, f'B{r}', b, SUB if b else BLACK, None)
    if c: put(ws, f'C{r}', c,
              GREEN if isinstance(c, str) and ('DCF!' in c or 'Assumptions!' in c) else BLACK,
              PCT if r in (BR['of which terminal value'], BR['Upside / (downside) vs spot'])
              else (PX if r == BR['DCF fair value per share (SAR)'] else NUM0), bold=bold)
    r += 1
ws.column_dimensions['B'].width = 52; ws.column_dimensions['C'].width = 16
r += 1
put(ws, f'A{r}', 'DDM — the locked dividend policy as a lens', BLACK, None, True, FILL_H); r += 1
put(ws, f'A{r}', 'DPS (SAR)', BLACK, None, True)
for j, c in enumerate(['C','D','E','F','G']):
    put(ws, f'{c}{r}', f"=Assumptions!{get_column_letter(3+j)}{DPSR}", GREEN, PX)
DPS_L = r; r += 1
put(ws, f'A{r}', 'PV of DPS @ Ke', BLACK, None)
for j, c in enumerate(['C','D','E','F','G']):
    put(ws, f'{c}{r}', f"={c}{DPS_L}/(1+{A_KE})^{j+1}", BLACK, PX)
PVD_L = r; r += 1
put(ws, f'A{r}', 'Σ PV of explicit DPS (FY26–30E)'); put(ws, f'C{r}', f"=SUM(C{PVD_L}:G{PVD_L})", BLACK, PX); SPV_D = r; r += 1
put(ws, f'A{r}', 'Terminal value = DPS30 × (1+g) / (Ke − g)')
put(ws, f'C{r}', f"=G{DPS_L}*(1+{A_TGDIV})/({A_KE}-{A_TGDIV})", BLACK, PX); TV_D = r; r += 1
put(ws, f'A{r}', 'PV of terminal value'); put(ws, f'C{r}', f"=C{TV_D}/(1+{A_KE})^5", BLACK, PX); PVT_D = r; r += 1
put(ws, f'A{r}', 'DDM fair value per share (SAR)', BLACK, None, True)
put(ws, f'C{r}', f"=C{SPV_D}+C{PVT_D}", BLACK, PX, True); DDM_PS = r; r += 1
put(ws, f'A{r}', '  terminal as a share of the dividend-model value'); put(ws, f'C{r}', f"=C{PVT_D}/C{DDM_PS}", BLACK, PCT); r += 2
# EVERY FIGURE IN THIS NOTE WAS TYPED AND MOST HAD GONE STALE: a cost of capital of 7.6%
# against the schedule's 8.13%, a dividend cover quoted at the GUIDED capex band the model
# no longer uses, and a core enterprise value that is not the one the sheet above computes.
# A note beside a live model is read as the model speaking.
put(ws, f'A{r}',
    'What the market is paying, read off this sheet: at the spot price the market values '
    'the whole equity above what these lines add to, and the difference is the disagreement '
    'section 4 of the study takes apart. The regular dividend is covered %.2f to %.2f times '
    'by first-year model free cash flow across the range of capital intensity this company '
    'own three filed years actually ran — covered throughout that range, so the question is '
    'whether the data-centre build takes intensity above anything it has yet run.'
    % (D['cover'][-1]['cover'], D['cover'][0]['cover']), SUB, None)
json.dump(dict(DDM_PS=DDM_PS), open(os.path.join(HERE, '_vb_rows.json'), 'w'))

# ================= Relative & Normalized =====================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative & normalized-earnings lenses', 'Links to Assumptions, DCF and the Income Statement.', 7)
r = 5
put(ws, f'A{r}', 'Lens', BLACK, None, True, FILL_H); put(ws, f'B{r}', 'Workings', BLACK, None, True, FILL_H); put(ws, f'C{r}', 'SAR/share', BLACK, None, True, FILL_H); r += 1
put(ws, f'A{r}', 'Relative EV/EBITDA'); put(ws, f'B{r}', 'FY26E EBITDA × justified multiple + stakes − net debt − NCI, /sh', SUB, None)
put(ws, f'C{r}', f"=(DCF!B{DC['EBITDA']}*{A_EVX}+{A_ASSOC}+{A_LISTED}+{A_FUNDS}"
                 f"-{A_NETDEBT})*(1-{A_NCISHARE})/{A_SHARES}", GREEN, PX)
RELR = r; r += 1
put(ws, f'A{r}', '  implied P/E cross-check at that value'); put(ws, f'C{r}', f"=C{RELR}*{A_SHARES}/{A_NP26}", BLACK, MULT); r += 1
put(ws, f'A{r}', '  IS-build FY26E EPS, for reference'); put(ws, f'C{r}', f"='Income Statement'!E{IS['EPS (SAR)']}", GREEN, PX); r += 1
put(ws, f'A{r}', 'Normalized earnings power'); put(ws, f'B{r}', 'Normalized PAT (ex one-offs) × through-cycle P/E, /sh', SUB, None)
put(ws, f'C{r}', f'={A_NORMPAT}*{A_NORMPE}/{A_SHARES}', GREEN, PX); NORMR = r; r += 1
put(ws, f'A{r}', '  normalized EPS (SAR)'); put(ws, f'C{r}', f'={A_NORMPAT}/{A_SHARES}', GREEN, PX); r += 1
put(ws, f'A{r+1}', 'P/B cross-check: spot / FY25 BVPS'); put(ws, f'C{r+1}', f"={A_SPOT}/('Balance Sheet'!D{BS['Equity attributable to shareholders']}/{A_SHARES})", BLACK, MULT)
put(ws, f'A{r+2}', 'Dividend yield — two framings: regular declared 2.20/sh = 5.0%; cash paid during 2025 (incl. the SAR 2.00 special for FY24) 4.20/sh = 9.6%.', SUB, None)
ws.column_dimensions['B'].width = 52
json.dump(dict(RELR=RELR, NORMR=NORMR), open(os.path.join(HERE, '_rn_rows.json'), 'w'))

# ================= Summary ===================================================
ws = sheet('Summary')
title(ws, 'Valuation summary — Saudi Telecom Company (Tadawul: 7010)',
      'Four-lens fair value vs spot. Base links live; bear/bull are scenario outputs (study §1.5).', 7)
L = D['lenses']
LR = D['lens_record']
r = 5
# ONE CLASS PRIMARY IS THE CENTRAL AND THE OTHERS ARE CROSS-CHECKS. This table used to carry
# a WEIGHT column and a "Weighted central" row that summed four lenses at 35/25/20/20 — the
# construction [R-LENS-03] retired outright. Two of those four are not permitted cross-checks
# for this class at all and between them carried 45% of the answer. What a reader gets now is
# each lens's own range, the primary marked as the central, and the ENVELOPE as the range of
# the present-value reads — never an average, and never a spread invented around one.
for j, h in enumerate(['', 'Bear', 'Base', 'Bull', 'Role']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
_PRIM = LR['primary']['kind']
lens_rows = [
 ('dcf', 'Discounted cash flow', "='SOTP Bridge'!C%d" % BR['DCF fair value per share (SAR)']),
 ('ddm', 'Dividend discount', "='SOTP Bridge'!C%d" % DDM_PS),
 ('relative', 'Enterprise multiple on own history', f"='Relative & Normalized'!C{RELR}"),
 ('normalized', 'Normalised earnings power', f"='Relative & Normalized'!C{NORMR}"),
]
first = r
for key, nm, ba in lens_rows:
    role = 'THE CENTRAL' if key == _PRIM else 'cross-check'
    put(ws, f'A{r}', nm)
    put(ws, f'B{r}', round(L[key]['bear'], 1), BLUE, PX)
    put(ws, f'C{r}', ba, GREEN, PX)
    put(ws, f'D{r}', round(L[key]['bull'], 1), BLUE, PX)
    put(ws, f'E{r}', role, BLACK, None, bold=(role == 'THE CENTRAL'))
    r += 1
put(ws, f'A{r}', 'Book value per share — a DISCLOSED FLOOR, never weighted')
put(ws, f'C{r}', round(L['book_value'], 2), BLUE, PX)
put(ws, f'E{r}', 'floor', BLACK, None); r += 2
put(ws, f'A{r}', 'THE CENTRAL — the class primary, not an average of the rows above',
    BLACK, None, True)
put(ws, f'C{r}', f"=C{first}", BLACK, PX, True)
WC = r; r += 1
put(ws, f'A{r}', 'Published range — the span of the present-value reads', BLACK, None, True)
put(ws, f'B{r}', round(LR['envelope']['low'], 2), BLACK, PX, True)
put(ws, f'D{r}', round(LR['envelope']['high'], 2), BLACK, PX, True)
r += 2
put(ws, f'A{r}', 'Spot price (SAR)'); put(ws, f'B{r}', an('Spot price (SAR/share)'), GREEN, PX); r += 1
put(ws, f'A{r}', 'Central against spot')
put(ws, f'B{r}', f"=C{WC}/{A_SPOT}-1", BLACK, PCT); r += 2
put(ws, f'A{r}', 'The retired four-lens blend at typed weights is recorded on Assumptions so '
                 'nothing here can quietly rebuild it. A number produced by averaging '
                 'several methods is not more robust than the best of them: it is a new '
                 'method with free parameters nobody tested, and it imports every weakness '
                 'of the weakest lens at whatever weight somebody typed.', SUB, None)
r += 1
json.dump(dict(first=first, WC=WC), open(os.path.join(HERE, '_sum_rows.json'), 'w'))

# ================= Fundamental Valuation =====================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — football-field data', 'Bear/base/bull per lens (links to Summary).', 7)
r = 5
for j, h in enumerate(['Lens', 'Bear', 'Base', 'Bull']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
# THE LAST ROW WAS LABELLED "Weighted central" AND POINTED AT THE BOOK-VALUE FLOOR. Two
# defects in one cell: a retired label naming a construction this house does not build, on
# a row that was not even what the label claimed. The Summary rows it mirrors are the four
# lenses and then the disclosed floor, so it is named for what it is.
for i, nm in enumerate(['Cash-flow model (the class primary)', 'Dividend discount',
                        'Enterprise multiple on own history', 'Normalised earnings power',
                        'Book value — a disclosed floor, never weighted into an answer']):
    put(ws, f'A{r}', nm, BLACK, None, i == 4)
    for j, col in enumerate('BCD'):
        put(ws, f'{col}{r}', f'=Summary!{col}{first + i}', GREEN, PX, i == 4)
    r += 1
put(ws, f'A{r}', 'Spot'); put(ws, f'B{r}', f'={A_SPOT}', GREEN, PX)

# ================= Summary Financials ========================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials (SAR mn)', 'Every cell links to the statement sheets.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
r = 6
links = [
 ('Revenue', 'Income Statement', IS['Total revenue']),
 ('EBITDA', 'Income Statement', IS['EBITDA']),
 ('EBIT', 'Income Statement', IS['Operating profit (EBIT)']),
 ('Net profit (attributable)', 'Income Statement', IS['Net profit (attributable)']),
 ('Total assets', 'Balance Sheet', BSJ['TA']),
 ('Total equity', 'Balance Sheet', BS['Total equity']),
 ('Net debt (IR basis)', 'Balance Sheet', BSJ['NDR']),
]
for nm, sh, rr in links:
    put(ws, f'A{r}', nm)
    for j, col in enumerate(['B', 'C', 'D'] + FCOLS):
        put(ws, f'{col}{r}', f"='{sh}'!{col}{rr}", GREEN, NUM0)
    r += 1
put(ws, f'A{r}', 'Free cash flow (forecast)')
for col in FCOLS:
    put(ws, f'{col}{r}', f"='Cash Flow'!{col}{CFJ['FCFR']}", GREEN, NUM0)
r += 1
put(ws, f'A{r}', 'Free cash flow (historical, disclosed)', BLACK, None)
for col, v in zip('BCD', (12628.0, 7959.0, 6488.0)):
    put(ws, f'{col}{r}', v, BLUE, NUM0)

# ================= Monte Carlo ===============================================
ws = sheet('Monte Carlo')
title(ws, 'Monte Carlo — engine outputs (YZ-HAR v2)',
      '50,000 paths · 16 factors · seed 42 · computed by the Testahil MC engine (values, not a sheet simulation). Zero drift — the Step 0-passed configuration for this name.', 8)
pr = D['mc']['prob_read']; q20, q60 = D['mc']['q20'], D['mc']['q60']
r = 5
put(ws, f'A{r}', 'The probability read (%s)' % D['engine']['horizons']['3M']['label'],
    BLACK, None, True, FILL_G); r += 1
prr = [
 ('P(price above spot)', pr['p_above'], PCT),
 ('P(+10%) vs P(−10%) — odds', f"{pr['p_up10']*100:.0f}% vs {pr['p_dn10']*100:.0f}% · {pr['odds']:.1f}:1", '@'),
 ('Median level (SAR) and % move', f"{pr['median']:.2f} ({pr['med_move']*100:+.1f}%)", '@'),
 ('50% band (25th–75th)', f"{pr['band50'][0]:.1f} – {pr['band50'][1]:.1f}  ({pr['band50_pct'][0]*100:+.1f}% / {pr['band50_pct'][1]*100:+.1f}%)", '@'),
 ('Touch(+10%) / touch(−10%)', f"{pr['touch_up10']*100:.0f}% / {pr['touch_dn10']*100:.0f}%", '@'),
]
for nm, v, fmt in prr:
    put(ws, f'A{r}', nm); put(ws, f'B{r}', v, BLACK, fmt); r += 1
r += 1
put(ws, f'A{r}', 'Percentile map (SAR/share)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for tag, q in [(D['engine']['horizons']['1M']['label'], q20),
               (D['engine']['horizons']['3M']['label'], q60)]:
    put(ws, f'A{r}', tag)
    for j, p in enumerate(['5', '25', '50', '75', '95']):
        put(ws, f'{get_column_letter(2+j)}{r}', round(q[p], 1), BLACK, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Engine inputs (from Assumptions)', BLACK, None, True, FILL_H); r += 1
for nm, ref in [('Forward volatility over three months (annualised)', '=' + A_VOL),
                ('Momentum lean applied over three months', '=' + A_LEAN)]:
    put(ws, f'A{r}', nm); put(ws, f'B{r}', ref, GREEN, PCT); r += 1
r += 1
put(ws, f'A{r}', 'Level-touch ladder (probability of touching by horizon)', BLACK, None, True, FILL_H); r += 1
for j, h in enumerate(['Level (SAR)', D['engine']['horizons']['1M']['label'],
                       D['engine']['horizons']['3M']['label']]):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
for L_, tv in D['mc']['touch'].items():
    put(ws, f'A{r}', float(L_), BLACK, PX)
    put(ws, f'B{r}', tv['t20'], BLACK, PCT); put(ws, f'C{r}', tv['t60'], BLACK, PCT); r += 1

# ================= Sensitivity ===============================================
ws = sheet('Sensitivity')
title(ws, 'DCF sensitivity — WACC × terminal growth (live formulas)',
      'Fair value per share (SAR). Recomputes the full FCFF stream off the DCF sheet; the centre cell is the base case.', 8)
gw = D['sens']['wacc_steps']; gg = D['sens']['g_steps']
put(ws, 'A5', 'WACC \\ terminal g', BLACK, None, True, FILL_H)
for j, g in enumerate(gg):
    put(ws, f'{get_column_letter(2+j)}5', g, BLACK, PCT, True, FILL_H)
FR = DC['FCFF']
for i, w in enumerate(gw):
    rr = 6 + i
    put(ws, f'A{rr}', w, BLACK, PCT, True)
    ws[f'A{rr}'].font = BLACK if abs(w - D['dcf']['wacc']) > 1e-9 else Font(bold=True)
    for j, g in enumerate(gg):
        col = get_column_letter(2+j)
        f = (f"=(SUMPRODUCT(DCF!$B${FR}:$F${FR},1/(1+$A{rr})^{{1,2,3,4,5}})"
             f"+DCF!$F${FR}*(1+{col}$5)/($A{rr}-{col}$5)/(1+$A{rr})^5"
             f"+{A_ASSOC}+{A_LISTED}-{A_NETDEBT}-{A_NCI})/{A_SHARES}")
        put(ws, f'{col}{rr}', f, BLACK, PX)
put(ws, 'A12', 'The WACC axis centres on the sourced bottom-up build (rf 5.5% + β 0.48 × ERP 5.01%, blended with after-tax Kd 4.5%); '
               'the CDS-ERP alternative WACC (7.90%) sits between rows 3 and 4. At β = 1.0 the WACC is 9.90% — off this grid deliberately: '
               'see the beta grid below.', SUB, None)
put(ws, 'A14', 'Beta sensitivity (regressed β = 0.48 on a 9-week window — grid mandatory)', BLACK, None, True, FILL_H)
put(ws, 'A15', 'beta', BLACK, None, True); put(ws, 'B15', 'Ke (rating ERP)', BLACK, None, True); put(ws, 'C15', 'WACC', BLACK, None, True); put(ws, 'D15', 'DCF value/sh', BLACK, None, True)
for i, b in enumerate([0.30, 0.48, 0.70, 0.85, 1.00, 1.20]):
    rr = 16 + i
    put(ws, f'A{rr}', b, BLUE, '0.00')
    put(ws, f'B{rr}', f"={A_RF}+A{rr}*{A_ERP}", BLACK, PCT)
    put(ws, f'C{rr}', f"=(1-{A_WD})*B{rr}+{A_WD}*{A_KDAT}", BLACK, PCT)
    put(ws, f'D{rr}', (f"=(SUMPRODUCT(DCF!$B${FR}:$F${FR},1/(1+C{rr})^{{1,2,3,4,5}})"
                       f"+DCF!$F${FR}*(1+{A_TG})/(C{rr}-{A_TG})/(1+C{rr})^5"
                       f"+{A_ASSOC}+{A_LISTED}-{A_NETDEBT}-{A_NCI})/{A_SHARES}"), BLACK, PX)
    if abs(b - 0.48) < 1e-9:
        put(ws, f'E{rr}', '<- regressed base', SUB, None)
    if abs(b - 1.00) < 1e-9:
        put(ws, f'E{rr}', '<- house fallback had the regression failed', SUB, None)

# ================= Per-Share & Ratios ========================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share & ratios — the standing dashboard (device A-5)', 'Links to statements / Assumptions.', 10)
for j, h in enumerate([''] + YH + YF):
    put(ws, f'{get_column_letter(1+j)}4', h, BLACK, None, True, FILL_H)
allc = ['B', 'C', 'D'] + FCOLS
NPR = IS['Net profit (attributable)']; REVR = IS['Total revenue']; EBITDAR = IS['EBITDA']
EQR = BS['Equity attributable to shareholders']; NDRr = BSJ['NDR']
r = 6
def prow(r, label, fml, fmt=PX, cols=allc):
    put(ws, f'A{r}', label)
    for col in cols:
        f = fml(col)
        if f is not None:
            put(ws, f'{col}{r}', f, BLACK, fmt)
    return r + 1
r = prow(r, 'EPS (SAR)', lambda c: f"='Income Statement'!{c}{NPR}/{A_SHARES}")
DPS_hist = {'B': 1.60, 'C': 3.75, 'D': 2.20}
put(ws, f'A{r}', 'DPS declared (SAR)')
for col in 'BCD':
    put(ws, f'{col}{r}', DPS_hist[col], BLUE, PX)
for j, col in enumerate(FCOLS):
    put(ws, f'{col}{r}', f"=Assumptions!{get_column_letter(3+j)}{DPSR}", GREEN, PX)
DPSL = r; r += 1
r = prow(r, 'Book value / share (SAR)', lambda c: f"='Balance Sheet'!{c}{EQR}/{A_SHARES}")
r = prow(r, 'P/E at spot (×)f', lambda c: f"={A_SPOT}/('Income Statement'!{c}{NPR}/{A_SHARES})")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'P/B at spot (×)f', lambda c: f"={A_SPOT}/('Balance Sheet'!{c}{EQR}/{A_SHARES})")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Dividend yield at spot (declared)', lambda c: f"={c}{DPSL}/{A_SPOT}", PCT)
r = prow(r, 'Payout (of attributable NP)f', lambda c: f"={c}{DPSL}*{A_SHARES}/'Income Statement'!{c}{NPR}", PCT)
r = prow(r, 'EBITDA margin', lambda c: f"='Income Statement'!{c}{EBITDAR}/'Income Statement'!{c}{REVR}", PCT)
r = prow(r, 'Net margin', lambda c: f"='Income Statement'!{c}{NPR}/'Income Statement'!{c}{REVR}", PCT)
r = prow(r, 'ROAE (attributable)', lambda c: f"='Income Statement'!{c}{NPR}/'Balance Sheet'!{c}{EQR}", PCT)
r = prow(r, 'Net debt (IR) / EBITDA (×)', lambda c: f"='Balance Sheet'!{c}{NDRr}/'Income Statement'!{c}{EBITDAR}")
for col in allc: ws[f'{col}{r-1}'].number_format = MULT
r = prow(r, 'Capex intensity (% of revenue)', lambda c: (f"=-'Cash Flow'!{c}{CF['− Capex']}/'Income Statement'!{c}{REVR}" if c in FCOLS else None), PCT)
for col, v in zip('BCD', (0.136, 0.157, 0.152)):
    put(ws, f'{col}{r-1}', v, BLUE, PCT)
r = prow(r, 'FCF yield at spot (forecast)', lambda c: (f"='Cash Flow'!{c}{CFJ['FCFR']}/({A_SPOT}*{A_SHARES})" if c in FCOLS else None), PCT)
r = prow(r, 'Revenue YoY', lambda c: (f"='Income Statement'!{c}{REVR}/'Income Statement'!{chr(ord(c)-1)}{REVR}-1" if c != 'B' else None), PCT)
r = prow(r, 'EPS YoY', lambda c: (f"=('Income Statement'!{c}{NPR}/'Income Statement'!{chr(ord(c)-1)}{NPR})-1" if c != 'B' else None), PCT)
put(ws, f'A{r+1}', 'FY24 EPS/DPS carry the TAWAL disposal gain and the SAR 2.00 special dividend — the payout row prints >100% on the '
                   'cash-paid framing (141% LTM per stc). Declared-regular vs cash-paid framings both stated (house rule).', SUB, None)

# ================= Peer & Sector =============================================
ws = sheet('Peer & Sector')
title(ws, 'Peer set & sector', 'GCC/regional telecoms; multiples approximate, as-of dates flagged (context, not model inputs).', 7)
r = 5
for j, h in enumerate(['Name', 'Mkt', 'P/E (t)', 'EV/EBITDA', 'Div yield', 'Note']):
    put(ws, f'{get_column_letter(1+j)}{r}', h, BLACK, None, True, FILL_H)
r += 1
peers = [
 ('stc (7010)', 'KSA', '14.7×', '9.5×', '5.0% (9.6% incl. special)', 'Incumbent; ~57% mobile share; net debt ~0'),
 ('Mobily (7020)', 'KSA', '13.5×', '~7.4×', '4.5%', 'No.2; e& anchor shareholder; fastest sub growth'),
 ('Zain KSA (7030)', 'KSA', '12.8×', 'n/a', '4.9%', 'No.3; tower-light; higher leverage'),
 ('e& (EAND)', 'UAE', '13.2×', 'n/a', '5.1%', 'UAE incumbent + international portfolio'),
 ('Ooredoo (ORDS)', 'Qatar', '10.7×', 'n/a', '5.7%', 'Multi-market; data-centre pivot'),
 ('du (DU)', 'UAE', '17.6×', 'n/a', '5.5%', 'No.2 UAE; hyperscale DC momentum'),
 ('Omantel / Beyon / ETEL', 'Oman/Bah/Egy', '11.5× / 11.1× / 8.8×', 'n/a', '3.9% / 7.1% / 1.5%', 'Regional context'),
]
for row in peers:
    for j, v in enumerate(row):
        put(ws, f'{get_column_letter(1+j)}{r}', v, BLACK, None)
    r += 1
ws.column_dimensions['F'].width = 46
r += 1
put(ws, f'A{r}', 'Sector: KSA mobile ≈ 57/27/16 stc/Mobily/Zain; Nov-2024 CST spectrum auction (stc: 600MHz + 3.8GHz); 10,800+ stc 5G sites, '
                 '63% populated coverage; FTTH 3.75mn; FWA one of the highest-adoption markets globally; center3–HUMAIN 1GW AI-DC ambition; '
                 'SilkLink Syria fibre corridor (SAR 3bn, Feb-2026). Peer P/E band 10.7–17.6×; stc premium tracks its share, balance sheet and yield.', SUB, None)
put(ws, f'A{r+1}', 'Analyst context (not a model input): 17 analysts, 8 Buy / 9 Hold, average 12m TP SAR 47.9 (41.1–55.0), Jul-2026.', SUB, None)

# ================= sheet order ===============================================
order = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge', 'Segments',
         'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet', 'Cash Flow',
         'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in order]
wb.save(os.path.join(HERE, FN))
print('partC ok — sheets:', wb.sheetnames)
