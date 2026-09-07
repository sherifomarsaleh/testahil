"""STC_Valuation_Model_05092026_public.xlsx — part A: READ FIRST, Assumptions, Segments,
DCF, Income Statement, Balance Sheet, Cash Flow. 16 sheets mirroring the TMPV canonical model.
Blue = inputs · black = formulas · green = cross-sheet links. All inputs live on Assumptions."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

HERE = os.path.dirname(os.path.abspath(__file__))
# PATHS ARE ABSOLUTE AGAINST THIS FILE'S OWN DIRECTORY. They were relative to the
# working directory, so running the build from the repository root — which is how
# every gate and the CI runner invoke things — read no inputs and scattered outputs.
# A path relative to cwd is a path that depends on who ran it.


D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
_BETA_REC = json.load(open(os.path.join(HERE, 'beta_result.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM = '#,##0.0;(#,##0.0);"-"'; NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'; MULT = '0.00x'; PX = '0.00'
FN = 'STC_Valuation_Model_05092026_public.xlsx'


def longdate(iso):
    """'2026-09-03' -> '3 September 2026'. A DATE STATED IN PROSE IS A FIGURE AND IS
    COMPUTED, NOT TYPED: this builder carried a spot date and a balance-sheet date that
    had both been superseded by a re-strike, and each read as ordinary sentence text."""
    y, m, d = (int(x) for x in iso.split('-'))
    return '%d %s %d' % (d, ('January', 'February', 'March', 'April', 'May', 'June', 'July',
                             'August', 'September', 'October', 'November', 'December')[m - 1], y)

wb = Workbook()

def sheet(name):
    ws = wb.create_sheet(name) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = name
    return ws

def title(ws, text, sub=None, width=10):
    ws['A1'] = text; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, width + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if sub:
        ws['A2'] = sub; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = 44
    for c in range(2, width + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12.5

def put(ws, addr, val, font=BLACK, fmt=None, bold=False, fill=None):
    c = ws[addr]; c.value = val
    c.font = Font(color=(font.color if font else '000000'), bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

YH = ['FY23', 'FY24', 'FY25']; YF = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']

# ============ READ FIRST =====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Saudi Telecom Company (Tadawul: 7010)', width=9)
lines = [
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the stc valuation study. Every blue cell is an input; every',
 'black cell is a formula; green cells link across sheets. All inputs live on the Assumptions sheet — change one',
 '(the EBITDA margin path, capex intensity, the beta, terminal growth, a stake mark) and the whole model reprices.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs shown',
 'as ranges. The preparer is not licensed by any securities regulator and may hold a position in the security.', '',
 'Entity note. Saudi Telecom Company (stc Group) is the Kingdom’s incumbent operator: stc KSA (consumer,',
 'enterprise, wholesale) plus subsidiaries — solutions by stc (79%, listed 7202), stc bank (85%, SAMA-licensed',
 'Jan-2025), stc Kuwait (51.8%), stc Bahrain (100%), center3 (data centres), sirar, iot squared (50% PIF JV) —',
 'and minority stakes: 43.06% of Digital Infrastructure Co (TAWAL+GLIC towers, PIF-controlled) and 9.97% of',
 'Telefónica. House lens: going-concern FCFF DCF (primary), cross-checked by the dividend-policy DDM,',
 'relative multiples and normalized earnings; the tower and Telefónica stakes are bridge items marked separately.', '',
 'Currency. SAR million unless stated. Spot SAR %.2f (%s close, the latest known price this study is\n'
 ' struck against; the Monte Carlo cone starts at the last session in the price library, %s, which is a\n'
 ' different date and is stated as one).' % (D['spot'], longdate(D['spot_date']), longdate(D['cone_anchor_date'])),
 'Historical financials are the company’s own audited annual and reviewed interim statements, FY2023',
 'onward, each fetched from its own investor-relations archive and listed with that URL in the study’s',
 'bibliography. The balance sheet the bridge stands on is the reviewed interim to %s.' %
 longdate(D['bridge_record']['balance_sheet_date']), '',
 'Sheets: Summary · Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized ·',
 'DCF · Income Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity ·',
 'Per-Share & Ratios · Peer & Sector.']
for i, ln in enumerate(lines, start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 118

# ============ ASSUMPTIONS ====================================================
wa = sheet('Assumptions')
title(wa, 'Assumptions — the single input layer', 'All blue cells are inputs. Every other sheet links here.', 9)
r = 4
def hdr(ws_, row, text):
    put(ws_, f'A{row}', text, bold=True, fill=FILL_H); return row + 1
#: EVERY ANCHOR'S ROW IS RECORDED BY NAME. The other sheets used to reach into this one by
#: hardcoded cell number — Assumptions!$B$16 for the cost of capital, $B$17 for terminal
#: growth, $B$19 to $B$22 for the bridge — so ADDING ONE ROW here silently re-pointed all of
#: them at their neighbours. It did: publishing a second tax rate shifted the block down one
#: and the workbook's terminal value came out 58% high while every formula recalculated
#: without an error. A reference by position is a reference that moves when anything above
#: it does.
ANCHOR_ROWS = {}


def inp(ws_, row, label, val, fmt=NUM, note=None):
    put(ws_, f'A{row}', label)
    put(ws_, f'B{row}', val, BLUE, fmt)
    if note: put(ws_, f'C{row}', note, SUB)
    ANCHOR_ROWS[label] = row
    return row + 1


def anchor(label, row):
    """Record a row written directly rather than through inp()."""
    ANCHOR_ROWS[label] = row
    return row
# EVERY ANCHOR BELOW COMES OUT OF THE COMMITTED RECORD. This block used to be fourteen
# typed financial numerals — the numeric-traceability standard's own prohibition — and every
# one of them was the PRE-REBUILD study's: a spot four weeks stale, the 40-session daily
# beta the rule refuses as a study beta, associates at 4,641 against a filed 12,909.648, a
# net debt that omitted the spectrum-licence liability, a minority at book rather than at
# its share of value, and a terminal growth typed as a nominal rate. A builder that types
# its anchors does not go stale loudly; it goes stale silently and recalculates perfectly.
_CR = D['coc_record']
_BR = D['bridge_record']
_TR = D['terminal_record']
_FIS = D['forecast_is']
_IS = json.load(open(os.path.join(HERE, 'income_statement.json')))
# the shield the sanctioned schedule itself used, recovered from its own record
_KD_SHIELD = 1.0 - _CR['kd_aftertax'] / _CR['kd_pretax']
_REL = D['rel_basis']

r = hdr(wa, r, 'ANCHORS')                                                  # 4
r = inp(wa, r, 'Spot price (SAR/share)', D['spot'], PX,
        'the latest known close, %s, and the price the published gap is measured against'
        % D['spot_date'])                                                  # B5
r = inp(wa, r, 'Shares outstanding (mn)', _BR['shares_mn'], NUM0,
        'issued capital divided by par, less treasury, FOOTED against the count note 17 of '
        'the reviewed interim itself states')                             # B6
# TWO RATES, EACH ON ITS OWN BASE, and B7 is the one after-tax operating profit needs. A
# rate measured against profit before zakat cannot be applied to EBIT: the lines between
# them are a net charge on this book, so the two differ. And BOTH exclude the reversal of
# prior years' zakat that note 33(a) names on its own line — a rate carrying it forward
# assumes the company keeps discovering it has over-provided, for ever.
r = inp(wa, r, 'Effective zakat rate ON EBIT (used for after-tax operating profit)',
        _IS['effective_zakat_rate_on_ebit'], PCT,
        'three filed years together, with the disclosed prior-year reversal of %s put back. '
        'Carrying that reversal forward would read %.2f%%.'
        % (f"{_IS['zakat_reversal_fy2025']:,}",
           100 * _IS['zakat_rate_carrying_the_reversal']))                 # B7
r = inp(wa, r, 'Effective zakat rate on profit before zakat (income statement)',
        _FIS['zakat_rate'], PCT,
        'the same charge on the base the income statement applies it to')
r = hdr(wa, r, 'COST OF CAPITAL — the sanctioned schedule, not built on this sheet')  # 8
r = inp(wa, r, 'Risk-free rate, normalised by the sovereign default spread',
        _CR['rf_star'], PCT2,
        'the observed sovereign yield of %.2f%% less this sovereign\'s OWN default spread '
        'of %.2f%%. Country risk enters ONCE, through the premium; the raw yield plus a '
        'premium already carrying it would count it twice.'
        % (100 * _CR['rf_observed'], 100 * _CR['default_spread']))         # B9
r = inp(wa, r, 'Equity beta — own-stock weekly regression vs the published index',
        _CR['beta'], '0.0000',
        'against the index of the exchange the stock is listed on, through the sanctioned '
        'resolver. The delivered study used a 40-session DAILY regression at 0.48, which is '
        'not one of the three tiers and may stand only as a flagged interim.')  # B10
r = inp(wa, r, 'Equity risk premium (%s basis, central)' % _CR['erp_basis'], _CR['erp'],
        PCT2, 'both bases are published; the swap basis is central because it is the '
              "market's own live pricing of the sovereign's credit against an agency "
              'judgement updated in steps')                                # B11
put(wa, f'A{r}', 'Cost of equity Ke = rf* + beta x ERP')
put(wa, f'B{r}', '=B%d+B%d*B%d'
    % (ANCHOR_ROWS['Risk-free rate, normalised by the sovereign default spread'],
       ANCHOR_ROWS['Equity beta — own-stock weekly regression vs the published index'],
       ANCHOR_ROWS['Equity risk premium (%s basis, central)' % _CR['erp_basis']]),
    BLACK, PCT2)
put(wa, f'C{r}', 'reproduces the schedule\'s own %.3f%% from the three cells above it'
                 % (100 * _CR['ke_exp']), SUB)
r = anchor('Cost of equity', r) + 1                                        # B12
r = inp(wa, r, 'Pre-tax cost of debt — the company\'s own latest issue', _CR['kd_pretax'],
        PCT2,
        'the January 2026 sukuk in two tranches, weighted. Held against an effective rate '
        'computed independently over two periods from the finance cost on the borrowings '
        'that actually bear it.')                                          # B13
put(wa, f'A{r}', 'After-tax Kd')
# THE DEBT SHIELD IS NOT THE ZAKAT RATE. Interest is deductible against income tax, and the
# schedule uses the statutory rate for it; applying the effective zakat rate here would
# shield the debt at a rate no authority allows.
put(wa, f'B{r}', '=B%d*(1-%.6f)' % (ANCHOR_ROWS["Pre-tax cost of debt — the company's own "
                                                "latest issue"], _KD_SHIELD), BLACK, PCT2)
r = anchor('After-tax Kd', r) + 1                                          # B14
put(wa, f'A{r}', 'Debt weight D/(D+E) — MARKET-VALUE equity, never book')
put(wa, f'B{r}', '=%.3f/(B%d*B%d+%.3f)'
    % (_CR['gross_debt'] / 1000.0, ANCHOR_ROWS['Spot price (SAR/share)'],
       ANCHOR_ROWS['Shares outstanding (mn)'], _CR['gross_debt'] / 1000.0), BLACK, PCT2)
put(wa, f'C{r}', 'gross borrowings from the LATEST DISCLOSED balance sheet', SUB)
r = anchor('Debt weight', r) + 1                                           # B15
put(wa, f'A{r}', 'WACC — FLAT, because the riyal is pegged and today is already the terminal')
put(wa, f'B{r}', '=(1-B%d)*B%d+B%d*B%d'
    % (ANCHOR_ROWS['Debt weight'], ANCHOR_ROWS['Cost of equity'],
       ANCHOR_ROWS['Debt weight'], ANCHOR_ROWS['After-tax Kd']), BLACK, PCT2)
put(wa, f'C{r}', 'the module returns a flat ladder here rather than manufacturing a glide '
                 'the peg forbids; explicit %.3f%% and terminal %.3f%%'
                 % (100 * _CR['wacc_exp'], 100 * _CR['wacc_terminal']), SUB)
r = anchor('WACC', r) + 1                                                  # B16
r = inp(wa, r, 'Terminal growth — DERIVED, terminal inflation + stated real growth',
        _TR['nominal_growth'], PCT2,
        'real growth of %.1f%% on the house terminal inflation. A typed nominal rate is '
        'unfalsifiable: nobody can tell whether 2.5%% meant inflation plus half a point or '
        'something else.' % (100 * _TR['real_growth']))                    # B17
r = hdr(wa, r, 'EV -> EQUITY BRIDGE — every line from the committed bridge record')  # 18
r = inp(wa, r, 'Investments in associates and joint ventures', _BR['associates']['value'],
        NUM0,
        'the FILED carrying value. The delivered study carried 4,641 — a figure from before '
        'the towers business was contributed to DIIC in February 2025 — and correcting it '
        'raised the answer by 3.55%.')                                     # B19
r = inp(wa, r, 'Listed equity investment at its disclosed fair value',
        D['dcf']['telefonica'], NUM0, 'marked at the disclosed fair value, not at cost')  # B20
# THE BRIDGE HAD A LINE THE SHEET DID NOT CARRY, and the reconciliation is what found it:
# everything down to enterprise value matched the model to zero while the equity value came
# out 2.7% short. Investment funds and unlisted equity investments at fair value are 5,163mn
# — a real asset outside the telecom cash flows the model discounts, so they are added
# rather than left out, exactly as the listed stake above is.
r = inp(wa, r, 'Investment funds and unlisted equity investments, at fair value',
        [l['value'] for l in _BR['lines']
         if l['name'].startswith('plus investment funds')][0], NUM0,
        'note 9.1, outside the operating cash flows this model discounts')
r = inp(wa, r, 'Net debt', _BR['net_debt_build']['net'], NUM0,
        'borrowings %.0f + leases %.0f + spectrum licences %.0f less cash %.0f, murabahas '
        '%.0f, sukuk %.0f and treasury bills %.0f. THE SPECTRUM LICENCE LIABILITY is '
        'consideration owed for licences already capitalised and it is disclosed OUTSIDE '
        'borrowings, so a bridge reading the borrowings lines does not see it.'
        % (_BR['net_debt_build']['borrowings'], _BR['net_debt_build']['leases'],
           _BR['net_debt_build']['spectrum_licences'],
           _BR['net_debt_build']['cash_non_bank'], _BR['net_debt_build']['murabahas'],
           _BR['net_debt_build']['sukuk'], _BR['net_debt_build']['treasury_bills']))  # B21
r = inp(wa, r, 'Non-controlling interests, at their share of equity value',
        -[l['value'] for l in _BR['lines']
          if l['name'].startswith('less the minority')][0], NUM0,
        'the model capitalises 100%% of subsidiary cash flow, so the minority\'s claim is '
        'worth its SHARE OF THAT VALUE and not its historical cost. Book is %.0f and is '
        'published beside it so a reader sees the choice.' % _BR['nci']['book'])  # B22
r = hdr(wa, r, 'DDM (the locked SAR 0.55/quarter policy)')                  # 23
put(wa, f'A{r}', 'DPS path FY26E–FY30E (SAR)', BLACK);
for j, v in enumerate([2.20, 2.20, 2.30, 2.40, 2.55]):
    put(wa, f'{get_column_letter(3+j)}{r}', v, BLUE, PX)
DPS_ROW = r; r += 1                                                          # 24
r = inp(wa, r, 'DDM terminal dividend growth', 0.030, PCT2, 'post-policy (2027+) payout growth ≈ EPS growth at ~75% payout')  # B25
r = inp(wa, r, 'Minority share of equity value (proportional)',
        _BR['nci']['profit_share'], PCT,
        "the minority's own share of profit, which is how the cross-check lenses apply it: "
        'proportionally to the whole equity value including the stakes, rather than as a '
        'fixed deduction')
r = hdr(wa, r, 'CROSS-CHECKS — published beside the central, never averaged into it')  # 26
# THE MULTIPLE IS THE COMPANY'S OWN, NOT A BAND SOMEBODY TYPED. This was 9.0x with a note
# reading "GCC band ~8-10x" — a judgement about peers, and the rebuild found it happened to
# be almost exactly what the stock traded at, which is the circularity a relative lens must
# not have: a multiple taken from the current price values the company at what it already
# trades at. It is now the company's OWN trailing enterprise multiple at each of the last
# three financial year ends.
r = inp(wa, r, 'EV/EBITDA — the company\'s own trailing multiple, three year ends',
        _REL['evx']['base'], MULT,
        'bear %.3fx to bull %.3fx, the range its own three year ends span. NOT a peer band '
        'and NOT the current multiple.' % (_REL['evx']['bear'], _REL['evx']['bull']))
r = inp(wa, r, 'FY26E net profit for the relative cross-check', _REL['np26'], NUM0,
        'from the projected income statement, not aligned by hand')
r = inp(wa, r, 'Normalized PAT (ex one-offs)', _REL['norm_pat'], NUM0)
r = inp(wa, r, 'Justified through-cycle P/E', 15.0, MULT,
        'a stated judgement rather than a measurement, and it is a CROSS-CHECK: it carries '
        'no weight in the central')
# THE FOUR SYNTHESIS WEIGHTS ARE GONE. They were 35/25/20/20 and they are the construction
# [R-LENS-03] retired: one class primary IS the central, every other lens is a cross-check
# published beside it, and the envelope is the RANGE of the present-value reads. A number
# produced by averaging several methods is not more robust than the best of them — it is a
# new method with free parameters nobody tested, wearing the appearance of caution.
r = hdr(wa, r, 'THE RETIRED BLEND — recorded so nothing here can quietly rebuild it')
put(wa, f'A{r}', 'Former weights: cash flow 35%, dividend discount 25%, relative 20%, '
                 'normalised earnings 20%', BLACK, None)
put(wa, f'C{r}', 'RETIRED. Two of those four lenses are not permitted cross-checks for this '
                 'class at all and between them carried 45% of a central.', SUB)
r += 1
# THESE ROWS DESCRIBED A MODEL THIS HOUSE DOES NOT RUN. They named an engine version that
# is not the production one, an anchor volatility from a study-local simulation, a "secular
# drift failed (-4.8%)" note — a score against a naive benchmark, which is the retired
# verdict machinery and may not reach a reader in any form — and a sixteen-factor stack of
# typed event probabilities. Every row below is the production strike's own record.
_EN = D['engine']
_H3 = _EN['horizons']['3M']
r = hdr(wa, r, 'THE PRICE MAP — the production engine, reproduced (outputs on the Monte Carlo sheet)')
r = inp(wa, r, 'Forward volatility over three months (annualised)',
        round(_H3['anchor_vol_ann'], 4), PCT,
        'Projected from this stock\u2019s own recent, medium and longer-run variation')
r = inp(wa, r, 'Tail parameter and width calibration', '%.1f / %.3f' % (_EN['nu'], _EN['width_cal']), '@',
        'The market panel\u2019s fitted pair. Neither is quoted alone: they trade off, and the honest '
        'object is the cone they jointly produce')
r = inp(wa, r, 'Momentum lean applied over three months', round(_H3['signal_alpha'], 6), PCT2,
        'Direction call %s (reading %+.2f against a %.2f threshold); capped at the strength measured'
        % (_EN['call'].upper(), _EN['signal_z'], _EN['dead_zone']))
r = inp(wa, r, 'Paths / seed', '%s / %d' % (format(_EN['n_paths'], ','), _EN['seed']), '@')
r = hdr(wa, r, 'FORECAST DRIVERS (FY26E–FY30E) — top-down (§3.5-C gate: subs × ARPU not disclosed)')  # 41
put(wa, f'A{r}', 'Driver \\ year', bold=True)
for j, y in enumerate(YF):
    put(wa, f'{get_column_letter(3+j)}{r}', y, bold=True, fill=FILL_H)
r += 1
DRV = {}
def drv(row, label, vals, fmt=PCT):
    put(wa, f'A{row}', label)
    for j, v in enumerate(vals):
        put(wa, f'{get_column_letter(3+j)}{row}', v, BLUE, fmt)
    DRV[label] = row
    return row + 1
dr = D['drivers']
# THE DISCLOSED OPERATING SEGMENTS, EACH ON ITS OWN REAL RATE. This sheet used to carry four
# BUSINESS UNITS — consumer, enterprise, wholesale and a "subsidiaries" residual — which are
# not what the company reports. Note 9 discloses eleven segments and the model is built on
# them, so this is the finest sourced level rather than a grouping the study chose.
#
# THE RATES ARE REAL AND THE NOMINAL IS DERIVED. A typed nominal growth rate is
# unfalsifiable: nobody reading the sheet can tell whether it meant inflation plus a point
# or inflation minus three. Each segment carries its own measured real rate and the house
# Saudi inflation ladder sits beside it, so every nominal in the model recomputes from two
# figures on this page.
# THE FADE IS PUBLISHED PER YEAR, not as one rate repeated. The model fades each segment's
# real rate to zero by the last explicit year — a segment still growing in real terms in
# year five would be capitalised at a rate it never reached — and a sheet carrying ONE rate
# compounded it flat for five years and reached FY2030 revenue three per cent above the
# model. IT RECALCULATED PERFECTLY TO THE WRONG ANSWER, which is why a clean recalculation
# is necessary and not sufficient.
_seg_path = dr['segment_real_growth_path']
for _name in sorted(_seg_path, key=lambda k: -abs(_seg_path[k][0])):
    r = drv(r, '%s — real revenue growth' % _name, _seg_path[_name])
# The first forecast year is scaled onto the reviewed half's own annualised revenue. Without
# this line the sheet cannot reach the model's first year at all.
r = drv(r, 'First-year scale onto the reviewed half (annualised)',
        [dr['h1_anchor_scale']] + [1.0] * 4, NUM3 if 'NUM3' in dir() else PX)
r = drv(r, 'House Saudi inflation ladder (nominal = real x this)', dr['inflation_ladder'])
# THE ELIMINATION IS A DRIVER, NOT A RESIDUAL SOMEBODY TYPED. Inter-segment revenue is
# eleven and a half billion on seventy-eight, so a model that summed the eleven segments and
# stopped would overstate group revenue by fifteen per cent. Held at its FY2025 share.
r = drv(r, 'Inter-segment eliminations (% of gross segment revenue)',
        [dr['elimination_share']] * 5)
r = drv(r, 'Group EBITDA margin', dr['ebitda_m'])
r = drv(r, 'D&A (% of revenue)', dr['dna_pct'])
r = drv(r, 'Capex intensity (% of revenue)', dr['capex_pct'])
# WORKING CAPITAL IS NO LONGER A DRAG TYPED AS A SHARE OF REVENUE. It is projected from the
# asset-conversion cycle — receivable, inventory, contract and payable days each against
# their own driver — and what appears here is the RESULT of that projection, not an input to
# it. The days themselves are on the Assumptions sheet below.
r = drv(r, 'Working-capital movement (% of revenue, an OUTPUT of the cycle)',
        dr['wc_out_pct'])
# NO TYPED FINANCIAL NUMERALS. These three rows used to be literal arrays — associates at
# 500 rising to 620, net finance at 200 rising to 280, and a flat 2.5% minority share — none
# of which came from the model, which is the numeric-traceability standard's own example of
# what a builder may not contain. Two of them are now computed and the third is not
# forecast at all and says so.
_fis = D['forecast_is']
r = drv(r, 'Finance income (SAR mn)', [D['forecast'][y]['fin_income'] for y in D['forecast']],
        NUM0)
r = drv(r, 'Finance cost (SAR mn)', [D['forecast'][y]['fin_cost'] for y in D['forecast']],
        NUM0)
r = drv(r, 'Early retirement programme (SAR mn, three-year mean escalated)',
        [D['forecast'][y]['early_retirement'] for y in D['forecast']], NUM0)
r = drv(r, 'Effective zakat rate (three filed years together)', [_fis['zakat_rate']] * 5)
r = drv(r, 'DPS declared (SAR/share)', dr['payout_dps'], PX)
wa.column_dimensions['C'].width = 11

# ===== WACC BUILD — FULL DETAIL & SOURCING (rows 84-92) ======================
r = hdr(wa, 84, 'WACC BUILD — FULL DETAIL & SOURCING (feeds rows 9-17; reference only)')
put(wa, f'A{r}', 'Risk-free rate — where the quote comes from')
put(wa, f'C{r}',
    'One dated sovereign quote is held for every Saudi company this desk covers, so no single study carries a '
    'rate of its own: %.2f%% as at %s, less this sovereign\u2019s own default spread of %.2f%%, giving the '
    'normalised %.2f%% the cost of equity is built on. Country risk is then charged once, inside the equity '
    'risk premium, rather than twice. %s'
    % (D['coc_record']['rf_observed'] * 100, D['macro_record']['path_as_of'],
       D['coc_record']['default_spread'] * 100, D['coc_record']['rf_star'] * 100,
       D['coc_record']['sovereign_staleness_disclosed']), SUB)
r += 1
r = inp(wa, r, 'ERP — CDS-based (Damodaran, "more current" alternative)', 0.0572, PCT2,
        'Same original file, Saudi row, CDS column: sovereign CDS 0.98% → ERP 5.72%. For Saudi the CDS basis is ABOVE the '
        'rating basis (CDS prices more risk than Aa3 implies) — the alternative WACC below is therefore the HIGHER one.')   # B86
put(wa, f'A{r}', 'Ke, alternative (CDS-based ERP)'); put(wa, f'B{r}', '=B9+B10*B86', BLACK, PCT2); r += 1   # B87
put(wa, f'A{r}', 'WACC, alternative (CDS-based ERP)'); put(wa, f'B{r}', '=(1-B15)*B87+B15*B14', BLACK, PCT2); r += 1  # B88
put(wa, f'A{r}', 'Beta — regression detail')
# THIS CELL DESCRIBED A REGRESSION THE MODEL NO LONGER USES. It told a reader the beta came
# from a nine-week DAILY window on a commercial data vendor's index quotes — which the
# standing preference order says is not one of its tiers at all — while the model had long
# since moved to the sanctioned weekly regression against the exchange's own published
# index. The provenance is now read from the beta record itself, and the retired reading is
# named as retired rather than left standing as the method.
_BETAG = [g for g in D['sens']['beta_grid'] if abs(g['beta'] - 1.0) < 1e-9][0]
put(wa, f'C{r}',
    'A %.2f-year weekly regression of stc against the published index of the exchange it is listed on, the '
    'Tadawul All Share Index, read as of %s: %d paired weeks on that exchange\u2019s own trading week, beta '
    '%.4f, standard error %.4f, R\u00b2 %.1f%%, and a confidence interval running %.4f to %.4f. It is the '
    'first tier of the house preference order rather than a stopgap. A superseded reading of %.4f, taken from '
    'a nine-week daily window on a commercial vendor\u2019s index quotes, is recorded in the bibliography and '
    'is not used. Sensitivity: at a beta of 1.00 the cost of equity is %.2f%% and the cost of capital %.2f%%, '
    'which the beta grid on the Sensitivity sheet prices in full.'
    % (_BETA_REC['window_years'], _BETA_REC['index_asof'], _BETA_REC['n'], _BETA_REC['beta'],
       _BETA_REC['se'], _BETA_REC['r2'] * 100, _BETA_REC['ci90'][0], _BETA_REC['ci90'][1],
       _BETA_REC['superseded_beta'], _BETAG['ke'] * 100, _BETAG['wacc'] * 100), SUB)
r += 1
put(wa, f'A{r}', 'Kd source & currency-mix evidence')
put(wa, f'C{r}', 'Named instruments: May-2019 $1.25bn 10y sukuk at 3.89% (matures 2029); Jan-2026 $2bn dual-tranche sukuk '
                  '4.489% (5y, T+75) / 5.083% (10y, T+90) under the $5bn programme (books $5.4bn); Mar-2021 ECA loan $584mn; '
                  'remainder SAR murabaha/facilities (3M SAIBOR 4.79% Apr-2026 + 60–100bp). USD-linked ≈ 55–60% of gross '
                  'debt; SAR pegged 3.75 → USD legs economically quasi-SAR, single blended Kd 5.0% used (no floating-FX '
                  'tranche modelled).', SUB)
r += 1
put(wa, f'A{r}', 'Weights source')
put(wa, f'C{r}', D['coc_record']['weights_source'] +
     ' E/(D+E) = %.1f%%.' % (100.0 * D['coc_record']['weight_equity']), SUB)
r += 1
put(wa, f'A{r}', 'Both premium bases')
# THIS CELL POINTED A READER AT A REPOSITORY FILE AND A CODE MODULE. Neither is a source
# an outside reader can open, and a file path in a delivered document is caught by shape.
put(wa, f'C{r}',
    'The cost of capital is built the same way for every market this desk covers: the sovereign yield for the '
    'country, less that same sovereign\u2019s own default spread so country risk is charged once rather than '
    'twice, plus beta times an equity risk premium published for Saudi Arabia specifically. Both premium bases '
    'are published — %.2f%% central and %.2f%% on the alternative — giving costs of capital of %.2f%% and '
    '%.2f%%, so the choice is visible rather than made silently.'
    % (D['dcf']['wacc_build']['erp_market'] * 100, D['dcf']['wacc_build']['erp_rating'] * 100,
       D['dcf']['wacc'] * 100, D['dcf']['wacc_rating_basis'] * 100), SUB)

json.dump(DRV, open(os.path.join(HERE, '_asm_rows.json'), 'w'))
json.dump(dict(DPS_ROW=DPS_ROW, ANCHOR_ROWS=ANCHOR_ROWS),
          open(os.path.join(HERE, '_asm_extra.json'), 'w'))
wb.save(os.path.join(HERE, FN))
print('partA1 ok; drivers:', len(DRV))
