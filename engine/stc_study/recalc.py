#!/usr/bin/env python3
"""STC — recalculate the DELIVERED workbook and compare its VALUES to the model.

WHY THIS FILE EXISTS, AND WHAT ITS ABSENCE COST. Twenty-two of twenty-four study
directories carry a recalculation script; STC did not, and the workbook it delivered
published a central of SAR 28.94 against a study publishing 38.14 — a label collision put
the gross margin of 0.485 into every segment's revenue formula, so group revenue halved
every forecast year, profit turned negative from FY28E and closing cash reached minus
thirty-three billion riyals.

NOTHING ABOUT THAT LOOKED BROKEN. Every formula parsed, every cell computed, and a
recalculation run that afternoon reported "16 sheets, 1,507 cells, 0 formula errors" and
was called clean. It searched for error strings. IT NEVER ASKED WHETHER THE NUMBERS WERE
RIGHT. The depth bar says so in its own words — a clean recalculation is necessary but
NOT sufficient — and this script is that sentence made arithmetic: it opens the delivered
file, evaluates it in an engine that is not this repository's code, and holds the answers
it computes against the answers the model committed.

A FORMULA MODEL THAT RECOMPUTES THE WRONG NUMBER IS WORSE THAN A PASTED ONE, because it
looks live.

Run:  python3 engine/stc_study/recalc.py
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = 'STC_Valuation_Model_05092026_public.xlsx'
TOL = 0.005          # half a per cent: the workbook rounds, the model does not

#: (sheet, cell, the committed figure it must reproduce, what it is)
#: Every target is read from study_numbers.json — none is typed here, because a target
#: typed by hand is a second copy of the answer and goes stale the way the workbook did.
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))


def targets():
    d = D
    return [
        ('Summary', 'C12', d['central'], 'the central'),
        ('DCF', 'B27', d['dcf']['ev'], 'enterprise value'),
        ('DCF', 'B28', d['dcf']['tv_pct'], 'terminal share of enterprise value'),
        # BY LABEL, NOT BY POSITION. This target read 'C12' and the first fix to the
        # bridge moved that row onto the minority interest, so the check reported a
        # 10,397% disagreement that was entirely its own. A check that hardcodes a row
        # number has the defect it is looking for.
        ('SOTP Bridge', ('DCF fair value per share (SAR)', 'C'), d['central'],
         'the bridge per share'),
        ('Relative & Normalized', 'C6', d['lenses']['relative']['base'],
         'the enterprise-multiple cross-check'),
    ]


def evaluate(path):
    """Recompute the file in LibreOffice — an evaluator that is not this repo's code."""
    # CONVERT FROM THE ORIGINAL INTO A SEPARATE DIRECTORY. Copying the file into the
    # output directory first makes LibreOffice overwrite its own input, and the result
    # carries no cached values at all — every formula cell reads back as None, which a
    # careless script would report as "nothing to compare" rather than as a broken run.
    out = tempfile.mkdtemp(prefix='stc_recalc_')
    r = subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx',
                        '--outdir', out, path],
                       capture_output=True, text=True, timeout=600)
    if r.returncode != 0:
        raise SystemExit('the workbook could not be recalculated: %s' % r.stderr[-400:])
    return os.path.join(out, os.path.basename(path))


def main():
    src = os.path.join(HERE, WORKBOOK)
    if not os.path.exists(src):
        raise SystemExit('no delivered workbook at %s' % src)
    wb = load_workbook(evaluate(src), data_only=True)

    problems, unparseable, checked = [], 0, 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                checked += 1
                if isinstance(c.value, str) and c.value.startswith('#'):
                    unparseable += 1
                    problems.append('%s!%s evaluates to %s' % (ws.title, c.coordinate,
                                                               c.value))

    # THE HALF THAT WAS MISSING. An error-free workbook that computes a different answer
    # from the model is the failure this study actually shipped.
    wrong = []
    for sheet, cell, want, what in targets():
        if sheet not in wb.sheetnames:
            wrong.append('%s is not a sheet in the delivered workbook' % sheet)
            continue
        ws = wb[sheet]
        if isinstance(cell, tuple):
            label, col = cell
            hits = [c.row for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.strip() == label]
            if len(hits) != 1:
                wrong.append('%s carries %d rows labelled %r, so the target is ambiguous'
                             % (sheet, len(hits), label))
                continue
            cell = '%s%d' % (col, hits[0])
        got = ws[cell].value
        if not isinstance(got, (int, float)):
            wrong.append('%s!%s (%s) computes %r, which is not a number'
                         % (sheet, cell, what, got))
            continue
        base = abs(want) or 1.0
        if abs(got - want) / base > TOL:
            wrong.append('%s!%s (%s) computes %.4f, the model commits %.4f — %.1f%% apart'
                         % (sheet, cell, what, got, want, 100 * abs(got - want) / base))

    # AN EMPTY RESULT IS NOT A CLEAN RESULT. A conversion that produced no cached values
    # would leave every target reading None, which must fail loudly as a broken run rather
    # than quietly as a disagreement.
    numeric = sum(1 for ws in wb.worksheets for row in ws.iter_rows()
                  for c in row if isinstance(c.value, (int, float)))
    if numeric < 200:
        raise SystemExit('only %d numeric cells came back — the recalculation did not '
                         'run, and that is a broken check rather than a failing one'
                         % numeric)
    print('recalculated %d sheets, %d populated cells, %d numeric'
          % (len(wb.worksheets), checked, numeric))
    print('  formula errors      : %d' % unparseable)
    print('  answers checked     : %d' % len(targets()))
    print('  answers disagreeing : %d' % len(wrong))
    for p in problems[:10] + wrong:
        print('   ! %s' % p)
    if problems or wrong:
        print('\nFAIL — the delivered workbook does not reproduce the study.')
        return 1
    print('\nOK — the delivered workbook reproduces the study to within %.1f%%.'
          % (TOL * 100))
    return 0


if __name__ == '__main__':
    sys.exit(main())
