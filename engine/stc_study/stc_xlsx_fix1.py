"""Fix pass: BS restructure to disclosed anchors, IS PBT bridge, GP driver, CF link."""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FN = 'STC_Valuation_Model_09072026_public.xlsx'
wb = load_workbook(FN)
A = json.load(open('_asm_rows.json'))
BSJ = json.load(open('_bs_rows.json')); BS = BSJ['BS']
IS = json.load(open('_is_rows.json'))
CFJ = json.load(open('_cf_rows.json')); CF = CFJ['CF']; CLOSE = CFJ['CLOSE']
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
SUB = Font(size=9, color='6E7B77')
NUM0 = '#,##0;(#,##0);"-"'; PCT = '0.0%;(0.0%);"-"'
FCOLS = ['E', 'F', 'G', 'H', 'I']; ACOLS = ['C', 'D', 'E', 'F', 'G']

# ---- 1) Assumptions: add Gross margin driver row 55 -------------------------
wa = wb['Assumptions']
grow = 55
wa[f'A{grow}'] = 'Gross margin (% of revenue)'
for j, v in enumerate([0.485]*5):
    c = wa[f'{get_column_letter(3+j)}{grow}']; c.value = v; c.font = BLUE; c.number_format = PCT
A['Gross margin (% of revenue)'] = grow
json.dump(A, open('_asm_rows.json', 'w'))

def ac(label, j): return f"Assumptions!${ACOLS[j]}${A[label]}"

# ---- 2) Income Statement fixes ----------------------------------------------
wi = wb['Income Statement']
# GP forecast linked to the new driver
GP = IS['Gross profit']; REV = IS['Total revenue']
for j, c in enumerate(FCOLS):
    wi[f'{c}{GP}'] = f"={c}{REV}*{ac('Gross margin (% of revenue)', j)}"
    wi[f'{c}{GP}'].font = BLACK; wi[f'{c}{GP}'].number_format = NUM0
# associates/net-finance bridge row: rename + correct history
ASSR = IS['Share of associates & JVs + net finance & other']
wi[f'A{ASSR}'] = 'Associates, net finance, impairments & other'
for col, v in [('B', 826.0), ('C', -2292.0), ('D', 285.0)]:
    wi[f'{col}{ASSR}'] = v; wi[f'{col}{ASSR}'].font = BLUE; wi[f'{col}{ASSR}'].number_format = NUM0
EBTR = IS['Profit before zakat & income tax']
for col, v in [('B', 13987.0), ('C', 12134.0), ('D', 14723.0)]:
    wi[f'{col}{EBTR}'] = v; wi[f'{col}{EBTR}'].font = BLUE; wi[f'{col}{EBTR}'].number_format = NUM0
IS['Associates, net finance, impairments & other'] = ASSR
del IS['Share of associates & JVs + net finance & other']
json.dump(IS, open('_is_rows.json', 'w'))

# ---- 3) Balance Sheet restructure -------------------------------------------
wbs = wb['Balance Sheet']
PPER = BS['PP&E, ROU, intangibles & inv. property']
ASSOC = BS['Investments in associates & JVs']; FA = BS['Financial assets (incl. Telefónica 9.97%)']
TRR = BS['Trade receivables, net']; CASH = BSJ['CASH']; OTHA = BS['Other assets']
TA = BSJ['TA']; TLE = BSJ['TLE']; CHK = BSJ['CHK']; NDR = BSJ['NDR']
EQA = BS['Equity attributable to shareholders']; NCI2 = BS['Non-controlling interests']
TEQ = BS['Total equity']; DEBT = BS['Borrowings & sukuk (excl. leases)']
LEAS = BS['Lease liabilities']; PAY = BS['Trade payables & financial liabilities']
OTHL = BS['Zakat payable, provisions & other liabilities']
wbs[f'A{PPER}'] = 'Net fixed & intangible assets (PP&E, ROU, intangibles, inv. prop.)'
# history: disclosed anchors blue; PP&E and Other-liabilities are the balancing formulas
hist_vals = {
    ASSOC: (3800.0, 4200.0, 4641.0), FA: (22000.0, 23500.0, 24893.0),
    TRR: (24500.0, 25800.0, 26727.0), OTHA: (2600.0, 2600.0, 2952.0),
    EQA: (78985.0, 89417.0, 83414.0), NCI2: (2530.0, 3069.0, 3482.0),
    DEBT: (21958.0, 15132.0, 15191.0), LEAS: (6985.0, 4580.0, 2253.0),
    PAY: (25000.0, 26500.0, 29610.0),
}
for row, (b, c, d) in hist_vals.items():
    for col, v in zip('BCD', (b, c, d)):
        cell = wbs[f'{col}{row}']; cell.value = v; cell.font = BLUE; cell.number_format = NUM0
for col, v in zip('BCD', (28138.0, 30755.0, 15080.0)):
    cell = wbs[f'{col}{CASH}']; cell.value = v; cell.font = BLUE; cell.number_format = NUM0
for col, v in zip('BCD', (159646.0, 160638.0, 157477.0)):
    cell = wbs[f'{col}{TA}']; cell.value = v; cell.font = Font(color='0000FF', bold=True); cell.number_format = NUM0
for col in 'BCD':   # PP&E = disclosed TA − all other asset lines (balancing)
    wbs[f'{col}{PPER}'] = f"={col}{TA}-{col}{ASSOC}-{col}{FA}-{col}{TRR}-{col}{CASH}-{col}{OTHA}"
    wbs[f'{col}{PPER}'].font = BLACK; wbs[f'{col}{PPER}'].number_format = NUM0
for col in 'BCD':   # Other liabilities = disclosed TA − equity − debt − leases − payables (balancing)
    wbs[f'{col}{OTHL}'] = f"={col}{TA}-{col}{TEQ}-{col}{DEBT}-{col}{LEAS}-{col}{PAY}"
    wbs[f'{col}{OTHL}'].font = BLACK; wbs[f'{col}{OTHL}'].number_format = NUM0
# forecast TA = SUM of asset lines (PP&E rolls from FY25 balancing value)
for c in FCOLS:
    wbs[f'{c}{TA}'] = f"=SUM({c}{PPER}:{c}{OTHA})"
    wbs[f'{c}{TA}'].font = Font(bold=True); wbs[f'{c}{TA}'].number_format = NUM0
# cash forecast links to the correct Cash Flow closing row
for c in FCOLS:
    wbs[f'{c}{CASH}'] = f"='Cash Flow'!{c}{CLOSE}"
    wbs[f'{c}{CASH}'].font = GREEN; wbs[f'{c}{CASH}'].number_format = NUM0
# footnote refresh
wbs[f'A{CHK+2}'] = ('Historic mapping: disclosed IR anchors (total assets, cash+murabaha, total debt, attributable equity) are blue as '
                    'reported; FY25 line detail (associates 4,641 · financial assets 24,893 · receivables 26,727 · leases 2,253 · payables+financial '
                    'liabilities 29,610 · NCI 3,482) is from the Q1-2026 FS 31-Dec-25 comparatives; FY23/FY24 line detail is estimated to tie to the '
                    'disclosed totals (flagged). "Net fixed & intangible assets" and "Zakat payable, provisions & other liabilities" are the balancing '
                    'lines — shown as formulas off the disclosed totals, so the balance check is exact in every column. FY26E borrowings step up '
                    'SAR +7,284mn (the completed Jan-2026 $2bn sukuk net of amortisation); gross debt flat thereafter. Two cash framings: FS cash incl. '
                    'stc bank (21,442 at Q1-26) vs IR core-group cash (15,412) — the model rolls the IR-basis series (FY25: 15,080).')
wbs[f'A{CHK+2}'].font = SUB
wb.save(FN)
print('fix1 ok')
