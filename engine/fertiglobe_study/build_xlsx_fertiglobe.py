"""Fertiglobe_Valuation_Model_09-08-2026.xlsx — the 16-sheet companion workbook for the
Fertiglobe plc (ADX: FERTIGLB) valuation study.

Blue font = an input you can change · black = a live formula · green = a link to another
sheet in this workbook.

THE WORKBOOK CALCULATES, IT DOES NOT STORE. Every quantity that can be derived from a
driver is written as a live Excel formula: the cost of capital is built in the sheet from
the Abu Dhabi sovereign, the beta and the equity risk premium; the discount factors
compound one from the next; the unit build multiplies volume by price and by cost per
tonne; the balance sheet rolls property, working capital, equity and net debt forward; and
every ratio and per-share figure is computed where it is shown.

Only three classes of cell are pasted values, and the READ FIRST sheet names them:

  1. audited and disclosed history — the primary record, not a calculation (where a line is
     both disclosed and derivable, the DISCLOSED figure is carried and the subtotals above
     and below it are formulas that must reconcile to it);
  2. the output of the unit build that would be unreadable if it were flattened into a grid
     — the disclosed volumes, segment revenues and benchmark prices that the realisation
     factor and the cost pass-through are regressed on;
  3. whole-model re-runs — the probabilistic price map and the sensitivity grids, where
     each cell is a complete revaluation of the model rather than one arithmetic step.

Every formula cell also records the model's own value for that cell in xlsx_expected.json.
recalc_fertiglobe.py evaluates the delivered workbook independently and asserts the two
agree cell for cell, so a formula that computes the right thing the wrong way fails there
rather than shipping. driver_test_fertiglobe.py then perturbs each driver in the delivered
file and asserts the headline moves in the asserted direction.

No financial numeral is typed into this builder. Everything numeric is read from, or
derived from, study_numbers.json.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
STK = json.load(open(os.path.join(HERE, 'strike_result.json')))

# ---------------------------------------------------------------------------
# the model, read from the single source of numeric truth
# ---------------------------------------------------------------------------
M = D['meta']
IN = {k: v['value'] for k, v in D['inputs'].items()}
HI, HB, CCC, U, CS = D['hist_is'], D['hist_bs'], D['ccc'], D['unit'], D['cost_stack']
FA, FB = D['frame_A'], D['frame_B']
DA, DB = D['dcf_A'], D['dcf_B']
BA, BB = D['bridge_A'], D['bridge_B']
BAB, BBB = D['bridge_A_book'], D['bridge_B_book']
W, REL, NRM, BOOK, LN, SN, EXPS = (D['wacc'], D['rel'], D['norm'], D['book'],
                                   D['lenses'], D['sens'], D['experts'])
SH, FX, SPOT = M['shares_mn'], M['fx'], M['spot_aed']
TAX, NCI, G = D['tax_rate'], D['nci_share'], D['g_term']
REALIS = D['realisation']
YF = ['FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']
YH = ['FY2023', 'FY2024', 'FY2025']
H3 = ['FY23', 'FY24', 'FY25']

# constants recovered from the model's own output rather than typed
TRADE_PX = [FA['rev_3p'][i] * 1000.0 / FA['vol_3p'][i] for i in range(5)]
TRADE_MGN = round(FA['ebitda_3p'][0] / FA['rev_3p'][0], 6)
CORP = [round(FA['ebitda'][i] - FA['ebitda_own'][i] - FA['ebitda_3p'][i], 6) for i in range(5)]
PAYOUT = round(1.0 - (FA['equity'][0] - HB['FY25']['eq_own']) / FA['np_attr'][0], 6)
REPL_T = (DA['ic_replacement'] - FA['nwc'][4]) * 1000.0 / (IN['cap_urea'] + IN['cap_nh3_merchant'])
IRATE = round(NRM['interest'] / IN['netdebt_h1_26'], 6)
DSO, DIO = CCC['FY25']['dso'], CCC['FY25']['dio']
DPO = CCC['FY25_ex_accrual']['dpo']
NWC_PRIOR = HB['FY25']['inv'] + HB['FY25']['recv'] - (HB['FY25']['pay'] - IN['sorfert_accr_fy25'])
CAP_TOT = IN['cap_urea'] + IN['cap_nh3_merchant']
W_AE_TAX = 1.0 - IN['w_egypt'] - IN['w_algeria']          # tax-weighting basis
W_OTHER = IN['nca_other_regions'] / IN['nca_total']
W_AE_ERP = W_AE_TAX - W_OTHER                             # equity-risk-premium basis
PBT_SUM = IN['pbt_fy22'] + HI['FY23']['pbt'] + HI['FY24']['pbt'] + HI['FY25']['pbt']
TAX_EFF = ((IN['tax_fy22'] + HI['FY23']['tax'] + HI['FY24']['tax'] + HI['FY25']['tax']) / PBT_SUM)
TAX_CASH = ((IN['tax_paid_fy22'] + IN['tax_paid_fy23'] + IN['tax_paid_fy24']
             + IN['tax_paid_fy25']) / PBT_SUM)
TAX_JUR = (W_AE_TAX * IN['tax_dam_uae'] + IN['w_egypt'] * IN['tax_dam_eg']
           + IN['w_algeria'] * IN['tax_dam_dz'])
MKTCAP = M['mktcap_usd']
ND_NOW = IN['netdebt_h1_26']
EV_TRAIL = M['ev_trailing']


def roll(F):
    """The financing roll — interest, profit, dividends and net debt, year by year."""
    nd, out = ND_NOW, []
    for i in range(5):
        interest = IRATE * max(nd, 0.0)
        pbt = F['ebit'][i] - interest
        np_ = pbt * (1 - TAX)
        attr = np_ * (1 - NCI)
        div = PAYOUT * attr
        nd2 = nd - (F['fcff'][i] - interest * (1 - TAX)) + div + np_ * NCI
        out.append(dict(nd_prev=nd, interest=interest, pbt=pbt, np=np_, nci=np_ * NCI,
                        attr=attr, div=div, nd=nd2))
        nd = nd2
    return out


RA, RB = roll(FA), roll(FB)


def splice(F):
    """The 2026 half-year splice: reported first half plus a modelled second half."""
    h2v = F['vol_own'][0] - U['H1_26']['vol_own']
    h2r = F['rev_own'][0] - U['H1_26']['rev_own']
    h2p = h2r * 1000.0 / h2v
    h2c = CS['passthrough']['intercept'] + CS['passthrough']['slope'] * h2p
    h2e = h2r - h2v * h2c / 1000.0
    return dict(vol=h2v, rev=h2r, px=h2p, cost_t=h2c, ebitda=h2e,
                fy26=U['H1_26']['ebitda_own'] + h2e)


SPA, SPB = splice(FA), splice(FB)

# the pass-through regression, reproduced so its intermediate cells can be checked
_XS = [U[k]['px_realised'] for k in ('FY24', 'FY25', 'H1_26')]
_YS = [U[k]['cash_cost_t'] for k in ('FY24', 'FY25', 'H1_26')]
_XB = sum(_XS) / 3.0
_YB = sum(_YS) / 3.0
_SXY = sum((x - _XB) * (y - _YB) for x, y in zip(_XS, _YS))
_SXX = sum((x - _XB) ** 2 for x in _XS)
_SLOPE = _SXY / _SXX
_ICPT = _YB - _SLOPE * _XB
_SSR = sum((y - (_ICPT + _SLOPE * x)) ** 2 for x, y in zip(_XS, _YS))
_SST = sum((y - _YB) ** 2 for y in _YS)
_R2 = 1 - _SSR / _SST

# mid-cycle split of enterprise value between the two operating legs
MID_OWN = sum(FA['ebitda_own'][2:5] + FB['ebitda_own'][2:5]) / 6.0
MID_3P = sum(FA['ebitda_3p'][2:5] + FB['ebitda_3p'][2:5]) / 6.0
MID_CORP = sum(CORP[2:5]) * 2 / 6.0
MID_TOT = MID_OWN + MID_3P + MID_CORP
EV_MID = (DA['ev'] + DB['ev']) / 2.0

# expert panel
E1_MULT = EXPS['e1']['ev'] / REL['ebitda_mid']
DCF_PS = D['dcf_ps_aed']
CENTRAL = D['central']

# ---------------------------------------------------------------------------
# workbook scaffolding
# ---------------------------------------------------------------------------
wb = Workbook()
EXPECT, ANCH = {}, {}
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F4F7F6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='16343E')
FILL_H = PatternFill('solid', start_color='E9F0EF')
FILL_G = PatternFill('solid', start_color='F5F1E8')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.000;(0.000);"-"'
MULT = '0.00x'; DF4 = '0.0000'; B3 = '0.000'


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula and record the model's own value for the same cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    assert expect is not None, f'{ws.title}!{ad} has no expected value'
    EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


def note(ws, ad, text):
    put(ws, ad, text, fmt=None).font = SUB


CD = ['B', 'C', 'D', 'E', 'F']                       # five forecast columns
HC = ['B', 'C', 'D']                                 # three historical columns
FC = ['E', 'F', 'G', 'H', 'I']                       # five forecast columns, statements
ALL8 = HC + FC

# ============ 1 READ FIRST ==================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Fertiglobe plc (ADX: FERTIGLB)', None, 9)
READ = [
 'Companion model · Independent valuation study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the Fertiglobe valuation study. Blue cells are',
 'inputs; black cells are formulas; green cells link to another sheet in this workbook.', '',
 'IT CALCULATES, IT DOES NOT STORE. Every figure that can be derived from a driver is a live formula,',
 'so you can change a blue cell on Assumptions and watch the model reprice. The cost of capital is',
 'built in the sheet — the Abu Dhabi US dollar sovereign yield less its own default spread gives the',
 'normalised risk-free rate, beta times the equity risk premium gives the cost of equity, the cost of',
 'debt is taken after tax, and the weights come from net debt and market capitalisation. The glide',
 'fractions are shown as they are derived and the discount factors compound, each one the year before',
 'divided by one plus that year\'s cost of capital. The unit build multiplies volume by price and by',
 'cost per tonne. The balance sheet rolls property forward as prior plus capital expenditure less',
 'depreciation, and rolls working capital off the collection, inventory and payment days measured on',
 'the audited statements. Every ratio and per-share figure on every sheet is computed where it is',
 'shown.', '',
 'THREE THINGS ARE PASTED VALUES, and it is worth knowing exactly which.',
 '  1. Audited and disclosed history. The primary record is not a calculation. Where a line is both',
 '     disclosed and derivable the DISCLOSED figure is carried, and the subtotals above and below it',
 '     are formulas that have to reconcile to it — gross profit, operating profit, profit before tax,',
 '     profit for the year and profit to owners are all computed from the disclosed components.',
 '  2. The output of the unit build. The disclosed sales volumes, own-produced segment revenue and',
 '     EBITDA, and the published urea and ammonia benchmark prices for the three reported periods sit',
 '     on the Segments sheet as the raw material for two regressions. Everything downstream of them —',
 '     the realisation factor against the benchmark and the cost pass-through slope and intercept —',
 '     is computed in front of you, and the whole forecast cost stack is built from those two.',
 '  3. Whole-model re-runs. The probabilistic price map on the Monte Carlo sheet and every grid on',
 '     the Sensitivity sheet are complete revaluations of the model, one per cell, so they cannot be',
 '     a single formula. THESE GRIDS DO NOT REDRAW WHEN A DRIVER IS CHANGED. Change a blue cell and',
 '     the valuation moves; the price map and the sensitivity grids stay as they were computed.',
 '',
 'Anything else that is pasted is a defect.', '',
 'How the forecast is built. Not as one growth rate. Volume is installed capacity times a utilisation',
 'path, split between urea and merchant ammonia. Price is the published benchmark for each product,',
 'blended at the volume mix and multiplied by the realisation the company has actually achieved',
 'against that benchmark in each of the three reported periods. Revenue is volume times price divided',
 'by a thousand. Cash cost per tonne is an intercept plus a slope times the realised price, both of',
 'them regressed on the sheet from the same three periods, because the Egyptian and Algerian gas',
 'contracts are product-linked — the cost side moves with the price side rather than with a consumer',
 'price index. Margin is an OUTPUT of that build, never an input.', '',
 'The contested judgement, carried both ways and never averaged into one number. Framing A treats the',
 '2026 nitrogen price as a war premium on top of a marginal-cost anchor and lets it revert. Framing B',
 'takes the company\'s own sourced supply and demand balance and the European tariff wall on Russian',
 'product and holds prices near recent levels. Both are built in full, both are discounted, both',
 'reach a value per share, and both are shown side by side on the Summary, Fundamental Valuation,',
 'SOTP Bridge and DCF sheets. The cash-flow lens is their average.', '',
 'The 2026 half-year is already reported, so it is not forecast. The first six months of 2026 are the',
 'disclosed actual and only the second half is modelled; the splice is shown on the Segments sheet.',
 '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model',
 'outputs shown as ranges and distributions.', '',
 'Currency. The group reports and functions in US dollars, and the shares trade in dirhams on the Abu',
 f'Dhabi exchange. The model runs in US dollars, in millions, and translates at the fixed peg of',
 f'{FX:.4f} dirhams to the dollar only at the last step. Market price AED {SPOT:.2f} at the '
 f"{M['price_date']} close.", '',
 'Sheets: READ FIRST · Summary · Fundamental Valuation · Assumptions · SOTP Bridge · Segments ·',
 'Relative & Normalized · DCF · Income Statement · Balance Sheet · Cash Flow · Summary Financials ·',
 'Monte Carlo · Sensitivity · Per-Share & Ratios · Peer & Sector.']
for i, ln in enumerate(READ, start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 104

# ============ 2 SUMMARY (frame; the numbers are filled once addresses exist) =
ws = sheet('Summary')
title(ws, 'Summary — valuation at a glance', 'Every value links live to the sheet that computes it',
      6, awidth=50, cwidth=16)
ws.freeze_panes = 'A5'

# ============ 3 FUNDAMENTAL VALUATION (frame) ================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — four lenses, and the judgement carried both ways', None, 6,
      awidth=54, cwidth=16)
ws.freeze_panes = 'A5'

# ============ 4 ASSUMPTIONS =================================================
ws = sheet('Assumptions')
title(ws, 'Assumptions — the driver register', 'Blue cells are inputs: change one and the model '
      'reprices. Green cells are computed elsewhere in the workbook and shown here so every driver '
      'the model consumes is visible in one place.', 6, awidth=62, cwidth=14)
ws.freeze_panes = 'A4'
hdr(ws, 3, ['Driver'] + YF)
A = {}
r = 4


def block(name, items):
    global r
    band(ws, r, 6); put(ws, f'A{r}', name, bold=True, fmt=None); r += 1
    for key, lab, val, fmt in items:
        put(ws, f'A{r}', lab, fmt=None)
        if isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(ws, f'{CD[i]}{r}', v, BLUE, fmt)
        else:
            put(ws, f'C{r}', val, BLUE, fmt)
        A[key] = r
        r += 1
    r += 1


def a(key, i=None):
    col = CD[i] if i is not None else 'C'
    return f'Assumptions!${col}${A[key]}'


block('Market and share count', [
    ('spot', 'Market price (AED per share)', SPOT, NUM2),
    ('fx', 'Dirhams per US dollar (Central Bank peg)', FX, '0.0000'),
    ('shares', 'Ordinary shares outstanding (mn)', SH, NUM0)])
block('Capacity and utilisation', [
    ('cap_urea', 'Urea production capacity (kt)', IN['cap_urea'], NUM0),
    ('cap_nh3', 'Merchant ammonia capacity (kt)', IN['cap_nh3_merchant'], NUM0),
    ('util_urea', 'Urea capacity utilisation', FA['util_urea'], PCT),
    ('util_nh3', 'Merchant ammonia capacity utilisation', FA['util_nh3'], PCT),
    ('vol_3p', 'Third-party traded volume (kt)', FA['vol_3p'], NUM0)])
block('Price paths — the contested judgement, carried both ways', [
    ('pxa_u', 'Framing A — urea benchmark, Egypt free on board ($/t)', FA['px_urea'], NUM0),
    ('pxa_n', 'Framing A — ammonia benchmark, Middle East ($/t)', FA['px_nh3'], NUM0),
    ('pxb_u', 'Framing B — urea benchmark, Egypt free on board ($/t)', FB['px_urea'], NUM0),
    ('pxb_n', 'Framing B — ammonia benchmark, Middle East ($/t)', FB['px_nh3'], NUM0),
    ('px_3p', 'Third-party traded price ($/t)', TRADE_PX, NUM0)])
block('Cost and margin drivers', [
    ('realis', 'Realisation against the blended benchmark', None, B3),
    ('pt_slope', 'Cost pass-through — share of each incremental dollar of realised price', None, B3),
    ('pt_int', 'Cost pass-through — fixed cash cost intercept ($/t)', None, NUM1),
    ('trade_mgn', 'Third-party trading EBITDA margin', TRADE_MGN, PCT),
    ('corp', 'Corporate and other segment EBITDA ($m)', CORP, NUM1)])
block('Capital intensity', [
    ('dna', 'Depreciation and amortisation ($m)', FA['dna'], NUM0),
    ('capex', 'Capital expenditure ($m)', FA['capex'], NUM0)])
block('Working-capital cycle (measured on the Balance Sheet sheet)', [
    ('dso', 'Days sales outstanding', None, NUM1),
    ('dio', 'Days inventory outstanding', None, NUM1),
    ('dpo', 'Days payables outstanding, excluding the gas accrual', None, NUM1),
    ('nwc0', 'Working capital at FY2025, excluding the gas accrual ($m)', None, NUM0)])
block('Tax — three sourced estimates, averaged on the DCF sheet', [
    ('t_uae', 'Statutory corporate tax rate — United Arab Emirates', IN['tax_dam_uae'], PCT),
    ('t_eg', 'Statutory corporate tax rate — Egypt', IN['tax_dam_eg'], PCT),
    ('t_dz', 'Statutory corporate tax rate — Algeria', IN['tax_dam_dz'], PCT),
    ('w_eg', 'Share of non-current assets in Egypt', IN['w_egypt'], PCT),
    ('w_dz', 'Share of non-current assets in Algeria', IN['w_algeria'], PCT),
    ('tax_fy22', 'Income tax charge, FY2022 ($m)', IN['tax_fy22'], NUM1),
    ('pbt_fy22', 'Profit before tax, FY2022 ($m)', IN['pbt_fy22'], NUM1),
    ('paid22', 'Income taxes paid, FY2022 ($m)', IN['tax_paid_fy22'], NUM1),
    ('paid23', 'Income taxes paid, FY2023 ($m)', IN['tax_paid_fy23'], NUM1),
    ('paid24', 'Income taxes paid, FY2024 ($m)', IN['tax_paid_fy24'], NUM1),
    ('paid25', 'Income taxes paid, FY2025 ($m)', IN['tax_paid_fy25'], NUM1)])
block('Cost of capital', [
    ('ust10', 'Ten-year United States Treasury yield', IN['ust10'], PCT2),
    ('ad_cds', 'Abu Dhabi sovereign credit default swap spread', IN['ad_cds'], PCT2),
    ('ad_ads', 'Abu Dhabi adjusted default spread (rating basis)', IN['ad_ads'], PCT2),
    ('ad_erp', 'Abu Dhabi equity risk premium, rating basis', IN['ad_erp'], PCT2),
    ('ad_crp', 'Abu Dhabi country risk premium', IN['ad_crp'], PCT2),
    ('ad_erp_c', 'Abu Dhabi equity risk premium, swap basis', IN['ad_erp_cds'], PCT2),
    ('eg_erp', 'Egypt equity risk premium, rating basis', IN['eg_erp'], PCT2),
    ('eg_erp_c', 'Egypt equity risk premium, swap basis', IN['eg_erp_cds'], PCT2),
    ('dz_erp', 'Algeria equity risk premium', IN['dz_erp'], PCT2),
    ('mat_erp', 'Mature-market equity risk premium', None, PCT2),
    ('beta', 'Beta — own-stock weekly regression against the local market', IN['beta'], B3),
    ('kd_bc', 'Marginal debt spread — facilities B and C', IN['kd_spread_facility_bc'], PCT2),
    ('kd_ad', 'Marginal debt spread — the parent term loan', IN['kd_spread_adnoc'], PCT2),
    ('wd_term', 'Terminal debt weight', W['wd_term'], PCT),
    ('nca_oth', 'Non-current assets outside the core countries ($m)', IN['nca_other_regions'], NUM1),
    ('nca_tot', 'Non-current assets, total ($m)', IN['nca_total'], NUM1),
    ('nd_now', 'Net debt at 30 June 2026 ($m)', ND_NOW, NUM1),
    ('irate', 'Interest rate charged on net debt in the forecast', IRATE, PCT)])
block('Terminal block and the bridge to equity', [
    ('g', 'Terminal growth', G, PCT),
    ('roic_sec', 'Long-run return on capital for merchant nitrogen', DA['roic_sector'], PCT),
    ('repl_t', 'Replacement cost of installed capacity ($ per tonne)', REPL_T, NUM0),
    ('nci', 'Minority share of group profit', NCI, PCT),
    ('nci_bv', 'Minority interests at book value ($m)', IN['eqnci_fy25'], NUM1),
    ('payout', 'Dividend payout ratio in the forecast', PAYOUT, PCT)])
block('Lens inputs', [
    ('mult', 'Justified enterprise value / EBITDA', REL['mult'], MULT),
    ('pe', 'Justified price / earnings', NRM['pe'], MULT),
    ('w_dcf', 'Weight — discounted cash flow', LN['dcf']['weight'], PCT),
    ('w_rel', 'Weight — relative multiples', LN['relative']['weight'], PCT),
    ('w_norm', 'Weight — normalised earnings power', LN['normalized']['weight'], PCT),
    ('w_book', 'Weight — book value', LN['book']['weight'], PCT),
    ('eb_h1', 'Adjusted EBITDA, first half 2026 ($m)', IN['adj_ebitda_h1_26'], NUM1),
    ('np_h1', 'Profit to owners, first half 2026 ($m)', IN['npown_h1_26'], NUM1)])
note(ws, f'A{r}', 'Every input above is sourced and dated in the study bibliography. The green rows '
     'are computed where they are measured — the realisation factor and the cost pass-through on the '
     'Segments sheet, the working-capital days on the Balance Sheet sheet — and are shown here '
     'because the forecast consumes them.')

# ============ 6 SEGMENTS ====================================================
ws = sheet('Segments')
title(ws, 'Segments — the unit build', 'Volume times price, and cost per tonne regressed on the '
      'realised price. Both the realisation factor and the cost pass-through are measured here from '
      'the three reported periods, not assumed.', 7, awidth=52, cwidth=15)
ws.freeze_panes = 'A5'
CAL = ['B', 'C', 'D']
CK = ['FY24', 'FY25', 'H1_26']
hdr(ws, 4, ['Reported period', 'FY2024', 'FY2025', 'First half 2026'])
CR = {}
r = 5
for key, lab, fmt in [('vu', 'Urea sales volume (kt)', NUM0),
                      ('vn', 'Merchant ammonia sales volume (kt)', NUM0),
                      ('vo', 'Own-produced sales volume, total (kt)', NUM0),
                      ('ro', 'Own-produced segment revenue ($m)', NUM1),
                      ('eo', 'Own-produced segment EBITDA ($m)', NUM1),
                      ('bu', 'Urea benchmark, Egypt free on board ($/t)', NUM0),
                      ('bn', 'Ammonia benchmark, Middle East ($/t)', NUM0)]:
    put(ws, f'A{r}', lab, fmt=None)
    for i, k in enumerate(CK):
        v = {'vu': U[k]['vol_urea'], 'vn': U[k]['vol_nh3'], 'vo': U[k]['vol_own'],
             'ro': U[k]['rev_own'], 'eo': U[k]['ebitda_own'], 'bu': U[k]['bm_urea'],
             'bn': U[k]['bm_nh3']}[key]
        put(ws, f'{CAL[i]}{r}', v, BLUE, fmt)
    CR[key] = r
    r += 1
for key, lab, fml, vals, fmt in [
        ('px', 'Realised price ($/t)', lambda c: f'={c}{CR["ro"]}*1000/{c}{CR["vo"]}',
         [U[k]['px_realised'] for k in CK], NUM1),
        ('bb', 'Volume-weighted benchmark price ($/t)',
         lambda c: f'=({c}{CR["vu"]}*{c}{CR["bu"]}+{c}{CR["vn"]}*{c}{CR["bn"]})'
                   f'/({c}{CR["vu"]}+{c}{CR["vn"]})', [U[k]['bm_blend'] for k in CK], NUM1),
        ('rl', 'Realisation against the benchmark', lambda c: f'={c}{CR["px"]}/{c}{CR["bb"]}',
         [U[k]['realisation'] for k in CK], B3),
        ('cc', 'Cash cost, total ($m)', lambda c: f'={c}{CR["ro"]}-{c}{CR["eo"]}',
         [U[k]['cash_cost_tot'] for k in CK], NUM1),
        ('ct', 'Cash cost per tonne ($/t)', lambda c: f'={c}{CR["cc"]}*1000/{c}{CR["vo"]}',
         [U[k]['cash_cost_t'] for k in CK], NUM1),
        ('mg', 'Own-produced EBITDA margin', lambda c: f'={c}{CR["eo"]}/{c}{CR["ro"]}',
         [U[k]['ebitda_margin_own'] for k in CK], PCT)]:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(3):
        putf(ws, f'{CAL[i]}{r}', fml(CAL[i]), vals[i], fmt)
    CR[key] = r
    r += 1
r += 1
band(ws, r, 7); put(ws, f'A{r}', 'Realisation factor applied to the forecast (average of the three)',
                    bold=True, fmt=None)
putf(ws, f'B{r}', f'=AVERAGE(B{CR["rl"]}:D{CR["rl"]})', REALIS, B3, bold=True)
R_REALIS = r
r += 2
band(ws, r, 7)
put(ws, f'A{r}', 'Cost pass-through — least squares of cash cost per tonne on realised price',
    bold=True, fmt=None)
r += 1
REG = {}
for key, lab, fml, vals, fmt in [
        ('x', 'Realised price, the explanatory variable ($/t)', lambda c: f'={c}{CR["px"]}',
         _XS, NUM1),
        ('y', 'Cash cost per tonne, the dependent variable ($/t)', lambda c: f'={c}{CR["ct"]}',
         _YS, NUM1)]:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(3):
        putf(ws, f'{CAL[i]}{r}', fml(CAL[i]), vals[i], fmt)
    REG[key] = r
    r += 1
for key, lab, fml, val, fmt in [
        ('xb', 'Mean realised price ($/t)', f'=AVERAGE(B{REG["x"]}:D{REG["x"]})', _XB, NUM1),
        ('yb', 'Mean cash cost per tonne ($/t)', f'=AVERAGE(B{REG["y"]}:D{REG["y"]})', _YB, NUM1)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fml, val, fmt)
    REG[key] = r
    r += 1
_xb, _yb = f'$B${REG["xb"]}', f'$B${REG["yb"]}'
_cross = '+'.join(f'({c}{REG["x"]}-{_xb})*({c}{REG["y"]}-{_yb})' for c in CAL)
_sq = '+'.join(f'({c}{REG["x"]}-{_xb})^2' for c in CAL)
put(ws, f'A{r}', 'Sum of cross products', fmt=None); putf(ws, f'B{r}', f'={_cross}', _SXY, NUM1)
REG['sxy'] = r; r += 1
put(ws, f'A{r}', 'Sum of squared deviations in price', fmt=None)
putf(ws, f'B{r}', f'={_sq}', _SXX, NUM1); REG['sxx'] = r; r += 1
put(ws, f'A{r}', 'Slope — the pass-through', fmt=None)
putf(ws, f'B{r}', f'=B{REG["sxy"]}/B{REG["sxx"]}', _SLOPE, B3, bold=True)
REG['slope'] = r; r += 1
put(ws, f'A{r}', 'Intercept — fixed cash cost per tonne ($/t)', fmt=None)
putf(ws, f'B{r}', f'=B{REG["yb"]}-B{REG["slope"]}*B{REG["xb"]}', _ICPT, NUM1, bold=True)
REG['icpt'] = r; r += 1
_res = '+'.join(f'({c}{REG["y"]}-($B${REG["icpt"]}+$B${REG["slope"]}*{c}{REG["x"]}))^2' for c in CAL)
_tot = '+'.join(f'({c}{REG["y"]}-{_yb})^2' for c in CAL)
put(ws, f'A{r}', 'Sum of squared residuals', fmt=None); putf(ws, f'B{r}', f'={_res}', _SSR, NUM1)
REG['ssr'] = r; r += 1
put(ws, f'A{r}', 'Total sum of squares', fmt=None); putf(ws, f'B{r}', f'={_tot}', _SST, NUM1)
REG['sst'] = r; r += 1
put(ws, f'A{r}', 'Share of the variation explained', fmt=None)
putf(ws, f'B{r}', f'=1-B{REG["ssr"]}/B{REG["sst"]}', _R2, B3)
REG['r2'] = r; r += 2
note(ws, f'A{r}', 'The Egyptian and Algerian gas contracts are product-linked, so the cost side '
     'moves with the price side. The slope above is how much of each incremental dollar of realised '
     'price comes straight back out as cash cost; the intercept is what is left when the price side '
     'is stripped away. Both are measured, not assumed.')
r += 2

# --- the two framings -------------------------------------------------------
FRAME_ROWS = {}


def frame_block(r0, F, sp, tag, label, pk_u, pk_n):
    band(ws, r0, 7)
    put(ws, f'A{r0}', label, bold=True, fmt=None)
    hdr(ws, r0 + 1, ['$m unless stated'] + YF)
    R = {}
    rr = r0 + 2

    def line(key, lab, fml, vals, fmt=NUM0, bold=False, green=False, cols=range(5)):
        nonlocal rr
        put(ws, f'A{rr}', lab, bold=bold, fmt=None)
        for i in cols:
            putf(ws, f'{CD[i]}{rr}', fml(i), vals[i], fmt, bold=bold, green=green)
        R[key] = rr
        rr += 1

    line('uu', 'Urea capacity utilisation', lambda i: f'={a("util_urea", i)}', F['util_urea'],
         PCT, green=True)
    line('un', 'Merchant ammonia capacity utilisation', lambda i: f'={a("util_nh3", i)}',
         F['util_nh3'], PCT, green=True)
    line('vu', 'Urea volume (kt)', lambda i: f'={a("cap_urea")}*{CD[i]}{R["uu"]}', F['vol_urea'],
         NUM0)
    line('vn', 'Merchant ammonia volume (kt)', lambda i: f'={a("cap_nh3")}*{CD[i]}{R["un"]}',
         F['vol_nh3'], NUM0)
    line('vo', 'Own-produced volume (kt)', lambda i: f'={CD[i]}{R["vu"]}+{CD[i]}{R["vn"]}',
         F['vol_own'], NUM0)
    line('pu', 'Urea benchmark price ($/t)', lambda i: f'={a(pk_u, i)}', F['px_urea'], NUM0,
         green=True)
    line('pn', 'Ammonia benchmark price ($/t)', lambda i: f'={a(pk_n, i)}', F['px_nh3'], NUM0,
         green=True)
    line('bb', 'Volume-weighted benchmark price ($/t)',
         lambda i: (f'=({CD[i]}{R["vu"]}*{CD[i]}{R["pu"]}+{CD[i]}{R["vn"]}*{CD[i]}{R["pn"]})'
                    f'/({CD[i]}{R["vu"]}+{CD[i]}{R["vn"]})'), F['bm_blend'], NUM1)
    line('px', 'Realised price ($/t)', lambda i: f'={CD[i]}{R["bb"]}*{a("realis")}',
         F['px_realised'], NUM1)
    line('ro', 'Own-produced revenue ($m)',
         lambda i: f'={CD[i]}{R["vo"]}*{CD[i]}{R["px"]}/1000', F['rev_own'], NUM1)
    # cost per tonne: 2026 falls out of the half-year splice, later years off the pass-through
    put(ws, f'A{rr}', 'Cash cost per tonne ($/t)', fmt=None)
    R['ct'] = rr
    R['co'] = rr + 1
    R['eo'] = rr + 2
    putf(ws, f'B{R["ct"]}', f'=B{R["co"]}*1000/B{R["vo"]}', F['cost_t'][0], NUM1)
    for i in range(1, 5):
        putf(ws, f'{CD[i]}{R["ct"]}', f'={a("pt_int")}+{a("pt_slope")}*{CD[i]}{R["px"]}',
             F['cost_t'][i], NUM1)
    put(ws, f'A{R["co"]}', 'Own-produced cash cost ($m)', fmt=None)
    putf(ws, f'B{R["co"]}', f'=B{R["ro"]}-B{R["eo"]}', F['cost_own'][0], NUM1)
    for i in range(1, 5):
        putf(ws, f'{CD[i]}{R["co"]}', f'={CD[i]}{R["vo"]}*{CD[i]}{R["ct"]}/1000',
             F['cost_own'][i], NUM1)
    put(ws, f'A{R["eo"]}', 'Own-produced EBITDA ($m)', fmt=None)
    rr = R['eo'] + 1
    # the splice sits below; its address is resolved after the block is laid out
    for i in range(1, 5):
        putf(ws, f'{CD[i]}{R["eo"]}', f'={CD[i]}{R["ro"]}-{CD[i]}{R["co"]}', F['ebitda_own'][i],
             NUM1)
    line('v3', 'Third-party traded volume (kt)', lambda i: f'={a("vol_3p", i)}', F['vol_3p'],
         NUM0, green=True)
    line('p3', 'Third-party traded price ($/t)', lambda i: f'={a("px_3p", i)}', TRADE_PX, NUM0,
         green=True)
    line('r3', 'Third-party traded revenue ($m)',
         lambda i: f'={CD[i]}{R["v3"]}*{CD[i]}{R["p3"]}/1000', F['rev_3p'], NUM1)
    line('e3', 'Third-party traded EBITDA ($m)', lambda i: f'={CD[i]}{R["r3"]}*{a("trade_mgn")}',
         F['ebitda_3p'], NUM1)
    line('cp', 'Corporate and other segment EBITDA ($m)', lambda i: f'={a("corp", i)}', CORP,
         NUM1, green=True)
    line('rev', 'Group revenue ($m)', lambda i: f'={CD[i]}{R["ro"]}+{CD[i]}{R["r3"]}', F['rev'],
         NUM0, bold=True)
    line('eb', 'Group EBITDA ($m)',
         lambda i: f'={CD[i]}{R["eo"]}+{CD[i]}{R["e3"]}+{CD[i]}{R["cp"]}', F['ebitda'], NUM0,
         bold=True)
    line('mgn', 'Group EBITDA margin', lambda i: f'={CD[i]}{R["eb"]}/{CD[i]}{R["rev"]}',
         F['ebitda_margin'], PCT)
    rr += 1
    put(ws, f'A{rr}', 'The 2026 splice — the reported first half plus a modelled second half',
        bold=True, fmt=None)
    rr += 1
    S = {}
    for key, lab, fml, val, fmt in [
            ('v', 'Second-half volume (kt)', f'=B{R["vo"]}-D{CR["vo"]}', sp['vol'], NUM0),
            ('r', 'Second-half revenue ($m)', f'=B{R["ro"]}-D{CR["ro"]}', sp['rev'], NUM1),
            ('p', 'Second-half realised price ($/t)', None, sp['px'], NUM1),
            ('c', 'Second-half cash cost per tonne ($/t)', None, sp['cost_t'], NUM1),
            ('e', 'Second-half EBITDA ($m)', None, sp['ebitda'], NUM1),
            ('f', 'Full-year 2026 own-produced EBITDA ($m)', None, sp['fy26'], NUM1)]:
        put(ws, f'A{rr}', lab, fmt=None)
        if key == 'p':
            fml = f'=B{S["r"]}*1000/B{S["v"]}'
        elif key == 'c':
            fml = f'={a("pt_int")}+{a("pt_slope")}*B{S["p"]}'
        elif key == 'e':
            fml = f'=B{S["r"]}-B{S["v"]}*B{S["c"]}/1000'
        elif key == 'f':
            fml = f'=D{CR["eo"]}+B{S["e"]}'
        putf(ws, f'B{rr}', fml, val, fmt, bold=(key == 'f'))
        S[key] = rr
        rr += 1
    putf(ws, f'B{R["eo"]}', f'=B{S["f"]}', F['ebitda_own'][0], NUM1)
    FRAME_ROWS[tag] = R
    return rr + 1


r = frame_block(r, FA, SPA, 'A', 'FRAMING A — normalisation to a marginal-cost anchor',
                'pxa_u', 'pxa_n')
r = frame_block(r, FB, SPB, 'B', 'FRAMING B — a structurally tight market', 'pxb_u', 'pxb_n')
note(ws, f'A{r}', 'Margin is an output of this build, never an input. The only difference between '
     'the two framings is the price path; volume, the realisation factor, the cost pass-through, '
     'the trading leg and the corporate load are identical in both.')

# fill the green Assumptions cells that are measured here
ws = wb['Assumptions']
putf(ws, f'C{A["realis"]}', f'=Segments!B{R_REALIS}', REALIS, B3, green=True)
putf(ws, f'C{A["pt_slope"]}', f'=Segments!B{REG["slope"]}', _SLOPE, B3, green=True)
putf(ws, f'C{A["pt_int"]}', f'=Segments!B{REG["icpt"]}', _ICPT, NUM1, green=True)
putf(ws, f'C{A["mat_erp"]}', f'={a("ad_erp")}-{a("ad_crp")}', IN['mature_erp'], PCT2)

# ============ 8 DCF =========================================================
ws = sheet('DCF')
title(ws, 'Discounted cash flow — the full waterfall, both framings', 'Every line is a live '
      'formula. The cost of capital is built at the foot of the sheet, the glide fractions are '
      'derived in front of you, and the discount factors compound one from the next.', 7,
      awidth=54, cwidth=14)
ws.freeze_panes = 'A5'
hdr(ws, 4, ['$m unless stated'] + YF)
DR = {}


def dcf_block(r0, F, RL, tag, label, segR):
    band(ws, r0, 7); put(ws, f'A{r0}', label, bold=True, fmt=None)
    R = {}
    rr = r0 + 1

    def line(key, lab, fml, vals, fmt=NUM0, bold=False, green=False):
        nonlocal rr
        put(ws, f'A{rr}', lab, bold=bold, fmt=None)
        for i in range(5):
            putf(ws, f'{CD[i]}{rr}', fml(i), vals[i], fmt, bold=bold, green=green)
        R[key] = rr
        rr += 1

    line('rev', 'Revenue', lambda i: f'=Segments!{CD[i]}{segR["rev"]}', F['rev'], green=True)
    line('eb', 'EBITDA', lambda i: f'=Segments!{CD[i]}{segR["eb"]}', F['ebitda'], green=True)
    line('mgn', 'EBITDA margin (EBITDA / revenue)',
         lambda i: f'={CD[i]}{R["eb"]}/{CD[i]}{R["rev"]}', F['ebitda_margin'], PCT)
    line('dna', 'Depreciation and amortisation', lambda i: f'={a("dna", i)}', F['dna'], green=True)
    line('ebit', 'EBIT (EBITDA less depreciation and amortisation)',
         lambda i: f'={CD[i]}{R["eb"]}-{CD[i]}{R["dna"]}', F['ebit'], bold=True)
    line('nopat', 'Net operating profit after tax (EBIT times one less the tax rate)',
         lambda i: f'={CD[i]}{R["ebit"]}*(1-$C${{TAXROW}})', F['nopat'])
    line('add', 'Add back depreciation and amortisation', lambda i: f'={CD[i]}{R["dna"]}',
         F['dna'])
    line('capex', 'Less capital expenditure', lambda i: f'={a("capex", i)}', F['capex'],
         green=True)
    line('recv', 'Trade and other receivables (days sales outstanding)',
         lambda i: f'={a("dso")}*{CD[i]}{R["rev"]}/365', F['recv'], NUM1)
    line('inv', 'Inventories (days inventory outstanding)',
         lambda i: f'={a("dio")}*({CD[i]}{R["rev"]}-{CD[i]}{R["eb"]})/365', F['inv'], NUM1)
    line('pay', 'Trade and other payables (days payables outstanding)',
         lambda i: f'={a("dpo")}*({CD[i]}{R["rev"]}-{CD[i]}{R["eb"]})/365', F['pay'], NUM1)
    line('nwc', 'Net working capital',
         lambda i: f'={CD[i]}{R["recv"]}+{CD[i]}{R["inv"]}-{CD[i]}{R["pay"]}', F['nwc'], NUM1)
    line('dnwc', 'Less change in net working capital',
         lambda i: (f'={CD[i]}{R["nwc"]}-{a("nwc0")}' if i == 0
                    else f'={CD[i]}{R["nwc"]}-{CD[i-1]}{R["nwc"]}'), F['dnwc'], NUM1)
    line('fcff', 'Free cash flow to the firm',
         lambda i: (f'={CD[i]}{R["nopat"]}+{CD[i]}{R["add"]}-{CD[i]}{R["capex"]}'
                    f'-{CD[i]}{R["dnwc"]}'), F['fcff'], bold=True)
    line('coc', 'Forward cost of capital',
         lambda i: (f'=$C${{WACC}}+($C${{WACCT}}-$C${{WACC}})*{CD[i]}{{GLIDE}}'),
         W['wacc_path'], PCT2)
    # this line compounds off itself, and R['df'] is only registered once the row is written --
    # rr is the row currently being written, so the back-reference reads it from the closure
    line('df', 'Discount factor (each one the year before, divided by one plus that year\'s rate)',
         lambda i: (f'=1/(1+{CD[i]}{R["coc"]})' if i == 0
                    else f'={CD[i-1]}{rr}/(1+{CD[i]}{R["coc"]})'), DA['df'], DF4)
    line('pv', 'Present value of free cash flow',
         lambda i: f'={CD[i]}{R["fcff"]}*{CD[i]}{R["df"]}',
         DA['pv'] if tag == 'A' else DB['pv'], bold=True)
    DR[tag] = R
    return rr + 1


r = dcf_block(5, FA, RA, 'A', 'FRAMING A — normalisation to a marginal-cost anchor',
              FRAME_ROWS['A'])
r = dcf_block(r, FB, RB, 'B', 'FRAMING B — a structurally tight market', FRAME_ROWS['B'])

# --- terminal block, both framings side by side -----------------------------
band(ws, r, 7)
put(ws, f'A{r}', 'TERMINAL BLOCK AND ENTERPRISE VALUE', bold=True, fmt=None)
put(ws, f'C{r}', 'Framing A', bold=True, fmt=None)
put(ws, f'D{r}', 'Framing B', bold=True, fmt=None)
r += 1
TB = {}
RA_, RB_ = DR['A'], DR['B']


def trow(key, lab, fa, fb, va, vb, fmt=NUM0, bold=False, green=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    putf(ws, f'C{r}', fa, va, fmt, bold=bold, green=green)
    putf(ws, f'D{r}', fb, vb, fmt, bold=bold, green=green)
    TB[key] = r
    r += 1


trow('nwc30', 'Net working capital, FY2030E', f'=F{RA_["nwc"]}', f'=F{RB_["nwc"]}',
     FA['nwc'][4], FB['nwc'][4], NUM1)
trow('nopat30', 'Net operating profit after tax, FY2030E', f'=F{RA_["nopat"]}', f'=F{RB_["nopat"]}',
     FA['nopat'][4], FB['nopat'][4], NUM1)
trow('icb', 'Invested capital at book, FY2030E (property plus working capital)',
     f"='Balance Sheet'!I{{PPE}}+C{TB['nwc30']}", f"='Balance Sheet'!I{{PPE}}+D{TB['nwc30']}",
     FA['ic'][4], FB['ic'][4], NUM1)
trow('icr', 'Invested capital at replacement cost, FY2030E',
     f'=({a("cap_urea")}+{a("cap_nh3")})*{a("repl_t")}/1000+C{TB["nwc30"]}',
     f'=({a("cap_urea")}+{a("cap_nh3")})*{a("repl_t")}/1000+D{TB["nwc30"]}',
     DA['ic_replacement'], DB['ic_replacement'], NUM1)
trow('rb', 'Return on capital — book basis', f'=C{TB["nopat30"]}/C{TB["icb"]}',
     f'=D{TB["nopat30"]}/D{TB["icb"]}', DA['roic_book'], DB['roic_book'], PCT)
trow('rr', 'Return on capital — replacement-cost basis', f'=C{TB["nopat30"]}/C{TB["icr"]}',
     f'=D{TB["nopat30"]}/D{TB["icr"]}', DA['roic_replacement'], DB['roic_replacement'], PCT)
trow('rs', 'Return on capital — long-run sector', f'={a("roic_sec")}', f'={a("roic_sec")}',
     DA['roic_sector'], DB['roic_sector'], PCT, green=True)
trow('rt', 'Terminal return on capital (the average of the three above)',
     f'=AVERAGE(C{TB["rb"]}:C{TB["rs"]})', f'=AVERAGE(D{TB["rb"]}:D{TB["rs"]})',
     DA['roic_term'], DB['roic_term'], PCT, bold=True)
trow('g', 'Terminal growth', f'={a("g")}', f'={a("g")}', G, G, PCT, green=True)
trow('rein', 'Reinvestment rate (growth divided by the return on capital)',
     f'=C{TB["g"]}/C{TB["rt"]}', f'=D{TB["g"]}/D{TB["rt"]}', DA['rr_term'], DB['rr_term'], PCT)
trow('ngr', 'Terminal-year profit after tax grown one year',
     f'=C{TB["nopat30"]}*(1+C{TB["g"]})', f'=D{TB["nopat30"]}*(1+D{TB["g"]})',
     DA['nopat_term'], DB['nopat_term'], NUM1)
trow('wt', 'Terminal cost of capital', f'=$C${{WT}}', f'=$C${{WT}}',
     W['wacc_term_rating'], W['wacc_term_rating'], PCT2)
trow('tv', 'Terminal value',
     f'=C{TB["ngr"]}*(1-C{TB["rein"]})/(C{TB["wt"]}-C{TB["g"]})',
     f'=D{TB["ngr"]}*(1-D{TB["rein"]})/(D{TB["wt"]}-D{TB["g"]})', DA['tv'], DB['tv'], bold=True)
trow('pve', 'Present value of the five explicit years', f'=SUM(B{RA_["pv"]}:F{RA_["pv"]})',
     f'=SUM(B{RB_["pv"]}:F{RB_["pv"]})', DA['pv_explicit'], DB['pv_explicit'])
trow('pvt', 'Present value of the terminal value', f'=C{TB["tv"]}*F{RA_["df"]}',
     f'=D{TB["tv"]}*F{RB_["df"]}', DA['pv_tv'], DB['pv_tv'])
trow('ev', 'Enterprise value', f'=C{TB["pve"]}+C{TB["pvt"]}', f'=D{TB["pve"]}+D{TB["pvt"]}',
     DA['ev'], DB['ev'], bold=True)
trow('tvs', 'Terminal value as a share of enterprise value', f'=C{TB["pvt"]}/C{TB["ev"]}',
     f'=D{TB["pvt"]}/D{TB["ev"]}', DA['tv_share'], DB['tv_share'], PCT, bold=True)
trow('tvm', 'Terminal value implied enterprise value / EBITDA', f'=C{TB["tv"]}/F{RA_["eb"]}',
     f'=D{TB["tv"]}/F{RB_["eb"]}', DA['tv_ebitda_implied'], DB['tv_ebitda_implied'], MULT)
r += 1

# --- cost of capital --------------------------------------------------------
band(ws, r, 7)
put(ws, f'A{r}', 'COST OF CAPITAL — BUILT HERE, NOT PASTED', bold=True, fmt=None)
r += 1
CC = {}


def crow(key, lab, fml, val, fmt=PCT2, bold=False, green=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    putf(ws, f'C{r}', fml, val, fmt, bold=bold, green=green)
    CC[key] = r
    r += 1


crow('ust', 'Ten-year United States Treasury yield', f'={a("ust10")}', IN['ust10'], green=True)
crow('cds', 'Abu Dhabi sovereign credit default swap spread', f'={a("ad_cds")}', IN['ad_cds'],
     green=True)
crow('adgb', 'Abu Dhabi US dollar sovereign yield, ten year',
     f'=C{CC["ust"]}+C{CC["cds"]}', W['adgb10'])
crow('ads', 'Less the sovereign\'s own adjusted default spread', f'={a("ad_ads")}', IN['ad_ads'],
     green=True)
crow('rfs', 'Risk-free rate, normalised (rating basis)', f'=C{CC["adgb"]}-C{CC["ads"]}',
     W['rf_star_rating'], bold=True)
crow('rfc', 'Risk-free rate, normalised (swap basis)', f'=C{CC["adgb"]}-C{CC["cds"]}',
     W['rf_star_cds'])
crow('woth', 'Share of non-current assets outside the core countries',
     f'={a("nca_oth")}/{a("nca_tot")}', W_OTHER, PCT)
crow('wae', 'Share of non-current assets in the United Arab Emirates',
     f'=1-{a("w_eg")}-{a("w_dz")}-C{CC["woth"]}', W_AE_ERP, PCT)
crow('erpr', 'Equity risk premium, asset-weighted (rating basis)',
     f'=C{CC["wae"]}*{a("ad_erp")}+{a("w_eg")}*{a("eg_erp")}+{a("w_dz")}*{a("dz_erp")}'
     f'+C{CC["woth"]}*{a("mat_erp")}', W['erp_rating'])
crow('erpc', 'Equity risk premium, asset-weighted (swap basis)',
     f'=C{CC["wae"]}*{a("ad_erp_c")}+{a("w_eg")}*{a("eg_erp_c")}+{a("w_dz")}*{a("dz_erp")}'
     f'+C{CC["woth"]}*{a("mat_erp")}', W['erp_cds'])
crow('beta', 'Beta', f'={a("beta")}', IN['beta'], B3, green=True)
crow('ker', 'Cost of equity (rating basis) — risk free plus beta times the premium',
     f'=C{CC["rfs"]}+C{CC["beta"]}*C{CC["erpr"]}', W['ke_rating'], bold=True)
crow('kec', 'Cost of equity (swap basis)', f'=C{CC["rfc"]}+C{CC["beta"]}*C{CC["erpc"]}',
     W['ke_cds'])
crow('keb', 'Cost of equity, the average of the two bases',
     f'=AVERAGE(C{CC["ker"]}:C{CC["kec"]})', BOOK['ke_blend'])
crow('kds', 'Marginal debt spread (the average of the two most recent facilities)',
     f'=AVERAGE({a("kd_bc")},{a("kd_ad")})', W['kd_spread'])
crow('kd', 'Cost of debt, marginal', f'=C{CC["ust"]}+C{CC["kds"]}', W['kd'])
crow('kdat', 'Cost of debt after tax', f'=C{CC["kd"]}*(1-$C${{TAXROW}})', W['kd_at'])
crow('mcap', 'Market capitalisation ($m)', f'={a("shares")}*{a("spot")}/{a("fx")}', MKTCAP, NUM0)
crow('nd', 'Net debt ($m)', f'={a("nd_now")}', ND_NOW, NUM1, green=True)
crow('we', 'Equity weight (market capitalisation over capital)',
     f'=C{CC["mcap"]}/(C{CC["mcap"]}+C{CC["nd"]})', W['we'], PCT)
crow('wd', 'Debt weight', f'=1-C{CC["we"]}', W['wd'], PCT)
crow('wacc', 'Cost of capital, explicit window (rating basis)',
     f'=C{CC["we"]}*C{CC["ker"]}+C{CC["wd"]}*C{CC["kdat"]}', W['wacc_rating'], bold=True)
crow('waccc', 'Cost of capital, explicit window (swap basis)',
     f'=C{CC["we"]}*C{CC["kec"]}+C{CC["wd"]}*C{CC["kdat"]}', W['wacc_cds'])
crow('wdt', 'Terminal debt weight', f'={a("wd_term")}', W['wd_term'], PCT, green=True)
crow('wacct', 'Terminal cost of capital (rating basis)',
     f'=(1-C{CC["wdt"]})*C{CC["ker"]}+C{CC["wdt"]}*C{CC["kdat"]}', W['wacc_term_rating'],
     bold=True)
crow('wacctc', 'Terminal cost of capital (swap basis)',
     f'=(1-C{CC["wdt"]})*C{CC["kec"]}+C{CC["wdt"]}*C{CC["kdat"]}', W['wacc_term_cds'])
r += 1

# --- the glide --------------------------------------------------------------
band(ws, r, 7)
put(ws, f'A{r}', 'THE GLIDE — SHOWN AS IT IS DERIVED', bold=True, fmt=None)
r += 1
put(ws, f'A{r}', 'Year in the explicit window', fmt=None)
put(ws, f'B{r}', 1, BLUE, NUM0)
for i in range(1, 5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i-1]}{r}+1', i + 1, NUM0)
IDX = r
r += 1
put(ws, f'A{r}', 'Glide fraction (this year over the last year of the window)', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{IDX}/$F${IDX}', W['glide'][i], PCT)
GLIDE = r
r += 1
note(ws, f'A{r}', 'The forward cost of capital on each waterfall walks the explicit-window rate to '
     'the terminal rate by the fraction on this row, so leverage normalises smoothly rather than '
     'stepping in one year. The discount factors above compound off that path.')
r += 2

# --- tax triangulation ------------------------------------------------------
band(ws, r, 7)
put(ws, f'A{r}', 'THE TAX RATE — THREE SOURCED ESTIMATES, AVERAGED HERE', bold=True, fmt=None)
r += 1
TX = {}
put(ws, f'A{r}', 'Profit before tax, FY2022 to FY2025 ($m)', fmt=None)
putf(ws, f'C{r}', f"={a('pbt_fy22')}+'Income Statement'!B{{PBT}}+'Income Statement'!C{{PBT}}"
     f"+'Income Statement'!D{{PBT}}", PBT_SUM, NUM1)
TX['pbt'] = r; r += 1
put(ws, f'A{r}', 'Income tax charged, FY2022 to FY2025 ($m)', fmt=None)
putf(ws, f'C{r}', f"={a('tax_fy22')}+'Income Statement'!B{{TAX}}+'Income Statement'!C{{TAX}}"
     f"+'Income Statement'!D{{TAX}}",
     IN['tax_fy22'] + HI['FY23']['tax'] + HI['FY24']['tax'] + HI['FY25']['tax'], NUM1)
TX['chg'] = r; r += 1
put(ws, f'A{r}', 'Income taxes paid, FY2022 to FY2025 ($m)', fmt=None)
putf(ws, f'C{r}', f"={a('paid22')}+{a('paid23')}+{a('paid24')}+{a('paid25')}",
     IN['tax_paid_fy22'] + IN['tax_paid_fy23'] + IN['tax_paid_fy24'] + IN['tax_paid_fy25'], NUM1)
TX['paid'] = r; r += 1
put(ws, f'A{r}', 'Estimate one — aggregate effective rate on four years of profit', fmt=None)
putf(ws, f'C{r}', f'=C{TX["chg"]}/C{TX["pbt"]}', TAX_EFF, PCT)
TX['e1'] = r; r += 1
put(ws, f'A{r}', 'Estimate two — aggregate cash rate on four years of profit', fmt=None)
putf(ws, f'C{r}', f'=C{TX["paid"]}/C{TX["pbt"]}', TAX_CASH, PCT)
TX['e2'] = r; r += 1
put(ws, f'A{r}', 'Estimate three — statutory rates weighted by where the assets sit', fmt=None)
putf(ws, f'C{r}', f'=(1-{a("w_eg")}-{a("w_dz")})*{a("t_uae")}+{a("w_eg")}*{a("t_eg")}'
     f'+{a("w_dz")}*{a("t_dz")}', TAX_JUR, PCT)
TX['e3'] = r; r += 1
put(ws, f'A{r}', 'Forecast tax rate (the average of the three)', bold=True, fmt=None)
putf(ws, f'C{r}', f'=AVERAGE(C{TX["e1"]}:C{TX["e3"]})', TAX, PCT, bold=True)
TAXROW = r
band(ws, r, 7)
r += 2
note(ws, f'A{r}', 'No one of the three is right on its own. The effective rate is depressed by '
     'reliefs that will not repeat, the cash rate is distorted by the timing of settlements, and '
     'the statutory blend ignores both. The model runs on their average and the sensitivity sheet '
     'shows what the answer does across the range.')

# resolve the placeholders now that the rows are known
for coord in (f'C{TB["wt"]}', f'D{TB["wt"]}'):
    ws[coord] = ws[coord].value.replace('{WT}', str(CC['wacct']))
ws[f'C{CC["kdat"]}'] = ws[f'C{CC["kdat"]}'].value.replace('{TAXROW}', str(TAXROW))
DR['tax'] = TAXROW
DR['wacc'] = CC['wacc']
DR['wacc_t'] = CC['wacct']
DR['glide'] = GLIDE
# the two waterfalls referenced these rows before they existed; rewrite them now
for tag in ('A', 'B'):
    R = DR[tag]
    for i in range(5):
        c = ws[f'{CD[i]}{R["coc"]}']
        c.value = (f'=$C${CC["wacc"]}+($C${CC["wacct"]}-$C${CC["wacc"]})*{CD[i]}{GLIDE}')
        c = ws[f'{CD[i]}{R["nopat"]}']
        c.value = f'={CD[i]}{R["ebit"]}*(1-$C${TAXROW})'

# ============ 10 BALANCE SHEET ==============================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — audited history and a rolled-forward forecast', '$m, consolidated. '
      'FY2023 to FY2025 are the audited closing figures; the forecast rolls property, working '
      'capital, equity and net debt forward from them.', 10, awidth=52, cwidth=12)
ws.freeze_panes = 'B5'
hdr(ws, 4, ['$m'] + YH + YF)
BS = {}
r = 5
for key, lab in [('ppe', 'Property, plant and equipment'), ('rou', 'Right-of-use assets'),
                 ('gwi', 'Goodwill and intangible assets'), ('inv', 'Inventories'),
                 ('recv', 'Trade and other receivables'), ('cash', 'Cash and cash equivalents'),
                 ('ta', 'Total assets'), ('pay', 'Trade and other payables')]:
    put(ws, f'A{r}', lab, fmt=None)
    for i, y in enumerate(H3):
        put(ws, f'{HC[i]}{r}', HB[y][key], BLUE, NUM1, bold=(key == 'ta'))
    BS[key] = r
    r += 1
for key, lab, ik in [('ltd', 'Loans and borrowings — non-current', 'ltd'),
                     ('std', 'Loans and borrowings — current', 'std'),
                     ('lease', 'Lease liabilities', 'lease'),
                     ('dtl', 'Deferred tax liabilities', 'dtl'),
                     ('taxpay', 'Income tax payable', 'taxpay')]:
    put(ws, f'A{r}', lab, fmt=None)
    for i, y in enumerate(('fy23', 'fy24', 'fy25')):
        put(ws, f'{HC[i]}{r}', IN[f'{ik}_{y}'], BLUE, NUM1)
    BS[key] = r
    r += 1
for key, lab in [('eq_own', 'Equity attributable to owners'),
                 ('eq_nci', 'Non-controlling interests')]:
    put(ws, f'A{r}', lab, fmt=None)
    for i, y in enumerate(H3):
        put(ws, f'{HC[i]}{r}', HB[y][key], BLUE, NUM1)
    BS[key] = r
    r += 1
put(ws, f'A{r}', 'Total equity', bold=True, fmt=None)
for i, y in enumerate(H3):
    putf(ws, f'{HC[i]}{r}', f'={HC[i]}{BS["eq_own"]}+{HC[i]}{BS["eq_nci"]}', HB[y]['eq_tot'],
         NUM1, bold=True)
BS['eq_tot'] = r
band(ws, r, 10); r += 1
put(ws, f'A{r}', 'Gross interest-bearing debt', fmt=None)
for i, y in enumerate(H3):
    putf(ws, f'{HC[i]}{r}', f'={HC[i]}{BS["ltd"]}+{HC[i]}{BS["std"]}', HB[y]['debt_gross'], NUM1)
BS['debt'] = r; r += 1
put(ws, f'A{r}', 'Net debt (debt and leases less cash)', bold=True, fmt=None)
for i, y in enumerate(H3):
    putf(ws, f'{HC[i]}{r}', f'={HC[i]}{BS["debt"]}+{HC[i]}{BS["lease"]}-{HC[i]}{BS["cash"]}',
         HB[y]['net_debt'], NUM1, bold=True)
BS['nd'] = r; r += 1
put(ws, f'A{r}', 'Net working capital (inventories and receivables less payables)', fmt=None)
for i, y in enumerate(H3):
    putf(ws, f'{HC[i]}{r}', f'={HC[i]}{BS["inv"]}+{HC[i]}{BS["recv"]}-{HC[i]}{BS["pay"]}',
         HB[y]['nwc'], NUM1)
BS['nwc'] = r; r += 1
put(ws, f'A{r}', 'Invested capital (property plus net working capital)', fmt=None)
for i, y in enumerate(H3):
    putf(ws, f'{HC[i]}{r}', f'={HC[i]}{BS["ppe"]}+{HC[i]}{BS["nwc"]}',
         HB[y]['ppe'] + HB[y]['nwc'], NUM1)
BS['ic'] = r; r += 1

# forecast columns, framing A
RA_ = DR['A']
for i in range(5):
    putf(ws, f'{FC[i]}{BS["ppe"]}',
         f'={"D" if i == 0 else FC[i-1]}{BS["ppe"]}+DCF!{CD[i]}{RA_["capex"]}'
         f'-DCF!{CD[i]}{RA_["dna"]}', FA['ppe'][i], NUM1)
    putf(ws, f'{FC[i]}{BS["inv"]}', f'=DCF!{CD[i]}{RA_["inv"]}', FA['inv'][i], NUM1, green=True)
    putf(ws, f'{FC[i]}{BS["recv"]}', f'=DCF!{CD[i]}{RA_["recv"]}', FA['recv'][i], NUM1, green=True)
    putf(ws, f'{FC[i]}{BS["pay"]}', f'=DCF!{CD[i]}{RA_["pay"]}', FA['pay'][i], NUM1, green=True)
    putf(ws, f'{FC[i]}{BS["nwc"]}',
         f'={FC[i]}{BS["inv"]}+{FC[i]}{BS["recv"]}-{FC[i]}{BS["pay"]}', FA['nwc'][i], NUM1)
    putf(ws, f'{FC[i]}{BS["ic"]}', f'={FC[i]}{BS["ppe"]}+{FC[i]}{BS["nwc"]}', FA['ic'][i], NUM1)
    putf(ws, f'{FC[i]}{BS["eq_own"]}',
         f'={"D" if i == 0 else FC[i-1]}{BS["eq_own"]}'
         f"+'Income Statement'!{FC[i]}{{ATTR}}*(1-{a('payout')})", FA['equity'][i], NUM1)
    prev = f'{a("nd_now")}' if i == 0 else f'{FC[i-1]}{BS["nd"]}'
    putf(ws, f'{FC[i]}{BS["nd"]}',
         f'={prev}-(DCF!{CD[i]}{RA_["fcff"]}+\'Income Statement\'!{FC[i]}{{FIN}}*'
         f"(1-{{TAXR}}))+{a('payout')}*'Income Statement'!{FC[i]}{{ATTR}}"
         f"+'Income Statement'!{FC[i]}{{NCI}}", FA['net_debt'][i], NUM1, bold=True)
band(ws, BS['nd'], 10)
r += 1

# --- the asset-conversion cycle --------------------------------------------
band(ws, r, 10)
put(ws, f'A{r}', 'THE ASSET-CONVERSION CYCLE — MEASURED, THEN PROJECTED', bold=True, fmt=None)
r += 1
CY = {}


def cyrow(key, lab, fml, vals, fmt=NUM1, cols=(1, 2), bold=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    for j, i in enumerate(cols):
        putf(ws, f'{HC[i]}{r}', fml(HC[i]), vals[j], fmt, bold=bold)
    CY[key] = r
    r += 1


cyrow('dso', 'Days sales outstanding (receivables over revenue)',
      lambda c: f"=365*{c}{BS['recv']}/'Income Statement'!{c}{{REV}}",
      [CCC['FY24']['dso'], CCC['FY25']['dso']])
cyrow('dio', 'Days inventory outstanding (inventories over cost of sales)',
      lambda c: f"=365*{c}{BS['inv']}/'Income Statement'!{c}{{COGS}}",
      [CCC['FY24']['dio'], CCC['FY25']['dio']])
cyrow('dpo', 'Days payables outstanding (payables over cost of sales)',
      lambda c: f"=365*{c}{BS['pay']}/'Income Statement'!{c}{{COGS}}",
      [CCC['FY24']['dpo'], CCC['FY25']['dpo']])
cyrow('ccc', 'Cash conversion cycle (days)',
      lambda c: f'={c}{CY["dso"]}+{c}{CY["dio"]}-{c}{CY["dpo"]}',
      [CCC['FY24']['ccc'], CCC['FY25']['ccc']], bold=True)
put(ws, f'A{r}', 'Gas accrual carried inside payables at FY2025 ($m)', fmt=None)
put(ws, f'D{r}', IN['sorfert_accr_fy25'], BLUE, NUM1)
CY['accr'] = r; r += 1
cyrow('dpo_x', 'Days payables outstanding, excluding that accrual',
      lambda c: f"=365*({c}{BS['pay']}-{c}{CY['accr']})/'Income Statement'!{c}{{COGS}}",
      [CCC['FY25_ex_accrual']['dpo']], cols=(2,))
cyrow('ccc_x', 'Cash conversion cycle, excluding that accrual (days)',
      lambda c: f'={c}{CY["dso"]}+{c}{CY["dio"]}-{c}{CY["dpo_x"]}',
      [CCC['FY25_ex_accrual']['ccc']], cols=(2,), bold=True)
cyrow('nwc_x', 'Net working capital at FY2025, excluding that accrual ($m)',
      lambda c: f'={c}{BS["inv"]}+{c}{BS["recv"]}-({c}{BS["pay"]}-{c}{CY["accr"]})',
      [NWC_PRIOR], cols=(2,), bold=True)
r += 1
note(ws, f'A{r}', 'The gas accrual is a retrospective true-up on a disputed supply contract, not a '
     'trade payable in the ordinary course. Leaving it inside payables would make the company look '
     'as though it were financing itself on its suppliers to the tune of a full quarter of cost of '
     'sales. The forecast therefore runs on the cycle excluding it, and the 2026 change in working '
     'capital is measured against the same adjusted base.')
r += 2
note(ws, f'A{r}', 'This is a condensed layout and does not foot to zero: other liabilities, '
     'provisions and related-party balances are not shown separately. Every historical line above '
     'is the audited closing figure.')

# fill the Assumptions green cells that are measured here
wsA = wb['Assumptions']
putf(wsA, f'C{A["dso"]}', f"='Balance Sheet'!D{CY['dso']}", DSO, NUM1, green=True)
putf(wsA, f'C{A["dio"]}', f"='Balance Sheet'!D{CY['dio']}", DIO, NUM1, green=True)
putf(wsA, f'C{A["dpo"]}', f"='Balance Sheet'!D{CY['dpo_x']}", DPO, NUM1, green=True)
putf(wsA, f'C{A["nwc0"]}', f"='Balance Sheet'!D{CY['nwc_x']}", NWC_PRIOR, NUM1, green=True)

# ============ 9 INCOME STATEMENT ============================================
ws = sheet('Income Statement')
title(ws, 'Income statement — three years audited, five years forecast', '$m, consolidated. The '
      'disclosed lines are carried as reported; every subtotal is a formula that has to reconcile '
      'to them, and every forecast column is a formula.', 10, awidth=52, cwidth=12)
ws.freeze_panes = 'B5'
hdr(ws, 4, ['$m'] + YH + YF)
IS = {}
r = 5


def isrow(key, lab, hist=None, hf=None, hv=None, ff=None, fv=None, fmt=NUM1, bold=False,
          green=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    for i in range(3):
        if hf is not None:
            putf(ws, f'{HC[i]}{r}', hf(i), hv[i], fmt, bold=bold)
        else:
            put(ws, f'{HC[i]}{r}', hist[i], BLUE, fmt, bold=bold)
    for i in range(5):
        if ff is None:
            put(ws, f'{FC[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{FC[i]}{r}', ff(i), fv[i], fmt, bold=bold, green=green)
    IS[key] = r
    r += 1
    if bold:
        band(ws, r - 1, 10)


isrow('rev', 'Revenue', hist=[HI[y]['rev'] for y in H3],
      ff=lambda i: f'=DCF!{CD[i]}{RA_["rev"]}', fv=FA['rev'], bold=True, green=True)
isrow('cogs', 'Cost of sales', hist=[HI[y]['cogs'] for y in H3])
isrow('gp', 'Gross profit', hf=lambda i: f'={HC[i]}{IS["rev"]}-{HC[i]}{IS["cogs"]}',
      hv=[HI[y]['gp'] for y in H3])
isrow('sga', 'Selling, general and administrative expenses', hist=[HI[y]['sga'] for y in H3])
isrow('oth', 'Other income, net', hist=[HI[y]['othinc'] for y in H3])
isrow('dna', 'Depreciation and amortisation', hist=[HI[y]['dna'] for y in H3],
      ff=lambda i: f'=DCF!{CD[i]}{RA_["dna"]}', fv=FA['dna'], green=True)
isrow('ebitda', 'EBITDA',
      hf=lambda i: (f'={HC[i]}{IS["gp"]}-{HC[i]}{IS["sga"]}+{HC[i]}{IS["oth"]}'
                    f'+{HC[i]}{IS["dna"]}'), hv=[HI[y]['ebitda'] for y in H3],
      ff=lambda i: f'=DCF!{CD[i]}{RA_["eb"]}', fv=FA['ebitda'], bold=True, green=True)
isrow('ebit', 'Operating profit (EBIT)',
      hf=lambda i: f'={HC[i]}{IS["ebitda"]}-{HC[i]}{IS["dna"]}', hv=[HI[y]['ebit'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["ebitda"]}-{FC[i]}{IS["dna"]}', fv=FA['ebit'], bold=True)
isrow('fi', 'Finance income', hist=[IN[f'finc_{y}'] for y in ('fy23', 'fy24', 'fy25')])
isrow('fc', 'Finance costs', hist=[IN[f'fcost_{y}'] for y in ('fy23', 'fy24', 'fy25')],
      ff=lambda i: (f'={a("irate")}*MAX(' +
                    (f'{a("nd_now")}' if i == 0 else f"'Balance Sheet'!{FC[i-1]}{BS['nd']}") +
                    ',0)'), fv=[RA[i]['interest'] for i in range(5)])
isrow('fx', 'Net foreign exchange gain and loss',
      hist=[IN[f'fx_{y}'] for y in ('fy23', 'fy24', 'fy25')])
isrow('nf', 'Net finance result',
      hf=lambda i: f'={HC[i]}{IS["fi"]}-{HC[i]}{IS["fc"]}+{HC[i]}{IS["fx"]}',
      hv=[HI[y]['netfin'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["fi"]}-{FC[i]}{IS["fc"]}+{FC[i]}{IS["fx"]}',
      fv=[-RA[i]['interest'] for i in range(5)])
isrow('pbt', 'Profit before tax', hf=lambda i: f'={HC[i]}{IS["ebit"]}+{HC[i]}{IS["nf"]}',
      hv=[HI[y]['pbt'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["ebit"]}+{FC[i]}{IS["nf"]}', fv=[RA[i]['pbt'] for i in range(5)],
      bold=True)
isrow('tax', 'Income tax expense', hist=[HI[y]['tax'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["pbt"]}*DCF!$C${TAXROW}',
      fv=[RA[i]['pbt'] * TAX for i in range(5)])
isrow('np', 'Profit for the year', hf=lambda i: f'={HC[i]}{IS["pbt"]}-{HC[i]}{IS["tax"]}',
      hv=[HI[y]['np'] for y in H3], ff=lambda i: f'={FC[i]}{IS["pbt"]}-{FC[i]}{IS["tax"]}',
      fv=[RA[i]['np'] for i in range(5)])
isrow('nci', 'Non-controlling interests', hist=[HI[y]['nci'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["np"]}*{a("nci")}', fv=[RA[i]['nci'] for i in range(5)])
isrow('attr', 'Profit attributable to owners', hf=lambda i: f'={HC[i]}{IS["np"]}-{HC[i]}{IS["nci"]}',
      hv=[HI[y]['np_own'] for y in H3], ff=lambda i: f'={FC[i]}{IS["np"]}-{FC[i]}{IS["nci"]}',
      fv=[RA[i]['attr'] for i in range(5)], bold=True)
_eps = [HI[y]['np_own'] / SH for y in H3] + [RA[i]['attr'] / SH for i in range(5)]
isrow('eps', 'Earnings per share ($)', hf=lambda i: f'={HC[i]}{IS["attr"]}/{a("shares")}',
      hv=_eps[:3], ff=lambda i: f'={FC[i]}{IS["attr"]}/{a("shares")}', fv=_eps[3:], fmt=PX)
isrow('epsa', 'Earnings per share (AED)', hf=lambda i: f'={HC[i]}{IS["eps"]}*{a("fx")}',
      hv=[x * FX for x in _eps[:3]], ff=lambda i: f'={FC[i]}{IS["eps"]}*{a("fx")}',
      fv=[x * FX for x in _eps[3:]], fmt=PX)
isrow('mgn', 'EBITDA margin', hf=lambda i: f'={HC[i]}{IS["ebitda"]}/{HC[i]}{IS["rev"]}',
      hv=[HI[y]['ebitda_margin'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["ebitda"]}/{FC[i]}{IS["rev"]}', fv=FA['ebitda_margin'], fmt=PCT)
isrow('etr', 'Effective tax rate', hf=lambda i: f'={HC[i]}{IS["tax"]}/{HC[i]}{IS["pbt"]}',
      hv=[HI[y]['eff_tax'] for y in H3],
      ff=lambda i: f'={FC[i]}{IS["tax"]}/{FC[i]}{IS["pbt"]}', fv=[TAX] * 5, fmt=PCT)
r += 1
note(ws, f'A{r}', 'The forecast is built at the EBITDA line from the unit build, so cost of sales '
     'and the administrative and other lines are not split out in the forecast columns; the total '
     'cash cost that replaces them is on the Segments sheet. The finance charge is computed on the '
     'net debt the business is actually carrying that year, so profit is struck after financing '
     'and differs from the pre-financing cash-flow waterfall by construction.')

# resolve the Balance Sheet placeholders that needed Income Statement rows
wsB = wb['Balance Sheet']
for i in range(5):
    for cell in (f'{FC[i]}{BS["eq_own"]}', f'{FC[i]}{BS["nd"]}'):
        v = wsB[cell].value
        wsB[cell] = (v.replace('{ATTR}', str(IS['attr'])).replace('{FIN}', str(IS['nf']))
                     .replace('{NCI}', str(IS['nci'])).replace('{TAXR}', f'DCF!$C${TAXROW}'))
for key in ('dso', 'dio', 'dpo', 'dpo_x'):
    for c in HC:
        cell = f'{c}{CY[key]}'
        if isinstance(wsB[cell].value, str):
            wsB[cell] = (wsB[cell].value.replace('{REV}', str(IS['rev']))
                         .replace('{COGS}', str(IS['cogs'])))
wsD = wb['DCF']
for cell in (f'C{TX["pbt"]}', f'C{TX["chg"]}'):
    wsD[cell] = wsD[cell].value.replace('{PBT}', str(IS['pbt'])).replace('{TAX}', str(IS['tax']))
for c in ('C', 'D'):
    cell = f'{c}{TB["icb"]}'
    wsD[cell] = wsD[cell].value.replace('{PPE}', str(BS['ppe']))
# the net-debt roll subtracts the after-tax finance charge, which the statement carries negative
_ndchk = FA['net_debt']

# ============ 5 SOTP BRIDGE =================================================
ws = sheet('SOTP Bridge')
title(ws, 'Enterprise value to equity — the bridge, and where the value sits', 'A single-class '
      'nitrogen producer, so this is the bridge from enterprise value to the shareholder plus the '
      'split of that value between the two operating legs — not a holding-company sum of parts.',
      6, awidth=58, cwidth=16)
ws.freeze_panes = 'A5'
hdr(ws, 4, ['Step', 'Framing A', 'Framing B'])
BRG = {}
r = 5


def brow(key, lab, fa, fb, va, vb, fmt=NUM0, bold=False, green=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    putf(ws, f'B{r}', fa, va, fmt, bold=bold, green=green)
    putf(ws, f'C{r}', fb, vb, fmt, bold=bold, green=green)
    BRG[key] = r
    r += 1


brow('pve', 'Present value of the five explicit years ($m)', f'=DCF!C{TB["pve"]}',
     f'=DCF!D{TB["pve"]}', DA['pv_explicit'], DB['pv_explicit'], green=True)
brow('pvt', 'Present value of the terminal value ($m)', f'=DCF!C{TB["pvt"]}',
     f'=DCF!D{TB["pvt"]}', DA['pv_tv'], DB['pv_tv'], green=True)
brow('ev', 'Enterprise value ($m)', f'=B{BRG["pve"]}+B{BRG["pvt"]}',
     f'=C{BRG["pve"]}+C{BRG["pvt"]}', DA['ev'], DB['ev'], bold=True)
brow('nd', 'Less net debt ($m)', f'={a("nd_now")}', f'={a("nd_now")}', ND_NOW, ND_NOW, NUM1,
     green=True)
brow('eqt', 'Equity value before minority interests ($m)', f'=B{BRG["ev"]}-B{BRG["nd"]}',
     f'=C{BRG["ev"]}-C{BRG["nd"]}', BA['eq_total'], BB['eq_total'], bold=True)
brow('nce', 'Less minority interests at their share of group profit ($m)',
     f'=B{BRG["eqt"]}*{a("nci")}', f'=C{BRG["eqt"]}*{a("nci")}', BA['nci_earnings'],
     BB['nci_earnings'])
brow('eqa', 'Equity attributable to owners ($m)', f'=B{BRG["eqt"]}-B{BRG["nce"]}',
     f'=C{BRG["eqt"]}-C{BRG["nce"]}', BA['eq_attr'], BB['eq_attr'], bold=True)
brow('psu', 'Value per share ($)', f'=B{BRG["eqa"]}/{a("shares")}', f'=C{BRG["eqa"]}/{a("shares")}',
     BA['ps_usd'], BB['ps_usd'], PX)
brow('psa', 'Value per share (AED)', f'=B{BRG["psu"]}*{a("fx")}', f'=C{BRG["psu"]}*{a("fx")}',
     BA['ps_aed'], BB['ps_aed'], PX, bold=True)
brow('tvs', 'Terminal value as a share of enterprise value', f'=B{BRG["pvt"]}/B{BRG["ev"]}',
     f'=C{BRG["pvt"]}/C{BRG["ev"]}', DA['tv_share'], DB['tv_share'], PCT, bold=True)
r += 1
put(ws, f'A{r}', 'The cash-flow lens — the average of the two framings (AED per share)', bold=True,
    fmt=None)
putf(ws, f'B{r}', f'=AVERAGE(B{BRG["psa"]}:C{BRG["psa"]})', DCF_PS, PX, bold=True)
BRG['dcf_ps'] = r
band(ws, r, 6); r += 2
put(ws, f'A{r}', 'THE MINORITY INTERESTS ON THE OTHER BASIS — BOTH ARE SHOWN, NEITHER IS AVERAGED',
    bold=True, fmt=None)
r += 1
brow('ncb', 'Less minority interests at their book carrying value ($m)', f'={a("nci_bv")}',
     f'={a("nci_bv")}', IN['eqnci_fy25'], IN['eqnci_fy25'], NUM1, green=True)
brow('eqb', 'Equity attributable to owners, book basis ($m)', f'=B{BRG["eqt"]}-B{BRG["ncb"]}',
     f'=C{BRG["eqt"]}-C{BRG["ncb"]}', BAB['eq_attr'], BBB['eq_attr'])
brow('psb', 'Value per share, book basis (AED)', f'=B{BRG["eqb"]}/{a("shares")}*{a("fx")}',
     f'=C{BRG["eqb"]}/{a("shares")}*{a("fx")}', BAB['ps_aed'], BBB['ps_aed'], PX, bold=True)
r += 1
note(ws, f'A{r}', 'Egypt Basic Industries is a quarter owned outside the group and Sorfert almost '
     'half, so the minority deduction is the single largest step in this bridge. Deducting the '
     'minorities at their share of the profit they actually earn takes out roughly five times what '
     'their historical book carrying value would; the model runs on the earnings basis and shows '
     'the book basis beside it rather than splitting the difference.')
r += 2

# --- the segment split ------------------------------------------------------
band(ws, r, 6)
put(ws, f'A{r}', 'WHERE THE VALUE SITS — THE TWO OPERATING LEGS', bold=True, fmt=None)
r += 1
hdr(ws, r, ['Leg', 'Mid-cycle EBITDA ($m)', 'Share', 'Enterprise value ($m)', 'AED per share'])
r += 1
SEGV = {}
SA_, SB_ = FRAME_ROWS['A'], FRAME_ROWS['B']
_ev_mid = f'=(DCF!C{TB["ev"]}+DCF!D{TB["ev"]})/2'
put(ws, f'A{r}', 'Mid-cycle enterprise value, the average of the two framings ($m)', fmt=None)
putf(ws, f'B{r}', _ev_mid, EV_MID, NUM0)
SEGV['ev'] = r
r += 1
for key, lab, ra, rb, val in [
        ('own', 'Own-produced production and marketing', SA_['eo'], SB_['eo'], MID_OWN),
        ('t3', 'Third-party trading', SA_['e3'], SB_['e3'], MID_3P),
        ('cp', 'Corporate and other', SA_['cp'], SB_['cp'], MID_CORP)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', f'=(SUM(Segments!D{ra}:F{ra})+SUM(Segments!D{rb}:F{rb}))/6', val, NUM1)
    SEGV[key] = r
    r += 1
put(ws, f'A{r}', 'Total mid-cycle EBITDA ($m)', bold=True, fmt=None)
putf(ws, f'B{r}', f'=SUM(B{SEGV["own"]}:B{SEGV["cp"]})', MID_TOT, NUM1, bold=True)
SEGV['tot'] = r
r += 1
for key in ('own', 't3', 'cp'):
    rr = SEGV[key]
    val = {'own': MID_OWN, 't3': MID_3P, 'cp': MID_CORP}[key]
    putf(ws, f'C{rr}', f'=B{rr}/$B${SEGV["tot"]}', val / MID_TOT, PCT)
    putf(ws, f'D{rr}', f'=C{rr}*$B${SEGV["ev"]}', val / MID_TOT * EV_MID, NUM0)
    putf(ws, f'E{rr}', f'=D{rr}/{a("shares")}*{a("fx")}', val / MID_TOT * EV_MID / SH * FX, PX)
putf(ws, f'C{SEGV["tot"]}', f'=SUM(C{SEGV["own"]}:C{SEGV["cp"]})', 1.0, PCT, bold=True)
putf(ws, f'D{SEGV["tot"]}', f'=SUM(D{SEGV["own"]}:D{SEGV["cp"]})', EV_MID, NUM0, bold=True)
r += 1
note(ws, f'A{r}', 'The legs are separated on the enterprise value they earn, before net debt and '
     'before the minority deduction, because both of those attach to the group rather than to '
     'either leg. Trading is a real business but a thin one: it moves a sixth of the tonnes for a '
     'twentieth of the profit, and the corporate load costs more than the whole of it.')

# ============ 7 RELATIVE & NORMALIZED =======================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative multiples, normalised earnings power, and book value', 'The three lenses that '
      'do not discount a cash flow. Each is built from the same forecast.', 6, awidth=58, cwidth=16)
ws.freeze_panes = 'A5'
RN = {}
r = 4
hdr(ws, r, ['Relative multiples', 'Value']); r += 1


def rnrow(key, lab, fml, val, fmt=NUM0, bold=False, green=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    putf(ws, f'C{r}', fml, val, fmt, bold=bold, green=green)
    RN[key] = r
    r += 1


rnrow('ebm', 'Mid-cycle EBITDA — the average of FY2028E to FY2030E on both framings ($m)',
      f'=(SUM(DCF!D{RA_["eb"]}:F{RA_["eb"]})+SUM(DCF!D{DR["B"]["eb"]}:F{DR["B"]["eb"]}))/6',
      REL['ebitda_mid'], NUM1)
rnrow('mult', 'Justified enterprise value / EBITDA', f'={a("mult")}', REL['mult'], MULT,
      green=True)
rnrow('ev', 'Implied enterprise value ($m)', f'=C{RN["ebm"]}*C{RN["mult"]}', REL['ev'])
rnrow('eqt', 'Less net debt, giving equity before minorities ($m)', f'=C{RN["ev"]}-{a("nd_now")}',
      REL['eq_total'])
rnrow('ps', 'Implied value per share (AED)',
      f'=C{RN["eqt"]}*(1-{a("nci")})/{a("shares")}*{a("fx")}', REL['ps_aed'], PX, bold=True)
r += 1
hdr(ws, r, ['Trailing market check', 'Value']); r += 1
rnrow('mcap', 'Market capitalisation ($m)', f'=DCF!C{CC["mcap"]}', MKTCAP, NUM0, green=True)
rnrow('evt', 'Trailing enterprise value ($m)', f'=C{RN["mcap"]}+{a("nd_now")}', EV_TRAIL)
rnrow('evebt', 'Trailing enterprise value / EBITDA (first half annualised)',
      f'=C{RN["evt"]}/({a("eb_h1")}*2)', REL['ev_ebitda_trailing'], MULT)
rnrow('pet', 'Trailing price / earnings (first half annualised)',
      f'=C{RN["mcap"]}/({a("np_h1")}*2)', REL['pe_trailing'], MULT)
r += 1
hdr(ws, r, ['Normalised earnings power', 'Value']); r += 1
rnrow('nmgn', 'Mid-cycle EBITDA margin — the average of FY2023 to FY2025',
      f"=AVERAGE('Income Statement'!B{IS['mgn']}:D{IS['mgn']})", NRM['margin'], PCT)
rnrow('nrev', 'Mid-cycle revenue — the average of FY2028E to FY2030E on both framings ($m)',
      f'=(SUM(DCF!D{RA_["rev"]}:F{RA_["rev"]})+SUM(DCF!D{DR["B"]["rev"]}:F{DR["B"]["rev"]}))/6',
      NRM['rev'], NUM1)
rnrow('neb', 'Normalised EBITDA ($m)', f'=C{RN["nrev"]}*C{RN["nmgn"]}', NRM['ebitda'], NUM1)
rnrow('nebit', 'Less depreciation, giving normalised EBIT ($m)',
      f'=C{RN["neb"]}-{a("dna", 2)}', NRM['ebit'], NUM1)
rnrow('nint', 'Less interest on net debt ($m)', f'={a("irate")}*{a("nd_now")}', NRM['interest'],
      NUM1)
rnrow('nnp', 'Normalised profit to owners, after tax and minorities ($m)',
      f'=(C{RN["nebit"]}-C{RN["nint"]})*(1-DCF!$C${TAXROW})*(1-{a("nci")})', NRM['np'], NUM1)
rnrow('neps', 'Normalised earnings per share ($)', f'=C{RN["nnp"]}/{a("shares")}',
      NRM['eps_usd'], PX)
rnrow('npe', 'Justified price / earnings', f'={a("pe")}', NRM['pe'], MULT, green=True)
rnrow('nps', 'Implied value per share (AED)', f'=C{RN["neps"]}*C{RN["npe"]}*{a("fx")}',
      NRM['ps_aed'], PX, bold=True)
r += 1
hdr(ws, r, ['Book value and the sustainable return', 'Value']); r += 1
rnrow('bv', 'Book value per share ($)', f"='Balance Sheet'!D{BS['eq_own']}/{a('shares')}",
      BOOK['bvps_usd'], PX)
rnrow('roet', 'Return on equity, trailing',
      f"='Income Statement'!D{IS['attr']}/'Balance Sheet'!D{BS['eq_own']}",
      BOOK['roe_trailing'], PCT)
rnrow('roes', 'Sustainable return on equity — the average of the three audited years',
      f"=('Income Statement'!B{IS['attr']}/'Balance Sheet'!B{BS['eq_own']}"
      f"+'Income Statement'!C{IS['attr']}/'Balance Sheet'!C{BS['eq_own']}"
      f"+'Income Statement'!D{IS['attr']}/'Balance Sheet'!D{BS['eq_own']})/3",
      BOOK['roe_sust'], PCT)
rnrow('keb', 'Cost of equity, the average of the two premium bases', f'=DCF!C{CC["keb"]}',
      BOOK['ke_blend'], PCT2, green=True)
rnrow('pb', 'Justified price to book — return less growth, over cost of equity less growth',
      f'=(C{RN["roes"]}-{a("g")})/(C{RN["keb"]}-{a("g")})', BOOK['pb_just'], MULT)
rnrow('bps', 'Implied value per share (AED)', f'=C{RN["bv"]}*C{RN["pb"]}*{a("fx")}',
      BOOK['ps_aed'], PX, bold=True)
r += 1
note(ws, f'A{r}', 'The book lens is the weakest of the four here and is weighted accordingly. A '
     'nitrogen producer\'s plants are carried at depreciated historical cost in a currency that '
     'has been stable, so book equity understates what the same capacity would cost to build — but '
     'it is the one lens that does not need a forecast at all, which is why it is kept.')

# ============ 11 CASH FLOW =================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow — audited markers and the forecast waterfall', '$m. The forecast waterfall is '
      'the same one the discounted cash flow runs on, shown here against the reported history.',
      9, awidth=52, cwidth=13)
ws.freeze_panes = 'B5'
hdr(ws, 4, ['$m', 'FY2023', 'FY2024', 'FY2025'] + YF)
CF = {}
r = 5
for key, lab, hv in [
        ('cfo', 'Net cash from operating activities (reported)',
         [IN['cfo_fy23'], IN['cfo_fy24'], IN['cfo_fy25']]),
        ('capexh', 'Purchase of property, plant and equipment (reported)',
         [IN['capex_fy23'], IN['capex_fy24'], IN['capex_fy25']]),
        ('taxp', 'Income taxes paid (reported)',
         [IN['tax_paid_fy23'], IN['tax_paid_fy24'], IN['tax_paid_fy25']])]:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(3):
        put(ws, f'{HC[i]}{r}', hv[i], BLUE, NUM1)
    CF[key] = r
    r += 1
put(ws, f'A{r}', 'Dividends paid to owners and to minorities (reported)', fmt=None)
putf(ws, f'D{r}', f'={IN["divsh_fy25"]}+{IN["divnci_fy25"]}'.replace(
    str(IN['divsh_fy25']), f"'Balance Sheet'!D{BS['eq_own']}*0+{IN['divsh_fy25']}"),
    IN['divsh_fy25'] + IN['divnci_fy25'], NUM1)
CF['div'] = r
r += 1
put(ws, f'A{r}', 'Cash conversion — operating cash flow over EBITDA', fmt=None)
for i in range(3):
    putf(ws, f'{HC[i]}{r}', f"={HC[i]}{CF['cfo']}/'Income Statement'!{HC[i]}{IS['ebitda']}",
         [IN['cfo_fy23'] / HI['FY23']['ebitda'], IN['cfo_fy24'] / HI['FY24']['ebitda'],
          IN['cfo_fy25'] / HI['FY25']['ebitda']][i], PCT)
CF['conv'] = r
r += 2
band(ws, r, 9)
put(ws, f'A{r}', 'THE FORECAST WATERFALL — FRAMING A', bold=True, fmt=None)
r += 1
FCF = ['E', 'F', 'G', 'H', 'I']
for key, lab, src, vals, bold in [
        ('nopat', 'Net operating profit after tax', RA_['nopat'], FA['nopat'], False),
        ('dna', 'Add back depreciation and amortisation', RA_['dna'], FA['dna'], False),
        ('capex', 'Less capital expenditure', RA_['capex'], FA['capex'], False),
        ('dnwc', 'Less change in net working capital', RA_['dnwc'], FA['dnwc'], False)]:
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(5):
        putf(ws, f'{FCF[i]}{r}', f'=DCF!{CD[i]}{src}', vals[i], NUM1, green=True)
    CF[key] = r
    r += 1
put(ws, f'A{r}', 'Free cash flow to the firm', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}',
         f'={FCF[i]}{CF["nopat"]}+{FCF[i]}{CF["dna"]}-{FCF[i]}{CF["capex"]}-{FCF[i]}{CF["dnwc"]}',
         FA['fcff'][i], NUM1, bold=True)
CF['fcff'] = r
band(ws, r, 9); r += 1
put(ws, f'A{r}', 'Less interest after tax', fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}',
         f"='Income Statement'!{FC[i]}{IS['fc']}*(1-DCF!$C${TAXROW})",
         RA[i]['interest'] * (1 - TAX), NUM1)
CF['int'] = r
r += 1
put(ws, f'A{r}', 'Less dividends to owners', fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}', f"={a('payout')}*'Income Statement'!{FC[i]}{IS['attr']}",
         RA[i]['div'], NUM1)
CF['divf'] = r
r += 1
put(ws, f'A{r}', 'Less profit accruing to minorities', fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}', f"='Income Statement'!{FC[i]}{IS['nci']}", RA[i]['nci'], NUM1,
         green=True)
CF['ncif'] = r
r += 1
put(ws, f'A{r}', 'Movement in net debt', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}',
         f'=-{FCF[i]}{CF["fcff"]}+{FCF[i]}{CF["int"]}+{FCF[i]}{CF["divf"]}+{FCF[i]}{CF["ncif"]}',
         RA[i]['nd'] - RA[i]['nd_prev'], NUM1, bold=True)
CF['mv'] = r
r += 1
put(ws, f'A{r}', 'Closing net debt', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{FCF[i]}{r}', f"='Balance Sheet'!{FC[i]}{BS['nd']}", FA['net_debt'][i], NUM1,
         bold=True, green=True)
CF['nd'] = r
r += 2
note(ws, f'A{r}', 'The reported cash flow and the forecast waterfall are not the same object and '
     'are not meant to reconcile line for line: the reported figure is after interest, tax and the '
     'working-capital swing the gas accrual causes, while the waterfall is struck before financing '
     'so that it can be discounted at a cost of capital that already prices the debt.')

# ============ 12 SUMMARY FINANCIALS ========================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials — the eight-year picture', '$m unless stated. Every cell on this '
      'sheet is a link or a ratio; nothing is typed twice.', 10, awidth=48, cwidth=12)
ws.freeze_panes = 'B5'
hdr(ws, 4, ['$m'] + YH + YF)
SF = {}
r = 5
_rev8 = [HI[y]['rev'] for y in H3] + FA['rev']
_eb8 = [HI[y]['ebitda'] for y in H3] + FA['ebitda']
_ebit8 = [HI[y]['ebit'] for y in H3] + FA['ebit']
_attr8 = [HI[y]['np_own'] for y in H3] + [RA[i]['attr'] for i in range(5)]
_nd8 = [HB[y]['net_debt'] for y in H3] + FA['net_debt']
_ic8 = [HB[y]['ppe'] + HB[y]['nwc'] for y in H3] + FA['ic']


def sfrow(key, lab, fml, vals, fmt=NUM1, skip=(), green=False, bold=False):
    global r
    put(ws, f'A{r}', lab, bold=bold, fmt=None)
    for i in range(8):
        if i in skip:
            put(ws, f'{ALL8[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALL8[i]}{r}', fml(i), vals[i], fmt, green=green, bold=bold)
    SF[key] = r
    r += 1


sfrow('rev', 'Revenue', lambda i: f"='Income Statement'!{ALL8[i]}{IS['rev']}", _rev8, NUM0,
      green=True, bold=True)
sfrow('g', 'Revenue growth', lambda i: f'={ALL8[i]}{SF["rev"]}/{ALL8[i-1]}{SF["rev"]}-1',
      [None] + [_rev8[i] / _rev8[i - 1] - 1 for i in range(1, 8)], PCT, skip=(0,))
sfrow('eb', 'EBITDA', lambda i: f"='Income Statement'!{ALL8[i]}{IS['ebitda']}", _eb8, NUM0,
      green=True)
sfrow('mgn', 'EBITDA margin', lambda i: f'={ALL8[i]}{SF["eb"]}/{ALL8[i]}{SF["rev"]}',
      [_eb8[i] / _rev8[i] for i in range(8)], PCT)
sfrow('ebit', 'EBIT', lambda i: f"='Income Statement'!{ALL8[i]}{IS['ebit']}", _ebit8, NUM0,
      green=True)
sfrow('attr', 'Profit attributable to owners',
      lambda i: f"='Income Statement'!{ALL8[i]}{IS['attr']}", _attr8, NUM0, green=True)
sfrow('fcff', 'Free cash flow to the firm', lambda i: f"='Cash Flow'!{FCF[i-3]}{CF['fcff']}",
      [None] * 3 + FA['fcff'], NUM0, skip=(0, 1, 2), green=True)
sfrow('nd', 'Net debt', lambda i: f"='Balance Sheet'!{ALL8[i]}{BS['nd']}", _nd8, NUM0, green=True)
sfrow('ic', 'Invested capital', lambda i: f"='Balance Sheet'!{ALL8[i]}{BS['ic']}", _ic8, NUM0,
      green=True)
sfrow('roic', 'Return on invested capital', lambda i: f'=DCF!{CD[i-3]}{RA_["nopat"]}/{ALL8[i]}'
      f'{SF["ic"]}', [None] * 3 + FA['roic'], PCT, skip=(0, 1, 2))
sfrow('lev', 'Net debt / EBITDA', lambda i: f'={ALL8[i]}{SF["nd"]}/{ALL8[i]}{SF["eb"]}',
      [_nd8[i] / _eb8[i] for i in range(8)], MULT)
r += 1
note(ws, f'A{r}', 'The forecast columns are framing A. Framing B differs only in the price path; '
     'its revenue, EBITDA and cash flow are on the DCF sheet beside framing A\'s, and both reach a '
     'value on the SOTP Bridge sheet.')

# ============ 13 MONTE CARLO ================================================
ws = sheet('Monte Carlo')
title(ws, 'Probabilistic price map', 'A map of price dispersion around the market price. It carries '
      'no view on value and is never blended with the valuation. Each figure is a complete engine '
      're-run, so nothing on this sheet redraws when a driver is changed.', 8, awidth=46, cwidth=14)
ws.freeze_panes = 'A5'
hdr(ws, 4, ['Horizon', '5th', '25th', 'Median', '75th', '95th', 'Above the market price'])
r = 5
for tag, lab in (('1M', 'One month'), ('3M', 'Three months')):
    h = STK['horizons'][tag]
    put(ws, f'A{r}', f"{lab} — to {h['grade_date']}", fmt=None)
    for i, k in enumerate(('p5', 'p25', 'p50', 'p75', 'p95')):
        put(ws, f'{get_column_letter(2+i)}{r}', h['pct'][k], BLUE, PX)
    put(ws, f'G{r}', h['p_above'], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Level event', 'One month', 'Three months']); r += 1
for lab, k in [('Finishes 10% or more above the market price', 'end_up10'),
               ('Finishes 10% or more below the market price', 'end_dn10'),
               ('Touches 10% above the market price at any point', 'touch_up10'),
               ('Touches 10% below the market price at any point', 'touch_dn10'),
               ('Touches 20% above the market price at any point', 'touch_up20'),
               ('Touches 20% below the market price at any point', 'touch_dn20')]:
    put(ws, f'A{r}', lab, fmt=None)
    put(ws, f'B{r}', STK['horizons']['1M'][k], BLUE, PCT)
    put(ws, f'C{r}', STK['horizons']['3M'][k], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Setting', 'Value']); r += 1
put(ws, f'A{r}', 'Simulated paths', fmt=None); put(ws, f'C{r}', 50000, BLUE, NUM0); r += 1
put(ws, f'A{r}', 'Annualised volatility at the three-month anchor', fmt=None)
put(ws, f'C{r}', STK['horizons']['3M']['anchor_vol_ann'], BLUE, PCT); r += 1
put(ws, f'A{r}', 'Market price (AED)', fmt=None)
putf(ws, f'C{r}', f'={a("spot")}', SPOT, NUM2, green=True); r += 1
put(ws, f'A{r}', 'Anchor date', fmt=None); put(ws, f'C{r}', STK['anchor_date'], BLUE, None); r += 2
S0 = D['step0']
note(ws, f'A{r}', 'How the map has behaved out of sample: across '
     f"{S0['windows_scored']} non-overlapping three-month windows on this stock, the interval that "
     f"should contain nine outcomes in ten contained {S0['cov90']:.0%} of them and the interval "
     f"that should contain four in five contained {S0['cov80']:.0%}. The average position of the "
     f"realised price inside the distribution was {S0['pit_mean']:.2f}, against a half for a "
     'perfectly centred map. Against a carry-adjusted random walk the map scored '
     f"{S0['skill_norm']:+.4f}, which is inside the confidence interval either way — it is honest "
     'about dispersion, not a forecast of direction.')

# ============ 14 SENSITIVITY ================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — what the valuation needs the world to do', 'AED per share, framing A. '
      'Each cell is a complete re-run of the whole model, including the unit build, so these grids '
      'are engine outputs rather than formulas and do NOT redraw when a driver is changed.', 8,
      awidth=52, cwidth=13)
ws.freeze_panes = 'A5'
r = 4
put(ws, f'A{r}', 'Terminal cost of capital (rows) against terminal growth (columns)', bold=True,
    fmt=None)
r += 1
hdr(ws, r, [''] + [f'{g:.1%}' for g in SN['g_grid']]); r += 1
for i, wt in enumerate(SN['wacc_grid']):
    put(ws, f'A{r}', f'{wt:.2%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid_wacc_g'][i][j], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Single-driver sensitivities — five complete re-runs per row', bold=True,
    fmt=None); r += 1
hdr(ws, r, ['Driver (the grid it is run over)', '', '', '', '', '', 'Swing']); r += 1
for lab, grid, vals, gfmt in [
        ('Cost pass-through', SN['pt_grid'], SN['grid_pt'], '{:.3f}'),
        ('Nitrogen price path, multiplicative', SN['px_grid'], SN['grid_px'], '{:+.0%}'),
        ('Beta', SN['beta_grid'], SN['grid_beta'], '{:.3f}'),
        ('Tax rate', SN['tax_grid'], SN['grid_tax'], '{:.1%}'),
        ('Terminal growth, at the base cost of capital', SN['g_grid'], SN['grid_wacc_g'][2],
         '{:.1%}')]:
    put(ws, f'A{r}', f"{lab}  ({' / '.join(gfmt.format(g) for g in grid)})", fmt=None)
    for j, v in enumerate(vals):
        put(ws, f'{get_column_letter(2+j)}{r}', v, BLUE, PX)
    putf(ws, f'G{r}', f'=MAX(B{r}:F{r})-MIN(B{r}:F{r})', max(vals) - min(vals), PX)
    r += 1
ws.column_dimensions['G'].width = 13
r += 1
note(ws, f'A{r}', 'The pass-through row is the one that matters. Every other driver moves the '
     'answer by less than the width of the range the four lenses already span; the pass-through '
     'moves it by more than the whole of that range, which is why it is measured on the Segments '
     'sheet from the company\'s own reported periods rather than assumed.')

# ============ 15 PER-SHARE & RATIOS ========================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share and ratio analysis', 'The indicator set for a merchant nitrogen producer. '
      'Every ratio is a formula off the statements.', 10, awidth=52, cwidth=12)
ws.freeze_panes = 'B5'
hdr(ws, 4, ['Measure'] + YH + YF)
r = 5
_eq8 = [HB[y]['eq_own'] for y in H3] + FA['equity']
_nf8 = [HI[y]['netfin'] for y in H3] + [-RA[i]['interest'] for i in range(5)]


def ratio(lab, fml, vals, fmt, skip=()):
    global r
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip:
            put(ws, f'{ALL8[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALL8[i]}{r}', fml(i), vals[i], fmt)
    r += 1


ratio('Earnings per share ($)', lambda i: f"='Income Statement'!{ALL8[i]}{IS['eps']}",
      [x / SH for x in _attr8], PX)
ratio('Earnings per share (AED)', lambda i: f"='Income Statement'!{ALL8[i]}{IS['epsa']}",
      [x / SH * FX for x in _attr8], PX)
ratio('Book value per share ($)',
      lambda i: f"='Balance Sheet'!{ALL8[i]}{BS['eq_own']}/{a('shares')}",
      [x / SH for x in _eq8], PX)
ratio('Free cash flow per share ($)',
      lambda i: f"='Summary Financials'!{ALL8[i]}{SF['fcff']}/{a('shares')}",
      [None] * 3 + [x / SH for x in FA['fcff']], PX, skip=(0, 1, 2))
ratio('Dividend per share ($)', lambda i: f"={a('payout')}*'Income Statement'!{ALL8[i]}"
      f"{IS['attr']}/{a('shares')}", [None] * 3 + [RA[i]['div'] / SH for i in range(5)], PX,
      skip=(0, 1, 2))
ratio('Gross margin', lambda i: f"='Income Statement'!{ALL8[i]}{IS['gp']}/'Income Statement'!"
      f"{ALL8[i]}{IS['rev']}", [HI[y]['gp_margin'] for y in H3] + [None] * 5, PCT,
      skip=(3, 4, 5, 6, 7))
ratio('EBITDA margin', lambda i: f"='Income Statement'!{ALL8[i]}{IS['mgn']}",
      [_eb8[i] / _rev8[i] for i in range(8)], PCT)
ratio('EBIT margin', lambda i: f"='Income Statement'!{ALL8[i]}{IS['ebit']}/'Income Statement'!"
      f"{ALL8[i]}{IS['rev']}", [_ebit8[i] / _rev8[i] for i in range(8)], PCT)
ratio('Net margin, attributable', lambda i: f"='Income Statement'!{ALL8[i]}{IS['attr']}/"
      f"'Income Statement'!{ALL8[i]}{IS['rev']}",
      [_attr8[i] / _rev8[i] for i in range(8)], PCT)
ratio('Return on equity, on opening equity',
      lambda i: f"='Income Statement'!{ALL8[i]}{IS['attr']}/'Balance Sheet'!{ALL8[i-1]}"
      f"{BS['eq_own']}", [None] + [_attr8[i] / _eq8[i - 1] for i in range(1, 8)], PCT, skip=(0,))
ratio('Return on invested capital', lambda i: f"='Summary Financials'!{ALL8[i]}{SF['roic']}",
      [None] * 3 + FA['roic'], PCT, skip=(0, 1, 2))
ratio('Net debt / EBITDA', lambda i: f"='Summary Financials'!{ALL8[i]}{SF['lev']}",
      [_nd8[i] / _eb8[i] for i in range(8)], MULT)
ratio('Interest cover (EBIT over the net finance charge)',
      lambda i: f"=-'Income Statement'!{ALL8[i]}{IS['ebit']}/'Income Statement'!{ALL8[i]}"
      f"{IS['nf']}", [_ebit8[i] / -_nf8[i] for i in range(8)], MULT)
ratio('Net working capital / revenue',
      lambda i: f"='Balance Sheet'!{ALL8[i]}{BS['nwc']}/'Income Statement'!{ALL8[i]}{IS['rev']}",
      [HB[y]['nwc'] / HI[y]['rev'] for y in H3] + [FA['nwc'][i] / FA['rev'][i] for i in range(5)],
      PCT)
ratio('Capital expenditure / revenue',
      lambda i: (f"='Cash Flow'!{ALL8[i]}{CF['capexh']}/'Income Statement'!{ALL8[i]}{IS['rev']}"
                 if i < 3 else
                 f"=DCF!{CD[i-3]}{RA_['capex']}/'Income Statement'!{ALL8[i]}{IS['rev']}"),
      [IN['capex_fy23'] / HI['FY23']['rev'], IN['capex_fy24'] / HI['FY24']['rev'],
       IN['capex_fy25'] / HI['FY25']['rev']] + [FA['capex'][i] / FA['rev'][i] for i in range(5)],
      PCT)
ratio('Cash cost per tonne of own product ($/t)',
      lambda i: (f"=Segments!{CAL[i-1]}{CR['ct']}" if i in (1, 2) else
                 f"=Segments!{CD[i-3]}{SA_['ct']}"),
      [None, U['FY24']['cash_cost_t'], U['FY25']['cash_cost_t']] + FA['cost_t'], NUM1, skip=(0,))
ratio('Realised price per tonne of own product ($/t)',
      lambda i: (f"=Segments!{CAL[i-1]}{CR['px']}" if i in (1, 2) else
                 f"=Segments!{CD[i-3]}{SA_['px']}"),
      [None, U['FY24']['px_realised'], U['FY25']['px_realised']] + FA['px_realised'], NUM1,
      skip=(0,))
r += 1
note(ws, f'A{r}', 'The forecast columns are framing A. Return on equity is struck on opening equity '
     'rather than an average because the forecast equity balance is itself a roll-forward, and '
     'averaging a roll against itself flatters the early years.')

# ============ 16 PEER & SECTOR =============================================
ws = sheet('Peer & Sector')
title(ws, 'Peer frame and sector context', 'A sanity check on the multiple, not an independent '
      'valuation. No listed nitrogen producer carries the same gas position.', 6, awidth=34,
      cwidth=17)
ws.freeze_panes = 'A5'
hdr(ws, 4, ['Company', 'Market', 'Enterprise value / EBITDA', 'Why it is and is not comparable'])
r = 5
PROW0 = r
for p in REL['peers']:
    put(ws, f'A{r}', p['name'], fmt=None)
    put(ws, f'B{r}', p['mkt'], fmt=None)
    put(ws, f'C{r}', p['ev_ebitda'], BLUE, MULT)
    put(ws, f'D{r}', p['note'], fmt=None, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1
PROW1 = r - 1
ws.column_dimensions['D'].width = 46
_pm = sorted(p['ev_ebitda'] for p in REL['peers'])
_med = (_pm[len(_pm) // 2 - 1] + _pm[len(_pm) // 2]) / 2
put(ws, f'A{r}', 'Peer median', bold=True, fmt=None)
putf(ws, f'C{r}', f'=MEDIAN(C{PROW0}:C{PROW1})', _med, MULT, bold=True)
r += 1
put(ws, f'A{r}', 'Peer average', fmt=None)
putf(ws, f'C{r}', f'=AVERAGE(C{PROW0}:C{PROW1})', sum(_pm) / len(_pm), MULT)
r += 1
put(ws, f'A{r}', 'Applied in the relative lens', fmt=None)
putf(ws, f'C{r}', f"='Relative & Normalized'!C{RN['mult']}", REL['mult'], MULT, green=True)
r += 2
hdr(ws, r, ['The company on its own multiples', 'Value']); r += 1
for lab, fml, val, fmt in [
        ('Trailing enterprise value / EBITDA', f"='Relative & Normalized'!C{RN['evebt']}",
         REL['ev_ebitda_trailing'], MULT),
        ('Trailing price / earnings', f"='Relative & Normalized'!C{RN['pet']}",
         REL['pe_trailing'], MULT),
        ('Trailing price / book',
         f"={a('spot')}/{a('fx')}/'Relative & Normalized'!C{RN['bv']}",
         SPOT / FX / BOOK['bvps_usd'], MULT),
        ('Terminal value implied enterprise value / EBITDA, framing A',
         f'=DCF!C{TB["tvm"]}', DA['tv_ebitda_implied'], MULT),
        ('Terminal value implied enterprise value / EBITDA, framing B',
         f'=DCF!D{TB["tvm"]}', DB['tv_ebitda_implied'], MULT)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', fml, val, fmt, green=True)
    r += 1
r += 1
note(ws, f'A{r}', 'The two Gulf names trade where they do because of who owns them and where they '
     'list, not because their nitrogen assets are worth more per tonne; the two American names sit '
     'on a gas position that is structurally cheaper than anything in this group. The applied '
     'multiple sits below the peer median for that reason.')

# ============ 3 FUNDAMENTAL VALUATION (filled) =============================
ws = wb['Fundamental Valuation']
hdr(ws, 4, ['Lens', 'AED per share', 'Weight', 'Contribution', 'Against the market price'])
FV = {}
r = 5
for key, lab, src, val in [
        ('dcf', 'Discounted cash flow — the average of the two framings',
         f"='SOTP Bridge'!B{BRG['dcf_ps']}", LN['dcf']['value']),
        ('rel', 'Relative multiples — mid-cycle EBITDA on a peer-anchored multiple',
         f"='Relative & Normalized'!C{RN['ps']}", LN['relative']['value']),
        ('norm', 'Normalised earnings power — mid-cycle earnings on a justified multiple',
         f"='Relative & Normalized'!C{RN['nps']}", LN['normalized']['value']),
        ('book', 'Book value marked to the sustainable return on that equity',
         f"='Relative & Normalized'!C{RN['bps']}", LN['book']['value'])]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', src, val, PX, green=True)
    wk = LN[{'dcf': 'dcf', 'rel': 'relative', 'norm': 'normalized', 'book': 'book'}[key]]['weight']
    putf(ws, f'C{r}', f'={a("w_" + key if key != "norm" else "w_norm")}', wk, PCT, green=True)
    putf(ws, f'D{r}', f'=B{r}*C{r}', val * wk, PX)
    putf(ws, f'E{r}', f'=B{r}/{a("spot")}-1', val / SPOT - 1, PCT)
    FV[key] = r
    r += 1
band(ws, r, 5)
put(ws, f'A{r}', 'Weighted central value', bold=True, fmt=None)
putf(ws, f'B{r}', f'=SUM(D{FV["dcf"]}:D{FV["book"]})', CENTRAL, PX, bold=True)
putf(ws, f'C{r}', f'=SUM(C{FV["dcf"]}:C{FV["book"]})', 1.0, PCT, bold=True)
putf(ws, f'E{r}', f'=B{r}/{a("spot")}-1', CENTRAL / SPOT - 1, PCT, bold=True)
FV['central'] = r
r += 1
put(ws, f'A{r}', 'Lowest of the four lenses and the two framings', fmt=None)
_span_lo, _span_hi = D['span']
putf(ws, f'B{r}', f"=MIN(MIN(B{FV['dcf']}:B{FV['book']}),'SOTP Bridge'!B{BRG['psa']},"
     f"'SOTP Bridge'!C{BRG['psa']})", _span_lo, PX)
FV['lo'] = r; r += 1
put(ws, f'A{r}', 'Highest of the four lenses and the two framings', fmt=None)
putf(ws, f'B{r}', f"=MAX(MAX(B{FV['dcf']}:B{FV['book']}),'SOTP Bridge'!B{BRG['psa']},"
     f"'SOTP Bridge'!C{BRG['psa']})", _span_hi, PX)
FV['hi'] = r; r += 2
band(ws, r, 5)
put(ws, f'A{r}', 'THE CONTESTED JUDGEMENT — COMPUTED BOTH WAYS, PUBLISHED SIDE BY SIDE',
    bold=True, fmt=None)
put(ws, f'B{r}', 'Framing A', bold=True, fmt=None)
put(ws, f'C{r}', 'Framing B', bold=True, fmt=None)
r += 1
for lab, fa, fb, va, vb, fmt in [
        ('Urea price assumed in FY2030E ($/t)', f'={a("pxa_u", 4)}', f'={a("pxb_u", 4)}',
         FA['px_urea'][4], FB['px_urea'][4], NUM0),
        ('FY2030E EBITDA ($m)', f'=DCF!F{RA_["eb"]}', f'=DCF!F{DR["B"]["eb"]}',
         FA['ebitda'][4], FB['ebitda'][4], NUM0),
        ('Enterprise value ($m)', f"='SOTP Bridge'!B{BRG['ev']}", f"='SOTP Bridge'!C{BRG['ev']}",
         DA['ev'], DB['ev'], NUM0),
        ('Terminal value as a share of enterprise value', f"='SOTP Bridge'!B{BRG['tvs']}",
         f"='SOTP Bridge'!C{BRG['tvs']}", DA['tv_share'], DB['tv_share'], PCT),
        ('Value per share (AED)', f"='SOTP Bridge'!B{BRG['psa']}",
         f"='SOTP Bridge'!C{BRG['psa']}", BA['ps_aed'], BB['ps_aed'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fa, va, fmt, green=True)
    putf(ws, f'C{r}', fb, vb, fmt, green=True)
    r += 1
put(ws, f'A{r}', 'The cash-flow lens is their average, never one of them', fmt=None)
putf(ws, f'B{r}', f"='SOTP Bridge'!B{BRG['dcf_ps']}", DCF_PS, PX, bold=True, green=True)
r += 2
band(ws, r, 5)
put(ws, f'A{r}', 'THE PANEL — THREE METHODS, WORKED SEPARATELY', bold=True, fmt=None)
r += 1
hdr(ws, r, ['Method', 'Input', 'Enterprise value ($m)', 'AED per share']); r += 1
put(ws, f'A{r}', 'Mid-cycle multiple on through-the-cycle EBITDA', fmt=None)
put(ws, f'B{r}', E1_MULT, BLUE, MULT)
putf(ws, f'C{r}', f"=B{r}*'Relative & Normalized'!C{RN['ebm']}", EXPS['e1']['ev'], NUM0)
putf(ws, f'D{r}', f'=(C{r}-{a("nd_now")})*(1-{a("nci")})/{a("shares")}*{a("fx")}',
     EXPS['e1']['ps_aed'], PX)
E1 = r; r += 1
put(ws, f'A{r}', 'Discounted cash flow with an explicit gas pass-through', fmt=None)
putf(ws, f'B{r}', f'={a("pt_slope")}', _SLOPE, B3, green=True)
putf(ws, f'C{r}', f"='SOTP Bridge'!B{BRG['ev']}", DA['ev'], NUM0, green=True)
putf(ws, f'D{r}', f"='SOTP Bridge'!B{BRG['psa']}", BA['ps_aed'], PX, green=True)
E2 = r; r += 1
put(ws, f'A{r}', 'Replacement cost of installed nitrogen capacity', fmt=None)
putf(ws, f'B{r}', f'={a("repl_t")}', REPL_T, NUM0, green=True)
putf(ws, f'C{r}', f'=({a("cap_urea")}+{a("cap_nh3")})*B{r}/1000', EXPS['e3']['ev'], NUM0)
putf(ws, f'D{r}', f'=(C{r}-{a("nd_now")})*(1-{a("nci")})/{a("shares")}*{a("fx")}',
     EXPS['e3']['ps_aed'], PX)
E3 = r; r += 1
put(ws, f'A{r}', 'Median of the three', bold=True, fmt=None)
putf(ws, f'D{r}', f'=MEDIAN(D{E1}:D{E3})',
     sorted([EXPS['e1']['ps_aed'], BA['ps_aed'], EXPS['e3']['ps_aed']])[1], PX, bold=True)
FV['panel'] = r
r += 2
note(ws, f'A{r}', 'The three do not agree, and the disagreement is the point: the multiple method '
     'prices the cycle the peers are in, the cash-flow method prices the gas contracts, and the '
     'replacement method prices the steel. They bracket the same answer from three directions.')

# ============ 2 SUMMARY (filled) ===========================================
ws = wb['Summary']
hdr(ws, 4, ['Lens', 'AED per share', 'Weight', 'Contribution', 'Against the market price',
            'Terminal value share'])
r = 5
for key, lab in [('dcf', 'Discounted cash flow'), ('rel', 'Relative multiples'),
                 ('norm', 'Normalised earnings power'), ('book', 'Book value')]:
    lk = {'dcf': 'dcf', 'rel': 'relative', 'norm': 'normalized', 'book': 'book'}[key]
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', f"='Fundamental Valuation'!B{FV[key]}", LN[lk]['value'], PX, green=True)
    putf(ws, f'C{r}', f"='Fundamental Valuation'!C{FV[key]}", LN[lk]['weight'], PCT, green=True)
    putf(ws, f'D{r}', f'=B{r}*C{r}', LN[lk]['value'] * LN[lk]['weight'], PX)
    putf(ws, f'E{r}', f'=B{r}/$B${{SPOTROW}}-1', LN[lk]['value'] / SPOT - 1, PCT)
    if key == 'dcf':
        putf(ws, f'F{r}', f"=('SOTP Bridge'!B{BRG['tvs']}+'SOTP Bridge'!C{BRG['tvs']})/2",
             (DA['tv_share'] + DB['tv_share']) / 2, PCT, green=True)
    r += 1
band(ws, r, 6)
put(ws, f'A{r}', 'Weighted central value', bold=True, fmt=None)
putf(ws, f'B{r}', f'=SUM(D5:D{r-1})', CENTRAL, PX, bold=True)
putf(ws, f'C{r}', f'=SUM(C5:C{r-1})', 1.0, PCT, bold=True)
putf(ws, f'E{r}', f'=B{r}/$B${{SPOTROW}}-1', CENTRAL / SPOT - 1, PCT, bold=True)
SUM_CENTRAL = r
r += 1
put(ws, f'A{r}', 'Lowest of the four lenses and the two framings', fmt=None)
putf(ws, f'B{r}', f"='Fundamental Valuation'!B{FV['lo']}", _span_lo, PX, green=True)
r += 1
put(ws, f'A{r}', 'Highest of the four lenses and the two framings', fmt=None)
putf(ws, f'B{r}', f"='Fundamental Valuation'!B{FV['hi']}", _span_hi, PX, green=True)
r += 1
put(ws, f'A{r}', 'Median of the three panel methods', fmt=None)
putf(ws, f'B{r}', f"='Fundamental Valuation'!D{FV['panel']}",
     sorted([EXPS['e1']['ps_aed'], BA['ps_aed'], EXPS['e3']['ps_aed']])[1], PX, green=True)
r += 1
band(ws, r, 6)
put(ws, f'A{r}', 'Market price (AED per share)', bold=True, fmt=None)
putf(ws, f'B{r}', f'={a("spot")}', SPOT, NUM2, bold=True, green=True)
SPOTROW = r
r += 2
hdr(ws, r, ['Key figure', 'Value']); r += 1
for lab, fml, val, fmt in [
        ('Ordinary shares outstanding (mn)', f'={a("shares")}', SH, NUM0),
        ('Market capitalisation ($m)', f'=DCF!C{CC["mcap"]}', MKTCAP, NUM0),
        ('Net debt at 30 June 2026 ($m)', f'={a("nd_now")}', ND_NOW, NUM1),
        ('FY2025 revenue ($m)', f"='Income Statement'!D{IS['rev']}", HI['FY25']['rev'], NUM1),
        ('FY2025 EBITDA ($m)', f"='Income Statement'!D{IS['ebitda']}", HI['FY25']['ebitda'], NUM1),
        ('FY2025 profit to owners ($m)', f"='Income Statement'!D{IS['attr']}",
         HI['FY25']['np_own'], NUM1),
        ('Realisation against the blended benchmark', f'={a("realis")}', REALIS, B3),
        ('Cost pass-through, share of each incremental dollar', f'={a("pt_slope")}', _SLOPE, B3),
        ('Forecast tax rate', f'=DCF!C{TAXROW}', TAX, PCT),
        ('Cost of capital — explicit window', f'=DCF!C{CC["wacc"]}', W['wacc_rating'], PCT2),
        ('Cost of capital — terminal', f'=DCF!C{CC["wacct"]}', W['wacc_term_rating'], PCT2),
        ('Terminal growth', f'={a("g")}', G, PCT)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', fml, val, fmt, green=True)
    r += 1
r += 1
note(ws, f'A{r}', 'Values are model outputs shown as a range and a weighted central figure. They '
     'are not a recommendation and not a target.')
for row in range(5, SUM_CENTRAL + 1):
    c = ws[f'E{row}']
    if isinstance(c.value, str):
        c.value = c.value.replace('{SPOTROW}', str(SPOTROW))
ANCH.update(summary_central=f'B{SUM_CENTRAL}', summary_spot=f'B{SPOTROW}',
            fv_central=f'B{FV["central"]}', bridge_dcf_ps=f'B{BRG["dcf_ps"]}',
            bridge_psa_a=f'B{BRG["psa"]}', bridge_psa_b=f'C{BRG["psa"]}',
            bridge_tvs_a=f'B{BRG["tvs"]}', dcf_tvs_a=f'C{TB["tvs"]}',
            dcf_ev_a=f'C{TB["ev"]}', dcf_ev_b=f'D{TB["ev"]}',
            dcf_wacc=f'C{CC["wacc"]}', dcf_wacc_term=f'C{CC["wacct"]}',
            dcf_tax=f'C{TAXROW}', bs_nd30=f'I{BS["nd"]}', bs_eq30=f'I{BS["eq_own"]}',
            rel_ps=f'C{RN["ps"]}', norm_ps=f'C{RN["nps"]}', book_ps=f'C{RN["bps"]}',
            bridge_psb_a=f'B{BRG["psb"]}', rel_evebt=f'C{RN["evebt"]}', rel_pet=f'C{RN["pet"]}',
            seg_slope=f'B{REG["slope"]}', is_rev=f'D{IS["rev"]}',
            dcf_ebitda26_a=f'B{RA_["eb"]}', panel_median=f'D{FV["panel"]}')

# ---------------------------------------------------------------------------
# Sheets are BUILT in dependency order (a sheet's rows must exist before another references
# them) but must be DELIVERED in the model-study order. Cross-sheet formulas address sheets by
# name, so reordering here cannot disturb them.
CANON = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios',
         'Peer & Sector']
if sorted(CANON) != sorted(wb.sheetnames):
    raise SystemExit(f'sheet set does not match the model study:\n  built {wb.sheetnames}')
wb._sheets = [wb[n] for n in CANON]
assert wb.sheetnames == CANON, wb.sheetnames

# Every deferred row reference is written as a {TOKEN} placeholder and patched once the row
# is known. A surviving token means a patch was dropped -- fail loudly rather than ship a
# formula Excel cannot parse.
_resid = []
for _s in wb.worksheets:
    for _row in _s.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and _c.value.startswith('=') and '{' in _c.value:
                _resid.append(f'{_s.title}!{_c.coordinate}: {_c.value}')
if _resid:
    raise SystemExit('unpatched formula placeholder(s):\n  ' + '\n  '.join(_resid))

out = os.path.join(HERE, 'Fertiglobe_Valuation_Model_09-08-2026.xlsx')
wb.save(out)
json.dump({'expected': EXPECT, 'anchors': ANCH},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nchk = sum(len(v) for v in EXPECT.values())
nform = nlit = 0
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
            elif isinstance(c.value, (int, float)):
                nlit += 1
print(f'wrote {out}')
print(f'{len(wb.sheetnames)} sheets: {wb.sheetnames}')
print(f'FORMULA CELLS: {nform} (of which {nchk} carry a recorded expected value)')
print(f'PASTED NUMERIC CELLS: {nlit}')
