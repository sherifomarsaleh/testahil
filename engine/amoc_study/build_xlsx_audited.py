"""AMOC_Valuation_Model_06082026_public.xlsx — formula-first workbook on the AUDITED statements.

Rewritten from the ground up for the audited emit. The previous builder was written against a
per-tonne cost construction (crack multiples, a solved feedstock differential, house yields) that
the filings replaced outright, so it is superseded rather than patched.

Design rule is unchanged and is the point of the file: IT CALCULATES, IT DOES NOT STORE. Every
figure derivable from a driver is a live formula. Only two classes of cell are pasted, and READ
FIRST names them: (1) audited/disclosed history, and (2) whole-model re-runs that cannot be
expressed as a formula inside a grid.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
IN = {k: v['value'] for k, v in D['inputs'].items()}
SRC = {k: v['source'] for k, v in D['inputs'].items()}
DAT = {k: v['date'] for k, v in D['inputs'].items()}
RNG = {k: v['ring'] for k, v in D['inputs'].items()}
AU, U, F, W, DCF = D['audited'], D['unit'], D['fcst'], D['wacc'], D['dcf']
BASE, LN, SN, EXP = D['base'], D['lenses'], D['sens'], D['experts']
REL, NRM, BK, TR = D['rel'], D['norm'], D['book'], D['terminal_recon']
STK, BT = D['strike'], D['backtest']
SH, SPOT = IN['shares_mn'], IN['spot']
YRS = F['years']
LINES = U['lines']; LBL = U['labels']

INK = '1C3A36'; PANEL = 'EAF0EE'; CREAM = 'F6F1E6'; GOLD = 'C0A45F'
BLUE = Font(color='1F4E9C', size=10)          # input
BLACK = Font(color='1C3A36', size=10)         # formula
GREEN = Font(color='2E6B4F', size=10)         # cross-sheet link
NUM0 = '#,##0'; NUM1 = '#,##0.0'; NUM3 = '#,##0.000'; PX = '#,##0.00'
PCT = '0.0%'; PCT2 = '0.00%'; PCT3 = '0.000%'; MULT = '0.0"x"'

wb = Workbook(); wb.remove(wb.active)
EXPECT = {}; ANCH = {}; TOK = {}


def sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = INK
    ws.column_dimensions['A'].width = 52
    for c in 'BCDEFGHIJ':
        ws.column_dimensions[c].width = 15
    ws.sheet_view.showGridLines = False
    return ws


def put(ws, ref, val, font=BLACK, fmt=None, bold=False):
    c = ws[ref]; c.value = val
    c.font = Font(color=font.color.rgb, size=font.size, bold=bold)
    if fmt:
        c.number_format = fmt
    return c


def putf(ws, ref, formula, value, fmt=None, bold=False, green=False):
    """Write a FORMULA and record the model's own value for it, so an independent evaluator can
    assert cell by cell that the workbook reproduces the model."""
    c = ws[ref]; c.value = formula
    c.font = Font(color=(GREEN if green else BLACK).color.rgb, size=10, bold=bold)
    if fmt:
        c.number_format = fmt
    EXPECT.setdefault(ws.title, {})[ref] = float(value)
    return c


def band(ws, row, ncol=8, fill=PANEL):
    for i in range(1, ncol + 1):
        ws.cell(row=row, column=i).fill = PatternFill('solid', fgColor=fill)


def title(ws, t, sub=None, ncol=8):
    ws['A1'] = t; ws['A1'].font = Font(bold=True, size=13, color=INK)
    if sub:
        ws['A2'] = sub; ws['A2'].font = Font(size=9.5, color='6E7B77', italic=True)
        ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        ws.row_dimensions[2].height = 30
    band(ws, 3, ncol, INK)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True, size=9.5, color=INK)
        c.fill = PatternFill('solid', fgColor=PANEL)
        c.alignment = Alignment(wrap_text=True, vertical='bottom')


def note(ws, row, text, ncol=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=8.5, color='6E7B77', italic=True)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    ws.row_dimensions[row].height = max(28, 12 * (len(text) // 110 + 1))


# ============ ASSUMPTIONS (built first — every other sheet points here) =======
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every driver in the model, with its source',
      'Blue cells are inputs. Change one and the whole workbook reprices. Company-layer inputs '
      'carry the audited or reviewed filing they were read from, with the note number where the '
      'filing gives one.', 6)
wsA.column_dimensions['A'].width = 46
wsA.column_dimensions['B'].width = 10
wsA.column_dimensions['C'].width = 18
wsA.column_dimensions['D'].width = 13
wsA.column_dimensions['E'].width = 11
wsA.column_dimensions['F'].width = 82
r = 4
hdr(wsA, r, ['Driver', 'Unit', 'Value', 'Date', 'Layer', 'Source']); r += 1
AR = {}


def drv(key, label, unit, fmt, ring=None):
    global r
    v = IN[key]
    put(wsA, f'A{r}', label, BLACK)
    put(wsA, f'B{r}', unit, BLACK)
    if isinstance(v, (int, float)):
        put(wsA, f'C{r}', v, BLUE, fmt)
    else:
        put(wsA, f'C{r}', str(v), BLUE)
    put(wsA, f'D{r}', DAT.get(key, ''), BLACK)
    put(wsA, f'E{r}', ring or RNG.get(key, ''), BLACK)
    c = put(wsA, f'F{r}', SRC.get(key, ''), BLACK)
    c.font = Font(size=8, color='6E7B77')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    wsA.row_dimensions[r].height = max(14, 9.5 * (len(SRC.get(key, '')) // 100 + 1))
    AR[key] = r
    r += 1


def A(key):
    return f"Assumptions!$C${AR[key]}"


band(wsA, r, 6, CREAM); put(wsA, f'A{r}', 'MARKET', BLACK, bold=True); r += 1
drv('spot', 'Share price', 'EGP', PX)
drv('shares_mn', 'Shares outstanding', 'mn', NUM1)
r += 1
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'THE AUDITED PERIODS — read off the filings, nothing triangulated', BLACK,
    bold=True); r += 1
for k, lab in (('rev_h2_25', 'Net sales, 6M to 31-Dec-2025'),
               ('cogs_h2_25', 'Cost of sales, 6M to 31-Dec-2025'),
               ('ga_h2_25', 'G&A expenses, 6M to 31-Dec-2025'),
               ('mkt_h2_25', 'Marketing and selling, 6M to 31-Dec-2025'),
               ('othexp_h2_25', 'Other operating expenses, 6M to 31-Dec-2025'),
               ('othrev_h2_25', 'Other revenue, 6M to 31-Dec-2025'),
               ('invinc_h2_25', 'Revenue from investments, 6M to 31-Dec-2025'),
               ('prov_h2_25', 'Formed provisions, 6M to 31-Dec-2025'),
               ('fin_h2_25', 'Finance expenses, 6M to 31-Dec-2025'),
               ('tax_h2_25', 'Current income tax, 6M to 31-Dec-2025'),
               ('dtax_h2_25', 'Deferred tax credit, 6M to 31-Dec-2025'),
               ('pat_h2_25', 'Net profit after tax, 6M to 31-Dec-2025'),
               ('nci_h2_25', 'Minority share of profit, 6M to 31-Dec-2025'),
               ('emp_h2_25', 'Employees profit share and BOD bonuses, 6M to Dec-2025'),
               ('rev_q1_26', 'Net sales, 3M to 31-Mar-2026'),
               ('cogs_q1_26', 'Cost of sales, 3M to 31-Mar-2026'),
               ('ga_q1_26', 'G&A expenses, 3M to 31-Mar-2026'),
               ('mkt_q1_26', 'Marketing and selling, 3M to 31-Mar-2026'),
               ('othexp_q1_26', 'Other operating expenses, 3M to 31-Mar-2026'),
               ('othrev_q1_26', 'Other revenue, 3M to 31-Mar-2026'),
               ('prov_q1_26', 'Formed provisions, 3M to 31-Mar-2026'),
               ('ecl_q1_26', 'Expected credit losses formed, 3M to 31-Mar-2026'),
               ('fin_q1_26', 'Finance cost, 3M to 31-Mar-2026'),
               ('tax_q1_26', 'Current income tax, 3M to 31-Mar-2026'),
               ('dtax_q1_26', 'Deferred tax, 3M to 31-Mar-2026'),
               ('pat_q1_26', 'Net profit after tax, 3M to 31-Mar-2026'),
               ('nci_q1_26', 'Minority share of profit, 3M to 31-Mar-2026'),
               ('rev_h2_24', 'Net sales, 6M to 31-Dec-2024'),
               ('cogs_h2_24', 'Cost of sales, 6M to 31-Dec-2024'),
               ('pat_h2_24', 'Net profit after tax, 6M to 31-Dec-2024'),
               ('rev_q1_25', 'Net sales, 3M to 31-Mar-2025'),
               ('cogs_q1_25', 'Cost of sales, 3M to 31-Mar-2025'),
               ('pat_fy25_full', 'Majority profit, year to 30-Jun-2025')):
    drv(k, lab, 'EGP', NUM0)
r += 1
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'NOTE 15-A — THE AUDITED COST STACK', BLACK, bold=True); r += 1
for k, lab in (('cos_raw', 'Cost of sales — raw materials, 6M to Dec-2025'),
               ('cos_salaries', 'Cost of sales — salaries, 6M to Dec-2025'),
               ('cos_other', 'Cost of sales — other (gas, power, water, spares, EPROM)'),
               ('cos_support', 'Cost of sales — supporting materials'),
               ('cos_dep', 'Cost of sales — depreciation'),
               ('cos_raw_24', 'Cost of sales — raw materials, 6M to Dec-2024'),
               ('cos_salaries_24', 'Cost of sales — salaries, 6M to Dec-2024'),
               ('cos_other_24', 'Cost of sales — other, 6M to Dec-2024'),
               ('cos_support_24', 'Cost of sales — supporting materials, 6M to Dec-2024'),
               ('cos_dep_24', 'Cost of sales — depreciation, 6M to Dec-2024')):
    drv(k, lab, 'EGP', NUM0)
r += 1
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'AUDITED BALANCE SHEET AND CASH FLOW', BLACK, bold=True); r += 1
for k, lab, f_ in (('ppe_net', 'Fixed assets, net (note 6)', NUM0),
                   ('puc', 'Projects under construction (note 7)', NUM0),
                   ('inventory', 'Inventory, net (note 9-A)', NUM0),
                   ('recv', 'Accounts receivable, net (note 9-B)', NUM0),
                   ('debtors', 'Debtors and other debit balances', NUM0),
                   ('cash', 'Cash at banks and on hand (note 9-E)', NUM0),
                   ('fin_inv', 'Pledged deposits — NOT free cash', NUM0),
                   ('payables', 'Accounts and notes payable (note 10-3)', NUM0),
                   ('creditors', 'Creditors and other credit balances (note 11)', NUM0),
                   ('provisions', 'Provisions — tax disputes and claims (note 10-1)', NUM0),
                   ('debt_lt', 'Long-term loans', NUM0),
                   ('debt_st', 'Short-term loans and facilities', NUM0),
                   ('assets_snap', 'Total assets', NUM0),
                   ('liab_snap', 'Total liabilities', NUM0),
                   ('eq_parent', 'Total AMOC (parent) equity', NUM0),
                   ('eq_nci', 'Non-controlling interest carrying amount', NUM0),
                   ('eq_parent_jun25', 'Total AMOC equity at 30-Jun-2025', NUM0),
                   ('dep_h2_25', 'Depreciation and RoU amortisation, 6M to Dec-2025', NUM0),
                   ('dep_q1_26', 'Depreciation and RoU amortisation, 3M to Mar-2026', NUM0),
                   ('capex_h2_25', 'CASH capex, 6M to Dec-2025', NUM0),
                   ('capex_q1_26', 'CASH capex, 3M to Mar-2026', NUM0),
                   ('credint_h2_25', 'Credit interest earned, 6M to Dec-2025', NUM0),
                   ('credint_q1_26', 'Credit interest earned, 3M to Mar-2026', NUM0),
                   ('div_h2_25', 'Cash dividends PAID, 6M to Dec-2025', NUM0),
                   ('div_q1_26', 'Cash dividends PAID, 3M to Mar-2026', NUM0),
                   ('nci_share', 'Minority share of group profit — DISCLOSED', PCT3),
                   ('alexpet_stake', 'Alexandria Petroleum Company shareholding', PCT),
                   ('egpc_sales', 'Sales to EGPC, 6M to Dec-2025', NUM0)):
    drv(k, lab, 'EGP' if f_ == NUM0 else '%', f_)
r += 1
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'NOTE 14-A — THE AUDITED PRODUCT TABLE', BLACK, bold=True); r += 1
hdr(wsA, r, ['Product line', 'Unit', 'Tonnes', 'Value (EGP)', 'Layer', 'Source']); r += 1
PTR = {}
for k in LINES:
    put(wsA, f'A{r}', LBL[k], BLACK)
    put(wsA, f'B{r}', 'tonnes', BLACK)
    put(wsA, f'C{r}', IN['prod_t'][k], BLUE, NUM3)
    put(wsA, f'D{r}', IN['prod_v'][k], BLUE, NUM0)
    put(wsA, f'E{r}', 'Company', BLACK)
    c = put(wsA, f'F{r}', 'Note 14-A, audited transition period 1-Jul-2025 to 31-Dec-2025', BLACK)
    c.font = Font(size=8, color='6E7B77')
    PTR[k] = r
    r += 1
PT_FIRST, PT_LAST = PTR[LINES[0]], PTR[LINES[-1]]
put(wsA, f'A{r}', 'TOTAL (must equal net sales for the period)', BLACK, bold=True)
putf(wsA, f'C{r}', f"=SUM(C{PT_FIRST}:C{PT_LAST})", U['tot_t'], NUM3, bold=True)
putf(wsA, f'D{r}', f"=SUM(D{PT_FIRST}:D{PT_LAST})", U['tot_v'] * 1e6, NUM0, bold=True)
PT_TOT = r; r += 2
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'FORECAST DRIVERS — the only free operating parameters left', BLACK,
    bold=True); r += 1
hdr(wsA, r, ['Driver'] + [y.replace('E', '') for y in YRS], 1); r += 1
PATHR = {}


def path(key, label, vals, fmt):
    global r
    put(wsA, f'A{r}', label, BLACK)
    for i, v in enumerate(vals):
        put(wsA, f'{get_column_letter(2+i)}{r}', v, BLUE, fmt)
    PATHR[key] = r
    r += 1


def P(key, i):
    return f"Assumptions!${get_column_letter(2+i)}${PATHR[key]}"


for k in LINES:
    path('vg_' + k, f'Volume growth — {LBL[k]}', IN['line_vol_growth'][k], PCT2)
path('pg', 'Realised price growth per tonne, all lines', IN['line_price_growth'], PCT2)
path('infl', 'Egyptian inflation factor (pound-denominated costs)',
     IN['fixed_cost_infl'][3:8], '0.000')
path('kd', 'Cost of debt path', IN['kd_path'], PCT2)
path('cash_yield', 'Yield on cash path', IN['cash_yield_path'], PCT2)
path('fx', 'USD/EGP average rate path', IN['fx_path'], NUM1)
r += 1
band(wsA, r, 6, CREAM)
put(wsA, f'A{r}', 'COST OF CAPITAL AND TERMINAL', BLACK, bold=True); r += 1
for k, lab, f_ in (('raw_pass', 'Raw-material pass-through factor', '0.000'),
                   ('rf', 'Risk-free rate — 10y EGP government bond', PCT2),
                   ('sov_spread_cds', 'Sovereign default spread (CDS basis)', PCT2),
                   ('erp_cds', 'Equity risk premium (CDS basis)', PCT2),
                   ('sov_spread_rating', 'Sovereign default spread (rating basis)', PCT2),
                   ('erp_rating', 'Equity risk premium (rating basis)', PCT2),
                   ('beta', 'Beta — own-stock regression', '0.000'),
                   ('kd', 'Cost of debt', PCT2),
                   ('cash_yield', 'Yield on cash', PCT2),
                   ('rf_term', 'Terminal risk-free rate', PCT2),
                   ('erp_term', 'Terminal equity risk premium', PCT2),
                   ('kd_term', 'Terminal cost of debt', PCT2),
                   ('wd_term', 'Terminal debt weight', PCT),
                   ('g_term', 'Terminal growth rate', PCT),
                   ('tax_stat', 'Statutory corporate tax rate', PCT),
                   ('egypt_nominal_growth', 'Egyptian long-run nominal growth', PCT),
                   ('ev_ebitda_just', 'Justified EV / EBITDA', MULT),
                   ('pe_just', 'Justified price / earnings', MULT),
                   ('e1_pe', 'Expert 1 justified price / earnings', MULT),
                   ('roe_sust', 'Sustainable return on equity', PCT)):
    drv(k, lab, '%' if 'PCT' in str(f_) or f_ == PCT else 'x', f_)
r += 1
band(wsA, r, 6, CREAM); put(wsA, f'A{r}', 'LENS WEIGHTS', BLACK, bold=True); r += 1
LW = {}
for k, lab in (('dcf', 'Discounted cash flow'), ('relative', 'Relative multiples'),
               ('normalized', 'Normalised earnings power'),
               ('book', 'Book value and sustainable return')):
    put(wsA, f'A{r}', lab, BLACK)
    put(wsA, f'C{r}', IN['lens_weights'][k], BLUE, PCT)
    LW[k] = r; r += 1


def LWR(k):
    return f"Assumptions!$C${LW[k]}"


# ============ PRODUCT AND COST — the audited build, live =====================
wsP = sheet('Product and Cost')
title(wsP, 'Revenue and cost of sales, built off the audited notes',
      'Note 14-A gives eight product lines with tonnes AND value. Note 15-A gives the cost stack '
      'in five components. Nothing on this sheet is reconstructed: the realisation per tonne is '
      'one disclosed number divided by another, and the cost composition is as filed.', 8)
r = 4
band(wsP, r, 8); put(wsP, f'A{r}', 'NOTE 14-A — AS FILED, 6 MONTHS TO 31-DEC-2025', BLACK,
                     bold=True); r += 1
hdr(wsP, r, ['Product line', 'Tonnes', 'Value (EGP)', 'EGP / tonne', 'Share of tonnage',
             'Share of value']); r += 1
PROW = {}
for k in LINES:
    put(wsP, f'A{r}', LBL[k], BLACK)
    putf(wsP, f'B{r}', f"=Assumptions!$C${PTR[k]}", IN['prod_t'][k], NUM3, green=True)
    putf(wsP, f'C{r}', f"=Assumptions!$D${PTR[k]}", IN['prod_v'][k], NUM0, green=True)
    putf(wsP, f'D{r}', f"=C{r}/B{r}", U['px'][k], NUM0)
    PROW[k] = r; r += 1
P_FIRST, P_LAST = PROW[LINES[0]], PROW[LINES[-1]]
put(wsP, f'A{r}', 'TOTAL', BLACK, bold=True)
putf(wsP, f'B{r}', f"=SUM(B{P_FIRST}:B{P_LAST})", U['tot_t'], NUM3, bold=True)
putf(wsP, f'C{r}', f"=SUM(C{P_FIRST}:C{P_LAST})", U['tot_v'] * 1e6, NUM0, bold=True)
putf(wsP, f'D{r}', f"=C{r}/B{r}", U['tot_v'] * 1e6 / U['tot_t'], NUM0, bold=True)
PTOTR = r
for k in LINES:
    putf(wsP, f'E{PROW[k]}', f"=B{PROW[k]}/B${PTOTR}", U['mix_t'][k], PCT2)
    putf(wsP, f'F{PROW[k]}', f"=C{PROW[k]}/C${PTOTR}", U['mix_v'][k], PCT2)
r += 1
put(wsP, f'A{r}', 'Specialty slate (oils and wax)', BLACK, bold=True)
putf(wsP, f'E{r}', f"=E{PROW['oils']}+E{PROW['wax']}", U['spec_share_t'], PCT2, bold=True)
putf(wsP, f'F{r}', f"=F{PROW['oils']}+F{PROW['wax']}", U['spec_share_v'], PCT2, bold=True)
r += 2
note(wsP, r, 'Every realisation in column D is one disclosed number divided by another. The '
             'previous edition of this model used a THREE-line table obtained from a reviewer '
             'rather than from the filing, and had to reconstruct realisations through a crude '
             'parity and a solved crack multiple. None of that is needed: the note states tonnes '
             'and value for eight lines, and fuel oil (mix) alone is 57% of the tonnage, which '
             'the old three-line build had merged into a single slate with gas oil, naphtha and '
             'LPG.', 8)
r += 2
band(wsP, r, 8); put(wsP, f'A{r}', 'NOTE 15-A — THE AUDITED COST STACK', BLACK, bold=True); r += 1
hdr(wsP, r, ['Component', '6M to Dec-2025', '6M to Dec-2024', 'Share of cost of sales',
             'Share of net sales']); r += 1
CROW = {}
CKEYS = [('raw', 'Raw materials', 'cos_raw', 'cos_raw_24'),
         ('salaries', 'Salaries', 'cos_salaries', 'cos_salaries_24'),
         ('other', 'Other — natural gas, electricity, water, spares, EPROM contract',
          'cos_other', 'cos_other_24'),
         ('support', 'Supporting materials (chemicals and additives)', 'cos_support',
          'cos_support_24'),
         ('dep', 'Depreciation', 'cos_dep', 'cos_dep_24')]
for key, lab, a25, a24 in CKEYS:
    put(wsP, f'A{r}', lab, BLACK)
    putf(wsP, f'B{r}', f"={A(a25)}", IN[a25], NUM0, green=True)
    putf(wsP, f'C{r}', f"={A(a24)}", IN[a24], NUM0, green=True)
    CROW[key] = r; r += 1
C_FIRST, C_LAST = CROW['raw'], CROW['dep']
put(wsP, f'A{r}', 'COST OF SALES', BLACK, bold=True)
putf(wsP, f'B{r}', f"=SUM(B{C_FIRST}:B{C_LAST})", IN['cogs_h2_25'], NUM0, bold=True)
putf(wsP, f'C{r}', f"=SUM(C{C_FIRST}:C{C_LAST})", IN['cogs_h2_24'], NUM0, bold=True)
CTOT = r; r += 1
put(wsP, f'A{r}', 'Check: note 15-A foots to the filed cost of sales (must be zero)', BLACK)
putf(wsP, f'B{r}', f"=B{CTOT}-{A('cogs_h2_25')}", 0.0, '0.00')
putf(wsP, f'C{r}', f"=C{CTOT}-{A('cogs_h2_24')}", 0.0, '0.00')
r += 1
for key, _, _, _ in CKEYS:
    putf(wsP, f'D{CROW[key]}', f"=B{CROW[key]}/B${CTOT}", U['cost_share'][key], PCT2)
    putf(wsP, f'E{CROW[key]}', f"=B{CROW[key]}/{A('rev_h2_25')}",
         IN[dict((k, a) for k, _, a, _ in CKEYS)[key]] / IN['rev_h2_25'], PCT2)
put(wsP, f'A{r}', 'GROSS PROFIT as filed', BLACK, bold=True)
putf(wsP, f'B{r}', f"={A('rev_h2_25')}-B{CTOT}", IN['rev_h2_25'] - IN['cogs_h2_25'], NUM0,
     bold=True)
putf(wsP, f'C{r}', f"={A('rev_h2_24')}-C{CTOT}", IN['rev_h2_24'] - IN['cogs_h2_24'], NUM0,
     bold=True)
GPR = r; r += 1
put(wsP, f'A{r}', 'GROSS MARGIN as filed', BLACK, bold=True)
putf(wsP, f'B{r}', f"=B{GPR}/{A('rev_h2_25')}",
     (IN['rev_h2_25'] - IN['cogs_h2_25']) / IN['rev_h2_25'], PCT2, bold=True)
putf(wsP, f'C{r}', f"=C{GPR}/{A('rev_h2_24')}",
     (IN['rev_h2_24'] - IN['cogs_h2_24']) / IN['rev_h2_24'], PCT2, bold=True)
r += 2
note(wsP, r, 'Raw materials are 90.7% of cost of sales and 85.1% of net sales. That single line '
             'is what this company is: a pass-through processor whose value sits in the spread '
             'between what feedstock costs and what the slate fetches, and in the tonnage that '
             'spread is earned on. The previous edition BUILT this stack from house estimates of '
             'yields, energy intensity and a solved feedstock differential; it carried no '
             'salaries line inside cost of sales at all, and put chemicals at roughly five times '
             'the disclosed figure.', 8)
r += 2

band(wsP, r, 8); put(wsP, f'A{r}', 'THE BASE YEAR — nine audited months, annualised',
                     BLACK, bold=True); r += 1
hdr(wsP, r, ['', '6M to Dec-2025', '3M to Mar-2026', 'Nine months', 'Annualised (x 4/3)']); r += 1
put(wsP, f'A{r}', 'Net sales', BLACK)
putf(wsP, f'B{r}', f"={A('rev_h2_25')}", IN['rev_h2_25'], NUM0, green=True)
putf(wsP, f'C{r}', f"={A('rev_q1_26')}", IN['rev_q1_26'], NUM0, green=True)
putf(wsP, f'D{r}', f"=B{r}+C{r}", AU['rev9'] * 1e6, NUM0)
putf(wsP, f'E{r}', f"=D{r}*4/3", AU['base_rev'] * 1e6, NUM0, bold=True)
BREV = r; r += 1
put(wsP, f'A{r}', 'Cost of sales', BLACK)
putf(wsP, f'B{r}', f"={A('cogs_h2_25')}", IN['cogs_h2_25'], NUM0, green=True)
putf(wsP, f'C{r}', f"={A('cogs_q1_26')}", IN['cogs_q1_26'], NUM0, green=True)
putf(wsP, f'D{r}', f"=B{r}+C{r}", (IN['cogs_h2_25'] + IN['cogs_q1_26']), NUM0)
putf(wsP, f'E{r}', f"=D{r}*4/3", (IN['cogs_h2_25'] + IN['cogs_q1_26']) * 4 / 3, NUM0)
BCOGS = r; r += 1
put(wsP, f'A{r}', 'GROSS PROFIT', BLACK, bold=True)
putf(wsP, f'D{r}', f"=D{BREV}-D{BCOGS}", AU['gp9'] * 1e6, NUM0, bold=True)
putf(wsP, f'E{r}', f"=E{BREV}-E{BCOGS}", AU['gp9'] * 1e6 * 4 / 3, NUM0, bold=True)
BGP = r; r += 1
put(wsP, f'A{r}', 'GROSS MARGIN — the base-year margin, as filed', BLACK, bold=True)
putf(wsP, f'D{r}', f"=D{BGP}/D{BREV}", AU['base_gm'], PCT3, bold=True)
BGM = r; r += 1
put(wsP, f'A{r}', 'Operating expense (G&A + marketing + other)', BLACK)
putf(wsP, f'B{r}', f"={A('ga_h2_25')}+{A('mkt_h2_25')}+{A('othexp_h2_25')}",
     IN['ga_h2_25'] + IN['mkt_h2_25'] + IN['othexp_h2_25'], NUM0)
putf(wsP, f'C{r}', f"={A('ga_q1_26')}+{A('mkt_q1_26')}+{A('othexp_q1_26')}",
     IN['ga_q1_26'] + IN['mkt_q1_26'] + IN['othexp_q1_26'], NUM0)
putf(wsP, f'D{r}', f"=B{r}+C{r}", AU['opex_ann'] * 1e6 * 3 / 4, NUM0)
putf(wsP, f'E{r}', f"=D{r}*4/3", AU['opex_ann'] * 1e6, NUM0, bold=True)
BOPX = r; r += 1
put(wsP, f'A{r}', 'Depreciation and right-of-use amortisation', BLACK)
putf(wsP, f'B{r}', f"={A('dep_h2_25')}", IN['dep_h2_25'], NUM0, green=True)
putf(wsP, f'C{r}', f"={A('dep_q1_26')}", IN['dep_q1_26'], NUM0, green=True)
putf(wsP, f'D{r}', f"=B{r}+C{r}", AU['dep_ann'] * 1e6 * 3 / 4, NUM0)
putf(wsP, f'E{r}', f"=D{r}*4/3", AU['dep_ann'] * 1e6, NUM0, bold=True)
BDEP = r; r += 1
put(wsP, f'A{r}', 'CASH capital expenditure', BLACK)
putf(wsP, f'B{r}', f"={A('capex_h2_25')}", IN['capex_h2_25'], NUM0, green=True)
putf(wsP, f'C{r}', f"={A('capex_q1_26')}", IN['capex_q1_26'], NUM0, green=True)
putf(wsP, f'D{r}', f"=B{r}+C{r}", AU['capex_ann'] * 1e6 * 3 / 4, NUM0)
putf(wsP, f'E{r}', f"=D{r}*4/3", AU['capex_ann'] * 1e6, NUM0, bold=True)
BCAP = r; r += 1
put(wsP, f'A{r}', 'Capital expenditure as a multiple of depreciation', BLACK, bold=True)
putf(wsP, f'E{r}', f"=E{BCAP}/E{BDEP}", AU['capex_ann'] / AU['dep_ann'], '0.00"x"', bold=True)
r += 1
put(wsP, f'A{r}', 'Effective tax rate (both filed periods)', BLACK)
putf(wsP, f'E{r}', f"=({A('tax_h2_25')}-{A('dtax_h2_25')}+{A('tax_q1_26')}-{A('dtax_q1_26')})"
                   f"/({A('pat_h2_25')}+{A('pat_q1_26')}+{A('tax_h2_25')}-{A('dtax_h2_25')}"
                   f"+{A('tax_q1_26')}-{A('dtax_q1_26')})", AU['tax_eff'], PCT2, bold=True)
BTAX = r; r += 2
note(wsP, r, 'There is no clean audited twelve-month period: the year-end moved from 30 June to '
             '31 December and the April-to-June 2025 quarter is not separately filed. The base '
             'year is therefore the nine CONTIGUOUS audited months, annualised by four thirds. '
             'That scaling is the only step between the filings and the base year and it is on '
             'the face of this sheet. Note the capital-expenditure multiple: cash capex is '
             'running BELOW the depreciation charge, so the plant is being run rather than '
             'renewed. That single number does more to the valuation than any assumption in it.',
     8)
ANCH['prod'] = dict(rows=PROW, tot=PTOTR, cost=CROW, ctot=CTOT, gp=GPR, gm=BGM,
                    brev=BREV, bcogs=BCOGS, bgp=BGP, bopx=BOPX, bdep=BDEP, bcap=BCAP, btax=BTAX)


# ============ FORECAST =======================================================
wsF = sheet('Forecast')
title(wsF, 'The forecast — volume by audited line, cost by audited composition',
      'The only free operating parameters are volume growth per line and the growth in the '
      'realised price per tonne. The cost stack keeps its filed composition: raw materials and '
      'supporting materials are pass-through; salaries and the other line are pound-denominated '
      'and inflate; depreciation is the asset-register charge.', 7)
FC = ['B', 'C', 'D', 'E', 'F']
r = 4
band(wsF, r, 7); put(wsF, f'A{r}', 'VOLUME BY LINE (tonnes, annualised)', BLACK, bold=True); r += 1
hdr(wsF, r, [''] + [y.replace('E', '') for y in YRS]); r += 1
FV = {}
for k in LINES:
    put(wsF, f'A{r}', LBL[k], BLACK)
    prev = f"'Product and Cost'!B{PROW[k]}*2"
    for i, c in enumerate(FC):
        putf(wsF, f'{c}{r}', f"={prev}*(1+{P('vg_'+k, i)})",
             U['lines_vol'][k][i] * 1e6, NUM0)
        prev = f'{c}{r}'
    FV[k] = r; r += 1
put(wsF, f'A{r}', 'TOTAL TONNES', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', "=" + "+".join(f"{c}{FV[k]}" for k in LINES),
         U['vol'][i] * 1e6, NUM0, bold=True)
FVT = r; r += 1
put(wsF, f'A{r}', 'Realised price index (base year = 1.00)', BLACK)
prev = None
for i, c in enumerate(FC):
    f_ = f"=(1+{P('pg', i)})" if i == 0 else f"={FC[i-1]}{r}*(1+{P('pg', i)})"
    _v = 1.0
    for j in range(i + 1):
        _v *= (1 + IN['line_price_growth'][j])
    putf(wsF, f'{c}{r}', f_, _v, '0.0000')
FPX = r; r += 1
band(wsF, r, 7); put(wsF, f'A{r}', 'REVENUE BY LINE (EGP)', BLACK, bold=True); r += 1
FR = {}
for k in LINES:
    put(wsF, f'A{r}', LBL[k], BLACK)
    for i, c in enumerate(FC):
        putf(wsF, f'{c}{r}', f"={c}{FV[k]}*'Product and Cost'!$D${PROW[k]}*{c}${FPX}",
             U['lines_rev'][k][i] * 1e6, NUM0)
    FR[k] = r; r += 1
put(wsF, f'A{r}', 'TOTAL REVENUE', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', "=" + "+".join(f"{c}{FR[k]}" for k in LINES), F['rev'][i] * 1e6, NUM0,
         bold=True)
FREV = r; r += 1
put(wsF, f'A{r}', 'Specialty share of revenue', BLACK)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"=({c}{FR['oils']}+{c}{FR['wax']})/{c}{FREV}",
         (U['lines_rev']['oils'][i] + U['lines_rev']['wax'][i]) / F['rev'][i], PCT2)
r += 1
put(wsF, f'A{r}', 'Volume index vs base year', BLACK)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"={c}{FVT}/('Product and Cost'!B${PTOTR}*2)",
         U['vol'][i] * 1e6 / (U['tot_t'] * 2), '0.0000')
FVI = r; r += 1
put(wsF, f'A{r}', 'Inflation index (pound-denominated costs)', BLACK)
for i, c in enumerate(FC):
    f_ = f"={P('infl', 0)}" if i == 0 else f"={FC[i-1]}{r}*{P('infl', i)}"
    _v = 1.0
    for j in range(i + 1):
        _v *= IN['fixed_cost_infl'][3 + j]
    putf(wsF, f'{c}{r}', f_, _v, '0.0000')
FINF = r; r += 1
band(wsF, r, 7); put(wsF, f'A{r}', 'COST OF SALES — the audited composition, rolled forward',
                     BLACK, bold=True); r += 1
_cos0 = (IN['cogs_h2_25'] + IN['cogs_q1_26']) * 4 / 3
CSH = {}
for key, lab in (('raw', 'Raw materials (pass-through: volume x price)'),
                 ('support', 'Supporting materials (pass-through)'),
                 ('salaries', 'Salaries (pound-denominated: inflates)'),
                 ('other', 'Other — gas, power, water, spares, EPROM (inflates)'),
                 ('dep', 'Depreciation (asset register)')):
    put(wsF, f'A{r}', lab, BLACK)
    for i, c in enumerate(FC):
        base = f"(('Product and Cost'!D${BREV}-'Product and Cost'!D${BGP})*4/3)"
        shr = f"'Product and Cost'!D${CROW[key]}"
        if key in ('raw', 'support'):
            f_ = f"={base}*{shr}*{c}{FVI}*{c}{FPX}" + (f"*{A('raw_pass')}" if key == 'raw' else "")
            _v = _cos0 * U['cost_share'][key] * (U['vol'][i] * 1e6 / (U['tot_t'] * 2))
            _p = 1.0
            for j in range(i + 1):
                _p *= (1 + IN['line_price_growth'][j])
            _v *= _p
        elif key == 'dep':
            f_ = f"={base}*{shr}"
            _v = _cos0 * U['cost_share'][key]
        else:
            f_ = f"={base}*{shr}*{c}{FINF}"
            _v = _cos0 * U['cost_share'][key]
            for j in range(i + 1):
                _v *= IN['fixed_cost_infl'][3 + j]
        putf(wsF, f'{c}{r}', f_, _v, NUM0)
    CSH[key] = r; r += 1
put(wsF, f'A{r}', 'COST OF SALES', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', "=" + "+".join(f"{c}{CSH[k]}" for k in CSH),
         U['cogs'][i] * 1e6, NUM0, bold=True)
FCOGS = r; r += 1
put(wsF, f'A{r}', 'GROSS PROFIT', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"={c}{FREV}-{c}{FCOGS}", F['gp'][i] * 1e6, NUM0, bold=True)
FGP = r; r += 1
put(wsF, f'A{r}', 'GROSS MARGIN — an OUTPUT of the audited composition', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"={c}{FGP}/{c}{FREV}", F['gm'][i], PCT2, bold=True)
FGM = r; r += 1
put(wsF, f'A{r}', 'Operating expense (inflates)', BLACK)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"='Product and Cost'!E${BOPX}*{c}{FINF}", F['opex'][i] * 1e6, NUM0)
FOPX = r; r += 1
put(wsF, f'A{r}', 'EBITDA', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"={c}{FGP}-{c}{FOPX}+'Product and Cost'!E${BDEP}",
         F['ebitda'][i] * 1e6, NUM0, bold=True)
FEBITDA = r; r += 1
put(wsF, f'A{r}', 'less depreciation and amortisation', BLACK)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"='Product and Cost'!E${BDEP}", F['dna'][i] * 1e6, NUM0)
FDNA = r; r += 1
put(wsF, f'A{r}', 'EBIT', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsF, f'{c}{r}', f"={c}{FEBITDA}-{c}{FDNA}", F['ebit'][i] * 1e6, NUM0, bold=True)
FEBIT = r; r += 1
ANCH['fcst'] = dict(vol=FV, volt=FVT, rev=FR, revt=FREV, cogs=FCOGS, gp=FGP, gm=FGM,
                    opex=FOPX, ebitda=FEBITDA, dna=FDNA, ebit=FEBIT, inf=FINF, px=FPX, vi=FVI)


# ============ DCF ============================================================
wsD = sheet('DCF')
title(wsD, 'Discounted cash flow — the waterfall and the terminal block',
      'The cost of capital slides from an explicit-window rate to a terminal rate, each year '
      'discounted at its own forward rate. Capital expenditure is held at the ACTUAL cash run '
      'rate from the filings, inflated.', 7)
r = 4
band(wsD, r, 7); put(wsD, f'A{r}', 'COST OF CAPITAL', BLACK, bold=True); r += 1
put(wsD, f'A{r}', 'Risk-free rate less sovereign default spread', BLACK)
putf(wsD, f'B{r}', f"={A('rf')}-{A('sov_spread_cds')}", W['rf_star'], PCT2)
RFS = r; r += 1
put(wsD, f'A{r}', 'COST OF EQUITY, explicit window', BLACK, bold=True)
putf(wsD, f'B{r}', f"=B{RFS}+{A('beta')}*{A('erp_cds')}", W['ke_exp'], PCT2, bold=True)
KEE = r; r += 1
put(wsD, f'A{r}', 'Net debt (negative = net cash)', BLACK)
putf(wsD, f'B{r}', f"={A('debt_lt')}+{A('debt_st')}-{A('cash')}", DCF['nd'] * 1e6, NUM0)
NDR = r; r += 1
put(wsD, f'A{r}', 'Market capitalisation', BLACK)
putf(wsD, f'B{r}', f"={A('spot')}*{A('shares_mn')}*1000000", SPOT * SH * 1e6, NUM0)
MCR = r; r += 1
put(wsD, f'A{r}', 'Debt weight (NEGATIVE — the company is net cash)', BLACK)
putf(wsD, f'B{r}', f"=B{NDR}/(B{NDR}+B{MCR})", W['wd_exp'], PCT2)
WDR = r; r += 1
put(wsD, f'A{r}', 'Equity weight', BLACK)
putf(wsD, f'B{r}', f"=1-B{WDR}", W['we_exp'], PCT2)
WER = r; r += 1
put(wsD, f'A{r}', 'Cost of net debt, after tax', BLACK)
putf(wsD, f'B{r}', f"=({A('kd')}*({A('debt_lt')}+{A('debt_st')})-{A('cash_yield')}*{A('cash')})"
                   f"/(({A('debt_lt')}+{A('debt_st')})-{A('cash')})*(1-'Product and Cost'!E${BTAX})",
     W['k_nd_at'], PCT2)
KND = r; r += 1
put(wsD, f'A{r}', 'WEIGHTED COST OF CAPITAL — explicit window', BLACK, bold=True)
putf(wsD, f'B{r}', f"=B{WER}*B{KEE}+B{WDR}*B{KND}", W['wacc_exp'], PCT2, bold=True)
WEXP = r; r += 1
put(wsD, f'A{r}', 'Terminal cost of equity', BLACK)
putf(wsD, f'B{r}', f"={A('rf_term')}+{A('beta')}*{A('erp_term')}", W['ke_term'], PCT2)
KET = r; r += 1
put(wsD, f'A{r}', 'WEIGHTED COST OF CAPITAL — terminal', BLACK, bold=True)
putf(wsD, f'B{r}', f"=(1-{A('wd_term')})*B{KET}+{A('wd_term')}*{A('kd_term')}"
                   f"*(1-'Product and Cost'!E${BTAX})", W['wacc_term'], PCT2, bold=True)
WTRM = r; r += 2
band(wsD, r, 7); put(wsD, f'A{r}', 'THE GLIDE AND THE WATERFALL', BLACK, bold=True); r += 1
hdr(wsD, r, [''] + [y.replace('E', '') for y in YRS]); r += 1
put(wsD, f'A{r}', 'Cumulative progress along the cost-of-debt path', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"=({P('kd', 0)}-{P('kd', i)})/({P('kd', 0)}-{P('kd', 4)})",
         F['glide_frac'][i], '0.000')
GLR = r; r += 1
put(wsD, f'A{r}', 'Forward cost of capital', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"=$B${WEXP}-($B${WEXP}-$B${WTRM})*{c}{GLR}", F['fwd_wacc'][i], PCT2)
FWR = r; r += 1
put(wsD, f'A{r}', 'Cumulative discount factor', BLACK)
for i, c in enumerate(FC):
    f_ = f"=1/(1+{c}{FWR})" if i == 0 else f"={FC[i-1]}{r}/(1+{c}{FWR})"
    putf(wsD, f'{c}{r}', f_, F['df'][i], '0.0000')
DFR = r; r += 1
put(wsD, f'A{r}', 'EBIT', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"=Forecast!{c}{FEBIT}", F['ebit'][i] * 1e6, NUM0, green=True)
DEBIT = r; r += 1
put(wsD, f'A{r}', 'NOPAT = EBIT x (1 - effective tax)', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"={c}{DEBIT}*(1-'Product and Cost'!E${BTAX})", F['nopat'][i] * 1e6,
         NUM0)
DNOP = r; r += 1
put(wsD, f'A{r}', 'add back depreciation and amortisation', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"=Forecast!{c}{FDNA}", F['dna'][i] * 1e6, NUM0, green=True)
DDNA = r; r += 1
put(wsD, f'A{r}', 'less capital expenditure (ACTUAL run rate, inflated)', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"='Product and Cost'!E${BCAP}*Forecast!{c}{FINF}",
         F['capex'][i] * 1e6, NUM0)
DCAP = r; r += 1
put(wsD, f'A{r}', 'Net working capital', BLACK)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"=Forecast!{c}{FREV}*(({A('inventory')}+{A('recv')}+{A('debtors')}"
                         f"-{A('payables')}-{A('creditors')})/'Product and Cost'!E${BREV})",
         F['nwc'][i] * 1e6, NUM0)
DNWC = r; r += 1
put(wsD, f'A{r}', 'less change in net working capital', BLACK)
for i, c in enumerate(FC):
    prev = (f"({A('inventory')}+{A('recv')}+{A('debtors')}-{A('payables')}-{A('creditors')})"
            if i == 0 else f"{FC[i-1]}{DNWC}")
    putf(wsD, f'{c}{r}', f"={c}{DNWC}-{prev}", F['dnwc'][i] * 1e6, NUM0)
DDNW = r; r += 1
put(wsD, f'A{r}', 'FREE CASH FLOW TO THE FIRM', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"={c}{DNOP}+{c}{DDNA}-{c}{DCAP}-{c}{DDNW}", F['fcff'][i] * 1e6, NUM0,
         bold=True)
DFCFF = r; r += 1
put(wsD, f'A{r}', 'PRESENT VALUE OF FREE CASH FLOW', BLACK, bold=True)
for i, c in enumerate(FC):
    putf(wsD, f'{c}{r}', f"={c}{DFCFF}*{c}{DFR}", F['pv'][i] * 1e6, NUM0, bold=True)
DPV = r; r += 2
band(wsD, r, 7); put(wsD, f'A{r}', 'THE TERMINAL BLOCK', BLACK, bold=True); r += 1
put(wsD, f'A{r}', 'Invested capital at the end of the window', BLACK)
putf(wsD, f'B{r}', f"=F{DNWC}+({A('ppe_net')}+{A('puc')})+SUM(B{DCAP}:F{DCAP})-SUM(B{DDNA}:F{DDNA})",
     (F['nwc'][4] + F['ppe'][4]) * 1e6, NUM0)
DIC = r; r += 1
put(wsD, f'A{r}', 'Terminal return on invested capital', BLACK)
putf(wsD, f'B{r}', f"=F{DNOP}*(1+{A('g_term')})/B{DIC}", DCF['roic_term'], PCT2)
DROIC = r; r += 1
put(wsD, f'A{r}', 'Required reinvestment rate = growth / return', BLACK)
putf(wsD, f'B{r}', f"={A('g_term')}/B{DROIC}", DCF['rr_term'], PCT2)
DRR = r; r += 1
put(wsD, f'A{r}', 'TERMINAL VALUE', BLACK, bold=True)
putf(wsD, f'B{r}', f"=F{DNOP}*(1+{A('g_term')})*(1-B{DRR})/(B{WTRM}-{A('g_term')})",
     DCF['tv'] * 1e6, NUM0, bold=True)
DTV = r; r += 1
put(wsD, f'A{r}', 'Present value of the terminal block', BLACK, bold=True)
putf(wsD, f'B{r}', f"=B{DTV}*F{DFR}", DCF['pv_tv'] * 1e6, NUM0, bold=True)
DPVTV = r; r += 1
put(wsD, f'A{r}', 'Present value of the explicit window', BLACK)
putf(wsD, f'B{r}', f"=SUM(B{DPV}:F{DPV})", DCF['pv_explicit'] * 1e6, NUM0)
DPVE = r; r += 1
put(wsD, f'A{r}', 'ENTERPRISE VALUE', BLACK, bold=True)
putf(wsD, f'B{r}', f"=B{DPVE}+B{DPVTV}", DCF['ev'] * 1e6, NUM0, bold=True)
DEV = r; r += 1
put(wsD, f'A{r}', 'TERMINAL VALUE AS A SHARE OF ENTERPRISE VALUE', BLACK, bold=True)
putf(wsD, f'B{r}', f"=B{DPVTV}/B{DEV}", DCF['tv_share'], PCT2, bold=True)
DTVS = r; r += 1
ANCH['dcf'] = dict(ev=DEV, tv=DTV, pvtv=DPVTV, pve=DPVE, tvs=DTVS, wexp=WEXP, wtrm=WTRM,
                   fcff=DFCFF, pv=DPV, nd=NDR, roic=DROIC, rr=DRR, keе=KEE)


# ============ EV BRIDGE ======================================================
wsB = sheet('EV Bridge')
title(wsB, 'Enterprise value to equity',
      'Minorities come off the OPERATING enterprise value, before the cash is added. The cash '
      'belongs to the parent and is added in full.', 4)
r = 4
hdr(wsB, r, ['', 'EGP', 'per share']); r += 1
put(wsB, f'A{r}', 'ENTERPRISE VALUE (the operating assets)', BLACK)
putf(wsB, f'B{r}', f"=DCF!$B${DEV}", DCF['ev'] * 1e6, NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/({A('shares_mn')}*1000000)", DCF['ev'] / SH, PX)
BEV = r; r += 1
put(wsB, f'A{r}', 'Minority share of group profit — DISCLOSED (note: 86.45% of the wax subsidiary)',
    BLACK)
putf(wsB, f'B{r}', f"={A('nci_share')}", IN['nci_share'], PCT3, green=True)
BNCS = r; r += 1
put(wsB, f'A{r}', 'less minority interests, ON THE ENTERPRISE VALUE', BLACK)
putf(wsB, f'B{r}', f"=B{BEV}*B{BNCS}", DCF['nci_val'] * 1e6, NUM0)
putf(wsB, f'C{r}', f"=B{r}/({A('shares_mn')}*1000000)", DCF['nci_val'] / SH, PX)
BNCI = r; r += 1
put(wsB, f'A{r}', 'Operating assets attributable to shareholders', BLACK)
putf(wsB, f'B{r}', f"=B{BEV}-B{BNCI}", (DCF['ev'] - DCF['nci_val']) * 1e6, NUM0)
BPRE = r; r += 1
put(wsB, f'A{r}', 'less net debt (negative = net cash ADDED in full)', BLACK)
putf(wsB, f'B{r}', f"=DCF!$B${NDR}", DCF['nd'] * 1e6, NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/({A('shares_mn')}*1000000)", DCF['nd'] / SH, PX)
BND = r; r += 1
put(wsB, f'A{r}', 'EQUITY ATTRIBUTABLE TO SHAREHOLDERS', BLACK, bold=True)
putf(wsB, f'B{r}', f"=B{BPRE}-B{BND}", DCF['eq_attr'] * 1e6, NUM0, bold=True)
BEQ = r; r += 1
put(wsB, f'A{r}', 'FAIR VALUE PER SHARE (EGP)', BLACK, bold=True)
putf(wsB, f'B{r}', f"=B{BEQ}/({A('shares_mn')}*1000000)", DCF['ps'], PX, bold=True)
BPS = r; TOK['BR_PS'] = BPS; r += 1
put(wsB, f'A{r}', 'TERMINAL VALUE AS A SHARE OF ENTERPRISE VALUE', BLACK, bold=True)
putf(wsB, f'B{r}', f"=DCF!$B${DTVS}", DCF['tv_share'], PCT2, bold=True, green=True)
r += 1
put(wsB, f'A{r}', 'Spot price (EGP)', BLACK)
putf(wsB, f'B{r}', f"={A('spot')}", SPOT, PX, green=True)
BSP = r; r += 1
put(wsB, f'A{r}', 'Implied against spot', BLACK)
putf(wsB, f'B{r}', f"=B{BPS}/B{BSP}-1", DCF['ps'] / SPOT - 1, PCT)
r += 2
put(wsB, f'A{r}', 'MEMO — pledged deposits, NOT counted as free cash', BLACK)
putf(wsB, f'B{r}', f"={A('fin_inv')}", IN['fin_inv'], NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/({A('shares_mn')}*1000000)", IN['fin_inv'] / (SH * 1e6), PX)
r += 1
put(wsB, f'A{r}', 'MEMO — tax-disputes and claims provision (note 10-1), NOT deducted here',
    BLACK)
putf(wsB, f'B{r}', f"={A('provisions')}", IN['provisions'], NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/({A('shares_mn')}*1000000)", IN['provisions'] / (SH * 1e6), PX)
r += 2
note(wsB, r, 'Two disclosed balances sit beside this bridge and neither is inside it. The pledged '
             'deposits are real cash but are charged against credit facilities, so they are not '
             'free and are excluded from net debt. The tax-disputes provision is a liability the '
             'company has recognised; it is not deducted from equity value here because the '
             'earnings the model discounts are struck after the tax charge that provision '
             'relates to, but a reader who thinks the exposure is incremental should subtract it '
             'and the per-share figure is given so they can.', 4)
ANCH['bridge'] = dict(ev=BEV, nci=BNCI, nd=BND, eq=BEQ, ps=BPS)


# ============ OTHER LENSES ===================================================
wsL = sheet('Other Lenses')
title(wsL, 'The three cross-check lenses', None, 3)
r = 4
band(wsL, r, 3); put(wsL, f'A{r}', 'LENS 2 — RELATIVE MULTIPLES', BLACK, bold=True); r += 1
put(wsL, f'A{r}', f"{REL['year']} EBITDA", BLACK)
putf(wsL, f'B{r}', f"=Forecast!C{FEBITDA}", REL['ebitda_mid'] * 1e6, NUM0, green=True)
LE = r; r += 1
put(wsL, f'A{r}', 'Justified EV / EBITDA', BLACK)
putf(wsL, f'B{r}', f"={A('ev_ebitda_just')}", IN['ev_ebitda_just'], MULT, green=True)
LM = r; r += 1
put(wsL, f'A{r}', 'Forward enterprise value', BLACK)
putf(wsL, f'B{r}', f"=B{LE}*B{LM}", REL['ev_rel_fwd'] * 1e6, NUM0)
LF = r; r += 1
put(wsL, f'A{r}', 'discounted back at the year-2 factor', BLACK)
putf(wsL, f'B{r}', f"=B{LF}*DCF!C{DFR}", REL['ev_rel_fwd'] * REL['df_rel'] * 1e6, NUM0)
LD = r; r += 1
put(wsL, f'A{r}', 'add interim present value of free cash flow', BLACK)
putf(wsL, f'B{r}', f"=DCF!B{DPV}+DCF!C{DPV}", REL['pv_interim'] * 1e6, NUM0)
LI = r; r += 1
put(wsL, f'A{r}', 'IMPLIED VALUE PER SHARE', BLACK, bold=True)
putf(wsL, f'B{r}', f"=((B{LD}+B{LI})*(1-{A('nci_share')})-DCF!B{NDR})/({A('shares_mn')}*1000000)",
     LN['relative']['base'], PX, bold=True)
LREL = r; r += 2
band(wsL, r, 3); put(wsL, f'A{r}', 'LENS 3 — NORMALISED EARNINGS POWER', BLACK, bold=True); r += 1
put(wsL, f'A{r}', f"{NRM['year']} EBIT", BLACK)
putf(wsL, f'B{r}', f"=Forecast!D{FEBIT}", NRM['ebit'] * 1e6, NUM0, green=True)
NE = r; r += 1
put(wsL, f'A{r}', 'add net finance income', BLACK)
putf(wsL, f'B{r}', f"=({A('credint_h2_25')}+{A('credint_q1_26')})*4/3", NRM['interest'] * 1e6,
     NUM0)
NF = r; r += 1
put(wsL, f'A{r}', 'Attributable normalised earnings', BLACK)
putf(wsL, f'B{r}', f"=(B{NE}+B{NF})*(1-'Product and Cost'!E${BTAX})*(1-{A('nci_share')})",
     NRM['np'] * 1e6, NUM0)
NN = r; r += 1
put(wsL, f'A{r}', 'Normalised earnings per share', BLACK)
putf(wsL, f'B{r}', f"=B{NN}/({A('shares_mn')}*1000000)", NRM['eps'], PX)
NPS = r; r += 1
put(wsL, f'A{r}', 'IMPLIED VALUE PER SHARE', BLACK, bold=True)
putf(wsL, f'B{r}', f"=B{NPS}*{A('pe_just')}", LN['normalized']['base'], PX, bold=True)
LNRM = r; r += 2
band(wsL, r, 3); put(wsL, f'A{r}', 'LENS 4 — BOOK VALUE AND SUSTAINABLE RETURN', BLACK,
                     bold=True); r += 1
put(wsL, f'A{r}', 'Attributable book value per share (AUDITED)', BLACK)
putf(wsL, f'B{r}', f"={A('eq_parent')}/({A('shares_mn')}*1000000)", BK['bvps'], PX)
BV = r; r += 1
put(wsL, f'A{r}', 'Justified price / book = (ROE - g) / (Ke - g)', BLACK)
putf(wsL, f'B{r}', f"=({A('roe_sust')}-{A('g_term')})/(DCF!B{KET}-{A('g_term')})", BK['pb_just'],
     MULT)
PB = r; r += 1
put(wsL, f'A{r}', 'IMPLIED VALUE PER SHARE', BLACK, bold=True)
putf(wsL, f'B{r}', f"=B{BV}*B{PB}", LN['book']['base'], PX, bold=True)
LBK = r; r += 1
ANCH['lenses'] = dict(rel=LREL, norm=LNRM, book=LBK)


# ============ SUMMARY ========================================================
wsS = sheet('Summary')
title(wsS, 'Fair value — four lenses, weighted',
      'Bear and bull are whole-model re-runs and are values, not formulas. Everything else is '
      'live.', 6)
r = 4
hdr(wsS, r, ['Lens', 'Bear (EGP)', 'Base (EGP)', 'Bull (EGP)', 'Weight', 'Contribution']); r += 1
SRW = {}
for key, lab, ref in (('dcf', 'Discounted cash flow (primary)', f"='EV Bridge'!$B${BPS}"),
                      ('relative', 'Relative multiples', f"='Other Lenses'!$B${LREL}"),
                      ('normalized', 'Normalised earnings power', f"='Other Lenses'!$B${LNRM}"),
                      ('book', 'Book value and sustainable return', f"='Other Lenses'!$B${LBK}")):
    put(wsS, f'A{r}', lab, BLACK)
    put(wsS, f'B{r}', LN[key]['bear'], BLUE, PX)
    putf(wsS, f'C{r}', ref, LN[key]['base'], PX, green=True)
    put(wsS, f'D{r}', LN[key]['bull'], BLUE, PX)
    putf(wsS, f'E{r}', f"={LWR(key)}", IN['lens_weights'][key], PCT)
    putf(wsS, f'F{r}', f"=C{r}*E{r}", LN[key]['base'] * IN['lens_weights'][key], PX)
    SRW[key] = r; r += 1
S_FIRST, S_LAST = SRW['dcf'], SRW['book']
put(wsS, f'A{r}', 'WEIGHTED CENTRAL FAIR VALUE', BLACK, bold=True)
putf(wsS, f'B{r}', f"=MIN(B{S_FIRST}:B{S_LAST})", D['span'][0], PX, bold=True)
putf(wsS, f'C{r}', f"=SUM(F{S_FIRST}:F{S_LAST})", D['central'], PX, bold=True)
putf(wsS, f'D{r}', f"=MAX(D{S_FIRST}:D{S_LAST})", D['span'][1], PX, bold=True)
putf(wsS, f'E{r}', f"=SUM(E{S_FIRST}:E{S_LAST})", 1.0, PCT, bold=True)
SC = r; r += 2
put(wsS, f'A{r}', 'Spot price', BLACK)
putf(wsS, f'C{r}', f"={A('spot')}", SPOT, PX, green=True)
SS = r; r += 1
put(wsS, f'A{r}', 'Implied against spot', BLACK, bold=True)
putf(wsS, f'C{r}', f"=C{SC}/C{SS}-1", D['central'] / SPOT - 1, PCT, bold=True)
r += 2
band(wsS, r, 6, CREAM); put(wsS, f'A{r}', 'BASE-YEAR CHECKS — all live off the filings', BLACK,
                            bold=True); r += 1
for lab, f_, v, fmt in (
        ('Base-year revenue (9 audited months x 4/3)', f"='Product and Cost'!E${BREV}",
         AU['base_rev'] * 1e6, NUM0),
        ('Base-year gross margin, as filed', f"='Product and Cost'!D${BGM}", AU['base_gm'], PCT3),
        ('Effective tax rate, computed', f"='Product and Cost'!E${BTAX}", AU['tax_eff'], PCT2),
        ('Capital expenditure / depreciation', f"='Product and Cost'!E${BCAP}/'Product and Cost'!E${BDEP}",
         AU['capex_ann'] / AU['dep_ann'], '0.00"x"'),
        ('Net cash per share', f"=-DCF!B{NDR}/({A('shares_mn')}*1000000)", -DCF['nd'] / SH, PX),
        ('Terminal value as a share of enterprise value', f"=DCF!B{DTVS}", DCF['tv_share'], PCT2),
        ('Minority share of group profit, DISCLOSED', f"={A('nci_share')}", IN['nci_share'],
         PCT3)):
    put(wsS, f'A{r}', lab, BLACK)
    putf(wsS, f'C{r}', f_, v, fmt)
    r += 1
ANCH['sum'] = dict(rows=SRW, central=SC, spot=SS)


# ============ STATEMENTS =====================================================
wsI = sheet('Statements')
title(wsI, 'The filed periods, and the forecast beside them',
      'The four historical columns are AS FILED. No line in them is reconstructed, driven off a '
      'days assumption, or rolled back through profit and dividends.', 10)
HC = ['B', 'C', 'D', 'E']
HL = ['6M Dec-2024', '3M Mar-2025', '6M Dec-2025', '3M Mar-2026']
r = 4
band(wsI, r, 10); put(wsI, f'A{r}', 'PROFIT OR LOSS — AS FILED', BLACK, bold=True); r += 1
hdr(wsI, r, [''] + HL + [y.replace('E', '') for y in YRS]); r += 1
FCOL = ['F', 'G', 'H', 'I', 'J']
put(wsI, f'A{r}', 'Net sales', BLACK)
for c, k in zip(HC, ('rev_h2_24', 'rev_q1_25', 'rev_h2_25', 'rev_q1_26')):
    putf(wsI, f'{c}{r}', f"={A(k)}", IN[k], NUM0, green=True)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"=Forecast!{FC[i]}{FREV}", F['rev'][i] * 1e6, NUM0, green=True)
IREV = r; r += 1
put(wsI, f'A{r}', 'Cost of sales', BLACK)
for c, k in zip(HC, ('cogs_h2_24', 'cogs_q1_25', 'cogs_h2_25', 'cogs_q1_26')):
    putf(wsI, f'{c}{r}', f"=-{A(k)}", -IN[k], NUM0, green=True)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"=-Forecast!{FC[i]}{FCOGS}", -U['cogs'][i] * 1e6, NUM0)
ICOGS = r; r += 1
put(wsI, f'A{r}', 'GROSS PROFIT', BLACK, bold=True)
for c in HC + FCOL:
    _v = None
    putf(wsI, f'{c}{r}', f"={c}{IREV}+{c}{ICOGS}", 0.0, NUM0, bold=True)
IGP = r
for c, k1, k2 in zip(HC, ('rev_h2_24', 'rev_q1_25', 'rev_h2_25', 'rev_q1_26'),
                     ('cogs_h2_24', 'cogs_q1_25', 'cogs_h2_25', 'cogs_q1_26')):
    EXPECT['Statements'][f'{c}{r}'] = IN[k1] - IN[k2]
for i, c in enumerate(FCOL):
    EXPECT['Statements'][f'{c}{r}'] = F['gp'][i] * 1e6
r += 1
put(wsI, f'A{r}', 'GROSS MARGIN', BLACK, bold=True)
for c, k1, k2 in zip(HC, ('rev_h2_24', 'rev_q1_25', 'rev_h2_25', 'rev_q1_26'),
                     ('cogs_h2_24', 'cogs_q1_25', 'cogs_h2_25', 'cogs_q1_26')):
    putf(wsI, f'{c}{r}', f"={c}{IGP}/{c}{IREV}", (IN[k1] - IN[k2]) / IN[k1], PCT2, bold=True)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"={c}{IGP}/{c}{IREV}", F['gm'][i], PCT2, bold=True)
IGM = r; r += 1
put(wsI, f'A{r}', 'Operating expense (G&A, marketing, other)', BLACK)
putf(wsI, f'D{r}', f"=-({A('ga_h2_25')}+{A('mkt_h2_25')}+{A('othexp_h2_25')})",
     -(IN['ga_h2_25'] + IN['mkt_h2_25'] + IN['othexp_h2_25']), NUM0)
putf(wsI, f'E{r}', f"=-({A('ga_q1_26')}+{A('mkt_q1_26')}+{A('othexp_q1_26')})",
     -(IN['ga_q1_26'] + IN['mkt_q1_26'] + IN['othexp_q1_26']), NUM0)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"=-Forecast!{FC[i]}{FOPX}", -F['opex'][i] * 1e6, NUM0)
IOPX = r; r += 1
put(wsI, f'A{r}', 'EBITDA', BLACK, bold=True)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"=Forecast!{FC[i]}{FEBITDA}", F['ebitda'][i] * 1e6, NUM0, bold=True,
         green=True)
IEBITDA = r; r += 1
put(wsI, f'A{r}', 'EBIT', BLACK, bold=True)
for i, c in enumerate(FCOL):
    putf(wsI, f'{c}{r}', f"=Forecast!{FC[i]}{FEBIT}", F['ebit'][i] * 1e6, NUM0, bold=True,
         green=True)
IEBIT = r; r += 1
put(wsI, f'A{r}', 'Net profit after tax — AS FILED', BLACK, bold=True)
for c, k in zip(HC, ('pat_h2_24', None, 'pat_h2_25', 'pat_q1_26')):
    if k:
        putf(wsI, f'{c}{r}', f"={A(k)}", IN[k], NUM0, bold=True, green=True)
IPAT = r; r += 1
put(wsI, f'A{r}', 'Minority interest — AS FILED', BLACK)
for c, k in zip(('D', 'E'), ('nci_h2_25', 'nci_q1_26')):
    putf(wsI, f'{c}{r}', f"={A(k)}", IN[k], NUM0, green=True)
INCI = r; r += 1
put(wsI, f'A{r}', 'Attributable to AMOC shareholders', BLACK, bold=True)
for c in ('D', 'E'):
    _k = 'pat_h2_25' if c == 'D' else 'pat_q1_26'
    _n = 'nci_h2_25' if c == 'D' else 'nci_q1_26'
    putf(wsI, f'{c}{r}', f"={c}{IPAT}-{c}{INCI}", IN[_k] - IN[_n], NUM0, bold=True)
r += 2
band(wsI, r, 10); put(wsI, f'A{r}', 'BALANCE SHEET AT 31-DEC-2025 — AS FILED', BLACK,
                      bold=True); r += 1
BSL = [('Fixed assets, net (note 6)', 'ppe_net'), ('Projects under construction (note 7)', 'puc'),
       ('Inventory, net (note 9-A)', 'inventory'), ('Accounts receivable, net (note 9-B)', 'recv'),
       ('Debtors and other debit balances', 'debtors'),
       ('Cash at banks and on hand (note 9-E)', 'cash'),
       ('Pledged deposits (other financial investments)', 'fin_inv'),
       ('Accounts and notes payable (note 10-3)', 'payables'),
       ('Creditors and other credit balances (note 11)', 'creditors'),
       ('Provisions (note 10-1)', 'provisions'), ('Long-term loans', 'debt_lt'),
       ('Short-term loans and facilities', 'debt_st'), ('Total assets', 'assets_snap'),
       ('Total liabilities', 'liab_snap'), ('Total AMOC equity', 'eq_parent'),
       ('Non-controlling interest', 'eq_nci')]
for lab, k in BSL:
    put(wsI, f'A{r}', lab, BLACK)
    putf(wsI, f'B{r}', f"={A(k)}", IN[k], NUM0, green=True)
    r += 1
put(wsI, f'A{r}', 'CHECK: assets less liabilities less total equity (must be zero)', BLACK,
    bold=True)
putf(wsI, f'B{r}', f"={A('assets_snap')}-{A('liab_snap')}-{A('eq_parent')}-{A('eq_nci')}", 0.0,
     '0.00', bold=True)
r += 1
put(wsI, f'A{r}', 'NET WORKING CAPITAL', BLACK, bold=True)
putf(wsI, f'B{r}', f"={A('inventory')}+{A('recv')}+{A('debtors')}-{A('payables')}-{A('creditors')}",
     AU['nwc'] * 1e6, NUM0, bold=True)
INWC = r; r += 1
put(wsI, f'A{r}', 'INVESTED CAPITAL', BLACK, bold=True)
putf(wsI, f'B{r}', f"=B{INWC}+{A('ppe_net')}+{A('puc')}", AU['ic'] * 1e6, NUM0, bold=True)
r += 1
put(wsI, f'A{r}', 'NET CASH', BLACK, bold=True)
putf(wsI, f'B{r}', f"={A('cash')}-{A('debt_lt')}-{A('debt_st')}", -DCF['nd'] * 1e6, NUM0,
     bold=True)
r += 1
ANCH['stm'] = dict(rev=IREV, gp=IGP, gm=IGM, ebit=IEBIT, pat=IPAT, nwc=INWC)


# ============ READ FIRST =====================================================
wsRF = sheet('READ FIRST')
wsRF.column_dimensions['A'].width = 118
title(wsRF, 'Testahil — Alexandria Mineral Oils Company S.A.E. (EGX: AMOC)', None, 4)
LINES_RF = [
 'Companion model · Independent valuation study · Educational analysis · Not investment advice',
 '',
 'THIS EDITION IS BUILT ON THE AUDITED FINANCIAL STATEMENTS.',
 'The consolidated statements for the transition period 1 July 2025 to 31 December 2025, audited by',
 'Crowe (Dr A. M. Hegazy & Co) with an UNQUALIFIED opinion signed at Giza on 18 February 2026; the',
 'limited-review statements for the six months to 31 December 2024; and the reviewed statements for',
 'the three months to 31 March 2026. Every company figure on the Assumptions sheet carries one of',
 'those three as its source, with the note number where the filing gives one.',
 '',
 'WHAT THAT REPLACED. The first edition of this model was built without the filings, which could not',
 'be reached from the build environment. It triangulated revenue across press reports, reconstructed',
 'the balance sheet from four disclosed lines and a set of days assumptions, and BUILT a cost stack',
 'from house estimates of yields, energy intensity and a solved feedstock differential. All of that',
 'is gone. Twelve published assumptions were overturned, and the fair value moved from EGP 9.38 to',
 'the figure on the Summary sheet.',
 '',
 'IT CALCULATES — IT DOES NOT STORE. Blue cells are inputs; black cells are formulas; green cells',
 'link across sheets. Change a blue cell and the model reprices: the cost of equity is built from the',
 'risk-free rate net of the sovereign default spread, beta and the equity premium; the glide',
 'fractions are derived from the cost-of-debt path and the discount factors compound; the waterfall',
 'chains gross profit to EBITDA to EBIT to NOPAT to free cash flow to present value; the terminal',
 'block chains reinvestment = growth over return; every ratio and per-share figure is a formula.',
 '',
 'ONLY TWO CLASSES OF CELL ARE PASTED.',
 '  1. AUDITED AND DISCLOSED HISTORY — the primary record, not a calculation. Every one of these sits',
 '     on the Assumptions sheet with its filing and note reference.',
 '  2. WHOLE-MODEL RE-RUNS — the bear and bull columns on Summary. Each is a complete revaluation of',
 '     the model at a different set of drivers, which cannot be expressed as a formula inside a grid.',
 '     THESE DO NOT REDRAW WHEN YOU CHANGE A DRIVER. Everything else does.',
 '',
 'THREE THINGS TO KNOW BEFORE READING THE NUMBERS.',
 '',
 '  · THERE IS NO CLEAN AUDITED TWELVE-MONTH PERIOD. The year-end moved from 30 June to 31 December',
 '    and the April-to-June 2025 quarter is not separately filed. The base year is the NINE contiguous',
 '    audited months from 1 July 2025 to 31 March 2026, annualised by four thirds. That scaling is the',
 '    only step between the filings and the base year and it is shown on Product and Cost.',
 '',
 '  · CAPITAL EXPENDITURE IS RUNNING BELOW DEPRECIATION. Cash paid for fixed assets and projects under',
 '    construction was EGP 30.4mn in the audited half and EGP 65.0mn in the reviewed quarter. The plant',
 '    is being run, not renewed. That single fact does more to this valuation than any assumption in',
 '    it, and it cuts both ways: it lifts free cash flow now and it makes the terminal growth rate',
 '    hard to defend, because growth = return x reinvestment implies a NEGATIVE steady-state rate on',
 '    the recent record.',
 '',
 '  · THE COMPANY IS NET CASH. Cash of EGP 2,463.5mn against loans of EGP 21.0mn. Net debt therefore',
 '    enters the bridge as a NEGATIVE, and the weighting on net debt RAISES the discount rate above',
 '    the cost of equity rather than lowering it. A further EGP 508.8mn of deposits is PLEDGED against',
 '    credit facilities and is deliberately NOT counted as free cash.',
 '',
 'NO RATING AND NO PRICE TARGET. A fair-value range and a distribution.',
]
rr = 4
for ln in LINES_RF:
    c = wsRF.cell(row=rr, column=1, value=ln)
    c.font = Font(size=10, bold=ln.isupper() and len(ln) > 20)
    rr += 1
rr += 1
band(wsRF, rr, 4); put(wsRF, f'A{rr}', 'HEADLINE FIGURES — ALL LIVE LINKS', BLACK, bold=True)
rr += 1
for lab, fml, val, fmt in (
        ('Spot price (EGP)', f"={A('spot')}", SPOT, PX),
        ('Weighted central fair value (EGP)', f"=Summary!C{SC}", D['central'], PX),
        ('Implied against spot', f"=Summary!C{SS+1}", D['central'] / SPOT - 1, PCT),
        ('Discounted cash flow lens (EGP)', f"='EV Bridge'!$B${BPS}", DCF['ps'], PX),
        ('Base-year revenue (EGP)', f"='Product and Cost'!E${BREV}", AU['base_rev'] * 1e6, NUM0),
        ('Base-year gross margin, as filed', f"='Product and Cost'!D${BGM}", AU['base_gm'], PCT3),
        ('Terminal value as a share of enterprise value', f"=DCF!$B${DTVS}", DCF['tv_share'], PCT2),
        ('Weighted cost of capital — explicit', f"=DCF!$B${WEXP}", W['wacc_exp'], PCT2),
        ('Weighted cost of capital — terminal', f"=DCF!$B${WTRM}", W['wacc_term'], PCT2)):
    put(wsRF, f'A{rr}', lab, BLACK)
    putf(wsRF, f'C{rr}', fml, val, fmt, green=True)
    rr += 1

ORDER = ['READ FIRST', 'Summary', 'Assumptions', 'Product and Cost', 'Forecast', 'DCF',
         'EV Bridge', 'Other Lenses', 'Statements']
wb._sheets = [wb[n] for n in ORDER]

# ---- resolve cross-sheet tokens ---------------------------------------------
import re as _re
_pat = _re.compile(r'@([A-Z0-9_]+)@')
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and '@' in _c.value:
                _c.value = _pat.sub(lambda m: str(TOK[m.group(1)]), _c.value)

OUTP = os.path.join(HERE, 'AMOC_Valuation_Model_06082026_public.xlsx')
wb.save(OUTP)

# ---- inject cached values so the DELIVERED file carries its own results ------
import re as _re2, shutil as _sh, zipfile as _zf
nform_pre = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith('='))
_tmp = OUTP + '.tmp'
_sh.move(OUTP, _tmp)
_zin = _zf.ZipFile(_tmp); _zout = _zf.ZipFile(OUTP, 'w', _zf.ZIP_DEFLATED)
_xml_sheet = {f'xl/worksheets/sheet{i+1}.xml': ws.title for i, ws in enumerate(wb.worksheets)}
for _it in _zin.infolist():
    _data = _zin.read(_it.filename)
    if _it.filename in _xml_sheet:
        _vals = EXPECT.get(_xml_sheet[_it.filename], {})

        def _fix(m, _v=_vals):
            _ref = _re2.search(r'r="([A-Z]+[0-9]+)"', m.group(0))
            if not _ref or _ref.group(1) not in _v:
                return m.group(0)
            return m.group(0).replace('</f>', f'</f><v>{_v[_ref.group(1)]!r}</v>')

        _data = _re2.sub(r'<c [^>]*>\s*<f>[^<]*</f>\s*</c>', _fix,
                         _data.decode('utf-8')).encode('utf-8')
    _zout.writestr(_it, _data)
_zin.close(); _zout.close(); os.remove(_tmp)
_z = _zf.ZipFile(OUTP)
_cached = sum(len(_re2.findall(r'</f><v>', _z.read(n).decode('utf-8')))
              for n in _z.namelist() if n.startswith('xl/worksheets/sheet'))
_z.close()

nform = nform_pre
nval = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
           if c.value is not None and not (isinstance(c.value, str) and c.value.startswith('=')))
json.dump(dict(expected=EXPECT, anchors=ANCH, counts=dict(formulas=nform, values=nval)),
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
import openpyxl as _op
_wbv = _op.load_workbook(OUTP)
_num = sum(1 for ws in _wbv.worksheets for row in ws.iter_rows() for c in row
           if isinstance(c.value, (int, float)))
_txt = sum(1 for ws in _wbv.worksheets for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and not c.value.startswith('='))
json.dump(dict(formulas=nform, numeric_values=_num, text=_txt,
               formula_share=nform / (nform + _num)),
          open(os.path.join(HERE, 'formula_count.json'), 'w'), indent=1)
print(f'wrote {OUTP}')
print(f'cached values present in {_cached} of {nform} formula cells (verified on the saved file)')
print(f'CENSUS: {nform} formulas / {_num} pasted numeric ({nform/(nform+_num):.1%} live), '
      f'{_txt} text labels')
