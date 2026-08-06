"""Prove the delivered workbook is a LIVE DRIVER model, not a pasted register.

READ FIRST tells the reader that changing a blue cell on Assumptions reprices the model.
That is a claim about the delivered file, so it is tested on the delivered file: each driver
below is perturbed in place, the whole workbook is re-evaluated from scratch, and the test
asserts the named headline moves in the asserted DIRECTION by a non-trivial amount. A
dead-input sweep then bumps every remaining driver and requires it to move something.

Two of the directions below are counter-intuitive and are asserted deliberately, because a
model that got them backwards would still look plausible:

  · A HIGHER COST OF DEBT RAISES the cost of capital here even though the company is net
    cash. Net debt is negative, so the debt WEIGHT is negative; and the cost of net debt is
    (cost of borrowing x debt less cash yield x cash) / (debt less cash), whose denominator
    is negative. Raising the borrowing rate therefore LOWERS the cost of net debt, and a
    negative weight on a lower rate raises the blend. Two sign flips, one result.

  · A HIGHER YIELD ON CASH LOWERS the cost of capital, by the same mechanism running the
    other way. It also raises the valuation, which is the intuitive half.

  · Depreciation pulls the two halves of the valuation in OPPOSITE directions and both legs
    are asserted, so the workbook cannot quietly lose either. In the explicit window a
    higher charge is a pure tax shield and lifts present value. In the terminal state capex
    is unchanged, so a permanently higher charge is a business consuming its own asset base.
"""
import json, os, sys
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import openpyxl
import xlcalc

wb = openpyxl.load_workbook(os.path.join(HERE, 'AMOC_Valuation_Model_06082026_public.xlsx'))
AN = json.load(open(os.path.join(HERE, 'xlsx_expected.json')))['anchors']
DC, BR, BS_, SU, RN = AN['dcf'], AN['bridge'], AN['bs'], AN['sum'], AN['rn']
FCC = AN['cols']['fcst']

A = {}
for row in wb['Assumptions'].iter_rows(min_col=1, max_col=1):
    c = row[0]
    if isinstance(c.value, str):
        A.setdefault(c.value, c.row)


def row_of(label):
    if label not in A:
        raise KeyError(f'no Assumptions row labelled {label!r}')
    return A[label]


def read(overrides=None):
    bk = xlcalc.Book(wb, overrides)
    return dict(
        dcf=bk.cell_value('EV Bridge', f"B{BR['ps']}"),
        central=bk.cell_value('Summary', f"C{SU['central']}"),
        pv_expl=bk.cell_value('DCF', f"B{DC['pv_explicit']}"),
        tv=bk.cell_value('DCF', f"B{DC['tv']}"),
        pat_cy25=bk.cell_value('Product Lines', f"B{AN['legs']['pat_cy25']}"),
        ppe25=bk.cell_value('Balance Sheet', f"E{BS_['ppe']}"),
        ev=bk.cell_value('DCF', f"B{DC['ev']}"),
        rev26=bk.cell_value('DCF', f"B{DC['rev']}"),
        ebitda26=bk.cell_value('DCF', f"B{DC['ebitda']}"),
        wacc=bk.cell_value('DCF', f"B{DC['wacc_exp']}"),
        wacc_term=bk.cell_value('DCF', f"B{DC['wacc_term']}"),
        nd30=bk.cell_value('Balance Sheet', f"{FCC[4]}{BS_['nd']}"),
        bvps=bk.cell_value('Relative & Normalized', f"B{RN['book'] - 5}"),
        nwc=bk.cell_value('Balance Sheet', f"E{BS_['nwc']}"),
        eq23=bk.cell_value('Balance Sheet', f"B{BS_['eq']}"),
        hist_ebitda=bk.cell_value('Income Statement', f"E{AN['is']['ebitda']}"),
        hist_other=bk.cell_value('Income Statement', f"E{AN['is']['other']}"),
        gm30=bk.cell_value('Product Lines', f"{AN['cols']['uc'][4]}{AN['legs']['gm']}"),
        oil_margin=bk.cell_value('Product Lines', f"{AN['cols']['uc'][0]}{AN['legs']['m_spec']}"),
        feed_diff=bk.cell_value('Product Lines', f"B{AN['legs']['feed_diff']}"),
        cogs26=bk.cell_value('Product Lines', f"{AN['cols']['uc'][0]}{AN['legs']['cogs']}"),
        hist_gm_fy25=bk.cell_value('Product Lines', f"D{AN['legs']['hist_gm']}"),
        recon=bk.cell_value('Product Lines', f"E{AN['legs']['recon']}"),
        divyield=bk.cell_value('Summary', f"C{SU['central'] + 11}"),
        expert1=bk.cell_value('Fundamental Valuation', f"C{AN['fv']['e1']}"),
        expert2=bk.cell_value('Fundamental Valuation', f"C{AN['fv']['e2']}"),
        panel=bk.cell_value('Fundamental Valuation', f"C{AN['fv']['panel']}"),
    )


base = read()
print('base:  ' + ' · '.join(f'{k} {v:,.4f}' for k, v in base.items()))

# label, column, bump, headline it must move, required direction, why
CASES = [
    ('Terminal growth', 'C', +0.01, 'dcf', +1,
     'a higher terminal growth rate must raise the discounted cash flow value'),
    ('Beta', 'C', +0.20, 'dcf', -1,
     'a higher beta raises the cost of equity in both windows and must lower the value'),
    ('Terminal risk-free rate', 'C', +0.02, 'dcf', -1,
     'a higher terminal risk-free rate must lower the value'),
    ('Risk-free rate — Egypt 10-year', 'C', +0.02, 'wacc', +1,
     'a higher local risk-free rate must raise the explicit-window cost of capital'),
    ('Equity risk premium — Egypt', 'C', +0.02, 'dcf', -1,
     'a wider equity risk premium must lower the value'),
    ('Sovereign default spread', 'C', +0.01, 'wacc', -1,
     'a wider sovereign spread is NETTED OUT of the risk-free rate, so it lowers the rate'),
    ('Effective tax rate', 'C', +0.05, 'dcf', -1,
     'a higher tax rate must lower NOPAT and the value'),
    ('Receivable days', 'C', +5.0, 'dcf', -1,
     'slower collection absorbs cash into working capital and must lower the value'),
    ('Inventory days on cost of sales', 'C', +5.0, 'nwc', +1,
     'more inventory must raise net working capital'),
    ('Payable days on cost of sales', 'C', +5.0, 'dcf', +1,
     'longer payment terms fund the cycle and must raise the value'),
    ('Chemicals and catalyst — oil', 'C', +20.0, 'gm30', -1,
     'THE COST BUILD IS LIVE: a chemicals charge is a COST, so raising it must NARROW the '
     'blended margin. In the previous edition no cost input existed to bump — the margin was '
     'reached by assumption and this test could not be written'),
    ('Fixed conversion cost, year to Jun-2023', 'C', +200.0, 'gm30', -1,
     'and the fixed conversion leg does the same, through the calibration: a larger fixed leg '
     'leaves LESS of the disclosed FY2022/23 cost of sales for feedstock, which lowers the '
     'solved differential, which reprices feedstock in every forecast year'),
    ('Process loss and internal fuel burn', 'C', +0.01, 'gm30', 0,
     'EXPECTATION CORRECTED. I asserted this would narrow the margin, on the reasoning that a '
     'higher loss rate means more feedstock bought per tonne sold. The model says it moves the '
     'margin by four ten-millionths, and the model is right: the loss rate scales the feedstock '
     'AND energy charge in the calibration year identically, so the solved differential falls by '
     'the same proportion that intake rises, and the two cancel in every forecast year. It is a '
     'genuine invariance and it means the one yield parameter I could not source is one the '
     'margin does not depend on'),
    ('Process loss and internal fuel burn', 'C', +0.01, 'feed_diff', -1,
     'and here is where it DOES show up — the solved differential absorbs it, which is the '
     'mechanism that makes the margin invariant'),
    ('USD/EGP average, year to Jun-2023', 'C', +2.0, 'feed_diff', 0,
     'DECOMPOSED, and this one surprised me too. The calibration-year exchange rate divides the '
     'implied throughput and multiplies every cost component, so it cancels out of the EGP cost '
     'of sales exactly. The differential is solved against a DISCLOSED EGP figure, so it cannot '
     'depend on the rate used to look at it. Sourcing this rate better would not change one '
     'number in the study'),
    ('USD/EGP average, year to Jun-2025', 'C', +2.0, 'hist_gm_fy25', -1,
     'THIS ONE IS NOT INVARIANT, AND THE SWEEP MISSED IT. The June-2025 rate prices that '
     'column of the historical cost build, so a weaker pound narrows that year\'s built margin. '
     'It moves nothing in the forecast, which is why the old probe set scored it dead — the '
     'probe set was the defect, not the input'),
    ('Brent deck, 2030E', 'C', +10.0, 'gm30', +1,
     'THE INPUT THAT DROVE NOTHING NOW DRIVES BOTH SIDES. Crude prices the product AND the '
     'feed, so most of a crude move cancels; what does not cancel is the EGP-denominated fixed '
     'leg, whose share of revenue falls when the dollar side rises. Net: the margin widens '
     'slightly. That this is a SMALL move is the point of the identifiability test'),
    ('Specialty fixed-cost intensity vs the fuel slate', 'C', +1.0, 'gm30', 0,
     'DECOMPOSED: this input allocates the fixed leg BETWEEN lines and cannot change the total. '
     'It must move the per-line margins and leave the blended margin untouched — an invariance '
     'the old assumed-ratio construction could not offer'),
    ('Specialty fixed-cost intensity vs the fuel slate', 'C', +1.0, 'oil_margin', -1,
     'and it must lower the base-oil margin, because it loads more of the fixed cost onto the '
     'specialty lines'),
    ('Paraffin-wax volume growth', 'B', +0.02, 'gm30', +1,
     'THE MIX TEST: growing the high-margin leg faster must widen the BLENDED gross margin with '
     'no leg margin changing. This is the assertion the previous build could not make, because '
     'it had no per-leg margins for a mix to act on'),
    ('Fuel-slate volume growth', 'B', +0.02, 'gm30', -1,
     'and growing the low-margin leg faster must narrow it'),
    ('Base oils, sales value', 'C', +500.0, 'gm30', +1,
     'DECOMPOSED: it is the disclosed VALUE, not the tonnage, that moves the mix. Raising the '
     'base-oil TONNAGE at a fixed disclosed value just reallocates the same revenue between '
     'price and volume and moves nothing — an invariance worth having, not a dead input'),
    ('USD/EGP average, year to Jun-2024', 'C', +2.0, 'rev26', -1,
     'EXPECTATION CORRECTED, NOT THE MODEL. In the previous edition this input was provably '
     'immaterial: it divided the disclosed realisations to get dollar prices and a blanket '
     'reconciliation factor multiplied by exactly the same amount, so the two cancelled. The '
     'cost build removed that factor. This rate now sets the DOLLAR realisations from which the '
     'crack multiples are solved, and those multiples price every forecast year on both sides '
     'of the margin — so a higher rate means lower dollar realisations, lower cracks and lower '
     'revenue. The old assertion was right about the old model and would have been a false pass '
     'here'),
    ('Fixed conversion cost, year to Jun-2023', 'C', +200.0, 'hist_ebitda', -1,
     'THE HISTORICAL MARGIN IS NOW BUILT, SO IT CAN ONLY BE MOVED THROUGH A COST. This row was '
     'previously a blue input holding the four historical gross margins; it is now a formula '
     'reading the cost build, and the only way to reach it is to change what production costs. '
     'A larger fixed conversion leg narrows the base-year margin and shrinks the operating '
     'result'),
    ('Fixed conversion cost, year to Jun-2023', 'C', +200.0, 'hist_other', +1,
     'and because DISCLOSED profit after tax is fixed, a smaller operating result must ENLARGE '
     'the non-operating residual by exactly the same amount — the identity that keeps the '
     'reconstruction honest, now tested through the cost build rather than around it'),
    ('Energy and utilities', 'C', +5.0, 'feed_diff', -1,
     'THE CALIBRATION IS LIVE. The feedstock differential is SOLVED against the disclosed '
     'FY2022/23 cost of sales, so charging more for energy leaves less of that disclosed total '
     'for feedstock and the solved differential must FALL. This is the assertion that proves '
     'the solve is a solve and not a stored number'),
    ('Brent average, year to Jun-2023', 'C', +5.0, 'feed_diff', +1,
     'and a higher calibration-year crude price means the same disclosed feedstock spend buys '
     'the same tonnes at a HIGHER parity, so the differential rises toward it'),
    ('Operating cost load, % of revenue', 'B', +0.005, 'ebitda26', -1,
     'a heavier operating load must cut 2026E EBITDA'),

    ('USD/EGP average rate path', 'B', +3.0, 'rev26', +1,
     'a weaker pound raises the pound value of dollar-benchmarked product and must lift revenue'),
    ('Depreciation, % of revenue', 'C', +0.005, 'pv_expl', +1,
     'in the explicit window a higher charge is a tax shield and must lift the present value'),
    ('Depreciation, % of revenue', 'C', +0.005, 'tv', -1,
     'in the terminal state, against unchanged capex, it must lower the terminal value'),
    ('Capital expenditure, % of revenue', 'B', +0.005, 'dcf', -1,
     'heavier capital spending must lower free cash flow and the value'),
    ('Justified EV / EBITDA', 'C', +1.0, 'central', +1,
     'a higher justified multiple must raise the weighted central'),
    ('Justified price / earnings', 'C', +1.0, 'central', +1,
     'a higher justified price-to-earnings must raise the weighted central'),
    ('Sustainable return on equity', 'C', +0.03, 'central', +1,
     'a higher sustainable return must raise the book lens and the central'),
    ('Cash and equivalents', 'C', +1000.0, 'dcf', +1,
     'more cash reaches the shareholder through the bridge and must raise the value'),
    ('Cost of debt', 'C', +0.03, 'wacc', +1,
     'COUNTER-INTUITIVE: a negative debt weight on a lower cost of net debt raises the blend'),
    ('Yield on cash', 'C', +0.02, 'wacc', -1,
     'COUNTER-INTUITIVE: the same mechanism in reverse lowers the blended rate'),
    ('Yield on cash', 'C', +0.02, 'dcf', +1,
     'and a lower cost of capital must raise the value'),
    ('Terminal debt weight', 'C', +0.10, 'wacc_term', -1,
     'more of the cheaper after-tax debt must lower the terminal cost of capital'),
    ('Terminal cost of debt', 'C', +0.03, 'wacc_term', +1,
     'a higher terminal borrowing rate must raise the terminal cost of capital'),
    ('Dividend payout ratio', 'C', +0.15, 'nd30', +1,
     'paying more out must leave less net cash (higher net debt) at the end of the forecast'),
    ('Total assets', 'C', +500.0, 'dcf', -1,
     'a larger residual asset base raises invested capital, lowers the terminal return and so '
     'raises the reinvestment the terminal growth requires'),
    ('Shares outstanding', 'C', +100.0, 'dcf', -1,
     'the same equity spread over more shares must be worth less per share'),
    ('Share price', 'C', +1.0, 'wacc', -1,
     'a higher market capitalisation shrinks the NEGATIVE net-debt weight toward zero, pulling '
     'the blended rate back down toward the cost of equity'),

    ('Jul-Dec 2025 profit after tax', 'C', +100.0, 'pat_cy25', +1,
     'the transition half is one of the two legs of the constructed calendar-2025 base'),
    ('FY2024/25 profit after tax', 'C', +100.0, 'pat_cy25', +1,
     'the June year is the other leg of that construction'),
    ('Jul-Dec 2024 profit after tax', 'C', +100.0, 'pat_cy25', -1,
     'the prior-year half is SUBTRACTED to isolate January-June 2025, so it moves the base down'),
    ('Total liabilities', 'C', +200.0, 'bvps', -1,
     'DECOMPOSED: total liabilities does NOT touch the cash-flow lens — net working capital comes '
     'from days drivers and the asset base from total assets. What it drives is disclosed equity, '
     'and therefore book value per share and the book lens'),
    ('Total liabilities', 'C', +200.0, 'central', -1,
     'and through the book lens, the weighted central'),
    ('Expert 1 justified price / earnings', 'C', +1.0, 'expert1', +1,
     'THE PANEL IS LIVE. Expert 1 is an independent opinion with its own multiple, so raising it '
     'must raise Expert 1 and NOTHING in the primary model'),
    ('Expert 1 justified price / earnings', 'C', +1.0, 'dcf', 0,
     'and the cash-flow lens must not move by a piastre, because the panel reads FROM the model '
     'and never back into it'),
    ('Expert 1 justified price / earnings', 'C', +1.0, 'panel', 0,
     'DECOMPOSED: the panel MEDIAN does not move either, and that is the finding rather than a '
     'defect. Expert 1 already sits ABOVE the median, so pushing it further up leaves the middle '
     'value where it was — Expert 3, which is the cash-flow lens by identity. This assertion is '
     'the workbook proving in one bump that the panel median is not an independent check on the '
     'model it contains'),
    ('Dividend per share', 'C', +0.20, 'eq23', +1,
     'DECOMPOSED: closing equity is disclosed, so the dividend drives the roll-BACK, not the '
     'roll-forward. A larger dividend means more was paid out of each historical year, so opening '
     'equity three years ago must have been HIGHER, not lower'),
    ('Dividend per share', 'C', +0.20, 'divyield', +1,
     'and it raises the disclosed dividend yield'),
    ('USD/EGP spot', 'C', +2.0, 'central', 0,
     'the spot exchange rate translates the answer into dollars and must not move the pound answer'),
]

fails, rows = [], []
for label, col, bump, key, sign, why in CASES:
    r = row_of(label)
    cur = wb['Assumptions'][f'{col}{r}'].value
    out = read({('Assumptions', f'{col}{r}'): cur + bump})
    delta = out[key] - base[key]
    rel = delta / abs(base[key]) if base[key] else 0.0
    if sign == 0:
        # An invariance claim is a claim about MATERIALITY, not about floating point. A
        # hundred-thousandth of the base cannot reach the second decimal of a price per
        # share, and demanding bit-exactness would fail invariances that are real. The
        # movement branch below is held an order of magnitude clear of this line so that
        # no bump can satisfy both tests.
        ok = abs(rel) < 1e-5
    else:
        ok = (delta * sign > 0) and abs(rel) > 1e-4
    rows.append((label, bump, key, base[key], out[key], rel, ok, why))
    print(f"  [{'OK ' if ok else 'BAD'}] {label} {bump:+g} -> {key} {base[key]:,.4f} -> "
          f"{out[key]:,.4f} ({rel:+.3%})   {why}")
    if not ok:
        fails.append((label, key, delta, why))

# A driver that moves NOTHING anywhere is a dead input.
DEAD_OK = {
    # display-scope by design: the two house bounds sit on the peer sheet, which is context and
    # is deliberately NOT in the valuation chain. Naming them here is the fix the reachability
    # finding asked for — publish the set rather than force history into the chain.
    'House low bound on the multiple', 'House high bound on the multiple',
    # disclosed history that the CURRENT-year model does not consume downstream
    'FY2022/23 revenue', 'FY2022/23 gross profit', 'FY2022/23 cost of sales',
    'FY2022/23 profit after tax', 'FY2023/24 revenue — method A',
    'FY2023/24 revenue — method B', 'FY2023/24 profit after tax',
    'FY2024/25 sales volume', 'FY2024/25 oils and waxes output',
}
print('\nDEAD-INPUT SWEEP — every driver not covered above is bumped and must move something')
dead = []
covered = {c[0] for c in CASES}
for label, r in sorted(A.items(), key=lambda kv: kv[1]):
    if label in covered or label in DEAD_OK:
        continue
    for col in ('C', 'B'):
        cell = wb['Assumptions'][f'{col}{r}']
        if isinstance(cell.value, (int, float)):
            break
    else:
        continue
    out = read({('Assumptions', f'{col}{r}'): cell.value * 1.10 + 1e-6})
    if all(abs(out[k] - base[k]) < 1e-9 for k in base):
        dead.append(label)
if dead:
    print('  inputs that changed nothing:', dead)
else:
    print('  none — every remaining driver reprices the model')

# ---- REACHABILITY: an input can be dead by ABSENCE, and the sweep above cannot see it ------
# The sweep bumps cells that EXIST on the Assumptions sheet. An input that is registered in the
# model, published in the source register as a driver, and never written to the workbook at all
# is invisible to it — which is exactly how seven retired inputs from an earlier revenue build
# survived a gate that reported "0 dead inputs". Absence now counts as death.
SN = json.load(open(os.path.join(HERE, 'study_numbers.json')))
SHEET_LABELS = {c.value for c in
                [row[0] for row in wb['Assumptions'].iter_rows(min_col=1, max_col=1)]
                if isinstance(c.value, str)}
_model_src = open(os.path.join(HERE, 'compute.py')).read()
_doc_src = ''.join(open(os.path.join(HERE, f)).read()
                   for f in ('docx_amoc.py', 'docx_register.py', 'build_xlsx_amoc.py'))
DRIVES = {k for k in SN['inputs'] if f"V['{k}']" in _model_src}
QUOTED = {k for k in SN['inputs'] if f"IN['{k}']" in _doc_src}
UNREACHED = sorted(k for k in SN['inputs'] if k not in DRIVES and k not in QUOTED)
print('\nREACHABILITY SWEEP — every registered input must DRIVE the model or be QUOTED in a '
      'deliverable')
print(f'  {len(DRIVES)} drive the model | {len(QUOTED - DRIVES)} are disclosed context quoted in '
      f'a document | {len(UNREACHED)} reach nothing')
if UNREACHED:
    print('  registered but reaching nothing:', UNREACHED)
else:
    print(f'  none — all {len(SN["inputs"])} registered inputs are accounted for')
assert not UNREACHED, (
    f'{len(UNREACHED)} inputs are registered and published as drivers but drive nothing: '
    f'{UNREACHED}. Delete them or wire them; do not ship a source register that lists inputs '
    f'the model never reads.')

json.dump([dict(driver=l, bump=b, headline=k, base=bv, bumped=ov, move=rel, ok=ok, why=w)
           for l, b, k, bv, ov, rel, ok, w in rows],
          open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1)

assert not fails, f'{len(fails)} drivers failed to move the model as asserted: {fails}'
assert not dead, f'dead inputs: {dead}'
print(f'\nDRIVER TEST OK — {len(CASES)} driver assertions, all in the asserted direction; '
      f'0 dead inputs')
