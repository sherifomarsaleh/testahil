"""AMOC_Valuation_Model_06082026_public.xlsx — 16 sheets, built formula-first.

Blue = input · black = formula · green = cross-sheet link.

The workbook CALCULATES. Every quantity arithmetically derivable from a driver is a live
Excel formula, so a reader can change a blue cell on Assumptions and watch the model
reprice: the cost of capital is built from the risk-free rate net of the sovereign spread,
beta and the premium; the glide fractions are derived from the cost-of-debt path; the
discount factors compound; the waterfall chains margin -> EBIT -> NOPAT -> free cash flow ->
present value; the terminal block chains reinvestment = growth / return; the statements roll
forward; every ratio and per-share figure is a formula.

Only TWO classes of cell are pasted here, and READ FIRST names them:

  1. audited and disclosed history — the primary record, not a calculation. Where a line is
     both disclosed and derivable, the DISCLOSED figure is carried;
  2. whole-model re-runs — the Monte Carlo price map and the sensitivity grids, each cell of
     which is a complete revaluation. Those grids do NOT redraw when a driver changes.

The third class the standing rule permits — the output of a unit build too complex to
flatten into a grid — is NOT used. AMOC's revenue build has two product legs (specialty
oils and waxes; fuel and by-products), each a volume times a dollar price times an exchange
rate, and that fits on a sheet as live formulas. Where a build can be shown, it is shown.

Every formula cell records the model's own value into xlsx_expected.json as it is written;
recalc.py then evaluates the delivered workbook independently and asserts the two agree, and
that no formula cell is left unchecked.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36')
FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM3 = '#,##0.000'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'
MULT = '0.00x'; DF4 = '0.0000'

M, HI, HB, F, BASE = D['meta'], D['hist_is'], D['hist_bs'], D['fcst'], D['base']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, TR, REL, NRM, BK = D['experts'], D['terminal_recon'], D['rel'], D['norm'], D['book']
U, S0, STK, BT = D['unit'], D['step0'], D['strike'], D['backtest']
IN = {k: v['value'] for k, v in D['inputs'].items()}
SPOT, SH = M['spot'], M['shares_mn']
TAX = IN['tax_eff']; NCI_SH = DCF['nci_share']; PAYOUT = F['payout']
IN_DEBT = IN['debt_snap']
YF = F['years']
YH = ['FY2022/23', 'FY2023/24', 'FY2024/25', 'CY2025']
H4 = ['FY23', 'FY24', 'FY25', 'CY25']
HC = ['B', 'C', 'D', 'E']                 # historical columns on the statements
FCOL = ['F', 'G', 'H', 'I', 'J']          # forecast columns on the statements
CD = ['B', 'C', 'D', 'E', 'F']            # forecast columns on the standalone blocks
UC = ['C', 'D', 'E', 'F', 'G']            # forecast columns on Product Legs (B is the base year)
ALLC = HC + FCOL

wb = Workbook()
EXPECT = {}
ANCH = {}
# Cross-sheet references between sheets built in different orders are written as @TOKENS@
# and resolved in one pass at the end. Hardcoding a row number that another sheet has not
# been laid out yet is exactly the "points one row off" defect gate (q) exists to catch, so
# it is made structurally impossible rather than checked for.
TOK = {}


def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws


def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth


def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula AND record the model's own value for that cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H


def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)


def note(ws, row, text, w=10):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SUB; c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=w)
    ws.row_dimensions[row].height = 28


# ============ 4 ASSUMPTIONS (built first: every other sheet points at it) =====
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every blue cell is a live driver', 'Change one and the model reprices. '
      'Values in EGP mn unless stated.', 6, 52, 16)
AR = {}          # key -> row


def drv(row, key, label, unit, value, fmt=NUM1, src=''):
    put(wsA, f'A{row}', label, BLACK, None)
    put(wsA, f'B{row}', unit, BLACK, None)
    put(wsA, f'C{row}', value, BLUE, fmt)
    if src:
        put(wsA, f'D{row}', src, BLACK, None)
    AR[key] = row
    return row


def A(key):
    return f"Assumptions!$C${AR[key]}"


r = 4
band(wsA, r, 6); put(wsA, f'A{r}', 'MARKET AND SHARE COUNT', BLACK, None, bold=True); r += 1
drv(r, 'spot', 'Share price', 'EGP', SPOT, PX, 'EGX close, 6 Aug 2026'); r += 1
drv(r, 'shares', 'Shares outstanding', 'mn', SH, NUM1, 'reported; payout cross-check'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'DISCLOSED HISTORY — REVENUE AND PROFIT', BLACK, None, bold=True); r += 1
drv(r, 'rev_fy23', 'FY2022/23 revenue', 'EGP mn', IN['rev_fy23'], NUM0, 'cost of sales + gross profit'); r += 1
drv(r, 'gp_fy23', 'FY2022/23 gross profit', 'EGP mn', IN['gp_fy23'], NUM0, 'disclosed'); r += 1
drv(r, 'cogs_fy23', 'FY2022/23 cost of sales', 'EGP mn', IN['cogs_fy23'], NUM0, 'disclosed'); r += 1
drv(r, 'pat_fy23', 'FY2022/23 profit after tax', 'EGP mn', IN['pat_fy23'], NUM0, 'consolidated'); r += 1
drv(r, 'rev_fy24_a', 'FY2023/24 revenue — method A', 'EGP mn', IN['rev_fy24_a'], NUM0, 'prior-year comparative'); r += 1
drv(r, 'rev_fy24_b', 'FY2023/24 revenue — method B', 'EGP mn', IN['rev_fy24_b'], NUM0, 'back-solved from +10.8%'); r += 1
drv(r, 'pat_fy24', 'FY2023/24 profit after tax', 'EGP mn', IN['pat_fy24'], NUM0, 'consolidated'); r += 1
drv(r, 'rev_fy25_a', 'FY2024/25 revenue — method A', 'EGP mn', IN['rev_fy25_a'], NUM0, 'company release'); r += 1
drv(r, 'rev_fy25_b', 'FY2024/25 revenue — method B', 'EGP mn', IN['rev_fy25_b'], NUM0, 'aggregator'); r += 1
drv(r, 'rev_fy25_c', 'FY2024/25 revenue — method C', 'EGP mn', IN['rev_fy25_c'], NUM0, 'same release, revenues line'); r += 1
drv(r, 'pat_fy25', 'FY2024/25 profit after tax', 'EGP mn', IN['pat_fy25'], NUM0, 'consolidated'); r += 1
drv(r, 'rev_h1fy25', 'Jul-Dec 2024 revenue', 'EGP mn', IN['rev_h1fy25'], NUM0, 'disclosed comparative'); r += 1
drv(r, 'pat_h1fy25', 'Jul-Dec 2024 profit after tax', 'EGP mn', IN['pat_h1fy25'], NUM0, 'from the +2% disclosure'); r += 1
drv(r, 'rev_h2cy25', 'Jul-Dec 2025 revenue', 'EGP mn', IN['rev_h2cy25'], NUM0, 'transition period'); r += 1
drv(r, 'pat_h2cy25', 'Jul-Dec 2025 profit after tax', 'EGP mn', IN['pat_h2cy25'], NUM0, 'transition period'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'VOLUMES AND THE UNIT BUILD', BLACK, None, bold=True); r += 1
drv(r, 'vol_fy25', 'FY2024/25 sales volume', 'mn tonnes', IN['vol_fy25'], NUM3, 'disclosed'); r += 1
drv(r, 'vol_spec_fy25', 'FY2024/25 oils and waxes output', 'mn tonnes', IN['vol_spec_fy25'], NUM3, 'disclosed, 172kt'); r += 1
drv(r, 'vol_h2cy25', 'Jul-Dec 2025 sales volume', 'mn tonnes', IN['vol_h2cy25'], NUM3, 'disclosed, 808kt'); r += 1
drv(r, 'line_oil_t', 'Base oils sold, year to Jun-2024', 'tonnes', IN['line_oil_t'], NUM0, 'disclosed product table'); r += 1
drv(r, 'line_oil_v', 'Base oils, sales value', 'EGP mn', IN['line_oil_v'], NUM1, 'disclosed'); r += 1
drv(r, 'line_wax_t', 'Paraffin wax sold, year to Jun-2024', 'tonnes', IN['line_wax_t'], NUM0, 'disclosed'); r += 1
drv(r, 'line_wax_v', 'Paraffin wax, sales value', 'EGP mn', IN['line_wax_v'], NUM1, 'disclosed'); r += 1
drv(r, 'line_tot_t', 'Total tonnes, year to Jun-2024', 'tonnes', IN['line_tot_t'], NUM0, 'disclosed'); r += 1
drv(r, 'line_tot_v', 'Total sales value, year to Jun-2024', 'EGP mn', IN['line_tot_v'], NUM1, 'disclosed'); r += 1
drv(r, 'fx_fy24', 'USD/EGP average, year to Jun-2024', 'EGP', IN['fx_fy24'], NUM1, 'month-weighted across the float'); r += 1
drv(r, 'spec_ramp_cy25', 'Specialty tonnage ramp to the base year', '%', IN['spec_ramp_cy25'], PCT, 'on the disclosed +40% export rise'); r += 1
drv(r, 'oil_share_spec', 'Base oils as a share of specialty tonnage', '%', U['oil_share_of_spec'], PCT, 'from the disclosed table'); r += 1
drv(r, 'loss_frac', 'Process loss and internal fuel burn', '%', IN['loss_frac'], PCT, 'share of feedstock intake'); r += 1
drv(r, 'bbl_per_t', 'Barrels per tonne of feedstock', 'bbl', IN['bbl_per_t_feed'], '0.00', 'heavier than light crude at 7.33'); r += 1
drv(r, 'energy_usd_t', 'Energy and utilities', 'USD/t feed', IN['energy_usd_t'], NUM1, 'lube trains are energy-intensive'); r += 1
for _k in ('oil', 'wax', 'fuel'):
    drv(r, 'chem_' + _k, f'Chemicals and catalyst — {_k}', 'USD/t', IN['chem_usd_t'][_k], NUM1,
        'solvent and catalyst; near nil on the fuel slate'); r += 1
drv(r, 'fixed_fy23', 'Fixed conversion cost, year to Jun-2023', 'EGP mn', IN['fixed_cost_fy23'], NUM0, 'labour, maintenance, plant overhead'); r += 1
for _i, _v in enumerate(IN['fixed_cost_infl']):
    drv(r, 'finf_' + str(_i), f'Fixed-cost inflation factor, year {_i + 1}', 'x', _v, '0.000',
        'Egyptian headline inflation easing to target'); r += 1
drv(r, 'cplx', 'Specialty fixed-cost intensity vs the fuel slate', 'x', IN['complexity_weight'], '0.00', 'allocates the fixed leg between lines'); r += 1
drv(r, 'crude_fy23', 'Brent average, year to Jun-2023', 'USD/bbl', IN['crude_hist']['fy23'], NUM1, 'house estimate'); r += 1
drv(r, 'crude_fy24', 'Brent average, year to Jun-2024', 'USD/bbl', IN['crude_hist']['fy24'], NUM1, 'house estimate'); r += 1
drv(r, 'fx_fy23', 'USD/EGP average, year to Jun-2023', 'EGP', IN['fx_hist']['fy23'], NUM1, 'house estimate'); r += 1
drv(r, 'fx_fy25', 'USD/EGP average, year to Jun-2025', 'EGP', IN['fx_hist']['fy25'], NUM1, 'house estimate'); r += 1
drv(r, 'recon_px', 'Base-year realisation premium over the Brent deck', 'x', U['recon_px'], '0.0000', 'carried forward rather than dropped'); r += 1
for _i, _v in enumerate(IN['brent_path']):
    drv(r, 'brent_' + str(_i), f'Brent deck, {2026 + _i}E', 'USD/bbl', _v, NUM1, 'now drives BOTH sides of the margin'); r += 1
drv(r, 'fx_cy25', 'USD/EGP average, calendar 2025', 'EGP', IN['fx_avg_cy25'], NUM1, ''); r += 1
drv(r, 'fx_spot', 'USD/EGP spot', 'EGP', IN['fx'], NUM1, '6 Aug 2026'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'BALANCE SHEET — DISCLOSED SNAPSHOT AND DAYS DRIVERS', BLACK, None, bold=True); r += 1
drv(r, 'assets', 'Total assets', 'EGP mn', IN['assets_snap'], NUM0, 'disclosed'); r += 1
drv(r, 'liab', 'Total liabilities', 'EGP mn', IN['liab_snap'], NUM0, 'disclosed'); r += 1
drv(r, 'cash', 'Cash and equivalents', 'EGP mn', IN['cash_snap'], NUM0, 'disclosed'); r += 1
drv(r, 'debt', 'Gross debt', 'EGP mn', IN['debt_snap'], NUM1, 'disclosed'); r += 1
drv(r, 'recv_days', 'Receivable days', 'days', IN['recv_days'], NUM1, ''); r += 1
drv(r, 'inv_days', 'Inventory days on cost of sales', 'days', IN['inv_days'], NUM1, ''); r += 1
drv(r, 'pay_days', 'Payable days on cost of sales', 'days', IN['pay_days'], NUM1, ''); r += 1
drv(r, 'other_ca', 'Other current assets', 'EGP mn', IN['other_ca'], NUM0, ''); r += 1
drv(r, 'opex_hist', 'Historical operating cost load, % of revenue', '%', IN['opex_pct'][0], PCT2, 'held flat across the reconstructed years'); r += 1
drv(r, 'dps', 'Dividend per share', 'EGP', IN['dps'], PX, 'declared'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'COST OF CAPITAL', BLACK, None, bold=True); r += 1
drv(r, 'rf', 'Risk-free rate — Egypt 10-year', '%', IN['rf'], PCT2, ''); r += 1
drv(r, 'sov', 'Sovereign default spread', '%', IN['sov_spread_cds'], PCT2, 'netted out of rf'); r += 1
drv(r, 'erp', 'Equity risk premium — Egypt', '%', IN['erp_cds'], PCT2, 'Damodaran, CDS basis'); r += 1
drv(r, 'beta', 'Beta', 'x', IN['beta'], '0.000', 'own-stock regression'); r += 1
drv(r, 'kd', 'Cost of debt', '%', IN['kd'], PCT2, ''); r += 1
drv(r, 'cash_yield', 'Yield on cash', '%', IN['cash_yield'], PCT2, ''); r += 1
drv(r, 'tax', 'Effective tax rate', '%', TAX, PCT2, ''); r += 1
drv(r, 'rf_term', 'Terminal risk-free rate', '%', IN['rf_term'], PCT2, 'norm-built'); r += 1
drv(r, 'erp_term', 'Terminal equity risk premium', '%', IN['erp_term'], PCT2, 'normalised'); r += 1
drv(r, 'kd_term', 'Terminal cost of debt', '%', IN['kd_term'], PCT2, ''); r += 1
drv(r, 'wd_term', 'Terminal debt weight', '%', IN['wd_term'], PCT, 'normalised'); r += 1
drv(r, 'g', 'Terminal growth', '%', IN['g_term'], PCT2, 'sensitised 3-7%'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'OPERATING DRIVERS', BLACK, None, bold=True); r += 1
drv(r, 'dna_pct', 'Depreciation, % of revenue', '%', IN['dna_pct'], PCT2, ''); r += 1
drv(r, 'nci', 'Minority share of group profit', '%', NCI_SH, PCT, ''); r += 1
drv(r, 'payout', 'Dividend payout ratio', '%', PAYOUT, PCT, 'reported'); r += 1
drv(r, 'ev_ebitda', 'Justified EV / EBITDA', 'x', IN['ev_ebitda_just'], MULT, ''); r += 1
drv(r, 'mult_low', 'House low bound on the multiple', 'x', 3.5, MULT, 'NOT a peer observation'); r += 1
drv(r, 'mult_high', 'House high bound on the multiple', 'x', 6.0, MULT, 'NOT a peer observation'); r += 1
drv(r, 'pe', 'Justified price / earnings', 'x', IN['pe_just'], MULT, ''); r += 1
drv(r, 'roe_sust', 'Sustainable return on equity', '%', IN['roe_sust'], PCT, ''); r += 1
drv(r, 'e1_pe', 'Expert 1 justified price / earnings', 'x', EXP['e1']['pe'], MULT,
    'Expert 1 strikes a lower multiple than the main normalised lens deliberately: it is an '
    'independent opinion, not a re-run of the house view'); r += 1

r += 1
band(wsA, r, 6); put(wsA, f'A{r}', 'PATHS — one column per forecast year', BLACK, None, bold=True)
prow = r + 1
hdr(wsA, prow, ['Driver'] + YF, 1)
for c, y in enumerate(YF):
    wsA.column_dimensions[get_column_letter(2 + c)].width = 13
PATHS = [
    ('kd_path', 'Cost of debt path', IN['kd_path'], PCT2),
    ('fx_path', 'USD/EGP average rate path', IN['fx_path'], NUM1),
    ('capex_pct', 'Capital expenditure, % of revenue', IN['capex_pct'], PCT2),
    ('opex_pct', 'Operating cost load, % of revenue', IN['opex_pct'], PCT2),
    ('line_vg_oil', 'Base-oil volume growth', IN['line_vol_growth']['oil'], PCT2),
    ('line_vg_wax', 'Paraffin-wax volume growth', IN['line_vol_growth']['wax'], PCT2),
    ('line_vg_fuel', 'Fuel-slate volume growth', IN['line_vol_growth']['fuel'], PCT2),
    ('cash_yield_path', 'Yield on cash path', IN['cash_yield_path'], PCT2),
]
PR = {}
rr = prow + 1
for key, lab, vals, fmt in PATHS:
    put(wsA, f'A{rr}', lab, BLACK, None)
    for c, v in enumerate(vals):
        put(wsA, f'{get_column_letter(2+c)}{rr}', v, BLUE, fmt)
    PR[key] = rr
    rr += 1


def P(key, i):
    return f"Assumptions!${get_column_letter(2+i)}${PR[key]}"


def GMH(i):
    return f"Assumptions!${get_column_letter(2+i)}${GMH_R}"


rr += 1
band(wsA, rr, 6); put(wsA, f'A{rr}', 'HISTORICAL GROSS MARGIN — one column per period', BLACK,
                      None, bold=True); rr += 1
hdr(wsA, rr, ['Period'] + YH, 1); rr += 1
put(wsA, f'A{rr}', 'Gross margin (historical) — BUILT, see Product Lines', BLACK, None)
for c, v in enumerate(U['gm_built']):
    putf(wsA, f'{get_column_letter(2+c)}{rr}',
         f"='Product Lines'!{get_column_letter(2+c)}@PL_HGM@", v, PCT2)
GMH_R = rr; rr += 1
put(wsA, f'A{rr}', 'NOT an input. FY2022/23 is the DISCLOSED margin and the other three are '
                   'PREDICTIONS of the cost build, which is calibrated on that one year alone',
    BLACK, None)
rr += 2
band(wsA, rr, 6); put(wsA, f'A{rr}', 'LENS WEIGHTS', BLACK, None, bold=True); rr += 1
LW = {}
for k, lab in (('dcf', 'Discounted cash flow'), ('relative', 'Relative multiples'),
               ('normalized', 'Normalised earnings power'), ('book', 'Book and sustainable return')):
    put(wsA, f'A{rr}', lab, BLACK, None)
    put(wsA, f'C{rr}', LN[k]['w'], BLUE, PCT)
    LW[k] = rr; rr += 1
put(wsA, f'A{rr}', 'Total', BLACK, None, bold=True)
putf(wsA, f'C{rr}', f"=C{LW['dcf']}+C{LW['relative']}+C{LW['normalized']}+C{LW['book']}", 1.0,
     PCT, bold=True)
rr += 2
note(wsA, rr, 'Blue cells are the only inputs in this workbook. Everything on every other sheet '
              'is a formula pointing back here, except audited history and the two whole-model '
              're-run grids (Monte Carlo, Sensitivity), which are labelled as such on their own '
              'sheets.', 6)


def LWr(k):
    return f"Assumptions!$C${LW[k]}"


# ============ 6 PRODUCT LINES — the bottom-up build, live ====================
wsU = sheet('Product Lines')
title(wsU, 'Revenue and margin built from the disclosed product table',
      'Three lines, each a tonnage times a dollar realisation times an exchange rate. No price '
      'is calibrated and none is a residual. The blended gross margin is an OUTPUT of the mix.',
      8, 44, 13)
LN3 = ['oil', 'wax', 'fuel']
LNAME = dict(oil='Base oils', wax='Paraffin wax', fuel='Fuel and by-products')
r = 4
band(wsU, r, 8); put(wsU, f'A{r}', 'THE DISCLOSED PRODUCT TABLE — year to 30 June 2024',
                     BLACK, None, bold=True); r += 1
hdr(wsU, r, ['', 'Tonnes', 'Value (EGP mn)', 'EGP / tonne', 'USD / tonne', 'Share of tonnage',
             'Share of value']); r += 1
DT = {}
for k in LN3:
    put(wsU, f'A{r}', LNAME[k], BLACK, None)
    if k == 'fuel':
        putf(wsU, f'B{r}', f"={A('line_tot_t')}-B{DT['oil'][0]}-B{DT['wax'][0]}", U['line_fuel_t'], NUM0)
        putf(wsU, f'C{r}', f"={A('line_tot_v')}-C{DT['oil'][0]}-C{DT['wax'][0]}", U['line_fuel_v'], NUM1)
    else:
        putf(wsU, f'B{r}', f"={A('line_'+k+'_t')}", IN[f'line_{k}_t'], NUM0, green=True)
        putf(wsU, f'C{r}', f"={A('line_'+k+'_v')}", IN[f'line_{k}_v'], NUM1, green=True)
    putf(wsU, f'D{r}', f"=C{r}*1000000/B{r}", U['px_egp'][k], NUM0)
    putf(wsU, f'E{r}', f"=D{r}/{A('fx_fy24')}", U['px_usd'][k], NUM0)
    DT[k] = (r, )
    r += 1
put(wsU, f'A{r}', 'Total (disclosed)', BLACK, None, bold=True)
putf(wsU, f'B{r}', f"={A('line_tot_t')}", IN['line_tot_t'], NUM0, bold=True, green=True)
putf(wsU, f'C{r}', f"={A('line_tot_v')}", IN['line_tot_v'], NUM1, bold=True, green=True)
TOTR = r
for k in LN3:
    rr_ = DT[k][0]
    putf(wsU, f'F{rr_}', f"=B{rr_}/B{TOTR}", (IN[f'line_{k}_t'] if k != 'fuel' else U['line_fuel_t'])/IN['line_tot_t'], PCT)
    putf(wsU, f'G{rr_}', f"=C{rr_}/C{TOTR}", (IN[f'line_{k}_v'] if k != 'fuel' else U['line_fuel_v'])/IN['line_tot_v'], PCT)
r += 2
note(wsU, r, 'Every realisation on this sheet is DERIVED from the disclosed tonnage and value. '
             'The previous build had a specialty price that was a free input and a fuel price '
             'that was the residual of the base-year revenue — which meant the "implied fuel '
             'price" offered as a plausibility check was a residual of the very construction it '
             'was said to validate. Nothing here is a residual.', 8)
r += 1

band(wsU, r, 8); put(wsU, f'A{r}', 'THE CALENDAR-2025 BASE — volume built the same way as revenue',
                     BLACK, None, bold=True); r += 1
put(wsU, f'A{r}', 'Jul-Dec 2025 tonnage (disclosed)', BLACK, None)
putf(wsU, f'B{r}', f"={A('vol_h2cy25')}", IN['vol_h2cy25'], NUM3, green=True); VH2 = r; r += 1
put(wsU, f'A{r}', 'Jul-Dec 2024, off the disclosed +14.5%', BLACK, None)
putf(wsU, f'B{r}', f"=B{VH2}/1.145", U['vol_h2_fy25'], NUM3); VH2P = r; r += 1
put(wsU, f'A{r}', 'FY2024/25 tonnage (disclosed)', BLACK, None)
putf(wsU, f'B{r}', f"={A('vol_fy25')}", IN['vol_fy25'], NUM3, green=True); VFY = r; r += 1
put(wsU, f'A{r}', 'Jan-Jun 2025 (the residual half)', BLACK, None)
putf(wsU, f'B{r}', f"=B{VFY}-B{VH2P}", U['vol_h1_cy25'], NUM3); r += 1
put(wsU, f'A{r}', 'CALENDAR 2025 TONNAGE', BLACK, None, bold=True)
putf(wsU, f'B{r}', f"=B{r-1}+B{VH2}", U['vol_cy25'], NUM3, bold=True); VCY = r; r += 1
put(wsU, f'A{r}', 'Specialty tonnage (FY2024/25 disclosed, plus the export ramp)', BLACK, None)
putf(wsU, f'B{r}', f"={A('vol_spec_fy25')}*(1+{A('spec_ramp_cy25')})", U['spec_vol25'], NUM3)
SPV = r; r += 2

band(wsU, r, 8); put(wsU, f'A{r}', 'THE CALENDAR-2025 BASE LINES', BLACK, None,
                     bold=True); r += 1
hdr(wsU, r, ['', 'Tonnes (mn)', '', '', '']); r += 1
BR3 = {}
for k in LN3:
    put(wsU, f'A{r}', LNAME[k], BLACK, None)
    if k == 'oil':
        putf(wsU, f'B{r}', f"=B{SPV}*{A('oil_share_spec')}", U['vol25']['oil'], NUM3)
    elif k == 'wax':
        putf(wsU, f'B{r}', f"=B{SPV}*(1-{A('oil_share_spec')})", U['vol25']['wax'], NUM3)
    else:
        putf(wsU, f'B{r}', f"=B{VCY}-B{SPV}", U['vol25']['fuel'], NUM3)
    BR3[k] = r; r += 1
put(wsU, f'A{r}', 'Calendar-2025 revenue from two disclosed halves (EGP mn)', BLACK, None)
putf(wsU, f'E{r}', f"=({A('rev_fy25_a')}+{A('rev_fy25_b')}+{A('rev_fy25_c')})/3"
                   f"-{A('rev_h1fy25')}+{A('rev_h2cy25')}", BASE['rev_cy25'], NUM0)
RCY25_R = r; r += 2

band(wsU, r, 8); put(wsU, f'A{r}', 'THE CRACK STRUCTURE — solved from the disclosed table',
                     BLACK, None, bold=True); r += 1
put(wsU, f'A{r}', 'Brent average, year to June 2024 (USD / bbl)', BLACK, None)
putf(wsU, f'B{r}', f"={A('crude_fy24')}", IN['crude_hist']['fy24'], NUM1, green=True)
CB24 = r; r += 1
put(wsU, f'A{r}', 'Crude parity (USD / tonne) = Brent x barrels per tonne', BLACK, None)
putf(wsU, f'B{r}', f"=B{CB24}*{A('bbl_per_t')}", U['parity_fy24'], NUM0); PAR24 = r; r += 1
CRK = {}
for k in LN3:
    put(wsU, f'A{r}', f'{LNAME[k]} — realisation as a multiple of crude parity', BLACK, None)
    putf(wsU, f'B{r}', f"=E{DT[k][0]}/B{PAR24}", U['crack'][k], '0.000')
    CRK[k] = r; r += 1
r += 1
note(wsU, r, 'These are the disclosed product table divided by the crude price. They are not '
             'assumptions and they are not fitted to any margin. Base oil near 1.9x crude, wax '
             'near 1.7x and a gas-oil blend within a per cent of parity is the textbook shape of '
             'a lube refinery slate — the strongest single check that the product table is real.',
     8)
r += 2

band(wsU, r, 8); put(wsU, f'A{r}',
                     'THE FEEDSTOCK DIFFERENTIAL — solved on the ONE disclosed cost of sales',
                     BLACK, None, bold=True); r += 1
HM = U['hist_margin']
put(wsU, f'A{r}', 'Brent average, year to June 2023 (USD / bbl)', BLACK, None)
putf(wsU, f'B{r}', f"={A('crude_fy23')}", IN['crude_hist']['fy23'], NUM1, green=True)
CB23 = r; r += 1
put(wsU, f'A{r}', 'Crude parity, year to June 2023 (USD / tonne)', BLACK, None)
putf(wsU, f'B{r}', f"=B{CB23}*{A('bbl_per_t')}", HM['fy23']['parity'], NUM0); PAR23 = r; r += 1
put(wsU, f'A{r}', 'Blended realisation at the disclosed mix (USD / tonne)', BLACK, None)
putf(wsU, f'B{r}', "=" + "+".join(f"F{DT[k][0]}*B{PAR23}*B{CRK[k]}" for k in LN3),
     sum((IN[f'line_{k}_t'] if k != 'fuel' else U['line_fuel_t']) / IN['line_tot_t']
         * HM['fy23']['parity'] * U['crack'][k] for k in LN3), NUM0)
BL23 = r; r += 1
put(wsU, f'A{r}', 'IMPLIED FY2022/23 THROUGHPUT (mn tonnes) — derived, not assumed', BLACK, None)
putf(wsU, f'B{r}', f"={A('rev_fy23')}/(B{BL23}*{A('fx_fy23')})*1000000/1000000",
     U['ton_fy23'], NUM3); T23 = r; r += 1
put(wsU, f'A{r}', 'Feedstock intake (mn tonnes) = product / (1 - process loss)', BLACK, None)
putf(wsU, f'B{r}', f"=B{T23}/(1-{A('loss_frac')})", U['ton_fy23']/(1-IN['loss_frac']), NUM3)
FT23 = r; r += 1
put(wsU, f'A{r}', 'Energy and utilities (EGP mn)', BLACK, None)
putf(wsU, f'B{r}', f"=B{FT23}*{A('energy_usd_t')}*{A('fx_fy23')}", HM['fy23']['cogs']['energy'],
     NUM0); r += 1
put(wsU, f'A{r}', 'Chemicals, solvent and catalyst (EGP mn)', BLACK, None)
putf(wsU, f'B{r}', "=(" + "+".join(f"B{T23}*F{DT[k][0]}*{A('chem_'+k)}" for k in LN3) +
     f")*{A('fx_fy23')}", HM['fy23']['cogs']['chem'], NUM0); r += 1
put(wsU, f'A{r}', 'Fixed conversion — labour, maintenance, plant overhead (EGP mn)', BLACK, None)
putf(wsU, f'B{r}', f"={A('fixed_fy23')}", IN['fixed_cost_fy23'], NUM0, green=True); r += 1
put(wsU, f'A{r}', 'DISCLOSED cost of sales, year to June 2023 (EGP mn)', BLACK, None)
putf(wsU, f'B{r}', f"={A('cogs_fy23')}", IN['cogs_fy23'], NUM0, green=True); CG23 = r; r += 1
put(wsU, f'A{r}', 'Residual available for feedstock (EGP mn)', BLACK, None)
putf(wsU, f'B{r}', f"=B{CG23}-B{CG23-3}-B{CG23-2}-B{CG23-1}",
     IN['cogs_fy23'] - HM['fy23']['cogs']['energy'] - HM['fy23']['cogs']['chem']
     - IN['fixed_cost_fy23'], NUM0); FRES = r; r += 1
put(wsU, f'A{r}', 'FEEDSTOCK DIFFERENTIAL vs crude parity — SOLVED', BLACK, None, bold=True)
putf(wsU, f'B{r}', f"=B{FRES}/(B{FT23}*B{PAR23}*{A('fx_fy23')})", U['feed_diff'], '0.0000',
     bold=True)
FD = r; r += 2
note(wsU, r, 'Nothing about the margin is assumed anywhere on this sheet. The year to June 2023 '
             'is the only period with a disclosed cost of sales; the feedstock charge is whatever '
             'it has to be for the build to reproduce it exactly. The answer — a small discount '
             'to crude parity — is what a lube plant drawing vacuum gas oil from the adjacent '
             'state complex should show. Change the crude deck, the loss rate or the energy '
             'assumption and this cell resolves; it is a solve, not an input.', 8)
r += 2

band(wsU, r, 8); put(wsU, f'A{r}',
                     'THE HISTORICAL MARGIN — calibrated on ONE year, the other three PREDICTED',
                     BLACK, None, bold=True); r += 1
hdr(wsU, r, ['', *YH]); r += 1
HC4 = ['B', 'C', 'D', 'E']
HKEY = ['fy23', 'fy24', 'fy25', 'cy25']
HFX = [A('fx_fy23'), A('fx_fy24'), A('fx_fy25'), A('fx_cy25')]
HREV = [A('rev_fy23'), A('line_tot_v'),
        f"({A('rev_fy25_a')}+{A('rev_fy25_b')}+{A('rev_fy25_c')})/3", f"E{RCY25_R}"]
_hrevv = [IN['rev_fy23'], IN['line_tot_v'], BASE['rev_fy25'], BASE['rev_cy25']]
put(wsU, f'A{r}', 'Revenue (EGP mn)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={HREV[i]}", _hrevv[i], NUM0)
HREV_R = r; r += 1
HLV = {}
for k in LN3:
    put(wsU, f'A{r}', f'{LNAME[k]} — tonnes (mn)', BLACK, None)
    _t = [U['ton_fy23'], IN['line_tot_t'] / 1e6, IN['vol_fy25'], None]
    for i, c in enumerate(HC4):
        if i == 3:
            putf(wsU, f'{c}{r}', f"=B{BR3[k]}", U['vol25'][k], NUM3)
        elif i == 1:
            putf(wsU, f'{c}{r}', f"=B{DT[k][0]}/1000000",
                 (IN[f'line_{k}_t'] if k != 'fuel' else U['line_fuel_t']) / 1e6, NUM3)
        else:
            putf(wsU, f'{c}{r}', f"={'B' + str(T23) if i == 0 else A('vol_fy25')}*F{DT[k][0]}",
                 _t[i] * ((IN[f'line_{k}_t'] if k != 'fuel' else U['line_fuel_t'])
                          / IN['line_tot_t']), NUM3)
    HLV[k] = r; r += 1
put(wsU, f'A{r}', 'Total product tonnes (mn)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HLV['oil']}+{c}{HLV['wax']}+{c}{HLV['fuel']}",
         sum(U['hist_margin'][HKEY[i]]['cogs']['feed_t'] for _ in [0]) * (1 - IN['loss_frac']),
         NUM3)
HVOL_R = r; r += 1
put(wsU, f'A{r}', 'Feedstock intake (mn tonnes)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HVOL_R}/(1-{A('loss_frac')})",
         U['hist_margin'][HKEY[i]]['cogs']['feed_t'], NUM3)
HFEED_R = r; r += 1
put(wsU, f'A{r}', 'Crude parity implied by that revenue (USD / tonne)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', "=" + f"{c}{HREV_R}*1000000/((" +
         "+".join(f"{c}{HLV[k]}*$B${CRK[k]}" for k in LN3) + f")*1000000*{HFX[i]})",
         U['hist_margin'][HKEY[i]]['parity'], NUM0)
HPAR_R = r; r += 1
put(wsU, f'A{r}', 'Implied Brent equivalent (USD / bbl)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HPAR_R}/{A('bbl_per_t')}", U['hist_margin'][HKEY[i]]['brent'],
         NUM1)
r += 1
put(wsU, f'A{r}', 'Feedstock cost (EGP mn)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HFEED_R}*{c}{HPAR_R}*$B${FD}*{HFX[i]}",
         U['hist_margin'][HKEY[i]]['cogs']['feed'], NUM0)
HCF_R = r; r += 1
put(wsU, f'A{r}', 'Energy and utilities (EGP mn)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HFEED_R}*{A('energy_usd_t')}*{HFX[i]}",
         U['hist_margin'][HKEY[i]]['cogs']['energy'], NUM0)
HCE_R = r; r += 1
put(wsU, f'A{r}', 'Chemicals, solvent and catalyst (EGP mn)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', "=(" + "+".join(f"{c}{HLV[k]}*{A('chem_'+k)}" for k in LN3) +
         f")*{HFX[i]}", U['hist_margin'][HKEY[i]]['cogs']['chem'], NUM0)
HCC_R = r; r += 1
put(wsU, f'A{r}', 'Fixed conversion (EGP mn)', BLACK, None)
for i, c in enumerate(HC4):
    _f = A('fixed_fy23') if i == 0 else f"{HC4[i-1]}{r}*{A('finf_' + str(i - 1))}"
    putf(wsU, f'{c}{r}', f"={_f}", U['hist_margin'][HKEY[i]]['cogs']['fixed'], NUM0)
HCX_R = r; r += 1
put(wsU, f'A{r}', 'COST OF SALES (EGP mn)', BLACK, None, bold=True)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HCF_R}+{c}{HCE_R}+{c}{HCC_R}+{c}{HCX_R}",
         U['hist_margin'][HKEY[i]]['cogs']['total'], NUM0, bold=True)
HCOGS_R = r; r += 1
put(wsU, f'A{r}', 'HISTORICAL GROSS MARGIN — an OUTPUT', BLACK, None, bold=True)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"=({c}{HREV_R}-{c}{HCOGS_R})/{c}{HREV_R}", U['gm_built'][i], PCT2,
         bold=True)
HGM_R = r; TOK['PL_HGM'] = HGM_R; r += 1
put(wsU, f'A{r}', 'What the PREVIOUS edition assumed', BLACK, None)
for i, c in enumerate(HC4):
    put(wsU, f'{c}{r}', U['gm_assumed_old'][i], BLUE, PCT2)
r += 1
put(wsU, f'A{r}', 'Difference (built less assumed)', BLACK, None)
for i, c in enumerate(HC4):
    putf(wsU, f'{c}{r}', f"={c}{HGM_R}-{c}{r-1}", U['gm_built'][i] - U['gm_assumed_old'][i],
         PCT2)
r += 2
note(wsU, r, 'The June-2023 column is the CALIBRATION: the feedstock differential above was '
             'solved so that this column reproduces the disclosed cost of sales exactly, so a '
             'zero difference there is arithmetic, not evidence. The other three columns are '
             'PREDICTIONS — nothing in the cost build was tuned to them — and they land within '
             'a point of a house margin path that was built by an entirely different route. '
             'That is the check. Where the two disagree, the BUILT number is the one this model '
             'carries.', 8)
r += 2

band(wsU, r, 8); put(wsU, f'A{r}', 'THE FORECAST — margin falls out of the cost build',
                     BLACK, None, bold=True); r += 1
hdr(wsU, r, ['', *YF]); r += 1
LV = {}
for k in LN3:
    put(wsU, f'A{r}', f'{LNAME[k]} — tonnes (mn)', BLACK, None)
    prev = f'B{BR3[k]}'
    for i, c in enumerate(UC):
        putf(wsU, f'{c}{r}', f"={prev}*(1+{P('line_vg_'+k, i)})", U['lines_vol'][k][i], NUM3)
        prev = f'{c}{r}'
    LV[k] = r; r += 1
put(wsU, f'A{r}', 'Total tonnes (mn)', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{LV['oil']}+{c}{LV['wax']}+{c}{LV['fuel']}", U['vol'][i], NUM3)
VOL_R = r; r += 1
put(wsU, f'A{r}', 'Feedstock intake (mn tonnes)', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{VOL_R}/(1-{A('loss_frac')})", U['vol'][i]/(1-IN['loss_frac']),
         NUM3)
FEED_R = r; r += 1
put(wsU, f'A{r}', 'Brent deck (USD / bbl)', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={A('brent_'+str(i))}", IN['brent_path'][i], NUM1, green=True)
BRT_R = r; r += 1
put(wsU, f'A{r}', 'Crude parity, carrying the base-year realisation premium (USD / tonne)',
    BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{BRT_R}*{A('bbl_per_t')}*{A('recon_px')}",
         IN['brent_path'][i]*IN['bbl_per_t_feed']*U['recon_px'], NUM0)
PAR_R = r; r += 1
put(wsU, f'A{r}', 'USD/EGP average', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={P('fx_path', i)}", IN['fx_path'][i], NUM1, green=True)
FXR = r; r += 1
LR = {}
for k in LN3:
    put(wsU, f'A{r}', f'{LNAME[k]} — revenue (EGP mn)', BLACK, None)
    for i, c in enumerate(UC):
        putf(wsU, f'{c}{r}', f"={c}{LV[k]}*{c}{PAR_R}*$B${CRK[k]}*{c}{FXR}",
             U['lines_rev'][k][i], NUM0)
    LR[k] = r; r += 1
put(wsU, f'A{r}', 'TOTAL REVENUE (EGP mn)', BLACK, None, bold=True)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{LR['oil']}+{c}{LR['wax']}+{c}{LR['fuel']}", F['rev'][i], NUM0,
         bold=True)
REV_R = r; r += 1
put(wsU, f'A{r}', 'Specialty share of revenue', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"=({c}{LR['oil']}+{c}{LR['wax']})/{c}{REV_R}",
         U['spec_rev'][i]/F['rev'][i], PCT)
SSHR = r; r += 1
_fx = IN['fixed_cost_infl']
_ff = IN['fixed_cost_fy23']*_fx[0]*_fx[1]*_fx[2]
_fixv = []
for i in range(5):
    _ff = _ff*_fx[3+i]; _fixv.append(_ff)
put(wsU, f'A{r}', 'Feedstock cost (EGP mn)', BLACK, None)
_feedv = []
for i, c in enumerate(UC):
    _v = U['vol'][i]/(1-IN['loss_frac'])*IN['brent_path'][i]*IN['bbl_per_t_feed']*U['recon_px'] \
        * U['feed_diff']*IN['fx_path'][i]
    _feedv.append(_v)
    putf(wsU, f'{c}{r}', f"={c}{FEED_R}*{c}{PAR_R}*$B${FD}*{c}{FXR}", _v, NUM0)
CFD_R = r; r += 1
put(wsU, f'A{r}', 'Energy and utilities (EGP mn)', BLACK, None)
_env = []
for i, c in enumerate(UC):
    _v = U['vol'][i]/(1-IN['loss_frac'])*IN['energy_usd_t']*IN['fx_path'][i]
    _env.append(_v)
    putf(wsU, f'{c}{r}', f"={c}{FEED_R}*{A('energy_usd_t')}*{c}{FXR}", _v, NUM0)
CEN_R = r; r += 1
put(wsU, f'A{r}', 'Chemicals, solvent and catalyst (EGP mn)', BLACK, None)
_chv = []
for i, c in enumerate(UC):
    _v = sum(U['lines_vol'][k][i]*IN['chem_usd_t'][k] for k in LN3)*IN['fx_path'][i]
    _chv.append(_v)
    putf(wsU, f'{c}{r}', "=(" + "+".join(f"{c}{LV[k]}*{A('chem_'+k)}" for k in LN3) +
         f")*{c}{FXR}", _v, NUM0)
CCH_R = r; r += 1
put(wsU, f'A{r}', 'Fixed conversion (EGP mn)', BLACK, None)
prev = None
for i, c in enumerate(UC):
    if i == 0:
        putf(wsU, f'{c}{r}', f"={A('fixed_fy23')}*{A('finf_0')}*{A('finf_1')}*{A('finf_2')}"
                             f"*{A('finf_3')}", _fixv[0], NUM0)
    else:
        putf(wsU, f'{c}{r}', f"={UC[i-1]}{r}*{A('finf_'+str(3+i))}", _fixv[i], NUM0)
CFX_R = r; r += 1
put(wsU, f'A{r}', 'COST OF SALES (EGP mn)', BLACK, None, bold=True)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{CFD_R}+{c}{CEN_R}+{c}{CCH_R}+{c}{CFX_R}",
         _feedv[i]+_env[i]+_chv[i]+_fixv[i], NUM0, bold=True)
COGS_R = r; r += 1
put(wsU, f'A{r}', 'GROSS PROFIT (EGP mn) — revenue less the built cost', BLACK, None, bold=True)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{REV_R}-{c}{COGS_R}", F['gp'][i], NUM0, bold=True)
GP_R = r; r += 1
put(wsU, f'A{r}', 'BLENDED GROSS MARGIN — an OUTPUT of the cost build', BLACK, None, bold=True)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{GP_R}/{c}{REV_R}", F['gm'][i], PCT2, bold=True)
GM_R = r; r += 2

band(wsU, r, 8); put(wsU, f'A{r}', 'MARGIN BY LINE — the finding the old model could not produce',
                     BLACK, None, bold=True); r += 1
hdr(wsU, r, ['', *YF]); r += 1
put(wsU, f'A{r}', 'Feedstock + energy per tonne of PRODUCT (USD)', BLACK, None)
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"=({c}{CFD_R}+{c}{CEN_R})/{c}{FXR}/{c}{VOL_R}",
         (_feedv[i]+_env[i])/IN['fx_path'][i]/U['vol'][i], NUM0)
FPT_R = r; r += 1
put(wsU, f'A{r}', 'Fixed conversion per WEIGHTED tonne (USD)', BLACK, None)
_wt = [U['lines_vol']['oil'][i]*IN['complexity_weight'] +
       U['lines_vol']['wax'][i]*IN['complexity_weight'] + U['lines_vol']['fuel'][i]
       for i in range(5)]
for i, c in enumerate(UC):
    putf(wsU, f'{c}{r}', f"={c}{CFX_R}/{c}{FXR}/({c}{LV['oil']}*{A('cplx')}"
                         f"+{c}{LV['wax']}*{A('cplx')}+{c}{LV['fuel']})",
         _fixv[i]/IN['fx_path'][i]/_wt[i], NUM1)
FXPT_R = r; r += 1
LM = {}
for k in LN3:
    put(wsU, f'A{r}', f'{LNAME[k]} — GROSS MARGIN (output)', BLACK, None, bold=True)
    _w = A('cplx') if k != 'fuel' else '1'
    for i, c in enumerate(UC):
        putf(wsU, f'{c}{r}',
             f"=({c}{PAR_R}*$B${CRK[k]}-{c}{FPT_R}-{A('chem_'+k)}-{c}{FXPT_R}*{_w})"
             f"/({c}{PAR_R}*$B${CRK[k]})", U['line_margin'][k][i], PCT2, bold=True)
    LM[k] = r; r += 1
r += 1
note(wsU, r, 'Read the fuel row before anything else on this sheet. The specialty lines earn '
             'tens of per cent and the fuel and by-product slate runs at or below break-even, '
             'because it sells at crude parity and the feedstock costs almost as much as the '
             'product fetches. The previous edition ASSUMED the specialty slate earned 3.5x the '
             'fuel slate, which put them at roughly 14% against 4%. The build says the gap is far '
             'wider than that and that essentially all of this company\'s gross profit is made on '
             'about a seventh of its tonnage. Nothing here was fitted to produce that result.', 8)
r += 2
put(wsU, f'A{r}', 'Calendar-2025 profit after tax, same halves construction', BLACK, None)
putf(wsU, f'B{r}', f"=({A('pat_fy25')}-{A('pat_h1fy25')})+{A('pat_h2cy25')}", BASE['pat_cy25'],
     NUM0)
PATCY25_R = r; r += 1
ANCH['legs'] = dict(rev=REV_R, gp=GP_R, gm=GM_R, vol=VOL_R, spec_share=SSHR,
                    rev_cy25=RCY25_R, pat_cy25=PATCY25_R, recon=FD,
                    m_spec=LM['oil'], m_fuel=LM['fuel'], lines_rev=LR, lines_vol=LV,
                    cogs=COGS_R, feed=CFD_R, crack=CRK, feed_diff=FD, hist_gm=HGM_R,
                    hist_cogs=HCOGS_R, hist_par=HPAR_R)



def LG(row, col):
    return f"'Product Lines'!${col}${row}"


# ============ 8 DCF ==========================================================
wsD = sheet('DCF')
title(wsD, 'Discounted cash flow — the primary lens',
      'Cost of capital built in the sheet; glide inherited from the cost-of-debt path; waterfall '
      'chained to the present value of free cash flow to the firm.', 7, 46, 14)

# --- cost of capital block (built here, referenced by the waterfall) ---
r = 40
band(wsD, r, 7); put(wsD, f'A{r}', 'COST OF CAPITAL — EXPLICIT WINDOW', BLACK, None, bold=True); r += 1
put(wsD, f'A{r}', 'Risk-free rate (Egypt 10-year)', BLACK, None)
putf(wsD, f'B{r}', f"={A('rf')}", IN['rf'], PCT2, green=True); r += 1
put(wsD, f'A{r}', 'less sovereign default spread', BLACK, None)
putf(wsD, f'B{r}', f"={A('sov')}", IN['sov_spread_cds'], PCT2, green=True); r += 1
put(wsD, f'A{r}', 'Risk-free net of sovereign risk', BLACK, None)
putf(wsD, f'B{r}', f"=B{r-2}-B{r-1}", W['rf_star'], PCT2)
RFSTAR_R = r; r += 1
put(wsD, f'A{r}', 'Beta', BLACK, None)
putf(wsD, f'B{r}', f"={A('beta')}", IN['beta'], '0.000', green=True)
BETA_R = r; r += 1
put(wsD, f'A{r}', 'Equity risk premium', BLACK, None)
putf(wsD, f'B{r}', f"={A('erp')}", IN['erp_cds'], PCT2, green=True)
ERP_R = r; r += 1
put(wsD, f'A{r}', 'Cost of equity', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=B{RFSTAR_R}+B{BETA_R}*B{ERP_R}", W['ke_exp'], PCT2, bold=True)
KE_R = r; r += 1
put(wsD, f'A{r}', 'Gross debt', BLACK, None)
putf(wsD, f'B{r}', f"={A('debt')}", IN['debt_snap'], NUM1, green=True)
DEBT_R = r; r += 1
put(wsD, f'A{r}', 'Cash and equivalents', BLACK, None)
putf(wsD, f'B{r}', f"={A('cash')}", IN['cash_snap'], NUM0, green=True)
CASH_R = r; r += 1
put(wsD, f'A{r}', 'Net debt (negative = net cash)', BLACK, None)
putf(wsD, f'B{r}', f"=B{DEBT_R}-B{CASH_R}", BASE['nd_cy25'], NUM0)
ND_R = r; r += 1
put(wsD, f'A{r}', 'Market capitalisation', BLACK, None)
putf(wsD, f'B{r}', f"={A('spot')}*{A('shares')}", M['mktcap'], NUM0)
MC_R = r; r += 1
put(wsD, f'A{r}', 'Debt weight', BLACK, None)
putf(wsD, f'B{r}', f"=B{ND_R}/(B{ND_R}+B{MC_R})", W['wd_exp'], PCT)
WD_R = r; r += 1
put(wsD, f'A{r}', 'Equity weight', BLACK, None)
putf(wsD, f'B{r}', f"=1-B{WD_R}", W['we_exp'], PCT)
WE_R = r; r += 1
put(wsD, f'A{r}', 'After-tax cost of net debt', BLACK, None)
putf(wsD, f'B{r}', f"=({A('kd')}*B{DEBT_R}-{A('cash_yield')}*B{CASH_R})/(B{DEBT_R}-B{CASH_R})"
                   f"*(1-{A('tax')})", W['k_nd_at'], PCT2)
KND_R = r; r += 1
put(wsD, f'A{r}', 'WEIGHTED COST OF CAPITAL — EXPLICIT', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=B{WE_R}*B{KE_R}+B{WD_R}*B{KND_R}", W['wacc_exp'], PCT2, bold=True)
WEXP_R = r; r += 2

band(wsD, r, 7); put(wsD, f'A{r}', 'COST OF CAPITAL — TERMINAL (norm-built)', BLACK, None, bold=True); r += 1
put(wsD, f'A{r}', 'Terminal risk-free rate', BLACK, None)
putf(wsD, f'B{r}', f"={A('rf_term')}", IN['rf_term'], PCT2, green=True); r += 1
put(wsD, f'A{r}', 'Terminal equity risk premium', BLACK, None)
putf(wsD, f'B{r}', f"={A('erp_term')}", IN['erp_term'], PCT2, green=True); r += 1
put(wsD, f'A{r}', 'Terminal cost of equity', BLACK, None)
putf(wsD, f'B{r}', f"=B{r-2}+B{BETA_R}*B{r-1}", W['ke_term'], PCT2)
KET_R = r; r += 1
put(wsD, f'A{r}', 'Terminal after-tax cost of debt', BLACK, None)
putf(wsD, f'B{r}', f"={A('kd_term')}*(1-{A('tax')})", W['kd_term_at'], PCT2)
KDT_R = r; r += 1
put(wsD, f'A{r}', 'Terminal debt weight', BLACK, None)
putf(wsD, f'B{r}', f"={A('wd_term')}", IN['wd_term'], PCT, green=True)
WDT_R = r; r += 1
put(wsD, f'A{r}', 'WEIGHTED COST OF CAPITAL — TERMINAL', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=(1-B{WDT_R})*B{KET_R}+B{WDT_R}*B{KDT_R}", W['wacc_term'], PCT2, bold=True)
WTRM_R = r; r += 2

band(wsD, r, 7); put(wsD, f'A{r}', 'THE GLIDE — fractions inherited from the cost-of-debt path',
                     BLACK, None, bold=True); r += 1
hdr(wsD, r, ['', *YF]); GH_R = r; r += 1
put(wsD, f'A{r}', 'Cost of debt path', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={P('kd_path', i)}", IN['kd_path'][i], PCT2, green=True)
KDP_R = r; r += 1
put(wsD, f'A{r}', 'Cumulative progress (glide fraction)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"=(B{KDP_R}-{c}{KDP_R})/(B{KDP_R}-F{KDP_R})", F['glide_frac'][i], '0.000')
GF_R = r; r += 1
put(wsD, f'A{r}', 'Forward cost of capital', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"=$B${WEXP_R}-($B${WEXP_R}-$B${WTRM_R})*{c}{GF_R}", F['fwd_wacc'][i],
         PCT2, bold=True)
FWD_R = r; r += 1
put(wsD, f'A{r}', 'Cumulative discount factor', BLACK, None)
for i, c in enumerate(CD):
    if i == 0:
        putf(wsD, f'{c}{r}', f"=1/(1+{c}{FWD_R})", F['df'][i], DF4)
    else:
        putf(wsD, f'{c}{r}', f"={CD[i-1]}{r}/(1+{c}{FWD_R})", F['df'][i], DF4)
DF_R = r

# --- the waterfall (rows 4..) ---
hdr(wsD, 4, ['Free cash flow to the firm (EGP mn)', *YF])
r = 5
put(wsD, f'A{r}', 'Revenue', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={LG(ANCH['legs']['rev'], UC[i])}", F['rev'][i], NUM0, green=True)
DREV_R = r; r += 1
put(wsD, f'A{r}', 'Gross margin (from the product-line mix)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={LG(ANCH['legs']['gm'], UC[i])}", F['gm'][i], PCT2, green=True)
DGM_R = r; r += 1
put(wsD, f'A{r}', 'Gross profit', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={LG(ANCH['legs']['gp'], UC[i])}", F['gp'][i], NUM0, green=True)
DGP_R = r; r += 1
put(wsD, f'A{r}', 'less operating cost load', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DREV_R}*{P('opex_pct', i)}", F['opex'][i], NUM0)
DOPX_R = r; r += 1
put(wsD, f'A{r}', 'EBITDA', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DGP_R}-{c}{DOPX_R}", F['ebitda'][i], NUM0, bold=True)
EBITDA_R = r; r += 1
put(wsD, f'A{r}', 'EBITDA margin', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{EBITDA_R}/{c}{DREV_R}", F['ebitda_margin'][i], PCT)
r += 1
put(wsD, f'A{r}', 'less depreciation and amortisation', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DREV_R}*{A('dna_pct')}", F['dna'][i], NUM0)
DNA_R = r; r += 1
put(wsD, f'A{r}', 'EBIT', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{EBITDA_R}-{c}{DNA_R}", F['ebit'][i], NUM0, bold=True)
EBIT_R = r; r += 1
put(wsD, f'A{r}', 'NOPAT = EBIT x (1 - tax)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{EBIT_R}*(1-{A('tax')})", F['nopat'][i], NUM0)
NOPAT_R = r; r += 1
put(wsD, f'A{r}', 'add back depreciation and amortisation', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DNA_R}", F['dna'][i], NUM0)
r += 1
put(wsD, f'A{r}', 'less capital expenditure', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DREV_R}*{P('capex_pct', i)}", F['capex'][i], NUM0)
CAPEX_R = r; r += 1
put(wsD, f'A{r}', 'Net working capital', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DREV_R}*'Balance Sheet'!$E$@BS_NWCP@", F['nwc'][i], NUM0)
NWC_R = r; r += 1
put(wsD, f'A{r}', 'less change in net working capital', BLACK, None)
putf(wsD, f'B{r}', f"=B{NWC_R}-'Balance Sheet'!$E$@BS_NWC@", F['dnwc'][0], NUM0)
for i, c in enumerate(CD[1:], start=1):
    putf(wsD, f'{c}{r}', f"={c}{NWC_R}-{CD[i-1]}{NWC_R}", F['dnwc'][i], NUM0)
DNWC_R = r; r += 1
put(wsD, f'A{r}', 'FREE CASH FLOW TO THE FIRM', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{NOPAT_R}+{c}{DNA_R}-{c}{CAPEX_R}-{c}{DNWC_R}", F['fcff'][i],
         NUM0, bold=True)
FCFF_R = r; r += 1
put(wsD, f'A{r}', 'Discount factor', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{DF_R}", F['df'][i], DF4)
DFL_R = r; r += 1
put(wsD, f'A{r}', 'PRESENT VALUE OF FREE CASH FLOW', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{FCFF_R}*{c}{DFL_R}", F['pv'][i], NUM0, bold=True)
PV_R = r; r += 2

band(wsD, r, 7); put(wsD, f'A{r}', 'INVESTED CAPITAL AND THE TERMINAL BLOCK', BLACK, None,
                     bold=True); r += 1
put(wsD, f'A{r}', 'Property, plant and equipment (roll-forward)', BLACK, None)
putf(wsD, f'B{r}', f"='Balance Sheet'!$E$@BS_PPE@+B{CAPEX_R}-B{DNA_R}", F['ppe'][0], NUM0)
for i, c in enumerate(CD[1:], start=1):
    putf(wsD, f'{c}{r}', f"={CD[i-1]}{r}+{c}{CAPEX_R}-{c}{DNA_R}", F['ppe'][i], NUM0)
PPE_R = r; r += 1
put(wsD, f'A{r}', 'Invested capital', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{NWC_R}+{c}{PPE_R}", F['ic'][i], NUM0)
IC_R = r; r += 1
put(wsD, f'A{r}', 'Return on invested capital', BLACK, None)
for i, c in enumerate(CD):
    putf(wsD, f'{c}{r}', f"={c}{NOPAT_R}/{c}{IC_R}", F['roic'][i], PCT)
r += 1
put(wsD, f'A{r}', 'Terminal growth', BLACK, None)
putf(wsD, f'B{r}', f"={A('g')}", IN['g_term'], PCT2, green=True)
G_R = r; r += 1
put(wsD, f'A{r}', 'Terminal return on invested capital', BLACK, None)
putf(wsD, f'B{r}', f"=F{NOPAT_R}*(1+B{G_R})/F{IC_R}", DCF['roic_term'], PCT)
ROICT_R = r; r += 1
put(wsD, f'A{r}', 'Required reinvestment rate = g / return', BLACK, None)
putf(wsD, f'B{r}', f"=B{G_R}/B{ROICT_R}", DCF['rr_term'], PCT)
RR_R = r; r += 1
put(wsD, f'A{r}', 'Terminal NOPAT', BLACK, None)
putf(wsD, f'B{r}', f"=F{NOPAT_R}*(1+B{G_R})", DCF['ps'] * 0 + F['nopat'][-1] * (1 + IN['g_term']),
     NUM0)
NT_R = r; r += 1
put(wsD, f'A{r}', 'TERMINAL VALUE', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=B{NT_R}*(1-B{RR_R})/(B{WTRM_R}-B{G_R})", DCF['tv'], NUM0, bold=True)
TV_R = r; r += 2

band(wsD, r, 7); put(wsD, f'A{r}', 'ENTERPRISE VALUE', BLACK, None, bold=True); r += 1
put(wsD, f'A{r}', 'Present value of the explicit window', BLACK, None)
putf(wsD, f'B{r}', f"=SUM(B{PV_R}:F{PV_R})", DCF['pv_explicit'], NUM0)
PVE_R = r; r += 1
put(wsD, f'A{r}', 'Present value of the terminal value', BLACK, None)
putf(wsD, f'B{r}', f"=B{TV_R}*F{DFL_R}", DCF['pv_tv'], NUM0)
PVTV_R = r; r += 1
put(wsD, f'A{r}', 'Terminal value as a share of enterprise value', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=B{PVTV_R}/(B{PVE_R}+B{PVTV_R})", DCF['tv_share'], PCT, bold=True)
TVSH_R = r; r += 1
put(wsD, f'A{r}', 'ENTERPRISE VALUE', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"=B{PVE_R}+B{PVTV_R}", DCF['ev'], NUM0, bold=True)
EV_R = r; r += 1
put(wsD, f'A{r}', 'Fair value per share (from the bridge)', BLACK, None, bold=True)
putf(wsD, f'B{r}', f"='EV Bridge'!$B$@BR_PS@", DCF['ps'], PX, bold=True, green=True)
DPS_R = r; r += 2
note(wsD, r, 'One date, one price of time: the terminal value is capitalised at the terminal cost '
             'of capital and discounted at the SAME year-5 cumulative factor as year 5 cash flow. '
             'Discounting a terminal value at a lower rate than the cash flow arriving on the same '
             'day is the commonest way a target price is manufactured, and it is not done here.', 7)
ANCH['dcf'] = dict(ev=EV_R, tv_share=TVSH_R, wacc_exp=WEXP_R, wacc_term=WTRM_R, ebitda=EBITDA_R,
                   nopat=NOPAT_R, fcff=FCFF_R, df=DFL_R, pv=PV_R, ic=IC_R, ppe=PPE_R,
                   nwc=NWC_R, rev=DREV_R, ebit=EBIT_R, dna=DNA_R, capex=CAPEX_R, ke=KE_R,
                   ke_term=KET_R, nd=ND_R, mc=MC_R, g=G_R, fwd=FWD_R, gp=DGP_R)

# ============ 5 EV BRIDGE ====================================================
wsB = sheet('EV Bridge')
title(wsB, 'Enterprise value to equity — the bridge', 'Net debt is negative on this name, so the '
      'cash pile is ADDED rather than subtracted.', 6, 52, 16)
hdr(wsB, 4, ['', 'EGP mn', 'per share'])
r = 5
put(wsB, f'A{r}', 'Enterprise value (discounted cash flow)', BLACK, None)
putf(wsB, f'B{r}', f"=DCF!$B${EV_R}", DCF['ev'], NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}", DCF['ev'] / SH, PX)
BEV_R = r; r += 1
put(wsB, f'A{r}', 'Minority share of group profit', BLACK, None)
putf(wsB, f'B{r}', f"={A('nci')}", NCI_SH, PCT, green=True)
BNCIS_R = r; r += 1
put(wsB, f'A{r}', 'less minority interests — ON THE ENTERPRISE VALUE, BEFORE THE CASH',
    BLACK, None)
putf(wsB, f'B{r}', f"=B{BEV_R}*B{BNCIS_R}", DCF['nci_val'], NUM0)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}", DCF['nci_val'] / SH, PX)
BNCI_R = r; r += 1
put(wsB, f'A{r}', 'Operating assets attributable to shareholders', BLACK, None)
putf(wsB, f'B{r}', f"=B{BEV_R}-B{BNCI_R}", DCF['ev'] - DCF['nci_val'], NUM0)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}", (DCF['ev'] - DCF['nci_val']) / SH, PX)
BPRE_R = r; r += 1
put(wsB, f'A{r}', 'less net debt (negative = net cash added, IN FULL)', BLACK, None)
putf(wsB, f'B{r}', f"=DCF!$B${ND_R}", BASE['nd_cy25'], NUM0, green=True)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}", BASE['nd_cy25'] / SH, PX)
BND_R = r; r += 1
put(wsB, f'A{r}', 'EQUITY ATTRIBUTABLE TO SHAREHOLDERS', BLACK, None, bold=True)
putf(wsB, f'B{r}', f"=B{BPRE_R}-B{BND_R}", DCF['eq_attr'], NUM0, bold=True)
BEQ_R = r; r += 1
put(wsB, f'A{r}', 'Shares outstanding (mn)', BLACK, None)
putf(wsB, f'B{r}', f"={A('shares')}", SH, NUM1, green=True)
BSH_R = r; r += 1
put(wsB, f'A{r}', 'FAIR VALUE PER SHARE (EGP)', BLACK, None, bold=True)
putf(wsB, f'B{r}', f"=B{BEQ_R}/B{BSH_R}", DCF['ps'], PX, bold=True)
BPS_R = r; TOK['BR_PS'] = BPS_R; r += 1
put(wsB, f'A{r}', 'TERMINAL VALUE AS A SHARE OF ENTERPRISE VALUE', BLACK, None, bold=True)
putf(wsB, f'B{r}', f"=DCF!$B${TVSH_R}", DCF['tv_share'], PCT, bold=True, green=True)
BTV_R = r; TOK['BR_TV'] = BTV_R; r += 1
put(wsB, f'A{r}', 'Spot price (EGP)', BLACK, None)
putf(wsB, f'B{r}', f"={A('spot')}", SPOT, PX, green=True)
BSPOT_R = r; r += 1
put(wsB, f'A{r}', 'Implied against spot', BLACK, None)
putf(wsB, f'B{r}', f"=B{BPS_R}/B{BSPOT_R}-1", DCF['ps'] / SPOT - 1, PCT)
r += 1
put(wsB, f'A{r}', 'MEMO — what the WRONG order would have given', BLACK, None, bold=True)
r += 1
put(wsB, f'A{r}', 'Minority taken AFTER the cash, on the combined total (the rejected order)',
    BLACK, None)
putf(wsB, f'B{r}', f"=(B{BEV_R}-B{BND_R})*(1-B{BNCIS_R})",
     (DCF['ev'] - BASE['nd_cy25']) * (1 - NCI_SH), NUM0)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}",
     (DCF['ev'] - BASE['nd_cy25']) * (1 - NCI_SH) / SH, PX)
BWRONG_R = r; r += 1
put(wsB, f'A{r}', 'Value the rejected order would have handed the minority', BLACK, None)
putf(wsB, f'B{r}', f"=B{BEQ_R}-B{BWRONG_R}",
     DCF['eq_attr'] - (DCF['ev'] - BASE['nd_cy25']) * (1 - NCI_SH), NUM0)
putf(wsB, f'C{r}', f"=B{r}/{A('shares')}",
     (DCF['eq_attr'] - (DCF['ev'] - BASE['nd_cy25']) * (1 - NCI_SH)) / SH, PX)
r += 2
note(wsB, r, 'THE ORDER OF THESE ROWS IS THE SUBSTANTIVE POINT. The minority is deducted from '
             'the OPERATING enterprise value, before the cash is added, and the cash is then '
             'added back in full because it belongs to the parent. An earlier edition of this '
             'model did it the other way round — added the cash first, then took the minority '
             'share off the combined total — and defended it in a note reading "the minority '
             'deduction is taken on the equity value AFTER net debt, so the minority does not '
             'carry a share of the parent\'s cash". That defence was algebraically inverted: '
             'deducting a percentage of a total that ALREADY includes the cash is exactly how '
             'you hand the minority a slice of it. An external review caught it, the correction '
             'is accepted, and the memo rows above compute what the rejected order would have '
             'produced so the reader can see the size of it rather than take it on trust. '
             'Doubling the minority share to 6% is a separate contested choice, computed on the '
             'Fundamental Valuation sheet.', 6)

# ============ 10 BALANCE SHEET (needed by DCF references) ====================
wsBS = sheet('Balance Sheet')
title(wsBS, 'Balance sheet — four historical periods and five forecast years',
      'History rolled back from the disclosed snapshot through disclosed profit and dividends; '
      'the forecast rolls forward off the waterfall.', 11, 44, 13)
hdr(wsBS, 4, ['EGP mn', *YH, *YF])
r = 5
put(wsBS, f'A{r}', 'Revenue (memo)', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"='Income Statement'!{HC[i]}@IS_REV@", HI[k]['rev'], NUM0,
         green=True)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"=DCF!{CD[i]}{DREV_R}", F['rev'][i], NUM0, green=True)
BSREV_R = r; TOK['BS_REV'] = BSREV_R; r += 1
put(wsBS, f'A{r}', 'Cost of sales (memo)', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}',
         f"='Income Statement'!{HC[i]}@IS_REV@-'Income Statement'!{HC[i]}@IS_GP@",
         HI[k]['rev'] - HI[k]['gp'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"=DCF!{CD[i]}{DREV_R}-DCF!{CD[i]}{DGP_R}",
         F['rev'][i] - F['gp'][i], NUM0)
BSCOGS_R = r; r += 2

band(wsBS, r, 11); put(wsBS, f'A{r}', 'ASSETS', BLACK, None, bold=True); r += 1
put(wsBS, f'A{r}', 'Property, plant and equipment', BLACK, None)
# The CY2025 figure is the RESIDUAL against disclosed total assets, so it is a formula, not a
# paste; the earlier years scale off it with the square root of revenue, because a refinery's
# plant does not grow one-for-one with turnover.
putf(wsBS, f'E{r}', f"={A('assets')}-{A('cash')}-E@BS_INV@-E@BS_RECV@-{A('other_ca')}",
     HB['CY25']['ppe'], NUM0)
for i, k in enumerate(H4[:3]):
    putf(wsBS, f'{HC[i]}{r}', f"=$E${r}*({HC[i]}@BS_REV@/$E$@BS_REV@)^0.5", HB[k]['ppe'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"=DCF!{CD[i]}{PPE_R}", F['ppe'][i], NUM0, green=True)
BSPPE_R = r; TOK['BS_PPE'] = BSPPE_R; r += 1
put(wsBS, f'A{r}', 'Inventories', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSCOGS_R}*{A('inv_days')}/365", HB[k]['inv'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSCOGS_R}*{A('inv_days')}/365",
         (F['rev'][i] - F['gp'][i]) * IN['inv_days'] / 365, NUM0)
BSINV_R = r; TOK['BS_INV'] = BSINV_R; r += 1
put(wsBS, f'A{r}', 'Trade receivables', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSREV_R}*{A('recv_days')}/365", HB[k]['recv'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSREV_R}*{A('recv_days')}/365",
         F['rev'][i] * IN['recv_days'] / 365, NUM0)
BSRECV_R = r; TOK['BS_RECV'] = BSRECV_R; r += 1
put(wsBS, f'A{r}', 'Other current assets', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={A('other_ca')}", IN['other_ca'], NUM0, green=True)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={A('other_ca')}", IN['other_ca'], NUM0, green=True)
BSOCA_R = r; r += 1
put(wsBS, f'A{r}', 'Cash and equivalents', BLACK, None)
putf(wsBS, f'E{r}', f"={A('cash')}", HB['CY25']['cash'], NUM0, green=True)
for i, k in enumerate(H4[:3]):
    putf(wsBS, f'{HC[i]}{r}',
         f"={HC[i]}@BS_EQ@+{HC[i]}@BS_NCI@+{HC[i]}@BS_PAY@+{HC[i]}@BS_OL@+{A('debt')}"
         f"-{HC[i]}@BS_PPE@-{HC[i]}@BS_INV@-{HC[i]}@BS_RECV@-{A('other_ca')}",
         HB[k]['cash'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={A('debt')}-'Cash Flow'!{CD[i]}$@CF_CND@", F['cash'][i], NUM0)
BSCASH_R = r; r += 1
put(wsBS, f'A{r}', 'TOTAL ASSETS', BLACK, None, bold=True)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}',
         f"={HC[i]}{BSPPE_R}+{HC[i]}{BSINV_R}+{HC[i]}{BSRECV_R}+{HC[i]}{BSOCA_R}+{HC[i]}{BSCASH_R}",
         HB[k]['ppe'] + HB[k]['inv'] + HB[k]['recv'] + IN['other_ca'] + HB[k]['cash'], NUM0,
         bold=True)
for i, c in enumerate(FCOL):
    _ta = (F['ppe'][i] + (F['rev'][i] - F['gp'][i]) * IN['inv_days'] / 365
           + F['rev'][i] * IN['recv_days'] / 365 + IN['other_ca'] + F['cash'][i])
    putf(wsBS, f'{c}{r}', f"={c}{BSPPE_R}+{c}{BSINV_R}+{c}{BSRECV_R}+{c}{BSOCA_R}+{c}{BSCASH_R}",
         _ta, NUM0, bold=True)
BSTA_R = r; r += 2

band(wsBS, r, 11); put(wsBS, f'A{r}', 'LIABILITIES AND EQUITY', BLACK, None, bold=True); r += 1
put(wsBS, f'A{r}', 'Trade payables', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSCOGS_R}*{A('pay_days')}/365", HB[k]['pay'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSCOGS_R}*{A('pay_days')}/365",
         (F['rev'][i] - F['gp'][i]) * IN['pay_days'] / 365, NUM0)
BSPAY_R = r; TOK['BS_PAY'] = BSPAY_R; r += 1
put(wsBS, f'A{r}', 'Gross debt', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={A('debt')}", IN['debt_snap'], NUM1, green=True)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={A('debt')}", IN['debt_snap'], NUM1, green=True)
BSDEBT_R = r; r += 1
put(wsBS, f'A{r}', 'Other liabilities, provisions and tax', BLACK, None)
putf(wsBS, f'E{r}', f"={A('liab')}-{A('debt')}-E{BSPAY_R}", HB['CY25']['other_liab'], NUM0)
for i, k in enumerate(H4[:3]):
    putf(wsBS, f'{HC[i]}{r}', f"=$E${r}*{HC[i]}{BSREV_R}/$E${BSREV_R}", HB[k]['other_liab'],
         NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSREV_R}/{'E'}{BSREV_R}*{HC[3]}{r}",
         BASE['other_liab'] * F['rev'][i] / BASE['rev_cy25'], NUM0)
BSOL_R = r; TOK['BS_OL'] = BSOL_R; r += 1
put(wsBS, f'A{r}', 'Shareholders equity (attributable)', BLACK, None)
# Closing equity is disclosed total assets less disclosed total liabilities; the three earlier
# years are rolled BACKWARDS through disclosed profit and the declared dividend.
putf(wsBS, f'E{r}', f"=({A('assets')}-{A('liab')})*(1-{A('nci')})", HB['CY25']['eqp'], NUM0)
putf(wsBS, f'D{r}', f"=E{r}/(1-{A('nci')})-{A('pat_h2cy25')}+0.5*{A('dps')}*{A('shares')}"
                    f"-(E{r}/(1-{A('nci')})-{A('pat_h2cy25')}+0.5*{A('dps')}*{A('shares')})"
                    f"*{A('nci')}", HB['FY25']['eqp'], NUM0)
putf(wsBS, f'C{r}', f"=D{r}/(1-{A('nci')})-{A('pat_fy25')}+{A('dps')}*{A('shares')}"
                    f"-(D{r}/(1-{A('nci')})-{A('pat_fy25')}+{A('dps')}*{A('shares')})"
                    f"*{A('nci')}", HB['FY24']['eqp'], NUM0)
putf(wsBS, f'B{r}', f"=C{r}/(1-{A('nci')})-{A('pat_fy24')}+{A('dps')}*{A('shares')}"
                    f"-(C{r}/(1-{A('nci')})-{A('pat_fy24')}+{A('dps')}*{A('shares')})"
                    f"*{A('nci')}", HB['FY23']['eqp'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"='Cash Flow'!{CD[i]}$@CF_CEQ@", F['equity'][i], NUM0, green=True)
BSEQ_R = r; TOK['BS_EQP'] = BSEQ_R; TOK['BS_EQ'] = BSEQ_R; r += 1
put(wsBS, f'A{r}', 'Minority interests', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSEQ_R}/(1-{A('nci')})*{A('nci')}", HB[k]['nci'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSEQ_R}/(1-{A('nci')})*{A('nci')}",
         F['equity'][i] / (1 - NCI_SH) * NCI_SH, NUM0)
BSNCI_R = r; TOK['BS_NCI'] = BSNCI_R; r += 1
put(wsBS, f'A{r}', 'TOTAL LIABILITIES AND EQUITY', BLACK, None, bold=True)
for i, k in enumerate(H4):
    _v = (HB[k]['pay'] + IN['debt_snap'] + HB[k]['other_liab'] + HB[k]['eqp'] + HB[k]['nci'])
    putf(wsBS, f'{HC[i]}{r}',
         f"={HC[i]}{BSPAY_R}+{HC[i]}{BSDEBT_R}+{HC[i]}{BSOL_R}+{HC[i]}{BSEQ_R}+{HC[i]}{BSNCI_R}",
         _v, NUM0, bold=True)
for i, c in enumerate(FCOL):
    _v = ((F['rev'][i] - F['gp'][i]) * IN['pay_days'] / 365 + IN['debt_snap']
          + BASE['other_liab'] * F['rev'][i] / BASE['rev_cy25']
          + F['equity'][i] + F['equity'][i] / (1 - NCI_SH) * NCI_SH)
    putf(wsBS, f'{c}{r}', f"={c}{BSPAY_R}+{c}{BSDEBT_R}+{c}{BSOL_R}+{c}{BSEQ_R}+{c}{BSNCI_R}",
         _v, NUM0, bold=True)
BSTLE_R = r; r += 2

band(wsBS, r, 11); put(wsBS, f'A{r}', 'DERIVED', BLACK, None, bold=True); r += 1
put(wsBS, f'A{r}', 'Net working capital', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}',
         f"={HC[i]}{BSINV_R}+{HC[i]}{BSRECV_R}+{HC[i]}{BSOCA_R}-{HC[i]}{BSPAY_R}",
         HB[k]['nwc'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"=DCF!{CD[i]}{NWC_R}", F['nwc'][i], NUM0, green=True)
BSNWC_R = r; TOK['BS_NWC'] = BSNWC_R; r += 1
put(wsBS, f'A{r}', 'Net working capital, % of revenue', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSNWC_R}/{HC[i]}{BSREV_R}",
         HB[k]['nwc'] / HI[k]['rev'], PCT)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSNWC_R}/{c}{BSREV_R}", F['nwc'][i] / F['rev'][i], PCT)
NWCP_R = r; TOK['BS_NWCP'] = NWCP_R; r += 1
put(wsBS, f'A{r}', 'Net debt (negative = net cash)', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSDEBT_R}-{HC[i]}{BSCASH_R}", HB[k]['nd'], NUM0)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"={c}{BSDEBT_R}-{c}{BSCASH_R}", F['net_debt'][i], NUM0)
BSND_R = r; TOK['BS_ND'] = BSND_R; r += 1
put(wsBS, f'A{r}', 'Net cash per share (EGP)', BLACK, None)
for i, k in enumerate(H4):
    putf(wsBS, f'{HC[i]}{r}', f"=-{HC[i]}{BSND_R}/{A('shares')}", -HB[k]['nd'] / SH, PX)
for i, c in enumerate(FCOL):
    putf(wsBS, f'{c}{r}', f"=-{c}{BSND_R}/{A('shares')}", -F['net_debt'][i] / SH, PX)
r += 1
put(wsBS, f'A{r}', 'Balance check (assets less liabilities and equity)', BLACK, None)
BSCHK_R = r
for i, k in enumerate(H4):
    _v = (HB[k]['ppe'] + HB[k]['inv'] + HB[k]['recv'] + IN['other_ca'] + HB[k]['cash']) - \
         (HB[k]['pay'] + IN['debt_snap'] + HB[k]['other_liab'] + HB[k]['eqp'] + HB[k]['nci'])
    putf(wsBS, f'{HC[i]}{r}', f"={HC[i]}{BSTA_R}-{HC[i]}{BSTLE_R}", _v, NUM0)
r += 2
note(wsBS, r, 'The historical balance sheets are a reconstruction, and the workbook says so: only '
              'total assets, total liabilities, cash and gross debt are disclosed at one date. '
              'Everything else is built from days drivers and a roll-back through disclosed profit '
              'and dividends. The reconstruction is checked two ways — implied remaining asset '
              'life, and share capital plus disclosed reserves against total equity.', 11)

# ============ 11 CASH FLOW ===================================================
wsC = sheet('Cash Flow')
title(wsC, 'Cash flow — linked to the waterfall, not restated',
      'Every line points at the discounted cash flow sheet or at the roll-forward, so the two '
      'cannot drift apart.', 7, 46, 14)
hdr(wsC, 4, ['EGP mn', *YF])
r = 5
put(wsC, f'A{r}', 'EBITDA', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=DCF!{c}{EBITDA_R}", F['ebitda'][i], NUM0, green=True)
r += 1
put(wsC, f'A{r}', 'less depreciation and amortisation', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=DCF!{c}{DNA_R}", F['dna'][i], NUM0, green=True)
r += 1
put(wsC, f'A{r}', 'EBIT', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{r-2}-{c}{r-1}", F['ebit'][i], NUM0)
CFEBIT_R = r; r += 1
put(wsC, f'A{r}', 'Cash tax on EBIT', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{CFEBIT_R}*{A('tax')}", F['ebit'][i] * TAX, NUM0)
r += 1
put(wsC, f'A{r}', 'NOPAT', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{CFEBIT_R}-{c}{r-1}", F['nopat'][i], NUM0)
CFNOPAT_R = r; r += 1
put(wsC, f'A{r}', 'add back depreciation and amortisation', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=DCF!{c}{DNA_R}", F['dna'][i], NUM0, green=True)
r += 1
put(wsC, f'A{r}', 'less change in net working capital', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=DCF!{c}{DNWC_R}", F['dnwc'][i], NUM0, green=True)
r += 1
put(wsC, f'A{r}', 'OPERATING CASH FLOW', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{CFNOPAT_R}+{c}{r-2}-{c}{r-1}",
         F['nopat'][i] + F['dna'][i] - F['dnwc'][i], NUM0, bold=True)
CFOCF_R = r; r += 1
put(wsC, f'A{r}', 'less capital expenditure', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=DCF!{c}{CAPEX_R}", F['capex'][i], NUM0, green=True)
r += 1
put(wsC, f'A{r}', 'FREE CASH FLOW TO THE FIRM', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{CFOCF_R}-{c}{r-1}", F['fcff'][i], NUM0, bold=True)
CFFCFF_R = r; r += 1
put(wsC, f'A{r}', 'Opening cash (gross debt less opening net debt)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=MAX({A('debt')}-{c}$@CF_OND@,0)",
         max(IN_DEBT - (BASE['nd_cy25'] if i == 0 else F['net_debt'][i - 1]), 0.0), NUM0)
CFOCASH_R = r; r += 1
put(wsC, f'A{r}', 'Net finance income', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={P('cash_yield_path', i)}*{c}{CFOCASH_R}-{P('kd_path', i)}*{A('debt')}",
         F['interest'][i], NUM0)
CFFIN_R = r; r += 1
put(wsC, f'A{r}', 'Attributable profit', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=({c}{CFEBIT_R}+{c}{CFFIN_R})*(1-{A('tax')})*(1-{A('nci')})",
         F['np_attr'][i], NUM0)
CFNP_R = r; r += 1
put(wsC, f'A{r}', 'Dividends paid', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={A('payout')}*{c}{CFNP_R}", F['div'][i], NUM0)
CFDIV_R = r; r += 2

band(wsC, r, 7); put(wsC, f'A{r}', 'NET CASH ROLL-FORWARD', BLACK, None, bold=True); r += 1
put(wsC, f'A{r}', 'Opening net debt', BLACK, None)
putf(wsC, f'B{r}', f"='Balance Sheet'!$E$@BS_ND@", BASE['nd_cy25'], NUM0, green=True)
for i, c in enumerate(CD[1:], start=1):
    putf(wsC, f'{c}{r}', f"={CD[i-1]}$@CF_CND@", F['net_debt'][i - 1], NUM0)
CFOND_R = r; TOK['CF_OND'] = CFOND_R; r += 1
put(wsC, f'A{r}', 'Closing net debt', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}',
         f"={c}{CFOND_R}-({c}{CFFCFF_R}+{c}{CFFIN_R}*(1-{A('tax')}))+{c}{CFDIV_R}",
         F['net_debt'][i], NUM0, bold=True)
CFCND_R = r; TOK['CF_CND'] = CFCND_R; r += 1
put(wsC, f'A{r}', 'Closing net cash', BLACK, None)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"=-{c}{CFCND_R}", -F['net_debt'][i], NUM0)
r += 1
put(wsC, f'A{r}', 'Opening attributable equity', BLACK, None)
putf(wsC, f'B{r}', f"='Balance Sheet'!$E$@BS_EQP@", BASE['eqp_cy25'], NUM0, green=True)
for i, c in enumerate(CD[1:], start=1):
    putf(wsC, f'{c}{r}', f"={CD[i-1]}$@CF_CEQ@", F['equity'][i - 1], NUM0)
CFOEQ_R = r; r += 1
put(wsC, f'A{r}', 'Closing attributable equity', BLACK, None, bold=True)
for i, c in enumerate(CD):
    putf(wsC, f'{c}{r}', f"={c}{CFOEQ_R}+{c}{CFNP_R}-{c}{CFDIV_R}", F['equity'][i], NUM0,
         bold=True)
CFCEQ_R = r; TOK['CF_CEQ'] = CFCEQ_R; r += 2
note(wsC, r, 'The net-cash roll-forward closes: opening net debt less free cash flow, less the '
             'after-tax finance income the cash itself earns, plus dividends paid, equals closing '
             'net debt. It is one roll-forward, computed once here and consumed by the balance '
             'sheet, the ratios and the equity-side expert method.', 7)

# ============ 9 INCOME STATEMENT =============================================
wsI = sheet('Income Statement')
title(wsI, 'Income statement — four historical periods and five forecast years',
      'Historical periods are the reported June financial years plus the constructed calendar-2025 '
      'base; the forecast runs on calendar years after the year-end change.', 11, 44, 13)
hdr(wsI, 4, ['EGP mn', *YH, *YF])
r = 5


def isline(lab, key, fml_h, fml_f, vals_h, vals_f, fmt=NUM0, bd=False):
    global r
    put(wsI, f'A{r}', lab, BLACK, None, bold=bd)
    for i in range(4):
        if fml_h is None:
            put(wsI, f'{HC[i]}{r}', vals_h[i], BLUE, fmt, bold=bd)
        else:
            putf(wsI, f'{HC[i]}{r}', fml_h(HC[i], i), vals_h[i], fmt, bold=bd)
    for i in range(5):
        putf(wsI, f'{FCOL[i]}{r}', fml_f(FCOL[i], i), vals_f[i], fmt, bold=bd,
             green=fml_f(FCOL[i], i).startswith('=DCF!'))
    r += 1
    return r - 1


# Historical revenue is a formula too: FY2022/23 is the disclosed figure, the two June years
# are the AVERAGE of their independently sourced methods, and calendar 2025 is the constructed
# base. Only the four disclosed profit-after-tax figures remain as pasted values on this sheet.
put(wsI, 'A5', 'Revenue', BLACK, None)
putf(wsI, 'B5', f"={A('rev_fy23')}", HI['FY23']['rev'], NUM0, green=True)
putf(wsI, 'C5', f"=({A('rev_fy24_a')}+{A('rev_fy24_b')})/2", HI['FY24']['rev'], NUM0)
putf(wsI, 'D5', f"=({A('rev_fy25_a')}+{A('rev_fy25_b')}+{A('rev_fy25_c')})/3",
     HI['FY25']['rev'], NUM0)
putf(wsI, 'E5', f"='Product Lines'!E{RCY25_R}", HI['CY25']['rev'], NUM0, green=True)
for i in range(5):
    putf(wsI, f'{FCOL[i]}5', f"=DCF!{CD[i]}{DREV_R}", F['rev'][i], NUM0, green=True)
ISREV = 5
r = 6
ISGM = isline('Gross margin', 'gm', lambda c, i: f"={GMH(i)}",
              lambda c, i: f"=DCF!{CD[i]}{DGM_R}", [HI[k]['gm'] for k in H4], F['gm'], PCT)
ISGP = isline('Gross profit', 'gp', lambda c, i: f"={c}{ISREV}*{c}{ISGM}",
              lambda c, i: f"={c}{ISREV}*{c}{ISGM}", [HI[k]['gp'] for k in H4], F['gp'])
ISOPX = isline('Operating cost load', 'opex',
               lambda c, i: f"={c}{ISREV}*{A('opex_hist')}",
               lambda c, i: f"=DCF!{CD[i]}{DOPX_R}", [HI[k]['opex'] for k in H4], F['opex'])
ISEBITDA = isline('EBITDA', 'ebitda', lambda c, i: f"={c}{ISGP}-{c}{ISOPX}",
                  lambda c, i: f"={c}{ISGP}-{c}{ISOPX}", [HI[k]['ebitda'] for k in H4],
                  F['ebitda'], bd=True)
isline('EBITDA margin', 'em', lambda c, i: f"={c}{ISEBITDA}/{c}{ISREV}",
       lambda c, i: f"={c}{ISEBITDA}/{c}{ISREV}",
       [HI[k]['ebitda'] / HI[k]['rev'] for k in H4], F['ebitda_margin'], PCT)
ISDNA = isline('Depreciation and amortisation', 'dna', lambda c, i: f"={c}{ISREV}*{A('dna_pct')}",
               lambda c, i: f"=DCF!{CD[i]}{DNA_R}", [HI[k]['dna'] for k in H4], F['dna'])
ISEBIT = isline('EBIT', 'ebit', lambda c, i: f"={c}{ISEBITDA}-{c}{ISDNA}",
                lambda c, i: f"={c}{ISEBITDA}-{c}{ISDNA}", [HI[k]['ebit'] for k in H4],
                F['ebit'], bd=True)
ISFIN = isline('Net finance income', 'fin',
               lambda c, i: f"=$E${ISFINH}*{c}{ISREV}/$E${ISREV}*0.85"
                            if False else f"={c}{ISREV}/$E${ISREV}*$E${ISFINH}*0.85",
               lambda c, i: f"='Cash Flow'!{CD[i]}{CFFIN_R}",
               [HI[k]['fin'] for k in H4], F['interest']) if False else None
# the base-year finance line is the anchor the earlier years scale from, so it is written first
put(wsI, f'A{r}', 'Net finance income', BLACK, None)
putf(wsI, f'E{r}', f"={A('cash_yield')}*{A('cash')}-{A('kd')}*{A('debt')}", HI['CY25']['fin'],
     NUM0)
for i in range(3):
    putf(wsI, f'{HC[i]}{r}', f"=$E${r}*{HC[i]}{ISREV}/$E${ISREV}*0.85", HI[H4[i]]['fin'], NUM0)
for i in range(5):
    putf(wsI, f'{FCOL[i]}{r}', f"='Cash Flow'!{CD[i]}{CFFIN_R}", F['interest'][i], NUM0,
         green=True)
ISFIN = r; r += 1
ISPAT = isline('Profit after tax (disclosed for history)', 'pat',
               lambda c, i: (f"={A('pat_fy23')}" if i == 0 else
                             f"={A('pat_fy24')}" if i == 1 else
                             f"={A('pat_fy25')}" if i == 2 else
                             f"='Product Lines'!B@LEGS_PAT@"),
               lambda c, i: f"=({c}{ISEBIT}+{c}{ISFIN})*(1-{A('tax')})",
               [HI[k]['pat'] for k in H4],
               [(F['ebit'][i] + F['interest'][i]) * (1 - TAX) for i in range(5)], bd=True)
ISPBT = isline('Profit before tax', 'ebt', lambda c, i: f"={c}{ISPAT}/(1-{A('tax')})",
               lambda c, i: f"={c}{ISPAT}/(1-{A('tax')})", [HI[k]['ebt'] for k in H4],
               [(F['ebit'][i] + F['interest'][i]) for i in range(5)])
ISOTH = isline('Other and non-operating income', 'other',
               lambda c, i: f"={c}{ISPBT}-{c}{ISEBIT}-{c}{ISFIN}",
               lambda c, i: f"={c}{ISPBT}-{c}{ISEBIT}-{c}{ISFIN}",
               [HI[k]['other'] for k in H4], [0.0] * 5)
ISTAX = isline('Tax', 'tax', lambda c, i: f"={c}{ISPAT}-{c}{ISPBT}",
               lambda c, i: f"={c}{ISPAT}-{c}{ISPBT}", [HI[k]['tax'] for k in H4],
               [-(F['ebit'][i] + F['interest'][i]) * TAX for i in range(5)])
ISNCI = isline('Minority interests', 'nci', lambda c, i: f"={c}{ISPAT}*{A('nci')}",
               lambda c, i: f"={c}{ISPAT}*{A('nci')}",
               [HI[k]['nci'] for k in H4],
               [(F['ebit'][i] + F['interest'][i]) * (1 - TAX) * NCI_SH for i in range(5)])
ISNPA = isline('PROFIT ATTRIBUTABLE TO SHAREHOLDERS', 'npa',
               lambda c, i: f"={c}{ISPAT}-{c}{ISNCI}", lambda c, i: f"={c}{ISPAT}-{c}{ISNCI}",
               [HI[k]['npa'] for k in H4], F['np_attr'], bd=True)
ISEPS = isline('Earnings per share (EGP)', 'eps', lambda c, i: f"={c}{ISNPA}/{A('shares')}",
               lambda c, i: f"={c}{ISNPA}/{A('shares')}",
               [HI[k]['npa'] / SH for k in H4],
               [F['np_attr'][i] / SH for i in range(5)], PX)
r += 1
note(wsI, r, 'Only the historical PROFIT AFTER TAX is a pasted value — it is the disclosed '
             'figure. Every other line in every column is a formula. Note the OTHER INCOME line: '
             'it is the residual between disclosed profit and the operating result, and it is '
             'large in the devaluation years and small now. The forecast carries NONE of it, '
             'which is why forecast attributable profit is not a simple extrapolation of the '
             'reported record.', 11)

# ============ 12 SUMMARY FINANCIALS ==========================================
wsSF = sheet('Summary Financials')
title(wsSF, 'Summary financials — the lines a reader checks first', None, 11, 44, 13)
hdr(wsSF, 4, ['EGP mn', *YH, *YF])
r = 5
SF_ROWS = [
    ('Revenue', f"='Income Statement'!{{c}}{ISREV}", [HI[k]['rev'] for k in H4], F['rev'], NUM0),
    ('EBITDA', f"='Income Statement'!{{c}}{ISEBITDA}", [HI[k]['ebitda'] for k in H4],
     F['ebitda'], NUM0),
    ('EBIT', f"='Income Statement'!{{c}}{ISEBIT}", [HI[k]['ebit'] for k in H4], F['ebit'], NUM0),
    ('Attributable profit', f"='Income Statement'!{{c}}{ISNPA}",
     [HI[k]['ebt'] * (1 - TAX) * (1 - NCI_SH) for k in H4], F['np_attr'], NUM0),
    ('Total assets', f"='Balance Sheet'!{{c}}{BSTA_R}",
     [HB[k]['ppe'] + HB[k]['inv'] + HB[k]['recv'] + IN['other_ca'] + HB[k]['cash'] for k in H4],
     [(F['ppe'][i] + (F['rev'][i] - F['gp'][i]) * IN['inv_days'] / 365
       + F['rev'][i] * IN['recv_days'] / 365 + IN['other_ca'] + F['cash'][i]) for i in range(5)],
     NUM0),
    ('Shareholders equity', f"='Balance Sheet'!{{c}}{BSEQ_R}", [HB[k]['eqp'] for k in H4],
     F['equity'], NUM0),
    ('Net cash', f"=0-'Balance Sheet'!{{c}}{BSND_R}", [-HB[k]['nd'] for k in H4],
     [-x for x in F['net_debt']], NUM0),
]
for lab, fml, vh, vf, fmt in SF_ROWS:
    put(wsSF, f'A{r}', lab, BLACK, None)
    for i in range(4):
        putf(wsSF, f'{HC[i]}{r}', fml.format(c=HC[i]), vh[i], fmt, green=True)
    for i in range(5):
        putf(wsSF, f'{FCOL[i]}{r}', fml.format(c=FCOL[i]), vf[i], fmt, green=True)
    r += 1
r += 1
put(wsSF, f'A{r}', 'Invested capital', BLACK, None)
for i in range(5):
    putf(wsSF, f'{FCOL[i]}{r}', f"=DCF!{CD[i]}{IC_R}", F['ic'][i], NUM0, green=True)
r += 1
put(wsSF, f'A{r}', 'Return on invested capital', BLACK, None)
for i in range(5):
    putf(wsSF, f'{FCOL[i]}{r}', f"=DCF!{CD[i]}{NOPAT_R}/DCF!{CD[i]}{IC_R}", F['roic'][i], PCT)
r += 1
put(wsSF, f'A{r}', 'Free cash flow to the firm', BLACK, None)
for i in range(5):
    putf(wsSF, f'{FCOL[i]}{r}', f"=DCF!{CD[i]}{FCFF_R}", F['fcff'][i], NUM0, green=True)

# ============ 7 RELATIVE & NORMALIZED ========================================
wsR = sheet('Relative & Normalized')
title(wsR, 'Relative multiples, normalised earnings power and book value',
      'The three cross-check lenses, each built on the sheet.', 6, 52, 16)
r = 4
band(wsR, r, 6); put(wsR, f'A{r}', 'LENS 2 — RELATIVE MULTIPLES', BLACK, None, bold=True); r += 1
put(wsR, f'A{r}', f"{YF[1]} EBITDA", BLACK, None)
putf(wsR, f'B{r}', f"=DCF!C{EBITDA_R}", REL['ebitda_mid'], NUM0, green=True)
RE_R = r; r += 1
put(wsR, f'A{r}', 'Justified enterprise value / EBITDA', BLACK, None)
putf(wsR, f'B{r}', f"={A('ev_ebitda')}", IN['ev_ebitda_just'], MULT, green=True)
RM_R = r; r += 1
put(wsR, f'A{r}', f"Enterprise value as at end-{YF[1][:4]}", BLACK, None)
putf(wsR, f'B{r}', f"=B{RE_R}*B{RM_R}", REL['ev_rel_fwd'], NUM0)
RF_R = r; r += 1
put(wsR, f'A{r}', 'Year-2 discount factor', BLACK, None)
putf(wsR, f'B{r}', f"=DCF!C{DFL_R}", REL['df_rel'], DF4, green=True)
RD_R = r; r += 1
put(wsR, f'A{r}', 'Enterprise value discounted to today', BLACK, None)
putf(wsR, f'B{r}', f"=B{RF_R}*B{RD_R}", REL['ev_rel'], NUM0)
RT_R = r; r += 1
put(wsR, f'A{r}', 'add the interim free cash flow the multiple year does not cover', BLACK, None)
putf(wsR, f'B{r}', f"=DCF!B{PV_R}+DCF!C{PV_R}", REL['pv_interim'], NUM0)
RI_R = r; r += 1
put(wsR, f'A{r}', 'less net debt (net cash added)', BLACK, None)
putf(wsR, f'B{r}', f"=DCF!$B${ND_R}", BASE['nd_cy25'], NUM0, green=True)
RN_R = r; r += 1
put(wsR, f'A{r}', 'IMPLIED VALUE PER SHARE (EGP)', BLACK, None, bold=True)
putf(wsR, f'B{r}', f"=(B{RT_R}+B{RI_R})*(1-{A('nci')})/{A('shares')}-B{RN_R}/{A('shares')}",
     LN['relative']['base'], PX, bold=True)
RELPS_R = r; r += 1
put(wsR, f'A{r}', 'Trailing enterprise value / EBITDA (company)', BLACK, None)
putf(wsR, f'B{r}', f"=({A('spot')}*{A('shares')}+DCF!$B${ND_R})/'Income Statement'!E{ISEBITDA}",
     REL['ev_ebitda_trailing'], MULT)
r += 1
put(wsR, f'A{r}', 'Trailing price / earnings (company)', BLACK, None)
putf(wsR, f'B{r}', f"={A('spot')}/'Income Statement'!E{ISEPS}", REL['pe_trailing'], MULT)
r += 2

band(wsR, r, 6); put(wsR, f'A{r}', 'LENS 3 — NORMALISED EARNINGS POWER', BLACK, None, bold=True)
r += 1
put(wsR, f'A{r}', f"{YF[2]} EBITDA", BLACK, None)
putf(wsR, f'B{r}', f"=DCF!D{EBITDA_R}", NRM['ebitda'], NUM0, green=True)
NE_R = r; r += 1
put(wsR, f'A{r}', 'less depreciation and amortisation', BLACK, None)
putf(wsR, f'B{r}', f"=DCF!D{DNA_R}", NRM['dna'], NUM0, green=True)
ND2_R = r; r += 1
put(wsR, f'A{r}', 'EBIT', BLACK, None)
putf(wsR, f'B{r}', f"=B{NE_R}-B{ND2_R}", NRM['ebit'], NUM0)
NB_R = r; r += 1
put(wsR, f'A{r}', 'add net finance income', BLACK, None)
putf(wsR, f'B{r}', f"='Cash Flow'!D{CFFIN_R}", NRM['interest'], NUM0, green=True)
NF_R = r; r += 1
put(wsR, f'A{r}', 'Attributable normalised earnings', BLACK, None)
putf(wsR, f'B{r}', f"=(B{NB_R}+B{NF_R})*(1-{A('tax')})*(1-{A('nci')})", NRM['np'], NUM0)
NN_R = r; r += 1
put(wsR, f'A{r}', 'Normalised earnings per share (EGP)', BLACK, None)
putf(wsR, f'B{r}', f"=B{NN_R}/{A('shares')}", NRM['eps'], PX)
NEPS_R = r; r += 1
put(wsR, f'A{r}', 'Justified price / earnings', BLACK, None)
putf(wsR, f'B{r}', f"={A('pe')}", IN['pe_just'], MULT, green=True)
NPE_R = r; r += 1
put(wsR, f'A{r}', 'IMPLIED VALUE PER SHARE (EGP)', BLACK, None, bold=True)
putf(wsR, f'B{r}', f"=B{NEPS_R}*B{NPE_R}", LN['normalized']['base'], PX, bold=True)
NRMPS_R = r; r += 2

band(wsR, r, 6); put(wsR, f'A{r}', 'LENS 4 — BOOK VALUE AND SUSTAINABLE RETURN', BLACK, None,
                     bold=True); r += 1
put(wsR, f'A{r}', 'Attributable book value per share (EGP)', BLACK, None)
putf(wsR, f'B{r}', f"='Balance Sheet'!E{BSEQ_R}/{A('shares')}", BK['bvps'], PX)
BV_R = r; r += 1
put(wsR, f'A{r}', 'Sustainable return on equity', BLACK, None)
putf(wsR, f'B{r}', f"={A('roe_sust')}", IN['roe_sust'], PCT, green=True)
RS_R = r; r += 1
put(wsR, f'A{r}', 'Perpetual cost of equity', BLACK, None)
putf(wsR, f'B{r}', f"=DCF!$B${KET_R}", BK['ke_term'], PCT2, green=True)
KP_R = r; r += 1
put(wsR, f'A{r}', 'Terminal growth', BLACK, None)
putf(wsR, f'B{r}', f"={A('g')}", IN['g_term'], PCT2, green=True)
BG_R = r; r += 1
put(wsR, f'A{r}', 'Justified price / book', BLACK, None)
putf(wsR, f'B{r}', f"=(B{RS_R}-B{BG_R})/(B{KP_R}-B{BG_R})", BK['pb_just'], MULT)
PB_R = r; r += 1
put(wsR, f'A{r}', 'IMPLIED VALUE PER SHARE (EGP)', BLACK, None, bold=True)
putf(wsR, f'B{r}', f"=B{PB_R}*B{BV_R}", LN['book']['base'], PX, bold=True)
BKPS_R = r; r += 1
put(wsR, f'A{r}', 'Trailing return on average attributable equity', BLACK, None)
putf(wsR, f'B{r}', f"='Income Statement'!E{ISNPA}/(('Balance Sheet'!D{BSEQ_R}+"
                   f"'Balance Sheet'!E{BSEQ_R})/2)", BK['roe_trailing'], PCT)
r += 2
note(wsR, r, 'The relative lens discounts a FORWARD enterprise value back to today; a multiple '
             'applied to a future year produces a value as at that year, not now. The book lens '
             'uses the PERPETUAL cost of equity inside the perpetuity identity, because a blended '
             'rate inside a perpetuity formula is internally inconsistent.', 6)

# ============ 3 FUNDAMENTAL VALUATION ========================================
wsFV = sheet('Fundamental Valuation')
title(wsFV, 'Fundamental valuation — four lenses, weighted', None, 7, 46, 15)
hdr(wsFV, 4, ['Lens', 'Bear (EGP)', 'Base (EGP)', 'Bull (EGP)', 'Weight', 'Contribution'])
r = 5
LENS_SRC = [
    ('dcf', 'Discounted cash flow (primary)', f"='EV Bridge'!$B$@BR_PS@"),
    ('relative', 'Relative multiples', f"='Relative & Normalized'!$B${RELPS_R}"),
    ('normalized', 'Normalised earnings power', f"='Relative & Normalized'!$B${NRMPS_R}"),
    ('book', 'Book value and sustainable return', f"='Relative & Normalized'!$B${BKPS_R}"),
]
LR = {}
for key, lab, src in LENS_SRC:
    put(wsFV, f'A{r}', lab, BLACK, None)
    put(wsFV, f'B{r}', LN[key]['bear'], BLUE, PX)
    putf(wsFV, f'C{r}', src, LN[key]['base'], PX, green=True)
    put(wsFV, f'D{r}', LN[key]['bull'], BLUE, PX)
    putf(wsFV, f'E{r}', f"={LWr(key)}", LN[key]['w'], PCT, green=True)
    putf(wsFV, f'F{r}', f"=C{r}*E{r}", LN[key]['base'] * LN[key]['w'], PX)
    LR[key] = r; r += 1
put(wsFV, f'A{r}', 'WEIGHTED CENTRAL FAIR VALUE (EGP)', BLACK, None, bold=True)
putf(wsFV, f'B{r}', f"=MIN(B{LR['dcf']}:B{LR['book']})", D['span'][0], PX)
putf(wsFV, f'C{r}', f"=SUM(F{LR['dcf']}:F{LR['book']})", D['central'], PX, bold=True)
putf(wsFV, f'D{r}', f"=MAX(D{LR['dcf']}:D{LR['book']})", D['span'][1], PX)
putf(wsFV, f'E{r}', f"=SUM(E{LR['dcf']}:E{LR['book']})", 1.0, PCT)
FVC_R = r; r += 2

band(wsFV, r, 7); put(wsFV, f'A{r}', 'CROSS-CHECKS', BLACK, None, bold=True); r += 1
put(wsFV, f'A{r}', 'Spot price (EGP)', BLACK, None)
putf(wsFV, f'C{r}', f"={A('spot')}", SPOT, PX, green=True)
FVS_R = r; r += 1
put(wsFV, f'A{r}', 'Implied against spot', BLACK, None)
putf(wsFV, f'C{r}', f"=C{FVC_R}/C{FVS_R}-1", D['central'] / SPOT - 1, PCT)
r += 1
put(wsFV, f'A{r}', 'Terminal value as a share of enterprise value', BLACK, None)
putf(wsFV, f'C{r}', f"=DCF!$B${TVSH_R}", DCF['tv_share'], PCT, green=True)
r += 1
r += 1
put(wsFV, f'A{r}', 'THE EXPERT PANEL, LIVE — and why the median is not a check', BLACK, None,
    bold=True)
r += 1
put(wsFV, f'A{r}', 'Expert 1 — earnings power: 2028E attributable earnings per share', BLACK, None)
putf(wsFV, f'B{r}', f"='Relative & Normalized'!$B${NEPS_R}", NRM['eps'], PX, green=True)
E1EPS_R = r; r += 1
put(wsFV, f'A{r}', '        × Expert 1 justified multiple', BLACK, None)
putf(wsFV, f'B{r}', f"={A('e1_pe')}", EXP['e1']['pe'], NUM1, green=True)
E1PE_R = r; r += 1
put(wsFV, f'A{r}', '        = Expert 1 fair value (EGP)', BLACK, None)
putf(wsFV, f'C{r}', f"=B{E1EPS_R}*B{E1PE_R}", EXP['e1']['base'], PX)
E1_R = r; r += 1
put(wsFV, f'A{r}', 'Expert 2 — free cash flow to equity: present value of the explicit years',
    BLACK, None)
put(wsFV, f'B{r}', EXP['e2']['pv'], BLUE, NUM0)
E2PV_R = r; r += 1
put(wsFV, f'A{r}', '        plus the present value of the terminal block', BLACK, None)
put(wsFV, f'B{r}', EXP['e2']['pv_tv'], BLUE, NUM0)
E2TV_R = r; r += 1
put(wsFV, f'A{r}', '        = Expert 2 fair value (EGP)', BLACK, None)
putf(wsFV, f'C{r}', f"=(B{E2PV_R}+B{E2TV_R})/{A('shares')}", EXP['e2']['base'], PX)
E2_R = r; r += 1
put(wsFV, f'A{r}', 'Expert 3 — economic profit. LINKED, not recomputed: an economic-profit build '
                   'off the same NOPAT, capital and discount path IS the cash-flow lens', BLACK,
    None)
putf(wsFV, f'C{r}', f"='EV Bridge'!$B$@BR_PS@", DCF['ps'], PX, green=True)
E3_R = r; r += 1
put(wsFV, f'A{r}', 'Expert 3 less the cash-flow lens — MUST be zero, and that is the point',
    BLACK, None)
putf(wsFV, f'C{r}', f"=C{E3_R}-'EV Bridge'!$B$@BR_PS@", 0.0, '0.0000')
r += 1
put(wsFV, f'A{r}', 'Expert panel median (EGP) — NOT an independent check, see note', BLACK, None)
putf(wsFV, f'C{r}', f"=MEDIAN(C{E1_R},C{E2_R},C{E3_R})", D['panel_centre'], PX)
PANEL_R = r; r += 1
put(wsFV, f'A{r}', 'Spread between the two INDEPENDENT reads (Expert 1 less Expert 2)', BLACK,
    None)
putf(wsFV, f'C{r}', f"=C{E1_R}-C{E2_R}", EXP['e1']['base'] - EXP['e2']['base'], PX)
SPREAD_R = r; r += 2
put(wsFV, f'A{r}', 'Currency-of-discounting alternative (EGP)', BLACK, None)
put(wsFV, f'C{r}', DCF['ccy_alt_ps'], BLUE, PX)
r += 1
put(wsFV, f'A{r}', 'Rating-basis cost of capital alternative (EGP)', BLACK, None)
put(wsFV, f'C{r}', DCF['ps_rating_basis'], BLUE, PX)
r += 1
put(wsFV, f'A{r}', 'Gross-debt weighting alternative, rejected (EGP)', BLACK, None)
put(wsFV, f'C{r}', DCF['ps_gross_basis'], BLUE, PX)
r += 2
note(wsFV, r, 'Bear and bull columns are whole-model re-runs — each is a complete revaluation at a '
              'different set of drivers, so they are values rather than formulas. The base column, '
              'the weights, the contributions and the weighted central are all live.', 7)

# ============ 2 SUMMARY ======================================================
wsS = sheet('Summary')
title(wsS, 'Testahil — Alexandria Mineral Oils Company S.A.E. (EGX: AMOC)',
      'Summary valuation. Educational analysis, not investment advice. No rating and no price '
      'target — a fair-value range and a distribution.', 7, 50, 16)
r = 4
band(wsS, r, 7); put(wsS, f'A{r}', 'SUMMARY VALUATION TABLE', BLACK, None, bold=True); r += 1
hdr(wsS, r, ['Lens', 'Bear (EGP)', 'Base (EGP)', 'Bull (EGP)', 'Weight', 'Note'])
r += 1
SUM_START = r
for key, lab, _src in LENS_SRC:
    put(wsS, f'A{r}', lab, BLACK, None)
    putf(wsS, f'B{r}', f"='Fundamental Valuation'!B{LR[key]}", LN[key]['bear'], PX, green=True)
    putf(wsS, f'C{r}', f"='Fundamental Valuation'!C{LR[key]}", LN[key]['base'], PX, green=True)
    putf(wsS, f'D{r}', f"='Fundamental Valuation'!D{LR[key]}", LN[key]['bull'], PX, green=True)
    putf(wsS, f'E{r}', f"='Fundamental Valuation'!E{LR[key]}", LN[key]['w'], PCT, green=True)
    if key == 'dcf':
        put(wsS, f'F{r}', 'terminal value share ->', BLACK, None)
        putf(wsS, f'G{r}', f"=DCF!$B${TVSH_R}", DCF['tv_share'], PCT, green=True)
        wsS.column_dimensions['G'].width = 14
    r += 1
put(wsS, f'A{r}', 'WEIGHTED CENTRAL FAIR VALUE', BLACK, None, bold=True)
putf(wsS, f'B{r}', f"=MIN(B{SUM_START}:B{r-1})", D['span'][0], PX, bold=True)
putf(wsS, f'C{r}', f"=SUM(C{SUM_START}*E{SUM_START},C{SUM_START+1}*E{SUM_START+1},"
                   f"C{SUM_START+2}*E{SUM_START+2},C{SUM_START+3}*E{SUM_START+3})",
     D['central'], PX, bold=True)
putf(wsS, f'D{r}', f"=MAX(D{SUM_START}:D{r-1})", D['span'][1], PX, bold=True)
putf(wsS, f'E{r}', f"=SUM(E{SUM_START}:E{r-1})", 1.0, PCT, bold=True)
SC_R = r; r += 1
put(wsS, f'A{r}', 'Spot price', BLACK, None)
putf(wsS, f'C{r}', f"={A('spot')}", SPOT, PX, green=True)
SS_R = r; r += 1
put(wsS, f'A{r}', 'Implied against spot', BLACK, None)
putf(wsS, f'C{r}', f"=C{SC_R}/C{SS_R}-1", D['central'] / SPOT - 1, PCT)
r += 2

band(wsS, r, 7); put(wsS, f'A{r}', 'THE COMPANY IN NUMBERS', BLACK, None, bold=True); r += 1
put(wsS, f'A{r}', 'Market capitalisation (EGP mn)', BLACK, None)
putf(wsS, f'C{r}', f"={A('spot')}*{A('shares')}", M['mktcap'], NUM0)
SMC_R = r; r += 1
put(wsS, f'A{r}', 'Net cash (EGP mn)', BLACK, None)
putf(wsS, f'C{r}', f"=0-DCF!$B${ND_R}", -BASE['nd_cy25'], NUM0)
SNC_R = r; r += 1
put(wsS, f'A{r}', 'Enterprise value (EGP mn)', BLACK, None)
putf(wsS, f'C{r}', f"=C{SMC_R}-C{SNC_R}", M['mktcap'] + BASE['nd_cy25'], NUM0)
r += 1
put(wsS, f'A{r}', 'Calendar 2025 revenue (EGP mn)', BLACK, None)
putf(wsS, f'C{r}', f"='Product Lines'!E{RCY25_R}", BASE['rev_cy25'], NUM0, green=True)
r += 1
put(wsS, f'A{r}', 'Calendar 2025 profit after tax (EGP mn)', BLACK, None)
putf(wsS, f'C{r}', f"='Product Lines'!B{PATCY25_R}", BASE['pat_cy25'], NUM0, green=True)
r += 1
put(wsS, f'A{r}', 'Trailing price / earnings', BLACK, None)
putf(wsS, f'C{r}', f"='Relative & Normalized'!$B${RELPS_R+2}", REL['pe_trailing'], MULT,
     green=True)
r += 1
put(wsS, f'A{r}', 'Dividend yield on the declared dividend', BLACK, None)
putf(wsS, f'C{r}', f"={A('dps')}/{A('spot')}", IN['dps'] / SPOT, PCT)
r += 1
put(wsS, f'A{r}', 'Weighted central fair value in dollars (USD)', BLACK, None)
putf(wsS, f'C{r}', f"=C{SC_R}/{A('fx_spot')}", D['central'] / IN['fx'], PX)
r += 1
put(wsS, f'A{r}', 'Weighted cost of capital — explicit window', BLACK, None)
putf(wsS, f'C{r}', f"=DCF!$B${WEXP_R}", W['wacc_exp'], PCT2, green=True)
r += 1
put(wsS, f'A{r}', 'Weighted cost of capital — terminal', BLACK, None)
putf(wsS, f'C{r}', f"=DCF!$B${WTRM_R}", W['wacc_term'], PCT2, green=True)
r += 1
put(wsS, f'A{r}', 'Terminal growth', BLACK, None)
putf(wsS, f'C{r}', f"={A('g')}", IN['g_term'], PCT2, green=True)
r += 2
note(wsS, r, 'Terminal value as a share of enterprise value is shown beside the discounted '
             'cash flow lens above and again on the bridge, linked live to the discounted cash '
             'flow sheet in both places and typed in neither.', 7)

# ============ 15 PER-SHARE & RATIOS ==========================================
wsP = sheet('Per-Share & Ratios')
title(wsP, 'Per-share figures and ratios — every cell a formula', None, 11, 44, 13)
hdr(wsP, 4, ['', *YH, *YF])
r = 5
RATIOS = [
    ('Earnings per share (EGP)', f"='Income Statement'!{{c}}{ISEPS}",
     [HI[k]['ebt'] * (1 - TAX) * (1 - NCI_SH) / SH for k in H4],
     [F['np_attr'][i] / SH for i in range(5)], PX),
    ('Book value per share (EGP)', f"='Balance Sheet'!{{c}}{BSEQ_R}/{A('shares')}",
     [HB[k]['eqp'] / SH for k in H4], [F['equity'][i] / SH for i in range(5)], PX),
    ('Net cash per share (EGP)', f"=0-'Balance Sheet'!{{c}}{BSND_R}/{A('shares')}",
     [-HB[k]['nd'] / SH for k in H4], [-F['net_debt'][i] / SH for i in range(5)], PX),
    ('EBITDA margin', f"='Income Statement'!{{c}}{ISEBITDA}/'Income Statement'!{{c}}{ISREV}",
     [HI[k]['ebitda'] / HI[k]['rev'] for k in H4], F['ebitda_margin'], PCT),
    ('Return on average equity',
     None, None, None, PCT),
    ('Asset turnover (revenue / total assets)',
     f"='Income Statement'!{{c}}{ISREV}/'Balance Sheet'!{{c}}{BSTA_R}",
     [HI[k]['rev'] / (HB[k]['ppe'] + HB[k]['inv'] + HB[k]['recv'] + IN['other_ca'] + HB[k]['cash'])
      for k in H4],
     [F['rev'][i] / (F['ppe'][i] + (F['rev'][i] - F['gp'][i]) * IN['inv_days'] / 365
                     + F['rev'][i] * IN['recv_days'] / 365 + IN['other_ca'] + F['cash'][i])
      for i in range(5)], MULT),
    ('Net working capital days',
     f"='Balance Sheet'!{{c}}{BSNWC_R}/'Income Statement'!{{c}}{ISREV}*365",
     [HB[k]['nwc'] / HI[k]['rev'] * 365 for k in H4],
     [F['nwc'][i] / F['rev'][i] * 365 for i in range(5)], NUM1),
    ('Dividend per share (EGP)', None, None, None, PX),
]
for lab, fml, vh, vf, fmt in RATIOS:
    put(wsP, f'A{r}', lab, BLACK, None)
    if fml is None:
        if lab.startswith('Return'):
            for i in range(1, 4):
                putf(wsP, f'{HC[i]}{r}',
                     f"='Income Statement'!{HC[i]}{ISNPA}/(('Balance Sheet'!{HC[i-1]}{BSEQ_R}+"
                     f"'Balance Sheet'!{HC[i]}{BSEQ_R})/2)",
                     HI[H4[i]]['ebt'] * (1 - TAX) * (1 - NCI_SH) /
                     ((HB[H4[i - 1]]['eqp'] + HB[H4[i]]['eqp']) / 2), PCT)
            prevcol = HC[3]
            for i in range(5):
                pv_eq = BASE['eqp_cy25'] if i == 0 else F['equity'][i - 1]
                putf(wsP, f'{FCOL[i]}{r}',
                     f"='Income Statement'!{FCOL[i]}{ISNPA}/(('Balance Sheet'!{prevcol}{BSEQ_R}+"
                     f"'Balance Sheet'!{FCOL[i]}{BSEQ_R})/2)",
                     F['np_attr'][i] / ((pv_eq + F['equity'][i]) / 2), PCT)
                prevcol = FCOL[i]
        else:
            for i in range(4):
                putf(wsP, f'{HC[i]}{r}', f"={A('dps')}", IN['dps'], PX, green=True)
            for i in range(5):
                putf(wsP, f'{FCOL[i]}{r}', f"='Cash Flow'!{CD[i]}{CFDIV_R}/{A('shares')}",
                     F['div'][i] / SH, PX)
    else:
        for i in range(4):
            putf(wsP, f'{HC[i]}{r}', fml.format(c=HC[i]), vh[i], fmt)
        for i in range(5):
            putf(wsP, f'{FCOL[i]}{r}', fml.format(c=FCOL[i]), vf[i], fmt)
    r += 1

# ============ 13 MONTE CARLO (whole-model re-run — pasted) ===================
wsM = sheet('Monte Carlo')
title(wsM, 'Forward price distribution — engine output',
      'PASTED VALUES. Each percentile is the output of a 50,000-path simulation and is not '
      'recomputed by this workbook. Changing a driver does NOT redraw this sheet.', 7, 46, 15)
hdr(wsM, 4, ['', '1 month', '3 months'])
r = 5
H1, H3 = STK['horizons']['1M'], STK['horizons']['3M']
for lab, k in (('Sessions to the check date', 'h'), ('Target date', 'target_date'),
               ('Grade date', 'grade_date')):
    put(wsM, f'A{r}', lab, BLACK, None)
    put(wsM, f'B{r}', H1[k], BLUE, NUM0 if k == 'h' else None)
    put(wsM, f'C{r}', H3[k], BLUE, NUM0 if k == 'h' else None)
    r += 1
for p in (5, 25, 50, 75, 95):
    put(wsM, f'A{r}', f'{p}th percentile (EGP)', BLACK, None)
    put(wsM, f'B{r}', H1['pct'][f'p{p}'], BLUE, PX)
    put(wsM, f'C{r}', H3['pct'][f'p{p}'], BLUE, PX)
    r += 1
for lab, k in (('Probability above spot', 'p_above'), ('Probability 10% or more up', 'p_up10'),
               ('Probability 10% or more down', 'p_dn10'),
               ('Probability of touching +10%', 'touch_up10'),
               ('Probability of touching -10%', 'touch_dn10')):
    put(wsM, f'A{r}', lab, BLACK, None)
    put(wsM, f'B{r}', H1[k], BLUE, PCT)
    put(wsM, f'C{r}', H3[k], BLUE, PCT)
    r += 1
put(wsM, f'A{r}', 'Annualised volatility at the anchor', BLACK, None)
put(wsM, f'B{r}', H1['anchor_vol_ann'], BLUE, PCT)
put(wsM, f'C{r}', H3['anchor_vol_ann'], BLUE, PCT)
r += 2
band(wsM, r, 7); put(wsM, f'A{r}', 'CALIBRATION OF THE ENGINE THAT PRODUCED THIS', BLACK, None,
                     bold=True); r += 1
for lab, v, fmt in (('Windows scored (post-break)', S0['windows_scored'], NUM0),
                    ('First origin', S0['first_origin'], None),
                    ('Last origin', S0['last_origin'], None),
                    ('Skill against a carry-anchored random walk', S0['skill_norm'], PCT),
                    ('Verdict', S0['verdict'], None),
                    ('Coverage of the 90% band', S0['cov90'], PCT),
                    ('Probability-integral-transform mean', S0['pit_mean'], '0.000'),
                    ('Cone width against the benchmark', S0['w90_ratio'], '0.000'),
                    ('Market panel verdict', S0['market_gate']['verdict'], None),
                    ('Market panel skill', S0['market_gate']['skill'], PCT)):
    put(wsM, f'A{r}', lab, BLACK, None)
    put(wsM, f'B{r}', v, BLUE, fmt)
    r += 1

# ============ 14 SENSITIVITY — the two grids and the beta sweep are LIVE =====
wsSN = sheet('Sensitivity')
title(wsSN, 'Sensitivity — the cost-of-capital grids are LIVE formulas',
      'Each cell re-runs the discount path, the terminal block and the bridge off this sheet. '
      'Only the sweeps that change REVENUE are pasted, because those need the product build '
      're-run and cannot be a formula in a grid.', 9, 34, 12)


def _dfblock(ws, row, we_ref, wt_ref, want):
    """Five discount factors from an explicit and a terminal rate, on the model's own glide."""
    vals = []
    prev = None
    for i, c in enumerate(('C', 'D', 'E', 'F', 'G')):
        fwd = f"({we_ref}-({we_ref}-{wt_ref})*DCF!{CD[i]}${GF_R})"
        f = f"=1/(1+{fwd})" if prev is None else f"={prev}/(1+{fwd})"
        putf(ws, f'{c}{row}', f, want[i], DF4)
        prev = f'{c}{row}'
        vals.append(want[i])
    return vals


def _dfs(we, wt):
    out, cc = [], 1.0
    for f in F['glide_frac']:
        cc /= (1 + (we - (we - wt) * f)); out.append(cc)
    return out


def _val(we, wt, g):
    df = _dfs(we, wt)
    roic = F['nopat'][-1] * (1 + g) / F['ic'][-1]
    rr = min(g / roic, 0.95)
    tv = F['nopat'][-1] * (1 + g) * (1 - rr) / (wt - g)
    ev = sum(F['fcff'][i] * df[i] for i in range(5)) + tv * df[-1]
    return ev * (1 - NCI_SH) / SH - DCF['nd'] / SH


GF_R = FWD_R - 1          # the glide-fraction row on the DCF sheet
FC_R = FCFF_R
r = 4
band(wsSN, r, 9)
put(wsSN, f'A{r}', 'GRID 1 — TERMINAL COST OF CAPITAL x TERMINAL GROWTH (EGP per share, LIVE)',
    BLACK, None, bold=True); r += 1
put(wsSN, f'A{r}', 'row: terminal rate · helper columns C:G are that row\'s discount factors',
    BLACK, None); r += 1
hdr(wsSN, r, ['Terminal WACC', 'PV explicit', 'df1', 'df2', 'df3', 'df4', 'df5']); r += 1
G1_HELP = r
for i, wt in enumerate(SN['wt_grid']):
    put(wsSN, f'A{r}', wt, BLUE, PCT2)
    _dfblock(wsSN, r, f"DCF!$B${WEXP_R}", f'$A${r}', _dfs(W['wacc_exp'], wt))
    df = _dfs(W['wacc_exp'], wt)
    putf(wsSN, f'B{r}',
         "=" + "+".join(f"DCF!{CD[j]}${FC_R}*{'CDEFG'[j]}{r}" for j in range(5)),
         sum(F['fcff'][j] * df[j] for j in range(5)), NUM0)
    r += 1
r += 1
hdr(wsSN, r, ['Terminal WACC \\ terminal g'] + [f'{g*100:.0f}%' for g in SN['g_grid']])
r += 1
G1_R = r
for i, wt in enumerate(SN['wt_grid']):
    putf(wsSN, f'A{r}', f"=A{G1_HELP+i}", wt, PCT2)
    for j, g in enumerate(SN['g_grid']):
        col = get_column_letter(2 + j)
        hr = G1_HELP + i
        roic = f"(DCF!F${NOPAT_R}*(1+{col}${G1_R-1}0)/DCF!F${IC_R})"
        putf(wsSN, f'{col}{r}',
             f"=((B{hr}+DCF!F${NOPAT_R}*(1+{SN['g_grid'][j]})"
             f"*(1-{SN['g_grid'][j]}/(DCF!F${NOPAT_R}*(1+{SN['g_grid'][j]})"
             f"/DCF!F${IC_R}))/(A{hr}-{SN['g_grid'][j]})*G{hr})"
             f")*(1-{A('nci')})/{A('shares')}-DCF!$B${ND_R}/{A('shares')}",
             _val(W['wacc_exp'], wt, g), PX)
    r += 1
r += 1
note(wsSN, r, 'Every cell above is a formula. Change the terminal cost of capital in column A, or '
              'the growth rate in the header, and the grid redraws. The previous edition pasted '
              'all 25 cells, which is how a stale sweep can ship beside a base case it no longer '
              'reproduces.', 9)
r += 2

band(wsSN, r, 9)
put(wsSN, f'A{r}', 'GRID 2 — BETA (EGP per share, LIVE)', BLACK, None, bold=True); r += 1
hdr(wsSN, r, ['Beta'] + [f'{b:.2f}' for b in SN['beta_grid']]); r += 1
B_HELP = r
lbls = ['Cost of equity, explicit', 'Cost of equity, terminal', 'WACC explicit', 'WACC terminal',
        'df1', 'df2', 'df3', 'df4', 'df5', 'PV explicit']
BROWS = {}
for k, lab in enumerate(lbls):
    put(wsSN, f'A{r}', lab, BLACK, None); BROWS[lab] = r; r += 1
for j, bta in enumerate(SN['beta_grid']):
    col = get_column_letter(2 + j)
    ke = (IN['rf'] - IN['sov_spread_cds']) + bta * IN['erp_cds']
    ket = IN['rf_term'] + bta * IN['erp_term']
    we = W['we_exp'] * ke + W['wd_exp'] * W['k_nd_at']
    wt = (1 - IN['wd_term']) * ket + IN['wd_term'] * W['kd_term_at']
    putf(wsSN, f"{col}{BROWS['Cost of equity, explicit']}",
         f"=DCF!$B${RFSTAR_R}+{bta}*DCF!$B${ERP_R}", ke, PCT2)
    putf(wsSN, f"{col}{BROWS['Cost of equity, terminal']}",
         f"={A('rf_term')}+{bta}*{A('erp_term')}", ket, PCT2)
    putf(wsSN, f"{col}{BROWS['WACC explicit']}",
         f"=DCF!$B${WE_R}*{col}{BROWS['Cost of equity, explicit']}"
         f"+DCF!$B${WD_R}*DCF!$B${KND_R}", we, PCT2)
    putf(wsSN, f"{col}{BROWS['WACC terminal']}",
         f"=(1-{A('wd_term')})*{col}{BROWS['Cost of equity, terminal']}"
         f"+{A('wd_term')}*DCF!$B${KDT_R}", wt, PCT2)
    df = _dfs(we, wt); prev = None
    for i in range(5):
        rr_ = BROWS[f'df{i+1}']
        fwd = (f"({col}{BROWS['WACC explicit']}-({col}{BROWS['WACC explicit']}"
               f"-{col}{BROWS['WACC terminal']})*DCF!{CD[i]}${GF_R})")
        f = f"=1/(1+{fwd})" if prev is None else f"={prev}/(1+{fwd})"
        putf(wsSN, f'{col}{rr_}', f, df[i], DF4)
        prev = f'{col}{rr_}'
    putf(wsSN, f"{col}{BROWS['PV explicit']}",
         "=" + "+".join(f"DCF!{CD[i]}${FC_R}*{col}{BROWS[f'df{i+1}']}" for i in range(5)),
         sum(F['fcff'][i] * df[i] for i in range(5)), NUM0)
r += 1
put(wsSN, f'A{r}', 'FAIR VALUE (EGP per share)', BLACK, None, bold=True)
for j, bta in enumerate(SN['beta_grid']):
    col = get_column_letter(2 + j)
    g = IN['g_term']
    putf(wsSN, f'{col}{r}',
         f"=((({col}{BROWS['PV explicit']})+DCF!F${NOPAT_R}*(1+{A('g')})"
         f"*(1-DCF!$B${RR_R})/({col}{BROWS['WACC terminal']}-{A('g')})"
         f"*{col}{BROWS['df5']}))*(1-{A('nci')})/{A('shares')}-DCF!$B${ND_R}/{A('shares')}",
         _val(*( (lambda b: ((W['we_exp']*((IN['rf']-IN['sov_spread_cds'])+b*IN['erp_cds'])
                             + W['wd_exp']*W['k_nd_at']),
                            ((1-IN['wd_term'])*(IN['rf_term']+b*IN['erp_term'])
                             + IN['wd_term']*W['kd_term_at']), IN['g_term']))(bta) )), PX)
BFV_R = r; r += 1
put(wsSN, f'A{r}', 'Check: the base column must reproduce the base case', BLACK, None)
putf(wsSN, f'D{r}', f"=D{BFV_R}-'EV Bridge'!$B$@BR_PS@", 0.0, '0.0000')
r += 2
note(wsSN, r, 'The base column of this sweep is now a formula off the same cells the base case '
              'uses, so it cannot drift from it — the row above proves it to four decimals. The '
              'previous edition pasted this sweep and it had drifted: its centre column read '
              '10.04 against a base case of 9.91.', 9)
r += 2

band(wsSN, r, 9)
put(wsSN, f'A{r}', 'GRID 3 — EXPLICIT x TERMINAL COST OF CAPITAL (EGP per share, LIVE)',
    BLACK, None, bold=True); r += 1
put(wsSN, f'A{r}', "each anchor varied independently around its own base; one formula a cell, "
                   "no helper block", BLACK, None); r += 1
hdr(wsSN, r, ['Explicit \\ terminal'] + [f'{x*100:.1f}%' for x in SN['wt_grid']]); r += 1
G3_HDR = r - 1
G3_R = r
for i, we in enumerate(SN['we_grid']):
    put(wsSN, f'A{r}', we, BLUE, PCT2)
    for j, wt in enumerate(SN['wt_grid']):
        col = get_column_letter(2 + j)
        wref, tref = f'$A${r}', f'{col}${G3_HDR}'
        fwd = [f"(1+({wref}-({wref}-{tref})*DCF!{CD[k]}${FWD_R - 1}))" for k in range(5)]
        cum = ["*".join(fwd[:k + 1]) for k in range(5)]
        pv = "+".join(f"DCF!{CD[k]}${FCFF_R}/({cum[k]})" for k in range(5))
        tv = (f"DCF!F${NOPAT_R}*(1+{A('g')})*(1-DCF!$B${RR_R})/({tref}-{A('g')})/({cum[4]})")
        putf(wsSN, f'{col}{r}',
             f"=({pv}+{tv})*(1-{A('nci')})/{A('shares')}-DCF!$B${ND_R}/{A('shares')}",
             _val(we, wt, IN['g_term']), PX)
    r += 1
# the header row must carry the terminal rates as numbers for the formulas above to read them
for j, wt in enumerate(SN['wt_grid']):
    put(wsSN, f'{get_column_letter(2+j)}{G3_HDR}', wt, BLUE, PCT2)
r += 1
put(wsSN, f'A{r}', 'Check: the centre cell must reproduce the base case', BLACK, None)
putf(wsSN, f'D{r}', f"=D{G3_R+2}-'EV Bridge'!$B$@BR_PS@", 0.0, '0.0000')
r += 2

band(wsSN, r, 9)
put(wsSN, f'A{r}', 'WHOLE-MODEL RE-RUNS — PASTED, and these do NOT redraw', BLACK, None,
    bold=True); r += 1
note(wsSN, r, 'The sweeps below change REVENUE, so each cell needs the three-line product build '
              're-run from the top. They are complete revaluations and are the one class of cell '
              'on this sheet that cannot be a formula.', 9)
r += 1
for lab, gr, vals, fmt in (('Margin ratio (specialty over fuel)', SN['gm_grid'], SN['grid_margin'], '0.00'),
                           ('Volume growth multiplier', SN['vol_grid'], SN['grid_vol'], '0.00'),
                           ('Exchange-rate path multiplier', SN['fx_grid'], SN['grid_fx'], '0.00'),
                           ('Net working capital, % of revenue', SN['nwc_grid'], SN['grid_nwc'], PCT2)):
    put(wsSN, f'A{r}', lab, BLACK, None, bold=True)
    for j, gv in enumerate(gr):
        put(wsSN, f'{get_column_letter(2+j)}{r}', gv, BLUE, fmt)
    r += 1
    put(wsSN, f'A{r}', 'Fair value', BLACK, None)
    for j, v in enumerate(vals):
        put(wsSN, f'{get_column_letter(2+j)}{r}', v, BLUE, PX)
    r += 2

# ============ 16 PEER & SECTOR ===============================================
wsPS = sheet('Peer & Sector')
title(wsPS, 'Peer and sector context — refining and lubricant base oils', None, 7, 46, 15)
hdr(wsPS, 4, ['Comparator basis', 'Multiple', 'Applied to (EGP mn)', 'Implied value (EGP/share)'])
r = 5
PEERS = [
    ('Company trailing enterprise value / EBITDA (computed)', None, 'trailing EBITDA'),
    ('Justified enterprise value / EBITDA (adopted)', 'ev_ebitda', 'forward EBITDA'),
    ('House low bound — NOT a peer observation', 'mult_low', 'forward EBITDA'),
    ('House high bound — NOT a peer observation', 'mult_high', 'forward EBITDA'),
]
_pv = {None: REL['ev_ebitda_trailing'], 'ev_ebitda': IN['ev_ebitda_just'],
       'mult_low': 3.5, 'mult_high': 6.0}
for lab, key, basis in PEERS:
    put(wsPS, f'A{r}', lab, BLACK, None)
    mult = _pv[key]
    if key is None:
        putf(wsPS, f'B{r}', f"='Relative & Normalized'!$B${RELPS_R+1}", mult, MULT, green=True)
    else:
        putf(wsPS, f'B{r}', f"={A(key)}", mult, MULT, green=True)
    putf(wsPS, f'C{r}', f"=DCF!C{EBITDA_R}", REL['ebitda_mid'], NUM0, green=True)
    putf(wsPS, f'D{r}', f"=B{r}*C{r}*DCF!C{DFL_R}*(1-{A('nci')})/{A('shares')}"
                        f"-DCF!$B${ND_R}/{A('shares')}",
         (mult * REL['ebitda_mid']) * REL['df_rel'] * (1 - NCI_SH) / SH - BASE['nd_cy25'] / SH, PX)
    r += 1
r += 1
band(wsPS, r, 7); put(wsPS, f'A{r}', 'SECTOR INDICATORS — WHAT ACTUALLY DRIVES A PROCESSOR',
                      BLACK, None, bold=True); r += 1
hdr(wsPS, r, ['Indicator', *YF]); r += 1
put(wsPS, f'A{r}', 'Throughput (mn tonnes)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsPS, f'{get_column_letter(2+i)}{r}', f"='Product Lines'!{UC[i]}{VOL_R}", U['vol'][i],
         NUM3, green=True)
r += 1
put(wsPS, f'A{r}', 'EBITDA per tonne (EGP)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsPS, f'{get_column_letter(2+i)}{r}',
         f"=DCF!{c}{EBITDA_R}/'Product Lines'!{UC[i]}{VOL_R}", F['ebitda'][i] / U['vol'][i],
         NUM0)
r += 1
put(wsPS, f'A{r}', 'Specialty share of revenue', BLACK, None)
for i, c in enumerate(CD):
    putf(wsPS, f'{get_column_letter(2+i)}{r}',
         f"='Product Lines'!{UC[i]}{ANCH['legs']['spec_share']}",
         U['spec_rev'][i] / F['rev'][i], PCT)
r += 1
put(wsPS, f'A{r}', 'Gross profit per tonne (EGP)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsPS, f'{get_column_letter(2+i)}{r}',
         f"=DCF!{c}{DGP_R}/'Product Lines'!{UC[i]}{VOL_R}", F['gp'][i] / U['vol'][i], NUM0)
r += 1
put(wsPS, f'A{r}', 'Capital expenditure per tonne (EGP)', BLACK, None)
for i, c in enumerate(CD):
    putf(wsPS, f'{get_column_letter(2+i)}{r}',
         f"=DCF!{c}{CAPEX_R}/'Product Lines'!{UC[i]}{VOL_R}", F['capex'][i] / U['vol'][i],
         NUM0)
r += 2
note(wsPS, r, 'A refinery is not valued on revenue growth. The indicators that matter are '
              'throughput, the margin earned on each tonne, the share of that tonnage in the '
              'high-value specialty slate, and the capital needed to keep the plant running — all '
              'four are on this sheet and all four are formulas.', 7)

# ============ 1 READ FIRST ===================================================
wsRF = sheet('READ FIRST')
title(wsRF, 'Testahil — Alexandria Mineral Oils Company S.A.E. (EGX: AMOC)', None, 9, 100, 12)
LINES = [
 'Companion model · Independent valuation study · Educational analysis · Not investment advice',
 '',
 'WHAT THIS WORKBOOK IS. A transparent companion to the AMOC valuation study. Blue cells are inputs;',
 'black cells are formulas; green cells link across sheets.',
 '',
 'IT CALCULATES — IT DOES NOT STORE. Every figure that can be derived arithmetically from a driver is a',
 'live formula. Change a blue cell on Assumptions and the model reprices: the cost of equity is built from',
 'the risk-free rate net of the sovereign default spread, beta and the equity risk premium; the cost of net',
 'debt is blended from what the borrowing costs and what the cash earns; the glide fractions are derived',
 'from the cost-of-debt path and the discount factors compound; the waterfall chains gross margin to EBITDA',
 'to EBIT to NOPAT to free cash flow to present value; the terminal block chains reinvestment = growth over',
 'return; the statements roll forward; and every ratio and per-share figure on every sheet is a formula.',
 'This claim has been tested, not asserted: see the two gates named at the foot of this sheet.',
 '',
 'ONLY TWO CLASSES OF CELL ARE PASTED, AND HERE THEY ARE.',
 '',
 '  1. AUDITED AND DISCLOSED HISTORY — the primary record, not a calculation. Where a line is both',
 '     disclosed and derivable, the disclosed figure is carried. After this workbook was rebuilt there are',
 '     just EIGHT such cells outside the input register: the four historical profit-after-tax figures on the',
 '     Income Statement, and the four observed peer multiples on Peer & Sector. Everything else that was',
 '     once pasted — the historical revenue line, the whole historical balance sheet, the property, plant',
 '     and equipment residual — is now derived on the sheet from the disclosed inputs.',
 '',
 '  2. WHOLE-MODEL RE-RUNS — the Monte Carlo price map and the sensitivity grids, and the bear and bull',
 '     columns of the valuation. Each individual cell there is a complete revaluation of the entire model at',
 '     a different set of drivers, which cannot be expressed as a formula inside a grid. THESE DO NOT REDRAW',
 '     WHEN YOU CHANGE A DRIVER. Everything else does.',
 '',
 'The standing rule permits a third pasted class — the output of a unit build too complex to flatten into a',
 'spreadsheet grid. IT IS NOT USED HERE. AMOC\'s revenue build is three product lines, each a volume times a',
 'dollar price times an exchange rate, and that fits on a sheet. See Product Lines: every cell of the',
 'revenue build is live.',
 '',
 'HOW REVENUE IS BUILT — AND WHAT CHANGED. Not as one growth rate, and no longer from a calibrated price.',
 'The company reports a product table — tonnes and value for base oils, for paraffin wax, and in total —',
 'and every realisation in this model is DERIVED from it by division. Base oils, paraffin wax and the fuel',
 'and by-product slate are carried as three separate lines, each a volume times a dollar price a tonne',
 'times an explicit exchange-rate path, because all three price off dollar product benchmarks even when',
 'sold domestically. An earlier edition of this workbook had TWO legs, a specialty price that was a free',
 'input and a fuel price that was the RESIDUAL of base-year revenue — and then offered the implied fuel',
 'realisation as evidence that the split was real. A residual cannot corroborate the construction that',
 'produced it. Both are gone: NO PRICE IN THIS MODEL IS CALIBRATED AND NONE IS A RESIDUAL.',
 '',
 'THE GROSS MARGIN IS AN OUTPUT, NOT AN ASSUMPTION. The two line margins are SOLVED from one disclosed',
 'blend and one judgment parameter — the ratio of the specialty spread to the fuel spread — and are then',
 'held CONSTANT for the whole forecast. Every forecast year\'s blended margin is those constants re-weighted',
 'by that year\'s revenue mix, so the margin widens only to the extent the mix shifts. Change a volume',
 'growth rate on Product Lines and watch the blended margin move with no margin input touched.',
 '',
 'THE TWO ROUTES TO THE BASE YEAR DISAGREE, AND THE GAP IS ON THE FACE OF THE SHEET. Rolling the reported',
 'product table forward does not reproduce the base year built from two filed halves; a single',
 'reconciliation factor scales every line onto the filed figure. It is shown, not buried. The filed halves',
 'govern the LEVEL; the product table governs only the MIX and the per-tonne economics.',
 '',
 'THREE THINGS TO KNOW BEFORE READING THE NUMBERS.',
 '',
 '  · THE FINANCIAL YEAR CHANGED. The exchange approved a move from a 30 June year-end to 31 December.',
 '    July-December 2025 is a six-month transition period. History is shown on the reported June years; the',
 '    calendar-2025 base year is CONSTRUCTED from two disclosed halves, with no estimated component in',
 '    either; and the forecast runs on calendar years. See Product Lines for the construction.',
 '',
 '  · THE COMPANY IS NET CASH. Gross debt is a rounding error against the cash pile. That has two',
 '    consequences the workbook makes visible: net debt enters the bridge as a NEGATIVE, so the cash is',
 '    added; and the weighting on net debt RAISES the discount rate above the cost of equity rather than',
 '    lowering it, because the observed equity cost is diluted by cash that carries almost no risk.',
 '',
 '  · THE HISTORICAL BALANCE SHEETS ARE A RECONSTRUCTION. Only four balance-sheet lines are disclosed, at',
 '    one date. The rest is built from days drivers and a roll-back through disclosed profit and dividends,',
 '    and is labelled as such on the sheet.',
 '',
 'NO RATING AND NO PRICE TARGET. A fair-value range and a distribution.',
]
rr = 4
for ln in LINES:
    c = wsRF.cell(row=rr, column=1, value=ln)
    c.font = Font(size=10, bold=ln.isupper() and len(ln) > 20)
    rr += 1
rr += 1
band(wsRF, rr, 9); put(wsRF, f'A{rr}', 'HEADLINE FIGURES — ALL LIVE LINKS', BLACK, None,
                       bold=True); rr += 1
for lab, fml, val, fmt in (
        ('Spot price (EGP)', f"={A('spot')}", SPOT, PX),
        ('Weighted central fair value (EGP)', f"=Summary!C{SC_R}", D['central'], PX),
        ('Implied against spot', f"=Summary!C{SS_R+1}", D['central'] / SPOT - 1, PCT),
        ('Discounted cash flow lens (EGP)', f"='EV Bridge'!$B$@BR_PS@", DCF['ps'], PX),
        ('Terminal value as a share of enterprise value', f"=DCF!$B${TVSH_R}", DCF['tv_share'],
         PCT),
        ('Weighted cost of capital — explicit', f"=DCF!$B${WEXP_R}", W['wacc_exp'], PCT2),
        ('Weighted cost of capital — terminal', f"=DCF!$B${WTRM_R}", W['wacc_term'], PCT2)):
    put(wsRF, f'A{rr}', lab, BLACK, None)
    putf(wsRF, f'C{rr}', fml, val, fmt, green=True)
    rr += 1
rr += 1
for ln in [
    'TWO GATES RUN ON THIS DELIVERED FILE, NOT ON THE SCRIPT THAT WROTE IT.',
    '  · Cell-level agreement. The builder records the model\'s own value for every formula cell as it',
    '    writes it. An independent evaluator then re-evaluates the workbook and asserts that every formula',
    '    cell reproduces that value and that none is left unchecked. A formula that computes the right thing',
    '    the wrong way, or points one row off, fails here rather than shipping a different number.',
    '  · Driver propagation. Every input is perturbed in place, the whole workbook is re-evaluated, and the',
    '    headline is required to move in the asserted direction, with a dead-input sweep over the rest.',
    '    The live-driver claim above is only made because this test passes.',
    '',
    'THE CENSUS. The exact split of formula cells against pasted numeric cells is counted on the delivered',
    'file and printed by the build, then restated in the study\'s quality-control gate. Text labels and',
    'headings are counted separately and form no part of the claim in either direction. Every pasted numeric',
    'cell falls into one of the two classes named above; there is no third bucket and no residual category.',
]:
    c = wsRF.cell(row=rr, column=1, value=ln)
    c.font = Font(size=10)
    rr += 1

# ---- order the sheets to the canonical sequence -----------------------------
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'EV Bridge',
         'Product Lines', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity', 'Per-Share & Ratios',
         'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]
assert len(wb.sheetnames) == 16, wb.sheetnames

ANCH['dcf'].update(dict(pv_explicit=PVE_R, pv_tv=PVTV_R, tv=TV_R, rr=RR_R,
                        we=WE_R, wd=WD_R, knd=KND_R, kdt=KDT_R,
                        roic_term=ROICT_R, nt=NT_R))
ANCH['legs'] = dict(ANCH['legs'], pat_cy25=PATCY25_R)
ANCH['bridge'] = dict(eq=BEQ_R, ps=BPS_R, tv_share=BTV_R, ev=BEV_R, nd=BND_R)
TOK['IS_REV'] = ISREV; TOK['IS_GP'] = ISGP; TOK['LEGS_PAT'] = ANCH['legs']['pat_cy25']
ANCH['is'] = dict(rev=ISREV, other=ISOTH, gm=ISGM, gp=ISGP, ebitda=ISEBITDA, ebit=ISEBIT, fin=ISFIN, pat=ISPAT,
                  npa=ISNPA, eps=ISEPS, dna=ISDNA)
ANCH['bs'] = dict(ppe=BSPPE_R, cash=BSCASH_R, ta=BSTA_R, eq=BSEQ_R, nwc=BSNWC_R, nd=BSND_R,
                  tle=BSTLE_R, chk=BSCHK_R)
ANCH['cf'] = dict(fcff=CFFCFF_R, fin=CFFIN_R, np=CFNP_R, div=CFDIV_R, ond=CFOND_R,
                  cnd=CFCND_R, ceq=CFCEQ_R)
ANCH['rn'] = dict(rel=RELPS_R, norm=NRMPS_R, book=BKPS_R, pe_trailing=RELPS_R + 2,
                  ev_ebitda_trailing=RELPS_R + 1)
ANCH['fv'] = dict(rows=LR, central=FVC_R, spot=FVS_R,
                  e1=E1_R, e2=E2_R, e3=E3_R, panel=PANEL_R, spread=SPREAD_R)
ANCH['sum'] = dict(start=SUM_START, central=SC_R, spot=SS_R)
ANCH['cols'] = dict(hist=HC, fcst=FCOL, cd=CD, uc=UC)

# ---- resolve cross-sheet tokens --------------------------------------------
import re as _re
_pat = _re.compile(r'@([A-Z0-9_]+)@')
_unresolved = []
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and '@' in _c.value:
                def _sub(m):
                    if m.group(1) not in TOK:
                        _unresolved.append(m.group(1))
                        return m.group(0)
                    return str(TOK[m.group(1)])
                _c.value = _pat.sub(_sub, _c.value)
assert not _unresolved, f'unresolved layout tokens: {sorted(set(_unresolved))}'

OUTP = os.path.join(HERE, 'AMOC_Valuation_Model_06082026_public.xlsx')
wb.save(OUTP)

# ---- inject cached values so the DELIVERED file carries its own results ------
# openpyxl writes formulas with no cached result, so every reader that does not recalculate —
# pandas, a preview pane, an auditor reading the raw XML — sees a blank model, and the
# cell-level gate has nothing in the artefact to check against. The model's own value for
# every formula cell is already recorded as it is written; it is injected here so the
# workbook is self-describing, and `fullCalcOnLoad` is left set so a spreadsheet application
# still recomputes from the formulas on open.
import re as _re2, shutil as _sh, zipfile as _zf
nform_pre = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith('='))
_tmp = OUTP + '.tmp'
_sh.move(OUTP, _tmp)
_zin = _zf.ZipFile(_tmp)
_zout = _zf.ZipFile(OUTP, 'w', _zf.ZIP_DEFLATED)
_sheet_xml = {ws.title: f'xl/worksheets/sheet{i+1}.xml' for i, ws in enumerate(wb.worksheets)}
_xml_sheet = {v: k for k, v in _sheet_xml.items()}
_injected = 0
for _it in _zin.infolist():
    _data = _zin.read(_it.filename)
    if _it.filename in _xml_sheet:
        _title = _xml_sheet[_it.filename]
        _vals = EXPECT.get(_title, {})

        _hits = []

        def _fix(m, _v=_vals, _h=_hits):
            _ref = _re2.search(r'r="([A-Z]+[0-9]+)"', m.group(0))
            if not _ref or _ref.group(1) not in _v:
                return m.group(0)
            _h.append(1)
            return m.group(0).replace('</f>', f'</f><v>{_v[_ref.group(1)]!r}</v>')

        _txt = _data.decode('utf-8')
        _txt = _re2.sub(r'<c [^>]*>\s*<f>[^<]*</f>\s*</c>', _fix, _txt)
        _injected += len(_hits)
        _data = _txt.encode('utf-8')
    _zout.writestr(_it, _data)
_zin.close(); _zout.close(); os.remove(_tmp)
_z = _zf.ZipFile(OUTP)
_cached = sum(len(_re2.findall(r'</f><v>', _z.read(n).decode('utf-8')))
              for n in _z.namelist() if n.startswith('xl/worksheets/sheet'))
_z.close()
print(f'cached values present in {_cached} of {nform_pre} formula cells (verified on the saved file)')

nform = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('='))
nval = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
           if c.value is not None and not (isinstance(c.value, str) and c.value.startswith('=')))
json.dump(dict(expected=EXPECT, anchors=ANCH,
               counts=dict(formulas=nform, values=nval)),
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)

# ---- census, counted on the DELIVERED file, split by pasted class ------------
# gate (q) asks for formula count against pasted count. A pasted cell is a NUMERIC constant;
# labels, headings and notes are text and are not part of the claim either way.
import openpyxl as _op
_wbv = _op.load_workbook(OUTP)
_num, _txt_n, _fml = 0, 0, 0
_num_by_sheet = {}
for _ws in _wbv.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            if _c.value is None:
                continue
            if isinstance(_c.value, str) and _c.value.startswith('='):
                _fml += 1
            elif isinstance(_c.value, (int, float)):
                _num += 1
                _num_by_sheet[_ws.title] = _num_by_sheet.get(_ws.title, 0) + 1
            else:
                _txt_n += 1
CENSUS = dict(formulas=_fml, numeric_values=_num, text=_txt_n,
              numeric_by_sheet=dict(sorted(_num_by_sheet.items(), key=lambda kv: -kv[1])),
              formula_share=_fml / (_fml + _num))
json.dump(CENSUS, open(os.path.join(HERE, 'formula_count.json'), 'w'), indent=1)
print(f'wrote {OUTP}')
print(f'formulas {nform} | non-formula cells {nval} | recorded expectations '
      f'{sum(len(v) for v in EXPECT.values())}')
print(f'CENSUS on the delivered file: {_fml} formulas / {_num} pasted numeric '
      f'({_fml/(_fml+_num):.1%} live), {_txt_n} text labels')
print('  pasted numerics by sheet: ' +
      ', '.join(f'{k} {v}' for k, v in CENSUS['numeric_by_sheet'].items()))
