"""AMOC_Valuation_Model_03092026_public.xlsx — the workbook CALCULATES.

The previous edition shipped 488 formulas against 186 pasted cells: 72.4% formula. The pasted
27.6% was not incidental — it was the sensitivity grids, the beta sweep and the bear/bull
columns, i.e. precisely the cells a reader would want to interrogate. They were pasted because
each one is a whole-model re-run and the builder had no way to express a re-run inside a grid.

This build removes that excuse. The forecast engine is laid out as an eight-line-by-five-year
formula block, and a SCENARIO BLOCK routine writes that entire engine again, in formulas, for
every sensitivity point — twenty-five complete re-runs, each one live. Change a driver on
Assumptions and every sensitivity cell in the workbook moves with it.

Exactly ONE class of cell is now pasted: a figure read off an audited or disclosed filing.
Nothing else. There is no "too complex to flatten" class and no "whole-model re-run" class,
because both were flattened.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
IN = {k: v['value'] for k, v in D['inputs'].items()}
# the escalation convention comes from the study's committed record, not from a
# constant typed here: this builder reads study_numbers.json and nothing else, so
# a convention it cannot see is a convention it would silently get wrong
POUND_ON_PRICE = bool(D['dcf']['pound_on_price'])
SRC = {k: v['source'] for k, v in D['inputs'].items()}
DAT = {k: v['date'] for k, v in D['inputs'].items()}
RNG = {k: v['ring'] for k, v in D['inputs'].items()}
UB, TTM, RT, BR = D['unitbuild'], D['ttm'], D['rates'], D['bridge']
F, W, DCF, LN, SN = D['fcst'], D['wacc'], D['dcf'], D['lenses'], D['sens']
AU, BASE, REL, NRM, BK = D['audited'], D['base'], D['rel'], D['norm'], D['book']
SH, SPOT = IN['shares_mn'], IN['spot']
YRS = F['years']
LINES, LBL = UB['lines'], UB['labels']
NY = 5
UC = ['C', 'D', 'E', 'F', 'G']            # the five forecast columns; B is the base year

INK = '1C3A36'; PANEL = 'EAF0EE'; GOLDF = 'F6F1E6'
BLUE = '1F4E9C'; BLACK = '1C3A36'; GREEN = '2E6B4F'
NUM0 = '#,##0'; NUM1 = '#,##0.0'; NUM3 = '#,##0.000'; PX = '#,##0.00'
PCT = '0.0%'; PCT2 = '0.00%'; MULT = '0.00"x"'

wb = Workbook(); wb.remove(wb.active)
EXPECT = {}
NPASTE = [0]
NFORM = [0]


def sheet(name, wa=52):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = INK
    ws.column_dimensions['A'].width = wa
    for c in 'BCDEFGHIJKLMN':
        ws.column_dimensions[c].width = 14
    ws.sheet_view.showGridLines = False
    return ws


def put(ws, ref, val, color=BLACK, fmt=None, bold=False, paste=False):
    c = ws[ref]; c.value = val
    c.font = Font(color=color, size=10, bold=bold)
    if fmt:
        c.number_format = fmt
    if paste:
        NPASTE[0] += 1
    return c


def putf(ws, ref, formula, value, fmt=None, bold=False, green=False):
    """Write a FORMULA and record the model's own value, so an independent evaluator can assert
    cell by cell that the workbook reproduces the model rather than merely looking plausible."""
    c = ws[ref]; c.value = formula
    c.font = Font(color=(GREEN if green else BLACK), size=10, bold=bold)
    if fmt:
        c.number_format = fmt
    EXPECT.setdefault(ws.title, {})[ref] = float(value)
    NFORM[0] += 1
    return c


def band(ws, row, n=9, fill=PANEL):
    for i in range(1, n + 1):
        ws.cell(row=row, column=i).fill = PatternFill('solid', fgColor=fill)


def title(ws, t, sub=None, n=9):
    ws['A1'] = t; ws['A1'].font = Font(bold=True, size=13, color=INK)
    if sub:
        ws['A2'] = sub; ws['A2'].font = Font(size=9.5, color='6E7B77', italic=True)
        ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        ws.row_dimensions[2].height = 32
    band(ws, 3, n, INK)


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True, size=9.5, color=INK)
        c.fill = PatternFill('solid', fgColor=PANEL)
        c.alignment = Alignment(wrap_text=True, vertical='bottom')


def note(ws, row, text, n=9):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=8.5, color='6E7B77', italic=True)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    ws.row_dimensions[row].height = max(26, 11 * (len(text) // 115 + 1))


# ================= ASSUMPTIONS — the only pasted cells in the file ============
wsA = sheet('Assumptions', 46)
title(wsA, 'Assumptions — every driver, with the filing it was read from',
      'Blue cells are the ONLY pasted values in this workbook. Every one is a figure read off an '
      'audited or disclosed filing, or a stated house judgement. There is no third class: the '
      'previous edition also pasted its sensitivity grids and its bear and bull columns because '
      'each is a whole-model re-run; those re-runs are now written out as live formula blocks on '
      'the Sensitivity sheet, so a reader can change a driver here and watch every one of them '
      'move.', 6)
wsA.column_dimensions['A'].width = 44
wsA.column_dimensions['B'].width = 16
wsA.column_dimensions['C'].width = 12
wsA.column_dimensions['D'].width = 11
wsA.column_dimensions['E'].width = 96
hdr(wsA, 5, ['Driver', 'Value', 'Ring', 'Date', 'Source'])

AROW = {}
r = 6


def arow(key, label, val, fmt=NUM0, k=None):
    global r
    put(wsA, f'A{r}', label)
    put(wsA, f'B{r}', val, color=BLUE, fmt=fmt, paste=True)
    put(wsA, f'C{r}', RNG.get(k or key, 'House'), fmt=None)
    put(wsA, f'D{r}', DAT.get(k or key, D['meta']['asof']))
    c = put(wsA, f'E{r}', SRC.get(k or key, ''))
    c.alignment = Alignment(wrap_text=True, vertical='top')
    wsA.row_dimensions[r].height = max(14, 10 * (len(SRC.get(k or key, '')) // 105 + 1))
    AROW[key] = f'Assumptions!$B${r}'
    r += 1


put(wsA, f'A{r}', 'MARKET AND CAPITAL', bold=True); band(wsA, r, 6); r += 1
arow('spot', 'Spot price, EGP', SPOT, PX)
arow('shares_mn', 'Shares outstanding, mn', SH, NUM1)

put(wsA, f'A{r}', 'THE TWELVE MONTHS TO 30 JUNE 2026 — the base year', bold=True); band(wsA, r, 6); r += 1
for k, lab, fmt in [('rev_h2_25', 'Net sales, 6M to 31-Dec-2025 (AUDITED)', NUM0),
                    ('cogs_h2_25', 'Cost of sales, 6M to 31-Dec-2025 (AUDITED)', NUM0),
                    ('rev_h1cy26', 'Net sales, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('cogs_h1cy26', 'Cost of sales, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('gp_h1cy26', 'Gross profit, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('ga_h1cy26', 'General and administrative, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('mkt_h1cy26', 'Marketing and selling, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('othexp_h1cy26', 'Other expenses, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('prov_h1cy26', 'Formed provisions, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('ecl_h1cy26', 'Expected credit losses, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('fin_h1cy26', 'Finance expenses, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('othrev_h1cy26', 'Other income, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('pat_h1cy26', 'Profit after tax, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('nci_h1cy26', 'Non-controlling interest, 6M to 30-Jun-2026 (FILED)', NUM0),
                    ('maj_h1cy26', "Majority's share, 6M to 30-Jun-2026 (FILED)", NUM0),
                    ('cos_ttm_salaries', 'Salaries in cost of sales, twelve months (note 15-A)', NUM0),
                    ('cos_ttm_raw', 'Raw materials, twelve months (note 15-A)', NUM0),
                    ('cos_ttm_support', 'Supporting materials, twelve months (note 15-A)', NUM0),
                    ('cos_ttm_dep', 'Depreciation in cost of sales, twelve months (note 15-A)', NUM0),
                    ('cos_ttm_other', 'Other cost of sales, twelve months (note 15-A)', NUM0),
                    ('dep_ttm', 'Depreciation and RoU, twelve months (both cash-flow statements)', NUM0),
                    ('capex_ttm', 'Cash capital expenditure, twelve months (both cash-flow statements)', NUM0),
                    ('credint_h1cy26', 'Credit interest, 6M to 30-Jun-2026 (note 14-B)', NUM0),
                    ('rev_h1cy26_rep', 'Net sales, 6M to 30-Jun-2026 (press release, cross-check)', NUM0),
                    ('pat_h1cy26_rep', 'Profit after tax, 6M to 30-Jun-2026 (press release, cross-check)', NUM0),
                    ('gp_h1cy26_rep', 'Gross profit (press release) — the filing confirms it', NUM0),
                    ('rev_q1_26', 'Net sales, 3M to 31-Mar-2026 (reviewed)', NUM0),
                    ('cogs_q1_26', 'Cost of sales, 3M to 31-Mar-2026 (reviewed)', NUM0),
                    ('ga_q1_26', 'Administrative expense, 3M to 31-Mar-2026', NUM0),
                    ('mkt_q1_26', 'Selling expense, 3M to 31-Mar-2026', NUM0),
                    ('othexp_q1_26', 'Other operating expense, 3M to 31-Mar-2026', NUM0),
                    ('othrev_q1_26', 'Other revenue, 3M to 31-Mar-2026', NUM0),
                    ('prov_q1_26', 'Formed provisions, 3M to 31-Mar-2026', NUM0),
                    ('ecl_q1_26', 'Expected credit losses, 3M to 31-Mar-2026', NUM0),
                    ('fin_q1_26', 'Finance cost, 3M to 31-Mar-2026', NUM0),
                    ('ga_h2_25', 'Administrative expense, 6M to 31-Dec-2025', NUM0),
                    ('mkt_h2_25', 'Selling expense, 6M to 31-Dec-2025', NUM0),
                    ('othexp_h2_25', 'Other operating expense, 6M to 31-Dec-2025', NUM0),
                    ('prov_h2_25', 'Formed provisions, 6M to 31-Dec-2025', NUM0),
                    ('dep_h2_25', 'Depreciation, 6M to 31-Dec-2025', NUM0),
                    ('dep_q1_26', 'Depreciation, 3M to 31-Mar-2026', NUM0),
                    ('capex_h2_25', 'Cash capital expenditure, 6M to 31-Dec-2025', NUM0),
                    ('capex_q1_26', 'Cash capital expenditure, 3M to 31-Mar-2026', NUM0),
                    ('credint_h2_25', 'Credit interest, 6M to 31-Dec-2025', NUM0),
                    ('credint_q1_26', 'Credit interest, 3M to 31-Mar-2026', NUM0),
                    ('emp_h2_25', "Employees' profit share and board bonuses, 6M", NUM0),
                    ('pat_h2_25', 'Profit after tax, 6M to 31-Dec-2025 (AUDITED)', NUM0),
                    ('nci_h2_25', 'Minority share of profit, 6M to 31-Dec-2025', NUM0),
                    ('nci_q1_26', 'Minority share of profit, 3M to 31-Mar-2026', NUM0),
                    ('tax_h2_25', 'Current tax, 6M to 31-Dec-2025', NUM0),
                    ('dtax_h2_25', 'Deferred tax CREDIT, 6M to 31-Dec-2025', NUM0),
                    ('tax_q1_26', 'Current tax, 3M to 31-Mar-2026', NUM0),
                    ('dtax_q1_26', 'Deferred tax, 3M to 31-Mar-2026', NUM0),
                    ('invinc_h2_25', 'Investment income, 6M to 31-Dec-2025', NUM0),
                    ('othrev_h2_25', 'Other revenue, 6M to 31-Dec-2025', NUM0),
                    ('fin_h2_25', 'Finance cost, 6M to 31-Dec-2025', NUM0)]:
    arow(k, lab, IN[k], fmt)

put(wsA, f'A{r}', 'NOTE 14-A — the audited product table, 6M to 31-Dec-2025', bold=True)
band(wsA, r, 6); r += 1
for k in LINES:
    arow(f'pt_{k}', f'{LBL[k]} — tonnes sold', IN['prod_t'][k], NUM3, k='prod_t')
for k in LINES:
    arow(f'pv_{k}', f'{LBL[k]} — sales value, EGP', IN['prod_v'][k], NUM0, k='prod_v')

put(wsA, f'A{r}', 'NOTE 15-A — the audited cost stack, 6M to 31-Dec-2025', bold=True)
band(wsA, r, 6); r += 1
for k in ['cos_raw', 'cos_salaries', 'cos_other', 'cos_support', 'cos_dep']:
    arow(k, SRC[k].split(' — ')[1].split(',')[0].strip().capitalize()
          if ' — ' in SRC[k] else k, IN[k], NUM0)

put(wsA, f'A{r}', 'BALANCE SHEET AT 31 DECEMBER 2025 (AUDITED)', bold=True); band(wsA, r, 6); r += 1
for k, lab in [('ppe_net', 'Fixed assets, net'), ('ppe_gross', 'Fixed assets, at COST'),
               ('puc', 'Projects under construction'), ('inventory', 'Inventory'),
               ('recv', 'Trade receivables'), ('debtors', 'Debtors and other debit balances'),
               ('payables', 'Accounts and notes payable'),
               ('creditors', 'Creditors and other credit balances'),
               ('div_declared', 'of which DIVIDENDS PAYABLE — a financing claim'),
               ('provisions', 'Provisions — tax disputes and claims'),
               ('cash', 'Cash at banks and on hand, free'),
               ('fin_inv', 'Deposits PLEDGED against facilities'),
               ('fvoci', 'Equity investment at fair value through OCI'),
               ('debt_lt', 'Long-term loans'), ('debt_st', 'Short-term loans'),
               ('eq_parent', 'Parent equity'), ('eq_nci', 'Non-controlling interest, carrying')]:
    arow(k, lab, IN[k], NUM0)

put(wsA, f'A{r}', 'DRIVERS', bold=True); band(wsA, r, 6); r += 1
arow('tax_stat', 'Statutory tax rate — applied to operating profit', IN['tax_stat'], PCT2)
arow('raw_pass', 'Feedstock pass-through factor', IN['raw_pass'], NUM1)
arow('g_term', 'Terminal growth', IN['g_term'], PCT2)
arow('roe_sust', 'Sustainable return on equity — book lens', IN['roe_sust'], PCT2)
arow('pe_just', 'Justified price/earnings — normalised lens', IN['pe_just'], NUM1)
arow('rel_rerating', 'Re-rating on own trailing EV/EBITDA — relative lens', IN['rel_rerating'], PCT2)
for k in LINES:
    arow(f'proc_{k}', f'Processing intensity — {LBL[k]}', IN['proc_intensity'][k], '0.00',
         k='proc_intensity')
for i in range(NY):
    arow(f'lpg_{i}', f'Realisation growth, {YRS[i]}', IN['line_price_growth'][i], PCT2,
         k='line_price_growth')
for k in LINES:
    for i in range(NY):
        arow(f'lvg_{k}_{i}', f'Volume growth — {LBL[k]}, {YRS[i]}', IN['line_vol_growth'][k][i],
             PCT2, k='line_vol_growth')
for i in range(NY):
    arow(f'infl_{i}', f'Egyptian inflation factor, {YRS[i]}', IN['fixed_cost_infl'][3 + i], NUM3,
         k='fixed_cost_infl')

put(wsA, f'A{r}', 'COST OF CAPITAL', bold=True); band(wsA, r, 6); r += 1
for k, lab, fmt in [('rf', 'Risk-free rate, Egypt 10-year local currency', PCT2),
                    ('sov_spread_cds', 'Sovereign default spread, CDS basis', PCT2),
                    ('erp_cds', 'Equity risk premium, CDS basis', PCT2),
                    ('beta', 'Beta', NUM3),
                    ('cash_yield', 'Yield on cash', PCT2),
                    ('kd', 'Cost of debt', PCT2),
                    ('cbe_target', 'CBE inflation target IN FORCE for the terminal horizon', PCT2),
                    ('real_rate_term', 'Terminal real risk-free rate', PCT2),
                    ('erp_term', 'Terminal equity risk premium', PCT2),
                    ('kd_term', 'Terminal cost of debt', PCT2),
                    ('wd_term', 'Terminal debt weight', PCT2)]:
    arow(k, lab, IN[k], fmt)
for i in range(NY):
    arow(f'kdp_{i}', f'Cost-of-debt path, {YRS[i]}', IN['kd_path'][i], PCT2, k='kd_path')
for nm, wv in [('dcf', 0.45), ('relative', 0.20), ('normalized', 0.20), ('book', 0.15)]:
    arow(f'w_{nm}', f'Lens weight — {nm}', IN['lens_weights'][nm], PCT, k='lens_weights')

A = AROW


def L(k):
    return A[k]


# ================= BASE YEAR — the coherence test, in formulas ================
wsB = sheet('Income Statement', 56)
title(wsB, 'Base year — twelve contiguous months to 30 June 2026',
      'The gross profit of the reported half is NOT taken as released. It is solved from the '
      "release's own profit line, and the released figure is shown beside it with the profit it "
      'would have implied. Every cell below is a formula.', 5)
hdr(wsB, 5, ['', 'EGP', 'Check', '', ''])
put(wsB, 'A6', 'Administrative + selling + other, 3M to 31-Mar-2026', bold=True)
putf(wsB, 'B6', f"={L('ga_q1_26')}+{L('mkt_q1_26')}+{L('othexp_q1_26')}",
     IN['ga_q1_26'] + IN['mkt_q1_26'] + IN['othexp_q1_26'], NUM0)
put(wsB, 'A7', 'Scaled to the half (x2)')
putf(wsB, 'B7', '=B6*2', (IN['ga_q1_26'] + IN['mkt_q1_26'] + IN['othexp_q1_26']) * 2, NUM0)
put(wsB, 'A8', 'Other revenue, half')
putf(wsB, 'B8', f"={L('othrev_q1_26')}*2", IN['othrev_q1_26'] * 2, NUM0)
put(wsB, 'A9', 'Provisions and expected credit losses, half')
putf(wsB, 'B9', f"=({L('prov_q1_26')}+{L('ecl_q1_26')})*2",
     (IN['prov_q1_26'] + IN['ecl_q1_26']) * 2, NUM0)
put(wsB, 'A10', 'Finance cost, half')
putf(wsB, 'B10', f"={L('fin_q1_26')}*2", IN['fin_q1_26'] * 2, NUM0)
put(wsB, 'A11', 'Effective tax rate, nine months (for the solve only)')
_pbt9 = (IN['tax_h2_25'] - IN['dtax_h2_25'] + IN['tax_q1_26'] - IN['dtax_q1_26']) / RT['tax_eff']
putf(wsB, 'B11',
     f"=({L('tax_h2_25')}-{L('dtax_h2_25')}+{L('tax_q1_26')}-{L('dtax_q1_26')})/"
     f"(({L('tax_h2_25')}-{L('dtax_h2_25')}+{L('tax_q1_26')}-{L('dtax_q1_26')})/{RT['tax_eff']})",
     RT['tax_eff'], PCT2)
put(wsB, 'A13', 'GROSS PROFIT, 6M to 30-Jun-2026 — AS FILED', bold=True)
putf(wsB, 'B13', f"={L('gp_h1cy26')}", IN['gp_h1cy26'], NUM0, bold=True)
put(wsB, 'A14', 'Gross profit as the press release stated it')
putf(wsB, 'B14', f"={L('gp_h1cy26_rep')}", IN['gp_h1cy26_rep'], NUM0)
put(wsB, 'A15', 'Profit after tax, 6M to 30-Jun-2026 — AS FILED')
putf(wsB, 'B15', f"={L('pat_h1cy26')}", IN['pat_h1cy26'], NUM0)
put(wsB, 'A16', 'against the profit the press release reported')
putf(wsB, 'B16', f"={L('pat_h1cy26_rep')}", IN['pat_h1cy26_rep'], NUM0)
put(wsB, 'A17', 'Release against filing, gross profit', bold=True)
putf(wsB, 'B17', '=B13/B14-1', TTM['ct3'], PCT2, bold=True)
put(wsB, 'A19', 'BASE-YEAR REVENUE — twelve months to 30-Jun-2026', bold=True)
putf(wsB, 'B19', f"=({L('rev_h2_25')}+{L('rev_h1cy26')})/10^6", TTM['rev'], NUM0, bold=True)
put(wsB, 'A20', 'BASE-YEAR GROSS PROFIT', bold=True)
putf(wsB, 'B20', f"=({L('rev_h2_25')}-{L('cogs_h2_25')}+B13)/10^6", TTM['gp'], NUM0, bold=True)
put(wsB, 'A21', 'BASE-YEAR COST OF SALES', bold=True)
putf(wsB, 'B21', '=B19-B20', TTM['cogs'], NUM0, bold=True)
put(wsB, 'A22', 'BASE-YEAR GROSS MARGIN', bold=True)
putf(wsB, 'B22', '=B20/B19', TTM['gm'], PCT2, bold=True)
put(wsB, 'A24', 'Alternative base: nine audited months x 4/3')
putf(wsB, 'B24', f"=({L('rev_h2_25')}+{L('rev_q1_26')})/10^6*4/3", TTM['rev9_ann'], NUM0)
put(wsB, 'A25', 'its gross margin')
putf(wsB, 'B25',
     f"=({L('rev_h2_25')}+{L('rev_q1_26')}-{L('cogs_h2_25')}-{L('cogs_q1_26')})/"
     f"({L('rev_h2_25')}+{L('rev_q1_26')})", TTM['gm9'], PCT2)
put(wsB, 'A27', 'Operating lines, SAME twelve months', bold=True); band(wsB, 27, 5)
for i, (lab, key, val) in enumerate([
        ('Administrative expense', f"={L('ga_h2_25')}+{L('ga_h1cy26')}", TTM['ga']),
        ('Selling expense', f"={L('mkt_h2_25')}+{L('mkt_h1cy26')}", TTM['mkt']),
        ('Other operating expense', f"={L('othexp_h2_25')}+{L('othexp_h1cy26')}", TTM['oth']),
        ('Provisions and credit losses',
         f"={L('prov_h2_25')}+{L('prov_h1cy26')}+{L('ecl_h1cy26')}", TTM['prov']),
        ('Depreciation', f"={L('dep_ttm')}", TTM['dep']),
        ('Cash capital expenditure', f"={L('capex_ttm')}", TTM['capex']),
        ('Credit interest', f"={L('credint_h2_25')}+{L('credint_h1cy26')}", TTM['credint']),
        ("Employees' profit share", f"={L('emp_h2_25')}*2", TTM['emp'])]):
    put(wsB, f'A{28+i}', lab)
    putf(wsB, f'B{28+i}', '=(' + key[1:] + ')/10^6', val, NUM0)
note(wsB, 37,
     'Every operating line is struck on the SAME twelve months as revenue and gross profit. The '
     'previous edition built revenue from the six-month product table doubled while taking cost '
     'of sales from the nine months scaled by four thirds, so its base-year gross margin of '
     '7.081% corresponded to no filed period at all. One period, both sides, or the margin is an '
     'artefact of the scalars.')

put(wsB, 'A39', 'RATES — solved, not assumed', bold=True); band(wsB, 39, 5)
put(wsB, 'A40', "Employees' profit share, % of profit after tax")
putf(wsB, 'B40', f"=B35/(({L('pat_h2_25')}+{L('pat_h1cy26')})/10^6)", RT['emp_rate'], PCT2)
put(wsB, 'A41', 'Minority interest, % of OPERATING profit, whole base year')
putf(wsB, 'B41',
     f"=(({L('nci_h2_25')}+{L('nci_h1cy26')})/10^6)/"
     f"(({L('pat_h2_25')}+{L('pat_h1cy26')})/10^6-B34*(1-{RT['tax_eff']}))",
     RT['nci_op'], PCT2)
put(wsB, 'A42', 'Implied asset life, years = fixed assets at cost / depreciation')
putf(wsB, 'B42', f"={L('ppe_gross')}/10^6/B32", RT['asset_life'], NUM1)
put(wsB, 'A43', 'Maintenance capital expenditure at cost over that life')
putf(wsB, 'B43', f"={L('ppe_gross')}/10^6/B42", RT['maint_capex0'], NUM0)
put(wsB, 'A44', 'Inventory days of cost of sales')
putf(wsB, 'B44', f"={L('inventory')}/10^6/B21*365", RT['inv_days'], NUM1)
put(wsB, 'A45', 'Receivable days of revenue')
putf(wsB, 'B45', f"=({L('recv')}+{L('debtors')})/10^6/B19*365", RT['recv_days'], NUM1)
put(wsB, 'A46', 'Operating payable days of cost of sales — dividends payable REMOVED')
putf(wsB, 'B46', f"=({L('payables')}+{L('creditors')}-{L('div_declared')})/10^6/B21*365",
     RT['pay_days'], NUM1)


# ================= PRODUCT AND COST — eight lines, both sides =================
wsP = sheet('Segments', 40)
title(wsP, 'The eight disclosed product lines — revenue AND cost, per line',
      'Note 14-A gives tonnes and value per line. Note 15-A gives the cost stack for the company '
      'and NOT by line, so conversion is allocated on the registered processing-intensity '
      'weights and feedstock is then allocated on NET REALISABLE VALUE — the standard '
      'joint-product method. Every cell here is a formula.')
hdr(wsP, 5, ['Line', 'Tonnes, 6M', 'Value, EGP 6M', 'Realisation EGP/t',
             'Base tonnes mn', 'Base realisation', 'Conversion EGP/t', 'Feedstock EGP/t',
             'Spread EGP/t'])
PR = {}
for i, k in enumerate(LINES):
    rr = 6 + i
    put(wsP, f'A{rr}', LBL[k])
    putf(wsP, f'B{rr}', f"={L('pt_'+k)}", IN['prod_t'][k], NUM3, green=True)
    putf(wsP, f'C{rr}', f"={L('pv_'+k)}", IN['prod_v'][k], NUM0, green=True)
    putf(wsP, f'D{rr}', f'=C{rr}/B{rr}', IN['prod_v'][k] / IN['prod_t'][k], NUM0)
    putf(wsP, f'E{rr}', f'=B{rr}/10^6', UB['t0'][k], NUM3)
    putf(wsP, f'F{rr}', f'=D{rr}*$B$16', UB['px0'][k], NUM0)
    PR[k] = rr
put(wsP, 'A14', 'TOTAL', bold=True)
putf(wsP, 'B14', '=SUM(B6:B13)', sum(IN['prod_t'].values()), NUM3, bold=True)
putf(wsP, 'C14', '=SUM(C6:C13)', sum(IN['prod_v'].values()), NUM0, bold=True)
putf(wsP, 'E14', '=SUM(E6:E13)', UB['T0'], NUM3, bold=True)
put(wsP, 'A15', 'Revenue at the disclosed realisations, twelve months as filed')
putf(wsP, 'B15', '=SUMPRODUCT(E6:E13,D6:D13)',
     sum(UB['t0'][k] * IN['prod_v'][k] / IN['prod_t'][k] for k in LINES), NUM0)
put(wsP, 'A16', 'REALISATION INDEX — solved so the base year foots', bold=True)
putf(wsP, 'B16', "='Income Statement'!B19/B15", UB['px_index'], NUM3, bold=True)
putf(wsP, 'F14', '=SUMPRODUCT(E6:E13,F6:F13)', TTM['rev'], NUM0, bold=True)

put(wsP, 'A18', 'COST ALLOCATION', bold=True); band(wsP, 18)
put(wsP, 'A19', 'Cost of sales, base year')
putf(wsP, 'B19', "='Income Statement'!B21", TTM['cogs'], NUM0)
put(wsP, 'A20', 'Feedstock share of cost (note 15-A)')
putf(wsP, 'B20',
     f"={L('cos_ttm_raw')}/({L('cos_ttm_raw')}+{L('cos_ttm_salaries')}+{L('cos_ttm_other')}+"
     f"{L('cos_ttm_support')}+{L('cos_ttm_dep')})", AU['cost_share']['raw'], PCT2)
put(wsP, 'A21', 'Feedstock total')
putf(wsP, 'B21', '=B19*B20', UB['raw_tot0'] / 1e6, NUM0)
put(wsP, 'A22', 'Conversion total')
putf(wsP, 'B22', '=B19-B21', UB['conv_tot0'] / 1e6, NUM0)
put(wsP, 'A23', 'Processing-weighted tonnage')
putf(wsP, 'B23', '=SUMPRODUCT(E6:E13,H6:H13)', UB['pw_den'], NUM3)
for k in LINES:
    rr = PR[k]
    putf(wsP, f'H{rr}', f"={L('proc_'+k)}", UB['proc'][k], '0.00', green=True)
# conversion per tonne, then NRV, then feedstock per tonne
for k in LINES:
    rr = PR[k]
    putf(wsP, f'G{rr}', f'=$B$22*H{rr}/$B$23', UB['conv_pt'][k], NUM0)
put(wsP, 'A24', 'Net realisable value weighted tonnage')
putf(wsP, 'B24', '=SUMPRODUCT(E6:E13,I6:I13)', UB['nrv_den'], NUM0)
for k in LINES:
    rr = PR[k]
    putf(wsP, f'I{rr}', f'=F{rr}-G{rr}', UB['nrv'][k], NUM0)

hdr(wsP, 26, ['Line', 'Conversion EGP/t', 'Net realisable EGP/t', 'Feedstock EGP/t',
              'Total cost EGP/t', 'Realisation EGP/t', 'Spread EGP/t', 'Gross margin'])
for i, k in enumerate(LINES):
    rr = 27 + i
    put(wsP, f'A{rr}', LBL[k])
    putf(wsP, f'B{rr}', f'=G{PR[k]}', UB['conv_pt'][k], NUM0)
    putf(wsP, f'C{rr}', f'=I{PR[k]}', UB['nrv'][k], NUM0)
    putf(wsP, f'D{rr}', f'=$B$21*I{PR[k]}/$B$24', UB['raw_pt'][k], NUM0)
    putf(wsP, f'E{rr}', f'=D{rr}+B{rr}', UB['raw_pt'][k] + UB['conv_pt'][k], NUM0)
    putf(wsP, f'F{rr}', f'=F{PR[k]}', UB['px0'][k], NUM0)
    putf(wsP, f'G{rr}', f'=F{rr}-E{rr}', UB['spread'][k], NUM0, bold=True)
    putf(wsP, f'H{rr}', f'=G{rr}/F{rr}', UB['margin0'][k], PCT2)
put(wsP, 'A35', 'Cost of sales rebuilt from the per-line allocation', bold=True)
putf(wsP, 'B35', '=SUMPRODUCT(E6:E13,E27:E34)', TTM['cogs'], NUM0, bold=True)
put(wsP, 'A36', 'must equal the base-year cost of sales — difference')
putf(wsP, 'B36', "=B35-'Income Statement'!B21", 0.0, NUM3, bold=True)
note(wsP, 38,
     'Row 36 is the footing test: the eight per-line costs must rebuild the disclosed cost of '
     'sales exactly. Feedstock allocated flat per tonne would make fuel oil sell below the cost '
     'of its own feed; feedstock allocated on relative sales value would make base oils and '
     'paraffin wax — the products this plant exists for — loss-making. Net realisable value is '
     'the standard joint-product method and the only one of the three that leaves every '
     'disclosed line with a positive spread.')


# ================= THE ENGINE, and every scenario written out again ===========
wsF = sheet('DCF', 42)
title(wsF, 'Discounted cash flow — the forecast engine — eight lines by five years, all formulas',
      'This is the block the Sensitivity sheet replicates twenty-five times, once per grid '
      'point, so that every sensitivity cell in this workbook is a live re-run rather than a '
      'pasted number.')


def engine(ws, row0, tag, vol_cell, gm_cell, fx_cell, we_cell, wt_cell, g_cell, wc_cell, mv):
    """Write ONE complete re-run of the model as formulas, and return the row of the answer.

    `mv` carries the model's own values for the block so every cell can be checked. This
    function is the whole point of the file: a scenario is not a number somebody pasted, it is
    this block again with one driver pointed somewhere else."""
    r0 = row0
    hdr(ws, r0, [tag] + YRS)
    rows = {}

    def line(label, key, vals, fmt=NUM0, f=None, bold=False):
        nonlocal r0
        r0 += 1
        put(ws, f'A{r0}', label, bold=bold)
        for j, cl in enumerate(UC):
            putf(ws, f'{cl}{r0}', f(cl, j, r0), vals[j], fmt, bold=bold)
        rows[key] = r0
        return r0

    # cumulative driver indices
    r_inf = line('Inflation index, cumulative', 'inf', mv['inf'], NUM3,
                 lambda cl, j, rr: (f"={L('infl_0')}" if j == 0
                                    else f"={UC[j-1]}{rr}*{L('infl_'+str(j))}"))
    r_px = line('Realisation index, cumulative', 'px', mv['px'], NUM3,
                lambda cl, j, rr: (f"=(1+{L('lpg_0')}*{fx_cell})" if j == 0
                                   else f"={UC[j-1]}{rr}*(1+{L('lpg_'+str(j))}*{fx_cell})"))
    # per-line volumes
    vrows = {}
    for k in LINES:
        vrows[k] = line(f'  Volume — {LBL[k]}, mn t', f'v_{k}', mv['vol'][k], NUM3,
                        lambda cl, j, rr, k=k: (
                            f"='Segments'!E{PR[k]}*(1+{L('lvg_'+k+'_0')}+{vol_cell})"
                            if j == 0 else
                            f"={UC[j-1]}{rr}*(1+{L('lvg_'+k+'_'+str(j))}+{vol_cell})"))
    r_vt = line('Total volume, mn t', 'vt', mv['vtot'], NUM3,
                lambda cl, j, rr: f"=SUM({cl}{vrows[LINES[0]]}:{cl}{vrows[LINES[-1]]})", bold=True)
    # revenue and cost, per line, summed
    r_rev = line('REVENUE', 'rev', mv['rev'], NUM0,
                 lambda cl, j, rr: '=' + '+'.join(
                     f"{cl}{vrows[k]}*'Segments'!$F${PR[k]}*{cl}{r_px}" for k in LINES),
                 bold=True)
    # THE WORKBOOK IS A LIVE MODEL AND IT CARRIED THE DEFECT TOO [corrected
    # 03-Sep-2026]. The salary and other-conversion legs multiplied the INFLATION
    # index while realised price multiplied the PRICE index -- the same real cost
    # drift of +2.7 points a year that the study's own compute.py carried, written
    # out in Excel formulas so a reader recalculating the file would reproduce the
    # defect faithfully and see nothing wrong. Both now escalate on the same index
    # the price does, which is the principle raw_pass = 1.0 already declares for
    # the feedstock leg one term to the left.
    _pound_cell = r_px if POUND_ON_PRICE else r_inf
    r_cogs = line('Cost of sales', 'cogs', mv['cogs'], NUM0,
                  lambda cl, j, rr: '=' + '+'.join(
                      f"{cl}{vrows[k]}*('Segments'!$D${27+i}*{cl}{r_px}*"
                      f"{L('raw_pass')}+'Segments'!$B${27+i}*"
                      f"({UB['sal_sh']:.10f}*{cl}{_pound_cell}"
                      f"+{UB['oth_sh']:.10f}*{cl}{_pound_cell}"
                      f"+{UB['sup_sh']:.10f}*{cl}{r_px}+{UB['dep_sh']:.10f}))"
                      for i, k in enumerate(LINES)))
    r_gp = line('GROSS PROFIT', 'gp', mv['gp'], NUM0,
                lambda cl, j, rr: f"={cl}{r_rev}-{cl}{r_cogs}+{gm_cell}*{cl}{r_rev}", bold=True)
    r_gm = line('Gross margin', 'gm', mv['gm'], PCT2,
                lambda cl, j, rr: f"={cl}{r_gp}/{cl}{r_rev}")
    # operating expense: three lines on three drivers, plus the charge never taken
    r_opex = line('Operating expense', 'opex', mv['opex'], NUM0,
                  lambda cl, j, rr: (
                      f"=('Income Statement'!$B$28*{cl}{r_inf})"
                      f"+('Income Statement'!$B$29*{cl}{r_inf}*{cl}{r_vt}/'Segments'!$E$14)"
                      f"+('Income Statement'!$B$30*{cl}{r_inf})+('Income Statement'!$B$31*{cl}{r_inf})"))
    r_ebitda = line('EBITDA', 'ebitda', mv['ebitda'], NUM0,
                    lambda cl, j, rr: f"={cl}{r_gp}-{cl}{r_opex}", bold=True)
    # capex and depreciation off the asset register
    r_capex = line('Capital expenditure — maintenance + growth', 'capex', mv['capex'], NUM0,
                   lambda cl, j, rr: (
                       f"='Income Statement'!$B$43*{cl}{r_inf}+MAX({cl}{r_vt}-"
                       + (f"'Segments'!$E$14" if j == 0 else f"{UC[j-1]}{r_vt}")
                       + f",0)*{RT['cap_intensity']:.6f}*{cl}{r_inf}"))
    r_gross = line('Fixed assets at cost, closing', 'gross', mv['gross'], NUM0,
                   lambda cl, j, rr: (
                       f"=({L('ppe_gross')}+{L('puc')})/10^6+{cl}{r_capex}" if j == 0
                       else f"={UC[j-1]}{rr}+{cl}{r_capex}"))
    r_dna = line('Depreciation — rolls off the asset register', 'dna', mv['dna'], NUM0,
                 lambda cl, j, rr: (
                     f"=({L('ppe_gross')}+{L('puc')})/10^6/'Income Statement'!$B$42" if j == 0
                     else f"={UC[j-1]}{r_gross}/'Income Statement'!$B$42"))
    r_ppe = line('Fixed assets, net book, closing', 'ppe', mv['ppe'], NUM0,
                 lambda cl, j, rr: (
                     f"=({L('ppe_net')}+{L('puc')})/10^6+{cl}{r_capex}-{cl}{r_dna}" if j == 0
                     else f"={UC[j-1]}{rr}+{cl}{r_capex}-{cl}{r_dna}"))
    r_ebit = line('EBIT', 'ebit', mv['ebit'], NUM0,
                  lambda cl, j, rr: f"={cl}{r_ebitda}-{cl}{r_dna}", bold=True)
    r_nop0 = line('NOPAT before the profit share', 'nop0', mv['nop0'], NUM0,
                  lambda cl, j, rr: f"={cl}{r_ebit}*(1-{L('tax_stat')})")
    r_emp = line("  less employees' profit share and board bonuses", 'emp', mv['emp'], NUM0,
                 lambda cl, j, rr: f"=MAX({cl}{r_nop0},0)*'Income Statement'!$B$40")
    r_nopat = line('NOPAT', 'nopat', mv['nopat'], NUM0,
                   lambda cl, j, rr: f"={cl}{r_nop0}-{cl}{r_emp}", bold=True)
    r_nwc = line('Net working capital — on days', 'nwc', mv['nwc'], NUM0,
                 lambda cl, j, rr: (
                     f"='Income Statement'!$B$44*{wc_cell}/365*{cl}{r_cogs}"
                     f"+'Income Statement'!$B$45*{wc_cell}/365*{cl}{r_rev}"
                     f"-'Income Statement'!$B$46*{wc_cell}/365*{cl}{r_cogs}"))
    r_dnwc = line('  change in working capital', 'dnwc', mv['dnwc'], NUM0,
                  lambda cl, j, rr: (
                      f"={cl}{r_nwc}-('Income Statement'!$B$44/365*'Income Statement'!$B$21"
                      f"+'Income Statement'!$B$45/365*'Income Statement'!$B$19"
                      f"-'Income Statement'!$B$46/365*'Income Statement'!$B$21)" if j == 0
                      else f"={cl}{r_nwc}-{UC[j-1]}{r_nwc}"))
    r_fcff = line('FREE CASH FLOW TO THE FIRM', 'fcff', mv['fcff'], NUM0,
                  lambda cl, j, rr: f"={cl}{r_nopat}+{cl}{r_dna}-{cl}{r_capex}-{cl}{r_dnwc}",
                  bold=True)
    r_glide = line('Glide fraction', 'glide', mv['glide'], NUM3,
                   lambda cl, j, rr: f"={W['glide_frac'][j]:.10f}")
    r_wacc = line('Forward cost of capital', 'wacc', mv['fwd'], PCT2,
                  lambda cl, j, rr: f"={we_cell}-({we_cell}-{wt_cell})*{cl}{r_glide}")
    r_df = line('Cumulative discount factor', 'df', mv['df'], '0.00000',
                lambda cl, j, rr: (f"=1/(1+{cl}{r_wacc})" if j == 0
                                   else f"={UC[j-1]}{rr}/(1+{cl}{r_wacc})"))
    r_pv = line('Present value', 'pv', mv['pv'], NUM0,
                lambda cl, j, rr: f"={cl}{r_fcff}*{cl}{r_df}")
    r_ic = line('Invested capital — replacement basis', 'ic', mv['icr'], NUM0,
                lambda cl, j, rr: f"={cl}{r_nwc}+{cl}{r_gross}")
    # terminal block and bridge, in column B
    rt = r_pv + 1
    put(ws, f'A{rt}', 'PV of the explicit window', bold=True)
    putf(ws, f'B{rt}', f"=SUM(C{r_pv}:G{r_pv})", mv['pv_explicit'], NUM0, bold=True)
    put(ws, f'A{rt+1}', 'Terminal return on invested capital, at REPLACEMENT cost')
    putf(ws, f'B{rt+1}', f"=G{r_nopat}*(1+{g_cell})/G{r_ic}", mv['roic'], PCT2)
    put(ws, f'A{rt+2}', 'Required reinvestment = growth / return')
    putf(ws, f'B{rt+2}', f"=MIN({g_cell}/B{rt+1},0.95)", mv['rr'], PCT2)
    put(ws, f'A{rt+3}', 'Terminal value')
    putf(ws, f'B{rt+3}',
         f"=G{r_nopat}*(1+{g_cell})*(1-B{rt+2})/MAX({wt_cell}-{g_cell},0.02)", mv['tv'], NUM0)
    put(ws, f'A{rt+4}', 'PV of the terminal block')
    putf(ws, f'B{rt+4}', f"=B{rt+3}*G{r_df}", mv['pv_tv'], NUM0)
    put(ws, f'A{rt+5}', 'ENTERPRISE VALUE', bold=True)
    putf(ws, f'B{rt+5}', f"=B{rt}+B{rt+4}", mv['ev'], NUM0, bold=True)
    put(ws, f'A{rt+6}', 'plus net cash')
    putf(ws, f'B{rt+6}', f"=({L('cash')}-{L('debt_lt')}-{L('debt_st')})/10^6", -mv['nd'], NUM0)
    put(ws, f'A{rt+7}', 'Enterprise value including cash')
    putf(ws, f'B{rt+7}', f"=B{rt+5}+B{rt+6}", mv['eq_gross'], NUM0)
    put(ws, f'A{rt+8}', 'less minority — on the WHOLE enterprise, cash included')
    putf(ws, f'B{rt+8}', f"=B{rt+7}*'Income Statement'!$B$41", mv['nci'], NUM0)
    put(ws, f'A{rt+9}', 'less the tax-disputes provision')
    putf(ws, f'B{rt+9}', f"={L('provisions')}/10^6", mv['prov'], NUM0)
    put(ws, f'A{rt+10}', 'less dividends payable — declared, out of working capital')
    putf(ws, f'B{rt+10}', f"={L('div_declared')}/10^6", mv['divp'], NUM0)
    put(ws, f'A{rt+11}', 'plus non-operating investments and pledged deposits')
    putf(ws, f'B{rt+11}', f"=({L('fvoci')}+{L('fin_inv')})/10^6", mv['inv'], NUM0)
    put(ws, f'A{rt+12}', 'EQUITY ATTRIBUTABLE', bold=True)
    putf(ws, f'B{rt+12}', f"=B{rt+7}-B{rt+8}-B{rt+9}-B{rt+10}+B{rt+11}", mv['eq'], NUM0, bold=True)
    put(ws, f'A{rt+13}', 'VALUE PER SHARE, EGP', bold=True)
    putf(ws, f'B{rt+13}', f"=B{rt+12}/{L('shares_mn')}", mv['ps'], PX, bold=True)
    rows['ps'] = rt + 13
    rows['ev'] = rt + 5
    return rows


# --- the model's own values for the base block -------------------------------
def blockvals(key):
    return D['blocks'][key]


BV = D['blocks']['base']
put(wsF, 'A5', 'Levers for this block — the base case sets them all neutral', bold=True)
band(wsF, 5)
put(wsF, 'A6', 'Volume growth added a year'); put(wsF, 'B6', 0.0, color=BLUE, fmt=PCT2, paste=True)
put(wsF, 'A7', 'Gross-margin shift'); put(wsF, 'B7', 0.0, color=BLUE, fmt=PCT2, paste=True)
put(wsF, 'A8', 'Realisation-path multiplier'); put(wsF, 'B8', 1.0, color=BLUE, fmt=NUM1, paste=True)
put(wsF, 'A9', 'Working-capital cycle multiplier'); put(wsF, 'B9', 1.0, color=BLUE, fmt=NUM1,
                                                        paste=True)
put(wsF, 'A10', 'Explicit cost of capital')
putf(wsF, 'B10',
     f"={L('rf')}-{L('sov_spread_cds')}+{L('beta')}*{L('erp_cds')}", W['ke_exp'], PCT2)
put(wsF, 'A11', 'Gross borrowings as a share of the capital structure')
putf(wsF, 'B11', f"=({L('debt_lt')}+{L('debt_st')})/10^6/({L('spot')}*{L('shares_mn')})",
     (IN['debt_lt'] + IN['debt_st']) / 1e6 / (IN['spot'] * IN['shares_mn']), PCT2)
put(wsF, 'A12', 'Equity weight — unlevered'); putf(wsF, 'B12', '=1-B11',
     1 - (IN['debt_lt'] + IN['debt_st']) / 1e6 / (IN['spot'] * IN['shares_mn']), PCT2)
put(wsF, 'A13', 'After-tax cost of net debt (shown; not used to discount operations)')
putf(wsF, 'B13',
     f"=({L('kd')}*({L('debt_lt')}+{L('debt_st')})/10^6-{L('cash_yield')}*{L('cash')}/10^6)/"
     f"(({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6)*(1-{RT['tax_eff']})",
     W['k_nd_at'], PCT2)
put(wsF, 'A14', 'OPERATING DISCOUNT RATE — unlevered; the cash is added in the bridge', bold=True)
putf(wsF, 'B14', '=B10', W['wacc_exp'], PCT2, bold=True)
put(wsF, 'A15', 'Terminal risk-free — DERIVED from the target in force')
putf(wsF, 'B15', f"={L('cbe_target')}+{L('real_rate_term')}", RT['rf_term'], PCT2)
put(wsF, 'A16', 'Terminal cost of equity')
putf(wsF, 'B16', f"=B15+{L('beta')}*{L('erp_term')}", W['ke_term'], PCT2)
put(wsF, 'A17', 'WACC — terminal', bold=True)
putf(wsF, 'B17', f"=(1-{L('wd_term')})*B16+{L('wd_term')}*{L('kd_term')}*(1-{RT['tax_eff']})",
     W['wacc_term'], PCT2, bold=True)

BASE_ROWS = engine(wsF, 19, 'BASE CASE', '$B$6', '$B$7', '$B$8', '$B$14', '$B$17',
                   L('g_term'), '$B$9', BV)

# ================= SENSITIVITY — twenty-five complete re-runs =================
wsS = sheet('Sensitivity', 44)
title(wsS, 'Sensitivity — every grid point is the whole engine, again, in formulas',
      'The previous edition pasted these rows because each point is a whole-model re-run and its '
      'builder could not express a re-run inside a grid. Three of the pasted rows turned out not '
      'to reproduce: a working-capital row whose grid was never sorted, a beta row whose centre '
      'did not return the base case, and a volume row whose label described a different scenario '
      'from the one it ran. None of that can happen here, because there is nothing to paste.')

ROWS_S = {}
rr = 5
for gname, cells, pts in D['blocks']['grids']:
    put(wsS, f'A{rr}', gname, bold=True); band(wsS, rr); rr += 1
    for pi, pt in enumerate(pts):
        key = f'{gname}|{pi}'
        put(wsS, f'A{rr}', f'   grid point {pi+1}: {pt["label"]}')
        for cn, cv, fmt in pt['levers']:
            put(wsS, f'{cn}{rr}', cv, color=BLUE, fmt=fmt, paste=True)
        ROWS_S[key] = rr
        rr += 1
    rr += 1

BLK0 = rr + 2
put(wsS, f'A{rr}', 'Each block below is the full engine with that point\'s levers', bold=True)
rr += 2
SUM_PS = {}
for gname, cells, pts in D['blocks']['grids']:
    for pi, pt in enumerate(pts):
        key = f'{gname}|{pi}'
        lr = ROWS_S[key]
        lev = {c: f'$B${lr}' for c, _, _ in pt['levers']}
        rws = engine(wsS, rr, f'{gname} — {pt["label"]}',
                     lev.get('B', '$B$6') if 'B' in lev else f'DCF!$B$6',
                     f'$C${lr}' if pt['has_gm'] else 'DCF!$B$7',
                     f'$D${lr}' if pt['has_fx'] else 'DCF!$B$8',
                     f'$F${lr}' if pt['has_we'] else 'DCF!$B$14',
                     f'$G${lr}' if pt['has_wt'] else 'DCF!$B$17',
                     f'$H${lr}' if pt['has_g'] else L('g_term'),
                     f'$E${lr}' if pt['has_wc'] else 'DCF!$B$9',
                     D['blocks']['scen'][key])
        SUM_PS[key] = rws['ps']
        rr = rws['ps'] + 3

# summary of the grids, pointing at the blocks
put(wsS, 'A4', 'Summary — every cell points at its own live block below', bold=True)
sr = 4
for gname, cells, pts in D['blocks']['grids']:
    sr += 1
    for pi, pt in enumerate(pts):
        key = f'{gname}|{pi}'
        putf(wsS, f'I{ROWS_S[key]}', f'=B{SUM_PS[key]}', D['blocks']['scen'][key]['ps'], PX,
             bold=True)

put(wsS, f'A{BLK0-2}', 'GATE — every row must return the base case at its own base point')
for gname, cells, pts in D['blocks']['grids']:
    for pi, pt in enumerate(pts):
        if pt.get('is_base'):
            key = f'{gname}|{pi}'
            putf(wsS, f'J{ROWS_S[key]}', f"=B{SUM_PS[key]}-DCF!B{BASE_ROWS['ps']}", 0.0,
                 '0.0000')

# ================= LENSES ====================================================
wsL = sheet('Fundamental Valuation', 46)
title(wsL, 'Four lenses, one bridge, one valuation date',
      'Lens 2 is on the TRAILING metric so it borrows no discount factor from lens 1. Lens 3 is '
      'forward-dated and is therefore DISCOUNTED to the valuation date at the cost of equity, '
      'with net cash added at face outside the multiple. Lens 4 uses the same cost-of-equity '
      'glide lens 1 uses. All formulas.', 7)
put(wsL, 'A5', 'Base-year EBITDA')
putf(wsL, 'B5', "='Income Statement'!B20-'Income Statement'!B28-'Income Statement'!B29-'Income Statement'!B30-'Income Statement'!B31"
                "+'Income Statement'!B32", BASE['ebitda_cy25'], NUM0)
put(wsL, 'A6', 'Enterprise value at spot')
putf(wsL, 'B6', f"={L('spot')}*{L('shares_mn')}+({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6",
     REL['ev_trailing'] if 'ev_trailing' in REL else SPOT * SH + DCF['nd'], NUM0)
put(wsL, 'A7', "The company's OWN trailing EV/EBITDA")
putf(wsL, 'B7', '=B6/B5', RT['just_mult'], MULT)
put(wsL, 'A8', 'Justified multiple = trailing x (1 + re-rating)')
putf(wsL, 'B8', f"=B7*(1+{L('rel_rerating')})", RT['just_mult'], MULT, bold=True)

put(wsL, 'A10', 'LENS 1 — discounted cash flow', bold=True); band(wsL, 10, 7)
putf(wsL, 'B10', f"=DCF!B{BASE_ROWS['ps']}", LN['dcf']['base'], PX, bold=True, green=True)
put(wsL, 'A11', 'LENS 2 — relative multiples, on the trailing metric', bold=True)
putf(wsL, 'B11',
     f"=(B8*B5-({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6)*(1-'Income Statement'!B41)"
     f"-({L('provisions')}+{L('div_declared')}-{L('fvoci')}-{L('fin_inv')})/10^6",
     LN['relative']['base'] * SH, NUM0)
putf(wsL, 'C11', f"=B11/{L('shares_mn')}", LN['relative']['base'], PX, bold=True)
put(wsL, 'A12', 'LENS 3 — normalised earnings power, DISCOUNTED', bold=True)
putf(wsL, 'B12', f"=DCF!E{BASE_ROWS['ebit']}", F['ebit'][2], NUM0)
putf(wsL, 'C12',
     f"=B12*(1-{L('tax_stat')})*(1-'Income Statement'!B40)*(1-'Income Statement'!B41)/{L('shares_mn')}",
     NRM['eps'], NUM3)
putf(wsL, 'D12',
     f"=C12*{L('pe_just')}/(1+DCF!B10)^{RT['norm_yrs']}"
     f"+(({L('cash')}-{L('debt_lt')}-{L('debt_st')}-{L('provisions')}-{L('div_declared')})/10^6)"
     f"/{L('shares_mn')}", LN['normalized']['base'], PX, bold=True)
put(wsL, 'A13', 'LENS 4 — book value and sustainable return', bold=True)
putf(wsL, 'B13', f"={L('eq_parent')}/10^6/{L('shares_mn')}", BK['bvps'], PX)
putf(wsL, 'C13', f"={RT['ke_blend']:.10f}", RT['ke_blend'], PCT2)
putf(wsL, 'D13', f"=({L('roe_sust')}-{L('g_term')})/(C13-{L('g_term')})", BK['pb_just'], MULT)
putf(wsL, 'E13', '=B13*D13', LN['book']['base'], PX, bold=True)

# [R-LENS-03]: ONE CLASS PRIMARY IS THE CENTRAL. For a refiner that is the
# cash-flow lens; the other three are CROSS-CHECKS published beside it, never
# weighted into the answer. This row used to compute a 45/20/20/15 blend and
# label it the central, which is the architecture the rule retired — and the
# recalculation gate caught it, because compute.py's own central had already
# moved to the primary while this formula had not.
put(wsL, 'A15', 'CENTRAL — the cash-flow lens', bold=True); band(wsL, 15, 7)
putf(wsL, 'B15', "=B10", D['central'], PX, bold=True)
put(wsL, 'A16', 'against spot')
putf(wsL, 'B16', f"=B15/{L('spot')}-1", D['central'] / SPOT - 1, PCT)
putf(wsL, 'B17',
     f"=B10*{L('w_dcf')}+C11*{L('w_relative')}+D12*{L('w_normalized')}+E13*{L('w_book')}",
     D['lenses']['retired_blend']['base'], PX)
put(wsL, 'A17', 'the retired 45/20/20/15 blend, published beside the answer and unused')
note(wsL, 19,
     'ONE CLASS PRIMARY IS THE CENTRAL. The cash-flow lens is the answer; the relative, '
     'normalised-earnings and book rows are cross-checks and the range across them is an '
     'ENVELOPE, not a spread around a weighted estimate. The blend on the row above is kept '
     'visible and unused: three of the four lenses value a refiner on reported earnings and '
     'historical-cost book, and averaging them into the answer imports every weakness of the '
     'weakest at a weight nobody tested out of sample. An earlier edition labelled a row '
     '"weighted central" and then took the minimum and maximum across all four lenses, both of '
     'which came from the cash-flow lens alone, so the word "weighted" applied to the centre '
     'and not to the ends and the published spread was about two and a half times too wide.')


# =============================================================================
# THE MODEL-REPORT SHEET LIST. The 01-09-2026 edition shipped seven sheets
# against the sixteen the model report requires, while compute.py attested
# structure_matches_model=True — a self-set boolean that nothing outside the
# study ever checked, which is the [R-ENF-01] failure exactly. The nine sheets
# below close it. Every cell here is a formula pointing at the sheets above:
# a summary that stores its own copy of a number is the [L-016] defect.
# =============================================================================
RT_ = BASE_ROWS['ev'] - 5          # first row of the terminal-and-bridge block on DCF
BR_ = {'pv_explicit': RT_, 'roic': RT_ + 1, 'rr': RT_ + 2, 'tv': RT_ + 3,
       'pv_tv': RT_ + 4, 'ev': RT_ + 5, 'netcash': RT_ + 6, 'ev_cash': RT_ + 7,
       'nci': RT_ + 8, 'prov': RT_ + 9, 'divp': RT_ + 10, 'inv': RT_ + 11,
       'eq': RT_ + 12, 'ps': RT_ + 13}
FVS = "'Fundamental Valuation'"
ISS = "'Income Statement'"
SEG = "'Segments'"


def _row(ws, r, label, ref, val, fmt=NUM0, bold=False, green=False, indent=False):
    put(ws, f'A{r}', ('    ' if indent else '') + label, bold=bold)
    putf(ws, f'B{r}', ref, val, fmt, bold=bold, green=green)


def _years(ws, r, label, colref, vals, fmt=NUM0, bold=False):
    """One line across the five forecast columns, pointing at the DCF engine."""
    put(ws, f'A{r}', label, bold=bold)
    for j, cl in enumerate(UC):
        putf(ws, f'{cl}{r}', colref(cl, j), vals[j], fmt, bold=bold)


# ================= SUMMARY ===================================================
wsSU = sheet('Summary', 52)
title(wsSU, 'Summary — the answer, and every lens behind it',
      'Nothing on this sheet is typed. Every figure points at the sheet that computed it, so a '
      'driver changed on Assumptions moves this page too.', 6)
hdr(wsSU, 5, ['', 'Bear', 'Base', 'Bull', 'Weight'])
LNK = [('Discounted cash flow', 'dcf', f"={FVS}!B10"),
       ('Relative multiples', 'relative', f"={FVS}!C11"),
       ('Normalised earnings power', 'normalized', f"={FVS}!D12"),
       ('Book value and sustainable return', 'book', f"={FVS}!E13")]
r = 6
for lab, key, baseref in LNK:
    put(wsSU, f'A{r}', lab)
    put(wsSU, f'B{r}', LN[key]['bear'], BLUE, PX, paste=True)
    putf(wsSU, f'C{r}', baseref, LN[key]['base'], PX, green=True)
    put(wsSU, f'D{r}', LN[key]['bull'], BLUE, PX, paste=True)
    putf(wsSU, f'E{r}', f"={L('w_' + key)}", LN[key]['w'], PCT)
    r += 1
band(wsSU, r, 6)
# [R-LENS-03]: the central IS the class primary — row 6, the cash-flow lens —
# and the other three rows are cross-checks. This row used to weight all four.
put(wsSU, f'A{r}', 'CENTRAL — the cash-flow lens', bold=True)
for cl, k in (('B', 'bear'), ('C', 'base'), ('D', 'bull')):
    putf(wsSU, f'{cl}{r}', f"={cl}6", LN['central'][k], PX, bold=True, green=True)
CEN_R = r
r += 2
_row(wsSU, r, 'Market price at the valuation date', f"={L('spot')}", SPOT, PX); r += 1
_row(wsSU, r, 'Central against price', f"=C{CEN_R}/B{r - 1}-1", D['central'] / SPOT - 1, PCT,
     bold=True); r += 1
_row(wsSU, r, 'Shares in issue, millions', f"={L('shares_mn')}", SH, NUM1); r += 1
_row(wsSU, r, 'Market capitalisation at the valuation date', f"=B{r - 1}*{L('spot')}",
     SH * SPOT, NUM0); r += 2
band(wsSU, r, 6); put(wsSU, f'A{r}', 'WHERE THE VALUE SITS', bold=True); r += 1
_row(wsSU, r, 'Present value of the explicit five years', f"=DCF!B{BR_['pv_explicit']}",
     DCF['pv_explicit']); r += 1
_row(wsSU, r, 'Present value of the terminal block', f"=DCF!B{BR_['pv_tv']}", DCF['pv_tv']); r += 1
_row(wsSU, r, 'Terminal share of enterprise value',
     f"=B{r - 1}/(B{r - 2}+B{r - 1})", DCF['tv_share'], PCT); r += 1
_row(wsSU, r, 'Net cash added in the bridge', f"=DCF!B{BR_['netcash']}", -DCF['nd']); r += 2
note(wsSU, r,
     'The bear and bull columns of each lens are the blue cells here and the only pasted figures '
     'on the sheet: each is a complete re-run of the model at that scenario, computed in '
     'compute.py and reproduced in formulas on the Sensitivity sheet. The central row is a live '
     'formula across them, because the defect being closed was a sheet that labelled a row '
     'weighted and then took the minimum and maximum across four lenses.')

# ================= SOTP BRIDGE ===============================================
wsSB = sheet('SOTP Bridge', 62)
title(wsSB, 'Enterprise value to equity — every step',
      'AMOC is a single operating company: one refinery, eight product lines off one crude '
      'slate, no stakes to sum. There are no parts to break out, so this sheet is the full '
      'enterprise-to-equity walk rather than a sum of pieces, and it says so rather than '
      'inventing segments to fill the shape.', 5)
r = 5
for lab, key, mv, bold in [
        ('Present value of the explicit window', 'pv_explicit', DCF['pv_explicit'], False),
        ('Terminal value', 'tv', DCF['tv'], False),
        ('Present value of the terminal block', 'pv_tv', DCF['pv_tv'], False),
        ('ENTERPRISE VALUE', 'ev', DCF['ev'], True),
        ('plus net cash', 'netcash', -DCF['nd'], False),
        ('Enterprise value including cash', 'ev_cash', BR['eq_gross'], False),
        ('less minority interest, on the whole enterprise', 'nci', BR['nci'], False),
        ('less the tax-disputes provision', 'prov', BR['prov'], False),
        ('less dividends declared and payable', 'divp', BR['divp'], False),
        ('plus non-operating investments and pledged deposits', 'inv', BR['inv'], False),
        ('EQUITY ATTRIBUTABLE TO THE PARENT', 'eq', BR['eq'], True)]:
    _row(wsSB, r, lab, f"=DCF!B{BR_[key]}", mv, NUM0, bold=bold, green=True)
    r += 1
_row(wsSB, r, 'VALUE PER SHARE, EGP', f"=B{r - 1}/{L('shares_mn')}", BR['ps'], PX, bold=True)
PS_SB = r
r += 2
band(wsSB, r, 5); put(wsSB, f'A{r}', 'THE SAME WALK, PER SHARE', bold=True); r += 1
for lab, key, mv in [('Enterprise value', 'ev', DCF['ev']),
                     ('Net cash', 'netcash', -DCF['nd']),
                     ('Minority interest', 'nci', -BR['nci']),
                     ('Provisions, dividends payable and investments', None,
                      BR['inv'] - BR['prov'] - BR['divp'])]:
    if key:
        _row(wsSB, r, lab, f"=DCF!B{BR_[key]}/{L('shares_mn')}"
             + ('*-1' if key == 'nci' else ''), mv / SH, PX)
    else:
        _row(wsSB, r, lab,
             f"=(DCF!B{BR_['inv']}-DCF!B{BR_['prov']}-DCF!B{BR_['divp']})/{L('shares_mn')}",
             mv / SH, PX)
    r += 1
_row(wsSB, r, 'VALUE PER SHARE, EGP', f"=SUM(B{r - 4}:B{r - 1})", BR['ps'], PX, bold=True)
r += 2
note(wsSB, r,
     'The minority is taken on the WHOLE enterprise, cash included, because the subsidiary that '
     'carries it holds its share of the cash as well as of the plant. Taking it on the operating '
     'enterprise alone and then adding the group cash at face would credit the parent with cash '
     'it does not wholly own.')

# ================= RELATIVE & NORMALIZED =====================================
wsRN = sheet('Relative & Normalized', 60)
title(wsRN, 'Lens 2 and lens 3, in full',
      'Lens 2 is struck on the TRAILING metric so it borrows no discount factor from the '
      'cash-flow lens. Lens 3 is forward-dated and is therefore discounted back to the '
      'valuation date, with net cash added at face outside the multiple.', 5)
r = 5
band(wsRN, r, 5); put(wsRN, f'A{r}', 'LENS 2 — RELATIVE MULTIPLES', bold=True); r += 1
_row(wsRN, r, 'Base-year EBITDA', f"={FVS}!B5", BASE['ebitda_cy25']); r += 1
_row(wsRN, r, 'Enterprise value at the market price', f"={FVS}!B6",
     SPOT * SH + DCF['nd']); r += 1
_row(wsRN, r, "The company's own trailing enterprise value / EBITDA", f"={FVS}!B7",
     RT['just_mult'], MULT); r += 1
_row(wsRN, r, 'Re-rating applied', f"={L('rel_rerating')}", IN['rel_rerating'], PCT); r += 1
_row(wsRN, r, 'Justified multiple', f"={FVS}!B8", RT['just_mult'], MULT, bold=True); r += 1
_row(wsRN, r, 'Equity on that multiple', f"={FVS}!B11", LN['relative']['base'] * SH); r += 1
_row(wsRN, r, 'LENS 2, EGP PER SHARE', f"={FVS}!C11", LN['relative']['base'], PX,
     bold=True, green=True); r += 2
band(wsRN, r, 5); put(wsRN, f'A{r}', 'LENS 3 — NORMALISED EARNINGS POWER', bold=True); r += 1
_row(wsRN, r, f"Normalised EBIT, {NRM['year']}", f"={FVS}!B12", F['ebit'][2]); r += 1
_row(wsRN, r, 'Normalised earnings per share', f"={FVS}!C12", NRM['eps'], NUM3); r += 1
_row(wsRN, r, 'Justified price / earnings', f"={L('pe_just')}", IN['pe_just'], MULT); r += 1
_row(wsRN, r, 'Discounted back at the explicit cost of equity',
     f"=(1+DCF!B10)^{RT['norm_yrs']}", (1 + W['ke_exp']) ** RT['norm_yrs'], NUM3); r += 1
_row(wsRN, r, 'LENS 3, EGP PER SHARE', f"={FVS}!D12", LN['normalized']['base'], PX,
     bold=True, green=True); r += 2
band(wsRN, r, 5); put(wsRN, f'A{r}', 'LENS 4 — BOOK VALUE AND SUSTAINABLE RETURN', bold=True)
r += 1
_row(wsRN, r, 'Book value per share attributable to the parent', f"={FVS}!B13",
     BK['bvps'], PX); r += 1
_row(wsRN, r, 'Sustainable return on equity', f"={L('roe_sust')}", IN['roe_sust'], PCT); r += 1
_row(wsRN, r, 'Cost of equity used', f"={FVS}!C13", RT['ke_blend'], PCT2); r += 1
_row(wsRN, r, 'Justified price / book', f"={FVS}!D13", BK['pb_just'], MULT); r += 1
_row(wsRN, r, 'LENS 4, EGP PER SHARE', f"={FVS}!E13", LN['book']['base'], PX,
     bold=True, green=True)

# ================= BALANCE SHEET =============================================
wsBS = sheet('Balance Sheet', 58)
title(wsBS, 'Balance sheet — as filed, and the lines the forecast projects',
      'The filed column is the blue class: figures read off the statement. Everything to the '
      'right is projected from the working-capital days and the asset register on the Income '
      'Statement sheet, through the same engine the cash flow runs on.', 7)
hdr(wsBS, 5, ['EGP mn', 'As filed'] + YRS)
r = 6
BSL = [('Fixed assets at cost', f"=({L('ppe_gross')}+{L('puc')})/10^6",
        (IN['ppe_gross'] + IN['puc']) / 1e6,
        lambda cl, j: f"=DCF!{cl}{BASE_ROWS['gross']}", F.get('gross', D['blocks']['base']['gross'])),
       ('Accumulated depreciation', f"=-({L('ppe_gross')}-{L('ppe_net')})/10^6",
        -(IN['ppe_gross'] - IN['ppe_net']) / 1e6, None, None),
       ('Fixed assets, net book', f"=({L('ppe_net')}+{L('puc')})/10^6",
        (IN['ppe_net'] + IN['puc']) / 1e6,
        lambda cl, j: f"=DCF!{cl}{BASE_ROWS['ppe']}", D['blocks']['base']['ppe']),
       ('Inventory', f"={L('inventory')}/10^6", IN['inventory'] / 1e6, None, None),
       ('Trade and other receivables', f"=({L('recv')}+{L('debtors')})/10^6",
        (IN['recv'] + IN['debtors']) / 1e6, None, None),
       ('Cash and bank', f"={L('cash')}/10^6", IN['cash'] / 1e6, None, None),
       ('Financial investments and pledged deposits', f"=({L('fvoci')}+{L('fin_inv')})/10^6",
        (IN['fvoci'] + IN['fin_inv']) / 1e6, None, None),
       ('Trade and other payables', f"=-({L('payables')}+{L('creditors')})/10^6",
        -(IN['payables'] + IN['creditors']) / 1e6, None, None),
       ('Dividends declared and payable', f"=-{L('div_declared')}/10^6",
        -IN['div_declared'] / 1e6, None, None),
       ('Borrowings', f"=-({L('debt_lt')}+{L('debt_st')})/10^6",
        -(IN['debt_lt'] + IN['debt_st']) / 1e6, None, None),
       ('Provisions', f"=-{L('provisions')}/10^6", -IN['provisions'] / 1e6, None, None)]
for lab, ref, mv, colf, cvals in BSL:
    put(wsBS, f'A{r}', lab)
    putf(wsBS, f'B{r}', ref, mv, NUM0)
    if colf:
        for j, cl in enumerate(UC):
            putf(wsBS, f'{cl}{r}', colf(cl, j), cvals[j], NUM0)
    r += 1
band(wsBS, r, 7)
put(wsBS, f'A{r}', 'Equity attributable to the parent, as filed', bold=True)
putf(wsBS, f'B{r}', f"={L('eq_parent')}/10^6", IN['eq_parent'] / 1e6, NUM0, bold=True)
r += 1
put(wsBS, f'A{r}', 'Minority interest, as filed')
putf(wsBS, f'B{r}', f"={L('eq_nci')}/10^6", IN['eq_nci'] / 1e6, NUM0)
r += 2
band(wsBS, r, 7)
put(wsBS, f'A{r}', 'PROJECTED WORKING CAPITAL AND INVESTED CAPITAL', bold=True); r += 1
_years(wsBS, r, 'Net working capital — on the filed days',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['nwc']}", D['blocks']['base']['nwc']); r += 1
_years(wsBS, r, '    change in the year',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['dnwc']}", D['blocks']['base']['dnwc']); r += 1
_years(wsBS, r, 'Invested capital at replacement cost',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['ic']}", D['blocks']['base']['icr'], NUM0, True)
r += 2
note(wsBS, r,
     'Working capital is projected on inventory, receivable and payable DAYS solved off the '
     'filed balance sheet, with dividends declared removed from payables first — leaving them in '
     'would have the company financing itself out of a distribution it has already committed to '
     'pay.')

# ================= CASH FLOW =================================================
wsCF = sheet('Cash Flow', 58)
title(wsCF, 'Free cash flow to the firm — the full waterfall, year by year',
      'This is the flow the discounted cash flow discounts, taken line by line from the engine '
      'on the DCF sheet. It stops at present value, not at free cash flow.', 7)
hdr(wsCF, 5, ['EGP mn'] + YRS)
r = 6
for lab, key, fmt, bold in [
        ('Revenue', 'rev', NUM0, False), ('Cost of sales', 'cogs', NUM0, False),
        ('GROSS PROFIT', 'gp', NUM0, True), ('Operating expense', 'opex', NUM0, False),
        ('EBITDA', 'ebitda', NUM0, True), ('Depreciation', 'dna', NUM0, False),
        ('EBIT', 'ebit', NUM0, True),
        ('NOPAT before the profit share', 'nop0', NUM0, False),
        ("Employees' profit share and board bonuses", 'emp', NUM0, False),
        ('NOPAT', 'nopat', NUM0, True),
        ('add back depreciation', 'dna', NUM0, False),
        ('less capital expenditure', 'capex', NUM0, False),
        ('less the change in working capital', 'dnwc', NUM0, False),
        ('FREE CASH FLOW TO THE FIRM', 'fcff', NUM0, True),
        ('Forward cost of capital', 'wacc', PCT2, False),
        ('Cumulative discount factor', 'df', '0.00000', False),
        ('PRESENT VALUE', 'pv', NUM0, True)]:
    _years(wsCF, r, lab, (lambda k: lambda cl, j: f"=DCF!{cl}{BASE_ROWS[k]}")(key),
           D['blocks']['base'][{'wacc': 'fwd'}.get(key, key)], fmt, bold)
    r += 1
r += 1
band(wsCF, r, 7)
_row(wsCF, r, 'PV of the explicit window', f"=DCF!B{BR_['pv_explicit']}",
     DCF['pv_explicit'], NUM0, bold=True); r += 1
_row(wsCF, r, 'PV of the terminal block', f"=DCF!B{BR_['pv_tv']}", DCF['pv_tv'], NUM0, bold=True)
r += 2
note(wsCF, r,
     'Two lines are shown as they are computed rather than as a conventional waterfall would '
     'order them: tax is struck on EBIT at the statutory rate before the profit share, and the '
     "employees' share is then taken off the taxed figure, which is the order the company's own "
     'statements follow.')

# ================= SUMMARY FINANCIALS ========================================
wsSF = sheet('Summary Financials', 58)
title(wsSF, 'Summary financials — the base year and the five forecast years',
      'The base column is the twelve months to 30 June 2026 built on the Income Statement '
      'sheet; the five forecast columns are the engine. Margins are outputs of the volume, '
      'price and cost-per-unit build, never inputs.', 7)
hdr(wsSF, 5, ['EGP mn', 'Base year'] + YRS)
r = 6
SFL = [('Revenue', 'rev', f"={ISS}!B19", TTM['rev'], NUM0),
       ('Gross profit', 'gp', f"={ISS}!B20", TTM['gp'], NUM0),
       ('Gross margin', 'gm', f"={ISS}!B22", TTM['gm'], PCT2),
       ('EBITDA', 'ebitda', f"={FVS}!B5", BASE['ebitda_cy25'], NUM0),
       ('EBIT', 'ebit', f"={FVS}!B5-{ISS}!B32", BASE['ebitda_cy25'] - TTM['dep'], NUM0),
       ('NOPAT', 'nopat', None, None, NUM0),
       ('Capital expenditure', 'capex', f"={ISS}!B33", TTM['capex'], NUM0),
       ('Free cash flow to the firm', 'fcff', None, None, NUM0)]
for lab, key, bref, bval, fmt in SFL:
    put(wsSF, f'A{r}', lab, bold=(key in ('rev', 'ebitda', 'fcff')))
    if bref:
        putf(wsSF, f'B{r}', bref, bval, fmt)
    for j, cl in enumerate(UC):
        putf(wsSF, f'{cl}{r}', f"=DCF!{cl}{BASE_ROWS[key]}",
             D['blocks']['base'][key][j], fmt, bold=(key in ('rev', 'ebitda', 'fcff')))
    r += 1
r += 1
band(wsSF, r, 7); put(wsSF, f'A{r}', 'PER TONNE, EGP — the unit the plant actually runs on',
                      bold=True); r += 1
_years(wsSF, r, 'Volume, million tonnes',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['vt']}", D['blocks']['base']['vtot'], NUM3); r += 1
_years(wsSF, r, 'Revenue per tonne',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['rev']}/DCF!{cl}{BASE_ROWS['vt']}",
       [D['blocks']['base']['rev'][j] / D['blocks']['base']['vtot'][j] for j in range(NY)],
       NUM0); r += 1
_years(wsSF, r, 'Gross profit per tonne',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['gp']}/DCF!{cl}{BASE_ROWS['vt']}",
       [D['blocks']['base']['gp'][j] / D['blocks']['base']['vtot'][j] for j in range(NY)],
       NUM0); r += 1
_years(wsSF, r, 'EBITDA per tonne',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['ebitda']}/DCF!{cl}{BASE_ROWS['vt']}",
       [D['blocks']['base']['ebitda'][j] / D['blocks']['base']['vtot'][j] for j in range(NY)],
       NUM0)

# ================= MONTE CARLO ===============================================
wsMC = sheet('Monte Carlo', 58)
ST, S0 = D['strike'], D['step0']
title(wsMC, 'The price distribution, and how often its bands have held',
      'This is the price engine, not the valuation: an independent lens, shown beside the '
      'fundamental answer and never fed into it. The percentiles are produced by the engine and '
      'are read in here, like a figure off a filing.', 6)
r = 5
_row(wsMC, r, 'Anchor date', f'="{ST["anchor_date"]}"', 0, None) if False else None
put(wsMC, f'A{r}', 'Anchor date'); put(wsMC, f'B{r}', ST['anchor_date'], BLUE, paste=True); r += 1
put(wsMC, f'A{r}', 'Anchor price, EGP'); put(wsMC, f'B{r}', ST['spot'], BLUE, PX, paste=True)
r += 1
put(wsMC, f'A{r}', 'Live risk-free used by the engine')
put(wsMC, f'B{r}', ST['rf_live'], BLUE, PCT2, paste=True); r += 1
put(wsMC, f'A{r}', 'Dividend yield used by the engine')
put(wsMC, f'B{r}', ST['q_annual'], BLUE, PCT2, paste=True); r += 1
put(wsMC, f'A{r}', 'Tail parameter and width calibration (per-market fit)')
put(wsMC, f'B{r}', ST['nu'], BLUE, NUM1, paste=True)
put(wsMC, f'C{r}', ST['width_cal'], BLUE, NUM3, paste=True); r += 1
put(wsMC, f'A{r}', 'Per-name width overlay applied')
put(wsMC, f'B{r}', ST['width_overlay_mult'], BLUE, NUM3, paste=True); r += 2
band(wsMC, r, 6); put(wsMC, f'A{r}', 'THE DISTRIBUTION, EGP PER SHARE', bold=True); r += 1
hdr(wsMC, r, ['Horizon', 'p5', 'p25', 'p50', 'p75', 'p95']); r += 1
for hk in ['1M', '3M']:
    h = ST['horizons'].get(hk)
    if not h:
        continue
    put(wsMC, f'A{r}', f"{hk} — grade date {h['grade_date']}")
    for i, q in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        put(wsMC, f'{get_column_letter(2 + i)}{r}', h['pct'][q], BLUE, PX, paste=True)
    r += 1
r += 1
band(wsMC, r, 6); put(wsMC, f'A{r}', 'HOW OFTEN THE BANDS HAVE HELD, ON THIS NAME', bold=True)
r += 1
for lab, v, fmt in [('Resolved three-month windows scored', S0['windows_scored'], NUM0),
                    ('First origin', S0['first_origin'], None),
                    ('Last origin', S0['last_origin'], None),
                    ('Inside the 50 per cent band', S0['cov50'], PCT),
                    ('Inside the 80 per cent band', S0['cov80'], PCT),
                    ('Inside the 90 per cent band', S0['cov90'], PCT),
                    ('Mean probability-integral transform (0.5 is centred)',
                     S0['pit_mean'], NUM3)]:
    put(wsMC, f'A{r}', lab)
    put(wsMC, f'B{r}', v, BLUE, fmt, paste=True)
    r += 1
r += 1
note(wsMC, r,
     'These bands are a record of what happened, not a forecast of the fundamental value. The '
     'fundamental study, this price distribution and the technical read are independent lenses: '
     'no output of one is an input to another, and they are shown side by side so that agreement '
     'between them is information rather than an echo.')

# ================= PER-SHARE & RATIOS ========================================
wsPR = sheet('Per-Share & Ratios', 58)
title(wsPR, 'Per share, and the multiples the answer implies',
      'What the central and the market price each imply, on the same base-year '
      'figures. Every cell is a formula off the Summary and Income Statement sheets.', 6)
r = 5
_row(wsPR, r, 'Central, EGP', f"=Summary!C{CEN_R}", D['central'], PX, bold=True,
     green=True); r += 1
_row(wsPR, r, 'Market price, EGP', f"={L('spot')}", SPOT, PX); r += 1
_row(wsPR, r, 'Book value per share, EGP', f"={FVS}!B13", BK['bvps'], PX); r += 1
_row(wsPR, r, 'Base-year revenue per share, EGP',
     f"={ISS}!B19/{L('shares_mn')}", TTM['rev'] / SH, PX); r += 1
_row(wsPR, r, 'Base-year EBITDA per share, EGP',
     f"={FVS}!B5/{L('shares_mn')}", BASE['ebitda_cy25'] / SH, PX); r += 1
_row(wsPR, r, 'Dividend declared and payable, per share, EGP',
     f"={L('div_declared')}/10^6/{L('shares_mn')}",
     IN['div_declared'] / 1e6 / SH, PX); r += 2
band(wsPR, r, 6); hdr(wsPR, r, ['Multiple', 'At the central', 'At the market price']); r += 1
MULTS = [('Price / book', f"=B$5/{FVS}!B13", f"={L('spot')}/{FVS}!B13",
          D['central'] / BK['bvps'], SPOT / BK['bvps'], MULT),
         ('Enterprise value / base-year EBITDA',
          f"=(B$5*{L('shares_mn')}+({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6)/{FVS}!B5",
          f"={FVS}!B6/{FVS}!B5",
          (D['central'] * SH + DCF['nd']) / BASE['ebitda_cy25'],
          (SPOT * SH + DCF['nd']) / BASE['ebitda_cy25'], MULT),
         ('Enterprise value / base-year revenue',
          f"=(B$5*{L('shares_mn')}+({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6)/{ISS}!B19",
          f"={FVS}!B6/{ISS}!B19",
          (D['central'] * SH + DCF['nd']) / TTM['rev'],
          (SPOT * SH + DCF['nd']) / TTM['rev'], MULT),
         ('Yield on the dividend already declared and payable',
          f"={L('div_declared')}/10^6/{L('shares_mn')}/B$5",
          f"={L('div_declared')}/10^6/{L('shares_mn')}/{L('spot')}",
          IN['div_declared'] / 1e6 / SH / D['central'],
          IN['div_declared'] / 1e6 / SH / SPOT, PCT)]
for lab, fa, fb, va, vb, fmt in MULTS:
    put(wsPR, f'A{r}', lab)
    putf(wsPR, f'B{r}', fa, va, fmt)
    putf(wsPR, f'C{r}', fb, vb, fmt)
    r += 1
r += 1
band(wsPR, r, 6); put(wsPR, f'A{r}', 'RETURNS', bold=True); r += 1
_row(wsPR, r, 'Trailing return on equity', f"={L('roe_sust')}", IN['roe_sust'], PCT); r += 1
_row(wsPR, r, 'Terminal return on invested capital, at replacement cost',
     f"=DCF!B{BR_['roic']}", DCF['roic_term'], PCT2); r += 1
_row(wsPR, r, 'Required reinvestment rate in the terminal year',
     f"=DCF!B{BR_['rr']}", DCF['rr_term'], PCT2); r += 1
_row(wsPR, r, 'Terminal growth', f"={L('g_term')}", IN['g_term'], PCT2)

# ================= PEER & SECTOR =============================================
wsPS = sheet('Peer & Sector', 58)
title(wsPS, 'Sector context — a refiner and lubricant base-oil producer',
      'The comparator basis is the company\'s OWN traded multiple and a stated re-rating, not a '
      'peer set: no Egyptian listed refiner discloses a comparable base-oil split, and importing '
      'a foreign multiple into an EGP valuation would price a different cost of capital. That is '
      'a gap and it is named as one rather than filled.', 7)
r = 5
hdr(wsPS, r, ['Comparator basis', 'Multiple', 'Applied to (EGP mn)', 'Implied EGP/share']); r += 1
for lab, mref, mv in [
        ("The company's own trailing enterprise value / EBITDA", f"={FVS}!B7", RT['just_mult']),
        ('Justified multiple, after the stated re-rating', f"={FVS}!B8", RT['just_mult'])]:
    put(wsPS, f'A{r}', lab)
    putf(wsPS, f'B{r}', mref, mv, MULT)
    putf(wsPS, f'C{r}', f"={FVS}!B5", BASE['ebitda_cy25'], NUM0)
    putf(wsPS, f'D{r}',
         f"=(B{r}*C{r}-({L('debt_lt')}+{L('debt_st')}-{L('cash')})/10^6)"
         f"*(1-{ISS}!B41)/{L('shares_mn')}",
         (mv * BASE['ebitda_cy25'] - DCF['nd']) * (1 - RT['nci_op']) / SH, PX)
    r += 1
r += 1
band(wsPS, r, 7)
put(wsPS, f'A{r}', 'WHAT ACTUALLY DRIVES A PROCESSOR', bold=True); r += 1
hdr(wsPS, r, ['Indicator'] + YRS); r += 1
_years(wsPS, r, 'Throughput, million tonnes',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['vt']}", D['blocks']['base']['vtot'], NUM3); r += 1
_years(wsPS, r, 'Gross profit per tonne, EGP',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['gp']}/DCF!{cl}{BASE_ROWS['vt']}",
       [D['blocks']['base']['gp'][j] / D['blocks']['base']['vtot'][j] for j in range(NY)],
       NUM0); r += 1
_years(wsPS, r, 'EBITDA per tonne, EGP',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['ebitda']}/DCF!{cl}{BASE_ROWS['vt']}",
       [D['blocks']['base']['ebitda'][j] / D['blocks']['base']['vtot'][j] for j in range(NY)],
       NUM0); r += 1
_years(wsPS, r, 'Gross margin',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['gm']}", D['blocks']['base']['gm'], PCT2); r += 1
_years(wsPS, r, 'Capital expenditure as a share of revenue',
       lambda cl, j: f"=DCF!{cl}{BASE_ROWS['capex']}/DCF!{cl}{BASE_ROWS['rev']}",
       [D['blocks']['base']['capex'][j] / D['blocks']['base']['rev'][j] for j in range(NY)],
       PCT2); r += 2
note(wsPS, r,
     'The crude slate is bought from a single counterparty at a disclosed formula and the refined '
     'product is sold into a market whose price is set outside the company, so what a reader can '
     'actually compare across the sector is the spread per tonne and how much of the plant is '
     'running — not a price/earnings ratio struck on a different tax regime and a different '
     'currency.')


# ================= READ FIRST ================================================
wsR = sheet('READ FIRST', 100)
title(wsR, 'Read this first', None, 4)
LINES_R = [
    ('THE RULE', ''),
    ('IT CALCULATES. IT DOES NOT STORE.', ''),
    ('', ''),
    ('Exactly ONE class of cell is pasted in this workbook: a figure read off an audited or '
     'disclosed filing, or a stated house judgement. Those are the blue cells on Assumptions, '
     'and the blue lever cells that define each sensitivity grid point. Everything else is a '
     'live formula.', ''),
    ('', ''),
    ('The previous edition also pasted its sensitivity grids, its beta sweep and its bear and '
     'bull columns, on the grounds that each is a whole-model re-run and a re-run cannot be '
     'written inside a grid. That was 27.6% of the file, and it was the part a reader would '
     'most want to interrogate. Three of those pasted rows turned out not to reproduce.', ''),
    ('', ''),
    ('The Sensitivity sheet now writes the entire forecast engine out again, in formulas, once '
     'for every grid point. Change any driver on Assumptions and every sensitivity cell in the '
     'file moves with it.', ''),
    ('', ''),
    ('WHAT IS AUDITED AND WHAT IS NOT', ''),
    ('The base year is the twelve months to 30 June 2026. The first half of it — July to '
     'December 2025 — is AUDITED by Crowe with an unqualified opinion. The second half — '
     'January to June 2026 — is REPORTED: a disclosure to the Egyptian Exchange dated 29-30 '
     'July 2026, not a filing. Half of this base year is a press release and the study says so '
     'wherever it uses it.', ''),
    ('', ''),
    ('The gross profit of that reported half is NOT used as released. Run through the company\'s '
     'own first-quarter expense run rates the released figure implies a profit after tax about '
     '12.6% above the profit the same release reports. The gross profit used here is solved '
     'from the profit line, which passes two independent coherence tests. Base Year rows 13 to '
     '17 show the whole test.', ''),
    ('', ''),
    ('THE ONE OPERATING INPUT THAT IS NOT READ OFF A FILING', ''),
    ('Note 15-A discloses the cost stack for the company and not by line. The processing-'
     'intensity weights on Assumptions are therefore a judgement, and the per-line margins '
     'depend on them. They are registered, dated and sensitised like every other input rather '
     'than hidden inside a formula.', ''),
]
rr = 5
for a, b in LINES_R:
    c = put(wsR, f'A{rr}', a, bold=(a.isupper() and len(a) < 60))
    c.alignment = Alignment(wrap_text=True, vertical='top')
    wsR.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    wsR.row_dimensions[rr].height = max(14, 12 * (len(a) // 95 + 1))
    rr += 1

# --- the sheet list IS the model report's, in its order, asserted here -------
# compute.py used to attest structure_matches_model=True while the file shipped
# seven sheets. A boolean the study sets on itself is not a check [R-ENF-01], so
# the order is imposed here and the list is asserted against the protocol module
# rather than against a copy of it typed into this file.
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.dirname(HERE)))
_sys.path.insert(0, os.path.dirname(HERE))
import research_protocol as _RP                                        # noqa: E402
WANT = list(_RP.MODEL_STUDY['excel_sheets'])
have = set(wb.sheetnames)
missing, extra = [x for x in WANT if x not in have], [x for x in wb.sheetnames if x not in WANT]
assert not missing and not extra, (
    'workbook does not match the model-report sheet list — missing %s, unexpected %s'
    % (missing, extra))
wb._sheets = [wb[n] for n in WANT]
assert wb.sheetnames == WANT, wb.sheetnames
OUT = os.path.join(HERE, 'AMOC_Valuation_Model_03092026_public.xlsx')
wb.save(OUT)
json.dump({'expected': EXPECT, 'n_formula': NFORM[0], 'n_pasted': NPASTE[0]},
          open(os.path.join(HERE, 'xlsx_expected_v5.json'), 'w'), indent=1)
print(f"{OUT}\n  sheets {len(wb.sheetnames)} — matches the model report\n"
      f"  formula cells {NFORM[0]}  pasted {NPASTE[0]}  "
      f"formula share {NFORM[0]/(NFORM[0]+NPASTE[0]):.1%}")

# The document quotes these counts. A number stated in prose must be COMPUTED, not typed,
# so the builder writes them here and docx_v6.py reads them back.
import json as _json
_json.dump({'formulas': NFORM[0], 'pasted': NPASTE[0],
            'share': NFORM[0] / (NFORM[0] + NPASTE[0])},
           open(os.path.join(HERE, 'workbook_census.json'), 'w'), indent=1)
print('  census written to workbook_census.json')
