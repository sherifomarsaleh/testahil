"""GBCO refresh — 16-sheet Excel model builder, BOTTOM-UP AND FORMULA-PURE.

Reads study_numbers.json ONLY. Blue = input: a disclosed figure, a sourced market quote,
or a stated judgement, each annotated. Every other numeric cell is a live FORMULA — the
builder ends by walking the computation sheets and failing if any derived cell is a
pasted value. FY26E is not typed anywhere: it is H1 disclosed + H2 built by formula on
the FY25 measured seasonal split with explicit tempering inputs. The terminal value pays
for its growth: reinvestment = g / terminal ROIC, both visible cells.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
OUT = os.path.join(HERE, 'GBCO_Valuation_Model_19082026_public.xlsx')

INK = '1C3A36'; BLUE = '1F4EB0'; GREY = '6E7B77'
F_HDR = PatternFill('solid', fgColor='EAF0EE')

SHEETS = ["READ FIRST", "Summary", "Fundamental Valuation", "Assumptions", "SOTP Bridge",
          "Segments", "Relative & Normalized", "DCF", "Income Statement", "Balance Sheet",
          "Cash Flow", "Summary Financials", "Monte Carlo", "Sensitivity",
          "Per-Share & Ratios", "Peer & Sector"]
wb = openpyxl.Workbook(); wb.remove(wb.active)
WS = {s: wb.create_sheet(s) for s in SHEETS}

def W(sn, coord, v, kind='f', fmt='#,##0.0', bold=False):
    c = WS[sn][coord]; c.value = v
    if kind == 'in':
        c.font = Font(color=BLUE, size=10, bold=bold)
        if isinstance(v, (int, float)): c.number_format = fmt
    elif kind == 'f':
        c.font = Font(color=INK, size=10, bold=bold)
        c.number_format = fmt
    elif kind == 'h':
        c.font = Font(color=INK, size=10, bold=True); c.fill = F_HDR
    elif kind == 't':
        c.font = Font(color=INK, size=10, bold=bold)
    elif kind == 'n':
        c.font = Font(color=GREY, size=8.5, italic=True)
    return c

def widths(sn, w):
    for col, x in w.items():
        WS[sn].column_dimensions[col].width = x

a1 = D['auto_h1']; h1 = D['h1']; lob1 = D['lob_h1']; hist = D['hist']; dr = D['drivers']
wac = D['wacc']; dcf = D['dcf']; legs = D['legs']; L = D['lenses']; BW = D['both_ways']
fs = D['fs_forecast']; pub = D['published']; mnt = D['mnt']; bs = D['bs']
YRS = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
COLS = ['E', 'F', 'G', 'H', 'I']

# =====================================================================================
# ASSUMPTIONS — central coordinate registry A (every other sheet references via A[...])
# =====================================================================================
AS = "Assumptions"
A = {}
def a_in(key, row, label, v, note, fmt='0.0000'):
    A[key] = f"Assumptions!$B${row}"
    W(AS, f'A{row}', label, 't'); W(AS, f'B{row}', v, 'in', fmt=fmt); W(AS, f'C{row}', note, 'n')
def a_f(key, row, label, formula, note, fmt='0.0000'):
    A[key] = f"Assumptions!$B${row}"
    W(AS, f'A{row}', label, 't'); W(AS, f'B{row}', formula, 'f', fmt=fmt); W(AS, f'C{row}', note, 'n')

W(AS, 'B2', "Assumptions — every driver, with source and date. Blue = input, black = formula.", 't', bold=True)
W(AS, 'B4', "Capital, market & terminal", 'h')
a_in('shares', 5, "Shares outstanding (mn)", D['shares'], "2Q/1H26 release, shareholder information", '#,##0.0')
a_in('spot', 6, "Last library close (22-Jul-26)", D['spot'], "published page anchor", '0.00')
a_in('spot_ir', 7, "EGX quote, company IR page (19-Aug-26)", D['spot_ir'], "market-value weight basis", '0.00')
a_in('tax', 8, "Statutory corporate tax", D['tax_statutory'], "FS note 11-C", '0.000')
a_in('tg', 9, "Terminal growth, nominal EGP", wac['tg'], "long-run inflation 9-10% + 1.5-2% real", '0.000')
a_in('roic_t', 10, "Terminal return on invested capital", dcf['roic_t'],
     "anchored on the leg's realized LTM ROCE 22.9% (release table 10); sensitized 15-25%", '0.000')
W(AS, 'B12', "Cost of capital (rf* = observed yield − the sovereign's own default spread)", 'h')
a_in('rf', 13, "Egypt 10Y local yield (19-Aug-26)", wac['rf_obs'], "investing.com")
a_in('ds_rat', 14, "Adjusted default spread, rating basis", D['ds_rating'], "Damodaran Jan-2026")
a_in('ds_cds', 15, "Sovereign CDS spread", D['ds_cds'], "Damodaran Jan-2026")
a_in('erp_rat', 16, "Total equity risk premium, rating basis", D['erp_rating'], "Damodaran Jan-2026")
a_in('erp_cds', 17, "Total equity risk premium, CDS basis", D['erp_cds'], "Damodaran Jan-2026")
a_in('beta', 18, "Beta vs the exchange's published index", wac['beta'],
     "own weekly 5y Dimson regression vs EGX 30: n=255, R2 0.24, SE 0.202", '0.000')
a_in('kd_egp', 19, "Avg rate, current EGP borrowings H1-26", wac['kd_pretax_local'],
     "FS note 26; the book is variable-rate (note 29), so the average IS marginal")
a_in('kd_fx', 20, "USD tranche, local-equivalent", wac['kd_fx_local_equiv'],
     "7.78% coupon (note 26) + expected depreciation (CPI differential)")
a_in('pct_loc', 21, "EGP share of auto-leg debt", wac['pct_local'],
     "per-tranche split not disclosed; bounded <=17% USD from note 29; flagged", '0.00')
a_in('debt', 22, "Auto gross debt + leasing notes (30-Jun-26)", a1['debt']+a1['notes'],
     "release table 7", '#,##0.1')
a_f('mktcap', 23, "Market capitalisation", f"={A['spot_ir']}*{A['shares']}", "IR quote x shares", '#,##0.0')
a_f('we', 24, "Equity weight", f"={A['mktcap']}/({A['mktcap']}+{A['debt']})", "market values", '0.000')
a_f('wd', 25, "Debt weight", f"=1-{A['we']}", "", '0.000')
a_f('ke_cds', 26, "Cost of equity, CDS basis", f"={A['rf']}-{A['ds_cds']}+{A['beta']}*{A['erp_cds']}", "")
a_f('ke_rat', 27, "Cost of equity, rating basis", f"={A['rf']}-{A['ds_rat']}+{A['beta']}*{A['erp_rat']}", "")
a_f('kd_bl', 28, "Blended pre-tax cost of debt", f"={A['pct_loc']}*{A['kd_egp']}+(1-{A['pct_loc']})*{A['kd_fx']}", "")
a_f('kd_at', 29, "After-tax cost of debt", f"={A['kd_bl']}*(1-{A['tax']})", "")
a_f('wacc', 30, "WACC — CDS basis (primary)", f"={A['we']}*{A['ke_cds']}+{A['wd']}*{A['kd_at']}", "")
a_f('wacc_rat', 31, "WACC — rating basis (shown alongside)", f"={A['we']}*{A['ke_rat']}+{A['wd']}*{A['kd_at']}", "")

W(AS, 'B33', "Growth, cost and operating paths (FY27E-FY30E; FY26E is DERIVED on Segments)", 'h')
PATHS = [
 ('pc_vol_g', 'PC total volume growth', dr['pc_vol_g'], '0.000'),
 ('pc_asp_g', 'PC price growth', dr['pc_asp_g'], '0.000'),
 ('ckd_g',   'CKD volume growth (the localization driver)', dr['ckd_g'], '0.000'),
 ('cv_vol_g', 'CV&CE volume growth', dr['cv_vol_g'], '0.000'),
 ('cv_asp_g', 'CV&CE price growth', dr['cv_asp_g'], '0.000'),
 ('bus_g',   'Bus volume growth (the recovery driver)', dr['bus_g'], '0.000'),
 ('ce_path', 'Construction-equipment units (planned)', dr['ce_path'], '#,##0'),
 ('lm_vol_g', 'Light-Mobility volume growth', dr['lm_vol_g'], '0.000'),
 ('lm_asp_g', 'Light-Mobility price growth', dr['lm_asp_g'], '0.000'),
 ('tr_g',    'Trading revenue growth', dr['tr_g'], '0.000'),
 ('oth_g',   'Other-auto revenue growth', dr['oth_g'], '0.000'),
 ('fx',      'Imported-cost escalator (currency path)', dr['fx_path'], '0.000'),
 ('cpi',     'Domestic-cost escalator (inflation path)', dr['cpi_path'], '0.000'),
 ('gsa',     'GS&A % of revenue', dr['gsa_pct'], '0.0000'),
 ('dna',     'D&A % of revenue', dr['dna_pct'], '0.0000'),
 ('etr',     'Effective tax path', dr['etr_path'], '0.000'),
 ('capex',   'Capital expenditure (EGP mn)', dr['capex'], '#,##0'),
 ('wc',      'Working capital % of revenue', dr['wc_pct'], '0.000'),
 ('kd_fwd',  'Auto funding-cost path', dr['kd_fwd'], '0.000'),
 ('cap_g',   'GB Capital revenue growth', dr['cap_rev_g'], '0.000'),
 ('cap_np',  'GB Capital net profit path (incl associates)', dr['cap_np_path'], '#,##0'),
 ('pickup',  'Associates equity pickup path', dr['assoc_pickup_path'], '#,##0'),
]
P_ROW = {}
r = 34
for key, lab, vals, fmt in PATHS:
    W(AS, f'A{r}', lab, 't')
    for j, col in enumerate(COLS):
        v = vals[j]
        if v is None:
            W(AS, f'{col}{r}', "derived", 'n')
        else:
            W(AS, f'{col}{r}', v, 'in', fmt=fmt)
    P_ROW[key] = r; r += 1
def PR(key, col):
    return f"Assumptions!{col}${P_ROW[key]}"

r += 1
W(AS, f'B{r}', "Single conventions & held ratios", 'h'); r += 1
singles = [
 ('oth_pct', 'Other operating income % of revenue (held at H1-26)', dr['oth_pct'], '0.0000'),
 ('prov_pct', 'Net provisions % of revenue (held at H1-26)', dr['prov_pct'], '0.0000'),
 ('elim', 'Inter-segment eliminations % of summed revenue (H1-26 actual)', dr['elim_pct'], '0.0000'),
 ('div_ps', 'Dividend per share, held at the FY25 payout', dr['div_ps'], '0.00'),
 ('div_auto', 'Share of the dividend funded by the auto leg (sweep)', dr['div_auto_share'], '0.00'),
 ('fin_h1', 'Auto finance cost H1-26 (release table 8)', a1['fin'], '#,##0.1'),
 ('fin_norm', 'Normalized funding rate (eased-cycle endpoint)', dr['fin_norm_rate'], '0.000'),
 ('imp_pc', 'Imported share of unit cost — PC', dr['imp_share']['pc'], '0.00'),
 ('imp_cv', 'Imported share of unit cost — CV&CE', dr['imp_share']['cv'], '0.00'),
 ('imp_lm', 'Imported share of unit cost — LM', dr['imp_share']['lm'], '0.00'),
 ('imp_tr', 'Imported share of unit cost — trading/other', dr['imp_share']['tr'], '0.00'),
 ('tr_pg', 'Trading/other price inflation (differential year)', dr['tr_price_g'], '0.00'),
 ('tr_reset', 'Tires unit-cost reset FY27E (one-off exhausts)', dr['tr_reset'], '0.00'),
 ('nd', 'Auto net debt, 30-Jun-26', a1['nd'], '#,##0.1'),
 ('nci', 'Auto segment NCI, 30-Jun-26', a1['nci_bs'], '#,##0.1'),
 ('eq_auto', 'Auto segment equity before NCI, 30-Jun-26', a1['eq_seg'], '#,##0.1'),
 ('eq_cap', 'GB Capital segment equity before NCI, 30-Jun-26', D['capital_h1']['eq'], '#,##0.1'),
 ('eq_elim', 'Inter-segment equity eliminations, 30-Jun-26', D['eq_elim'], '#,##0.1'),
 ('parent_eq', 'Parent equity, 30-Jun-26 (restated basis)', bs['parent_eq'], '#,##0.1'),
 ('assoc_bv', 'Investments in associates, carrying 30-Jun-26', D['assoc_total'], '#,##0.1'),
 ('mnt_bv', 'MNT B.V. equity-method carrying (review qualified)', mnt['carrying'], '#,##0.1'),
 ('mnt_stake', 'MNT B.V. stake after the June-26 first close', mnt['stake'], '0.0000'),
 ('round_usd', 'MNT-Halan round valuation (USD mn, first close)', mnt['round_usd'], '#,##0'),
 ('fx_spot', 'USD/EGP (19-Aug-26)', D['usdegp'], '0.00'),
 ('other_assoc', 'Other associates carrying (Mier+Bedaia+Kaf)', D['other_assoc'], '#,##0.1'),
 ('fvoci', 'Investments at fair value through OCI', D['fvoci'], '#,##0.1'),
 ('disc', 'Holding-company discount', legs['disc'], '0.00'),
 ('w_sotp', 'Lens weight — sum of the parts', L['weights']['sotp'], '0.00'),
 ('w_book', 'Lens weight — book & sustainable return', L['weights']['book'], '0.00'),
 ('w_rel', 'Lens weight — relative multiples', L['weights']['relative'], '0.00'),
 ('w_norm', 'Lens weight — normalised earnings', L['weights']['normalized'], '0.00'),
 ('book_bear_m', 'Book-lens bear multiple', 0.80, '0.00'),
 ('pe_bear', 'Relative P/E — bear', L['relative']['pe']['bear'], '0.0'),
 ('pe_base', 'Relative P/E — base', L['relative']['pe']['base'], '0.0'),
 ('pe_bull', 'Relative P/E — bull', L['relative']['pe']['bull'], '0.0'),
 ('nm_bear', 'Normalised multiple — bear', dr['norm_mult']['bear'], '0.0'),
 ('nm_base', 'Normalised multiple — base', dr['norm_mult']['base'], '0.0'),
 ('nm_bull', 'Normalised multiple — bull', dr['norm_mult']['bull'], '0.0'),
 ('ns_bear', 'Normalised profit scalar — bear', dr['norm_scal']['bear'], '0.00'),
 ('ns_bull', 'Normalised profit scalar — bull', dr['norm_scal']['bull'], '0.00'),
 ('mnt_bear_m', 'Bear-case multiplier on the round mark', 0.80, '0.00'),
 ('mnt_bull_m', 'Bull-case multiplier on the round mark', 1.10, '0.00'),
 ('bookmark_bear_m', 'Bear-case multiplier on the carrying value', 0.85, '0.00'),
 ('bookmark_bull_m', 'Bull-case multiplier on the carrying value', 1.10, '0.00'),
 ('cap_bear_m', 'Bear-case multiple on GB Capital operating book', 0.85, '0.00'),
 ('cap_bull_m', 'Bull-case multiple on GB Capital operating book', 1.15, '0.00'),
 ('disc_bear', 'Bear-case holding discount', 0.18, '0.00'),
 ('disc_bull', 'Bull-case holding discount', 0.04, '0.00'),
]
for key, lab, v, fmt in singles:
    a_in(key, r, lab, v, "", fmt=fmt); r += 1
a_f('cap_oper', r, "GB Capital operating equity ex-associates",
    f"={A['eq_cap']}-{A['assoc_bv']}", "segment equity less the associates carrying", '#,##0.1')
r += 1
a_f('mnt_round', r, "MNT stake at the round mark (EGP mn)",
    f"={A['mnt_stake']}*{A['round_usd']}*{A['fx_spot']}", "stake x round x FX", '#,##0.1')
widths(AS, {'A': 42, 'B': 13, 'C': 66, 'D': 9, 'E': 9, 'F': 9, 'G': 9, 'H': 9, 'I': 9})

# =====================================================================================
# SEGMENTS — the bottom-up build: subtype units, seasonal H2 formulas, per-unit costs
# =====================================================================================
S = "Segments"
W(S, 'B2', "The bottom-up build — every disclosed unit grown on its own driver; FY26E derived "
           "from the H1 actual + the FY25 measured seasonal split; margins are OUTPUTS", 't', bold=True)
for j, y in enumerate(['FY23', 'FY24', 'FY25'] + YRS):
    W(S, f'{get_column_letter(2+j)}4', y, 'h')

# --- disclosed anchor block (blue), rows 40+ -----------------------------------------
W(S, 'A40', "Disclosed anchors (statements & release, 13-Aug-2026)", 'h')
anchors = [
 ('h1_pc_u', 41, 'PC volumes H1-26 (units)', lob1['pc_u'], '#,##0'),
 ('h1_ckd_u', 42, 'CKD volumes H1-26', lob1['ckd_u'], '#,##0'),
 ('h1_pc_r', 43, 'PC revenue H1-26', lob1['pc_r'], '#,##0.1'),
 ('h1_cv_u', 44, 'CV&CE volumes H1-26', lob1['cv_u'], '#,##0'),
 ('h1_bus_u', 45, 'Bus volumes H1-26', lob1['bus_u'], '#,##0'),
 ('h1_cv_r', 46, 'CV&CE revenue H1-26', lob1['cv_r'], '#,##0.1'),
 ('h1_lm_u', 47, 'Light-Mobility volumes H1-26', lob1['lm_u'], '#,##0'),
 ('h1_lm_r', 48, 'Light-Mobility revenue H1-26', lob1['lm_r'], '#,##0.1'),
 ('h1_tr_r', 49, 'Trading revenue H1-26', lob1['tr_r'], '#,##0.1'),
 ('h1_oth_r', 50, 'Other-auto revenue H1-26 (residual to the release total)', lob1['oth_r'], '#,##0.1'),
 ('h1_auto_r', 51, 'Auto revenue H1-26', a1['rev'], '#,##0.1'),
 ('h1_auto_gp', 52, 'Auto gross profit H1-26', a1['gp'], '#,##0.1'),
 ('h125_pc_u', 53, 'PC volumes H1-25', lob1['pc_u_h125'], '#,##0'),
 ('h125_cv_u', 54, 'CV&CE volumes H1-25', lob1['cv_u_h125'], '#,##0'),
 ('h125_lm_u', 55, 'Light-Mobility volumes H1-25', lob1['lm_u_h125'], '#,##0'),
 ('h125_tr_r', 56, 'Trading revenue H1-25', lob1['tr_r_h125'], '#,##0.1'),
 ('h125_oth_r', 57, 'Other-auto revenue H1-25 (residual)', lob1['oth_r_h125'], '#,##0.1'),
 ('h125_auto_r', 58, 'Auto revenue H1-25', a1['h1_25_rev'], '#,##0.1'),
 ('h125_auto_gp', 59, 'Auto gross profit H1-25', a1['h1_25_gp'], '#,##0.1'),
 ('m_pc', 60, 'PC statement gross margin H1-26 (segment note 5-B)', dr['lob_margins_h1']['pc'], '0.0000'),
 ('m_cv', 61, 'Buses & trucks statement gross margin H1-26', dr['lob_margins_h1']['cv'], '0.0000'),
 ('m_lm', 62, '2-3-4W statement gross margin H1-26', dr['lob_margins_h1']['lm'], '0.0000'),
 ('m_tr', 63, 'Tires + other trading statement gross margin H1-26', dr['lob_margins_h1']['tr'], '0.0000'),
 ('m_oth', 64, 'Other-auto gross margin (residual)', dr['lob_margins_h1']['oth'], '0.0000'),
 ('t_pc', 65, 'H2-26E temper — PC', dr['h2_temper']['pc'], '0.00'),
 ('t_cv', 66, 'H2-26E temper — CV&CE (2Q surge partly timing)', dr['h2_temper']['cv'], '0.00'),
 ('t_lm', 67, 'H2-26E temper — LM (Qute supply cap)', dr['h2_temper']['lm'], '0.00'),
 ('t_tr', 68, 'H2-26E temper — trading (H2 restock)', dr['h2_temper']['tr'], '0.00'),
 ('t_oth', 69, 'H2-26E temper — other', dr['h2_temper']['oth'], '0.00'),
 ('st_pc', 70, 'H2-26E price step — PC (post-deval increases)', dr['h2_step']['pc'], '0.000'),
 ('st_cv', 71, 'H2-26E price step — CV&CE', dr['h2_step']['cv'], '0.000'),
 ('st_lm', 72, 'H2-26E price step — LM', dr['h2_step']['lm'], '0.000'),
 ('ckd_mix', 73, 'CKD H2 mix drift (Sadat launches)', dr['ckd_mix'], '0.00'),
 ('bus_mix', 74, 'Bus H2 mix drift', dr['bus_mix'], '0.00'),
]
SC = {}
for key, rw, lab, v, fmt in anchors:
    SC[key] = f"$B${rw}"
    W(S, f'A{rw}', lab, 't'); W(S, f'B{rw}', v, 'in', fmt=fmt)
# derived seasonal shares & H1 prices (formulas off blue anchors + D-column FY25)
for _k, _r in [('s_pc', 75), ('s_cv', 76), ('s_lm', 77), ('s_tr', 78), ('s_oth', 79),
               ('asp_pc_h1', 80), ('asp_cv_h1', 81), ('asp_lm_h1', 82),
               ('gpm_h1_26', 83), ('gpm_h1_25', 84), ('gpm_h2_25', 85), ('gpm_h2_26', 86)]:
    SC[_k] = f"$B${_r}"
derived = [
 ('s_pc', 75, 'H1 share of FY25 — PC units', f"={SC['h125_pc_u']}/$D$7", '0.0000'),
 ('s_cv', 76, 'H1 share of FY25 — CV&CE units', f"={SC['h125_cv_u']}/$D$15", '0.0000'),
 ('s_lm', 77, 'H1 share of FY25 — LM units', f"={SC['h125_lm_u']}/$D$20", '0.0000'),
 ('s_tr', 78, 'H1 share of FY25 — trading revenue', f"={SC['h125_tr_r']}/$D$25", '0.0000'),
 ('s_oth', 79, 'H1 share of FY25 — other revenue', f"={SC['h125_oth_r']}/$D$28", '0.0000'),
 ('asp_pc_h1', 80, 'PC realized price H1-26', f"={SC['h1_pc_r']}/{SC['h1_pc_u']}", '0.0000'),
 ('asp_cv_h1', 81, 'CV&CE realized price H1-26', f"={SC['h1_cv_r']}/{SC['h1_cv_u']}", '0.0000'),
 ('asp_lm_h1', 82, 'LM realized price H1-26', f"={SC['h1_lm_r']}/{SC['h1_lm_u']}", '0.00000'),
 ('gpm_h1_26', 83, 'Auto gross margin H1-26', f"={SC['h1_auto_gp']}/{SC['h1_auto_r']}", '0.0000'),
 ('gpm_h1_25', 84, 'Auto gross margin H1-25', f"={SC['h125_auto_gp']}/{SC['h125_auto_r']}", '0.0000'),
 ('gpm_h2_25', 85, 'Auto gross margin H2-25 (FY25 less H1-25)',
  f"=($D$31-{SC['h125_auto_gp']})/($D$30-{SC['h125_auto_r']})", '0.0000'),
 ('gpm_h2_26', 86, 'H2-26E gross margin (H1 less the measured seasonal gap)',
  f"={SC['gpm_h1_26']}-({SC['gpm_h1_25']}-{SC['gpm_h2_25']})", '0.0000'),
]
for key, rw, lab, f, fmt in derived:
    SC[key] = f"$B${rw}"
    W(S, f'A{rw}', lab, 't'); W(S, f'B{rw}', f, 'f', fmt=fmt)

# --- the forecast table (rows 5-33) ---------------------------------------------------
LR = dict(ckd=5, cbu=6, pc_u=7, pc_asp=8, pc_r=9, pc_c=10, pc_gp=11,
          bus=12, truck=13, ce=14, cv_u=15, cv_asp=16, cv_r=17, cv_c=18, cv_gp=19,
          lm_u=20, lm_asp=21, lm_r=22, lm_c=23, lm_gp=24,
          tr_r=25, tr_c=26, tr_gp=27, oth_r=28, oth_c=29, oth_gp=30 - 1)
LR = {'ckd': 5, 'cbu': 6, 'pc_u': 7, 'pc_asp': 8, 'pc_r': 9, 'pc_c': 10, 'pc_gp': 11,
      'bus': 12, 'truck': 13, 'ce': 14, 'cv_u': 15, 'cv_asp': 16, 'cv_r': 17, 'cv_c': 18,
      'cv_gp': 19, 'lm_u': 20, 'lm_asp': 21, 'lm_r': 22, 'lm_c': 23, 'lm_gp': 24,
      'tr_r': 25, 'tr_c': 26, 'tr_gp': 27, 'oth_r': 28, 'oth_c': 29,
      'auto_r': 30, 'auto_gp': 31, 'gpm': 32}
labels = {
 'ckd': 'PC — locally assembled units (CKD, the launch-calendar driver)',
 'cbu': 'PC — imported units (CBU, residual import requirement)',
 'pc_u': 'PC — total units', 'pc_asp': 'PC — price per unit (EGP mn)',
 'pc_r': 'PC — revenue', 'pc_c': 'PC — cost per unit (class-escalated)',
 'pc_gp': 'PC — gross profit',
 'bus': 'CV&CE — bus units (the recovery driver)', 'truck': 'CV&CE — truck units (residual)',
 'ce': 'CV&CE — construction-equipment units (planned)', 'cv_u': 'CV&CE — total units',
 'cv_asp': 'CV&CE — price per unit', 'cv_r': 'CV&CE — revenue',
 'cv_c': 'CV&CE — cost per unit (class-escalated)', 'cv_gp': 'CV&CE — gross profit',
 'lm_u': '2-3-4W — units', 'lm_asp': '2-3-4W — price per unit', 'lm_r': '2-3-4W — revenue',
 'lm_c': '2-3-4W — cost per unit (class-escalated)', 'lm_gp': '2-3-4W — gross profit',
 'tr_r': 'Trading — revenue', 'tr_c': 'Trading — cost ratio (class-escalated)',
 'tr_gp': 'Trading — gross profit', 'oth_r': 'Other automotive — revenue',
 'oth_c': 'Other automotive — cost ratio', 'auto_r': 'AUTOMOTIVE REVENUE',
 'auto_gp': 'AUTOMOTIVE GROSS PROFIT', 'gpm': 'Automotive gross margin (an OUTPUT)'}
for k, rw in LR.items():
    W(S, f'A{rw}', labels[k], 't', bold=(k in ('auto_r', 'auto_gp')))
# history columns (blue where disclosed; ASP/GPM formulas; subtypes not disclosed pre-2026)
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    c = get_column_letter(2+j); Hy = hist[y]
    W(S, f'{c}5', "n/d", 'n'); W(S, f'{c}6', "n/d", 'n')
    W(S, f'{c}7', Hy['pc_u'], 'in', fmt='#,##0')
    W(S, f'{c}8', f"={c}9/{c}7", 'f', fmt='0.0000')
    W(S, f'{c}9', Hy['pc_r'], 'in')
    W(S, f'{c}12', "n/d", 'n'); W(S, f'{c}13', "n/d", 'n'); W(S, f'{c}14', "n/d", 'n')
    W(S, f'{c}15', Hy['cv_u'], 'in', fmt='#,##0')
    W(S, f'{c}16', f"={c}17/{c}15", 'f', fmt='0.0000')
    W(S, f'{c}17', Hy['cv_r'], 'in')
    W(S, f'{c}20', Hy['lm_u'], 'in', fmt='#,##0')
    W(S, f'{c}21', f"={c}22/{c}20", 'f', fmt='0.00000')
    W(S, f'{c}22', Hy['lm_r'], 'in')
    W(S, f'{c}25', Hy['tr_r'], 'in'); W(S, f'{c}28', Hy['oth_r'], 'in')
    W(S, f'{c}30', f"={c}9+{c}17+{c}22+{c}25+{c}28", 'f', bold=True)
    W(S, f'{c}31', Hy['auto_gp'], 'in', bold=True)
    W(S, f'{c}32', f"={c}31/{c}30", 'f', fmt='0.000')
# FY26E — every cell a formula off the blue anchors
E = 'E'
W(S, 'E7', f"={SC['h1_pc_u']}+{SC['h1_pc_u']}*(1-{SC['s_pc']})/{SC['s_pc']}*{SC['t_pc']}", 'f', fmt='#,##0')
W(S, 'E5', f"={SC['h1_ckd_u']}*(E7/{SC['h1_pc_u']})*{SC['ckd_mix']}", 'f', fmt='#,##0')
W(S, 'E6', "=E7-E5", 'f', fmt='#,##0')
W(S, 'E9', f"={SC['h1_pc_r']}+(E7-{SC['h1_pc_u']})*{SC['asp_pc_h1']}*{SC['st_pc']}", 'f')
W(S, 'E8', "=E9/E7", 'f', fmt='0.0000')
W(S, 'E10', f"=E8*(1-{SC['m_pc']})", 'f', fmt='0.0000')
W(S, 'E15', f"={SC['h1_cv_u']}+{SC['h1_cv_u']}*(1-{SC['s_cv']})/{SC['s_cv']}*{SC['t_cv']}", 'f', fmt='#,##0')
W(S, 'E12', f"={SC['h1_bus_u']}*(E15/{SC['h1_cv_u']})*{SC['bus_mix']}", 'f', fmt='#,##0')
W(S, 'E14', f"={PR('ce_path', 'E')}", 'f', fmt='#,##0')
W(S, 'E13', "=E15-E12-E14", 'f', fmt='#,##0')
W(S, 'E17', f"={SC['h1_cv_r']}+(E15-{SC['h1_cv_u']})*{SC['asp_cv_h1']}*{SC['st_cv']}", 'f')
W(S, 'E16', "=E17/E15", 'f', fmt='0.0000')
W(S, 'E18', f"=E16*(1-{SC['m_cv']})", 'f', fmt='0.0000')
W(S, 'E20', f"={SC['h1_lm_u']}+{SC['h1_lm_u']}*(1-{SC['s_lm']})/{SC['s_lm']}*{SC['t_lm']}", 'f', fmt='#,##0')
W(S, 'E22', f"={SC['h1_lm_r']}+(E20-{SC['h1_lm_u']})*{SC['asp_lm_h1']}*{SC['st_lm']}", 'f')
W(S, 'E21', "=E22/E20", 'f', fmt='0.00000')
W(S, 'E23', f"=E21*(1-{SC['m_lm']})", 'f', fmt='0.00000')
W(S, 'E25', f"={SC['h1_tr_r']}*(1+(1-{SC['s_tr']})/{SC['s_tr']}*{SC['t_tr']})", 'f')
W(S, 'E26', f"=1-{SC['m_tr']}", 'f', fmt='0.0000')
W(S, 'E28', f"={SC['h1_oth_r']}*(1+(1-{SC['s_oth']})/{SC['s_oth']}*{SC['t_oth']})", 'f')
W(S, 'E29', f"=1-{SC['m_oth']}", 'f', fmt='0.0000')
W(S, 'E30', "=E9+E17+E22+E25+E28", 'f', bold=True)
W(S, 'E31', f"={SC['h1_auto_gp']}+(E30-{SC['h1_auto_r']})*{SC['gpm_h2_26']}", 'f', bold=True)
W(S, 'E32', "=E31/E30", 'f', fmt='0.000')
# per-LOB GP rows in FY26E: shown for the H2 construction consistency (formulas)
W(S, 'E11', f"=E9-E7*E10", 'f')
W(S, 'E19', f"=E17-E15*E18", 'f')
W(S, 'E24', f"=E22-E20*E23", 'f')
W(S, 'E27', "=E25*(1-E26)", 'f')
# FY27+ — growth and class-escalated unit costs; cost moves with price from FY28E
for j, col in enumerate(COLS[1:], start=1):
    pv = COLS[j-1]
    W(S, f'{col}5', f"={pv}5*(1+{PR('ckd_g', col)})", 'f', fmt='#,##0')
    W(S, f'{col}7', f"={pv}7*(1+{PR('pc_vol_g', col)})", 'f', fmt='#,##0')
    W(S, f'{col}6', f"={col}7-{col}5", 'f', fmt='#,##0')
    W(S, f'{col}8', f"={pv}8*(1+{PR('pc_asp_g', col)})", 'f', fmt='0.0000')
    W(S, f'{col}9', f"={col}7*{col}8", 'f')
    W(S, f'{col}12', f"={pv}12*(1+{PR('bus_g', col)})", 'f', fmt='#,##0')
    W(S, f'{col}14', f"={PR('ce_path', col)}", 'f', fmt='#,##0')
    W(S, f'{col}15', f"={pv}15*(1+{PR('cv_vol_g', col)})", 'f', fmt='#,##0')
    W(S, f'{col}13', f"={col}15-{col}12-{col}14", 'f', fmt='#,##0')
    W(S, f'{col}16', f"={pv}16*(1+{PR('cv_asp_g', col)})", 'f', fmt='0.0000')
    W(S, f'{col}17', f"={col}15*{col}16", 'f')
    W(S, f'{col}20', f"={pv}20*(1+{PR('lm_vol_g', col)})", 'f', fmt='#,##0')
    W(S, f'{col}21', f"={pv}21*(1+{PR('lm_asp_g', col)})", 'f', fmt='0.00000')
    W(S, f'{col}22', f"={col}20*{col}21", 'f')
    W(S, f'{col}25', f"={pv}25*(1+{PR('tr_g', col)})", 'f')
    W(S, f'{col}28', f"={pv}28*(1+{PR('oth_g', col)})", 'f')
    # unit costs: FY27E carries the class differential once; held with price after
    if j == 1:
        W(S, f'{col}10', f"=E10*(1+{A['imp_pc']}*{PR('fx', col)}+(1-{A['imp_pc']})*{PR('cpi', col)})", 'f', fmt='0.0000')
        W(S, f'{col}18', f"=E18*(1+{A['imp_cv']}*{PR('fx', col)}+(1-{A['imp_cv']})*{PR('cpi', col)})", 'f', fmt='0.0000')
        W(S, f'{col}23', f"=E23*(1+{A['imp_lm']}*{PR('fx', col)}+(1-{A['imp_lm']})*{PR('cpi', col)})", 'f', fmt='0.00000')
        W(S, f'{col}26', f"=E26*(1+{A['imp_tr']}*{PR('fx', col)}+(1-{A['imp_tr']})*{PR('cpi', col)})"
                         f"/(1+{A['tr_pg']})*{A['tr_reset']}", 'f', fmt='0.0000')
        W(S, f'{col}29', f"=E29*(1+{A['imp_tr']}*{PR('fx', col)}+(1-{A['imp_tr']})*{PR('cpi', col)})"
                         f"/(1+{A['tr_pg']})", 'f', fmt='0.0000')
    else:
        W(S, f'{col}10', f"={pv}10*(1+{PR('pc_asp_g', col)})", 'f', fmt='0.0000')
        W(S, f'{col}18', f"={pv}18*(1+{PR('cv_asp_g', col)})", 'f', fmt='0.0000')
        W(S, f'{col}23', f"={pv}23*(1+{PR('lm_asp_g', col)})", 'f', fmt='0.00000')
        W(S, f'{col}26', f"={pv}26", 'f', fmt='0.0000')
        W(S, f'{col}29', f"={pv}29", 'f', fmt='0.0000')
    W(S, f'{col}11', f"=({col}8-{col}10)*{col}7", 'f')
    W(S, f'{col}19', f"=({col}16-{col}18)*{col}15", 'f')
    W(S, f'{col}24', f"=({col}21-{col}23)*{col}20", 'f')
    W(S, f'{col}27', f"={col}25*(1-{col}26)", 'f')
    W(S, f'{col}30', f"={col}9+{col}17+{col}22+{col}25+{col}28", 'f', bold=True)
    W(S, f'{col}31', f"={col}11+{col}19+{col}24+{col}27+{col}28*(1-{col}29)", 'f', bold=True)
    W(S, f'{col}32', f"={col}31/{col}30", 'f', fmt='0.000')
W(S, 'A34', "GB Capital revenue", 't')
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    W(S, f'{get_column_letter(2+j)}34', hist[y]['cap_rev'], 'in')
W(S, 'E34', dr['cap_rev_26'], 'in')
for j, col in enumerate(COLS[1:], start=1):
    W(S, f'{col}34', f"={COLS[j-1]}34*(1+{PR('cap_g', col)})", 'f')
W(S, 'A35', "GB Capital net profit (path input, incl associates)", 't')
for j, col in enumerate(COLS):
    W(S, f'{col}35', f"={PR('cap_np', col)}", 'f')
W(S, 'A37', "FY27E unit costs carry each class's escalator against its price once (the measured "
            "compression year); from FY28E cost moves with price — held flat both directions. "
            "Trading FY27E divides out its price growth and carries the 1.04 inventory-reset. "
            "'n/d' = the subtype split is not disclosed for that year.", 'n')
widths(S, {'A': 46, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12})

# =====================================================================================
# DCF — waterfall, H1 realization block, disciplined terminal value, case engine
# =====================================================================================
C = "DCF"
W(C, 'B2', "Automotive leg — free cash flow to the firm; the terminal growth is PAID FOR "
           "(reinvestment = growth / terminal return)", 't', bold=True)
for j, col in enumerate(COLS):
    W(C, f'{col}4', YRS[j], 'h')
dcf_rows = [
 (5,  'Revenue', lambda col: f"=Segments!{col}30"),
 (6,  'Gross profit', lambda col: f"=Segments!{col}31"),
 (7,  'GS&A', lambda col: f"=-{col}5*{PR('gsa', col)}"),
 (8,  'Other operating income', lambda col: f"={col}5*{A['oth_pct']}"),
 (9,  'Net provisions', lambda col: f"={col}5*{A['prov_pct']}"),
 (10, 'Operating profit', lambda col: f"={col}6+{col}7+{col}8+{col}9"),
 (11, 'Effective tax rate', lambda col: f"={PR('etr', col)}"),
 (12, 'Net operating profit after tax', lambda col: f"={col}10*(1-{col}11)"),
 (13, 'Depreciation & amortisation', lambda col: f"={col}5*{PR('dna', col)}"),
 (14, 'EBITDA', lambda col: f"={col}10+{col}13"),
 (15, 'Capital expenditure', lambda col: f"=-{PR('capex', col)}"),
 (16, 'Working capital (level)', lambda col: f"={col}5*{PR('wc', col)}"),
 (17, 'Change in working capital', None),
 (18, 'Free cash flow to the firm', lambda col: f"={col}12+{col}13+{col}15-{col}17"),
]
for rw, lab, fn in dcf_rows:
    W(C, f'A{rw}', lab, 't', bold=(rw in (10, 14, 18)))
    if fn:
        for col in COLS:
            W(C, f'{col}{rw}', fn(col), 'f', fmt='0.000' if rw == 11 else '#,##0.0',
              bold=(rw in (10, 14, 18)))
W(C, 'E17', "=E16-B22", 'f')
for col, pv in zip(COLS[1:], COLS[:-1]):
    W(C, f'{col}17', f"={col}16-{pv}16", 'f')
W(C, 'A20', "The realized first half (model definition, from the disclosed actuals)", 'h')
h1blk = [
 (21, 'H1-26 operating profit (release)', a1['op'], 'in'),
 (22, 'FY25 working capital (base)', hist['FY25']['wc'], 'in'),
 (23, 'H1-26 working capital', a1['wc'], 'in'),
 (24, 'H1-26 effective tax rate', D['etr_h1'], 'in'),
 (25, 'H1-26 depreciation & amortisation', a1['dna'], 'in'),
 (26, 'H1-26 capital expenditure', a1['capex'], 'in'),
]
for rw, lab, v, kind in h1blk:
    W(C, f'A{rw}', lab, 't'); W(C, f'B{rw}', v, kind, fmt='#,##0.0000' if rw == 24 else '#,##0.1')
W(C, 'A27', 'H1-26 realized free cash flow', 't', bold=True)
W(C, 'B27', "=B21*(1-B24)+B25-B26-(B23-B22)", 'f', bold=True)
W(C, 'A28', 'H2-26E free cash flow', 't'); W(C, 'B28', "=E18-B27", 'f')
W(C, 'A30', 'Discount period (years, from 30-Jun-26, mid-period)', 't')
W(C, 'A31', 'Discount factor'); W(C, 'A31', 'Discount factor', 't'); W(C, 'A32', 'Present value', 't')
for j, col in enumerate(COLS):
    W(C, f'{col}30', [0.5, 1.5, 2.5, 3.5, 4.5][j], 'in', fmt='0.0')
    W(C, f'{col}31', f"=1/(1+{A['wacc']})^{col}30", 'f', fmt='0.0000')
    W(C, f'{col}32', f"={'B28' if j == 0 else col+'18'}*{col}31", 'f')
tvrows = [
 (34, 'Sum of discounted flows', "=SUM(E32:I32)", '#,##0.0', True),
 (35, 'Terminal reinvestment rate (growth / terminal return)', f"={A['tg']}/{A['roic_t']}", '0.000', False),
 (36, 'Terminal cash flow (FY30E NOPAT grown, less forced reinvestment)',
      f"=I12*(1+{A['tg']})*(1-B35)", '#,##0.0', False),
 (37, 'Terminal value', f"=B36/({A['wacc']}-{A['tg']})", '#,##0.0', False),
 (38, 'Present value of the terminal value', "=B37*I31", '#,##0.0', False),
 (39, 'Enterprise value — automotive', "=B34+B38", '#,##0.0', True),
 (40, 'Less net debt (30-Jun-26)', f"=-{A['nd']}", '#,##0.0', False),
 (41, 'Less non-controlling interests', f"=-{A['nci']}", '#,##0.0', False),
 (42, 'Automotive equity value', "=B39+B40+B41", '#,##0.0', True),
 (43, 'Per share', f"=B42/{A['shares']}", '0.00', True),
 (44, 'Terminal share of enterprise value', "=B38/B39", '0.0%', False),
 (45, "Alternative (the July construction): Gordon on year-five free cash flow",
      f"=B34+I18*(1+{A['tg']})/({A['wacc']}-{A['tg']})*I31-{A['nd']}-{A['nci']}", '#,##0.0', False),
 (46, "  its implied terminal return on capital (growth/(1-FCFF/NOPAT)) — disclosed, not hidden",
      f"={A['tg']}/(1-I18/I12)", '0.000', False),
]
for rw, lab, f, fmt, b in tvrows:
    W(C, f'A{rw}', lab, 't', bold=b); W(C, f'B{rw}', f, 'f', fmt=fmt, bold=b)
# --- case engine (bear / bull), same construction, shifted by blue inputs -------------
W(C, 'A49', "Case engine — the same model under shifted drivers (blue shifts)", 'h')
W(C, 'B50', "bear", 'h'); W(C, 'C50', "bull", 'h')
case_in = [(51, 'Gross-margin shift', -0.010, 0.008, '0.000'),
           (52, 'Discount-rate shift', 0.020, -0.015, '0.000'),
           (53, 'Terminal-growth shift', -0.010, 0.010, '0.000'),
           (54, 'Tax-rate shift', 0.04, -0.02, '0.000')]
for rw, lab, vb, vu, fmt in case_in:
    W(C, f'A{rw}', lab, 't'); W(C, f'B{rw}', vb, 'in', fmt=fmt); W(C, f'C{rw}', vu, 'in', fmt=fmt)
W(C, 'A55', 'Case cost of capital', 't')
W(C, 'B55', f"={A['wacc']}+B52", 'f'); W(C, 'C55', f"={A['wacc']}+C52", 'f')
W(C, 'A56', 'Case terminal growth', 't')
W(C, 'B56', f"={A['tg']}+B53", 'f'); W(C, 'C56', f"={A['tg']}+C53", 'f')
for cs, base_col in (('bear', 'B'), ('bull', 'C')):
    r0 = 58 if cs == 'bear' else 62
    W(C, f'A{r0}', f'{cs} — NOPAT', 't'); W(C, f'A{r0+1}', f'{cs} — free cash flow', 't')
    W(C, f'A{r0+2}', f'{cs} — present value', 't')
    for j, col in enumerate(COLS):
        W(C, f'{col}{r0}', f"=({col}10+{col}5*${base_col}$51)*(1-({col}11+${base_col}$54))", 'f')
        W(C, f'{col}{r0+1}', f"={col}{r0}+{col}13+{col}15-{col}17", 'f')
        src = f"({col}{r0+1}-$B$27)" if j == 0 else f"{col}{r0+1}"
        W(C, f'{col}{r0+2}', f"={src}/(1+${base_col}$55)^{col}30", 'f')
    W(C, f'A{r0+3}', f'{cs} — automotive equity value', 't', bold=True)
    W(C, f'{base_col}66' if False else f'B{r0+3}' if cs == 'bear' else f'B{r0+3}',
      f"=SUM(E{r0+2}:I{r0+2})+I{r0}*(1+{base_col}56)*(1-{base_col}56/{A['roic_t']})"
      f"/({base_col}55-{base_col}56)/(1+{base_col}55)^I30-{A['nd']}-{A['nci']}", 'f', bold=True)
widths(C, {'A': 46, 'B': 13, 'C': 13, 'D': 11, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12})

# =====================================================================================
# SOTP BRIDGE — both framings and the cases, all formulas
# =====================================================================================
B = "SOTP Bridge"
W(B, 'B2', "The group as the sum of its parts — the contested stake both ways, never averaged", 't', bold=True)
W(B, 'B4', "Round mark", 'h'); W(B, 'C4', "Balance-sheet mark", 'h')
rows_b = [
 (5, 'Automotive equity value (cash-flow model)', "=DCF!B42", "=DCF!B42"),
 (6, 'GB Capital operating equity (ex-associates)', f"={A['cap_oper']}", f"={A['cap_oper']}"),
 (7, 'MNT-Halan stake', f"={A['mnt_round']}", f"={A['mnt_bv']}"),
 (8, 'Other associates', f"={A['other_assoc']}", f"={A['other_assoc']}"),
 (9, 'Investments at fair value (OCI)', f"={A['fvoci']}", f"={A['fvoci']}"),
 (10, 'Sum of the parts', "=SUM(B5:B9)", "=SUM(C5:C9)"),
 (11, 'Holding-company discount', f"=-B10*{A['disc']}", f"=-C10*{A['disc']}"),
 (12, 'Equity value', "=B10+B11", "=C10+C11"),
 (13, 'Per share', f"=B12/{A['shares']}", f"=C12/{A['shares']}"),
]
for rw, lab, fa, fb in rows_b:
    b = rw in (10, 12, 13)
    W(B, f'A{rw}', lab, 't', bold=b)
    W(B, f'B{rw}', fa, 'f', fmt='0.00' if rw == 13 else '#,##0.1', bold=b)
    W(B, f'C{rw}', fb, 'f', fmt='0.00' if rw == 13 else '#,##0.1', bold=b)
W(B, 'A15', 'Bear per share (bear equity, leg haircuts, bear discount)', 't')
W(B, 'B15', f"=(DCF!B61+{A['cap_oper']}*{A['cap_bear_m']}+{A['mnt_round']}*{A['mnt_bear_m']}"
            f"+{A['other_assoc']}+{A['fvoci']})*(1-{A['disc_bear']})/{A['shares']}", 'f', fmt='0.00')
W(B, 'C15', f"=(DCF!B61+{A['cap_oper']}*{A['cap_bear_m']}+({A['mnt_bv']}+{A['other_assoc']}"
            f"+{A['fvoci']})*{A['bookmark_bear_m']})*(1-{A['disc_bear']})/{A['shares']}", 'f', fmt='0.00')
W(B, 'A16', 'Bull per share (bull equity, leg uplifts, bull discount)', 't')
W(B, 'B16', f"=(DCF!B65+{A['cap_oper']}*{A['cap_bull_m']}+{A['mnt_round']}*{A['mnt_bull_m']}"
            f"+{A['other_assoc']}+{A['fvoci']})*(1-{A['disc_bull']})/{A['shares']}", 'f', fmt='0.00')
W(B, 'C16', f"=(DCF!B65+{A['cap_oper']}*{A['cap_bull_m']}+({A['mnt_bv']}+{A['other_assoc']}"
            f"+{A['fvoci']})*{A['bookmark_bull_m']})*(1-{A['disc_bull']})/{A['shares']}", 'f', fmt='0.00')
W(B, 'A18', f"The whole gap between the two columns is one judgement about one stake. The "
            f"reviewer could not verify the carrying value (qualified conclusion, second "
            f"period); the round is real bank-led money but a first close. Neither is "
            f"averaged away — the study carries both to the end.", 'n')
widths(B, {'A': 48, 'B': 16, 'C': 18})

# =====================================================================================
# INCOME STATEMENT / BALANCE SHEET / CASH FLOW
# =====================================================================================
IS = "Income Statement"
W(IS, 'B2', "Group income statement — three disclosed years and the five-year model", 't', bold=True)
for j, y in enumerate(['FY23', 'FY24', 'FY25'] + YRS):
    W(IS, f'{get_column_letter(2+j)}4', y, 'h')
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    c = get_column_letter(2+j); Hy = hist[y]
    W(IS, f'{c}5', f"=Segments!{c}30", 'f')
    W(IS, f'{c}6', Hy['cap_rev'], 'in')
    W(IS, f'{c}7', round(Hy['group_rev']-Hy['auto_rev']-Hy['cap_rev'], 1), 'in')
    W(IS, f'{c}8', f"=SUM({c}5:{c}7)", 'f', bold=True)
    W(IS, f'{c}10', Hy['auto_ebit'], 'in')
    W(IS, f'{c}16', Hy['np'], 'in', bold=True)
    W(IS, f'{c}17', f"={c}16/{A['shares']}", 'f', fmt='0.00')
gis = [(5, 'Automotive revenue', lambda col: f"=Segments!{col}30"),
       (6, 'GB Capital revenue', lambda col: f"=Segments!{col}34"),
       (7, 'Inter-segment eliminations', lambda col: f"=-({col}5+{col}6)*{A['elim']}"),
       (8, 'Group revenue', lambda col: f"=SUM({col}5:{col}7)"),
       (10, 'Automotive operating profit', lambda col: f"=DCF!{col}10"),
       (11, 'Automotive finance cost', None),
       (12, 'Automotive profit before tax', lambda col: f"={col}10+{col}11"),
       (13, 'Automotive tax', lambda col: f"=-{col}12*DCF!{col}11"),
       (14, 'Automotive net profit', lambda col: f"={col}12+{col}13"),
       (15, 'GB Capital net profit (incl associates)', lambda col: f"=Segments!{col}35"),
       (16, 'Group net profit (attributable)', lambda col: f"={col}14+{col}15"),
       (17, 'Earnings per share (EGP)', lambda col: f"={col}16/{A['shares']}")]
for rw, lab, fn in gis:
    W(IS, f'A{rw}', lab, 't', bold=(rw in (8, 16)))
    if fn:
        for col in COLS:
            W(IS, f'{col}{rw}', fn(col), 'f', fmt='0.00' if rw == 17 else '#,##0.1', bold=(rw == 16))
W(IS, 'E11', f"=-({A['fin_h1']}+0.5*{A['nd']}*{PR('kd_fwd', 'E')})", 'f')
for col, pv in zip(COLS[1:], COLS[:-1]):
    W(IS, f'{col}11', f"=-'Balance Sheet'!{pv}10*{PR('kd_fwd', col)}", 'f')
W(IS, 'A19', "History: the model's FY26E includes the realized half (revenue 48,474.4, "
             "attributable profit 1,262.0, earnings per share 1.163 — all disclosed).", 'n')
widths(IS, {'A': 40, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12})

BS = "Balance Sheet"
W(BS, 'B2', "Condensed balance-sheet model — every roll a formula", 't', bold=True)
for j, y in enumerate(['FY25'] + YRS):
    W(BS, f'{get_column_letter(4+j)}4', y, 'h')
bs_anchor = [(5, 'Automotive working capital', hist['FY25']['wc']),
             (6, 'Net PP&E and intangibles', 13389.1),
             (7, 'Investments in associates', None),
             (10, 'Automotive net debt', None),
             (11, 'Parent equity (restated basis)', None)]
W(BS, 'A5', 'Automotive working capital', 't'); W(BS, 'D5', hist['FY25']['wc'], 'in')
W(BS, 'A6', 'Net PP&E and intangibles (roll: + capex − D&A)', 't'); W(BS, 'D6', hist['FY25']['bs']['ppe'], 'in')
W(BS, 'A7', 'Investments in associates (roll: + equity pickup)', 't')
W(BS, 'D7', hist['FY25']['bs']['assoc'], 'in')
W(BS, 'A8', 'GB Capital operating net assets (held)', 't')
W(BS, 'A10', 'Automotive net debt (cash-sweep roll)', 't', bold=True)
W(BS, 'D10', a1['nd'], 'in', bold=True)
W(BS, 'A11', 'Parent equity (roll: + profit − dividend)', 't', bold=True)
W(BS, 'D11', bs['parent_eq'], 'in', bold=True)
W(BS, 'A12', 'Dividends paid', 't')
W(BS, 'A16', 'Book value per share', 't')
for j, col in enumerate(COLS):
    pv = 'D' if j == 0 else COLS[j-1]
    W(BS, f'{col}5', f"=DCF!{col}16", 'f')
    W(BS, f'{col}6', f"={pv}6-DCF!{col}15-DCF!{col}13", 'f')
    W(BS, f'{col}7', f"={pv}7+{PR('pickup', col)}", 'f')
    W(BS, f'{col}8', f"={A['cap_oper']}", 'f')
    W(BS, f'{col}12', f"={A['div_ps']}*{A['shares']}", 'f')
    W(BS, f'{col}10', f"={pv}10-(DCF!{col}18+'Income Statement'!{col}11*(1-DCF!{col}11)"
                      f"-{col}12*{A['div_auto']})", 'f')
    W(BS, f'{col}11', f"={pv}11+'Income Statement'!{col}16-{col}12", 'f')
    W(BS, f'{col}16', f"={col}11/{A['shares']}", 'f', fmt='0.00')
W(BS, 'A14', "The D column carries the disclosed anchors: FY25 restated associates (15,732.4 "
             "after the +2,460.2 restatement) and the 30-Jun-26 net debt / parent equity. "
             "Conversion cycle measured at 30-Jun-26: 122 days inventory, 27 receivables, "
             "82 payables — the working-capital ratio path projects that cycle.", 'n')
widths(BS, {'A': 46, 'B': 6, 'C': 6, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12})

CF = "Cash Flow"
W(CF, 'B2', "Condensed group cash-flow model", 't', bold=True)
for j, y in enumerate(YRS):
    W(CF, f'{COLS[j]}4', y, 'h')
cf_rows = [(5, 'Group net profit', lambda col: f"='Income Statement'!{col}16"),
           (6, 'Depreciation & amortisation', lambda col: f"=DCF!{col}13"),
           (7, 'Change in working capital', lambda col: f"=-DCF!{col}17"),
           (8, 'Capital expenditure', lambda col: f"=DCF!{col}15"),
           (9, 'Dividends paid', lambda col: f"=-'Balance Sheet'!{col}12"),
           (10, 'Cash flow after dividends', lambda col: f"=SUM({col}5:{col}9)"),
           (11, 'Change in automotive net debt (sweep)', None)]
for rw, lab, fn in cf_rows:
    W(CF, f'A{rw}', lab, 't', bold=(rw == 10))
    if fn:
        for col in COLS:
            W(CF, f'{col}{rw}', fn(col), 'f', bold=(rw == 10))
for j, col in enumerate(COLS):
    base = A['nd'] if j == 0 else f"'Balance Sheet'!{COLS[j-1]}10"
    W(CF, f'{col}11', f"='Balance Sheet'!{col}10-{base}", 'f')
widths(CF, {'A': 42, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12})

# =====================================================================================
# RELATIVE & NORMALIZED / FUNDAMENTAL VALUATION
# =====================================================================================
RN = "Relative & Normalized"
W(RN, 'B2', "Relative multiples and normalised earnings power", 't', bold=True)
W(RN, 'A4', 'FY26E group earnings per share', 't'); W(RN, 'B4', "='Income Statement'!E17", 'f', fmt='0.00')
W(RN, 'A5', 'Multiple band (bear / base / bull)', 't')
W(RN, 'B5', f"={A['pe_bear']}", 'f', fmt='0.0'); W(RN, 'C5', f"={A['pe_base']}", 'f', fmt='0.0')
W(RN, 'D5', f"={A['pe_bull']}", 'f', fmt='0.0')
W(RN, 'A6', 'Relative lens value', 't')
for c2, m in (('B', 'B5'), ('C', 'C5'), ('D', 'D5')):
    W(RN, f'{c2}6', f"=B4*{m}", 'f', fmt='0.00')
W(RN, 'A8', "Normalisation walk (mid-cycle)", 'h')
W(RN, 'A9', 'FY27E automotive operating profit', 't'); W(RN, 'B9', "=DCF!F10", 'f')
W(RN, 'A10', 'Financing cost at the eased-cycle rate on today\'s net debt', 't')
W(RN, 'B10', f"=-{A['nd']}*{A['fin_norm']}", 'f')
W(RN, 'A11', 'Tax at the statutory rate (no loss-stranding)', 't')
W(RN, 'B11', f"=-(B9+B10)*{A['tax']}", 'f')
W(RN, 'A12', 'GB Capital net profit, FY27E (incl associates)', 't'); W(RN, 'B12', "=Segments!F35", 'f')
W(RN, 'A13', 'Normalised group profit', 't', bold=True); W(RN, 'B13', "=B9+B10+B11+B12", 'f', bold=True)
W(RN, 'A14', 'Normalised earnings lens (bear / base / bull)', 't')
W(RN, 'B14', f"=B13*{A['ns_bear']}/{A['shares']}*{A['nm_bear']}", 'f', fmt='0.00')
W(RN, 'C14', f"=B13/{A['shares']}*{A['nm_base']}", 'f', fmt='0.00')
W(RN, 'D14', f"=B13*{A['ns_bull']}/{A['shares']}*{A['nm_bull']}", 'f', fmt='0.00')
W(RN, 'A16', 'The two distortions, priced (per share, FY26E basis)', 'h')
W(RN, 'A17', 'Tax above statute on unshielded regional losses', 't')
W(RN, 'B17', f"=({PR('etr', 'E')}-{A['tax']})*('Income Statement'!E14/(1-{PR('etr', 'E')}))/{A['shares']}", 'f', fmt='0.00')
W(RN, 'A18', 'Funding cost above the eased-cycle rate', 't')
W(RN, 'B18', f"={A['nd']}*({PR('kd_fwd', 'E')}-{A['fin_norm']})*(1-{A['tax']})/{A['shares']}", 'f', fmt='0.00')
widths(RN, {'A': 46, 'B': 12, 'C': 12, 'D': 12})

FV = "Fundamental Valuation"
W(FV, 'B2', "Four lenses, one field — both framings carried to the end", 't', bold=True)
W(FV, 'A4', 'Lens', 'h'); W(FV, 'B4', 'Value / share', 'h'); W(FV, 'C4', 'Weight', 'h')
W(FV, 'A5', 'Sum of the parts — round mark', 't'); W(FV, 'B5', "='SOTP Bridge'!B13", 'f', fmt='0.00')
W(FV, 'A6', 'Sum of the parts — balance-sheet mark', 't'); W(FV, 'B6', "='SOTP Bridge'!C13", 'f', fmt='0.00')
W(FV, 'A7', 'Book value & sustainable return', 't')
W(FV, 'B7', f"={A['parent_eq']}/{A['shares']}", 'f', fmt='0.00')
W(FV, 'A8', 'Relative multiples', 't'); W(FV, 'B8', "='Relative & Normalized'!C6", 'f', fmt='0.00')
W(FV, 'A9', 'Normalised earnings power', 't'); W(FV, 'B9', "='Relative & Normalized'!C14", 'f', fmt='0.00')
for rw, wk in ((5, 'w_sotp'), (7, 'w_book'), (8, 'w_rel'), (9, 'w_norm')):
    W(FV, f'C{rw}', f"={A[wk]}", 'f', fmt='0.00')
W(FV, 'C6', "same weight — the framings alternate, they never blend", 'n')
W(FV, 'A11', 'Weighted central — round mark', 't', bold=True)
W(FV, 'B11', "=B5*C5+B7*C7+B8*C8+B9*C9", 'f', fmt='0.00', bold=True)
W(FV, 'A12', 'Weighted central — balance-sheet mark', 't', bold=True)
W(FV, 'B12', "=B6*C5+B7*C7+B8*C8+B9*C9", 'f', fmt='0.00', bold=True)
W(FV, 'A14', 'Published range — all three formulas', 'h')
W(FV, 'A15', 'Bear (balance-sheet-mark world: bear legs, 0.8x book, bear multiples)', 't')
W(FV, 'B15', f"=C5*'SOTP Bridge'!C15+C7*B7*{A['book_bear_m']}+C8*'Relative & Normalized'!B6"
             f"+C9*'Relative & Normalized'!B14", 'f', fmt='0.00')
W(FV, 'A16', 'Base (round-mark central)', 't'); W(FV, 'B16', "=B11", 'f', fmt='0.00')
W(FV, 'A17', 'Bull (round-mark world: bull legs, marked book, bull multiples)', 't')
W(FV, 'B17', f"=C5*'SOTP Bridge'!B16+C7*({A['parent_eq']}+{A['mnt_round']}-{A['mnt_bv']})"
             f"/{A['shares']}+C8*'Relative & Normalized'!D6+C9*'Relative & Normalized'!D14", 'f', fmt='0.00')
widths(FV, {'A': 52, 'B': 13, 'C': 10})

# =====================================================================================
# SUMMARY / SUMMARY FINANCIALS / MONTE CARLO / SENSITIVITY / PER-SHARE / PEER
# =====================================================================================
SU = "Summary"
W(SU, 'B2', "GB Corp — fair value at a glance (refresh of 19 August 2026)", 't', bold=True)
su_rows = [
 (4, 'Last library close (22-Jul-26)', f"={A['spot']}", '0.00'),
 (5, 'Fair value — bear', "='Fundamental Valuation'!B15", '0.00'),
 (6, 'Fair value — base (round mark)', "='Fundamental Valuation'!B16", '0.00'),
 (7, 'Central under the balance-sheet mark', "='Fundamental Valuation'!B12", '0.00'),
 (8, 'Fair value — bull', "='Fundamental Valuation'!B17", '0.00'),
 (10, 'MNT-Halan gap per share (round vs book, post-discount)',
      f"=({A['mnt_round']}-{A['mnt_bv']})*(1-{A['disc']})/{A['shares']}", '0.00'),
 (11, 'Terminal share of automotive enterprise value', "=DCF!B44", '0.0%'),
]
for rw, lab, f, fmt in su_rows:
    W(SU, f'A{rw}', lab, 't'); W(SU, f'B{rw}', f, 'f', fmt=fmt)
h1_facts = [(13, 'H1-26 group revenue (+35.2%)', h1['rev']),
            (14, 'H1-26 group net profit (−24.5%)', h1['np']),
            (15, 'Automotive net debt 30-Jun-26 (2.14x EBITDA)', a1['nd'])]
for rw, lab, v in h1_facts:
    W(SU, f'A{rw}', lab, 't'); W(SU, f'B{rw}', v, 'in')
W(SU, 'A17', f"The July study's published range was {pub['fair']['bear']} / {pub['fair']['base']} / "
             f"{pub['fair']['full']}. This refresh re-built the model bottom-up on the "
             f"first-half accounts and made the terminal value pay for its growth.", 'n')
widths(SU, {'A': 52, 'B': 14})

SF = "Summary Financials"
W(SF, 'B2', "Summary financials (links)", 't', bold=True)
for j, y in enumerate(['FY23', 'FY24', 'FY25'] + YRS):
    W(SF, f'{get_column_letter(2+j)}4', y, 'h')
W(SF, 'A5', 'Group revenue', 't'); W(SF, 'A6', 'Automotive EBITDA', 't')
W(SF, 'A7', 'Group net profit (attributable)', 't'); W(SF, 'A8', 'Earnings per share (EGP)', 't')
for j, y in enumerate(['FY23', 'FY24', 'FY25']):
    c = get_column_letter(2+j)
    W(SF, f'{c}5', f"='Income Statement'!{c}8", 'f')
    W(SF, f'{c}6', hist[y]['auto_ebitda'], 'in')
    W(SF, f'{c}7', f"='Income Statement'!{c}16", 'f')
    W(SF, f'{c}8', f"='Income Statement'!{c}17", 'f', fmt='0.00')
for col in COLS:
    W(SF, f'{col}5', f"='Income Statement'!{col}8", 'f')
    W(SF, f'{col}6', f"=DCF!{col}14", 'f')
    W(SF, f'{col}7', f"='Income Statement'!{col}16", 'f')
    W(SF, f'{col}8', f"='Income Statement'!{col}17", 'f', fmt='0.00')
widths(SF, {'A': 34, 'B': 11, 'C': 11, 'D': 11, 'E': 11, 'F': 11, 'G': 11, 'H': 11, 'I': 11})

MC = "Monte Carlo"
W(MC, 'B2', "Published price-probability map — reproduced unchanged from the live page", 't', bold=True)
W(MC, 'A4', f"A copy, not a computation: struck on the closing library of {pub['asof']['mc']['data']}, "
            f"computed {pub['asof']['mc']['computed']}, anchored at the {D['spot']:.2f} close. This "
            f"refresh changes the fundamental range only; the map was NOT re-struck and its "
            f"1-month window resolves {pub['dist']['t20']['resolve']}. A fresh price data-set "
            f"re-strikes it on its own clock.", 'n')
W(MC, 'A6', 'Horizon', 'h')
for j, p in enumerate(['5th', '25th', 'median', '75th', '95th']):
    W(MC, f'{get_column_letter(3+j)}6', p, 'h')
for i, (hz, lab) in enumerate([('t20', '1 month (resolves 23-Aug-26)'),
                               ('t60', '3 months (resolves 22-Oct-26)')]):
    W(MC, f'A{7+i}', lab, 't')
    dd = pub['dist'][hz]
    for j, k in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        W(MC, f'{get_column_letter(3+j)}{7+i}', dd[k], 'in', fmt='0.00')
W(MC, 'A10', 'Touch ladder (level, probability within 1 month / 3 months, %)', 't')
for i, (lvl, p1, p3) in enumerate(pub['touch']):
    W(MC, f'B{11+i}', lvl, 'in', fmt='0.00')
    W(MC, f'C{11+i}', p1, 'in', fmt='0'); W(MC, f'D{11+i}', p3, 'in', fmt='0')
widths(MC, {'A': 44, 'B': 10, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10})

SE = "Sensitivity"
W(SE, 'B2', "Live sensitivities — every cell a formula", 't', bold=True)
W(SE, 'A4', "Sum of the parts per share: stake mark x holding discount", 'h')
gm = D['sens']['grid_mult']; gd = D['sens']['grid_disc']
for j, d_ in enumerate(gd):
    W(SE, f'{get_column_letter(3+j)}5', d_, 'in', fmt='0%')
for i, m_ in enumerate(gm):
    W(SE, f'B{6+i}', m_, 'in', fmt='0%')
    for j in range(len(gd)):
        col = get_column_letter(3+j)
        W(SE, f'{col}{6+i}', f"=(DCF!$B$42+{A['cap_oper']}+{A['mnt_round']}*$B{6+i}"
                             f"+{A['other_assoc']}+{A['fvoci']})*(1-{col}$5)/{A['shares']}", 'f', fmt='0.0')
W(SE, 'A12', "The crux as a ladder: the second close, priced through the whole synthesis", 'h')
W(SE, 'A13', 'Round valuation (USD mn)', 'h'); W(SE, 'B13', 'Stake value (EGP mn)', 'h')
W(SE, 'C13', 'Sum of the parts / share', 'h'); W(SE, 'D13', 'Weighted central / share', 'h')
for i, x in enumerate([1000, 1200, 1400, 1600]):
    rr2 = 14+i
    W(SE, f'A{rr2}', x, 'in', fmt='#,##0')
    W(SE, f'B{rr2}', f"={A['mnt_stake']}*A{rr2}*{A['fx_spot']}", 'f')
    W(SE, f'C{rr2}', f"=(DCF!$B$42+{A['cap_oper']}+B{rr2}+{A['other_assoc']}+{A['fvoci']})"
                     f"*(1-{A['disc']})/{A['shares']}", 'f', fmt='0.00')
    W(SE, f'D{rr2}', f"='Fundamental Valuation'!C5*C{rr2}+'Fundamental Valuation'!C7*'Fundamental Valuation'!B7"
                     f"+'Fundamental Valuation'!C8*'Fundamental Valuation'!B8"
                     f"+'Fundamental Valuation'!C9*'Fundamental Valuation'!B9", 'f', fmt='0.00')
W(SE, 'A18', 'the balance-sheet carrying instead', 't')
W(SE, 'B18', f"={A['mnt_bv']}", 'f')
W(SE, 'C18', "='SOTP Bridge'!C13", 'f', fmt='0.00'); W(SE, 'D18', "='Fundamental Valuation'!B12", 'f', fmt='0.00')
W(SE, 'A20', "Terminal-value construction (the disciplined base vs the alternatives)", 'h')
W(SE, 'A21', 'Terminal return on capital', 'h'); W(SE, 'B21', 'Automotive equity (EGP mn)', 'h')
W(SE, 'C21', 'Per share', 'h')
for i, rv in enumerate([0.15, 0.20, 0.225, 0.25]):
    rr3 = 22+i
    W(SE, f'A{rr3}', rv, 'in', fmt='0.000')
    W(SE, f'B{rr3}', f"=DCF!B34+DCF!I12*(1+{A['tg']})*(1-{A['tg']}/A{rr3})/({A['wacc']}-{A['tg']})"
                     f"*DCF!I31-{A['nd']}-{A['nci']}", 'f')
    W(SE, f'C{rr3}', f"=B{rr3}/{A['shares']}", 'f', fmt='0.00')
W(SE, 'A26', 'July construction (Gordon on year-five cash flow)', 't')
W(SE, 'B26', "=DCF!B45", 'f'); W(SE, 'C26', f"=B26/{A['shares']}", 'f', fmt='0.00')
widths(SE, {'A': 34, 'B': 16, 'C': 14, 'D': 16, 'E': 10, 'F': 10, 'G': 10})

PS = "Per-Share & Ratios"
W(PS, 'B2', "Per-share values and ratios", 't', bold=True)
for j, y in enumerate(YRS):
    W(PS, f'{COLS[j]}4', y, 'h')
ps_rows = [
 (5, 'Earnings per share', lambda col: f"='Income Statement'!{col}17", '0.00'),
 (6, 'Book value per share', lambda col: f"='Balance Sheet'!{col}16", '0.00'),
 (7, 'Return on equity', lambda col: f"='Income Statement'!{col}16/'Balance Sheet'!{col}11", '0.0%'),
 (8, 'Net debt / automotive EBITDA', lambda col: f"='Balance Sheet'!{col}10/DCF!{col}14", '0.00'),
 (9, 'Price / earnings at the last close', lambda col: f"={A['spot']}/'Income Statement'!{col}17", '0.0'),
 (10, 'Price / book at the last close', lambda col: f"={A['spot']}/'Balance Sheet'!{col}16", '0.00'),
 (11, 'Dividend yield at the last close', lambda col: f"={A['div_ps']}/{A['spot']}", '0.0%'),
]
for rw, lab, fn, fmt in ps_rows:
    W(PS, f'A{rw}', lab, 't')
    for col in COLS:
        W(PS, f'{col}{rw}', fn(col), 'f', fmt=fmt)
widths(PS, {'A': 40, 'E': 11, 'F': 11, 'G': 11, 'H': 11, 'I': 11})

PE = "Peer & Sector"
W(PE, 'B2', "Peer frame (context only — never a source for the subject's numbers)", 't', bold=True)
W(PE, 'A4', 'Peer', 'h'); W(PE, 'B4', 'Trailing P/E', 'h'); W(PE, 'C4', 'Note', 'h')
peers_rows = [
 ('Contact Financial (EGX) — consumer/NBFS', D['peers']['CNFN'], 'the direct Egyptian financing peer'),
 ('Dogus Otomotiv (Istanbul) — auto distribution', D['peers']['DOAS'], 'P/B 0.62 on inflation-restated book'),
 ('AutoNation (US) — auto retail', D['peers']['AN'], 'mature-market anchor; forward 8.35'),
 ('Bajaj Auto (India) — two/three-wheelers', D['peers']['BAJAJ'], 'partner brand; far richer market'),
]
for i, (nm, pe_, note) in enumerate(peers_rows):
    W(PE, f'A{5+i}', nm, 't'); W(PE, f'B{5+i}', pe_, 'in', fmt='0.00'); W(PE, f'C{5+i}', note, 'n')
W(PE, 'A10', 'GB Corp at the last close on FY26E earnings', 't')
W(PE, 'B10', f"={A['spot']}/'Income Statement'!E17", 'f', fmt='0.0')
W(PE, 'A11', 'GB Corp at the last close on restated book', 't')
W(PE, 'B11', f"={A['spot']}/'Fundamental Valuation'!B7", 'f', fmt='0.00')
W(PE, 'A13', "Peer marks 19-Aug-2026, aggregator-sourced (market data only). Logged discrepancy: "
             "Contact Financial's market capitalisation is quoted between EGP 4.5bn and 6.6bn "
             "across services on different dates.", 'n')
widths(PE, {'A': 44, 'B': 12, 'C': 52})

# =====================================================================================
# READ FIRST
# =====================================================================================
rf_rows = [
 ("GB Corp (GBCO.CA / EGX) — valuation model, fundamental refresh of 19 August 2026", 't', True),
 ("Built bottom-up and formula-pure. Blue cells are the only values in the computation "
  "sheets: each is a disclosed figure from the 30-Jun-2026 reviewed statements or the "
  "2Q/1H26 release, a sourced market quote, or a stated judgement with its rationale. "
  "Every other number is a live formula — FY26E itself is derived (H1 actual + a second "
  "half built on the FY25 measured seasonal split with explicit tempering inputs), "
  "volumes run at the finest disclosed level (locally assembled vs imported cars; buses, "
  "trucks and equipment), unit costs escalate on their own physical classes, and margins "
  "are outputs.", 't', False),
 ("The terminal value pays for its growth: reinvestment is forced to growth divided by "
  "the terminal return on capital (both visible cells on Assumptions/DCF). The July "
  "construction — Gordon on year-five cash flow, whose hidden implied return was 25% — "
  "is shown as a labelled alternative and priced on the Sensitivity sheet.", 't', False),
 ("The contested stake is computed both ways, side by side, never averaged: MNT-Halan at "
  "the June-2026 round mark and at the company's own carrying value, which the reviewer "
  "could not verify (qualified conclusion, second consecutive period).", 't', False),
 ("The price-probability sheet reproduces the published page unchanged (struck 22 July "
  "2026) — this refresh moves the fundamental range only.", 't', False),
 ("Change any blue driver and the fair value reprices through Segments, the cash-flow "
  "model, the bridge and the synthesis. All amounts EGP millions unless stated.", 't', False),
]
r = 2
for txt, kind, bold in rf_rows:
    W("READ FIRST", f"B{r}", txt, kind, bold=bold); r += 2
widths("READ FIRST", {'A': 2, 'B': 110})

# =====================================================================================
# print setup + save + purity audit
# =====================================================================================
for ws in wb.worksheets:
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if ws.title in ("Segments", "Income Statement", "DCF", "Summary Financials",
                    "Balance Sheet", "Assumptions"):
        ws.page_setup.orientation = 'landscape'
wb.save(OUT)

wb2 = openpyxl.load_workbook(OUT)
assert wb2.sheetnames == SHEETS
n_f = n_in = 0
COMPUTE_SHEETS = ["Summary", "Fundamental Valuation", "Assumptions", "SOTP Bridge", "Segments",
                  "Relative & Normalized", "DCF", "Income Statement", "Balance Sheet",
                  "Cash Flow", "Summary Financials", "Sensitivity", "Per-Share & Ratios",
                  "Peer & Sector", "Monte Carlo"]
violations = []
for sn in COMPUTE_SHEETS:
    for row in wb2[sn].iter_rows():
        for c2 in row:
            v = c2.value
            if isinstance(v, str) and v.startswith('='):
                n_f += 1
            elif isinstance(v, (int, float)):
                color = c2.font.color.rgb if c2.font and c2.font.color else None
                if color and str(color)[-6:] == BLUE:
                    n_in += 1
                else:
                    violations.append((sn, c2.coordinate, v))
print(f"saved {OUT}")
print(f"formula cells: {n_f} | blue input cells: {n_in} | non-input value cells: {len(violations)}")
for v in violations[:12]:
    print("  VALUE-CELL VIOLATION:", v)
assert not violations, "derived value cells found — the workbook must be formula-pure"
print("FORMULA-PURITY AUDIT PASS: every numeric cell is a formula or an annotated blue input")
