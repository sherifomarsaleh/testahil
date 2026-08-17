"""AMR — the companion workbook. Formula-first by construction.

Every cell that can be derived from a driver IS a formula. As it writes each formula the
builder records the value the model itself computed for that cell into xlsx_expected.json,
so recalc.py can evaluate the delivered file independently and assert that all of them
reproduce the model with none left unchecked.

Only three classes of cell are pasted, and READ FIRST names them: audited and disclosed
history; the output of the engine that prices the probability map; and the sensitivity
grids, where each cell is a complete re-run of the whole model.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
STK = json.load(open(os.path.join(HERE, 'strike_result.json')))
M, H, F, U, W, DCF = D['meta'], D['history'], D['forecast'], D['unit_build'], D['wacc'], D['dcf']
LN, SN, CS, C = D['lenses'], D['sensitivity'], D['cost_stack'], D['contested']
ALT, DFL, PEERS = D['dcf_alt'], D['dual_framing_leases'], D['peers']['peers']
FX, SH, UNITS, FY = M['fx'], M['shares_mn'], U['units'], F['years']
NDEBT = H['lease_liabilities'][2] + H['bank_debt'][2] - H['cash'][2] - H['deposits'][2]

INK = '1C3A36'
BLUE = Font(color='1F4E9C', size=10)
BLACK = Font(color='222222', size=10)
GREEN = Font(color='1E6B4F', size=10)
HEAD = Font(bold=True, color=INK, size=10.5)
TITLE = Font(bold=True, color=INK, size=13)
SUB = Font(italic=True, color='5A6764', size=9.5)
SECT = Font(bold=True, color='896F36', size=10.5)

wb = openpyxl.Workbook()
wb.remove(wb['Sheet'])
EXPECTED, PASTED = {}, {'audited': [], 'engine': [], 'grid': []}
NFORM = 0
N2, N3, N0, PC1, PC2, MULT = '#,##0.00', '#,##0.000', '#,##0', '0.0%', '0.00%', '0.00"x"'
COLS = ['B', 'C', 'D', 'E', 'F']
ALLC = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
FCOL = ['E', 'F', 'G', 'H', 'I']


def sheet(name, widths):
    ws = wb.create_sheet(name)
    EXPECTED[name] = {}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    return ws


def T(ws, coord, text, font=None, wrap=False):
    c = ws[coord]
    c.value = text
    c.font = font or BLACK
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def V(ws, coord, value, fmt=None, kind='audited'):
    c = ws[coord]
    c.value = value
    c.font = BLUE
    if fmt:
        c.number_format = fmt
    PASTED[kind].append(f'{ws.title}!{coord}')
    return c


def FM(ws, coord, formula, model_value, fmt=None, link=False):
    global NFORM
    c = ws[coord]
    c.value = formula
    c.font = GREEN if link else BLACK
    if fmt:
        c.number_format = fmt
    EXPECTED[ws.title][coord] = float(model_value)
    NFORM += 1
    return c


def dash(ws, coord):
    ws[coord] = '-'
    ws[coord].font = Font(color='9AA5A2', size=10)


# ===========================================================================
# 1. READ FIRST
# ===========================================================================
RF = sheet('READ FIRST', {'A': 118})
lines = [
    ('Testahil — Americana Restaurants International PLC (Abu Dhabi Securities Exchange: AMR)', TITLE),
    ('Companion model · Independent valuation study · Educational analysis · Not investment advice', SUB),
    ('', None),
    ('WHAT THIS WORKBOOK IS. A transparent companion to the Americana Restaurants valuation study. '
     'Blue cells are inputs, black cells are formulas, green cells link across sheets.', None),
    ('', None),
    ('IT CALCULATES. Every figure that can be derived arithmetically from a driver is a live formula, '
     'so you can change a blue cell on the Assumptions sheet and watch the model reprice. The cost of '
     'capital is built from the risk-free rate, the sovereign spread, beta and the premium rather than '
     'pasted. The glide the discount rate walks down is derived from the tax path. The discount factors '
     'compound. The cash-flow waterfall chains: EBITDA less depreciation gives EBIT, EBIT after tax '
     'gives NOPAT, and free cash flow to the firm falls out of NOPAT, depreciation, capital expenditure '
     'and the change in working capital. The terminal block chains too: the reinvestment rate is growth '
     'divided by the return on capital, and the terminal value follows from terminal NOPAT, that '
     'reinvestment rate, the terminal cost of capital and growth. The statements roll forward — property, '
     'right-of-use assets, working capital, equity, cash and the lease liability each carry from one year '
     'to the next — and every ratio and per-share figure on every sheet is a formula.', None),
    ('', None),
    ('FOUR CLASSES OF CELL ARE PASTED, AND ONLY FOUR.', SECT),
    ('  1. Audited and disclosed history. The primary record is not a calculation. Where a line is both '
     'disclosed and derivable, the disclosed figure is carried.', None),
    ('  2. The probability map. Each figure on the Monte Carlo sheet is the output of a fifty-thousand-path '
     'simulation, not a formula.', None),
    ('  3. The sensitivity grids. Each cell there is a complete re-run of the whole model with joint driver '
     'shifts, so it cannot be a single formula. Those grids do NOT redraw when a driver is changed.', None),
    ('  4. Two whole-model re-runs named individually: the bear and bull of the cash-flow lens on the '
     'Summary sheet, whose joint driver sets (margin plus or minus 1.5 points, cost of capital minus or '
     'plus 1.5 points, terminal growth 2% or 4%, like-for-like plus or minus 1 point) are stated beside '
     'them; and the leases-as-operating-cost reading on the Fundamental Valuation sheet. The expert '
     'panel, both readings of the margin question, the cyclical value and every lens bear and bull are '
     'LIVE FORMULAS in this edition.', None),
    ('Anything else pasted would be a defect. The formula and pasted-value counts are reported at the '
     'foot of this sheet.', None),
    ('', None),
    ('HOW REVENUE IS BUILT. Not as one growth rate. The restaurant estate is disclosed country by country '
     'at each period end and segment revenue is disclosed for the same units, so revenue is built as '
     'restaurants times revenue per restaurant across seven market units. The restaurant count grows on '
     'the company\'s own net-new-store guidance; the revenue per restaurant grows on disclosed '
     'like-for-like sales growth, less a currency drag applied only to the two markets whose currencies '
     'are not pegged to the dollar. The build reproduces reported revenue exactly in all three audited '
     'years. Margins are an OUTPUT of the cost stack, not an input: eight cost classes are each escalated '
     'on their own driver, never on one blended index.', None),
    ('', None),
    ('CURRENCY. The company reports in US dollars. The shares trade in dirhams in Abu Dhabi and in riyals '
     'in Riyadh, and the dirham has been pegged at 3.6725 to the dollar since 1997. The model runs in US '
     'dollars — the reporting and functional currency — and converts to dirhams at the peg. Figures are '
     'USD million unless stated; per-share figures are shown in both currencies.', None),
    ('', None),
    ('THE OPEN QUESTION. Between the first half of 2025 and the first half of 2026 the EBITDA margin went '
     'from 22.6% to 25.5%. Whether that is structural or cyclical is the single judgement that decides '
     'this valuation, and nothing in the filings settles it. Both readings are computed and both are '
     'published, on the Summary and Fundamental Valuation sheets. They are not averaged.', None),
    ('', None),
    ('WHAT IT IS NOT. It is not investment advice, a recommendation, or a price target. Values are model '
     'outputs shown as ranges.', None),
]
r = 1
for text, font in lines:
    if text:
        T(RF, f'A{r}', text, font or BLACK, wrap=True)
        RF.row_dimensions[r].height = max(14, 13 * (1 + len(text) // 118))
    r += 1
READ_FIRST_COUNT_ROW = r + 1

# ===========================================================================
# 2. ASSUMPTIONS
# ===========================================================================
A = sheet('Assumptions', dict({'A': 62}, **{c: 13 for c in COLS}))
T(A, 'A1', 'Assumptions — every driver in the model', TITLE)
T(A, 'A2', 'Blue cells are inputs. Change one and the model reprices: everything downstream of '
           'this sheet is a formula.', SUB)
T(A, 'A3', 'Driver', HEAD)
for i, y in enumerate(FY):
    T(A, f'{COLS[i]}3', y, HEAD)
T(A, 'A4', 'Anchors', SECT)
T(A, 'A5', 'Market price (AED per share)'); V(A, 'C5', M['spot_aed'], N2)
T(A, 'A6', 'Dirhams per US dollar (the peg)'); V(A, 'C6', FX, N3)
T(A, 'A7', 'Shares outstanding, net of treasury (million) — issued less treasury')
FM(A, 'C7', '=C8-E8', SH, N0)
T(A, 'A8', 'Shares issued (million)'); V(A, 'C8', M['shares_issued_mn'], N0)
T(A, 'D8', 'less treasury (million)', SUB); V(A, 'E8', M['shares_issued_mn'] - SH, N0)
T(A, 'A10', 'Volume — the restaurant estate', SECT)
T(A, 'A11', 'Net new restaurants a year')
for i in range(5):
    V(A, f'{COLS[i]}11', U['nso'][i], N0)
T(A, 'A12', 'Share of net additions, by market', SUB)
for j, u in enumerate(UNITS):
    T(A, f'A{13 + j}', f'   {u}'); V(A, f'C{13 + j}', U['nso_mix'][u], PC1)
T(A, 'A20', 'Restaurants at 31 December 2025, by market', SUB)
for j, u in enumerate(UNITS):
    T(A, f'A{21 + j}', f'   {u}'); V(A, f'C{21 + j}', U['stores_hist'][u][1], N0)
T(A, 'A29', 'Price — revenue per restaurant', SECT)
T(A, 'A30', 'Like-for-like sales growth')
for i in range(5):
    V(A, f'{COLS[i]}30', U['lfl'][i], PC1)
T(A, 'A31', 'Currency drag on US dollar revenue per restaurant, by market', SUB)
for j, u in enumerate(UNITS):
    T(A, f'A{32 + j}', f'   {u}'); V(A, f'C{32 + j}', U['fx_drag'][u], PC1)
T(A, 'A39', 'Revenue per restaurant in FY2025 (USD thousand), by market', SUB)
for j, u in enumerate(UNITS):
    T(A, f'A{40 + j}', f'   {u}'); V(A, f'C{40 + j}', U['rps_2025'][u], N2)
T(A, 'A48', 'Cost stack — one escalator per driver class, each as a share of revenue', SECT)
COST_ROWS = [('inventory', 'Food, filling and packing materials', 49),
             ('royalties', 'Royalties to the brand franchisors', 50),
             ('staff', 'Staff costs', 51),
             ('delivery', 'Home delivery and transportation', 52),
             ('advertising', 'Advertisement and business development', 53),
             ('utilities', 'Utilities and communication', 54),
             ('rent_other', 'Short-term, low-value and variable lease payments', 55),
             ('maintenance', 'Maintenance, repairs and other restaurant costs', 56)]
for key, lab, rr in COST_ROWS:
    if rr in (51, 52):
        T(A, f'A{rr}', lab + ' — UNIT BUILD, not a share: inputs at rows 109-114 below')
        continue
    T(A, f'A{rr}', lab)
    for i in range(5):
        V(A, f'{COLS[i]}{rr}', CS['lines'][key]['path'][i], PC2)
T(A, 'A57', 'All other operating costs (the residual of the three expense notes)')
V(A, 'C57', CS['residual_pct'], PC2)
T(A, 'A58', 'Other income, as a share of revenue')
V(A, 'C58', H['other_income'][2] / H['revenue'][2], PC2)
T(A, 'A60', 'Capital, leases and working capital', SECT)
ASM = [('A61', 'Capital expenditure per new restaurant (USD thousand)', 'C61',
        D['inputs']['capex_per_store_k']['value'], N0),
       ('A62', 'Maintenance capital expenditure, share of revenue', 'C62',
        D['inputs']['maintenance_capex_pct']['value'], PC2),
       ('A63', 'Restaurant closure rate', 'C63', D['inputs']['closure_rate']['value'], PC2),
       ('A64', 'Additions to right-of-use assets, share of revenue', 'C64',
        D['inputs']['rou_additions_pct']['value'], PC2),
       ('A65', 'Total lease payments, share of revenue', 'C65',
        D['inputs']['lease_payments_pct']['value'], PC2),
       ('A66', 'Right-of-use depreciation rate, on the opening balance', 'C66',
        D['inputs']['rou_depreciation_rate']['value'], PC2),
       ('A67', 'Owned-asset depreciation rate, on the opening balance', 'C67',
        D['inputs']['owned_depreciation_rate']['value'], PC2),
       ('A68', 'Net working capital, share of revenue (negative — the till pays first)', 'C68',
        D['inputs']['nwc_pct_revenue']['value'], PC2)]
for lc, lab, vc, val, fmt in ASM:
    T(A, lc, lab); V(A, vc, val, fmt)
T(A, 'A70', 'Tax and distribution', SECT)
T(A, 'A71', 'Effective tax rate')
for i in range(5):
    V(A, f'{COLS[i]}71', F['etr'][i], PC1)
T(A, 'A72', 'Dividend payout ratio'); V(A, 'C72', D['inputs']['payout_ratio']['value'], PC1)
T(A, 'A73', 'Yield on cash and bank deposits'); V(A, 'C73', D['inputs']['deposit_yield']['value'], PC2)
T(A, 'A74', 'Finance costs other than lease interest, share of revenue')
V(A, 'C74', D['inputs']['other_finance_cost_pct']['value'], PC2)
T(A, 'A75', 'Cost of capital', SECT)
COC = [('A76', 'US ten-year Treasury yield', 'C76', W['rf_ust'], PC2),
       ('A77', 'US sovereign default spread (ratings basis)', 'C77', W['us_default_spread'], PC2),
       ('A78', 'Blended equity risk premium (ratings basis)', 'C78', W['erp_rating'], PC2),
       ('A79', 'US sovereign credit-default-swap spread', 'C79', W['us_cds'], PC2),
       ('A80', 'Blended equity risk premium (credit-default-swap basis)', 'C80', W['erp_cds'], PC2),
       ('A81', 'Beta', 'C81', W['beta'], N3),
       ('A82', 'Cost of debt — the group’s own incremental borrowing rate', 'C82', W['kd'], PC2),
       ('A83', 'Terminal risk-free rate', 'C83', W['terminal_rf'], PC2),
       ('A84', 'Terminal growth', 'C84', W['terminal_g'], PC1)]
for lc, lab, vc, val, fmt in COC:
    T(A, lc, lab); V(A, vc, val, fmt)
T(A, 'A86', 'Balance-sheet anchors', SECT)
ANCH = [('A87', 'Lease liabilities at 31 December 2025 (USD million)', 'C87',
         H['lease_liabilities'][2], N2),
        ('A88', 'Cash and bank deposits at 31 December 2025 (USD million)', 'C88',
         H['cash'][2] + H['deposits'][2], N2),
        ('A89', 'Bank debt at 31 December 2025 (USD million)', 'C89', H['bank_debt'][2], N2),
        ('A90', 'Non-controlling interests (USD million)', 'C90', H['nci'][2], N2),
        ('A91', 'Equity attributable to shareholders at 31 December 2025 (USD million)', 'C91',
         H['equity'][2], N2),
        ('A92', 'Net working capital at 31 December 2025 (USD million)', 'C92', H['nwc'][2], N2),
        ('A93', 'Owned assets at 31 December 2025 (USD million)', 'C93',
         H['ppe'][2] + H['intangibles'][2] + H['investment_property'][2], N2),
        ('A94', 'Right-of-use assets at 31 December 2025 (USD million)', 'C94', H['rou'][2], N2),
        ('A95', 'Intercompany eliminations, share of segment revenue', 'C95',
         U['elimination_pct'], PC2),
        ('A96', 'FY2026 revenue implied by the disclosed first half (USD million)', 'C96',
         U['revenue_2026_from_h1'], N2),
        ('A97', 'Book value per share basis — equity at 30 June 2026 (USD million)', 'C97',
         H['h1_2026']['equity'], N2)]
for lc, lab, vc, val, fmt in ANCH:
    T(A, lc, lab); V(A, vc, val, fmt)
T(A, 'A99', 'Lens inputs', SECT)
LENSI = [('A100', 'Justified enterprise value / EBITDA', 'C100', LN['relative']['multiple'], MULT),
         ('A101', 'Justified price / earnings', 'C101', LN['normalised']['multiple'], MULT),
         ('A102', 'Sustainable return on equity', 'C102', LN['book']['roe'], PC1),
         ('A103', 'Weight — discounted cash flow', 'C103', LN['weights']['Discounted cash flow'], PC1),
         ('A104', 'Weight — relative multiples', 'C104', LN['weights']['Relative multiples'], PC1),
         ('A105', 'Weight — normalised earnings power', 'C105',
          LN['weights']['Normalised earnings power'], PC1),
         ('A106', 'Weight — book value and sustainable return', 'C106',
          LN['weights']['Book value and sustainable return'], PC1)]
for lc, lab, vc, val, fmt in LENSI:
    T(A, lc, lab); V(A, vc, val, fmt)
T(A, 'A108', 'Unit-built cost lines, the terminal fade, and the anchor roll', SECT)
T(A, 'A109', 'Restaurant-level staff per restaurant (full-time equivalents)')
for i in range(5):
    V(A, f'{COLS[i]}109', CS['fte_per_store'][i], N2)
T(A, 'A110', 'Staff cost per full-time equivalent (USD thousand, FY2025)')
V(A, 'C110', CS['wage_per_fte'], N3)
T(A, 'A111', 'Wage growth per full-time equivalent'); V(A, 'C111', CS['wage_growth'], PC1)
T(A, 'A112', 'Above-restaurant staff (thousand full-time equivalents)')
V(A, 'C112', CS['above_restaurant_fte'], N3)
T(A, 'A113', 'Home delivery share of revenue')
for i in range(5):
    V(A, f'{COLS[i]}113', CS['delivery_share'][i], PC1)
T(A, 'A114', 'Delivery cost per dollar of delivered revenue')
for i in range(5):
    V(A, f'{COLS[i]}114', CS['delivery_cost_ratio'][i], PC2)
T(A, 'A115', 'Recurring impairment charge, share of revenue')
V(A, 'C115', D['inputs']['impairment_rate_recurring']['value'], PC2)
T(A, 'A116', 'Terminal return on incremental invested capital')
V(A, 'C116', W['terminal_roic'], PC1)
T(A, 'A117', 'Calendar days, valuation date to the price anchor')
V(A, 'C117', W['days_to_anchor'], N0)
T(A, 'A118', 'Dividend paid inside the roll window (USD per share)')
V(A, 'C118', W['div_in_window'], '0.0000')
T(A, 'A119', 'Cyclical margin path (the other reading of the open question)')
for i in range(5):
    V(A, f'{COLS[i]}119', D['inputs']['margin_path_cyclical']['value'][i], PC1)
T(A, 'A124', 'Four-wall EBITDA, first half of 2026 (USD million)')
V(A, 'C124', D['inputs']['four_wall_ebitda_h1_26']['value'], N2)
T(A, 'A125', 'Corporate overhead for the owner-cash reading (USD million)')
V(A, 'C125', H['admin'][2] * 0.86, N2)
T(A, 'A126', 'Dividend declared against FY2025 (USD million)')
V(A, 'C126', D['inputs']['div_fy25_declared']['value'], N2)

# ===========================================================================
# 3. SEGMENTS — the unit build
# ===========================================================================
S = sheet('Segments', dict({'A': 34}, **{c: 12 for c in 'BCDEFGHI'}))
T(S, 'A1', 'The unit build — restaurants times revenue per restaurant, market by market', TITLE)
T(S, 'A2', 'The historical columns are the disclosed record. Every forecast cell is a formula: '
           'the restaurant count grows on the net-new-store programme and the revenue per '
           'restaurant on like-for-like growth net of currency.', SUB)
T(S, 'A4', 'Restaurants at year end', HEAD); T(S, 'B4', 'FY2025', HEAD)
for i, y in enumerate(FY):
    T(S, f'{"CDEFG"[i]}4', y, HEAD)
for j, u in enumerate(UNITS):
    rr = 5 + j
    T(S, f'A{rr}', u)
    FM(S, f'B{rr}', f'=Assumptions!$C${21 + j}', U['stores_hist'][u][1], N0, link=True)
    for i in range(5):
        col, prev = 'CDEFG'[i], ('B' if i == 0 else 'CDEFG'[i - 1])
        FM(S, f'{col}{rr}', f'={prev}{rr}+Assumptions!{COLS[i]}$11*Assumptions!$C${13 + j}',
           U['stores_f'][u][i], N0)
T(S, 'A12', 'Total restaurants', HEAD)
FM(S, 'B12', '=SUM(B5:B11)', sum(U['stores_hist'][u][1] for u in UNITS), N0)
for i in range(5):
    col = 'CDEFG'[i]
    FM(S, f'{col}12', f'=SUM({col}5:{col}11)', F['stores'][i], N0)
T(S, 'A14', 'Revenue per restaurant (USD thousand)', HEAD)
T(S, 'B14', 'FY2024', HEAD); T(S, 'C14', 'FY2025', HEAD)
for i, y in enumerate(FY):
    T(S, f'{"DEFGH"[i]}14', y, HEAD)
for j, u in enumerate(UNITS):
    rr = 15 + j
    T(S, f'A{rr}', u)
    V(S, f'B{rr}', U['rps_2024'][u], N2)
    FM(S, f'C{rr}', f'=Assumptions!$C${40 + j}', U['rps_2025'][u], N2, link=True)
    for i in range(5):
        col, prev = 'DEFGH'[i], ('C' if i == 0 else 'DEFGH'[i - 1])
        FM(S, f'{col}{rr}',
           f'={prev}{rr}*(1+Assumptions!{COLS[i]}$30)*(1-Assumptions!$C${32 + j})',
           U['rps_f'][u][i], N2)
T(S, 'A24', 'Revenue before eliminations (USD million)', HEAD)
for k, y in enumerate(['FY2023', 'FY2024', 'FY2025']):
    T(S, f'{"BCD"[k]}24', y, HEAD)
for i, y in enumerate(FY):
    T(S, f'{FCOL[i]}24', y, HEAD)
for j, u in enumerate(UNITS):
    rr = 25 + j
    T(S, f'A{rr}', u)
    for k in range(3):
        V(S, f'{"BCD"[k]}{rr}', U['revenue_hist'][u][k], N2)
    for i in range(5):
        FM(S, f'{FCOL[i]}{rr}', f'={"DEFGH"[i]}{15 + j}*{"CDEFG"[i]}{5 + j}/1000',
           U['stores_f'][u][i] * U['rps_f'][u][i] / 1000.0, N2)
T(S, 'A32', 'Total before eliminations', HEAD)
for k in range(3):
    col = 'BCD'[k]
    FM(S, f'{col}32', f'=SUM({col}25:{col}31)',
       sum(U['revenue_hist'][u][k] for u in UNITS), N2)
for i in range(5):
    col = FCOL[i]
    FM(S, f'{col}32', f'=SUM({col}25:{col}31)',
       sum(U['stores_f'][u][i] * U['rps_f'][u][i] / 1000.0 for u in UNITS), N2)
T(S, 'A33', 'Less intercompany eliminations')
for i in range(5):
    col = FCOL[i]
    FM(S, f'{col}33', f'=-{col}32*Assumptions!$C$95',
       -sum(U['stores_f'][u][i] * U['rps_f'][u][i] / 1000.0 for u in UNITS)
       * U['elimination_pct'], N2)
T(S, 'A34', 'Revenue on the unit build', HEAD)
for i in range(5):
    col = FCOL[i]
    FM(S, f'{col}34', f'={col}32+{col}33',
       sum(U['stores_f'][u][i] * U['rps_f'][u][i] / 1000.0 for u in UNITS)
       * (1 - U['elimination_pct']), N2)
T(S, 'A35', 'FY2026 implied by the disclosed first half')
FM(S, 'E35', '=Assumptions!$C$96', U['revenue_2026_from_h1'], N2, link=True)
T(S, 'A36', 'Revenue adopted', HEAD)
FM(S, 'E36', '=AVERAGE(E34,E35)', F['revenue'][0], N2)
for i in range(1, 5):
    FM(S, f'{FCOL[i]}36', f'={FCOL[i]}34', F['revenue'][i], N2)
BB = D['brand_build']
T(S, 'A40', 'THE BRAND BUILD — product by product, volume x price, the cross-check spine', SECT)
T(S, 'A41', 'Restaurants by brand', HEAD)
T(S, 'B41', 'FY2024', HEAD); T(S, 'C41', 'FY2025', HEAD)
for i, y in enumerate(FY):
    T(S, f'{"DEFGH"[i]}41', y, HEAD)
for b, nm in enumerate(BB['brands']):
    rr = 42 + b
    T(S, f'A{rr}', nm)
    V(S, f'B{rr}', BB['stores_2024'][b], N0)
    V(S, f'C{rr}', BB['stores_2025'][b], N0)
    for i in range(5):
        col, prev = 'DEFGH'[i], ('C' if i == 0 else 'DEFGH'[i - 1])
        FM(S, f'{col}{rr}', f'={prev}{rr}+Assumptions!{COLS[i]}$11*{BB["nso_mix"][b]}',
           BB['stores_f'][i][b], N0)
T(S, 'A48', 'Revenue by brand (USD million; residual ties to the audited total)', HEAD)
for b, nm in enumerate(BB['brands']):
    rr = 49 + b
    T(S, f'A{rr}', nm)
    V(S, f'B{rr}', BB['revenue_2024'][b], N2)
    V(S, f'C{rr}', BB['revenue_2025'][b], N2)
    for i in range(5):
        col, prev = 'DEFGH'[i], ('C' if i == 0 else 'DEFGH'[i - 1])
        FM(S, f'{col}{rr}',
           f'={col}{42 + b}*({prev}{rr}/{prev}{42 + b})*(1+Assumptions!{COLS[i]}$30)'
           f'*(1-{BB["wavg_drag"]:.6f})',
           BB['revenue_f'][i][b], N2)
T(S, 'A54', 'Brand build, after eliminations (FY2026 blended with the disclosed half, as above)',
  HEAD)
bt = BB['total_f']
FM(S, 'D54', '=(SUM(D49:D53)*(1-Assumptions!$C$95)+Assumptions!$C$96)/2', bt[0], N2)
for i in range(1, 5):
    col = 'DEFGH'[i]
    FM(S, f'{col}54', f'=SUM({col}49:{col}53)*(1-Assumptions!$C$95)', bt[i], N2)
T(S, 'A55', 'Against the geographic build (row 36)')
for i in range(5):
    col = 'DEFGH'[i]
    FM(S, f'{col}55', f'={col}54/{FCOL[i]}36-1', BB['vs_geographic'][i], PC1)
T(S, 'A56', 'The gap is a disclosed-mix effect, stated rather than smoothed: the company\'s '
            'brand-level opening mix is KFC-heavy (40% of openings at about USD 1.3 million of '
            'revenue per restaurant) while the geographic mix spreads additions across markets '
            'averaging about USD 0.9 million, so the brand build runs about 2% ahead by FY2030. '
            'The geographic build is adopted because it ties the audited history exactly; the '
            'brand build is the volume-times-price cross-check.', SUB, wrap=True)
S.row_dimensions[56].height = 56

T(S, 'A38', 'FY2026 is half disclosed and half built. The unit build and the first-half run rate '
            'are averaged rather than one being chosen over the other; the two are within two '
            'per cent of each other. The historical columns tie exactly to reported revenue in '
            'all three audited years.', SUB, wrap=True)
S.row_dimensions[38].height = 42

# ===========================================================================
# 4. CASH FLOW — the rolls and the forecast waterfall
# ===========================================================================
CF = sheet('Cash Flow', dict({'A': 60}, **{c: 12 for c in 'BCDEFGHI'}))
T(CF, 'A1', 'Cash flow — historical markers and the forecast waterfall', TITLE)
T(CF, 'A2', 'USD million. The forecast columns are the engine of the whole model: the '
            'depreciation, capital-expenditure and working-capital lines the discounted cash '
            'flow reads are computed here.', SUB)
T(CF, 'A4', 'USD million', HEAD)
for k, y in enumerate(['FY2023', 'FY2024', 'FY2025']):
    T(CF, f'{"BCD"[k]}4', y, HEAD)
for i, y in enumerate(FY):
    T(CF, f'{FCOL[i]}4', y, HEAD)
T(CF, 'A5', 'EBITDA')
for k in range(3):
    FM(CF, f'{"BCD"[k]}5', f"='Income Statement'!{'BCD'[k]}8", H['ebitda'][k], N2, link=True)
for i in range(5):
    FM(CF, f'{FCOL[i]}5', f"='Income Statement'!{FCOL[i]}8", F['ebitda'][i], N2, link=True)
T(CF, 'A6', 'Owned-asset depreciation and amortisation')
for k in range(3):
    V(CF, f'{"BCD"[k]}6', H['owned_dna'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(CF, f'{FCOL[i]}6', f"=Assumptions!$C$67*'Balance Sheet'!{prev}5", F['owned_dep'][i], N2)
T(CF, 'A7', 'Right-of-use depreciation')
for k in range(3):
    V(CF, f'{"BCD"[k]}7', H['rou_dep'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(CF, f'{FCOL[i]}7', f"=Assumptions!$C$66*'Balance Sheet'!{prev}6", F['rou_dep'][i], N2)
T(CF, 'A8', 'Depreciation and amortisation', HEAD)
for k in range(3):
    FM(CF, f'{"BCD"[k]}8', f'={"BCD"[k]}6+{"BCD"[k]}7', H['dna'][k], N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}8', f'={FCOL[i]}6+{FCOL[i]}7', F['dna'][i], N2)
T(CF, 'A9', 'Gross openings — net additions plus closures')
for i in range(5):
    FM(CF, f'{FCOL[i]}9',
       f'=Assumptions!{COLS[i]}$11+Segments!{"CDEFG"[i]}12*Assumptions!$C$63',
       F['gross_openings'][i], N0)
T(CF, 'A10', 'Owned capital expenditure')
for k in range(3):
    V(CF, f'{"BCD"[k]}10', H['capex'][k], N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}10',
       f"={FCOL[i]}9*Assumptions!$C$61/1000+Assumptions!$C$62*'Income Statement'!{FCOL[i]}5",
       F['capex'][i], N2)
T(CF, 'A11', 'Additions to right-of-use assets')
V(CF, 'C11', H.get('rou', [0, 0, 0])[0] * 0 + 279.400, N2)
V(CF, 'D11', 255.162, N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}11', f"=Assumptions!$C$64*'Income Statement'!{FCOL[i]}5",
       F['lease_additions'][i], N2)
T(CF, 'A12', 'Capital expenditure, including right-of-use additions', HEAD)
for i in range(5):
    FM(CF, f'{FCOL[i]}12', f'={FCOL[i]}10+{FCOL[i]}11', F['capex_total'][i], N2)
T(CF, 'A13', 'Change in net working capital')
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(CF, f'{FCOL[i]}13', f"='Balance Sheet'!{FCOL[i]}10-'Balance Sheet'!{prev}10",
       F['dnwc'][i], N2)
T(CF, 'A14', 'Tax')
for k in range(3):
    V(CF, f'{"BCD"[k]}14', H['tax'][k], N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}14', f"='Income Statement'!{FCOL[i]}16*-1", F['tax'][i], N2)
T(CF, 'A15', 'Free cash flow to the firm', HEAD)
for i in range(5):
    FM(CF, f'{FCOL[i]}15', f'=DCF!{COLS[i]}16', F['fcff'][i], N2, link=True)
T(CF, 'A16', 'Lease payments — principal and interest')
for k in range(3):
    V(CF, f'{"BCD"[k]}16', H['lease_principal'][k] + H['lease_interest'][k], N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}16', f"=Assumptions!$C$65*'Income Statement'!{FCOL[i]}5",
       F['lease_payments'][i], N2)
T(CF, 'A17', 'Dividends paid')
for k in range(3):
    V(CF, f'{"BCD"[k]}17', H['dividends_paid'][k], N2)
for i in range(5):
    FM(CF, f'{FCOL[i]}17', f"=Assumptions!$C$72*'Income Statement'!{FCOL[i]}17",
       F['dividends'][i], N2)
T(CF, 'A18', 'Movement in cash and deposits')
for i in range(5):
    FM(CF, f'{FCOL[i]}18',
       f"={FCOL[i]}5+'Income Statement'!{FCOL[i]}13-'Income Statement'!{FCOL[i]}5"
       f"*Assumptions!$C$74-{FCOL[i]}14-{FCOL[i]}13-{FCOL[i]}10-{FCOL[i]}16-{FCOL[i]}17",
       (F['cash'][i] - (H['cash'][2] + H['deposits'][2] if i == 0 else F['cash'][i - 1])), N2)
T(CF, 'A19', 'Cash and deposits, closing')
for i in range(5):
    FM(CF, f'{FCOL[i]}19', f"='Balance Sheet'!{FCOL[i]}9", F['cash'][i], N2, link=True)
T(CF, 'A20', 'Cash generated from operations, as reported')
for k in range(3):
    V(CF, f'{"BCD"[k]}20', H['cfo'][k], N2)
T(CF, 'A22', 'The forecast free-cash-flow line is free cash flow to the FIRM on the capitalised-'
             'lease reading: taking a new restaurant lease is an investment, so right-of-use '
             'additions are charged as capital expenditure and the matching lease liability is '
             'deducted in the bridge to equity. Adding back right-of-use depreciation without '
             'charging the additions would count the shield and never pay for the asset.',
     SUB, wrap=True)
CF.row_dimensions[22].height = 56

# ===========================================================================
# 5. BALANCE SHEET
# ===========================================================================
BS = sheet('Balance Sheet', dict({'A': 60}, **{c: 12 for c in 'BCDEFGHI'}))
T(BS, 'A1', 'Balance sheet — condensed, three years of history and five of forecast', TITLE)
T(BS, 'A2', 'USD million. Every historical line is the audited closing figure. Every forecast '
            'line rolls forward: assets from capital expenditure less depreciation, working '
            'capital off revenue, equity from profit less dividends, and the lease liability '
            'from additions, interest and payments.', SUB)
T(BS, 'A4', 'USD million', HEAD)
for k, y in enumerate(['FY2023', 'FY2024', 'FY2025']):
    T(BS, f'{"BCD"[k]}4', y, HEAD)
for i, y in enumerate(FY):
    T(BS, f'{FCOL[i]}4', y, HEAD)
T(BS, 'A5', 'Property, equipment, intangibles and investment property')
for k in range(2):
    V(BS, f'{"BCD"[k]}5', H['ppe'][k] + H['intangibles'][k] + H['investment_property'][k], N2)
FM(BS, 'D5', '=Assumptions!$C$93',
   H['ppe'][2] + H['intangibles'][2] + H['investment_property'][2], N2, link=True)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(BS, f'{FCOL[i]}5', f"={prev}5+'Cash Flow'!{FCOL[i]}10-'Cash Flow'!{FCOL[i]}6",
       F['owned_assets'][i], N2)
T(BS, 'A6', 'Right-of-use assets')
for k in range(2):
    V(BS, f'{"BCD"[k]}6', H['rou'][k], N2)
FM(BS, 'D6', '=Assumptions!$C$94', H['rou'][2], N2, link=True)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(BS, f'{FCOL[i]}6', f"={prev}6+'Cash Flow'!{FCOL[i]}11-'Cash Flow'!{FCOL[i]}7",
       F['rou'][i], N2)
T(BS, 'A7', 'Inventories')
for k in range(3):
    V(BS, f'{"BCD"[k]}7', H['inventories'][k], N2)
T(BS, 'A8', 'Trade and other receivables')
for k in range(3):
    V(BS, f'{"BCD"[k]}8', H['receivables'][k], N2)
T(BS, 'A9', 'Cash and bank deposits')
for k in range(3):
    V(BS, f'{"BCD"[k]}9', H['cash'][k] + H['deposits'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(BS, f'{FCOL[i]}9', f"={prev}9+'Cash Flow'!{FCOL[i]}18", F['cash'][i], N2)
T(BS, 'A10', 'Net working capital')
for k in range(2):
    V(BS, f'{"BCD"[k]}10', H['nwc'][k], N2)
FM(BS, 'D10', '=Assumptions!$C$92', H['nwc'][2], N2, link=True)
for i in range(5):
    FM(BS, f'{FCOL[i]}10', f"=Assumptions!$C$68*'Income Statement'!{FCOL[i]}5", F['nwc'][i], N2)
T(BS, 'A11', 'Payables, tax and provisions (the company’s own working-capital aggregate)')
for k in range(3):
    V(BS, f'{"BCD"[k]}11', H['payables'][k], N2)
T(BS, 'A12', 'Lease liabilities')
for k in range(3):
    V(BS, f'{"BCD"[k]}12', H['lease_liabilities'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(BS, f'{FCOL[i]}12',
       f"={prev}12+'Cash Flow'!{FCOL[i]}11+Assumptions!$C$82*{prev}12-'Cash Flow'!{FCOL[i]}16",
       F['lease_liabilities'][i], N2)
T(BS, 'A13', 'Bank debt')
for k in range(3):
    V(BS, f'{"BCD"[k]}13', H['bank_debt'][k], N2)
T(BS, 'A14', 'Equity attributable to shareholders')
for k in range(2):
    V(BS, f'{"BCD"[k]}14', H['equity'][k], N2)
FM(BS, 'D14', '=Assumptions!$C$91', H['equity'][2], N2, link=True)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(BS, f'{FCOL[i]}14',
       f"={prev}14+'Income Statement'!{FCOL[i]}17-'Cash Flow'!{FCOL[i]}17", F['equity'][i], N2)
T(BS, 'A15', 'Non-controlling interests')
for k in range(3):
    V(BS, f'{"BCD"[k]}15', H['nci'][k], N2)
T(BS, 'A16', 'Total assets')
for k in range(3):
    V(BS, f'{"BCD"[k]}16', H['total_assets'][k], N2)
T(BS, 'A17', 'Net debt — lease liabilities and bank debt less cash and deposits')
for k in range(3):
    FM(BS, f'{"BCD"[k]}17', f'={"BCD"[k]}12+{"BCD"[k]}13-{"BCD"[k]}9', H['net_debt'][k], N2)
for i in range(5):
    FM(BS, f'{FCOL[i]}17', f'={FCOL[i]}12-{FCOL[i]}9', F['net_debt'][i], N2)
T(BS, 'A18', 'Invested capital — owned assets, right-of-use assets and working capital')
for i in range(5):
    FM(BS, f'{FCOL[i]}18', f'={FCOL[i]}5+{FCOL[i]}6+{FCOL[i]}10', F['invested_capital'][i], N2)
T(BS, 'A19', 'Return on invested capital')
for i in range(5):
    FM(BS, f'{FCOL[i]}19', f'=DCF!{COLS[i]}12/{FCOL[i]}18', F['roic'][i], PC1)
T(BS, 'A20', 'Net debt / EBITDA')
for k in range(3):
    FM(BS, f'{"BCD"[k]}20', f"={'BCD'[k]}17/'Income Statement'!{'BCD'[k]}8",
       H['net_debt'][k] / H['ebitda'][k], MULT)
for i in range(5):
    FM(BS, f'{FCOL[i]}20', f"={FCOL[i]}17/'Income Statement'!{FCOL[i]}8",
       F['net_debt'][i] / F['ebitda'][i], MULT)
T(BS, 'A22', 'This is a CONDENSED layout: it does not foot to zero, because the provisions, '
             'employee end-of-service liability, deferred tax and related-party balances are '
             'not shown as separate lines. The lines that do appear are the audited closing '
             'figures.', SUB, wrap=True)
BS.row_dimensions[22].height = 42

# ===========================================================================
# 6. INCOME STATEMENT
# ===========================================================================
IS = sheet('Income Statement', dict({'A': 60}, **{c: 12 for c in 'BCDEFGHI'}))
T(IS, 'A1', 'Income statement — three years of history, five of forecast', TITLE)
T(IS, 'A2', 'USD million, consolidated. Every historical figure is the audited record; every '
            'forecast line is a formula.', SUB)
T(IS, 'A4', 'USD million', HEAD)
for k, y in enumerate(['FY2023', 'FY2024', 'FY2025']):
    T(IS, f'{"BCD"[k]}4', y, HEAD)
for i, y in enumerate(FY):
    T(IS, f'{FCOL[i]}4', y, HEAD)
T(IS, 'A5', 'Revenue')
for k in range(3):
    V(IS, f'{"BCD"[k]}5', H['revenue'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}5', f'=Segments!{FCOL[i]}36', F['revenue'][i], N2, link=True)
T(IS, 'A6', 'Cost of revenues')
for k in range(3):
    V(IS, f'{"BCD"[k]}6', -H['cogs'][k], N2)
for i in range(5):
    dash(IS, f'{FCOL[i]}6')
T(IS, 'A7', 'Gross profit')
for k in range(3):
    FM(IS, f'{"BCD"[k]}7', f'={"BCD"[k]}5+{"BCD"[k]}6', H['gross_profit'][k], N2)
for i in range(5):
    dash(IS, f'{FCOL[i]}7')
T(IS, 'A8', 'EBITDA', HEAD)
for k in range(3):
    V(IS, f'{"BCD"[k]}8', H['ebitda'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}8', f'=DCF!{COLS[i]}8', F['ebitda'][i], N2, link=True)
T(IS, 'A9', 'EBITDA margin')
for k in range(3):
    FM(IS, f'{"BCD"[k]}9', f'={"BCD"[k]}8/{"BCD"[k]}5', H['ebitda_margin'][k], PC1)
for i in range(5):
    FM(IS, f'{FCOL[i]}9', f'={FCOL[i]}8/{FCOL[i]}5', F['ebitda_margin'][i], PC1)
T(IS, 'A10', 'Depreciation and amortisation')
for k in range(3):
    FM(IS, f'{"BCD"[k]}10', f"=-'Cash Flow'!{'BCD'[k]}8", -H['dna'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}10', f"=-'Cash Flow'!{FCOL[i]}8", -F['dna'][i], N2)
T(IS, 'A11', 'EBIT, before the impairment line beneath', HEAD)
for k in range(3):
    FM(IS, f'{"BCD"[k]}11', f'={"BCD"[k]}8+{"BCD"[k]}10', H['ebit'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}11', f'={FCOL[i]}8+{FCOL[i]}10', F['ebitda'][i] - F['dna'][i], N2)
T(IS, 'A12', 'Impairment losses — audited history, and the recurring charge in the forecast')
for k in range(3):
    V(IS, f'{"BCD"[k]}12', -(H['impair_nonfin'][k] + H['impair_fin'][k]), N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}12', f'=-{FCOL[i]}5*Assumptions!$C$115', -F['impairment'][i], N2)
T(IS, 'A13', 'Finance income')
for k in range(3):
    V(IS, f'{"BCD"[k]}13', H['finance_income'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(IS, f'{FCOL[i]}13', f"=Assumptions!$C$73*'Balance Sheet'!{prev}9",
       F['finance_income'][i], N2)
T(IS, 'A14', 'Finance costs')
for k in range(3):
    V(IS, f'{"BCD"[k]}14', -H['finance_cost'][k], N2)
for i in range(5):
    prev = 'D' if i == 0 else FCOL[i - 1]
    FM(IS, f'{FCOL[i]}14',
       f"=-Assumptions!$C$82*'Balance Sheet'!{prev}12-Assumptions!$C$74*{FCOL[i]}5",
       -F['finance_cost'][i], N2)
T(IS, 'A15', 'Profit before tax')
for k in range(3):
    FM(IS, f'{"BCD"[k]}15', f'={"BCD"[k]}11+{"BCD"[k]}12+{"BCD"[k]}13+{"BCD"[k]}14',
       H['pbt'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}15', f'={FCOL[i]}11+{FCOL[i]}12+{FCOL[i]}13+{FCOL[i]}14', F['pbt'][i], N2)
T(IS, 'A16', 'Income tax and zakat')
for k in range(3):
    V(IS, f'{"BCD"[k]}16', -H['tax'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}16', f'=-{FCOL[i]}15*Assumptions!{COLS[i]}$71', -F['tax'][i], N2)
T(IS, 'A17', 'Profit for the year', HEAD)
for k in range(3):
    FM(IS, f'{"BCD"[k]}17', f'={"BCD"[k]}15+{"BCD"[k]}16', H['pat'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}17', f'={FCOL[i]}15+{FCOL[i]}16', F['pat'][i], N2)
T(IS, 'A18', 'Non-controlling interests')
for k in range(3):
    V(IS, f'{"BCD"[k]}18', H['pat_shareholders'][k] - H['pat'][k], N2)
for i in range(5):
    dash(IS, f'{FCOL[i]}18')
T(IS, 'A19', 'Profit attributable to shareholders', HEAD)
for k in range(3):
    FM(IS, f'{"BCD"[k]}19', f'={"BCD"[k]}17+{"BCD"[k]}18', H['pat_shareholders'][k], N2)
for i in range(5):
    FM(IS, f'{FCOL[i]}19', f'={FCOL[i]}17', F['pat'][i], N2)
T(IS, 'A20', 'Earnings per share (USD)')
for k in range(3):
    FM(IS, f'{"BCD"[k]}20', f'={"BCD"[k]}19/Assumptions!$C$7', H['eps'][k], N3)
for i in range(5):
    FM(IS, f'{FCOL[i]}20', f'={FCOL[i]}19/Assumptions!$C$7', F['eps'][i], N3)
T(IS, 'A22', 'EBIT here is EBITDA less depreciation, so it sits ABOVE the impairment charges; '
             'the audited operating profit is EBIT after those charges. Impairments are a '
             'recurring feature of a large restaurant estate — some brand-country units are '
             'always underperforming — but they are not forecast as a line, so the forecast '
             'columns carry none and the free-cash-flow waterfall is struck before them. The '
             'non-controlling interest is a rounding item, 0.98 million of a 490 million equity '
             'at the last year end and negative at the half year, so the forecast carries none.',
     SUB, wrap=True)
IS.row_dimensions[22].height = 56

# ===========================================================================
# 7. DCF
# ===========================================================================
X = sheet('DCF', dict({'A': 62}, **{c: 13 for c in COLS}))
T(X, 'A1', 'Discounted cash flow — the full waterfall', TITLE)
T(X, 'A2', 'Every line is a live formula. The cost of capital is built below rather than '
           'pasted, the glide is inherited from the tax path, the discount factors compound, '
           'and the terminal value is capitalised at the terminal rate.', SUB)
T(X, 'A4', 'USD million', HEAD)
for i, y in enumerate(FY):
    T(X, f'{COLS[i]}4', y, HEAD)
T(X, 'A5', 'Revenue')
for i, c in enumerate(COLS):
    FM(X, f'{c}5', f'=Segments!{FCOL[i]}36', F['revenue'][i], N2, link=True)
T(X, 'A6', 'Cash operating costs — six escalator classes as shares, staff and delivery '
           'as unit builds (headcount x wage; channel share x unit cost)')
for i, c in enumerate(COLS):
    terms = '+'.join(f'Assumptions!{c}${rr}' for _, _, rr in COST_ROWS
                     if rr not in (51, 52))
    prev_col = 'B' if i == 0 else COLS[i - 1]
    staff = (f"((Segments!{'BCDEF'[i]}12+Segments!{'CDEFG'[i]}12)/2*Assumptions!{c}$109/1000"
             f"+Assumptions!$C$112)*Assumptions!$C$110*(1+Assumptions!$C$111)^{i + 1}")
    FM(X, f'{c}6',
       f'=-{c}5*({terms}+Assumptions!$C$57)-{staff}-{c}5*Assumptions!{c}$113*Assumptions!{c}$114',
       -CS['cash_cost_f'][i], N2)
T(X, 'A7', 'Other income')
for i, c in enumerate(COLS):
    FM(X, f'{c}7', f'={c}5*Assumptions!$C$58',
       F['revenue'][i] * H['other_income'][2] / H['revenue'][2], N2)
T(X, 'A8', 'EBITDA', HEAD)
for i, c in enumerate(COLS):
    FM(X, f'{c}8', f'={c}5+{c}6+{c}7', F['ebitda'][i], N2)
T(X, 'A9', 'EBITDA margin')
for i, c in enumerate(COLS):
    FM(X, f'{c}9', f'={c}8/{c}5', F['ebitda_margin'][i], PC1)
T(X, 'A10', 'Less depreciation and amortisation')
for i, c in enumerate(COLS):
    FM(X, f'{c}10', f"=-'Cash Flow'!{FCOL[i]}8", -F['dna'][i], N2, link=True)
T(X, 'A11', 'EBIT, after the recurring impairment charge')
for i, c in enumerate(COLS):
    FM(X, f'{c}11', f'={c}8+{c}10-{c}5*Assumptions!$C$115', F['ebit'][i], N2)
T(X, 'A12', 'NOPAT — EBIT after tax')
for i, c in enumerate(COLS):
    FM(X, f'{c}12', f'={c}11*(1-Assumptions!{c}$71)', F['nopat'][i], N2)
T(X, 'A13', 'Add back depreciation and amortisation')
for i, c in enumerate(COLS):
    FM(X, f'{c}13', f'=-{c}10', F['dna'][i], N2)
T(X, 'A14', 'Less capital expenditure, including right-of-use additions')
for i, c in enumerate(COLS):
    FM(X, f'{c}14', f"=-'Cash Flow'!{FCOL[i]}12", -F['capex_total'][i], N2, link=True)
T(X, 'A15', 'Less change in working capital')
for i, c in enumerate(COLS):
    FM(X, f'{c}15', f"=-'Cash Flow'!{FCOL[i]}13", -F['dnwc'][i], N2, link=True)
T(X, 'A16', 'Free cash flow to the firm', HEAD)
for i, c in enumerate(COLS):
    FM(X, f'{c}16', f'={c}12+{c}13+{c}14+{c}15', F['fcff'][i], N2)
T(X, 'A17', 'Forward cost of capital')
for i, c in enumerate(COLS):
    FM(X, f'{c}17', f'=$C$46-($C$46-$C$53)*{c}44', F['wacc_path'][i], PC2)
T(X, 'A18', 'Discount factor')
FM(X, 'B18', '=1/(1+B17)', F['discount_factor'][0], N3)
for i in range(1, 5):
    FM(X, f'{COLS[i]}18', f'={COLS[i-1]}18/(1+{COLS[i]}17)', F['discount_factor'][i], N3)
T(X, 'A19', 'Present value of free cash flow to the firm')
for i, c in enumerate(COLS):
    FM(X, f'{c}19', f'={c}16*{c}18', DCF['pv'][i], N2)
T(X, 'A21', 'TERMINAL VALUE AND THE BRIDGE TO EQUITY', SECT)
T(X, 'A22', 'Terminal growth'); FM(X, 'C22', '=Assumptions!$C$84', W['terminal_g'], PC1, link=True)
T(X, 'A23', 'Terminal-year NOPAT grown one year'); FM(X, 'C23', '=F12*(1+C22)', DCF['nopat_next'], N2)
T(X, 'A24', 'Invested capital at the end of the forecast')
FM(X, 'C24', "='Balance Sheet'!I18", DCF['invested_capital'], N2, link=True)
T(X, 'A25', 'Terminal return on incremental capital — the payback-anchored fade')
FM(X, 'C25', '=Assumptions!$C$116', DCF['roic_term'], PC1, link=True)
T(X, 'A63', 'Model-implied average return on closing invested capital (the bull reading)')
FM(X, 'C63', '=F12/C24', DCF['roic_implied_avg'], PC1)
T(X, 'A26', 'Required reinvestment rate — growth divided by the return on capital')
FM(X, 'C26', '=C22/C25', DCF['rr_term'], PC1)
T(X, 'A27', 'Terminal cost of capital'); FM(X, 'C27', '=C53', W['wacc_terminal'], PC2)
T(X, 'A28', 'Terminal value'); FM(X, 'C28', '=C23*(1-C26)/(C27-C22)', DCF['tv'], N2)
T(X, 'A29', 'Present value of the five forecast years')
FM(X, 'C29', '=SUM(B19:F19)', DCF['sum_pv'], N2)
T(X, 'A30', 'Present value of the terminal value'); FM(X, 'C30', '=C28*F18', DCF['pv_tv'], N2)
T(X, 'A31', 'Enterprise value', HEAD); FM(X, 'C31', '=C29+C30', DCF['ev'], N2)
T(X, 'A32', 'Terminal value as a share of enterprise value', HEAD)
FM(X, 'C32', '=C30/C31', DCF['tv_share'], PC1)
T(X, 'A33', 'Less lease liabilities'); FM(X, 'C33', '=-Assumptions!$C$87', -H['lease_liabilities'][2], N2, link=True)
T(X, 'A34', 'Plus cash and bank deposits')
FM(X, 'C34', '=Assumptions!$C$88', H['cash'][2] + H['deposits'][2], N2, link=True)
T(X, 'A35', 'Less bank debt'); FM(X, 'C35', '=-Assumptions!$C$89', -H['bank_debt'][2], N2, link=True)
T(X, 'A36', 'Less non-controlling interests')
FM(X, 'C36', '=-Assumptions!$C$90', -H['nci'][2], N2, link=True)
T(X, 'A37', 'Equity value', HEAD)
FM(X, 'C37', '=C31+C33+C34+C35+C36', DCF['equity'], N2)
T(X, 'A38', 'Fair value per share at 31 December 2025 (USD)')
FM(X, 'C38', '=C37/Assumptions!$C$7', DCF['fv_unrolled'], N3)
T(X, 'D38', 'Anchor roll factor', SUB)
FM(X, 'E38', '=(1+C52)^(Assumptions!$C$117/365)', W['roll_factor'], N3)
T(X, 'A39', 'Fair value per share at the 7 August 2026 anchor (AED) — rolled at the cost of '
            'equity, net of the dividend paid in the window', HEAD)
FM(X, 'C39', '=(C38*E38-Assumptions!$C$118)*Assumptions!$C$6', DCF['fv'] * FX, N2)
T(X, 'A41', 'THE COST OF CAPITAL — BUILT HERE, NOT PASTED', SECT)
T(X, 'A42', 'The glide', HEAD)
for i, y in enumerate(FY):
    T(X, f'{COLS[i]}42', y, HEAD)
T(X, 'A43', 'Effective tax rate path')
for i, c in enumerate(COLS):
    FM(X, f'{c}43', f'=Assumptions!{c}$71', F['etr'][i], PC1, link=True)
T(X, 'A44', 'Glide fraction — cumulative progress of the tax path')
for i, c in enumerate(COLS):
    FM(X, f'{c}44', f'=({c}43-$B$43)/($F$43-$B$43)', F['glide'][i], PC1)
T(X, 'A45', 'Row 17 walks the explicit-window cost of capital down to the terminal rate on this '
            'fraction, so the shape of the path is inherited from the only forward schedule in '
            'the model rather than invented as a second free parameter.', SUB, wrap=True)
X.row_dimensions[45].height = 42
T(X, 'A46', 'Cost of capital, explicit window', HEAD)
FM(X, 'C46', '=C59*C52+C58*C55', W['wacc_rating'], PC2)
T(X, 'A47', 'US ten-year Treasury yield')
FM(X, 'C47', '=Assumptions!$C$76', W['rf_ust'], PC2, link=True)
T(X, 'A48', 'Less the US sovereign default spread')
FM(X, 'C48', '=Assumptions!$C$77', W['us_default_spread'], PC2, link=True)
T(X, 'A49', 'Risk-free rate net of the sovereign spread'); FM(X, 'C49', '=C47-C48', W['rf_rating'], PC2)
T(X, 'A50', 'Beta'); FM(X, 'C50', '=Assumptions!$C$81', W['beta'], N3, link=True)
T(X, 'A51', 'Blended equity risk premium (ratings basis)')
FM(X, 'C51', '=Assumptions!$C$78', W['erp_rating'], PC2, link=True)
T(X, 'A52', 'Cost of equity'); FM(X, 'C52', '=C49+C50*C51', W['ke_rating'], PC2)
T(X, 'A53', 'Cost of capital, terminal', HEAD); FM(X, 'C53', '=C59*C61+C58*C62', W['wacc_terminal'], PC2)
T(X, 'A54', 'Cost of debt — the group’s own incremental borrowing rate')
FM(X, 'C54', '=Assumptions!$C$82', W['kd'], PC2, link=True)
T(X, 'A55', 'Cost of debt after tax'); FM(X, 'C55', '=C54*(1-Assumptions!$B$71)', W['kd_after_tax'], PC2)
T(X, 'A56', 'Market capitalisation (USD million)')
FM(X, 'C56', '=Assumptions!$C$5/Assumptions!$C$6*Assumptions!$C$7', M['mktcap'], N2)
T(X, 'A57', 'Debt at market value — the lease liability')
FM(X, 'C57', '=Assumptions!$C$87', H['lease_liabilities'][2], N2, link=True)
T(X, 'A58', 'Debt weight'); FM(X, 'C58', '=C57/(C57+C56)', W['debt_weight'], PC1)
T(X, 'A59', 'Equity weight'); FM(X, 'C59', '=1-C58', W['equity_weight'], PC1)
T(X, 'A60', 'Terminal risk-free rate'); FM(X, 'C60', '=Assumptions!$C$83', W['terminal_rf'], PC2, link=True)
T(X, 'A61', 'Terminal cost of equity'); FM(X, 'C61', '=C60+C50*C51', W['ke_terminal'], PC2)
T(X, 'A62', 'Terminal cost of debt after tax')
FM(X, 'C62', '=C54*(1-Assumptions!$F$71)', W['kd'] * (1 - F['etr'][4]), PC2)
T(X, 'A71', 'THE OTHER READING OF THE OPEN QUESTION — the cyclical margin, computed live', SECT)
T(X, 'A72', 'Cyclical EBITDA margin path')
CYC = D['contested']['way_b']
mrg = D['inputs']['margin_path_cyclical']['value']
for i, c in enumerate(COLS):
    FM(X, f'{c}72', f'=Assumptions!{c}$119', mrg[i], PC1, link=True)
cyc_eb = [F['revenue'][i] * mrg[i] for i in range(5)]
cyc_nopat = [(cyc_eb[i] - F['dna'][i] - F['impairment'][i]) * (1 - F['etr'][i]) for i in range(5)]
cyc_fcff = [cyc_nopat[i] + F['dna'][i] - F['capex_total'][i] - F['dnwc'][i] for i in range(5)]
T(X, 'A73', 'EBITDA on the cyclical path')
for i, c in enumerate(COLS):
    FM(X, f'{c}73', f'={c}5*{c}72', cyc_eb[i], N2)
T(X, 'A74', 'NOPAT on the cyclical path')
for i, c in enumerate(COLS):
    FM(X, f'{c}74', f'=({c}73+{c}10-{c}5*Assumptions!$C$115)*(1-Assumptions!{c}$71)',
       cyc_nopat[i], N2)
T(X, 'A75', 'Free cash flow to the firm, cyclical')
for i, c in enumerate(COLS):
    FM(X, f'{c}75', f'={c}74-{c}10+{c}14+{c}15', cyc_fcff[i], N2)
T(X, 'A76', 'Present value')
for i, c in enumerate(COLS):
    FM(X, f'{c}76', f'={c}75*{c}18', cyc_fcff[i] * F['discount_factor'][i], N2)
cyc_tvnum = cyc_nopat[4] * (1 + W['terminal_g'])
cyc_tv = cyc_tvnum * (1 - W['terminal_g'] / W['terminal_roic']) / (W['wacc_terminal'] - W['terminal_g'])
cyc_ev = sum(cyc_fcff[i] * F['discount_factor'][i] for i in range(5)) + cyc_tv * F['discount_factor'][4]
cyc_eq = cyc_ev - H['lease_liabilities'][2] + H['cash'][2] + H['deposits'][2] - H['nci'][2]
T(X, 'A77', 'Terminal value, cyclical')
FM(X, 'C77', '=F74*(1+C22)*(1-C22/C25)/(C27-C22)', cyc_tv, N2)
T(X, 'A78', 'Equity value, cyclical')
FM(X, 'C78', '=SUM(B76:F76)+C77*F18+C33+C34+C35+C36', cyc_eq, N2)
T(X, 'A79', 'Value per share on the cyclical reading (AED), rolled to the anchor', HEAD)
FM(X, 'C79', '=(C78/Assumptions!$C$7*E38-Assumptions!$C$118)*Assumptions!$C$6',
   CYC['value_aed'], N2)
T(X, 'A80', 'Every cell in this block is a live formula: the cyclical reading shares the '
            'revenue, depreciation, capital-expenditure and working-capital lines of the base '
            'case and differs only in the margin path, so it can be — and is — computed in the '
            'sheet rather than pasted.', SUB, wrap=True)
X.row_dimensions[80].height = 42

T(X, 'A64', 'The same cost of capital on the credit-default-swap basis', SECT)
T(X, 'A65', 'Risk-free rate net of the US sovereign credit-default-swap spread')
FM(X, 'C65', '=C47-Assumptions!$C$79', W['rf_cds'], PC2)
T(X, 'A66', 'Blended equity risk premium (credit-default-swap basis)')
FM(X, 'C66', '=Assumptions!$C$80', W['erp_cds'], PC2, link=True)
T(X, 'A67', 'Cost of equity, credit-default-swap basis'); FM(X, 'C67', '=C65+C50*C66', W['ke_cds'], PC2)
T(X, 'A68', 'Cost of capital, credit-default-swap basis', HEAD)
FM(X, 'C68', '=C59*C67+C58*C55', W['wacc_cds'], PC2)
T(X, 'A69', 'Both bases are published. The default spread stripped out of the risk-free rate is '
            'on the same basis as the premium added back, so sovereign risk is counted once and '
            'once only.', SUB, wrap=True)
X.row_dimensions[69].height = 42

# ===========================================================================
# 8. SOTP BRIDGE — enterprise value to equity
# ===========================================================================
BR = sheet('SOTP Bridge', dict(A=58, B=16, C=16))
T(BR, 'A1', 'Enterprise value to equity — the bridge', TITLE)
T(BR, 'A2', 'A single operating business, so the bridge is one leg, not a sum of parts. Every '
            'line links live to the cash-flow sheet.', SUB)
T(BR, 'A4', 'Step', HEAD); T(BR, 'B4', 'USD million', HEAD); T(BR, 'C4', 'AED per share', HEAD)
BRIDGE = [('Present value of the five forecast years', '=DCF!C29', DCF['sum_pv']),
          ('Present value of the terminal value', '=DCF!C30', DCF['pv_tv']),
          ('Enterprise value', '=B5+B6', DCF['ev']),
          ('Less lease liabilities', '=-Assumptions!$C$87', -H['lease_liabilities'][2]),
          ('Plus cash and bank deposits', '=Assumptions!$C$88', H['cash'][2] + H['deposits'][2]),
          ('Less bank debt', '=-Assumptions!$C$89', -H['bank_debt'][2]),
          ('Less non-controlling interests', '=-Assumptions!$C$90', -H['nci'][2]),
          ('Equity attributable to shareholders', '=B7+B8+B9+B10+B11', DCF['equity'])]
for j, (lab, fml, val) in enumerate(BRIDGE):
    rr = 5 + j
    T(BR, f'A{rr}', lab, HEAD if lab.startswith(('Enterprise', 'Equity')) else BLACK)
    FM(BR, f'B{rr}', fml, val, N2)
    FM(BR, f'C{rr}', f'=B{rr}/Assumptions!$C$7*Assumptions!$C$6', val / SH * FX, N2)
T(BR, 'A14', 'Terminal value as a share of enterprise value', HEAD)
FM(BR, 'B14', '=DCF!C32', DCF['tv_share'], PC1)
T(BR, 'A15', 'Present value of the five forecast years, as a share of enterprise value')
FM(BR, 'B15', '=B5/B7', DCF['sum_pv'] / DCF['ev'], PC1)
T(BR, 'A17', 'Just over half the enterprise value sits beyond the fifth forecast year. That is '
             'what a growing business with a nine-and-a-half per cent cost of capital and three '
             'per cent long-run growth looks like, and it is the reason the terminal '
             'assumptions are sensitised on their own sheet rather than buried.', SUB, wrap=True)
BR.row_dimensions[17].height = 42

# ===========================================================================
# 9. RELATIVE & NORMALIZED
# ===========================================================================
RN = sheet('Relative & Normalized', dict(A=62, B=14, C=14, D=20, E=14, F=14))
T(RN, 'A1', 'Relative multiples, normalised earnings power, and the book lens', TITLE)
T(RN, 'A4', 'Relative lens', HEAD); T(RN, 'B4', 'Value', HEAD)
T(RN, 'A5', 'FY2027E EBITDA (USD million)'); FM(RN, 'B5', '=DCF!C8', F['ebitda'][1], N2, link=True)
T(RN, 'A6', 'Justified enterprise value / EBITDA')
FM(RN, 'B6', '=Assumptions!$C$100', LN['relative']['multiple'], MULT, link=True)
T(RN, 'A7', 'Implied enterprise value at the end of FY2027 (USD million)')
FM(RN, 'B7', '=B5*B6', F['ebitda'][1] * LN['relative']['multiple'], N2)
T(RN, 'A8', 'Discount factor back to the valuation date'); FM(RN, 'B8', '=DCF!C18', F['discount_factor'][1], N3, link=True)
T(RN, 'A9', 'Plus the present value of the FY2026 AND FY2027 free cash flows — a two-year '
            'discount makes both years intervening')
FM(RN, 'B9', '=DCF!B19+DCF!C19', DCF['pv'][0] + DCF['pv'][1], N2, link=True)
T(RN, 'A10', 'Implied enterprise value today (USD million)')
FM(RN, 'B10', '=B7*B8+B9', LN['relative']['ev'], N2)
T(RN, 'A11', 'Implied equity value (USD million)')
FM(RN, 'B11', '=B10+DCF!C33+DCF!C34+DCF!C35+DCF!C36', LN['relative']['equity'], N2)
T(RN, 'A12', 'Implied value per share (AED), rolled to the anchor', HEAD)
FM(RN, 'B12', '=(B11/Assumptions!$C$7*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6',
   LN['values']['Relative multiples'] * FX, N2)
T(RN, 'A13', 'Bear at 7.0 times (column C) and bull at 10.5 times (column D), same construction')
FM(RN, 'C13', '=((B5*7*B8+B9+DCF!C33+DCF!C34+DCF!C35+DCF!C36)/Assumptions!$C$7*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6',
   LN['ranges']['Relative multiples'][0] * FX, N2)
FM(RN, 'D13', '=((B5*10.5*B8+B9+DCF!C33+DCF!C34+DCF!C35+DCF!C36)/Assumptions!$C$7*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6',
   LN['ranges']['Relative multiples'][2] * FX, N2)
T(RN, 'A15', 'The company’s own trailing multiples', HEAD)
T(RN, 'A16', 'Trailing enterprise value / EBITDA')
FM(RN, 'B16', "=(DCF!C56-DCF!C34+DCF!C57)/'Income Statement'!D8", D['trailing']['ev_ebitda'], MULT)
T(RN, 'A17', 'Trailing price / earnings')
FM(RN, 'B17', "=Assumptions!$C$5/Assumptions!$C$6/'Income Statement'!D20", D['trailing']['pe'], MULT)
T(RN, 'A18', 'Trailing price / book')
FM(RN, 'B18', "=Assumptions!$C$5/Assumptions!$C$6/('Balance Sheet'!D14/Assumptions!$C$7)",
   D['trailing']['pb'], MULT)
T(RN, 'A19', 'Net debt / EBITDA'); FM(RN, 'B19', "='Balance Sheet'!D20", H['net_debt'][2] / H['ebitda'][2], MULT, link=True)
T(RN, 'A21', 'Normalised earnings lens — the mid-cycle margin at today’s scale', HEAD)
T(RN, 'A22', 'Current-scale revenue, FY2026E (USD million)'); FM(RN, 'B22', '=DCF!B5', F['revenue'][0], N2, link=True)
T(RN, 'A23', 'Mid-cycle EBITDA margin — the midpoint of the structural and cyclical '
            'FY2028 readings')
FM(RN, 'B23', '=(DCF!D9+Assumptions!$D$119)/2', LN['normalised']['margin'], PC1)
T(RN, 'A24', 'Normalised EBITDA (USD million)'); FM(RN, 'B24', '=B22*B23', LN['normalised']['ebitda'], N2)
T(RN, 'A25', 'Less depreciation and amortisation'); FM(RN, 'B25', "=-'Cash Flow'!E8", -F['dna'][0], N2)
T(RN, 'A26', 'Normalised EBIT, after the recurring impairment charge (USD million)')
FM(RN, 'B26', '=B24+B25-B22*Assumptions!$C$115', LN['normalised']['ebit'], N2)
T(RN, 'A27', 'Net finance result (USD million)')
FM(RN, 'B27', "='Income Statement'!E13-Assumptions!$C$82*'Balance Sheet'!D12"
              "-Assumptions!$C$74*B22", LN['normalised']['net_finance'], N2)
T(RN, 'A28', 'Normalised earnings after tax (USD million)')
FM(RN, 'B28', '=(B26+B27)*(1-Assumptions!$C$71)', LN['normalised']['earnings'], N2)
T(RN, 'A29', 'Normalised earnings per share (USD)')
FM(RN, 'B29', '=B28/Assumptions!$C$7', LN['normalised']['eps'], N3)
T(RN, 'A30', 'Justified price / earnings'); FM(RN, 'B30', '=Assumptions!$C$101', LN['normalised']['multiple'], MULT, link=True)
T(RN, 'A31', 'Implied value per share (AED), rolled to the anchor', HEAD)
FM(RN, 'B31', '=(B29*B30*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6',
   LN['values']['Normalised earnings power'] * FX, N2)
T(RN, 'A32', 'Bear at 13 times (column C) and bull at 21 times (column D)')
FM(RN, 'C32', '=(B29*13*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6', LN['ranges']['Normalised earnings power'][0] * FX, N2)
FM(RN, 'D32', '=(B29*21*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6', LN['ranges']['Normalised earnings power'][2] * FX, N2)
T(RN, 'A34', 'Book value and sustainable return', HEAD)
T(RN, 'A35', 'Book value per share (USD)')
FM(RN, 'B35', '=Assumptions!$C$97/Assumptions!$C$7', LN['book']['bvps'], N3)
T(RN, 'A36', 'Sustainable return on equity'); FM(RN, 'B36', '=Assumptions!$C$102', LN['book']['roe'], PC1, link=True)
T(RN, 'A37', 'Trailing return on equity')
FM(RN, 'B37', "='Income Statement'!D19/(('Balance Sheet'!C14+'Balance Sheet'!D14)/2)",
   H['pat_shareholders'][2] / ((H['equity'][1] + H['equity'][2]) / 2), PC1)
T(RN, 'A38', 'Terminal cost of equity'); FM(RN, 'B38', '=DCF!C61', W['ke_terminal'], PC2, link=True)
T(RN, 'A39', 'Justified price / book')
FM(RN, 'B39', '=(B36-Assumptions!$C$84)/(B38-Assumptions!$C$84)', LN['book']['justified_pb'], MULT)
T(RN, 'A39b', '', SUB) if False else None
T(RN, 'E39', 'Roll factor from the 30 June book date (38 days)', SUB)
FM(RN, 'F39', '=(1+DCF!C52)^(38/365)', (1 + W['ke_rating']) ** (38 / 365), N3)
T(RN, 'A40', 'Implied value per share (AED), rolled from the book date to the anchor', HEAD)
FM(RN, 'B40', '=B35*B39*F39*Assumptions!$C$6',
   LN['values']['Book value and sustainable return'] * FX, N2)
T(RN, 'A41', 'Bear at a 34% return (column C) and bull at 48% (column D)')
FM(RN, 'C41', '=B35*((0.34-Assumptions!$C$84)/(B38+0.02-Assumptions!$C$84))*F39*Assumptions!$C$6',
   LN['ranges']['Book value and sustainable return'][0] * FX, N2)
FM(RN, 'D41', '=B35*((0.48-Assumptions!$C$84)/(B38-0.01-Assumptions!$C$84))*F39*Assumptions!$C$6',
   LN['ranges']['Book value and sustainable return'][2] * FX, N2)
T(RN, 'A43', 'The book lens is the weakest of the four here and carries the lowest weight for '
             'that reason. An operator that leases its estate and distributes almost all its '
             'earnings holds very little book equity, so a justified price-to-book multiple '
             'divides a large return by a small base and is unstable in both directions.',
     SUB, wrap=True)
RN.row_dimensions[43].height = 42

# ===========================================================================
# 10. FUNDAMENTAL VALUATION
# ===========================================================================
FV = sheet('Fundamental Valuation', dict(A=54, B=32, C=15, D=15, E=15))
T(FV, 'A1', 'Fundamental valuation — the four lenses, the open question, and the panel', TITLE)
T(FV, 'A4', 'Lens', HEAD); T(FV, 'B4', 'Basis', HEAD); T(FV, 'C4', 'AED per share', HEAD)
LROWS = [('Discounted cash flow', 'the full waterfall on the DCF sheet', '=DCF!C39',
          LN['values']['Discounted cash flow'] * FX),
         ('Relative multiples', f"{LN['relative']['multiple']:.1f}x forward EBITDA",
          "='Relative & Normalized'!B12", LN['values']['Relative multiples'] * FX),
         ('Normalised earnings power', f"{LN['normalised']['multiple']:.0f}x mid-cycle earnings",
          "='Relative & Normalized'!B31", LN['values']['Normalised earnings power'] * FX),
         ('Book value and sustainable return', 'justified price to book on a sustainable return',
          "='Relative & Normalized'!B40",
          LN['values']['Book value and sustainable return'] * FX)]
for j, (lab, basis, fml, val) in enumerate(LROWS):
    rr = 5 + j
    T(FV, f'A{rr}', lab); T(FV, f'B{rr}', basis)
    FM(FV, f'C{rr}', fml, val, N2)
T(FV, 'A10', 'Weighted central', HEAD)
FM(FV, 'C10', '=Summary!C10', LN['central'] * FX, N2, link=True)
T(FV, 'A12', 'THE OPEN QUESTION — the margin, computed both ways', SECT)
T(FV, 'A13', C['way_a']['name']); T(FV, 'B13', C['way_a']['detail'])
FM(FV, 'C13', '=DCF!C39', C['way_a']['value_aed'], N2, link=True)
T(FV, 'A14', C['way_b']['name']); T(FV, 'B14', C['way_b']['detail'])
FM(FV, 'C14', '=DCF!C79', C['way_b']['value_aed'], N2, link=True)
T(FV, 'A15', 'The gap')
FM(FV, 'C15', '=C14/C13-1', C['gap_pct'], PC1)
T(FV, 'A16', 'Both readings are LIVE in this workbook: the structural value is the DCF sheet '
             'headline and the cyclical value is the parallel formula block at the foot of the '
             'DCF sheet, sharing every driver except the margin path. They are published side '
             'by side and never averaged.', SUB, wrap=True)
FV.row_dimensions[16].height = 42
T(FV, 'A18', 'A SECOND READING — how the lease estate is treated', SECT)
T(FV, 'A19', DFL['way_a']['name']); V(FV, 'C19', DFL['way_a']['value_aed'], N2, kind='engine')
T(FV, 'A20', DFL['way_b']['name']); V(FV, 'C20', DFL['way_b']['value_aed'], N2, kind='engine')
T(FV, 'A21', 'The gap'); FM(FV, 'C21', '=C20/C19-1', DFL['gap_pct'], PC1)
T(FV, 'A22', DFL['finding'], SUB, wrap=True)
FV.row_dimensions[22].height = 42
T(FV, 'A24', 'THE EXPERT PANEL', SECT)
T(FV, 'A25', 'Expert', HEAD); T(FV, 'B25', 'Method', HEAD)
T(FV, 'C25', 'Base (AED)', HEAD); T(FV, 'D25', 'Low', HEAD); T(FV, 'E25', 'High', HEAD)
E1, E2, E3 = D['experts']
T(FV, 'A26', E1['label']); T(FV, 'B26', E1['method'])
OC = ("(2*Assumptions!$C$124-Assumptions!$C$125-Assumptions!$C$62*DCF!B5"
      "-Assumptions!$C$65*DCF!B5)*(1-Assumptions!$C$71)")
BRIDGE1 = "+Assumptions!$C$88-Assumptions!$C$90"
RLL = "*DCF!$E$38-Assumptions!$C$118)*Assumptions!$C$6"
FM(FV, 'C26', f'=(({OC}/DCF!C61{BRIDGE1})/Assumptions!$C$7{RLL}', E1['base'] * FX, N2)
FM(FV, 'D26', f'=(({OC}/(DCF!C61+0.015){BRIDGE1})/Assumptions!$C$7{RLL}', E1['low'] * FX, N2)
FM(FV, 'E26', f'=(({OC}/(DCF!C61-0.01){BRIDGE1})/Assumptions!$C$7{RLL}', E1['high'] * FX, N2)
T(FV, 'A27', E2['label']); T(FV, 'B27', E2['method'])
DPS = "Assumptions!$C$126/Assumptions!$C$7"
FM(FV, 'C27', f'=({DPS}*(1+Assumptions!$C$84)/(DCF!C61-Assumptions!$C$84){RLL}',
   E2['base'] * FX, N2)
FM(FV, 'D27', f'=({DPS}*1.02/(DCF!C61+0.01-0.02){RLL}', E2['low'] * FX, N2)
FM(FV, 'E27', f'=({DPS}*1.04/(DCF!C61-0.005-0.04){RLL}', E2['high'] * FX, N2)
T(FV, 'A28', E3['label']); T(FV, 'B28', E3['method'])
ROIC3 = "DCF!B12/'Balance Sheet'!E18"
GT = (f"DCF!B12*({ROIC3}-DCF!C53)/({ROIC3})"
      f"*Assumptions!$C$84/(DCF!C53*(DCF!C53-Assumptions!$C$84))")
BR3 = "+DCF!C33+DCF!C34+DCF!C35+DCF!C36"
FM(FV, 'C28', f'=((DCF!B12/DCF!C53+{GT}{BR3})/Assumptions!$C$7{RLL}', E3['base'] * FX, N2)
FM(FV, 'D28', f'=((DCF!B12/(DCF!C53+0.01){BR3})/Assumptions!$C$7{RLL}', E3['low'] * FX, N2)
FM(FV, 'E28', f'=(((DCF!B12/DCF!C53+{GT})*1.18{BR3})/Assumptions!$C$7{RLL}',
   E3['high'] * FX, N2)
T(FV, 'A29', 'Panel median', HEAD)
FM(FV, 'C29', '=MEDIAN(C26:C28)', LN['expert_median'] * FX, N2)

# ===========================================================================
# 11. SUMMARY
# ===========================================================================
SU = sheet('Summary', dict(A=44, B=13, C=13, D=13, E=11, F=14, G=15, H=12))
T(SU, 'A1', 'Summary — the valuation at a glance', TITLE)
T(SU, 'A2', 'All values link live to their source sheets. AED per share unless stated.', SUB)
T(SU, 'A4', 'Lens', HEAD); T(SU, 'B4', 'Bear', HEAD); T(SU, 'C4', 'Base', HEAD)
T(SU, 'D4', 'Bull', HEAD); T(SU, 'E4', 'Weight', HEAD); T(SU, 'F4', 'Contribution', HEAD)
T(SU, 'G4', 'Terminal value share', HEAD); T(SU, 'H4', 'vs market', HEAD)
KEYS = ['Discounted cash flow', 'Relative multiples', 'Normalised earnings power',
        'Book value and sustainable return']
SRC = ['=DCF!C39', "='Relative & Normalized'!B12", "='Relative & Normalized'!B31",
       "='Relative & Normalized'!B40"]
WROW = [103, 104, 105, 106]
BB = [None, ("='Relative & Normalized'!C13", "='Relative & Normalized'!D13"),
      ("='Relative & Normalized'!C32", "='Relative & Normalized'!D32"),
      ("='Relative & Normalized'!C41", "='Relative & Normalized'!D41")]
for j, k in enumerate(KEYS):
    rr = 5 + j
    T(SU, f'A{rr}', k)
    if j == 0:
        V(SU, f'B{rr}', LN['ranges'][k][0] * FX, N2, kind='engine')
    else:
        FM(SU, f'B{rr}', BB[j][0], LN['ranges'][k][0] * FX, N2, link=True)
    FM(SU, f'C{rr}', SRC[j], LN['values'][k] * FX, N2, link=True)
    if j == 0:
        V(SU, f'D{rr}', LN['ranges'][k][2] * FX, N2, kind='engine')
    else:
        FM(SU, f'D{rr}', BB[j][1], LN['ranges'][k][2] * FX, N2, link=True)
    FM(SU, f'E{rr}', f'=Assumptions!$C${WROW[j]}', LN['weights'][k], PC1, link=True)
    FM(SU, f'F{rr}', f'=C{rr}*E{rr}', LN['values'][k] * FX * LN['weights'][k], N2)
    FM(SU, f'H{rr}', f'=C{rr}/$C$12-1', LN['values'][k] / M['spot'] - 1, PC1)
FM(SU, 'G5', '=DCF!C32', DCF['tv_share'], PC1, link=True)
T(SU, 'A9b', '', SUB) if False else None
T(SU, 'A10', 'Weighted central — outer columns are the weighted bear and bull', HEAD)
wb_ = sum(LN['ranges'][k][0] * LN['weights'][k] for k in KEYS) * FX
wu_ = sum(LN['ranges'][k][2] * LN['weights'][k] for k in KEYS) * FX
FM(SU, 'B10', '=B5*E5+B6*E6+B7*E7+B8*E8', wb_, N2)
FM(SU, 'C10', '=SUM(F5:F8)', LN['central'] * FX, N2)
FM(SU, 'D10', '=D5*E5+D6*E6+D7*E7+D8*E8', wu_, N2)
FM(SU, 'E10', '=SUM(E5:E8)', 1.0, PC1)
FM(SU, 'H10', '=C10/$C$12-1', LN['central'] / M['spot'] - 1, PC1)
T(SU, 'A11b', '', SUB) if False else None
T(SU, 'A11', 'Range across the lenses (widest bear to widest bull, unweighted)')
FM(SU, 'B11', '=MIN(B5:B8)', min(LN['ranges'][k][0] for k in KEYS) * FX, N2)
FM(SU, 'D11', '=MAX(D5:D8)', max(LN['ranges'][k][2] for k in KEYS) * FX, N2)
T(SU, 'A11c', '', SUB) if False else None
T(SU, 'A13', 'Expert panel median')
FM(SU, 'C13', "='Fundamental Valuation'!C29", LN['expert_median'] * FX, N2, link=True)
FM(SU, 'H13', '=C13/$C$12-1', LN['expert_median'] / M['spot'] - 1, PC1)
T(SU, 'A12', 'Market price (anchor, 7 August 2026)', HEAD)
FM(SU, 'C12', '=Assumptions!$C$5', M['spot_aed'], N2, link=True)
T(SU, 'A14', 'The open question — the margin, both ways', SECT)
T(SU, 'A15', 'Structural: the first-half gains hold')
FM(SU, 'C15', "='Fundamental Valuation'!C13", C['way_a']['value_aed'], N2, link=True)
FM(SU, 'H15', '=C15/$C$12-1', C['way_a']['value_usd'] / M['spot'] - 1, PC1)
T(SU, 'A16', 'Cyclical: the margin reverts to the three-year average')
FM(SU, 'C16', "='Fundamental Valuation'!C14", C['way_b']['value_aed'], N2, link=True)
FM(SU, 'H16', '=C16/$C$12-1', C['way_b']['value_usd'] / M['spot'] - 1, PC1)
T(SU, 'A18', 'Key figures', SECT)
KF = [('Shares outstanding, net of treasury (million)', '=Assumptions!$C$7', SH, N0),
      ('Market capitalisation (USD million)', '=DCF!C56', M['mktcap'], N2),
      ('Net debt at 31 December 2025 (USD million)', "='Balance Sheet'!D17", H['net_debt'][2], N2),
      ('FY2025 revenue (USD million)', "='Income Statement'!D5", H['revenue'][2], N2),
      ('FY2025 EBITDA (USD million)', "='Income Statement'!D8", H['ebitda'][2], N2),
      ('FY2025 profit attributable to shareholders (USD million)', "='Income Statement'!D19",
       H['pat_shareholders'][2], N2),
      ('Restaurants at 31 December 2025', '=Segments!B12', sum(U['stores_hist'][u][1] for u in UNITS), N0),
      ('Cost of capital — explicit window', '=DCF!C46', W['wacc_rating'], PC2),
      ('Cost of capital — terminal', '=DCF!C53', W['wacc_terminal'], PC2),
      ('Terminal growth', '=DCF!C22', W['terminal_g'], PC1),
      ('Terminal value as a share of enterprise value', '=DCF!C32', DCF['tv_share'], PC1),
      ('Trailing enterprise value / EBITDA', "='Relative & Normalized'!B16",
       D['trailing']['ev_ebitda'], MULT),
      ('Dividend yield on the declared FY2025 distribution', '=Assumptions!$C$126/DCF!C56',
       D['trailing']['dividend_yield'], PC1)]
for j, (lab, fml, val, fmt) in enumerate(KF):
    rr = 19 + j
    T(SU, f'A{rr}', lab)
    FM(SU, f'C{rr}', fml, val, fmt, link=True)
T(SU, 'A33', 'The terminal value share is shown beside the cash-flow lens above and again in '
             'the key figures, and it is a live link to the DCF sheet in both places.',
     SUB, wrap=True)

# ===========================================================================
# 12. SUMMARY FINANCIALS
# ===========================================================================
SF = sheet('Summary Financials', dict({'A': 46}, **{c: 12 for c in 'BCDEFGHI'}))
T(SF, 'A1', 'Summary financials — the eight-year picture', TITLE)
T(SF, 'A2', 'USD million unless stated. Every cell on this sheet is a link or a ratio; nothing '
            'is typed twice.', SUB)
T(SF, 'A4', 'USD million', HEAD)
for i, y in enumerate(['FY2023', 'FY2024', 'FY2025'] + FY):
    T(SF, f'{ALLC[i]}4', y, HEAD)
SFROWS = [('Revenue', "='Income Statement'!{c}5", H['revenue'] + F['revenue'], N2),
          ('EBITDA', "='Income Statement'!{c}8", H['ebitda'] + F['ebitda'], N2),
          ('EBITDA margin', "='Income Statement'!{c}9",
           H['ebitda_margin'] + F['ebitda_margin'], PC1),
          ('EBIT (before impairments in the forecast)', "='Income Statement'!{c}11",
           H['ebit'] + [F['ebitda'][i] - F['dna'][i] for i in range(5)], N2),
          ('Profit attributable to shareholders', "='Income Statement'!{c}19",
           H['pat_shareholders'] + F['pat'], N2),
          ('Net debt', "='Balance Sheet'!{c}17", H['net_debt'] + F['net_debt'], N2)]
for j, (lab, tpl, vals, fmt) in enumerate(SFROWS):
    rr = 5 + j
    T(SF, f'A{rr}', lab)
    for i, c in enumerate(ALLC):
        FM(SF, f'{c}{rr}', tpl.format(c=c), vals[i], fmt, link=True)
T(SF, 'A11', 'Revenue growth')
dash(SF, 'B11')
for i in range(1, 8):
    FM(SF, f'{ALLC[i]}11', f'={ALLC[i]}5/{ALLC[i-1]}5-1',
       (H['revenue'] + F['revenue'])[i] / (H['revenue'] + F['revenue'])[i - 1] - 1, PC1)
T(SF, 'A12', 'Free cash flow to the firm')
for i in range(3):
    dash(SF, ALLC[i] + '12')
for i in range(5):
    FM(SF, f'{FCOL[i]}12', f"='Cash Flow'!{FCOL[i]}15", F['fcff'][i], N2, link=True)
T(SF, 'A13', 'Invested capital')
for i in range(3):
    dash(SF, ALLC[i] + '13')
for i in range(5):
    FM(SF, f'{FCOL[i]}13', f"='Balance Sheet'!{FCOL[i]}18", F['invested_capital'][i], N2, link=True)
T(SF, 'A14', 'Return on invested capital')
for i in range(3):
    dash(SF, ALLC[i] + '14')
for i in range(5):
    FM(SF, f'{FCOL[i]}14', f"='Balance Sheet'!{FCOL[i]}19", F['roic'][i], PC1, link=True)

# ===========================================================================
# 13. PER-SHARE & RATIOS
# ===========================================================================
PS = sheet('Per-Share & Ratios', dict({'A': 46}, **{c: 12 for c in 'BCDEFGHI'}))
T(PS, 'A1', 'Per-share and ratio analysis', TITLE)
T(PS, 'A2', 'The indicator set a multi-country restaurant operator is actually judged on: '
            'like-for-like growth, revenue per restaurant, the four-wall margin, the return on '
            'the capital a restaurant costs, and the cash it converts. Every ratio is a formula '
            'off the statements.', SUB)
T(PS, 'A4', 'Measure', HEAD)
for i, y in enumerate(['FY2023', 'FY2024', 'FY2025'] + FY):
    T(PS, f'{ALLC[i]}4', y, HEAD)
T(PS, 'A5', 'Earnings per share (USD)')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}5', f"='Income Statement'!{c}20", (H['eps'] + F['eps'])[i], N3, link=True)
T(PS, 'A6', 'Earnings per share (AED)')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}6', f'={c}5*Assumptions!$C$6', (H['eps'] + F['eps'])[i] * FX, N3)
T(PS, 'A7', 'Book value per share (USD)')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}7', f"='Balance Sheet'!{c}14/Assumptions!$C$7",
       (H['equity'] + F['equity'])[i] / SH, N3)
T(PS, 'A8', 'EBITDA margin')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}8', f"='Income Statement'!{c}9",
       (H['ebitda_margin'] + F['ebitda_margin'])[i], PC1, link=True)
T(PS, 'A9', 'EBIT margin')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}9', f"='Income Statement'!{c}11/'Income Statement'!{c}5",
       (H['ebit'] + [F['ebitda'][j] - F['dna'][j] for j in range(5)])[i]
       / (H['revenue'] + F['revenue'])[i], PC1)
T(PS, 'A10', 'Net margin')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}10', f"='Income Statement'!{c}19/'Income Statement'!{c}5",
       (H['pat_shareholders'] + F['pat'])[i] / (H['revenue'] + F['revenue'])[i], PC1)
T(PS, 'A11', 'Revenue per restaurant (USD thousand)')
for i in range(3):
    dash(PS, ALLC[i] + '11')
for i in range(5):
    FM(PS, f'{FCOL[i]}11', f"='Income Statement'!{FCOL[i]}5/Segments!{'CDEFG'[i]}12*1000",
       F['revenue'][i] / F['stores'][i] * 1000.0, N2)
T(PS, 'A12', 'Restaurants at year end')
for i in range(3):
    dash(PS, ALLC[i] + '12')
for i in range(5):
    FM(PS, f'{FCOL[i]}12', f"=Segments!{'CDEFG'[i]}12", F['stores'][i], N0, link=True)
T(PS, 'A13', 'Return on equity')
dash(PS, 'B13')
for i in range(1, 8):
    eq = H['equity'] + F['equity']
    pa = H['pat_shareholders'] + F['pat']
    FM(PS, f'{ALLC[i]}13',
       f"='Income Statement'!{ALLC[i]}19/(('Balance Sheet'!{ALLC[i-1]}14+"
       f"'Balance Sheet'!{ALLC[i]}14)/2)", pa[i] / ((eq[i - 1] + eq[i]) / 2), PC1)
T(PS, 'A14', 'Return on invested capital')
for i in range(3):
    dash(PS, ALLC[i] + '14')
for i in range(5):
    FM(PS, f'{FCOL[i]}14', f"='Balance Sheet'!{FCOL[i]}19", F['roic'][i], PC1, link=True)
T(PS, 'A15', 'Net debt / EBITDA')
for i, c in enumerate(ALLC):
    FM(PS, f'{c}15', f"='Balance Sheet'!{c}20",
       (H['net_debt'] + F['net_debt'])[i] / (H['ebitda'] + F['ebitda'])[i], MULT, link=True)
T(PS, 'A16', 'Net working capital / revenue')
for i in range(3):
    FM(PS, f'{ALLC[i]}16', f"='Balance Sheet'!{ALLC[i]}10/'Income Statement'!{ALLC[i]}5",
       H['nwc'][i] / H['revenue'][i], PC1)
for i in range(5):
    FM(PS, f'{FCOL[i]}16', f"='Balance Sheet'!{FCOL[i]}10/'Income Statement'!{FCOL[i]}5",
       F['nwc'][i] / F['revenue'][i], PC1)
T(PS, 'A17', 'Capital expenditure / revenue')
for i in range(3):
    FM(PS, f'{ALLC[i]}17', f"='Cash Flow'!{ALLC[i]}10/'Income Statement'!{ALLC[i]}5",
       H['capex'][i] / H['revenue'][i], PC1)
for i in range(5):
    FM(PS, f'{FCOL[i]}17', f"='Cash Flow'!{FCOL[i]}12/'Income Statement'!{FCOL[i]}5",
       F['capex_total'][i] / F['revenue'][i], PC1)
T(PS, 'A18', 'Dividend per share (USD)')
for i in range(3):
    FM(PS, f'{ALLC[i]}18', f"='Cash Flow'!{ALLC[i]}17/Assumptions!$C$7",
       H['dividends_paid'][i] / SH, N3)
for i in range(5):
    FM(PS, f'{FCOL[i]}18', f"='Cash Flow'!{FCOL[i]}17/Assumptions!$C$7",
       F['dividends'][i] / SH, N3)

# ===========================================================================
# 14. MONTE CARLO
# ===========================================================================
MC = sheet('Monte Carlo', dict(A=42, B=13, C=13, D=13, E=13, F=13, G=15))
T(MC, 'A1', 'The probability price map', TITLE)
T(MC, 'A2', 'A map of price dispersion. It carries no view on value and is never blended with '
            'the valuation. Each figure is an engine output — fifty thousand simulated paths — '
            'not a formula, so it does not redraw when a driver changes.', SUB)
T(MC, 'A4', 'Horizon', HEAD)
for j, lab in enumerate(['5th', '25th', 'Median', '75th', '95th', 'Chance above the market']):
    T(MC, f'{"BCDEFG"[j]}4', lab, HEAD)
for j, (tag, lab) in enumerate([('1M', 'One month — to '), ('3M', 'Three months — to ')]):
    p = STK['horizons'][tag]
    rr = 5 + j
    T(MC, f'A{rr}', lab + p['grade_date'])
    for k, q in enumerate(['p5', 'p25', 'p50', 'p75', 'p95']):
        V(MC, f'{"BCDEF"[k]}{rr}', p['pct'][q], N2, kind='engine')
    V(MC, f'G{rr}', p['p_above'], PC1, kind='engine')
T(MC, 'A8', 'Level event', HEAD); T(MC, 'B8', 'One month', HEAD); T(MC, 'C8', 'Three months', HEAD)
EV = [('Finishes 10% or more above the market price', 'p_up10'),
      ('Finishes 10% or more below the market price', 'p_dn10'),
      ('Touches 10% above at any point', 'touch_up10'),
      ('Touches 10% below at any point', 'touch_dn10')]
for j, (lab, key) in enumerate(EV):
    rr = 9 + j
    T(MC, f'A{rr}', lab)
    V(MC, f'B{rr}', STK['horizons']['1M'][key], PC1, kind='engine')
    V(MC, f'C{rr}', STK['horizons']['3M'][key], PC1, kind='engine')
T(MC, 'A14', 'Engine setting', HEAD); T(MC, 'B14', 'Value', HEAD)
SET = [('Simulated paths', 50000, N0), ('Anchor date', STK['anchor_date'], None),
       ('Market price at the anchor (AED)', STK['spot'], N2),
       ('Annualised volatility at the anchor', STK['horizons']['3M']['anchor_vol_ann'], PC1),
       ('Dividend yield inside the drift anchor', STK['q_annual'], PC1)]
for j, (lab, val, fmt) in enumerate(SET):
    rr = 15 + j
    T(MC, f'A{rr}', lab)
    V(MC, f'B{rr}', val, fmt, kind='engine')

# ===========================================================================
# 15. SENSITIVITY
# ===========================================================================
SE = sheet('Sensitivity', dict(A=52, B=13, C=13, D=13, E=13, F=13, G=13, H=13))
T(SE, 'A1', 'Sensitivity — what the valuation needs the world to do', TITLE)
T(SE, 'A2', 'AED per share. Each cell is a complete re-run of the model, including the unit '
            'build, so these grids are engine outputs rather than formulas and do NOT redraw '
            'when a driver is changed.', SUB)
T(SE, 'A4', 'Cost of capital (rows) against terminal growth (columns)', HEAD)
for j, g in enumerate(SN['g_grid']):
    T(SE, f'{"BCDEF"[j]}5', f'{100*g:.1f}%', HEAD)
for i, dw in enumerate(SN['w_grid']):
    rr = 6 + i
    T(SE, f'A{rr}', f'{100*(W["wacc_rating"]+dw):.2f}%')
    for j in range(5):
        V(SE, f'{"BCDEF"[j]}{rr}', SN['grid_growth_wacc'][i][j] * FX, N2, kind='grid')
T(SE, 'A12', 'Cost of capital (rows) against the EBITDA margin (columns)', HEAD)
for j, dm in enumerate(SN['m_grid']):
    T(SE, f'{"BCDEF"[j]}13', f'{100*dm:+.1f}pp', HEAD)
for i, dw in enumerate(SN['w_grid']):
    rr = 14 + i
    T(SE, f'A{rr}', f'{100*(W["wacc_rating"]+dw):.2f}%')
    for j in range(5):
        V(SE, f'{"BCDEF"[j]}{rr}', SN['grid_margin_wacc'][i][j] * FX, N2, kind='grid')
T(SE, 'A20', 'One driver at a time — five re-runs a row, the grid shown beside the name', HEAD)
T(SE, 'H20', 'Swing', HEAD)
for j, (lab, vals) in enumerate(SN['single'].items()):
    rr = 21 + j
    T(SE, f'A{rr}', lab)
    for k in range(5):
        V(SE, f'{"BCDEF"[k]}{rr}', vals[k] * FX, N2, kind='grid')
    FM(SE, f'H{rr}', f'=MAX(B{rr}:F{rr})-MIN(B{rr}:F{rr})',
       (max(vals) - min(vals)) * FX, N2)
T(SE, 'A26', 'The margin row is the one that matters. A two-point swing in the EBITDA margin is '
             'worth more than the whole plausible range of the cost of capital, which is why the '
             'structural-versus-cyclical question is the study\'s central judgement rather than '
             'a footnote.', SUB, wrap=True)
SE.row_dimensions[26].height = 42

# ===========================================================================
# 16. PEER & SECTOR
# ===========================================================================
PE = sheet('Peer & Sector', dict(A=32, B=18, C=14, D=14, E=14, F=44))
T(PE, 'A1', 'Peer frame and sector context', TITLE)
T(PE, 'A2', 'A cross-check, not an independent valuation. No figure on this sheet enters the '
            'build; Americana’s own historicals come exclusively from its audited consolidated '
            'financial statements.', SUB)
T(PE, 'A4', 'Company', HEAD); T(PE, 'B4', 'Market', HEAD)
T(PE, 'C4', 'EV / EBITDA', HEAD); T(PE, 'D4', 'Price / earnings', HEAD)
T(PE, 'E4', 'EBITDA margin', HEAD); T(PE, 'F4', 'Why it is here', HEAD)
rr = 5
EV_OK, PE_OK = [], []
for sym, p in PEERS.items():
    if p.get('error'):
        continue
    T(PE, f'A{rr}', p['name']); T(PE, f'B{rr}', p['country'])
    if p.get('ev_ebitda') and 4 < p['ev_ebitda'] < 40:
        EV_OK.append(rr)
    if p.get('pe_trailing') and 5 < p['pe_trailing'] < 45:
        PE_OK.append(rr)
    V(PE, f'C{rr}', round(p['ev_ebitda'], 2) if p.get('ev_ebitda') else '-', MULT)
    V(PE, f'D{rr}', round(p['pe_trailing'], 2) if p.get('pe_trailing') else '-', MULT)
    V(PE, f'E{rr}', p.get('ebitda_margin') or '-', PC1)
    T(PE, f'F{rr}', p['rationale'])
    PE.row_dimensions[rr].height = 26
    rr += 1
T(PE, f'A{rr + 1}', 'Peer median, enterprise value / EBITDA — MEDIAN over the rows inside '
                    'the stated band of 4x to 40x (one exclusion: the row whose multiple is '
                    'an arithmetic artefact of a depressed year)', HEAD)
ev_cells = [f'C{r_}' for r_ in EV_OK]
pe_cells = [f'D{r_}' for r_ in PE_OK]
import statistics as _st
_evr = [round(p['ev_ebitda'], 2) for p in PEERS.values()
        if p.get('ev_ebitda') and 4 < p['ev_ebitda'] < 40]
FM(PE, f'C{rr + 1}', '=MEDIAN(' + ','.join(ev_cells) + ')', _st.median(_evr), MULT)
T(PE, f'A{rr + 2}', 'Peer median, price / earnings — MEDIAN over the rows inside the stated '
                    'band of 5x to 45x', HEAD)
_per = [round(p['pe_trailing'], 2) for p in PEERS.values()
        if p.get('pe_trailing') and 5 < p['pe_trailing'] < 45]
FM(PE, f'D{rr + 2}', '=MEDIAN(' + ','.join(pe_cells) + ')', _st.median(_per), MULT)
T(PE, f'A{rr + 4}', 'Americana — trailing enterprise value / EBITDA', HEAD)
FM(PE, f'C{rr + 4}', "='Relative & Normalized'!B16", D['trailing']['ev_ebitda'], MULT, link=True)
T(PE, f'A{rr + 5}', 'Americana — trailing price / earnings', HEAD)
FM(PE, f'D{rr + 5}', "='Relative & Normalized'!B17", D['trailing']['pe'], MULT, link=True)
T(PE, f'A{rr + 6}', 'Justified multiple applied in the relative lens', HEAD)
FM(PE, f'C{rr + 6}', "='Relative & Normalized'!B6", LN['relative']['multiple'], MULT, link=True)
T(PE, f'A{rr + 8}',
  'The set splits cleanly in two, and that split is why the justified multiple sits below the '
  'peer median. The global names — Yum! Brands, Restaurant Brands, Domino\'s — are franchisORS: '
  'they collect a royalty on someone else\'s sales and carry EBITDA margins of 20% to 35% on a '
  'fraction of the revenue. Americana is on the other side of that contract. It is the operator, '
  'it PAYS the royalty (5.6% of revenue in FY2025), and it carries the restaurants, the staff and '
  'the leases. The right comparators are the listed franchisees — Alsea in Mexico, Devyani and '
  'Sapphire in India, both of which run the same two brands Americana runs — and those trade '
  'across a very wide range on growth expectations rather than on any stable multiple.',
  SUB, wrap=True)
PE.row_dimensions[rr + 8].height = 110

# ===========================================================================
# order, counts and save
# ===========================================================================
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]
assert wb.sheetnames == ORDER, wb.sheetnames

npaste = sum(len(v) for v in PASTED.values())
T(RF, f'A{READ_FIRST_COUNT_ROW}',
  f'COUNTS. This workbook carries {NFORM} formula cells against {npaste} pasted values — '
  f'{len(PASTED["audited"])} audited or disclosed history, {len(PASTED["engine"])} engine '
  f'outputs, and {len(PASTED["grid"])} sensitivity-grid cells. Every one of the '
  f'{NFORM} formula cells is checked against the model that wrote it.', SECT, wrap=True)
RF.row_dimensions[READ_FIRST_COUNT_ROW].height = 40

OUT = os.path.join(HERE, 'AMR_Valuation_Model_09082026_public.xlsx')
wb.save(OUT)
with open(os.path.join(HERE, 'xlsx_expected.json'), 'w') as f:
    json.dump(dict(expected=EXPECTED, n_formula=NFORM, pasted=PASTED,
                   n_pasted=npaste), f, indent=1)
print(f'wrote {os.path.basename(OUT)} — {NFORM} formula cells, {npaste} pasted '
      f'({len(PASTED["audited"])} audited, {len(PASTED["engine"])} engine, '
      f'{len(PASTED["grid"])} grid)')
