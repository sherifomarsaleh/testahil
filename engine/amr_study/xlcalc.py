"""A small, explicit spreadsheet evaluator.

The delivered model is verified by evaluating it here rather than by handing it back to the
application that wrote it: arithmetic, SUM/MIN/MAX/MEDIAN/AVERAGE over ranges or over
comma-separated scalars, absolute and relative cell references, and cross-sheet references.
Anything the evaluator does not understand raises rather than being silently skipped.

This started as a workaround — LibreOffice could not load any document here, because only
libreoffice-core was installed and the Writer/Calc import filters were absent. That is fixed
(see engine/make_pdf.py), but the evaluator is KEPT deliberately: an independent
reimplementation that has to agree with the model cell-for-cell is a stronger check than
asking a spreadsheet engine to confirm its own arithmetic, and it is what carries the
expected-value gate in recalc.py.

Used by recalc.py — does the workbook reproduce the model? — and by driver_test.py — does
changing a driver on the Assumptions sheet actually reprice the workbook?
"""
import re
from openpyxl.utils import range_boundaries, get_column_letter

FUNC = re.compile(r'\b(SUM|MIN|MAX|MEDIAN|AVERAGE)\(([^()]*)\)')
# An UNQUOTED sheet name may not contain a hyphen or a space: every sheet in this workbook
# whose name does is written quoted. Allowing them made "C34-Assumptions!$C$45" parse as a
# reference to a sheet called "C34-Assumptions", which silently swallowed the subtraction.
SHEETREF = re.compile(r"(?<![A-Za-z0-9_$!.])(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))"
                      r"!(\$?[A-Z]{1,3}\$?\d+)")
CELLREF = re.compile(r'(?<![A-Z0-9_!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![\d(])')
RANGE = re.compile(r"^(?:(?:'[^']+'|[A-Za-z][A-Za-z0-9_]*)!)?\$?[A-Z]{1,3}\$?\d+"
                   r":\$?[A-Z]{1,3}\$?\d+$")


class Book:
    """Evaluates a workbook's formulas, with an optional override layer over input cells."""

    def __init__(self, wb, overrides=None):
        self.wb = wb
        self.overrides = {(s, c): v for (s, c), v in (overrides or {}).items()}
        self.cache = {}
        self.stack = []

    def cell_value(self, sheet, coord):
        key = (sheet, coord)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise ValueError(f'circular reference at {sheet}!{coord}')
        v = self.wb[sheet][coord].value
        if isinstance(v, str) and v.startswith('='):
            self.stack.append(key)
            try:
                v = self.evaluate(v[1:], sheet)
            finally:
                self.stack.pop()
        elif isinstance(v, str):
            v = 0.0 if v == '-' else v
        elif v is None:
            v = 0.0
        self.cache[key] = v
        return v

    def range_values(self, sheet, rng):
        c1, r1, c2, r2 = range_boundaries(rng)
        out = []
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                out.append(self.cell_value(sheet, f'{get_column_letter(cc)}{rr}'))
        return [x for x in out if isinstance(x, (int, float))]

    def arg_values(self, arg, sheet):
        """A function argument is either one range, or a comma-separated list of scalars."""
        arg = arg.strip()
        if RANGE.match(arg):
            tgt, rng = sheet, arg
            sm = re.match(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))!(.+)", arg)
            if sm:
                tgt = sm.group(1) or sm.group(2); rng = sm.group(3)
            return self.range_values(tgt, rng.replace('$', ''))
        return [float(self.evaluate(part, sheet)) for part in arg.split(',') if part.strip()]

    def evaluate(self, expr, sheet):
        e = expr
        # functions over ranges or scalar lists (possibly cross-sheet)
        while True:
            m = FUNC.search(e)
            if not m:
                break
            fn = m.group(1)
            vals = self.arg_values(m.group(2), sheet)
            if not vals:
                val = 0.0
            elif fn == 'SUM':
                val = sum(vals)
            elif fn == 'MIN':
                val = min(vals)
            elif fn == 'MAX':
                val = max(vals)
            elif fn == 'AVERAGE':
                val = sum(vals) / len(vals)
            else:
                vs = sorted(vals); n = len(vs)
                val = (vs[n // 2] if n % 2 else (vs[n // 2 - 1] + vs[n // 2]) / 2)
            e = e[:m.start()] + repr(float(val)) + e[m.end():]
        # cross-sheet single-cell references
        while True:
            m = SHEETREF.search(e)
            if not m:
                break
            tgt = m.group(1) or m.group(2)
            v = self.cell_value(tgt, m.group(3).replace('$', ''))
            e = e[:m.start()] + repr(float(v or 0)) + e[m.end():]
        # same-sheet references
        while True:
            m = CELLREF.search(e)
            if not m:
                break
            v = self.cell_value(sheet, f'{m.group(2)}{m.group(4)}')
            e = e[:m.start()] + repr(float(v or 0)) + e[m.end():]
        # Excel's caret is exponentiation. Translate AFTER all references have been
        # substituted, so a '^' inside a sheet name can never be caught by this.
        e = e.replace('^', '**')
        if not re.fullmatch(r'[-+*/(). 0-9eE]+|[-+*/(). 0-9eE*]+', e.replace('**', '*')):
            raise ValueError(f'unparsed formula fragment: {expr!r} -> {e!r}')
        return eval(e)

    def formula_cells(self):
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith('='):
                        yield ws.title, c.coordinate
