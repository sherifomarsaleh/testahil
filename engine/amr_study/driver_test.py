"""AMR — does the delivered workbook actually reprice when a driver changes?

Two passes, both run on the DELIVERED file, not on the builder:

  1. DIRECTIONAL. Each named driver is perturbed in place on the Assumptions sheet, the
     whole workbook is re-evaluated through the independent evaluator, and the headline
     (the weighted central value on Summary!C10) must move in the asserted direction by a
     non-trivial amount. If a test fails, the first hypothesis is that the EXPECTATION is
     wrong, not the model — so the failure prints the mechanism, not just the mismatch.

  2. DEAD-INPUT SWEEP. Every remaining blue cell on the Assumptions sheet is perturbed and
     must move at least one formula cell somewhere in the workbook. An input that changes
     nothing is a defect: it is either wired to nothing or documenting a number the model
     does not use.
"""
import json, os
import openpyxl
import xlcalc

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'AMR_Valuation_Model_09082026_public.xlsx')
HEADLINE = ('Summary', 'C10')

wb = openpyxl.load_workbook(XLSX)
BASE_BOOK = xlcalc.Book(wb)
BASE = BASE_BOOK.cell_value(*HEADLINE)
ALL_FORMULA = list(BASE_BOOK.formula_cells())
BASE_ALL = {}
for sh, cd in ALL_FORMULA:
    try:
        BASE_ALL[(sh, cd)] = BASE_BOOK.cell_value(sh, cd)
    except Exception:
        BASE_ALL[(sh, cd)] = None
print(f'headline (weighted central, AED per share): {BASE:.4f}')


def headline_with(overrides):
    bk = xlcalc.Book(openpyxl.load_workbook(XLSX), overrides=overrides)
    return bk.cell_value(*HEADLINE)


def all_with(overrides):
    bk = xlcalc.Book(openpyxl.load_workbook(XLSX), overrides=overrides)
    out = {}
    for sh, cd in ALL_FORMULA:
        try:
            out[(sh, cd)] = bk.cell_value(sh, cd)
        except Exception:
            out[(sh, cd)] = None
    return out


ASM = wb['Assumptions']


def cur(coord):
    return float(ASM[coord].value)


# ---- pass 1: directional ---------------------------------------------------
# (label, cells, multiplier or delta, expected direction, why)
MULT, DELTA = 'x', '+'
CASES = [
    ('Like-for-like sales growth', [f'{c}30' for c in 'BCDEF'], (DELTA, 0.01), +1,
     'higher revenue per restaurant lifts revenue, EBITDA and free cash flow'),
    ('Net new restaurants', [f'{c}11' for c in 'BCDEF'], (MULT, 1.20), +1,
     'more restaurants means more revenue; the capital they cost is well inside the '
     'cash they generate at a three-year payback'),
    ('Food and packaging cost', [f'{c}49' for c in 'BCDEF'], (DELTA, 0.01), -1,
     'a costlier input basket takes a point of margin straight out of EBITDA'),
    ('Staff per restaurant (the unit build volume side)', [f'{c}109' for c in 'BCDEF'],
     (MULT, 1.05), -1, 'more heads per restaurant at the same wage is more cost'),
    ('Wage growth per full-time equivalent', ['C111'], (DELTA, 0.01), -1,
     'the price side of the staff unit build'),
    ('Delivery channel share', [f'{c}113' for c in 'BCDEF'], (DELTA, 0.02), -1,
     'a costlier channel taking a larger share of revenue'),
    ('Delivery cost per delivered dollar', [f'{c}114' for c in 'BCDEF'], (DELTA, 0.005), -1,
     'the price side of the delivery unit build'),
    ('Recurring impairment rate', ['C115'], (DELTA, 0.002), -1,
     'a straight charge against EBIT and the terminal cash flow'),
    ('Terminal return on incremental capital', ['C116'], (DELTA, 0.10), +1,
     'a higher terminal return needs less reinvestment for the same growth'),
    ('Cyclical margin path', [f'{c}119' for c in 'BCDEF'], (DELTA, 0.005), +1,
     'a kinder cyclical reading lifts the normalised lens midpoint margin'),
    ('Royalty rate', [f'{c}50' for c in 'BCDEF'], (DELTA, 0.005), -1,
     'a higher royalty is a direct transfer to the franchisor'),
    ('Beta', ['C81'], (DELTA, 0.20), -1,
     'a higher beta raises the cost of equity and discounts the same cash flows harder'),
    ('Equity risk premium (ratings basis)', ['C78'], (DELTA, 0.01), -1, 'the same mechanism'),
    ('US ten-year Treasury yield', ['C76'], (DELTA, 0.01), -1,
     'a higher risk-free rate raises the whole cost of capital'),
    ('Terminal growth', ['C84'], (DELTA, 0.005), +1,
     'faster perpetual growth raises the terminal value, and three-quarters of the '
     'enterprise value sits there'),
    ('Terminal risk-free rate', ['C83'], (DELTA, 0.01), -1,
     'raises the terminal cost of capital, which capitalises the terminal cash flow'),
    ('Cost of debt', ['C82'], (DELTA, 0.01), -1,
     'raises the after-tax cost of debt inside the weighted cost, and the lease interest '
     'charge on the income statement with it'),
    ('Effective tax rate', [f'{c}71' for c in 'BCDEF'], (DELTA, 0.02), -1,
     'less of every dollar of EBIT survives to NOPAT'),
    ('Capital expenditure per new restaurant', ['C61'], (MULT, 1.25), -1,
     'each opening costs more, so free cash flow falls'),
    ('Maintenance capital expenditure', ['C62'], (DELTA, 0.005), -1, 'the same, on the estate'),
    ('Additions to right-of-use assets', ['C64'], (DELTA, 0.01), -1,
     'taking leases is investment on the capitalised reading, so more of it is cash out'),
    ('Net working capital', ['C68'], (DELTA, 0.01), -1,
     'working capital is negative here, so making it LESS negative consumes cash'),
    ('Justified enterprise value / EBITDA', ['C100'], (DELTA, 1.0), +1,
     'lifts the relative lens, which carries a fifth of the weight'),
    ('Justified price / earnings', ['C101'], (DELTA, 2.0), +1, 'lifts the normalised lens'),
    ('Sustainable return on equity', ['C102'], (DELTA, 0.05), +1, 'lifts the book lens'),
    ('Market price', ['C5'], (DELTA, 0.20), -1,
     'a higher market price raises the equity weight in the cost of capital, and since '
     'equity is dearer than debt here the weighted cost rises and the value falls — the '
     'valuation does not chase the price'),
    ('Restaurants at 31 December 2025', [f'C{21 + j}' for j in range(7)], (MULT, 1.05), +1,
     'a larger opening estate carries through the whole forecast'),
    ('Revenue per restaurant, FY2025 base', [f'C{40 + j}' for j in range(7)], (MULT, 1.05), +1,
     'the price side of the unit build, lifted at the base year'),
]

rows, failures = [], []
for label, cells, (mode, amt), direction, why in CASES:
    ov = {}
    for cd in cells:
        base = cur(cd)
        ov[('Assumptions', cd)] = base * amt if mode == MULT else base + amt
    got = headline_with(ov)
    move = got - BASE
    pct = move / BASE
    ok = (abs(pct) > 1e-4) if direction == 0 else (move * direction > 0 and abs(pct) > 1e-4)
    rows.append(dict(driver=label, cells=cells, perturbation=(mode, amt),
                     expected='higher' if direction > 0 else
                              ('lower' if direction < 0 else 'not asserted, must not be dead'),
                     base=BASE, moved_to=got, change_pct=pct, passed=bool(ok), mechanism=why))
    flag = 'OK ' if ok else 'BAD'
    print(f'  [{flag}] {label:44s} {BASE:.4f} -> {got:.4f}  ({pct:+.2%})')
    if not ok:
        failures.append(label)
        print(f'         mechanism as understood: {why}')

# The dividend payout ratio is deliberately NOT in the directional list. It was tested
# there first and moved the headline by exactly zero, so the mechanism was decomposed
# rather than the model changed: the enterprise value is struck before financing, the
# bridge subtracts the AUDITED 31-December-2025 net debt, and each lens is anchored on a
# year whose opening cash predates any forecast distribution. Paying more out therefore
# moves the balance sheet, the cash roll and later-year finance income — all of which the
# sweep below confirms — without moving the value of the firm. That is correct behaviour
# for a free-cash-flow-to-the-firm model, not a wiring fault, so it is asserted in the
# sweep (must not be dead) rather than the directional pass (must move the headline).

# ---- pass 2: dead-input sweep ---------------------------------------------
tested = {cd for _, cells, _, _, _ in CASES for cd in cells}
blue = []
for row in ASM.iter_rows():
    for c in row:
        if c.value is None or isinstance(c.value, str):
            continue
        if c.font and c.font.color and c.font.color.rgb and 'IF4E9C' != c.font.color.rgb:
            pass
        if c.column_letter in ('B', 'C', 'D', 'E', 'F') and isinstance(c.value, (int, float)):
            blue.append(c.coordinate)
untested = [cd for cd in blue if cd not in tested]
dead = []
for cd in untested:
    base = cur(cd)
    newv = base * 1.10 if abs(base) > 1e-12 else 0.01
    after = all_with({('Assumptions', cd): newv})
    moved = sum(1 for k in BASE_ALL
                if isinstance(BASE_ALL[k], (int, float)) and isinstance(after[k], (int, float))
                and abs(after[k] - BASE_ALL[k]) > 1e-9)
    if moved == 0:
        dead.append(cd)
print(f'\ndead-input sweep: {len(untested)} further inputs perturbed, {len(dead)} dead')
for cd in dead:
    print('   DEAD:', cd, ASM[f'A{cd[1:]}'].value)

res = dict(headline_cell=f'{HEADLINE[0]}!{HEADLINE[1]}', headline_base=BASE,
           directional=rows, n_directional=len(rows),
           n_directional_failed=len(failures),
           dead_sweep_inputs=len(untested), dead_inputs=dead)
json.dump(res, open(os.path.join(HERE, 'driver_test_result.json'), 'w'), indent=1)
assert not failures, f'{len(failures)} directional driver tests failed: {failures}'
assert not dead, f'{len(dead)} dead inputs: {dead}'
print(f'\nDRIVER TEST OK — {len(rows)} drivers perturbed in place on the delivered file, every '
      f'one moves the headline as asserted; {len(untested)} further inputs swept, zero dead')
