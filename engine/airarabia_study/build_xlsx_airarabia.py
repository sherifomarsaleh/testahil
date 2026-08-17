"""AIRARABIA_Valuation_Model_09082026_public.xlsx — 16 sheets mirroring the house
canonical model (operating-company variant). Blue = inputs · black = formulas ·
green = cross-sheet links.

The workbook is FORMULA-DRIVEN. Every quantity that is arithmetically derivable
from an input is written as a live Excel formula. Only three classes of cell are
pasted values:

  1. audited and disclosed historical figures (the primary record);
  2. the unit build's disclosed FY2025 bases (passengers, revenue per passenger,
     cost per passenger, the FY2025 revenue-line levels) — the unit build itself
     is LIVE: forecast revenue = passengers x rate, cost = passengers x unit
     cost, all as formulas off the Assumptions sheet;
  3. whole-model re-runs: the Monte Carlo price map, the sensitivity grids, the
     DCF scenario bear/bull bounds, the high-fuel alternative valuation and the
     expert-panel legs — each figure there is a complete revaluation and cannot
     be a single formula.

Every formula cell also carries the model's own value into xlsx_expected.json,
and recalc.py evaluates the workbook independently and asserts the two agree.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'; NUM2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, REL, NRM, BK = D['experts'], D['rel'], D['norm'], D['book']
BU, S0, STK = D['bottomup'], D['step0'], D['strike']
IN = {k: v['value'] for k, v in D['inputs'].items()}
SPOT, SH = M['spot'], M['shares_mn']
TAX = IN['tax_eff']
NCI_SH = DCF['nci_share']
YF = F['years']
H3 = ['FY23', 'FY24', 'FY25']
CD = ['B', 'C', 'D', 'E', 'F']              # forecast columns (Segments / DCF / Cash Flow)
HC = ['B', 'C', 'D']                        # historical columns on the statements
FCOL = ['E', 'F', 'G', 'H', 'I']            # forecast columns on the statements
ALL = HC + FCOL

nd25 = HB['FY25']['nd']
FLC = IN['fleet_cons']
NCIB = IN['nci_book']
ROLLC = DCF['roll_cash']
debt25 = HB['FY25']['debt']
liq25 = HB['FY25']['cash'] + HB['FY25']['dep']
nwc25 = HB['FY25']['nwc']
fleet25 = F['fleet_assets_fy25']
intang25 = HB['FY25']['intang']
nonop25 = DCF['non_op']
jv_book, jv_cap = DCF['jv_book'], DCF['jv_cap']
GA_CASH_25 = IN['ga_fy25'] - 40.081
FUEL_BASE = [IN['fuel_intensity'] * p for p in IN['jet_eff_base']]

# split cargo/service expected paths (compute carries their sum)
cargo_e, svc_e, hotel_e, lease_e = [], [], [], []
c_, s_, h_, l_ = 186.948, 223.323, 59.842, 219.755
for i in range(5):
    c_ *= 1 + IN['cargo_g'][i]; s_ *= 1 + IN['svc_g'][i]
    h_ *= 1 + IN['hotel_g'][i]; l_ *= 1 + IN['lease_g'][i]
    cargo_e.append(c_); svc_e.append(s_); hotel_e.append(h_); lease_e.append(l_)
assoc_e = F['assoc']
ga_e, sm_e, oth_e = F['ga'], F['sm'], F['other_inc']

wb = Workbook()
EXPECT = {}
ANCH = {}

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Air Arabia PJSC (DFM: AIRARABIA)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the Air Arabia valuation study. Every blue cell is an',
 'input; every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice: the cost of capital is built from',
 'the dirham government-bond yield net of the sovereign spread, beta and the premium rather than pasted; the',
 'discount factors compound from the cost-of-debt glide; revenue is passengers times revenue per passenger;',
 'the cost stack is passengers times cost per passenger, line by line; and the income statement, balance',
 'sheet, cash flow, ratios and all four lenses chain off the same cells.', '',
 'THREE THINGS ARE PASTED VALUES, and it is worth knowing exactly which. First, audited and disclosed',
 'history — the primary record, not a calculation; where a line is both disclosed and derivable the',
 'DISCLOSED figure is carried. Second, the unit build\'s disclosed bases: FY2025 passengers, the per-',
 'passenger revenue and cost rates, and the FY2025 revenue-line levels are pasted from the audited',
 'disclosures and grown on their own drivers — the build itself is live formula from there. Third,',
 'whole-model re-runs, where each figure is a complete revaluation of the entire model and so cannot be a',
 'single formula: the Monte Carlo price map, the sensitivity grids, the DCF scenario bear/bull bounds, the',
 'high-fuel alternative valuation, and the expert-panel legs. Everything else — every lens base value, the',
 'joint-venture bridge on both framings, and the anchor-date roll — is a live formula. Changing a driver',
 'reprices the model but does NOT redraw the engine outputs.', '',
 'How revenue is built. Not as one growth rate. Passenger and baggage revenue is passengers times revenue',
 'per passenger; ancillary revenue is passengers times an ancillary rate; cargo, service, hotel and',
 'aircraft-lease income grow on their own paths. Costs are built the same way: fuel, staff, maintenance,',
 'landing, handling and other costs are each per-passenger rates with their OWN escalator (fuel on the',
 'commodity path, staff on wages, airport charges on UAE inflation). Margins are OUTPUTS.', '',
 'The beta is switchable, and both regressions are in the sheet. The adopted beta is this stock\'s own',
 'five-year weekly regression against the Dubai market index — the exchange every filing says the shares',
 'are listed on, and the index the company\'s own annual report benchmarks itself against. The same',
 'regression run against the Abu Dhabi market index is carried beside it as a cross-check: it is a lower',
 'beta, but it explains only a third as much of the share\'s weekly movement, which is why it is published',
 'rather than adopted. Set the benchmark switch on Assumptions to 1 and the whole workbook reprices on it.', '',
 'Two judgements are shown both ways, never averaged. The fuel path: the base case follows the official',
 'US energy-agency curve (relief from 2027); the alternative holds the airline-association high-fuel',
 'assumption. And the joint-venture network (Abu Dhabi, Egypt, Pakistan, Morocco, the new Saudi venture):',
 'the base carries it at its audited balance-sheet value of AED 363mn; the alternative capitalises its',
 'profit share at a growth multiple. Both appear on the Summary and the SOTP Bridge.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model',
 'outputs shown as ranges.', '',
 'Sourcing note, up front. FY2023, FY2024 and FY2025 all come from the company\'s own audited consolidated',
 'financial statements, read from the company\'s investor-relations page; FY2024 comparatives are as',
 'restated in the FY2025 filing. The interim quarter is the reviewed Q1-2026 filing. Passengers, load',
 'factor and fleet come from the company\'s own results presentations. Every input is listed with source',
 'and date in the companion bibliography document.', '',
 'Discount convention. Each explicit year is discounted at its own forward cost of capital, gliding',
 f"{W['wacc_exp']*100:.2f}% -> {W['wacc_term']*100:.2f}% on the same easing calendar as the cost-of-debt path; the terminal value is",
 'capitalised at the terminal rate and discounted at the year-5 cumulative factor. One date, one price of',
 'time: values are struck at 31-Dec-2025 and rolled to the 7-Aug-2026 anchor at the cost of equity, net of',
 'the AED 0.30 dividend paid inside the window.', '',
 f"Currency. AED million unless stated. Spot AED {SPOT:.2f} ({M['asof']} close). Sheets: READ FIRST · Summary ·",
 'Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income',
 'Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share &',
 'Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 112

# ============ 4 ASSUMPTIONS (built early so referenced rows are fixed) =========
wsA = sheet('Assumptions')
title(wsA, 'Assumptions — every driver of the model', 'Blue cells are inputs. Change one and the model reprices.',
      8, awidth=64, cwidth=11)
hdr(wsA, 4, ['Driver'] + YF)
PATHS = [
    ('Passengers (millions)', IN['pax_path'], NUM2),
    ('Passenger + baggage revenue per passenger (AED)', IN['fare_path'], NUM1),
    ('Ancillary revenue per passenger (AED)', IN['anc_path'], NUM1),
    ('Cargo revenue growth', IN['cargo_g'], PCT),
    ('Service revenue growth', IN['svc_g'], PCT),
    ('Hotel revenue growth', IN['hotel_g'], PCT),
    ('Aircraft-lease revenue growth', IN['lease_g'], PCT),
    ('Effective jet fuel price (USD/bbl) — base path', IN['jet_eff_base'], NUM1),
    ('Effective jet fuel price (USD/bbl) — high-fuel alternative', IN['jet_eff_alt'], NUM1),
    ('Staff cost per passenger (AED)', IN['staff_per_pax'], NUM1),
    ('Maintenance cost per passenger (AED)', IN['maint_per_pax'], NUM1),
    ('Landing & overflying charges per passenger (AED)', IN['landing_per_pax'], NUM1),
    ('Handling charges per passenger (AED)', IN['handling_per_pax'], NUM1),
    ('Other direct cost per passenger (AED)', IN['other_per_pax'], NUM1),
    ('Administrative cost growth', IN['ga_g'], PCT),
    ('Other-income growth', IN['other_g'], PCT),
    ('Depreciation & amortisation (AED mn)', IN['dna_path'], NUM0),
    ('Fleet capital expenditure incl. pre-delivery payments (AED mn)', IN['capex_path'], NUM0),
    ('Owned aircraft additions (units)', FLC['owned_adds'], NUM0),
    ('Leased aircraft additions (units)', FLC['leased_adds'], NUM0),
    ('Consolidated fleet, year-end (aircraft)', FLC['ends'], NUM0),
    ('Booked finance-cost rate on the debt book', IN['kd_booked_path'], PCT2),
    ('Debt amortisation (AED mn)', IN['debt_amort'], NUM0),
    ('Marginal cost of debt path', IN['kd_path'], PCT2),
    ('Deposit yield path', IN['dep_rate_path'], PCT2),
    ('Growth in the share of joint-venture and associate profit', IN['assoc_g'], PCT),
    ('Dividend payout ratio', IN['payout'], PCT),
]
AROW = {}
r = 5
for label, vals, fmt in PATHS:
    put(wsA, f'A{r}', label, fmt=None)
    for j, cc in enumerate(CD):
        put(wsA, f'{cc}{r}', vals[j], BLUE, fmt)
    AROW[label] = r
    r += 1
r += 1
hdr(wsA, r, ['Scalar driver', 'Value']); r += 1
SCALARS = [
    ('Selling & marketing / revenue', IN['sm_pct'], PCT2),
    ('Working capital / revenue', IN['nwc_pct'], PCT),
    ('Tax rate', TAX, PCT),
    ('AED government bond yield (January-2031 tranche)', IN['rf'], PCT2),
    ('Observed sovereign spread at the auction (netted out)', IN['sov_spread_obs'], PCT2),
    ('Rating-table sovereign spread (alternative netting, disclosed)', IN['sov_spread_rating'], PCT2),
    ('Beta — adopted (own stock, five-year weekly vs the Dubai market index)',
     IN['beta_used'], NUM2),
    ('Beta — alternative benchmark (same regression vs the Abu Dhabi market index)',
     IN['beta_alt_benchmark'], NUM2),
    ('Equity risk premium', IN['erp_rating'], PCT2),
    ('Terminal risk-free rate', IN['rf_term'], PCT2),
    ('Terminal equity risk premium', IN['erp_term'], PCT2),
    ('Marginal cost of debt', IN['kd'], PCT2),
    ('Terminal cost of debt', IN['kd_term'], PCT2),
    ('Terminal debt weight', IN['wd_term'], PCT),
    ('Terminal growth', IN['g_term'], PCT),
    ('Joint-venture capitalisation multiple', IN['jv_pe'], MULT),
    ('Justified EV/EBITDA', IN['ev_ebitda_just'], MULT),
    ('Justified price/earnings', IN['pe_just'], MULT),
    ('Sustainable return on equity', IN['roe_sust'], PCT),
    ('Spot price (AED)', SPOT, PX),
    ('Shares outstanding (mn)', SH, NUM0),
    ('FY2025 dividend per share (AED, approved 12 March 2026)', IN['dps_fy25'], PX),
    ('Days from 31-Dec-2025 to the 7-Aug-2026 anchor', IN['anchor_days'], NUM0),
    ('Fuel intensity (AED per passenger per USD/bbl of effective jet price)', IN['fuel_intensity'], '0.000'),
    ('Right-of-use value per leased aircraft (AED mn)', FLC['ac_rou'], NUM0),
    ('Loan drawn per owned aircraft (AED mn)', FLC['loan_per_owned'], NUM0),
    ('Dividend floor (AED mn = 30 fils on the full share count)', 1400.0, NUM0),
    ('Minority interests at carrying value (AED mn, audited)', NCIB, NUM2),
    ('Scenario: passenger multiplier', 1.0, NUM2),
    ('Scenario: fare multiplier', 1.0, NUM2),
    ('Scenario: fuel multiplier', 1.0, NUM2),
    ('Scenario: capital-expenditure multiplier', 1.0, NUM2),
    ('Scenario: cost-of-capital shift', 0.0, PCT2),
    ('Scenario: high-fuel switch (0 = base path, 1 = alternative)', 0.0, NUM2),
    ('Scenario: beta benchmark switch (0 = adopted index, 1 = alternative index)', 0.0, NUM2),
    ('Cash administrative costs, FY2025 (AED mn, audited less its depreciation)', GA_CASH_25, NUM1),
    ('Cargo revenue, FY2025 (AED mn, disclosed)', 186.948, NUM1),
    ('Service revenue, FY2025 (AED mn, disclosed)', 223.323, NUM1),
    ('Hotel revenue, FY2025 (AED mn, disclosed)', 59.842, NUM1),
    ('Aircraft-lease revenue, FY2025 (AED mn, disclosed)', 219.755, NUM1),
    ('Other income, FY2025 (AED mn, disclosed)', IN['other_fy25'], NUM1),
    ('Share of JV and associate profit, FY2025 (AED mn, disclosed)', IN['assoc_fy25'], NUM1),
    ('Working capital, FY2025 (AED mn, audited balance sheet)', nwc25, NUM1),
    ('Gross debt, FY2025 (AED mn, audited)', debt25, NUM1),
    ('Cash and fixed deposits, FY2025 (AED mn, audited)', liq25, NUM1),
    ('Fleet assets, FY2025 (PP&E + right-of-use + aircraft advances, AED mn, audited)', fleet25, NUM1),
    ('Intangible assets (AED mn, audited)', intang25, NUM1),
    ('JV and associates at carrying value (AED mn, audited)', jv_book, NUM1),
    ('Non-operating assets (investments + investment property + net investment in lease, AED mn)', nonop25, NUM1),
    ('Minority share of profit', NCI_SH, '0.000%'),
    ('Equity attributable to owners, FY2025 (AED mn, audited)', IN['eqp_fy25'], NUM1),
    ('Weight — discounted cash flow', IN['lens_weights']['dcf'], PCT),
    ('Weight — relative', IN['lens_weights']['relative'], PCT),
    ('Weight — normalised', IN['lens_weights']['normalized'], PCT),
    ('Weight — book', IN['lens_weights']['book'], PCT),
    ('Relative lens — bear multiple', 5.0, MULT),
    ('Relative lens — bull multiple', 8.0, MULT),
    ('Normalised lens — bear P/E', 10.0, MULT),
    ('Normalised lens — bull P/E', 16.0, MULT),
]
for label, v, fmt in SCALARS:
    put(wsA, f'A{r}', label, fmt=None)
    put(wsA, f'C{r}', v, BLUE, fmt)
    AROW[label] = r
    r += 1

def AR(label, col='C'):
    return f"Assumptions!${col}${AROW[label]}"
def ARP(label, j):     # path driver, year j
    return f"Assumptions!{CD[j]}{AROW[label]}"

# ============ 6 SEGMENTS (unit build — LIVE) ==================================
wsS = sheet('Segments')
title(wsS, 'Segments — the unit build, live', 'Revenue = passengers x rate; costs = passengers x unit cost. All formula.',
      8, awidth=52, cwidth=12)
hdr(wsS, 4, ['Revenue build (AED mn)'] + YF)
SR = {}
def srow(r_, label, mk_formula, expects, fmt=NUM0, green=False, bold=False):
    put(wsS, f'A{r_}', label, fmt=None, bold=bold)
    for j, cc in enumerate(CD):
        putf(wsS, f'{cc}{r_}', mk_formula(j, cc), expects[j], fmt, green=green, bold=bold)
    SR[label] = r_

srow(5, 'Passengers (millions)',
     lambda j, cc: f"={ARP('Passengers (millions)', j)}*{AR('Scenario: passenger multiplier')}",
     F['pax'], NUM2)
srow(6, 'Revenue per passenger — fare + baggage (AED)',
     lambda j, cc: f"={ARP('Passenger + baggage revenue per passenger (AED)', j)}"
                   f"*{AR('Scenario: fare multiplier')}",
     IN['fare_path'], NUM1)
srow(7, 'Passenger and baggage revenue',
     lambda j, cc: f"={cc}5*{cc}6", [F['seg_rev'][j]['pax'] for j in range(5)])
srow(8, 'Ancillary rate (AED per passenger)',
     lambda j, cc: f"={ARP('Ancillary revenue per passenger (AED)', j)}",
     IN['anc_path'], NUM1, green=True)
srow(9, 'Ancillary revenue', lambda j, cc: f"={cc}5*{cc}8",
     [F['seg_rev'][j]['ancillary'] for j in range(5)])
srow(10, 'Cargo revenue',
     lambda j, cc: (f"={AR('Cargo revenue, FY2025 (AED mn, disclosed)')}*(1+{ARP('Cargo revenue growth', j)})"
                    if j == 0 else f"={CD[j-1]}10*(1+{ARP('Cargo revenue growth', j)})"),
     cargo_e, NUM1)
srow(11, 'Service revenue',
     lambda j, cc: (f"={AR('Service revenue, FY2025 (AED mn, disclosed)')}*(1+{ARP('Service revenue growth', j)})"
                    if j == 0 else f"={CD[j-1]}11*(1+{ARP('Service revenue growth', j)})"),
     svc_e, NUM1)
srow(12, 'Hotel revenue',
     lambda j, cc: (f"={AR('Hotel revenue, FY2025 (AED mn, disclosed)')}*(1+{ARP('Hotel revenue growth', j)})"
                    if j == 0 else f"={CD[j-1]}12*(1+{ARP('Hotel revenue growth', j)})"),
     hotel_e, NUM1)
srow(13, 'Aircraft-lease revenue (JV network)',
     lambda j, cc: (f"={AR('Aircraft-lease revenue, FY2025 (AED mn, disclosed)')}*(1+{ARP('Aircraft-lease revenue growth', j)})"
                    if j == 0 else f"={CD[j-1]}13*(1+{ARP('Aircraft-lease revenue growth', j)})"),
     lease_e, NUM1)
srow(14, 'Total revenue', lambda j, cc: f"={cc}7+{cc}9+{cc}10+{cc}11+{cc}12+{cc}13",
     F['rev'], NUM0, bold=True)
hdr(wsS, 16, ['Cost build (AED per passenger)'] + YF)
srow(17, 'Fuel cost per passenger (intensity x effective jet price)',
     lambda j, cc: (f"={AR('Fuel intensity (AED per passenger per USD/bbl of effective jet price)')}"
                    f"*({ARP('Effective jet fuel price (USD/bbl) — base path', j)}"
                    f"*(1-{AR('Scenario: high-fuel switch (0 = base path, 1 = alternative)')})"
                    f"+{ARP('Effective jet fuel price (USD/bbl) — high-fuel alternative', j)}"
                    f"*{AR('Scenario: high-fuel switch (0 = base path, 1 = alternative)')})"
                    f"*{AR('Scenario: fuel multiplier')}"),
     FUEL_BASE, NUM1)
srow(18, 'Staff', lambda j, cc: f"={ARP('Staff cost per passenger (AED)', j)}",
     IN['staff_per_pax'], NUM1, green=True)
srow(19, 'Maintenance', lambda j, cc: f"={ARP('Maintenance cost per passenger (AED)', j)}",
     IN['maint_per_pax'], NUM1, green=True)
srow(20, 'Landing & overflying', lambda j, cc: f"={ARP('Landing & overflying charges per passenger (AED)', j)}",
     IN['landing_per_pax'], NUM1, green=True)
srow(21, 'Handling', lambda j, cc: f"={ARP('Handling charges per passenger (AED)', j)}",
     IN['handling_per_pax'], NUM1, green=True)
srow(22, 'Other direct', lambda j, cc: f"={ARP('Other direct cost per passenger (AED)', j)}",
     IN['other_per_pax'], NUM1, green=True)
srow(23, 'Cash cost per passenger', lambda j, cc: f"=SUM({cc}17:{cc}22)",
     F['cash_cost_pax'], NUM1, bold=True)
srow(24, 'Direct cash operating costs (AED mn)', lambda j, cc: f"={cc}5*{cc}23",
     F['dcost_cash'], NUM0)
srow(25, 'Administrative costs (cash)',
     lambda j, cc: (f"={AR('Cash administrative costs, FY2025 (AED mn, audited less its depreciation)')}*(1+{ARP('Administrative cost growth', j)})"
                    if j == 0 else f"={CD[j-1]}25*(1+{ARP('Administrative cost growth', j)})"),
     ga_e, NUM1)
srow(26, 'Selling & marketing', lambda j, cc: f"={AR('Selling & marketing / revenue')}*{cc}14",
     sm_e, NUM1)
srow(27, 'EBITDA', lambda j, cc: f"={cc}14-{cc}24-{cc}25-{cc}26", F['ebitda'], NUM0, bold=True)
srow(28, 'Other income (management fees from the JV network and misc.)',
     lambda j, cc: (f"={AR('Other income, FY2025 (AED mn, disclosed)')}*(1+{ARP('Other-income growth', j)})"
                    if j == 0 else f"={CD[j-1]}28*(1+{ARP('Other-income growth', j)})"),
     oth_e, NUM1)
srow(29, 'EBITDA including fees and other income', lambda j, cc: f"={cc}27+{cc}28",
     F['ebitda_incl'], NUM0, bold=True)
srow(30, 'EBITDA margin', lambda j, cc: f"={cc}27/{cc}14", F['ebitda_margin'], PCT)
ANCH['seg_rev_row'] = 14; ANCH['seg_ebitda_row'] = 27; ANCH['seg_margin_row'] = 30

# history block (pasted, audited/disclosed)
hdr(wsS, 33, ['Disclosed history', 'FY2023', 'FY2024', 'FY2025'])
HISTU = [
    ('Passengers (millions)', [BU['pax_hist']['FY23'], BU['pax_hist']['FY24'], BU['pax_hist']['FY25']], NUM2),
    ('Seat load factor', [BU['lf_hist']['FY23'], BU['lf_hist']['FY24'], BU['lf_hist']['FY25']], PCT),
    ('Passenger + baggage revenue (AED mn)', [None, BU['rev_lines_fy24']['pax'] + BU['rev_lines_fy24']['baggage'],
                                              BU['rev_lines_fy25']['pax'] + BU['rev_lines_fy25']['baggage']], NUM1),
    ('Ancillary revenue (AED mn)', [None, BU['rev_lines_fy24']['ancillary'], BU['rev_lines_fy25']['ancillary']], NUM1),
    ('Fuel cost (AED mn)', [BU['dcost_lines_fy23']['fuel'], BU['dcost_lines_fy24']['fuel'], BU['dcost_lines_fy25']['fuel']], NUM1),
    ('Staff cost, direct (AED mn)', [BU['dcost_lines_fy23']['staff'], BU['dcost_lines_fy24']['staff'], BU['dcost_lines_fy25']['staff']], NUM1),
    ('Maintenance (AED mn)', [BU['dcost_lines_fy23']['maint'], BU['dcost_lines_fy24']['maint'], BU['dcost_lines_fy25']['maint']], NUM1),
    ('Landing & overflying (AED mn)', [BU['dcost_lines_fy23']['landing'], BU['dcost_lines_fy24']['landing'], BU['dcost_lines_fy25']['landing']], NUM1),
    ('Handling (AED mn)', [BU['dcost_lines_fy23']['handling'], BU['dcost_lines_fy24']['handling'], BU['dcost_lines_fy25']['handling']], NUM1),
]
r = 34
for label, vals, fmt in HISTU:
    put(wsS, f'A{r}', label, fmt=None)
    for j, cc in enumerate(['B', 'C', 'D']):
        if vals[j] is not None:
            put(wsS, f'{cc}{r}', vals[j], BLUE, fmt)
        else:
            put(wsS, f'{cc}{r}', 'n/d', BLACK, None)
    r += 1
put(wsS, f'A{r+1}', 'FY2023 revenue lines are on the pre-restatement basis (commissions netted); '
    'seat/ASK data is not disclosed by the company — passengers x rate is the finest disclosed level.',
    fmt=None, wrap=True)

# ============ 8 DCF ===========================================================
wsD = sheet('DCF')
title(wsD, 'DCF — cost of capital, glide, waterfall, terminal', 'Everything here is built, not pasted.',
      8, awidth=52, cwidth=12)
put(wsD, 'A4', 'Cost of capital — explicit window', bold=True, fmt=None)
putf(wsD, 'C5', f"={AR('AED government bond yield (January-2031 tranche)')}", IN['rf'], PCT2, green=True)
put(wsD, 'A5', 'AED government bond yield', fmt=None)
putf(wsD, 'C6', f"={AR('Observed sovereign spread at the auction (netted out)')}", IN['sov_spread_obs'], PCT2, green=True)
put(wsD, 'A6', 'less the observed auction spread over US Treasuries', fmt=None)
putf(wsD, 'C7', '=C5-C6', W['rf_star'], PCT2)
put(wsD, 'A7', 'Net risk-free rate', fmt=None)
_BSW = AR('Scenario: beta benchmark switch (0 = adopted index, 1 = alternative index)')
putf(wsD, 'C8',
     f"={AR('Beta — adopted (own stock, five-year weekly vs the Dubai market index)')}"
     f"*(1-{_BSW})"
     f"+{AR('Beta — alternative benchmark (same regression vs the Abu Dhabi market index)')}"
     f"*{_BSW}",
     IN['beta_used'], NUM2)
put(wsD, 'A8', 'Beta (switchable: adopted index / alternative benchmark)', fmt=None)
putf(wsD, 'C9', f"={AR('Equity risk premium')}", IN['erp_rating'], PCT2, green=True)
put(wsD, 'A9', 'Equity risk premium', fmt=None)
putf(wsD, 'C10', '=C7+C8*C9', W['ke_exp'], PCT2, bold=True)
put(wsD, 'A10', 'Cost of equity', bold=True, fmt=None)
putf(wsD, 'C11', f"={AR('Marginal cost of debt')}", IN['kd'], PCT2, green=True)
put(wsD, 'A11', 'Marginal cost of debt', fmt=None)
putf(wsD, 'C12', f"={AR('Tax rate')}", TAX, PCT, green=True)
put(wsD, 'A12', 'Tax rate', fmt=None)
putf(wsD, 'C13', '=C11*(1-C12)', W['kd_at'], PCT2)
put(wsD, 'A13', 'Cost of debt after tax', fmt=None)
putf(wsD, 'C14', f"={AR('Spot price (AED)')}*{AR('Shares outstanding (mn)')}", M['mktcap'], NUM0)
put(wsD, 'A14', 'Market capitalisation (AED mn)', fmt=None)
putf(wsD, 'C15', f"={AR('Gross debt, FY2025 (AED mn, audited)')}", debt25, NUM1, green=True)
put(wsD, 'A15', 'Gross debt (audited)', fmt=None)
putf(wsD, 'C16', '=C15/(C15+C14)', W['wd_exp'], PCT2)
put(wsD, 'A16', 'Debt weight (gross)', fmt=None)
putf(wsD, 'C17', f"=(1-C16)*C10+C16*C13+{AR('Scenario: cost-of-capital shift')}", W['wacc_exp'], PCT2, bold=True)
put(wsD, 'A17', 'Cost of capital — explicit window', bold=True, fmt=None)
put(wsD, 'A19', 'Cost of capital — terminal', bold=True, fmt=None)
putf(wsD, 'C20', f"={AR('Terminal risk-free rate')}", IN['rf_term'], PCT2, green=True)
put(wsD, 'A20', 'Terminal risk-free rate', fmt=None)
putf(wsD, 'C21', f"=C20+C8*{AR('Terminal equity risk premium')}",
     W['ke_term'], PCT2)
put(wsD, 'A21', 'Terminal cost of equity', fmt=None)
putf(wsD, 'C22', f"={AR('Terminal cost of debt')}*(1-C12)", W['kd_term_at'], PCT2)
put(wsD, 'A22', 'Terminal cost of debt after tax', fmt=None)
putf(wsD, 'C23', f"={AR('Terminal debt weight')}", IN['wd_term'], PCT, green=True)
put(wsD, 'A23', 'Terminal debt weight', fmt=None)
putf(wsD, 'C24', f"=(1-C23)*C21+C23*C22+{AR('Scenario: cost-of-capital shift')}", W['wacc_term'], PCT2, bold=True)
put(wsD, 'A24', 'Cost of capital — terminal', bold=True, fmt=None)
hdr(wsD, 26, ['Glide & discounting'] + YF)
put(wsD, 'A27', 'Cost-of-debt path', fmt=None)
put(wsD, 'A28', 'Glide fraction (path progress)', fmt=None)
put(wsD, 'A29', 'Forward cost of capital', fmt=None)
put(wsD, 'A30', 'Cumulative discount factor', fmt=None)
kd1 = ARP('Marginal cost of debt path', 0); kd5 = ARP('Marginal cost of debt path', 4)
for j, cc in enumerate(CD):
    putf(wsD, f'{cc}27', f"={ARP('Marginal cost of debt path', j)}", IN['kd_path'][j], PCT2, green=True)
    putf(wsD, f'{cc}28', f"=({kd1.replace('!B', '!$B').replace('!F','!$F')}-{cc}27)/({kd1}-{kd5})",
         F['glide_frac'][j], DF4)
    putf(wsD, f'{cc}29', f"=$C$17-($C$17-$C$24)*{cc}28", F['fwd_wacc'][j], PCT2)
    if j == 0:
        putf(wsD, f'{cc}30', f"=1/(1+{cc}29)", F['df'][j], DF4)
    else:
        putf(wsD, f'{cc}30', f"={CD[j-1]}30/(1+{cc}29)", F['df'][j], DF4)
hdr(wsD, 32, ['FCFF waterfall (AED mn)'] + YF)
WROWS = [
    ('Revenue', 33, lambda j, cc: f"=Segments!{cc}14", F['rev'], NUM0, True),
    ('EBITDA incl. fees and other income', 34, lambda j, cc: f"=Segments!{cc}29", F['ebitda_incl'], NUM0, True),
    ('less depreciation & amortisation', 35, lambda j, cc: f"={ARP('Depreciation & amortisation (AED mn)', j)}",
     F['dna'], NUM0, True),
    ('EBIT', 36, lambda j, cc: f"={cc}34-{cc}35", F['ebit_incl'], NUM0, False),
    ('NOPAT (EBIT x (1 - tax))', 37, lambda j, cc: f"={cc}36*(1-$C$12)", F['nopat'], NUM0, False),
    ('add back depreciation & amortisation', 38, lambda j, cc: f"={cc}35", F['dna'], NUM0, False),
    ('less owned capex + pre-delivery payments', 39,
     lambda j, cc: f"={ARP('Fleet capital expenditure incl. pre-delivery payments (AED mn)', j)}"
                   f"*{AR('Scenario: capital-expenditure multiplier')}",
     F['capex'], NUM0, False),
    ('Working capital balance', 40, lambda j, cc: f"={AR('Working capital / revenue')}*{cc}33", F['nwc'], NUM0, False),
    ('less change in working capital', 41,
     lambda j, cc: (f"={cc}40-{AR('Working capital, FY2025 (AED mn, audited balance sheet)')}" if j == 0
                    else f"={cc}40-{CD[j-1]}40"), F['dnwc'], NUM0, False),
    ('Free cash flow to the firm', 42, lambda j, cc: f"={cc}37+{cc}38-{cc}39-{cc}41-{cc}45", F['fcff'], NUM0, False),
    ('Discount factor', 43, lambda j, cc: f"={cc}30", F['df'], DF4, False),
    ('PV of FCFF', 44, lambda j, cc: f"={cc}42*{cc}43", F['pv'], NUM0, False),
]
for label, rr, mk, exps, fmt, green in WROWS:
    put(wsD, f'A{rr}', label, fmt=None, bold=(rr in (33, 34, 42, 44)))
    for j, cc in enumerate(CD):
        putf(wsD, f'{cc}{rr}', mk(j, cc), exps[j], fmt, green=green,
             bold=(rr in (42, 44)))
put(wsD, 'A45', 'less leased-fleet additions (gross right-of-use value)', fmt=None)
for j, cc in enumerate(CD):
    putf(wsD, f'{cc}45',
         f"={ARP('Leased aircraft additions (units)', j)}"
         f"*{AR('Right-of-use value per leased aircraft (AED mn)')}"
         f"*{AR('Scenario: capital-expenditure multiplier')}",
         F['leased_gross'][j], NUM0)
put(wsD, 'A46', 'Fleet assets roll-forward', fmt=None)
for j, cc in enumerate(CD):
    if j == 0:
        fm = (f"={AR('Fleet assets, FY2025 (PP&E + right-of-use + aircraft advances, AED mn, audited)')}"
              f"+{cc}39+{cc}45-{cc}35")
    else:
        fm = f"={CD[j-1]}46+{cc}39+{cc}45-{cc}35"
    putf(wsD, f'{cc}46', fm, F['ppe'][j], NUM0)
put(wsD, 'A47', 'Invested capital (fleet + intangibles + working capital)', fmt=None)
for j, cc in enumerate(CD):
    putf(wsD, f'{cc}47', f"={cc}46+{AR('Intangible assets (AED mn, audited)')}+{cc}40", F['ic'][j], NUM0)
put(wsD, 'A49', 'Terminal block', bold=True, fmt=None)
putf(wsD, 'C50', f"=F37*(1+{AR('Terminal growth')})", F['nopat'][4] * (1 + IN['g_term']), NUM1)
put(wsD, 'A50', 'Terminal NOPAT (FY2031E)', fmt=None)
putf(wsD, 'C51', '=C50/F47', DCF['roic_term'], PCT2)
put(wsD, 'A51', 'Terminal return on invested capital', fmt=None)
putf(wsD, 'C52', f"={AR('Terminal growth')}/C51", DCF['rr_term'], PCT2)
put(wsD, 'A52', 'Reinvestment rate = growth / return on capital', fmt=None)
putf(wsD, 'C53', f"=C50*(1-C52)/(C24-{AR('Terminal growth')})", DCF['tv'], NUM0)
put(wsD, 'A53', 'Terminal value', fmt=None)
putf(wsD, 'C54', '=C53*F30', DCF['pv_tv'], NUM0)
put(wsD, 'A54', 'PV of terminal value', fmt=None)
putf(wsD, 'C55', '=SUM(B44:F44)', DCF['pv_explicit'], NUM0)
put(wsD, 'A55', 'PV of explicit years', fmt=None)
putf(wsD, 'C56', '=C54+C55', DCF['ev'], NUM0, bold=True)
put(wsD, 'A56', 'Enterprise value of the airline', bold=True, fmt=None)
putf(wsD, 'C57', '=C54/C56', DCF['tv_share'], PCT, bold=True)
put(wsD, 'A57', 'Terminal value share of enterprise value', bold=True, fmt=None)
putf(wsD, 'C59', "='SOTP Bridge'!C13", DCF['eq_attr'], NUM0, green=True)
put(wsD, 'A59', 'Equity attributable (from the SOTP Bridge)', fmt=None)
putf(wsD, 'C60', f"=C59/{AR('Shares outstanding (mn)')}", DCF['ps_dec'], PX)
put(wsD, 'A60', 'Per share at 31-Dec-2025', fmt=None)
putf(wsD, 'C61', f"=(1+C10)^({AR('Days from 31-Dec-2025 to the 7-Aug-2026 anchor')}/365)",
     DCF['roll'], DF4)
put(wsD, 'A61', 'Anchor accretion factor (at the cost of equity)', fmt=None)
putf(wsD, 'C62', f"=C60*C61-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     DCF['ps_dec'] * DCF['roll'] - IN['dps_fy25'], PX)
put(wsD, 'A62', 'Per share, whole equity rolled at Ke (single-rate view)', fmt=None)
putf(wsD, 'C63', f"=(1+{ARP('Deposit yield path', 0)})^({AR('Days from 31-Dec-2025 to the 7-Aug-2026 anchor')}/365)",
     ROLLC, DF4)
put(wsD, 'A63', 'Cash-leg accretion factor (deposit yield)', fmt=None)
putf(wsD, 'C64', "='SOTP Bridge'!C15", DCF['ps'], PX, bold=True, green=True)
put(wsD, 'A64', 'Fair value per share at the anchor (split roll — the published figure)', bold=True, fmt=None)

# ============ 5 SOTP BRIDGE ====================================================
wsB = sheet('SOTP Bridge')
title(wsB, 'SOTP / EV-to-equity bridge — both JV framings', 'The joint-venture network is the contested judgement: both framings shown, never averaged.',
      7, awidth=56, cwidth=14)
hdr(wsB, 4, ['Bridge — base framing (JV at audited carrying value)', '', 'AED mn'])
putf(wsB, 'C5', '=DCF!C56', DCF['ev'], NUM0, green=True)
put(wsB, 'A5', 'Enterprise value of the airline (DCF)', fmt=None)
putf(wsB, 'C6', f"={AR('Cash and fixed deposits, FY2025 (AED mn, audited)')}-{AR('Gross debt, FY2025 (AED mn, audited)')}",
     -nd25, NUM0)
put(wsB, 'A6', 'plus net cash (cash + deposits - borrowings - leases)', fmt=None)
putf(wsB, 'C7', f"={AR('Non-operating assets (investments + investment property + net investment in lease, AED mn)')}",
     nonop25, NUM0, green=True)
put(wsB, 'A7', 'plus non-operating assets', fmt=None)
putf(wsB, 'C8', f"={AR('JV and associates at carrying value (AED mn, audited)')}", jv_book, NUM0, green=True)
put(wsB, 'A8', 'plus JV network at audited carrying value', fmt=None)
putf(wsB, 'C9', '=SUM(C5:C8)', DCF['eq_attr'] + DCF['nci_val'], NUM0, bold=True)
put(wsB, 'A9', 'Equity before minorities', bold=True, fmt=None)
putf(wsB, 'C11', f"={AR('Minority interests at carrying value (AED mn, audited)')}", NCIB, NUM2, green=True)
put(wsB, 'A11', 'less minorities at audited carrying value', fmt=None)
putf(wsB, 'C13', '=C9-C11', DCF['eq_attr'], NUM0, bold=True)
put(wsB, 'A13', 'Equity attributable to shareholders', bold=True, fmt=None)
putf(wsB, 'C14', f"=C13/{AR('Shares outstanding (mn)')}", DCF['ps_dec'], PX)
put(wsB, 'A14', 'Per share at 31-Dec-2025', fmt=None)
putf(wsB, 'C15', f"=(C5*DCF!C61+(C6+C7+C8)*DCF!C63-C11)/{AR('Shares outstanding (mn)')}"
     f"-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     DCF['ps'], PX, bold=True)
put(wsB, 'A15', 'Per share at the anchor — operating equity rolled at the cost of equity, '
    'cash and near-cash legs at the deposit yield', bold=True, fmt=None)
hdr(wsB, 17, ['Alternative framing (JV capitalised at a growth multiple)', '', 'AED mn'])
putf(wsB, 'C18', f"={AR('Joint-venture capitalisation multiple')}*{AR('Share of JV and associate profit, FY2025 (AED mn, disclosed)')}",
     jv_cap, NUM0)
put(wsB, 'A18', 'JV network capitalised (multiple x FY2025 profit share)', fmt=None)
putf(wsB, 'C19', '=C5+C6+C7+C18-C11', DCF['ev'] - nd25 + nonop25 + jv_cap - NCIB, NUM0)
put(wsB, 'A19', 'Equity attributable on this framing', fmt=None)
putf(wsB, 'C20', f"=(C5*DCF!C61+(C6+C7+C18)*DCF!C63-C11)/{AR('Shares outstanding (mn)')}"
     f"-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     DCF['ps_jvcap'], PX, bold=True)
put(wsB, 'A20', 'Per share at the anchor — JV capitalised', bold=True, fmt=None)
hdr(wsB, 22, ['Non-operating assets, itemised (audited FY2025)', '', 'AED mn'])
for i, (lbl, v) in enumerate([
        ('Investments at fair value through OCI', HB['FY25']['fvoci']),
        ('Investment property (book; disclosed fair value AED 334mn)', HB['FY25']['invprop']),
        ('Net investment in lease (aircraft subleased to the JV airlines)', HB['FY25']['nil'])]):
    put(wsB, f'A{23+i}', lbl, fmt=None)
    put(wsB, f'C{23+i}', v, BLUE, NUM1)
putf(wsB, 'C26', '=SUM(C23:C25)', nonop25, NUM1, bold=True)
put(wsB, 'A26', 'Total non-operating assets', bold=True, fmt=None)
put(wsB, 'A28', 'The JV network: Air Arabia Abu Dhabi (49%), Air Arabia Egypt (49%, raised from 40% in 2025), '
    'Fly Jinnah (45%), Air Arabia Maroc (44.13%), Air Arabia DMM — Saudi Arabia (49%, pre-operational). '
    'FY2025 profit share AED 190.0mn; audited carrying value AED 363.4mn.', fmt=None, wrap=True)

# ============ 3 FUNDAMENTAL VALUATION =========================================
wsF = sheet('Fundamental Valuation')
title(wsF, 'Fundamental valuation — the four lenses and the alternatives', None, 6, awidth=56, cwidth=14)
hdr(wsF, 4, ['Lens', 'Basis', 'AED per share'])
putf(wsF, 'C5', "='SOTP Bridge'!C15", DCF['ps'], PX, green=True, bold=True)
put(wsF, 'A5', 'Discounted cash flow (primary)', fmt=None)
put(wsF, 'B5', 'FCFF, glide-discounted, JV at carrying value', fmt=None)
putf(wsF, 'C6', "='Relative & Normalized'!C11", LN['relative']['base'], PX, green=True)
put(wsF, 'A6', 'Relative multiples', fmt=None)
put(wsF, 'B6', f"{IN['ev_ebitda_just']:.1f}x FY2027E EBITDA, discounted", fmt=None)
putf(wsF, 'C7', "='Relative & Normalized'!C29", LN['normalized']['base'], PX, green=True)
put(wsF, 'A7', 'Normalised earnings power', fmt=None)
put(wsF, 'B7', f"{IN['pe_just']:.0f}x mid-cycle EPS at current scale", fmt=None)
putf(wsF, 'C8', "='Relative & Normalized'!C37", LN['book']['base'], PX, green=True)
put(wsF, 'A8', 'Book value and sustainable return', fmt=None)
put(wsF, 'B8', 'Justified price-to-book on FY2025 equity', fmt=None)
band(wsF, 10, 6)
putf(wsF, 'C10', "=Summary!C9", D['central'], PX, bold=True)
put(wsF, 'A10', 'Weighted central', bold=True, fmt=None)
hdr(wsF, 12, ['Alternative readings (whole-model re-runs unless linked)', '', 'AED per share'])
putf(wsF, 'C13', "='SOTP Bridge'!C20", DCF['ps_jvcap'], PX, green=True)
put(wsF, 'A13', 'DCF with the JV network capitalised (the contested judgement, other framing)', fmt=None)
put(wsF, 'C14', DCF['ps_iata_fuel'], BLUE, PX)
put(wsF, 'A14', 'DCF on the high-fuel alternative (airline-association path held)', fmt=None)
put(wsF, 'C15', DCF['bear'], BLUE, PX)
put(wsF, 'A15', 'DCF bear scenario (high fuel + weaker traffic + tighter money)', fmt=None)
put(wsF, 'C16', DCF['bull'], BLUE, PX)
put(wsF, 'A16', 'DCF bull scenario (fuel relief + stronger traffic + JV capitalised)', fmt=None)
hdr(wsF, 18, ['Expert panel (each leg is a complete model of its own)', '', 'AED per share'])
DPSR = AR('FY2025 dividend per share (AED, approved 12 March 2026)')
SHR = AR('Shares outstanding (mn)')
NCIR = AR('Minority share of profit')
TAXR = AR('Tax rate')
# Expert 1 — LIVE: FY2028E earnings power x justified multiple, rolled
E1_EPS = (f"(('Segments'!D30*'Segments'!D14+'Segments'!D28-{ARP('Depreciation & amortisation (AED mn)', 2)}"
          f"+('Cash Flow'!D13-'Cash Flow'!D14)+'Income Statement'!G15)*(1-{TAXR})*(1-{NCIR})/{SHR})")
put(wsF, 'A19', 'Expert 1 — earnings power at a justified multiple (live formula)', fmt=None)
putf(wsF, 'C19', f"=13*{E1_EPS}*DCF!C61-{DPSR}", EXP['e1']['base'], PX)
putf(wsF, 'D19', f"=10*{E1_EPS}*DCF!C61-{DPSR}", EXP['e1']['rng'][0], PX)
putf(wsF, 'E19', f"=16*{E1_EPS}*DCF!C61-{DPSR}", EXP['e1']['rng'][1], PX)
# Expert 2 — LIVE: owner cash earnings capitalised at the terminal rate, half net cash
E2_FCFE = (f"((AVERAGE('Cash Flow'!D9:F9)+('Cash Flow'!E13-'Cash Flow'!E14)*(1-{TAXR})"
           f"+'Income Statement'!H15*(1-{TAXR})*0.4)*(1-{NCIR}))")
E2_HALFCASH = (f"0.5*({AR('Cash and fixed deposits, FY2025 (AED mn, audited)')}-{AR('Gross debt, FY2025 (AED mn, audited)')})")
put(wsF, 'A20', 'Expert 2 — owner cash earnings capitalised (live formula)', fmt=None)
putf(wsF, 'C20', f"=({E2_FCFE}*(1+{AR('Terminal growth')})/(DCF!C21-{AR('Terminal growth')})+{E2_HALFCASH})/{SHR}*DCF!C61-{DPSR}",
     EXP['e2']['base'], PX)
putf(wsF, 'D20', f"={E2_FCFE}*1.015/(0.5*(DCF!C10+DCF!C21)-0.015)/{SHR}*DCF!C61-{DPSR}",
     EXP['e2']['rng'][0], PX)
putf(wsF, 'E20', f"=({E2_FCFE}*1.035/(DCF!C21-0.035)+2*{E2_HALFCASH})/{SHR}*DCF!C61-{DPSR}",
     EXP['e2']['rng'][1], PX)
# Expert 3 — economic-profit legs stay whole-model outputs (their PV chain lives in the
# study's Appendix C table line by line); pasted, named as such
put(wsF, 'A21', 'Expert 3 — cash returns vs the cost of capital (whole-model re-run; full chain in Appendix C)', fmt=None)
put(wsF, 'C21', EXP['e3']['base'], BLUE, PX)
put(wsF, 'D21', EXP['e3']['rng'][0], BLUE, PX)
put(wsF, 'E21', EXP['e3']['rng'][1], BLUE, PX)
putf(wsF, 'C23', '=MEDIAN(C19,C20,C21)', D['panel_centre'], PX, bold=True)
put(wsF, 'A23', 'Panel median', bold=True, fmt=None)
put(wsF, 'A25', 'Columns D and E carry each expert\'s own low and high. The spot price on the anchor date is '
    f'AED {SPOT:.2f}.', fmt=None, wrap=True)

# ============ 2 SUMMARY =======================================================
wsU = sheet('Summary')
title(wsU, 'Summary — valuation at a glance', 'All values link live to their source sheets', 7,
      awidth=48, cwidth=14)
hdr(wsU, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Weight', 'Contribution', 'vs spot'])
r = 5
LSRC = {'dcf': "='Fundamental Valuation'!C5", 'relative': "='Relative & Normalized'!C11",
        'normalized': "='Relative & Normalized'!C29", 'book': "='Relative & Normalized'!C37"}
BSRC = {'relative': "='Relative & Normalized'!C12", 'normalized': "='Relative & Normalized'!C31",
        'book': "='Relative & Normalized'!C39"}
USRC = {'relative': "='Relative & Normalized'!C13", 'normalized': "='Relative & Normalized'!C32",
        'book': "='Relative & Normalized'!C40"}
WLBL = {'dcf': 'Weight — discounted cash flow', 'relative': 'Weight — relative',
        'normalized': 'Weight — normalised', 'book': 'Weight — book'}
for k in ['dcf', 'relative', 'normalized', 'book']:
    l = LN[k]
    put(wsU, f'A{r}', l['name'], fmt=None)
    if k in BSRC:
        putf(wsU, f'B{r}', BSRC[k], l['bear'], PX, green=True)
    else:
        put(wsU, f'B{r}', l['bear'], BLUE, PX)
    putf(wsU, f'C{r}', LSRC[k], l['base'], PX, green=True)
    if k in USRC:
        putf(wsU, f'D{r}', USRC[k], l['bull'], PX, green=True)
    else:
        put(wsU, f'D{r}', l['bull'], BLUE, PX)
    putf(wsU, f'E{r}', f"={AR(WLBL[k])}", l['w'], PCT, green=True)
    putf(wsU, f'F{r}', f'=C{r}*E{r}', l['base'] * l['w'], PX)
    putf(wsU, f'G{r}', f'=C{r}/$C$16-1', l['base'] / SPOT - 1, PCT)
    r += 1
band(wsU, 9, 7)
putf(wsU, 'B9', '=B5*E5+B6*E6+B7*E7+B8*E8', LN['central']['bear'], PX, bold=True)
putf(wsU, 'C9', '=SUM(F5:F8)', D['central'], PX, bold=True)
putf(wsU, 'D9', '=D5*E5+D6*E6+D7*E7+D8*E8', LN['central']['bull'], PX, bold=True)
putf(wsU, 'E9', '=SUM(E5:E8)', 1.0, PCT, bold=True)
putf(wsU, 'G9', '=C9/$C$16-1', D['central'] / SPOT - 1, PCT, bold=True)
put(wsU, 'A9', 'Weighted central (range weighted like the base)', bold=True, fmt=None)
putf(wsU, 'B15', '=MIN(B5:B8)', D['span_widest'][0], PX)
putf(wsU, 'D15', '=MAX(D5:D8)', D['span_widest'][1], PX)
put(wsU, 'A15', 'Widest single-lens span (the DCF scenarios) — labelled, not the weighted range', fmt=None)
putf(wsU, 'C10', "='SOTP Bridge'!C20", DCF['ps_jvcap'], PX, green=True)
putf(wsU, 'G10', '=C10/$C$16-1', DCF['ps_jvcap'] / SPOT - 1, PCT)
put(wsU, 'A10', 'DCF — JV network capitalised (contested judgement, other framing)', fmt=None)
putf(wsU, 'C11', '=C9+' + AR('Weight — discounted cash flow') + "*('SOTP Bridge'!C20-'Fundamental Valuation'!C5)",
     D['central_jvcap'], PX)
putf(wsU, 'G11', '=C11/$C$16-1', D['central_jvcap'] / SPOT - 1, PCT)
put(wsU, 'A11', 'Weighted central on that framing', fmt=None)
putf(wsU, 'C12', '=DCF!C57', DCF['tv_share'], PCT, green=True)
put(wsU, 'A12', 'Terminal value share of DCF enterprise value', fmt=None)
putf(wsU, 'C13', "='Fundamental Valuation'!C23", D['panel_centre'], PX, green=True)
putf(wsU, 'G13', '=C13/$C$16-1', D['panel_centre'] / SPOT - 1, PCT)
put(wsU, 'A13', 'Expert panel median', fmt=None)
put(wsU, 'C14', DCF['ps_iata_fuel'], BLUE, PX)
putf(wsU, 'G14', '=C14/$C$16-1', DCF['ps_iata_fuel'] / SPOT - 1, PCT)
put(wsU, 'A14', 'DCF on the high-fuel alternative (whole-model re-run)', fmt=None)
band(wsU, 16, 7)
putf(wsU, 'C16', f"={AR('Spot price (AED)')}", SPOT, PX, bold=True, green=True)
put(wsU, 'A16', 'Market price (7-Aug-2026 close)', bold=True, fmt=None)
hdr(wsU, 18, ['Key figure', '', 'Value'])
KEY = [('Shares outstanding (mn)', f"={AR('Shares outstanding (mn)')}", SH, NUM0),
       ('Market capitalisation (AED mn)', '=DCF!C14', M['mktcap'], NUM0),
       ('Net cash incl. fixed deposits, FY2025 (AED mn)', "='SOTP Bridge'!C6", -nd25, NUM0),
       ('FY2025 revenue (AED mn)', "='Income Statement'!D5", HI['FY25']['rev'], NUM0),
       ('FY2025 EBITDA (AED mn)', "='Income Statement'!D9", HI['FY25']['ebitda'], NUM0),
       ('FY2025 attributable profit (AED mn)', "='Income Statement'!D20", HI['FY25']['npa'], NUM0),
       ('Cost of capital — explicit window', '=DCF!C17', W['wacc_exp'], PCT2),
       ('Cost of capital — terminal', '=DCF!C24', W['wacc_term'], PCT2),
       ('Terminal growth', f"={AR('Terminal growth')}", IN['g_term'], PCT)]
r = 19
for lbl, fm, ex, fmt in KEY:
    put(wsU, f'A{r}', lbl, fmt=None)
    putf(wsU, f'C{r}', fm, ex, fmt, green=True)
    r += 1
ANCH['summary_spot'] = 'C16'

# ============ 9 INCOME STATEMENT ==============================================
wsI = sheet('Income Statement')
title(wsI, 'Income statement — 3 years audited + 5 years forecast',
      'FY2023 as reported; FY2024 as restated in the FY2025 filing; forecast chains from the Segments unit build.',
      10, awidth=44, cwidth=11)
hdr(wsI, 4, [''] + ['FY2023', 'FY2024', 'FY2025'] + YF)
dc23 = BU['dcost_lines_fy23']; dc24 = BU['dcost_lines_fy24']; dc25 = BU['dcost_lines_fy25']
dep23 = dc23['dep_ppe'] + dc23['dep_rou'] + dc23['amort']
dep24 = dc24['dep_ppe'] + dc24['dep_rou'] + dc24['amort']
dep25 = dc25['dep_ppe'] + dc25['dep_rou'] + dc25['amort']
HVAL = {
    5:  ('Revenue', [HI[y]['rev'] for y in H3], NUM0),
    6:  ('Direct operating costs (cash)', [HI['FY23']['dcost'] - dep23, HI['FY24']['dcost'] - dep24,
                                           HI['FY25']['dcost'] - dep25], NUM0),
    7:  ('Administrative expenses (cash)', [HI['FY23']['ga'] - (HI['FY23']['dna'] - dep23),
                                            HI['FY24']['ga'] - (HI['FY24']['dna'] - dep24),
                                            HI['FY25']['ga'] - (HI['FY25']['dna'] - dep25)], NUM1),
    8:  ('Selling & marketing', [HI[y]['sm'] for y in H3], NUM1),
    10: ('Depreciation & amortisation', [HI[y]['dna'] for y in H3], NUM0),
    12: ('Other income', [HI[y]['other'] for y in H3], NUM1),
    13: ('Finance income', [HI[y]['fininc'] for y in H3], NUM1),
    14: ('Finance costs', [HI[y]['fincost'] for y in H3], NUM1),
    15: ('Share of JV and associate profit', [HI[y]['assoc'] for y in H3], NUM1),
    17: ('Income tax', [HI[y]['tax'] for y in H3], NUM1),
    20: ('Attributable to owners', [HI[y]['npa'] for y in H3], NUM0),
}
for rr, (lbl, vals, fmt) in HVAL.items():
    put(wsI, f'A{rr}', lbl, fmt=None)
    for j, cc in enumerate(HC):
        put(wsI, f'{cc}{rr}', vals[j], BLUE, fmt)
put(wsI, 'A9', 'EBITDA', bold=True, fmt=None)
put(wsI, 'A11', 'Operating profit (EBIT)', bold=True, fmt=None)
put(wsI, 'A16', 'Profit before tax (before the JV share, forecast)', bold=True, fmt=None)
put(wsI, 'A18', 'Profit for the year', bold=True, fmt=None)
put(wsI, 'A19', 'Minorities', fmt=None)
put(wsI, 'A21', 'Earnings per share (AED)', fmt=None)
put(wsI, 'A22', 'EBITDA margin', fmt=None)
for j, cc in enumerate(HC):
    y = H3[j]
    putf(wsI, f'{cc}9', f'={cc}5-{cc}6-{cc}7-{cc}8', HI[y]['ebitda'], NUM0, bold=True)
    putf(wsI, f'{cc}11', f'={cc}9-{cc}10', HI[y]['ebit'], NUM0, bold=True)
    putf(wsI, f'{cc}16', f'={cc}11+{cc}12+{cc}13-{cc}14+{cc}15', HI[y]['ebt'], NUM0, bold=True)
    putf(wsI, f'{cc}18', f'={cc}16-{cc}17', HI[y]['pat'], NUM0, bold=True)
    # history keeps the audited presentation (JV share inside profit before tax); the
    # forecast carries it below the tax line — each column is labelled by its basis
    putf(wsI, f'{cc}19', f'={cc}18-{cc}20', HI[y]['nci'], NUM2)
    putf(wsI, f'{cc}21', f"={cc}20/{AR('Shares outstanding (mn)')}", HI[y]['npa'] / SH, PX)
    putf(wsI, f'{cc}22', f'={cc}9/{cc}5', HI[y]['ebitda'] / HI[y]['rev'], PCT)
# forecast columns
np_f = F['np_attr']; pat_f = [n / (1 - NCI_SH) for n in np_f]
pretax_ex_jv = [(pat_f[j] - F['assoc'][j]) / (1 - TAX) for j in range(5)]
for j, cc in enumerate(FCOL):
    sc = CD[j]     # matching column on Segments / DCF / Cash Flow
    putf(wsI, f'{cc}5', f'=Segments!{sc}14', F['rev'][j], NUM0, green=True)
    putf(wsI, f'{cc}6', f'=Segments!{sc}24', F['dcost_cash'][j], NUM0, green=True)
    putf(wsI, f'{cc}7', f'=Segments!{sc}25', ga_e[j], NUM1, green=True)
    putf(wsI, f'{cc}8', f'=Segments!{sc}26', sm_e[j], NUM1, green=True)
    putf(wsI, f'{cc}9', f'={cc}5-{cc}6-{cc}7-{cc}8', F['ebitda'][j], NUM0, bold=True)
    putf(wsI, f'{cc}10', f"={ARP('Depreciation & amortisation (AED mn)', j)}", F['dna'][j], NUM0, green=True)
    putf(wsI, f'{cc}11', f'={cc}9-{cc}10', F['ebitda'][j] - F['dna'][j], NUM0, bold=True)
    putf(wsI, f'{cc}12', f'=Segments!{sc}28', oth_e[j], NUM1, green=True)
    putf(wsI, f'{cc}13', f"='Cash Flow'!{sc}13", F['fininc'][j], NUM1, green=True)
    putf(wsI, f'{cc}14', f"='Cash Flow'!{sc}14", F['interest'][j], NUM1, green=True)
    if j == 0:
        putf(wsI, f'{cc}15', f"={AR('Share of JV and associate profit, FY2025 (AED mn, disclosed)')}"
             f"*(1+{ARP('Growth in the share of joint-venture and associate profit', j)})",
             assoc_e[j], NUM1)
    else:
        putf(wsI, f'{cc}15', f"={FCOL[j-1]}15*(1+{ARP('Growth in the share of joint-venture and associate profit', j)})",
             assoc_e[j], NUM1)
    putf(wsI, f'{cc}16', f'={cc}11+{cc}12+{cc}13-{cc}14', pretax_ex_jv[j], NUM0, bold=True)
    putf(wsI, f'{cc}17', f"={cc}16*{AR('Tax rate')}", pretax_ex_jv[j] * TAX, NUM1)
    putf(wsI, f'{cc}18', f'={cc}16-{cc}17+{cc}15', pat_f[j], NUM0, bold=True)
    putf(wsI, f'{cc}19', f"={cc}18*{AR('Minority share of profit')}", pat_f[j] * NCI_SH, NUM2)
    putf(wsI, f'{cc}20', f'={cc}18-{cc}19', np_f[j], NUM0, bold=True)
    putf(wsI, f'{cc}21', f"={cc}20/{AR('Shares outstanding (mn)')}", np_f[j] / SH, PX)
    putf(wsI, f'{cc}22', f'={cc}9/{cc}5', F['ebitda_margin'][j], PCT)
put(wsI, 'A24', 'The forecast income statement carries EBIT excluding fees/other income on row 11; the DCF '
    'taxes EBIT including that recurring line (row 12), exactly as the study text describes.', fmt=None, wrap=True)

# ============ 11 CASH FLOW =====================================================
wsC = sheet('Cash Flow')
title(wsC, 'Cash flow — FCFF and the financing walk', 'Links to the DCF waterfall; net debt rolls forward.',
      8, awidth=52, cwidth=12)
hdr(wsC, 4, ['FCFF build (AED mn)'] + YF)
CROWS = [
    (5, 'NOPAT', lambda j, cc: f'=DCF!{cc}37', F['nopat'], True),
    (6, 'add depreciation & amortisation', lambda j, cc: f'=DCF!{cc}38', F['dna'], True),
    (7, 'less owned capex + pre-delivery payments', lambda j, cc: f'=DCF!{cc}39', F['capex'], True),
    (8, 'less change in working capital', lambda j, cc: f'=DCF!{cc}41', F['dnwc'], True),
    (10, 'less leased-fleet additions (gross)', lambda j, cc: f'=DCF!{cc}45', F['leased_gross'], True),
]
for rr, lbl, mk, exps, green in CROWS:
    put(wsC, f'A{rr}', lbl, fmt=None)
    for j, cc in enumerate(CD):
        putf(wsC, f'{cc}{rr}', mk(j, cc), exps[j], NUM0, green=green)
put(wsC, 'A9', 'Free cash flow to the firm', bold=True, fmt=None)
for j, cc in enumerate(CD):
    putf(wsC, f'{cc}9', f'={cc}5+{cc}6-{cc}7-{cc}8-{cc}10', F['fcff'][j], NUM0, bold=True)
hdr(wsC, 11, ['Financing walk (AED mn)'] + YF)
put(wsC, 'A12', 'Opening net debt (negative = net cash)', fmt=None)
put(wsC, 'A13', 'Finance income (deposit yield x opening cash and deposits)', fmt=None)
put(wsC, 'A14', 'Finance costs (booked rate x average gross debt)', fmt=None)
put(wsC, 'A15', 'Dividends paid (payout, floored at 30 fils)', fmt=None)
put(wsC, 'A16', 'Closing net debt', fmt=None)
put(wsC, 'A17', 'Opening gross debt (borrowings + leases)', fmt=None)
put(wsC, 'A18', 'Closing gross debt (+ new leases + aircraft loans - amortisation)', fmt=None)
nd_open = [nd25] + F['net_debt'][:-1]
debt_open = [HB['FY25']['debt']] + F['debt'][:-1]
for j, cc in enumerate(CD):
    if j == 0:
        putf(wsC, f'{cc}12', f"={AR('Gross debt, FY2025 (AED mn, audited)')}-{AR('Cash and fixed deposits, FY2025 (AED mn, audited)')}",
             nd25, NUM0)
        putf(wsC, f'{cc}17', f"={AR('Gross debt, FY2025 (AED mn, audited)')}", debt_open[j], NUM0, green=True)
    else:
        putf(wsC, f'{cc}12', f'={CD[j-1]}16', nd_open[j], NUM0)
        putf(wsC, f'{cc}17', f'={CD[j-1]}18', debt_open[j], NUM0)
    putf(wsC, f'{cc}18',
         f"={cc}17+DCF!{cc}45+{ARP('Owned aircraft additions (units)', j)}"
         f"*{AR('Loan drawn per owned aircraft (AED mn)')}-{ARP('Debt amortisation (AED mn)', j)}",
         F['debt'][j], NUM0)
    putf(wsC, f'{cc}13', f"={ARP('Deposit yield path', j)}*({cc}17-{cc}12)",
         F['fininc'][j], NUM1)
    putf(wsC, f'{cc}14', f"={ARP('Booked finance-cost rate on the debt book', j)}*({cc}17+{cc}18)/2",
         F['interest'][j], NUM1)
    putf(wsC, f'{cc}15', f"=MAX({ARP('Dividend payout ratio', j)}*'Income Statement'!{FCOL[j]}20,"
         f"{AR('Dividend floor (AED mn = 30 fils on the full share count)')})",
         F['div'][j], NUM0)
    putf(wsC, f'{cc}16', f"={cc}12-({cc}9+({cc}13-{cc}14)*(1-{AR('Tax rate')}))+{cc}15",
         F['net_debt'][j], NUM0, bold=True)
hdr(wsC, 20, ['Audited history (AED mn)', 'FY2023', 'FY2024', 'FY2025'])
CH = [('Net cash from operating activities', [IN['ocf_fy23'], IN['ocf_fy24'], IN['ocf_fy25']]),
      ('Fleet capex incl. aircraft advances', [IN['capex_fy23'], IN['capex_fy24'], IN['capex_fy25']]),
      ('Dividends paid to owners', [IN['div_fy23'], IN['div_fy24'], IN['div_fy25']]),
      ('Depreciation & amortisation', [HI['FY23']['dna'], HI['FY24']['dna'], HI['FY25']['dna']])]
r = 21
for lbl, vals in CH:
    put(wsC, f'A{r}', lbl, fmt=None)
    for j, cc in enumerate(['B', 'C', 'D']):
        put(wsC, f'{cc}{r}', vals[j], BLUE, NUM0)
    r += 1

# ============ 10 BALANCE SHEET =================================================
wsBS = sheet('Balance Sheet')
title(wsBS, 'Balance sheet — 3 years audited + roll-forward',
      'FY2023 column is the restated 1-Jan-2024 position from the FY2025 filing.',
      10, awidth=46, cwidth=11)
hdr(wsBS, 4, [''] + ['FY2023', 'FY2024', 'FY2025'] + YF)
BROWS = [
    (5, 'Fleet assets (PP&E + right-of-use + aircraft advances)',
     [HB[y]['ppe'] + HB[y]['rou'] + HB[y]['adv'] for y in H3]),
    (6, 'Intangible assets', [None, None, HB['FY25']['intang']]),
    (7, 'Other non-current assets', None),
    (8, 'Inventories', [HB[y]['inv'] for y in H3]),
    (9, 'Trade and other receivables (current)', [HB[y]['recv'] for y in H3]),
    (10, 'Cash and fixed deposits', [HB[y]['cash'] + HB[y]['dep'] for y in H3]),
    (12, 'Borrowings + lease liabilities', [HB[y]['debt'] for y in H3]),
    (13, 'Payables, deferred income, provisions (operating)',
     [HB[y]['pay'] + HB[y]['definc'] + HB[y]['maint'] + HB[y]['staffb'] for y in H3]),
    (15, 'Equity attributable to owners', [HB[y]['eqp'] for y in H3]),
]
for rr, lbl, vals in BROWS:
    put(wsBS, f'A{rr}', lbl, fmt=None)
    if vals:
        for j, cc in enumerate(HC):
            if vals[j] is not None:
                put(wsBS, f'{cc}{rr}', vals[j], BLUE, NUM0)
put(wsBS, 'A16', 'Working capital (from the lines above)', fmt=None)
put(wsBS, 'A17', 'Net debt (negative = net cash)', fmt=None)
put(wsBS, 'A18', 'Net debt / EBITDA', fmt=None)
for j, cc in enumerate(HC):
    y = H3[j]
    putf(wsBS, f'{cc}16', f'={cc}8+{cc}9-{cc}13', HB[y]['nwc'], NUM0)
    putf(wsBS, f'{cc}17', f'={cc}12-{cc}10', HB[y]['nd'], NUM0)
    putf(wsBS, f'{cc}18', f"={cc}17/'Income Statement'!{cc}9", HB[y]['nd'] / HI[y]['ebitda'], MULT)
for j, cc in enumerate(FCOL):
    sc = CD[j]
    putf(wsBS, f'{cc}5', f'=DCF!{sc}46', F['ppe'][j], NUM0, green=True)
    putf(wsBS, f'{cc}6', f"={AR('Intangible assets (AED mn, audited)')}", intang25, NUM0, green=True)
    putf(wsBS, f'{cc}16', f'=DCF!{sc}40', F['nwc'][j], NUM0, green=True)
    putf(wsBS, f'{cc}17', f"='Cash Flow'!{sc}16", F['net_debt'][j], NUM0, green=True)
    putf(wsBS, f'{cc}18', f"={cc}17/'Income Statement'!{cc}9", F['net_debt'][j] / F['ebitda'][j], MULT)
    if j == 0:
        putf(wsBS, f'{cc}15', f"={AR('Equity attributable to owners, FY2025 (AED mn, audited)')}"
             f"+'Income Statement'!{cc}20-'Cash Flow'!{sc}15", F['equity'][j], NUM0)
    else:
        putf(wsBS, f'{cc}15', f"={FCOL[j-1]}15+'Income Statement'!{cc}20-'Cash Flow'!{sc}15",
             F['equity'][j], NUM0)
put(wsBS, 'A20', 'The forecast rolls fleet assets (capex less depreciation plus leased-fleet additions), '
    'working capital (ratio x revenue), net debt (opening less retained cash generation plus dividends) '
    'and equity (profit less dividends). Deferred income, maintenance provisions and staff benefits are '
    'inside working capital — never double-counted against net debt.', fmt=None, wrap=True)

# ============ 7 RELATIVE & NORMALIZED =========================================
wsR = sheet('Relative & Normalized')
title(wsR, 'Relative multiples · normalised earnings · book value', None, 7, awidth=56, cwidth=13)
hdr(wsR, 4, ['Relative lens', '', 'Value'])
putf(wsR, 'C5', '=Segments!C27', F['ebitda'][1], NUM0, green=True)
put(wsR, 'A5', 'FY2027E EBITDA EXCLUDING fees/other income (AED mn) — peer basis', fmt=None)
putf(wsR, 'C6', f"={AR('Justified EV/EBITDA')}", IN['ev_ebitda_just'], MULT, green=True)
put(wsR, 'A6', 'Justified EV/EBITDA (peer median, primary filings)', fmt=None)
putf(wsR, 'C7', f"=Segments!C28*(1-{AR('Tax rate')})/(DCF!C24-{AR('Terminal growth')})",
     REL['fee_value'], NUM0)
put(wsR, 'A7', 'Fee/other-income stream valued separately (after-tax annuity)', fmt=None)
putf(wsR, 'C8', '=C5*C6+C7', REL['ev_rel_fwd'], NUM0)
put(wsR, 'A8', 'Enterprise value at end-FY2027', fmt=None)
putf(wsR, 'C9', '=DCF!C30', F['df'][1], DF4, green=True)
put(wsR, 'A9', 'Year-2 discount factor', fmt=None)
putf(wsR, 'C10', '=DCF!B44+DCF!C44', F['pv'][0] + F['pv'][1], NUM0, green=True)
put(wsR, 'A10', 'PV of FY2026-27 free cash flow', fmt=None)
BRIDGE_CASH = (f"({AR('Cash and fixed deposits, FY2025 (AED mn, audited)')}"
               f"-{AR('Gross debt, FY2025 (AED mn, audited)')}"
               f"+{AR('Non-operating assets (investments + investment property + net investment in lease, AED mn)')}"
               f"+{AR('JV and associates at carrying value (AED mn, audited)')})")
def rel_formula(mult_ref):
    return (f"=((C5*{mult_ref}+C7)*C9+C10)*DCF!C61/{AR('Shares outstanding (mn)')}"
            f"+({BRIDGE_CASH}*DCF!C63-{AR('Minority interests at carrying value (AED mn, audited)')})"
            f"/{AR('Shares outstanding (mn)')}"
            f"-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}")
putf(wsR, 'C11', rel_formula('C6'), LN['relative']['base'], PX, bold=True)
put(wsR, 'A11', 'Implied value per share at the anchor (split roll)', bold=True, fmt=None)
putf(wsR, 'C12', rel_formula(AR('Relative lens — bear multiple')), LN['relative']['bear'], PX)
put(wsR, 'A12', 'Bear (5.0x)', fmt=None)
putf(wsR, 'C13', rel_formula(AR('Relative lens — bull multiple')), LN['relative']['bull'], PX)
put(wsR, 'A13', 'Bull (8.0x)', fmt=None)
put(wsR, 'A15', 'Trailing context (formulas at spot)', bold=True, fmt=None)
putf(wsR, 'C16', f"=(DCF!C14+{AR('Gross debt, FY2025 (AED mn, audited)')}-{AR('Cash and fixed deposits, FY2025 (AED mn, audited)')})/'Income Statement'!D9",
     (M['mktcap'] + nd25) / HI['FY25']['ebitda'], MULT)
put(wsR, 'A16', 'EV / FY2025 EBITDA at spot — EX fees (peer basis)', fmt=None)
putf(wsR, 'C18', f"=(DCF!C14+{AR('Gross debt, FY2025 (AED mn, audited)')}-{AR('Cash and fixed deposits, FY2025 (AED mn, audited)')})/('Income Statement'!D9+'Income Statement'!D12)",
     (M['mktcap'] + nd25) / HI['FY25']['ebitda_incl'], MULT)
put(wsR, 'A18', 'EV / FY2025 EBITDA at spot — INCLUDING fees (both bases published)', fmt=None)
putf(wsR, 'C17', f"={AR('Spot price (AED)')}/'Income Statement'!D21", REL['pe_trailing'], MULT)
put(wsR, 'A17', 'Price / FY2025 earnings at spot', fmt=None)
hdr(wsR, 20, ['Normalised earnings power', '', 'Value'])
putf(wsR, 'C21', '=Segments!D30', F['ebitda_margin'][2], PCT, green=True)
put(wsR, 'A21', 'Mid-cycle EBITDA margin (FY2028E)', fmt=None)
putf(wsR, 'C22', '=Segments!B14', F['rev'][0], NUM0, green=True)
put(wsR, 'A22', 'FY2026E revenue (current scale)', fmt=None)
putf(wsR, 'C23', '=C21*C22+Segments!B28', NRM['ebitda'], NUM0)
put(wsR, 'A23', 'Normalised EBITDA incl. fees and other income', fmt=None)
putf(wsR, 'C24', f"={ARP('Depreciation & amortisation (AED mn)', 0)}", F['dna'][0], NUM0, green=True)
put(wsR, 'A24', 'less depreciation & amortisation', fmt=None)
putf(wsR, 'C25', '=C23-C24', NRM['ebit'], NUM0)
put(wsR, 'A25', 'Normalised EBIT', fmt=None)
putf(wsR, 'C26', f"='Cash Flow'!B13-'Cash Flow'!B14", F['fininc'][0] - F['interest'][0], NUM1, green=True)
put(wsR, 'A26', 'Net finance income (FY2026E)', fmt=None)
put(wsR, 'A27', 'JV share EXCLUDED from the multiplied base (enters at book below) — '
    'the base framing carries the JV at carrying value in every lens', fmt=None)
putf(wsR, 'C28', f"=(C25+C26)*(1-{AR('Tax rate')})*(1-{AR('Minority share of profit')})/{AR('Shares outstanding (mn)')}",
     NRM['eps'], '0.000')
put(wsR, 'A28', 'Normalised earnings per share (ex-JV)', fmt=None)
NORM_TAIL = (f"*C28*DCF!C61+{AR('JV and associates at carrying value (AED mn, audited)')}"
             f"/{AR('Shares outstanding (mn)')}*DCF!C63"
             f"-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}")
putf(wsR, 'C29', f"={AR('Justified price/earnings')}{NORM_TAIL}",
     LN['normalized']['base'], PX, bold=True)
put(wsR, 'A29', 'Normalised value per share at the anchor', bold=True, fmt=None)
putf(wsR, 'C31', f"={AR('Normalised lens — bear P/E')}{NORM_TAIL}", LN['normalized']['bear'], PX)
put(wsR, 'A31', 'Bear (10x)', fmt=None)
putf(wsR, 'C32', f"={AR('Normalised lens — bull P/E')}{NORM_TAIL}", LN['normalized']['bull'], PX)
put(wsR, 'A32', 'Bull (16x)', fmt=None)
hdr(wsR, 35, ['Book value and sustainable return', '', 'Value'])
putf(wsR, 'C36', f"={AR('Equity attributable to owners, FY2025 (AED mn, audited)')}/{AR('Shares outstanding (mn)')}",
     BK['bvps'], PX)
put(wsR, 'A36', 'Book value per share (FY2025, audited)', fmt=None)
putf(wsR, 'C37', f"=(({AR('Sustainable return on equity')}-{AR('Terminal growth')})/(DCF!C21-{AR('Terminal growth')}))*C36*DCF!C61-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     LN['book']['base'], PX, bold=True)
put(wsR, 'A37', 'Justified price-to-book value per share at the anchor', bold=True, fmt=None)
putf(wsR, 'C38', f"=({AR('Sustainable return on equity')}-{AR('Terminal growth')})/(DCF!C21-{AR('Terminal growth')})",
     BK['pb_just'], MULT)
put(wsR, 'A38', 'Justified P/B multiple', fmt=None)
putf(wsR, 'C39', f"=(({AR('Sustainable return on equity')}-0.02-0.015)/(0.5*(DCF!C10+DCF!C21)-0.015))*C36*DCF!C61-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     LN['book']['bear'], PX)
put(wsR, 'A41', 'The bear leg holds the same Gordon identity as base and bull: '
    '(ROE_bear - g_bear)/(k_bear - g_bear).', fmt=None, wrap=True)
put(wsR, 'A39', 'Bear construction', fmt=None)
putf(wsR, 'C40', f"=(({AR('Sustainable return on equity')}+0.02-{AR('Terminal growth')})/(DCF!C21-{AR('Terminal growth')}))*C36*DCF!C61-{AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
     LN['book']['bull'], PX)
put(wsR, 'A40', 'Bull construction', fmt=None)

# ============ 12 SUMMARY FINANCIALS ===========================================
wsSF = sheet('Summary Financials')
title(wsSF, 'Summary financials — history and forecast in one grid', None, 10, awidth=44, cwidth=11)
hdr(wsSF, 4, [''] + ['FY2023', 'FY2024', 'FY2025'] + YF)
def sf_row(rr, lbl, hist_fm, fc_fm, hist_ex, fc_ex, fmt):
    put(wsSF, f'A{rr}', lbl, fmt=None)
    for j, cc in enumerate(HC):
        if hist_fm:
            putf(wsSF, f'{cc}{rr}', hist_fm(j, cc), hist_ex[j], fmt, green=True)
    for j, cc in enumerate(FCOL):
        putf(wsSF, f'{cc}{rr}', fc_fm(j, cc), fc_ex[j], fmt, green=True)
sf_row(5, 'Revenue (AED mn)', lambda j, cc: f"='Income Statement'!{cc}5",
       lambda j, cc: f"='Income Statement'!{cc}5", [HI[y]['rev'] for y in H3], F['rev'], NUM0)
put(wsSF, 'A6', 'Revenue growth', fmt=None)
for j, cc in enumerate(['C', 'D']):
    putf(wsSF, f'{cc}6', f"={cc}5/{HC[j]}5-1", HI[H3[j+1]]['rev'] / HI[H3[j]]['rev'] - 1, PCT)
for j, cc in enumerate(FCOL):
    prev = 'D' if j == 0 else FCOL[j-1]
    putf(wsSF, f'{cc}6', f"={cc}5/{prev}5-1",
         F['rev'][j] / (HI['FY25']['rev'] if j == 0 else F['rev'][j-1]) - 1, PCT)
sf_row(7, 'EBITDA (AED mn)', lambda j, cc: f"='Income Statement'!{cc}9",
       lambda j, cc: f"='Income Statement'!{cc}9", [HI[y]['ebitda'] for y in H3], F['ebitda'], NUM0)
sf_row(8, 'EBITDA margin', lambda j, cc: f"='Income Statement'!{cc}22",
       lambda j, cc: f"='Income Statement'!{cc}22",
       [HI[y]['ebitda'] / HI[y]['rev'] for y in H3], F['ebitda_margin'], PCT)
sf_row(9, 'Attributable profit (AED mn)', lambda j, cc: f"='Income Statement'!{cc}20",
       lambda j, cc: f"='Income Statement'!{cc}20", [HI[y]['npa'] for y in H3], np_f, NUM0)
sf_row(10, 'Earnings per share (AED)', lambda j, cc: f"='Income Statement'!{cc}21",
       lambda j, cc: f"='Income Statement'!{cc}21", [HI[y]['npa'] / SH for y in H3],
       [n / SH for n in np_f], PX)
put(wsSF, 'A11', 'Dividend per share (AED)', fmt=None)
for j, (cc, v) in enumerate(zip(HC, [0.20, 0.25, 0.30])):
    put(wsSF, f'{cc}11', v, BLUE, PX)
for j, cc in enumerate(FCOL):
    putf(wsSF, f'{cc}11', f"='Cash Flow'!{CD[j]}15/{AR('Shares outstanding (mn)')}",
         F['div'][j] / SH, PX)
sf_row(12, 'Net debt (AED mn, negative = net cash)', lambda j, cc: f"='Balance Sheet'!{cc}17",
       lambda j, cc: f"='Balance Sheet'!{cc}17", [HB[y]['nd'] for y in H3], F['net_debt'], NUM0)
put(wsSF, 'A13', 'Invested capital (AED mn)', fmt=None)
for j, cc in enumerate(FCOL):
    putf(wsSF, f'{cc}13', f'=DCF!{CD[j]}47', F['ic'][j], NUM0, green=True)
put(wsSF, 'A14', 'Return on invested capital', fmt=None)
for j, cc in enumerate(FCOL):
    putf(wsSF, f'{cc}14', f'=DCF!{CD[j]}37/DCF!{CD[j]}47', F['roic'][j], PCT)
put(wsSF, 'A15', 'Return on equity (attributable, opening equity)', fmt=None)
for j, cc in enumerate(FCOL):
    open_eq = IN['eqp_fy25'] if j == 0 else F['equity'][j-1]
    src = AR('Equity attributable to owners, FY2025 (AED mn, audited)') if j == 0 \
        else f"'Balance Sheet'!{FCOL[j-1]}15"
    putf(wsSF, f'{cc}15', f"='Income Statement'!{cc}20/{src}", np_f[j] / open_eq, PCT)
put(wsSF, 'A16', 'Passengers (millions)', fmt=None)
for j, cc in enumerate(HC):
    put(wsSF, f'{cc}16', BU['pax_hist'][H3[j]], BLUE, NUM2)
for j, cc in enumerate(FCOL):
    putf(wsSF, f'{cc}16', f'=Segments!{CD[j]}5', F['pax'][j], NUM2, green=True)
put(wsSF, 'A17', 'Revenue per passenger (AED, total revenue basis)', fmt=None)
for j, cc in enumerate(ALL):
    if j < 3:
        putf(wsSF, f'{cc}17', f'={cc}5/{cc}16', HI[H3[j]]['rev'] / BU['pax_hist'][H3[j]], NUM1)
    else:
        putf(wsSF, f'{cc}17', f'={cc}5/{cc}16', F['rev'][j-3] / F['pax'][j-3], NUM1)

# ============ 13 MONTE CARLO ===================================================
wsM = sheet('Monte Carlo')
title(wsM, 'Monte Carlo price map — 50,000 simulated paths', 'Whole-model engine output: pasted, does not redraw with drivers.',
      8, awidth=48, cwidth=12)
hdr(wsM, 4, ['Horizon', '1 month', '3 months'])
H1, H3m = STK['horizons']['1M'], STK['horizons']['3M']
MC_ROWS = [('Check date', H1['grade_date'], H3m['grade_date'], None),
           ('5th percentile (AED)', H1['pct']['p5'], H3m['pct']['p5'], PX),
           ('25th percentile', H1['pct']['p25'], H3m['pct']['p25'], PX),
           ('Median', H1['pct']['p50'], H3m['pct']['p50'], PX),
           ('75th percentile', H1['pct']['p75'], H3m['pct']['p75'], PX),
           ('95th percentile', H1['pct']['p95'], H3m['pct']['p95'], PX),
           ('Probability of finishing above spot', H1['p_above'], H3m['p_above'], PCT),
           ('Probability of +10% or better at the check date', H1['p_up10'], H3m['p_up10'], PCT),
           ('Probability of -10% or worse at the check date', H1['p_dn10'], H3m['p_dn10'], PCT),
           ('Probability of TOUCHING +10% at any point', H1['touch_up10'], H3m['touch_up10'], PCT),
           ('Probability of TOUCHING -10% at any point', H1['touch_dn10'], H3m['touch_dn10'], PCT)]
r = 5
for lbl, v1, v3, fmt in MC_ROWS:
    put(wsM, f'A{r}', lbl, fmt=None)
    put(wsM, f'B{r}', v1, BLUE, fmt)
    put(wsM, f'C{r}', v3, BLUE, fmt)
    r += 1
put(wsM, f'A{r+1}', f"Struck at AED {STK['spot']:.2f} on {STK['anchor_date']} with a 5.7% trailing dividend "
    "yield netted from the drift. How much to trust these bands: over the stock's full history the "
    "simulation's probability bands beat a naive random-walk benchmark by a small margin on a standard "
    "probabilistic accuracy score (+0.7% over the full 58-window history, 2012-2026), and over the "
    "recent four-plus years the two are statistically indistinguishable — the bands are honest, not "
    "clairvoyant. Realised outcomes landed inside the 80% band 79% of the time and inside the 90% band "
    "86% of the time over that same full history — ONE window set, quoted identically in the study.",
    fmt=None, wrap=True)

# ============ 14 SENSITIVITY ===================================================
wsX = sheet('Sensitivity')
title(wsX, 'Sensitivity — each cell is a complete revaluation', 'Whole-model re-runs: pasted, do not redraw with drivers.',
      8, awidth=40, cwidth=12)
r = 4
def grid(lbl, rows_lbl, cols_lbl, rows, cols, table, fmt=PX):
    global r
    put(wsX, f'A{r}', lbl, bold=True, fmt=None); r += 1
    hdr(wsX, r, [rows_lbl + ' \\ ' + cols_lbl] + [f'{c}' for c in cols]); r += 1
    for i, rl in enumerate(rows):
        put(wsX, f'A{r}', rl, fmt=None)
        for j, v in enumerate(table[i]):
            put(wsX, f'{get_column_letter(2+j)}{r}', v, BLUE, fmt)
        r += 1
    r += 1
grid('Fair value per share: terminal cost of capital x terminal growth',
     'terminal cost of capital', 'terminal growth',
     [f"{w:.2%}" for w in SN['wt_grid']], [f"{g:.1%}" for g in SN['g_grid']], SN['grid_wacc_g'])
grid('Fair value per share: explicit x terminal cost of capital',
     'explicit', 'terminal', [f"{w:.2%}" for w in SN['we_grid']],
     [f"{w:.2%}" for w in SN['wt_grid']], SN['grid_exp_term'])
def vec(lbl, xs, vals, xfmt='{}'):
    global r
    put(wsX, f'A{r}', lbl, bold=True, fmt=None); r += 1
    hdr(wsX, r, [''] + [xfmt.format(x) for x in xs]); r += 1
    put(wsX, f'A{r}', 'Fair value per share (AED)', fmt=None)
    for j, v in enumerate(vals):
        put(wsX, f'{get_column_letter(2+j)}{r}', v, BLUE, PX)
    r += 2
vec('Beta (leftmost column = the alternative-benchmark regression, published not adopted)',
    SN['beta_grid'], SN['grid_beta'])
vec('Fuel cost per passenger (multiplier on the base path)', SN['fuel_grid'], SN['grid_fuel'])
vec('Passenger volumes (multiplier)', SN['paxg_grid'], SN['grid_pax'])
vec('Fare per passenger (multiplier)', SN['fare_grid'], SN['grid_fare'])
vec('Fleet capex (multiplier)', SN['capex_grid'], SN['grid_capex'])
vec('JV network value in the bridge (AED mn)', [f"{x:,.0f}" for x in SN['jv_grid']], SN['grid_jv'])
vec('Working capital / revenue', [f"{x:.0%}" for x in SN['nwc_grid']], SN['grid_nwc'])
put(wsX, f'A{r}', 'Published scenario driver vectors — type these into the Scenario cells on '
    'Assumptions to reproduce each pasted output LIVE in this workbook', bold=True, fmt=None); r += 1
hdr(wsX, r, ['Scenario', 'Passenger x', 'Fare x', 'Fuel path', 'Capex x', 'Rate shift',
             'Terminal g', 'Output (AED/sh)']); r += 1
for nm, px_, fx_, fu, cx, sh_, g_, outv in [
        ('High-fuel alternative', 1.0, 1.0, 'alternative (switch=1)', 1.0, '0bp', '2.5%', DCF['ps_iata_fuel']),
        ('Bear', 0.94, 0.97, 'alternative (switch=1)', 1.15, '+100bp', '1.5%', DCF['bear']),
        ('Bull (JV capitalised)', 1.05, 1.03, 'base (switch=0)', 0.90, '-100bp', '3.5%', DCF['bull'])]:
    put(wsX, f'A{r}', nm, fmt=None)
    for col, v in zip(['B', 'C', 'D', 'E', 'F', 'G', 'H'],
                      [px_, fx_, fu, cx, sh_, g_, outv]):
        put(wsX, f'{col}{r}', v, BLUE, PX if col == 'H' else (NUM2 if isinstance(v, float) else None))
    r += 1
put(wsX, f'A{r}', 'The bull additionally switches the SOTP Bridge to the JV-capitalised row. '
    'Every other figure in the workbook reprices live when the Scenario cells change.', fmt=None); r += 1

# ============ 15 PER-SHARE & RATIOS ===========================================
wsP = sheet('Per-Share & Ratios')
title(wsP, 'Per-share figures and ratios — all formulas', None, 8, awidth=48, cwidth=12)
hdr(wsP, 4, ['Per share (AED)', 'FY2025', 'FY2026E', 'FY2028E', 'FY2030E'])
PS = [
    ('Earnings per share', ["='Income Statement'!D21", "='Income Statement'!E21",
                            "='Income Statement'!G21", "='Income Statement'!I21"],
     [HI['FY25']['npa'] / SH, np_f[0] / SH, np_f[2] / SH, np_f[4] / SH]),
    ('Dividend per share', [f"={AR('FY2025 dividend per share (AED, approved 12 March 2026)')}",
                            f"='Cash Flow'!B15/{AR('Shares outstanding (mn)')}",
                            f"='Cash Flow'!D15/{AR('Shares outstanding (mn)')}",
                            f"='Cash Flow'!F15/{AR('Shares outstanding (mn)')}"],
     [IN['dps_fy25'], F['div'][0] / SH, F['div'][2] / SH, F['div'][4] / SH]),
    ('Book value per share', [f"={AR('Equity attributable to owners, FY2025 (AED mn, audited)')}/{AR('Shares outstanding (mn)')}",
                              f"='Balance Sheet'!E15/{AR('Shares outstanding (mn)')}",
                              f"='Balance Sheet'!G15/{AR('Shares outstanding (mn)')}",
                              f"='Balance Sheet'!I15/{AR('Shares outstanding (mn)')}"],
     [IN['eqp_fy25'] / SH, F['equity'][0] / SH, F['equity'][2] / SH, F['equity'][4] / SH]),
    ('Free cash flow to the firm per share', [None, f"='Cash Flow'!B9/{AR('Shares outstanding (mn)')}",
                                              f"='Cash Flow'!D9/{AR('Shares outstanding (mn)')}",
                                              f"='Cash Flow'!F9/{AR('Shares outstanding (mn)')}"],
     [None, F['fcff'][0] / SH, F['fcff'][2] / SH, F['fcff'][4] / SH]),
    ('Net cash per share', ["='SOTP Bridge'!C6/" + AR('Shares outstanding (mn)').split('!')[1].join(["Assumptions!", ""]) if False else f"='SOTP Bridge'!C6/{AR('Shares outstanding (mn)')}",
                            None, None, None],
     [-nd25 / SH, None, None, None]),
]
r = 5
for lbl, fms, exs in PS:
    put(wsP, f'A{r}', lbl, fmt=None)
    for j, cc in enumerate(['B', 'C', 'D', 'E']):
        if fms[j] is not None:
            putf(wsP, f'{cc}{r}', fms[j], exs[j], PX, green=True)
    r += 1
hdr(wsP, r + 1, ['At the spot price', '', 'Value'])
r += 2
RAT = [
    ('Trailing price/earnings', f"={AR('Spot price (AED)')}/'Income Statement'!D21", REL['pe_trailing'], MULT),
    ('Forward price/earnings (FY2027E)', f"={AR('Spot price (AED)')}/'Income Statement'!F21",
     SPOT / (np_f[1] / SH), MULT),
    ('Price / book (FY2025)', f"={AR('Spot price (AED)')}/(({AR('Equity attributable to owners, FY2025 (AED mn, audited)')})/{AR('Shares outstanding (mn)')})",
     SPOT / (IN['eqp_fy25'] / SH), MULT),
    ('EV / EBITDA (trailing, incl. fees)', f"=(DCF!C14+{AR('Gross debt, FY2025 (AED mn, audited)')}-{AR('Cash and fixed deposits, FY2025 (AED mn, audited)')})/('Income Statement'!D9+'Income Statement'!D12)",
     (M['mktcap'] + nd25) / HI['FY25']['ebitda_incl'], MULT),
    ('Dividend yield (FY2025 dividend at spot)', f"={AR('FY2025 dividend per share (AED, approved 12 March 2026)')}/{AR('Spot price (AED)')}",
     IN['dps_fy25'] / SPOT, PCT),
]
for lbl, fm, ex, fmt in RAT:
    put(wsP, f'A{r}', lbl, fmt=None)
    putf(wsP, f'C{r}', fm, ex, fmt)
    r += 1

# ============ 16 PEER & SECTOR =================================================
wsPS = sheet('Peer & Sector')
title(wsPS, 'Peer frame — cross-check only, never a build source',
      'Peer figures from public reporting and reputable aggregators, as dated; Air Arabia rows are live formulas.',
      8, awidth=40, cwidth=13)
hdr(wsPS, 4, ['Carrier', 'Trailing P/E', 'EV/EBITDA', 'Basis / date'])
PEERS = [
    ('Ryanair', 12.60, 7.81, 'TTM USD ADR, 09-Aug-2026 (aggregator)'),
    ('easyJet', 12.39, 3.48, 'TTM GBP, 07-Aug-2026 (aggregator)'),
    ('IndiGo', None, 11.09, 'TTM INR, 09-Aug-2026; P/E n/m (FX loss year)'),
    ('Pegasus', 8.90, 6.58, 'TRY statutory basis — treat with caution'),
    ('Wizz Air', None, None, 'n/m — near-zero FY2026 earnings'),
    ('Jazeera Airways', 17.7, None, 'computed: mcap / FY2025 net, KWD, 07-Aug-2026'),
    ('Air transport sector (Damodaran, Jan-2026)', 12.87, 7.58, 'profitable firms / positive-EBITDA firms'),
]
r = 5
for nm, pe, ev, basis in PEERS:
    put(wsPS, f'A{r}', nm, fmt=None)
    put(wsPS, f'B{r}', pe if pe is not None else 'n/m', BLUE, MULT if pe is not None else None)
    put(wsPS, f'C{r}', ev if ev is not None else 'n/m', BLUE, MULT if ev is not None else None)
    put(wsPS, f'D{r}', basis, fmt=None)
    r += 1
band(wsPS, r, 4)
put(wsPS, f'A{r}', 'Air Arabia (live formulas)', bold=True, fmt=None)
putf(wsPS, f'B{r}', f"='Per-Share & Ratios'!C{r+7-r+5}" if False else f"={AR('Spot price (AED)')}/'Income Statement'!D21",
     REL['pe_trailing'], MULT, bold=True)
putf(wsPS, f'C{r}', f"=(DCF!C14+{AR('Gross debt, FY2025 (AED mn, audited)')}-{AR('Cash and fixed deposits, FY2025 (AED mn, audited)')})/('Income Statement'!D9+'Income Statement'!D12)",
     (M['mktcap'] + nd25) / HI['FY25']['ebitda_incl'], MULT, bold=True)
put(wsPS, f'D{r}', 'FY2025 audited, spot 07-Aug-2026', fmt=None)
r += 2
put(wsPS, f'A{r}', 'Justified multiples adopted: EV/EBITDA 7.5x (sector centre), P/E 13x (between the '
    'mature European carriers ~12.5x and Jazeera ~17.7x). Air Arabia trades above the peer set on both '
    'trailing multiples — the relative lens question is whether its net cash, franchise margin and '
    'growing JV network justify the premium.', fmt=None, wrap=True)

# ---- finalize ----------------------------------------------------------------
ORDER = ['READ FIRST', 'Summary', 'Fundamental Valuation', 'Assumptions', 'SOTP Bridge',
         'Segments', 'Relative & Normalized', 'DCF', 'Income Statement', 'Balance Sheet',
         'Cash Flow', 'Summary Financials', 'Monte Carlo', 'Sensitivity',
         'Per-Share & Ratios', 'Peer & Sector']
wb._sheets = [wb[n] for n in ORDER]
assert wb.sheetnames == ORDER and len(ORDER) == 16
OUT = os.path.join(HERE, 'AIRARABIA_Valuation_Model_09082026_public.xlsx')
wb.save(OUT)
nf = sum(len(v) for v in EXPECT.values())
with open(os.path.join(HERE, 'xlsx_expected.json'), 'w') as f:
    json.dump(dict(expected=EXPECT, anchors=ANCH), f, indent=1)
print(f'wrote {os.path.basename(OUT)} — 16 sheets, {nf} formula cells with recorded expectations')
