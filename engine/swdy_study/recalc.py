"""Recalculate the delivered xlsx and reconcile it cell-by-cell against
study_numbers.json. Fails loudly on any unresolvable formula or mismatch.

LibreOffice is not usable in this environment (it cannot load any spreadsheet,
including a trivial CSV and the previous study's own workbook), so the
recalculation is done by an explicit evaluator over the formula set this builder
actually emits: arithmetic, SUM/MIN/MAX/MEDIAN over ranges, absolute and
relative cell references, and cross-sheet references. Anything the evaluator
does not understand is reported as a failure rather than skipped.
"""
import json, os, re, sys
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'SWDY_Valuation_Model_05082026_public.xlsx')
wb = openpyxl.load_workbook(XLSX)
D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
DCF, LN, HI, HB, F = D['dcf'], D['lenses'], D['hist_is'], D['hist_bs'], D['fcst']
SH = D['meta']['shares_mn']

CACHE = {}
STACK = []

def cell_value(sheet, coord):
    key = (sheet, coord)
    if key in CACHE:
        return CACHE[key]
    if key in STACK:
        raise ValueError(f'circular reference at {sheet}!{coord}')
    v = wb[sheet][coord].value
    if isinstance(v, str) and v.startswith('='):
        STACK.append(key)
        try:
            v = evaluate(v[1:], sheet)
        finally:
            STACK.pop()
    elif isinstance(v, str):
        v = 0.0 if v == '-' else v
    elif v is None:
        v = 0.0
    CACHE[key] = v
    return v

def range_values(sheet, rng):
    c1, r1, c2, r2 = range_boundaries(rng)
    out = []
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            out.append(cell_value(sheet, f'{get_column_letter(cc)}{rr}'))
    return [x for x in out if isinstance(x, (int, float))]

FUNC = re.compile(r'\b(SUM|MIN|MAX|MEDIAN|AVERAGE)\(([^()]*)\)')
SHEETREF = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_-]*))!(\$?[A-Z]{1,3}\$?\d+)")
CELLREF = re.compile(r'(?<![A-Z0-9_!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![\d(])')

def evaluate(expr, sheet):
    e = expr
    # functions over ranges (possibly cross-sheet)
    while True:
        m = FUNC.search(e)
        if not m:
            break
        fn, arg = m.group(1), m.group(2).strip()
        tgt, rng = sheet, arg
        sm = re.match(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &_-]*))!(.+)", arg)
        if sm:
            tgt = sm.group(1) or sm.group(2); rng = sm.group(3)
        vals = range_values(tgt, rng.replace('$', ''))
        if not vals:
            val = 0.0
        elif fn == 'SUM':
            val = sum(vals)
        elif fn == 'MIN':
            val = min(vals)
        elif fn == 'MAX':
            val = max(vals)
        else:
            vs = sorted(vals); n = len(vs)
            val = (vs[n // 2] if n % 2 else (vs[n // 2 - 1] + vs[n // 2]) / 2)
        e = e[:m.start()] + repr(float(val)) + e[m.end():]
    # cross-sheet single-cell references
    while True:
        m = SHEETREF.search(e)
        if not m:
            break
        tgt = m.group(1) or m.group(2)
        v = cell_value(tgt, m.group(3).replace('$', ''))
        e = e[:m.start()] + repr(float(v or 0)) + e[m.end():]
    # same-sheet references
    while True:
        m = CELLREF.search(e)
        if not m:
            break
        v = cell_value(sheet, f'{m.group(2)}{m.group(4)}')
        e = e[:m.start()] + repr(float(v or 0)) + e[m.end():]
    if not re.fullmatch(r'[-+*/(). 0-9eE]+', e):
        raise ValueError(f'unparsed formula fragment: {expr!r} -> {e!r}')
    return eval(e)

# ---- 1: every formula must evaluate ------------------------------------------
nform, errors = 0, []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
                try:
                    cell_value(ws.title, c.coordinate)
                except Exception as ex:
                    errors.append(f'{ws.title}!{c.coordinate}: {ex}')
print(f'formulas: {nform}, unresolvable: {len(errors)}')
for e in errors[:20]:
    print('  ', e)

# ---- 2: cell-level reconciliation against study_numbers.json ------------------
def g(sheet, cell):
    return cell_value(sheet, cell)

checks = [
    ('DCF enterprise value', g('DCF', 'C29'), DCF['ev'], 1.0),
    ('DCF present value of explicit years', g('DCF', 'C26'), DCF['pv_explicit'], 1.0),
    ('DCF present value of terminal value', g('DCF', 'C27'), DCF['pv_tv'], 1.0),
    ('DCF terminal value share', g('DCF', 'C28'), DCF['tv_share'], 0.002),
    ('DCF fair value per share', g('DCF', 'C31'), DCF['ps'], 0.02),
    ('Bridge equity attributable', g('SOTP Bridge', 'C12'), DCF['eq_attr'], 1.0),
    ('Bridge enterprise value', g('SOTP Bridge', 'C7'), DCF['ev'], 1.0),
    ('Fundamental — DCF lens', g('Fundamental Valuation', 'C5'), DCF['ps'], 0.02),
    ('Fundamental — relative lens', g('Fundamental Valuation', 'C8'), LN['relative']['base'], 0.02),
    ('Fundamental — normalised lens', g('Fundamental Valuation', 'C9'), LN['normalized']['base'], 0.02),
    ('Fundamental — book lens', g('Fundamental Valuation', 'C10'), LN['book']['base'], 0.02),
    ('Fundamental — panel median', g('Fundamental Valuation', 'C26'), D['panel_centre'], 0.02),
    ('Fundamental — currency alternative', g('Fundamental Valuation', 'C18'), DCF['ccy_alt_ps'], 0.02),
    ('Summary weighted central', g('Summary', 'C9'), D['central'], 0.02),
    ('Summary terminal value share', g('Summary', 'C12'), DCF['tv_share'], 0.002),
    ('Summary panel median', g('Summary', 'C13'), D['panel_centre'], 0.02),
    ('Summary market capitalisation', g('Summary', 'C18'), D['meta']['mktcap'], 1.0),
    ('Relative lens implied value', g('Relative & Normalized', 'C9'), LN['relative']['base'], 0.02),
    ('Normalised lens implied value', g('Relative & Normalized', 'C24'), LN['normalized']['base'], 0.02),
    ('Book lens implied value', g('Relative & Normalized', 'C32'), LN['book']['base'], 0.02),
    ('Segments total FY2025 revenue', g('Segments', 'B12'), HI['FY25']['rev'], 1.0),
    ('Segments group EBITDA FY2026E', g('Segments', 'B22'), F['ebitda'][0], 1.0),
    ('Segments FY2025 share sums to 100%', g('Segments', 'C12'), 1.0, 0.0005),
    ('Segments group EBITDA margin FY2026E', g('Segments', 'B23'), F['ebitda_margin'][0], 0.0005),
    ('Income statement FY2025 EBITDA', g('Income Statement', 'D7'), HI['FY25']['ebitda'], 1.0),
    ('Income statement FY2030E attributable profit', g('Income Statement', 'I17'), F['np_attr'][4], 1.0),
    ('Income statement FY2024 EBITDA margin', g('Income Statement', 'C8'),
     HI['FY24']['ebitda'] / HI['FY24']['rev'], 0.001),
    ('Balance sheet FY2024 net debt / EBITDA', g('Balance Sheet', 'C18'),
     HB['FY24']['nd'] / HI['FY24']['ebitda'], 0.01),
]
bad = 0
for name, got, want, tol in checks:
    ok = got is not None and abs(float(got) - float(want)) <= tol
    if not ok:
        bad += 1
    print(f"  [{'OK ' if ok else 'BAD'}] {name}: workbook={got:,.4f} model={float(want):,.4f}")

# The condensed balance sheet does not foot to zero by construction; the residual is the
# block of other liabilities, provisions and deferred tax. Check it against the audited
# FY2024 figures rather than asserting a cosmetic zero.
bc = [g('Balance Sheet', f'{c}19') for c in ('B', 'C')]   # FY2025 column is n/a
audited_other_fy24 = 13439.559 + 942.646 + 3669.893 + 2050.078 + 94.612 + 2631.482
print(f'condensed-balance residual FY2023/FY2024: {[round(float(v)) for v in bc]}')
print(f'  FY2024 residual {float(bc[1]):,.0f} vs audited other liabilities '
      f'{audited_other_fy24:,.0f} (gap {float(bc[1])-audited_other_fy24:+,.0f})')
assert abs(float(bc[1]) - audited_other_fy24) < 5.0, 'FY2024 condensed residual does not reconcile'
assert not errors, f'{len(errors)} unresolvable formulas'
assert bad == 0, f'{bad} reconciliation mismatches'
print(f'RECALC OK — {nform} formulas, 0 unresolvable, {len(checks)} reconciliation checks passed')
