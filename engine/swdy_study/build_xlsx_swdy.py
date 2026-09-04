"""SWDY_Valuation_Model_05082026_public.xlsx — 16 sheets mirroring the house canonical
model (operating-company variant). Blue = inputs · black = formulas · green = cross-sheet
links.

The workbook is FORMULA-DRIVEN. Every quantity that is arithmetically derivable from an
input is written as a live Excel formula, not as a pasted number, so the reader can trace
each figure back to the drivers on the Assumptions sheet and change one to see the model
reprice. Only three classes of cell are pasted values:

  1. audited and disclosed historical figures (the primary record);
  2. the three-segment revenue-and-margin build (Cables, Constructions and infrastructure,
     Electrical products and digital solutions — the company's own disclosed segments) — its
     OUTPUT (segment revenue and gross profit) is pasted and everything from there down is
     formula;
  3. engine outputs that are whole-model re-runs by construction: the Monte Carlo price
     map and the sensitivity grids, each cell of which is a complete revaluation.

Every formula cell also carries the model's own value for that cell into
xlsx_expected.json, and recalc.py evaluates the workbook independently and asserts the two
agree. A formula that computes the right thing the wrong way therefore fails the gate.
"""
import json, os
HERE = os.path.dirname(os.path.abspath(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open(os.path.join(HERE, 'study_numbers.json')))
BLUE = Font(color='0000FF'); GREEN = Font(color='008000'); BLACK = Font(color='000000')
TITLE = Font(bold=True, size=13, color='F6F1E6'); SUB = Font(size=9, color='6E7B77')
FILL_T = PatternFill('solid', start_color='1C3A36'); FILL_H = PatternFill('solid', start_color='EAF0EE')
FILL_G = PatternFill('solid', start_color='F6F1E6')
NUM0 = '#,##0;(#,##0);"-"'; NUM1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0.0%;(0.0%);"-"'; PCT2 = '0.00%'; PX = '0.00;(0.00);"-"'; MULT = '0.00x'; DF4 = '0.0000'
YH = ['FY2023', 'FY2024', 'FY2025']
YF = D['fcst']['years']
M, HI, HB, F = D['meta'], D['hist_is'], D['hist_bs'], D['fcst']
W, DCF, LN, SN = D['wacc'], D['dcf'], D['lenses'], D['sens']
EXP, TR, REL, NRM, BK = D['experts'], D['terminal_recon'], D['rel'], D['norm'], D['book']
SEG, S0, STK = D['seg_fy25'], D['step0'], D['strike']
IN = {k: v['value'] for k, v in D['inputs'].items()}
SPOT, SH = M['spot'], M['shares_mn']
SEGS = D['bottomup']['subs']
BU = D['bottomup']
TAX = IN['tax_eff']
NCI_SH = DCF['nci_share']
PAYOUT, ASSOC_G = F['payout'], F['assoc_g']
H3 = ['FY23', 'FY24', 'FY25']
CD = ['B', 'C', 'D', 'E', 'F']              # forecast columns on the DCF / Segments EBITDA blocks
HC = ['B', 'C', 'D']                        # historical columns on the statements
FCOL = ['E', 'F', 'G', 'H', 'I']            # forecast columns on the statements
CFF = ['D', 'E', 'F', 'G', 'H']             # forecast columns on the cash-flow sheet
ALL = HC + FCOL

wb = Workbook()

# ---- expected-value ledger: every formula cell records what the model says it is -------
EXPECT = {}
ANCH = {}

def sheet(n):
    ws = wb.create_sheet(n) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = n
    return ws

def title(ws, t, s=None, w=10, awidth=46, cwidth=13):
    ws['A1'] = t; ws['A1'].font = TITLE; ws['A1'].fill = FILL_T
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = FILL_T
    if s:
        ws['A2'] = s; ws['A2'].font = SUB
    ws.column_dimensions['A'].width = awidth
    for c in range(2, w + 1):
        ws.column_dimensions[get_column_letter(c)].width = cwidth

def put(ws, ad, v, font=BLACK, fmt=NUM0, bold=False, fill=None, wrap=False):
    c = ws[ad]; c.value = v
    c.font = Font(color=font.color, bold=bold)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

def putf(ws, ad, formula, expect, fmt=NUM0, bold=False, green=False):
    """Write a live formula and record the model's own value for the same cell."""
    put(ws, ad, formula, GREEN if green else BLACK, fmt, bold=bold)
    if expect is not None:
        EXPECT.setdefault(ws.title, {})[ad] = float(expect)

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = Font(bold=True); c.fill = FILL_H

def band(ws, row, w=10):
    for c in range(1, w + 1):
        ws.cell(row=row, column=c).fill = FILL_G
        ws.cell(row=row, column=c).font = Font(bold=True)

# ============ 1 READ FIRST ====================================================
ws = sheet('READ FIRST')
title(ws, 'Testahil — Elsewedy Electric Company S.A.E. (EGX: SWDY)', None, 9)
for i, ln in enumerate([
 'Companion model · Independent Valuation Study · Educational analysis · Not investment advice', '',
 'What this workbook is. A transparent companion to the SWDY valuation study. Every blue cell is an input;',
 'every black cell is a formula; green cells link across sheets.', '',
 'IT IS FORMULA-DRIVEN. Every figure that can be derived arithmetically from a driver is a live formula, so',
 'you can change a blue cell on Assumptions and watch the model reprice: the cost of capital is built from',
 'the risk-free rate, beta and the premium rather than pasted; the discount factors compound from the glide;',
 'and the income statement, balance sheet, cash flow, ratios and all four lenses chain off the same cells.', '',
 'THREE THINGS ARE PASTED VALUES, and it is worth knowing exactly which. First, audited and disclosed history',
 '— the primary record, not a calculation. Second, the segment build: revenue and profit for the three',
 'segments the company itself discloses (Cables and its accessories, Constructions and infrastructure,',
 'Electrical products and digital solutions) are pasted for FY2025 and grown on their own drivers; only its',
 'OUTPUT is carried here and everything downstream of it is formula. Third, whole-model engine outputs, where',
 'each figure is a complete re-run of the entire valuation and so cannot be a single formula: the Monte Carlo',
 'price map, the sensitivity grids, the DCF scenario bear/bull bounds, the multi-leg currency-of-discounting',
 'alternative (its USD cost of capital IS a live formula; the leg-by-leg USD discounting is the engine\'s),',
 'and the expert-panel legs. Everything else — including every lens base value, the relative/normalised/book',
 'bear and bull bounds, and the anchor-date roll — is a live formula. Changing a driver reprices the model',
 'but does NOT redraw the engine outputs.', '',
 'How revenue is built. Not as one growth rate. Each of the three disclosed segments is grown on its own',
 'driver — Cables on copper-price growth times FX-translation growth times a modest real-volume assumption,',
 'Constructions and Electrical products on a taper of their own recent revenue CAGR — because none of the',
 'audited filings discloses a tonnage, order-book or backlog figure to build a literal unit model from.',
 'Margins come from the same segment build; group EBITDA margin is an OUTPUT of it, not an input.', '',
 'What it is not. It is not investment advice, a recommendation, or a price target. Values are model outputs',
 'shown as ranges.', '',
 'Sourcing note, up front. FY2023, FY2024 and FY2025 all come from the company\'s own audited consolidated',
 'financial statements — every income-statement, balance-sheet and segment line is the audited figure, not a',
 'derivation or a triangulation. Segment revenue ties EXACTLY to consolidated revenue in every year (Note',
 '5-3); segment profit reconciles to consolidated operating profit through an explicit, exactly-reconciling',
 'corporate cost load (Note 16 less G&A, net impairment, other expenses and other income). Every input is',
 'annotated where it appears and listed with source and date in the companion bibliography document.', '',
 'Discount convention. Each explicit year is discounted at its own forward cost of capital, gliding',
 f"{W['wacc_exp']*100:.1f}% -> {W['wacc_term']*100:.1f}% on the same easing calendar as the interest forecast; the terminal value is",
 'capitalised at the terminal rate and discounted at the year-5 cumulative factor. One date, one price of time.',
 'The glide fractions are not a free parameter: they are the cost-of-debt path\'s own cumulative progress, and',
 'the DCF sheet computes them in front of you.', '',
 'The open question. This company earns just over half its revenue on a hard-currency-linked basis but',
 'reports, lists and borrows in Egyptian pounds. The primary model charges the full Egyptian cost of capital.',
 'The Fundamental Valuation sheet also shows what the same cash flows are worth if the hard-currency leg is',
 'discounted at a hard-currency rate. Both are shown; they are not averaged.', '',
 f"Currency. EGP million unless stated. Spot EGP {SPOT:.2f} ({M['asof']} close). Sheets: READ FIRST · Summary ·",
 'Fundamental Valuation · Assumptions · SOTP Bridge · Segments · Relative & Normalized · DCF · Income',
 'Statement · Balance Sheet · Cash Flow · Summary Financials · Monte Carlo · Sensitivity · Per-Share &',
 'Ratios · Peer & Sector.'], start=3):
    ws.cell(row=i, column=1, value=ln).font = Font(size=10)
ws.column_dimensions['A'].width = 112

# ============ 2 SUMMARY =======================================================
ws = sheet('Summary')
title(ws, 'Summary — valuation at a glance', 'All values link live to their source sheets', 7,
      awidth=44, cwidth=15)
hdr(ws, 4, ['Lens', 'Bear', 'Base', 'Bull', 'Role', '', 'vs price'])
LENS_SRC = {'dcf': '=DCF!C62', 'relative': "='Relative & Normalized'!C11",
            'normalized': "='Relative & Normalized'!C28", 'book': "='Relative & Normalized'!C36"}
BEAR_SRC = {'relative': "='Relative & Normalized'!C12",
            'normalized': "='Relative & Normalized'!E28", 'book': "='Relative & Normalized'!E36"}
BULL_SRC = {'relative': "='Relative & Normalized'!D12",
            'normalized': "='Relative & Normalized'!F28", 'book': "='Relative & Normalized'!F36"}
RETW = D['lens_record']['retired']['blend']
# The 'vs price' column divides by the market price, which sits at the FOOT of this block
# — below the rows that reference it — so its address is computed here from the layout and
# ASSERTED when the row is actually written. Retiring the blend added two rows, and on the
# sister study every absolute reference left behind silently re-pointed at the expert-panel
# median instead of the price: a formula naming a cell by address moves with the re-issue,
# and only an assertion notices when it does not [L-067].
SPOT_ROW = 5 + 4 + 7
ROLE = {'dcf': 'THE CENTRAL — the class primary', 'relative': 'cross-check',
        'normalized': 'REMOVED — not a lens this class publishes',
        'book': 'a disclosed floor, never weighted'}
r = 5
for k in ['dcf', 'relative', 'normalized', 'book']:
    l = LN[k]
    put(ws, f'A{r}', l['name'], fmt=None)
    if k in BEAR_SRC:
        putf(ws, f'B{r}', BEAR_SRC[k], l['bear'], PX, green=True)
    else:
        put(ws, f'B{r}', l['bear'], BLUE, PX)   # DCF bear/bull are whole-model scenario re-runs
    putf(ws, f'C{r}', LENS_SRC[k], l['base'], PX, green=True)
    if k in BULL_SRC:
        putf(ws, f'D{r}', BULL_SRC[k], l['bull'], PX, green=True)
    else:
        put(ws, f'D{r}', l['bull'], BLUE, PX)
    put(ws, f'E{r}', ROLE[k], fmt=None)
    putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', l['base'] / SPOT - 1, PCT)
    r += 1
band(ws, r, 7)
LK = ['dcf', 'relative', 'normalized', 'book']
# THE WEIGHT AND CONTRIBUTION COLUMNS WENT WITH THE BLEND, and a role column replaces
# them: which lens is the answer, which sit beside it, and which this class does not
# publish at all.
put(ws, f'A{r}', 'THE CENTRAL — the cash-flow lens, not an average', bold=True, fmt=None)
putf(ws, f'B{r}', '=B5', LN['dcf']['bear'], PX, bold=True)
putf(ws, f'C{r}', '=C5', D['central'], PX, bold=True)
putf(ws, f'D{r}', '=D5', LN['dcf']['bull'], PX, bold=True)
put(ws, f'E{r}', 'the class primary', fmt=None)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['central'] / SPOT - 1, PCT, bold=True)
r += 1
put(ws, f'A{r}', 'NOT AVERAGED — the retired 45/20/20/15 blend, published unused',
    bold=True, fmt=None)
putf(ws, f'C{r}', '=' + '+'.join('C%d*%g' % (5 + i, RETW[k]) for i, k in enumerate(LK)),
     D['retired_blend_value'], PX)
put(ws, f'E{r}', 'retired 04-Sep-2026', fmt=None)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['retired_blend_value'] / SPOT - 1, PCT)
r += 1
put(ws, f'A{r}', 'Span across the lenses (min/max) — a spread between METHODS, not a '
                 'range around the answer', fmt=None)
putf(ws, f'B{r}', '=MIN(B5:B8)', min(LN[k]['bear'] for k in LK), PX)
putf(ws, f'D{r}', '=MAX(D5:D8)', max(LN[k]['bull'] for k in LK), PX)
r += 2
put(ws, f'A{r}', 'Alternative reading — currency of discounting', fmt=None)
putf(ws, f'C{r}', "='Fundamental Valuation'!C18", DCF['ccy_alt_ps'], PX, green=True)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', DCF['ccy_alt_ps'] / SPOT - 1, PCT)
r += 1
put(ws, f'A{r}', 'Terminal value share of DCF enterprise value', fmt=None)
putf(ws, f'C{r}', '=DCF!C28', DCF['tv_share'], PCT, green=True)
r += 1
put(ws, f'A{r}', 'Expert panel median', fmt=None)
putf(ws, f'C{r}', "='Fundamental Valuation'!C26", D['panel_centre'], PX, green=True)
putf(ws, f'G{r}', f'=C{r}/$C${SPOT_ROW}-1', D['panel_centre'] / SPOT - 1, PCT)
r += 1
band(ws, r, 7)
assert r == SPOT_ROW, (f'market price landed on row {r}, not the row {SPOT_ROW} that every '
                       f'"vs price" formula on this sheet divides by')
put(ws, f'A{r}', 'Market price (anchor)', bold=True, fmt=None)
put(ws, f'C{r}', SPOT, BLUE, PX, bold=True)     # row 14
r += 2
hdr(ws, r, ['Key figure', 'Value'])
KEY = [('Shares outstanding (mn)', 'SHARES', SH, NUM0),
       ('Market capitalisation (EGP mn)', 'MKTCAP', M['mktcap'], NUM0),
       ('Net bank debt, FY2025 (EGP mn)', "='Balance Sheet'!D17", IN['nd_fy25'], NUM0),
       ('FY2025 revenue (EGP mn)', "='Income Statement'!D5", HI['FY25']['rev'], NUM0),
       ('FY2025 EBITDA (EGP mn)', "='Income Statement'!D7", HI['FY25']['ebitda'], NUM0),
       ('FY2025 attributable profit (EGP mn)', "='Income Statement'!D17", HI['FY25']['npa'], NUM0),
       ('Cost of capital — explicit window', '=DCF!C46', W['wacc_exp'], PCT),
       ('Cost of capital — terminal', '=DCF!C53', W['wacc_term'], PCT),
       ('Terminal growth', '=DCF!C23', IN['g_term'], PCT)]
SUMMARY_KEY_START = r + 1

# ============ 3 FUNDAMENTAL VALUATION =========================================
ws = sheet('Fundamental Valuation')
title(ws, 'Fundamental valuation — the four lenses and the alternative reading', None, 6,
      awidth=52, cwidth=15)
hdr(ws, 4, ['Lens / step', 'Basis', 'EGP per share'])
rows = [
    ('Discounted cash flow', 'links to the DCF sheet', "=DCF!C62", DCF['ps']),
    ('  bear', 'margin −1.5pp, weaker currency path, +2pp cost of capital, g 3%', LN['dcf']['bear'], None),
    ('  bull', 'margin +1.5pp, stronger currency path, −2pp cost of capital, g 6%', LN['dcf']['bull'], None),
    ('Relative multiples', f"{IN['ev_ebitda_just']}x mid-cycle EV/EBITDA",
     "='Relative & Normalized'!C11", LN['relative']['base']),
    ('Normalised earnings power', f"{IN['pe_just']}x normalised earnings per share",
     "='Relative & Normalized'!C28", LN['normalized']['base']),
    ('Book value and sustainable return', 'justified price-to-book on sustainable return on equity',
     "='Relative & Normalized'!C36", LN['book']['base']),
]
r = 5
for a_, b_, c_, xp in rows:
    put(ws, f'A{r}', a_, fmt=None); put(ws, f'B{r}', b_, fmt=None)
    if isinstance(c_, str):
        putf(ws, f'C{r}', c_, xp, PX, green=True)
    else:
        put(ws, f'C{r}', c_, BLUE, PX)
    r += 1
r += 1
band(ws, r, 3)
put(ws, f'A{r}', 'THE CENTRAL — the cash-flow lens itself, not an average of the four',
    bold=True, fmt=None)
putf(ws, f'C{r}', '=Summary!C9', D['central'], PX, bold=True, green=True)
r += 2
put(ws, f'A{r}', 'THE CURRENCY-OF-DISCOUNTING QUESTION', bold=True, fmt=None); r += 1
for lab, val, fmt, xp in [
        ('Share of forecast revenue earned in hard currency', F['fgn_egp'][-1] / F['rev'][-1], PCT, None),
        ('Cost of capital applied in the primary model (explicit → terminal)',
         f"{W['wacc_exp']*100:.1f}% → {W['wacc_term']*100:.1f}%", None, None),
        ('Hard-currency cost of capital for the foreign leg', W['wacc_usd_alt'], PCT, None),
        ('Fair value if the foreign leg is discounted at the hard-currency rate', DCF['ccy_alt_ps'], PX, None),
        ('Market price', f'=Summary!C{SPOT_ROW}', PX, SPOT)]:
    put(ws, f'A{r}', lab, fmt=None)
    if xp is not None:
        putf(ws, f'C{r}', val, xp, fmt, green=True)
    else:
        put(ws, f'C{r}', val, BLUE, fmt)
    r += 1   # the currency alternative lands on row 18
r += 1
put(ws, f'A{r}', 'EXPERT PANEL', bold=True, fmt=None); r += 1
hdr(ws, r, ['Expert', 'Method', 'Base (EGP/share)', 'Low', 'High']); r += 1
for k, nm in [('e1', 'Expert 1'), ('e2', 'Expert 2'), ('e3', 'Expert 3')]:
    e = EXP[k]
    put(ws, f'A{r}', nm, fmt=None); put(ws, f'B{r}', e['method_short'], fmt=None)
    put(ws, f'C{r}', e['base'], BLUE, PX); put(ws, f'D{r}', e['rng'][0], BLUE, PX)
    put(ws, f'E{r}', e['rng'][1], BLUE, PX); r += 1
band(ws, r, 5); put(ws, f'A{r}', 'Panel median', bold=True, fmt=None)
putf(ws, f'C{r}', '=MEDIAN(C23:C25)', D['panel_centre'], PX, bold=True)   # row 26

# ============ 4 ASSUMPTIONS ====================================================
ws = sheet('Assumptions')
title(ws, 'Assumptions — every input in the model', 'Blue cells are inputs. Change one and the '
      'model reprices: everything downstream is a formula.', 8, awidth=52, cwidth=13)
r = 4
A = {}          # key -> row on Assumptions

def block(name, items):
    global r
    band(ws, r, 8); put(ws, f'A{r}', name, bold=True, fmt=None); r += 1
    for key, lab, val, fmt in items:
        put(ws, f'A{r}', lab, fmt=None)
        if isinstance(val, (list, tuple)):
            for i, v in enumerate(val):
                put(ws, f'{get_column_letter(2+i)}{r}', v, BLUE, fmt)
        else:
            put(ws, f'C{r}', val, BLUE, fmt)
        A[key] = r
        r += 1
    r += 1

def a(key, i=None):
    """Absolute reference to an Assumptions cell; i selects a year column for list inputs."""
    col = get_column_letter(2 + i) if i is not None else 'C'
    return f"Assumptions!${col}${A[key]}"

hdr(ws, 3, ['Input', YF[0], YF[1], YF[2], YF[3], YF[4]])
block('Anchors', [
    ('spot', 'Spot price (EGP)', SPOT, PX),
    ('shares', 'Shares outstanding (mn)', SH, NUM0),
    ('tax_eff', 'Effective tax rate', IN['tax_eff'], PCT),
    ('tax_stat', 'Statutory corporate tax rate', IN['tax_stat'], PCT),
    ('fx_fy25', 'FY2025 average USD/EGP', IN['fx_hist']['FY25'], NUM1)])
block('Revenue drivers — the three disclosed segments', [
    ('copper', 'Copper (USD/tonne)', IN['copper_fcst'], NUM0),
    ('fx_path', 'USD/EGP path', IN['fx_path'], NUM1),
    ('cab_real_g', 'Cables — real (volume) growth over copper x FX', IN['cables_real_growth'], PCT),
    ('con_g', 'Constructions and infrastructure — revenue growth', IN['construct_growth'], PCT),
    ('ele_g', 'Electrical products and digital solutions — revenue growth', IN['elecprod_growth'],
     PCT)])
block('Segment gross margins and corporate cost load', [
    ('cab_mgn', 'Cables — gross/segment margin', IN['cables_margin'], PCT),
    ('con_mgn', 'Constructions and infrastructure — segment margin', IN['construct_margin'], PCT),
    ('ele_mgn', 'Electrical products and digital solutions — segment margin', IN['elecprod_margin'],
     PCT),
    ('opex_pct', 'Corporate cost load (% of revenue) — the bridge from segment profit to EBIT',
     IN['opex_pct'], PCT)])
block('Capital intensity', [
    ('nwc_pct', 'Working capital / revenue', IN['nwc_pct'], PCT),
    ('capex_pct', 'Capital expenditure / revenue', IN['capex_pct'], PCT),
    ('dna_pct', 'Depreciation and amortisation / revenue', IN['dna_pct'], PCT)])
block('Cost of capital', [
    ('rf', 'Risk-free rate (10-year local currency)', IN['rf'], PCT),
    ('sov', 'Sovereign default spread (netted out)', IN['sov_spread_cds'], PCT),
    ('erp', 'Equity risk premium', IN['erp_cds'], PCT),
    ('beta', 'Beta', IN['beta'], '0.000'),
    ('kd', 'Cost of debt, blended', IN['kd'], PCT),
    ('kd_path', 'Cost of debt path', IN['kd_path'], PCT),
    ('rf_term', 'Terminal risk-free rate', IN['rf_term'], PCT),
    ('erp_term', 'Terminal equity risk premium', IN['erp_term'], PCT),
    ('kd_term', 'Terminal cost of debt', IN['kd_term'], PCT),
    ('wd_term', 'Terminal debt weight', IN['wd_term'], PCT),
    ('g_term', 'Terminal growth', IN['g_term'], PCT)])
block('Balance-sheet and bridge anchors', [
    ('nd_fy25', 'Net bank debt at FY2025 (EGP mn, disclosed)', IN['nd_fy25'], NUM0),
    ('assoc_bv', 'Equity-accounted investees at carrying value (EGP mn)', IN['assoc_bv_fy25'], NUM0),
    ('intang', 'Intangible assets and goodwill (EGP mn)', IN['intang_fy25'], NUM0),
    ('pat_fy25', 'FY2025 profit after tax (EGP mn, disclosed)', IN['pat_fy25'], NUM0),
    ('npa_fy25', 'FY2025 profit after minority interests (EGP mn, disclosed)', IN['npa_fy25'], NUM0),
    ('dps_fy24', 'FY2024 dividend per share (EGP)', IN['dps_fy24'], PX),
    ('dps_fy25', 'FY2025 dividend per share (EGP, ratified 6 May 2026, paid 4 June 2026)',
     IN['dps_fy25'], PX),
    ('payout', 'Forecast dividend payout ratio (struck at the actual FY2025 rate)', PAYOUT, PCT),
    ('assoc_g', 'Growth in the share of equity-accounted investees', ASSOC_G, PCT),
    ('cash_yield', 'Yield assumed on surplus cash (blend of EGP deposit and hard-currency rates)',
     0.10, PCT),
    ('anchor_days', 'Days from the 31-Dec-2025 valuation date to the 5-Aug-2026 anchor',
     IN['anchor_days'], NUM0)])
block('Currency-of-discounting alternative', [
    ('usd_rf', 'US dollar risk-free rate', IN['usd_rf'], PCT),
    ('usd_erp', 'Hard-currency-leg equity risk premium', IN['usd_erp'], PCT),
    ('usd_kd', 'US dollar cost of debt', IN['usd_kd'], PCT),
    ('usd_wd', 'Debt weight, USD leg', IN['usd_wd'], PCT),
    ('usd_g', 'Terminal growth of the USD leg', IN['usd_g_term'], PCT)])
block('Lens inputs', [
    ('ev_ebitda_just', 'Justified EV/EBITDA', IN['ev_ebitda_just'], MULT),
    ('pe_just', 'Justified price/earnings', IN['pe_just'], MULT),
    ('roe_sust', 'Sustainable return on equity', IN['roe_sust'], PCT),
    ])
# The four lens-weight rows that stood here went with the blend. They are REMOVED rather
# than zeroed: a weight of zero is still a weight, and four zeros read as a scheme
# somebody switched off rather than one that no longer exists. The retired weights are
# published once, in the Summary's own labelled retired row.

# now that the Assumptions addresses exist, finish the Summary key-figure block
ws = wb['Summary']
rr = SUMMARY_KEY_START
SHARES_ROW = SUMMARY_KEY_START            # the first key figure is the share count
for lab, fml, val, fmt in KEY:
    put(ws, f'A{rr}', lab, fmt=None)
    if fml == 'SHARES':
        fml = f'={a("shares")}'
    elif fml == 'MKTCAP':
        fml = f'=$C${SPOT_ROW}*C{SHARES_ROW}'
    put(ws, f'A{rr}', lab, fmt=None)
    putf(ws, f'C{rr}', fml, val, fmt, green=True)
    rr += 1
ANCH['summary_shares'] = f'C{SHARES_ROW}'
ANCH['summary_mktcap'] = f'C{SHARES_ROW + 1}'

# ...and upgrade the Fundamental Valuation sheet's hard-currency cost of capital from a
# pasted value to a live formula off the newly-addressable USD-alternative inputs
ws = wb['Fundamental Valuation']
putf(ws, 'C17',
     f"=(1-{a('usd_wd')})*({a('usd_rf')}+{a('beta')}*{a('usd_erp')})"
     f"+{a('usd_wd')}*{a('usd_kd')}*(1-{a('tax_eff')})", W['wacc_usd_alt'], PCT2, green=True)

# ============ 5 SOTP BRIDGE ====================================================
ws = sheet('SOTP Bridge')
title(ws, 'Enterprise value to equity — the bridge', None, 5, awidth=52, cwidth=16)
hdr(ws, 4, ['Step', 'EGP mn', 'Per share (EGP)'])
brows = [('Present value of the five forecast years', '=DCF!C26', DCF['pv_explicit']),
         ('Present value of the terminal value', '=DCF!C27', DCF['pv_tv']),
         ('Enterprise value', '=C5+C6', DCF['ev']),
         ('Less net bank debt', f'=-{a("nd_fy25")}', -IN['nd_fy25']),
         ('Plus equity-accounted investees at carrying value', f'={a("assoc_bv")}', DCF['assoc']),
         ('Equity before minority interests', '=C7+C8+C9', DCF['ev'] - IN['nd_fy25'] + DCF['assoc']),
         ('Less minority interests at their share of group profit', '=-C10*$C$15', -DCF['nci_val']),
         ('Equity attributable to shareholders', '=C10+C11', DCF['eq_attr'])]
r = 5
for lab, v, xp in brows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, NUM0, bold=(r in (7, 12)),
         green=v.startswith(('=DCF', '=Assumptions', '=-Assumptions')))
    putf(ws, f'D{r}', f'=C{r}/{a("shares")}', xp / SH, PX, bold=(r in (7, 12)))
    r += 1
band(ws, 12, 4)
r += 1
put(ws, f'A{r}', 'Terminal value as a share of enterprise value', fmt=None)
putf(ws, f'C{r}', '=DCF!C28', DCF['tv_share'], PCT, green=True)
r += 1
put(ws, f'A{r}', 'Minority share of group profit', fmt=None)
putf(ws, f'C{r}', f'=({a("pat_fy25")}-{a("npa_fy25")})/{a("pat_fy25")}', NCI_SH, PCT)  # row 15

# ============ 6 SEGMENTS =======================================================
ws = sheet('Segments')
title(ws, 'Segments — the margin build', 'FY2025 disclosed shares and gross margins; forecast '
      'revenue and gross profit come from the unit build, and EBITDA is computed from them', 9,
      awidth=34, cwidth=13)
hdr(ws, 4, ['Segment', 'FY2025 revenue', 'Share', 'Segment margin'] + YF)
r = 5
_seg_rev_tot_row = r + len(SEGS)
for s in SEGS:
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    put(ws, f'B{r}', SEG['rev'][s], BLUE, NUM0)
    putf(ws, f'C{r}', f'=B{r}/$B${_seg_rev_tot_row}', SEG['rev'][s] / IN['rev_fy25'], PCT)
    put(ws, f'D{r}', SEG['gp_margin'][s], BLUE, PCT)
    for i in range(5):
        put(ws, f'{get_column_letter(5+i)}{r}', F['seg_rev'][i][s], BLUE, NUM0)
    r += 1
_last = r - 1                                   # 11
band(ws, r, 9); put(ws, f'A{r}', 'Total revenue', bold=True, fmt=None)
putf(ws, f'B{r}', f'=SUM(B5:B{_last})', IN['rev_fy25'], NUM0, bold=True)
putf(ws, f'C{r}', f'=SUM(C5:C{_last})', 1.0, PCT, bold=True)
for i in range(5):
    col = get_column_letter(5 + i)
    putf(ws, f'{col}{r}', f'=SUM({col}5:{col}{_last})', F['rev'][i], NUM0, bold=True)
REV_TOT = r                                     # 12
r += 2
hdr(ws, r, ['Segment profit by segment (Note 16 basis)'] + YF); r += 1
first_g = r                                     # 15
for s in SEGS:
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    for i in range(5):
        put(ws, f'{CD[i]}{r}', F['seg_gp'][i][s], BLUE, NUM0)
    r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Group segment profit', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'=SUM({CD[i]}{first_g}:{CD[i]}{r-1})', BU['gp'][i], NUM0, bold=True)
GP_TOT = r                                      # 22
r += 1
put(ws, f'A{r}', 'Group segment-profit margin', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{GP_TOT}/{get_column_letter(5+i)}${REV_TOT}',
         BU['gp'][i] / F['rev'][i], PCT)
r += 2
hdr(ws, r, ['Segment EBIT contribution — segment profit less the pro-rata corporate load'] + YF)
r += 1
first_e = r
for j, s in enumerate(SEGS):
    put(ws, f'A{r}', SEG['names'][s], fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}',
             f'={CD[i]}{first_g + j}-{a("opex_pct", i)}*{get_column_letter(5+i)}{5 + j}',
             F['seg_ebit'][i][s], NUM0)
    r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Group EBIT', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'=SUM({CD[i]}{first_e}:{CD[i]}{r-1})', F['ebit'][i], NUM0, bold=True)
EBIT_TOT = r
r += 1
put(ws, f'A{r}', 'Add back depreciation and amortisation', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={a("dna_pct")}*{get_column_letter(5+i)}{REV_TOT}',
         F['dna'][i], NUM0)
r += 1
band(ws, r, 6); put(ws, f'A{r}', 'Group EBITDA', bold=True, fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{EBIT_TOT}+{CD[i]}{r-1}', F['ebitda'][i], NUM0, bold=True)
EBITDA_TOT = r
r += 1
put(ws, f'A{r}', 'Group EBITDA margin', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}{r}', f'={CD[i]}{EBITDA_TOT}/{get_column_letter(5+i)}${REV_TOT}',
         F['ebitda_margin'][i], PCT)
put(ws, f'A{r+1}', 'The corporate load is stated on the segment-profit-to-EBIT basis — the same '
    'basis as the audited historical bridge — so EBIT falls out first and EBITDA is EBIT plus '
    'D&A, not the other way round.', fmt=None).font = SUB
ANCH.update(seg_rev_tot=REV_TOT, seg_gp_tot=GP_TOT, seg_ebitda_tot=EBITDA_TOT, seg_ebitda_mgn=r)

# ============ 7 RELATIVE & NORMALIZED ==========================================
ws = sheet('Relative & Normalized')
title(ws, 'Relative multiples and normalised earnings power', None, 5, awidth=52, cwidth=16)
hdr(ws, 4, ['Relative lens', 'Value'])
ebitda_mid = REL['ebitda_mid']
ev_rel_fwd = IN['ev_ebitda_just'] * ebitda_mid
rel_rows = [
    ('FY2027E EBITDA (EGP mn)', '=DCF!C6', ebitda_mid, NUM0),
    ('Justified enterprise value / EBITDA', f'={a("ev_ebitda_just")}', IN['ev_ebitda_just'], MULT),
    ('Implied enterprise value AS AT end-FY2027 (EGP mn)', '=C5*C6', ev_rel_fwd, NUM0),
    ('Discount factor back to the valuation date (year-2)', '=DCF!C16', F['df'][1], DF4),
    ('Plus present value of the interim FY26-27 free cash flows (EGP mn)', '=DCF!B17+DCF!C17',
     REL['pv_interim'], NUM0),
    ('Implied enterprise value at 31-Dec-2025 (EGP mn)', '=C7*C8+C9', REL['ev_rel'], NUM0),
    ('Implied value per share, rolled to the anchor (EGP)',
     f"=((C10-{a('nd_fy25')}+{a('assoc_bv')})*(1-'SOTP Bridge'!$C$15)/{a('shares')})"
     f"*DCF!$C$61-{a('dps_fy25')}", LN['relative']['base'], PX)]
r = 5
for lab, v, xp, fmt in rel_rows:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=('DCF' in v or v.startswith('=Assumptions')))
    r += 1
band(ws, 11, 3)
put(ws, 'A12', 'Bear at 5.5× (C) / bull at 8.0× (D), same construction', fmt=None)
for cell, mult, xp in (('C12', 5.5, LN['relative']['bear']), ('D12', 8.0, LN['relative']['bull'])):
    putf(ws, cell,
         f"=(({mult}*C5*C8+C9-{a('nd_fy25')}+{a('assoc_bv')})*(1-'SOTP Bridge'!$C$15)"
         f"/{a('shares')})*DCF!$C$61-{a('dps_fy25')}", xp, PX)
r = 13
mktcap_f = f'({a("spot")}*{a("shares")})'
for lab, v, xp, fmt in [
        ('Trailing enterprise value / EBITDA',
         f"=({mktcap_f}+{a('nd_fy25')})/'Income Statement'!D7", REL['ev_ebitda_trailing'], MULT),
        ('Trailing price / earnings', f"={a('spot')}/'Income Statement'!D18", REL['pe_trailing'], MULT),
        ('Trailing price / book', f"={a('spot')}/C31", SPOT / BK['bvps'], MULT),
        ('Net bank debt / EBITDA', f"={a('nd_fy25')}/'Income Statement'!D7",
         IN['nd_fy25'] / HI['FY25']['ebitda'], MULT)]:
    put(ws, f'A{r}', lab, fmt=None); putf(ws, f'C{r}', v, xp, fmt); r += 1
r += 1
hdr(ws, r, ['Normalised earnings lens — mid-cycle margin (FY2028E) at current (FY2026E) scale',
            'Value'])
r += 1                                            # r = 19
norm_ebitda = NRM['margin'] * NRM['rev']
for lab, v, xp, fmt in [
        ('Current-scale revenue (FY2026E, EGP mn)', '=DCF!B5', NRM['rev'], NUM0),
        ('Mid-cycle EBITDA margin (FY2028E)', '=DCF!D7', NRM['margin'], PCT),
        ('Normalised EBITDA (EGP mn)', '=C19*C20', norm_ebitda, NUM0),
        ('Normalised EBIT (EGP mn)', f'=C21-C19*{a("dna_pct")}', NRM['ebit'], NUM0),
        ('Net finance cost (FY2026E, EGP mn)', "='Income Statement'!E11", -NRM['interest'], NUM0),
        ('Share of equity-accounted investees (FY2026E, EGP mn)', "='Income Statement'!E12",
         NRM['assoc'], NUM0),
        ('Normalised attributable earnings (EGP mn)',
         f"=(C22+C23+C24)*(1-{a('tax_eff')})*(1-'SOTP Bridge'!$C$15)", NRM['np'], NUM0),
        ('Normalised earnings per share (EGP)', f'=C25/{a("shares")}', NRM['eps'], PX),
        ('Justified price / earnings', f'={a("pe_just")}', IN['pe_just'], MULT),
        ('Implied value per share, rolled to the anchor (EGP)',
         f'=C26*C27*DCF!$C$61-{a("dps_fy25")}', LN['normalized']['base'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=('DCF' in v or 'Income Statement' in v))
    r += 1
band(ws, r - 1, 3)                                 # implied value lands on row 28
put(ws, 'D28', 'bear 7.0× (E) / bull 11.5× (F):', fmt=None)
putf(ws, 'E28', f'=C26*7*DCF!$C$61-{a("dps_fy25")}', LN['normalized']['bear'], PX)
putf(ws, 'F28', f'=C26*11.5*DCF!$C$61-{a("dps_fy25")}', LN['normalized']['bull'], PX)
r += 1
hdr(ws, r, ['Book lens', 'Value']); r += 1         # r = 31
for lab, v, xp, fmt in [
        ('Book value per share (EGP)', f"='Balance Sheet'!D14/{a('shares')}", BK['bvps'], PX),
        ('Sustainable return on equity', f'={a("roe_sust")}', BK['roe_sust'], PCT),
        ('Trailing return on equity',
         "='Income Statement'!D17/(('Balance Sheet'!C14+'Balance Sheet'!D14)/2)",
         BK['roe_trailing'], PCT),
        ('Perpetual (terminal) cost of equity — a steady-state multiple takes a steady-state rate',
         '=DCF!C49', BK['ke_blend'], PCT),
        ('Justified price / book', f'=(C32-{a("g_term")})/(C34-{a("g_term")})', BK['pb_just'], MULT),
        ('Implied value per share, rolled to the anchor (EGP)',
         f'=C31*C35*DCF!$C$61-{a("dps_fy25")}', LN['book']['base'], PX)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt,
         green=('DCF' in v or 'Balance Sheet' in v or 'Income Statement' in v))
    r += 1
band(ws, r - 1, 3)                                 # implied value lands on row 36
put(ws, 'D36', 'bear / bull constructions (E / F):', fmt=None)
putf(ws, 'E36', f"=(({a('roe_sust')}-0.03)/((DCF!C39+DCF!C49)/2-0.03))*C31"
     f"*DCF!$C$61-{a('dps_fy25')}", LN['book']['bear'], PX)
putf(ws, 'F36', f"=(({a('roe_sust')}+0.02-{a('g_term')})/(DCF!C49-{a('g_term')}))*C31"
     f"*DCF!$C$61-{a('dps_fy25')}", LN['book']['bull'], PX)

# ============ 8 DCF =============================================================
ws = sheet('DCF')
title(ws, 'Discounted cash flow — the full waterfall', 'Every line is a live formula: the cost of '
      'capital is built below, the glide is derived from the cost-of-debt path, and the terminal '
      'value is capitalised at the terminal rate and discounted at the year-5 factor', 6,
      awidth=46, cwidth=15)
hdr(ws, 4, ['EGP mn'] + YF)

def wf(r, lab, fmls, vals, fmt=NUM0, bd=False, green=False):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i in range(5):
        putf(ws, f'{CD[i]}{r}', fmls(i), vals[i], fmt, bold=bd, green=green)
    if bd: band(ws, r, 6)

wf(5, 'Revenue', lambda i: f'=Segments!{get_column_letter(5+i)}{REV_TOT}', F['rev'], green=True)
wf(6, 'EBITDA', lambda i: f'=Segments!{CD[i]}{EBITDA_TOT}', F['ebitda'], green=True)
wf(7, 'EBITDA margin', lambda i: f'={CD[i]}6/{CD[i]}5', F['ebitda_margin'], PCT)
wf(8, 'Less depreciation and amortisation', lambda i: f'=-{CD[i]}5*{a("dna_pct")}',
   [-x for x in F['dna']])
wf(9, 'EBIT', lambda i: f'={CD[i]}6+{CD[i]}8', F['ebit'], bd=True)
wf(10, f"NOPAT — EBIT x (1 - {IN['tax_eff']:.0%})", lambda i: f'={CD[i]}9*(1-{a("tax_eff")})',
   F['nopat'])
wf(11, 'Add back depreciation and amortisation', lambda i: f'=-{CD[i]}8', F['dna'])
wf(12, 'Less capital expenditure', lambda i: f'=-{CD[i]}5*{a("capex_pct", i)}',
   [-x for x in F['capex']])
wf(13, 'Less change in working capital',
   lambda i: (f"=-({CD[i]}5*{a('nwc_pct')}-'Balance Sheet'!D16)" if i == 0
              else f'=-({CD[i]}5-{CD[i-1]}5)*{a("nwc_pct")}'), [-x for x in F['dnwc']])
wf(14, 'Free cash flow to the firm',
   lambda i: f'={CD[i]}10+{CD[i]}11+{CD[i]}12+{CD[i]}13', F['fcff'], bd=True)
wf(15, 'Forward cost of capital', lambda i: f'=$C$46-($C$46-$C$53)*{CD[i]}57', F['fwd_wacc'], PCT2)
wf(16, 'Discount factor', lambda i: (f'=1/(1+{CD[i]}15)' if i == 0
                                     else f'={CD[i-1]}16/(1+{CD[i]}15)'), F['df'], DF4)
wf(17, 'Present value of FCFF', lambda i: f'={CD[i]}14*{CD[i]}16', F['pv'], bd=True)

put(ws, 'A19', 'TERMINAL VALUE, BRIDGE AND THE ANCHOR ROLL', bold=True, fmt=None)
nopat_grown = F['nopat'][-1] * (1 + IN['g_term'])
tv_block = [
    ('Terminal-year NOPAT grown one year (EGP mn)', '=F10*(1+C23)', nopat_grown, NUM0),
    ('Terminal return on invested capital', "=F10*(1+C23)/'Summary Financials'!I13",
     DCF['roic_term'], PCT),
    ('Required reinvestment rate (g / return on capital)', '=C23/C21', DCF['rr_term'], PCT),
    ('Terminal growth', f'={a("g_term")}', IN['g_term'], PCT),
    ('Terminal cost of capital', '=C53', W['wacc_term'], PCT),
    ('Terminal value — terminal-year FCFF C20×(1−C22), capitalised (EGP mn)',
     '=C20*(1-C22)/(C24-C23)', DCF['tv'], NUM0),
    ('Present value of the five forecast years (EGP mn)', '=SUM(B17:F17)', DCF['pv_explicit'], NUM0),
    ('Present value of the terminal value (EGP mn)', '=C25*F16', DCF['pv_tv'], NUM0),
    ('Terminal value as a share of enterprise value', '=C27/(C26+C27)', DCF['tv_share'], PCT),
    ('Enterprise value (EGP mn)', '=C26+C27', DCF['ev'], NUM0),
    ('Equity attributable to shareholders (EGP mn)', "='SOTP Bridge'!C12", DCF['eq_attr'], NUM0),
    ('Fair value per share at 31-Dec-2025 (EGP)', f'=C30/{a("shares")}', DCF['ps_dec'], PX)]
r = 20
for lab, v, xp, fmt in tv_block:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt,
         green=('SOTP' in v or 'Summary Financials' in v or v.startswith('=Assumptions')))
    r += 1
band(ws, 31, 4)

put(ws, 'A33', 'COST OF CAPITAL — BUILT HERE, NOT ASSUMED', bold=True, fmt=None)
coc = [
    ('Risk-free rate, 10-year local currency', f'={a("rf")}', IN['rf'], PCT2),
    ('Less sovereign default spread (removed to avoid double-counting)', f'={a("sov")}',
     IN['sov_spread_cds'], PCT2),
    ('Risk-free rate net of the sovereign spread', '=C34-C35', W['rf_star'], PCT2),
    ('Beta', f'={a("beta")}', IN['beta'], '0.000'),
    ('Equity risk premium', f'={a("erp")}', IN['erp_cds'], PCT2),
    ('Cost of equity, explicit window', '=C36+C37*C38', W['ke_exp'], PCT2),
    ('Cost of debt, blended', f'={a("kd")}', IN['kd'], PCT2),
    ('Cost of debt after tax', f'=C40*(1-{a("tax_eff")})', W['kd_at'], PCT2),
    ('Market capitalisation (EGP mn)', f'={a("spot")}*{a("shares")}', SPOT * SH, NUM0),
    ('Net bank debt (EGP mn)', f'={a("nd_fy25")}', IN['nd_fy25'], NUM0),
    ('Debt weight (net debt / (net debt + market capitalisation))', '=C43/(C43+C42)',
     W['wd_exp'], PCT2),
    ('Equity weight', '=1-C44', W['we_exp'], PCT2),
    ('Cost of capital, explicit window', '=C45*C39+C44*C41', W['wacc_exp'], PCT2),
    ('Terminal risk-free rate', f'={a("rf_term")}', IN['rf_term'], PCT2),
    ('Terminal equity risk premium', f'={a("erp_term")}', IN['erp_term'], PCT2),
    ('Terminal cost of equity', '=C47+C37*C48', W['ke_term'], PCT2),
    ('Terminal cost of debt', f'={a("kd_term")}', IN['kd_term'], PCT2),
    ('Terminal cost of debt after tax', f'=C50*(1-{a("tax_eff")})', W['kd_term_at'], PCT2),
    ('Terminal debt weight', f'={a("wd_term")}', IN['wd_term'], PCT2),
    ('Terminal cost of capital', '=(1-C52)*C49+C52*C51', W['wacc_term'], PCT2)]
r = 34
for lab, v, xp, fmt in coc:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'C{r}', v, xp, fmt, green=v.startswith('=Assumptions'))
    r += 1
band(ws, 46, 4); band(ws, 53, 4)
hdr(ws, 55, ['The glide — inherited from the cost-of-debt path, not invented'] + YF)
put(ws, 'A56', 'Cost of debt path', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}56', f'={a("kd_path", i)}', IN['kd_path'][i], PCT2, green=True)
put(ws, 'A57', 'Glide fraction — cumulative progress of the easing', fmt=None)
for i in range(5):
    putf(ws, f'{CD[i]}57', f'=($B$56-{CD[i]}56)/($B$56-$F$56)', F['glide_frac'][i], PCT2)
put(ws, 'A58', 'Note: row 15 above is the explicit-window cost of capital walked down to the '
    'terminal rate by the glide fraction on row 57, so the shape of the easing is inherited from '
    'the cost-of-debt path rather than being a second free parameter.', fmt=None).font = SUB
put(ws, 'A60', 'THE ANCHOR ROLL — one date, one price of time', bold=True, fmt=None)
put(ws, 'A61', 'Anchor accretion factor — (1 + cost of equity)^(days to anchor / 365)', fmt=None)
putf(ws, 'C61', f'=(1+C39)^({a("anchor_days")}/365)', DCF['roll'], DF4)
put(ws, 'A62', 'Fair value per share at the 5-Aug-2026 anchor (EGP)', fmt=None)
putf(ws, 'C62', f'=C31*C61-{a("dps_fy25")}', DCF['ps'], PX, bold=True)
band(ws, 62, 4)
put(ws, 'A63', 'The bridge on row 31 is dated 31-Dec-2025 (the audited balance-sheet date it '
    'subtracts net debt at). Row 62 rolls it to the anchor at the cost of equity, net of the '
    'EGP 1.85 FY2025 dividend paid in the window. Every lens on every sheet is rolled the same '
    'way.', fmt=None).font = SUB

# ============ 9 INCOME STATEMENT =================================================
ws = sheet('Income Statement')
title(ws, 'Income statement — three years historical, five years forecast', 'EGP mn, consolidated. '
      'History is the audited record; every forecast line is a formula', 9, awidth=40, cwidth=12)
hdr(ws, 4, ['EGP mn'] + YH + YF)

def isline(r, lab, hist, fc_f, fc_v, fmt=NUM0, bd=False, hist_f=None, hist_v=None, green=False):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i in range(3):
        if hist_f is not None:
            putf(ws, f'{HC[i]}{r}', hist_f(i), hist_v[i], fmt, bold=bd)
        else:
            put(ws, f'{HC[i]}{r}', hist[i], BLUE, fmt, bold=bd)
    if fc_f is not None:
        for i in range(5):
            putf(ws, f'{FCOL[i]}{r}', fc_f(i), fc_v[i], fmt, bold=bd, green=green)
    if bd: band(ws, r, 9)

pbt_f = [F['ebit'][i] - F['interest'][i] + F['assoc'][i] for i in range(5)]
pat_f = [x * (1 - TAX) for x in pbt_f]
isline(5, 'Revenue', [HI[y]['rev'] for y in H3], lambda i: f'=DCF!{CD[i]}5', F['rev'], bd=True,
       green=True)
isline(6, 'Gross profit', [HI[y]['gp'] for y in H3],
       lambda i: f'=Segments!{CD[i]}{GP_TOT}', BU['gp'], green=True)
isline(7, 'EBITDA', [HI[y]['ebitda'] for y in H3], lambda i: f'=DCF!{CD[i]}6', F['ebitda'],
       bd=True, green=True)
isline(8, 'EBITDA margin', None, lambda i: f'={FCOL[i]}7/{FCOL[i]}5', F['ebitda_margin'], PCT,
       hist_f=lambda i: f'={HC[i]}7/{HC[i]}5',
       hist_v=[HI[y]['ebitda'] / HI[y]['rev'] for y in H3])
isline(9, 'Depreciation and amortisation', [-abs(HI[y]['dna']) for y in H3],
       lambda i: f'=DCF!{CD[i]}8', [-x for x in F['dna']], green=True)
isline(10, 'EBIT', None, lambda i: f'={FCOL[i]}7+{FCOL[i]}9', F['ebit'], bd=True,
       hist_f=lambda i: f'={HC[i]}7+{HC[i]}9', hist_v=[HI[y]['ebit'] for y in H3])
isline(11, 'Net finance costs', [-abs(HI[y]['fin']) for y in H3],
       lambda i: (f"=-({a('kd_path', i)}*'Balance Sheet'!$D$11-{a('cash_yield')}*"
                  f"MAX('Balance Sheet'!$D$11-'Balance Sheet'!{'D' if i == 0 else FCOL[i-1]}17,0))"),
       [-x for x in F['interest']])
isline(12, 'Share of equity-accounted investees', [HI[y]['assoc'] for y in H3],
       lambda i: f'={"D" if i == 0 else FCOL[i-1]}12*(1+{a("assoc_g")})', F['assoc'])
isline(13, 'Profit before tax', None, lambda i: f'={FCOL[i]}10+{FCOL[i]}11+{FCOL[i]}12', pbt_f,
       hist_f=lambda i: f'={HC[i]}10+{HC[i]}11+{HC[i]}12', hist_v=[HI[y]['ebt'] for y in H3])
isline(14, 'Income tax', [-abs(HI[y]['tax']) for y in H3],
       lambda i: f'=-{FCOL[i]}13*{a("tax_eff")}', [-x * TAX for x in pbt_f])
# Profit after tax and profit after minority interests are DISCLOSED for all three
# historical years, so they are carried as reported rather than re-derived: closing the
# audited FY2023 account arithmetically lands 0.1 below the printed 11,138.0, and the
# printed figure is the primary record. The forecast columns are formulas.
isline(15, 'Profit for the year', [HI[y]['pat'] for y in H3],
       lambda i: f'={FCOL[i]}13+{FCOL[i]}14', pat_f)
isline(16, 'Non-controlling interests', [-abs(HI[y]['nci']) for y in H3],
       lambda i: f"=-{FCOL[i]}15*'SOTP Bridge'!$C$15", [-x * NCI_SH for x in pat_f])
isline(17, 'Profit attributable to shareholders', [HI[y]['npa'] for y in H3],
       lambda i: f'={FCOL[i]}15+{FCOL[i]}16', F['np_attr'], bd=True)
put(ws, 'A18', 'Earnings per share (EGP)', fmt=None)
npa_all = [HI[y]['npa'] for y in H3] + F['np_attr']
for i in range(8):
    putf(ws, f'{ALL[i]}18', f'={ALL[i]}17/{a("shares")}', npa_all[i] / SH, PX)
put(ws, 'A20', 'Note: every FY2023-25 statement line is the audited figure — no closure or '
    'derivation is used for any historical year (the EBITDA and EPS rows are labelled house '
    'derivations). In the forecast the finance charge is computed on the gross debt book less '
    'the cash the business is accumulating, so it moves with net debt rather than being frozen '
    'at the FY2025 balance; profit is therefore struck after interest and differs from the '
    'pre-financing DCF waterfall by construction.', fmt=None).font = SUB

# ============ 10 BALANCE SHEET ====================================================
ws = sheet('Balance Sheet')
title(ws, 'Balance sheet — condensed', 'EGP mn, consolidated. Every FY2023-25 line is the audited '
      'closing figure — no triangulation is needed for any year, including FY2025, because the full '
      'filing is in hand', 9, awidth=40, cwidth=12)
hdr(ws, 4, ['EGP mn'] + YH + YF)

def bsline(r, lab, key, fc_f=None, fc_v=None, bd=False, fmt=NUM0, d_f=None, d_v=None):
    put(ws, f'A{r}', lab, bold=bd, fmt=None)
    for i, y in enumerate(['FY23', 'FY24']):
        v = HB[y].get(key)
        put(ws, f'{HC[i]}{r}', v if v is not None else '-', BLUE, fmt, bold=bd)
    if d_f is not None:
        putf(ws, f'D{r}', d_f, d_v, fmt, bold=bd, green=str(d_f).startswith('=Assumptions'))
    else:
        v = HB['FY25'].get(key)
        put(ws, f'D{r}', v if v is not None else '-', BLUE, fmt, bold=bd)
    if fc_f is not None:
        for i in range(5):
            putf(ws, f'{FCOL[i]}{r}', fc_f(i), fc_v[i], fmt, bold=bd)
    if bd: band(ws, r, 9)

bsline(5, 'Property, plant and equipment', 'ppe',
       fc_f=lambda i: f'={"D" if i == 0 else FCOL[i-1]}5-DCF!{CD[i]}12+DCF!{CD[i]}8', fc_v=F['ppe'])
bsline(6, 'Inventories', 'inv')
bsline(7, 'Contract assets', 'ca')
bsline(8, 'Trade and other receivables', 'recv')
bsline(9, 'Cash and cash equivalents', 'cash', d_f='=D11-D17', d_v=HB['FY25']['cash'])
bsline(10, 'Total assets', 'assets', bd=True)
bsline(11, 'Loans and borrowings', 'debt')
bsline(12, 'Trade and other payables', 'pay')
bsline(13, 'Contract liabilities', 'cl')
bsline(14, 'Equity attributable to shareholders', 'eqp',
       fc_f=lambda i: (f'={"D" if i == 0 else FCOL[i-1]}14+'
                       f"'Income Statement'!{FCOL[i]}17*(1-{a('payout')})"), fc_v=F['equity'],
       bd=True)
bsline(15, 'Non-controlling interests', 'nci')
bsline(16, 'Net working capital', 'nwc',
       fc_f=lambda i: f"='Income Statement'!{FCOL[i]}5*{a('nwc_pct')}", fc_v=F['nwc'])
bsline(17, 'Net bank debt', 'nd',
       fc_f=lambda i: (f'={"D" if i == 0 else FCOL[i-1]}17-DCF!{CD[i]}14'
                       f"-'Income Statement'!{FCOL[i]}11*(1-{a('tax_eff')})"
                       f"+{a('payout')}*'Income Statement'!{FCOL[i]}17"), fc_v=F['net_debt'],
       d_f=f'={a("nd_fy25")}', d_v=IN['nd_fy25'], bd=True)
put(ws, 'A18', 'Net debt / EBITDA', fmt=None)
nd_all = [HB[y]['nd'] for y in H3] + F['net_debt']
eb_all = [HI[y]['ebitda'] for y in H3] + F['ebitda']
for i in range(8):
    putf(ws, f'{ALL[i]}18', f"={ALL[i]}17/'Income Statement'!{ALL[i]}7", nd_all[i] / eb_all[i], MULT)
put(ws, 'A19', 'Residual: other liabilities, provisions and deferred tax not shown separately '
    '(total assets less the lines above)', fmt=None)
for i in range(3):
    col, y = HC[i], H3[i]
    resid = HB[y]['assets'] - (HB[y]['debt'] + HB[y]['pay'] + HB[y]['cl'] + HB[y]['eqp'] + HB[y]['nci'])
    putf(ws, f'{col}19', f'={col}10-({col}11+{col}12+{col}13+{col}14+{col}15)', resid, NUM0)
put(ws, 'A21', 'Note: this is a CONDENSED layout, so it does not foot to zero — the residual row '
    'above is the block of other liabilities, provisions, deferred tax and related-party balances that '
    'is not shown separately. For FY2024 that residual is 22,828, which reconciles to the audited '
    'statements (provisions 13,440 + 943, deferred tax 3,670, related parties 2,050 + 95, other '
    'liabilities 2,631 = 22,829). Every line, including FY2025, is now the audited closing figure — '
    'no triangulation or roll-forward is used for any historical year.', fmt=None).font = SUB

# ============ 11 CASH FLOW =========================================================
ws = sheet('Cash Flow')
title(ws, 'Cash flow — historical markers and the forecast waterfall', 'EGP mn', 9,
      awidth=44, cwidth=12)
hdr(ws, 4, ['EGP mn', 'FY2024', 'FY2025'] + YF)
put(ws, 'A5', 'EBITDA', fmt=None)
putf(ws, 'B5', "='Income Statement'!C7", HI['FY24']['ebitda'], NUM0, green=True)
putf(ws, 'C5', "='Income Statement'!D7", HI['FY25']['ebitda'], NUM0, green=True)
for i in range(5):
    putf(ws, f'{CFF[i]}5', f"='Income Statement'!{FCOL[i]}7", F['ebitda'][i], NUM0, green=True)
for r, lab, v in [(6, 'Interest paid', -IN['int_paid_fy25']),
                  (7, 'Income tax paid', -IN['tax_paid_fy25']),
                  (8, 'Operating cash flow after interest and tax', IN['ocf_fy25'])]:
    put(ws, f'A{r}', lab, fmt=None)
    put(ws, f'B{r}', '-', BLACK, NUM0)
    put(ws, f'C{r}', v, BLUE, NUM0)
put(ws, 'A9', 'NOPAT', fmt=None)
put(ws, 'B9', '-', BLACK, NUM0); put(ws, 'C9', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}9', f'=DCF!{CD[i]}10', F['nopat'][i], NUM0, green=True)
put(ws, 'A10', 'Add back depreciation and amortisation', fmt=None)
put(ws, 'B10', '-', BLACK, NUM0); put(ws, 'C10', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}10', f'=DCF!{CD[i]}11', F['dna'][i], NUM0, green=True)
put(ws, 'A11', 'Capital expenditure', fmt=None)
put(ws, 'B11', -IN['capex_fy24'], BLUE, NUM0); put(ws, 'C11', -IN['capex_fy25'], BLUE, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}11', f'=DCF!{CD[i]}12', -F['capex'][i], NUM0, green=True)
put(ws, 'A12', 'Change in working capital', fmt=None)
put(ws, 'B12', '-', BLACK, NUM0)
putf(ws, 'C12', "=-('Balance Sheet'!D16-'Balance Sheet'!C16)",
     -(HB['FY25']['nwc'] - HB['FY24']['nwc']), NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}12', f'=DCF!{CD[i]}13', -F['dnwc'][i], NUM0, green=True)
put(ws, 'A13', 'Free cash flow to the firm', bold=True, fmt=None)
put(ws, 'B13', '-', BLACK, NUM0); put(ws, 'C13', '-', BLACK, NUM0)
for i in range(5):
    putf(ws, f'{CFF[i]}13', f'={CFF[i]}9+{CFF[i]}10+{CFF[i]}11+{CFF[i]}12', F['fcff'][i], NUM0,
         bold=True)
band(ws, 13, 9)
put(ws, 'A15', 'Cash conversion — operating cash flow as a share of EBITDA, FY2025', fmt=None)
putf(ws, 'C15', '=C8/C5', IN['ocf_fy25'] / HI['FY25']['ebitda'], PCT)
put(ws, 'A16', 'This single ratio is the crux of the valuation: the company earned far more '
    'EBITDA than it converted to cash, because growth was funded through working capital.',
    fmt=None).font = SUB

# ============ 12 SUMMARY FINANCIALS =================================================
ws = sheet('Summary Financials')
title(ws, 'Summary financials — the eight-year picture', 'EGP mn unless stated. Every cell on this '
      'sheet is a link or a ratio; nothing is typed twice', 9, awidth=40, cwidth=12)
hdr(ws, 4, ['EGP mn'] + YH + YF)
rev_all = [HI[y]['rev'] for y in H3] + F['rev']
ebit_all = [HI[y]['ebit'] for y in H3] + F['ebit']

def sfline(r, lab, fml, vals, fmt=NUM0, skip=()):
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{r}', '-', BLACK, fmt)
        else:
            f_ = fml(i)
            putf(ws, f'{ALL[i]}{r}', f_, vals[i], fmt,
                 green=f_.startswith(("='I", "='B", '=DCF', "='C")))

sfline(5, 'Revenue', lambda i: f"='Income Statement'!{ALL[i]}5", rev_all)
sfline(6, 'Revenue growth', lambda i: f'={ALL[i]}5/{ALL[i-1]}5-1',
       [None] + [rev_all[i] / rev_all[i - 1] - 1 for i in range(1, 8)], PCT, skip=(0,))
sfline(7, 'EBITDA', lambda i: f"='Income Statement'!{ALL[i]}7", eb_all)
sfline(8, 'EBITDA margin', lambda i: f'={ALL[i]}7/{ALL[i]}5',
       [eb_all[i] / rev_all[i] for i in range(8)], PCT)
sfline(9, 'EBIT', lambda i: f"='Income Statement'!{ALL[i]}10", ebit_all)
sfline(10, 'Attributable profit', lambda i: f"='Income Statement'!{ALL[i]}17", npa_all)
sfline(11, 'Free cash flow to the firm', lambda i: f"='Cash Flow'!{CFF[i-3]}13",
       [None] * 3 + F['fcff'], skip=(0, 1, 2))
sfline(12, 'Net bank debt', lambda i: f"='Balance Sheet'!{ALL[i]}17", nd_all)
sfline(13, 'Invested capital',
       lambda i: f"='Balance Sheet'!{ALL[i]}16+'Balance Sheet'!{ALL[i]}5+{a('intang')}",
       [None, None, F['ic_fy25']] + F['ic'], skip=(0, 1))
sfline(14, 'Return on invested capital', lambda i: f'=DCF!{CD[i-3]}10/{ALL[i]}13',
       [None] * 3 + F['roic'], PCT, skip=(0, 1, 2))
put(ws, 'A16', "Return on invested capital is the DCF's NOPAT over the same year's invested "
    'capital, which is why the FY2025 column carries capital but no return: FY2025 NOPAT is struck '
    'on a different, post-financing basis and is reported in the terminal-growth reconciliation on '
    'the Per-Share & Ratios sheet.', fmt=None).font = SUB

# ============ 13 MONTE CARLO ==========================================================
ws = sheet('Monte Carlo')
title(ws, 'Probabilistic price map', 'A map of price dispersion. It carries no view on value and '
      'is never blended with the valuation. Each figure is an engine output, not a formula.', 8,
      awidth=40, cwidth=14)
hdr(ws, 4, ['Horizon', '5th', '25th', 'Median', '75th', '95th', 'P(above spot)'])
r = 5
for tag in ('1M', '3M'):
    h = STK['horizons'][tag]
    put(ws, f'A{r}', f"{'One month' if tag=='1M' else 'Three months'} — to {h['grade_date']}", fmt=None)
    for i, k in enumerate(('p5', 'p25', 'p50', 'p75', 'p95')):
        put(ws, f'{get_column_letter(2+i)}{r}', h['pct'][k], BLUE, PX)
    put(ws, f'G{r}', h['p_above'], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Level event', 'One month', 'Three months']); r += 1
for lab, k in [('Finishes 10% or more above spot', 'p_up10'),
               ('Finishes 10% or more below spot', 'p_dn10'),
               ('Touches 10% above spot at any point', 'touch_up10'),
               ('Touches 10% below spot at any point', 'touch_dn10')]:
    put(ws, f'A{r}', lab, fmt=None)
    put(ws, f'B{r}', STK['horizons']['1M'][k], BLUE, PCT)
    put(ws, f'C{r}', STK['horizons']['3M'][k], BLUE, PCT)
    r += 1
r += 1
hdr(ws, r, ['Engine setting', 'Value']); r += 1
for lab, v, fmt, green in [('Simulated paths', 50000, NUM0, False),
                           ('Annualised volatility (3-month anchor)',
                            STK['horizons']['3M']['anchor_vol_ann'], PCT, False),
                           ('Spot price (EGP)', f'=Summary!C{SPOT_ROW}', PX, True),
                           ('Anchor date', STK['anchor_date'], None, False)]:
    put(ws, f'A{r}', lab, fmt=None)
    if green:
        putf(ws, f'C{r}', v, SPOT, fmt, green=True)
    else:
        put(ws, f'C{r}', v, BLUE, fmt)
    r += 1

# ============ 14 SENSITIVITY ============================================================
ws = sheet('Sensitivity')
title(ws, 'Sensitivity — what the valuation needs the world to do', 'EGP per share. Each cell is a '
      'complete re-run of the model, including the unit build, so these grids are engine outputs '
      'rather than formulas and do NOT redraw when a driver is changed.', 8, awidth=40, cwidth=13)
r = 4
put(ws, f'A{r}', 'Terminal cost of capital (rows) x terminal growth (columns)', bold=True, fmt=None)
r += 1
hdr(ws, r, [''] + [f'{g:.0%}' for g in SN['g_grid']]); r += 1
for i, wt in enumerate(SN['wt_grid']):
    put(ws, f'A{r}', f'{wt:.1%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid_wacc_g'][i][j], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Explicit-window cost of capital (columns) x terminal cost of capital (rows)',
    bold=True, fmt=None); r += 1
hdr(ws, r, [''] + [f'{x:.1%}' for x in SN['we_grid']]); r += 1
for i, wt in enumerate(SN['wt_grid']):
    put(ws, f'A{r}', f'terminal {wt:.1%}', fmt=None)
    for j in range(5):
        put(ws, f'{get_column_letter(2+j)}{r}', SN['grid_exp_term'][j][i], BLUE, PX)
    r += 1
r += 1
put(ws, f'A{r}', 'Single-driver sensitivities — five engine re-runs per row; the parameter grid '
    'for each row is shown beside its name', bold=True, fmt=None); r += 1
hdr(ws, r, ['Driver (parameter grid)', '', '', '', '', '', '', 'Swing']); r += 1
for lab, grid, vals, gfmt in [
        ('Beta', SN['beta_grid'], SN['grid_beta'], '{:.2f}'),
        ('Exchange-rate path multiplier', SN['fx_grid'], SN['grid_fx'], '{:.2f}x'),
        ('Segment margins, multiplicative', SN['mg_grid'], SN['grid_margin'], '{:.3f}x'),
        ('Copper price multiplier', SN['cu_grid'], SN['grid_copper'], '{:.3f}x'),
        ('Working capital / revenue', SN['nwc_grid'], SN['grid_nwc'], '{:.1%}'),
        ('Terminal return on invested capital', SN['roic_grid'], SN['grid_roic'], '{:.1%}'),
        ('Terminal growth (at base cost of capital)', SN['g_grid'], SN['grid_wacc_g'][2], '{:.0%}')]:
    put(ws, f'A{r}', f"{lab}  ({' / '.join(gfmt.format(g) for g in grid)})", fmt=None)
    vv = vals[:6]
    for j, v in enumerate(vv):
        put(ws, f'{get_column_letter(2+j)}{r}', v, BLUE, PX)
    last = get_column_letter(1 + len(vv))
    putf(ws, f'H{r}', f'=MAX(B{r}:{last}{r})-MIN(B{r}:{last}{r})', max(vv) - min(vv), PX)
    r += 1
ws.column_dimensions['H'].width = 13
ws.column_dimensions['A'].width = 52

# ============ 15 PER-SHARE & RATIOS =========================================================
ws = sheet('Per-Share & Ratios')
title(ws, 'Per-share and ratio analysis', 'The indicator set for a diversified industrial with a '
      'contracting arm. Every ratio is a formula off the statements.', 9, awidth=44, cwidth=12)
hdr(ws, 4, ['Measure'] + YH + YF)
r = 5

def ratio(lab, fml, vals, fmt, skip=()):
    global r
    put(ws, f'A{r}', lab, fmt=None)
    for i in range(8):
        if i in skip or vals[i] is None:
            put(ws, f'{ALL[i]}{r}', '-', BLACK, fmt)
        else:
            putf(ws, f'{ALL[i]}{r}', fml(i), vals[i], fmt)
    r += 1

eq_all = [HB[y]['eqp'] for y in H3] + F['equity']
fin_all = [HI[y]['fin'] for y in H3] + [-x for x in F['interest']]
ratio('Earnings per share (EGP)', lambda i: f"='Income Statement'!{ALL[i]}18",
      [x / SH for x in npa_all], PX)
ratio('Book value per share (EGP)', lambda i: f"='Balance Sheet'!{ALL[i]}14/{a('shares')}",
      [x / SH for x in eq_all], PX)
ratio('Free cash flow per share (EGP)', lambda i: f"='Summary Financials'!{ALL[i]}11/{a('shares')}",
      [None] * 3 + [x / SH for x in F['fcff']], PX, skip=(0, 1, 2))
ratio('Gross margin', lambda i: f"='Income Statement'!{ALL[i]}6/'Income Statement'!{ALL[i]}5",
      [HI[y]['gp'] / HI[y]['rev'] for y in H3] + [BU['gp'][i] / F['rev'][i] for i in range(5)], PCT)
ratio('EBITDA margin', lambda i: f"='Income Statement'!{ALL[i]}8",
      [eb_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('EBIT margin', lambda i: f"='Income Statement'!{ALL[i]}10/'Income Statement'!{ALL[i]}5",
      [ebit_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Net margin (attributable)',
      lambda i: f"='Income Statement'!{ALL[i]}17/'Income Statement'!{ALL[i]}5",
      [npa_all[i] / rev_all[i] for i in range(8)], PCT)
ratio('Return on equity',
      lambda i: (f"='Income Statement'!{ALL[i]}17/(('Balance Sheet'!{ALL[i-1]}14+"
                 f"'Balance Sheet'!{ALL[i]}14)/2)"),
      [None] + [npa_all[i] / ((eq_all[i - 1] + eq_all[i]) / 2) for i in range(1, 8)], PCT, skip=(0,))
ratio('Return on invested capital', lambda i: f"='Summary Financials'!{ALL[i]}14",
      [None] * 3 + F['roic'], PCT, skip=(0, 1, 2))
ratio('Net debt / EBITDA', lambda i: f"='Balance Sheet'!{ALL[i]}18",
      [nd_all[i] / eb_all[i] for i in range(8)], MULT)
ratio('Interest cover (EBIT / net interest)',
      lambda i: f"=-'Income Statement'!{ALL[i]}10/'Income Statement'!{ALL[i]}11",
      [None] + [ebit_all[i] / abs(fin_all[i]) for i in range(1, 8)], MULT, skip=(0,))
ratio('Working capital / revenue',
      lambda i: f"='Balance Sheet'!{ALL[i]}16/'Income Statement'!{ALL[i]}5",
      [HB[y]['nwc'] / HI[y]['rev'] for y in H3] + [F['nwc'][i] / F['rev'][i] for i in range(5)], PCT)
_CFCAPCOL = [None, 'B', 'C'] + CFF   # Cash Flow sheet's historical columns are now FY2024/FY2025
ratio('Capital expenditure / revenue',
      lambda i: f"=-'Cash Flow'!{_CFCAPCOL[i]}11/'Income Statement'!{ALL[i]}5",
      [None, IN['capex_fy24'] / HI['FY24']['rev'], IN['capex_fy25'] / HI['FY25']['rev']]
      + [F['capex'][i] / F['rev'][i] for i in range(5)], PCT, skip=(0,))
# the last two rows are historical reconciliation constants from the terminal-growth check
put(ws, f'A{r}', 'Reinvestment rate', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{ALL[i]}{r}', TR['rr'][y], BLUE, PCT)
for i in range(3, 8):
    put(ws, f'{ALL[i]}{r}', '-', BLACK, PCT)
r += 1
put(ws, f'A{r}', 'Implied growth (return on capital x reinvestment)', fmt=None)
for i, y in enumerate(H3):
    put(ws, f'{ALL[i]}{r}', TR['implied_g'][y], BLUE, PCT)
for i in range(3, 8):
    put(ws, f'{ALL[i]}{r}', '-', BLACK, PCT)
r += 2
put(ws, f'A{r}', 'Terminal growth reconciliation: actual NOPAT compound growth FY2023–FY2025 was '
    f"{TR['nopat_cagr']:.1%}; the implied growth from stable years only (FY2024 excluded as a "
    f"debt-funded capacity burst) is {TR['stable_g']:.1%}; the adopted terminal growth is "
    f"{IN['g_term']:.1%}, below the blended nominal ceiling of {TR['ceiling']:.1%}.",
    fmt=None).font = SUB

# ============ 16 PEER & SECTOR ===============================================================
ws = sheet('Peer & Sector')
title(ws, 'Peer frame and sector context', 'No clean comparable exists; this is a sanity check, '
      'not an independent valuation', 6, awidth=34, cwidth=22)
hdr(ws, 4, ['Company / frame', 'Market', 'Relevance', 'Caution'])
r = 5
for a1, a2, a3, a4 in [
    ('Riyadh Cables', 'Saudi Arabia', 'nearest listed regional cable manufacturer',
     'far smaller, no contracting arm, lighter balance sheet, pegged currency'),
    ('Electro Cable Egypt', 'Egypt', 'the only other listed Egyptian cable manufacturer',
     'much smaller, domestic, heavily levered, currently loss-making'),
    ('European cable majors', 'Europe', 'closest match on business model — cables plus projects',
     'developed-market cost of capital; no convertibility risk'),
    ('Regional engineering and construction contractors', 'Gulf and North Africa',
     'the right frame for the roughly 32% that is the Constructions and infrastructure segment',
     'project accounting and backlog quality are not comparable'),
]:
    put(ws, f'A{r}', a1, fmt=None); put(ws, f'B{r}', a2, fmt=None)
    put(ws, f'C{r}', a3, fmt=None, wrap=True); put(ws, f'D{r}', a4, fmt=None, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1
ws.column_dimensions['C'].width = 40; ws.column_dimensions['D'].width = 44
r += 1
hdr(ws, r, ['Own multiples', 'Value']); r += 1
for lab, v, xp, fmt in [
        ('Trailing enterprise value / EBITDA', "='Relative & Normalized'!C13",
         REL['ev_ebitda_trailing'], MULT),
        ('Trailing price / earnings', "='Relative & Normalized'!C14", REL['pe_trailing'], MULT),
        ('Trailing price / book', "='Relative & Normalized'!C15", SPOT / BK['bvps'], MULT),
        ('Justified enterprise value / EBITDA applied', "='Relative & Normalized'!C6",
         IN['ev_ebitda_just'], MULT),
        ('Justified price / earnings applied', "='Relative & Normalized'!C27", IN['pe_just'], MULT)]:
    put(ws, f'A{r}', lab, fmt=None)
    putf(ws, f'B{r}', v, xp, fmt, green=True); r += 1

out = os.path.join(HERE, 'SWDY_Valuation_Model_05082026_public.xlsx')
wb.save(out)
json.dump({'expected': EXPECT, 'anchors': ANCH},
          open(os.path.join(HERE, 'xlsx_expected.json'), 'w'), indent=1)
nchk = sum(len(v) for v in EXPECT.values())
nform = nlit = 0
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                nform += 1
            elif isinstance(c.value, (int, float)):
                nlit += 1
print(f"wrote {out} | {len(wb.sheetnames)} sheets: {wb.sheetnames}")
print(f"formulas: {nform} (of which {nchk} carry a checked expected value) | numeric literals: {nlit}")
